"""Pruebas del verificador de cargue de SIIFA.

Lo que se cuida acá es que la hoja de verificación diga la verdad: que el
bot haya contestado OK no prueba nada en una conciliación —lo que vale es
lo que SIIFA tiene guardado—. Y sobre todo la FECHA: una respuesta que
quedó registrada con la fecha corrida es una respuesta que la EPS puede
alegar como dada fuera del término.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

TOOLS = Path(__file__).resolve().parents[2] / "tools"
sys.path.insert(0, str(TOOLS))

import siifa_verificar_cargue as ver  # noqa: E402


def _excel_cargado(tmp_path: Path, filas: list[dict]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "RESPUESTAS"
    cols = [
        "ID_SEGUIMIENTO_FACTURA_GLOSA",
        "NUMERO_FACTURA",
        "CODIGO_GLOSA",
        "VALOR_GLOSA",
        "CODIGO_RESPUESTA",
        "FECHA_RESPUESTA",
        "OBSERVACION_RESPUESTA",
    ]
    ws.append(cols)
    for f in filas:
        ws.append([f.get(c) for c in cols])
    ruta = tmp_path / "respuestas_GLOSAS.xlsx"
    wb.save(ruta)
    return ruta


def _cargada(ident=15110544, factura="HUS454747", codigo="RE9602", fecha="2026-05-11"):
    return {
        "ID_SEGUIMIENTO_FACTURA_GLOSA": ident,
        "NUMERO_FACTURA": factura,
        "CODIGO_GLOSA": "TA0801",
        "VALOR_GLOSA": 22200,
        "CODIGO_RESPUESTA": codigo,
        "FECHA_RESPUESTA": fecha,
        "OBSERVACION_RESPUESTA": "ESE HUS NO ACEPTA...",
    }


def _en_siifa(codigo="RE9602", fecha="2026-05-11T00:00:00", tiene="SI"):
    return {
        "tiene_respuesta": tiene,
        "codigo_respuesta": codigo,
        "fecha_respuesta": fecha,
        "observacion_respuesta": "ESE HUS NO ACEPTA...",
    }


def test_lo_que_quedo_igual_sale_como_registrada():
    fila = ver.comparar(
        {
            "id": 1,
            "factura": "HUS454747",
            "codigo_glosa": "TA0801",
            "valor_glosa": 22200,
            "codigo": "RE9602",
            "fecha": "2026-05-11",
        },
        _en_siifa(),
    )

    assert fila["RESULTADO"] == ver.REGISTRADA
    assert fila["DIFERENCIAS"] == ""


def test_una_glosa_que_sigue_sin_respuesta_se_marca_para_volver_a_subir():
    fila = ver.comparar(
        {
            "id": 1,
            "factura": "HUS454747",
            "codigo_glosa": "TA0801",
            "valor_glosa": 22200,
            "codigo": "RE9602",
            "fecha": "2026-05-11",
        },
        _en_siifa(tiene="NO", codigo="", fecha=""),
    )

    assert fila["RESULTADO"] == ver.SIN_RESPUESTA
    assert "volver a subirla" in fila["DIFERENCIAS"]


def test_si_la_fecha_quedo_corrida_se_avisa_del_termino():
    """El riesgo concreto: que la EPS alegue respuesta fuera de término."""
    fila = ver.comparar(
        {
            "id": 1,
            "factura": "HUS454747",
            "codigo_glosa": "TA0801",
            "valor_glosa": 22200,
            "codigo": "RE9602",
            "fecha": "2026-05-11",
        },
        _en_siifa(fecha="2026-08-03T00:00:00"),
    )

    assert fila["RESULTADO"] == ver.CON_DIFERENCIAS
    assert "2026-08-03" in fila["DIFERENCIAS"]
    assert "término" in fila["DIFERENCIAS"]


def test_si_el_codigo_quedo_distinto_tambien_se_avisa():
    fila = ver.comparar(
        {
            "id": 1,
            "factura": "HUS454747",
            "codigo_glosa": "TA0801",
            "valor_glosa": 22200,
            "codigo": "RE9602",
            "fecha": "2026-05-11",
        },
        _en_siifa(codigo="RE9901"),
    )

    assert fila["RESULTADO"] == ver.CON_DIFERENCIAS
    assert "RE9901" in fila["DIFERENCIAS"]


def test_una_glosa_que_siifa_no_devuelve_no_se_da_por_buena():
    fila = ver.comparar(
        {
            "id": 99,
            "factura": "HUS454747",
            "codigo_glosa": "TA0801",
            "valor_glosa": 22200,
            "codigo": "RE9602",
            "fecha": "2026-05-11",
        },
        None,
    )

    assert fila["RESULTADO"] == ver.NO_ESTA


def test_se_lee_el_mismo_excel_que_uso_el_bot_de_cargue(tmp_path):
    ruta = _excel_cargado(tmp_path, [_cargada(), _cargada(ident=15110545)])

    cargado = ver.leer_cargado(ruta)

    assert [c["id"] for c in cargado] == [15110544, 15110545]
    assert cargado[0]["fecha"] == "2026-05-11"
    assert cargado[0]["codigo"] == "RE9602"


def test_con_el_reporte_del_cargue_solo_se_verifica_lo_que_dio_ok(tmp_path):
    reporte = tmp_path / "reporte.csv"
    reporte.write_text(
        "id_seguimiento_factura_glosa,numero_factura,estado,detalle,fecha_hora\n"
        "15110544,HUS454747,OK,,2026-08-03T16:43:43\n"
        "15110545,HUS454747,ERROR,se cayo la conexion,2026-08-03T16:43:44\n",
        encoding="utf-8",
    )

    assert ver.ids_del_reporte(reporte) == {15110544}


def test_la_hoja_pinta_de_rojo_lo_que_no_quedo(tmp_path):
    """El auditor tiene que ver de un vistazo qué falta por volver a subir."""
    filas = [
        ver.comparar(
            {
                "id": 1,
                "factura": "HUS454747",
                "codigo_glosa": "TA0801",
                "valor_glosa": 22200,
                "codigo": "RE9602",
                "fecha": "2026-05-11",
            },
            _en_siifa(),
        ),
        ver.comparar(
            {
                "id": 2,
                "factura": "HUS454747",
                "codigo_glosa": "TA0801",
                "valor_glosa": 22200,
                "codigo": "RE9602",
                "fecha": "2026-05-11",
            },
            _en_siifa(tiene="NO"),
        ),
    ]
    salida = tmp_path / "verificacion.xlsx"

    ver.escribir(filas, salida)

    ws = load_workbook(salida).active
    col = [c.value for c in ws[1]].index("RESULTADO") + 1
    assert ws.cell(row=2, column=col).value == ver.REGISTRADA
    assert ws.cell(row=3, column=col).value == ver.SIN_RESPUESTA
    assert (
        ws.cell(row=2, column=col).fill.fgColor.rgb != ws.cell(row=3, column=col).fill.fgColor.rgb
    )


def test_la_constancia_sale_una_por_factura(tmp_path):
    """Es lo que reemplaza al pantallazo: se anexa a los soportes."""
    import pytest

    pytest.importorskip("reportlab")

    filas = [
        ver.comparar(
            {
                "id": i,
                "factura": factura,
                "codigo_glosa": "TA0801",
                "valor_glosa": 22200,
                "codigo": "RE9602",
                "fecha": "2026-05-11",
            },
            _en_siifa(),
        )
        for i, factura in ((1, "HUS454747"), (2, "HUS454747"), (3, "HUS497119"))
    ]
    carpeta = tmp_path / "EVIDENCIAS"

    hechas = ver.constancias(filas, carpeta, "03/08/2026 17:00")

    assert hechas == 2
    assert (carpeta / "CONSTANCIA_SIIFA_HUS454747.pdf").is_file()
    assert (carpeta / "CONSTANCIA_SIIFA_HUS497119.pdf").is_file()
