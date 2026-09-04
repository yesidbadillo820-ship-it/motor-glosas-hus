"""La directriz del 03-09-2026: las objeciones CL de glosas Médica/Mixta NO se
responden con el bot (van a gestión manual del equipo médico, para que la nota
crédito pueda cruzar). El generador debe: omitirlas de "Respuestas Glosa",
aislarlas en la hoja "PARA GESTION MEDICA" y conservar la numeración de la
grilla para las que sí se responden."""

import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

GEN = Path(__file__).resolve().parents[2] / "tools" / "glosas_dispensario" / "gen_lote.py"

COLS = [
    "EstadoCxCObjecion",
    "TipoObjecionTramite",
    "FacturaCartera.PlanBeneficio.Contrato.Entidad.NombreEntidad",
    "FacturaCartera.Factura",
    "ListadoConceptos.ConceptoObjecion.Codigo",
    "ListadoConceptos.ServicioProductoFactura.Codigo",
    "ListadoConceptos.ServicioProductoFactura.Descripcion",
    "ListadoConceptos.ValorObjecion",
    "ValorObjecion",
    "ListadoConceptos.Observaciones",
]

ENT = "DISPENSARIO MEDICO BUCARAMANGA MEBUG"


def _fila(tipo, code, valor, obs):
    return [
        "Glosa_Inicial",
        tipo,
        ENT,
        "HUS0000500001",
        code,
        "890201",
        "CONSULTA",
        valor,
        valor,
        obs,
    ]


def _correr_gen(tmp_path, filas):
    src = tmp_path / "export.xlsx"
    out = tmp_path / "salida.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "i"
    ws.append(COLS)
    for f in filas:
        ws.append(f)
    wb.save(src)
    r = subprocess.run(
        [sys.executable, str(GEN), str(src), str(out)],
        capture_output=True,
        text=True,
        cwd=GEN.parent,
    )
    assert r.returncode == 0, r.stderr
    return load_workbook(out), r.stdout


def test_cl_medica_y_mixta_se_excluyen_y_la_numeracion_no_se_corre(tmp_path):
    wb, stdout = _correr_gen(
        tmp_path,
        [
            _fila("Administrativo", "TA0601", 1000, "SE GLOSA MAYOR VALOR"),
            _fila("Mixta", "CL0801", 2000, "SE GLOSA EXAMEN NO PERTINENTE"),
            _fila("Medico", "CL0101", 3000, "SE GLOSA ESTANCIA"),
            _fila("Administrativo", "TA0201", 4000, "SE GLOSA MAYOR VALOR COBRADO"),
        ],
    )
    filas = list(wb["Respuestas Glosa"].iter_rows(min_row=2, values_only=True))
    numeros = [(f[1], f[2]) for f in filas]
    # solo las 2 administrativas, con su numero REAL de grilla (1 y 4, no 1 y 2)
    assert numeros == [(1, "TA0601"), (4, "TA0201")]
    assert all("CL" not in str(f[2]) for f in filas)
    # las 2 CL quedan aisladas para el equipo medico, con su numero de grilla
    med = list(wb["PARA GESTION MEDICA"].iter_rows(min_row=2, values_only=True))
    assert [(m[1], m[2]) for m in med] == [(2, "CL0801"), (3, "CL0101")]
    assert "EXCLUIDAS PARA GESTION MEDICA" in stdout


def test_cl_administrativa_si_se_responde(tmp_path):
    wb, _ = _correr_gen(
        tmp_path,
        [_fila("Administrativo", "CL0301", 5000, "SE GLOSA ESTANCIA COBRADA")],
    )
    filas = list(wb["Respuestas Glosa"].iter_rows(min_row=2, values_only=True))
    assert [(f[1], f[2]) for f in filas] == [(1, "CL0301")]
    assert "PARA GESTION MEDICA" not in wb.sheetnames
