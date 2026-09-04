"""Pruebas del bot RPA del lote Dispensario: cruce tarifario del contrato 440,
lectura en cascada de la factura, búsqueda de soportes y enriquecimiento del
Excel sin inventar datos."""

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

RAIZ = Path(__file__).resolve().parents[2] / "tools" / "glosas_dispensario"
sys.path.insert(0, str(RAIZ))

from bot_lote_dispensario import buscar_pdf_factura, enriquecer_excel, frase_anclaje  # noqa: E402
from tarifario_440 import cargar_tarifario, frase_tarifaria, tarifa_de  # noqa: E402


# ── Tarifario del contrato 440 ──────────────────────────────────────────────
def _tarifario_servicios(ruta):
    wb = Workbook()
    ws = wb.active
    ws.title = "SERVICIOS DE PROCEDIMIENTOS"
    ws.append(
        [
            "ITEM",
            "CUPS",
            "DESCRIPCION CUPS",
            "CODIGO IPS",
            "DESCRIPCION IPS",
            "PRECIO DE REFERENCIA",
            "TARIFA A LA QUE CORRESPONDE",
        ]
    )
    ws.append(
        [1, "039001", "INSERCION DE CATETER", "039001H", "INSERCION DE CATETER", 1689585, "PROPIA"]
    )
    ws.append([2, "902210", "HEMOGRAMA IV", "902210AMB", "HEMOGRAMA IV", 17556.5, "PROPIA"])
    wb.save(ruta)


def _tarifario_medicamentos(ruta):
    wb = Workbook()
    ws = wb.active
    ws.title = "ANEXO 01-MX REGULADOS"
    ws.append(["CODIGO CUM", "CÓDIGO ATC", "NOMBRE", "PRECIO DE VENTA", "NORMATIVA VIGENTE"])
    ws.append(["20103720-02", "J05AR02", "ABACAVIR 600 MG", 31361, "CIRCULAR 19 DE 2024"])
    wb.save(ruta)


def test_tarifario_cruza_por_codigo_ips_cups_y_cum(tmp_path):
    serv, mx = tmp_path / "serv.xlsx", tmp_path / "mx.xlsx"
    _tarifario_servicios(serv)
    _tarifario_medicamentos(mx)
    idx = cargar_tarifario(serv, mx)

    assert tarifa_de(idx, "039001H")["precio"] == 1689585  # codigo IPS exacto
    assert tarifa_de(idx, "039001")["precio"] == 1689585  # CUPS pelado
    assert (
        tarifa_de(idx, "902210amb")["precio"] == 17557
        or tarifa_de(idx, "902210amb")["precio"] == 17556
    )
    assert tarifa_de(idx, "20103720-02")["precio"] == 31361  # CUM completo
    assert tarifa_de(idx, "20103720")["precio"] == 31361  # CUM sin presentacion
    assert tarifa_de(idx, "999999") is None  # sin pacto: no se cita nada
    frase = frase_tarifaria("039001H", tarifa_de(idx, "039001H"))
    assert "CONTRATO 440-DIGSA-DMBUG-2025" in frase and "$1.689.585" in frase


def test_tarifario_sin_archivos_devuelve_indice_vacio(tmp_path):
    assert cargar_tarifario(tmp_path / "no.xlsx", None) == {}


# ── Busqueda de soportes en las carpetas de radicacion ──────────────────────
def test_busca_el_pdf_en_el_mes_mas_reciente_primero(tmp_path):
    viejo = tmp_path / "3. MARZO 2026 - SOPORTES RADICACION"
    nuevo = tmp_path / "9.SEPTIEMBRE - SOPORTES RADICACION"
    (viejo / "sub").mkdir(parents=True)
    nuevo.mkdir()
    (viejo / "sub" / "FEV_HUS0000540273.pdf").write_bytes(b"%PDF viejo")
    (nuevo / "HUS0000540273 RADICADO.pdf").write_bytes(b"%PDF nuevo")
    hallado = buscar_pdf_factura("HUS0000540273", [str(viejo), str(nuevo)])
    assert hallado is not None and hallado.parent == nuevo
    assert buscar_pdf_factura("HUS0000999999", [str(viejo), str(nuevo)]) is None


def test_indice_toma_la_factura_electronica_de_la_carpeta_de_la_factura(tmp_path):
    """La estructura real de la unidad Y: una carpeta por factura con varios
    soportes adentro. El que sirve para el cotejo es el FEV (la factura)."""
    from bot_lote_dispensario import buscar_en_indice, indexar_soportes

    mes = tmp_path / "9.SEPTIEMBRE - SOPORTES RADICACION"
    carpeta = mes / "DISPENSARIO" / "LILIANA" / "ENV-233972-OK" / "HUS552002"
    carpeta.mkdir(parents=True)
    (carpeta / "epicrisis.pdf").write_bytes(b"%PDF")
    (carpeta / "FEV_900006037_HUS552002.pdf").write_bytes(b"%PDF")
    (carpeta / "anexo autorizacion HUS552002.pdf").write_bytes(b"%PDF")

    indice = indexar_soportes([str(mes)])
    hallado = buscar_en_indice(indice, "HUS0000552002")
    assert hallado is not None and hallado.name == "FEV_900006037_HUS552002.pdf"
    # los demás soportes de la carpeta quedan como alternativas, no se pierden
    assert len(indice["552002"]) == 3
    assert buscar_en_indice(indice, "HUS0000552003") is None


# ── No invencion y enriquecimiento del Excel ────────────────────────────────
def test_sin_datos_leidos_no_hay_frase_de_anclaje():
    assert frase_anclaje("HUS0000500001", dict(paciente=None, total=None)) is None
    frase = frase_anclaje("HUS0000500001", dict(paciente="PEREZ JUAN", total=123456))
    assert "PEREZ JUAN" in frase and "$123.456" in frase


def test_enriquecer_inserta_antes_del_cierre_y_es_idempotente(tmp_path):
    ruta = tmp_path / "resp.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Respuestas Glosa"
    ws.append(
        [
            "Factura",
            "# Objeción",
            "Cód.",
            "Servicio",
            "Valor Objetado",
            "Valor Aceptado",
            "Cod Respuesta",
            "Detalle Respuesta",
        ]
    )
    detalle = "ESE HUS NO ACEPTA LA GLOSA... POR LO EXPUESTO, SE SOLICITA EL LEVANTAMIENTO TOTAL."
    ws.append(["HUS0000500001", 1, "TA0601", "890201 - CONSULTA", 1000, 0, "RE9901", detalle])
    ws.append(["HUS0000500001", 3, "TA0201", "890301 - CONSULTA", 2000, 0, "RE9901", detalle])
    wb.save(ruta)

    tocadas = enriquecer_excel(
        ruta,
        por_factura={"HUS0000500001": "LA FACTURA SE ANEXA."},
        por_linea={("HUS0000500001", 3): "LA TARIFA PACTADA ES DE $9.999."},
    )
    assert tocadas == 2
    filas = list(load_workbook(ruta)["Respuestas Glosa"].iter_rows(min_row=2, values_only=True))
    # la frase queda ANTES del cierre institucional
    assert filas[0][7].index("LA FACTURA SE ANEXA.") < filas[0][7].index("POR LO EXPUESTO")
    assert "LA TARIFA PACTADA" not in filas[0][7]
    assert filas[1][7].index("LA TARIFA PACTADA ES DE $9.999.") < filas[1][7].index(
        "LA FACTURA SE ANEXA."
    )
    # repetir no duplica
    assert enriquecer_excel(ruta, {"HUS0000500001": "LA FACTURA SE ANEXA."}, {}) == 0


# ── Lectura en cascada (pdfplumber como primer intento) ─────────────────────
def test_cascada_lee_paciente_y_total_con_pdfplumber(tmp_path):
    pymupdf = pytest.importorskip("pymupdf")
    pytest.importorskip("pdfplumber")
    from extraer_factura_pdf import extraer_datos_factura

    ruta = tmp_path / "factura.pdf"
    doc = pymupdf.open()
    pagina = doc.new_page()
    pagina.insert_text((72, 100), "FACTURA ELECTRONICA DE VENTA HUS0000540273")
    pagina.insert_text((72, 130), "PACIENTE: RODRIGUEZ GOMEZ MARIA FERNANDA")
    pagina.insert_text((72, 160), "VALOR TOTAL: $ 1.234.567")
    doc.save(ruta)
    doc.close()

    datos = extraer_datos_factura(ruta)
    assert datos["ok"] and datos["metodo"] in ("pdfplumber", "PyPDF2")
    assert datos["paciente"] == "RODRIGUEZ GOMEZ MARIA FERNANDA"
    assert datos["total"] == 1234567


def test_cascada_sin_texto_ni_ocr_no_inventa(tmp_path):
    pymupdf = pytest.importorskip("pymupdf")
    from extraer_factura_pdf import extraer_datos_factura

    ruta = tmp_path / "escaneada.pdf"
    doc = pymupdf.open()
    doc.new_page()  # pagina en blanco: nada que leer sin OCR
    doc.save(ruta)
    doc.close()

    datos = extraer_datos_factura(ruta)
    assert datos["paciente"] is None and datos["total"] is None
    assert not datos["ok"] and datos["metodo"].startswith("SIN_LECTURA")


# ── El constructor con trazabilidad (parrafo de evidencia) ──────────────────
def _datos_pdf(**kw):
    base = dict(
        archivo="HUS0000500001.pdf",
        metodo="pdfplumber",
        ok=True,
        paciente=None,
        total=None,
        servicios=[],
    )
    base.update(kw)
    return base


def test_parrafo_completo_cita_pdf_paciente_servicio_y_coincidencia():
    from bot_lote_dispensario import parrafo_evidencia

    tarifa = dict(precio=180000, descripcion="CONSULTA NEUROLOGIA", fuente="anexo 6.2 · SOAT")
    p = parrafo_evidencia(
        _datos_pdf(paciente="RODRIGUEZ GOMEZ MARIA", total=500000),
        ("890275H CONSULTA NEUROLOGIA 180000", 180000),
        tarifa,
        valor_objetado=999,
        cups="890275H",
    )
    assert p.index("AL REVISAR EL DOCUMENTO DE SOPORTE HUS0000500001.PDF") == 0
    assert "SE EVIDENCIA LA ATENCION PRESTADA AL USUARIO RODRIGUEZ GOMEZ MARIA" in p
    assert (
        "EL DOCUMENTO DETALLA EL COBRO DEL SERVICIO 890275H CONSULTA NEUROLOGIA 180000 POR VALOR DE $180.000"
        in p
    )
    assert "AL CRUZAR ESTA EVIDENCIA CON EL CONTRATO 440-DIGSA-DMBUG-2025" in p
    assert "CORRESPONDE EXACTAMENTE A LA TARIFA PACTADA EN EL ANEXO 6.2 ($180.000" in p
    assert "CAUSAL DE GLOSA ES INFUNDADA" in p


def test_parrafo_sin_coincidencia_no_afirma_exactitud():
    from bot_lote_dispensario import parrafo_evidencia

    tarifa = dict(precio=180000, descripcion="CONSULTA", fuente="anexo 6.2 · SOAT")
    p = parrafo_evidencia(
        _datos_pdf(paciente="PEREZ JUAN"),
        ("CONSULTA 175000", 175000),  # valor leido DISTINTO de la tarifa
        tarifa,
        valor_objetado=999,
        cups="890275H",
    )
    assert "EXACTAMENTE" not in p and "INFUNDADA" not in p
    # sin cotejo NO se proclama la cifra del anexo (evita ratificacion con
    # el propio soporte): se cita la fila pactada, sin valor
    assert "$180.000" not in p
    assert "SE ENCUENTRA PACTADO EN EL ANEXO TARIFARIO" in p and "890275H" in p


def test_parrafo_valor_objetado_igual_a_tarifa_sin_pdf():
    from bot_lote_dispensario import parrafo_evidencia

    tarifa = dict(precio=35136, descripcion="HEMOGRAMA", fuente="anexo 6.2 · LAB")
    p = parrafo_evidencia(None, None, tarifa, valor_objetado=35136, cups="902210AMB")
    assert p.startswith("AL VERIFICAR EL CONTRATO 440-DIGSA-DMBUG-2025")
    assert "VALOR OBJETADO CORRESPONDE EXACTAMENTE" in p


def test_parrafo_sin_datos_ni_tarifa_es_none():
    from bot_lote_dispensario import parrafo_evidencia

    assert parrafo_evidencia(None, None, None, 1000, "X") is None
    assert parrafo_evidencia(_datos_pdf(), None, None, 1000, "X") is None  # PDF sin datos legibles


def test_hallar_servicio_en_pdf_por_codigo_y_por_descripcion():
    from bot_lote_dispensario import hallar_servicio_en_pdf

    filas = [
        ["1", "890201", "CONSULTA GENERAL", "35.136"],
        ["2", "873501H", "FLUOROSCOPIA COMO GUIA", "137.591"],
    ]
    s = hallar_servicio_en_pdf(filas, "873501H", "OTRA COSA")
    assert s and "FLUOROSCOPIA" in s[0] and s[1] == 137591
    s2 = hallar_servicio_en_pdf(filas, "999999", "CONSULTA GENERAL MEDICINA")
    assert s2 and "CONSULTA GENERAL" in s2[0] and s2[1] == 35136
    assert hallar_servicio_en_pdf(filas, "111111", "NADA QUE VER AQUI") is None


def test_el_codigo_del_servicio_no_se_lee_como_si_fuera_plata():
    """890275H no son $890.275: confundirlos hacía ver sobrecobros falsos."""
    from bot_lote_dispensario import hallar_servicio_en_pdf

    fila = [["890275H", "CONSULTA", "DE", "PRIMERA", "VEZ", "1", "192.600", "192.600"]]
    s = hallar_servicio_en_pdf(fila, "890275H", "CONSULTA DE PRIMERA VEZ")
    assert s and s[2] == [192600, 192600] and s[1] == 192600
    # el número de la factura tampoco es un valor
    fila2 = [["HUS0000542497", "890201", "CONSULTA", "35.136"]]
    s2 = hallar_servicio_en_pdf(fila2, "890201", "CONSULTA")
    assert s2 and s2[2] == [35136]


def test_lee_el_detalle_aunque_el_pdf_no_dibuje_la_tabla(tmp_path):
    """Muchas facturas electrónicas no tienen tabla: el renglón del cobro se
    rescata del texto, o el cotejo se quedaría sin con qué comparar."""
    pymupdf = pytest.importorskip("pymupdf")
    pytest.importorskip("pdfplumber")
    from bot_lote_dispensario import hallar_servicio_en_pdf
    from extraer_factura_pdf import extraer_datos_factura

    ruta = tmp_path / "factura.pdf"
    doc = pymupdf.open()
    pagina = doc.new_page()
    pagina.insert_text((50, 80), "FACTURA ELECTRONICA DE VENTA HUS0000542497")
    pagina.insert_text((50, 110), "890275H CONSULTA DE PRIMERA VEZ 1 192.600 192.600")
    doc.save(ruta)
    doc.close()

    datos = extraer_datos_factura(ruta)
    assert datos["servicios"], "sin tabla dibujada el detalle debe salir del texto"
    s = hallar_servicio_en_pdf(datos["servicios"], "890275H", "CONSULTA DE PRIMERA VEZ")
    assert s and s[2] == [192600, 192600]


def test_apertura_entra_despues_del_encabezado_y_antes_del_argumento(tmp_path):
    ruta = tmp_path / "resp.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Respuestas Glosa"
    ws.append(
        [
            "Factura",
            "# Objeción",
            "Cód.",
            "Servicio",
            "Valor Objetado",
            "Valor Aceptado",
            "Cod Respuesta",
            "Detalle Respuesta",
        ]
    )
    detalle = (
        "ESE HUS NO ACEPTA LA GLOSA APLICADA A LA FACTURA HUS0000500001. FRENTE AL CARGO "
        "OBJETADO BAJO EL CONCEPTO TA0601: EL ARGUMENTO DEL MOTOR. POR LO EXPUESTO, "
        "SE SOLICITA EL LEVANTAMIENTO TOTAL."
    )
    ws.append(["HUS0000500001", 1, "TA0601", "890201 - CONSULTA", 1000, 0, "RE9901", detalle])
    wb.save(ruta)

    n = enriquecer_excel(
        ruta, apertura_por_linea={("HUS0000500001", 1): "AL REVISAR EL SOPORTE X."}
    )
    assert n == 1
    d = list(load_workbook(ruta)["Respuestas Glosa"].iter_rows(min_row=2, values_only=True))[0][7]
    k_enc = d.index("BAJO EL CONCEPTO TA0601:")
    assert k_enc < d.index("AL REVISAR EL SOPORTE X.") < d.index("EL ARGUMENTO DEL MOTOR")
    # idempotente
    assert (
        enriquecer_excel(
            ruta, apertura_por_linea={("HUS0000500001", 1): "AL REVISAR EL SOPORTE X."}
        )
        == 0
    )
