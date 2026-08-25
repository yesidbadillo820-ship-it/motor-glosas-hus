"""Pruebas del puente entre el informe de SIIFA y el bot de cargue.

La razón de ser de este script: en el informe real del 03-08-2026 había
2.579 líneas pendientes que se responden con 272 textos distintos, porque
SIIFA registra la misma glosa muchas veces. Lo que se prueba acá es
justamente eso: que agrupe bien, y que al repartir no se pierda ni se
duplique ninguna línea.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

TOOLS = Path(__file__).resolve().parents[2] / "tools"
sys.path.insert(0, str(TOOLS))

import siifa_preparar_respuestas as prep  # noqa: E402

COLS_INFORME = [
    "id_seguimiento_factura_glosa",
    "tipo_seguimiento",
    "numero_factura",
    "valor_glosa",
    "codigo_glosa",
    "descripcion_glosa",
    "observacion_glosa",
    "fecha_reporte",
    "tiene_respuesta",
]


def _informe(tmp_path: Path, filas: list[dict]) -> Path:
    """Arma un informe con la misma forma que el que baja el bot."""
    wb = Workbook()
    ws = wb.active
    ws.title = "SEGUIMIENTOS"
    ws.append(COLS_INFORME)
    for f in filas:
        ws.append([f.get(c) for c in COLS_INFORME])
    ruta = tmp_path / "informe.xlsx"
    wb.save(ruta)
    return ruta


def _linea(id_, factura, codigo, valor=1000, obs="null", respondida="NO", tipo="GLOSA"):
    return {
        "id_seguimiento_factura_glosa": id_,
        "tipo_seguimiento": tipo,
        "numero_factura": factura,
        "valor_glosa": valor,
        "codigo_glosa": codigo,
        "descripcion_glosa": f"Descripción de {codigo}",
        "observacion_glosa": obs,
        "fecha_reporte": "2026-06-09T17:07:47",
        "tiene_respuesta": respondida,
    }


def _leer(ruta: Path) -> tuple[list[str], list[dict]]:
    ws = load_workbook(ruta)[load_workbook(ruta).sheetnames[0]]
    filas = ws.iter_rows(values_only=True)
    cols = [str(c) for c in next(filas)]
    return cols, [dict(zip(cols, f)) for f in filas]


def _llenar(plantilla: Path, respuestas: dict[str, tuple[str, str]]) -> Path:
    """Escribe respuestas en la hoja de trabajo, como haría el auditor.

    `respuestas` va por factura+código, que es como piensa el auditor.
    """
    wb = load_workbook(plantilla)
    ws = wb.active
    hdr = [c.value for c in ws[1]]
    for fila in ws.iter_rows(min_row=2):
        clave = (fila[hdr.index("NUMERO_FACTURA")].value, fila[hdr.index("CODIGO_GLOSA")].value)
        if clave in respuestas:
            cod, obs = respuestas[clave]
            fila[hdr.index("CODIGO_RESPUESTA")].value = cod
            fila[hdr.index("OBSERVACION_RESPUESTA")].value = obs
    wb.save(plantilla)
    return plantilla


def test_las_lineas_repetidas_se_responden_una_sola_vez(tmp_path):
    """Lo del caso real: la misma glosa registrada 80 veces es UNA respuesta."""
    filas = [_linea(i, "HUS454747", "TA5701", valor=16480) for i in range(1, 81)]
    filas.append(_linea(999, "HUS454747", "CO2301", valor=4700))
    datos = prep.leer_informe(_informe(tmp_path, filas))

    grupos = prep.agrupar(datos)

    assert len(grupos) == 2, "80 líneas iguales + 1 distinta = 2 respuestas por escribir"
    assert sum(len(v) for v in grupos.values()) == 81


def test_la_misma_glosa_con_observacion_distinta_se_responde_aparte(tmp_path):
    """Si la EPS puso otra observación, el sustento del hospital cambia."""
    filas = [
        _linea(1, "HUS475438", "DE5601", obs="NO HABILITADO"),
        _linea(2, "HUS475438", "DE5601", obs="El XML no fue enviado al correo"),
    ]
    datos = prep.leer_informe(_informe(tmp_path, filas))

    assert len(prep.agrupar(datos)) == 2


def test_lo_ya_respondido_no_vuelve_a_la_hoja_de_trabajo(tmp_path):
    """Responder dos veces la misma glosa embolata el trámite."""
    filas = [
        _linea(1, "HUS520221", "CO2301", respondida="SI"),
        _linea(2, "HUS520221", "TA0801", respondida="NO"),
    ]
    datos = prep.leer_informe(_informe(tmp_path, filas))

    grupos = prep.agrupar(datos)

    assert len(grupos) == 1
    assert grupos[list(grupos)[0]][0]["codigo_glosa"] == "TA0801"


def test_del_informe_a_la_hoja_de_trabajo_y_de_vuelta_a_todas_las_lineas(tmp_path):
    """El camino completo: agrupar, llenar dos columnas, repartir.

    Es la prueba que importa: el auditor escribe 2 respuestas y el bot
    recibe las 5 líneas que le corresponden, cada una con su id.
    """
    filas = [
        _linea(11, "HUS454747", "TA5701"),
        _linea(12, "HUS454747", "TA5701"),
        _linea(13, "HUS454747", "TA5701"),
        _linea(21, "HUS514051", "CO2301"),
        _linea(22, "HUS514051", "CO2301"),
    ]
    informe = _informe(tmp_path, filas)
    datos = prep.leer_informe(informe)
    grupos = prep.agrupar(datos)

    plantilla = tmp_path / "para_llenar.xlsx"
    prep.escribir_plantilla(grupos, plantilla)
    _, filas_plantilla = _leer(plantilla)
    assert len(filas_plantilla) == 2, "dos respuestas por escribir, no cinco"

    _llenar(
        plantilla,
        {
            ("HUS454747", "TA5701"): ("RE9901", "ESE HUS NO ACEPTA LA GLOSA (apoyo terapéutico)."),
            ("HUS514051", "CO2301"): ("RE9901", "ESE HUS NO ACEPTA LA GLOSA (procedimientos)."),
        },
    )

    resultado, sin_responder = prep.expandir(grupos, prep.leer_plantilla_llena(plantilla))

    assert not sin_responder
    assert len(resultado) == 5, "las 5 líneas originales, ninguna perdida ni duplicada"
    assert {r["ID_SEGUIMIENTO_FACTURA_GLOSA"] for r in resultado} == {11, 12, 13, 21, 22}
    de_454747 = [r for r in resultado if r["NUMERO_FACTURA"] == "HUS454747"]
    assert len(de_454747) == 3
    assert all(r["CODIGO_RESPUESTA"] == "RE9901" for r in de_454747)
    assert all("apoyo terapéutico" in r["OBSERVACION_RESPUESTA"] for r in de_454747)


def test_lo_que_el_auditor_dejo_en_blanco_no_se_sube(tmp_path):
    """Media respuesta no se carga: se queda para la próxima pasada."""
    filas = [_linea(11, "HUS454747", "TA5701"), _linea(21, "HUS514051", "CO2301")]
    informe = _informe(tmp_path, filas)
    grupos = prep.agrupar(prep.leer_informe(informe))

    plantilla = tmp_path / "para_llenar.xlsx"
    prep.escribir_plantilla(grupos, plantilla)
    _llenar(plantilla, {("HUS454747", "TA5701"): ("RE9901", "No acepta.")})

    resultado, sin_responder = prep.expandir(grupos, prep.leer_plantilla_llena(plantilla))

    assert len(resultado) == 1
    assert resultado[0]["NUMERO_FACTURA"] == "HUS454747"
    assert len(sin_responder) == 1


def test_se_puede_trabajar_solo_las_glosas_o_solo_una_factura(tmp_path):
    """Para ir por partes, que es como se trabaja un lote grande."""
    filas = [
        _linea(1, "HUS454747", "TA5701"),
        _linea(2, "HUS475438", "DE5601", tipo="DEVOLUCION"),
        _linea(3, "HUS514051", "CO2301"),
    ]
    datos = prep.leer_informe(_informe(tmp_path, filas))

    assert len(prep.agrupar(datos, tipo="GLOSA")) == 2
    assert len(prep.agrupar(datos, tipo="DEVOLUCION")) == 1
    assert len(prep.agrupar(datos, factura="HUS454747")) == 1


def test_avisa_claro_si_el_archivo_no_es_el_informe(tmp_path):
    """El auditor puede pasar el Excel equivocado: hay que decírselo bien."""
    wb = Workbook()
    wb.active.title = "OTRA COSA"
    otro = tmp_path / "otro.xlsx"
    wb.save(otro)

    with pytest.raises(SystemExit) as e:
        prep.leer_informe(otro)

    assert "SEGUIMIENTOS" in str(e.value)


@pytest.mark.parametrize(
    "observacion",
    [
        "Glosa Calculada Afiliado - Validacion: (CODIGO DEL ",  # termina en espacio
        "Primera línea\nSegunda línea",  # salto de línea
        "  Con espacios al principio y al final  ",
        "Observación con tabulación\ty más texto",
    ],
)
def test_una_observacion_rara_no_hace_perder_la_respuesta(tmp_path, observacion):
    """Regresión: en el caso real se perdían 27 respuestas en silencio.

    Las observaciones de la EPS traen espacios al final, saltos de línea y
    caracteres raros. Cuando el grupo se identificaba con ese mismo texto,
    Excel se lo devolvía cambiado y la fila ya no calzaba con su grupo: la
    respuesta se caía sin decir nada. Ahora el grupo se identifica con una
    huella, que sobrevive el viaje.
    """
    filas = [_linea(11, "HUS454747", "TA5701", obs=observacion)]
    informe = _informe(tmp_path, filas)
    grupos = prep.agrupar(prep.leer_informe(informe))

    plantilla = tmp_path / "para_llenar.xlsx"
    prep.escribir_plantilla(grupos, plantilla)
    _llenar(plantilla, {("HUS454747", "TA5701"): ("RE9901", "No acepta.")})

    resultado, sin_responder = prep.expandir(grupos, prep.leer_plantilla_llena(plantilla))

    assert not sin_responder, "la respuesta se perdió al pasar por Excel"
    assert len(resultado) == 1
    assert resultado[0]["CODIGO_RESPUESTA"] == "RE9901"


def test_dos_glosas_que_solo_se_diferencian_al_final_no_se_confunden(tmp_path):
    """Las observaciones largas de Sanitas comparten los primeros 180 caracteres."""
    base = "Glosa Calculada Afiliado - Tipo y Numero de Identificacion: (CC, 12345678, PEPITO PEREZ), Fecha Atencion: (14/04/2025), "
    filas = [
        _linea(11, "HUS527098", "CO2301", obs=base + "Procedimiento: (890208)"),
        _linea(12, "HUS527098", "CO2301", obs=base + "Procedimiento: (903437)"),
    ]
    grupos = prep.agrupar(prep.leer_informe(_informe(tmp_path, filas)))

    assert len(grupos) == 2, "son dos glosas distintas, no una"


def test_la_hoja_de_trabajo_arranca_por_lo_mas_viejo(tmp_path):
    """Lo vencido primero: es lo que puede costarle plata al hospital."""
    viejo = _linea(1, "HUS400000", "CO2301")
    viejo["fecha_reporte"] = "2026-04-22T10:00:00"
    nuevo = _linea(2, "HUS500000", "CO2301")
    nuevo["fecha_reporte"] = "2026-07-31T10:00:00"
    datos = prep.leer_informe(_informe(tmp_path, [nuevo, viejo]))

    plantilla = tmp_path / "para_llenar.xlsx"
    prep.escribir_plantilla(prep.agrupar(datos), plantilla)

    _, filas = _leer(plantilla)
    assert filas[0]["NUMERO_FACTURA"] == "HUS400000"
