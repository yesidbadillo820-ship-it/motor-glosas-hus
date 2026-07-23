"""Tests del exportador de la hoja de conciliacion
(tools/exportar_conciliacion_dispensario.py).

Verifica que, a partir de una carpeta de piloto YA corrida, arma el
CONCILIACION.xlsx con una fila por glosa y con el valor objetado tomado del
HUS.xlsx (incluyendo glosas repetidas de mismo servicio y valor).
"""

from __future__ import annotations

import sys
from pathlib import Path

_TOOLS = Path(__file__).resolve().parent.parent.parent / "tools"
sys.path.insert(0, str(_TOOLS))

import exportar_conciliacion_dispensario as expo  # noqa: E402
import piloto_conciliacion_dispensario as pil  # noqa: E402


def _excel(ruta: Path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    ws.append(
        [
            "FACTURA",
            "RADICADO",
            "FECHA ATENCION",
            "NOMBRES Y APELLIDOS",
            "IDENTIFICACION",
            "VALOR FACTURA",
            "# GLOSA",
            "CODIGO CONCEPTO DE GLOSA",
            "MOTIVO DE GLOSA",
            "VALOR OBJETADO",
            "SERVICIO OBJETADO",
        ]
    )
    # dos glosas de mismo servicio y valor pero # distinto (los "hemocultivos")
    ws.append(
        [
            "468966",
            "700",
            "2026-02-10",
            "GUSTAVO M",
            "88",
            900000,
            "1",
            "CL08 01 CALIDAD",
            "x",
            250811,
            "HEMOCULTIVO",
        ]
    )
    ws.append(
        [
            "468966",
            "700",
            "2026-02-10",
            "GUSTAVO M",
            "88",
            900000,
            "2",
            "CL08 01 CALIDAD",
            "x",
            250811,
            "HEMOCULTIVO",
        ]
    )
    ws.append(
        [
            "468966",
            "700",
            "2026-02-10",
            "GUSTAVO M",
            "88",
            900000,
            "3",
            "TA02 01 TARIFAS",
            "y",
            23400,
            "CONSULTA",
        ]
    )
    wb.save(str(ruta))


def test_exporta_desde_carpeta_piloto(tmp_path):
    xls = tmp_path / "HUS.xlsx"
    _excel(xls)
    piloto = tmp_path / "Piloto"
    pil.correr_piloto(
        excel=xls,
        indice_ruta=None,
        cartera_ruta=None,
        facturas=["HUS0000468966"],
        auto=False,
        conocido=None,
        salida_dir=piloto,
    )
    salida = tmp_path / "CONCILIACION.xlsx"
    rc = expo.main(["--piloto-dir", str(piloto), "--excel", str(xls), "--salida", str(salida)])
    assert rc == 0
    assert salida.exists()

    import openpyxl

    ws = openpyxl.load_workbook(salida).active
    encab = [c.value for c in ws[1]]
    assert "Valor objetado" in encab and "Decision" in encab
    filas = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(filas) == 3  # las 3 glosas (los dos hemocultivos incluidos)
    # el valor objetado se tomo del HUS.xlsx para cada glosa (repetidas incluidas)
    valores = sorted(f[7] for f in filas)
    assert valores == [23400, 250811, 250811]
