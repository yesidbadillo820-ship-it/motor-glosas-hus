"""Pruebas del informe final del cargue.

Lo que se cuida acá es el caso silencioso: que el bot diga OK y SIIFA no
tenga la respuesta. Si eso se reporta como cargado, el auditor da por
respondida una glosa cuyo plazo sigue corriendo — y se entera cuando ya se
venció.
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

TOOLS = Path(__file__).resolve().parents[2] / "tools"
sys.path.insert(0, str(TOOLS))

import siifa_informe_del_cargue as inf  # noqa: E402
import siifa_perfiles as perfiles  # noqa: E402

GIRON = perfiles.buscar("GIRON")


def _csv(tmp_path: Path, filas: list[tuple], nombre="reporte.csv") -> Path:
    ruta = tmp_path / nombre
    with open(ruta, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "id_seguimiento_factura_glosa",
                "numero_factura",
                "estado",
                "detalle",
                "fecha_hora",
            ],
        )
        w.writeheader()
        for ident, estado, detalle in filas:
            w.writerow(
                {
                    "id_seguimiento_factura_glosa": ident,
                    "numero_factura": "FE102938",
                    "estado": estado,
                    "detalle": detalle,
                    "fecha_hora": "2026-08-14T10:00:00",
                }
            )
    return ruta


def _seg(ident, respondida="SI", tipo="GLOSA", valor=25000, factura="FE102938"):
    return {
        "id_seguimiento_factura_glosa": ident,
        "numero_factura": factura,
        "tipo_seguimiento": tipo,
        "razon_social_eps": "E.P.S. FAMISANAR SAS",
        "valor_glosa": valor,
        "codigo_glosa": "TA0201",
        "codigo_respuesta": "RE9602",
        "fecha_respuesta": "2026-08-14T10:00:00",
        "tiene_respuesta": respondida,
    }


# ----------------------------------------------- los tres estados


def test_una_respuesta_que_siifa_tiene_queda_registrada(tmp_path):
    reportes = inf.leer_reportes([_csv(tmp_path, [(1, "OK", "")])])

    cruzadas = inf.cruzar(reportes, [_seg(1, "SI")])

    assert cruzadas[0]["estado"] == inf.REGISTRADA


def test_el_bot_dijo_ok_pero_siifa_no_la_tiene(tmp_path):
    """EL CASO PELIGROSO. Reportarla como cargada haría dar por respondida
    una glosa cuyo plazo sigue corriendo."""
    reportes = inf.leer_reportes([_csv(tmp_path, [(1, "OK", "")])])

    cruzadas = inf.cruzar(reportes, [_seg(1, "NO")])

    assert cruzadas[0]["estado"] == inf.NO_QUEDO
    assert "no la trae respondida" in cruzadas[0]["detalle"]


def test_una_que_fallo_queda_como_error_con_su_causa(tmp_path):
    reportes = inf.leer_reportes([_csv(tmp_path, [(1, "ERROR", "HTTP 500: entity changes")])])

    cruzadas = inf.cruzar(reportes, [_seg(1, "NO")])

    assert cruzadas[0]["estado"] == inf.CON_ERROR
    assert "HTTP 500" in cruzadas[0]["detalle"]


def test_una_que_no_aparece_en_el_informe_se_marca(tmp_path):
    """Puede ser que la EPS la reformulara —pasó con FAMISANAR— o que el
    informe se bajara antes del cargue. En los dos casos hay que mirarla."""
    reportes = inf.leer_reportes([_csv(tmp_path, [(999, "OK", "")])])

    cruzadas = inf.cruzar(reportes, [_seg(1, "SI")])

    assert cruzadas[0]["estado"] == inf.NO_QUEDO
    assert "reformulada" in cruzadas[0]["detalle"]


# ----------------------------------------------- juntar varias tandas


def test_el_reintento_manda_sobre_el_intento_que_fallo(tmp_path):
    """Un cargue de 48.000 va en tandas y con reintentos. Si contara el primer
    intento, el informe reportaría errores ya resueltos."""
    primero = _csv(tmp_path, [(1, "ERROR", "HTTP 500")], "reporte1.csv")
    reintento = _csv(tmp_path, [(1, "OK", "")], "reporte2.csv")

    reportes = inf.leer_reportes([primero, reintento])

    assert reportes[1]["ok"] is True


def test_una_fila_saltada_por_saltar_csv_cuenta_como_cargada(tmp_path):
    """YA_OK_PREVIO significa que se cargó en una corrida anterior."""
    reportes = inf.leer_reportes([_csv(tmp_path, [(1, "YA_OK_PREVIO", "saltada")])])

    assert reportes[1]["ok"] is True


def test_un_reporte_que_no_existe_se_dice_claro(tmp_path):
    with pytest.raises(SystemExit, match="No encuentro el reporte"):
        inf.leer_reportes([tmp_path / "no_existe.csv"])


def test_las_filas_sin_id_no_tumban_el_informe(tmp_path):
    ruta = _csv(tmp_path, [(1, "OK", "")])
    with open(ruta, "a", encoding="utf-8") as f:
        f.write(",FE1,OK,,\n")

    assert list(inf.leer_reportes([ruta])) == [1]


# ----------------------------------------------- las cuentas


def test_las_devoluciones_no_inflan_el_valor(tmp_path):
    """La regla del repo, también acá: una devolución vale una vez por
    factura aunque tenga mil líneas."""
    filas = [(i, "OK", "") for i in range(1, 101)]
    reportes = inf.leer_reportes([_csv(tmp_path, filas)])
    informe = [_seg(i, "SI", tipo="DEVOLUCION", valor=3_000_000) for i in range(1, 101)]

    cruzadas = inf.cruzar(reportes, informe)

    from siifa_novedades import valor_total

    assert valor_total(cruzadas) == 3_000_000


def test_el_desglose_agrupa_por_entidad_y_cuenta_facturas(tmp_path):
    reportes = inf.leer_reportes([_csv(tmp_path, [(1, "OK", ""), (2, "OK", ""), (3, "OK", "")])])
    informe = [
        _seg(1, valor=1000, factura="FE1"),
        _seg(2, valor=2000, factura="FE1"),
        _seg(3, valor=5000, factura="FE2"),
    ]

    cruzadas = inf.cruzar(reportes, informe)
    (nombre, n, nf, valor) = inf._por(cruzadas, lambda f: f["razon_social_eps"])[0]

    assert (n, nf, valor) == (3, 2, 8000)


# ----------------------------------------------- el archivo que se entrega


def test_el_excel_trae_resumen_y_detalle(tmp_path):
    reportes = inf.leer_reportes([_csv(tmp_path, [(1, "OK", ""), (2, "ERROR", "HTTP 500")])])
    cruzadas = inf.cruzar(reportes, [_seg(1, "SI"), _seg(2, "NO")])
    ruta = tmp_path / "INFORME.xlsx"

    inf.escribir(cruzadas, ruta, GIRON)

    wb = load_workbook(ruta)
    assert wb.sheetnames == ["RESUMEN", "DETALLE"]
    assert "Clínica Girón" in str(wb["RESUMEN"].cell(row=1, column=1).value)
    assert wb["DETALLE"].max_row == 3


def test_lo_que_no_quedo_sale_de_primero_en_el_detalle(tmp_path):
    """Es lo accionable: si va al final de 48.000 filas, nadie lo ve."""
    reportes = inf.leer_reportes(
        [_csv(tmp_path, [(1, "OK", ""), (2, "OK", ""), (3, "ERROR", "x")])]
    )
    cruzadas = inf.cruzar(reportes, [_seg(1, "SI"), _seg(2, "NO"), _seg(3, "NO")])
    ruta = tmp_path / "INFORME.xlsx"

    inf.escribir(cruzadas, ruta, GIRON)

    estados = [c.value for c in load_workbook(ruta)["DETALLE"]["A"][1:]]
    assert estados[0] == inf.NO_QUEDO
    assert estados[-1] == inf.REGISTRADA


def test_un_error_de_ya_respondida_cuenta_como_registrada(tmp_path):
    """Pasó con 756 filas de Socorro: el timeout cortó la confirmación, la
    escritura sí entró, y el reintento recibió «ya tiene un registro previo».
    Lo que manda es el estado de la plataforma, no el error del intento."""
    reportes = inf.leer_reportes(
        [_csv(tmp_path, [(1, "ERROR", "ya tiene un registro previo de respuesta")])]
    )

    cruzadas = inf.cruzar(reportes, [_seg(1, "SI")])

    assert cruzadas[0]["estado"] == inf.REGISTRADA
    assert "SIIFA la tiene registrada" in cruzadas[0]["detalle"]


def test_sin_reportes_sale_el_censo_completo():
    """Para cuando el CSV del cargue se perdió (pasó: una corrida
    interrumpida lo sobreescribió): el censo dice cuánto de lo que SIIFA
    tiene está respondido —lo del bot y lo cargado antes— y qué falta."""
    cruzadas = inf.censar([_seg(1, "SI"), _seg(2, "SI"), _seg(3, "NO")])

    estados = [c["estado"] for c in cruzadas]
    assert estados.count(inf.REGISTRADA) == 2
    assert estados.count(inf.SIN_RESPONDER) == 1


def test_el_censo_tambien_se_escribe_en_excel(tmp_path):
    cruzadas = inf.censar([_seg(1, "SI"), _seg(2, "NO")])
    ruta = tmp_path / "CENSO.xlsx"

    inf.escribir(cruzadas, ruta, GIRON)

    estados = [c.value for c in load_workbook(ruta)["DETALLE"]["A"][1:]]
    assert estados[0] == inf.SIN_RESPONDER, "lo pendiente sale de primero"
    assert estados[-1] == inf.REGISTRADA
