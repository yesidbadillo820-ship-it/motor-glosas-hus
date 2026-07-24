"""Tests del orquestador del piloto (tools/piloto_conciliacion_dispensario.py).

Cubre: seleccion de casos representativos, la corrida extremo a extremo
(expediente -> evidencia -> hechos -> decision) generando una carpeta por
expediente con cada etapa + METRICAS, y la evaluacion de umbrales de aceptacion.
"""

from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

_TOOLS = Path(__file__).resolve().parent.parent.parent / "tools"
sys.path.insert(0, str(_TOOLS))

import piloto_conciliacion_dispensario as pil  # noqa: E402


def test_seleccionar_casos_representativos():
    exps = [
        {
            "factura_corta": "1",
            "valor_objetado_total": 100,
            "glosas": [{}],
            "contrato": {"codigo": "287"},
        },
        {
            "factura_corta": "2",
            "valor_objetado_total": 50,
            "glosas": [{}, {}, {}],
            "contrato": {"codigo": "287"},
        },
        {
            "factura_corta": "3",
            "valor_objetado_total": 10,
            "glosas": [{}],
            "contrato": {"codigo": "440"},
        },
    ]
    sel = pil.seleccionar_casos(exps, conocido=None)
    assert "1" in sel  # mayor valor
    assert "2" in sel  # mas glosas
    assert "3" in sel  # unico 440
    assert len(sel) == len(set(sel))  # sin repetidos


def _excel(ruta: Path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    ws.append(
        [
            "FACTURA",
            "RADICADO",
            "FECHA ATENCION",
            "NOMBRES Y APELLIDOS",
            "IDENTIFICACION",
            "VALOR FACTURA",
            "# GLOSA",
            "CODIGO CONCEPTO DE GLOSA",
            "MOTIVO DE GLOSA",
            "VALOR OBJETADO",
            "SERVICIO OBJETADO",
        ]
    )
    ws.append(
        [
            "436483",
            "644561",
            datetime(2025, 10, 28),
            "VALERY MOSQUERA",
            "1029407046",
            231461,
            "1",
            "TA02 01 TARIFAS",
            "Mayor valor",
            156061,
            "CONSULTA 890201",
        ]
    )
    ws.append(
        [
            "455554",
            "700000",
            datetime(2026, 2, 15),
            "JOHN VELEZ",
            "88888888",
            14405349,
            "1",
            "CL01 01 CALIDAD",
            "Estancia",
            14405349,
            "UCI 999999",
        ]
    )
    wb.save(str(ruta))


def _indice(ruta: Path):
    payload = {
        "documentos": [
            {"factura": "436483", "tipo": "RIPS", "nombre": "RIPS.json", "ruta": "Y:\\x"},
            {
                "factura": "436483",
                "tipo": "Factura electronica (XML/CUFE)",
                "nombre": "XML.xml",
                "ruta": "Y:\\y",
            },
            {
                "factura": "999999",
                "tipo": "RIPS",
                "nombre": "otro.json",
                "ruta": "Y:\\z",
            },  # huerfano
        ]
    }
    ruta.write_text(json.dumps(payload), encoding="utf-8")


def test_correr_piloto_genera_carpetas_y_metricas(tmp_path):
    xls = tmp_path / "HUS.xlsx"
    _excel(xls)
    idx = tmp_path / "indice.json"
    _indice(idx)
    out = tmp_path / "Piloto"

    metricas = pil.correr_piloto(
        excel=xls,
        indice_ruta=idx,
        cartera_ruta=None,
        facturas=["HUS0000436483", "HUS0000455554"],
        auto=False,
        conocido=None,
        salida_dir=out,
    )
    # Carpeta por expediente con las 4 etapas + resumen + log
    for fac in ("HUS0000436483", "HUS0000455554"):
        carpeta = out / fac
        for nombre in (
            "expediente.json",
            "evidencia.json",
            "hechos.json",
            "decision.json",
            "resumen.txt",
            "log.txt",
        ):
            assert (carpeta / nombre).exists(), f"falta {fac}/{nombre}"
    assert (out / "METRICAS.json").exists()
    # hoja de trabajo consolidada para conciliar
    assert (out / "CONCILIACION.xlsx").exists()
    import openpyxl

    wb = openpyxl.load_workbook(out / "CONCILIACION.xlsx")
    ws = wb.active
    encabezados = [c.value for c in ws[1]]
    assert "Decision" in encabezados and "Defendibilidad %" in encabezados
    assert ws.max_row >= 2  # al menos una glosa

    # Metricas basicas
    assert metricas["indice"]["facturas_piloto"] == 2
    assert metricas["indice"]["documentos_huerfanos_en_indice"] == 1  # la factura 999999
    assert "aceptacion" in metricas and "cumple_umbrales" in metricas
    assert metricas["indice"]["documentos_encontrados"] == 2  # solo los de 436483

    # decision.json contiene el bloque decision
    dec = json.loads((out / "HUS0000436483" / "decision.json").read_text(encoding="utf-8"))
    assert dec["glosas"][0]["decision"]["decision"] in (
        "SOLICITAR_LEVANTAMIENTO",
        "ACEPTAR_PARCIAL",
        "SOLICITAR_SOPORTE",
        "CONCILIAR",
        "ESCALAR",
    )


def test_guardarrail_sin_levantamiento_sin_prueba(tmp_path):
    """Ninguna decision de 'solicitar levantamiento' debe salir sin un hecho probado."""
    xls = tmp_path / "HUS.xlsx"
    _excel(xls)
    out = tmp_path / "Piloto"
    metricas = pil.correr_piloto(
        excel=xls,
        indice_ruta=None,
        cartera_ruta=None,
        facturas=["HUS0000436483", "HUS0000455554"],
        auto=False,
        conocido=None,
        salida_dir=out,
    )
    assert metricas["aceptacion"]["decisiones_sin_soporte_probatorio"] == 0


def test_pct_hechos_evaluados_no_supera_100(tmp_path):
    """El indicador es % de glosas evaluadas (0-100), no hechos/glosa."""
    xls = tmp_path / "HUS.xlsx"
    _excel(xls)
    out = tmp_path / "Piloto"
    metricas = pil.correr_piloto(
        excel=xls,
        indice_ruta=None,
        cartera_ruta=None,
        facturas=["HUS0000436483", "HUS0000455554"],
        auto=False,
        conocido=None,
        salida_dir=out,
    )
    assert 0 <= metricas["aceptacion"]["pct_hechos_evaluados"] <= 100.0
