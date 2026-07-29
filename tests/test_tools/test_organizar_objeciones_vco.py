"""Tests del organizador de objeciones VCO (tools/organizar_objeciones_vco.py).

Cubre: autodetección de formato por encabezados (tolerante a tildes y
espacios), mapeo consolidado → cargue OBJECIONES (consecutivo por factura,
partición del código de glosa, valores numéricos), mapeo inverso cargue →
CONSOLIDADO VCO (parseo de "(DESC CANTIDAD n)"), flags --sin-prefijo y
--detalle-servicio, y el CLI end-to-end con archivos reales.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

# El script vive en tools/ (sin __init__.py): lo importamos por ruta.
_TOOLS = Path(__file__).resolve().parent.parent.parent / "tools"
sys.path.insert(0, str(_TOOLS))

import organizar_objeciones_vco as org  # noqa: E402

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import Workbook, load_workbook  # noqa: E402

# ---------------------------------------------------------------------------
# Fixtures de archivos
# ---------------------------------------------------------------------------


def _xlsx_consolidado(ruta: Path, encabezados=None, filas=None) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    ws.append(encabezados or org.COLUMNAS_CONSOLIDADO)
    for fila in (
        filas
        if filas is not None
        else [
            [
                "VCO-SAVIA-2026-R1-001",
                "HUS521454",
                42800,
                "TA2901",
                "NO SE EVIDENCIAN TARIFAS DEL INSUMO",
                "FMQ0158",
                "EQUIPO DE BOMBA DE INFUSION",
                1,
                42800,
                42800,
            ],
            [
                "VCO-SAVIA-2026-R1-001",
                "HUS521454",
                22000,
                "TA0701",
                "MAYOR VALOR COBRADO (CEFRADINA AMP X 1 GR CANTIDAD 5)",
                "19977336-1",
                "CEFRADINA 1 G",
                5,
                8500,
                42500,
            ],
            [
                "VCO-SAVIA-2026-R1-002",
                "HUS523499",
                94695,
                "FA0201",
                "INTERCONSULTA NO FACTURABLE",
                890426,
                "INTERCONSULTA ANESTESIOLOGIA",
                1,
                94695,
                94695,
            ],
        ]
    ):
        ws.append(fila)
    wb.save(str(ruta))
    return ruta


def _xlsx_cargue(ruta: Path, filas=None) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "OBJECIONES"
    ws.append(org.COLUMNAS_CARGUE)
    for fila in (
        filas
        if filas is not None
        else [
            [
                1,
                "17/07/2026",
                "HUS521454",
                "17/07/2026",
                "VCO-SAVIA-2026-R1-001",
                "VCO-SAVIA-2026-R1-001",
                "TA",
                29,
                "CARTERA",
                "TA2901",
                "FMQ0158",
                "",
                "",
                42800,
                "MAYOR VALOR COBRADO (CATETER INTRAVENOSO 18 CANTIDAD 1)",
                "",
            ],
        ]
    ):
        ws.append(fila)
    wb.save(str(ruta))
    return ruta


# ---------------------------------------------------------------------------
# Utilidades puras
# ---------------------------------------------------------------------------


def test_norm_tolerante():
    assert org._norm("  Número   Factura ") == "NUMERO FACTURA"
    assert org._norm(None) == ""


@pytest.mark.parametrize(
    "crudo, esperado",
    [
        (42800, 42800),
        ("42800", 42800),
        ("$ 1.234.567", 1234567),
        ("1.234,56", 1234.56),
        ("1,234.56", 1234.56),
        ("", None),
        (None, None),
        ("N/A", None),
    ],
)
def test_numero_tolerante(crudo, esperado):
    assert org._numero(crudo) == esperado


@pytest.mark.parametrize(
    "codigo, clase, general",
    [
        ("TA2901", "TA", 29),
        ("ta 29 01", "TA", 29),
        ("FA0201", "FA", 2),
        ("SO0801", "SO", 8),
        ("RAROX", "", ""),
        ("", "", ""),
    ],
)
def test_partir_codigo_glosa(codigo, clase, general):
    assert org._partir_codigo_glosa(codigo) == (clase, general)


# ---------------------------------------------------------------------------
# Detección de formato
# ---------------------------------------------------------------------------


def test_detecta_consolidado(tmp_path):
    ruta = _xlsx_consolidado(tmp_path / "cons.xlsx")
    formato, filas, idx = org.leer_entrada(ruta)
    assert formato == "consolidado"
    assert len(filas) == 3
    assert idx["factura"] == 1 and idx["valor_glosa"] == 2


def test_detecta_consolidado_con_alias(tmp_path):
    encabezados = [
        "ACTA",
        "Nro Factura",
        "Vlr Glosa",
        "Cód Glosa",
        "Observaciones",
        "Cod Servicio",
        "Descripción del Servicio",
        "Cant",
        "Vlr Unitario",
        "Vlr Total",
    ]
    ruta = _xlsx_consolidado(tmp_path / "alias.xlsx", encabezados=encabezados)
    formato, filas, idx = org.leer_entrada(ruta)
    assert formato == "consolidado"
    assert idx["acta"] == 0 and idx["codigo_glosa"] == 3


def test_detecta_consolidado_variante_radicado(tmp_path):
    """Encabezados reales del CONSOLIDADO_VCO_FIDUPREVISORA.xlsx del usuario."""
    encabezados = [
        "NUMERO RADICADO",
        "NUMERO FACTURA",
        "VALOR GLOSA",
        "CODIGO GLOSA ESPECIFICA",
        "DESCRIPCION GLOSA AUDITOR",
        "CODIGO SERVICIO",
        "DESCRIPCION SERVICIO",
        "CANTIDAD",
        "VALOR UNITARIO SERVICIO",
        "VALOR TOTAL SERVICIO",
    ]
    ruta = _xlsx_consolidado(tmp_path / "radicado.xlsx", encabezados=encabezados)
    formato, _, idx = org.leer_entrada(ruta)
    assert formato == "consolidado"
    assert idx["acta"] == 0  # NUMERO RADICADO → acta
    assert idx["observacion"] == 4  # DESCRIPCION GLOSA AUDITOR → observación
    assert idx["descripcion_servicio"] == 6  # no se confunde con la col E


def test_detecta_cargue(tmp_path):
    ruta = _xlsx_cargue(tmp_path / "cargue.xlsx")
    formato, filas, idx = org.leer_entrada(ruta)
    assert formato == "cargue"
    assert idx["CRNCXC"] == 2 and idx["CROVALOBJ"] == 13


def test_formato_desconocido(tmp_path):
    wb = Workbook()
    wb.active.append(["A", "B", "C"])
    wb.active.append([1, 2, 3])
    ruta = tmp_path / "raro.xlsx"
    wb.save(str(ruta))
    with pytest.raises(SystemExit):
        org.leer_entrada(ruta)


def test_hoja_por_nombre_tolerante(tmp_path):
    wb = Workbook()
    wb.active.title = "Otra"
    ws = wb.create_sheet("  objeciones ")
    ws.append(org.COLUMNAS_CARGUE)
    ws.append([1, "", "HUS1", "", "", "", "", "", "", "", "", "", "", 100, "", ""])
    ruta = tmp_path / "hojas.xlsx"
    wb.save(str(ruta))
    formato, filas, _ = org.leer_entrada(ruta, "OBJECIONES")
    assert formato == "cargue" and len(filas) == 1
    with pytest.raises(SystemExit):
        org.leer_entrada(ruta, "NoExiste")


# ---------------------------------------------------------------------------
# consolidado → cargue
# ---------------------------------------------------------------------------


def _correr_cli(argv):
    return org.main(argv)


def test_cli_consolidado_a_cargue(tmp_path, capsys):
    entrada = _xlsx_consolidado(tmp_path / "CONSOLIDADO_VCO.xlsx")
    salida = tmp_path / "OBJECIONES_SAVIA_SALUD.xlsx"
    rc = _correr_cli(
        [
            "--entrada",
            str(entrada),
            "--salida",
            str(salida),
            "--entidad",
            "SAVIA SALUD",
            "--fecha-documento",
            "17/07/2026",
        ]
    )
    assert rc == 0
    assert salida.exists()

    wb = load_workbook(str(salida))
    ws = wb["OBJECIONES"]
    assert [c.value for c in ws[1]] == org.COLUMNAS_CARGUE
    assert ws.freeze_panes == "A2"

    filas = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(filas) == 3
    # Consecutivo por factura: HUS521454 → 1 (dos filas), HUS523499 → 2.
    assert [f[0] for f in filas] == [1, 1, 2]
    assert filas[0][2] == "HUS521454"
    assert filas[0][4] == filas[0][5] == "VCO-SAVIA-2026-R1-001"
    assert filas[0][6] == "TA" and filas[0][7] == 29  # clase / concepto general
    assert filas[0][9] == "TA2901" and filas[0][10] == "FMQ0158"
    assert filas[0][13] == 42800
    assert filas[2][6] == "FA" and filas[2][7] == 2
    assert filas[2][10] == "890426"  # código servicio numérico → texto
    # Fecha del documento respetada (Excel la devuelve como datetime).
    import datetime as dt

    assert filas[0][1] == dt.datetime(2026, 7, 17)

    resumen = capsys.readouterr().out
    assert "3 objeciones" in resumen
    assert "159.495" in resumen or "159,495" in resumen


def test_consecutivo_inicial_y_sin_prefijo(tmp_path):
    entrada = _xlsx_consolidado(tmp_path / "cons.xlsx")
    salida = tmp_path / "out.xlsx"
    rc = _correr_cli(
        [
            "--entrada",
            str(entrada),
            "--salida",
            str(salida),
            "--consecutivo-inicial",
            "100",
            "--sin-prefijo",
            "--usuario",
            "AUDITORIA",
            "--tipo-objecion",
            "G",
        ]
    )
    assert rc == 0
    ws = load_workbook(str(salida))["OBJECIONES"]
    filas = list(ws.iter_rows(min_row=2, values_only=True))
    assert [f[0] for f in filas] == [100, 100, 101]
    assert filas[0][2] == "521454"  # sin prefijo HUS
    assert all(f[8] == "AUDITORIA" and f[15] == "G" for f in filas)


def test_detalle_servicio_en_observacion(tmp_path):
    entrada = _xlsx_consolidado(tmp_path / "cons.xlsx")
    salida = tmp_path / "out.xlsx"
    rc = _correr_cli(["--entrada", str(entrada), "--salida", str(salida), "--detalle-servicio"])
    assert rc == 0
    ws = load_workbook(str(salida))["OBJECIONES"]
    obs = ws.cell(row=2, column=15).value
    assert "SERVICIO EQUIPO DE BOMBA DE INFUSION" in obs
    assert "CANT 1" in obs and "VLR UNIT 42800" in obs


def test_referencia_flag_cuando_no_hay_acta(tmp_path):
    encabezados = list(org.COLUMNAS_CONSOLIDADO[1:])  # sin columna CROOBSERV
    filas = [["HUS1", 100, "TA2901", "OBS", "X", "D", 1, 100, 100]]
    ruta = _xlsx_consolidado(tmp_path / "sin_acta.xlsx", encabezados, filas)
    salida = tmp_path / "out.xlsx"
    rc = _correr_cli(
        [
            "--entrada",
            str(ruta),
            "--salida",
            str(salida),
            "--referencia",
            "VCO-SAVIA-2026-R2-009",
        ]
    )
    assert rc == 0
    ws = load_workbook(str(salida))["OBJECIONES"]
    assert ws.cell(row=2, column=5).value == "VCO-SAVIA-2026-R2-009"
    assert ws.cell(row=2, column=6).value == "VCO-SAVIA-2026-R2-009"


def test_filas_sin_factura_se_omiten(tmp_path):
    filas = [
        ["ACTA1", "", 100, "TA2901", "OBS", "X", "D", 1, 100, 100],
        ["ACTA1", "HUS9", 200, "TA2901", "OBS", "X", "D", 1, 200, 200],
    ]
    ruta = _xlsx_consolidado(tmp_path / "c.xlsx", filas=filas)
    salida = tmp_path / "out.xlsx"
    rc = _correr_cli(["--entrada", str(ruta), "--salida", str(salida)])
    assert rc == 0
    ws = load_workbook(str(salida))["OBJECIONES"]
    assert ws.max_row == 2  # encabezado + 1 fila válida


# ---------------------------------------------------------------------------
# cargue → consolidado
# ---------------------------------------------------------------------------


def test_cli_cargue_a_consolidado(tmp_path, capsys):
    entrada = _xlsx_cargue(tmp_path / "OBJECIONES.xlsx")
    salida = tmp_path / "CONSOLIDADO_VCO_SAVIA_SALUD.xlsx"
    rc = _correr_cli(["--entrada", str(entrada), "--salida", str(salida)])
    assert rc == 0

    ws = load_workbook(str(salida))["Hoja1"]
    assert [c.value for c in ws[1]] == org.COLUMNAS_CONSOLIDADO
    fila = list(ws.iter_rows(min_row=2, values_only=True))[0]
    assert fila[0] == "VCO-SAVIA-2026-R1-001"
    assert fila[1] == "HUS521454"
    assert fila[2] == 42800
    assert fila[3] == "TA2901"
    assert fila[5] == "FMQ0158"
    # Detalle parseado de "(CATETER INTRAVENOSO 18 CANTIDAD 1)".
    assert fila[6] == "CATETER INTRAVENOSO 18" and fila[7] == 1
    assert "1 facturas" in capsys.readouterr().out


def test_cargue_sin_concepto_usa_clase(tmp_path):
    filas = [
        [1, "", "HUS7", "", "REF", "REF", "SO", 8, "U", "", "879301", "", "", 500, "OBS", ""],
    ]
    entrada = _xlsx_cargue(tmp_path / "c.xlsx", filas=filas)
    salida = tmp_path / "out.xlsx"
    assert _correr_cli(["--entrada", str(entrada), "--salida", str(salida)]) == 0
    ws = load_workbook(str(salida))["Hoja1"]
    assert ws.cell(row=2, column=4).value == "SO08"


def test_cargue_vacio_da_error_claro(tmp_path, capsys):
    entrada = _xlsx_cargue(tmp_path / "vacio.xlsx", filas=[])
    rc = _correr_cli(["--entrada", str(entrada)])
    assert rc == 2
    assert "solo" in capsys.readouterr().err.lower()


def test_entrada_inexistente(tmp_path):
    rc = _correr_cli(["--entrada", str(tmp_path / "no_existe.xlsx")])
    assert rc == 2


def test_dimension_mentirosa_se_relee(tmp_path):
    """Exports del ERP a veces declaran <dimension A1:P1> con datos reales
    más abajo — el lector debe releer las dimensiones."""
    import zipfile

    entrada = _xlsx_cargue(tmp_path / "dim.xlsx")
    # Reescribir el XML de la hoja con dimension A1:P1 (mentirosa).
    with zipfile.ZipFile(entrada) as z:
        contenido = {n: z.read(n) for n in z.namelist()}
    hoja = [n for n in contenido if n.startswith("xl/worksheets/sheet")][0]
    xml = contenido[hoja].decode("utf-8")
    import re as _re

    xml = _re.sub(r'<dimension ref="[^"]+"/>', '<dimension ref="A1:P1"/>', xml)
    contenido[hoja] = xml.encode("utf-8")
    with zipfile.ZipFile(entrada, "w") as z:
        for nombre, datos in contenido.items():
            z.writestr(nombre, datos)

    formato, filas, _ = org.leer_entrada(entrada)
    assert formato == "cargue" and len(filas) == 1
