"""Pruebas del balance de un corte (qué había, qué se respondió, qué falta,
qué llegó nuevo).

Lo delicado acá son dos confusiones que dejarían al auditor tranquilo con
plazos corriendo:

1. Contar como «respondida en este trabajo» una que ya venía respondida de
   antes (inflaría el logro y esconde quién la trabajó).
2. Que una devolución de mil líneas infle el valor de lo pendiente y lo
   nuevo: la regla del repo —una vez por factura— aplica también acá.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

TOOLS = Path(__file__).resolve().parents[2] / "tools"
sys.path.insert(0, str(TOOLS))

import siifa_balance as bal  # noqa: E402
import siifa_perfiles as perfiles  # noqa: E402

GUANE = perfiles.buscar("GUANE")


def _fila(ident, respondida="SI", tipo="GLOSA", valor=10_000, factura="FE1"):
    return {
        "id_seguimiento_factura_glosa": ident,
        "numero_factura": factura,
        "tipo_seguimiento": tipo,
        "razon_social_eps": "E.P.S. FAMISANAR SAS",
        "valor_glosa": valor,
        "codigo_glosa": "TA0201",
        "tiene_respuesta": respondida,
    }


# ----------------------------------------------- los seis cajones


def test_lo_que_se_respondio_en_este_trabajo():
    """Estaba sin responder al corte y hoy aparece respondida: eso es lo
    que se cargó."""
    cajones = bal.balancear([_fila(1, "NO")], [_fila(1, "SI")])

    assert [bal._id(f) for f in cajones[bal.RESPONDIDA]] == ["1"]


def test_lo_que_ya_venia_respondido_no_se_cuenta_como_logro():
    cajones = bal.balancear([_fila(1, "SI")], [_fila(1, "SI")])

    assert cajones[bal.RESPONDIDA] == []
    assert len(cajones[bal.VENIA_RESPONDIDA]) == 1


def test_lo_del_corte_que_sigue_sin_responder_es_lo_urgente():
    cajones = bal.balancear([_fila(1, "NO")], [_fila(1, "NO")])

    assert len(cajones[bal.SIN_RESPONDER]) == 1


def test_lo_nuevo_se_separa_de_lo_viejo():
    cajones = bal.balancear(
        [_fila(1, "NO")],
        [_fila(1, "SI"), _fila(2, "NO"), _fila(3, "SI")],
    )

    assert [bal._id(f) for f in cajones[bal.NUEVA_SIN_RESPONDER]] == ["2"]
    assert [bal._id(f) for f in cajones[bal.NUEVA_RESPONDIDA]] == ["3"]


def test_lo_que_desaparecio_del_informe_queda_marcado():
    """Puede ser una reformulación de la EPS: hay que verlo, no perderlo."""
    cajones = bal.balancear([_fila(1, "NO"), _fila(2, "NO")], [_fila(1, "SI")])

    assert [bal._id(f) for f in cajones[bal.YA_NO_ESTA]] == ["2"]


def test_el_detalle_que_se_entrega_es_el_estado_de_hoy():
    """Si se entregara la fila del corte, saldría «sin responder» algo que
    ya está respondido en la plataforma."""
    hoy = _fila(1, "SI")
    hoy["codigo_respuesta"] = "RE9602"

    cajones = bal.balancear([_fila(1, "NO")], [hoy])

    assert cajones[bal.RESPONDIDA][0].get("codigo_respuesta") == "RE9602"


# ----------------------------------------------- las cuentas


def test_una_devolucion_de_muchas_lineas_vale_una_sola_vez():
    hoy = [_fila(i, "NO", tipo="DEVOLUCION", valor=3_000_000, factura="FE9") for i in range(1, 51)]

    cajones = bal.balancear([], hoy)
    items, facturas, valor = bal._cuenta(cajones[bal.NUEVA_SIN_RESPONDER])

    assert (items, facturas, valor) == (50, 1, 3_000_000)


def test_las_glosas_si_se_suman_completas():
    cajones = bal.balancear([], [_fila(1, "NO", valor=100), _fila(2, "NO", valor=200)])

    assert bal._cuenta(cajones[bal.NUEVA_SIN_RESPONDER])[2] == 300


# ----------------------------------------------- el archivo que se entrega


def test_el_excel_trae_resumen_y_solo_las_hojas_con_contenido(tmp_path):
    cajones = bal.balancear([_fila(1, "NO")], [_fila(1, "SI"), _fila(2, "NO")])
    ruta = tmp_path / "BALANCE.xlsx"

    bal.escribir(cajones, ruta, GUANE)

    wb = load_workbook(ruta)
    assert wb.sheetnames == ["RESUMEN", "PENDIENTES", "NUEVAS"]
    assert "Guane" in str(wb["RESUMEN"].cell(row=1, column=1).value)


def test_sin_pendientes_ni_novedades_solo_queda_el_resumen(tmp_path):
    cajones = bal.balancear([_fila(1, "SI")], [_fila(1, "SI")])
    ruta = tmp_path / "BALANCE.xlsx"

    bal.escribir(cajones, ruta, GUANE)

    assert load_workbook(ruta).sheetnames == ["RESUMEN"]


def test_los_pendientes_juntan_lo_viejo_y_lo_nuevo(tmp_path):
    cajones = bal.balancear(
        [_fila(1, "NO", factura="VIEJA")],
        [_fila(1, "NO", factura="VIEJA"), _fila(2, "NO", factura="NUEVA")],
    )
    ruta = tmp_path / "BALANCE.xlsx"

    bal.escribir(cajones, ruta, GUANE)

    facturas = [c.value for c in load_workbook(ruta)["PENDIENTES"]["C"][1:]]
    assert sorted(facturas) == ["NUEVA", "VIEJA"]


def test_el_resumen_marca_en_rojo_lo_sin_responder(tmp_path):
    cajones = bal.balancear([_fila(1, "NO")], [_fila(1, "NO")])
    ruta = tmp_path / "BALANCE.xlsx"

    bal.escribir(cajones, ruta, GUANE)

    ws = load_workbook(ruta)["RESUMEN"]
    fila_sin_responder = next(
        r for r in range(4, ws.max_row + 1) if ws.cell(row=r, column=1).value == bal.SIN_RESPONDER
    )
    assert ws.cell(row=fila_sin_responder, column=1).font.bold


def test_mostrar_no_se_cae_y_dice_lo_pendiente(capsys):
    cajones = bal.balancear([_fila(1, "NO")], [_fila(1, "NO"), _fila(2, "NO")])

    bal.mostrar(cajones, GUANE, limite=25)

    salida = capsys.readouterr().out
    assert "BALANCE DE" in salida
    assert "PENDIENTE DE RESPONDER" in salida


def test_mostrar_celebra_cuando_no_queda_nada(capsys):
    cajones = bal.balancear([_fila(1, "NO")], [_fila(1, "SI")])

    bal.mostrar(cajones, GUANE, limite=25)

    assert "NADA PENDIENTE" in capsys.readouterr().out
