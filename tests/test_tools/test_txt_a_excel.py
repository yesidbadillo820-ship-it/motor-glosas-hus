"""Pruebas del motor tools/txt_a_excel.py (usado por TXT_A_EXCEL.cmd).

Cubre los dos mundos de entrada: .txt ya delimitados (comas tipo FURIPS,
punto y coma, TAB) y reportes de ancho fijo (columnas alineadas con
espacios), y las dos salidas (.csv delimitado por comas y .xlsx).
"""

from __future__ import annotations

import csv
import importlib.util
from pathlib import Path

_RAIZ = Path(__file__).resolve().parents[2]
_MOTOR = _RAIZ / "tools" / "txt_a_excel.py"

# Línea de detalle estilo FURIPS: comas, sin encabezado, campos vacíos.
_FURIPS = (
    "HUS311371,169895,2,19624,,1,69700,69700,69700\r\n"
    "HUS311371,169895,2,21708,,1,643400,643400,643400\r\n"
)

# Reporte de ancho fijo: columnas alineadas con espacios; el texto interior
# tiene espacios sencillos que NO deben partirse.
_ANCHO_FIJO = (
    "FACTURA        NIT         VALOR     DETALLE\n"
    "HUS0000533470  900156264   480200    CONSULTA DE URGENCIAS\n"
    "HUS0000533550  900156264   1912100   SALA DE OBSERVACION\n"
)


def _cargar_modulo():
    spec = importlib.util.spec_from_file_location("txt_a_excel", _MOTOR)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _celdas(ruta_xlsx: Path) -> list[tuple]:
    from openpyxl import load_workbook

    wb = load_workbook(ruta_xlsx)
    try:
        return list(wb.active.values)
    finally:
        wb.close()


def _filas_csv(ruta_csv: Path) -> list[list[str]]:
    with open(ruta_csv, encoding="cp1252", newline="") as fh:
        return list(csv.reader(fh))


def test_txt_por_comas_estilo_furips(tmp_path):
    txt = tmp_path / "FURIPS2.txt"
    txt.write_bytes(_FURIPS.encode("cp1252"))
    mod = _cargar_modulo()
    assert mod.main([str(tmp_path)]) == 0

    filas = _filas_csv(tmp_path / "FURIPS2.csv")
    assert filas[0] == ["HUS311371", "169895", "2", "19624", "", "1", "69700", "69700", "69700"]
    # el .csv reproduce el contenido tal cual (fiel para la plataforma)
    assert (tmp_path / "FURIPS2.csv").read_bytes().decode("cp1252").strip().splitlines()[0] == (
        "HUS311371,169895,2,19624,,1,69700,69700,69700"
    )

    celdas = _celdas(tmp_path / "FURIPS2.xlsx")
    assert celdas[0][0] == "HUS311371"  # código: texto
    assert celdas[0][1] == 169895  # número limpio: numérico
    assert celdas[0][4] is None  # campo vacío
    assert txt.read_bytes() == _FURIPS.encode("cp1252")  # el original ni se toca


def test_ancho_fijo_columnas_alineadas(tmp_path):
    (tmp_path / "reporte.txt").write_text(_ANCHO_FIJO, encoding="utf-8")
    mod = _cargar_modulo()
    assert mod.main([str(tmp_path)]) == 0

    filas = _filas_csv(tmp_path / "reporte.csv")
    assert filas[0] == ["FACTURA", "NIT", "VALOR", "DETALLE"]
    assert filas[1] == ["HUS0000533470", "900156264", "480200", "CONSULTA DE URGENCIAS"]
    assert filas[2][3] == "SALA DE OBSERVACION"

    celdas = _celdas(tmp_path / "reporte.xlsx")
    assert celdas[1][0] == "HUS0000533470"  # ceros a la izquierda: texto
    assert celdas[1][2] == 480200  # valor: numérico


def test_valor_celda_protege_los_datos_de_facturacion():
    mod = _cargar_modulo()
    v = mod.valor_celda
    assert v("HUS0000533470") == "HUS0000533470"
    assert v("001") == "001"  # ceros a la izquierda
    assert v("1234") == 1234
    assert v("480200.00") == 480200.0
    assert v("480.200") == "480.200"  # ambiguo (miles): texto
    assert v("1.234.567") == "1.234.567"
    assert v("0.5") == 0.5
    assert v("12345678901234567890") == "12345678901234567890"  # >15 dígitos
    assert v("06/03/1962") == "06/03/1962"
    assert v("") is None


def test_delimitador_punto_y_coma_y_tab(tmp_path):
    (tmp_path / "pc.txt").write_text("a;b;c\n1;2;3\n", encoding="utf-8")
    (tmp_path / "tab.txt").write_text("a\tb\n1\t2\n", encoding="utf-8")
    mod = _cargar_modulo()
    assert mod.main([str(tmp_path)]) == 0
    assert _filas_csv(tmp_path / "pc.csv")[1] == ["1", "2", "3"]
    assert _filas_csv(tmp_path / "tab.csv")[1] == ["1", "2"]


def test_ignora_lineas_vacias_y_de_adorno(tmp_path):
    (tmp_path / "r.txt").write_text(
        "COL1  COL2\n------------\nx1    y1\n\n============\nx2    y2\n",
        encoding="utf-8",
    )
    mod = _cargar_modulo()
    assert mod.main([str(tmp_path)]) == 0
    filas = _filas_csv(tmp_path / "r.csv")
    assert filas == [["COL1", "COL2"], ["x1", "y1"], ["x2", "y2"]]


def test_recursivo_y_sin_recursion(tmp_path):
    (tmp_path / "a.txt").write_text("1,2\n3,4\n", encoding="utf-8")
    sub = tmp_path / "sub"
    sub.mkdir()
    (sub / "b.txt").write_text("5,6\n7,8\n", encoding="utf-8")

    mod = _cargar_modulo()
    assert mod.main([str(tmp_path), "--sin-recursion"]) == 0
    assert (tmp_path / "a.csv").exists()
    assert not (sub / "b.csv").exists()

    assert mod.main([str(tmp_path)]) == 0
    assert (sub / "b.csv").exists()
    assert (sub / "b.xlsx").exists()


def test_jamas_pisa_archivos_ajenos_y_refresca_los_propios(tmp_path):
    txt = tmp_path / "datos.txt"
    txt.write_text("a,b\n1,2\n", encoding="utf-8")
    ajeno = tmp_path / "datos.xlsx"
    ajeno.write_bytes(b"PK\x03\x04soy-un-excel-ajeno")

    mod = _cargar_modulo()
    assert mod.main([str(tmp_path)]) == 0
    assert ajeno.read_bytes() == b"PK\x03\x04soy-un-excel-ajeno"  # intacto
    assert not (tmp_path / "datos.csv").exists()  # se omitió completo

    # sin el ajeno: convierte, y en la segunda corrida refresca su propio par
    ajeno.unlink()
    assert mod.main([str(tmp_path)]) == 0
    assert (tmp_path / "datos.csv").exists()
    txt.write_text("a,b\n9,9\n", encoding="utf-8")
    assert mod.main([str(tmp_path)]) == 0
    assert _filas_csv(tmp_path / "datos.csv")[1] == ["9", "9"]


def test_simulacro_no_escribe(tmp_path):
    (tmp_path / "x.txt").write_text("1,2\n", encoding="utf-8")
    mod = _cargar_modulo()
    assert mod.main([str(tmp_path), "--simulacro"]) == 0
    assert not (tmp_path / "x.csv").exists()
    assert not (tmp_path / "x.xlsx").exists()


def test_codificacion_ansi_con_tildes(tmp_path):
    (tmp_path / "t.txt").write_bytes("CÓDIGO,DESCRIPCIÓN\n001,consulta médica\n".encode("cp1252"))
    mod = _cargar_modulo()
    assert mod.main([str(tmp_path)]) == 0
    filas = _filas_csv(tmp_path / "t.csv")
    assert filas[0] == ["CÓDIGO", "DESCRIPCIÓN"]
    assert filas[1] == ["001", "consulta médica"]
    celdas = _celdas(tmp_path / "t.xlsx")
    assert celdas[1][0] == "001"  # el cero a la izquierda sobrevive en el Excel


def test_dato_con_coma_queda_entre_comillas_en_el_csv(tmp_path):
    (tmp_path / "c.txt").write_text("COL1;COL2\nuno;dos, tres\n", encoding="utf-8")
    mod = _cargar_modulo()
    assert mod.main([str(tmp_path)]) == 0
    crudo = (tmp_path / "c.csv").read_bytes().decode("cp1252")
    assert '"dos, tres"' in crudo  # protegido: sigue siendo UNA columna
    assert _filas_csv(tmp_path / "c.csv")[1] == ["uno", "dos, tres"]


def test_muchas_filas_se_reparten_en_hojas(tmp_path):
    mod = _cargar_modulo()
    mod.MAX_FILAS_HOJA = 3  # límite chico solo para la prueba
    (tmp_path / "g.txt").write_text("".join(f"{i},x\n" for i in range(7)), encoding="utf-8")
    assert mod.main([str(tmp_path)]) == 0

    from openpyxl import load_workbook

    wb = load_workbook(tmp_path / "g.xlsx")
    assert wb.sheetnames == ["Datos", "Datos_2", "Datos_3"]
    assert sum(ws.max_row for ws in wb) == 7
    wb.close()


def test_comillas_de_pulgadas_no_danan_el_registro(tmp_path):
    """Una comilla suelta (5" de pulgada en un insumo) NO debe fusionar
    registros ni desaparecer: el corte es plano, sin interpretar comillas."""
    contenido = 'HUS311371,169895,2,"5 CATETER,1,69700,69700,69700\r\nHUS311372,169895,2,SONDA "ARROW",1,643400,643400,643400\r\n'
    (tmp_path / "f.txt").write_bytes(contenido.encode("cp1252"))
    mod = _cargar_modulo()
    assert mod.main([str(tmp_path)]) == 0

    # el .csv reproduce las lineas TAL CUAL, byte a byte (fiel para la plataforma)
    crudo = (tmp_path / "f.csv").read_bytes().decode("cp1252")
    assert crudo == contenido
    assert len(crudo.strip().splitlines()) == 2  # nada de fusionar registros
    celdas = _celdas(tmp_path / "f.xlsx")
    assert celdas[0][3] == '"5 CATETER'  # la comilla sobrevive
    assert celdas[1][3] == 'SONDA "ARROW"'


def test_nombre_con_coma_no_engana_al_detector(tmp_path):
    """'PEREZ, JUAN' (una coma por linea) no debe volverse el delimitador:
    el reporte de ancho fijo debe cortarse por sus columnas."""
    lineas = ["FACTURA        PACIENTE            VALOR"]
    for i in range(20):
        lineas.append(f"HUS000053{i:04d}  PEREZ, JUAN {i:03d}     480200")
    (tmp_path / "r.txt").write_text("\n".join(lineas) + "\n", encoding="utf-8")
    mod = _cargar_modulo()
    assert mod.main([str(tmp_path)]) == 0

    filas = _filas_csv(tmp_path / "r.csv")
    assert filas[0] == ["FACTURA", "PACIENTE", "VALOR"]
    assert filas[1] == ["HUS0000530000", "PEREZ, JUAN 000", "480200"]
    assert all(len(f) == 3 for f in filas)


def test_ancho_fijo_no_trunca_datos_tardios(tmp_path):
    """Una columna que solo trae datos despues de la linea 400 no se pierde
    (el perfil de columnas se calcula con TODAS las lineas)."""
    lineas = ["FACTURA        VALOR    OBS"]
    for i in range(440):
        lineas.append(f"HUS{i:010d}  480200   ")
    lineas.append("HUS9999999999  480649   GLOSA TOTAL 999")
    (tmp_path / "g.txt").write_text("\n".join(lineas) + "\n", encoding="utf-8")
    mod = _cargar_modulo()
    assert mod.main([str(tmp_path)]) == 0

    filas = _filas_csv(tmp_path / "g.csv")
    assert filas[-1] == ["HUS9999999999", "480649", "GLOSA TOTAL 999"]  # completo


def test_utf16_de_notepad_se_decodifica(tmp_path):
    (tmp_path / "u.txt").write_bytes("HUS311371,169895,2,19624\r\n".encode("utf-16"))
    mod = _cargar_modulo()
    assert mod.main([str(tmp_path)]) == 0
    assert _filas_csv(tmp_path / "u.csv")[0] == ["HUS311371", "169895", "2", "19624"]


def test_bom_utf8_no_contamina_el_primer_campo(tmp_path):
    (tmp_path / "b.txt").write_bytes(b"\xef\xbb\xbf" + b"HUS311371,169895,2\r\n")
    mod = _cargar_modulo()
    assert mod.main([str(tmp_path)]) == 0
    assert _filas_csv(tmp_path / "b.csv")[0][0] == "HUS311371"


def test_encabezado_negrita_y_freeze_solo_cuando_aplica(tmp_path):
    from openpyxl import load_workbook

    (tmp_path / "con.txt").write_text("FACTURA;VALOR\nHUS1;480200\n", encoding="utf-8")
    (tmp_path / "sin.txt").write_text("HUS1;480200\nHUS2;190000\n", encoding="utf-8")
    mod = _cargar_modulo()
    assert mod.main([str(tmp_path)]) == 0

    wb = load_workbook(tmp_path / "con.xlsx")
    assert wb.active["A1"].font.bold and wb.active.freeze_panes == "A2"
    wb.close()
    wb = load_workbook(tmp_path / "sin.xlsx")
    assert not wb.active["A1"].font.bold and wb.active.freeze_panes is None
    wb.close()


def test_csv_huerfano_sin_su_excel_no_se_pisa(tmp_path):
    (tmp_path / "d.txt").write_text("a,b,c\n1,2,3\n", encoding="utf-8")
    (tmp_path / "d.csv").write_text("csv ajeno", encoding="utf-8")  # sin d.xlsx
    mod = _cargar_modulo()
    assert mod.main([str(tmp_path)]) == 0
    assert (tmp_path / "d.csv").read_text(encoding="utf-8") == "csv ajeno"  # intacto
    assert not (tmp_path / "d.xlsx").exists()


def test_motor_embebido_en_cmd_identico_al_py():
    """La copia embebida tras #PYSTART# en TXT_A_EXCEL.cmd debe ser el .py exacto."""
    cmd_path = _RAIZ / "tools" / "TXT_A_EXCEL.cmd"
    lineas = cmd_path.read_text(encoding="utf-8").splitlines()
    marcador = "#PY" + "START#"  # partido para no autodetectarse
    idx = next(i for i, ln in enumerate(lineas) if marcador in ln)
    embebido = "\n".join(lineas[idx + 1 :]).strip("\n")
    fuente = _MOTOR.read_text(encoding="utf-8").replace("\r\n", "\n").strip("\n")
    assert embebido == fuente, (
        "La copia embebida en tools/TXT_A_EXCEL.cmd difiere de tools/txt_a_excel.py. "
        "Regenera el .cmd (encabezado batch + contenido del .py en CRLF)."
    )
