"""Pruebas del motor tools/revisar_xml_facturas.py (usado por REVISAR_XML.cmd).

Verifica que se extraiga la información de contrato de un XML de factura
electrónica en salud (AttachedDocument de la DIAN con el Invoice embebido en
el CDATA), que el filtro por lista de facturas funcione, y que la copia
embebida tras #PYSTART# en REVISAR_XML.cmd sea idéntica al .py fuente.
"""

from __future__ import annotations

import importlib.util
from pathlib import Path

_RAIZ = Path(__file__).resolve().parents[2]
_MOTOR = _RAIZ / "tools" / "revisar_xml_facturas.py"

# XML de prueba: AttachedDocument con el Invoice embebido en el CDATA, con los
# campos de interoperabilidad en salud (mismo formato que el que radica el HUS).
_XML_FACTURA = (
    '<?xml version="1.0" encoding="utf-8" standalone="no"?>'
    '<AttachedDocument xmlns="urn:oasis:names:specification:ubl:schema:xsd:AttachedDocument-2"'
    ' xmlns:cac="urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2"'
    ' xmlns:cbc="urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2">'
    "<cbc:ID>541589</cbc:ID>"
    "<cbc:ParentDocumentID>HUS533470</cbc:ParentDocumentID>"
    "<cac:ReceiverParty><cac:PartyTaxScheme>"
    "<cbc:RegistrationName>NUEVA EMPRESA PROMOTORA DE SALUD S.A.</cbc:RegistrationName>"
    "<cbc:CompanyID>900156264</cbc:CompanyID></cac:PartyTaxScheme></cac:ReceiverParty>"
    "<cac:Attachment><cac:ExternalReference><cbc:Description><![CDATA["
    '<?xml version="1.0" encoding="utf-8"?>'
    '<Invoice xmlns:cac="urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2"'
    ' xmlns:cbc="urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2"'
    ' xmlns:ext="urn:oasis:names:specification:ubl:schema:xsd:CommonExtensionComponents-2"'
    ' xmlns="urn:oasis:names:specification:ubl:schema:xsd:Invoice-2">'
    "<ext:UBLExtensions><ext:UBLExtension><ext:ExtensionContent><CustomTagGeneral>"
    "<Interoperabilidad><Group><Collection>"
    "<AdditionalInformation><Name>CODIGO_PRESTADOR</Name><Value>6800100792</Value></AdditionalInformation>"
    "<AdditionalInformation><Name>MODALIDAD_PAGO</Name>"
    '<Value schemeID="04">Pago por evento</Value></AdditionalInformation>'
    "<AdditionalInformation><Name>COBERTURA_PLAN_BENEFICIOS</Name>"
    '<Value schemeID="16">Plan de beneficios UPC Contributivo</Value></AdditionalInformation>'
    "<AdditionalInformation><Name>NUMERO_CONTRATO</Name>"
    "<Value>01_VCEVN_900006037</Value></AdditionalInformation>"
    "<AdditionalInformation><Name>NUMERO_POLIZA</Name><Value></Value></AdditionalInformation>"
    "<AdditionalInformation><Name>FACTURA_SIN_CONTRATO</Name>"
    '<Value schemeID="01">ATENCION DE URGENCIAS</Value></AdditionalInformation>'
    "</Collection></Group></Interoperabilidad>"
    "</CustomTagGeneral></ext:ExtensionContent></ext:UBLExtension></ext:UBLExtensions>"
    "<cbc:ID>HUS533470</cbc:ID>"
    "<cbc:IssueDate>2026-07-03</cbc:IssueDate>"
    '<cac:LegalMonetaryTotal><cbc:PayableAmount currencyID="COP">480200.00</cbc:PayableAmount>'
    "</cac:LegalMonetaryTotal>"
    "</Invoice>]]></cbc:Description></cac:ExternalReference></cac:Attachment>"
    "<cac:ResultOfVerification><cbc:ValidationResultCode>02</cbc:ValidationResultCode>"
    "</cac:ResultOfVerification>"
    "</AttachedDocument>"
)


def _cargar_modulo():
    spec = importlib.util.spec_from_file_location("revisar_xml_facturas", _MOTOR)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _escribir(ruta: Path, contenido: str = _XML_FACTURA) -> None:
    ruta.parent.mkdir(parents=True, exist_ok=True)
    ruta.write_text(contenido, encoding="utf-8")


def test_normalizar_factura_iguala_las_tres_formas():
    mod = _cargar_modulo()
    n = mod.normalizar_factura
    assert n("HUS0000533470") == n("HUS533470") == n("533470") == "533470"
    assert n("") == ""


def test_revisar_uno_extrae_la_info_de_contrato(tmp_path):
    ruta = tmp_path / "ad_HUS533470.xml"
    _escribir(ruta)
    mod = _cargar_modulo()
    d = mod.revisar_uno(ruta)

    assert d["error"] == ""
    assert d["factura"] == "HUS533470"
    assert d["eps"] == "NUEVA EMPRESA PROMOTORA DE SALUD S.A."
    assert d["fecha"] == "2026-07-03"
    assert d["valor"] == "480200.00"
    assert d["NUMERO_CONTRATO"] == "01_VCEVN_900006037"
    assert d["FACTURA_SIN_CONTRATO"] == "ATENCION DE URGENCIAS"
    assert d["MODALIDAD_PAGO"] == "Pago por evento"
    assert d["CODIGO_PRESTADOR"] == "6800100792"
    assert d["validacion_dian"].startswith("02")


def test_conclusion_arma_el_argumento_de_la_glosa(tmp_path):
    ruta = tmp_path / "ad.xml"
    _escribir(ruta)
    mod = _cargar_modulo()
    frase = mod.conclusion(mod.revisar_uno(ruta))
    assert "01_VCEVN_900006037" in frase
    assert "ATENCION DE URGENCIAS" in frase
    assert "DIAN" in frase


def test_procesar_con_lista_filtra_y_genera_excel(tmp_path):
    _escribir(tmp_path / "xml" / "ad_HUS533470.xml")
    # una factura que NO está en la lista: debe quedar fuera del informe
    otro = _XML_FACTURA.replace("HUS533470", "HUS999999")
    _escribir(tmp_path / "xml" / "ad_HUS999999.xml", otro)

    lista = tmp_path / "facturas.txt"
    lista.write_text("# glosa sin contrato\nHUS0000533470\n", encoding="utf-8")
    salida = tmp_path / "INFORME.xlsx"

    mod = _cargar_modulo()
    assert mod.procesar((tmp_path / "xml").resolve(), lista, salida) == 0
    assert salida.exists()

    from openpyxl import load_workbook

    ws = load_workbook(salida).active
    facturas = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert "HUS533470" in facturas
    assert "HUS999999" not in facturas  # filtrada por la lista


def test_xml_sin_invoice_embebido_no_revienta(tmp_path):
    ruta = tmp_path / "raro.xml"
    ruta.write_text(
        '<AttachedDocument xmlns="urn:x"><cbc xmlns="urn:y">nada</cbc></AttachedDocument>',
        encoding="utf-8",
    )
    mod = _cargar_modulo()
    d = mod.revisar_uno(ruta)
    assert d is not None
    assert d["error"]  # marca el problema, no lanza excepción


def test_motor_embebido_en_cmd_identico_al_py():
    """La copia embebida tras #PYSTART# en REVISAR_XML.cmd debe ser el .py exacto."""
    cmd_path = _RAIZ / "tools" / "REVISAR_XML.cmd"
    lineas = cmd_path.read_text(encoding="utf-8").splitlines()
    marcador = "#PY" + "START#"  # partido para no autodetectarse
    idx = next(i for i, ln in enumerate(lineas) if marcador in ln)
    embebido = "\n".join(lineas[idx + 1 :]).strip("\n")
    fuente = _MOTOR.read_text(encoding="utf-8").replace("\r\n", "\n").strip("\n")
    assert embebido == fuente, (
        "La copia embebida en tools/REVISAR_XML.cmd difiere de tools/revisar_xml_facturas.py. "
        "Regenera el .cmd (encabezado batch + contenido del .py en CRLF)."
    )
