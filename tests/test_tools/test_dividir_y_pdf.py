"""Tests del separador por factura y del conversor a PDF.

- `tools/dividir_detallado_por_factura.py`: un Excel por factura, con el
  formato intacto y listo para imprimir.
- `tools/excel_a_pdf.py`: conversión masiva a PDF, con o sin carpeta por
  archivo. La conversión real solo corre si hay LibreOffice instalado.
"""

from __future__ import annotations

import shutil
import sys
from pathlib import Path

import pytest

_TOOLS = Path(__file__).resolve().parent.parent.parent / "tools"
sys.path.insert(0, str(_TOOLS))
sys.path.insert(0, str(Path(__file__).resolve().parent))  # reusar el fixture del ajustador

openpyxl = pytest.importorskip("openpyxl")

import dividir_detallado_por_factura as div  # noqa: E402
import excel_a_pdf as pdf  # noqa: E402
from test_ajustar_detallado_glosas import (  # noqa: E402
    _factura,
    _membrete,
    _pie_archivo,
)


@pytest.fixture()
def detallado(tmp_path: Path) -> Path:
    """Una hoja con dos facturas apiladas, membrete y pie legal."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.column_dimensions["A"].width = 3.5
    _membrete(ws)
    fila = _factura(ws, 10, "HUS352890")
    # Un estilo distintivo, para comprobar que el separador lo conserva.
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    titulo = ws.cell(row=10, column=6)
    titulo.font = Font(name="Tahoma", size=14, bold=True)
    titulo.alignment = Alignment(horizontal="center")
    titulo.fill = PatternFill("solid", fgColor="FFFF00")
    titulo.border = Border(bottom=Side(style="thin"))
    fila = _factura(ws, fila, "HUS0000400001")
    _pie_archivo(ws, fila)
    ruta = tmp_path / "detallado.xlsx"
    wb.save(ruta)
    return ruta


# ─── Separar por factura ─────────────────────────────────────────────────────


class TestNombreArchivo:
    def test_corto_quita_los_ceros(self):
        assert div._nombre_archivo("HUS0000352890", "corto") == "HUS352890"
        assert div._nombre_archivo("HUS352890", "corto") == "HUS352890"

    def test_largo_respeta_el_original(self):
        assert div._nombre_archivo("HUS0000352890", "largo") == "HUS0000352890"

    def test_saca_los_caracteres_que_windows_no_admite(self):
        assert "/" not in div._nombre_archivo("HUS/352890", "largo")
        assert div._nombre_archivo("  ", "largo") == "SIN_NUMERO"


class TestDividir:
    def test_un_archivo_por_factura(self, detallado, tmp_path):
        salida = tmp_path / "por_factura"
        extraidas = div.dividir_libro(detallado, salida)
        assert sorted(e.archivo.name for e in extraidas) == ["HUS352890.xlsx", "HUS400001.xlsx"]
        assert all(e.archivo.exists() for e in extraidas)

    def test_cada_archivo_trae_solo_su_factura(self, detallado, tmp_path):
        div.dividir_libro(detallado, tmp_path / "por_factura")
        ws = openpyxl.load_workbook(tmp_path / "por_factura" / "HUS352890.xlsx").worksheets[0]
        texto = "\n".join(str(c.value) for f in ws.iter_rows() for c in f if c.value is not None)
        assert "HUS352890" in texto
        assert "HUS0000400001" not in texto
        # El membrete se queda afuera; el pie legal viaja con cada factura.
        assert "CUFE" not in texto and "Carrrera 33" not in texto
        assert "LICENCIADO A" in texto

    def test_sin_pie(self, detallado, tmp_path):
        div.dividir_libro(detallado, tmp_path / "sin_pie", incluir_pie=False)
        ws = openpyxl.load_workbook(tmp_path / "sin_pie" / "HUS352890.xlsx").worksheets[0]
        texto = "\n".join(str(c.value) for f in ws.iter_rows() for c in f if c.value is not None)
        assert "LICENCIADO A" not in texto

    def test_conserva_el_formato(self, detallado, tmp_path):
        div.dividir_libro(detallado, tmp_path / "por_factura")
        ws = openpyxl.load_workbook(tmp_path / "por_factura" / "HUS352890.xlsx").worksheets[0]
        assert ws.merged_cells.ranges, "se perdieron las celdas combinadas"
        assert ws.column_dimensions["A"].width == pytest.approx(3.5)
        titulo = ws.cell(row=1, column=6)
        assert titulo.value == "FACTURA ELECTRONICA DE VENTA"  # el separador no renombra
        assert (titulo.font.name, titulo.font.sz, titulo.font.b) == ("Tahoma", 14.0, True)
        assert titulo.alignment.horizontal == "center"
        assert titulo.fill.fgColor.rgb == "00FFFF00"
        assert titulo.border.bottom.style == "thin"

    def test_area_de_impresion_cubre_las_celdas_combinadas(self, detallado, tmp_path):
        """Medir solo las celdas con valor deja el PDF recortado a la derecha."""
        div.dividir_libro(detallado, tmp_path / "por_factura")
        ws = openpyxl.load_workbook(tmp_path / "por_factura" / "HUS352890.xlsx").worksheets[0]
        from openpyxl.utils import range_boundaries

        area = ws.print_area
        area = area[0] if isinstance(area, list) else area
        _, _, max_col, max_fila = range_boundaries(area.split("!")[-1].replace("$", ""))
        assert max_col >= max(r.max_col for r in ws.merged_cells.ranges)
        assert max_fila == ws.max_row
        assert ws.page_setup.fitToWidth == 1
        assert ws.sheet_properties.pageSetUpPr.fitToPage is True

    def test_carpeta_por_factura(self, detallado, tmp_path):
        salida = tmp_path / "carpetas"
        div.dividir_libro(detallado, salida, carpeta_por_factura=True)
        assert (salida / "HUS352890" / "HUS352890.xlsx").exists()
        assert (salida / "HUS400001" / "HUS400001.xlsx").exists()

    def test_solo_las_facturas_pedidas(self, detallado, tmp_path):
        extraidas = div.dividir_libro(detallado, tmp_path / "filtrado", facturas_pedidas={"352890"})
        assert [e.archivo.name for e in extraidas] == ["HUS352890.xlsx"]

    def test_no_sobrescribir(self, detallado, tmp_path):
        salida = tmp_path / "por_factura"
        div.dividir_libro(detallado, salida)
        marca = salida / "HUS352890.xlsx"
        marca.write_bytes(b"no me toques")
        div.dividir_libro(detallado, salida, sobrescribir=False)
        assert marca.read_bytes() == b"no me toques"

    def test_no_toca_el_original(self, detallado, tmp_path):
        antes = detallado.read_bytes()
        div.dividir_libro(detallado, tmp_path / "por_factura")
        assert detallado.read_bytes() == antes

    def test_cli(self, detallado, tmp_path):
        salida = tmp_path / "cli"
        reporte = tmp_path / "listado.csv"
        assert (
            div.main(
                [
                    "--detallado",
                    str(detallado),
                    "--salida",
                    str(salida),
                    "--reporte-csv",
                    str(reporte),
                ]
            )
            == 0
        )
        assert len(list(salida.glob("*.xlsx"))) == 2
        assert reporte.read_text(encoding="utf-8-sig").startswith("FACTURA;ARCHIVO")


# ─── Excel → PDF ─────────────────────────────────────────────────────────────


@pytest.fixture()
def carpeta_excels(tmp_path: Path) -> Path:
    carpeta = tmp_path / "excels"
    carpeta.mkdir()
    for nombre in ("HUS1.xlsx", "HUS2.xlsx"):
        wb = openpyxl.Workbook()
        wb.active["A1"] = nombre
        wb.save(carpeta / nombre)
    (carpeta / "~$HUS1.xlsx").write_bytes(b"temporal de Excel")
    (carpeta / "notas.txt").write_text("no es un Excel", encoding="utf-8")
    sub = carpeta / "sub"
    sub.mkdir()
    wb = openpyxl.Workbook()
    wb.active["A1"] = "HUS3"
    wb.save(sub / "HUS3.xlsx")
    return carpeta


class TestListar:
    def test_ignora_temporales_y_otros_formatos(self, carpeta_excels):
        nombres = [r.name for r in pdf.listar_excels([carpeta_excels])]
        assert nombres == ["HUS1.xlsx", "HUS2.xlsx"]

    def test_recursivo(self, carpeta_excels):
        nombres = sorted(r.name for r in pdf.listar_excels([carpeta_excels], recursivo=True))
        assert nombres == ["HUS1.xlsx", "HUS2.xlsx", "HUS3.xlsx"]

    def test_archivos_sueltos(self, carpeta_excels):
        uno = carpeta_excels / "HUS1.xlsx"
        assert pdf.listar_excels([uno]) == [uno]


class TestPlanificar:
    def test_todos_en_la_misma_carpeta(self, carpeta_excels, tmp_path):
        plan = pdf.planificar(pdf.listar_excels([carpeta_excels]), tmp_path / "pdf")
        assert [c.destino.name for c in plan] == ["HUS1.pdf", "HUS2.pdf"]
        assert {c.destino.parent for c in plan} == {tmp_path / "pdf"}

    def test_carpeta_por_archivo(self, carpeta_excels, tmp_path):
        plan = pdf.planificar(
            pdf.listar_excels([carpeta_excels]), tmp_path / "pdf", carpeta_por_archivo=True
        )
        assert plan[0].destino == tmp_path / "pdf" / "HUS1" / "HUS1.pdf"
        assert plan[1].destino == tmp_path / "pdf" / "HUS2" / "HUS2.pdf"

    def test_sin_salida_queda_junto_al_excel(self, carpeta_excels):
        plan = pdf.planificar(pdf.listar_excels([carpeta_excels]), None)
        assert plan[0].destino == carpeta_excels / "HUS1.pdf"

    def test_saltar_existentes(self, carpeta_excels, tmp_path):
        salida = tmp_path / "pdf"
        salida.mkdir()
        (salida / "HUS1.pdf").write_bytes(b"%PDF-")
        plan = pdf.planificar(pdf.listar_excels([carpeta_excels]), salida, saltar_existentes=True)
        assert plan[0].estado == "SALTADO"
        assert plan[1].estado == ""


class TestMotor:
    def test_pide_libreoffice_y_no_hay(self, monkeypatch):
        monkeypatch.setattr(pdf, "buscar_libreoffice", lambda ruta=None: None)
        with pytest.raises(SystemExit, match="No encontré LibreOffice"):
            pdf.elegir_motor("libreoffice", None)

    def test_pide_excel_y_no_hay(self, monkeypatch):
        monkeypatch.setattr(pdf, "hay_excel", lambda: False)
        with pytest.raises(SystemExit, match="No encontré Excel"):
            pdf.elegir_motor("excel", None)

    def test_auto_prefiere_excel(self, monkeypatch):
        monkeypatch.setattr(pdf, "hay_excel", lambda: True)
        assert pdf.elegir_motor("auto", None) == ("excel", None)

    def test_auto_cae_en_libreoffice(self, monkeypatch):
        monkeypatch.setattr(pdf, "hay_excel", lambda: False)
        monkeypatch.setattr(pdf, "buscar_libreoffice", lambda ruta=None: "/usr/bin/soffice")
        assert pdf.elegir_motor("auto", None) == ("libreoffice", "/usr/bin/soffice")

    def test_auto_sin_ninguno(self, monkeypatch):
        monkeypatch.setattr(pdf, "hay_excel", lambda: False)
        monkeypatch.setattr(pdf, "buscar_libreoffice", lambda ruta=None: None)
        with pytest.raises(SystemExit, match="No hay con qué convertir"):
            pdf.elegir_motor("auto", None)


class TestCLI:
    def test_dry_run_no_escribe_nada(self, carpeta_excels, tmp_path):
        salida = tmp_path / "pdf"
        assert (
            pdf.main(["--origen", str(carpeta_excels), "--salida", str(salida), "--dry-run"]) == 0
        )
        assert not salida.exists()

    def test_sin_archivos(self, tmp_path):
        vacia = tmp_path / "vacia"
        vacia.mkdir()
        assert pdf.main(["--origen", str(vacia)]) == 2


@pytest.mark.skipif(
    not shutil.which("soffice") and not Path("/usr/lib/libreoffice/program/soffice").exists(),
    reason="hace falta LibreOffice para convertir de verdad",
)
class TestConversionReal:
    def test_genera_los_pdf(self, carpeta_excels, tmp_path):
        salida = tmp_path / "pdf"
        assert (
            pdf.main(
                [
                    "--origen",
                    str(carpeta_excels),
                    "--salida",
                    str(salida),
                    "--motor",
                    "libreoffice",
                ]
            )
            == 0
        )
        generados = sorted(p.name for p in salida.glob("*.pdf"))
        assert generados == ["HUS1.pdf", "HUS2.pdf"]
        assert (salida / "HUS1.pdf").read_bytes().startswith(b"%PDF")

    def test_carpeta_por_archivo(self, carpeta_excels, tmp_path):
        salida = tmp_path / "pdf"
        assert (
            pdf.main(
                [
                    "--origen",
                    str(carpeta_excels),
                    "--salida",
                    str(salida),
                    "--carpeta-por-archivo",
                    "--motor",
                    "libreoffice",
                ]
            )
            == 0
        )
        assert (salida / "HUS1" / "HUS1.pdf").exists()
        assert (salida / "HUS2" / "HUS2.pdf").exists()
