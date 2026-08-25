"""Pruebas del comparador de informes de SIIFA.

Lo que se cuida acá: que la comparación diga bien QUÉ llegó nuevo. Si se
equivoca de más, el auditor trabaja sobre glosas ya respondidas; si se
equivoca de menos, una glosa nueva se vence sin contestar y queda aceptada
por el artículo 57 de la Ley 1438.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

TOOLS = Path(__file__).resolve().parents[2] / "tools"
sys.path.insert(0, str(TOOLS))

import siifa_novedades as nov  # noqa: E402

COLS = [
    "id_seguimiento_factura_glosa",
    "tipo_seguimiento",
    "numero_factura",
    "nit_eps",
    "razon_social_eps",
    "valor_glosa",
    "codigo_glosa",
    "descripcion_glosa",
    "fecha_formulacion",
    "tiene_respuesta",
]


def _informe(ruta: Path, filas: list[dict], titulo="SEGUIMIENTOS") -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = titulo
    ws.append(COLS)
    for f in filas:
        ws.append([f.get(c, "") for c in COLS])
    wb.save(ruta)
    return ruta


def _seg(ident, factura="HUS521868", tipo="GLOSA", eps="SANITAS", valor=3399053, resp="SI"):
    return {
        "id_seguimiento_factura_glosa": ident,
        "tipo_seguimiento": tipo,
        "numero_factura": factura,
        "nit_eps": "800251440",
        "razon_social_eps": eps,
        "valor_glosa": valor,
        "codigo_glosa": "FA0101",
        "descripcion_glosa": "Diferencia en el valor facturado",
        "fecha_formulacion": "2026-07-18",
        "tiene_respuesta": resp,
    }


def test_encuentra_lo_que_no_estaba_antes(tmp_path):
    antes = _informe(tmp_path / "antes.xlsx", [_seg(1), _seg(2)])
    ahora = _informe(tmp_path / "ahora.xlsx", [_seg(1), _seg(2), _seg(3, resp="NO")])

    res = nov.comparar(nov.leer_informe(ahora), nov.leer_informe(antes))

    assert res["modo"] == "comparacion"
    assert [n["id_seguimiento_factura_glosa"] for n in res["novedades"]] == [3]
    assert res["total_nuevo"] == 3
    assert res["total_anterior"] == 2


def test_una_novedad_ya_respondida_tambien_se_reporta(tmp_path):
    """Un seguimiento nuevo que llega YA respondido igual es novedad.

    Puede ser una reiteración de la EPS o un registro que otra área contestó.
    Si sólo miráramos «sin respuesta» se perdería, y el auditor no se enteraría
    de que hay plata nueva en discusión.
    """
    antes = _informe(tmp_path / "antes.xlsx", [_seg(1)])
    ahora = _informe(tmp_path / "ahora.xlsx", [_seg(1), _seg(9, resp="SI")])

    res = nov.comparar(nov.leer_informe(ahora), nov.leer_informe(antes))

    assert [n["id_seguimiento_factura_glosa"] for n in res["novedades"]] == [9]


def test_sin_informe_anterior_reporta_lo_no_respondido(tmp_path):
    ahora = _informe(
        tmp_path / "ahora.xlsx", [_seg(1, resp="SI"), _seg(2, resp="NO"), _seg(3, resp="")]
    )

    res = nov.comparar(nov.leer_informe(ahora), None)

    assert res["modo"] == "sin-respuesta"
    assert [n["id_seguimiento_factura_glosa"] for n in res["novedades"]] == [2, 3]


def test_avisa_si_desaparecio_algo_que_estaba(tmp_path):
    """Un informe anterior incompleto haría ver como «nuevo» lo que ya estaba."""
    antes = _informe(tmp_path / "antes.xlsx", [_seg(1), _seg(2)])
    ahora = _informe(tmp_path / "ahora.xlsx", [_seg(1)])

    res = nov.comparar(nov.leer_informe(ahora), nov.leer_informe(antes))

    assert [d["id_seguimiento_factura_glosa"] for d in res["desaparecidos"]] == [2]


def test_el_resumen_separa_glosas_de_devoluciones(tmp_path):
    """La devolución se responde por otra puerta: no puede ir mezclada."""
    filas = [
        _seg(1, factura="HUS521868", tipo="GLOSA", valor=1000),
        _seg(2, factura="HUS521868", tipo="GLOSA", valor=2000),
        _seg(3, factura="HUS600001", tipo="DEVOLUCION", eps="FAMISANAR", valor=5000),
    ]
    res = nov.resumen_por_entidad(filas)

    por_clave = {(e, t): (items, facturas, valor) for e, t, items, facturas, valor in res}
    assert por_clave[("SANITAS", "GLOSA")] == (2, 1, 3000)
    assert por_clave[("FAMISANAR", "DEVOLUCION")] == (1, 1, 5000)


def test_una_devolucion_no_multiplica_el_valor_de_la_factura(tmp_path):
    """SIIFA repite el valor de la factura en CADA línea de la devolución.

    Sumarlas infla la cifra hasta lo absurdo: las 1.340 líneas de SANITAS
    daban $24.917 millones cuando esas 9 facturas valen $111 millones. Un
    informe con esa cifra a gerencia es indefendible.
    """
    filas = [
        _seg(i, factura="HUS600001", tipo="DEVOLUCION", valor=10_000_000) for i in range(1, 101)
    ]

    (_, _, items, facturas, valor) = nov.resumen_por_entidad(filas)[0]

    assert (items, facturas) == (100, 1)
    assert valor == 10_000_000, "el valor de la factura devuelta se cuenta UNA vez"


def test_dos_facturas_devueltas_suman_sus_valores(tmp_path):
    filas = [
        _seg(1, factura="HUS600001", tipo="DEVOLUCION", valor=10_000_000),
        _seg(2, factura="HUS600001", tipo="DEVOLUCION", valor=10_000_000),
        _seg(3, factura="HUS600002", tipo="DEVOLUCION", valor=5_000_000),
    ]

    (_, _, items, facturas, valor) = nov.resumen_por_entidad(filas)[0]

    assert (items, facturas, valor) == (3, 2, 15_000_000)


def test_una_glosa_si_suma_item_por_item(tmp_path):
    """La regla de la devolución no puede contagiar a las glosas: una glosa
    objeta un ítem, y 35 ítems de una misma factura sí se suman."""
    filas = [_seg(i, factura="HUS521868", tipo="GLOSA", valor=1000) for i in range(1, 36)]

    (_, _, items, facturas, valor) = nov.resumen_por_entidad(filas)[0]

    assert (items, facturas, valor) == (35, 1, 35_000)


def test_cuenta_facturas_distintas_no_items(tmp_path):
    """35 glosas sobre una factura son 35 ítems y UNA factura."""
    filas = [_seg(i, factura="HUS521868", valor=100) for i in range(1, 36)]

    (_, _, items, facturas, valor) = nov.resumen_por_entidad(filas)[0]

    assert (items, facturas, valor) == (35, 1, 3500)


def test_el_valor_se_lee_con_el_lector_unico(tmp_path):
    """El valor puede venir como texto del Excel: "$3.399.053" son 3.399.053."""
    filas = [_seg(1, valor="$3.399.053")]

    (_, _, _, _, valor) = nov.resumen_por_entidad(filas)[0]

    assert valor == 3399053


def test_el_excel_de_salida_trae_las_novedades_y_el_resumen(tmp_path):
    res = {
        "modo": "comparacion",
        "novedades": [_seg(3, resp="NO")],
        "sin_respuesta": [],
        "desaparecidos": [],
        "total_nuevo": 3,
        "total_anterior": 2,
    }
    salida = tmp_path / "NOVEDADES_SIIFA.xlsx"

    nov.escribir_xlsx(res, salida)

    wb = load_workbook(salida)
    assert wb.sheetnames == ["NOVEDADES", "RESUMEN"]
    ws = wb["NOVEDADES"]
    assert ws.cell(row=1, column=1).value == "Entidad (EPS)"
    fila = [c.value for c in ws[2]]
    assert "HUS521868" in fila and "SANITAS" in fila and 3399053 in fila


def test_un_archivo_que_no_es_el_informe_se_rechaza(tmp_path):
    """Mejor un error claro que una comparación contra cualquier cosa."""
    import pytest

    wb = Workbook()
    wb.active.append(["cualquier", "cosa"])
    ruta = tmp_path / "otro.xlsx"
    wb.save(ruta)

    with pytest.raises(SystemExit, match="informe de seguimientos"):
        nov.leer_informe(ruta)


def test_lee_el_informe_aunque_la_hoja_se_llame_distinto(tmp_path):
    ruta = _informe(tmp_path / "x.xlsx", [_seg(1)], titulo="Hoja1")

    assert len(nov.leer_informe(ruta)) == 1
