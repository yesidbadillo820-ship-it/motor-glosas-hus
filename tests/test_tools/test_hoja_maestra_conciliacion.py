"""Tests del Expediente Inteligente de Conciliacion
(tools/hoja_maestra_conciliacion.py).

Cubre: normalizacion de factura, parseo de valores en formato colombiano, la
logica de contrato 287/440 por fecha, y el flujo completo construir() sobre
fixtures que imitan las tres fuentes reales (CARTERA+CRUCE ACTAS, RECEPCION,
TRAMITE): que la maestra tenga 1 fila por factura (espina = cartera), que cada
glosa cruce su motivo EPS con nuestra respuesta, que una glosa recibida sin
tramite quede 'SIN RESPUESTA', y que los cruces de consistencia detecten la
factura sin glosa.
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

_TOOLS = Path(__file__).resolve().parent.parent.parent / "tools"
sys.path.insert(0, str(_TOOLS))

import hoja_maestra_conciliacion as hm  # noqa: E402


# ---------------------------------------------------------------------------
# Unitarios
# ---------------------------------------------------------------------------
def test_normalizar_factura():
    assert hm.normalizar_factura("HUS0000316347") == "HUS0000316347"
    assert hm.normalizar_factura("316347") == "HUS0000316347"
    assert hm.normalizar_factura(316347) == "HUS0000316347"
    assert hm.normalizar_factura(None) == ""


def test_num_formato_colombiano():
    assert hm._num("$ 3.211.891") == 3211891.0
    assert hm._num("1605989") == 1605989.0
    assert hm._num("1.234,56") == 1234.56
    assert hm._num(None) == 0.0
    assert hm._num("-") == 0.0


def test_contrato_por_fecha():
    # atencion hasta 30-nov-2025 -> 287; desde dic-2025 -> 440 (aprox)
    assert hm.contrato_por_fecha(datetime(2025, 6, 1), "U22031").startswith("287")
    assert hm.contrato_por_fecha(datetime(2026, 1, 15), "C26001").startswith("440")
    assert hm.contrato_por_fecha(None, "") == "PENDIENTE"


# ---------------------------------------------------------------------------
# Fixtures que imitan las tres fuentes reales
# ---------------------------------------------------------------------------
def _fila(n: int, pares: dict) -> list:
    r = [None] * n
    for i, v in pares.items():
        r[i] = v
    return r


def _cartera(ruta: Path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CARTERA"
    ws.append(_fila(25, {0: "ENTIDAD: DIRECCION..."}))  # fila 0 = banner
    hdr = [None] * 25
    hdr[hm.CA_VIGENCIA], hdr[hm.CA_ESTADO], hdr[hm.CA_FACTURA] = "VIGENCIA", "ESTADO", "FACTURA"
    hdr[hm.CA_VALOR_FAC], hdr[hm.CA_VALOR_GLOSA] = "VALOR FACTURA", "VALOR GLOSA"
    ws.append(hdr)
    # Factura A: con glosa; Factura B: sin glosa (VALOR GLOSA = 0)
    ws.append(
        _fila(
            25,
            {
                hm.CA_VIGENCIA: "2025",
                hm.CA_ESTADO: "-",
                hm.CA_FACTURA: "HUS0000316347",
                hm.CA_FECHA_RAD: datetime(2025, 1, 5),
                hm.CA_VALOR_FAC: 38012636,
                hm.CA_SALDO_DGH: 6865663,
                hm.CA_VALOR_GLOSA: 28345,
                hm.CA_LEVANTADA: 0,
                hm.CA_ACEPTADA: 28345,
                hm.CA_RATIFICADA: 0,
                hm.CA_ACTAS: "AC000617",
                hm.CA_ESTADO_GLOSA: "CONCILIADA",
                hm.CA_EDADES: "0-90",
            },
        )
    )
    ws.append(
        _fila(
            25,
            {
                hm.CA_VIGENCIA: "2025",
                hm.CA_ESTADO: "-",
                hm.CA_FACTURA: "HUS0000400000",
                hm.CA_FECHA_RAD: datetime(2025, 2, 1),
                hm.CA_VALOR_FAC: 100000,
                hm.CA_SALDO_DGH: 0,
                hm.CA_VALOR_GLOSA: 0,
                hm.CA_EDADES: "0-90",
            },
        )
    )

    # Hoja CRUCE ACTAS (title + header + 1 par + totales)
    cr = wb.create_sheet("CRUCE ACTAS")
    cr.append(_fila(8, {0: "DETALLE FACTURA x ACTA"}))  # titulo
    cr.append(
        [
            "FACTURA",
            "ACTA",
            "TIPO",
            "GLOSA INICIAL",
            "ACEPTADA",
            "LEVANTADA",
            "RATIFICADA",
            "A PAGAR",
        ]
    )
    cr.append(["HUS0000316347", "AC000617", "CONCILIACION", 28345, 28345, 0, 0, 0])
    cr.append([None, None, None, 28345, 28345, 0, 0, 0])  # fila de totales (factura vacia)
    wb.save(str(ruta))


def _recepcion(ruta: Path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CRRPSEGUIMIENTORECEPCIONOBJECIO"
    ws.append(_fila(34, {hm.RE_FACTURA: "FACTURA", hm.RE_MOTIVO: "MOTIVO"}))
    # Consecutivo 104341 (factura A) -> tendra tramite; 200000 (factura A) -> SIN respuesta
    ws.append(
        _fila(
            34,
            {
                hm.RE_FACTURA: "HUS0000316347",
                hm.RE_CONSECUTIVO: "104341",
                hm.RE_FECHA_OBJ: datetime(2024, 12, 23),
                hm.RE_FECHA_FAC: datetime(2024, 10, 17),
                hm.RE_TIPO: "Mixta",
                hm.RE_COD_CONCEPTO: "604",
                hm.RE_NOM_CONCEPTO: "PERTINENCIA",
                hm.RE_CUPS: "890201",
                hm.RE_SERVICIO: "CONSULTA CONTROL",
                hm.RE_VALOR_OBJ: 28345,
                hm.RE_MOTIVO: "SE RECONOCE A TARIFA SOAT SIN CONTRATO",
            },
        )
    )
    ws.append(
        _fila(
            34,
            {
                hm.RE_FACTURA: "HUS0000316347",
                hm.RE_CONSECUTIVO: "200000",
                hm.RE_FECHA_OBJ: datetime(2025, 3, 1),
                hm.RE_FECHA_FAC: datetime(2024, 10, 17),
                hm.RE_TIPO: "Administrativo",
                hm.RE_COD_CONCEPTO: "701",
                hm.RE_NOM_CONCEPTO: "SOPORTES",
                hm.RE_VALOR_OBJ: 5000,
                hm.RE_MOTIVO: "FALTA SOPORTE",
            },
        )
    )
    wb.save(str(ruta))


def _tramite(ruta: Path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CRRPSEGUIMIENTOTRAMITEOBJECION"
    ws.append(_fila(33, {hm.TR_FACTURA: "FACTURA", hm.TR_ARGUMENTO: "OBS"}))
    # Respuesta a la glosa del consecutivo 104341 (factura A)
    ws.append(
        _fila(
            33,
            {
                hm.TR_FACTURA: "HUS0000316347",
                hm.TR_REC_CONSECUTIVO: "104341",
                hm.TR_FECHA_OBJ: datetime(2024, 12, 23),
                hm.TR_FECHA_FAC: datetime(2024, 10, 17),
                hm.TR_FECHA_DOC: datetime(2025, 1, 13),
                hm.TR_CONSECUTIVO: "105013",
                hm.TR_TIPO: "Mixta",
                hm.TR_CONTRATO_COD: "U22031",
                hm.TR_COD_GLOSA: "604",
                hm.TR_NOM_GLOSA: "PERTINENCIA",
                hm.TR_VALOR_OBJ: 28345,
                hm.TR_VALOR_ACEPTADO: 28345,
                hm.TR_ARGUMENTO: "604- ESE HUS ACEPTA OBJECION",
            },
        )
    )
    wb.save(str(ruta))


def test_construir_expediente_completo(tmp_path):
    cart = tmp_path / "cartera.xlsx"
    rec = tmp_path / "recepcion.xlsx"
    tra = tmp_path / "tramite.xlsx"
    _cartera(cart)
    _recepcion(rec)
    _tramite(tra)
    salida = tmp_path / "EXPEDIENTE.xlsx"

    r = hm.construir(cart, rec, tra, salida)

    # Maestra: 1 fila por factura (espina = cartera, 2 facturas)
    assert r["facturas_maestra"] == 2
    assert r["facturas_con_glosa"] == 1  # solo A tiene glosas
    # 2 glosas de A: la respondida (104341) + la recibida sin respuesta (200000)
    assert r["glosas"] == 2
    assert r["glosas_sin_respuesta"] == 1
    assert r["actas_filas"] == 1  # el par factura+acta (la fila de totales se excluye)
    assert r["total_aceptado"] == 28345
    assert salida.exists()

    # Verifica el contenido cruzado en la hoja de glosas
    import openpyxl

    wb = openpyxl.load_workbook(str(salida), data_only=True)
    assert wb.sheetnames == ["00_DASHBOARD", "01_MAESTRA", "02_GLOSAS", "03_ACTAS", "04_CRUCES"]
    gs = wb["02_GLOSAS"]
    filas = list(gs.iter_rows(values_only=True))
    hdr = list(filas[0])
    i_mot = hdr.index("MOTIVO EXACTO EPS")
    i_arg = hdr.index("ARGUMENTO ESE HUS")
    i_est = hdr.index("ESTADO RESPUESTA")
    respondida = next(f for f in filas[1:] if f[i_arg])
    # el motivo EPS (de la recepcion) y nuestra respuesta viajan en la MISMA fila
    assert "SOAT" in respondida[i_mot]
    assert "ESE HUS ACEPTA" in respondida[i_arg]
    assert respondida[i_est] == "RESPONDIDA"
    sin_resp = next(f for f in filas[1:] if f[i_est] == "SIN RESPUESTA")
    assert "SOPORTE" in sin_resp[i_mot]

    # La maestra marca la factura sin glosa
    ms = wb["01_MAESTRA"]
    mfilas = list(ms.iter_rows(values_only=True))
    mhdr = list(mfilas[0])
    i_fac = mhdr.index("FACTURA")
    i_res = mhdr.index("RESULTADO FINAL")
    fila_b = next(f for f in mfilas[1:] if f[i_fac] == "HUS0000400000")
    assert fila_b[i_res] == "SIN GLOSA"
    wb.close()

    # El cruce #1 (facturas sin glosa) detecta la factura B
    assert r["cruces"][1] == 1
