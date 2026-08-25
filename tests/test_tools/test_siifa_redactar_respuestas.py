"""Pruebas del redactor de respuestas de SIIFA.

Lo que se cuida acá es lo que puede costarle plata al hospital en una
conciliación: que la respuesta hable de LA factura y LA glosa que la EPS
objetó, que cite la norma correcta según la causal, y que no afirme hechos
clínicos que no salen de la base.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

TOOLS = Path(__file__).resolve().parents[2] / "tools"
sys.path.insert(0, str(TOOLS))

import siifa_redactar_respuestas as red  # noqa: E402

COLS = [
    "id_seguimiento_factura_glosa",
    "tipo_seguimiento",
    "numero_factura",
    "valor_glosa",
    "codigo_glosa",
    "descripcion_glosa",
    "observacion_glosa",
    "fecha_reporte",
    "tiene_respuesta",
]


def _linea(codigo, tipo="GLOSA", factura="HUS454747", valor=22200, obs="null", desc=None):
    return {
        "id_seguimiento_factura_glosa": 1,
        "tipo_seguimiento": tipo,
        "numero_factura": factura,
        "valor_glosa": valor,
        "codigo_glosa": codigo,
        "descripcion_glosa": desc or f"Causal del código {codigo}.",
        "observacion_glosa": obs,
        "fecha_reporte": "2026-06-09T17:07:47",
        "tiene_respuesta": "NO",
    }


def _informe(tmp_path: Path, filas: list[dict]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "SEGUIMIENTOS"
    ws.append(COLS)
    for f in filas:
        ws.append([f.get(c) for c in COLS])
    ruta = tmp_path / "informe.xlsx"
    wb.save(ruta)
    return ruta


def test_la_respuesta_nombra_la_factura_el_codigo_y_el_valor():
    """Una respuesta genérica que no identifica la glosa no sirve de nada."""
    r = red.redactar(_linea("TA0801", factura="HUS454747", valor=22200))

    texto = r["OBSERVACION_RESPUESTA"]
    assert "HUS454747" in texto
    assert "TA0801" in texto
    assert "$22.200" in texto
    assert r["CODIGO_RESPUESTA"] == "RE9901"


@pytest.mark.parametrize(
    "valor, esperado",
    [
        (1479360, "$1.479.360"),
        ("1.479.360", "$1.479.360"),  # como viene de un Excel con formato
        ("$ 1.479.360", "$1.479.360"),
        (22200.0, "$22.200"),
        (None, "SIN VALOR REGISTRADO"),
    ],
)
def test_el_valor_se_escribe_bien_venga_como_venga(valor, esperado):
    """El valor va escrito en una respuesta que se le manda a la EPS.

    Se lee con a_entero, el único lector de pesos del repo: leerlo aparte es
    justo lo que produjo cifras multiplicadas por cien en otros módulos.
    """
    texto = red.redactar(_linea("TA0801", valor=valor))["OBSERVACION_RESPUESTA"]

    assert esperado in texto


def test_la_respuesta_cita_lo_que_dijo_la_eps():
    """Responder sin recoger el argumento de la entidad es responder al aire."""
    r = red.redactar(
        _linea("TA0601", obs="SE OBJETA SISTEMA DE SUCCION CERRADA NO6 POR MAYOR VALOR COBRADO")
    )

    assert "SISTEMA DE SUCCION CERRADA" in r["OBSERVACION_RESPUESTA"]


def test_si_la_eps_no_sustento_nada_la_respuesta_lo_dice():
    """Una glosa sin motivación es, por sí sola, un argumento de defensa."""
    r = red.redactar(_linea("TA0801", obs="null"))

    assert "NO REGISTRÓ OBSERVACIÓN" in r["OBSERVACION_RESPUESTA"]


@pytest.mark.parametrize(
    "codigo, debe_citar",
    [
        ("TA0801", "RESOLUCIÓN 2284 DE 2023"),  # tarifas: motivación de la glosa
        ("AU5801", "LEY 1751 DE 2015"),  # autorización: autonomía médica
        ("CL0801", "LEY 1751 DE 2015"),  # pertinencia: autonomía médica
        ("SO0801", "RESOLUCIÓN 1995 DE 1999"),  # soportes: historia clínica
        ("CO2301", "LEY 1751 DE 2015"),  # cobertura
        ("DE5601", "ARTÍCULO 17 DE LA RESOLUCIÓN 2284"),  # los 22 días hábiles
    ],
)
def test_cada_causal_se_defiende_con_la_norma_que_le_corresponde(codigo, debe_citar):
    texto = red.redactar(_linea(codigo))["OBSERVACION_RESPUESTA"]

    assert debe_citar in texto


def test_un_codigo_desconocido_no_se_queda_sin_respuesta():
    """La familia del código (las dos primeras letras) alcanza para defender."""
    r = red.redactar(_linea("TA9999"))

    assert r["CODIGO_RESPUESTA"] == "RE9901"
    assert "TARIFA INSTITUCIONAL" in r["OBSERVACION_RESPUESTA"]


def test_una_devolucion_no_se_responde_como_una_glosa():
    glosa = red.redactar(_linea("TA0801", tipo="GLOSA"))["OBSERVACION_RESPUESTA"]
    devolucion = red.redactar(_linea("DE5601", tipo="DEVOLUCION"))["OBSERVACION_RESPUESTA"]

    assert "NO ACEPTA LA GLOSA APLICADA" in glosa
    assert "NO ACEPTA LA DEVOLUCIÓN APLICADA" in devolucion
    assert "REPROCESAR LA CUENTA" in devolucion


def test_toda_respuesta_cierra_pidiendo_conciliacion_y_con_el_correo():
    """Es la fórmula institucional: sin ella la respuesta queda coja."""
    texto = red.redactar(_linea("TA0801"))["OBSERVACION_RESPUESTA"]

    assert "MESA DE CONCILIACIÓN" in texto
    assert "CARTERA@HUS.GOV.CO" in texto
    assert "LEY 1438 DE 2011" in texto


def test_toda_respuesta_dice_que_hay_que_verificar_antes_de_subirla():
    """El auditor tiene que saber qué soporte sostiene cada respuesta."""
    for codigo in ("TA0801", "AU5801", "SO0801", "FA0603", "CO2301", "DE5601", "DE1601"):
        assert red.redactar(_linea(codigo))["REVISAR"], f"{codigo} sin nota de revisión"


def test_la_respuesta_de_soportes_avisa_que_sin_el_soporte_no_se_sostiene():
    """Es la única causal donde responder sin el papel es contraproducente."""
    assert "NO se puede sostener" in red.redactar(_linea("SO0801"))["REVISAR"]


def test_una_observacion_larguisima_no_se_come_el_cierre():
    """La cita de la EPS se recorta; el sustento propio nunca."""
    r = red.redactar(_linea("TA0801", obs="X" * 5000))

    texto = r["OBSERVACION_RESPUESTA"]
    assert len(texto) <= 1900
    assert "CARTERA@HUS.GOV.CO" in texto
    assert "TARIFA INSTITUCIONAL" in texto


def test_el_texto_de_la_eps_llega_limpio():
    """El export trae _x000D_ y saltos de línea que ensucian la respuesta."""
    r = red.redactar(_linea("TA0801", obs="PRIMERA LINEA_x000D_\nSEGUNDA   LINEA"))

    texto = r["OBSERVACION_RESPUESTA"]
    assert "_x000D_" not in texto
    assert "\n" not in texto
    assert "PRIMERA LINEA SEGUNDA LINEA" in texto


def test_lo_que_el_hospital_ya_respondio_manda_sobre_lo_redactado(tmp_path):
    """Si existe la respuesta real del HUS, esa es la que va — no una nueva."""
    informe = _informe(tmp_path, [_linea("TA0801", valor=22200)])
    wb = Workbook()
    ws = wb.active
    ws.title = "BD TRAMITE"
    ws.append(
        [
            "FACTURA",
            "ListadoConceptos.DetalleRecepcionObjecion.ConceptoObjecion.Codigo",
            "ListadoConceptos.DetalleRecepcionObjecion.ValorObjecion",
            "codigo_respuesta",
            "observacion_respuesta",
            "fecha_respuesta",
            "EstadoActual",
        ]
    )
    ws.append(
        [
            "HUS0000454747",
            "TA0801",
            22200,
            "RE9602",
            "RESPUESTA REAL DEL HUS.",
            "2026-01-06",
            "Confirmado",
        ]
    )
    tramites = tmp_path / "tramites.xlsx"
    wb.save(tramites)

    filas = red.armar(informe, tramites)

    assert len(filas) == 1
    assert filas[0]["CODIGO_RESPUESTA"] == "RE9602"
    assert filas[0]["OBSERVACION_RESPUESTA"] == "RESPUESTA REAL DEL HUS."
    assert filas[0]["ORIGEN_RESPUESTA"] == "EXACTO"


def test_la_respuesta_de_dgh_viaja_con_la_fecha_en_que_se_dio(tmp_path):
    """La fecha que se digita en SIIFA es la del día en que el HUS respondió.

    Si va con la fecha de hoy, en el histórico de SIIFA la respuesta queda
    registrada fuera del término y la EPS se agarra de ahí.
    """
    informe = _informe(tmp_path, [_linea("TA0801", valor=22200)])
    tramites = _tramites(
        tmp_path,
        [
            [
                "HUS0000454747",
                "TA0801",
                22200,
                "RE9602",
                "RESPUESTA REAL DEL HUS.",
                "2026-05-11 12:17:44",
                "Confirmado",
            ]
        ],
    )

    filas = red.armar(informe, tramites)

    assert filas[0]["FECHA_RESPUESTA"] == "2026-05-11"


def test_lo_redactado_va_sin_fecha_porque_se_responde_hoy(tmp_path):
    informe = _informe(tmp_path, [_linea("TA0801")])

    filas = red.armar(informe, None)

    assert filas[0]["FECHA_RESPUESTA"] is None


def test_sin_base_de_tramites_se_redacta_todo(tmp_path):
    informe = _informe(tmp_path, [_linea("TA0801"), _linea("DE5601", tipo="DEVOLUCION")])

    filas = red.armar(informe, None)

    assert len(filas) == 2
    assert all(f["ORIGEN_RESPUESTA"] == "REDACTADA" for f in filas)
    assert all(f["CODIGO_RESPUESTA"] == "RE9901" for f in filas)


def _tramites(tmp_path: Path, filas: list[list]) -> Path:
    """Base de DGH: una fila por respuesta que el hospital ya dio."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BD TRAMITE"
    ws.append(
        [
            "FACTURA",
            "ListadoConceptos.DetalleRecepcionObjecion.ConceptoObjecion.Codigo",
            "ListadoConceptos.DetalleRecepcionObjecion.ValorObjecion",
            "codigo_respuesta",
            "observacion_respuesta",
            "fecha_respuesta",
            "EstadoActual",
        ]
    )
    for f in filas:
        ws.append(f)
    ruta = tmp_path / "tramites.xlsx"
    wb.save(ruta)
    return ruta


def _correr(monkeypatch, tmp_path, informe, tramites, extra=()):
    glosas = tmp_path / "respuestas_GLOSAS.xlsx"
    devoluciones = tmp_path / "respuestas_DEVOLUCIONES.xlsx"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "siifa_redactar_respuestas.py",
            "--informe",
            str(informe),
            "--tramites",
            str(tramites),
            "--salida-glosas",
            str(glosas),
            "--salida-devoluciones",
            str(devoluciones),
            *extra,
        ],
    )
    red.main()
    return glosas, devoluciones


def _facturas(ruta: Path) -> list[str]:
    ws = load_workbook(ruta).active
    col = [c.value for c in ws[1]].index("NUMERO_FACTURA") + 1
    return [ws.cell(row=r, column=col).value for r in range(2, ws.max_row + 1)]


def test_solo_lo_ya_respondido_deja_fuera_del_cargue_lo_redactado(monkeypatch, tmp_path):
    """El archivo de cargue lleva sólo la respuesta real del hospital."""
    informe = _informe(
        tmp_path,
        [
            _linea("TA0801", factura="HUS454747", valor=22200),
            _linea("SO0101", factura="HUS532384", valor=51000),
        ],
    )
    tramites = _tramites(
        tmp_path,
        [
            [
                "HUS0000454747",
                "TA0801",
                22200,
                "RE9602",
                "RESPUESTA REAL DEL HUS.",
                "2026-01-06",
                "Confirmado",
            ]
        ],
    )

    glosas, _ = _correr(monkeypatch, tmp_path, informe, tramites, ["--solo-lo-ya-respondido"])

    assert _facturas(glosas) == ["HUS454747"]


def test_lo_redactado_no_se_bota_queda_en_su_propio_archivo(monkeypatch, tmp_path):
    """Sacarlo del cargue no es perderlo: hay que poder revisarlo y subirlo."""
    informe = _informe(
        tmp_path,
        [
            _linea("TA0801", factura="HUS454747", valor=22200),
            _linea("SO0101", factura="HUS532384", valor=51000),
        ],
    )
    tramites = _tramites(
        tmp_path,
        [
            [
                "HUS0000454747",
                "TA0801",
                22200,
                "RE9602",
                "RESPUESTA REAL DEL HUS.",
                "2026-01-06",
                "Confirmado",
            ]
        ],
    )

    glosas, _ = _correr(monkeypatch, tmp_path, informe, tramites, ["--solo-lo-ya-respondido"])

    aparte = red.con_sufijo(glosas, "REDACTADAS")
    assert aparte.exists()
    assert _facturas(aparte) == ["HUS532384"]


def test_sin_la_opcion_el_cargue_sigue_llevando_todo(monkeypatch, tmp_path):
    """Lo de siempre no cambia: sin la opción, van las dos en un solo archivo."""
    informe = _informe(
        tmp_path,
        [
            _linea("TA0801", factura="HUS454747", valor=22200),
            _linea("SO0101", factura="HUS532384", valor=51000),
        ],
    )
    tramites = _tramites(
        tmp_path,
        [
            [
                "HUS0000454747",
                "TA0801",
                22200,
                "RE9602",
                "RESPUESTA REAL DEL HUS.",
                "2026-01-06",
                "Confirmado",
            ]
        ],
    )

    glosas, _ = _correr(monkeypatch, tmp_path, informe, tramites)

    assert sorted(_facturas(glosas)) == ["HUS454747", "HUS532384"]
    assert not red.con_sufijo(glosas, "REDACTADAS").exists()


def test_el_archivo_que_sale_lo_puede_leer_el_bot_de_cargue(tmp_path):
    """Las columnas que exige responder_glosas_siifa.py tienen que estar."""
    informe = _informe(tmp_path, [_linea("TA0801")])
    filas = red.armar(informe, None)
    salida = tmp_path / "respuestas.xlsx"

    red.escribir(filas, salida, "Glosas")

    ws = load_workbook(salida).active
    encabezados = {c.value for c in ws[1]}
    assert {"ID_SEGUIMIENTO_FACTURA_GLOSA", "CODIGO_RESPUESTA", "OBSERVACION_RESPUESTA"} <= (
        encabezados
    )
    assert ws.max_row == 2
