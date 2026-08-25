"""Tests del organizador de objeciones del Dispensario
(tools/organizar_objeciones_dispensario.py).

Cubre: conversión de valores COP, limpieza de artefactos (cid:N), extracción
del código de servicio del motivo, número de CxC en formato DGH, armado de
las filas OBJECIONES, escritura del Excel con los formatos del ejemplo de
EMSSANAR, y el parseo end-to-end de un PDF sintético con el mismo layout de
columnas del "DETALLE DE AUDITORIA Y GLOSAS" (generado con reportlab).
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import pytest

# El script vive en tools/ (sin __init__.py): lo importamos por ruta.
_TOOLS = Path(__file__).resolve().parent.parent.parent / "tools"
sys.path.insert(0, str(_TOOLS))

import organizar_objeciones_dispensario as org  # noqa: E402

# ---------------------------------------------------------------------------
# Unidades puras
# ---------------------------------------------------------------------------


def test_to_int_formato_colombia():
    assert org._to_int("4.677.656,00") == 4677656
    assert org._to_int("3.700,00") == 3700
    assert org._to_int("0,00") == 0


def test_limpiar_remueve_cid_y_colapsa_espacios():
    assert org._limpiar("POR(cid:9) TOALLA  IMPREGNADA") == "POR TOALLA IMPREGNADA"


def test_servicio_desde_motivo():
    g = org.Glosa(
        codigo="CL0301",
        concepto="",
        valor=1,
        motivo="SE GLOSA CODIGO 21519H LOCALIZACION DE LESION",
    )
    assert org._servicio_de(g) == "21519H"
    g.motivo = "SE GLOSA ESTANCIA POR INOPORTUNIDAD"
    assert org._servicio_de(g) == ""


def test_cxc_formato_dgh():
    fac = org.Factura(prefijo="HUS", numero="530265")
    assert fac.cxc == "HUS0000530265"


def test_parse_fecha_dos_y_cuatro_digitos():
    assert org._parse_fecha("10/07/26") == datetime(2026, 7, 10)
    assert org._parse_fecha("10/07/2026") == datetime(2026, 7, 10)


def test_filas_objeciones_layout_dgh():
    fac = org.Factura(prefijo="HUS", numero="530265")
    fac.glosas.append(
        org.Glosa(
            codigo="SO0801",
            concepto="SOPORTES-APOYO DIAGNÓSTICO - EXISTE AUSENCIA",
            motivo="SE GLOSA CODIGO 20095751 3 ONDANSETRON AMPOLLA.",
            valor=3700,
        )
    )
    (fila,) = org.filas_objeciones(fac, datetime(2026, 7, 10))
    assert len(fila) == len(org.ENCABEZADOS) == 16
    assert fila[0] == "1"  # CDCONSEC
    assert fila[2] == "HUS0000530265"  # CRNCXC
    assert fila[8] == "999"  # GENUSUARIO4
    assert fila[9] == "SO0801"  # CRNCONOBJ
    assert fila[10] == "20095751"  # SLNSERPRO
    assert fila[13] == 3700  # CROVALOBJ
    assert fila[14].startswith("SO0801 SOPORTES-APOYO") and fila[14].endswith("$3700")
    assert fila[15] == 0  # CROTIPOBJ


def test_escribir_excel_formatos(tmp_path):
    fac = org.Factura(prefijo="HUS", numero="1")
    fac.glosas.append(org.Glosa(codigo="CL0101", concepto="C", motivo="M", valor=100))
    salida = tmp_path / "obj.xlsx"
    org.escribir_excel(org.filas_objeciones(fac, datetime(2026, 7, 10)), salida)

    from openpyxl import load_workbook

    ws = load_workbook(salida)["OBJECIONES"]
    assert [c.value for c in ws[1]] == org.ENCABEZADOS
    assert ws.cell(row=2, column=1).number_format == "@"
    assert ws.cell(row=2, column=2).number_format == org.FMT_FECHA
    assert ws.cell(row=2, column=14).number_format == org.FMT_CONTABLE
    assert ws.cell(row=2, column=16).number_format == "0"
    assert ws.cell(row=2, column=3).value == "HUS0000000001"


# ---------------------------------------------------------------------------
# End-to-end con un PDF sintético del mismo layout
# ---------------------------------------------------------------------------


def _pdf_sintetico(ruta: Path) -> None:
    """Dibuja un mini DETALLE DE AUDITORIA Y GLOSAS con las columnas reales."""
    reportlab = pytest.importorskip("reportlab")  # noqa: F841
    from reportlab.pdfgen import canvas

    c = canvas.Canvas(str(ruta), pagesize=(1029.45, 633.45))
    alto = 633.45
    c.setFont("Helvetica", 8)

    def linea(top, celdas):
        for x, txt in celdas:
            c.drawString(x, alto - top - 8, txt)

    # Encabezado de factura + primera glosa en el mismo renglón (como el real).
    linea(
        175,
        [
            (115, "25/05/2026"),
            (165, "JOSE PEREZ"),
            (331, "34.231.568,00"),
            (400, "622.967,00"),
            (443, "CL03 01 CALIDAD-HONORARIOS"),
            (661, "SE GLOSA CODIGO 21519H LOCALIZACION"),
        ],
    )
    linea(176, [(33, "HUS"), (69, "530265")])
    linea(187, [(443, "PROFESIONALES NO PERTINENTES"), (661, "VALOR A OBJETAR 622967 PESOS.")])
    # Segunda glosa.
    linea(
        261,
        [
            (408, "3.700,00"),
            (443, "SO08 01 SOPORTES-APOYO DIAGNÓSTICO"),
            (661, "SE GLOSA CODIGO 20095751 3 ONDANSETRON"),
        ],
    )
    # Cierre de factura con el total del PDF (622.967 + 3.700).
    linea(
        290,
        [
            (39, "Total"),
            (61, "Factura:"),
            (99, "HUS"),
            (129, "530265"),
            (395, "626.667,00"),
            (819, "Valor a pagar:"),
            (886, "33.604.901,00"),
        ],
    )
    c.drawString(420, alto - 599 - 8, "Fecha Impresión: 10/07/26 07:58:27")
    c.save()


def test_parsear_pdf_sintetico(tmp_path):
    pdf = tmp_path / "detalle.pdf"
    _pdf_sintetico(pdf)

    facturas, fecha_imp = org.parsear_pdf(pdf)
    assert fecha_imp == "10/07/26"
    assert len(facturas) == 1
    fac = facturas[0]
    assert fac.cxc == "HUS0000530265"
    assert fac.paciente == "JOSE PEREZ"
    assert fac.fecha_atencion == "25/05/2026"
    assert fac.valor_factura == 34231568
    assert fac.total_objetado_pdf == 626667
    assert [g.codigo for g in fac.glosas] == ["CL0301", "SO0801"]
    assert [g.valor for g in fac.glosas] == [622967, 3700]
    assert fac.total_extraido == fac.total_objetado_pdf
    assert "LOCALIZACION" in fac.glosas[0].motivo
    assert "NO PERTINENTES" in fac.glosas[0].concepto


def test_procesar_pdf_sintetico_genera_excel(tmp_path):
    pdf = tmp_path / "detalle.pdf"
    _pdf_sintetico(pdf)

    rc = org.procesar([pdf], tmp_path, "DISPENSARIO", None, None)
    assert rc == 0
    salida = tmp_path / "OBJECIONES_DISPENSARIO_HUS0000530265.xlsx"
    assert salida.exists()

    from openpyxl import load_workbook

    ws = load_workbook(salida)["OBJECIONES"]
    filas = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(filas) == 2
    assert filas[0][2] == "HUS0000530265"
    assert filas[0][1] == datetime(2026, 7, 10)  # CDFECDOC = Fecha Impresión
    assert sum(f[13] for f in filas) == 626667


def test_procesar_detecta_descuadre(tmp_path, caplog):
    """Si el total extraído no cuadra con el Total Factura del PDF → exit 1."""
    from reportlab.pdfgen import canvas

    pytest.importorskip("reportlab")

    pdf = tmp_path / "descuadre.pdf"
    c = canvas.Canvas(str(pdf), pagesize=(1029.45, 633.45))
    alto = 633.45
    c.setFont("Helvetica", 8)
    c.drawString(400, alto - 175 - 8, "100,00")
    c.drawString(443, alto - 175 - 8, "CL03 01 CALIDAD")
    c.drawString(661, alto - 175 - 8, "MOTIVO X")
    c.drawString(33, alto - 176 - 8, "HUS")
    c.drawString(69, alto - 176 - 8, "1")
    # El PDF declara un total distinto (999) al extraído (100).
    c.drawString(39, alto - 290 - 8, "Total")
    c.drawString(61, alto - 290 - 8, "Factura:")
    c.drawString(395, alto - 290 - 8, "999,00")
    c.save()

    rc = org.procesar([pdf], tmp_path, "DISPENSARIO", "10/07/2026", None)
    assert rc == 1
    assert "DESCUADRE" in caplog.text
