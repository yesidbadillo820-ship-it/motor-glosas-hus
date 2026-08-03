"""Pruebas del cruce entre lo pendiente en SIIFA y lo ya respondido en DGH.

En el caso real (03-08-2026), 1.082 de las 2.579 líneas pendientes en SIIFA
ya tenían respuesta escrita por el hospital en Dinámica Gerencial: nunca se
cargaron a la plataforma del Ministerio. Volver a redactarlas sería trabajo
perdido y, peor, dejaría dos respuestas distintas para la misma glosa.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

TOOLS = Path(__file__).resolve().parents[2] / "tools"
sys.path.insert(0, str(TOOLS))

import siifa_cruzar_tramites_dgh as cruce  # noqa: E402
import siifa_preparar_respuestas as prep  # noqa: E402

COLS_SIIFA = [
    "id_seguimiento_factura_glosa",
    "tipo_seguimiento",
    "numero_factura",
    "valor_glosa",
    "codigo_glosa",
    "descripcion_glosa",
    "observacion_glosa",
    "fecha_reporte",
    "tiene_respuesta",
]

# Los nombres largos que trae de verdad el export de DGH.
COLS_DGH = [
    "FACTURA",
    "ListadoConceptos.DetalleRecepcionObjecion.ConceptoObjecion.Codigo",
    "ListadoConceptos.DetalleRecepcionObjecion.ValorObjecion",
    "codigo_respuesta",
    "observacion_respuesta",
    "fecha_respuesta",
    "EstadoActual",
]


def _informe_siifa(tmp_path: Path, filas: list[dict]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "SEGUIMIENTOS"
    ws.append(COLS_SIIFA)
    for f in filas:
        ws.append([f.get(c) for c in COLS_SIIFA])
    ruta = tmp_path / "informe.xlsx"
    wb.save(ruta)
    return ruta


def _tramites_dgh(tmp_path: Path, filas: list[dict]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "BD TRAMITE"
    ws.append(COLS_DGH)
    for f in filas:
        ws.append([f.get(c) for c in COLS_DGH])
    ruta = tmp_path / "tramites.xlsx"
    wb.save(ruta)
    return ruta


def _glosa(id_, factura, codigo, valor=1000, obs="null"):
    return {
        "id_seguimiento_factura_glosa": id_,
        "tipo_seguimiento": "GLOSA",
        "numero_factura": factura,
        "valor_glosa": valor,
        "codigo_glosa": codigo,
        "descripcion_glosa": f"Descripción de {codigo}",
        "observacion_glosa": obs,
        "fecha_reporte": "2026-06-09T17:07:47",
        "tiene_respuesta": "NO",
    }


def _tramite(factura, codigo, valor, resp="RE9901", obs="ESE HUS NO ACEPTA.", fecha="2026-01-06"):
    return {
        "FACTURA": factura,
        "ListadoConceptos.DetalleRecepcionObjecion.ConceptoObjecion.Codigo": codigo,
        "ListadoConceptos.DetalleRecepcionObjecion.ValorObjecion": valor,
        "codigo_respuesta": resp,
        "observacion_respuesta": obs,
        "fecha_respuesta": fecha,
        "EstadoActual": "Confirmado",
    }


def _cruzar(informe: Path, tramites: Path):
    datos = prep.leer_informe(informe)
    grupos = prep.agrupar(datos)
    facturas = {cruce.normalizar_factura(v[0]["numero_factura"]) for v in grupos.values()}
    fino, grueso = cruce.indexar(cruce.leer_tramites(tramites, facturas))
    return grupos, cruce.cruzar(grupos, fino, grueso)


@pytest.mark.parametrize(
    "en_siifa, en_dgh",
    [
        ("HUS454747", "HUS0000454747"),  # el caso real: DGH rellena con ceros
        ("HUS454747", "HUS454747"),
        ("454747", "HUS0000454747"),
    ],
)
def test_la_factura_calza_aunque_venga_escrita_distinto(tmp_path, en_siifa, en_dgh):
    """SIIFA escribe HUS454747 y DGH escribe HUS0000454747: son la misma."""
    informe = _informe_siifa(tmp_path, [_glosa(1, en_siifa, "TA0801", valor=22200)])
    tramites = _tramites_dgh(tmp_path, [_tramite(en_dgh, "TA0801", 22200)])

    _, previas = _cruzar(informe, tramites)

    assert len(previas) == 1
    assert list(previas.values())[0]["CODIGO_RESPUESTA"] == "RE9901"
    assert list(previas.values())[0]["ORIGEN"] == "EXACTO"


def test_si_el_valor_no_calza_igual_sirve_pero_queda_marcado(tmp_path):
    """La respuesta al mismo código sirve, pero el auditor debe mirarla."""
    informe = _informe_siifa(tmp_path, [_glosa(1, "HUS454747", "TA0801", valor=22200)])
    tramites = _tramites_dgh(tmp_path, [_tramite("HUS0000454747", "TA0801", 99999)])

    _, previas = _cruzar(informe, tramites)

    resultado = list(previas.values())[0]
    assert resultado["ORIGEN"] == "POR_CODIGO"
    assert "no calza" in resultado["REVISAR"]


def test_si_dgh_tiene_dos_respuestas_distintas_queda_la_ultima_y_avisa(tmp_path):
    """Pasa cuando se corrigió la respuesta o la escribieron dos personas."""
    informe = _informe_siifa(tmp_path, [_glosa(1, "HUS454747", "TA0801", valor=22200)])
    tramites = _tramites_dgh(
        tmp_path,
        [
            _tramite("HUS0000454747", "TA0801", 22200, obs="PRIMERA VERSIÓN", fecha="2026-01-06"),
            _tramite("HUS0000454747", "TA0801", 22200, obs="VERSIÓN CORREGIDA", fecha="2026-03-20"),
        ],
    )

    _, previas = _cruzar(informe, tramites)

    resultado = list(previas.values())[0]
    assert resultado["OBSERVACION_RESPUESTA"] == "VERSIÓN CORREGIDA"
    assert "más de una respuesta" in resultado["REVISAR"]


def test_una_glosa_sin_tramite_previo_queda_para_escribir(tmp_path):
    informe = _informe_siifa(tmp_path, [_glosa(1, "HUS999999", "TA0801")])
    tramites = _tramites_dgh(tmp_path, [_tramite("HUS0000454747", "TA0801", 22200)])

    grupos, previas = _cruzar(informe, tramites)

    assert len(grupos) == 1
    assert previas == {}


def test_un_tramite_sin_respuesta_escrita_no_cuenta_como_respondido(tmp_path):
    """En DGH hay líneas con el trámite abierto y la respuesta en blanco."""
    informe = _informe_siifa(tmp_path, [_glosa(1, "HUS454747", "TA0801", valor=22200)])
    tramites = _tramites_dgh(tmp_path, [_tramite("HUS0000454747", "TA0801", 22200, resp="")])

    _, previas = _cruzar(informe, tramites)

    assert previas == {}


def test_la_hoja_de_trabajo_sale_con_la_respuesta_puesta_y_su_origen(tmp_path):
    """El camino completo: cruzar, escribir la hoja, y que quede lista."""
    informe = _informe_siifa(
        tmp_path,
        [
            _glosa(1, "HUS454747", "TA0801", valor=22200),
            _glosa(2, "HUS454747", "TA0801", valor=22200),
            _glosa(3, "HUS454747", "CO2301", valor=4700),
        ],
    )
    tramites = _tramites_dgh(
        tmp_path, [_tramite("HUS0000454747", "TA0801", 22200, obs="NO ACEPTA POR APOYO DX.")]
    )

    grupos, previas = _cruzar(informe, tramites)
    salida = tmp_path / "para_revisar.xlsx"
    prep.escribir_plantilla(grupos, salida, ya_respondidas=previas)

    ws = load_workbook(salida).active
    hdr = [c.value for c in ws[1]]
    filas = [dict(zip(hdr, [c.value for c in f])) for f in ws.iter_rows(min_row=2)]
    por_codigo = {f["CODIGO_GLOSA"]: f for f in filas}

    assert por_codigo["TA0801"]["CODIGO_RESPUESTA"] == "RE9901"
    assert por_codigo["TA0801"]["OBSERVACION_RESPUESTA"] == "NO ACEPTA POR APOYO DX."
    assert por_codigo["TA0801"]["ORIGEN_RESPUESTA"] == "EXACTO"
    assert por_codigo["TA0801"]["LINEAS_QUE_CUBRE"] == 2
    assert por_codigo["CO2301"]["ORIGEN_RESPUESTA"] == "ESCRIBIR"
    assert por_codigo["CO2301"]["CODIGO_RESPUESTA"] is None


def test_lo_cruzado_llega_completo_hasta_el_excel_del_bot(tmp_path):
    """Lo que importa de verdad: que la respuesta llegue a cada id de SIIFA."""
    informe = _informe_siifa(
        tmp_path,
        [
            _glosa(101, "HUS454747", "TA0801", valor=22200),
            _glosa(102, "HUS454747", "TA0801", valor=22200),
        ],
    )
    tramites = _tramites_dgh(tmp_path, [_tramite("HUS0000454747", "TA0801", 22200)])

    grupos, previas = _cruzar(informe, tramites)
    hoja = tmp_path / "para_revisar.xlsx"
    prep.escribir_plantilla(grupos, hoja, ya_respondidas=previas)

    filas, sin_responder = prep.expandir(grupos, prep.leer_plantilla_llena(hoja))

    assert not sin_responder
    assert {f["ID_SEGUIMIENTO_FACTURA_GLOSA"] for f in filas} == {101, 102}
    assert all(f["CODIGO_RESPUESTA"] == "RE9901" for f in filas)


def test_avisa_si_el_archivo_de_tramites_no_es_el_de_dgh(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "BD TRAMITE"
    ws.append(["COSA", "OTRA COSA"])
    otro = tmp_path / "otro.xlsx"
    wb.save(otro)

    with pytest.raises(SystemExit) as e:
        cruce.leer_tramites(otro, {"454747"})

    assert "factura" in str(e.value)
