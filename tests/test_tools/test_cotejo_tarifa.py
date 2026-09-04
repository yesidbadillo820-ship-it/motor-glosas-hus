"""Pruebas del COTEJO DE COBRO: ¿de verdad se está cobrando de más?

Lo que estas pruebas protegen es plata del hospital en las dos direcciones:
que el bot no sugiera aceptar una glosa que no procede, y que no deje pasar
un sobrecobro real.
"""

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

RAIZ = Path(__file__).resolve().parents[2] / "tools" / "glosas_dispensario"
sys.path.insert(0, str(RAIZ))

from cotejo_tarifa import (  # noqa: E402
    A_TARIFA,
    POR_DEBAJO,
    POR_VIGENCIA,
    SIN_COTEJO,
    SOBRECOBRO,
    cotejar,
    elegir_valor_facturado,
    factores_repetidos,
)


def _tarifa(precio, descripcion="CONSULTA", fuente="anexo 6.2 · SERVICIOS"):
    return dict(precio=precio, descripcion=descripcion, fuente=fuente)


# ── Qué número de la línea del PDF se compara ───────────────────────────────
def test_elige_el_valor_que_coincide_con_la_tarifa():
    # cantidad 1, unitario y total iguales: cualquiera sirve, pero se prefiere
    # el que coincide con lo pactado
    assert elegir_valor_facturado([180000, 180000], 180000) == (180000, "")
    # unitario + total de 3 unidades: el unitario es el que se cotej
    assert elegir_valor_facturado([106762, 320286], 106762) == (106762, "")


def test_sin_tarifa_deduce_el_unitario_por_la_cantidad():
    # 35.136 x 4 = 140.544: el menor es el unitario aunque no haya tarifa
    assert elegir_valor_facturado([35136, 140544], None) == (35136, "")
    # un solo importe: ese es
    assert elegir_valor_facturado([113500], None) == (113500, "")


def test_linea_ambigua_no_arriesga_una_cifra():
    valor, motivo = elegir_valor_facturado([113500, 192600], None)
    assert valor is None and "varios valores" in motivo
    assert elegir_valor_facturado([], 1000)[0] is None


# ── Los veredictos ──────────────────────────────────────────────────────────
def test_sin_pdf_o_sin_tarifa_no_hay_sugerencia():
    c = cotejar(None, _tarifa(180000), 113500, "890275H")
    assert c["veredicto"] == SIN_COTEJO and c["aceptar"] == 0
    assert "SE ACEPTA" not in c["respuesta"].upper()

    c2 = cotejar(192600, None, 113500, "890275H")
    assert c2["veredicto"] == SIN_COTEJO and c2["aceptar"] == 0
    assert "no aparece en el anexo tarifario" in c2["respuesta"]


def test_cobro_a_tarifa_es_glosa_infundada():
    c = cotejar(180000, _tarifa(180000), 113500, "890275H")
    assert c["veredicto"] == A_TARIFA and c["aceptar"] == 0
    assert c["respuesta"].startswith("NO ACEPTAR")
    # el redondeo de centavos del anexo no es un cobro de más
    assert cotejar(17557, _tarifa(17556), 5000, "902210AMB")["veredicto"] == A_TARIFA


def test_cobro_por_debajo_de_lo_pactado():
    c = cotejar(100000, _tarifa(180000), 50000, "890275H")
    assert c["veredicto"] == POR_DEBAJO and c["aceptar"] == 0
    assert "INFERIOR" in c["respuesta"]


def test_sobrecobro_aislado_sugiere_aceptar_la_diferencia():
    # la EPS objeta 113.500 pero el cobro de más real es de 12.600
    c = cotejar(192600, _tarifa(180000), 113500, "890275H")
    assert c["veredicto"] == SOBRECOBRO
    assert c["diferencia"] == 12600 and c["aceptar"] == 12600
    assert "SE ACEPTA PARCIALMENTE LA GLOSA POR $12.600" in c["respuesta"]
    assert "LEVANTAMIENTO DE LOS $100.900" in c["respuesta"]


def test_sobrecobro_mayor_que_lo_objetado_acepta_solo_lo_objetado():
    c = cotejar(300000, _tarifa(180000), 50000, "890275H")
    assert c["veredicto"] == SOBRECOBRO
    assert c["diferencia"] == 120000 and c["aceptar"] == 50000  # nunca más de lo glosado
    assert "SE ACEPTA LA GLOSA POR $50.000 POR MAYOR VALOR COBRADO" in c["respuesta"]


def test_diferencia_que_se_repite_en_el_lote_es_vigencia_y_no_se_acepta():
    # el lote 04-09-2026: 23 facturas facturadas al 7% sobre el anexo
    lote = [(192600, 180000), (247765, 231556), (114236, 106762), (148760, 139028)]
    factores = factores_repetidos(lote)
    assert 1.07 in factores

    c = cotejar(192600, _tarifa(180000), 113500, "890275H", factores)
    assert c["veredicto"] == POR_VIGENCIA and c["aceptar"] == 0
    assert "NO ACEPTAR AUTOMÁTICAMENTE" in c["respuesta"]
    assert "parágrafos 3 y 4 del contrato" in c["respuesta"]
    # sin el contexto del lote, el mismo caso sí sugeriría aceptar: por eso el
    # veredicto se dicta después de mirar todo el lote
    assert cotejar(192600, _tarifa(180000), 113500, "890275H")["veredicto"] == SOBRECOBRO


def test_un_caso_suelto_no_se_toma_como_vigencia():
    lote = [(192600, 180000), (247765, 231556)]  # solo dos: no alcanza
    assert factores_repetidos(lote) == set()
    lote_mixto = [(192600, 180000), (247765, 231556), (114236, 106762), (500000, 100000)]
    factores = factores_repetidos(lote_mixto)
    assert 1.07 in factores and 5.0 not in factores  # el caso raro sigue siendo sobrecobro
    assert cotejar(500000, _tarifa(100000), 400000, "X", factores)["veredicto"] == SOBRECOBRO


def test_los_porcentajes_se_reportan_para_que_el_auditor_verifique():
    c = cotejar(192600, _tarifa(180000), 113500, "890275H")
    assert c["porcentaje"] == 7.0
    assert c["valor_facturado"] == 192600 and c["tarifa"] == 180000


# ── La redacción por escenario ──────────────────────────────────────────────
def _pdf(archivo="FEV_900006037_HUS542497.PDF", paciente="RODRIGUEZ GOMEZ MARIA"):
    return dict(archivo=archivo, metodo="pdfplumber", ok=True, paciente=paciente, total=192600)


def test_escenario_1_cobro_a_tarifa_declara_causal_infundada():
    from bot_lote_dispensario import parrafo_por_escenario

    cot = cotejar(180000, _tarifa(180000), 113500, "890275H")
    p = parrafo_por_escenario(
        cot,
        _pdf(),
        ("890275H CONSULTA 1 180.000 180.000", 180000, [180000]),
        "890275H",
        "CONSULTA DE PRIMERA VEZ POR NEUROLOGIA",
    )
    assert p.startswith("AL REVISAR EL SOPORTE FEV_900006037_HUS542497.PDF")
    assert "SE EVIDENCIA LA ATENCION DEL USUARIO RODRIGUEZ GOMEZ MARIA" in p
    assert "EL VALOR FACTURADO DE $180.000" in p
    assert "TARIFA PACTADA EXACTA EN EL CONTRATO 440-DIGSA-DMBUG-2025" in p
    assert "CAUSAL DE GLOSA ES INFUNDADA" in p
    # el servicio se nombra como lo conoce la EPS, sin la fila cruda del PDF
    assert "890275H CONSULTA DE PRIMERA VEZ POR NEUROLOGIA" in p
    assert "1 180.000 180.000" not in p


def test_escenario_2_vigencia_2026_defiende_el_mayor_valor():
    from bot_lote_dispensario import parrafo_por_escenario

    factores = factores_repetidos([(192600, 180000), (247765, 231556), (114236, 106762)])
    cot = cotejar(192600, _tarifa(180000), 113500, "890275H", factores)
    p = parrafo_por_escenario(cot, _pdf(), None, "890275H", "CONSULTA NEUROLOGIA")
    assert "EL VALOR COBRADO DE $192.600" in p
    assert "ACTUALIZACION DE TARIFAS DE LA VIGENCIA 2026" in p
    assert "PARAGRAFOS 3 Y 4 DEL CONTRATO 440-DIGSA-DMBUG-2025" in p
    assert "EL COBRO ES CONTRACTUALMENTE VALIDO" in p
    # defiende, no acepta
    assert "SE ACEPTA" not in p


def test_escenario_3_sobrecobro_real_redacta_la_aceptacion_parcial():
    from bot_lote_dispensario import parrafo_por_escenario

    cot = cotejar(192600, _tarifa(180000), 113500, "890275H")  # caso aislado
    p = parrafo_por_escenario(cot, _pdf(), None, "890275H")
    assert p.startswith("VALIDADO EL SOPORTE FEV_900006037_HUS542497.PDF")
    assert "LA TARIFA PACTADA PARA EL CODIGO 890275H ES DE $180.000 Y SE FACTURO $192.600" in p
    assert "SE ACEPTA LA GLOSA POR EL MAYOR VALOR COBRADO DE $12.600" in p
    assert "LEVANTAMIENTO DE LOS $100.900 RESTANTES" in p


def test_fuera_de_los_tres_escenarios_no_se_redacta_nada():
    """Sin cotejo o cobrado por debajo: manda la redacción prudente."""
    from bot_lote_dispensario import parrafo_por_escenario

    assert parrafo_por_escenario(cotejar(None, _tarifa(180000), 113500, "X"), _pdf(), None) is None
    assert parrafo_por_escenario(cotejar(100000, _tarifa(180000), 50000, "X"), _pdf(), None) is None


def test_la_redaccion_no_inventa_el_paciente():
    from bot_lote_dispensario import parrafo_por_escenario

    cot = cotejar(180000, _tarifa(180000), 113500, "890275H")
    p = parrafo_por_escenario(cot, _pdf(paciente=None), None, "890275H")
    assert "USUARIO" not in p and "$180.000" in p
    # y sin PDF tampoco se cita un archivo que no existe
    sin_pdf = parrafo_por_escenario(cot, None, None, "890275H")
    assert "SOPORTE," in sin_pdf and ".PDF" not in sin_pdf


def test_el_texto_usa_las_mismas_cifras_que_las_columnas():
    """El texto y el Excel no pueden decir cifras distintas."""
    from bot_lote_dispensario import parrafo_por_escenario

    cot = cotejar(300000, _tarifa(180000), 50000, "890275H")
    p = parrafo_por_escenario(cot, _pdf(), None, "890275H")
    from _dinero import a_texto

    assert a_texto(cot["aceptar"]) in p and a_texto(cot["tarifa"]) in p
    assert a_texto(cot["valor_facturado"]) in p


# ── Las columnas en el Excel ────────────────────────────────────────────────
def _excel_respuestas(ruta, filas=1):
    wb = Workbook()
    ws = wb.active
    ws.title = "Respuestas Glosa"
    ws.append(
        [
            "Factura",
            "# Objeción",
            "Cód.",
            "Servicio",
            "Valor Objetado",
            "Valor Aceptado",
            "Cod Respuesta",
            "Detalle Respuesta",
        ]
    )
    for i in range(1, filas + 1):
        ws.append(
            ["HUS0000542497", i, "TA0201", "890275H - CONSULTA", 113500, 0, "RE9901", "TEXTO"]
        )
    wb.save(ruta)


def test_las_columnas_de_cotejo_no_tocan_el_valor_aceptado_del_cargue(tmp_path):
    from bot_lote_dispensario import agregar_cotejo_excel

    ruta = tmp_path / "resp.xlsx"
    _excel_respuestas(ruta)
    cot = cotejar(192600, _tarifa(180000), 113500, "890275H")
    cot["fuente"] = "FEV_900006037_HUS542497.pdf"

    assert agregar_cotejo_excel(ruta, {("HUS0000542497", 1): cot}) == 1
    ws = load_workbook(ruta)["Respuestas Glosa"]
    cabeceras = [c.value for c in ws[1]]
    fila = [c.value for c in ws[2]]
    # el cargue queda intacto: aceptado 0 y el detalle sin tocar
    assert fila[5] == 0 and fila[7] == "TEXTO"
    # y al lado aparece el cotejo
    assert cabeceras[8:] == [
        "VALOR FACTURADO (PDF)",
        "TARIFA PACTADA (440)",
        "DIFERENCIA",
        "¿SOBRECOBRO?",
        "VALOR SUGERIDO A ACEPTAR",
        "RESPUESTA SUGERIDA",
        "FUENTE DEL COTEJO",
    ]
    assert fila[8] == 192600 and fila[9] == 180000 and fila[10] == 12600
    assert fila[11].startswith(SOBRECOBRO) and "+7.0%" in fila[11]
    assert fila[12] == 12600 and "SE ACEPTA PARCIALMENTE" in fila[13]
    assert "FEV_900006037_HUS542497.pdf" in fila[14]


def test_repetir_el_cotejo_no_duplica_columnas(tmp_path):
    from bot_lote_dispensario import agregar_cotejo_excel

    ruta = tmp_path / "resp.xlsx"
    _excel_respuestas(ruta)
    cot = cotejar(180000, _tarifa(180000), 113500, "890275H")
    agregar_cotejo_excel(ruta, {("HUS0000542497", 1): cot})
    agregar_cotejo_excel(ruta, {("HUS0000542497", 1): cot})
    ws = load_workbook(ruta)["Respuestas Glosa"]
    assert [c.value for c in ws[1]].count("RESPUESTA SUGERIDA") == 1
    assert ws.max_column == 15


def test_la_hoja_de_trabajo_solo_trae_lo_que_hay_que_decidir(tmp_path):
    from bot_lote_dispensario import agregar_cotejo_excel

    ruta = tmp_path / "resp.xlsx"
    _excel_respuestas(ruta, filas=3)
    factores = factores_repetidos([(114236, 106762)] * 3)
    cotejos = {
        ("HUS0000542497", 1): cotejar(192600, _tarifa(180000), 113500, "890275H"),  # sobrecobro
        ("HUS0000542497", 2): cotejar(180000, _tarifa(180000), 113500, "890275H"),  # a tarifa
        ("HUS0000542497", 3): cotejar(
            114236, _tarifa(106762), 35136, "890226H", factores
        ),  # vigencia
    }
    agregar_cotejo_excel(ruta, cotejos)
    wb = load_workbook(ruta)
    hoja = wb["COTEJO DE COBRO"]
    numeros = [f[1] for f in hoja.iter_rows(min_row=2, values_only=True)]
    assert numeros == [1, 3]  # la línea cobrada a tarifa no da trabajo al auditor
    assert [f[5] for f in hoja.iter_rows(min_row=2, values_only=True)] == [SOBRECOBRO, POR_VIGENCIA]


def test_sin_nada_que_decidir_no_se_crea_la_hoja(tmp_path):
    from bot_lote_dispensario import agregar_cotejo_excel

    ruta = tmp_path / "resp.xlsx"
    _excel_respuestas(ruta)
    agregar_cotejo_excel(
        ruta, {("HUS0000542497", 1): cotejar(180000, _tarifa(180000), 113500, "890275H")}
    )
    assert "COTEJO DE COBRO" not in load_workbook(ruta).sheetnames


def test_el_robot_del_portal_sigue_leyendo_el_excel_con_las_columnas_nuevas(tmp_path):
    """El cargue lee sus columnas por nombre; agregar columnas no lo rompe."""
    sys.path.insert(0, str(RAIZ.parent))
    from bot_lote_dispensario import agregar_cotejo_excel
    from responder_glosas_simed import leer_excel_respuestas

    ruta = tmp_path / "resp.xlsx"
    _excel_respuestas(ruta, filas=2)
    cot = cotejar(192600, _tarifa(180000), 113500, "890275H")
    agregar_cotejo_excel(ruta, {("HUS0000542497", 1): cot, ("HUS0000542497", 2): cot})

    datos = leer_excel_respuestas(ruta)
    objeciones = list(datos.values())[0]
    assert [o["num"] for o in objeciones] == [1, 2]
    assert all(o["aceptado"] == 0 for o in objeciones)  # el bot no acepta nada solo
    assert all(o["detalle"] == "TEXTO" for o in objeciones)
