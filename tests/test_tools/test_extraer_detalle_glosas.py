"""Tests del extractor de detalle de glosas (tools/extraer_detalle_glosas.py).

Fixtures sintéticos que replican la estructura REAL de los adjuntos observados
(AUDITOOL csv, FAMISANAR txt, SURA txt, Sanitas/Policía/FACTRAMED xlsx, AXA pdf
tabular) sin datos de pacientes.
"""

from __future__ import annotations

import io
import json
import zipfile
from datetime import date
from pathlib import Path

import pytest

from tools import extraer_detalle_glosas as ext

GESTORES = ext.cargar_gestores(None)


# ─── Conversores ─────────────────────────────────────────────────────────────


@pytest.mark.parametrize(
    "entrada,esperado",
    [
        ("$9.576.693,00", 9576693.0),
        ("2,800", 2800.0),
        ("117300", 117300.0),
        (117300, 117300.0),
        ("1.046.700", 1046700.0),
        ("", None),
        (None, None),
    ],
)
def test_a_numero(entrada, esperado):
    assert ext.a_numero(entrada) == esperado


@pytest.mark.parametrize(
    "entrada,esperado",
    [
        ("18/06/26", date(2026, 6, 18)),
        ("01/07/2026 00:00", date(2026, 7, 1)),
        ("2026/06/25", date(2026, 6, 25)),
        ("26-06-2026", date(2026, 6, 26)),
        ("1900/01/01", None),  # 'sin fecha' de SURA
        ("", None),
    ],
)
def test_a_fecha(entrada, esperado):
    assert ext.a_fecha(entrada) == esperado


def test_sumar_habiles_con_festivos():
    # 7 hábiles desde 30/06/2026: sáb 4 y dom 5 de julio no cuentan
    assert ext.sumar_habiles(date(2026, 6, 30), 7) == date(2026, 7, 9)
    # 15 hábiles desde 26/06/2026: 29/06 y 20/07 son festivos en Colombia
    assert ext.sumar_habiles(date(2026, 6, 26), 15) == date(2026, 7, 21)


# ─── Parsers ─────────────────────────────────────────────────────────────────

CSV_AUDITOOL = (
    "\n"
    "IPS:EMPRESA SOCIAL DEL ESTADO HOSPITAL UNIVERSITARIO DE SANTANDER\n"
    "NIT: 900006037\n"
    "FEC. PRESENTACION FACTURA;PRESTADOR;PREFIJO;FACTURA;RADICADO FACTURA;"
    "FECHA EMISION FACTURA;VALOR DE LA FACTURA;FECHA GLOSA INICIAL;ID GLOSA;"
    "VALOR GLOSA INICIAL;CODIGO GLOSA INICIAL;FECHA DE NOTIFICACION AL PRESTADOR;\n"
    "18/06/26;ESE HUS;HUS;521123;1051374;02/06/26;117300;30/06/26;1;38200;"
    "tA02 01 TARIFAS;01/07/2026 00:00\n"
    "18/06/26;ESE HUS;HUS;521123;1051374;02/06/26;117300;30/06/26;2;10000;"
    "tA02 01 TARIFAS;01/07/2026 00:00\n"
).encode("latin-1")


def test_parser_auditool_por_servicio_y_agrupacion():
    filas = ext.parser_auditool_csv(CSV_AUDITOOL, "AUDITORIA_GLOSA_X.csv")
    assert len(filas) == 2
    assert filas[0]["categoria"] == "INICIAL"
    assert filas[0]["factura"] == "HUS521123"
    assert filas[0]["valor"] == 38200.0
    assert filas[0]["fecha_notificacion"] == date(2026, 7, 1)
    assert filas[0]["radicado"] == "1051374"
    for fila in filas:
        fila.update(empresa="DISPENSARIO", reglas={})
    grupos = ext.agrupar_por_factura(filas)
    assert len(grupos) == 1
    assert grupos[0]["valor"] == 48200.0  # 38200 + 10000


TXT_FAMISANAR = (
    "NRO_REGISTRO|NRO_RADICACION|NIT_IPS|FECHA_ENVIO|NRO_FACTURA|CONSECUTIVO_USUARIO|"
    "CONSECUTIVO_SERVICIO|TIPO_SERVICIO|CODIGO_DEVOLUCION|VALOR DEVOLUCION|OBSERVACION\n"
    "88751201_GL|88751201|900006037|01/07/2026 11:54|HUS518355|1|7|AT|CO0601|2400|obs\n"
    "88751202_GL|88751202|900006037|01/07/2026 11:54|HUS518999|1|1|AT|DE1601|9900|dev\n"
).encode("latin-1")


def test_parser_famisanar_separa_glosa_de_devolucion():
    filas = ext.parser_famisanar_txt(TXT_FAMISANAR, "DEVYGLOSAS0514737_900006037.txt")
    assert len(filas) == 2
    assert filas[0]["categoria"] is None  # CO* = glosa: categoría de la carpeta
    assert filas[1]["categoria"] == "DEVOLUCIONES"  # DE* = devolución siempre
    assert filas[0]["valor"] == 2400.0
    assert filas[0]["fecha_notificacion"] == date(2026, 7, 1)


TXT_SURA = (
    "NITEAPB;NITIPS;PREFIJOFACTURA;NUMEROFACTURA;FECHAFACTURA;OFICINARADICACION;"
    "NUMERORADICADO;FECHARADICACION;VALORFACTURA;NUMEROGLOSA;FECHAGLOSA;CAUSAGENERAL;"
    "CAUSAESPECIFICA;VALORGLOSA;OBSERVACIONES;OBSERVACIONESDETALLES\n"
    "800088702;900006037;HUS;282713;2024/05/24;1;139195615;2026/06/25;127000;11025925;"
    "1900/01/01;2;8;61700; ;\n"
).encode("latin-1")


def test_parser_sura():
    filas = ext.parser_sura_txt(TXT_SURA, "800088702_MASIVO.TXT")
    assert len(filas) == 1
    assert filas[0]["factura"] == "HUS282713"
    assert filas[0]["valor"] == 61700.0
    assert filas[0]["fecha_notificacion"] is None  # 1900/01/01
    assert filas[0]["fecha_radicacion"] == date(2026, 6, 25)


def _xlsx_bytes(encabezado: list, filas: list[list]) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(encabezado)
    for fila in filas:
        ws.append(fila)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_parser_objeciones_policia():
    contenido = _xlsx_bytes(
        ["CDCONSEC", "CDFECDOC", "CRNCXC", "CROFECOBJ", "CRNCONOBJ", "CROVALOBJ"],
        [[1, date(2026, 7, 1), "HUS0000503629", date(2026, 7, 1), "DE1601", 114236]],
    )
    filas = ext.parser_objeciones_xlsx(contenido, "OBJECIONES OK.xlsx")
    assert len(filas) == 1
    assert filas[0]["categoria"] == "DEVOLUCIONES"
    assert filas[0]["factura"] == "HUS0000503629"
    assert filas[0]["valor"] == 114236.0


def test_parser_sanitas():
    contenido = _xlsx_bytes(
        [
            "NUMERO DE FACTURA",
            "FECHA DE FACTURA",
            "ID GLOSA",
            "CODIGO CONCEPTO GENERAL",
            "VALOR GLOSADO GLOSA",
            "VALOR REAL GLOSA",
            "NUMERO DE RADICACION",
            "FECHA DE RADICACION",
        ],
        [["HUS519487", "28/05/26", 625624670, "TA", 131200, 131200, 1800577798, "04/06/2026"]],
    )
    filas = ext.parser_sanitas_xlsx(contenido, "900006037.xlsx")
    assert len(filas) == 1
    assert filas[0]["valor"] == 131200.0
    assert filas[0]["radicado"] == "1800577798"


def test_parser_factramed_filtra_por_ventana():
    contenido = _xlsx_bytes(
        [
            "Fecha_Hora_Radicacion",
            "Numero_Radicacion",
            "Numero_Factura",
            "Valor_Total_Factura",
            "Numero_Glosa_Dev",
            "Fecha_Hora_Glosa",
            "Codigo_Glosa_Dev",
            "Valor_Glosa_Dev",
        ],
        [
            [
                "2026-06-11 08:33",
                26771660,
                "HUS518073",
                37103002,
                320738557,
                date(2026, 6, 30),
                "SO3602",
                4700,
            ],
            [
                "2026-05-01 08:33",
                26770000,
                "HUS500000",
                1000,
                320000000,
                date(2026, 5, 2),
                "DE4401",
                999,
            ],
        ],
    )
    filas = ext.parser_factramed_xlsx(
        contenido, "FMED.xlsx", desde=date(2026, 6, 25), hasta=date(2026, 7, 2)
    )
    assert len(filas) == 1  # la de mayo queda fuera de la ventana
    assert filas[0]["factura"] == "HUS518073"
    assert filas[0]["categoria"] is None  # SO* = glosa
    assert filas[0]["valor"] == 4700.0
    assert filas[0]["valor_factura"] == 37103002.0


def test_parser_axa_texto_tabular():
    texto = (
        "PERIODO 2026-6-26\n"
        "No.reclamo FechaLiquidacion No.Póliza Id.accidentado No.factura Valorfactura "
        "CodigoProce Observaciones Descripción Cobrado Liquidado Glosado\n"
        "22946610 26-06-2026 3272899700 1005345731 HUS489150 71500 39143 "
        "SEOBJETAPORTOPEMAXIMO 71500 38023 33477\n"
    )
    filas = ext._parser_axa(texto)
    assert len(filas) == 1
    assert filas[0]["factura"] == "HUS489150"
    assert filas[0]["valor"] == 33477.0
    assert filas[0]["fecha_notificacion"] == date(2026, 6, 26)


# ─── Carpetas, entidades y gestores ──────────────────────────────────────────


@pytest.mark.parametrize(
    "carpeta,categoria",
    [
        ("DEVOLUCIONES OK", "DEVOLUCIONES"),
        ("INICIAL", "INICIAL"),
        ("RATIFICADA OK", "RATIFICADA"),
        ("RATIFICADAS", "RATIFICADA"),
        ("00-CONTROL AUTOMATICO", None),
    ],
)
def test_categoria_de_carpeta(carpeta, categoria):
    assert ext._categoria_de_carpeta(carpeta) == categoria


def test_info_entidad_con_marcas_manuales():
    empresa, reglas = ext.info_entidad("DISPENSARIO SOFIA OK", GESTORES)
    assert empresa == "DISPENSARIO MEDICO BUCARAMANGA"
    assert reglas["DEVOLUCIONES"] == "JOHANNA MORENO"
    empresa, reglas = ext.info_entidad("PEPITO SAS OK", GESTORES)
    assert empresa == "PEPITO SAS"
    assert reglas == {}


def test_responsable_por_categoria():
    _, reglas = ext.info_entidad("COOSALUD OK", GESTORES)
    assert ext.responsable_para(reglas, "INICIAL", GESTORES) == "DIANEYDA QUINTERO"
    assert ext.responsable_para(reglas, "RATIFICADA", GESTORES) == "JORGE GUARIN"
    assert ext.responsable_para(reglas, "DEVOLUCIONES", GESTORES) == "ALBA PEREZ"
    assert ext.responsable_para({}, "INICIAL", GESTORES) == "POR ASIGNAR"


def test_fecha_de_carpeta():
    assert ext.fecha_de_carpeta(Path("Z:/x/2026/07 JULIO/02")) == date(2026, 7, 2)
    assert ext.fecha_de_carpeta(Path("Z:/x/2026/07.JULIO/01 OK SOLO NUEVA")) == date(2026, 7, 1)


def test_gestores_ejemplo_sincronizado():
    ruta = Path(__file__).resolve().parents[2] / "tools" / "gestores_glosas.ejemplo.json"
    assert json.loads(ruta.read_text(encoding="utf-8")) == ext.GESTORES_DEFECTO


# ─── Flujo completo sobre una carpeta de día sintética ───────────────────────


def test_flujo_completo_genera_excel(tmp_path):
    dia = tmp_path / "2026" / "07 JULIO" / "02"
    carpeta_fami = dia / "INICIAL" / "FAMISANAR OK"
    carpeta_fami.mkdir(parents=True)
    # el TXT viaja dentro de un zip, como llega en el correo real
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w") as zf:
        zf.writestr("DEVYGLOSAS0514737_900006037_20260701.txt", TXT_FAMISANAR)
    (carpeta_fami / "FAMISANAR 7.23 OK.zip").write_bytes(buffer.getvalue())
    carpeta_disp = dia / "DEVOLUCIONES OK" / "DISPENSARIO OK"
    carpeta_disp.mkdir(parents=True)
    (carpeta_disp / "AUDITORIA_GLOSA_HUS.csv").write_bytes(CSV_AUDITOOL)

    filas, sin_parser = ext.recorrer_dia(dia, GESTORES, ventana_factramed=7)
    assert sin_parser == []
    categorias = {f["categoria"] for f in filas}
    assert categorias == {"INICIAL", "DEVOLUCIONES"}
    # el DE1601 de famisanar salta a DEVOLUCIONES aunque estaba en carpeta INICIAL
    dev_fami = [f for f in filas if f["factura"] == "HUS518999"]
    assert dev_fami[0]["categoria"] == "DEVOLUCIONES"

    salida = tmp_path / "ENTREGA.xlsx"
    ext.escribir_entrega(filas, sin_parser, salida, date(2026, 7, 6), GESTORES)
    from openpyxl import load_workbook

    wb = load_workbook(salida)
    assert wb.sheetnames == ["INICIAL", "RATIFICADA", "DEVOLUCIONES", "RESUMEN"]
    hoja = wb["INICIAL"]
    assert [c.value for c in hoja[1]] == ext.ENCABEZADOS["INICIAL"]
    por_factura = {f[6]: f for f in hoja.iter_rows(min_row=2, values_only=True)}
    # el CSV de AUDITOOL dice 'GLOSA INICIAL': su contenido manda sobre la
    # carpeta DEVOLUCIONES donde estaba guardado, y agrupa 38200+10000
    assert por_factura["HUS521123"][8] == 48200.0
    assert por_factura["HUS518355"][8] == 2400.0  # famisanar CO0601


# ─── Parsers de la estructura real de julio (COOSALUD, SAVIA, POLICÍA, MUTUAL) ─


def test_parser_coosalud_glosas_xlsx():
    contenido = _xlsx_bytes(
        [
            "id_glosa",
            "id_cuenta",
            "numero_factura",
            "codigo_glosa",
            "tipo_glosa",
            "valor_uni_glosa",
            "valor_total_glosa",
            "fecha_glosa",
        ],
        [
            [
                51111255,
                6799865,
                "HUS507721",
                "TA2901",
                "TARIFAS",
                125330,
                125330,
                date(2026, 5, 11),
            ],
            [51111256, 6799865, "HUS507721", "TA2901", "TARIFAS", 10580, 10580, date(2026, 5, 11)],
        ],
    )
    filas = ext.parser_coosalud_xlsx(contenido, "GLOSAS HUS507721.xlsx")
    assert len(filas) == 2
    assert filas[0]["factura"] == "HUS507721"
    assert filas[0]["valor"] == 125330.0
    assert filas[0]["no_glosa"] == "TA2901"
    assert filas[0]["fecha_notificacion"] == date(2026, 5, 11)


def test_parser_savia_xlsx():
    contenido = _xlsx_bytes(
        [
            "Numero_Radicado",
            "NIT_IPS",
            "Numero_factura",
            "Valor_Factura",
            "Fecha_Radicacion",
            "Valor_Glosa",
            "Motivo_Glosa_Valor_A",
        ],
        [[4468235, 900006037, "HUS464792", 9002220, "2026-06-11", 150000, "TA29"]],
    )
    filas = ext.parser_savia_xlsx(contenido, "SAVIA SALUD 4.07 OK.xlsx")
    assert len(filas) == 1
    assert filas[0]["factura"] == "HUS464792"
    assert filas[0]["valor"] == 150000.0
    assert filas[0]["radicado"] == "4468235"


def test_parser_policia_rad_encabezado_en_fila_intermedia():
    filas_xlsx = [
        ["ADMINISTRACIÓN"],
        [],
        ["No. RADICACION", 2190, "DEL", "2026-07-03"],
        [],
        [
            "N.",
            "No.FACTURA",
            "FECHA DE FACTURA",
            "NOMBRE PACIENTE",
            "VR. FACTURADO",
            "FECHA GLOSA",
            "VR. GLOSA ",
            "VALOR ACEPTADO IPS",
        ],
        [1, "HUS529263", date(2026, 6, 23), "PACIENTE X", 240101500, date(2026, 7, 3), 6165700, 0],
        [None, None, None, None, 240101500, None, 6165700, 0],
    ]
    contenido = _xlsx_bytes(filas_xlsx[0], filas_xlsx[1:])
    filas = ext.parser_policia_rad_xlsx(contenido, "GLOSA A RAD 2190.xlsx")
    assert len(filas) == 1  # la fila de totales sin factura no cuenta
    assert filas[0]["factura"] == "HUS529263"
    assert filas[0]["valor"] == 6165700.0
    assert filas[0]["radicado"] == "2190"
    assert filas[0]["fecha_notificacion"] == date(2026, 7, 3)


def test_xlsx_desconocido_cae_a_fila_minima_por_nombre():
    contenido = _xlsx_bytes(["Ítem"], [["1"], ["2"]])  # formulario MUTUAL de 1 columna
    filas = ext._parser_xlsx_auto(
        contenido, "Validación de subsanación Factura N° HUS492542 OK.xlsx"
    )
    assert len(filas) == 1
    assert filas[0]["factura"] == "HUS492542"
    assert "completar a mano" in filas[0]["observacion"]


def test_ligero_se_ignora():
    """Las copias *_LIGERO (hechas a mano para compartir) duplicarían valores."""
    ventana = (None, None)
    filas, tenia = ext.extraer_de_archivo("FAMISANAR 7.23 OK_LIGERO.zip", b"x", ventana)
    assert (filas, tenia) == ([], True)  # ignorado sin reportarse como 'sin parser'
    filas, _ = ext.extraer_de_archivo("DEVYGLOSAS0514737_900006037.txt", TXT_FAMISANAR, ventana)
    assert len(filas) == 2  # el original sí se procesa
