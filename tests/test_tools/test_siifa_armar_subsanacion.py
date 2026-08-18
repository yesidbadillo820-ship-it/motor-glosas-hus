"""Pruebas del armado de la subsanación (etapa 4 del trámite).

Dos cosas se cuidan acá:

1. Que **sólo** entren las glosas reiteradas y sin subsanar. Meter una ya
   subsanada la pisaría; meter una levantada reabriría algo ganado.
2. Que una **devolución** reiterada NUNCA salga lista para cargar. Su puerta
   de subsanación no está confirmada, y la de glosas escribiría sobre otro
   registro dejando un OK falso en el reporte —el mismo daño que costó
   descubrir en agosto con las respuestas—.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

TOOLS = Path(__file__).resolve().parents[2] / "tools"
sys.path.insert(0, str(TOOLS))

import siifa_armar_subsanacion as sub  # noqa: E402


def _fila(**kw):
    base = {
        "id_seguimiento_factura_glosa": 15110544,
        "id_seguimiento_factura_devolucion": "",
        "numero_factura": "HUS465198",
        "tipo_seguimiento": "GLOSA",
        "razon_social_eps": "SANITAS",
        "valor_glosa": 1_396_804,
        "codigo_glosa": "TA0801",
        "descripcion_glosa": "Los cargos por apoyo diagnóstico no son pertinentes.",
        "observacion_glosa": "No se evidencia orden médica.",
        "tiene_respuesta": "SI",
        "codigo_respuesta": "RE9901",
        "fecha_respuesta": "2026-08-05T14:07:00",
        "fecha_formulacion": "2026-01-29",
        "codigo_reiteracion": "GRE003",
        "descripcion_reiteracion": "Glosa Reiterada",
        "codigo_reiteracion_respuesta": "",
    }
    base.update(kw)
    return base


def test_solo_entra_lo_reiterado_y_sin_subsanar():
    filas = [
        _fila(),  # reiterada, sin subsanar -> entra
        _fila(codigo_reiteracion="", descripcion_reiteracion=""),  # esperando a la EPS
        _fila(descripcion_reiteracion="Glosa totalmente levantada"),  # ganada
        _fila(codigo_reiteracion_respuesta="RE9901"),  # ya subsanada
        _fila(tiene_respuesta="NO"),  # ni siquiera respondida
    ]

    glosas, devoluciones = sub.armar(filas, "2026-08-13")

    assert len(glosas) == 1
    assert devoluciones == []


def test_una_devolucion_reiterada_no_sale_lista_para_cargar():
    """Sin código ni observación, el bot no la manda a ninguna parte."""
    filas = [_fila(tipo_seguimiento="DEVOLUCION", descripcion_reiteracion="Devolución Reiterada")]

    glosas, devoluciones = sub.armar(filas, "2026-08-13")

    assert glosas == []
    assert len(devoluciones) == 1
    assert devoluciones[0]["CODIGO_RESPUESTA"] == ""
    assert devoluciones[0]["OBSERVACION_RESPUESTA"] == ""
    assert "NO cargar" in devoluciones[0]["REVISAR"]


def test_la_subsanacion_deja_constancia_de_la_respuesta_anterior():
    """Es el argumento de fondo: el hospital ya había contestado a tiempo."""
    r = sub.redactar_subsanacion(_fila())

    assert "05/08/2026" in r["OBSERVACION_RESPUESTA"]
    assert "HUS465198" in r["OBSERVACION_RESPUESTA"]
    assert "$1.396.804" in r["OBSERVACION_RESPUESTA"]


def test_la_subsanacion_no_repite_el_texto_de_la_primera_respuesta():
    """La EPS ya leyó ese texto y no lo aceptó: repetirlo no aporta nada."""
    from siifa_redactar_respuestas import redactar

    primera = redactar(_fila())["OBSERVACION_RESPUESTA"]
    segunda = sub.redactar_subsanacion(_fila())["OBSERVACION_RESPUESTA"]

    assert primera != segunda
    assert "SUBSANA LA GLOSA REITERADA" in segunda
    assert "NO APORTA ELEMENTO NUEVO" in segunda


def test_recoge_lo_que_decidio_la_eps():
    r = sub.redactar_subsanacion(_fila())

    assert "GLOSA REITERADA" in r["OBSERVACION_RESPUESTA"].upper()


def test_si_la_eps_no_motivo_la_reiteracion_se_deja_constancia():
    r = sub.redactar_subsanacion(_fila(descripcion_reiteracion=""))

    assert "SIN REGISTRAR DECISIÓN MOTIVADA" in r["OBSERVACION_RESPUESTA"]


def test_el_texto_cabe_en_el_limite_de_siifa():
    """1.500 caracteres. Pasarse devuelve un HTTP 500 que no dice la causa."""
    larga = _fila(
        descripcion_reiteracion="REITERADA. " * 400,
        descripcion_glosa="X" * 500,
    )

    r = sub.redactar_subsanacion(larga)

    assert len(r["OBSERVACION_RESPUESTA"]) <= sub.LIMITE_OBSERVACION


def test_al_recortar_no_se_pierde_ni_la_insistencia_ni_el_cierre():
    """Lo que se recorta es la cita de la EPS, no el argumento propio."""
    r = sub.redactar_subsanacion(_fila(descripcion_reiteracion="REITERADA. " * 400))

    texto = r["OBSERVACION_RESPUESTA"]
    assert "NO APORTA ELEMENTO NUEVO" in texto
    assert "SOLICITA SU LEVANTAMIENTO" in texto


def test_sostiene_la_no_aceptacion():
    """Si el hospital fuera a aceptar, el camino es la nota crédito, no esto."""
    assert sub.redactar_subsanacion(_fila())["CODIGO_RESPUESTA"] == "RE9901"


def test_el_excel_separa_las_devoluciones_en_su_propia_hoja(tmp_path):
    glosas, devoluciones = sub.armar(
        [
            _fila(),
            _fila(tipo_seguimiento="DEVOLUCION", descripcion_reiteracion="Devolución Reiterada"),
        ],
        "2026-08-13",
    )
    ruta = tmp_path / "SUBSANACION.xlsx"

    sub.escribir(glosas, devoluciones, ruta)

    wb = load_workbook(ruta)
    assert wb.sheetnames == ["RESPUESTAS", "DEVOLUCIONES_NO_CARGAR"]
    assert wb["RESPUESTAS"].max_row == 2
    assert wb["DEVOLUCIONES_NO_CARGAR"].max_row == 2


def test_sin_devoluciones_no_se_crea_la_hoja_de_aviso(tmp_path):
    glosas, devoluciones = sub.armar([_fila()], "2026-08-13")
    ruta = tmp_path / "SUBSANACION.xlsx"

    sub.escribir(glosas, devoluciones, ruta)

    assert load_workbook(ruta).sheetnames == ["RESPUESTAS"]


def test_el_excel_trae_las_columnas_que_lee_el_bot_de_cargue(tmp_path):
    """Si acá falta una columna, el cargue falla con el plazo encima."""
    import responder_glosas_siifa as bot

    glosas, _ = sub.armar([_fila()], "2026-08-13")
    ruta = tmp_path / "SUBSANACION.xlsx"
    sub.escribir(glosas, [], ruta)

    leidas = bot.leer_excel(ruta, None)

    assert len(leidas) == 1
    assert leidas[0]["id"] == 15110544
    assert leidas[0]["codigo"] == "RE9901"
    assert leidas[0]["observacion"]
    assert leidas[0]["fecha"], "sin fecha, SIIFA rechaza la subsanación"
    assert leidas[0]["tipo"] == "GLOSA", "el bot bloquea las devoluciones por este campo"


def test_el_valor_se_lee_con_el_lector_unico():
    glosas, _ = sub.armar([_fila(valor_glosa="$1.396.804")], "2026-08-13")

    assert glosas[0]["VALOR_GLOSA"] == 1_396_804


def test_la_plantilla_institucional_reemplaza_al_texto_del_motor():
    """El HUS radica un texto fijo ante la glosa ratificada (18-08-2026):
    RE9901, se mantiene la respuesta inicial, se pide conciliación."""
    plantilla = "ESE HUS NO ACEPTA GLOSA RATIFICADA; SE MANTIENE LA RESPUESTA DADA."

    glosas, _ = sub.armar([_fila()], "2026-08-18", texto_fijo=plantilla)

    assert glosas[0]["OBSERVACION_RESPUESTA"] == plantilla
    assert glosas[0]["CODIGO_RESPUESTA"] == "RE9901"


def test_la_plantilla_no_desbloquea_las_devoluciones():
    """La puerta de subsanación de devoluciones sigue sin confirmarse: la
    plantilla no cambia eso."""
    filas = [_fila(tipo_seguimiento="DEVOLUCION", descripcion_reiteracion="Devolución Reiterada")]

    glosas, devoluciones = sub.armar(filas, "2026-08-18", texto_fijo="TEXTO")

    assert glosas == []
    assert devoluciones[0]["OBSERVACION_RESPUESTA"] == ""


def test_una_plantilla_pasada_de_1500_se_rechaza_antes_de_armar_nada():
    import pytest

    with pytest.raises(SystemExit, match="1500"):
        sub.armar([_fila()], "2026-08-18", texto_fijo="X" * 1501)


def test_la_plantilla_del_hus_del_repo_cabe_y_dice_lo_que_debe():
    from pathlib import Path

    texto = " ".join(
        (Path(sub.__file__).parent / "plantillas" / "subsanacion_HUS.txt")
        .read_text(encoding="utf-8")
        .split()
    )

    assert len(texto) <= sub.LIMITE_OBSERVACION
    assert "NO ACEPTA GLOSA RATIFICADA" in texto
    assert "ARTÍCULO 57 DE LA LEY 1438 DE 2011" in texto
    assert "CONCILIACIÓN" in texto
