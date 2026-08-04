"""Pruebas del informe masivo de seguimientos SIIFA.

Cubren las dos fallas que aparecieron en la corrida real del auditor del
03-08-2026:

1. La consulta completa se cortó a mitad de camino (el servidor del
   Ministerio se sobrecargó) y el bot se conformó con lo poco que llevaba,
   en vez de completar el informe mes por mes.
2. Ese informe a medias se guardó ENCIMA del informe bueno de una corrida
   anterior, que se perdió sin aviso.
"""

from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

import pytest

TOOLS = Path(__file__).resolve().parents[2] / "tools"
sys.path.insert(0, str(TOOLS))

from siifa_client import SiifaApiError  # noqa: E402

import siifa_reporte_seguimientos as rep  # noqa: E402


def _registro(id_: int, mes: int) -> dict:
    """Un seguimiento como lo devuelve la API de SIIFA."""
    return {
        "idSeguimientoFactura": id_,
        "tipoSeguimiento": "GLOSA",
        "valor": 100000,
        "fechaCreacion": f"2026-{mes:02d}-10T08:00:00",
        "factura": {
            "numeroFactura": f"HUS{500000 + id_}",
            "valorBruto": 250000,
            "emisor": {"nitEmisor": "890201235", "razonSocial": "ESE HUS"},
            "adquiriente": {"nitAdquiriente": "900156264", "razonSocial": "EPS DE PRUEBA"},
        },
    }


class ClienteFalso:
    """Imita a SiifaClient. `guion` decide qué hace cada consulta."""

    def __init__(self, guion):
        self._guion = guion
        self.consultas: list[tuple[str | None, str | None]] = []

    def __enter__(self):
        return self

    def __exit__(self, *_):
        return False

    def login(self, usuario, password):
        return "token-de-prueba"

    def listar_seguimientos(self, **kw):
        ini = kw.get("fecha_creacion_inicio")
        fin = kw.get("fecha_creacion_final")
        self.consultas.append((ini, fin))
        yield from self._guion(ini, fin)


@pytest.fixture
def entorno(monkeypatch, tmp_path):
    """Deja listo el script para correr sin red ni credenciales reales."""
    monkeypatch.setattr(rep, "credenciales_desde_env", lambda: ("USUARIO", "CLAVE"))
    return tmp_path


def _correr(monkeypatch, cliente, salida: Path, extra: list[str] | None = None) -> None:
    monkeypatch.setattr(rep, "SiifaClient", lambda *a, **k: cliente)
    argv = ["siifa_reporte_seguimientos.py", "--salida", str(salida), "--desde", "2026-01-01"]
    argv += ["--hasta", "2026-03-31"]
    monkeypatch.setattr(sys, "argv", argv + (extra or []))
    rep.main()


def _filas_del_excel(ruta: Path) -> int:
    from openpyxl import load_workbook

    ws = load_workbook(ruta)["SEGUIMIENTOS"]
    return ws.max_row - 1  # menos el encabezado


def test_si_la_consulta_completa_se_corta_a_medias_sigue_mes_por_mes(monkeypatch, entorno):
    """La falla 1 de la corrida real.

    El servidor entrega un pedazo y después se cae. Antes, el bot guardaba
    ese pedazo y daba la corrida por terminada: el auditor se quedaba con 50
    de 2.597 registros. Ahora completa el informe mes por mes.
    """

    def guion(ini, fin):
        if ini == "2026-01-01" and fin == "2026-03-31":
            # Consulta de todo el rango: entrega 2 y se sobrecarga.
            yield _registro(1, 1)
            yield _registro(2, 1)
            raise SiifaApiError(500, "El servidor de SIIFA no alcanza a responder", None)
        # Mes por mes sí responde. Enero repite los 2 que ya habían bajado.
        mes = int(ini[5:7])
        if mes == 1:
            yield _registro(1, 1)
            yield _registro(2, 1)
            yield _registro(3, 1)
        elif mes == 2:
            yield _registro(4, 2)
        else:
            yield _registro(5, 3)

    cliente = ClienteFalso(guion)
    salida = entorno / "informe.xlsx"
    _correr(monkeypatch, cliente, salida)

    # El informe está completo y en el archivo que se pidió (no en un _PARCIAL).
    assert salida.exists()
    assert not (entorno / "informe_PARCIAL.xlsx").exists()
    # 5 registros distintos: los repetidos de enero se quitaron por id.
    assert _filas_del_excel(salida) == 5


def test_un_informe_incompleto_no_pisa_al_bueno(monkeypatch, entorno):
    """La falla 2 de la corrida real, que es la grave.

    El informe bueno (2.597 filas) se perdió porque una corrida fallida
    guardó 50 filas encima, en la misma ruta. Ahora el informe a medias se
    guarda al lado, con «_PARCIAL» en el nombre, y el bueno queda intacto.
    """
    salida = entorno / "informe.xlsx"
    salida.write_bytes(b"este es el informe bueno de la corrida anterior")
    antes = salida.read_bytes()

    def guion(ini, fin):
        if ini == "2026-01-01" and fin == "2026-03-31":
            yield _registro(1, 1)
            raise SiifaApiError(500, "sobrecargado", None)
        # Ni mes por mes responde: el informe no se puede completar.
        raise SiifaApiError(500, "sobrecargado", None)
        yield  # pragma: no cover  (hace de esto un generador)

    cliente = ClienteFalso(guion)
    _correr(monkeypatch, cliente, salida)

    # El bueno quedó tal cual: nadie lo tocó.
    assert salida.read_bytes() == antes
    # Lo bajado a medias quedó al lado, marcado.
    parcial = entorno / "informe_PARCIAL.xlsx"
    assert parcial.exists()
    assert _filas_del_excel(parcial) == 1


def test_dos_corridas_fallidas_no_se_pisan_entre_si(entorno):
    """Si ya hay un _PARCIAL, el siguiente se numera en vez de pisarlo."""
    pedido = entorno / "informe.xlsx"
    assert rep.ruta_para_informe_incompleto(pedido).name == "informe_PARCIAL.xlsx"

    (entorno / "informe_PARCIAL.xlsx").write_text("primera corrida fallida")
    assert rep.ruta_para_informe_incompleto(pedido).name == "informe_PARCIAL_2.xlsx"


def test_un_periodo_que_no_se_pudo_traer_deja_el_informe_marcado(monkeypatch, entorno):
    """Si un mes se pierde, el informe está incompleto aunque los otros
    hayan salido bien: tampoco puede pisar al bueno."""
    salida = entorno / "informe.xlsx"
    salida.write_text("informe bueno anterior")

    def guion(ini, fin):
        if ini == "2026-01-01" and fin == "2026-03-31":
            raise SiifaApiError(0, "sin respuesta", None)
            yield  # pragma: no cover
        mes = int(ini[5:7])
        if mes == 2:
            # Febrero no se puede traer ni partiéndolo en pedazos.
            raise SiifaApiError(500, "sobrecargado", None)
        yield _registro(mes, mes)

    cliente = ClienteFalso(guion)
    _correr(monkeypatch, cliente, salida)

    assert salida.read_text() == "informe bueno anterior"
    assert (entorno / "informe_PARCIAL.xlsx").exists()


def test_una_corrida_completa_si_reemplaza_al_informe_anterior(monkeypatch, entorno):
    """Lo normal: si la corrida sale entera, actualiza el archivo de siempre."""
    salida = entorno / "informe.xlsx"
    salida.write_text("informe viejo")

    def guion(ini, fin):
        yield _registro(1, 1)
        yield _registro(2, 2)

    cliente = ClienteFalso(guion)
    _correr(monkeypatch, cliente, salida)

    assert _filas_del_excel(salida) == 2
    assert not (entorno / "informe_PARCIAL.xlsx").exists()


def test_los_meses_del_rango_se_recorren_completos():
    """El barrido mes por mes cubre el rango sin dejar huecos ni solaparse."""
    tramos = rep._meses(date(2026, 1, 1), date(2026, 3, 31))
    assert tramos == [
        (date(2026, 1, 1), date(2026, 1, 31)),
        (date(2026, 2, 1), date(2026, 2, 28)),
        (date(2026, 3, 1), date(2026, 3, 31)),
    ]


def test_una_ruta_de_salida_que_no_sirve_se_avisa_antes_de_bajar_nada(tmp_path):
    """Bajar el informe toma minutos: la ruta se revisa primero.

    Pasó de verdad: se escribió un comando en vez de una carpeta, se bajaron
    los 2.598 seguimientos y todo se perdió al intentar guardarlos.
    """
    un_archivo = tmp_path / "no_soy_carpeta.txt"
    un_archivo.write_text("x", encoding="utf-8")

    with pytest.raises(SystemExit) as exc:
        rep.verificar_se_puede_guardar(un_archivo / "informe.xlsx")

    assert "no se va a poder guardar" in str(exc.value)


def test_una_ruta_buena_pasa_y_deja_la_carpeta_creada(tmp_path):
    destino = tmp_path / "SIIFA" / "informe_seguimientos.xlsx"

    rep.verificar_se_puede_guardar(destino)

    assert destino.parent.is_dir()
    assert not list(destino.parent.iterdir()), "no debe dejar basura"
