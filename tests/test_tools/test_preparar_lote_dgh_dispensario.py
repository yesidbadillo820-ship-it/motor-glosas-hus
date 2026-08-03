"""Tests del conversor de export DGH al formato del motor
(tools/preparar_lote_dgh_dispensario.py).

Cubre: deteccion automatica de la hoja de detalle, conversion al layout
HUS.xlsx (codigo+nombre de concepto, servicio+codigo CUPS, fecha de factura
como aproximacion documentada), el cruce con la hoja INICIAL (VENCE), y que el
resultado lo consuma cargar_glosas con familias/valores correctos.
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import pytest

_TOOLS = Path(__file__).resolve().parent.parent.parent / "tools"
sys.path.insert(0, str(_TOOLS))

import preparar_lote_dgh_dispensario as prep  # noqa: E402


def _export_dgh(ruta: Path):
    import openpyxl

    wb = openpyxl.Workbook()
    # Hoja INICIAL (control): FACTURA + VENCE + RADICADO
    ini = wb.active
    ini.title = "INICIAL"
    ini.append(["GESTOR", "FACTURA", "VALOR GLOSA", "VENCE", "RADICADO"])
    ini.append(["YESID", "HUS0000527398", 8735, datetime(2026, 8, 6), "R-111"])
    ini.append(["YESID", "HUS0000528436", 116260, datetime(2026, 8, 10), ""])
    # Hoja de detalle DGH (nombre arbitrario: se detecta por encabezados)
    det = wb.create_sheet("I")
    det.append(
        [
            "EstadoCxCObjecion",
            prep.COL_FACTURA,
            prep.COL_VALOR_FACTURA,
            prep.COL_FECHA_FACTURA,
            prep.COL_CONSECUTIVO,
            prep.COL_COD_CONCEPTO,
            prep.COL_NOM_CONCEPTO,
            prep.COL_COD_SERVICIO,
            prep.COL_DESC_SERVICIO,
            prep.COL_VALOR_OBJ,
            prep.COL_MOTIVO,
        ]
    )
    det.append(
        [
            "Glosa_Inicial",
            "HUS0000527398",
            87835,
            datetime(2026, 6, 18),
            "185828",
            "TA0201",
            "El cargo por consulta supera lo pactado",
            "890380H",
            "CONSULTA DE CONTROL POR ESPECIALISTA",
            8735,
            "TA0201 SE GENERA GLOSA POR MAYOR VALOR",
        ]
    )
    det.append(
        [
            "Glosa_Inicial",
            "HUS0000528436",
            148760,
            datetime(2026, 6, 20),
            "185830",
            "CL0801",
            "Pertinencia del servicio",
            "879101",
            "TOMOGRAFIA DE CUELLO",
            116260,
            "CL0801 SIN PERTINENCIA",
        ]
    )
    # fila sin concepto -> se descarta sin romper
    det.append(["Glosa_Inicial", "HUS0000999999", 1, None, "0", None, "", "", "", 0, ""])
    wb.save(str(ruta))


def test_convertir_y_consumir_con_cargar_glosas(tmp_path):
    dgh = tmp_path / "GLOSAS_DGH.xlsx"
    _export_dgh(dgh)
    salida = tmp_path / "HUS_LOTE.xlsx"
    r = prep.convertir(dgh, salida)
    assert r["glosas"] == 2 and r["facturas"] == 2
    assert r["valor_objetado_total"] == 8735 + 116260
    assert r["con_vence"] == 2  # ambas facturas cruzaron con INICIAL

    from asistente_conciliacion_dispensario import cargar_glosas, codigo_corto

    glosas = cargar_glosas(salida, None)
    assert len(glosas) == 2
    g1 = next(g for g in glosas if "527398" in g.factura)
    assert g1.familia == "TARIFAS"
    assert codigo_corto(g1.codigo) == "TA0201"
    assert "890380H" in g1.servicio  # el codigo CUPS viaja en el servicio
    assert g1.radicado == "R-111"  # radicado real de la hoja INICIAL
    assert isinstance(g1.fecha_atencion, datetime)  # fecha factura como aprox
    g2 = next(g for g in glosas if "528436" in g.factura)
    assert g2.familia == "CALIDAD"
    assert int(float(g2.valor_objetado)) == 116260


def test_detectar_hoja_falla_claro_si_no_es_dgh(tmp_path):
    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.append(["FACTURA", "VALOR"])  # no es un export DGH
    otro = tmp_path / "otro.xlsx"
    wb.save(str(otro))
    with pytest.raises(ValueError):
        prep.convertir(otro, tmp_path / "x.xlsx")
