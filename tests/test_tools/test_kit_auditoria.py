"""Pruebas de los 5 motores del kit de auditoría (bots de doble clic):

verificar_radicacion  — chequeo pre-radicación (previene glosas)
cruzar_glosas_xml     — Excel de glosas EPS ↔ XML + borradores por causal
semaforo_vencimientos — plazos en días hábiles CO con festivos
buscar_facturas       — recolector de soportes por lista
excel_a_csv           — cada Excel a .csv delimitado por comas
"""

from __future__ import annotations

import importlib.util
from datetime import date
from pathlib import Path

_RAIZ = Path(__file__).resolve().parents[2]

_XML_FACTURA = (
    '<?xml version="1.0" encoding="utf-8"?>'
    '<AttachedDocument xmlns="urn:ad" xmlns:cbc="urn:cbc" xmlns:cac="urn:cac">'
    "<cbc:ParentDocumentID>HUS533470</cbc:ParentDocumentID>"
    "<cac:ReceiverParty><cbc:RegistrationName>NUEVA EPS S.A.</cbc:RegistrationName></cac:ReceiverParty>"
    "<cbc:Description><![CDATA["
    '<?xml version="1.0"?><Invoice xmlns="urn:inv" xmlns:cbc="urn:cbc">'
    "<AdditionalInformation><Name>NUMERO_CONTRATO</Name><Value>01_VCEVN_900006037</Value></AdditionalInformation>"
    "<AdditionalInformation><Name>FACTURA_SIN_CONTRATO</Name><Value>ATENCION DE URGENCIAS</Value></AdditionalInformation>"
    "<AdditionalInformation><Name>COBERTURA_PLAN_BENEFICIOS</Name><Value>UPC Contributivo</Value></AdditionalInformation>"
    "<cbc:ID>HUS533470</cbc:ID><cbc:IssueDate>2026-07-03</cbc:IssueDate>"
    "<cbc:PayableAmount>480200.00</cbc:PayableAmount>"
    "</Invoice>]]></cbc:Description>"
    "<cbc:ValidationResultCode>02</cbc:ValidationResultCode>"
    "</AttachedDocument>"
)


def _cargar(nombre: str):
    ruta = _RAIZ / "tools" / f"{nombre}.py"
    spec = importlib.util.spec_from_file_location(nombre, ruta)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _excel_glosas(ruta: Path, filas: list[list]) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    for f in filas:
        ws.append(f)
    wb.save(ruta)


# ----------------------------- verificar_radicacion -----------------------


def test_verificar_detecta_lo_que_causa_glosas(tmp_path):
    (tmp_path / "ok.xml").write_text(_XML_FACTURA, encoding="utf-8")
    # XML sin contrato y sin validación DIAN
    malo = _XML_FACTURA.replace("HUS533470", "HUS999001")
    malo = malo.replace("<Value>01_VCEVN_900006037</Value>", "<Value></Value>")
    malo = malo.replace("<cbc:ValidationResultCode>02</cbc:ValidationResultCode>", "")
    (tmp_path / "malo.xml").write_text(malo, encoding="utf-8")
    (tmp_path / "HUS533470_HISTORIA.pdf").write_bytes(b"%PDF")

    mod = _cargar("verificar_radicacion")
    salida = tmp_path / "V.xlsx"
    assert mod.main([str(tmp_path), "--salida", str(salida)]) == 0

    from openpyxl import load_workbook

    ws = load_workbook(salida)["Facturas"]
    por_factura = {ws.cell(r, 2).value: ws.cell(r, 1).value for r in range(2, ws.max_row + 1)}
    assert por_factura["HUS533470"] == "LISTA"
    assert por_factura["HUS999001"] == "REVISAR"
    hallazgos = {ws.cell(r, 2).value: ws.cell(r, 9).value for r in range(2, ws.max_row + 1)}
    assert "NUMERO_CONTRATO vacio" in hallazgos["HUS999001"]
    assert "DIAN" in hallazgos["HUS999001"]
    assert "sin soportes" in hallazgos["HUS999001"]


def test_verificar_furips_lineas_inconsistentes(tmp_path):
    (tmp_path / "FURIPS2_lote.txt").write_text("a,b,c\n1,2,3\nx,y\n", encoding="utf-8")
    mod = _cargar("verificar_radicacion")
    res = mod.revisar_furips(tmp_path / "FURIPS2_lote.txt")
    assert res["campos"] == 3
    assert any("linea 3" in h for h in res["hallazgos"])


# ----------------------------- cruzar_glosas_xml --------------------------


def test_cruzar_arma_borrador_con_evidencia_real(tmp_path):
    (tmp_path / "xml").mkdir()
    (tmp_path / "xml" / "f.xml").write_text(_XML_FACTURA, encoding="utf-8")
    _excel_glosas(
        tmp_path / "G.xlsx",
        [
            ["Reporte de glosas"],  # fila basura antes del encabezado real
            ["FACTURA", "COD GLOSA", "VALOR GLOSADO", "OBSERVACION"],
            ["HUS0000533470", "DE4401", 480200, "FACTURA SIN CONTRATO"],
            ["HUS0000999999", "TA0201", 100000, "DIFERENCIA DE TARIFAS"],
        ],
    )
    mod = _cargar("cruzar_glosas_xml")
    salida = tmp_path / "C.xlsx"
    assert mod.main([str(tmp_path / "G.xlsx"), str(tmp_path / "xml"), "--salida", str(salida)]) == 0

    txt = (tmp_path / "BORRADORES_RESPUESTA.txt").read_text(encoding="utf-8")
    assert "01_VCEVN_900006037" in txt  # la evidencia va tal cual, sin dañarla
    assert "ATENCION DE URGENCIAS" in txt
    assert "[COMPLETAR" in txt  # siempre queda trabajo del auditor señalado

    from openpyxl import load_workbook

    ws = load_workbook(salida).active
    filas = {
        ws.cell(r, 1).value: [ws.cell(r, c).value for c in range(1, 11)]
        for r in range(2, ws.max_row + 1)
    }
    assert filas["HUS0000533470"][8] == "SI"  # XML hallado
    assert filas["HUS0000999999"][8] == "NO"
    assert filas["HUS0000533470"][2] == "Sin contrato / devolución"
    assert filas["HUS0000999999"][2] == "Tarifas"


def test_clasificar_causal_por_codigo_y_por_texto():
    mod = _cargar("cruzar_glosas_xml")
    assert mod.clasificar_causal("DE4401", "sin contrato")[0] == "contrato"
    assert mod.clasificar_causal("TA0201")[0] == "tarifas"
    assert mod.clasificar_causal("SO0101")[0] == "soportes"
    assert mod.clasificar_causal("", "FALTA AUTORIZACION")[0] == "autorizacion"
    assert mod.clasificar_causal("999", "sin datos")[0] == "general"


# ----------------------------- semaforo_vencimientos ----------------------


def test_semaforo_cuenta_habiles_con_festivos():
    mod = _cargar("semaforo_vencimientos")
    # 20 de julio de 2026 (lunes) es festivo: de viernes 17/07 el siguiente
    # hábil es el martes 21/07.
    assert mod.sumar_habiles(date(2026, 7, 17), 1) == date(2026, 7, 21)
    assert mod.habiles_entre(date(2026, 7, 17), date(2026, 7, 21)) == 1
    assert mod.habiles_entre(date(2026, 7, 21), date(2026, 7, 17)) == -1
    assert mod.semaforo(0) == "NEGRO (vencida)"
    assert mod.semaforo(4) == "ROJO"
    assert mod.semaforo(10) == "AMARILLO"
    assert mod.semaforo(11) == "VERDE"


def test_semaforo_excel_completo(tmp_path):
    _excel_glosas(
        tmp_path / "N.xlsx",
        [
            ["FACTURA", "FECHA NOTIFICACION GLOSA", "VALOR", "EPS"],
            ["HUS0000533470", "10/07/2026", 480200, "NUEVA EPS"],
            ["HUS0000111111", "01/06/2026", 99000, "NUEVA EPS"],  # vencida al corte
            ["HUS0000222222", "", 1, "NUEVA EPS"],  # sin fecha
        ],
    )
    mod = _cargar("semaforo_vencimientos")
    salida = tmp_path / "S.xlsx"
    assert mod.main([str(tmp_path / "N.xlsx"), "--hoy", "2026-07-16", "--salida", str(salida)]) == 0

    from openpyxl import load_workbook

    ws = load_workbook(salida).active
    por_factura = {ws.cell(r, 4).value: ws.cell(r, 1).value for r in range(2, ws.max_row + 1)}
    assert por_factura["HUS0000111111"] == "NEGRO (vencida)"
    assert por_factura["HUS0000533470"] == "VERDE"
    assert por_factura["HUS0000222222"] == "SIN FECHA"


# ----------------------------- buscar_facturas ----------------------------


def test_buscar_copia_organizado_y_reporta_faltantes(tmp_path):
    origen = tmp_path / "compartida"
    (origen / "sub").mkdir(parents=True)
    (origen / "HUS0000533470_HC.pdf").write_bytes(b"%PDF")
    (origen / "sub" / "533470 RIPS.txt").write_text("x", encoding="utf-8")
    (origen / "FURIPS268001007920.txt").write_text("x", encoding="utf-8")  # NO debe coincidir
    lista = tmp_path / "facturas.txt"
    lista.write_text("HUS0000533470\nHUS0000999999\n", encoding="utf-8")
    destino = tmp_path / "PAQUETE"

    mod = _cargar("buscar_facturas")
    assert mod.main([str(origen), "--lista", str(lista), "--destino", str(destino)]) == 0

    copiados = sorted(p.name for p in (destino / "HUS0000533470").iterdir())
    assert copiados == ["533470 RIPS.txt", "HUS0000533470_HC.pdf"]
    assert not (destino / "HUS0000999999").exists()
    resumen = (destino / "RESUMEN_BUSQUEDA.txt").read_text(encoding="utf-8")
    assert "HUS0000999999" in resumen  # reportada como sin soportes
    assert (origen / "HUS0000533470_HC.pdf").exists()  # el original ni se toca


def test_buscar_no_confunde_numeros_pegados(tmp_path):
    mod = _cargar("buscar_facturas")
    patrones = mod.compilar_patrones(["HUS533470"])
    p = patrones["HUS533470"]
    assert p.search("HUS0000533470_HC.pdf")
    assert p.search("factura 533470.pdf")
    assert not p.search("FURIPS533470999.txt")  # pegado a más dígitos: otro número


# ----------------------------- excel_a_csv --------------------------------


def test_excel_a_csv_formatos_y_no_pisa(tmp_path):
    from datetime import datetime

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["FACTURA", "FECHA", "VALOR", "NOTA"])
    ws.append(["HUS0000533470", datetime(2026, 7, 3), 480200.0, "atencion, urgencias"])
    wb.save(tmp_path / "L.xlsx")

    mod = _cargar("excel_a_csv")
    assert mod.main([str(tmp_path)]) == 0
    crudo = (tmp_path / "L.csv").read_bytes().decode("cp1252")
    assert "HUS0000533470,03/07/2026,480200," in crudo  # fecha dd/mm/aaaa y sin .0
    assert '"atencion, urgencias"' in crudo  # la coma interna va protegida

    # no pisa: si ya existe, se omite (y --forzar sí regenera)
    (tmp_path / "L.csv").write_text("no me pises", encoding="utf-8")
    assert mod.main([str(tmp_path)]) == 0
    assert (tmp_path / "L.csv").read_text(encoding="utf-8") == "no me pises"
    assert mod.main([str(tmp_path), "--forzar"]) == 0
    assert "HUS0000533470" in (tmp_path / "L.csv").read_bytes().decode("cp1252")


def test_excel_a_csv_todas_las_hojas(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.title = "Datos"
    wb.active.append(["a", "b"])
    ws2 = wb.create_sheet("Resumen 2026")
    ws2.append(["c", "d"])
    wb.save(tmp_path / "M.xlsx")

    mod = _cargar("excel_a_csv")
    assert mod.main([str(tmp_path), "--todas"]) == 0
    assert (tmp_path / "M_Datos.csv").exists()
    assert (tmp_path / "M_Resumen_2026.csv").exists()


# ----------------------------- motores embebidos en los .cmd --------------


def _embebido_igual_al_py(nombre_cmd: str, nombre_py: str) -> None:
    cmd_path = _RAIZ / "tools" / nombre_cmd
    lineas = cmd_path.read_text(encoding="utf-8").splitlines()
    marcador = "#PY" + "START#"
    idx = next(i for i, ln in enumerate(lineas) if marcador in ln)
    embebido = "\n".join(lineas[idx + 1 :]).strip("\n")
    fuente = (
        (_RAIZ / "tools" / nombre_py).read_text(encoding="utf-8").replace("\r\n", "\n").strip("\n")
    )
    assert embebido == fuente, f"regenerar {nombre_cmd}: difiere de {nombre_py}"


def test_motores_embebidos_identicos_al_py():
    _embebido_igual_al_py("VERIFICAR_RADICACION.cmd", "verificar_radicacion.py")
    _embebido_igual_al_py("CRUZAR_GLOSAS.cmd", "cruzar_glosas_xml.py")
    _embebido_igual_al_py("SEMAFORO_GLOSAS.cmd", "semaforo_vencimientos.py")
    _embebido_igual_al_py("BUSCAR_FACTURA.cmd", "buscar_facturas.py")
    _embebido_igual_al_py("EXCEL_A_CSV.cmd", "excel_a_csv.py")


def test_menu_motor_hus_es_ascii_crlf():
    """MOTOR_HUS.cmd es batch puro: 100% ASCII, CRLF, sin LF sueltos."""
    raw = (_RAIZ / "tools" / "MOTOR_HUS.cmd").read_bytes()
    assert raw.count(b"\n") == raw.count(b"\r\n")  # CRLF puro
    assert all(b < 128 for b in raw), "el batch debe ser 100% ASCII"
    assert b"REGISTRO_BOTS.csv" in raw  # la bitacora esta conectada


# ----------------------------- informe_gerencia ---------------------------


def test_informe_gerencia_desde_bitacora(tmp_path):
    (tmp_path / "REGISTRO_BOTS.csv").write_text(
        "fecha;hora;usuario;equipo;bot\n"
        "14/07/2026;08:12:33;yesid;AUD-PC01;UNIR_PDFS.cmd\n"
        "Wed 07/16/2026;11:45:00;maria;AUD-PC02;CRUZAR_GLOSAS.cmd\n",
        encoding="utf-8",
    )
    mod = _cargar("informe_gerencia")
    salida = tmp_path / "INF.html"
    assert mod.main([str(tmp_path / "REGISTRO_BOTS.csv"), "--salida", str(salida)]) == 0
    html = salida.read_text(encoding="utf-8")
    assert "UNIR_PDFS.cmd" in html and "maria" in html
    # el formato gringo Wed 07/16/2026 se interpreta bien (16 de julio)
    assert mod.parsear_fecha("Wed 07/16/2026").day == 16
    assert mod.parsear_fecha("14/07/2026").day == 14


def test_informe_gerencia_sin_bitacora(tmp_path):
    mod = _cargar("informe_gerencia")
    assert mod.main([str(tmp_path / "NO_EXISTE.csv")]) == 2


# ----------------------------- vigilante nocturno -------------------------


def test_vigilante_nocturno_estructura():
    """Batch puro con el RUN_NOCTURNO embebido tras #RUNSTART#, todo ASCII CRLF."""
    raw = (_RAIZ / "tools" / "VIGILANTE_NOCTURNO.cmd").read_bytes()
    assert raw.count(b"\n") == raw.count(b"\r\n")
    assert all(b < 128 for b in raw)
    marcador = ("#RUN" + "START#").encode()
    assert raw.count(marcador) == 1
    run = raw.decode("ascii").split("#RUN" + "START#\r\n", 1)[1]
    assert run.startswith("@echo off")
    assert "motor_nocturno.py" in run and "VIGILANTE_LOG.txt" in run
    assert "schtasks /create" in raw.decode("ascii")  # instala la tarea
    assert "schtasks /delete" in raw.decode("ascii")  # y se puede quitar
