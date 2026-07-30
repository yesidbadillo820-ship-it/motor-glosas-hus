"""
Pruebas de nucleo.office_tools (conversión Office ⇄ PDF — fase 2 de la caja).

Cada bloque se salta solo si le falta su motor, para no romper CI ni equipos
sin la Suite:
  · las conversiones PDF → Word/Excel/PowerPoint necesitan pdf2docx /
    pdfplumber / python-pptx (pesan; no van en CI, sí en el equipo del analista);
  · las conversiones Office → PDF y PDF/A necesitan LibreOffice instalado.
importar el módulo NO requiere ninguna de esas libs (se cargan dentro de cada
función), así que este archivo siempre se colecciona.
"""

import hashlib
import os
import sys

import pytest

SUITE = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "..", "tools", "suite_cartera_hus")
)
if SUITE not in sys.path:
    sys.path.insert(0, SUITE)

fitz = pytest.importorskip("fitz")

from nucleo import office_tools  # noqa: E402

sin_libreoffice = pytest.mark.skipif(
    not office_tools.hay_libreoffice(), reason="LibreOffice no está instalado"
)


def _silencio(_):
    pass


@pytest.fixture
def pdf(tmp_path):
    ruta = str(tmp_path / "soporte.pdf")
    doc = fitz.open()
    for i in range(2):
        doc.new_page().insert_text((72, 72), "Glosa %d Factura HUS99%d Valor 50000" % (i, i))
    doc.save(ruta)
    doc.close()
    return ruta


def _es_pdfa(ruta):
    doc = fitz.open(ruta)
    catalogo = doc.xref_object(doc.pdf_catalog())
    doc.close()
    return "OutputIntents" in catalogo


# --------------------------------------------------------- PDF → Office (libs)


def test_pdf_a_word(pdf, tmp_path):
    pytest.importorskip("pdf2docx")
    out = office_tools.pdf_a_word(pdf, str(tmp_path / "r.docx"), log=_silencio)
    assert os.path.exists(out) and os.path.getsize(out) > 0


def test_pdf_a_excel(pdf, tmp_path):
    pytest.importorskip("pdfplumber")
    pytest.importorskip("openpyxl")
    out = office_tools.pdf_a_excel(pdf, str(tmp_path / "r.xlsx"), log=_silencio)
    from openpyxl import load_workbook

    wb = load_workbook(out)
    assert len(wb.sheetnames) == 2  # una hoja por página


def test_pdf_a_powerpoint(pdf, tmp_path):
    pytest.importorskip("pptx")
    out = office_tools.pdf_a_powerpoint(pdf, str(tmp_path / "r.pptx"), log=_silencio)
    from pptx import Presentation

    assert len(Presentation(out).slides._sldIdLst) == 2


# --------------------------------------------------------- Office → PDF (LO)


@sin_libreoffice
def test_office_a_pdf_desde_html(tmp_path):
    html = tmp_path / "carta.html"
    html.write_text(
        "<html><body><h1>Objeción HUS</h1><p>COOSALUD</p></body></html>", encoding="utf-8"
    )
    out = office_tools.a_pdf(str(html), str(tmp_path), log=_silencio)
    assert os.path.exists(out) and out.endswith(".pdf")
    doc = fitz.open(out)
    assert "HUS" in doc[0].get_text()
    doc.close()


@sin_libreoffice
def test_office_a_pdf_salida_explicita(tmp_path):
    html = tmp_path / "c.html"
    html.write_text("<html><body>Prueba</body></html>", encoding="utf-8")
    destino = str(tmp_path / "MI_NOMBRE.pdf")
    out = office_tools.a_pdf(str(html), salida=destino, log=_silencio)
    assert out == destino and os.path.exists(destino)


@sin_libreoffice
def test_pdf_a_pdfa_es_pdfa(pdf, tmp_path):
    out = office_tools.pdf_a_pdfa(pdf, str(tmp_path), log=_silencio)
    assert _es_pdfa(out)


@sin_libreoffice
def test_pdf_a_pdfa_no_pisa_el_original(pdf):
    """PDF/A en la MISMA carpeta del PDF no debe sobrescribir el original."""
    antes = hashlib.md5(open(pdf, "rb").read()).hexdigest()
    out = office_tools.pdf_a_pdfa(pdf, os.path.dirname(pdf), log=_silencio)
    despues = hashlib.md5(open(pdf, "rb").read()).hexdigest()
    assert antes == despues
    assert os.path.abspath(out) != os.path.abspath(pdf)
    assert out.endswith("_pdfa.pdf")


# --------------------------------------------------------- sin LibreOffice


def test_mensaje_claro_sin_libreoffice(pdf, tmp_path, monkeypatch):
    monkeypatch.setattr(office_tools, "buscar_soffice", lambda: None)
    with pytest.raises(RuntimeError, match="LibreOffice"):
        office_tools.a_pdf(pdf, str(tmp_path), log=_silencio)
