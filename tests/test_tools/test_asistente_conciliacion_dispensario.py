"""Tests del asistente de conciliacion del Dispensario
(tools/asistente_conciliacion_dispensario.py).

Cubre: normalizacion de factura, codigo corto, familia, contrato por fecha,
clasificacion de soportes por token, indexacion de una carpeta que imita la
estructura real (ENV-.../HUS<f>/TIPO_900006037_HUS<f>.ext), y el flujo
end-to-end (cargar glosas de un Excel + procesar + escribir el Excel de
resultados), verificando que con soporte completo la glosa de TARIFAS concluye
'No procede' y sin soporte clinico la de SOPORTES queda 'Informacion insuficiente'.
"""

from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

_TOOLS = Path(__file__).resolve().parent.parent.parent / "tools"
sys.path.insert(0, str(_TOOLS))

import asistente_conciliacion_dispensario as asis  # noqa: E402


# ---------------------------------------------------------------------------
# Unidades puras
# ---------------------------------------------------------------------------
def test_norm_y_factura_corta():
    assert asis.norm_factura("436483") == "HUS0000436483"
    assert asis.norm_factura("HUS0000436483") == "HUS0000436483"
    assert asis.factura_corta("HUS0000436483") == "436483"


def test_codigo_corto():
    assert asis.codigo_corto("TA02 01 TARIFAS-CONSULTAS...") == "TA0201"
    assert asis.codigo_corto("CO46 01 COBERTURA-...") == "CO4601"
    assert asis.codigo_corto("re99 01") == "RE9901"


def test_familia_de():
    assert asis.familia_de("TA0201 ...") == "TARIFAS"
    assert asis.familia_de("SO36 01") == "SOPORTES"
    assert asis.familia_de("CO46") == "COBERTURA"


def test_contrato_por_fecha():
    assert asis.contrato_por_fecha(datetime(2025, 10, 28)) == "287"
    assert asis.contrato_por_fecha(datetime(2026, 3, 11)) == "440"
    assert asis.contrato_por_fecha(None) == "?"


def test_clasificar_por_token():
    assert asis.clasificar("RIPS_900006037_HUS436483.json") == "RIPS"
    assert asis.clasificar("HEV_900006037_HUS436483.pdf") == "Historia clinica / Evolucion"
    assert asis.clasificar("XML_900006037_HUS436483.xml") == "Factura electronica (XML/CUFE)"
    assert asis.clasificar("cualquiercosa.xml") == "Factura electronica (XML/CUFE)"


# ---------------------------------------------------------------------------
# Fixtures: carpeta de soportes que imita la estructura real
# ---------------------------------------------------------------------------
def _rips(fac, doc, cups):
    return {
        "numDocumentoIdObligado": "900006037",
        "numFactura": f"HUS{fac}",
        "usuarios": [
            {
                "tipoDocumentoIdentificacion": "CC",
                "numDocumentoIdentificacion": doc,
                "servicios": {"procedimientos": [{"codProcedimiento": cups, "vrServicio": 150000}]},
            }
        ],
    }


def _soportes(base: Path):
    root = base / "DISPENSARIO" / "LILIANA"
    # 436483 (TARIFAS): soporte completo
    d = root / "ENV-229912-OKDGH" / "HUS436483"
    d.mkdir(parents=True)
    (d / "RIPS_900006037_HUS436483.json").write_text(
        json.dumps(_rips("436483", "1029407046", "890201")), encoding="utf-8"
    )
    (d / "XML_900006037_HUS436483.xml").write_text(
        "<Invoice><ID>HUS436483</ID></Invoice>", encoding="utf-8"
    )
    (d / "CUV_900006037_HUS436483.json").write_text('{"ResultState":true}', encoding="utf-8")
    # 433131 (SOPORTES): SIN historia clinica (solo RIPS)
    d2 = root / "ENV-229900-OKDGH" / "HUS433131"
    d2.mkdir(parents=True)
    (d2 / "RIPS_900006037_HUS433131.json").write_text(
        json.dumps(_rips("433131", "91468363", "470700")), encoding="utf-8"
    )
    return base


def _excel(ruta: Path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    cols = [
        "FACTURA",
        "RADICADO",
        "FECHA ATENCION",
        "NOMBRES Y APELLIDOS",
        "IDENTIFICACION",
        "VALOR FACTURA",
        "# GLOSA",
        "CODIGO CONCEPTO DE GLOSA",
        "MOTIVO DE GLOSA",
        "VALOR OBJETADO",
        "SERVICIO OBJETADO",
    ]
    ws.append(cols)
    ws.append(
        [
            "436483",
            "644561",
            datetime(2025, 10, 28),
            "VALERY MOSQUERA",
            "1029407046",
            231461,
            "1",
            "TA02 01 TARIFAS-CONSULTAS",
            "Mayor valor cobrado",
            156061,
            "CONSULTA CONTROL 890201",
        ]
    )
    ws.append(
        [
            "433131",
            "640960",
            datetime(2025, 9, 11),
            "JUAN RODRIGUEZ",
            "91468363",
            12120600,
            "1",
            "SO36 01 SOPORTES-COPIA",
            "Glosa total soportes",
            12120600,
            "Glosa Total",
        ]
    )
    wb.save(str(ruta))


def test_indexar_soportes(tmp_path):
    _soportes(tmp_path)
    idx = asis.indexar_soportes(tmp_path, {"436483", "433131"}, ocr=False)
    assert len(idx["436483"]) == 3
    assert len(idx["433131"]) == 1
    tipos = {d.tipo for d in idx["436483"]}
    assert "RIPS" in tipos and "Factura electronica (XML/CUFE)" in tipos


def test_end_to_end(tmp_path):
    _soportes(tmp_path)
    xls = tmp_path / "glosas.xlsx"
    _excel(xls)

    glosas = asis.cargar_glosas(xls, solo=None)
    assert len(glosas) == 2
    idx = asis.indexar_soportes(
        tmp_path, {asis.factura_corta(g.factura) for g in glosas}, ocr=False
    )
    resultados = asis.procesar(glosas, idx, kb={})

    by_fac = {r.glosa.factura: r for r in resultados}
    tarifas = by_fac["HUS0000436483"]
    soportes = by_fac["HUS0000433131"]
    # TARIFAS con factura+RIPS+CUV y cruces OK -> No procede, alta confianza
    assert tarifas.conclusion == "No procede"
    assert tarifas.confianza >= 80
    assert tarifas.contrato == "287"
    assert any(c.startswith("OK") for c in tarifas.coherencia)
    # SOPORTES sin historia clinica -> no llega a 'No procede'
    assert soportes.conclusion in ("Informacion insuficiente", "Procede parcialmente")
    assert any(not m["encontrado"] for m in soportes.matriz)

    salida = tmp_path / "out.xlsx"
    asis.escribir_excel(resultados, salida)
    import openpyxl

    wb = openpyxl.load_workbook(salida)
    assert wb.sheetnames == ["Respuestas", "Matriz_Evidencia"]
    assert wb["Respuestas"].max_row == 3  # header + 2 glosas


def test_solo_filtra_una_factura(tmp_path):
    xls = tmp_path / "glosas.xlsx"
    _excel(xls)
    glosas = asis.cargar_glosas(xls, solo="HUS436483")
    assert len(glosas) == 1
    assert glosas[0].factura == "HUS0000436483"
