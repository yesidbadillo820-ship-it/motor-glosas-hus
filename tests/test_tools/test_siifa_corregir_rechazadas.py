"""Pruebas del corrector de respuestas rechazadas por SIIFA.

Sale de un caso real (05-08-2026): de 2.579 respuestas, 1.422 quedaron sin
subir. 927 porque el texto pasaba de 1.500 caracteres —SIIFA contesta un
error 500 que parece del servidor— y 495 porque el código RE9701 no lo
acepta la plataforma.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

TOOLS = Path(__file__).resolve().parents[2] / "tools"
sys.path.insert(0, str(TOOLS))

import siifa_corregir_rechazadas as cor  # noqa: E402

COLS = [
    "ID_SEGUIMIENTO_FACTURA_GLOSA",
    "NUMERO_FACTURA",
    "TIPO",
    "CODIGO_GLOSA",
    "VALOR_GLOSA",
    "ORIGEN_RESPUESTA",
    "CODIGO_RESPUESTA",
    "FECHA_RESPUESTA",
    "OBSERVACION_RESPUESTA",
]


def _fila(ident, codigo="RE9901", texto="ESE HUS NO ACEPTA.", origen="REDACTADA"):
    return {
        "ID_SEGUIMIENTO_FACTURA_GLOSA": ident,
        "NUMERO_FACTURA": "HUS454747",
        "TIPO": "GLOSA",
        "CODIGO_GLOSA": "TA0801",
        "VALOR_GLOSA": 22200,
        "ORIGEN_RESPUESTA": origen,
        "CODIGO_RESPUESTA": codigo,
        "FECHA_RESPUESTA": "2026-05-11",
        "OBSERVACION_RESPUESTA": texto,
    }


def _excel(tmp_path: Path, filas: list[dict]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "RESPUESTAS"
    ws.append(COLS)
    for f in filas:
        ws.append([f.get(c) for c in COLS])
    ruta = tmp_path / "respuestas.xlsx"
    wb.save(ruta)
    return ruta


def _reporte(tmp_path: Path, estados: dict[int, str], nombre="reporte.csv") -> Path:
    ruta = tmp_path / nombre
    lineas = ["id_seguimiento_factura_glosa,numero_factura,estado,detalle,fecha_hora"]
    for ident, estado in estados.items():
        lineas.append(f"{ident},HUS454747,{estado},,2026-08-05T09:00:00")
    ruta.write_text("\n".join(lineas), encoding="utf-8")
    return ruta


def test_el_texto_largo_se_recorta_al_limite_de_siifa():
    largo = "ESE HUS NO ACEPTA LA GLOSA. " * 100  # 2.800 caracteres

    texto, recortado = cor.recortar(largo)

    assert recortado
    assert len(texto) <= 1500


def test_el_recorte_no_parte_una_palabra_por_la_mitad():
    """El texto se le manda a la EPS: no puede quedar cortado a mitad de palabra."""
    largo = "PALABRA " * 400

    texto, _ = cor.recortar(largo, limite=100)

    assert not texto.rstrip(".").endswith("PALAB")
    assert texto.endswith("...") or texto.endswith(".")


def test_un_texto_que_ya_cabe_no_se_toca():
    texto, recortado = cor.recortar("ESE HUS NO ACEPTA.")

    assert texto == "ESE HUS NO ACEPTA."
    assert not recortado


def test_el_limite_es_el_que_mide_siifa():
    """1.500: entraron todas las de 1.499 y ninguna de 1.501."""
    assert cor.LIMITE_OBSERVACION == 1500


def test_solo_se_vuelven_a_subir_las_que_quedaron_en_error(tmp_path):
    excel = _excel(tmp_path, [_fila(1), _fila(2), _fila(3)])
    reporte = _reporte(tmp_path, {1: "OK", 2: "ERROR", 3: "ERROR"})
    salida = tmp_path / "corregidas.xlsx"

    _correr(excel, [reporte], salida)

    ws = load_workbook(salida).active
    ids = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert ids == [2, 3]


def test_lo_que_quedo_ok_en_cualquiera_de_los_reportes_no_se_repite(tmp_path):
    """El cargue se corrió varias veces: no se puede volver a subir lo bueno."""
    excel = _excel(tmp_path, [_fila(1), _fila(2)])
    primero = _reporte(tmp_path, {1: "ERROR", 2: "ERROR"}, "r1.csv")
    segundo = _reporte(tmp_path, {1: "OK"}, "r2.csv")
    salida = tmp_path / "corregidas.xlsx"

    _correr(excel, [primero, segundo], salida)

    ws = load_workbook(salida).active
    ids = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert ids == [2]


def test_el_codigo_que_siifa_no_acepta_se_cambia_y_queda_avisado(tmp_path):
    """RE9701 no existe en SIIFA. El cambio se marca: puede alterar el sentido."""
    excel = _excel(tmp_path, [_fila(1, codigo="RE9701", texto="ESE HUS ACEPTA DEVOLUCION.")])
    reporte = _reporte(tmp_path, {1: "ERROR"})
    salida = tmp_path / "corregidas.xlsx"

    _correr(excel, [reporte], salida, cambios=["RE9701=RE9901"])

    ws = load_workbook(salida).active
    hdr = [c.value for c in ws[1]]
    fila = dict(zip(hdr, [c.value for c in ws[2]]))
    assert fila["CODIGO_RESPUESTA"] == "RE9901"
    assert "REVISAR" in fila["CORRECCION"]


def test_el_archivo_que_sale_lo_puede_leer_el_bot_de_cargue(tmp_path):
    excel = _excel(tmp_path, [_fila(1, texto="X" * 2000)])
    reporte = _reporte(tmp_path, {1: "ERROR"})
    salida = tmp_path / "corregidas.xlsx"

    _correr(excel, [reporte], salida)

    ws = load_workbook(salida).active
    encabezados = {c.value for c in ws[1]}
    assert {
        "ID_SEGUIMIENTO_FACTURA_GLOSA",
        "CODIGO_RESPUESTA",
        "OBSERVACION_RESPUESTA",
        "FECHA_RESPUESTA",
    } <= encabezados
    fila = dict(zip([c.value for c in ws[1]], [c.value for c in ws[2]]))
    assert len(fila["OBSERVACION_RESPUESTA"]) <= 1500
    assert fila["FECHA_RESPUESTA"] == "2026-05-11", "la fecha de DGH no se puede perder"


def _correr(excel, reportes, salida, cambios=(), extra=()):
    """Corre el script como lo corre el auditor, por línea de comandos."""
    import contextlib
    import io

    argv = ["siifa_corregir_rechazadas.py", "--excel", str(excel), "--salida", str(salida)]
    for r in reportes:
        argv += ["--reporte", str(r)]
    for c in cambios:
        argv += ["--cambiar-codigo", c]
    argv += list(extra)
    viejo = sys.argv
    sys.argv = argv
    try:
        with contextlib.redirect_stdout(io.StringIO()):
            cor.main()
    finally:
        sys.argv = viejo


def test_se_pueden_apartar_las_respuestas_de_un_codigo(tmp_path):
    """Las de un código dudoso se tratan aparte, no se mezclan con el resto."""
    excel = _excel(tmp_path, [_fila(1, codigo="RE9901"), _fila(2, codigo="RE9701")])
    reporte = _reporte(tmp_path, {1: "ERROR", 2: "ERROR"})

    resto = tmp_path / "resto.xlsx"
    _correr(excel, [reporte], resto, extra=["--excluir-codigo", "RE9701"])
    aparte = tmp_path / "aparte.xlsx"
    _correr(excel, [reporte], aparte, extra=["--solo-codigo", "RE9701"])

    def ids(ruta):
        ws = load_workbook(ruta).active
        return [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]

    assert ids(resto) == [1]
    assert ids(aparte) == [2]
