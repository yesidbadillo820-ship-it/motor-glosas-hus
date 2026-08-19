"""Pruebas de la revisión previa al cargue.

Un cargue de 60.000 respuestas no se corrige sobre la marcha: lo que entra
mal, entra mal para siempre. Cada caso de acá salió del archivo real de la
Clínica Socorro (13-08-2026): 10.935 ya respondidas, 519 con el texto pasado
de largo y 16 sin texto.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

TOOLS = Path(__file__).resolve().parents[2] / "tools"
sys.path.insert(0, str(TOOLS))

import siifa_perfiles as perfiles  # noqa: E402
import siifa_revisar_antes_de_cargar as rev  # noqa: E402

GIRON = perfiles.buscar("GIRON")


def _fila(**kw):
    base = {
        "id_seguimiento_factura_glosa": 24123540,
        "id_seguimiento_factura_devolucion": None,
        "numero_factura": "FE102938",
        "nit_emisor": "890203242",
        "tipo_seguimiento": "GLOSA",
        "razon_social_eps": "E.P.S. FAMISANAR SAS",
        "valor_glosa": 25000,
        "codigo_glosa": "TA0201",
        "fecha_formulacion": "2026-08-06T12:29:32",
        "tiene_respuesta": "NO",
        "codigo_respuesta": "RE9602",
        "observacion_respuesta": "No se acepta glosa de mayor valor cobrado.",
        "fecha_respuesta": "2026-08-14 00:00:00",
    }
    base.update(kw)
    return base


def _revisar(filas, **kw):
    return rev.revisar(filas, GIRON, **kw)


# ------------------------------------------------------- lo que pasa limpio


def test_una_respuesta_completa_queda_lista():
    listas, aparte = _revisar([_fila()])

    assert len(listas) == 1 and aparte == []
    assert listas[0]["MOTIVO"] == ""
    assert listas[0]["CODIGO_RESPUESTA"] == "RE9602"


def test_el_none_del_excel_no_se_toma_por_dato():
    """La API devuelve el texto 'None' y 'null' en las celdas vacías; tomarlos
    por contenido haría pasar como buena una respuesta sin sustento."""
    listas, aparte = _revisar([_fila(observacion_respuesta="null")])

    assert listas == []
    assert "Sin texto" in aparte[0]["MOTIVO"]


# -------------------------------------------- lo que NO se puede cargar


def test_lo_ya_respondido_en_siifa_queda_fuera():
    """Volver a cargarlo PISA la respuesta registrada y su fecha —que es la
    que prueba que se respondió dentro del término—."""
    listas, aparte = _revisar([_fila(tiene_respuesta="SI")])

    assert listas == []
    assert "Ya está respondida" in aparte[0]["MOTIVO"]


def test_se_puede_pedir_expresamente_pisar_lo_ya_respondido():
    listas, aparte = _revisar([_fila(tiene_respuesta="SI")], incluir_ya_respondidas=True)

    assert len(listas) == 1 and aparte == []


def test_una_respuesta_sin_texto_queda_fuera():
    listas, aparte = _revisar([_fila(observacion_respuesta="")])

    assert listas == []
    assert "Sin texto" in aparte[0]["MOTIVO"]


def test_una_devolucion_con_codigo_de_glosa_queda_fuera():
    """RE9901 es de glosa. En una devolución, SIIFA lo rechaza diciendo que no
    pertenece al grupo — el error que costó descubrir en agosto."""
    listas, aparte = _revisar([_fila(tipo_seguimiento="DEVOLUCION", codigo_respuesta="RE9901")])

    assert listas == []
    assert "código de glosa" in aparte[0]["MOTIVO"]


def test_una_glosa_con_codigo_de_devolucion_queda_fuera():
    listas, aparte = _revisar([_fila(tipo_seguimiento="GLOSA", codigo_respuesta="RE9601")])

    assert listas == []
    assert "código de devolución" in aparte[0]["MOTIVO"]


def test_una_devolucion_con_su_codigo_pasa():
    for codigo in ("RE9501", "RE9601", "RE9701"):
        listas, aparte = _revisar([_fila(tipo_seguimiento="DEVOLUCION", codigo_respuesta=codigo)])
        assert len(listas) == 1, f"{codigo} es un código válido de devolución"


def test_una_respuesta_anterior_a_la_glosa_queda_fuera():
    """SIIFA rechaza una respuesta fechada antes de que la EPS formulara."""
    listas, aparte = _revisar([_fila(fecha_formulacion="2026-08-06", fecha_respuesta="2026-07-01")])

    assert listas == []
    assert "anterior a la glosa" in aparte[0]["MOTIVO"]


def test_sin_fecha_de_respuesta_no_se_bloquea():
    """48.766 filas de Socorro venían sin fecha: el bot pone la de hoy, que
    siempre es posterior a la formulación."""
    listas, aparte = _revisar([_fila(fecha_respuesta="")])

    assert len(listas) == 1 and aparte == []


def test_sin_codigo_de_respuesta_queda_fuera():
    listas, aparte = _revisar([_fila(codigo_respuesta="")])

    assert listas == []
    assert "Sin código" in aparte[0]["MOTIVO"]


# ------------------------------------------------- el límite de SIIFA


def test_el_texto_largo_se_recorta_y_queda_marcado():
    """519 filas de Socorro pasaban de 1.500. Sin recortar, rebotan con un
    HTTP 500 que no dice la causa; recortadas en silencio, el auditor no se
    entera de que su texto salió a medias."""
    largo = "NO SE ACEPTA LA GLOSA. " * 200

    listas, aparte = _revisar([_fila(observacion_respuesta=largo)])

    assert len(listas) == 1, "se carga igual: recortada es mejor que rechazada"
    assert len(listas[0]["OBSERVACION_RESPUESTA"]) <= rev.LIMITE_OBSERVACION
    assert "recortado" in listas[0]["MOTIVO"]


def test_el_recorte_no_parte_una_palabra():
    texto = "PALABRA " * 500

    recortado = rev.recortar(texto)

    assert len(recortado) <= rev.LIMITE_OBSERVACION
    assert recortado.endswith("...")
    assert "PALABR..." not in recortado


def test_un_texto_que_cabe_no_se_toca():
    assert rev.recortar("Corto y claro.") == "Corto y claro."


# ------------------------------------------------- el archivo de otra IPS


def test_un_archivo_de_otra_ips_no_se_puede_cargar():
    """La misma guarda del resto del repo, acá también: es el último punto
    antes de escribir en la plataforma."""
    with pytest.raises(SystemExit, match="ALTO"):
        perfiles.verificar_informe(GIRON, [_fila(nit_emisor="900190045")])


# ------------------------------------------------- lo que se entrega


def test_el_excel_de_cargue_lo_lee_el_bot(tmp_path):
    """Si las columnas no calzan, el cargue falla con 48.000 filas listas."""
    import responder_glosas_siifa as bot

    listas, _ = _revisar([_fila(), _fila(tipo_seguimiento="DEVOLUCION", codigo_respuesta="RE9601")])
    ruta = tmp_path / "CARGAR.xlsx"
    rev.escribir(listas, ruta, "RESPUESTAS")

    leidas = bot.leer_excel(ruta, None)

    assert len(leidas) == 2
    assert leidas[0]["id"] == 24123540
    assert leidas[0]["codigo"] == "RE9602"
    assert leidas[0]["observacion"]
    assert leidas[1]["tipo"] == "DEVOLUCION", "el bot necesita el TIPO para elegir la puerta"


def test_el_archivo_de_lo_apartado_dice_el_motivo(tmp_path):
    _, aparte = _revisar([_fila(tiene_respuesta="SI"), _fila(observacion_respuesta="")])
    ruta = tmp_path / "REVISAR.xlsx"

    rev.escribir(aparte, ruta, "REVISAR")

    ws = load_workbook(ruta)["REVISAR"]
    motivos = [ws.cell(row=r, column=rev.COLS_SALIDA.index("MOTIVO") + 1).value for r in (2, 3)]
    assert all(m for m in motivos), "cada fila apartada tiene que decir por qué"


def test_se_cuentan_bien_los_valores_con_devoluciones(tmp_path):
    """Una devolución vale una vez por factura: la regla del repo también
    aplica al resumen de esta revisión."""
    filas = [
        _fila(tipo_seguimiento="DEVOLUCION", codigo_respuesta="RE9601", valor_glosa=1_000_000)
        for _ in range(50)
    ]

    listas, _ = _revisar(filas)

    from siifa_novedades import valor_total

    assert valor_total(rev._para_valor(listas)) == 1_000_000
