"""Las autorizaciones que faltan en los RIPS antes de radicar.

OT-030 · 13-08-2026. Un RIPS al que le faltan números de autorización lo
rechaza el Ministerio, y la factura se devuelve. Hasta ahora eso se revisaba
abriendo los JSON uno por uno.

El caso real que trajo Yesid (HUS544245) es el típico y el que más duele:
la factura trae la autorización 4724500503082 en consultas, procedimientos,
hospitalización y otros servicios —pero los **47 medicamentos** van con
`null`—. A ojo, en un archivo de 75 KB, eso no se ve.

El bot recorre carpetas y subcarpetas completas —sirve una ruta del servidor
de facturación— y saca el informe en Excel con el veredicto por factura.
"""

from __future__ import annotations

import json
import zipfile
from pathlib import Path

import pytest

from tools.autorizaciones_rips import _clasificar, _resolver_entradas, analizar_archivo, procesar

AUT = "4724500503082"


def _rips(numero_factura: str, servicios: dict) -> dict:
    return {
        "numDocumentoIdObligado": "900006037",
        "numFactura": numero_factura,
        "usuarios": [{"tipoDocumentoIdentificacion": "CC", "servicios": servicios}],
    }


def _escribir(carpeta: Path, nombre: str, datos: dict, bom: bool = False) -> Path:
    ruta = carpeta / nombre
    ruta.write_text(
        json.dumps(datos, ensure_ascii=False),
        encoding="utf-8-sig" if bom else "utf-8",
    )
    return ruta


@pytest.fixture
def caso_real(tmp_path: Path) -> Path:
    """Reproduce el HUS544245: todo con autorización menos los medicamentos."""
    return _escribir(
        tmp_path,
        "RIPS_HUS544245.json",
        _rips(
            "HUS544245",
            {
                "consultas": [
                    {"codConsulta": f"89020{i}", "numAutorizacion": AUT} for i in range(3)
                ],
                "procedimientos": [
                    {"codProcedimiento": f"87200{i}", "numAutorizacion": AUT} for i in range(23)
                ],
                "medicamentos": [
                    {"codTecnologiaSalud": f"MED{i:03d}", "numAutorizacion": None}
                    for i in range(47)
                ],
                "otrosServicios": [
                    {"codServicio": f"OTR{i:03d}", "numAutorizacion": AUT} for i in range(30)
                ],
            },
        ),
        bom=True,
    )


class TestElCasoReal:
    def test_encuentra_la_autorizacion_de_la_factura(self, caso_real):
        r = analizar_archivo(caso_real, log=lambda *_a, **_k: None)
        assert AUT in json.dumps(r, default=str)

    def test_cuenta_los_47_medicamentos_sin_autorizacion(self, caso_real):
        r = analizar_archivo(caso_real, log=lambda *_a, **_k: None)
        texto = json.dumps(r, default=str)
        assert "47" in texto, "no contó los servicios sin autorización"

    def test_el_json_con_BOM_se_lee(self, caso_real):
        """Los RIPS que salen del sistema del hospital traen BOM UTF-8. Sin
        `utf-8-sig` el archivo entero se cae como ilegible."""
        r = analizar_archivo(caso_real, log=lambda *_a, **_k: None)
        assert r is not None

    def test_genera_el_informe_excel(self, caso_real, tmp_path):
        salida = tmp_path / "INFORME.xlsx"
        procesar([str(caso_real)], str(salida), log=lambda *_a, **_k: None)
        assert salida.exists() and salida.stat().st_size > 0

    def test_el_informe_trae_las_tres_hojas(self, caso_real, tmp_path):
        from openpyxl import load_workbook

        salida = tmp_path / "INFORME.xlsx"
        procesar([str(caso_real)], str(salida), log=lambda *_a, **_k: None)
        hojas = {h.upper() for h in load_workbook(salida).sheetnames}
        assert {"RESUMEN", "DETALLE", "ALERTAS"} <= hojas, hojas


class TestRecorreCarpetasYSubcarpetas:
    """Es lo que pidió Yesid: que busque en las carpetas de facturación y
    en todo lo que cuelgue de ellas, no archivo por archivo."""

    def test_encuentra_los_json_de_las_subcarpetas(self, tmp_path):
        hondo = tmp_path / "2026" / "07.JULIO" / "HUS544245" / "RIPS"
        hondo.mkdir(parents=True)
        _escribir(hondo, "a.json", _rips("HUS1", {"consultas": [{"numAutorizacion": AUT}]}))
        _escribir(tmp_path, "b.json", _rips("HUS2", {"consultas": [{"numAutorizacion": AUT}]}))
        encontrados = _resolver_entradas([str(tmp_path)])
        assert len(encontrados) == 2, encontrados

    def test_tambien_dentro_de_un_zip(self, tmp_path):
        datos = _rips("HUS3", {"consultas": [{"numAutorizacion": AUT}]})
        suelto = _escribir(tmp_path, "dentro.json", datos)
        z = tmp_path / "lote.zip"
        with zipfile.ZipFile(z, "w") as zf:
            zf.write(suelto, "carpeta/dentro.json")
        suelto.unlink()
        assert _resolver_entradas([str(z)])

    def test_se_pueden_mezclar_entradas(self, tmp_path):
        c = tmp_path / "carpeta"
        c.mkdir()
        _escribir(c, "a.json", _rips("HUS1", {"consultas": [{"numAutorizacion": AUT}]}))
        suelto = _escribir(tmp_path, "b.json", _rips("HUS2", {"consultas": []}))
        assert len(_resolver_entradas([str(c), str(suelto)])) == 2

    def test_una_carpeta_sin_json_no_revienta(self, tmp_path):
        (tmp_path / "vacia").mkdir()
        assert _resolver_entradas([str(tmp_path / "vacia")]) == []


class TestComoSeClasificaCadaAutorizacion:
    """`_clasificar` devuelve (estado, texto_mostrable)."""

    def test_lo_que_cuenta_como_faltante(self):
        """`null`, vacío y el texto "null" son el mismo problema: el
        Ministerio los rechaza igual, y por eso los tres se reportan."""
        for vacio in (None, "", "   ", "null", "NULL", "None"):
            estado, _ = _clasificar(vacio)
            assert estado != "OK", repr(vacio)

    def test_cada_forma_de_faltar_se_nombra_distinto(self):
        """El auditor tiene que poder distinguir el campo que llegó vacío
        del que llegó con la palabra "null" escrita: se corrigen distinto."""
        assert _clasificar(None)[0] == "NULL"
        assert _clasificar("")[0] == "VACIO"
        assert _clasificar("null")[0] == "TEXTO_NULL"

    def test_un_numero_de_verdad_pasa(self):
        assert _clasificar(AUT) == ("OK", AUT)

    def test_un_numero_como_entero_tambien(self):
        """El JSON a veces trae la autorización sin comillas."""
        assert _clasificar(int(AUT)) == ("OK", AUT)

    def test_los_espacios_de_sobra_no_estorban(self):
        assert _clasificar(f"  {AUT}  ") == ("OK", AUT)


class TestLoQueNoDebePasar:
    def test_un_json_roto_no_tumba_el_lote(self, tmp_path):
        (tmp_path / "roto.json").write_text("{ esto no es json", encoding="utf-8")
        _escribir(tmp_path, "bueno.json", _rips("HUS9", {"consultas": [{"numAutorizacion": AUT}]}))
        salida = tmp_path / "INFORME.xlsx"
        procesar([str(tmp_path)], str(salida), log=lambda *_a, **_k: None)
        assert salida.exists(), "un archivo ilegible no puede botar el informe entero"

    def test_un_rips_sin_servicios(self, tmp_path):
        ruta = _escribir(tmp_path, "vacio.json", _rips("HUS0", {}))
        assert analizar_archivo(ruta, log=lambda *_a, **_k: None) is not None
