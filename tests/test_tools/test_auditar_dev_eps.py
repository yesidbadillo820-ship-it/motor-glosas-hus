"""Pruebas del motor tools/auditar_devoluciones_eps.py (bot AUDITAR_DEV_EPS.cmd).

Audita devoluciones de Nueva EPS cruzando la factura del DGH (soportes), el
RIPS (JSON) y los soportes de autorización (PDF). El caso real HUS532392 (bebé
"Madre Canguro" con documento provisional en el JSON) debe reproducir la
observación del ejemplo de la guía.
"""

from __future__ import annotations

import importlib.util
from pathlib import Path

import pytest

_RAIZ = Path(__file__).resolve().parents[2]
_MOTOR = _RAIZ / "tools" / "auditar_devoluciones_eps.py"

# RIPS real de HUS532392 (recién nacido, documento provisional en el JSON).
_RIPS_JSON = (
    '{"numDocumentoIdObligado":"900006037","numFactura":"HUS532392",'
    '"usuarios":[{"tipoDocumentoIdentificacion":"RC","numDocumentoIdentificacion":"26066910192188",'
    '"servicios":{"procedimientos":[{"numAutorizacion":"316003877","codProcedimiento":"890315",'
    '"codServicio":371,"vrServicio":3397900,"codDiagnosticoPrincipal":"P073"}]}}]}'
)

# Fragmentos de texto tal como salen de los soportes OPF/PDE (con ruido OCR).
_TEXTO_SOPORTE = """MINISTERIO DE SALUD SOLICITUD DE AUTORIZACION
Hospital Universitario de Santander NIT 900006037-4 CODIGO: 680010079201
ENTIDAD A LA QUE SE SOLICITA (PAGADOR ): NUEVA EPS SUBSIDIADO
ASHLEY VALENTINA COLINA ROJAS tz
NUMERO DE IDENTIFICACION
1244781967 t---''
N Autorizacion: (POS) 5251-316003877
Afiliado: RC 1244781967 ROJAS ASHLEY VALENTINA A.
Factura N: HUS532392  Valor 3.397.900
Nombre del usuario: ASHLEY VALENTINA COLINA ROJAS
LIQUIDADOR KAROL AFANADOR
PAQUETE MENSUAL PROGRAMA MADRE CANGURO ETAPA II
Telefono Afiliado: 3002132174
"""


def _cargar():
    spec = importlib.util.spec_from_file_location("auditar_devoluciones_eps", _MOTOR)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _pdf(ruta: Path, texto: str) -> None:
    import fitz

    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((40, 60), texto, fontsize=9)
    doc.save(str(ruta))
    doc.close()


def _excel_dev(ruta: Path, facturas: list[tuple]) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(
        ["FACTURA DGH"] + [""] * 9 + ["RIPS - JSON", "", "", "SOPORTES AUTORIZACION", "", "", ""]
    )
    ws.append(
        [
            "FACTURA",
            "FAC",
            "VALOR FACTURA",
            "FECHA FACTURA",
            "FECHA RADICACION",
            "ENVIO",
            "TIPO",
            "NUMERO DE DOCUMENTO",
            "NOMBRE DEL PACIENTE",
            "SERVICIO",
            "TIPO",
            "NUMERO DE DOCUMENTO",
            "Nº AUTORIZACION",
            "Nº AUTORIZACION",
            "TIPO",
            "NUMERO DE DOCUMENTO",
            "OBSERVACION",
        ]
    )
    for fac, fecha, valor, envio in facturas:
        ws.append([f"HUS0000{fac[3:]}", fac, valor, fecha, "2026-07-09", envio] + [""] * 11)
    wb.save(ruta)


# --------------------------- extractores de texto -------------------------


def test_extraer_documento_ignora_institucionales():
    mod = _cargar()
    assert mod.extraer_documento(_TEXTO_SOPORTE) == ("RC", "1244781967")
    # el NIT y el codigo de prestador no son el documento del paciente
    assert mod.extraer_documento("NIT 900006037 CODIGO 680010079201")[1] == ""


def test_extraer_autorizaciones_todas_ultimos_9_digitos():
    mod = _cargar()
    # el PDE real trae varias, con prefijo (POS) 5251-/P071-: solo los ultimos 9
    texto = (
        "N° Autorización: (POS) 5251-313608762\n"
        "N° Autorización: (POS) 5251-313608676\n"
        "N° Autorización: (POS) P071-314624922\n"
    )
    assert mod.extraer_autorizaciones(texto) == ["313608762", "313608676", "314624922"]
    # ultimos 9 aunque el prefijo tenga letras (P071 -> 071...)
    assert mod._ultimos9("P071-314624922") == "314624922"
    assert mod._ultimos9("5251-316003877") == "316003877"


def test_extraer_autorizaciones_del_soporte_simple():
    mod = _cargar()
    assert mod.extraer_autorizaciones(_TEXTO_SOPORTE) == ["316003877"]
    # etiqueta en mayusculas y salto de linea
    assert mod.extraer_autorizaciones("NUMERO DE AUTORIZACION\n5251-316003877") == ["316003877"]


def test_json_autorizacion_null_se_reporta_tal_cual(tmp_path):
    mod = _cargar()
    j = tmp_path / "Rips_x.json"
    j.write_text(
        '{"numFactura":"HUS1","usuarios":[{"tipoDocumentoIdentificacion":"CC",'
        '"numDocumentoIdentificacion":"63508124","servicios":{'
        '"consultas":[{"numAutorizacion":null,"codConsulta":"890201"}],'
        '"procedimientos":[{"numAutorizacion":"313608762"}]}}]}',
        encoding="utf-8",
    )
    rips = mod.leer_rips(j)
    auts = mod.autorizaciones_del_json(rips)
    assert "null" in auts and "313608762" in auts


def test_extraer_nombre_no_cruza_lineas():
    mod = _cargar()
    assert mod.extraer_nombre(_TEXTO_SOPORTE) == "ASHLEY VALENTINA COLINA ROJAS"


def test_extraer_servicio():
    mod = _cargar()
    assert "MADRE CANGURO" in mod.extraer_servicio(_TEXTO_SOPORTE)


def test_extraer_documento_no_toma_de_preposicion_ni_valores():
    mod = _cargar()
    # 'DE' preposición + un valor en pesos no es un documento
    assert mod.extraer_documento("Valor total DE 1500000 pesos. Paciente: CC 79345612") == (
        "CC",
        "79345612",
    )
    # una fecha con 'DE' tampoco
    assert (
        mod.extraer_documento("VIGENCIA DE 20230101 AL 20230131. Afiliado RC 1098070312345")[1]
        == "1098070312345"
    )


def test_extraer_documento_prefiere_paciente_sobre_medico():
    mod = _cargar()
    texto = "Medico tratante CC 79111222\nPaciente: CC 52098765"
    assert mod.extraer_documento(texto) == ("CC", "52098765")


def test_extraer_documento_ignora_telefono():
    mod = _cargar()
    texto = "TELEFONO 3001234567\nAfiliado RC 1032456789"
    assert mod.extraer_documento(texto)[1] == "1032456789"


def test_leer_rips_estructura(tmp_path):
    mod = _cargar()
    j = tmp_path / "Rips_HUS532392.json"
    j.write_text(_RIPS_JSON, encoding="utf-8")
    rips = mod.leer_rips(j)
    assert rips["factura"] == "HUS532392"
    assert rips["usuarios"][0]["doc"] == "26066910192188"
    assert rips["usuarios"][0]["servicios"][0]["autoriz"] == "316003877"
    assert mod.autorizaciones_del_json(rips) == ["316003877"]


# --------------------------- observación ----------------------------------


def test_observacion_recien_nacido_reproduce_el_ejemplo():
    mod = _cargar()
    sop = {"doc": "1244781967", "autorizaciones": ["316003877"]}
    obs = mod.observacion("26066910192188", ["316003877"], sop)
    assert obs == "OK - DIFERENCIA DEL NUMERO DE DOCUMENTO DGH VS JSON"


def test_observacion_todo_coincide_es_ok():
    mod = _cargar()
    sop = {"doc": "1244781967", "autorizaciones": ["316003877"]}
    obs = mod.observacion("1244781967", ["316003877"], sop)
    assert obs == "OK"


def test_observacion_autorizacion_diferente():
    mod = _cargar()
    sop = {"doc": "111", "autorizaciones": ["999999999"]}
    obs = mod.observacion("111", ["316003877"], sop)
    assert "AUTORIZACION DIFERENTE" in obs


def test_observacion_varias_autorizaciones_coinciden_como_conjunto():
    mod = _cargar()
    # JSON y soporte traen el mismo conjunto de autorizaciones (en otro orden)
    sop = {"doc": "111", "autorizaciones": ["313608676", "313608762"]}
    obs = mod.observacion("111", ["313608762", "313608676"], sop)
    assert obs == "OK"


def test_observacion_avisa_json_null():
    mod = _cargar()
    sop = {"doc": "111", "autorizaciones": ["313608762"]}
    obs = mod.observacion("111", ["null", "313608762"], sop)
    assert obs.startswith("OK") and "NULL" in obs.upper()


def test_observacion_no_dice_ok_si_soporte_sin_documento():
    mod = _cargar()
    # autorización coincide pero el OCR no sacó documento: no debe quedar "OK" a secas
    sop = {"doc": "", "autorizaciones": ["316003877"]}
    obs = mod.observacion("1122334455", ["316003877"], sop)
    assert obs == "OK - SIN DOCUMENTO EN EL SOPORTE (no verificado)"


def test_observacion_sin_autorizacion_en_ambos():
    mod = _cargar()
    sop = {"doc": "1122334455", "autorizaciones": []}
    obs = mod.observacion("1122334455", [], sop)
    assert obs == "SIN AUTORIZACION EN JSON NI SOPORTE"


# --------------------------- end to end -----------------------------------


def test_procesar_end_to_end(tmp_path):
    pytest.importorskip("fitz")
    mod = _cargar()
    fbase = tmp_path / "facturas"
    sbase = tmp_path / "soportes"
    jdir = fbase / "202606" / "FACTURAS_SALUD" / "HUS532392" / "RIPS"
    jdir.mkdir(parents=True)
    (jdir / "Rips_HUS532392.json").write_text(_RIPS_JSON, encoding="utf-8")
    sdir = sbase / "NUEVA EPS" / "KARIN" / "ENV-230601" / "SOPORTES" / "HUS532392"
    sdir.mkdir(parents=True)
    _pdf(sdir / "OPF_900006037_HUS532392.pdf", _TEXTO_SOPORTE)
    _pdf(sdir / "PDE_900006037_HUS532392.pdf", _TEXTO_SOPORTE)

    excel = tmp_path / "DEV.xlsx"
    _excel_dev(excel, [("HUS532392", "2026-06-30", 3397900, "230601")])
    salida = tmp_path / "DEV_AUDITADO.xlsx"
    assert mod.procesar(excel, fbase, sbase, salida) == 0

    from openpyxl import load_workbook

    wb = load_workbook(salida)
    ws = wb.active
    fila = [c.value for c in ws[3]]  # primera factura
    # bloque RIPS-JSON
    assert fila[10] == "RC" and fila[11] == "26066910192188" and fila[12] == "316003877"
    # bloque soportes
    assert fila[13] == "316003877" and fila[14] == "RC" and fila[15] == "1244781967"
    # bloque DGH
    assert fila[7] == "1244781967"
    assert fila[8] == "ASHLEY VALENTINA COLINA ROJAS"
    # observación
    assert "DIFERENCIA DEL NUMERO DE DOCUMENTO DGH VS JSON" in fila[16]
    # hoja DETALLE con el servicio
    ws2 = wb["DETALLE"]
    filas = list(ws2.values)
    assert filas[1][0] == "HUS532392" and filas[1][5] == "316003877"


def test_factura_sin_datos_no_revienta(tmp_path):
    mod = _cargar()
    excel = tmp_path / "DEV.xlsx"
    _excel_dev(excel, [("HUS999999", "2026-06-30", 100, "230999")])
    salida = tmp_path / "OUT.xlsx"
    # bases vacías: no hay JSON ni soportes
    assert mod.procesar(excel, tmp_path / "no_facturas", tmp_path / "no_soportes", salida) == 0
    from openpyxl import load_workbook

    ws = load_workbook(salida).active
    assert "SIN JSON" in str(ws.cell(3, 17).value)


def test_indexar_no_confunde_facturas_por_substring(tmp_path):
    mod = _cargar()
    (tmp_path / "HUS532392").mkdir()
    (tmp_path / "HUS532392" / "Rips_HUS532392.json").write_text(_RIPS_JSON, encoding="utf-8")
    (tmp_path / "HUS5323921").mkdir()
    (tmp_path / "HUS5323921" / "Rips_HUS5323921.json").write_text(_RIPS_JSON, encoding="utf-8")
    idx = mod.indexar_arbol(tmp_path, {mod.normalizar_factura("HUS532392")})
    # solo la 532392 debe indexarse, no la 5323921
    assert idx[mod.normalizar_factura("HUS532392")]["json"] is not None
    assert mod.normalizar_factura("HUS5323921") not in idx


def test_indexar_estructura_plana_real_con_nit_en_el_nombre(tmp_path):
    """Los soportes reales se llaman OPF_<NIT>_<FAC>.pdf en una carpeta ENV-*;
    deben emparejarse pese a que el NIT también es un número en el nombre."""
    mod = _cargar()
    env = tmp_path / "ENV-230601-OK-C-DGH" / "SOPORTES"
    env.mkdir(parents=True)
    (env / "OPF_900006037_HUS532392.pdf").write_bytes(b"%PDF")
    (env / "PDE_900006037_HUS532392.pdf").write_bytes(b"%PDF")
    (env / "Rips_HUS532392.json").write_text(_RIPS_JSON, encoding="utf-8")
    norm = mod.normalizar_factura("HUS532392")
    idx = mod.indexar_arbol(tmp_path, {norm})
    assert len(idx[norm]["opf"]) == 1
    assert len(idx[norm]["pde"]) == 1
    assert idx[norm]["json"] is not None


def test_indexar_no_roba_soportes_de_otra_factura_en_carpeta_mal_rotulada(tmp_path):
    """Una carpeta rotulada HUS532392 con un OPF de HUS777888 NO debe asignar
    ese soporte a 532392 (el nombre del archivo manda)."""
    mod = _cargar()
    carpeta = tmp_path / "HUS532392"
    carpeta.mkdir()
    (carpeta / "OPF_900006037_HUS777888.pdf").write_bytes(b"%PDF")
    norm = mod.normalizar_factura("HUS532392")
    idx = mod.indexar_arbol(tmp_path, {norm})
    assert idx[norm]["opf"] == []  # no se lo roba


def test_leer_rips_json_no_dict_no_revienta(tmp_path):
    mod = _cargar()
    j = tmp_path / "Rips_x.json"
    j.write_text('[{"numFactura":"HUS1"}]', encoding="utf-8")  # arreglo de nivel superior
    rips = mod.leer_rips(j)
    assert rips["error"]  # se marca, no lanza excepción
    assert rips["usuarios"] == []


def test_leer_rips_ansi_con_enie(tmp_path):
    mod = _cargar()
    j = tmp_path / "Rips_x.json"
    contenido = '{"numFactura":"HUS1","usuarios":[{"tipoDocumentoIdentificacion":"CC","numDocumentoIdentificacion":"123456","servicios":{}}]}'
    # guardar en cp1252 (con un ñ en un comentario simulado no válido -> usar dato con ñ)
    j.write_bytes(contenido.encode("cp1252"))
    rips = mod.leer_rips(j)
    assert rips["usuarios"][0]["doc"] == "123456"


def test_filas_total_no_se_procesan(tmp_path):
    mod = _cargar()
    from openpyxl import load_workbook

    excel = tmp_path / "DEV.xlsx"
    _excel_dev(excel, [("HUS532392", "2026-06-30", 3397900, "230601")])
    # agregar una fila TOTAL al final
    wb = load_workbook(excel)
    ws = wb.active
    ws.append(["", "TOTAL GENERAL", 3397900, "", "", ""] + [""] * 11)
    wb.save(excel)

    salida = tmp_path / "OUT.xlsx"
    assert mod.procesar(excel, tmp_path / "nf", tmp_path / "ns", salida) == 0
    ws = load_workbook(salida).active
    # la fila TOTAL no debe tener observación (no se procesó como factura)
    assert ws.cell(4, 17).value in (None, "")


def test_motor_embebido_en_cmd_identico_al_py():
    cmd_path = _RAIZ / "tools" / "AUDITAR_DEV_EPS.cmd"
    lineas = cmd_path.read_text(encoding="utf-8").splitlines()
    marcador = "#PY" + "START#"
    idx = next(i for i, ln in enumerate(lineas) if marcador in ln)
    embebido = "\n".join(lineas[idx + 1 :]).strip("\n")
    fuente = _MOTOR.read_text(encoding="utf-8").replace("\r\n", "\n").strip("\n")
    assert embebido == fuente, "regenerar AUDITAR_DEV_EPS.cmd: difiere del .py"


def test_soportes_se_hallan_aunque_el_envio_no_coincida(tmp_path):
    """Regresion: antes se podaba por envio y si el envio de la carpeta no era
    el del Excel, no hallaba NINGUN soporte. Ahora debe hallarlos igual."""
    mod = _cargar()
    # el Excel dice envio 230601, pero el soporte esta archivado en ENV-999888
    sd = tmp_path / "NUEVA EPS" / "GESTOR" / "ENV-999888-XX" / "SOPORTES" / "HUS532392"
    sd.mkdir(parents=True)
    (sd / "OPF_900006037_HUS532392.pdf").write_bytes(b"%PDF")
    (sd / "PDE_900006037_HUS532392.pdf").write_bytes(b"%PDF")
    norm = mod.normalizar_factura("HUS532392")
    idx = mod.indexar_arbol(tmp_path, {norm})
    assert len(idx[norm]["opf"]) == 1
    assert len(idx[norm]["pde"]) == 1


def test_excel_incluye_rutas_de_los_archivos(tmp_path):
    pytest.importorskip("fitz")
    mod = _cargar()
    fbase = tmp_path / "facturas"
    sbase = tmp_path / "soportes"
    jd = fbase / "202606" / "FACTURAS_SALUD" / "HUS532392" / "RIPS"
    jd.mkdir(parents=True)
    (jd / "Rips_HUS532392.json").write_text(_RIPS_JSON, encoding="utf-8")
    sd = sbase / "NUEVA EPS" / "ENV-230601" / "SOPORTES" / "HUS532392"
    sd.mkdir(parents=True)
    _pdf(sd / "OPF_900006037_HUS532392.pdf", _TEXTO_SOPORTE)

    excel = tmp_path / "DEV.xlsx"
    _excel_dev(excel, [("HUS532392", "2026-06-30", 3397900, "230601")])
    salida = tmp_path / "OUT.xlsx"
    assert mod.procesar(excel, fbase, sbase, salida) == 0

    from openpyxl import load_workbook

    ws = load_workbook(salida).active
    assert ws.cell(2, 18).value == "VALIDACION SAT (PDE)"
    assert ws.cell(2, 19).value == "RUTA DEL JSON"
    assert ws.cell(2, 20).value == "RUTA DE LOS SOPORTES"
    assert ws.cell(3, 19).value.endswith("Rips_HUS532392.json")
    assert "OPF_900006037_HUS532392.pdf" in ws.cell(3, 20).value


def test_validacion_sat_detecta_proceso_exitoso():
    mod = _cargar()
    texto = (
        "Mi Seguridad Social - Confirmacion de la novedad\n"
        "Proceso exitoso. Su afiliacion al regimen subsidiado se ha reportado "
        "de manera exitosa realizada bajo el numero 074CC13514092080620260856000001\n"
        "https://sat.miseguridadsocial.gov.co/Novedades/Novedad/GuardarNovedad"
    )
    sat = mod.validacion_sat(texto)
    assert sat["es_sat"] and sat["exitoso"]
    assert sat["numero"] == "074CC13514092080620260856000001"
    assert "PROCESO EXITOSO" in mod.texto_validacion_sat(sat)
    # un soporte que NO es de SAT no reporta nada
    assert mod.validacion_sat("Autorizacion Nueva EPS 316003877")["es_sat"] is False
    assert mod.texto_validacion_sat({"es_sat": False}) == ""


def test_validacion_sat_sin_proceso_exitoso():
    mod = _cargar()
    texto = "Mi Seguridad Social. GuardarNovedad. La novedad no pudo procesarse."
    sat = mod.validacion_sat(texto)
    assert sat["es_sat"] and not sat["exitoso"]
    assert "NO se evidencia" in mod.texto_validacion_sat(sat)
