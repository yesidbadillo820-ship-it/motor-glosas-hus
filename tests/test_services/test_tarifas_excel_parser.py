"""Tests del parser de Excel de tarifas (tipo Famisanar 3 hojas)."""
from __future__ import annotations

from datetime import datetime
from io import BytesIO

import pytest
from openpyxl import Workbook

from app.services.tarifas_excel_parser import (
    _normalizar_valor,
    _parsear_fecha,
    _parsear_porcentaje,
    parsear_excel_tarifas,
)


class TestNormalizarValor:
    def test_numero_directo(self):
        assert _normalizar_valor(136746) == 136746.0

    def test_float_directo(self):
        assert _normalizar_valor(136746.50) == 136746.5

    def test_string_con_pesos(self):
        assert _normalizar_valor("$ 136.746") == 136746.0

    def test_string_con_millones(self):
        assert _normalizar_valor("$ 2.858.499") == 2858499.0

    def test_na(self):
        assert _normalizar_valor("N/A") == 0.0

    def test_none(self):
        assert _normalizar_valor(None) == 0.0

    def test_string_vacio(self):
        assert _normalizar_valor("") == 0.0


class TestParsearPorcentaje:
    def test_con_signo_menos(self):
        assert _parsear_porcentaje("-5%") == -5.0

    def test_con_signo_mas(self):
        assert _parsear_porcentaje("+10%") == 10.0

    def test_sin_signo(self):
        assert _parsear_porcentaje("-15%") == -15.0

    def test_numero_entero(self):
        assert _parsear_porcentaje(-5) == -5.0

    def test_fraccion_excel(self):
        # Excel a veces guarda 5% como 0.05
        assert _parsear_porcentaje(-0.05) == pytest.approx(-5.0)

    def test_cero(self):
        assert _parsear_porcentaje(0) == 0.0

    def test_vacio(self):
        assert _parsear_porcentaje("") == 0.0


class TestParsearFecha:
    def test_ya_datetime(self):
        d = datetime(2026, 4, 15)
        assert _parsear_fecha(d) == d

    def test_formato_ddmmyyyy(self):
        assert _parsear_fecha("15/04/2026") == datetime(2026, 4, 15)

    def test_formato_iso(self):
        assert _parsear_fecha("2026-04-15") == datetime(2026, 4, 15)

    def test_typo_doble_slash(self):
        # Caso real del Excel de Famisanar: "14//04/2027"
        assert _parsear_fecha("14//04/2027") == datetime(2027, 4, 14)

    def test_invalido(self):
        assert _parsear_fecha("abc") is None


# ─── Tests integración con Excel sintético ────────────────────────────────

def _crear_excel_famisanar_3_hojas() -> bytes:
    """Construye un xlsx en memoria con la estructura real Famisanar:
    hoja 1 (Anexo 3), hoja 2 (Anexo 3.1), hoja 3 (Anexo 3.2).
    """
    wb = Workbook()

    # ─── Hoja 1: Anexo 3 (Servicios CUPS) ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "Anexo 3"
    # Encabezado con metadata
    ws1["A1"] = "INFORMACIÓN EPS"
    ws1["A2"] = "NOMBRE DE LA EPS"
    ws1["F2"] = "FAMISANAR EPS"
    ws1["A3"] = "NÚMERO DE CONTRATO PRIMIGENIO"
    ws1["F3"] = "S-13-1-03-1-04958"
    ws1["A4"] = "VIGENCIA INICIO"
    ws1["F4"] = "15/04/2026"
    ws1["A5"] = "VIGENCIA FINAL"
    ws1["F5"] = "14/04/2027"
    # Fila encabezado de tabla (fila 10)
    headers1 = [
        "CUPS / CUMS / MIPRES", "DESCRIPCIÓN", "COD. REPS", "DESCRIPCIÓN REPS",
        "CÓDIGO PROPIO", "DESCRIPCIÓN CÓDIGO PROPIO", "INTERDEPENDENCIA",
        "TIPO TARIFA", "HOSPITALARIO", "AMBULATORIO", "URGENCIA",
        "CÓDIGO SEDE", "MARCA POR LISTADO", "OBSERVACIÓN",
    ]
    for i, h in enumerate(headers1, start=1):
        ws1.cell(row=10, column=i, value=h)
    # Datos
    ws1.append([])  # fila 11 vacía para probar skip
    ws1.cell(row=11, column=1, value="010101")
    ws1.cell(row=11, column=2, value="PUNCION CISTERNAL VIA LATERAL")
    ws1.cell(row=11, column=8, value="SOAT UVB VIGENTE")
    ws1.cell(row=11, column=9, value="-5%")
    ws1.cell(row=11, column=10, value="-5%")
    ws1.cell(row=11, column=11, value="-5%")
    ws1.cell(row=11, column=14, value="SE APLICA MANUAL SOAT VIGENTE")

    ws1.cell(row=12, column=1, value="010102")
    ws1.cell(row=12, column=2, value="PUNCION CISTERNAL VIA MEDIAL")
    ws1.cell(row=12, column=8, value="SOAT UVB VIGENTE")
    ws1.cell(row=12, column=9, value="-5%")

    # ─── Hoja 2: Anexo 3.1 (Medicamentos) ─────────────────────────────────
    ws2 = wb.create_sheet("Anexo 3.1")
    ws2["A1"] = "ANEXO TÉCNICO No 3.1: TARIFAS MEDICAMENTOS"
    ws2["A3"] = "NOMBRE DE LA EPS"
    ws2["F3"] = "FAMISANAR EPS"
    ws2["A4"] = "NUMERO DE CONTRATO"
    ws2["F4"] = "S-13-1-03-1-04958"
    headers2 = [
        "CONSECUTIVO", "CÓDIGO DCI", "DESCRIPCIÓN DCI", "CÓDIGO DEL PRESTADOR",
        "COD. REPS", "DESCRIPCIÓN REPS", "MAPIISS", "CUM/IUM", "DESCRIPCIÓN",
        "AGRUPADOR", "TARIFA UNITARIA", "APLICA IVA (SI-NO)", "TIPO PLAN",
        "OBSERVACIONES",
    ]
    for i, h in enumerate(headers2, start=1):
        ws2.cell(row=10, column=i, value=h)
    ws2.cell(row=11, column=1, value=1)
    ws2.cell(row=11, column=2, value="A10AE04SI-001")
    ws2.cell(row=11, column=3, value="INSULINA GLARGINA 1000 UI")
    ws2.cell(row=11, column=4, value="19914262-04")
    ws2.cell(row=11, column=7, value="P19914262-04")
    ws2.cell(row=11, column=10, value="MEDICAMENTOS")
    ws2.cell(row=11, column=11, value=136746)
    ws2.cell(row=11, column=12, value="NO")

    ws2.cell(row=12, column=1, value=2)
    ws2.cell(row=12, column=2, value="R03BB01SH-001")
    ws2.cell(row=12, column=3, value="IPRATROPIO BROMURO 20 MCG")
    ws2.cell(row=12, column=4, value="20066836-02")
    ws2.cell(row=12, column=10, value="MEDICAMENTOS")
    ws2.cell(row=12, column=11, value=20929)
    ws2.cell(row=12, column=12, value="NO")

    # ─── Hoja 3: Anexo 3.2 (Suministros) ──────────────────────────────────
    ws3 = wb.create_sheet("Anexo 3.2")
    ws3["A1"] = "ANEXO TÉCNICO No 3.2: TARIFAS SUMINISTROS"
    ws3["A3"] = "NOMBRE DE LA EPS"
    ws3["E3"] = "FAMISANAR EPS"
    headers3 = [
        "CONSECUTIVO", "MAPIISS", "CÓDIGO DEL PRESTADOR", "DESCRIPCIÓN DEL PRESTADOR",
        "AGRUPADOR", "TARIFA UNITARIA", "APLICA IVA (SI-NO)",
        "TARIFA FINAL (IVA INCLUIDO)", "OBSERVACIONES",
    ]
    for i, h in enumerate(headers3, start=1):
        ws3.cell(row=10, column=i, value=h)
    ws3.cell(row=11, column=1, value=1)
    ws3.cell(row=11, column=2, value="91010491")
    ws3.cell(row=11, column=3, value="FMQ6296")
    ws3.cell(row=11, column=4, value="STEN CORONARIO MEDICADO 2.25 MM X 14 MM")
    ws3.cell(row=11, column=5, value="SUMINISTROS CARDIOVASCULAR")
    ws3.cell(row=11, column=6, value=2858499)
    ws3.cell(row=11, column=7, value="NO")
    ws3.cell(row=11, column=8, value="N/A")

    ws3.cell(row=12, column=1, value=2)
    ws3.cell(row=12, column=2, value="91012378")
    ws3.cell(row=12, column=3, value="QX0106")
    ws3.cell(row=12, column=4, value="PINZAS LIGASURE LF537")
    ws3.cell(row=12, column=5, value="SUMINISTROS CIRUGIA GENERAL")
    ws3.cell(row=12, column=6, value=1660670)
    ws3.cell(row=12, column=7, value="NO")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestParsearExcelTarifas:
    @pytest.fixture
    def excel_bytes(self) -> bytes:
        return _crear_excel_famisanar_3_hojas()

    def test_detecta_metadata_eps(self, excel_bytes):
        r = parsear_excel_tarifas(excel_bytes, "famisanar.xlsx")
        assert r["eps"] == "FAMISANAR EPS"

    def test_detecta_contrato(self, excel_bytes):
        r = parsear_excel_tarifas(excel_bytes, "famisanar.xlsx")
        assert r["contrato"] == "S-13-1-03-1-04958"

    def test_detecta_vigencia(self, excel_bytes):
        r = parsear_excel_tarifas(excel_bytes, "famisanar.xlsx")
        assert r["vigencia_desde"] == datetime(2026, 4, 15)
        assert r["vigencia_hasta"] == datetime(2027, 4, 14)

    def test_detecta_3_hojas(self, excel_bytes):
        r = parsear_excel_tarifas(excel_bytes, "famisanar.xlsx")
        tipos = [h.split(":")[0] for h in r["hojas_detectadas"]]
        assert "ANEXO3" in tipos
        assert "ANEXO31" in tipos
        assert "ANEXO32" in tipos

    def test_total_filas(self, excel_bytes):
        r = parsear_excel_tarifas(excel_bytes, "famisanar.xlsx")
        # 2 CUPS + 2 medicamentos + 2 suministros
        assert len(r["filas"]) == 6

    def test_servicios_son_soat_porcentaje(self, excel_bytes):
        r = parsear_excel_tarifas(excel_bytes, "famisanar.xlsx")
        servicios = [f for f in r["filas"] if f["codigo_cups"].startswith("0101")]
        assert len(servicios) == 2
        for s in servicios:
            assert s["tipo_tarifa"] == "SOAT_PORCENTAJE"
            assert s["factor_ajuste"] == -5.0
            assert s["valor_pactado"] == 0.0

    def test_medicamentos_son_valor_fijo(self, excel_bytes):
        r = parsear_excel_tarifas(excel_bytes, "famisanar.xlsx")
        insulina = next((f for f in r["filas"] if f["codigo_cups"] == "19914262-04"), None)
        assert insulina is not None
        assert insulina["tipo_tarifa"] == "VALOR_FIJO"
        assert insulina["valor_pactado"] == 136746.0
        assert insulina["factor_ajuste"] == 0.0
        assert "INSULINA" in (insulina["descripcion"] or "")

    def test_suministros_valor_sin_iva(self, excel_bytes):
        r = parsear_excel_tarifas(excel_bytes, "famisanar.xlsx")
        sten = next((f for f in r["filas"] if f["codigo_cups"] == "FMQ6296"), None)
        assert sten is not None
        assert sten["tipo_tarifa"] == "VALOR_FIJO"
        # APLICA IVA = NO y TARIFA FINAL = "N/A" → usar TARIFA UNITARIA
        assert sten["valor_pactado"] == 2858499.0
        assert sten["modalidad"] == "SUMINISTROS CARDIOVASCULAR"

    def test_suministros_pinzas(self, excel_bytes):
        r = parsear_excel_tarifas(excel_bytes, "famisanar.xlsx")
        pinzas = next((f for f in r["filas"] if f["codigo_cups"] == "QX0106"), None)
        assert pinzas is not None
        assert pinzas["valor_pactado"] == 1660670.0

    def test_sin_errores_con_excel_bien_formado(self, excel_bytes):
        r = parsear_excel_tarifas(excel_bytes, "famisanar.xlsx")
        assert r["errores"] == []

    def test_archivo_invalido(self):
        r = parsear_excel_tarifas(b"not an xlsx", "bad.xlsx")
        assert r["filas"] == []
        assert r["errores"] != []


# ─── Formato Dispensario (hoja plana: CUPS + PRECIO DE REFERENCIA) ─────────

def _crear_excel_dispensario() -> bytes:
    """Layout plano tal como lo envía DISPENSARIO MEDICO BUCARAMANGA."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tarifas"
    # Sin metadata de encabezado: la tabla arranca en la fila 1 directamente
    headers = [
        "ITEM", "CUPS", "DESCRIPCION CUPS", "CODIGO IPS", "DESCRIPCION IPS",
        "PRECIO DE REFERENCIA", "TARIFA A LA QUE CORRESPONDE EL PRECIO DE REFERENCIA",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)

    # Filas de ejemplo reales del user
    datos = [
        (5081, "039001", "INSERCION DE CATETER EPIDURAL EN CANAL ESPINAL", "039001H",
         "INSERCION DE CATETER EPIDURAL EN CANAL ESPINAL", 1689585, "PROPIA"),
        (5082, "039002", "INSERCION DE CATETER EPIDURAL CON PUERTO IMPLANTABLE", "039002H",
         "INSERCION DE CATETER EPIDURAL CON PUERTO IMPLANTABLE", 1580085, "PROPIA"),
        (5086, "039204", "NEUROLISIS DE NERVIOS PERIFERICOS POR RADIOFRECUENCIA",
         "039204H", "NEUROLISIS DE NERVIOS PERIFERICOS", 5081895, "PROPIA"),
        (5087, "039306", "IMPLANTACION DE ELECTRODOS DE NEUROESTIMULACION ESPINAL",
         "039306H1", "IMPLANTACION DE ELECTRODOS", 8103986, "PROPIA"),
    ]
    for row_idx, tup in enumerate(datos, start=2):
        for col_idx, val in enumerate(tup, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestParsearExcelDispensario:
    @pytest.fixture
    def excel_bytes(self) -> bytes:
        return _crear_excel_dispensario()

    def test_detecta_simple_fijo(self, excel_bytes):
        r = parsear_excel_tarifas(excel_bytes, "dispensario.xlsx")
        tipos = [h.split(":")[0] for h in r["hojas_detectadas"]]
        assert "SIMPLE_FIJO" in tipos

    def test_numero_total_filas(self, excel_bytes):
        r = parsear_excel_tarifas(excel_bytes, "dispensario.xlsx")
        assert len(r["filas"]) == 4

    def test_valores_fijos(self, excel_bytes):
        r = parsear_excel_tarifas(excel_bytes, "dispensario.xlsx")
        for f in r["filas"]:
            assert f["tipo_tarifa"] == "VALOR_FIJO"
            assert f["factor_ajuste"] == 0.0

    def test_primer_cups_valor(self, excel_bytes):
        r = parsear_excel_tarifas(excel_bytes, "dispensario.xlsx")
        primero = next((f for f in r["filas"] if f["codigo_cups"] == "039001"), None)
        assert primero is not None
        assert primero["valor_pactado"] == 1689585.0
        assert "EPIDURAL" in (primero["descripcion"] or "")

    def test_modalidad_propia(self, excel_bytes):
        r = parsear_excel_tarifas(excel_bytes, "dispensario.xlsx")
        for f in r["filas"]:
            assert f["modalidad"] == "PROPIA"

    def test_codigo_ips_en_observacion(self, excel_bytes):
        r = parsear_excel_tarifas(excel_bytes, "dispensario.xlsx")
        primero = next((f for f in r["filas"] if f["codigo_cups"] == "039001"), None)
        # El código IPS "039001H" difiere del CUPS, debe guardarse
        assert "039001H" in (primero["observacion"] or "")

    def test_sin_eps_detectada(self, excel_bytes):
        """Dispensario no trae metadata de EPS; se espera eps=None."""
        r = parsear_excel_tarifas(excel_bytes, "dispensario.xlsx")
        assert r["eps"] is None


# ─── Casos reales de Famisanar: preamble largo + descripción fallback ──────

def _crear_excel_anexo3_con_preamble_largo() -> bytes:
    """Simula el Anexo 3 real donde hay ~55 filas de metadata antes del header."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Anexo 3"

    # Rellenar 55 filas de ruido (directorio, sedes, agrupadores, etc.)
    ws["A1"] = "DIRECTORIO DEL PRESTADOR"
    ws["A2"] = "CANALES DE ATENCIÓN AL USUARIO"
    ws["A15"] = "SEDES DE PRESTACIÓN DEL SERVICIO"
    ws["A20"] = "INFORMACIÓN EPS"
    ws["A21"] = "NOMBRE DE LA EPS"
    ws["F21"] = "FAMISANAR EPS"
    ws["A35"] = "AGRUPADORES OBJETO R. 2335/2023 - RIAS"
    ws["A36"] = "COD"
    ws["B36"] = "DESCRIPCIÓN"
    ws["C36"] = "MARCA"
    for i in range(12):
        ws.cell(row=37 + i, column=1, value=i + 1)
        ws.cell(row=37 + i, column=2, value=f"AGRUPADOR RIAS {i+1}")
    ws["A50"] = "AGRUPADORES OBJETO R. 2335/2023 - SERVICIOS"
    ws["A53"] = "DETALLE DE SERVICIOS CONTRATADOS (Servicios y Paquetes)"

    # Header real en fila 54 (más allá del viejo límite de 50)
    headers = [
        "CUPS / CUMS / MIPRES", "DESCRIPCIÓN CUPS / CUMS / MIPRES",
        "COD. REPS", "DESCRIPCIÓN REPS",
        "CÓDIGO PROPIO", "DESCRIPCIÓN CÓDIGO PROPIO",
        "INTERDEPENDENCIA", "TIPO TARIFA",
        "HOSPITALARIO", "AMBULATORIO", "URGENCIA",
        "CÓDIGO DE LA SEDE", "MARCA POR LISTADO SI/NO", "OBSERVACIÓN",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=54, column=i, value=h)
    # Dato real
    ws.cell(row=55, column=1, value="010101")
    ws.cell(row=55, column=2, value="PUNCION CISTERNAL VIA LATERAL")
    ws.cell(row=55, column=3, value=245)
    ws.cell(row=55, column=4, value="NEUROCIRUGIA")
    ws.cell(row=55, column=8, value="SOAT UVB VIGENTE")
    ws.cell(row=55, column=9, value="-5%")
    ws.cell(row=55, column=10, value="-5%")
    ws.cell(row=55, column=11, value="-5%")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestAnexo3ConPreambleLargo:
    def test_detecta_header_mas_alla_de_fila_50(self):
        data = _crear_excel_anexo3_con_preamble_largo()
        r = parsear_excel_tarifas(data, "famisanar.xlsx")
        tipos = [h.split(":")[0] for h in r["hojas_detectadas"]]
        assert "ANEXO3" in tipos, f"ANEXO3 no detectado. Hojas: {r['hojas_detectadas']}"

    def test_parsea_servicio_del_preamble_largo(self):
        data = _crear_excel_anexo3_con_preamble_largo()
        r = parsear_excel_tarifas(data, "famisanar.xlsx")
        servicios = [f for f in r["filas"] if f["codigo_cups"] == "010101"]
        assert len(servicios) == 1
        s = servicios[0]
        assert s["tipo_tarifa"] == "SOAT_PORCENTAJE"
        assert s["factor_ajuste"] == -5.0
        assert "PUNCION" in (s["descripcion"] or "")


def _crear_excel_medicamento_sin_desc_dci() -> bytes:
    """Medicamento donde 'DESCRIPCIÓN DCI' está vacía pero sí hay DESCRIPCIÓN general."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Anexo 3.1"
    ws["A3"] = "NOMBRE DE LA EPS"
    ws["F3"] = "FAMISANAR EPS"
    headers = [
        "CONSECUTIVO", "CÓDIGO DCI", "DESCRIPCIÓN DCI",
        "CÓDIGO DEL PRESTADOR", "COD. REPS", "DESCRIPCIÓN REPS",
        "MAPIISS", "CUM/IUM", "DESCRIPCIÓN",
        "AGRUPADOR", "TARIFA UNITARIA", "APLICA IVA (SI-NO)",
        "TIPO PLAN", "OBSERVACIONES",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=10, column=i, value=h)
    # Fila con DCI vacía pero DESCRIPCIÓN general llena
    ws.cell(row=11, column=1, value=1)
    ws.cell(row=11, column=2, value=None)  # DCI vacío
    ws.cell(row=11, column=3, value=None)  # DESC DCI vacío
    ws.cell(row=11, column=4, value="130105")
    ws.cell(row=11, column=6, value="SERVICIO FARMACÉUTICO")
    ws.cell(row=11, column=9, value="ACETAMINOFEN 500MG TABLETA")
    ws.cell(row=11, column=10, value="MEDICAMENTOS")
    ws.cell(row=11, column=11, value=42600)
    ws.cell(row=11, column=12, value="NO")
    # Fila donde DCI y DESCRIPCIÓN están vacías pero hay DESCRIPCIÓN REPS
    ws.cell(row=12, column=1, value=2)
    ws.cell(row=12, column=4, value="130106")
    ws.cell(row=12, column=6, value="SERVICIO FARMACÉUTICO - IBUPROFENO 400 MG")
    ws.cell(row=12, column=10, value="MEDICAMENTOS")
    ws.cell(row=12, column=11, value=20100)
    ws.cell(row=12, column=12, value="NO")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestDescripcionFallback:
    def test_fallback_dci_a_descripcion(self):
        data = _crear_excel_medicamento_sin_desc_dci()
        r = parsear_excel_tarifas(data, "med.xlsx")
        m1 = next((f for f in r["filas"] if f["codigo_cups"] == "130105"), None)
        assert m1 is not None
        assert m1["descripcion"] == "ACETAMINOFEN 500MG TABLETA"

    def test_fallback_dci_a_descripcion_reps(self):
        data = _crear_excel_medicamento_sin_desc_dci()
        r = parsear_excel_tarifas(data, "med.xlsx")
        m2 = next((f for f in r["filas"] if f["codigo_cups"] == "130106"), None)
        assert m2 is not None
        assert "IBUPROFENO" in (m2["descripcion"] or "")
