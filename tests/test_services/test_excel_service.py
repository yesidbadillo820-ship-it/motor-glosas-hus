"""Tests del exportador Excel — R51 P4."""
from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO

import pytest

from app.services.excel_service import EXCEL_DISPONIBLE, ExcelExporter


pytestmark = pytest.mark.skipif(
    not EXCEL_DISPONIBLE, reason="openpyxl no instalado"
)


class _GlosaFake:
    """Objeto minimal con los atributos que usa el exporter."""
    def __init__(self, **kw):
        for k, v in kw.items():
            setattr(self, k, v)


@pytest.fixture
def glosas_muestra():
    return [
        _GlosaFake(
            id=1, eps="FAMISANAR", paciente="PACIENTE A",
            factura="FE-001", numero_radicado="RAD-1",
            codigo_glosa="TA0201", valor_objetado=100000, valor_aceptado=0,
            estado="RADICADA", prioridad="ALTA",
            dias_restantes=3,
            creado_en=datetime(2026, 4, 1, tzinfo=timezone.utc),
        ),
        _GlosaFake(
            id=2, eps="SALUD TOTAL", paciente="PACIENTE B",
            factura="FE-002", numero_radicado="RAD-2",
            codigo_glosa="SO0101", valor_objetado=50000, valor_aceptado=50000,
            estado="LEVANTADA", prioridad="BAJA",
            dias_restantes=0,
            creado_en=datetime(2026, 4, 2, tzinfo=timezone.utc),
        ),
    ]


class TestReporteGlosas:
    def test_genera_bytes_io(self, glosas_muestra):
        out = ExcelExporter().generar_reporte_glosas(
            glosas_muestra, titulo="Test",
            fecha_inicio="2026-04-01", fecha_fin="2026-04-30",
        )
        assert isinstance(out, BytesIO)
        data = out.getvalue()
        # Excel (xlsx = zip-based), firma PK
        assert data[:2] == b"PK"
        assert len(data) > 1000  # no vacío

    def test_con_glosas_vacias_no_explota(self):
        out = ExcelExporter().generar_reporte_glosas([])
        assert isinstance(out, BytesIO)
        assert out.getvalue()[:2] == b"PK"

    def test_carga_el_workbook_tiene_hojas_esperadas(self, glosas_muestra):
        """Verifica que la hoja 'Glosas' tiene los datos sembrados."""
        from openpyxl import load_workbook
        out = ExcelExporter().generar_reporte_glosas(
            glosas_muestra, titulo="Reporte",
        )
        wb = load_workbook(out)
        assert "Glosas" in wb.sheetnames
        ws = wb["Glosas"]
        # Fila 5 = encabezados, fila 6+ = datos
        assert ws.cell(row=5, column=1).value == "ID"
        assert ws.cell(row=6, column=3).value == "FAMISANAR"
        assert ws.cell(row=7, column=3).value == "SALUD TOTAL"


class TestResumenMensual:
    def test_genera_reporte_tendencias(self):
        tendencias = [
            {"mes": "2026-03", "count": 10, "objetado": 500000,
             "aceptado": 400000, "recuperado": 400000},
            {"mes": "2026-04", "count": 8, "objetado": 300000,
             "aceptado": 100000, "recuperado": 100000},
        ]
        out = ExcelExporter().generar_resumen_mensual(tendencias, eps="FAMISANAR")
        assert isinstance(out, BytesIO)
        assert out.getvalue()[:2] == b"PK"


class TestFormatoFecha:
    def test_formato_datetime(self):
        e = ExcelExporter()
        dt = datetime(2026, 4, 25, 10, 30, tzinfo=timezone.utc)
        assert e._formato_fecha(dt) == "2026-04-25"

    def test_formato_none(self):
        assert ExcelExporter()._formato_fecha(None) == ""

    def test_formato_string_iso(self):
        assert ExcelExporter()._formato_fecha("2026-04-25T10:30:00Z") == "2026-04-25"

    def test_formato_string_invalido_se_devuelve_tal_cual(self):
        assert ExcelExporter()._formato_fecha("no-es-fecha") == "no-es-fecha"
