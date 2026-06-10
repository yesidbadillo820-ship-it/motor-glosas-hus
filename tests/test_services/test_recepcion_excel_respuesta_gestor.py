"""Excel-respuesta: columnas RESPUESTA IA / ESTADO IA / ID GLOSA /
GESTOR ASIGNADO y tolerancia a encabezados que no están en la fila 1.

El archivo real del DGH trae la hoja INICIAL con fila 1 de título
("ENTREGA GLOSA INICIAL") y los encabezados en la fila 2 (o más abajo,
con filas de banner). El generador debe localizar el header con la misma
ventana que el parser de importación (20 filas) y anotar las 4 columnas.
"""

from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from app.models.db import GlosaRecord
from app.services.recepcion_excel_response import (
    construir_respuestas_por_clave,
    generar_excel_con_respuestas,
)

_HEADERS = [
    "GESTOR",
    "ENTIDAD",
    "FACTURA",
    "CONSECUTIVO DGH",
    "VALOR GLOSA",
    "FECHA RECEPCION",
    "VENCE",
]


def _excel_original(filas, filas_basura_antes_del_header=1) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "INICIAL"
    for i in range(filas_basura_antes_del_header):
        ws.append([f"TITULO O BANNER {i + 1}"] + [""] * (len(_HEADERS) - 1))
    ws.append(_HEADERS)
    for f in filas:
        ws.append(f)
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


@pytest.fixture
def db():
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from app.database import Base

    engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=engine)
    s = sessionmaker(bind=engine)()
    try:
        yield s
    finally:
        s.close()


def _fila(factura, consecutivo, gestor="GESTOR PRUEBA"):
    return [gestor, "EPS DE PRUEBA", factura, consecutivo, "36402", "05/06/2026", "26/06/2026"]


def _glosa(db, factura, consecutivo, auditor_email=None, gestor_nombre="GESTOR PRUEBA"):
    g = GlosaRecord(
        eps="EPS DE PRUEBA",
        factura=factura,
        consecutivo_dgh=consecutivo,
        estado="RESPONDIDA",
        dictamen="Respuesta IA generada para la prueba",
        auditor_email=auditor_email,
        gestor_nombre=gestor_nombre,
    )
    db.add(g)
    db.commit()
    return g


def test_columna_gestor_asignado_presente_y_con_email(db):
    g = _glosa(db, "HUS0000000001", "100001", auditor_email="gestor.prueba@hus.gov.co")
    original = _excel_original([_fila("HUS0000000001", "100001")])

    respuestas = construir_respuestas_por_clave(db, [g.id])
    assert respuestas[("HUS0000000001", "100001")]["gestor_asignado"] == (
        "gestor.prueba@hus.gov.co"
    )

    xlsx = generar_excel_con_respuestas(original, respuestas)
    ws = load_workbook(BytesIO(xlsx))["INICIAL"]
    header = [c.value for c in ws[2]]  # header en fila 2 (fila 1 = título)
    assert header[-4:] == ["RESPUESTA IA", "ESTADO IA", "ID GLOSA", "GESTOR ASIGNADO"]
    fila_datos = [c.value for c in ws[3]]
    assert fila_datos[-4] == "Respuesta IA generada para la prueba"
    assert fila_datos[-3] == "RESPONDIDA"
    assert fila_datos[-2] == g.id
    assert fila_datos[-1] == "gestor.prueba@hus.gov.co"


def test_sin_asignacion_muestra_nombre_crudo_del_excel(db):
    g = _glosa(db, "HUS0000000002", "100002", auditor_email=None, gestor_nombre="GESTOR FANTASMA")
    original = _excel_original([_fila("HUS0000000002", "100002", gestor="GESTOR FANTASMA")])
    respuestas = construir_respuestas_por_clave(db, [g.id])
    xlsx = generar_excel_con_respuestas(original, respuestas)
    ws = load_workbook(BytesIO(xlsx))["INICIAL"]
    assert [c.value for c in ws[3]][-1] == "GESTOR FANTASMA"


def test_header_mas_alla_de_fila_5_tambien_se_anota(db):
    """Regresión: el generador usaba ventana de 5 filas (el parser de
    importación usa 20) — un INICIAL real con título+banners dejaba el
    Excel-respuesta SIN columnas IA."""
    g = _glosa(db, "HUS0000000003", "100003", auditor_email="gestor.prueba@hus.gov.co")
    original = _excel_original(
        [_fila("HUS0000000003", "100003")],
        filas_basura_antes_del_header=6,  # header queda en fila 7
    )
    respuestas = construir_respuestas_por_clave(db, [g.id])
    xlsx = generar_excel_con_respuestas(original, respuestas)
    ws = load_workbook(BytesIO(xlsx))["INICIAL"]
    header = [c.value for c in ws[7]]
    assert header[-4:] == ["RESPUESTA IA", "ESTADO IA", "ID GLOSA", "GESTOR ASIGNADO"]
    fila_datos = [c.value for c in ws[8]]
    assert fila_datos[-3] == "RESPONDIDA"
    assert fila_datos[-1] == "gestor.prueba@hus.gov.co"


def test_fila_sin_match_queda_no_procesada_y_gestor_vacio(db):
    original = _excel_original([_fila("HUS0000000099", "999999")])
    xlsx = generar_excel_con_respuestas(original, {})
    ws = load_workbook(BytesIO(xlsx))["INICIAL"]
    fila_datos = [c.value for c in ws[3]]
    assert fila_datos[-4] == "—"
    assert fila_datos[-3] == "NO_PROCESADA"
    assert fila_datos[-1] in ("", None)


def test_dictamen_html_sale_como_texto_plano_en_la_celda(db):
    """Regresión: la columna RESPUESTA IA traía `<table border="1"
    style=...>` literal. El dictamen HTML debe convertirse a texto
    legible (cabecera → 'CÓDIGO GLOSA: ... | VALOR OBJETADO: ...')."""
    g = _glosa(db, "HUS0000000004", "100004", auditor_email="gestor.prueba@hus.gov.co")
    g.dictamen = (
        '<table border="1" style="width:100%;border-collapse:collapse;">'
        "<tr><th>CÓDIGO GLOSA</th><th>VALOR OBJETADO</th><th>CÓDIGO RESPUESTA</th></tr>"
        "<tr><td>TA0801</td><td>$36.402</td><td><b>RE9901</b></td></tr></table>"
        '<div style="background:#f8fafc;">ARGUMENTACIÓN JURÍDICA<br/>'
        "El valor facturado corresponde al manual tarifario pactado.</div>"
    )
    db.commit()
    original = _excel_original([_fila("HUS0000000004", "100004")])
    respuestas = construir_respuestas_por_clave(db, [g.id])
    xlsx = generar_excel_con_respuestas(original, respuestas)
    ws = load_workbook(BytesIO(xlsx))["INICIAL"]
    celda = [c.value for c in ws[3]][-4]
    assert "<table" not in celda
    assert "style=" not in celda
    assert "CÓDIGO GLOSA: TA0801" in celda
    assert "VALOR OBJETADO: $36.402" in celda
    assert "CÓDIGO RESPUESTA: RE9901" in celda
    assert "El valor facturado corresponde al manual tarifario pactado." in celda
