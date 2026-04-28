"""Tests para los parsers FOMAG del Excel de tarifas contratadas.

El Excel de FOMAG (Acta No. 012, FIDUPREVISORA) tiene 5 hojas con
estructuras distintas — este test cubre las 3 formas que reconoce el
parser: ANEXO TARIFARIO, EXCLUIDOS (mismo formato), y PAQUETES.
"""
from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook

from app.services.tarifas_excel_parser import parsear_excel_tarifas


def _xlsx_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Cabecera real de FOMAG (ANEXO TARIFARIO + EXCLUIDOS comparten esta).
_HDR_FOMAG_TARIFARIO = [
    "NOMBRE PRESTADOR", "NIT", "CODIGO HABILITACION (12 DIGITOS)",
    "NOMBRE SEDE", "ATENCION", "DEPARTAMENTO", "MUNICIPIO",
    "CUPS RESOL 2641", "DESCRIPCION", "COD SERVICIO REPS",
    "NOMBRE SERVICIO REPS", "TARIFA SOAT PLENA 2025", "TARIFA SOAT  -20%",
    "TARIFAS PROPIAS", "TARIFAS ISS", "PROPUESTA IPS",
    "OBSERVACIÓN FOMAG", "RESOLUCIÓN", "REPS", "CRUCE",
    "TECHO", "COMPARACIÓN TECHO", "% MAYOR QUE TECHOS", "REPS2",
]

_HDR_FOMAG_AMBULATORIO = _HDR_FOMAG_TARIFARIO[:7] + [
    "CODIGO INSTITUCIONAL"
] + _HDR_FOMAG_TARIFARIO[7:]

_HDR_FOMAG_PAQUETES = [
    "MUNICIPIO", "RAZON SOCIAL", "NIT", "CODIGO", "SEDE",
    "PAQUETES DE GASTROENTEROLOGIA", "CUPS", "DESCRIPCION DEL CUPS",
    "VALOR PROPUESTO", "OBSERVACIONES",
]


def _fila_tarifario(cups: str, descripcion: str, propuesta: float, techo: float,
                    observacion: str, comparacion: str = "SOAT - 20%") -> list:
    return [
        "HOSPITAL UNIVERSITRIO DE SANTANDER", "900006037", "680010079201",
        "HOSPITAL UNIVERSITRIO DE SANTANDER", "", "SANTANDER", "BUCARAMANGA",
        cups, descripcion, "", "",
        propuesta * 1.25,  # SOAT pleno aprox
        propuesta,  # SOAT -20%
        "N/A", "N/A",
        propuesta,  # PROPUESTA IPS
        observacion,
        "", cups, "", techo, comparacion, "0%", "",
    ]


def _fila_paquete(codigo: str, cups: str, descripcion: str, valor: float,
                  obs: str = "INCLUYE: insumos, medicamentos") -> list:
    return [
        "BUCARAMANGA", "E.S.E. HOSPITAL UNIVERSITARIO DE SANTANDER",
        "900006037-4", codigo, "HOSPITAL UNIVERSITRIO DE SANTANDER",
        "", cups, descripcion, valor, obs,
    ]


class TestParserFomagTarifario:
    def test_detecta_y_carga_pactadas(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "ANEXO TARIFARIO"
        ws.append(_HDR_FOMAG_TARIFARIO)
        ws.append(_fila_tarifario(
            "010101", "PUNCION CISTERNAL, VIA LATERAL",
            732041, 732240, "SE ACEPTA TARIFA PROPUESTA POR LA IPS",
        ))
        ws.append(_fila_tarifario(
            "010103", "PUNCION CISTERNAL",
            732041, 732240, "SE ACEPTA TARIFA PROPUESTA POR LA IPS",
        ))
        resultado = parsear_excel_tarifas(_xlsx_bytes(wb), "fomag_test.xlsx")

        assert "FOMAG_TARIFARIO:ANEXO TARIFARIO" in resultado["hojas_detectadas"]
        assert resultado["eps"] == "FOMAG"
        assert len(resultado["filas"]) == 2
        f = resultado["filas"][0]
        assert f["codigo_cups"] == "010101"
        assert f["valor_pactado"] == 732041.0
        assert "FOMAG" in f["modalidad"]
        assert f["tipo_tarifa"] == "VALOR_FIJO"

    def test_excluidos_usan_techo_no_propuesta(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "EXCLUIDOS"
        ws.append(_HDR_FOMAG_TARIFARIO)
        # Propuesta IPS muy por encima del techo FOMAG
        ws.append(_fila_tarifario(
            "010902", "OTRA PUNCION CRANEAL",
            propuesta=3113506, techo=732240,
            observacion="TARIFA POR ENCIMA DEL TECHO",
        ))
        resultado = parsear_excel_tarifas(_xlsx_bytes(wb), "fomag.xlsx")

        assert len(resultado["filas"]) == 1
        f = resultado["filas"][0]
        # Debe usar el techo, no la propuesta
        assert f["valor_pactado"] == 732240.0
        assert "EXCLUIDO" in f["modalidad"]
        # Observación deja constancia del exceso
        assert "excede techo FOMAG" in (f["observacion"] or "")

    def test_ambulatorio_guarda_codigo_institucional(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "ANEXO TARIFARIO AMBULATORIOS"
        ws.append(_HDR_FOMAG_AMBULATORIO)
        fila = ([
            "HOSPITAL UNIVERSITRIO DE SANTANDER", "900006037", "680010079201",
            "HOSPITAL UNIVERSITRIO DE SANTANDER", "AMBULATORIA",
            "SANTANDER", "BUCARAMANGA",
            "902210AMB",  # CODIGO INSTITUCIONAL
            "902210", "HEMOGRAMA IV AUTOMATIZADO", "", "LABORATORIO",
            22000.0, 17600.0, "N/A", "N/A", 17600.0,
            "SE ACEPTA TARIFA PROPUESTA POR LA IPS",
            "", "902210", "706", 31520, "SOAT -20%", "79%", "706",
        ])
        ws.append(fila)
        resultado = parsear_excel_tarifas(_xlsx_bytes(wb), "fomag_amb.xlsx")
        assert len(resultado["filas"]) == 1
        f = resultado["filas"][0]
        assert f["codigo_cups"] == "902210"
        assert f["codigo_ips"] == "902210AMB"
        assert f["valor_pactado"] == 17600.0


class TestParserFomagPaquetes:
    def test_detecta_paquetes_con_codigo_institucional(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "PAQUETES"
        ws.append(_HDR_FOMAG_PAQUETES)
        ws.append(_fila_paquete(
            "423301H", "423301", "POLIPECTOMIA DE ESOFAGO VIA ENDOSCOPICA",
            3211891.0,
        ))
        ws.append(_fila_paquete(
            "511001H", "511001", "COLANGIOPANCREATOGRAFIA RETROGRADA ENDOSCOPICA",
            6932445.0,
        ))
        resultado = parsear_excel_tarifas(_xlsx_bytes(wb), "fomag.xlsx")
        assert "FOMAG_PAQUETES:PAQUETES" in resultado["hojas_detectadas"]
        assert resultado["eps"] == "FOMAG"
        assert len(resultado["filas"]) == 2
        f = resultado["filas"][0]
        assert f["codigo_cups"] == "423301"
        assert f["codigo_ips"] == "423301H"
        assert f["valor_pactado"] == 3211891.0
        assert "PAQUETE" in f["modalidad"]

    def test_subtitulo_de_seccion_actualiza_categoria(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "PAQUETES"
        ws.append(_HDR_FOMAG_PAQUETES)
        ws.append(_fila_paquete("423301H", "423301", "POLIPECTOMIA", 3211891.0))
        # Fila de subtítulo (sin CUPS ni VALOR) que cambia la categoría
        ws.append([None, None, None, None, None, "DE COLUMNA",
                   None, None, None, None])
        ws.append(_fila_paquete(
            "810001H", "810001",
            "CORRECCION DE DEFORMIDAD HASTA SEIS VERTEBRAS", 83876891.0,
        ))
        resultado = parsear_excel_tarifas(_xlsx_bytes(wb), "fomag.xlsx")
        assert len(resultado["filas"]) == 2
        # La segunda fila debe llevar la nueva categoría en modalidad
        cat_columna = resultado["filas"][1]["modalidad"]
        assert "COLUMNA" in cat_columna.upper()
