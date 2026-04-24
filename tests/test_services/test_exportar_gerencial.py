"""Tests del export Excel gerencial (Ronda 24)."""
from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app.models.db import GlosaRecord
from app.services.exportar_gerencial import (
    _color_estado,
    generar_reporte_gerencial,
)


@pytest.fixture
def db_session():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    s = Session()
    try:
        yield s
    finally:
        s.close()


class TestColorEstado:
    def test_lista_enviar_verde(self):
        assert _color_estado("LISTA_ENVIAR") == "dcfce7"

    def test_intervenir_rojo_claro(self):
        assert _color_estado("INTERVENIR") == "fee2e2"

    def test_desconocido_blanco(self):
        assert _color_estado("OTRO") == "FFFFFF"


class TestGenerarReporte:
    def test_bd_vacia_produce_excel_valido(self, db_session):
        buf = generar_reporte_gerencial(db_session, periodo="dia")
        assert isinstance(buf, BytesIO)
        buf.seek(0)
        wb = load_workbook(buf)
        assert set(wb.sheetnames) == {"Resumen", "Top EPS", "Autopilot", "Anomalías"}

    def test_incluye_kpis_en_hoja_resumen(self, db_session):
        db_session.add(GlosaRecord(
            eps="FAMISANAR EPS", paciente="X", factura="F-1",
            codigo_glosa="TA0201", valor_objetado=500_000,
            estado="PENDIENTE", creado_en=datetime.now(timezone.utc),
        ))
        db_session.commit()
        buf = generar_reporte_gerencial(db_session, periodo="dia")
        wb = load_workbook(buf)
        ws = wb["Resumen"]
        # Título institucional
        assert "ESE HUS" in (ws["A1"].value or "")
        # Debe haber al menos una celda con "Radicadas"
        textos = [
            c.value for row in ws.iter_rows(max_row=20) for c in row if c.value
        ]
        assert any("Radicadas" in str(t) for t in textos)

    def test_hoja_top_eps_tiene_cabecera(self, db_session):
        buf = generar_reporte_gerencial(db_session, periodo="dia")
        wb = load_workbook(buf)
        ws = wb["Top EPS"]
        assert ws["A1"].value == "#"
        assert ws["B1"].value == "EPS"

    def test_hoja_autopilot_con_una_glosa_pendiente(self, db_session):
        db_session.add(GlosaRecord(
            eps="SANITAS", paciente="X", factura="F",
            codigo_glosa="TA0201", valor_objetado=100_000,
            estado="PENDIENTE",
            dictamen="<p>Ley 1438/2011 Art 57. Resolución 2284/2023.</p>" * 20,
            creado_en=datetime.now(timezone.utc),
        ))
        db_session.commit()
        buf = generar_reporte_gerencial(db_session, periodo="dia")
        wb = load_workbook(buf)
        ws = wb["Autopilot"]
        # Debe haber al menos una fila de datos (row 2+)
        assert ws["A2"].value is not None or ws["B2"].value is not None

    def test_hoja_anomalias_existe(self, db_session):
        buf = generar_reporte_gerencial(db_session, periodo="dia")
        wb = load_workbook(buf)
        ws = wb["Anomalías"]
        assert "Anomalías" in str(ws["A1"].value or "")
