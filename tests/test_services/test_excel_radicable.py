"""Tests del generador del «Excel radicable» (formato institucional que
se radica ante la entidad pagadora — jun-2026).

Verifica contra el layout acordado con el dueño: hoja RESPUESTA GLOSAS
(título merged + línea de contrato + header de 12 columnas + una fila
por concepto + TOTAL GLOSADO) y hoja FUNDAMENTO JURÍDICO con 6
secciones numeradas generadas desde los metadatos del contrato en BD.

Todos los datos de los fixtures son FICTICIOS (facturas FAC-TEST-*,
NITs 900.000.00x-x) — acá no va ningún dato real de glosas.
"""

from datetime import datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base
from app.models.db import (
    ClausulaContrato,
    ConceptoGlosaRecord,
    ContratoRecord,
    GlosaRecord,
)
from app.services.excel_radicable import (
    _decision_hus,
    _entidad_corta,
    generar_excel_radicable,
    generar_excel_radicable_con_nombre,
)

ENCABEZADOS_ESPERADOS = [
    "No.",
    "FACTURA",
    "CONSEC. DGH",
    "FECHA FACTURA",
    "VR FACTURA",
    "CÓD. GLOSA",
    "CÓD. SERVICIO",
    "DESCRIPCIÓN SERVICIO",
    "VR GLOSADO",
    "OBSERVACIÓN AUDITORÍA DMBUG",
    "DECISIÓN HUS",
    "RESPUESTA ESE HUS",
]


@pytest.fixture
def db_session():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    s = sessionmaker(bind=engine)()
    try:
        yield s
    finally:
        s.close()
        engine.dispose()


def _sembrar_lote(db):
    """2 glosas: una con 2 conceptos, otra sin conceptos. Datos fake."""
    db.add(
        ContratoRecord(
            eps="DISPENSARIO MEDICO",
            detalles="CONTRATO TEST",
            numero_contrato="CONTRATO INTERADMINISTRATIVO No. 999-TEST/DM-2026",
            nit_eps="900.000.001-1",
            nit_ips="900.000.002-2",
            razon_social_eps="DISPENSARIO MEDICO DE PRUEBA",
            numero_proceso_secop="CD000-TEST-2026",
            fecha_suscripcion=datetime(2025, 12, 20),
            fecha_inicio=datetime(2026, 1, 1),
            fecha_fin=datetime(2026, 7, 30),
            anexos_descripcion='[{"nombre": "ANEXO No. 1", "descripcion": "TARIFARIO CUPS"}]',
            modalidades_tarifarias='["SOAT -20%", "TARIFA PROPIA ESE"]',
        )
    )
    db.add(
        ClausulaContrato(
            eps="DISPENSARIO MEDICO",
            numero_clausula="SEGUNDA",
            tema="TA",
            titulo="TARIFAS Y FORMA DE PAGO",
            texto_literal="Las tarifas pactadas se rigen por el ANEXO No. 1 del contrato.",
            pagina=3,
        )
    )
    # Glosa A — 2 conceptos (uno con respuesta propia, otro sin)
    db.add(
        GlosaRecord(
            id=1,
            eps="DISPENSARIO MEDICO",
            paciente="N/A",
            factura="FAC-TEST-001",
            consecutivo_dgh="100001",
            codigo_glosa="TA0801",
            valor_objetado=80000.0,
            valor_factura=200000.0,
            fecha_documento_dgh=datetime(2026, 4, 10),
            fecha_entrega=datetime(2026, 6, 10),
            estado="RESPONDIDA",
            dictamen="<p>Defensa glosa A por tarifas pactadas.</p>",
            codigo_respuesta="RE9901",
            score=0.9,
            observacion_eps="OBSERVACION GENERAL A",
        )
    )
    db.add(
        ConceptoGlosaRecord(
            id=11,
            glosa_id=1,
            factura="FAC-TEST-001",
            consecutivo_dgh="100001",
            codigo_glosa="TA0801",
            cups_codigo="111111",
            cups_descripcion="ECOGRAFIA DE PRUEBA",
            valor_objetado=50000.0,
            observacion_eps="MVC EN SERVICIO 111111",
            dictamen_html="<p>Respuesta concepto 1 con tarifa del anexo.</p>",
        )
    )
    db.add(
        ConceptoGlosaRecord(
            id=12,
            glosa_id=1,
            factura="FAC-TEST-001",
            consecutivo_dgh="100001",
            codigo_glosa="TA0802",
            cups_codigo="222222",
            cups_descripcion="LABORATORIO DE PRUEBA",
            valor_objetado=30000.0,
            observacion_eps="MVC EN SERVICIO 222222",
            dictamen_html=None,  # → debe caer al dictamen de la glosa
        )
    )
    # Glosa B — sin conceptos cargados, requiere soportes
    db.add(
        GlosaRecord(
            id=2,
            eps="DISPENSARIO MEDICO",
            paciente="N/A",
            factura="FAC-TEST-002",
            consecutivo_dgh="100002",
            codigo_glosa="SO0301",
            valor_objetado=70000.0,
            valor_factura=90000.0,
            fecha_documento_dgh=datetime(2026, 5, 2),
            fecha_entrega=datetime(2026, 6, 10),
            estado="REQUIERE_SOPORTES",
            dictamen="<p>Defensa glosa B condicionada a HC.</p>",
            codigo_respuesta="RE9901",
            cups_servicio="333333",
            servicio_descripcion="SERVICIO B DE PRUEBA",
            observacion_eps="FALTA SOPORTE DE HC",
        )
    )
    db.commit()


def _abrir(contenido: bytes):
    return load_workbook(BytesIO(contenido))


def test_estructura_hoja_respuesta_glosas(db_session):
    _sembrar_lote(db_session)
    wb = _abrir(generar_excel_radicable(db_session, glosa_ids=[1, 2]))

    assert wb.sheetnames == ["RESPUESTA GLOSAS", "FUNDAMENTO JURÍDICO"]
    ws = wb["RESPUESTA GLOSAS"]

    # Fila 1: título institucional merged con la entidad y su NIT (fake)
    assert "A1:L1" in {str(r) for r in ws.merged_cells.ranges}
    titulo = ws.cell(1, 1).value
    assert titulo.startswith("ESE HOSPITAL UNIVERSITARIO DE SANTANDER – RESPUESTA A GLOSA INICIAL")
    assert "DISPENSARIO MEDICO DE PRUEBA" in titulo
    assert "900.000.001-1" in titulo

    # Fila 2: línea del contrato desde ContratoRecord + fecha de entrega
    linea = ws.cell(2, 1).value
    assert "999-TEST/DM-2026" in linea
    assert "ENTREGA GLOSA INICIAL 10/06/2026" in linea
    assert "RESPUESTA ELABORADA" in linea

    # Fila 3: header con las 12 columnas exactas
    assert [ws.cell(3, c).value for c in range(1, 13)] == ENCABEZADOS_ESPERADOS

    # Una fila por concepto: 2 (glosa A) + 1 (glosa B sin conceptos) = 3
    assert ws.cell(4, 2).value == "FAC-TEST-001"
    assert ws.cell(4, 7).value == "111111"
    assert ws.cell(5, 7).value == "222222"
    assert ws.cell(6, 2).value == "FAC-TEST-002"
    assert ws.cell(6, 7).value == "333333"  # cae a cups_servicio de la glosa
    assert ws.cell(6, 8).value == "SERVICIO B DE PRUEBA"
    # No. consecutivo por fila y consecutivo DGH numérico
    assert [ws.cell(r, 1).value for r in (4, 5, 6)] == [1, 2, 3]
    assert ws.cell(4, 3).value == 100001
    # FECHA FACTURA en ISO y VR FACTURA repetido por concepto
    assert ws.cell(4, 4).value == "2026-04-10"
    assert ws.cell(4, 5).value == 200000
    assert ws.cell(5, 5).value == 200000

    # Fila TOTAL GLOSADO: suma de VR GLOSADO (50k + 30k + 70k)
    assert ws.cell(7, 8).value == "TOTAL GLOSADO"
    assert ws.cell(7, 9).value == 150000
    assert ws.cell(7, 9).number_format == "#,##0"

    # Estilos clave del formato del dueño
    assert ws.cell(3, 1).fill.fgColor.rgb == "FF1B5E3F"
    assert ws.cell(3, 1).font.color.rgb == "FFFFFFFF"
    assert ws.cell(1, 1).fill.fgColor.rgb == "FF1B5E3F"
    assert ws.column_dimensions["J"].width == 42
    assert ws.column_dimensions["L"].width == 95
    assert ws.cell(4, 12).alignment.wrap_text is True
    assert ws.cell(4, 12).alignment.vertical == "top"
    assert ws.cell(4, 1).border.left.style == "thin"
    assert ws.freeze_panes == "A4"


def test_respuesta_concepto_y_fallback_a_dictamen_glosa(db_session):
    _sembrar_lote(db_session)
    wb = _abrir(generar_excel_radicable(db_session, glosa_ids=[1, 2]))
    ws = wb["RESPUESTA GLOSAS"]

    # Concepto con dictamen propio → su texto plano (sin tags HTML)
    assert ws.cell(4, 12).value == "Respuesta concepto 1 con tarifa del anexo."
    # Concepto sin dictamen, código DISTINTO al principal y dictamen sin
    # sección para él: desde el fix del 12-jun-2026 (filas duplicadas en
    # el export) ya no se repite el dictamen completo — nota corta que
    # remite a la fila del código principal.
    assert "dictamen integral de la glosa #1" in ws.cell(5, 12).value
    assert "código principal TA0801" in ws.cell(5, 12).value
    assert "<p>" not in ws.cell(5, 12).value
    # Observación de la EPS por concepto
    assert ws.cell(4, 10).value == "MVC EN SERVICIO 111111"
    # Glosa sin conceptos → observación general de la glosa
    assert ws.cell(6, 10).value == "FALTA SOPORTE DE HC"


def test_decision_hus_en_filas(db_session):
    _sembrar_lote(db_session)
    wb = _abrir(generar_excel_radicable(db_session, glosa_ids=[1, 2]))
    ws = wb["RESPUESTA GLOSAS"]

    # Glosa A: defensa RE9901 con confianza alta → NO ACEPTA (rojo bold)
    assert ws.cell(4, 11).value == "NO ACEPTA"
    assert ws.cell(5, 11).value == "NO ACEPTA"
    assert ws.cell(4, 11).font.color.rgb == "FFC00000"
    assert ws.cell(4, 11).font.bold is True
    # Glosa B: REQUIERE_SOPORTES → marca de validación antes de radicar
    assert ws.cell(6, 11).value == "NO ACEPTA (VALIDAR SOPORTE HC ANTES DE RADICAR)"


def test_glosa_pendiente_sin_dictamen(db_session):
    db_session.add(
        GlosaRecord(
            id=3,
            eps="EPS DE PRUEBA XYZ",
            paciente="N/A",
            factura="FAC-TEST-003",
            consecutivo_dgh="100003",
            codigo_glosa="TA0101",
            valor_objetado=10000.0,
            estado="RADICADA",
            dictamen=None,
        )
    )
    db_session.commit()
    wb = _abrir(generar_excel_radicable(db_session, glosa_ids=[3]))
    ws = wb["RESPUESTA GLOSAS"]
    assert ws.cell(4, 12).value == "PENDIENTE DE ANÁLISIS — RADICADA"
    assert ws.cell(4, 11).value == "PENDIENTE"


def test_mapeo_decision_codigos_re():
    """Coherente con el catálogo RESPUESTAS_IPS (catalogo_glosas):
    RE97xx acepta 100%, RE98xx acepta parcial, RE99/95/96 defienden."""
    assert _decision_hus("RE9901", "RESPONDIDA", 0.9, True) == "NO ACEPTA"
    assert _decision_hus("RE9502", "RESPONDIDA", 0.9, True) == "NO ACEPTA"
    assert _decision_hus("RE9602", "RESPONDIDA", 0.9, True) == "NO ACEPTA"
    assert _decision_hus("RE9801", "RESPONDIDA", 0.9, True) == "ACEPTA PARCIAL"
    assert _decision_hus("RE9701", "RESPONDIDA", 0.9, True) == "ACEPTA"
    assert _decision_hus("RE9702", "RESPONDIDA", 0.9, True) == "ACEPTA"
    # REQUIERE_SOPORTES o confianza IA baja → validar HC antes de radicar
    esperado = "NO ACEPTA (VALIDAR SOPORTE HC ANTES DE RADICAR)"
    assert _decision_hus("RE9901", "REQUIERE_SOPORTES", 0.9, True) == esperado
    assert _decision_hus("RE9901", "RESPONDIDA", 0.30, True) == esperado
    # score 0.0 es «sin calificar», no confianza baja
    assert _decision_hus("RE9901", "RESPONDIDA", 0.0, True) == "NO ACEPTA"
    assert _decision_hus("", "RADICADA", None, False) == "PENDIENTE"


def test_hoja_fundamento_juridico_seis_secciones(db_session):
    _sembrar_lote(db_session)
    wb = _abrir(generar_excel_radicable(db_session, glosa_ids=[1, 2]))
    ws = wb["FUNDAMENTO JURÍDICO"]

    assert ws.column_dimensions["A"].width == 34
    assert ws.column_dimensions["B"].width == 130
    assert "A1:B1" in {str(r) for r in ws.merged_cells.ranges}
    assert "FUNDAMENTO JURÍDICO GENERAL" in ws.cell(1, 1).value
    assert "DMBUG" in ws.cell(1, 1).value
    assert ws.cell(1, 1).fill.fgColor.rgb == "FF1B5E3F"

    # 6 secciones numeradas en filas 3..8 (fila 2 en blanco)
    titulos = [ws.cell(r, 1).value for r in range(3, 9)]
    assert [t.split(".")[0] for t in titulos] == ["1", "2", "3", "4", "5", "6"]
    assert titulos[0] == "1. EXISTENCIA Y VIGENCIA DEL CONTRATO"
    assert titulos[3] == "4. EL AGOTAMIENTO PRESUPUESTAL NO EQUIVALE A AUSENCIA DE CONTRATO"
    assert titulos[5] == "6. TRÁMITE DE LA GLOSA"
    # Vigencia dinámica según la fecha de entrega del lote
    assert titulos[2] == "3. AJUSTE TARIFARIO VIGENCIA 2026"

    textos = {r: ws.cell(r, 2).value for r in range(3, 9)}
    # Sección 1: generada desde los metadatos REALES del ContratoRecord fake
    assert "999-TEST/DM-2026" in textos[3]
    assert "900.000.001-1" in textos[3]
    assert "CD000-TEST-2026" in textos[3]
    assert "FACTURAS DE ABRIL A MAYO DE 2026" in textos[3]
    # Sección 2: anexos + modalidades + cláusula tarifaria de BD
    assert "ANEXO NO. 1 (TARIFARIO CUPS)" in textos[4].upper()
    assert "SOAT -20%" in textos[4]
    assert "CLÁUSULA SEGUNDA" in textos[4]
    assert "111111" in textos[4]  # CUPS del lote
    # Sección 4: doctrina del agotamiento (cadena tripartita)
    assert "ART. 83" in textos[6]
    assert "1751 DE 2015" in textos[6]
    # Sección 5: SOAT pleno + régimen militar (DMBUG) sin instrucciones IA
    assert "DECRETO 4747 DE 2007" in textos[7]
    assert "DECRETO 1795 DE 2000" in textos[7]
    assert "NO CITES" not in textos[7].upper()
    # Sección 6: trámite con Art. 57 Ley 1438/2011 y plazo
    assert "57 DE LA LEY 1438 DE 2011" in textos[8]
    assert "QUINCE (15) DÍAS HÁBILES" in textos[8]
    assert "MESA DE CONCILIACIÓN" in textos[8]
    # Estilos de sección
    assert ws.cell(3, 1).font.color.rgb == "FF1B5E3F"
    assert ws.cell(3, 2).alignment.horizontal == "justify"


def test_fundamento_sin_metadatos_no_inventa_numeros(db_session):
    """EPS sin ContratoRecord ni catálogo → secciones genéricas, sin
    números de contrato/NIT inventados."""
    db_session.add(
        GlosaRecord(
            id=5,
            eps="PAGADOR GENERICO ZZZ",
            paciente="N/A",
            factura="FAC-TEST-005",
            consecutivo_dgh="100005",
            codigo_glosa="TA0101",
            valor_objetado=5000.0,
            estado="RESPONDIDA",
            dictamen="<p>Defensa.</p>",
            codigo_respuesta="RE9901",
        )
    )
    db_session.commit()
    wb = _abrir(generar_excel_radicable(db_session, glosa_ids=[5]))
    ws = wb["FUNDAMENTO JURÍDICO"]
    s1 = ws.cell(3, 2).value
    assert "PAGADOR GENERICO ZZZ" in s1
    assert "SE SUSCRIBIÓ" not in s1  # sin contrato no se afirma suscripción
    assert "NIT" not in s1.split("ESE HOSPITAL")[0]  # sin NIT de la EPS inventado
    # La hoja RESPUESTA omite el segmento de contrato pero mantiene la fila
    linea = wb["RESPUESTA GLOSAS"].cell(2, 1).value
    assert "CONTRATO" not in linea
    assert "ENTREGA GLOSA INICIAL" in linea


def test_nombre_archivo_institucional(db_session):
    _sembrar_lote(db_session)
    contenido, nombre = generar_excel_radicable_con_nombre(db_session, glosa_ids=[1, 2])
    assert nombre == "RESPUESTA_GLOSAS_DMBUG_10JUN2026_HUS.xlsx"
    assert _abrir(contenido).sheetnames == ["RESPUESTA GLOSAS", "FUNDAMENTO JURÍDICO"]


def test_entidad_corta_variantes():
    assert _entidad_corta("U220311 - DIRECCION DE SANIDAD EJERCITO - DISPENSARIO MEDICO") == "DMBUG"
    assert _entidad_corta("FOMAG - FONDO DEL MAGISTERIO") == "FOMAG"
    assert _entidad_corta("EPS FAMISANAR SAS") == "FAMISANAR"
    assert _entidad_corta("") == "ENTIDAD"


def test_respuesta_truncada_a_limite_excel(db_session):
    """Dictámenes monstruo no deben romper el límite de 32.767 chars."""
    db_session.add(
        GlosaRecord(
            id=7,
            eps="EPS DE PRUEBA XYZ",
            paciente="N/A",
            factura="FAC-TEST-007",
            consecutivo_dgh="100007",
            codigo_glosa="TA0101",
            valor_objetado=1000.0,
            estado="RESPONDIDA",
            dictamen="<p>" + ("ARGUMENTO MUY LARGO. " * 3000) + "</p>",
            codigo_respuesta="RE9901",
        )
    )
    db_session.commit()
    wb = _abrir(generar_excel_radicable(db_session, glosa_ids=[7]))
    celda = wb["RESPUESTA GLOSAS"].cell(4, 12).value
    assert len(celda) <= 32000
    assert "TRUNCADO" in celda
