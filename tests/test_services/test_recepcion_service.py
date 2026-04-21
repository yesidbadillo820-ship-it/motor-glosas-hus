"""Tests for RecepcionService (import Excel de recepción + conceptos DGH)."""
from io import BytesIO
from datetime import datetime

import pytest
from openpyxl import Workbook

from app.services.recepcion_service import (
    RecepcionService,
    _fix_mojibake,
    _split_entidad,
    _a_fecha,
    _buscar_fila_encabezado,
    COLUMN_ALIASES,
    CONCEPTO_COLS,
)


class TestMojibake:
    def test_latin1_utf8_bug_recuperable(self):
        # UTF-8 "ñ" = 0xC3 0xB1 — leído como Latin-1 da "Ã±". Recuperable.
        assert _fix_mojibake("AÃ±alisis") == "Añalisis"
        # UTF-8 "Ó" = 0xC3 0x93 — pero 0x93 en Latin-1 no es imprimible y
        # Python lo arroja por la decodificación estricta. Usamos "ó" (0xC3 0xB3).
        assert _fix_mojibake("CÃ³DIGO") == "CóDIGO"

    def test_mojibake_lossy_no_corrompe(self):
        # Caso degradado del DGH: "OBJECIÃ¿N" viene de conversión lossy
        # donde el accent se perdió. La función no debe reventar — devuelve
        # algo consistente (latin1→utf8 da "ÿ") en lugar de crashear.
        resultado = _fix_mojibake("OBJECIÃ¿N")
        assert resultado is not None
        assert isinstance(resultado, str)

    def test_texto_limpio_no_cambia(self):
        assert _fix_mojibake("GONADOTROPINA") == "GONADOTROPINA"
        assert _fix_mojibake("") == ""
        assert _fix_mojibake(None) is None

    def test_texto_con_a_acentuada_valida_no_cambia(self):
        # "á" solita no es mojibake — se respeta
        assert _fix_mojibake("Análisis") == "Análisis"


class TestSplitEntidad:
    def test_formato_codigo_plus_nombre(self):
        assert _split_entidad("U220181 - FAMISANAR EPS SUBSIDIADO") == (
            "U220181", "FAMISANAR EPS SUBSIDIADO"
        )
        assert _split_entidad("C230051 - SANITAS S.A.S. EPS CONTRIBUTIVO") == (
            "C230051", "SANITAS S.A.S. EPS CONTRIBUTIVO"
        )

    def test_sin_codigo(self):
        assert _split_entidad("FAMISANAR EPS") == ("", "FAMISANAR EPS")

    def test_vacio(self):
        assert _split_entidad("") == ("", "")
        assert _split_entidad(None) == ("", "")


class TestFechaParser:
    def test_formato_dd_mm_yyyy(self):
        assert _a_fecha("21/04/2026") == datetime(2026, 4, 21)

    def test_formato_con_hora(self):
        # Formato del DGH: "14/04/2026 15:27"
        assert _a_fecha("14/04/2026 15:27") == datetime(2026, 4, 14, 15, 27)

    def test_formato_con_hora_segundos(self):
        assert _a_fecha("14/04/2026 15:27:33") == datetime(2026, 4, 14, 15, 27, 33)

    def test_iso(self):
        assert _a_fecha("2026-04-21") == datetime(2026, 4, 21)

    def test_invalido(self):
        assert _a_fecha("no-es-fecha") is None
        assert _a_fecha(None) is None


# ─── Helpers para construir workbooks in-memory ──────────────────────────────

def _hoja_inicial(ws, filas_datos):
    """Poblar hoja INICIAL: fila 1 título, fila 2 headers, fila 3+ datos."""
    ws.append(["ENTREGA GLOSA INICIAL"] + [""] * 17)
    ws.append([
        "GESTOR", "FECHA DE ENTREGA", "FECHA RADICACION", "FECHA DOCUMENTO DGH",
        "FECHA RECEPCION", "ENTIDAD", "FACTURA", "CONSECUTIVO DGH", "VALOR GLOSA",
        "VENCE", "DEVOLUCION S/N", "DIAS RADICACION VS RECEPCION", "RADICADO",
        "REFERENCIA", "OBSERVACION TECNICO", "TECNICO QUE RECEPCIONO",
        "TIPO GLOSA", "PROFESIONAL(MEDICO)",
    ])
    for f in filas_datos:
        ws.append(f)


def _hoja_ratificada(ws, filas_datos):
    ws.append(["ENTREGA GLOSA RATIFICADA"] + [""] * 10)
    ws.append([
        "RESPONSABLE", "FECHA ENTREGA", "FECHA DE DOCUMENTO (DGH)",
        "FECHA NOTIFICACION OBJECIÓN", "EMPRESA", "NUMERO DE FACTURA",
        "CONSECUTIVO DGH", "VALOR GLOSA", "FECHA VENCIMIENTO",
        "OBSERVACION RECEPCION", "TECNICO QUE RECEPCIONO",
    ])
    for f in filas_datos:
        ws.append(f)


def _hoja_conceptos(ws, filas_datos):
    """Estructura de hojas I/R del DGH."""
    ws.append([
        "EstadoCxCObjecion", "TipoObjecionTramite", "Referencia",
        "FacturaCartera.PlanBeneficio.Contrato.Entidad.NombreEntidad",
        "FacturaCartera.Saldo", "UsuarioConfirmacion.Descripcion",
        "UsuarioConfirmacion.Nombre", "UsuarioCreacion.Descripcion",
        "UsuarioCreacion.Nombre",
        "FacturaCartera.PlanBeneficio.CodigoNombrePlanBeneficios",
        "FacturaCartera.PlanBeneficio.Contrato.Entidad.CodigoEntidad",
        "FacturaCartera.Factura", "FechaDocumento", "Consecutivo",
        "Observaciones", "EstadoActual", "FacturaCartera.Valor",
        "FacturaCartera.Fecha", "FacturaCartera.Tercero.Documento",
        "FacturaCartera.Tercero.NombreCompletoAN", "FechaObjecion",
        "ConceptoObjecion.Codigo", "Oid", "ConceptoObjecion.Nombre",
        "ValorObjecion", "FacturaCartera.Tercero.NombreCompletoNA",
        "ListadoConceptos.ConceptoObjecion.Codigo", "ListadoConceptos.Oid",
        "ListadoConceptos.ConceptoObjecion.Nombre",
        "ListadoConceptos.ServicioProductoFactura.Codigo",
        "ListadoConceptos.ServicioProductoFactura.Descripcion",
        "ListadoConceptos.ValorObjecion",
        "ListadoConceptos.ServicioProductoFactura.CentroCosto.CodigoNombreCentro",
        "ListadoConceptos.Observaciones",
    ])
    for f in filas_datos:
        ws.append(f)


def _bytes_wb(wb) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Fixtures ────────────────────────────────────────────────────────────────

@pytest.fixture
def db():
    """DB SQLite en memoria con todas las tablas creadas."""
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from app.database import Base
    import app.models.db  # registra todas las tablas

    engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=engine)
    Session = sessionmaker(bind=engine)
    s = Session()
    try:
        yield s
    finally:
        s.close()


# ─── Tests de detección de tipo de hoja ──────────────────────────────────────

class TestDeteccionTipoHoja:
    def test_hoja_inicial_con_titulo_detecta_header_en_fila_2(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "INICIAL"
        _hoja_inicial(ws, [])
        fila, idx = _buscar_fila_encabezado(ws, max_filas=5, mapa=COLUMN_ALIASES, min_aciertos=3)
        assert fila == 2
        # Las columnas clave del INICIAL deben estar mapeadas
        assert "gestor" in idx
        assert "factura" in idx
        assert "vence" in idx
        assert "entidad" in idx
        assert "tecnico_recepcion" in idx

    def test_hoja_ratificada_usa_aliases_nuevos(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "RATIFICADA"
        _hoja_ratificada(ws, [])
        fila, idx = _buscar_fila_encabezado(ws, max_filas=5, mapa=COLUMN_ALIASES, min_aciertos=3)
        assert fila == 2
        # Alias: RESPONSABLE=gestor, EMPRESA=entidad, NUMERO DE FACTURA=factura,
        # FECHA VENCIMIENTO=vence, FECHA NOTIFICACION OBJECION=fecha_recepcion
        assert "gestor" in idx
        assert "entidad" in idx
        assert "factura" in idx
        assert "vence" in idx
        assert "fecha_recepcion" in idx

    def test_hoja_conceptos_se_detecta_por_columnas_dgh(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "I"
        _hoja_conceptos(ws, [])
        fila, idx = _buscar_fila_encabezado(ws, max_filas=5, mapa=CONCEPTO_COLS, min_aciertos=4)
        assert fila == 1
        assert "concepto_codigo" in idx
        assert "concepto_valor" in idx
        assert "factura" in idx
        assert "consecutivo" in idx


# ─── Tests end-to-end con workbook real ──────────────────────────────────────

class TestImportacionCompleta:
    def test_import_inicial_crea_glosa(self, db):
        wb = Workbook()
        ws = wb.active
        ws.title = "INICIAL"
        _hoja_inicial(ws, [
            [
                "EQUIPO ASEGURADORAS", "21/04/2026", "01/04/2026", "15/04/2026",
                "15/04/2026", "U220181 - FAMISANAR EPS SUBSIDIADO",
                "HUS0000487351", "168185", 758300, "05/05/2026",
                "N", 8, "", "", "", "ELIAS ALFONSO CARVAJAL NAVARRO",
                "Administrativo", "",
            ]
        ])
        r = RecepcionService(db).procesar_excel(_bytes_wb(wb))
        assert r.creadas == 1
        assert r.errores == []

        from app.models.db import GlosaRecord
        g = db.query(GlosaRecord).filter(GlosaRecord.factura == "HUS0000487351").first()
        assert g is not None
        assert g.consecutivo_dgh == "168185"
        assert g.gestor_nombre == "EQUIPO ASEGURADORAS"
        assert g.eps_codigo == "U220181"
        assert g.tecnico_recepcion == "ELIAS ALFONSO CARVAJAL NAVARRO"
        assert g.estado == "RADICADA"

    def test_import_ratificada_marca_estado(self, db):
        wb = Workbook()
        ws = wb.active
        ws.title = "RATIFICADA"
        _hoja_ratificada(ws, [
            [
                "KAREN ORTIZ", "21/04/2026", "16/04/2026", "16/04/2026",
                "U220641 - FUNDACION SALUD MIA EPS SUBSIDIADO", "HUS0000473035",
                "168313", 24437, "27/04/2026", "", "ELIAS ALFONSO CARVAJAL NAVARRO",
            ]
        ])
        r = RecepcionService(db).procesar_excel(_bytes_wb(wb))
        assert r.creadas == 1
        assert r.ratificadas == 1

        from app.models.db import GlosaRecord
        g = db.query(GlosaRecord).filter(GlosaRecord.factura == "HUS0000473035").first()
        assert g.estado == "RATIFICADA"

    def test_import_completo_4_hojas(self, db):
        """Caso real: 4 hojas (INICIAL, RATIFICADA, I, R) en un solo archivo."""
        wb = Workbook()
        wb.remove(wb.active)

        ws1 = wb.create_sheet("INICIAL")
        _hoja_inicial(ws1, [
            [
                "EQUIPO ASEGURADORAS", "21/04/2026", "01/04/2026", "15/04/2026",
                "15/04/2026", "U220181 - FAMISANAR EPS SUBSIDIADO",
                "HUS0000487351", "168185", 758300, "05/05/2026",
                "N", 8, "", "", "", "ELIAS ALFONSO CARVAJAL NAVARRO",
                "Administrativo", "",
            ]
        ])

        ws2 = wb.create_sheet("RATIFICADA")
        _hoja_ratificada(ws2, [
            [
                "KAREN ORTIZ", "21/04/2026", "16/04/2026", "16/04/2026",
                "U220641 - FUNDACION SALUD MIA EPS SUBSIDIADO", "HUS0000473035",
                "168313", 24437, "27/04/2026", "", "ELIAS ALFONSO CARVAJAL NAVARRO",
            ]
        ])

        ws3 = wb.create_sheet("I")
        _hoja_conceptos(ws3, [
            [
                "Glosa_Inicial", "Administrativo", "", "FAMISANAR EPS", 758300,
                "ELIAS", "NCA", "ELIAS", "NCA",
                "U220181 - FAMISANAR EPS SUBSIDIADO", "EPS017",
                "HUS0000487351", "14/04/2026 15:27", "168185", "", "Confirmado",
                758300, "13/03/2026 16:44", "830003564",
                "ENTIDAD PROMOTORA DE SALUD FAMISANAR S A S", "14/04/2026",
                "", 564068, "", 0, "ENTIDAD PROMOTORA",
                "TA0801", "1038599",
                "Los cargos por apoyo diagnóstico presentan diferencias con los valores pactados",
                "906625",
                "GONADOTROPINA CORIONICA SUBUNIDAD BETA",
                22900, "734005 - LABORATORIO - INMUNOLOGIA",
                "SE REALIZA OBJECIÃ³N POR MAYOR VALOR",
            ],
            [
                "Glosa_Inicial", "Administrativo", "", "FAMISANAR EPS", 758300,
                "ELIAS", "NCA", "ELIAS", "NCA",
                "U220181 - FAMISANAR EPS SUBSIDIADO", "EPS017",
                "HUS0000487351", "14/04/2026 15:27", "168185", "", "Confirmado",
                758300, "13/03/2026 16:44", "830003564",
                "ENTIDAD PROMOTORA DE SALUD FAMISANAR S A S", "14/04/2026",
                "", 564068, "", 0, "ENTIDAD PROMOTORA",
                "TA0801", "1038598",
                "Los cargos por apoyo diagnóstico presentan diferencias",
                "902045", "TIEMPO DE PROTROMBINA [TP]",
                15400, "734009 - LABORATORIO - COAGULACION",
                "SE REALIZA OBJECION TP",
            ],
        ])

        ws4 = wb.create_sheet("R")
        _hoja_conceptos(ws4, [
            [
                "Glosa_Ratificada", "Administrativo", "",
                "FUNDACION SALUD MIA EPS SUBSIDIADO", 24437, "ELIAS", "NCA",
                "ELIAS", "NCA", "U220641 - FUNDACION SALUD MIA EPS SUBSIDIADO",
                "EPSC49", "HUS0000473035", "16/04/2026", "168313", "",
                "Confirmado", 131200, "06/02/2026 9:08", "900914254",
                "FUNDACION SALUD MIA EPS", "16/04/2026", "", 564356, "", 0,
                "FUNDACION SALUD MIA EPS",
                "TA0201", "1041754",
                "El cargo por consulta presenta diferencias con los valores pactados",
                "39143A-10",
                "CONSULTA DE PRIMERA VEZ POR ESPECIALISTA",
                24437, "731101 - CONSULTA EXTERNA ESPECIALIZADA",
                "SE RATIFICA - Se realiza glosa por mayor valor cobrado",
            ]
        ])

        r = RecepcionService(db).procesar_excel(_bytes_wb(wb))
        assert r.creadas == 2, f"Esperaba 2 glosas, obtuve {r.creadas}. Errores: {r.errores}"
        assert r.ratificadas == 1
        assert r.conceptos_creados == 3, f"Esperaba 3 conceptos, obtuve {r.conceptos_creados}"
        assert r.conceptos_huerfanos == []

        # Verificar que la glosa inicial tiene sus 2 conceptos
        from app.models.db import GlosaRecord, ConceptoGlosaRecord
        g_inicial = db.query(GlosaRecord).filter(GlosaRecord.factura == "HUS0000487351").first()
        conceptos = db.query(ConceptoGlosaRecord).filter(
            ConceptoGlosaRecord.glosa_id == g_inicial.id
        ).all()
        assert len(conceptos) == 2
        codigos = sorted([c.codigo_glosa for c in conceptos])
        assert codigos == ["TA0801", "TA0801"]
        cups = sorted([c.cups_codigo for c in conceptos])
        assert cups == ["902045", "906625"]
        # Mojibake corregido: "OBJECIÃ³N" debe volverse "OBJECIóN"
        observaciones_con_objecion = [c.observacion_eps for c in conceptos if "OBJECIóN" in (c.observacion_eps or "")]
        assert len(observaciones_con_objecion) == 1

    def test_concepto_huerfano_se_reporta(self, db):
        """Si no hay glosa padre (INICIAL antes), el concepto queda huérfano."""
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("I")
        _hoja_conceptos(ws, [
            [
                "Glosa_Inicial", "Administrativo", "", "FAMISANAR EPS", 0,
                "ELIAS", "NCA", "ELIAS", "NCA",
                "U220181 - FAMISANAR EPS SUBSIDIADO", "EPS017",
                "HUS9999999999", "14/04/2026", "999999", "", "Confirmado",
                0, "14/04/2026", "830003564", "FAMISANAR", "14/04/2026",
                "", 1, "", 0, "FAMISANAR",
                "TA0801", "9001", "Concepto huérfano",
                "906625", "GONADOTROPINA", 22900,
                "734005 - LABORATORIO", "Observación",
            ]
        ])
        r = RecepcionService(db).procesar_excel(_bytes_wb(wb))
        assert r.conceptos_creados == 0
        assert len(r.conceptos_huerfanos) == 1
        assert r.conceptos_huerfanos[0]["factura"] == "HUS9999999999"
        assert r.conceptos_huerfanos[0]["consecutivo_dgh"] == "999999"
        assert r.conceptos_huerfanos[0]["codigo_glosa"] == "TA0801"
