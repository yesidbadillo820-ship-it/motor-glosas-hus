"""El Excel de Glosas ADRES pone quién trabaja cada factura — de punta a punta.

02-09-2026. Con el paquete 31078 en la mano, las 84 facturas salían
«(sin gestor asignado)» aunque 16 estaban CERRADAS (con el correo de quien las
cerró en la columna «Cerrada por») y 41 EN PROCESO con glosas ya decididas.

Acá se arma un paquete de mentira en una base SQLite en memoria, se genera el
libro DE VERDAD con `construir_informe_paquete` y se lee con openpyxl:

  · FACTURAS: la columna Gestor trae a la persona y una columna nueva dice de
    dónde salió (quien la cerró / quien decidió sus glosas / la macro);
  · Hoja1: los renglones de esa factura llevan el mismo gestor, para que la
    hoja POR GESTOR reparta de verdad;
  · POR GESTOR: aparecen las personas, no solo «(sin gestor asignado)».
"""

from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.database import Base
from app.models.db import (
    FacturaAdresRecord,
    GlosaAdresRecord,
    PaqueteAdresRecord,
    UsuarioRecord,
)
from app.services.gestor_adres import (
    ORIGEN_CIERRE,
    ORIGEN_DECISIONES,
    ORIGEN_MACRO,
    SIN_GESTOR,
)
from app.services.glosas_adres_excel import construir_informe_paquete

CAROLINA = "auditorhus02@sinacsc.com"
JHON = "glosashus09@sinacsc.com"
SIN_NOMBRE = "glosashus03@sinacsc.com"  # está en la base, pero sin usuario creado


def _glosa(paquete_id, factura, n, **campos):
    base = dict(
        paquete_id=paquete_id,
        factura_clave=factura,
        factura=factura,
        radicacion="14408663",
        doc_victima="CC-91238287",
        consecutivo=str(n),
        tipo_elemento="PROCEDIMIENTO",
        codigo=f"39{n:03d}",
        descripcion=f"SERVICIO {n}",
        causal_codigo="3209",
        causal_texto="3209- La ayuda diagnóstica no tiene justificación",
        cant_reclamada=1.0,
        valor_reclamado=100.0,
        cant_aprobada=0.0,
        valor_aprobado=0.0,
        valor_glosado=100.0,
        clasificacion="PERTINENCIA",
        centro_costos="735101-BANCO DE SANGRE",
        cuenta_valor=True,
        glosa_total=False,
    )
    base.update(campos)
    return GlosaAdresRecord(**base)


@pytest.fixture
def libro():
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        db.add_all(
            [
                UsuarioRecord(
                    nombre="Carolina Rueda", email=CAROLINA, password_hash="x", rol="AUDITOR"
                ),
                UsuarioRecord(nombre="Jhon Alex", email=JHON, password_hash="x", rol="AUDITOR"),
            ]
        )
        paquete = PaqueteAdresRecord(numero_paquete="31078", archivo="reporte.xlsx")
        db.add(paquete)
        db.flush()
        pid = paquete.id
        db.add_all(
            [
                FacturaAdresRecord(
                    paquete_id=pid,
                    factura_clave="HUS1",
                    factura="HUS1",
                    estado="CERRADA",
                    cerrada_por=CAROLINA,
                ),
                FacturaAdresRecord(
                    paquete_id=pid,
                    factura_clave="HUS2",
                    factura="HUS2",
                    estado="EN PROCESO",
                ),
                FacturaAdresRecord(
                    paquete_id=pid,
                    factura_clave="HUS3",
                    factura="HUS3",
                    estado="PENDIENTE",
                ),
                FacturaAdresRecord(
                    paquete_id=pid,
                    factura_clave="HUS4",
                    factura="HUS4",
                    estado="CERRADA",
                    gestor="MARIA",
                    cerrada_por=CAROLINA,
                ),
                FacturaAdresRecord(
                    paquete_id=pid,
                    factura_clave="HUS5",
                    factura="HUS5",
                    estado="CERRADA",
                    cerrada_por=SIN_NOMBRE,
                ),
            ]
        )
        db.add_all(
            [
                _glosa(pid, "HUS1", 1, decision="SE SUBSANA", decidido_por=CAROLINA),
                _glosa(pid, "HUS2", 1, decision="SE SUBSANA", decidido_por=JHON),
                _glosa(pid, "HUS2", 2, decision="SE OBJETA", decidido_por=JHON),
                _glosa(pid, "HUS2", 3),  # sin decidir todavía
                _glosa(pid, "HUS3", 1),
                _glosa(
                    pid, "HUS4", 1, decision="SE ACEPTA", decidido_por=JHON, valor_aceptado=100.0
                ),
                _glosa(pid, "HUS5", 1, decision="SE OBJETA", decidido_por=SIN_NOMBRE),
            ]
        )
        db.commit()
        contenido = construir_informe_paquete(db, pid, generado_por="prueba")
    return load_workbook(BytesIO(contenido))


def _tabla(ws, fila_encabezado):
    filas = list(ws.iter_rows(values_only=True))
    cab = [str(c or "") for c in filas[fila_encabezado - 1]]
    # En Hoja1 la columna A (Código Habilitación) puede ir vacía: se filtra
    # por "la fila trae algo", no por la primera celda.
    datos = [
        dict(zip(cab, f))
        for f in filas[fila_encabezado:]
        if f and any(c not in (None, "") for c in f)
    ]
    return cab, datos


class TestLaHojaFacturas:
    def test_trae_la_columna_nueva(self, libro):
        cab, _ = _tabla(libro["FACTURAS"], 4)
        assert cab[-1] == "De dónde sale el gestor"
        assert len(cab) == 18

    def test_cerrada_dice_quien_la_cerro(self, libro):
        _, filas = _tabla(libro["FACTURAS"], 4)
        f = next(x for x in filas if x["Factura"] == "HUS1")
        assert f["Gestor"] == "Carolina Rueda"
        assert f["De dónde sale el gestor"] == ORIGEN_CIERRE
        assert f["Cerrada por"] == CAROLINA  # el correo sigue ahí, no se pierde

    def test_en_proceso_dice_quien_decidio_sus_glosas(self, libro):
        _, filas = _tabla(libro["FACTURAS"], 4)
        f = next(x for x in filas if x["Factura"] == "HUS2")
        assert f["Gestor"] == "Jhon Alex"
        assert f["De dónde sale el gestor"] == ORIGEN_DECISIONES

    def test_pendiente_sin_tocar_sigue_sin_gestor(self, libro):
        _, filas = _tabla(libro["FACTURAS"], 4)
        f = next(x for x in filas if x["Factura"] == "HUS3")
        assert f["Gestor"] == SIN_GESTOR
        assert not f["De dónde sale el gestor"]

    def test_la_macro_manda_cuando_trae_gestor(self, libro):
        _, filas = _tabla(libro["FACTURAS"], 4)
        f = next(x for x in filas if x["Factura"] == "HUS4")
        assert f["Gestor"] == "MARIA"
        assert f["De dónde sale el gestor"] == ORIGEN_MACRO

    def test_correo_sin_usuario_se_muestra_tal_cual(self, libro):
        """No se inventa un nombre: el correo es el dato real que hay."""
        _, filas = _tabla(libro["FACTURAS"], 4)
        f = next(x for x in filas if x["Factura"] == "HUS5")
        assert f["Gestor"] == SIN_NOMBRE

    def test_la_fila_total_no_se_corrio(self, libro):
        filas = list(libro["FACTURAS"].iter_rows(values_only=True))
        total = next(f for f in filas if f[0] == "TOTAL")
        assert str(total[6]).startswith("=SUM(G5:G")  # Glosas a responder
        assert str(total[15]).startswith("=SUM(P5:P")  # Sigue glosado
        assert libro["FACTURAS"].auto_filter.ref.startswith("A4:R")


class TestLaHojaDeDatosYPorGestor:
    def test_hoja1_lleva_el_gestor_de_su_factura(self, libro):
        cab, filas = _tabla(libro["Hoja1"], 1)
        de_hus2 = [f for f in filas if f["Número Factura"] == "HUS2"]
        assert len(de_hus2) == 3
        assert {f["GESTOR"] for f in de_hus2} == {"Jhon Alex"}
        de_hus3 = [f for f in filas if f["Número Factura"] == "HUS3"]
        assert {f["GESTOR"] for f in de_hus3} == {SIN_GESTOR}

    def test_las_26_columnas_de_la_macro_siguen_en_su_sitio(self, libro):
        """GESTOR es la columna 25 (Y) de la macro: cambia lo que dice, no dónde está."""
        cab, _ = _tabla(libro["Hoja1"], 1)
        assert cab[24] == "GESTOR"
        assert cab[2] == "Número Factura"

    def test_por_gestor_reparte_entre_personas(self, libro):
        _, filas = _tabla(libro["POR GESTOR"], 4)
        gestores = {f["Gestor"] for f in filas} - {"TOTAL"}
        assert gestores == {"Carolina Rueda", "Jhon Alex", "MARIA", SIN_NOMBRE, SIN_GESTOR}

    def test_como_leer_explica_de_donde_sale_el_gestor(self, libro):
        textos = " ".join(
            str(c) for f in libro["CÓMO LEER"].iter_rows(values_only=True) for c in f if c
        )
        assert "EL GESTOR" in textos
        assert "quien la cerró" in textos
