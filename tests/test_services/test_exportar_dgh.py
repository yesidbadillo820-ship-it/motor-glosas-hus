"""Tests del exportador formato DGH (Ronda 35)."""
from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app.models.db import ConceptoGlosaRecord, GlosaRecord
from app.services.exportar_dgh import (
    COLUMNAS_DGH,
    codigo_respuesta_efectivo,
    estado_cxc_objecion,
    generar_excel_dgh,
    generar_filas_dgh,
    limpiar_dictamen_para_dgh,
    resolver_tercero,
    tipo_objecion_tramite,
)


@pytest.fixture
def db():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    S = sessionmaker(bind=engine)
    s = S()
    try:
        yield s
    finally:
        s.close()


def _mk(db, **kw):
    d = dict(
        eps="FAMISANAR EPS", paciente="X", factura="HUS0000999",
        codigo_glosa="TA0201", valor_objetado=100_000,
        estado="PENDIENTE", creado_en=datetime.now(timezone.utc),
    )
    d.update(kw)
    g = GlosaRecord(**d)
    db.add(g); db.commit(); db.refresh(g)
    return g


# ─── Limpieza del dictamen ─────────────────────────────────────────────────

class TestLimpiarDictamen:
    def test_vacio_retorna_vacio(self):
        g = GlosaRecord()
        assert limpiar_dictamen_para_dgh("", g) == ""
        assert limpiar_dictamen_para_dgh(None, g) == ""

    def test_texto_fijo_ratificada_usa_canonico(self):
        from app.services.glosa_service import TEXTO_RATIFICADA
        g = GlosaRecord(modelo_ia="pre-analisis/texto_fijo/RATIFICADA",
                        dictamen="<div>RATIFICADA EPS: X | Factura: Y | Observación recepción: basura</div>")
        r = limpiar_dictamen_para_dgh(g.dictamen, g)
        assert r == TEXTO_RATIFICADA

    def test_texto_fijo_extemporanea_usa_canonico(self):
        g = GlosaRecord(
            modelo_ia="pre-analisis/texto_fijo/EXTEMPORANEA",
            dias_radicacion_dgh=22,
            dictamen="<div>GLOSA EXTEMPORÁNEA — 22 DÍAS</div>",
        )
        r = limpiar_dictamen_para_dgh(g.dictamen, g)
        assert "EXTEMPORÁNEA" in r.upper() or "EXTEMPORANEA" in r.upper()
        assert "22" in r

    def test_remueve_emojis(self):
        g = GlosaRecord(modelo_ia="anthropic/claude")
        html = "<p>📋 Tarifa pactada ESE HUS NO ACEPTA la glosa 🔍</p>"
        r = limpiar_dictamen_para_dgh(html, g)
        assert "📋" not in r
        assert "🔍" not in r
        assert "ESE HUS NO ACEPTA" in r

    def test_remueve_header_ratificada_debug(self):
        g = GlosaRecord(modelo_ia="anthropic/claude")
        html = ("RATIFICADA EPS: U240061 - FIDEICOMISOS | Factura: HUS001 | "
                "Observación recepción: ESE HUS DIJO XYZ "
                "ESE HUS NO ACEPTA LA RATIFICACIÓN")
        r = limpiar_dictamen_para_dgh(html, g)
        assert "RATIFICADA EPS:" not in r
        assert "Observación recepción:" not in r
        assert "NO ACEPTA LA RATIFICACIÓN" in r

    def test_remueve_banner_tarifa_pactada(self):
        g = GlosaRecord(modelo_ia="anthropic/claude")
        html = ("Tarifa pactada encontrada en el contrato · Defender "
                "CUPS: 890750 EPS: FAMISANAR Contrato: ABC "
                "ESE HUS NO ACEPTA la glosa por tarifa")
        r = limpiar_dictamen_para_dgh(html, g)
        assert "Tarifa pactada encontrada" not in r
        assert "ESE HUS NO ACEPTA" in r

    def test_strip_html(self):
        g = GlosaRecord(modelo_ia="anthropic/claude")
        html = "<div><p>Linea uno</p><br><b>Linea dos</b></div>"
        r = limpiar_dictamen_para_dgh(html, g)
        assert "<" not in r
        assert ">" not in r
        assert "Linea uno" in r and "Linea dos" in r


# ─── resolver_tercero ──────────────────────────────────────────────────────

class TestResolverTercero:
    def test_con_eps_codigo_y_tercero_nombre(self):
        g = GlosaRecord(
            eps="FAMISANAR EPS",
            eps_codigo="U220181",
            tercero_nombre="FAMISANAR EPS SUBSIDIADO",
            tercero_nit="900226715",
        )
        an, na, nit = resolver_tercero(g)
        assert an == "U220181 - FAMISANAR EPS SUBSIDIADO"
        assert na == "FAMISANAR EPS SUBSIDIADO"
        assert nit == "900226715"

    def test_otra_sin_definir_usa_tercero_nombre(self):
        g = GlosaRecord(
            eps="OTRA / SIN DEFINIR",
            tercero_nombre="SANITAS S.A.S. EPS CONTRIBUTIVO",
            tercero_nit="800251440",
        )
        an, na, nit = resolver_tercero(g)
        assert "SANITAS" in an
        assert na == "SANITAS S.A.S. EPS CONTRIBUTIVO"
        assert nit == "800251440"

    def test_eps_con_prefijo_en_mismo_string(self):
        g = GlosaRecord(eps="U220181 - FAMISANAR EPS SUBSIDIADO")
        an, na, nit = resolver_tercero(g)
        # Debe reconocer el prefijo y extraerlo
        assert "U220181" in an
        assert "FAMISANAR EPS SUBSIDIADO" in an
        assert na == "FAMISANAR EPS SUBSIDIADO"

    def test_sin_data_devuelve_sin_definir(self):
        g = GlosaRecord()
        an, na, nit = resolver_tercero(g)
        assert an == "SIN DEFINIR"
        assert na == "SIN DEFINIR"
        assert nit == ""


# ─── Clasificadores ────────────────────────────────────────────────────────

class TestClasificadores:
    def test_ratificada_por_estado(self):
        g = GlosaRecord(estado="RATIFICADA")
        assert estado_cxc_objecion(g) == "Glosa_Ratificada"

    def test_inicial_por_default(self):
        g = GlosaRecord(estado="PENDIENTE")
        assert estado_cxc_objecion(g) == "Glosa_Inicial"

    def test_tipo_administrativo_default(self):
        assert tipo_objecion_tramite("TA0201") == "Administrativo"
        assert tipo_objecion_tramite("SO0101") == "Administrativo"
        assert tipo_objecion_tramite("") == "Administrativo"

    def test_tipo_clinico(self):
        assert tipo_objecion_tramite("CL0101") == "Clínico"
        assert tipo_objecion_tramite("PE0301") == "Clínico"

    def test_codigo_respuesta_ratificada(self):
        g = GlosaRecord(modelo_ia="pre-analisis/texto_fijo/RATIFICADA")
        assert codigo_respuesta_efectivo(g) == "RE9901"

    def test_codigo_respuesta_existente_se_respeta(self):
        g = GlosaRecord(codigo_respuesta="RE9702")
        assert codigo_respuesta_efectivo(g) == "RE9702"


# ─── generar_filas_dgh ─────────────────────────────────────────────────────

class TestFilas:
    def test_glosa_sin_conceptos_fallback_una_fila(self, db):
        g = _mk(db, tercero_nombre="FAMISANAR SUBSIDIADO", tercero_nit="900226715",
                eps_codigo="U220181", cups_servicio="890701",
                servicio_descripcion="CONSULTA URGENCIAS",
                valor_objetado=50_000, valor_aceptado=0,
                dictamen="ESE HUS NO ACEPTA glosa por tarifa",
                codigo_respuesta="RE9901")
        filas = generar_filas_dgh(db, [g])
        assert len(filas) == 1
        f = filas[0]
        assert f["FacturaCartera.Factura"] == "HUS0000999"
        assert f["FacturaCartera.Tercero.Documento"] == "900226715"
        assert "U220181" in f["FacturaCartera.Tercero.NombreCompletoAN"]
        assert f["FacturaCartera.Tercero.NombreCompletoNA"] == "FAMISANAR SUBSIDIADO"
        assert f["ListadoConceptos.ConceptoObjecion.Codigo"] == "TA0201"
        assert f["ListadoConceptos.ServicioProductoFactura.Codigo"] == "890701"
        assert f["ListadoConceptos.ValorObjecion"] == 50_000.0
        assert f["CODIGO RESPUESTA"] == "RE9901"
        assert f["VALOR ACEPTADO"] == 0.0
        assert f["EstadoActual"] == "Confirmado"
        assert f["OBSERVACION"]  # no vacío

    def test_glosa_con_conceptos_emite_una_fila_por_concepto(self, db):
        g = _mk(db, tercero_nombre="FOMAG", tercero_nit="900500000")
        # Agregar 3 conceptos
        for i, (cups, desc, valor) in enumerate([
            ("890701", "CONSULTA URGENCIAS", 30000),
            ("873205", "RX CODO", 10000),
            ("890793", "CONSULTA ESP", 15000),
        ]):
            db.add(ConceptoGlosaRecord(
                glosa_id=g.id, factura=g.factura, codigo_glosa="TA0201",
                nombre_glosa="Tarifas", cups_codigo=cups, cups_descripcion=desc,
                valor_objetado=valor, observacion_eps="Mayor valor cobrado",
                oid_dgh=str(997490 + i),
            ))
        db.commit()
        filas = generar_filas_dgh(db, [g])
        assert len(filas) == 3
        cups_codes = sorted(f["ListadoConceptos.ServicioProductoFactura.Codigo"] for f in filas)
        assert cups_codes == ["873205", "890701", "890793"]

    def test_ratificada_emite_codigo_respuesta_9901(self, db):
        g = _mk(db, estado="RATIFICADA",
                modelo_ia="pre-analisis/texto_fijo/RATIFICADA",
                dictamen="<div>RATIFICADA EPS: X | Factura: Y | Observación recepción: basura ESE HUS NO ACEPTA</div>")
        filas = generar_filas_dgh(db, [g])
        f = filas[0]
        assert f["EstadoCxCObjecion"] == "Glosa_Ratificada"
        assert f["CODIGO RESPUESTA"] == "RE9901"
        # Observación limpia (texto canónico)
        from app.services.glosa_service import TEXTO_RATIFICADA
        assert f["OBSERVACION"] == TEXTO_RATIFICADA


# ─── Excel completo ────────────────────────────────────────────────────────

class TestExcel:
    def test_genera_excel_valido_con_26_columnas(self, db):
        _mk(db)
        buf = generar_excel_dgh(db, db.query(GlosaRecord).all())
        assert isinstance(buf, BytesIO)
        buf.seek(0)
        wb = load_workbook(buf)
        ws = wb.active
        assert ws.title == "Glosas_DGH"
        # Verificar que las 26 cabeceras coinciden exactamente
        for idx, col in enumerate(COLUMNAS_DGH, start=1):
            assert ws.cell(row=1, column=idx).value == col

    def test_celda_valor_objecion_es_numerica(self, db):
        _mk(db, valor_objetado=99200.0)
        buf = generar_excel_dgh(db, db.query(GlosaRecord).all())
        wb = load_workbook(buf)
        ws = wb.active
        # Columna ValorObjecion es la #19
        idx = COLUMNAS_DGH.index("ListadoConceptos.ValorObjecion") + 1
        assert ws.cell(row=2, column=idx).value == 99200.0
