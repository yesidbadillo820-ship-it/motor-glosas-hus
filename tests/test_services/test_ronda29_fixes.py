"""Regresiones de la ronda 29 (revisión ultracode 7-jul-2026).

Cubre los fixes aplicados sobre bugs confirmados por el workflow:
  1. Etiqueta "[degradado]" desbordaba el VARCHAR(100) de modelo en
     Postgres — el commit moría DESPUÉS de pagar la llamada de IA.
  2. placeholder_en_formula marcaba falsos positivos con incisos
     entre guiones ("... EL CUPS DE LA FACTURA - QUE LA EPS DEBE ...").
  3. cups_fantasma castigaba códigos VERIFICADOS por el motor (CUPS de
     la factura / tarifa contratada ausentes del texto de la glosa) y
     no reconocía códigos con separador de miles ("27.535").
  4. _extraer_valor: "VALOR OBJETADO 100%" devolvía "$ 100" y el
     formato "$ 6.446.000" (espacio tras $) no se capturaba.
  5. get_contrato: tildes en la EPS ("POLICÍA") rompían el match y las
     claves de un solo token matcheaban por subcadena dentro de otras
     palabras.
  6. Router /exportar/dgh reconectado (el botón "Formato DGH" del
     Historial daba 404 con el servicio completo huérfano).
"""

from __future__ import annotations

import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base, get_db
from app.models.db import AICacheRecord, GlosaRecord, UsuarioRecord
from app.services.validador_dictamen import detectar_defectos_criticos

MODELO_LARGO = (
    "claude-opus-4-7 [degradado] causa: Error code: 404 - model not found "
    "- intente con otro identificador de modelo o revise ANTHROPIC_MODEL_OPUS "
    "en el archivo .env del despliegue self-hosted del HUS"
)


@pytest.fixture
def db_session():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine)
    s = factory()
    try:
        yield s, factory
    finally:
        s.close()
        engine.dispose()


def _dictamen(cuerpo: str) -> str:
    relleno = " LA ESE HUS SOSTIENE LA FACTURACIÓN CONFORME AL CONTRATO." * 3
    return f"<argumento>ESE HUS NO ACEPTA LA GLOSA. {cuerpo}{relleno} COMUNICACIONES: CARTERA@HUS.GOV.CO.</argumento>"


# ── 1. modelo truncado a 80 en los caches de IA ──────────────────────────


class TestModeloTruncadoEnCache:
    def test_guardar_cache_ia_db_trunca(self, db_session, monkeypatch):
        s, factory = db_session
        import app.database as app_db

        monkeypatch.setattr(app_db, "SessionLocal", factory)
        from app.services.glosa_service import _guardar_cache_ia_db

        _guardar_cache_ia_db("clave-r29", "<argumento>X</argumento>", MODELO_LARGO)
        rec = s.query(AICacheRecord).filter(AICacheRecord.clave == "clave-r29").first()
        assert rec is not None
        assert len(rec.modelo) <= 80

    def test_guardar_cache_forense_trunca(self, db_session, monkeypatch):
        s, factory = db_session
        import app.database as app_db

        monkeypatch.setattr(app_db, "SessionLocal", factory)
        from app.services.auditor_forense import _guardar_cache_forense

        _guardar_cache_forense("clave-forense-r29", "<div>mapa</div>", MODELO_LARGO)
        rec = s.query(AICacheRecord).filter(AICacheRecord.clave == "clave-forense-r29").first()
        assert rec is not None
        assert len(rec.modelo) <= 80


# ── 2. placeholder_en_formula: inciso con guion NO es fórmula ────────────


class TestPlaceholderEnFormula:
    def _reglas(self, cuerpo: str) -> set:
        defectos = detectar_defectos_criticos(_dictamen(cuerpo), codigo_glosa="TA0801")
        return {d["regla"] for d in defectos}

    def test_inciso_con_guion_no_dispara(self):
        cuerpo = (
            "EL PROCEDIMIENTO SE LIQUIDÓ BAJO EL CUPS DE LA FACTURA - QUE LA EPS "
            "DEBE PRECISAR EN SU OBJECIÓN - CONFORME AL ANEXO TARIFARIO."
        )
        assert "placeholder_en_formula" not in self._reglas(cuerpo)

    def test_formula_con_numero_sigue_disparando(self):
        cuerpo = "SE LIQUIDA ASÍ: EL VALOR OBJETADO CONSIGNADO EN EL EXPEDIENTE - 10% = $54.000."
        assert "placeholder_en_formula" in self._reglas(cuerpo)

    def test_frase_igual_frase_sigue_disparando(self):
        cuerpo = (
            "(EL VALOR OBJETADO CONSIGNADO EN EL EXPEDIENTE = "
            "EL VALOR OBJETADO CONSIGNADO EN EL EXPEDIENTE)."
        )
        assert "placeholder_en_formula" in self._reglas(cuerpo)


# ── 3. cups_fantasma: códigos verificados y separadores de miles ─────────


class TestCupsFantasmaExtras:
    def test_codigo_verificado_extra_no_es_fantasma(self):
        defectos = detectar_defectos_criticos(
            _dictamen("EL SERVICIO CORRESPONDE AL CUPS 890201 DE LA FACTURA."),
            codigo_glosa="TA0801",
            texto_glosa="SE OBJETA TARIFA DEL PROCEDIMIENTO POR VALOR DE $100.000",
            codigos_validos_extra=["890201"],
        )
        assert "cups_fantasma" not in {d["regla"] for d in defectos}

    def test_codigo_con_separador_de_miles_en_glosa(self):
        defectos = detectar_defectos_criticos(
            _dictamen("EL SERVICIO CORRESPONDE AL CUPS 27535 SEGÚN CONTRATO."),
            codigo_glosa="TA0801",
            texto_glosa="SE OBJETA EL CODIGO 27.535 POR DIFERENCIA DE TARIFA",
        )
        assert "cups_fantasma" not in {d["regla"] for d in defectos}

    def test_cups_inventado_sigue_disparando(self):
        defectos = detectar_defectos_criticos(
            _dictamen("EL SERVICIO CORRESPONDE AL CUPS 99999 SEGÚN CONTRATO."),
            codigo_glosa="TA0801",
            texto_glosa="SE OBJETA TARIFA DEL PROCEDIMIENTO 890201 POR $100.000",
        )
        assert "cups_fantasma" in {d["regla"] for d in defectos}


# ── 4. _extraer_valor: etiquetado manda, % no contamina ──────────────────


class TestExtraerValorRonda29:
    @pytest.fixture
    def svc(self):
        from app.services.glosa_service import GlosaService

        return GlosaService.__new__(GlosaService)

    def test_objetado_100_por_ciento_no_devuelve_100(self, svc):
        texto = "VALOR OBJETADO 100% GLOSA TOTAL. VALOR FACTURADO $ 6.446.000"
        assert svc._extraer_valor(texto) != "$ 100"

    def test_espacio_despues_del_signo_pesos(self, svc):
        texto = "SE OBJETA POR VALOR OBJETADO: $ 6.446.000 SEGUN SOPORTES"
        assert svc._extraer_valor(texto) == "$ 6.446.000"

    def test_total_objetado_gana_sobre_subconceptos(self, svc):
        texto = (
            "CONCEPTO 1 VALOR OBJETADO $ 200.000. CONCEPTO 2 VALOR OBJETADO "
            "$ 300.000. TOTAL OBJETADO: $ 500.000"
        )
        assert svc._extraer_valor(texto) == "$ 500.000"


# ── 5. get_contrato: tildes y claves de un token ─────────────────────────


class TestGetContratoRonda29:
    def test_tilde_en_eps_matchea_catalogo(self):
        """La tilde no puede cambiar el resultado.

        La fecha va fija dentro de la vigencia del contrato de la Policía
        (hasta el 15-08-2026 según la malla): lo que se comprueba es que
        "POLICÍA" y "POLICIA" den lo mismo, no si el contrato sigue vigente
        hoy. Sin fecha fija, esta prueba se cayó sola al vencerse el contrato.
        """
        import datetime as dt

        from app.services.glosa_ia_prompts import get_contrato

        dia = dt.date(2026, 6, 1)
        con_tilde = get_contrato("POLICÍA NACIONAL", dia)
        sin_tilde = get_contrato("POLICIA NACIONAL", dia)
        assert con_tilde == sin_tilde
        assert "SIN CONTRATO" not in str(con_tilde.get("numero", "")).upper()

    def test_clave_un_token_exige_palabra_completa(self):
        from app.services.glosa_ia_prompts import get_contrato

        # "AURORA" incrustada dentro de otra palabra NO debe traer el
        # contrato ARL de AURORA.
        resultado = get_contrato("CLINICA LAURORA IPS")
        assert "GID-ARL-0090" not in str(resultado.get("numero", ""))

    def test_clave_un_token_como_palabra_si_matchea(self):
        """«ARL» como palabra completa dentro de «SEGUROS AURORA S.A.».

        01-09-2026 — se le pasa la fecha porque los contratos de AURORA
        vencieron el 31-08-2026 y sin fecha la ficha responde la advertencia de
        vigencia, no el número. Lo que esta prueba vigila es la RESOLUCIÓN del
        nombre, no el calendario.
        """
        import datetime as _dt

        from app.services.glosa_ia_prompts import get_contrato

        resultado = get_contrato("SEGUROS AURORA S.A.", _dt.date(2026, 5, 1))
        assert "GID-ARL-0090" in str(resultado.get("numero", ""))

    def test_la_resolucion_del_nombre_funciona_aunque_el_contrato_venciera(self):
        """Con el contrato vencido igual tiene que reconocer a AURORA.

        Si la resolución del nombre fallara, la ficha diría «SIN CONTRATO
        PACTADO» —como cualquier pagador desconocido— y no habría forma de
        distinguir «no lo encontré» de «lo encontré y está vencido».
        """
        from app.services.glosa_ia_prompts import get_contrato

        resultado = get_contrato("SEGUROS AURORA S.A.", None)
        assert resultado.get("_vigencia_vencida") is True
        assert "GID" in str(resultado.get("numero", ""))


# ── 6. GET /exportar/dgh reconectado ─────────────────────────────────────


class TestExportarDGH:
    @pytest.fixture
    def client(self, db_session):
        s, _ = db_session
        from app.api.deps import get_usuario_actual
        from app.main import app

        usuario = UsuarioRecord(id=1, email="auditor@hus.com", rol="AUDITOR", activo=1)
        app.dependency_overrides[get_db] = lambda: iter([s]).__next__()
        app.dependency_overrides[get_usuario_actual] = lambda: usuario
        with TestClient(app) as c:
            yield c
        app.dependency_overrides.clear()

    def test_descarga_xlsx_solo_respondidas(self, client, db_session):
        s, _ = db_session
        s.add(
            GlosaRecord(
                eps="FAMISANAR EPS",
                codigo_glosa="TA0801",
                estado="RESPONDIDA",
                factura="HUS123",
                dictamen="<argumento>ESE HUS NO ACEPTA LA GLOSA.</argumento>",
            )
        )
        s.add(GlosaRecord(eps="FAMISANAR EPS", codigo_glosa="SO0101", estado="PENDIENTE"))
        s.commit()

        r = client.get("/exportar/dgh?solo_respondidas=true")
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers["content-type"]

        from io import BytesIO

        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(r.content))
        ws = wb.active
        # encabezado + 1 fila (solo la RESPONDIDA; la PENDIENTE queda fuera)
        assert ws.max_row == 2
