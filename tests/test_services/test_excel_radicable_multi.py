"""Integración: Excel radicable × glosas multi-código («Opción A»).

Bug del 12-jun-2026 (export del dueño): cada ConceptoGlosaRecord genera
una fila, pero como `concepto.dictamen_html` casi nunca viene, TODAS las
filas de la misma factura caían al `glosa.dictamen` completo y salían
palabra por palabra idénticas. Estos tests verifican el fix de
presentación: la columna RESPUESTA ESE HUS se divide por los separadores
«═══ RESPUESTA AL CÓDIGO … ═══» de multi_codigo.py.

Datos 100 % ficticios (FAC-TEST-*).
"""

from datetime import datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base
from app.models.db import ConceptoGlosaRecord, GlosaRecord
from app.services.excel_radicable import generar_excel_radicable
from app.utils.html_a_texto import dictamen_a_texto_plano
from tests.test_services.test_dictamen_secciones import DICTAMEN_MULTI

_COL_RESPUESTA = 12  # columna L — RESPUESTA ESE HUS


@pytest.fixture
def db_session():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    s = sessionmaker(bind=engine)()
    try:
        yield s
    finally:
        s.close()
        engine.dispose()


def _glosa_multi(db, *, glosa_id: int = 1) -> GlosaRecord:
    """Glosa multi-código (principal TA0201) con dictamen de 3 secciones."""
    g = GlosaRecord(
        id=glosa_id,
        eps="DISPENSARIO MEDICO",
        paciente="N/A",
        factura="FAC-TEST-100",
        consecutivo_dgh="200001",
        codigo_glosa="TA0201",
        valor_objetado=120000.0,
        valor_factura=300000.0,
        fecha_documento_dgh=datetime(2026, 5, 5),
        fecha_entrega=datetime(2026, 6, 10),
        estado="RESPONDIDA",
        dictamen=DICTAMEN_MULTI,
        codigo_respuesta="RE9901",
        score=0.9,
    )
    db.add(g)
    return g


def _concepto(db, *, cid: int, glosa_id: int, codigo: str, dictamen_html=None):
    db.add(
        ConceptoGlosaRecord(
            id=cid,
            glosa_id=glosa_id,
            factura="FAC-TEST-100",
            consecutivo_dgh="200001",
            codigo_glosa=codigo,
            cups_codigo=f"99{cid:04d}",
            cups_descripcion=f"SERVICIO DE PRUEBA {cid}",
            valor_objetado=10000.0,
            observacion_eps=f"OBSERVACIÓN EPS CONCEPTO {cid}",
            dictamen_html=dictamen_html,
        )
    )


def _respuestas(contenido: bytes, n_filas: int) -> list[str]:
    ws = load_workbook(BytesIO(contenido))["RESPUESTA GLOSAS"]
    return [ws.cell(4 + i, _COL_RESPUESTA).value for i in range(n_filas)]


def test_tres_conceptos_reciben_respuestas_distintas(db_session):
    """El bug: filas 4 y 100 idénticas. Con el fix, cada código su sección."""
    _glosa_multi(db_session)
    _concepto(db_session, cid=11, glosa_id=1, codigo="TA0201")
    _concepto(db_session, cid=12, glosa_id=1, codigo="SO0601")
    _concepto(db_session, cid=13, glosa_id=1, codigo="AU0301")
    db_session.commit()

    r_ta, r_so, r_au = _respuestas(generar_excel_radicable(db_session, glosa_ids=[1]), 3)

    # Principal: cabecera del documento + SU argumentación, sin las otras.
    assert "CÓDIGO GLOSA: TA0201" in r_ta
    assert "LAS TARIFAS FACTURADAS" in r_ta
    assert "LOS SOPORTES DOCUMENTALES" not in r_ta
    assert "AUTORIZACIÓN EMITIDA" not in r_ta

    # Cada adicional: solo su sección (con su título declarativo).
    assert "RESPUESTA AL CÓDIGO SO0601" in r_so
    assert "LOS SOPORTES DOCUMENTALES" in r_so
    assert "LAS TARIFAS FACTURADAS" not in r_so
    assert "RESPUESTA AL CÓDIGO AU0301" in r_au
    assert "AUTORIZACIÓN EMITIDA" in r_au
    assert "LOS SOPORTES DOCUMENTALES" not in r_au

    # Ninguna fila duplicada (la firma del bug del 12-jun-2026).
    assert len({r_ta, r_so, r_au}) == 3
    # Y todas en texto plano (sin tags del HTML almacenado).
    assert all("<div" not in r and "<table" not in r for r in (r_ta, r_so, r_au))


def test_un_solo_concepto_mantiene_dictamen_completo(db_session):
    """Glosa NO multi-concepto: status quo intacto — el dictamen completo
    (todas las secciones) en su única fila."""
    _glosa_multi(db_session)
    _concepto(db_session, cid=21, glosa_id=1, codigo="TA0201")
    db_session.commit()

    (respuesta,) = _respuestas(generar_excel_radicable(db_session, glosa_ids=[1]), 1)
    assert respuesta == dictamen_a_texto_plano(DICTAMEN_MULTI)
    assert "LOS SOPORTES DOCUMENTALES" in respuesta  # secciones adicionales incluidas


def test_codigo_sin_seccion_muestra_nota_neutra(db_session):
    """Concepto cuyo código no está en el dictamen (la glosa se analizó
    como un todo) → nota corta que remite a la fila principal, NO el
    dictamen completo duplicado."""
    _glosa_multi(db_session)
    _concepto(db_session, cid=31, glosa_id=1, codigo="TA0201")
    _concepto(db_session, cid=32, glosa_id=1, codigo="CO0701")  # sin sección propia
    db_session.commit()

    r_ta, r_co = _respuestas(generar_excel_radicable(db_session, glosa_ids=[1]), 2)
    assert "dictamen integral de la glosa #1" in r_co
    assert "código principal TA0201" in r_co
    assert "Ver la fila del código principal" in r_co
    assert "LAS TARIFAS FACTURADAS" not in r_co  # no arrastra el dictamen
    assert r_co != r_ta


def test_concepto_con_dictamen_propio_no_usa_el_divisor(db_session):
    """Si el concepto trae su dictamen_html, ese manda — aunque su código
    tenga sección en el dictamen de la glosa."""
    _glosa_multi(db_session)
    _concepto(db_session, cid=41, glosa_id=1, codigo="TA0201")
    _concepto(
        db_session,
        cid=42,
        glosa_id=1,
        codigo="SO0601",
        dictamen_html="<p>Respuesta propia del concepto SO0601 cargada por el gestor.</p>",
    )
    db_session.commit()

    _r_ta, r_so = _respuestas(generar_excel_radicable(db_session, glosa_ids=[1]), 2)
    assert r_so == "Respuesta propia del concepto SO0601 cargada por el gestor."
    assert "RESPUESTA AL CÓDIGO SO0601" not in r_so
