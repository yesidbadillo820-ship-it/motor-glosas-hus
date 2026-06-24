"""Tests del endpoint GET /glosas/importar-recepcion/{id}/excel-radicable
(jun-2026 — descarga del formato institucional radicable).

Mismo patrón de auth y BD in-memory que el test del Excel-respuesta.
Datos 100 % ficticios (FAC-TEST-*, sin NITs reales).
"""

from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base, get_db
from app.models.db import (
    ConceptoGlosaRecord,
    GlosaRecord,
    ImportacionRecepcionRecord,
    UsuarioRecord,
)


@pytest.fixture
def db_session():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    s = sessionmaker(bind=engine)()
    try:
        yield s
    finally:
        s.close()
        engine.dispose()


def _client(db_session, con_auth=True):
    from app.api.deps import get_usuario_actual
    from app.main import app

    app.dependency_overrides[get_db] = lambda: iter([db_session]).__next__()
    if con_auth:
        u = UsuarioRecord(
            id=1,
            email="x@hus.com",
            nombre="Test",
            rol="AUDITOR",
            activo=1,
            password_hash="x",
        )
        app.dependency_overrides[get_usuario_actual] = lambda: u
    return TestClient(app), app


def _sembrar(db):
    db.add(
        GlosaRecord(
            id=10,
            eps="DISPENSARIO MEDICO",
            paciente="N/A",
            factura="FAC-TEST-010",
            consecutivo_dgh="200010",
            codigo_glosa="TA0801",
            valor_objetado=40000.0,
            valor_factura=120000.0,
            estado="RESPONDIDA",
            dictamen="<p>Defensa de prueba endpoint.</p>",
            codigo_respuesta="RE9901",
        )
    )
    db.add(
        ConceptoGlosaRecord(
            id=101,
            glosa_id=10,
            factura="FAC-TEST-010",
            consecutivo_dgh="200010",
            codigo_glosa="TA0801",
            cups_codigo="444444",
            cups_descripcion="SERVICIO ENDPOINT",
            valor_objetado=40000.0,
            observacion_eps="OBS ENDPOINT",
        )
    )
    db.add(
        ImportacionRecepcionRecord(
            id=1,
            usuario_email="x@hus.com",
            archivo_nombre="recepcion.xlsx",
            total_glosas=1,
            glosa_ids="[10]",
            estado="LISTO",
            ruta_original=None,  # el radicable NO depende del archivo original
        )
    )
    db.commit()


def test_descarga_excel_radicable_valido(db_session):
    _sembrar(db_session)
    c, app = _client(db_session)
    try:
        r = c.get("/glosas/importar-recepcion/1/excel-radicable")
        assert r.status_code == 200, r.text
        assert "spreadsheetml" in r.headers["content-type"]
        assert "RESPUESTA_GLOSAS_DMBUG_" in r.headers["content-disposition"]
        assert r.headers["content-disposition"].endswith('_HUS.xlsx"')

        wb = load_workbook(BytesIO(r.content))  # debe reabrir sin errores
        assert wb.sheetnames == ["RESPUESTA GLOSAS", "FUNDAMENTO JURÍDICO"]
        ws = wb["RESPUESTA GLOSAS"]
        assert ws.cell(3, 12).value == "RESPUESTA ESE HUS"
        assert ws.cell(4, 2).value == "FAC-TEST-010"
        assert ws.cell(4, 12).value == "Defensa de prueba endpoint."
        assert ws.cell(5, 8).value == "TOTAL GLOSADO"
        assert ws.cell(5, 9).value == 40000
    finally:
        app.dependency_overrides.clear()


def test_descarga_radicable_import_inexistente_404(db_session):
    c, app = _client(db_session)
    try:
        r = c.get("/glosas/importar-recepcion/999/excel-radicable")
        assert r.status_code == 404
    finally:
        app.dependency_overrides.clear()


def test_descarga_radicable_sin_auth_401(db_session):
    c, app = _client(db_session, con_auth=False)
    try:
        r = c.get("/glosas/importar-recepcion/1/excel-radicable")
        assert r.status_code == 401
    finally:
        app.dependency_overrides.clear()
