"""Tests del módulo de Pre-auditoría SINAC v2 (el consolidado como base de datos).

Cubre el flujo completo: subir fuentes (Radicación + DGReport) → registrar
oficio → escribir envío (autocompletar, dedup, subsanación) → auditar
(radicar/devolver, tope 3) → oficio de devolución PDF → estadísticas, más
el auto-sync (re-subir una fuente corregida se refleja en el consolidado).
"""

from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.auth import get_password_hash
from app.database import Base, get_db
from app.models.db import (
    DgReportRecord,
    FacturaPreauditoriaRecord,
    RadicacionCuentaRecord,
    UsuarioRecord,
)
from app.services import preauditoria_service as svc


@pytest.fixture
def db_session():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    s = sessionmaker(bind=engine)()
    try:
        yield s
    finally:
        s.close()
        engine.dispose()


@pytest.fixture
def usuario(db_session):
    u = UsuarioRecord(
        id=1,
        nombre="CLAUDIA",
        email="claudia@hus.gov.co",
        rol="AUDITOR",
        activo=1,
        password_hash=get_password_hash("xxxx"),
    )
    db_session.add(u)
    db_session.commit()
    return u


@pytest.fixture
def client(db_session, usuario):
    from app.api.deps import get_usuario_actual
    from app.main import app

    app.dependency_overrides[get_db] = lambda: iter([db_session]).__next__()
    app.dependency_overrides[get_usuario_actual] = lambda: usuario
    c = TestClient(app)
    yield c
    app.dependency_overrides.clear()


# ------------------------------------------------------------------
# Helpers: construir Excel de las fuentes reales
# ------------------------------------------------------------------


def _excel(headers, filas):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for f in filas:
        ws.append(f)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


RAD_HEADERS = [
    "Radicacion.Consecutivo",
    "CxC.Factura",
    "Radicacion.FechaDocumento",
    "CxC.Fecha",
    "CxC.Valor",
    "Radicacion.Tercero.Documento",
    "Radicacion.Tercero.NombreCompletoNA",
    "Radicacion.EstadoActual",
]

DG_HEADERS = ["FECHA DE ENVIO CORREO", "NUMERO DE FACTURA", "CUFE"]


def _rad_fila(
    envio, factura, valor, nit="860002400", entidad="AXA COLPATRIA ", estado="Radicado_Entidad"
):
    return [
        envio,
        factura,
        "2026-06-01 00:00:00",
        "2026-05-20 10:00:00",
        valor,
        nit,
        entidad,
        estado,
    ]


def _subir_radicacion(client, filas):
    return client.post(
        "/preauditoria/fuentes/radicacion",
        files={"archivo": ("radicacion.xlsx", _excel(RAD_HEADERS, filas), "application/xlsx")},
    )


def _subir_dgreport(client, facturas):
    filas = [["2026-07-10 09:00:00", f, "CUFE" + f] for f in facturas]
    return client.post(
        "/preauditoria/fuentes/dgreport",
        files={"archivo": ("dgreport.xlsx", _excel(DG_HEADERS, filas), "application/xlsx")},
    )


def _crear_oficio(client, radicado="FHUS-AS-I00877-26", fecha="2026-07-20T08:30"):
    r = client.post(
        "/preauditoria/oficios", json={"numero_radicado": radicado, "fecha_recibido": fecha}
    )
    assert r.status_code == 200, r.text
    return r.json()


def _escribir(client, oficio_id, envio):
    return client.post(f"/preauditoria/oficios/{oficio_id}/envios", json={"envio": str(envio)})


def _factura_id(client, factura):
    d = client.get(f"/preauditoria/facturas/{factura}/historial").json()
    return d["actual"]["id"]


ENV = "229304"
F1 = "HUS0000521680"
F2 = "HUS0000521070"
F3 = "HUS0000521071"


# ------------------------------------------------------------------
# Fuentes
# ------------------------------------------------------------------


class TestFuentes:
    def test_subir_radicacion_upsert(self, client):
        r = _subir_radicacion(client, [_rad_fila(ENV, F1, 250700), _rad_fila(ENV, F2, 10615224)])
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["nuevas"] == 2
        # re-subir sin cambios → sin_cambio
        r2 = _subir_radicacion(client, [_rad_fila(ENV, F1, 250700), _rad_fila(ENV, F2, 10615224)])
        assert r2.json()["sin_cambio"] == 2

    def test_radicacion_excluye_anulado(self, client):
        filas = [
            _rad_fila(ENV, F1, 250700, estado="Anulado"),
            _rad_fila("999999", F1, 300000, estado="Radicado_Entidad"),
        ]
        r = _subir_radicacion(client, filas)
        assert r.json()["facturas_validas"] == 1  # la Anulada se descarta
        # la factura quedó en el envío bueno
        lst = client.get("/preauditoria/fuentes/radicacion?envio=999999").json()
        assert lst["total"] == 1

    def test_entidad_se_recorta(self, client):
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700, entidad="AXA COLPATRIA   ")])
        d = client.get("/preauditoria/fuentes/radicacion").json()
        assert d["items"][0]["entidad"] == "AXA COLPATRIA"

    def test_el_resumen_dice_cual_fue_el_ultimo_archivo(self, client):
        """Sin esto no había forma de saber en la página si el Excel que se
        acaba de bajar del DGH ya estaba subido (pregunta real del 20-08)."""
        vacio = client.get("/preauditoria/fuentes/resumen").json()
        assert vacio["radicacion_ultimo_archivo"] is None
        assert vacio["radicacion_ultimo_cargue"] is None
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        _subir_dgreport(client, [F1])
        d = client.get("/preauditoria/fuentes/resumen").json()
        assert d["radicacion_ultimo_archivo"] == "radicacion.xlsx"
        assert d["radicacion_ultimo_cargue"]
        assert d["dgreport_ultimo_archivo"] == "dgreport.xlsx"
        assert d["dgreport_ultimo_cargue"]

    def test_dgreport_upsert(self, client):
        r = _subir_dgreport(client, [F1, F2])
        assert r.status_code == 200
        assert r.json()["nuevas"] == 2


# ------------------------------------------------------------------
# Escribir envío: autocompletar, varias facturas, dedup
# ------------------------------------------------------------------


class TestEscribirEnvio:
    def test_una_fila_por_factura(self, client):
        _subir_radicacion(
            client,
            [_rad_fila(ENV, F1, 250700), _rad_fila(ENV, F2, 10615224), _rad_fila(ENV, F3, 73500)],
        )
        _subir_dgreport(client, [F1])  # solo F1 tuvo correo
        o = _crear_oficio(client)
        r = _escribir(client, o["id"], ENV)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["nuevas"] == 3
        # el consolidado tiene 3 filas, autocompletadas
        cons = client.get("/preauditoria/consolidado").json()
        assert cons["total"] == 3
        porf = {i["factura"]: i for i in cons["items"]}
        assert porf[F1]["valor"] == 250700
        assert porf[F1]["correo_fe"] == "SI"
        assert porf[F2]["correo_fe"] == "NO"
        assert porf[F1]["entidad"] == "AXA COLPATRIA"
        assert porf[F1]["envio"] == ENV

    def test_dedup_envio(self, client):
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o = _crear_oficio(client)
        assert _escribir(client, o["id"], ENV).json()["nuevas"] == 1
        r2 = _escribir(client, o["id"], ENV)
        assert r2.json()["ya_cargado"] is True
        assert "ya fue cargado" in r2.json()["mensaje"].lower()
        # no duplicó
        assert client.get("/preauditoria/consolidado").json()["total"] == 1

    def test_envio_inexistente(self, client):
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o = _crear_oficio(client)
        r = _escribir(client, o["id"], "000000")
        assert r.status_code == 404
        assert "no existe" in r.json()["detail"].lower()

    def test_recarga_en_otro_oficio_reingresa_las_devueltas(self, client):
        """Caso real 30-07: facturación reenvía las subsanaciones con el MISMO
        número de envío dentro de un oficio nuevo. Antes el sistema bloqueaba
        con 'El envío ya fue cargado'; ahora la recarga reingresa las
        devueltas y omite las demás."""
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700), _rad_fila(ENV, F2, 100)])
        _subir_dgreport(client, [F1, F2])
        o1 = _crear_oficio(client, "FHUS-REC-1", "2026-07-27T14:46")
        _escribir(client, o1["id"], ENV)
        assert _devolver(client, _factura_id(client, F1)).status_code == 200
        assert _radicar(client, _factura_id(client, F2)).status_code == 200

        o2 = _crear_oficio(client, "FHUS-REC-2", "2026-07-30T14:02")
        r = _escribir(client, o2["id"], ENV)
        assert r.status_code == 200, r.text
        d = r.json()
        assert not d.get("ya_cargado")
        assert d["reingresos"] == 1 and d["nuevas"] == 0 and d["omitidas"] == 1
        # la devuelta quedó en subsanación, colgada del oficio NUEVO
        h = client.get(f"/preauditoria/facturas/{F1}/historial").json()["actual"]
        assert h["ronda"] == 2 and h["estado"] == "EN_SUBSANACION"
        assert h["oficio_fhus"] == "FHUS-REC-2"
        # la radicada no se tocó y ninguna factura se duplicó
        h2 = client.get(f"/preauditoria/facturas/{F2}/historial").json()["actual"]
        assert h2["estado"] == "RADICADA"
        assert client.get("/preauditoria/consolidado").json()["total"] == 2

    def test_recarga_sin_devueltas_se_permite_pero_no_mueve_nada(self, client):
        """La regla es hasta 3 oficios por envío: la recarga se permite aunque
        no haya devueltas, pero avisa que ninguna factura se movió."""
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        _subir_dgreport(client, [F1])
        o1 = _crear_oficio(client, "FHUS-REC-1", "2026-07-27T14:46")
        _escribir(client, o1["id"], ENV)
        assert _radicar(client, _factura_id(client, F1)).status_code == 200
        o2 = _crear_oficio(client, "FHUS-REC-2", "2026-07-30T14:02")
        r = _escribir(client, o2["id"], ENV)
        assert r.status_code == 200, r.text
        d = r.json()
        assert not d.get("ya_cargado")
        assert d["nuevas"] == 0 and d["reingresos"] == 0 and d["omitidas"] == 1
        assert any("no movió ninguna factura" in a for a in d["advertencias"])
        # la radicada sigue intacta y en su oficio original
        h = client.get(f"/preauditoria/facturas/{F1}/historial").json()["actual"]
        assert h["estado"] == "RADICADA" and h["oficio_fhus"] == "FHUS-REC-1"
        assert client.get("/preauditoria/consolidado").json()["total"] == 1

    def test_tope_de_3_oficios_por_envio(self, client):
        """El mismo envío se acepta en máximo 3 oficios; el 4.º se bloquea
        nombrando los radicados (no los ids internos)."""
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o1 = _crear_oficio(client, "FHUS-TOP-1", "2026-07-25T08:00")
        _escribir(client, o1["id"], ENV)
        assert _devolver(client, _factura_id(client, F1)).status_code == 200
        o2 = _crear_oficio(client, "FHUS-TOP-2", "2026-07-27T08:00")
        assert _escribir(client, o2["id"], ENV).json()["reingresos"] == 1
        assert _devolver(client, _factura_id(client, F1)).status_code == 200
        o3 = _crear_oficio(client, "FHUS-TOP-3", "2026-07-29T08:00")
        assert _escribir(client, o3["id"], ENV).json()["reingresos"] == 1
        o4 = _crear_oficio(client, "FHUS-TOP-4", "2026-07-31T08:00")
        r = _escribir(client, o4["id"], ENV)
        assert r.json()["ya_cargado"] is True
        assert "máximo 3" in r.json()["mensaje"]
        for radicado in ("FHUS-TOP-1", "FHUS-TOP-2", "FHUS-TOP-3"):
            assert radicado in r.json()["mensaje"]

    def test_recarga_doble_en_el_mismo_oficio_sigue_bloqueada(self, client):
        """El candado del doble clic no se pierde con la recarga permitida."""
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o1 = _crear_oficio(client, "FHUS-REC-1", "2026-07-27T14:46")
        _escribir(client, o1["id"], ENV)
        assert _devolver(client, _factura_id(client, F1)).status_code == 200
        o2 = _crear_oficio(client, "FHUS-REC-2", "2026-07-30T14:02")
        assert _escribir(client, o2["id"], ENV).json()["reingresos"] == 1
        # segunda vez en el MISMO oficio nuevo → bloqueada
        r = _escribir(client, o2["id"], ENV)
        assert r.json()["ya_cargado"] is True
        assert "este oficio" in r.json()["mensaje"]

    def _como_admin(self, db_session):
        # El borrado exige COORDINADOR/SUPER_ADMIN; el cliente base es AUDITOR.
        from app.api.deps import get_coordinador_o_admin
        from app.main import app

        admin = UsuarioRecord(
            id=7,
            nombre="ADMIN GUARD",
            email="admin-guard@hus.gov.co",
            rol="SUPER_ADMIN",
            activo=1,
            password_hash=get_password_hash("xxxx"),
        )
        db_session.add(admin)
        db_session.commit()
        app.dependency_overrides[get_coordinador_o_admin] = lambda: admin

    def test_eliminar_el_oficio_original_tras_la_recarga_se_bloquea_claro(self, client, db_session):
        """Tras recargar el envío en otro oficio, el oficio original guarda el
        historial de facturas que ya migraron: eliminarlo debe responder un
        mensaje claro (antes era un error 500 por integridad de la base)."""
        self._como_admin(db_session)
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o1 = _crear_oficio(client, "FHUS-DEL-1", "2026-07-27T14:46")
        _escribir(client, o1["id"], ENV)
        assert _devolver(client, _factura_id(client, F1)).status_code == 200
        o2 = _crear_oficio(client, "FHUS-DEL-2", "2026-07-30T14:02")
        assert _escribir(client, o2["id"], ENV).json()["reingresos"] == 1
        r = client.delete(f"/preauditoria/oficios/{o1['id']}")
        assert r.status_code == 409, r.text
        assert "subsanaron hacia otro oficio" in r.json()["detail"]
        assert F1 in r.json()["detail"]
        # el oficio nuevo (con la factura viva) tampoco pierde nada por esto
        h = client.get(f"/preauditoria/facturas/{F1}/historial").json()["actual"]
        assert h["oficio_fhus"] == "FHUS-DEL-2"

    def test_quitar_el_envio_del_oficio_original_tras_la_recarga_se_bloquea(
        self, client, db_session
    ):
        """Quitar el envío del oficio original tras la recarga era un no-op que
        respondía ok y le devolvía un cupo del tope de 3 al envío: ahora se
        bloquea explicando por qué."""
        self._como_admin(db_session)
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o1 = _crear_oficio(client, "FHUS-QUI-1", "2026-07-27T14:46")
        _escribir(client, o1["id"], ENV)
        assert _devolver(client, _factura_id(client, F1)).status_code == 200
        o2 = _crear_oficio(client, "FHUS-QUI-2", "2026-07-30T14:02")
        assert _escribir(client, o2["id"], ENV).json()["reingresos"] == 1
        r = client.delete(f"/preauditoria/oficios/{o1['id']}/envios/{ENV}")
        assert r.status_code == 409, r.text
        assert "subsanaron hacia otro oficio" in r.json()["detail"]
        # el registro del oficio original sigue contando para el tope
        p = client.get(f"/preauditoria/oficios/{o2['id']}/envios/{ENV}/preview").json()
        assert p["veces_cargado"] == 2

    def test_preview_de_la_recarga_avisa_el_oficio_anterior(self, client):
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o1 = _crear_oficio(client, "FHUS-REC-1", "2026-07-27T14:46")
        _escribir(client, o1["id"], ENV)
        assert _devolver(client, _factura_id(client, F1)).status_code == 200
        o2 = _crear_oficio(client, "FHUS-REC-2", "2026-07-30T14:02")
        p = client.get(f"/preauditoria/oficios/{o2['id']}/envios/{ENV}/preview").json()
        assert p["ya_cargado"] is False  # en ESTE oficio no se ha cargado
        assert p["cargado_en_otro_oficio"] == "FHUS-REC-1"
        assert p["reingresos"] == 1
        assert p["veces_cargado"] == 1 and p["max_cargas_envio"] == 3
        assert p["oficios_donde_cargado"] == ["FHUS-REC-1"]
        # y desde el oficio original sí dice que ya está cargado
        p1 = client.get(f"/preauditoria/oficios/{o1['id']}/envios/{ENV}/preview").json()
        assert p1["ya_cargado"] is True

    def test_preview_no_escribe(self, client):
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700), _rad_fila(ENV, F2, 100)])
        o = _crear_oficio(client)
        p = client.get(f"/preauditoria/oficios/{o['id']}/envios/{ENV}/preview").json()
        assert p["total_en_fuente"] == 2 and p["nuevas"] == 2
        assert client.get("/preauditoria/consolidado").json()["total"] == 0  # no escribió


# ------------------------------------------------------------------
# Auditar + subsanaciones + tope 3
# ------------------------------------------------------------------


def _devolver(client, fid, motivo="Falta FURIPS"):
    return client.patch(
        f"/preauditoria/facturas/{fid}/auditar",
        json={"resultado": "DEVUELTA", "motivo_devolucion": motivo},
    )


def _radicar(client, fid):
    return client.patch(f"/preauditoria/facturas/{fid}/auditar", json={"resultado": "RADICAR"})


def _actuar_como_admin(db_session):
    """Cambia el usuario del cliente a SUPER_ADMIN.

    Modificar una auditoría YA DECIDIDA (revertirla o corregir su observación,
    que también corrige el oficio de devolución) es solo de coordinación o
    administración; el cliente base es AUDITOR.
    """
    from app.api.deps import get_usuario_actual
    from app.main import app

    admin = UsuarioRecord(
        id=9,
        nombre="YUDY ADMIN",
        email="yudy-admin@hus.gov.co",
        rol="SUPER_ADMIN",
        activo=1,
        password_hash=get_password_hash("xxxx"),
    )
    db_session.add(admin)
    db_session.commit()
    app.dependency_overrides[get_usuario_actual] = lambda: admin
    return admin


def _texto_pdf(pdf: bytes) -> str:
    """Texto plano del PDF (espacios normalizados, como en los tests del PDF)."""
    import re
    from io import BytesIO

    from PyPDF2 import PdfReader

    crudo = "".join(p.extract_text() or "" for p in PdfReader(BytesIO(pdf)).pages)
    return re.sub(r"\s+", " ", crudo)


class TestAuditoria:
    def _setup(self, client, envio=ENV, factura=F1, valor=250700):
        _subir_radicacion(client, [_rad_fila(envio, factura, valor)])
        _subir_dgreport(client, [factura])  # con F.E.: se permite radicar
        o = _crear_oficio(client, f"FHUS-{envio}", "2026-07-20T08:00")
        _escribir(client, o["id"], envio)
        return o, _factura_id(client, factura)

    def test_sin_facturacion_electronica_no_se_radica(self, client):
        # Regla: CORREO F.E. = NO → solo se puede devolver.
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])  # SIN dgreport
        o = _crear_oficio(client, f"FHUS-{ENV}", "2026-07-20T08:00")
        _escribir(client, o["id"], ENV)
        fid = _factura_id(client, F1)
        r = _radicar(client, fid)
        assert r.status_code == 409
        assert (
            "facturación electrónica" in r.json()["detail"].lower()
            or "electr" in r.json()["detail"].lower()
        )
        # devolver sí se permite
        assert _devolver(client, fid).status_code == 200

    def test_radicar_registra_auditor(self, client):
        o, fid = self._setup(client)
        d = _radicar(client, fid).json()
        assert d["resultado"] == "RADICAR"
        assert d["estado"] == "RADICADA"
        assert d["auditor"] == "CLAUDIA"

    def test_devolver_sin_motivo_400(self, client):
        o, fid = self._setup(client)
        r = client.patch(f"/preauditoria/facturas/{fid}/auditar", json={"resultado": "DEVUELTA"})
        assert r.status_code == 400

    def test_devolver_incrementa_contador(self, client):
        o, fid = self._setup(client)
        d = _devolver(client, fid).json()
        assert d["resultado"] == "DEVUELTA"
        assert d["num_devoluciones"] == 1
        assert d["pendiente_subsanacion"] is True
        assert d["estado"] == "DEVUELTA_PEND_SUBSANACION"

    def test_subsanacion_reingreso_no_duplica(self, client):
        # F1 en envío ENV, devuelta; reingresa en un envío NUEVO (re-radicación)
        _subir_dgreport(client, [F1])
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o1 = _crear_oficio(client, "FHUS-1", "2026-07-20T08:00")
        _escribir(client, o1["id"], ENV)
        fid = _factura_id(client, F1)
        _devolver(client, fid)

        # la factura vuelve corregida en un envío nuevo
        _subir_radicacion(client, [_rad_fila("229999", F1, 250700)])
        o2 = _crear_oficio(client, "FHUS-2", "2026-07-21T08:00")
        r = _escribir(client, o2["id"], "229999")
        assert r.json()["reingresos"] == 1
        # NO se creó factura nueva: sigue habiendo una sola fila
        assert client.get("/preauditoria/consolidado").json()["total"] == 1
        d = client.get(f"/preauditoria/facturas/{F1}/historial").json()
        assert d["actual"]["ronda"] == 2
        assert d["actual"]["num_subsanacion"] == 1
        assert d["actual"]["estado"] == "EN_SUBSANACION"
        # radicar → SUBSANADA
        assert _radicar(client, d["actual"]["id"]).json()["estado"] == "SUBSANADA"

    def test_clics_repetidos_no_duplican_el_historial(self, client):
        """Si el servidor va lento el auditor hace varios clics: solo uno debe
        quedar. (En producción llegaron 5 eventos RADICADA de la misma factura.)"""
        self._setup(client)
        fid = _factura_id(client, F1)
        respuestas = [_radicar(client, fid) for _ in range(5)]
        assert respuestas[0].status_code == 200
        assert all(r.status_code == 409 for r in respuestas[1:])
        h = client.get(f"/preauditoria/facturas/{F1}/historial").json()
        radicadas = [e for e in h["eventos"] if e["tipo"] == "RADICADA"]
        assert len(radicadas) == 1

    def test_devolver_sin_motivo_no_cambia_el_estado(self, client):
        """La validación va ANTES de tomar la factura: un intento inválido no
        puede dejarla marcada como devuelta."""
        self._setup(client)
        fid = _factura_id(client, F1)
        r = client.patch(
            f"/preauditoria/facturas/{fid}/auditar",
            json={"resultado": "DEVUELTA", "motivo_devolucion": "   "},
        )
        assert r.status_code == 400
        d = client.get(f"/preauditoria/facturas/{fid}").json()
        assert d["resultado"] == "PENDIENTE" and d["num_devoluciones"] == 0
        # y después sí se puede decidir normalmente
        assert _radicar(client, fid).status_code == 200

    def test_la_observacion_se_guarda_y_queda_en_el_historial(self, client):
        """Lo que el auditor escribe al radicar ya no se pierde."""
        self._setup(client)
        fid = _factura_id(client, F1)
        r = client.patch(
            f"/preauditoria/facturas/{fid}/auditar",
            json={"resultado": "RADICAR", "observaciones": "Soportes revisados con la EPS"},
        )
        assert r.status_code == 200, r.text
        assert r.json()["observaciones"] == "Soportes revisados con la EPS"
        h = client.get(f"/preauditoria/facturas/{F1}/historial").json()
        assert h["actual"]["observaciones"] == "Soportes revisados con la EPS"
        radicada = [e for e in h["eventos"] if e["tipo"] == "RADICADA"][0]
        assert radicada["observaciones"] == "Soportes revisados con la EPS"

    def test_lo_escrito_en_motivo_al_radicar_no_se_pierde(self, client):
        """Caso real 29-07-2026: el auditor escribió "OKAY SOPORTES" en el
        recuadro de motivo (el de arriba) y radicó. El texto se descartaba en
        silencio y el historial salía vacío. Ahora se guarda como observación."""
        self._setup(client)
        fid = _factura_id(client, F1)
        r = client.patch(
            f"/preauditoria/facturas/{fid}/auditar",
            json={"resultado": "RADICAR", "motivo_devolucion": "OKAY SOPORTES"},
        )
        assert r.status_code == 200, r.text
        assert r.json()["observaciones"] == "OKAY SOPORTES"
        h = client.get(f"/preauditoria/facturas/{F1}/historial").json()
        assert h["actual"]["observaciones"] == "OKAY SOPORTES"
        radicada = [e for e in h["eventos"] if e["tipo"] == "RADICADA"][0]
        assert radicada["observaciones"] == "OKAY SOPORTES"

    def test_si_escribe_en_los_dos_recuadros_se_conservan_ambos(self, client):
        self._setup(client)
        fid = _factura_id(client, F1)
        r = client.patch(
            f"/preauditoria/facturas/{fid}/auditar",
            json={
                "resultado": "RADICAR",
                "observaciones": "Soportes completos",
                "motivo_devolucion": "OKAY SOPORTES",
            },
        )
        assert r.status_code == 200, r.text
        assert r.json()["observaciones"] == "Soportes completos — OKAY SOPORTES"

    def test_el_mismo_texto_en_los_dos_recuadros_no_se_duplica(self, client):
        self._setup(client)
        fid = _factura_id(client, F1)
        r = client.patch(
            f"/preauditoria/facturas/{fid}/auditar",
            json={
                "resultado": "RADICAR",
                "observaciones": "OKAY SOPORTES",
                "motivo_devolucion": "OKAY SOPORTES",
            },
        )
        assert r.json()["observaciones"] == "OKAY SOPORTES"

    def test_al_devolver_el_motivo_sigue_siendo_motivo(self, client):
        """El arreglo del radicar no puede contaminar la devolución: ahí el
        motivo es motivo y no debe colarse en la observación."""
        self._setup(client)
        fid = _factura_id(client, F1)
        r = client.patch(
            f"/preauditoria/facturas/{fid}/auditar",
            json={"resultado": "DEVUELTA", "motivo_devolucion": "Falta epicrisis"},
        )
        assert r.status_code == 200, r.text
        assert r.json()["motivo_devolucion"] == "Falta epicrisis"
        assert not r.json()["observaciones"]
        h = client.get(f"/preauditoria/facturas/{F1}/historial").json()
        dev = [e for e in h["eventos"] if e["tipo"] == "DEVUELTA"][0]
        assert dev["motivo"] == "Falta epicrisis"

    def test_anotar_la_observacion_de_una_factura_ya_radicada(self, client, db_session):
        """Las facturas que ya se radicaron sin observación se pueden anotar
        ahora (por administración), sin revertirlas ni volverlas a radicar."""
        self._setup(client)
        fid = _factura_id(client, F1)
        assert _radicar(client, fid).status_code == 200
        _actuar_como_admin(db_session)
        r = client.patch(
            f"/preauditoria/facturas/{fid}/observacion",
            json={"observaciones": "OKAY SOPORTES"},
        )
        assert r.status_code == 200, r.text
        assert r.json()["observaciones"] == "OKAY SOPORTES"
        # la decisión NO cambió
        assert r.json()["resultado"] == "RADICAR"
        assert r.json()["estado"] == "RADICADA"
        h = client.get(f"/preauditoria/facturas/{F1}/historial").json()
        anotacion = [e for e in h["eventos"] if e["tipo"] == "OBSERVACION"]
        assert len(anotacion) == 1
        assert anotacion[0]["observaciones"] == "OKAY SOPORTES"
        # y la RADICADA original sigue intacta (historial inmutable)
        assert len([e for e in h["eventos"] if e["tipo"] == "RADICADA"]) == 1

    def test_no_se_puede_anotar_una_observacion_vacia(self, client, db_session):
        self._setup(client)
        fid = _factura_id(client, F1)
        _radicar(client, fid)
        _actuar_como_admin(db_session)
        r = client.patch(f"/preauditoria/facturas/{fid}/observacion", json={"observaciones": "   "})
        assert r.status_code == 400

    def test_corregir_la_observacion_de_una_devuelta_corrige_el_motivo(self, client, db_session):
        """En una devuelta, la observación ES el motivo que saldrá en el
        oficio: corregirla antes de generar el oficio corrige lo que este
        imprimirá."""
        self._setup(client)
        fid = _factura_id(client, F1)
        assert _devolver(client, fid, motivo="Falta FURIPS").status_code == 200
        _actuar_como_admin(db_session)
        r = client.patch(
            f"/preauditoria/facturas/{fid}/observacion",
            json={"observaciones": "Falta FURIPS y epicrisis"},
        )
        assert r.status_code == 200, r.text
        assert r.json()["motivo_devolucion"] == "Falta FURIPS y epicrisis"
        # aún no hay oficio generado: no se corrigió ningún PDF
        assert r.json()["oficio_actualizado"] is False
        # el evento de devolución (el que se sellará en el oficio) quedó corregido
        h = client.get(f"/preauditoria/facturas/{F1}/historial").json()
        dev = [e for e in h["eventos"] if e["tipo"] == "DEVUELTA"][0]
        assert dev["motivo"] == "Falta FURIPS y epicrisis"

    def test_corregir_la_observacion_despues_del_oficio_corrige_el_pdf(self, client, db_session):
        """Caso real: se generó el oficio y el texto salió con un error.
        Administración corrige la observación y el PDF (que se arma al
        abrirlo) queda corregido de inmediato."""
        o, fid = self._setup(client)
        assert _devolver(client, fid, motivo="NOMBRE ERRADO EN FURIPS").status_code == 200
        dev = client.post(f"/preauditoria/oficios/{o['id']}/oficio-devolucion").json()
        pdf1 = client.get(dev["pdf_url"]).content
        _actuar_como_admin(db_session)
        r = client.patch(
            f"/preauditoria/facturas/{fid}/observacion",
            json={"observaciones": "VERIFICAR Y CORREGIR LOS CAMPOS DEL FURIPS"},
        )
        assert r.status_code == 200, r.text
        assert r.json()["oficio_actualizado"] is True
        # el evento sellado en el oficio quedó con el texto corregido
        h = client.get(f"/preauditoria/facturas/{F1}/historial").json()
        dev_evt = [e for e in h["eventos"] if e["tipo"] == "DEVUELTA"][0]
        assert dev_evt["motivo"] == "VERIFICAR Y CORREGIR LOS CAMPOS DEL FURIPS"
        # y el PDF se rearma con el texto nuevo (el viejo tenía el errado)
        pdf2 = client.get(dev["pdf_url"]).content
        assert pdf2[:5] == b"%PDF-"
        assert "NOMBRE ERRADO EN FURIPS" in _texto_pdf(pdf1)
        assert "VERIFICAR Y CORREGIR LOS CAMPOS DEL FURIPS" in _texto_pdf(pdf2)
        assert "NOMBRE ERRADO EN FURIPS" not in _texto_pdf(pdf2)
        # la decisión no cambió y la corrección quedó en el historial
        assert r.json()["resultado"] == "DEVUELTA"
        anot = [e for e in h["eventos"] if e["tipo"] == "OBSERVACION"]
        assert len(anot) == 1 and anot[0]["auditor"] == "YUDY ADMIN"

    def test_la_correccion_no_toca_los_oficios_de_rondas_anteriores(self, client, db_session):
        """Si la factura ya reingresó, corregir la observación corrige la
        devolución de ESTA ronda; el oficio viejo (ya entregado) no se toca."""
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o1 = _crear_oficio(client, "FHUS-COR-1", "2026-07-20T08:00")
        _escribir(client, o1["id"], ENV)
        _devolver(client, _factura_id(client, F1), motivo="Motivo ronda 1")
        client.post(f"/preauditoria/oficios/{o1['id']}/oficio-devolucion")
        # reingresa y vuelve a devolverse en la ronda 2
        _subir_radicacion(client, [_rad_fila("229994", F1, 250700)])
        o2 = _crear_oficio(client, "FHUS-COR-2", "2026-07-21T08:00")
        _escribir(client, o2["id"], "229994")
        fid = _factura_id(client, F1)
        assert _devolver(client, fid, motivo="Motivo ronda 2").status_code == 200
        _actuar_como_admin(db_session)
        r = client.patch(
            f"/preauditoria/facturas/{fid}/observacion",
            json={"observaciones": "Motivo ronda 2 corregido"},
        )
        assert r.status_code == 200, r.text
        assert r.json()["oficio_actualizado"] is False  # la ronda 2 aún no tiene oficio
        h = client.get(f"/preauditoria/facturas/{F1}/historial").json()
        # la devolución de la ronda 1 (sellada en el oficio viejo) sigue intacta
        assert [e for e in h["eventos"] if e["tipo"] == "DEVUELTA"][0]["motivo"] == "Motivo ronda 1"
        # la de la ronda 2 quedó corregida
        r2 = [e for e in h["eventos"] if e["tipo"] == "NUEVAMENTE_DEVUELTA"][0]
        assert r2["motivo"] == "Motivo ronda 2 corregido"

    def test_la_fila_revertida_muestra_la_observacion(self, client, db_session):
        """Antes la fila REVERTIDA salía siempre vacía en el historial."""
        self._setup(client)
        fid = _factura_id(client, F1)
        client.patch(
            f"/preauditoria/facturas/{fid}/auditar",
            json={"resultado": "RADICAR", "observaciones": "OKAY SOPORTES"},
        )
        _actuar_como_admin(db_session)  # revertir una decidida es de administración
        r = client.patch(f"/preauditoria/facturas/{fid}/auditar", json={"resultado": "PENDIENTE"})
        assert r.status_code == 200, r.text
        h = client.get(f"/preauditoria/facturas/{F1}/historial").json()
        rev = [e for e in h["eventos"] if e["tipo"] == "REVERTIDA"][0]
        assert rev["observaciones"] == "OKAY SOPORTES"

    def test_no_doble_devolucion_sin_reingreso(self, client):
        # BUG: devolver dos veces sin reingreso NO debe inflar el contador.
        o, fid = self._setup(client)
        assert _devolver(client, fid).json()["num_devoluciones"] == 1
        r = _devolver(client, fid, motivo="otra vez")
        assert r.status_code == 409
        # el contador sigue en 1
        d = client.get(f"/preauditoria/facturas/{fid}").json()
        assert d["num_devoluciones"] == 1

    def test_no_radicar_ya_radicada(self, client):
        o, fid = self._setup(client)
        assert _radicar(client, fid).status_code == 200
        assert _radicar(client, fid).status_code == 409  # ya radicada

    def test_pdf_inmutable_tras_reingreso(self, client):
        # BUG 1: el PDF de un oficio de devolución ya emitido no debe cambiar
        # aunque la factura reingrese después.
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o1 = _crear_oficio(client, "FHUS-INM-1", "2026-07-20T08:00")
        _escribir(client, o1["id"], ENV)
        _devolver(client, _factura_id(client, F1))
        dev = client.post(f"/preauditoria/oficios/{o1['id']}/oficio-devolucion").json()
        pdf1 = client.get(dev["pdf_url"]).content
        # la factura reingresa en un envío nuevo
        _subir_radicacion(client, [_rad_fila("229995", F1, 250700)])
        o2 = _crear_oficio(client, "FHUS-INM-2", "2026-07-21T08:00")
        _escribir(client, o2["id"], "229995")
        # el PDF del primer oficio SIGUE conteniendo la factura (mismo tamaño, %PDF)
        pdf2 = client.get(dev["pdf_url"]).content
        assert pdf2[:5] == b"%PDF-"
        assert len(pdf2) > 2000  # no quedó vacío
        assert abs(len(pdf1) - len(pdf2)) < 500  # contenido equivalente

    def test_tope_3_devoluciones(self, client):
        _subir_dgreport(client, [F1])
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        # ronda 1
        o = _crear_oficio(client, "FHUS-A", "2026-07-20T08:00")
        _escribir(client, o["id"], ENV)
        fid = _factura_id(client, F1)
        assert _devolver(client, fid).status_code == 200  # dev 1
        # reingresos y devoluciones 2 y 3
        for i, env in enumerate(["229991", "229992"], start=2):
            _subir_radicacion(client, [_rad_fila(env, F1, 250700)])
            oo = _crear_oficio(client, f"FHUS-{env}", f"2026-07-2{i}T08:00")
            _escribir(client, oo["id"], env)
            fid = _factura_id(client, F1)
            assert _devolver(client, fid).status_code == 200
        # 4.º reingreso: devolver debe bloquearse
        _subir_radicacion(client, [_rad_fila("229993", F1, 250700)])
        o4 = _crear_oficio(client, "FHUS-229993", "2026-07-24T08:00")
        _escribir(client, o4["id"], "229993")
        fid = _factura_id(client, F1)
        r = _devolver(client, fid)
        assert r.status_code == 409
        assert "máximo 3" in r.json()["detail"]
        # radicar sí se permite
        assert _radicar(client, fid).json()["estado"] == "SUBSANADA"


# ------------------------------------------------------------------
# Auto-sync: corregir la fuente se refleja en el consolidado
# ------------------------------------------------------------------


class TestAutoSync:
    def test_corregir_valor_y_nit_se_refleja(self, client):
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700, nit="111")])
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)
        # corregir la fuente (nuevo valor y NIT)
        r = _subir_radicacion(client, [_rad_fila(ENV, F1, 999999, nit="860002400")])
        assert r.json()["actualizadas"] == 1
        item = client.get("/preauditoria/consolidado").json()["items"][0]
        assert item["valor"] == 999999
        assert item["nit"] == "860002400"

    def test_fecha_factura_no_se_corre_un_dia(self, client):
        # HALLAZGO 1: F_FACTURA de la fuente (2026-05-20) no debe mostrarse 05-19.
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)
        item = client.get("/preauditoria/consolidado").json()["items"][0]
        assert item["f_factura"].startswith("2026-05-20")

    def test_agregar_correo_se_refleja(self, client):
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)
        assert client.get("/preauditoria/consolidado").json()["items"][0]["correo_fe"] == "NO"
        _subir_dgreport(client, [F1])
        assert client.get("/preauditoria/consolidado").json()["items"][0]["correo_fe"] == "SI"


# ------------------------------------------------------------------
# Oficio de devolución PDF + estadísticas
# ------------------------------------------------------------------


class TestDevolucionYStats:
    def _setup_devueltas(self, client):
        _subir_radicacion(
            client,
            [_rad_fila(ENV, F1, 250700), _rad_fila(ENV, F2, 10615224), _rad_fila(ENV, F3, 73500)],
        )
        _subir_dgreport(client, [F1, F2, F3])
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)
        for f in (F1, F2):
            _devolver(client, _factura_id(client, f))
        return o

    def test_consecutivo_escrito_por_el_auditor(self, client):
        """La numeración la lleva SINAC por fuera: el auditor escribe la que va."""
        anio = datetime.now().year
        # el sistema sugiere el siguiente según lo registrado aquí
        sug = client.get("/preauditoria/consecutivo-sugerido").json()
        assert sug["consecutivo"] == f"DEV-PRE-AUD-0001-{anio}"

        o = self._setup_devueltas(client)
        # …pero el auditor escribe el 89, que es el que le corresponde
        r = client.post(
            f"/preauditoria/oficios/{o['id']}/oficio-devolucion", json={"consecutivo": "89"}
        )
        assert r.status_code == 200, r.text
        assert r.json()["consecutivo"] == f"DEV-PRE-AUD-0089-{anio}"
        # el PDF sale con ese consecutivo
        pdf = client.get(r.json()["pdf_url"])
        assert pdf.status_code == 200 and pdf.content[:4] == b"%PDF"
        # y la sugerencia siguiente continúa desde ahí
        assert client.get("/preauditoria/consecutivo-sugerido").json()["consecutivo"] == (
            f"DEV-PRE-AUD-0090-{anio}"
        )

    def test_consecutivo_completo_y_repetido(self, client):
        anio = datetime.now().year
        o = self._setup_devueltas(client)
        r = client.post(
            f"/preauditoria/oficios/{o['id']}/oficio-devolucion",
            json={"consecutivo": f"dev-pre-aud-0120-{anio}"},  # completo, en minúsculas
        )
        assert r.status_code == 200, r.text
        assert r.json()["consecutivo"] == f"DEV-PRE-AUD-0120-{anio}"

        # el mismo consecutivo en otro oficio: se rechaza con mensaje claro
        _subir_radicacion(client, [_rad_fila("777777", "HUS0000700001", 5000)])
        o2 = _crear_oficio(client, "FHUS-OTRO-1", "2026-07-21T08:00")
        _escribir(client, o2["id"], "777777")
        _devolver(client, _factura_id(client, "HUS0000700001"))
        r2 = client.post(
            f"/preauditoria/oficios/{o2['id']}/oficio-devolucion",
            json={"consecutivo": f"DEV-PRE-AUD-0120-{anio}"},
        )
        assert r2.status_code == 409
        assert "ya fue usado" in r2.json()["detail"]

    def test_consecutivo_vacio_usa_el_sugerido(self, client):
        o = self._setup_devueltas(client)
        r = client.post(
            f"/preauditoria/oficios/{o['id']}/oficio-devolucion", json={"consecutivo": "  "}
        )
        assert r.status_code == 200, r.text
        assert r.json()["consecutivo"] == f"DEV-PRE-AUD-0001-{datetime.now().year}"

    def test_consecutivo_y_pdf(self, client):
        o = self._setup_devueltas(client)
        r = client.post(f"/preauditoria/oficios/{o['id']}/oficio-devolucion")
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["consecutivo"] == f"DEV-PRE-AUD-0001-{datetime.now().year}"
        assert d["total_facturas"] == 2
        pdf = client.get(d["pdf_url"])
        assert pdf.status_code == 200 and pdf.content[:5] == b"%PDF-"
        assert len(pdf.content) > 2000

    def test_sin_devueltas_400(self, client):
        _subir_radicacion(client, [_rad_fila(ENV, F3, 73500)])
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)
        assert client.post(f"/preauditoria/oficios/{o['id']}/oficio-devolucion").status_code == 400

    def test_estadisticas(self, client):
        self._setup_devueltas(client)
        _radicar(client, _factura_id(client, F3))
        d = client.get("/preauditoria/estadisticas").json()
        assert d["total_facturas"] == 3
        assert d["auditadas"] == 3
        assert d["devueltas"] == 2
        assert d["radicar"] == 1
        assert d["por_auditor"][0]["auditor"] == "CLAUDIA"
        assert d["valor_total"] == 250700 + 10615224 + 73500


# ------------------------------------------------------------------
# Eliminar oficios (solo admin/coordinador)
# ------------------------------------------------------------------


class TestEliminarOficio:
    def _admin(self, db_session):
        u = UsuarioRecord(
            id=2,
            nombre="ADMIN",
            email="admin@hus.gov.co",
            rol="SUPER_ADMIN",
            activo=1,
            password_hash=get_password_hash("xxxx"),
        )
        db_session.add(u)
        db_session.commit()
        return u

    def test_auditor_no_puede_eliminar(self, client):
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o = _crear_oficio(client)
        r = client.delete(f"/preauditoria/oficios/{o['id']}")
        assert r.status_code == 403  # el fixture es rol AUDITOR

    def test_admin_elimina_y_libera_envio(self, client, db_session):
        from app.api.deps import get_coordinador_o_admin
        from app.main import app

        admin = self._admin(db_session)
        app.dependency_overrides[get_coordinador_o_admin] = lambda: admin
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)
        r = client.delete(f"/preauditoria/oficios/{o['id']}")
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["facturas_borradas"] == 1 and d["envios_liberados"] == 1
        # el consolidado quedó limpio y el envío se puede volver a escribir
        assert client.get("/preauditoria/consolidado").json()["total"] == 0
        o2 = _crear_oficio(client, "FHUS-NUEVO-1", "2026-07-21T08:00")
        assert _escribir(client, o2["id"], ENV).json()["nuevas"] == 1

    def test_no_elimina_con_pdf_emitido(self, client, db_session):
        from app.api.deps import get_coordinador_o_admin
        from app.main import app

        admin = self._admin(db_session)
        app.dependency_overrides[get_coordinador_o_admin] = lambda: admin
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)
        _devolver(client, _factura_id(client, F1))
        client.post(f"/preauditoria/oficios/{o['id']}/oficio-devolucion")
        r = client.delete(f"/preauditoria/oficios/{o['id']}")
        assert r.status_code == 409
        assert "devolución" in r.json()["detail"]

    def test_masivo_reporta_rechazados(self, client, db_session):
        from app.api.deps import get_coordinador_o_admin
        from app.main import app

        admin = self._admin(db_session)
        app.dependency_overrides[get_coordinador_o_admin] = lambda: admin
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o = _crear_oficio(client)
        r = client.post("/preauditoria/oficios/eliminar-masivo", json={"ids": [o["id"], 9999]})
        d = r.json()
        assert len(d["eliminados"]) == 1
        assert d["rechazados"][0]["id"] == 9999


# ------------------------------------------------------------------
# Eliminar un oficio de DEVOLUCIÓN generado por error
# ------------------------------------------------------------------


class TestEliminarOficioDevolucion:
    def _admin(self, db_session):
        u = UsuarioRecord(
            id=3,
            nombre="ADMIN",
            email="admin2@hus.gov.co",
            rol="SUPER_ADMIN",
            activo=1,
            password_hash=get_password_hash("xxxx"),
        )
        db_session.add(u)
        db_session.commit()
        return u

    def _con_devolucion(self, client):
        """Deja una factura devuelta y su oficio de devolución generado."""
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)
        _devolver(client, _factura_id(client, F1))
        dev = client.post(f"/preauditoria/oficios/{o['id']}/oficio-devolucion").json()
        return o, dev

    def test_auditor_no_puede_eliminar(self, client):
        _o, dev = self._con_devolucion(client)
        r = client.delete(f"/preauditoria/oficios-devolucion/{dev['id']}")
        assert r.status_code == 403  # el fixture es rol AUDITOR

    def test_admin_elimina_y_libera_las_facturas(self, client, db_session):
        from app.api.deps import get_coordinador_o_admin
        from app.main import app

        app.dependency_overrides[get_coordinador_o_admin] = lambda: self._admin(db_session)
        _o, dev = self._con_devolucion(client)
        r = client.delete(f"/preauditoria/oficios-devolucion/{dev['id']}")
        assert r.status_code == 200, r.text
        assert r.json()["consecutivo"] == dev["consecutivo"]
        assert r.json()["facturas_liberadas"] == 1
        # desaparece del listado y su PDF ya no existe
        assert client.get("/preauditoria/oficios-devolucion").json()["total"] == 0
        assert client.get(f"/preauditoria/oficios-devolucion/{dev['id']}/pdf").status_code == 404
        # la factura SIGUE devuelta: no se le cambió la decisión
        f = client.get(f"/preauditoria/facturas/{_factura_id(client, F1)}").json()
        assert f["resultado"] == "DEVUELTA"
        assert f["num_devoluciones"] == 1
        assert f["oficio_devolucion_id"] is None

    def test_tras_eliminar_se_puede_generar_el_oficio_de_nuevo(self, client, db_session):
        """El caso de uso real: el consecutivo salió equivocado y hay que
        rehacer el oficio con el número correcto."""
        from app.api.deps import get_coordinador_o_admin
        from app.main import app

        app.dependency_overrides[get_coordinador_o_admin] = lambda: self._admin(db_session)
        o, dev = self._con_devolucion(client)
        client.delete(f"/preauditoria/oficios-devolucion/{dev['id']}")
        r = client.post(
            f"/preauditoria/oficios/{o['id']}/oficio-devolucion",
            json={"consecutivo": "77"},
        )
        assert r.status_code == 200, r.text
        assert r.json()["consecutivo"].startswith("DEV-PRE-AUD-0077-")
        assert r.json()["total_facturas"] == 1

    def test_el_consecutivo_borrado_se_puede_reutilizar(self, client, db_session):
        from app.api.deps import get_coordinador_o_admin
        from app.main import app

        app.dependency_overrides[get_coordinador_o_admin] = lambda: self._admin(db_session)
        o, dev = self._con_devolucion(client)
        usado = dev["consecutivo"]
        client.delete(f"/preauditoria/oficios-devolucion/{dev['id']}")
        r = client.post(
            f"/preauditoria/oficios/{o['id']}/oficio-devolucion",
            json={"consecutivo": usado},
        )
        assert r.status_code == 200, r.text
        assert r.json()["consecutivo"] == usado

    def test_tras_eliminar_la_factura_se_puede_revertir(self, client, db_session):
        """Con el oficio emitido la reversión está bloqueada; al eliminarlo,
        la factura vuelve a poder corregirse."""
        from app.api.deps import get_coordinador_o_admin
        from app.main import app

        app.dependency_overrides[get_coordinador_o_admin] = lambda: self._admin(db_session)
        _o, dev = self._con_devolucion(client)
        fid = _factura_id(client, F1)
        _actuar_como_admin(db_session)  # revertir una decidida es de administración
        bloqueada = client.patch(
            f"/preauditoria/facturas/{fid}/auditar", json={"resultado": "PENDIENTE"}
        )
        assert bloqueada.status_code == 409
        client.delete(f"/preauditoria/oficios-devolucion/{dev['id']}")
        r = client.patch(f"/preauditoria/facturas/{fid}/auditar", json={"resultado": "PENDIENTE"})
        assert r.status_code == 200, r.text
        assert client.get(f"/preauditoria/facturas/{fid}").json()["resultado"] == "PENDIENTE"

    def test_el_historial_de_la_devolucion_no_se_borra(self, client, db_session):
        """Queda constancia de que la factura sí fue devuelta ese día."""
        from app.api.deps import get_coordinador_o_admin
        from app.main import app

        app.dependency_overrides[get_coordinador_o_admin] = lambda: self._admin(db_session)
        _o, dev = self._con_devolucion(client)
        client.delete(f"/preauditoria/oficios-devolucion/{dev['id']}")
        h = client.get(f"/preauditoria/facturas/{F1}/historial").json()
        devueltas = [e for e in h["eventos"] if e["tipo"] == "DEVUELTA"]
        assert len(devueltas) == 1
        assert devueltas[0]["motivo"]

    def test_eliminar_uno_que_no_existe(self, client, db_session):
        from app.api.deps import get_coordinador_o_admin
        from app.main import app

        app.dependency_overrides[get_coordinador_o_admin] = lambda: self._admin(db_session)
        assert client.delete("/preauditoria/oficios-devolucion/9999").status_code == 404


# ------------------------------------------------------------------
# Eliminar UN envío cargado por error (solo admin/coordinador)
# ------------------------------------------------------------------


class TestEliminarEnvio:
    def _admin(self, db_session):
        u = UsuarioRecord(
            id=4,
            nombre="ADMIN",
            email="admin4@hus.gov.co",
            rol="SUPER_ADMIN",
            activo=1,
            password_hash=get_password_hash("xxxx"),
        )
        db_session.add(u)
        db_session.commit()
        return u

    def _como_admin(self, db_session):
        from app.api.deps import get_coordinador_o_admin
        from app.main import app

        app.dependency_overrides[get_coordinador_o_admin] = lambda: self._admin(db_session)

    def test_auditor_no_puede(self, client):
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)
        assert client.delete(f"/preauditoria/oficios/{o['id']}/envios/{ENV}").status_code == 403

    def test_borra_solo_ese_envio_y_lo_libera(self, client, db_session):
        self._como_admin(db_session)
        otro = "229999"
        _subir_radicacion(
            client,
            [_rad_fila(ENV, F1, 250700), _rad_fila(ENV, F2, 98000), _rad_fila(otro, F3, 55000)],
        )
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)
        _escribir(client, o["id"], otro)
        assert client.get("/preauditoria/consolidado").json()["total"] == 3

        r = client.delete(f"/preauditoria/oficios/{o['id']}/envios/{ENV}")
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["envio"] == ENV and d["facturas_borradas"] == 2

        # el otro envío quedó intacto…
        cons = client.get("/preauditoria/consolidado").json()
        assert cons["total"] == 1 and cons["items"][0]["factura"] == F3
        # …el oficio sigue existiendo con un solo envío…
        assert [e["envio"] for e in d["oficio"]["envios_escritos"]] == [otro]
        # …y el envío borrado se puede volver a escribir
        assert _escribir(client, o["id"], ENV).json()["nuevas"] == 2

    def test_envio_inexistente(self, client, db_session):
        self._como_admin(db_session)
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o = _crear_oficio(client)
        r = client.delete(f"/preauditoria/oficios/{o['id']}/envios/{ENV}")
        assert r.status_code == 404
        assert "no está cargado" in r.json()["detail"]

    def test_no_borra_con_pdf_emitido(self, client, db_session):
        self._como_admin(db_session)
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)
        _devolver(client, _factura_id(client, F1))
        client.post(f"/preauditoria/oficios/{o['id']}/oficio-devolucion")
        r = client.delete(f"/preauditoria/oficios/{o['id']}/envios/{ENV}")
        assert r.status_code == 409
        assert "devolución ya" in r.json()["detail"]

    def test_reingreso_vuelve_a_su_devolucion(self, client, db_session):
        """Si el envío traía una factura a subsanar, quitarlo la devuelve a su
        estado anterior sin borrarla ni perder el historial."""
        self._como_admin(db_session)
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o1 = _crear_oficio(client, "FHUS-1")
        _escribir(client, o1["id"], ENV)
        _devolver(client, _factura_id(client, F1), motivo="Falta epicrisis")
        # reingresa en otro oficio/envío para subsanar
        _subir_radicacion(client, [_rad_fila("229305", F1, 250700)])
        o2 = _crear_oficio(client, "FHUS-2", "2026-07-21T08:00")
        assert _escribir(client, o2["id"], "229305").json()["reingresos"] == 1

        r = client.delete(f"/preauditoria/oficios/{o2['id']}/envios/229305")
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["facturas_borradas"] == 0 and d["subsanaciones_revertidas"] == 1
        # la factura sigue viva, devuelta y en su oficio original
        h = client.get(f"/preauditoria/facturas/{F1}/historial").json()
        assert h["actual"]["resultado"] == "DEVUELTA"
        assert h["actual"]["oficio_fhus"] == "FHUS-1"
        assert h["actual"]["motivo_devolucion"] == "Falta epicrisis"


# ------------------------------------------------------------------
# Limpiar TODO el módulo (solo admin/coordinador)
# ------------------------------------------------------------------


class TestLimpiarTodo:
    def _admin(self, db_session):
        u = UsuarioRecord(
            id=3,
            nombre="ADMIN",
            email="admin2@hus.gov.co",
            rol="SUPER_ADMIN",
            activo=1,
            password_hash=get_password_hash("xxxx"),
        )
        db_session.add(u)
        db_session.commit()
        return u

    def _poblar(self, client):
        """Deja el módulo con datos de todas las tablas: 1 oficio, 1 envío,
        2 facturas (1 radicada + 1 devuelta), 4 eventos y 1 oficio de
        devolución emitido."""
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700), _rad_fila(ENV, F2, 98000)])
        _subir_dgreport(client, [F1])
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)
        _radicar(client, _factura_id(client, F1))
        _devolver(client, _factura_id(client, F2))
        assert client.post(f"/preauditoria/oficios/{o['id']}/oficio-devolucion").status_code == 200
        return o

    def test_auditor_no_puede(self, client):
        r = client.post("/preauditoria/admin/limpiar-todo", json={"confirmacion": "BORRAR TODO"})
        assert r.status_code == 403  # el fixture es rol AUDITOR

    def test_exige_confirmacion_exacta(self, client, db_session):
        from app.api.deps import get_coordinador_o_admin
        from app.main import app

        app.dependency_overrides[get_coordinador_o_admin] = lambda: self._admin(db_session)
        r = client.post("/preauditoria/admin/limpiar-todo", json={"confirmacion": "si, borrar"})
        assert r.status_code == 400
        assert "BORRAR TODO" in r.json()["detail"]

    def test_limpia_proceso_y_conserva_fuentes(self, client, db_session):
        from app.api.deps import get_coordinador_o_admin
        from app.main import app

        app.dependency_overrides[get_coordinador_o_admin] = lambda: self._admin(db_session)
        self._poblar(client)
        # La confirmación tolera minúsculas/espacios.
        r = client.post("/preauditoria/admin/limpiar-todo", json={"confirmacion": " borrar todo "})
        assert r.status_code == 200, r.text
        b = r.json()["borrado"]
        assert b["oficios_recepcion"] == 1
        assert b["facturas"] == 2
        assert b["eventos"] == 4  # 2 escritas + 1 radicada + 1 devuelta
        assert b["envios_cargados"] == 1
        assert b["oficios_devolucion"] == 1
        assert "fuente_radicacion" not in b  # las fuentes NO se tocaron
        # El proceso quedó en cero…
        assert client.get("/preauditoria/oficios").json()["total"] == 0
        assert client.get("/preauditoria/consolidado").json()["total"] == 0
        assert client.get("/preauditoria/oficios-devolucion").json()["total"] == 0
        # …pero las fuentes siguen cargadas.
        res = client.get("/preauditoria/fuentes/resumen").json()
        assert res["radicacion_facturas"] == 2
        assert res["dgreport_facturas"] == 1
        # Se puede empezar de cero: mismo envío re-escribible y consecutivo
        # SINAC reiniciado en 0001.
        o2 = _crear_oficio(client, "FHUS-DESPUES-1")
        assert _escribir(client, o2["id"], ENV).json()["nuevas"] == 2
        _devolver(client, _factura_id(client, F2))
        d = client.post(f"/preauditoria/oficios/{o2['id']}/oficio-devolucion").json()
        assert d["consecutivo"].startswith("DEV-PRE-AUD-0001-")

    def test_limpia_incluyendo_fuentes(self, client, db_session):
        from app.api.deps import get_coordinador_o_admin
        from app.main import app

        app.dependency_overrides[get_coordinador_o_admin] = lambda: self._admin(db_session)
        self._poblar(client)
        r = client.post(
            "/preauditoria/admin/limpiar-todo",
            json={"confirmacion": "BORRAR TODO", "incluir_fuentes": True},
        )
        assert r.status_code == 200, r.text
        b = r.json()["borrado"]
        assert b["fuente_radicacion"] == 2
        assert b["fuente_facturacion_electronica"] == 1
        res = client.get("/preauditoria/fuentes/resumen").json()
        assert res["radicacion_facturas"] == 0
        assert res["dgreport_facturas"] == 0


# ------------------------------------------------------------------
# Excel especial para oficios de ADRES (información completa)
# ------------------------------------------------------------------

ENTIDAD_ADRES = "ADMINISTRADORA DE LOS RECURSOS DEL SISTEMA GENERAL DE SEGURIDAD SOCIAL EN SALUD "


class TestExportAdres:
    def test_descarga_con_formato_del_consolidado(self, client):
        _subir_radicacion(
            client,
            [
                _rad_fila(ENV, F1, 250700, nit=svc.NIT_ADRES, entidad=ENTIDAD_ADRES),
                _rad_fila(ENV, F2, 98000, nit=svc.NIT_ADRES, entidad=ENTIDAD_ADRES),
                _rad_fila(ENV, F3, 55000),  # otra entidad: NO debe salir en el Excel
            ],
        )
        _subir_dgreport(client, [F1])
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)
        _radicar(client, _factura_id(client, F1))
        _devolver(client, _factura_id(client, F2), motivo="Falta FURIPS")
        assert client.post(f"/preauditoria/oficios/{o['id']}/oficio-devolucion").status_code == 200
        # La lista de oficios marca que este tiene facturas de ADRES.
        of = client.get("/preauditoria/oficios").json()["items"][0]
        assert of["tiene_adres"] is True
        r = client.get(f"/preauditoria/oficios/{o['id']}/export-adres.xlsx")
        assert r.status_code == 200
        from openpyxl import load_workbook

        ws = load_workbook(BytesIO(r.content)).active
        filas = list(ws.iter_rows(values_only=True))
        encabezado = list(filas[0])
        # Tramo SINAC en el orden del consolidado real…
        assert encabezado[:14] == [
            "Item",
            "Fecha_Recibido",
            "Envío",
            "AUD",
            "HUS",
            "Fecha_Factura",
            "Valor",
            "NIT",
            "Entidad",
            "Correo F.E.",
            "Observación Preauditoria Radicación SINAC",
            "Radicar_1",
            "Observaciones Adicionales",
            "Fecha_Entrega_Fact",
        ]
        # …seguido de las columnas de las otras áreas (vienen vacías).
        assert encabezado[14] == "Observación_FACTURACIÓN"
        assert encabezado[-1] == "INFOPOL"
        assert "Radicar_2" in encabezado
        assert len(filas) == 3  # encabezado + las 2 facturas ADRES (F3 no sale)
        assert [f[0] for f in filas[1:]] == [1, 2]  # Item consecutivo
        por_hus = {
            dict(zip(encabezado[:14], f))["HUS"]: dict(zip(encabezado[:14], f)) for f in filas[1:]
        }
        # F1 radicada
        f1 = por_hus[F1]
        assert f1["HUS"] == F1
        assert f1["AUD"] == "CLAUDIA"
        assert str(f1["NIT"]) == svc.NIT_ADRES
        assert f1["Entidad"] == ENTIDAD_ADRES.strip()
        assert f1["Correo F.E."] == "SI"
        assert f1["Valor"] == 250700
        assert f1["Radicar_1"] == "SI"
        assert f1["Observación Preauditoria Radicación SINAC"] == "SOPORTES COMPLETOS"
        assert f1["Fecha_Entrega_Fact"] in ("", None)  # la llena a mano quien entrega
        # F2 devuelta: motivo + consecutivo del oficio de devolución, Radicar_1=NO
        f2 = por_hus[F2]
        assert f2["HUS"] == F2
        assert f2["Correo F.E."] == "NO"
        assert f2["Radicar_1"] == "NO"
        obs2 = f2["Observación Preauditoria Radicación SINAC"]
        assert "Falta FURIPS" in obs2 and "DEV-PRE-AUD-0001" in obs2
        # el tramo de las otras áreas viene vacío
        assert all(v in ("", None) for v in filas[1][14:])

    def test_reconoce_adres_por_nombre_sin_nit(self, client):
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700, nit="901037916-6", entidad="ADRES ")])
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)
        assert client.get("/preauditoria/oficios").json()["items"][0]["tiene_adres"] is True

    def test_oficio_sin_adres_no_aplica(self, client):
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)
        of = client.get("/preauditoria/oficios").json()["items"][0]
        assert of["tiene_adres"] is False
        r = client.get(f"/preauditoria/oficios/{o['id']}/export-adres.xlsx")
        assert r.status_code == 404
        assert "ADRES" in r.json()["detail"]


# ------------------------------------------------------------------
# Cargue masivo por bloques (memoria acotada en el servidor de 1 GB)
# ------------------------------------------------------------------


class TestCargueMasivoPorBloques:
    """Los archivos grandes se guardan por bloques para no agotar la memoria.

    Estas pruebas fijan que trocear el cargue NO cambia el resultado: mismos
    conteos, sin duplicar y sin perder la idempotencia del upsert.
    """

    def _filas(self, n, valor=100.0):
        return [
            {
                "factura": f"HUS{i:07d}",
                "envio": ENV,
                "f_recibido": datetime(2026, 7, 20),
                "f_factura": datetime(2026, 7, 1),
                "valor": valor + i,
                "nit": svc.NIT_ADRES,
                "entidad": "ADRES",
                "estado_radicacion": "Registrado",
            }
            for i in range(n)
        ]

    def test_conteos_exactos_cruzando_bloques(self, db_session, monkeypatch):
        # 20-08-2026: se comparaba el diccionario COMPLETO con ==. El cargue
        # ahora devuelve además `filas_procesadas` y `completo`, para poder
        # decirle al auditor qué SÍ quedó guardado cuando un cargue se corta a
        # mitad. Lo que esta prueba cuida son los CONTEOS, así que se comparan
        # esos y no la forma exacta del diccionario.
        def _conteos(r: dict) -> dict:
            return {k: r[k] for k in ("nuevas", "actualizadas", "sin_cambio")}

        monkeypatch.setattr(svc, "TAM_BLOQUE_UPSERT", 3)  # 10 filas → 4 bloques
        filas = self._filas(10)
        primero = svc.upsert_radicacion(db_session, filas, "a.xlsx", "YESID")
        assert _conteos(primero) == {"nuevas": 10, "actualizadas": 0, "sin_cambio": 0}
        assert primero["completo"] is True
        assert primero["filas_procesadas"] == 10

        # re-subir el mismo archivo: todo sin cambio (idempotente)
        assert _conteos(svc.upsert_radicacion(db_session, filas, "a.xlsx", "YESID")) == {
            "nuevas": 0,
            "actualizadas": 0,
            "sin_cambio": 10,
        }
        # subir el archivo corregido: todo actualizado, sin duplicar
        for f in filas:
            f["valor"] += 1
        assert _conteos(svc.upsert_radicacion(db_session, filas, "b.xlsx", "YESID")) == {
            "nuevas": 0,
            "actualizadas": 10,
            "sin_cambio": 0,
        }
        assert db_session.query(RadicacionCuentaRecord).count() == 10

    def test_factura_repetida_en_otro_bloque_no_duplica(self, db_session, monkeypatch):
        monkeypatch.setattr(svc, "TAM_BLOQUE_UPSERT", 2)
        filas = self._filas(3)
        filas.append(dict(filas[0]))  # la misma factura, ya en otro bloque
        res = svc.upsert_radicacion(db_session, filas, "a.xlsx", "YESID")
        assert res["nuevas"] == 3 and res["sin_cambio"] == 1
        assert db_session.query(RadicacionCuentaRecord).count() == 3

    def test_dgreport_por_bloques(self, db_session, monkeypatch):
        monkeypatch.setattr(svc, "TAM_BLOQUE_UPSERT", 2)
        filas = [
            {"factura": f"HUS{i:07d}", "correo_fe": "SI", "numero_fe": f"CUFE{i}"} for i in range(7)
        ]
        assert svc.upsert_dgreport(db_session, filas, "d.xlsx", "YESID")["nuevas"] == 7
        assert svc.upsert_dgreport(db_session, filas, "d.xlsx", "YESID")["sin_cambio"] == 7
        assert db_session.query(DgReportRecord).count() == 7

    def test_archivo_mas_grande_que_un_bloque_real(self, db_session):
        """Con el tamaño de bloque real (2.000): un archivo de 2.500 filas."""
        filas = self._filas(2500)
        assert svc.upsert_radicacion(db_session, filas, "a.xlsx", "YESID")["nuevas"] == 2500
        assert db_session.query(RadicacionCuentaRecord).count() == 2500
        assert svc.upsert_radicacion(db_session, filas, "a.xlsx", "YESID")["sin_cambio"] == 2500

    def test_el_parseo_comparte_los_textos_repetidos(self, client):
        """ENTIDAD/NIT se repiten miles de veces: deben compartir una sola copia."""
        filas = [_rad_fila(ENV, f"HUS{i:07d}", 1000 + i) for i in range(50)]
        datos = svc.parsear_excel_radicacion(_excel(RAD_HEADERS, filas))
        entidades = [f["entidad"] for f in datos["facturas"]]
        assert len(entidades) == 50
        # todas las filas apuntan al MISMO objeto de texto, no a 50 copias
        assert len({id(e) for e in entidades}) == 1
        assert entidades[0] == "AXA COLPATRIA"  # y se sigue recortando


# ------------------------------------------------------------------
# Semáforo (unidad, sin cambios respecto de v1)
# ------------------------------------------------------------------


class TestSemaforo:
    RECIBIDO = datetime(2026, 7, 20, 8, 30)  # lunes

    def test_limite_tercer_dia_habil(self):
        assert (
            svc.calcular_semaforo(self.RECIBIDO, hoy=date(2026, 7, 20))["fecha_limite"]
            == "2026-07-23"
        )

    def test_amarillo_penultimo(self):
        assert svc.calcular_semaforo(self.RECIBIDO, hoy=date(2026, 7, 22))["estado"] == "AMARILLO"

    def test_rojo_ultimo(self):
        assert svc.calcular_semaforo(self.RECIBIDO, hoy=date(2026, 7, 23))["estado"] == "ROJO"

    def test_vencido(self):
        assert svc.calcular_semaforo(self.RECIBIDO, hoy=date(2026, 7, 24))["estado"] == "VENCIDO"

    def test_finde_no_cuenta(self):
        assert (
            svc.calcular_semaforo(datetime(2026, 7, 24, 16, 0), hoy=date(2026, 7, 24))[
                "fecha_limite"
            ]
            == "2026-07-29"
        )


# ------------------------------------------------------------------
# Parsers (unidad)
# ------------------------------------------------------------------


class TestParsers:
    def test_radicacion_mapea_columnas_reales(self):
        r = svc.parsear_excel_radicacion(_excel(RAD_HEADERS, [_rad_fila(ENV, F1, 250700)]))
        assert len(r["facturas"]) == 1
        f = r["facturas"][0]
        assert f["factura"] == F1 and f["envio"] == ENV and f["valor"] == 250700
        assert f["nit"] == "860002400"
        assert f["entidad"] == "AXA COLPATRIA"

    def test_dgreport_mapea(self):
        r = svc.parsear_excel_dgreport(_excel(DG_HEADERS, [["2026-07-10", F1, "CUFEX"]]))
        assert r["facturas"][0]["factura"] == F1
        assert r["facturas"][0]["correo_fe"] == "SI"

    def test_radicacion_sin_columnas_avisa(self):
        r = svc.parsear_excel_radicacion(_excel(["OTRA", "COSA"], [[1, 2]]))
        assert r["facturas"] == [] and r["advertencias"]

    def test_salta_encabezados_y_totales_repetidos(self):
        """El reporte real repite el encabezado y trae filas de TOTAL: se saltan.

        Cubre el atajo rápido del parseo (evitar normalizar 191.859 veces):
        los valores ASCII alfanuméricos se comparan directo y los que traen
        acentos o signos siguen pasando por la normalización completa.
        """
        filas = [
            _rad_fila(ENV, F1, 250700),
            _rad_fila(ENV, "FACTURA", 0),  # encabezado repetido
            _rad_fila(ENV, "Total", 999),  # fila de total
            _rad_fila(ENV, " total ", 999),  # con espacios
            _rad_fila(ENV, "FACTÚRA", 0),  # con acento → normalizada
            _rad_fila(ENV, F2, 98000),
        ]
        r = svc.parsear_excel_radicacion(_excel(RAD_HEADERS, filas))
        assert sorted(f["factura"] for f in r["facturas"]) == sorted([F1, F2])

    def test_anulado_en_cualquier_forma_se_descarta(self):
        """El estado llega escrito de varias formas; todas cuentan como anulada."""
        filas = [
            _rad_fila(ENV, F1, 1, estado="Anulado"),
            _rad_fila(ENV, F2, 2, estado="ANULADA"),
            _rad_fila(ENV, F3, 3, estado="anulado_entidad"),
            _rad_fila(ENV, "HUS0000999999", 4, estado="Radicado_Entidad"),
        ]
        r = svc.parsear_excel_radicacion(_excel(RAD_HEADERS, filas))
        assert [f["factura"] for f in r["facturas"]] == ["HUS0000999999"]
        assert "3 radicación(es) 'Anulado' descartada(s)" in r["advertencias"]

    def test_lee_archivos_sin_dimensiones_declaradas(self):
        """El reporte de DGH no declara <dimension>: debe leerse igual.

        Se omite el barrido previo de openpyxl (12 s en el archivo real), así
        que esta prueba fija que quitarlo no cambia lo que se lee.
        """
        import zipfile

        contenido = _excel(RAD_HEADERS, [_rad_fila(ENV, F1, 250700), _rad_fila(ENV, F2, 98000)])
        # el Excel de prueba sí declara dimensiones: se le quitan para simular
        # el reporte real de Dinámica Gerencial
        entrada = BytesIO()
        with zipfile.ZipFile(BytesIO(contenido)) as orig:
            with zipfile.ZipFile(entrada, "w", zipfile.ZIP_DEFLATED) as nuevo:
                for item in orig.infolist():
                    datos = orig.read(item.filename)
                    if item.filename.startswith("xl/worksheets/sheet"):
                        datos = re.sub(rb"<dimension[^>]*/>", b"", datos)
                    nuevo.writestr(item, datos)
        sin_dim = entrada.getvalue()
        assert b"<dimension" not in zipfile.ZipFile(BytesIO(sin_dim)).read(
            "xl/worksheets/sheet1.xml"
        )
        r = svc.parsear_excel_radicacion(sin_dim)
        assert sorted(f["factura"] for f in r["facturas"]) == sorted([F1, F2])
        assert r["leidas"] == 2


# ------------------------------------------------------------------
# Buscar en el consolidado (por factura, entidad o NIT)
# ------------------------------------------------------------------


class TestBuscarEnConsolidado:
    """La búsqueda cruza el consolidado con la fuente, que en producción tiene
    ~190.000 filas. La subconsulta se restringe primero a las facturas que ya
    están en el consolidado; estas pruebas fijan que el resultado sea el mismo
    que recorriendo la fuente entera."""

    def _preparar(self, client):
        _subir_radicacion(
            client,
            [
                _rad_fila(ENV, F1, 250700, nit="860002184", entidad="AXA COLPATRIA SEGUROS S.A."),
                _rad_fila(ENV, F2, 10615224, nit="860002400", entidad="LA PREVISORA S A"),
                # Este envío NO se escribe: la factura queda solo en la fuente.
                _rad_fila(
                    "999999", F3, 73500, nit="860002184", entidad="AXA COLPATRIA SEGUROS S.A."
                ),
            ],
        )
        _subir_dgreport(client, [F1])  # con correo F.E., para poder radicarla
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)

    def _buscar(self, client, q):
        d = client.get(f"/preauditoria/consolidado?q={q}").json()
        assert d["total"] == len(d["items"]), "el contador y las filas no coinciden"
        return sorted(i["factura"] for i in d["items"])

    def test_busca_por_entidad(self, client):
        self._preparar(client)
        assert self._buscar(client, "AXA") == [F1]

    def test_busca_por_nit(self, client):
        self._preparar(client)
        assert self._buscar(client, "860002400") == [F2]

    def test_busca_por_numero_de_factura(self, client):
        self._preparar(client)
        assert self._buscar(client, F2) == [F2]

    def test_no_trae_lo_que_solo_vive_en_la_fuente(self, client):
        """F3 tiene la MISMA entidad que F1, pero su envío nunca se escribió:
        no está en el consolidado y no debe aparecer. Es la razón por la que
        restringir la subconsulta no cambia el resultado."""
        self._preparar(client)
        assert F3 not in self._buscar(client, "AXA")

    def test_no_distingue_mayusculas(self, client):
        self._preparar(client)
        assert self._buscar(client, "axa") == self._buscar(client, "AXA") == [F1]

    def test_coincidencia_parcial(self, client):
        self._preparar(client)
        assert self._buscar(client, "PREVISORA") == [F2]

    def test_sin_resultados(self, client):
        self._preparar(client)
        assert self._buscar(client, "SANITAS") == []

    def test_se_combina_con_los_otros_filtros(self, client):
        """Buscar 'AXA' y filtrar por radicadas no debe traer la pendiente."""
        self._preparar(client)
        d = client.get("/preauditoria/consolidado?q=AXA&resultado=RADICAR").json()
        assert d["total"] == 0
        _radicar(client, _factura_id(client, F1))
        d = client.get("/preauditoria/consolidado?q=AXA&resultado=RADICAR").json()
        assert [i["factura"] for i in d["items"]] == [F1]


# ------------------------------------------------------------------
# Una factura devuelta no se repite en dos oficios de devolución
# ------------------------------------------------------------------


class TestDevueltasYaSalidas:
    """Caso real del 29-07: el oficio decía "Dev 3" y el PDF salió con 1
    factura, porque las otras dos ya habían salido en un oficio anterior. El
    sistema hacía lo correcto; lo que faltaba era decirlo en pantalla."""

    def _tres_devueltas(self, client):
        _subir_radicacion(
            client,
            [
                _rad_fila(ENV, F1, 175500),
                _rad_fila(ENV, F2, 13781171),
                _rad_fila(ENV, F3, 13229822),
            ],
        )
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)
        for f in (F1, F2, F3):
            _devolver(client, _factura_id(client, f))
        return o

    def test_cuenta_las_que_faltan_por_salir(self, client):
        o = self._tres_devueltas(client)
        d = client.get(f"/preauditoria/oficios/{o['id']}").json()
        assert d["devueltas"] == 3
        assert d["devueltas_sin_oficio"] == 3

    def test_tras_generar_no_quedan_por_salir(self, client):
        o = self._tres_devueltas(client)
        dev = client.post(f"/preauditoria/oficios/{o['id']}/oficio-devolucion").json()
        assert dev["total_facturas"] == 3
        d = client.get(f"/preauditoria/oficios/{o['id']}").json()
        # siguen devueltas, pero ya salieron: el botón debe quedar deshabilitado
        assert d["devueltas"] == 3
        assert d["devueltas_sin_oficio"] == 0

    def test_cada_factura_dice_en_que_oficio_salio(self, client):
        o = self._tres_devueltas(client)
        dev = client.post(f"/preauditoria/oficios/{o['id']}/oficio-devolucion").json()
        d = client.get(f"/preauditoria/oficios/{o['id']}").json()
        for f in d["facturas"]:
            assert f["consecutivo_devolucion"] == dev["consecutivo"]

    def test_el_segundo_oficio_solo_lleva_la_nueva(self, client):
        """La secuencia exacta del auditor: genera un oficio con 2, después
        devuelve una tercera y genera otro. El segundo lleva SOLO esa."""
        _subir_radicacion(
            client,
            [
                _rad_fila(ENV, F1, 175500),
                _rad_fila(ENV, F2, 13781171),
                _rad_fila(ENV, F3, 13229822),
            ],
        )
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)
        _devolver(client, _factura_id(client, F1))
        _devolver(client, _factura_id(client, F2))
        primero = client.post(f"/preauditoria/oficios/{o['id']}/oficio-devolucion").json()
        assert primero["total_facturas"] == 2

        _devolver(client, _factura_id(client, F3))
        d = client.get(f"/preauditoria/oficios/{o['id']}").json()
        assert d["devueltas"] == 3 and d["devueltas_sin_oficio"] == 1

        segundo = client.post(f"/preauditoria/oficios/{o['id']}/oficio-devolucion").json()
        assert segundo["total_facturas"] == 1
        assert segundo["consecutivo"] != primero["consecutivo"]
        # y cada una quedó sellada con SU oficio
        porf = {
            f["factura"]: f["consecutivo_devolucion"]
            for f in client.get(f"/preauditoria/oficios/{o['id']}").json()["facturas"]
        }
        assert porf[F1] == porf[F2] == primero["consecutivo"]
        assert porf[F3] == segundo["consecutivo"]

    def test_mensaje_claro_si_ya_salieron_todas(self, client):
        o = self._tres_devueltas(client)
        client.post(f"/preauditoria/oficios/{o['id']}/oficio-devolucion")
        r = client.post(f"/preauditoria/oficios/{o['id']}/oficio-devolucion")
        assert r.status_code == 400
        detalle = r.json()["detail"]
        assert "ya salieron" in detalle
        assert "no se repite" in detalle
        # NO debe decir que falte marcar devoluciones: eso confundía
        assert "primero marque" not in detalle

    def test_mensaje_distinto_si_no_hay_devoluciones(self, client):
        _subir_radicacion(client, [_rad_fila(ENV, F1, 175500)])
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)
        r = client.post(f"/preauditoria/oficios/{o['id']}/oficio-devolucion")
        assert r.status_code == 400
        assert "primero marque" in r.json()["detail"]


# ------------------------------------------------------------------
# Solo administración modifica auditorías ya decididas + informe de gestión
# ------------------------------------------------------------------


class TestSoloAdminModificaDecididas:
    def _setup(self, client):
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        _subir_dgreport(client, [F1])
        o = _crear_oficio(client, "FHUS-ROL-1", "2026-07-20T08:00")
        _escribir(client, o["id"], ENV)
        return o, _factura_id(client, F1)

    def test_auditor_no_corrige_observacion_de_radicada(self, client):
        _o, fid = self._setup(client)
        assert _radicar(client, fid).status_code == 200
        r = client.patch(
            f"/preauditoria/facturas/{fid}/observacion", json={"observaciones": "CAMBIO"}
        )
        assert r.status_code == 403, r.text
        assert "coordinación o administración" in r.json()["detail"]

    def test_auditor_no_corrige_observacion_de_devuelta(self, client):
        _o, fid = self._setup(client)
        assert _devolver(client, fid).status_code == 200
        r = client.patch(
            f"/preauditoria/facturas/{fid}/observacion", json={"observaciones": "CAMBIO"}
        )
        assert r.status_code == 403, r.text

    def test_auditor_no_revierte_la_decision_de_otra_persona(self, client, db_session):
        """Desde el 20-08 cada auditor deshace LO SUYO; lo de otro, no."""
        _o, fid = self._setup(client)
        assert _radicar(client, fid).status_code == 200
        f = db_session.get(FacturaPreauditoriaRecord, fid)
        f.auditor = "OTRA PERSONA"
        db_session.commit()
        r = client.patch(f"/preauditoria/facturas/{fid}/auditar", json={"resultado": "PENDIENTE"})
        assert r.status_code == 403, r.text
        assert "OTRA PERSONA" in r.json()["detail"]
        # la decisión no cambió
        assert client.get(f"/preauditoria/facturas/{fid}").json()["resultado"] == "RADICAR"

    def test_auditor_si_anota_mientras_este_pendiente(self, client):
        """La regla es sobre auditorías DECIDIDAS: mientras la factura siga
        pendiente, el auditor anota su observación con normalidad."""
        _o, fid = self._setup(client)
        r = client.patch(
            f"/preauditoria/facturas/{fid}/observacion", json={"observaciones": "NOTA PREVIA"}
        )
        assert r.status_code == 200, r.text
        assert r.json()["observaciones"] == "NOTA PREVIA"

    def test_admin_corrige_y_revierte(self, client, db_session):
        _o, fid = self._setup(client)
        assert _devolver(client, fid, motivo="Falta FURIPS").status_code == 200
        _actuar_como_admin(db_session)
        r = client.patch(
            f"/preauditoria/facturas/{fid}/observacion", json={"observaciones": "CORREGIDO"}
        )
        assert r.status_code == 200, r.text
        assert r.json()["motivo_devolucion"] == "CORREGIDO"
        r2 = client.patch(f"/preauditoria/facturas/{fid}/auditar", json={"resultado": "PENDIENTE"})
        assert r2.status_code == 200, r2.text


class TestInformeGestion:
    def test_informe_descargable_con_sus_hojas(self, client):
        from openpyxl import load_workbook

        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700), _rad_fila(ENV, F2, 100)])
        _subir_dgreport(client, [F1, F2])
        o = _crear_oficio(client, "FHUS-INF-1", "2026-07-20T08:00")
        _escribir(client, o["id"], ENV)
        assert _radicar(client, _factura_id(client, F1)).status_code == 200
        assert (
            _devolver(client, _factura_id(client, F2), motivo="Falta epicrisis").status_code == 200
        )
        client.post(f"/preauditoria/oficios/{o['id']}/oficio-devolucion")

        r = client.get("/preauditoria/estadisticas/informe.xlsx")
        assert r.status_code == 200, r.text
        assert "spreadsheetml" in r.headers["content-type"]
        wb = load_workbook(BytesIO(r.content))
        assert wb.sheetnames == [
            "RESUMEN",
            "POR AUDITOR",
            "POR OFICIO",
            "DEVOLUCIONES",
            "HISTORIAL",
        ]

        resumen = {f[0]: f[1] for f in wb["RESUMEN"].iter_rows(min_row=2, values_only=True)}
        assert resumen["Facturas en el consolidado"] == 2
        assert resumen["Radicadas / subsanadas"] == 1
        assert resumen["Devueltas (estado actual)"] == 1
        assert resumen["Oficios de devolución emitidos"] == 1
        assert resumen["Valor radicado"] == 250700

        auditores = list(wb["POR AUDITOR"].iter_rows(min_row=2, values_only=True))
        assert auditores[0][0] == "CLAUDIA" and auditores[0][1] == 2

        oficios = list(wb["POR OFICIO"].iter_rows(min_row=2, values_only=True))
        assert oficios[0][0] == "FHUS-INF-1" and oficios[0][3] == 2

        eventos = list(wb["HISTORIAL"].iter_rows(min_row=2, values_only=True))
        tipos = [e[4] for e in eventos]
        assert "ESCRITA" in tipos and "RADICADA" in tipos and "DEVUELTA" in tipos


# ------------------------------------------------------------------
# Quitar UNA factura del oficio (entró por error) — solo admin/coordinador
#
# Caso real 19-08-2026, envío 232047: la fuente de Radicación arrastró la
# factura HUS0000538528 de un cargue anterior; el DGH mostraba 12 facturas y
# el sistema traía 13. Antes tocaba quitar el envío entero y reescribirlo.
# ------------------------------------------------------------------


class TestQuitarFactura:
    def _admin(self, db_session):
        u = UsuarioRecord(
            id=7,
            nombre="ADMIN QF",
            email="admin7@hus.gov.co",
            rol="SUPER_ADMIN",
            activo=1,
            password_hash=get_password_hash("xxxx"),
        )
        db_session.add(u)
        db_session.commit()
        return u

    def _como_admin(self, db_session):
        from app.api.deps import get_coordinador_o_admin
        from app.main import app

        app.dependency_overrides[get_coordinador_o_admin] = lambda: self._admin(db_session)

    def test_auditor_no_puede(self, client):
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)
        fid = _factura_id(client, F1)
        assert client.delete(f"/preauditoria/oficios/{o['id']}/facturas/{fid}").status_code == 403

    def test_quita_solo_esa_factura_y_deja_el_resto(self, client, db_session):
        self._como_admin(db_session)
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700), _rad_fila(ENV, F2, 98000)])
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)
        fid = _factura_id(client, F2)

        r = client.delete(f"/preauditoria/oficios/{o['id']}/facturas/{fid}")
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["factura"] == F2 and d["borrada"] is True

        # La otra factura del mismo envío sigue intacta…
        cons = client.get("/preauditoria/consolidado").json()
        assert cons["total"] == 1 and cons["items"][0]["factura"] == F1
        # …el envío sigue cargado (con una factura menos en el conteo)…
        envios = {e["envio"]: e["total_facturas"] for e in d["oficio"]["envios_escritos"]}
        assert envios == {ENV: 1}
        # …y el oficio ya no la cuenta.
        assert d["oficio"]["total_facturas"] == 1

    def test_no_quita_una_factura_de_otro_oficio(self, client, db_session):
        self._como_admin(db_session)
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700), _rad_fila("229999", F3, 55000)])
        o1 = _crear_oficio(client)
        o2 = _crear_oficio(client, radicado="FHUS-AS-I00999-26")
        _escribir(client, o1["id"], ENV)
        _escribir(client, o2["id"], "229999")

        r = client.delete(f"/preauditoria/oficios/{o1['id']}/facturas/{_factura_id(client, F3)}")
        assert r.status_code == 409
        assert "ya no pertenece" in r.json()["detail"]
        assert client.get("/preauditoria/consolidado").json()["total"] == 2

    def test_no_quita_si_ya_salio_en_un_oficio_de_devolucion(self, client, db_session):
        self._como_admin(db_session)
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)
        fid = _factura_id(client, F1)
        _devolver(client, fid)
        client.post(f"/preauditoria/oficios/{o['id']}/oficio-devolucion")

        r = client.delete(f"/preauditoria/oficios/{o['id']}/facturas/{fid}")
        assert r.status_code == 409
        assert "oficio de devolución" in r.json()["detail"]
        assert client.get("/preauditoria/consolidado").json()["total"] == 1

    def test_reingresada_vuelve_a_su_devolucion_anterior(self, client, db_session):
        """Si la factura venía reingresada para subsanar, no se borra su
        historial: se revierte el reingreso y vuelve a quedar devuelta."""
        self._como_admin(db_session)
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        o1 = _crear_oficio(client)
        _escribir(client, o1["id"], ENV)
        fid = _factura_id(client, F1)
        _devolver(client, fid, motivo="Falta epicrisis")
        o2 = _crear_oficio(client, radicado="FHUS-AS-I00888-26")
        _escribir(client, o2["id"], ENV)  # reingreso para subsanar

        r = client.delete(f"/preauditoria/oficios/{o2['id']}/facturas/{fid}")
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["borrada"] is False and d["reingreso_revertido"] is True

        hist = client.get(f"/preauditoria/facturas/{F1}/historial").json()
        assert hist["actual"]["resultado"] == "DEVUELTA"
        assert hist["actual"]["oficio_fhus"] == o1["numero_radicado"]
        assert hist["actual"]["motivo_devolucion"] == "Falta epicrisis"
        # El historial de la devolución NO se perdió.
        assert "DEVUELTA" in [e["tipo"] for e in hist["eventos"]]


# ------------------------------------------------------------------
# Facturas que la fuente arrastra de un cargue anterior (la "fantasma")
# ------------------------------------------------------------------


class TestFuenteRezagada:
    def _envejecer(self, db_session, factura, dias=30):
        from datetime import timedelta as _td

        from app.core.tz import ahora_utc

        fila = (
            db_session.query(RadicacionCuentaRecord)
            .filter(RadicacionCuentaRecord.factura == factura)
            .one()
        )
        fila.actualizado_en = ahora_utc() - _td(days=dias)
        fila.importado_en = fila.actualizado_en
        db_session.commit()

    def test_preview_avisa_de_la_factura_vieja(self, client, db_session):
        _subir_radicacion(
            client,
            [_rad_fila(ENV, F1, 250700), _rad_fila(ENV, F2, 98000), _rad_fila(ENV, F3, 55000)],
        )
        self._envejecer(db_session, F3)
        o = _crear_oficio(client)

        p = client.get(f"/preauditoria/oficios/{o['id']}/envios/{ENV}/preview").json()
        assert p["facturas_rezagadas"] == [F3]
        assert len(p["avisos_fuente"]) == 1
        assert F3 in p["avisos_fuente"][0] and "DGH" in p["avisos_fuente"][0]

    def test_al_cargar_el_envio_queda_la_advertencia(self, client, db_session):
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700), _rad_fila(ENV, F3, 55000)])
        self._envejecer(db_session, F3)
        o = _crear_oficio(client)

        d = _escribir(client, o["id"], ENV).json()
        assert d["nuevas"] == 2  # se carga igual: la fuente manda
        assert any(F3 in a for a in d["advertencias"])

    def test_sin_diferencia_de_fechas_no_avisa_nada(self, client):
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700), _rad_fila(ENV, F2, 98000)])
        o = _crear_oficio(client)
        p = client.get(f"/preauditoria/oficios/{o['id']}/envios/{ENV}/preview").json()
        assert p["facturas_rezagadas"] == [] and p["avisos_fuente"] == []


class TestFuenteRezagadaEscenariosReales:
    """El aviso tiene que acertar en el flujo real de la Radicación.

    Lo probó el caso del 19-08-2026: la fuente llevaba 15 días sin
    re-subirse, así que las 13 facturas del envío 232047 (incluida la que
    facturación ya había sacado) tenían la MISMA fecha y el aviso no podía
    distinguirlas. Y al re-subir el Excel, si solo se sellaran las filas que
    cambian, las que siguen igual quedarían señaladas por error.
    """

    def _envejecer_todo(self, db_session, dias=30):
        from datetime import timedelta as _td

        from app.core.tz import ahora_utc

        for fila in db_session.query(RadicacionCuentaRecord).all():
            fila.importado_en = ahora_utc() - _td(days=dias)
            fila.actualizado_en = fila.importado_en
        db_session.commit()

    def _rezagadas(self, client, oficio_id, envio=ENV):
        return client.get(f"/preauditoria/oficios/{oficio_id}/envios/{envio}/preview").json()[
            "facturas_rezagadas"
        ]

    def test_resubir_el_excel_senala_solo_la_que_ya_no_viene(self, client, db_session):
        """El caso real: se re-sube el Excel y una factura ya no aparece."""
        _subir_radicacion(
            client,
            [_rad_fila(ENV, F1, 250700), _rad_fila(ENV, F2, 98000), _rad_fila(ENV, F3, 55000)],
        )
        self._envejecer_todo(db_session)
        # Excel de hoy: facturación ya sacó F3 del envío, así que no viene.
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700), _rad_fila(ENV, F2, 98000)])
        o = _crear_oficio(client)
        assert self._rezagadas(client, o["id"]) == [F3]

    def test_resubir_el_mismo_excel_con_una_correccion_no_acusa_a_las_demas(
        self, client, db_session
    ):
        """Sellar solo lo que cambia dejaría 2 falsos positivos por cada
        corrección de facturación."""
        _subir_radicacion(
            client,
            [_rad_fila(ENV, F1, 250700), _rad_fila(ENV, F2, 98000), _rad_fila(ENV, F3, 55000)],
        )
        self._envejecer_todo(db_session)
        # mismo archivo completo, con el valor de F1 corregido
        _subir_radicacion(
            client,
            [_rad_fila(ENV, F1, 250800), _rad_fila(ENV, F2, 98000), _rad_fila(ENV, F3, 55000)],
        )
        o = _crear_oficio(client)
        assert self._rezagadas(client, o["id"]) == []

    def test_agregar_una_factura_al_envio_no_acusa_a_las_viejas(self, client, db_session):
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700), _rad_fila(ENV, F2, 98000)])
        self._envejecer_todo(db_session)
        _subir_radicacion(
            client,
            [_rad_fila(ENV, F1, 250700), _rad_fila(ENV, F2, 98000), _rad_fila(ENV, F3, 55000)],
        )
        o = _crear_oficio(client)
        assert self._rezagadas(client, o["id"]) == []

    def test_fuente_vieja_sin_resubir_no_acusa_a_nadie(self, client, db_session):
        """Como el 19-08: la fuente lleva días sin re-subirse. Todas las filas
        son igual de viejas y el sistema NO puede señalar a ninguna."""
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700), _rad_fila(ENV, F2, 98000)])
        self._envejecer_todo(db_session)
        o = _crear_oficio(client)
        assert self._rezagadas(client, o["id"]) == []


class TestRevertirReingresoConservaSuOficioEmitido:
    """Al deshacer un reingreso, la factura vuelve amarrada al oficio de
    devolución que ya se le emitió — si no, el sistema dejaba emitir un
    segundo oficio con la tabla del PDF vacía."""

    def _como_admin(self, db_session):
        from app.api.deps import get_coordinador_o_admin
        from app.main import app

        u = UsuarioRecord(
            id=11,
            nombre="ADMIN REV",
            email="admin11@hus.gov.co",
            rol="SUPER_ADMIN",
            activo=1,
            password_hash=get_password_hash("xxxx"),
        )
        db_session.add(u)
        db_session.commit()
        app.dependency_overrides[get_coordinador_o_admin] = lambda: u

    def _escenario(self, client, db_session):
        """Factura devuelta con oficio emitido y reenviada a un oficio B."""
        self._como_admin(db_session)
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        _subir_dgreport(client, [F1])
        oA = _crear_oficio(client, radicado="FHUS-AS-I00001-26")
        _escribir(client, oA["id"], ENV)
        fid = _factura_id(client, F1)
        _devolver(client, fid)
        dev = client.post(f"/preauditoria/oficios/{oA['id']}/oficio-devolucion").json()
        oB = _crear_oficio(client, radicado="FHUS-AS-I00002-26")
        _escribir(client, oB["id"], ENV)  # reingreso para subsanar
        return oA, oB, fid, dev

    def test_al_quitar_la_factura_vuelve_a_su_oficio_de_devolucion(self, client, db_session):
        oA, oB, fid, dev = self._escenario(client, db_session)
        r = client.delete(f"/preauditoria/oficios/{oB['id']}/facturas/{fid}")
        assert r.status_code == 200, r.text

        fichaA = client.get(f"/preauditoria/oficios/{oA['id']}").json()
        suya = [f for f in fichaA["facturas"] if f["factura"] == F1][0]
        assert suya["estado"] in ("DEVUELTA_PEND_SUBSANACION", "NUEVAMENTE_DEVUELTA")
        assert suya["consecutivo_devolucion"] == dev["consecutivo"]
        # …y el oficio A no vuelve a ofrecerla para un segundo oficio.
        assert fichaA["devueltas_sin_oficio"] == 0

    def test_al_quitar_el_envio_pasa_lo_mismo(self, client, db_session):
        oA, oB, fid, dev = self._escenario(client, db_session)
        r = client.delete(f"/preauditoria/oficios/{oB['id']}/envios/{ENV}")
        assert r.status_code == 200, r.text
        fichaA = client.get(f"/preauditoria/oficios/{oA['id']}").json()
        suya = [f for f in fichaA["facturas"] if f["factura"] == F1][0]
        assert suya["consecutivo_devolucion"] == dev["consecutivo"]

    def test_no_deja_emitir_un_segundo_oficio_por_la_misma_factura(self, client, db_session):
        oA, oB, fid, dev = self._escenario(client, db_session)
        client.delete(f"/preauditoria/oficios/{oB['id']}/facturas/{fid}")
        r2 = client.post(f"/preauditoria/oficios/{oA['id']}/oficio-devolucion")
        assert r2.status_code != 200 or r2.json().get("total_facturas", 0) == 0


# ------------------------------------------------------------------
# Envíos que no dejaron todas sus facturas en el oficio
# (caso real 20-08-2026: la ventana mostraba 5 envíos y 4 facturas)
# ------------------------------------------------------------------


class TestEnvioSinTodasSusFacturas:
    """Cuando una factura del envío ya venía abierta (o radicada) en otro
    oficio, el sistema la omite: el envío queda cargado pero sin ella. La
    pantalla debe decir cuántas quedaron de verdad y dónde está la que falta."""

    ENV_B = "229999"

    def _dos_oficios(self, client):
        """F1 se queda abierta en el oficio A; el envío se vuelve a escribir en B."""
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700), _rad_fila(self.ENV_B, F2, 120000)])
        _subir_dgreport(client, [F1, F2])
        oA = _crear_oficio(client, radicado="FHUS-AS-I01188-26")
        _escribir(client, oA["id"], ENV)
        oB = _crear_oficio(client, radicado="FHUS-AS-I01190-26", fecha="2026-07-22T08:30")
        _escribir(client, oB["id"], ENV)  # F1 sigue pendiente en A: se omite
        _escribir(client, oB["id"], self.ENV_B)  # esta sí entra
        return oA, oB

    def _envio(self, ficha, envio):
        return [e for e in ficha["envios_escritos"] if e["envio"] == envio][0]

    def test_el_chip_dice_cuantas_quedaron_de_verdad(self, client):
        _, oB = self._dos_oficios(client)
        ficha = client.get(f"/preauditoria/oficios/{oB['id']}").json()
        assert ficha["total_facturas"] == 1  # solo F2
        vacio = self._envio(ficha, ENV)
        assert vacio["total_facturas"] == 1  # lo que traía la fuente
        assert vacio["en_este_oficio"] == 0  # lo que quedó aquí
        completo = self._envio(ficha, self.ENV_B)
        assert completo["en_este_oficio"] == completo["total_facturas"] == 1
        assert completo["faltantes"] == []

    def test_dice_cual_factura_falta_y_donde_quedo(self, client):
        _, oB = self._dos_oficios(client)
        ficha = client.get(f"/preauditoria/oficios/{oB['id']}").json()
        falta = self._envio(ficha, ENV)["faltantes"]
        assert len(falta) == 1
        assert falta[0]["factura"] == F1
        assert falta[0]["oficio"] == "FHUS-AS-I01188-26"
        assert "pendiente" in falta[0]["motivo"]
        assert "FHUS-AS-I01188-26" in falta[0]["motivo"]

    def test_la_factura_ya_radicada_se_explica_distinto(self, client):
        oA, _ = self._dos_oficios(client)
        _radicar(client, _factura_id(client, F1))  # queda cerrada en A
        oC = _crear_oficio(client, radicado="FHUS-AS-I01191-26", fecha="2026-07-23T08:30")
        _escribir(client, oC["id"], ENV)
        ficha = client.get(f"/preauditoria/oficios/{oC['id']}").json()
        falta = self._envio(ficha, ENV)["faltantes"]
        assert falta[0]["factura"] == F1
        assert "radicada" in falta[0]["motivo"]
        assert oA["numero_radicado"] in falta[0]["motivo"]

    def test_si_se_va_a_otro_oficio_el_de_origen_lo_dice(self, client):
        """Devuelta en A y reingresada en C: el chip de A deja de mentir."""
        oA, _ = self._dos_oficios(client)
        _devolver(client, _factura_id(client, F1))
        oC = _crear_oficio(client, radicado="FHUS-AS-I01192-26", fecha="2026-07-24T08:30")
        _escribir(client, oC["id"], ENV)  # reingreso: F1 se muda al oficio C
        fichaA = client.get(f"/preauditoria/oficios/{oA['id']}").json()
        assert fichaA["total_facturas"] == 0
        falta = self._envio(fichaA, ENV)["faltantes"]
        assert falta[0]["factura"] == F1
        assert "FHUS-AS-I01192-26" in falta[0]["motivo"]

    def test_la_lista_de_oficios_tambien_cuenta_lo_que_quedo(self, client):
        _, oB = self._dos_oficios(client)
        items = client.get("/preauditoria/oficios").json()["items"]
        fila = [o for o in items if o["id"] == oB["id"]][0]
        assert self._envio(fila, ENV)["en_este_oficio"] == 0
        # La lista no hace la consulta cara del detalle (no debe volverse lenta).
        assert "faltantes" not in self._envio(fila, ENV)

    def test_el_envio_normal_no_genera_ningun_aviso(self, client):
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700), _rad_fila(ENV, F2, 120000)])
        _subir_dgreport(client, [F1, F2])
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)
        ficha = client.get(f"/preauditoria/oficios/{o['id']}").json()
        e = self._envio(ficha, ENV)
        assert e["en_este_oficio"] == e["total_facturas"] == 2
        assert e["faltantes"] == []

    def test_quitar_una_factura_a_proposito_no_deja_avisos(self, client, db_session):
        """Si el auditor la quitó con la 🗑, el sistema no le insiste."""
        _actuar_como_admin(db_session)
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700), _rad_fila(ENV, F2, 120000)])
        _subir_dgreport(client, [F1, F2])
        o = _crear_oficio(client)
        _escribir(client, o["id"], ENV)
        fid = _factura_id(client, F1)
        assert client.delete(f"/preauditoria/oficios/{o['id']}/facturas/{fid}").status_code == 200
        ficha = client.get(f"/preauditoria/oficios/{o['id']}").json()
        e = self._envio(ficha, ENV)
        assert e["en_este_oficio"] == e["total_facturas"] == 1
        assert e["faltantes"] == []


# ------------------------------------------------------------------
# Corregir el número de un oficio mal digitado
# (caso real 20-08-2026: quedó FHUS-AS-101190-26, con un uno por la I)
# ------------------------------------------------------------------


class TestRenombrarOficio:
    def _con_factura(self, client, radicado="FHUS-AS-101190-26"):
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        _subir_dgreport(client, [F1])
        o = _crear_oficio(client, radicado=radicado)
        _escribir(client, o["id"], ENV)
        return o

    def test_el_auditor_no_puede_corregirlo(self, client):
        o = self._con_factura(client)
        r = client.patch(
            f"/preauditoria/oficios/{o['id']}", json={"numero_radicado": "FHUS-AS-I01190-26"}
        )
        assert r.status_code == 403

    def test_el_numero_queda_corregido_en_todas_partes(self, client, db_session):
        _actuar_como_admin(db_session)
        o = self._con_factura(client)
        r = client.patch(
            f"/preauditoria/oficios/{o['id']}", json={"numero_radicado": " fhus-as-i01190-26 "}
        )
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["antes"] == "FHUS-AS-101190-26"
        assert d["ahora"] == "FHUS-AS-I01190-26"
        assert d["facturas"] == 1
        assert d["eventos"] >= 1
        # …en el oficio, en la factura del consolidado y en el historial.
        assert d["oficio"]["numero_radicado"] == "FHUS-AS-I01190-26"
        hist = client.get(f"/preauditoria/facturas/{F1}/historial").json()
        assert hist["actual"]["oficio_fhus"] == "FHUS-AS-I01190-26"
        assert all(
            e["oficio_fhus"] == "FHUS-AS-I01190-26" for e in hist["eventos"] if e.get("oficio_fhus")
        )

    def test_no_deja_pisar_el_numero_de_otro_oficio(self, client, db_session):
        _actuar_como_admin(db_session)
        o = self._con_factura(client)
        otro = _crear_oficio(client, radicado="FHUS-AS-I01196-26", fecha="2026-07-25T08:30")
        r = client.patch(
            f"/preauditoria/oficios/{o['id']}", json={"numero_radicado": otro["numero_radicado"]}
        )
        assert r.status_code == 409
        assert "otro oficio" in r.json()["detail"]

    def test_el_numero_vacio_se_rechaza_con_mensaje(self, client, db_session):
        _actuar_como_admin(db_session)
        o = self._con_factura(client)
        r = client.patch(f"/preauditoria/oficios/{o['id']}", json={"numero_radicado": "   "})
        assert r.status_code == 400
        assert "Escriba el número" in r.json()["detail"]

    def test_dejar_el_mismo_numero_no_rompe_nada(self, client, db_session):
        _actuar_como_admin(db_session)
        o = self._con_factura(client)
        r = client.patch(
            f"/preauditoria/oficios/{o['id']}", json={"numero_radicado": o["numero_radicado"]}
        )
        assert r.status_code == 200
        assert r.json()["sin_cambios"] is True

    def test_el_oficio_que_no_existe_avisa(self, client, db_session):
        _actuar_como_admin(db_session)
        r = client.patch(
            "/preauditoria/oficios/9999", json={"numero_radicado": "FHUS-AS-I09999-26"}
        )
        assert r.status_code == 404


# ------------------------------------------------------------------
# Cada auditor deshace LO SUYO (regla acordada el 20-08-2026)
# ------------------------------------------------------------------


class TestDeshacerLoSuyo:
    """Antes «Dejar pendiente» era solo de coordinación y el auditor quedaba
    trancado con su propio error de dedo. Ahora deshace lo suyo, mientras esa
    auditoría no haya salido del hospital en un oficio de devolución."""

    def _factura(self, client, factura=F1, envio=ENV, radicado="FHUS-ROL-9"):
        _subir_radicacion(client, [_rad_fila(envio, factura, 250700)])
        _subir_dgreport(client, [factura])
        o = _crear_oficio(client, radicado, "2026-07-20T08:00")
        _escribir(client, o["id"], envio)
        return o, _factura_id(client, factura)

    def test_el_auditor_deshace_su_propia_radicacion(self, client):
        _o, fid = self._factura(client)
        assert _radicar(client, fid).status_code == 200
        r = client.patch(f"/preauditoria/facturas/{fid}/auditar", json={"resultado": "PENDIENTE"})
        assert r.status_code == 200, r.text
        assert client.get(f"/preauditoria/facturas/{fid}").json()["resultado"] == "PENDIENTE"

    def test_el_auditor_deshace_su_propia_devolucion(self, client):
        _o, fid = self._factura(client)
        assert _devolver(client, fid).status_code == 200
        r = client.patch(f"/preauditoria/facturas/{fid}/auditar", json={"resultado": "PENDIENTE"})
        assert r.status_code == 200, r.text
        d = client.get(f"/preauditoria/facturas/{fid}").json()
        assert d["resultado"] == "PENDIENTE"
        assert d["num_devoluciones"] == 0

    def test_no_deshace_lo_que_ya_salio_en_un_oficio_de_devolucion(self, client, db_session):
        """Ese PDF ya está en manos de la entidad: eso sí es de coordinación."""
        o, fid = self._factura(client)
        assert _devolver(client, fid).status_code == 200
        dev = client.post(f"/preauditoria/oficios/{o['id']}/oficio-devolucion")
        assert dev.status_code == 200, dev.text
        r = client.patch(f"/preauditoria/facturas/{fid}/auditar", json={"resultado": "PENDIENTE"})
        assert r.status_code == 403, r.text
        assert "oficio de devolución" in r.json()["detail"]
        # Tampoco coordinación: ese PDF ya está en manos de la entidad (regla
        # que ya existía). Para deshacerlo hay que eliminar antes el oficio de
        # devolución.
        _actuar_como_admin(db_session)
        r2 = client.patch(f"/preauditoria/facturas/{fid}/auditar", json={"resultado": "PENDIENTE"})
        assert r2.status_code == 409, r2.text
        assert "no se puede revertir" in r2.json()["detail"]

    def test_la_regla_sola_sin_pasar_por_la_pagina(self):
        """La decide el servicio, para que el router solo la aplique."""
        from types import SimpleNamespace

        from app.services import preauditoria_service as s

        propia = SimpleNamespace(factura=F1, auditor="CLAUDIA", oficio_devolucion_id=None)
        ajena = SimpleNamespace(factura=F1, auditor="EDGAR", oficio_devolucion_id=None)
        emitida = SimpleNamespace(factura=F1, auditor="CLAUDIA", oficio_devolucion_id=7)
        assert s.puede_revertir(propia, "claudia", False)[0] is True
        assert s.puede_revertir(propia, " CLAUDIA ", False)[0] is True
        assert s.puede_revertir(ajena, "CLAUDIA", False)[0] is False
        assert s.puede_revertir(emitida, "CLAUDIA", False)[0] is False
        # Coordinación pasa por encima de las dos restricciones.
        assert s.puede_revertir(ajena, "CLAUDIA", True)[0] is True
        assert s.puede_revertir(emitida, "CLAUDIA", True)[0] is True

    def test_la_factura_sin_auditor_no_la_deshace_cualquiera(self):
        from types import SimpleNamespace

        from app.services import preauditoria_service as s

        huerfana = SimpleNamespace(factura=F1, auditor=None, oficio_devolucion_id=None)
        puede, motivo = s.puede_revertir(huerfana, "CLAUDIA", False)
        assert puede is False
        assert "otra persona" in motivo


# ------------------------------------------------------------------
# El oficio del que la factura YA SALIÓ (caso real 21-08-2026)
# ------------------------------------------------------------------


class TestOficioQueYaCumplio:
    """La factura entró aquí, se devolvió, salió en el oficio de devolución y
    reingresó en el oficio siguiente. El oficio viejo pedía «resuélvala allá y
    vuelva a escribir el envío aquí» —cuando ya no hay nada que hacer— y encima
    se quedaba en rojo para siempre."""

    def _escenario(self, client):
        _subir_radicacion(client, [_rad_fila(ENV, F1, 3285631)])
        _subir_dgreport(client, [F1])
        viejo = _crear_oficio(client, radicado="FHUS-AS-I01196-26", fecha="2026-07-15T08:30")
        _escribir(client, viejo["id"], ENV)
        _devolver(client, _factura_id(client, F1), motivo="Adjuntar historia clínica")
        dev = client.post(f"/preauditoria/oficios/{viejo['id']}/oficio-devolucion").json()
        nuevo = _crear_oficio(client, radicado="FHUS-AS-I01212-26", fecha="2026-07-22T08:30")
        _escribir(client, nuevo["id"], ENV)  # reingresa: la factura se muda
        return viejo, nuevo, dev

    def test_el_aviso_dice_que_ya_siguio_su_camino(self, client):
        viejo, nuevo, dev = self._escenario(client)
        ficha = client.get(f"/preauditoria/oficios/{viejo['id']}").json()
        falta = ficha["envios_escritos"][0]["faltantes"][0]
        assert falta["factura"] == F1
        assert falta["cerrado"] is True
        assert "ya pasó por este oficio" in falta["motivo"]
        assert dev["consecutivo"] in falta["motivo"]
        assert "FHUS-AS-I01212-26" in falta["motivo"]
        # …y NO le pide al auditor que la vuelva a escribir aquí.
        assert "vuelva a escribir el envío aquí" not in falta["motivo"]

    def test_el_oficio_vacio_queda_completado_y_no_en_rojo(self, client):
        viejo, _nuevo, _dev = self._escenario(client)
        ficha = client.get(f"/preauditoria/oficios/{viejo['id']}").json()
        assert ficha["total_facturas"] == 0
        assert ficha["semaforo"]["estado"] == "COMPLETADO"

    def test_un_oficio_recien_registrado_sigue_corriendo_su_plazo(self, client):
        """El que nunca ha tenido facturas no está cumplido: le falta trabajar."""
        o = _crear_oficio(client, radicado="FHUS-AS-I01300-26", fecha="2026-07-15T08:30")
        ficha = client.get(f"/preauditoria/oficios/{o['id']}").json()
        assert ficha["total_facturas"] == 0
        assert ficha["semaforo"]["estado"] != "COMPLETADO"

    def test_la_que_nunca_entro_sigue_pidiendo_accion(self, client):
        """Contraste: si la factura jamás pasó por el oficio, el aviso sí pide
        resolverla en el otro oficio."""
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        _subir_dgreport(client, [F1])
        a = _crear_oficio(client, radicado="FHUS-AS-I01188-26")
        _escribir(client, a["id"], ENV)  # la factura queda abierta aquí
        b = _crear_oficio(client, radicado="FHUS-AS-I01190-26", fecha="2026-07-22T08:30")
        _escribir(client, b["id"], ENV)  # aquí no entra: sigue abierta en A
        ficha = client.get(f"/preauditoria/oficios/{b['id']}").json()
        falta = ficha["envios_escritos"][0]["faltantes"][0]
        assert falta["cerrado"] is False
        assert "resuélvala allá" in falta["motivo"]
        assert ficha["semaforo"]["estado"] != "COMPLETADO"


# ------------------------------------------------------------------
# ADRES devuelve el PAQUETE completo (regla de Cartera, 25-08-2026)
# ------------------------------------------------------------------

NIT_ADRES = "901037916"
ENT_ADRES = "ADMINISTRADORA DE LOS RECURSOS DEL SISTEMA GENERAL DE SEGURIDAD SOCIAL EN SALUD"


def _rad_adres(envio, factura, valor=250700):
    return _rad_fila(envio, factura, valor, nit=NIT_ADRES, entidad=ENT_ADRES)


class TestPaqueteAdresDevuelto:
    """ADRES no devuelve facturas sueltas: devuelve el paquete completo, así
    tenga facturas que pre-auditoría ya había dado por buenas. Esas tienen que
    poder volver a entrar. Para las demás entidades, una radicada NO se mueve."""

    def _radicada(self, client, filas, radicado="FHUS-AS-I01300-26"):
        _subir_radicacion(client, filas)
        _subir_dgreport(client, [f[1] for f in filas])
        o = _crear_oficio(client, radicado=radicado, fecha="2026-07-20T08:30")
        _escribir(client, o["id"], ENV)
        for f in filas:
            _radicar(client, _factura_id(client, f[1]))
        return o

    def test_la_radicada_de_adres_vuelve_a_entrar(self, client):
        a = self._radicada(client, [_rad_adres(ENV, F1)])
        b = _crear_oficio(client, radicado="FHUS-AS-I01301-26", fecha="2026-07-24T08:30")
        r = _escribir(client, b["id"], ENV)
        assert r.status_code == 200, r.text
        assert r.json()["reingresos_adres"] == 1
        assert r.json()["omitidas"] == 0

        ficha = client.get(f"/preauditoria/oficios/{b['id']}").json()
        assert [f["factura"] for f in ficha["facturas"]] == [F1]
        suya = ficha["facturas"][0]
        assert suya["resultado"] == "PENDIENTE"
        assert suya["estado"] == "EN_SUBSANACION"
        assert suya["ronda"] == 2
        # …y salió del oficio viejo.
        assert client.get(f"/preauditoria/oficios/{a['id']}").json()["total_facturas"] == 0

    def test_no_le_gasta_el_cupo_de_las_tres_devoluciones(self, client):
        """Ese contador es de las devoluciones que hace pre-auditoría a
        facturación. Que ADRES nos devuelva el paquete no es una devolución
        nuestra: si contara, tres paquetes devueltos bloquearían la factura."""
        self._radicada(client, [_rad_adres(ENV, F1)])
        b = _crear_oficio(client, radicado="FHUS-AS-I01301-26", fecha="2026-07-24T08:30")
        _escribir(client, b["id"], ENV)
        d = client.get(f"/preauditoria/facturas/{F1}/historial").json()["actual"]
        assert d["num_devoluciones"] == 0

    def test_la_de_otra_entidad_sigue_sin_poder_moverse(self, client):
        """La regla vieja se conserva para todas las demás: es la que evita
        que una factura ya radicada se cobre dos veces."""
        a = self._radicada(client, [_rad_fila(ENV, F1, 250700)])  # AXA COLPATRIA
        b = _crear_oficio(client, radicado="FHUS-AS-I01301-26", fecha="2026-07-24T08:30")
        r = _escribir(client, b["id"], ENV).json()
        assert r["reingresos_adres"] == 0
        assert r["omitidas"] == 1
        assert any("ya estaba radicada" in a for a in r["advertencias"])
        assert client.get(f"/preauditoria/oficios/{b['id']}").json()["total_facturas"] == 0
        assert client.get(f"/preauditoria/oficios/{a['id']}").json()["total_facturas"] == 1

    def test_ver_antes_lo_anuncia_antes_de_cargar(self, client):
        self._radicada(client, [_rad_adres(ENV, F1)])
        b = _crear_oficio(client, radicado="FHUS-AS-I01301-26", fecha="2026-07-24T08:30")
        p = client.get(f"/preauditoria/oficios/{b['id']}/envios/{ENV}/preview").json()
        assert p["reingresos_adres"] == 1
        assert p["ya_radicadas"] == 0  # ya no se cuenta como "no va a entrar"
        assert [f["clasificacion"] for f in p["facturas"]] == ["REINGRESO_ADRES"]

    def test_el_historial_deja_dicho_que_fue_por_adres(self, client):
        self._radicada(client, [_rad_adres(ENV, F1)])
        b = _crear_oficio(client, radicado="FHUS-AS-I01301-26", fecha="2026-07-24T08:30")
        _escribir(client, b["id"], ENV)
        eventos = client.get(f"/preauditoria/facturas/{F1}/historial").json()["eventos"]
        ultimo = eventos[-1]
        assert ultimo["tipo"] == "REINGRESO"  # el tipo NO cambia: el registro de envíos lo cuenta
        assert "ADRES devolvió el paquete completo" in (ultimo["motivo"] or "")
        assert "FHUS-AS-I01300-26" in (ultimo["motivo"] or "")

    def test_al_quitar_el_envio_la_factura_vuelve_a_RADICADA(self, client, db_session):
        """Y no a «devuelta»: nadie la devolvió nunca."""
        _actuar_como_admin(db_session)
        a = self._radicada(client, [_rad_adres(ENV, F1)])
        b = _crear_oficio(client, radicado="FHUS-AS-I01301-26", fecha="2026-07-24T08:30")
        _escribir(client, b["id"], ENV)
        r = client.delete(f"/preauditoria/oficios/{b['id']}/envios/{ENV}")
        assert r.status_code == 200, r.text
        d = client.get(f"/preauditoria/facturas/{F1}/historial").json()["actual"]
        assert d["resultado"] == "RADICAR"
        assert d["estado"] == "RADICADA"
        assert d["oficio_fhus"] == a["numero_radicado"]
        assert d["motivo_devolucion"] in (None, "")

    def test_al_quitar_la_factura_con_la_papelera_pasa_lo_mismo(self, client, db_session):
        _actuar_como_admin(db_session)
        self._radicada(client, [_rad_adres(ENV, F1)])
        b = _crear_oficio(client, radicado="FHUS-AS-I01301-26", fecha="2026-07-24T08:30")
        _escribir(client, b["id"], ENV)
        fid = _factura_id(client, F1)
        r = client.delete(f"/preauditoria/oficios/{b['id']}/facturas/{fid}")
        assert r.status_code == 200, r.text
        d = client.get(f"/preauditoria/facturas/{F1}/historial").json()["actual"]
        assert d["resultado"] == "RADICAR"
        assert d["estado"] == "RADICADA"

    def test_envio_con_facturas_de_adres_y_de_otra_entidad(self, client):
        """Caso borde real: en un mismo envío entra una de ADRES y una de otra
        entidad. Solo la de ADRES vuelve."""
        a = self._radicada(client, [_rad_adres(ENV, F1), _rad_fila(ENV, F2, 120000)])
        b = _crear_oficio(client, radicado="FHUS-AS-I01301-26", fecha="2026-07-24T08:30")
        r = _escribir(client, b["id"], ENV).json()
        assert r["reingresos_adres"] == 1
        assert r["omitidas"] == 1
        facturas_b = client.get(f"/preauditoria/oficios/{b['id']}").json()["facturas"]
        assert [f["factura"] for f in facturas_b] == [F1]
        facturas_a = client.get(f"/preauditoria/oficios/{a['id']}").json()["facturas"]
        assert [f["factura"] for f in facturas_a] == [F2]

    def test_una_devuelta_de_adres_sigue_reingresando_como_subsanacion(self, client):
        """No se rompió el camino normal: una ADRES devuelta reingresa como
        subsanación y SÍ cuenta su devolución."""
        _subir_radicacion(client, [_rad_adres(ENV, F1)])
        _subir_dgreport(client, [F1])
        a = _crear_oficio(client, radicado="FHUS-AS-I01300-26", fecha="2026-07-20T08:30")
        _escribir(client, a["id"], ENV)
        _devolver(client, _factura_id(client, F1), motivo="Falta FURIPS")
        b = _crear_oficio(client, radicado="FHUS-AS-I01301-26", fecha="2026-07-24T08:30")
        r = _escribir(client, b["id"], ENV).json()
        assert r["reingresos"] == 1
        assert r["reingresos_adres"] == 0
        d = client.get(f"/preauditoria/facturas/{F1}/historial").json()["actual"]
        assert d["num_devoluciones"] == 1


class TestObservacionLarga:
    """25-08-2026: al pegar el mensaje de error de ADRES SIA (larguísimo) en la
    observación, el servidor lo rechazaba y la pantalla mostraba
    «[object Object]». Ni se guardaba, ni se entendía por qué."""

    def _pendiente(self, client):
        _subir_radicacion(client, [_rad_fila(ENV, F1, 250700)])
        _subir_dgreport(client, [F1])
        o = _crear_oficio(client, radicado="FHUS-AS-I01400-26")
        _escribir(client, o["id"], ENV)
        return _factura_id(client, F1)

    def test_cabe_el_mensaje_de_error_de_adres(self, client):
        fid = self._pendiente(client)
        texto = (
            "ERROR EN ADRES SIA *Verificando el cumplimiento de requisitos mínimos, no es "
            "posible realizar la radicación debido a que el documento Formulario Único de "
            "Reclamaciones - FUR no fue anexado o el nombre del archivo no cumple con la "
            "nomenclatura establecida. " * 12
        )
        assert 2000 < len(texto) <= 4000
        r = client.patch(
            f"/preauditoria/facturas/{fid}/auditar",
            json={"resultado": "DEVUELTA", "motivo_devolucion": texto, "observaciones": texto},
        )
        assert r.status_code == 200, r.text
        assert (
            client.get(f"/preauditoria/facturas/{F1}/historial")
            .json()["actual"]["motivo_devolucion"]
            .startswith("ERROR EN ADRES SIA")
        )

    def test_pero_el_tope_sigue_existiendo(self, client):
        fid = self._pendiente(client)
        r = client.patch(
            f"/preauditoria/facturas/{fid}/auditar",
            json={"resultado": "DEVUELTA", "motivo_devolucion": "x" * 4100},
        )
        assert r.status_code == 422


class TestElNumeroDeFacturaSeCruzaAunqueVengaCorto:
    """26-08-2026: tres facturas salían con «Correo F.E.: NO» teniéndola. El
    Formato F.E. traía el número corto («544942») y la Radicación el largo
    («HUS0000544942»): el cruce fallaba en silencio y, sin correo, el sistema
    no deja radicar — el auditor tenía que devolver una factura buena."""

    def test_el_formato_fe_con_el_numero_corto_igual_cruza(self, client):
        _subir_radicacion(client, [_rad_fila(ENV, "HUS0000544942", 250700)])
        # El DGH exportó el número sin el HUS ni los ceros.
        r = client.post(
            "/preauditoria/fuentes/dgreport",
            files={
                "archivo": (
                    "dgreport.xlsx",
                    _excel(DG_HEADERS, [["2026-08-26 09:00", "544942", "CUFE1"]]),
                    "application/xlsx",
                )
            },
        )
        assert r.status_code == 200, r.text
        o = _crear_oficio(client, radicado="FHUS-AS-I01500-26")
        _escribir(client, o["id"], ENV)
        ficha = client.get(f"/preauditoria/oficios/{o['id']}").json()
        assert ficha["facturas"][0]["correo_fe"] == "SI", (
            "la factura sigue saliendo sin facturación electrónica: no se puede radicar"
        )

    def test_y_por_eso_ya_se_puede_radicar(self, client):
        _subir_radicacion(client, [_rad_fila(ENV, "HUS0000542599", 120000)])
        client.post(
            "/preauditoria/fuentes/dgreport",
            files={
                "archivo": (
                    "dgreport.xlsx",
                    _excel(DG_HEADERS, [["2026-08-26 09:00", "HUS542599", "CUFE2"]]),
                    "application/xlsx",
                )
            },
        )
        o = _crear_oficio(client, radicado="FHUS-AS-I01501-26")
        _escribir(client, o["id"], ENV)
        fid = _factura_id(client, "HUS0000542599")
        assert _radicar(client, fid).status_code == 200

    def test_la_radicacion_con_numero_corto_tambien_se_guarda_larga(self, client):
        """Así las dos fuentes hablan el mismo idioma."""
        _subir_radicacion(client, [_rad_fila(ENV, "544936", 99900)])
        d = client.get("/preauditoria/fuentes/radicacion").json()
        assert any(f["factura"] == "HUS0000544936" for f in d["items"])
