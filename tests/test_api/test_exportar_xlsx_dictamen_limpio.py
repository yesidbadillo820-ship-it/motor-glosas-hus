"""Regresión: la columna "Dictamen HUS" de GET /glosas/exportar-xlsx
debe traer texto plano legible, nunca el HTML crudo del dictamen."""

from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.core.tz import ahora_utc
from app.database import Base, get_db
from app.models.db import GlosaRecord, UsuarioRecord

DICTAMEN_HTML = (
    '<table border="1" style="width:100%;border-collapse:collapse;">'
    "<tr><th>CÓDIGO GLOSA</th><th>VALOR OBJETADO</th><th>CÓDIGO RESPUESTA</th></tr>"
    "<tr><td>TA0801</td><td>$36.402</td><td><b>RE9901</b></td></tr></table>"
    '<div style="background:#f8fafc;">ARGUMENTACIÓN JURÍDICA<br/>'
    "Se objeta la glosa por aplicar tarifa contractual vigente.</div>"
)


@pytest.fixture
def db_session():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    s = sessionmaker(bind=engine)()
    try:
        yield s
    finally:
        s.close()
        engine.dispose()


@pytest.fixture
def usuario():
    return UsuarioRecord(id=1, email="auditor@hus.com", rol="AUDITOR", activo=1)


@pytest.fixture
def client(db_session, usuario):
    from app.api.deps import get_usuario_actual
    from app.main import app

    app.dependency_overrides[get_db] = lambda: iter([db_session]).__next__()
    app.dependency_overrides[get_usuario_actual] = lambda: usuario
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.clear()


def test_dictamen_hus_sin_html_crudo(client, db_session):
    """Verifica que la respuesta del HUS en el Excel exportado salga como
    texto plano legible, NUNCA con HTML crudo.

    Cambio 12-jun-2026: el endpoint /exportar-xlsx ahora genera el formato
    RADICABLE institucional (PR #111) en vez del layout viejo de 22
    columnas planas. Estructura:
      fila 1 = título institucional (merged), fila 2 = línea del contrato,
      fila 3 = header de 12 columnas (col 12 = "RESPUESTA ESE HUS"),
      filas 4+ = una por concepto/glosa.
    """
    db_session.add(
        GlosaRecord(
            eps="SANITAS",
            paciente="Paciente X",
            factura="HUS999",
            codigo_glosa="TA0801",
            valor_objetado=36402,
            etapa="OBJECION",
            estado="RESPONDIDA",
            dictamen=DICTAMEN_HTML,
            creado_en=ahora_utc(),
        )
    )
    db_session.commit()

    r = client.get("/glosas/exportar-xlsx")
    assert r.status_code == 200, r.text
    ws = load_workbook(io.BytesIO(r.content)).active
    # Header está en fila 3 del formato radicable; la columna "RESPUESTA"
    # (la 12) lleva el dictamen — buscarla por nombre por si cambia el
    # orden en el futuro.
    headers_f3 = [str(c.value or "") for c in ws[3]]
    col_resp = next(
        (i for i, h in enumerate(headers_f3, start=1) if "RESPUESTA" in h.upper()),
        12,
    )
    # Primera fila de datos: fila 4
    celda = ws.cell(row=4, column=col_resp).value or ""
    assert "<table" not in celda
    assert "style=" not in celda
    # Encabezado tabular convertido a "CÓDIGO GLOSA: ... | VALOR OBJETADO: ..."
    assert "TA0801" in celda
    assert "36.402" in celda or "$36.402" in celda
    # Argumentación jurídica conservada en texto plano
    assert "Se objeta la glosa por aplicar tarifa contractual vigente." in celda
