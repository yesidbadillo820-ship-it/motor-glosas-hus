"""Tests del vault cifrado de credenciales EPS (/credenciales).

Datos 100% sintéticos — jamás credenciales reales. El xlsx se construye
in-memory replicando el layout del Excel maestro (hoja CONSOLIDADO,
banner en fila 1, encabezados en fila 2, bloques RADICACION/CARTERA/
DEVOLUCIONES).
"""

from __future__ import annotations

import io

import pytest
from cryptography.fernet import Fernet
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base, get_db
from app.models.db import (
    CredencialAccesoLog,
    EntidadCredencialRecord,
    UsuarioRecord,
)

# ─── Fixtures ────────────────────────────────────────────────────────────────

CLAVE_TEST = Fernet.generate_key().decode()

ENCABEZADOS = [
    "NIT",
    "EMPRESAS",
    "LINK PLATAFORMA",
    "USUARIO ",
    "CONTRASEÑA",
    "TELÉFONO",
    "CORREO",
    "NOMBRE",
    "CARGO",
    "LINK PLATAFORMA",
    "USUARIO",
    "CLAVE",
    "PLATAFORMA\nFISICO\nCORREO",
    "LINK PLATAFORMA\nDIRECCIÓN\nEMAIL",
    "USUARIO \nCONTRASEÑA",
    "MANUAL DE RADICACION",
    "MEDIO DE CONTACTO",
    "NOMBRE CONTACTO",
]


def _xlsx_sintetico() -> bytes:
    """Excel con 2 entidades ficticias + 1 fila sin NIT."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CONSOLIDADO"
    banner = [""] * 18
    banner[2], banner[9], banner[12] = "RADICACION", "CARTERA", "DEVOLUCIONES"
    ws.append(banner)
    ws.append(ENCABEZADOS)
    # Entidad 1: bloques RADICACION + CARTERA
    ws.append(
        [
            900000001,
            "ENTIDAD PRUEBA UNO",
            "https://radicar.prueba-uno.example.com",
            "user_demo",
            "pass_demo",
            3000000001,
            "soporte@prueba-uno.example.com",
            "Ana Contacto",
            "Analista cartera",
            "https://cartera.prueba-uno.example.com",
            "cart_demo",
            "clave_cart_demo",
            "",
            "",
            "",
            "Radicar antes del día 20 de cada mes",
            "Correo certificado",
            "Mesa de ayuda",
        ]
    )
    # Entidad 2: solo DEVOLUCIONES, celda combinada usuario\ncontraseña
    ws.append(
        [
            "900000002",
            "ENTIDAD PRUEBA DOS",
            *[""] * 10,
            "PLATAFORMA",
            "https://devoluciones.prueba-dos.example.com",
            "dev_demo\nclave_dev_demo",
            "",
            "",
            "Pedro Contacto",
        ]
    )
    # Fila con empresa pero SIN NIT → filas_sin_nit
    ws.append(["", "ENTIDAD SIN NIT", "https://x.example.com"] + [""] * 15)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def db_session():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    s = sessionmaker(bind=engine)()
    try:
        yield s
    finally:
        s.close()
        engine.dispose()


@pytest.fixture
def usuario_actual():
    """Usuario mutable: los tests cambian .rol para probar gates."""
    return UsuarioRecord(
        id=1,
        email="vault-tester@hus.gov.co",
        nombre="Vault Tester",
        rol="SUPER_ADMIN",
        activo=1,
        password_hash="x",
    )


@pytest.fixture
def client(db_session, usuario_actual, monkeypatch):
    monkeypatch.setenv("CRED_VAULT_KEY", CLAVE_TEST)
    from app.api.deps import get_usuario_actual
    from app.main import app

    def _db_override():
        yield db_session

    app.dependency_overrides[get_db] = _db_override
    app.dependency_overrides[get_usuario_actual] = lambda: usuario_actual
    # SIN context manager: no disparar el lifespan pesado (ver conftest).
    test_client = TestClient(app)
    yield test_client
    app.dependency_overrides.clear()


def _importar(client, contenido=None):
    return client.post(
        "/credenciales/importar",
        files={
            "archivo": (
                "credenciales_test.xlsx",
                contenido or _xlsx_sintetico(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


# ─── Tests ───────────────────────────────────────────────────────────────────


class TestImportar:
    def test_conteos(self, client):
        r = _importar(client)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["entidades"] == 2
        # UNO: RADICACION + CARTERA · DOS: DEVOLUCIONES
        assert d["credenciales_creadas"] == 3
        assert d["actualizadas"] == 0
        assert d["filas_sin_nit"] == 1

    def test_respuesta_sin_valores(self, client):
        r = _importar(client)
        assert "user_demo" not in r.text
        assert "pass_demo" not in r.text

    def test_cifrado_at_rest(self, client, db_session):
        from app.services.credenciales_vault import descifrar

        _importar(client)
        reg = (
            db_session.query(EntidadCredencialRecord)
            .filter_by(nit="900000001", bloque="RADICACION")
            .one()
        )
        # En BD jamás el texto plano — tokens Fernet (prefijo gAAAA)
        for col, plano in [
            ("usuario_cifrado", "user_demo"),
            ("password_cifrado", "pass_demo"),
            ("telefono_cifrado", "3000000001"),
            ("correo_cifrado", "soporte@prueba-uno.example.com"),
        ]:
            token = getattr(reg, col)
            assert token != plano
            assert token.startswith("gAAAA")
            assert descifrar(token) == plano
        # Lo no-secreto queda en claro
        assert reg.link_plataforma == "https://radicar.prueba-uno.example.com"
        assert reg.nombre_contacto == "Ana Contacto"
        assert reg.cargo == "Analista cartera"
        assert reg.manual_radicacion == "Radicar antes del día 20 de cada mes"

    def test_celda_combinada_devoluciones(self, client, db_session):
        from app.services.credenciales_vault import descifrar

        _importar(client)
        reg = (
            db_session.query(EntidadCredencialRecord)
            .filter_by(nit="900000002", bloque="DEVOLUCIONES")
            .one()
        )
        assert descifrar(reg.usuario_cifrado) == "dev_demo"
        assert descifrar(reg.password_cifrado) == "clave_dev_demo"
        assert reg.medio_contacto == "PLATAFORMA"
        # nombre de contacto de entidad (col R) aplica como fallback
        assert reg.nombre_contacto == "Pedro Contacto"

    def test_upsert_idempotente(self, client, db_session):
        _importar(client)
        r2 = _importar(client)
        d = r2.json()
        assert d["credenciales_creadas"] == 0
        assert d["actualizadas"] == 3
        assert db_session.query(EntidadCredencialRecord).count() == 3

    def test_audit_importar(self, client, db_session):
        _importar(client)
        log = db_session.query(CredencialAccesoLog).filter_by(accion="IMPORTAR").one()
        assert log.usuario_email == "vault-tester@hus.gov.co"
        assert "creadas=3" in log.motivo

    def test_solo_super_admin(self, client, usuario_actual):
        usuario_actual.rol = "COORDINADOR"
        assert _importar(client).status_code == 403
        usuario_actual.rol = "AUDITOR"
        assert _importar(client).status_code == 403

    def test_extension_invalida(self, client):
        r = client.post(
            "/credenciales/importar",
            files={"archivo": ("malo.csv", b"a,b", "text/csv")},
        )
        assert r.status_code == 400

    def test_excel_sin_hoja_consolidado(self, client):
        import openpyxl

        wb = openpyxl.Workbook()
        wb.active.title = "OTRA"
        buf = io.BytesIO()
        wb.save(buf)
        r = _importar(client, contenido=buf.getvalue())
        assert r.status_code == 400
        assert "CONSOLIDADO" in r.json()["detail"]

    def test_archivo_corrupto_400(self, client):
        r = _importar(client, contenido=b"esto no es un zip/xlsx")
        assert r.status_code == 400

    def test_nit_bloque_duplicado_en_archivo(self, client, db_session):
        """Mismo (nit, bloque) dos veces en el Excel → last-wins, sin
        chocar con el índice único ni inflar conteos."""
        import openpyxl

        from app.services.credenciales_vault import descifrar

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CONSOLIDADO"
        ws.append([""] * 18)
        ws.append(ENCABEZADOS)
        fila = lambda u, p: (
            [  # noqa: E731
                900000009,
                "ENTIDAD REPETIDA",
                "https://rep.example.com",
                u,
                p,
            ]
            + [""] * 13
        )
        ws.append(fila("user_v1", "pass_v1"))
        ws.append(fila("user_v2", "pass_v2"))
        buf = io.BytesIO()
        wb.save(buf)
        r = _importar(client, contenido=buf.getvalue())
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["entidades"] == 1  # NIT único aunque venga en 2 filas
        assert d["credenciales_creadas"] == 1
        assert d["actualizadas"] == 0
        regs = db_session.query(EntidadCredencialRecord).filter_by(nit="900000009").all()
        assert len(regs) == 1
        assert descifrar(regs[0].usuario_cifrado) == "user_v2"


class TestBuscar:
    def test_sin_secretos(self, client):
        _importar(client)
        r = client.get("/credenciales/buscar", params={"q": "PRUEBA UNO"})
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["total"] == 2  # RADICACION + CARTERA
        # ningún secreto, ni en claro ni cifrado
        assert "user_demo" not in r.text
        assert "pass_demo" not in r.text
        assert "gAAAA" not in r.text
        item = next(i for i in d["items"] if i["bloque"] == "RADICACION")
        assert item["tiene_usuario"] is True
        assert item["tiene_password"] is True
        assert item["link_plataforma"] == "https://radicar.prueba-uno.example.com"
        assert "usuario" not in item and "password" not in item

    def test_por_nit(self, client):
        _importar(client)
        r = client.get("/credenciales/buscar", params={"q": "900000002"})
        d = r.json()
        assert d["total"] == 1
        assert d["items"][0]["bloque"] == "DEVOLUCIONES"
        assert d["items"][0]["empresa"] == "ENTIDAD PRUEBA DOS"

    def test_cualquier_rol_autenticado(self, client, usuario_actual):
        _importar(client)
        usuario_actual.rol = "AUDITOR"
        assert client.get("/credenciales/buscar").status_code == 200
        usuario_actual.rol = "VIEWER"
        assert client.get("/credenciales/buscar").status_code == 200

    def test_funciona_sin_vault_key(self, client, usuario_actual, monkeypatch):
        _importar(client)
        monkeypatch.delenv("CRED_VAULT_KEY", raising=False)
        r = client.get("/credenciales/buscar", params={"q": "PRUEBA"})
        assert r.status_code == 200
        assert r.json()["total"] == 3


class TestRevelar:
    def _id_radicacion(self, client) -> int:
        r = client.get("/credenciales/buscar", params={"q": "PRUEBA UNO"})
        return next(i["id"] for i in r.json()["items"] if i["bloque"] == "RADICACION")

    def test_revelar_con_motivo_y_audit(self, client, db_session, usuario_actual):
        _importar(client)
        cred_id = self._id_radicacion(client)
        usuario_actual.rol = "COORDINADOR"
        r = client.post(
            f"/credenciales/{cred_id}/revelar",
            json={"motivo": "Radicación factura HUS123456"},
        )
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["usuario"] == "user_demo"
        assert d["password"] == "pass_demo"
        log = (
            db_session.query(CredencialAccesoLog)
            .filter_by(accion="REVELAR", credencial_id=cred_id)
            .one()
        )
        assert log.motivo == "Radicación factura HUS123456"
        assert log.usuario_email == "vault-tester@hus.gov.co"

    def test_motivo_obligatorio_y_minimo(self, client):
        _importar(client)
        cred_id = self._id_radicacion(client)
        assert client.post(f"/credenciales/{cred_id}/revelar", json={}).status_code == 422
        r = client.post(f"/credenciales/{cred_id}/revelar", json={"motivo": "abc"})
        assert r.status_code == 422

    def test_auditor_403(self, client, db_session, usuario_actual):
        _importar(client)
        cred_id = self._id_radicacion(client)
        usuario_actual.rol = "AUDITOR"
        r = client.post(
            f"/credenciales/{cred_id}/revelar",
            json={"motivo": "intento sin permiso"},
        )
        assert r.status_code == 403
        # el 403 ocurre ANTES del handler → no debe haber audit REVELAR
        n = db_session.query(CredencialAccesoLog).filter_by(accion="REVELAR").count()
        assert n == 0

    def test_404(self, client):
        r = client.post("/credenciales/99999/revelar", json={"motivo": "no existe el id"})
        assert r.status_code == 404


class TestVaultNoConfigurado:
    def test_importar_503(self, client, monkeypatch):
        monkeypatch.delenv("CRED_VAULT_KEY", raising=False)
        r = _importar(client)
        assert r.status_code == 503
        assert "CRED_VAULT_KEY" in r.json()["detail"]

    def test_revelar_503(self, client, monkeypatch):
        _importar(client)
        cred_id = client.get("/credenciales/buscar").json()["items"][0]["id"]
        monkeypatch.delenv("CRED_VAULT_KEY", raising=False)
        r = client.post(f"/credenciales/{cred_id}/revelar", json={"motivo": "vault apagado"})
        assert r.status_code == 503

    def test_clave_invalida_es_vault_off(self, monkeypatch):
        from app.services import credenciales_vault as v

        monkeypatch.setenv("CRED_VAULT_KEY", "no-es-una-clave-fernet")
        assert v.vault_disponible() is False
        monkeypatch.delenv("CRED_VAULT_KEY", raising=False)
        assert v.vault_disponible() is False
        monkeypatch.setenv("CRED_VAULT_KEY", CLAVE_TEST)
        assert v.vault_disponible() is True


class TestAudit:
    def test_solo_admin(self, client, usuario_actual):
        usuario_actual.rol = "COORDINADOR"
        assert client.get("/credenciales/audit").status_code == 403

    def test_lista_accesos(self, client):
        _importar(client)
        r = client.get("/credenciales/audit")
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["total"] >= 1
        assert d["items"][0]["accion"] == "IMPORTAR"
        assert d["items"][0]["usuario_email"] == "vault-tester@hus.gov.co"


class TestParser:
    """Unit tests del parser puro (sin HTTP)."""

    def test_separar_usuario_password(self):
        from app.services.credenciales_vault import _separar_usuario_password as sep

        assert sep("usr\npwd") == ("usr", "pwd")
        assert sep("USUARIO: usr\nCLAVE: pwd") == ("usr", "pwd")
        assert sep("usr\npwd parte1\npwd parte2") == ("usr", "pwd parte1 pwd parte2")
        assert sep("solo-una-linea") == ("solo-una-linea", "")
        assert sep("") == ("", "")

    def test_nit_numerico_normalizado(self):
        from app.services.credenciales_vault import parsear_consolidado

        res = parsear_consolidado(_xlsx_sintetico())
        nits = sorted(e["nit"] for e in res["entidades"])
        # 900000001 entra como número en Excel → sin '.0'
        assert nits == ["900000001", "900000002"]
        assert res["filas_sin_nit"] == 1

    def test_bloques_detectados(self):
        from app.services.credenciales_vault import parsear_consolidado

        res = parsear_consolidado(_xlsx_sintetico())
        por_nit = {e["nit"]: e for e in res["entidades"]}
        assert set(por_nit["900000001"]["bloques"]) == {"RADICACION", "CARTERA"}
        assert set(por_nit["900000002"]["bloques"]) == {"DEVOLUCIONES"}
