"""El acta de la mesa entra en Excel y sale cuadrada, optimizada y firmable.

Se prueba con un libro sintético que imita el formato real ACTA SINAC
(encabezado con NIT/razón/periodo, tabla desde la columna C, fila TOTAL,
firmas al pie): el lector encuentra todo por etiquetas, la revisión señala
exactamente lo que no cuadra, la optimización agrega sin sobreescribir y
el acta imprimible trae la cláusula de mérito ejecutivo y las firmas.
"""

from __future__ import annotations

import io
import json

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.database import Base, get_db
from app.models.db import AuditLogRecord, UsuarioRecord
from app.services import acta_conciliacion_excel as ax

ENCABEZADOS = [
    "ITEM",
    "RADICADO / ACTA ENTIDAD",
    "NUMERO FACTURA",
    "FECHA FACTURA",
    "TIPO DE GLOSA \nADM - MIX - MED",
    "TIPIFICACION",
    "COD GLOSA",
    "DESCRIPCION GLOSA",
    "VALOR FACTURA",
    "VALOR GLOSA INICIAL",
    "VALOR PENDIENTE POR CONCILIAR",
    "VALOR ACEPTA IPS",
    "VALOR LEVANTA ENTIDAD",
    "VALOR RATIFICADO",
    "DESCRIPCION DE CONCILIACION",
]


def _libro(
    lineas: list[dict],
    nit: str = "900123456",
    razon: str = "ENTIDAD DE PRUEBA",
    declara_facturas: int | None = None,
    declara_valor: float | None = None,
    total: dict | None = None,
    con_firmas: bool = True,
) -> bytes:
    """Arma un libro con la misma anatomía del acta real (tabla en C..Q)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ACTA"
    ws["C2"] = "NIT:"
    ws["D2"] = nit
    ws["G2"] = "Razon Social :"
    ws["H2"] = razon
    ws["C3"] = "PERIODO  DE CONCILIACION"
    ws["F3"] = "2026"
    ws["I3"] = "CANTIDAD FACTURAS"
    if declara_facturas is not None:
        ws["K3"] = declara_facturas
    ws["L3"] = "VALOR A CONCILIAR"
    if declara_valor is not None:
        ws["M3"] = declara_valor
    for j, et in enumerate(ENCABEZADOS, start=3):
        ws.cell(row=5, column=j, value=et)
    fila = 6
    for i, ln in enumerate(lineas, start=1):
        valores = [
            ln.get("item", i),
            ln.get("radicado", "AR001"),
            ln.get("factura", ""),
            "2026-01-15",
            "ADMINISTRATIVA",
            ln.get("tipificacion", "TARIFAS"),
            ln.get("cod", "TA0601"),
            ln.get("desc", "SE GLOSA EL SERVICIO DE PRUEBA"),
            ln.get("v_factura", 1_000_000),
            ln.get("glosa", 100_000),
            ln.get("pendiente", ln.get("glosa", 100_000)),
            ln.get("acepta", 0),
            ln.get("levanta", 0),
            ln.get("ratifica", 0),
            ln.get("texto", ""),
        ]
        for j, v in enumerate(valores, start=3):
            ws.cell(row=fila, column=j, value=v)
        fila += 1
    if total is not None:
        ws.cell(row=fila, column=9, value="TOTAL ")
        ws.cell(row=fila, column=11, value=total.get("v_factura", 0))
        ws.cell(row=fila, column=12, value=total.get("glosa", 0))
        fila += 1
    if con_firmas:
        base = fila + 2
        ws.cell(row=base, column=3, value="NOMBRE: ")
        ws.cell(row=base, column=4, value="AUDITORA DE PRUEBA")
        ws.cell(row=base + 1, column=3, value="CARGO: ")
        ws.cell(row=base + 1, column=4, value="MEDICO AUDITOR")
        ws.cell(row=base + 2, column=3, value="CORREO ")
        ws.cell(row=base + 2, column=4, value="auditora@hus.gov.co")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestLectura:
    def test_lee_encabezado_lineas_y_firmas(self):
        contenido = _libro(
            [
                {"factura": "HUS001", "glosa": 100_000},
                {"factura": "HUS002", "glosa": 50_000},
            ],
            declara_facturas=2,
            declara_valor=150_000,
        )
        acta = ax.leer_acta(contenido)
        assert acta.nit == "900123456"
        assert acta.razon_social == "ENTIDAD DE PRUEBA"
        assert acta.cantidad_facturas_declarada == 2
        assert acta.valor_a_conciliar_declarado == 150_000
        assert len(acta.lineas) == 2
        assert acta.firmas and acta.firmas[0]["nombre"] == "AUDITORA DE PRUEBA"

    def test_factura_combinada_se_hereda(self):
        """En el acta real la factura va en celda combinada: solo la primera
        fila del grupo la trae. Las siguientes la heredan."""
        contenido = _libro(
            [
                {"factura": "HUS001", "glosa": 100_000},
                {"factura": "", "glosa": 30_000},  # misma factura, celda vacía
            ]
        )
        acta = ax.leer_acta(contenido)
        assert [ln.factura for ln in acta.lineas] == ["HUS001", "HUS001"]

    def test_libro_sin_tabla_da_error_claro(self):
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(ValueError):
            ax.leer_acta(buf.getvalue())


class TestRevision:
    def test_resultados_por_linea(self):
        contenido = _libro(
            [
                {"factura": "F1", "glosa": 100, "levanta": 100, "texto": "levantada"},
                {"factura": "F2", "glosa": 100, "ratifica": 100, "texto": "ratificada"},
                {"factura": "F3", "glosa": 100, "acepta": 100, "texto": "acepta ips"},
                {"factura": "F4", "glosa": 100, "levanta": 40, "texto": "parcial"},
                {"factura": "F5", "glosa": 100},
            ]
        )
        acta = ax.leer_acta(contenido)
        assert [ln.resultado for ln in acta.lineas] == [
            "LEVANTADA TOTAL",
            "RATIFICADA TOTAL",
            "ACEPTADA POR LA IPS",
            "PARCIAL (queda pendiente)",
            "PENDIENTE",
        ]

    def test_acta_que_cuadra_queda_lista_para_firmar(self):
        contenido = _libro(
            [{"factura": "F1", "glosa": 100_000, "levanta": 100_000, "texto": "ok"}],
            declara_facturas=1,
            declara_valor=100_000,
        )
        r = ax.revisar(ax.leer_acta(contenido))
        assert r["hallazgos"] == []
        assert r["lista_para_firmar"] is True
        assert r["levanta_entidad"] == 100_000

    def test_hallazgos_concretos(self):
        contenido = _libro(
            [
                # reparte 150 sobre una glosa de 100
                {"factura": "F1", "glosa": 100, "levanta": 100, "acepta": 50, "texto": "x"},
                # conciliada pero sin texto de conciliación
                {"factura": "F2", "glosa": 100, "levanta": 100},
                # pendiente que no es ni lo glosado ni lo que queda tras mesa
                {"factura": "F3", "glosa": 100, "levanta": 100, "pendiente": 37, "texto": "x"},
            ],
            declara_facturas=5,  # el acta dice 5 y hay 3
            declara_valor=999_999,  # y declara otro valor
        )
        r = ax.revisar(ax.leer_acta(contenido))
        problemas = {h["problema"] for h in r["hallazgos"]}
        assert "reparte más de lo glosado" in problemas
        assert "conciliada sin texto" in problemas
        assert "el pendiente no cuadra" in problemas
        assert "la cantidad de facturas no cuadra" in problemas
        assert "el valor a conciliar no cuadra" in problemas
        assert r["lista_para_firmar"] is False

    def test_ambas_convenciones_de_pendiente_valen(self):
        """La columna de pendiente puede traer lo que ENTRÓ a mesa o lo que
        QUEDÓ: ninguna de las dos es hallazgo."""
        contenido = _libro(
            [
                {"factura": "F1", "glosa": 100, "levanta": 100, "pendiente": 100, "texto": "x"},
                {"factura": "F2", "glosa": 100, "levanta": 100, "pendiente": 0, "texto": "x"},
            ]
        )
        r = ax.revisar(ax.leer_acta(contenido))
        assert not [h for h in r["hallazgos"] if h["problema"] == "el pendiente no cuadra"]


class TestOptimizacion:
    def test_agrega_sin_sobreescribir(self):
        contenido = _libro(
            [{"factura": "F1", "glosa": 100, "levanta": 100, "texto": "LO ACORDADO EN MESA"}]
        )
        salida, resumen = ax.optimizar(contenido, es_xlsm=False)
        wb = openpyxl.load_workbook(io.BytesIO(salida))
        assert "REVISION" in wb.sheetnames
        ws = wb["ACTA"]
        # La columna nueva quedó al final con el resultado de la línea
        col = ws.max_column
        assert ws.cell(row=5, column=col).value == "RESULTADO DE LA MESA (auto)"
        assert ws.cell(row=6, column=col).value == "LEVANTADA TOTAL"
        # Y lo que escribió el auditor está intacto (columna Q = 17)
        assert ws.cell(row=6, column=17).value == "LO ACORDADO EN MESA"

    def test_revision_lista_los_hallazgos(self):
        contenido = _libro(
            [{"factura": "F1", "glosa": 100, "levanta": 100, "acepta": 50, "texto": "x"}]
        )
        salida, resumen = ax.optimizar(contenido, es_xlsm=False)
        wb = openpyxl.load_workbook(io.BytesIO(salida))
        rev = wb["REVISION"]
        texto = " ".join(str(c.value) for fila in rev.iter_rows() for c in fila if c.value)
        assert "reparte más de lo glosado" in texto
        assert "F1" in texto


class TestActaImprimible:
    def test_trae_lo_que_un_acta_debe_traer(self):
        contenido = _libro(
            [{"factura": "HUS0009", "glosa": 250_000, "levanta": 250_000, "texto": "SE LEVANTA"}]
        )
        acta = ax.leer_acta(contenido)
        html = ax.html_acta(acta)
        assert "900123456" in html  # NIT
        assert "ENTIDAD DE PRUEBA" in html
        assert "HUS0009" in html
        assert "$250.000" in html  # pesos con punto de miles
        assert "mérito ejecutivo" in html
        assert "AUDITORA DE PRUEBA" in html  # firma leída del libro
        assert "LEVANTADA TOTAL" in html  # resultado de la línea

    def test_acta_con_hallazgos_avisa_antes_de_firmar(self):
        contenido = _libro(
            [{"factura": "F1", "glosa": 100, "levanta": 100, "acepta": 50, "texto": "x"}]
        )
        acta = ax.leer_acta(contenido)
        html = ax.html_acta(acta)
        assert "hallazgo(s) de cuadre sin resolver" in html


# ─── Endpoints: la capacidad completa desde la pantalla Conciliación ─────


@pytest.fixture
def db_session():
    from sqlalchemy.pool import StaticPool

    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    S = sessionmaker(bind=engine)
    s = S()
    try:
        yield s
    finally:
        s.close()
        engine.dispose()


def _cliente(db_session, rol: str) -> TestClient:
    from app.api.deps import get_usuario_actual
    from app.main import app

    usuario = UsuarioRecord(id=1, email="auditor@hus.gov.co", nombre="Test", rol=rol, activo=1)
    app.dependency_overrides[get_db] = lambda: iter([db_session]).__next__()
    app.dependency_overrides[get_usuario_actual] = lambda: usuario
    return TestClient(app)


@pytest.fixture
def cliente_auditor(db_session):
    from app.main import app

    c = _cliente(db_session, "AUDITOR")
    with c:
        yield c
    app.dependency_overrides.clear()


@pytest.fixture
def cliente_viewer(db_session):
    from app.main import app

    c = _cliente(db_session, "VIEWER")
    with c:
        yield c
    app.dependency_overrides.clear()


def _subir(cliente, ruta: str, contenido: bytes, nombre: str = "acta.xlsx"):
    return cliente.post(
        ruta,
        files={
            "archivo": (
                nombre,
                contenido,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


class TestEndpoints:
    def test_revisar_devuelve_el_cuadre(self, cliente_auditor):
        contenido = _libro([{"factura": "F1", "glosa": 100_000, "levanta": 100_000, "texto": "ok"}])
        r = _subir(cliente_auditor, "/conciliaciones/acta-excel/revisar", contenido)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["entidad"] == "ENTIDAD DE PRUEBA"
        assert d["facturas"] == 1 and d["lineas"] == 1
        assert d["lista_para_firmar"] is True

    def test_optimizar_descarga_y_deja_constancia(self, cliente_auditor, db_session):
        contenido = _libro(
            [{"factura": "F1", "glosa": 100, "levanta": 100, "acepta": 50, "texto": "x"}]
        )
        r = _subir(cliente_auditor, "/conciliaciones/acta-excel/optimizar", contenido)
        assert r.status_code == 200, r.text
        assert "_OPTIMIZADA.xlsx" in r.headers["content-disposition"]
        assert r.headers["X-Hallazgos"] == "1"
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        assert "REVISION" in wb.sheetnames

        # Expediente: quedó la constancia con el cuadre
        fila = (
            db_session.query(AuditLogRecord)
            .filter(AuditLogRecord.accion == "ACTA_EXCEL_OPTIMIZADA")
            .one()
        )
        assert fila.campo == "CON_HALLAZGOS"
        assert json.loads(fila.detalle)["resumen"]["hallazgos"] == 1

    def test_pdf_devuelve_acta_firmable(self, cliente_auditor, db_session):
        contenido = _libro(
            [{"factura": "HUS0007", "glosa": 90_000, "ratifica": 90_000, "texto": "ratificada"}]
        )
        r = _subir(cliente_auditor, "/conciliaciones/acta-excel/pdf", contenido)
        assert r.status_code == 200, r.text
        assert r.headers["content-type"].startswith("text/html")
        assert "HUS0007" in r.text
        assert "mérito ejecutivo" in r.text
        assert (
            db_session.query(AuditLogRecord)
            .filter(AuditLogRecord.accion == "ACTA_EXCEL_PDF")
            .count()
            == 1
        )

    def test_viewer_no_puede(self, cliente_viewer):
        contenido = _libro([{"factura": "F1", "glosa": 100}])
        for ruta in (
            "/conciliaciones/acta-excel/revisar",
            "/conciliaciones/acta-excel/optimizar",
            "/conciliaciones/acta-excel/pdf",
        ):
            assert _subir(cliente_viewer, ruta, contenido).status_code == 403

    def test_archivo_que_no_es_excel_da_400(self, cliente_auditor):
        r = _subir(cliente_auditor, "/conciliaciones/acta-excel/revisar", b"hola", "acta.txt")
        assert r.status_code == 400

    def test_excel_que_no_es_acta_da_400(self, cliente_auditor):
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        r = _subir(cliente_auditor, "/conciliaciones/acta-excel/revisar", buf.getvalue())
        assert r.status_code == 400
        assert "ITEM" in r.json()["detail"]
