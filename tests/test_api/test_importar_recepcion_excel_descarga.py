"""Tests de la descarga del Excel-respuesta de una importación de
recepción (que quede descargable desde la app, no solo por correo)."""
from __future__ import annotations

from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base, get_db
from app.models.db import (
    GlosaRecord,
    ImportacionRecepcionRecord,
    UsuarioRecord,
)


@pytest.fixture
def db_session():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    s = sessionmaker(bind=engine)()
    try:
        yield s
    finally:
        s.close()
        engine.dispose()


def _client(db_session):
    from app.api.deps import get_usuario_actual
    from app.main import app

    u = UsuarioRecord(
        id=1, email="x@hus.com", nombre="Test", rol="AUDITOR", activo=1,
        password_hash="x",
    )
    app.dependency_overrides[get_db] = lambda: iter([db_session]).__next__()
    app.dependency_overrides[get_usuario_actual] = lambda: u
    c = TestClient(app)
    return c, app


def _excel_recepcion_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["GESTOR", "FACTURA", "CONSECUTIVO DGH", "VALOR GLOSA"])
    ws.append(["JUAN PEREZ", "FAC-001", "C-001", 12345])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def test_descarga_regenera_excel_con_columnas_ia(db_session, tmp_path):
    ruta = tmp_path / "1.xlsx"
    ruta.write_bytes(_excel_recepcion_bytes())

    g = GlosaRecord(
        id=10, eps="SURA EPS", paciente="N/A",
        factura="FAC-001", consecutivo_dgh="C-001",
        estado="RESPONDIDA", dictamen="Dictamen IA de prueba",
        modelo_ia="test",
    )
    db_session.add(g)
    rec = ImportacionRecepcionRecord(
        id=1, usuario_email="x@hus.com", archivo_nombre="recepcion.xlsx",
        total_glosas=1, glosa_ids="[10]", estado="LISTO",
        ruta_original=str(ruta),
    )
    db_session.add(rec)
    db_session.commit()

    c, app = _client(db_session)
    try:
        r = c.get("/glosas/importar-recepcion/1/excel-respuesta")
        assert r.status_code == 200, r.text
        assert "spreadsheetml" in r.headers["content-type"]

        wb = load_workbook(BytesIO(r.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        assert "RESPUESTA IA" in headers
        assert "ESTADO IA" in headers
        # La fila debe traer el dictamen persistido de la glosa.
        valores = [c.value for row in ws.iter_rows() for c in row]
        assert any("Dictamen IA de prueba" == v for v in valores)
    finally:
        app.dependency_overrides.clear()


def test_historial_lista_la_importacion(db_session):
    db_session.add(ImportacionRecepcionRecord(
        id=1, usuario_email="x@hus.com", archivo_nombre="recepcion.xlsx",
        total_glosas=3, glosa_ids="[1,2,3]", estado="LISTO",
        ruta_original="/data/recepcion/1.xlsx",
    ))
    db_session.commit()

    c, app = _client(db_session)
    try:
        r = c.get("/glosas/importar-recepcion/historial")
        assert r.status_code == 200, r.text
        imps = r.json()["importaciones"]
        assert len(imps) == 1
        assert imps[0]["id"] == 1
        assert imps[0]["total_glosas"] == 3
        assert imps[0]["descargable"] is True
    finally:
        app.dependency_overrides.clear()


def test_descarga_inexistente_404(db_session):
    c, app = _client(db_session)
    try:
        r = c.get("/glosas/importar-recepcion/999/excel-respuesta")
        assert r.status_code == 404
    finally:
        app.dependency_overrides.clear()


def test_descarga_sin_archivo_410(db_session):
    db_session.add(ImportacionRecepcionRecord(
        id=2, usuario_email="x@hus.com", archivo_nombre="r.xlsx",
        total_glosas=1, glosa_ids="[1]", estado="SIN_ARCHIVO",
        ruta_original=None,
    ))
    db_session.commit()
    c, app = _client(db_session)
    try:
        r = c.get("/glosas/importar-recepcion/2/excel-respuesta")
        assert r.status_code == 410
    finally:
        app.dependency_overrides.clear()
