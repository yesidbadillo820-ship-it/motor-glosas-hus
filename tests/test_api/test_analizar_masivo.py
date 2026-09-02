"""Procesamiento masivo: un CSV con varias glosas (+ ZIP de PDF) → un Excel
consolidado con todos los dictámenes (V2, Pilar 1, 02-09-2026).

Cada glosa del lote corre por la MISMA `_analizar_impl` del análisis single
(por eso no hay retroceso), unas cuantas a la vez, en el fondo. El endpoint
devuelve el job_id enseguida; el avance se consulta y el archivo se descarga.

Nota de las pruebas: la BD en memoria (StaticPool) es UNA sola conexión, que
SQLite no deja usar por varios hilos a la vez. En producción cada hilo abre su
propia sesión (SessionLocal), así que la concurrencia real sí funciona; aquí se
corre con concurrencia=1 —que igual pasa por el ThreadPoolExecutor— y siempre se
espera a que el lote termine antes de cerrar el test.
"""

from __future__ import annotations

import io
import time
import zipfile

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base, get_db
from app.models.db import UsuarioRecord
from app.models.schemas import GlosaResult


# ─────────────────────────── unidad: los ayudantes puros ───────────────────────────
class TestAyudantesDelLote:
    def test_parsea_csv_con_encabezados_tolerantes(self):
        from app.api.routers.analizar import _parsear_csv_glosas

        csv = b"EPS;Numero Factura;Texto;Valor Aceptado\nFAMISANAR;HUS1;TA0201 diferencia;0\n"
        filas = _parsear_csv_glosas(csv)
        assert len(filas) == 1
        assert filas[0]["eps"] == "FAMISANAR"
        assert filas[0]["numero_factura"] == "HUS1"
        assert filas[0]["tabla_excel"] == "TA0201 diferencia"

    def test_extraer_zip_indexa_por_basename(self):
        from app.api.routers.analizar import _extraer_zip

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("carpeta/HUS1.pdf", b"%PDF x")
        mapa = _extraer_zip(buf.getvalue())
        assert "hus1.pdf" in mapa

    def test_pdfs_de_fila_por_columna_y_por_factura(self):
        from app.api.routers.analizar import _pdfs_de_fila

        zip_map = {"hus1.pdf": b"a", "hus2_epicrisis.pdf": b"b"}
        # Por columna pdfs explícita
        f1 = _pdfs_de_fila({"pdfs": "HUS1.pdf"}, zip_map)
        assert [n for n, _, _ in f1] == ["HUS1.pdf"]
        # Sin columna: por coincidencia de factura en el nombre
        f2 = _pdfs_de_fila({"numero_factura": "HUS2"}, zip_map)
        assert [n for n, _, _ in f2] == ["hus2_epicrisis.pdf"]

    def test_texto_plano_quita_html(self):
        from app.api.routers.analizar import _texto_plano

        assert "<" not in _texto_plano("<div>hola<br/>mundo</div>")

    def test_construir_xlsx_una_hoja_con_encabezado(self):
        from app.api.routers.analizar import _construir_xlsx_masivo

        b = _construir_xlsx_masivo([{"numero_factura": "HUS1", "estado": "ok", "dictamen": "x"}])
        wb = load_workbook(io.BytesIO(b))
        assert wb.sheetnames == ["DICTAMENES"]
        filas = list(wb.active.iter_rows(values_only=True))
        assert filas[0][0] == "NUMERO FACTURA" and filas[1][0] == "HUS1"


# ─────────────────────────── integración end-to-end ───────────────────────────
@pytest.fixture
def engine_memoria():
    eng = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(eng)
    yield eng
    eng.dispose()


@pytest.fixture
def usuario_fake():
    return UsuarioRecord(id=1, email="a@hus.gov.co", nombre="Auditor", rol="AUDITOR", activo=1)


def _resultado_mock() -> GlosaResult:
    return GlosaResult(
        tipo="RESPUESTA RE9901",
        resumen="Defensa técnica",
        dictamen="<div>Dictamen del lote</div>",
        codigo_glosa="TA0201",
        valor_objetado="$ 100.000",
        paciente="N/A",
        mensaje_tiempo="EN TÉRMINOS",
        color_tiempo="green",
        score=88.0,
        dias_restantes=10,
        modelo_ia="mock/test",
        etapa_procesal="INICIAL",
    )


@pytest.fixture
def client(engine_memoria, usuario_fake):
    from unittest.mock import AsyncMock, MagicMock

    from app.api.deps import get_auditor_o_superior, get_usuario_actual
    from app.api.routers.analizar import get_glosa_service
    from app.main import app
    from app.services import resultados_masivo as rm

    Sesion = sessionmaker(bind=engine_memoria)

    def _get_db_fresco():
        db = Sesion()
        try:
            yield db
        finally:
            db.close()

    svc = MagicMock()
    svc.analizar = AsyncMock(return_value=_resultado_mock())

    rm._vaciar_para_pruebas()
    app.dependency_overrides[get_db] = _get_db_fresco
    app.dependency_overrides[get_usuario_actual] = lambda: usuario_fake
    app.dependency_overrides[get_auditor_o_superior] = lambda: usuario_fake
    app.dependency_overrides[get_glosa_service] = lambda: svc
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.clear()
    rm._vaciar_para_pruebas()


def _csv(filas: list[dict]) -> bytes:
    import csv as _csv

    buf = io.StringIO()
    w = _csv.DictWriter(buf, fieldnames=["eps", "etapa", "factura", "texto"])
    w.writeheader()
    for f in filas:
        w.writerow(f)
    return buf.getvalue().encode("utf-8")


def _esperar_listo(client, job_id, segundos=20.0) -> dict:
    fin = time.time() + segundos
    ultimo: dict = {}
    while time.time() < fin:
        r = client.get(f"/analizar/masivo/{job_id}")
        assert r.status_code == 200, r.text
        ultimo = r.json()
        if ultimo.get("estado") != "en_curso":
            return ultimo
        time.sleep(0.1)
    return ultimo


FILAS = [
    {
        "eps": "FAMISANAR",
        "etapa": "RESPUESTA A GLOSA",
        "factura": "HUS0000601111",
        "texto": "TA0201 diferencia tarifa CUPS 890750 valor objetado $100.000.",
    },
    {
        "eps": "COOSALUD",
        "etapa": "RESPUESTA A GLOSA",
        "factura": "HUS0000602222",
        "texto": "SO0101 falta epicrisis, valor objetado $50.000.",
    },
]


class TestElLoteSeProcesaYConsolida:
    def test_el_lote_termina_y_el_excel_trae_las_dos_glosas(self, client):
        job = client.post(
            "/analizar/masivo",
            files={"csv": ("glosas.csv", _csv(FILAS), "text/csv")},
            data={"concurrencia": "1"},
        ).json()["job_id"]

        estado = _esperar_listo(client, job)
        assert estado["estado"] == "listo", estado
        assert estado["total"] == 2 and estado["ok"] == 2 and estado["error"] == 0

        r = client.get(f"/analizar/masivo/{job}/resultado")
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers["content-type"]
        ws = load_workbook(io.BytesIO(r.content)).active
        filas = list(ws.iter_rows(values_only=True))
        assert filas[0][0] == "NUMERO FACTURA"
        assert len(filas) == 1 + 2
        assert {f[0] for f in filas[1:]} == {"HUS0000601111", "HUS0000602222"}
        assert all("Dictamen del lote" in (f[7] or "") for f in filas[1:])
        assert all("<div>" not in (f[7] or "") for f in filas[1:])

    def test_una_fila_sin_texto_es_error_y_no_tumba_el_lote(self, client):
        filas = FILAS + [
            {"eps": "SURA", "etapa": "RESPUESTA A GLOSA", "factura": "HUS0000603333", "texto": ""}
        ]
        job = client.post(
            "/analizar/masivo",
            files={"csv": ("g.csv", _csv(filas), "text/csv")},
            data={"concurrencia": "1"},
        ).json()["job_id"]
        estado = _esperar_listo(client, job)
        assert estado["estado"] == "listo"
        assert estado["ok"] == 2 and estado["error"] == 1


class TestBordes:
    def test_csv_vacio_es_400(self, client):
        r = client.post(
            "/analizar/masivo",
            files={"csv": ("v.csv", b"eps,etapa,factura,texto\n", "text/csv")},
            data={},
        )
        assert r.status_code == 400

    def test_lote_desconocido_es_404(self, client):
        assert client.get("/analizar/masivo/no-existe").status_code == 404
        assert client.get("/analizar/masivo/no-existe/resultado").status_code == 404

    def test_descargar_antes_de_terminar_es_409(self, client):
        from app.services import resultados_masivo as rm

        rm.abrir("job-x", 3)
        assert client.get("/analizar/masivo/job-x/resultado").status_code == 409
