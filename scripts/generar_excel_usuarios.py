"""Genera el Excel de usuarios + guía de uso IA para entrega al equipo.

Ejecutar:
    python scripts/generar_excel_usuarios.py

Genera: docs/usuarios_motor_glosas.xlsx con 3 hojas:
  1. Credenciales de Acceso (29 usuarios)
  2. Guía de Ingreso y Uso
  3. Recomendaciones para no saturar la IA
"""
import sys
from pathlib import Path

PROY = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROY))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Lista canónica de 29 usuarios (actualizada 22/04/2026).
# Password inicial de cada usuario = prefijo del email (parte antes del @).
USUARIOS = [
    ("Auditor Principal",               "admin@hus.gov.co",                     "SUPER_ADMIN"),
    ("YESID PEREZ",                     "glosashus09@sinacsc.com",              "SUPER_ADMIN"),
    ("YENFERSON ORTEGA",                "coordinadorglosashus01@sinacsc.com",   "COORDINADOR"),
    ("DIANEYDA QUINTERO",               "glosashus11@sinacsc.com",              "AUDITOR"),
    ("RUBY MILENA",                     "carterahus04@sinacsc.com",             "AUDITOR"),
    ("CAROLINA CIFUENTES",              "glosashus02@sinacsc.com",              "AUDITOR"),
    ("JHON JAIMES",                     "glosashus04@sinacsc.com",              "AUDITOR"),
    ("MARICELA ROJAS",                  "glosashus05@sinacsc.com",              "AUDITOR"),
    ("IRMA RIOS",                       "carterahus01@sinacsc.com",             "AUDITOR"),
    ("A_A_A_A (EQUIPO ASEGURADORAS)",   "glosashus12@sinacsc.com",              "AUDITOR"),
    ("A_A_A_A (EQUIPO ASEGURADORAS)",   "devoluciones02@sinacsc.com",           "AUDITOR"),
    ("A_A_A_A (EQUIPO ASEGURADORAS)",   "glosashus10@sinacsc.com",              "AUDITOR"),
    ("A_A_A_A (EQUIPO ASEGURADORAS)",   "glosashus16@sinacsc.com",              "AUDITOR"),
    ("KAREN ORTIZ",                     "radicadevoluciones@sinacsc.com",       "AUDITOR"),
    ("YUDY AMAYA",                      "coordinacioncartera@hus.gov.co",       "AUDITOR"),
    ("CLAUDIA SUAREZ",                  "glosashus08@sinacsc.com",              "AUDITOR"),
    ("SOFIA ORTEGA",                    "glosashus07@sinacsc.com",              "AUDITOR"),
    ("PATRICIA QUIÑONES",               "carterahus05@sinacsc.com",             "AUDITOR"),
    ("LAURA DIAZ",                      "auditorhus01@sinacsc.com",             "AUDITOR"),
    ("LEYDI ZULAY GONZALEZ",            "auditorhus03@sinacsc.com",             "AUDITOR"),
    ("LEIDY JHOANA SANGUINO",           "auditorhus02@sinacsc.com",             "AUDITOR"),
    ("JOHANNA MORENO",                  "devoluciones03@sinacsc.com",           "AUDITOR"),
    ("EDGAR SILVA",                     "devoluciones1@sinacsc.com",            "AUDITOR"),
    ("OSCAR VILLAMIZAR",                "glosashus03@sinacsc.com",              "AUDITOR"),
    ("SEBASTIAN SANCHES",               "devoluciones01@sinacsc.com",           "AUDITOR"),
    ("DANIEL FONCE",                    "glosashus01@sinacsc.com",              "AUDITOR"),
    ("ELIAS CARVAJAL",                  "glosashus15@sinacsc.com",              "AUDITOR"),
    ("IVAN ARCINIEGAS",                 "glosashus13@sinacsc.com",              "AUDITOR"),
    ("CAMILO CASTILLO",                 "glosashus14@sinacsc.com",              "AUDITOR"),
]


# Colores corporativos SINAC SC SAS
BRAND_NAVY = "0B1220"
BRAND_TEAL = "0369A1"
BRAND_GREEN = "10B981"
BRAND_AMBER = "F59E0B"


def _border_thin():
    return Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )


def _hoja_titulo(ws, texto: str, subtitulo: str = ""):
    """Pinta un encabezado institucional en la hoja."""
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "IA GLOSAS SINAC SC · ESE Hospital Universitario de Santander"
    c.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=BRAND_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = texto
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=BRAND_TEAL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    if subtitulo:
        ws.merge_cells("A3:F3")
        c = ws["A3"]
        c.value = subtitulo
        c.font = Font(name="Calibri", size=9, italic=True, color="475569")
        c.fill = PatternFill("solid", fgColor="F0F9FF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[3].height = 38


def crear_hoja_credenciales(wb):
    ws = wb.active
    ws.title = "Credenciales"

    _hoja_titulo(
        ws,
        "Credenciales de Acceso · 29 usuarios",
        "La CONTRASEÑA INICIAL es el PREFIJO del correo (parte antes del @). "
        "Ej.: glosashus04@sinacsc.com → contraseña: glosashus04 · "
        "Cambie su contraseña en el primer ingreso."
    )

    headers = ["#", "Nombre completo", "Correo institucional", "Contraseña inicial", "Rol"]
    header_fill = PatternFill("solid", fgColor=BRAND_TEAL)
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    border = _border_thin()

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[5].height = 26

    rol_colors = {
        "SUPER_ADMIN": ("FEF3C7", "92400E"),
        "COORDINADOR": ("E0E7FF", "3730A3"),
        "AUDITOR":     ("DBEAFE", "1E40AF"),
    }
    fill_alt = PatternFill("solid", fgColor="F8FAFC")

    for i, (nombre, email, rol) in enumerate(USUARIOS, start=1):
        row = 5 + i
        password = email.split("@")[0]
        rol_bg, rol_fg = rol_colors.get(rol, ("E2E8F0", "334155"))

        c = ws.cell(row=row, column=1, value=i)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(name="Calibri", size=10, bold=True, color="64748B")

        c = ws.cell(row=row, column=2, value=nombre)
        c.font = Font(name="Calibri", size=10, color="0F172A")
        c.alignment = Alignment(vertical="center", indent=1)

        c = ws.cell(row=row, column=3, value=email)
        c.font = Font(name="Calibri", size=10, color="334155")
        c.alignment = Alignment(vertical="center", indent=1)

        c = ws.cell(row=row, column=4, value=password)
        c.font = Font(name="Consolas", size=10, bold=True, color=BRAND_TEAL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", fgColor="EFF6FF")

        c = ws.cell(row=row, column=5, value=rol)
        c.font = Font(name="Calibri", size=9, bold=True, color=rol_fg)
        c.fill = PatternFill("solid", fgColor=rol_bg)
        c.alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border

        if i % 2 == 0:
            for col in [1, 2, 3, 5]:
                cur = ws.cell(row=row, column=col)
                if not cur.fill or cur.fill.fgColor.rgb == "00000000":
                    cur.fill = fill_alt
        ws.row_dimensions[row].height = 22

    # Anchos
    for col, w in {1: 5, 2: 36, 3: 36, 4: 30, 5: 14}.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    last_row = 5 + len(USUARIOS) + 2
    ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=5)
    c = ws.cell(row=last_row, column=1)
    c.value = (
        f"Total usuarios: {len(USUARIOS)}   ·   "
        "URL: https://motor-glosas-hus.onrender.com   ·   "
        "Soporte: cartera@hus.gov.co"
    )
    c.font = Font(name="Calibri", size=9, italic=True, color="64748B")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor="F8FAFC")
    ws.row_dimensions[last_row].height = 22

    nota = last_row + 2
    ws.merge_cells(start_row=nota, start_column=1, end_row=nota, end_column=5)
    c = ws.cell(row=nota, column=1)
    c.value = (
        "🔒 SEGURIDAD: (1) La contraseña inicial DEBE cambiarse en el primer login.  "
        "(2) Activar 2FA desde el menú de usuario si maneja datos sensibles.  "
        "(3) NO compartir credenciales — cada acción queda en el registro de auditoría individual."
    )
    c.font = Font(name="Calibri", size=9, color="991B1B")
    c.fill = PatternFill("solid", fgColor="FEF2F2")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ws.row_dimensions[nota].height = 38

    ws.freeze_panes = "A6"


def crear_hoja_guia(wb):
    """Guía paso a paso de ingreso y uso de la IA."""
    ws = wb.create_sheet("Guía de Uso")
    _hoja_titulo(
        ws,
        "Guía de Ingreso y Uso de la IA",
        "Lea esta guía antes del primer uso. Tiempo estimado: 5 minutos."
    )

    secciones = [
        ("🔐 1. INGRESAR AL SISTEMA", [
            ("URL", "https://motor-glosas-hus.onrender.com"),
            ("Correo", "El que aparece en la hoja 'Credenciales'."),
            ("Contraseña", "Prefijo del correo (ej. glosashus04@sinacsc.com → glosashus04)."),
            ("Primer login", "El sistema le pide cambiar la contraseña. Use una segura (8+ caracteres, números, símbolos)."),
            ("Problemas", "Si ve pantalla en blanco, visite /reset-sw.html y vuelva a entrar."),
        ]),
        ("⚖️ 2. ANALIZAR UNA GLOSA INDIVIDUAL", [
            ("Paso 1", "Panel 'Analizar glosa' (menú lateral izquierdo)."),
            ("Paso 2", "Seleccionar EPS / Pagador del desplegable. Si no aparece, elija 'OTRA / SIN DEFINIR'."),
            ("Paso 3", "Pegar el NÚMERO DE FACTURA. El sistema buscará conceptos ya cargados."),
            ("Paso 4", "Si aparece la caja azul con los conceptos, haga click en el que desea responder. El texto se autocompleta."),
            ("Paso 5", "Elegir etapa (RESPUESTA A GLOSA / RATIFICADA / ACLARACIÓN)."),
            ("Paso 6", "Tono: Conciliador para Salud Total/EPS confiables; Firme para cuando hay que exigir pago."),
            ("Paso 7", "Subir soportes PDF (opcional, hasta 10 archivos de 15MB)."),
            ("Paso 8", "Click en ANALIZAR. La IA demora 10-30 segundos."),
        ]),
        ("📥 3. IMPORTAR RECEPCIÓN (Excel 4 hojas)", [
            ("Dónde", "Menú lateral → 'Importar recepción'."),
            ("Formato", "Excel (.xlsx) con 4 hojas: INICIAL, RATIFICADA, I, R."),
            ("Qué hace", "Crea las glosas automáticamente, las asigna a cada gestor, carga todos los conceptos."),
            ("Después", "Cada gestor las ve en su bandeja 'Mis glosas'. Los conceptos aparecen al buscar la factura en Analizar."),
        ]),
        ("📋 4. IMPORTACIÓN MASIVA (copy-paste Excel)", [
            ("Dónde", "Menú lateral → 'Importación masiva'."),
            ("Cuándo usar", "Cuando tiene varias glosas en Excel y quiere procesar de golpe."),
            ("Separador", "Tab o pipe (|) — auto-detectado. Pegue directo desde Excel."),
            ("Columnas", "ENTIDAD | FACTURA | VALOR | CÓDIGO | CONCEPTO | CUPS | SERVICIO | MOTIVO."),
        ]),
        ("💚 5. SALUD TOTAL (archivo TXT)", [
            ("Dónde", "Menú lateral → 'Salud Total'."),
            ("Archivo", "TXT separado por TAB o pipe con 24 columnas."),
            ("Tipo respuesta", "Extemporánea (fuera de términos) / Ratificada / IA (análisis completo)."),
            ("Resultado", "Descarga un TXT formateado listo para radicar en el portal de Salud Total EPS."),
        ]),
        ("🔍 6. ATAJOS ÚTILES", [
            ("Cmd + K (Ctrl + K)", "Spotlight: busca glosas, usuarios, EPS, contratos o abre cualquier sección."),
            ("Mis glosas", "Sus glosas asignadas con semáforo de vencimiento (verde/amarillo/rojo/negro)."),
            ("Historial", "Todas las glosas procesadas (filtro por EPS, fecha, código)."),
            ("Alertas", "Glosas próximas a vencer (según días hábiles desde recepción)."),
            ("Dashboard", "KPIs del mes y tendencias de 6 meses (solo COORDINADOR/SUPER_ADMIN)."),
        ]),
        ("📄 7. DESCARGAR PDF DEL DICTAMEN", [
            ("Después de analizar", "Botón 'Imprimir PDF' en el resultado."),
            ("Contenido", "Cabecera institucional, tabla de códigos, argumento jurídico, mensaje de tiempo, firmas, pie de página SINAC."),
            ("Para radicar", "Firme manualmente el PDF impreso y adjunte al radicado en la EPS."),
        ]),
    ]

    r = 5
    for titulo, items in secciones:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1, value=titulo)
        c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BRAND_TEAL)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 24
        r += 1

        for etiqueta, detalle in items:
            c = ws.cell(row=r, column=1, value=etiqueta)
            c.font = Font(name="Calibri", size=10, bold=True, color="0F172A")
            c.alignment = Alignment(vertical="center", indent=1, wrap_text=True)
            c.fill = PatternFill("solid", fgColor="F8FAFC")
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            c = ws.cell(row=r, column=2, value=detalle)
            c.font = Font(name="Calibri", size=10, color="334155")
            c.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
            ws.row_dimensions[r].height = max(22, 18 + (len(detalle) // 70) * 16)
            r += 1
        r += 1

    ws.column_dimensions["A"].width = 22
    for col in "BCDEF":
        ws.column_dimensions[col].width = 18
    ws.freeze_panes = "A5"


def crear_hoja_recomendaciones(wb):
    """Recomendaciones para el equipo — cómo no saturar la IA."""
    ws = wb.create_sheet("Buenas prácticas IA")
    _hoja_titulo(
        ws,
        "Recomendaciones para el Equipo · Uso eficiente de la IA",
        "Seguir estas prácticas evita costos altos, respuestas genéricas y spam en el sistema."
    )

    bloques = [
        ("✅ QUÉ SÍ HACER", BRAND_GREEN, [
            "Usar el autocompletado de factura: al pegar el número, el sistema trae los conceptos ya cargados — evita redigitar.",
            "Seleccionar el concepto específico antes de analizar (no dejar el texto genérico 'glosa objetada por la entidad pagadora').",
            "Elegir el tono correcto: Firme solo cuando la EPS ya rechazó una vez; Conciliador en primera vuelta.",
            "Revisar el dictamen antes de radicar — la IA puede equivocarse en detalles; usted firma.",
            "Subir TODOS los soportes PDF relevantes (historia clínica, autorización, RIPS) al analizar. La IA los usa como contexto.",
            "Usar 'Refinar' en el modal cuando quiera ajustar tono o cita de una norma específica — no re-analizar desde cero.",
            "Guardar argumentos ganadores como 'Plantilla Gold' (⭐): la IA los usa de ejemplo en casos similares futuros.",
            "Agrupar glosas por factura y usar 'Analizar todos' en vez de ir una por una — más rápido y consistente.",
        ]),
        ("❌ QUÉ NO HACER", "DC2626", [
            "NO analizar la misma glosa múltiples veces seguidas para 'probar' respuestas. Cada clic cuesta tokens y queda en auditoría.",
            "NO pegar textos inútiles (ej. 'hola', '123', 'test', solo el código glosa TA0801 sin contexto). La IA genera respuesta mala.",
            "NO subir PDFs de 15MB si solo necesita 2-3 páginas. Comprímalos antes (ILovePDF, Smallpdf).",
            "NO usar el sistema desde 5 pestañas a la vez — se pisan las sesiones.",
            "NO spamear el botón Analizar si tarda — la primera petición ya está en curso. Espere la respuesta.",
            "NO modificar manualmente el dictamen para borrar citas legales. La IA calibra normativa; si quiere otra cita, use 'Refinar'.",
            "NO compartir credenciales con compañeros. Cada acción queda registrada al dueño del correo.",
            "NO usar glosas ajenas para 'aprender'. Cada auditor ve las suyas; el COORDINADOR ve todo.",
        ]),
        ("📈 PARA RESPUESTAS DE MEJOR CALIDAD", BRAND_TEAL, [
            "Dé contexto: Número de factura + Código glosa + Servicio + Valor. La IA rinde 3× mejor con datos completos.",
            "Si la EPS escribió una OBSERVACIÓN específica (ej. 'no concuerda cantidad facturada'), péguela completa. Ahí está la clave.",
            "Para glosas ratificadas use el tono Firme + cite la Ley 1438 Art. 57 y exija mesa de conciliación.",
            "Para glosas extemporáneas, la IA detecta automáticamente si superan 20 días hábiles y responde con plantilla RE9502.",
            "Si el dictamen cita una norma derogada o incorrecta (raro), use 'Validar' ✅ en el modal — el sistema detecta automáticamente.",
        ]),
        ("⚡ USO RESPONSABLE DE TOKENS IA", BRAND_AMBER, [
            "Cada análisis = ~2000-4000 tokens = costo real. Un uso responsable beneficia a todo el equipo.",
            "Lote grande (50+ conceptos)? Usar 'Analizar todos' con pausa automática entre llamadas (rate-limit interno).",
            "Antes de analizar, revise si la glosa YA tiene dictamen en Historial. No duplicar trabajo.",
            "Si necesita re-generar, use 'Regenerar' — mantiene contexto; NO borrar y analizar de nuevo.",
            "Horarios de menor carga (8–11 am) responden más rápido que horarios punta.",
        ]),
        ("🆘 SI ALGO NO FUNCIONA", "3730A3", [
            "Pantalla en blanco al entrar → visite /reset-sw.html y recargue con Ctrl+Shift+R.",
            "Mensaje 500 o 'Error interno' → intente de nuevo en 1-2 min; si persiste, reporte al COORDINADOR.",
            "Import masiva falla → verifique que el archivo Excel tenga las 4 hojas con los encabezados correctos.",
            "Contraseña olvidada → el SUPER_ADMIN (YESID o Auditor Principal) puede resetearla en el panel Usuarios.",
            "Sugerencias de mejora → anotarlas y enviarlas al COORDINADOR (YENFERSON ORTEGA).",
        ]),
    ]

    r = 5
    for titulo, color, items in bloques:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1, value=titulo)
        c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 26
        r += 1

        for idx, item in enumerate(items, start=1):
            c = ws.cell(row=r, column=1, value=str(idx))
            c.font = Font(name="Calibri", size=10, bold=True, color=color)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = PatternFill("solid", fgColor="F8FAFC")
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            c = ws.cell(row=r, column=2, value=item)
            c.font = Font(name="Calibri", size=10, color="334155")
            c.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
            ws.row_dimensions[r].height = max(22, 18 + (len(item) // 85) * 14)
            r += 1
        r += 1

    ws.column_dimensions["A"].width = 6
    for col in "BCDEF":
        ws.column_dimensions[col].width = 22
    ws.freeze_panes = "A5"


def main() -> None:
    wb = Workbook()
    crear_hoja_credenciales(wb)
    crear_hoja_guia(wb)
    crear_hoja_recomendaciones(wb)

    output = PROY / "docs" / "usuarios_motor_glosas.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"OK Excel generado: {output}")
    print(f"  Hoja 1: Credenciales ({len(USUARIOS)} usuarios)")
    print(f"  Hoja 2: Guia de Uso")
    print(f"  Hoja 3: Buenas practicas IA")


if __name__ == "__main__":
    main()
