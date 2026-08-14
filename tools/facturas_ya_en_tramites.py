"""facturas_ya_en_tramites.py — Arma la lista de facturas cuyo TRÁMITE ya se envió.

Recorre una carpeta con los archivos "MASIVO COOSALUD ..." que ya se le
mandaron a sistemas y escribe un TXT con todas las facturas que traen (una
por línea). Ese TXT se le pasa después a respuesta_tramites_dgh.py con
--omitir-facturas, para que el próximo masivo NO repita facturas.

Repetir una factura en el cargue de trámites hace que DGH la rechace, así que
este paso evita el retrabajo.

USO:
    py facturas_ya_en_tramites.py "D:\\...\\MASIVOS ENVIADOS"
    py facturas_ya_en_tramites.py "D:\\...\\MASIVOS ENVIADOS" --salida lista.txt

También acepta varias carpetas:
    py facturas_ya_en_tramites.py "D:\\carpeta1" "D:\\carpeta2"
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.stderr.write("Falta openpyxl. Corre: py -m pip install openpyxl\n")
    sys.exit(2)

COL_FACTURA = "FacturaCartera.Factura"
COL_ENTIDAD = "FacturaCartera.PlanBeneficio.Contrato.Entidad.NombreEntidad"
# Columna de respuesta que solo tienen los masivos YA diligenciados (ojo al
# espacio final: así viene en el formato real de DGH).
COLS_RESPUESTA = ("CODIGO RESPUESTA ", "CODIGO RESPUESTA")
PATRON = "*.xlsx"


def norm_factura(v: object) -> str | None:
    """HUS0000522511 / 0000522511 / 522511 -> 522511 (solo el número)."""
    d = re.sub(r"\D", "", str(v or "")).lstrip("0")
    return d or None


def facturas_de_archivo(path: Path, entidad: str | None) -> tuple[set[str], str]:
    """Facturas de un MASIVO de trámites YA diligenciado.

    Devuelve (facturas, motivo_si_se_descarta). Un export crudo de DGH, la
    plantilla base o un archivo de otra EPS NO cuentan como enviados:
      - debe traer la columna de CODIGO RESPUESTA **con datos** (si está vacía,
        el archivo todavía no se ha respondido, luego no se ha enviado);
      - si se pide --entidad, la entidad de la factura debe coincidir (evita
        mezclar Dispensario/SIMED con COOSALUD).
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # xlsx corrupto o protegido
        return set(), f"ilegible ({exc})"
    try:
        ws = wb[wb.sheetnames[0]]
        filas = ws.iter_rows(values_only=True)
        headers = list(next(filas, []) or [])
        if COL_FACTURA not in headers:
            return set(), "no es un archivo de trámites"
        idx = headers.index(COL_FACTURA)

        idx_rta = None
        for col in COLS_RESPUESTA:
            if col in headers:
                idx_rta = headers.index(col)
                break
        if idx_rta is None:
            return set(), "sin columna CODIGO RESPUESTA (export crudo / plantilla)"

        idx_ent = headers.index(COL_ENTIDAD) if COL_ENTIDAD in headers else None
        objetivo = entidad.upper() if entidad else None

        facturas: set[str] = set()
        con_respuesta = 0
        de_otra_entidad = 0
        for row in filas:
            if idx_rta >= len(row) or not str(row[idx_rta] or "").strip():
                continue  # fila sin respuesta diligenciada
            con_respuesta += 1
            if objetivo and idx_ent is not None:
                ent = str(row[idx_ent] or "").upper()
                if ent and objetivo not in ent:
                    de_otra_entidad += 1
                    continue
            if idx < len(row):
                f = norm_factura(row[idx])
                if f:
                    facturas.add(f)
        if not con_respuesta:
            return set(), "CODIGO RESPUESTA vacío (aún no se ha enviado)"
        if not facturas and de_otra_entidad:
            return set(), f"es de otra entidad (no {entidad})"
        return facturas, ""
    finally:
        wb.close()


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(
        description="Lista las facturas cuyos trámites ya se enviaron a sistemas."
    )
    p.add_argument("carpetas", nargs="+", help="Carpeta(s) con los MASIVO ya enviados.")
    p.add_argument(
        "--salida",
        default="FACTURAS YA EN TRAMITES.txt",
        help='TXT de salida. Def: "FACTURAS YA EN TRAMITES.txt"',
    )
    p.add_argument(
        "--entidad",
        default="COOSALUD",
        help="Solo cuenta las facturas de esta EPS (para no mezclar Dispensario "
        'con COOSALUD). Def: COOSALUD. Usa "" para no filtrar.',
    )
    p.add_argument(
        "--ver-descartados",
        action="store_true",
        help="Muestra también los archivos que NO se contaron y por qué.",
    )
    args = p.parse_args(argv)
    entidad = args.entidad.strip() or None

    todas: set[str] = set()
    por_archivo: list[tuple[str, int]] = []
    descartados: list[tuple[str, str]] = []
    revisados = 0
    for carpeta in args.carpetas:
        base = Path(carpeta.strip().strip('"'))
        if not base.is_dir():
            print(f"OJO: no es una carpeta, se salta: {base}")
            continue
        print(f"\nRevisando: {base}")
        for arch in sorted(base.rglob(PATRON)):
            if arch.name.startswith("~$"):
                continue
            revisados += 1
            facturas, motivo = facturas_de_archivo(arch, entidad)
            if facturas:
                nuevas = facturas - todas
                todas |= facturas
                por_archivo.append((arch.name, len(facturas)))
                print(f"  {arch.name:60} {len(facturas):>5} facturas ({len(nuevas)} nuevas)")
            elif motivo:
                descartados.append((arch.name, motivo))

    if descartados and args.ver_descartados:
        print(f"\n  --- No contados ({len(descartados)}) ---")
        for nombre, motivo in descartados:
            print(f"  {nombre:60} {motivo}")

    if not por_archivo:
        print(
            f"\nNo encontré ningún MASIVO de trámites ya enviado en lo revisado "
            f"({revisados} archivo(s) mirados).\n"
            f"Un masivo enviado trae '{COL_FACTURA}' y 'CODIGO RESPUESTA' con datos."
        )
        return 1

    salida = Path(args.salida)
    salida.write_text("\n".join(f"HUS{f}" for f in sorted(todas, key=int)) + "\n", encoding="utf-8")

    print("\n===== LISTO =====")
    print(f"  Masivos ya enviados encontrados:  {len(por_archivo)}")
    if entidad:
        print(f"  Entidad contada:                  {entidad}")
    if descartados:
        print(f"  Archivos no contados:             {len(descartados)} (usa --ver-descartados)")
    print(f"  Facturas distintas ya enviadas:   {len(todas)}")
    print(f"  Lista guardada en: {salida.resolve()}")
    print(
        "\n  Pásasela al bot de trámites así:\n"
        f'    py respuesta_tramites_dgh.py --omitir-facturas "{salida.resolve()}" ...'
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
