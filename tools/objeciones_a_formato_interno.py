"""objeciones_a_formato_interno.py — Excel de OBJECIONES del portal → formato interno (CRO*).

Convierte el Excel de "Objeciones de glosa" que se descarga del portal de la EPS
(columnas: Número de factura | Tecnología | Cantidad facturada | Valor Facturado |
Valor glosado | Código de glosa | Observacion) al Excel que ingiere el sistema interno
de cartera del HUS (hoja "OBJECIONES", columnas CRO*):

    CDCONSEC | CDFECDOC | CRNCXC | CROFECOBJ | CROREFERE | CROOBSERV | CROCLAOBJ |
    CRNCLAOBJ | GENUSUARIO4 | CRNCONOBJ | SLNSERPRO | IDRIPS | CTNCENCOS | CROVALOBJ |
    CRDOBSERV | CROTIPOBJ

Mapeo (deducido del ejemplo real OBJECIONES_EMSSANAR_HUS0000515948):
    CRNCXC     = factura en formato largo 'HUS' + 10 dígitos (HUS520567 -> HUS0000520567)
    CRNCONOBJ  = Código de glosa            (TA0201, FA1305, …)
    SLNSERPRO  = Tecnología / servicio      (890426, …)
    CROVALOBJ  = Valor glosado (entero)     (111400)
    CRDOBSERV  = "<código> <observación>"   (texto de la glosa, prefijado con el código)
    CDFECDOC = CROFECOBJ = fecha (--fecha, por defecto hoy)
Constantes (idénticas a cada fila del ejemplo):
    CDCONSEC=1  CROCLAOBJ=0  CROTIPOBJ=0  GENUSUARIO4=<--usuario, def 999>
    CROREFERE, CROOBSERV, CRNCLAOBJ, IDRIPS, CTNCENCOS = vacías

USO
---
    py objeciones_a_formato_interno.py ^
        --entrada "D:\\...\\Objeciones_de_glosa_Factura_HUS520567.xlsx" ^
        --salida  "D:\\...\\OBJECIONES_HUS520567.xlsx" ^
        --fecha 2026-07-21

DEPENDENCIAS:  py -m pip install openpyxl
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

logger = logging.getLogger("objeciones_formato_interno")

# Columnas de salida, en el orden EXACTO del sistema interno.
COLUMNAS_SALIDA = [
    "CDCONSEC", "CDFECDOC", "CRNCXC", "CROFECOBJ", "CROREFERE", "CROOBSERV",
    "CROCLAOBJ", "CRNCLAOBJ", "GENUSUARIO4", "CRNCONOBJ", "SLNSERPRO", "IDRIPS",
    "CTNCENCOS", "CROVALOBJ", "CRDOBSERV", "CROTIPOBJ",
]  # fmt: skip

# Encabezados aceptados del Excel de entrada (tolerante a tildes/mayúsculas).
ENTRADA_ALIAS = {
    "factura": {"numero de factura", "factura", "no factura", "num factura"},
    "tecnologia": {"tecnologia", "servicio", "codigo tecnologia", "cups"},
    "valor": {"valor glosado", "valor objetado", "valor glosa"},
    "codigo": {"codigo de glosa", "codigo glosa", "cod glosa", "codigo"},
    "obs": {"observacion", "observaciones", "observacion glosa", "descripcion"},
}


def setup_logging() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")


def _norm(s) -> str:
    s = unicodedata.normalize("NFKD", str(s or "").strip().lower())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.split())


def _to_int(v) -> int:
    """'$ 111.400' / '1.172.000' (formato Colombia) -> entero. Corta en la coma decimal."""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).split(",")[0]
    return int(re.sub(r"[^\d]", "", s) or "0")


def factura_larga(f: str) -> str:
    """'HUS520567' -> 'HUS0000520567' (HUS + 10 dígitos con ceros a la izquierda)."""
    n = re.sub(r"\D", "", re.sub(r"(?i)^hus", "", str(f or "").strip()))
    return "HUS" + n.zfill(10) if n else str(f or "").strip().upper()


def leer_objeciones(ruta: Path) -> list[dict]:
    """Lee el Excel de objeciones del portal. Localiza la fila de encabezados por sus
    nombres (puede no ser la primera) y devuelve una lista de dicts por objeción."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.stderr.write("ERROR: falta openpyxl. Instalalo con: py -m pip install openpyxl\n")
        sys.exit(2)

    wb = load_workbook(ruta, data_only=True, read_only=True)
    filas = list(wb.active.iter_rows(values_only=True))
    wb.close()

    hdr_idx: int | None = None
    idx: dict[str, int] = {}
    for i, fila in enumerate(filas):
        norm = [_norm(c) for c in fila]
        cand: dict[str, int] = {}
        for clave, alias in ENTRADA_ALIAS.items():
            for j, h in enumerate(norm):
                if h in alias:
                    cand[clave] = j
                    break
        if "factura" in cand and "codigo" in cand:  # fila de encabezados
            hdr_idx, idx = i, cand
            break
    if hdr_idx is None:
        raise ValueError(
            "no encontré la fila de encabezados (esperaba 'Número de factura' y "
            "'Código de glosa'). ¿Es el Excel de objeciones del portal?"
        )

    objeciones: list[dict] = []
    for fila in filas[hdr_idx + 1 :]:
        fac = fila[idx["factura"]] if idx.get("factura") is not None else None
        cod = fila[idx["codigo"]] if idx.get("codigo") is not None else None
        if not fac or not cod:
            continue
        objeciones.append(
            {
                "factura": str(fac).strip(),
                "codigo": str(cod).strip(),
                "tecnologia": str(fila[idx["tecnologia"]]).strip()
                if idx.get("tecnologia") is not None and fila[idx["tecnologia"]]
                else "",
                "valor": _to_int(fila[idx["valor"]]) if idx.get("valor") is not None else 0,
                "obs": str(fila[idx["obs"]]).strip()
                if idx.get("obs") is not None and fila[idx["obs"]]
                else "",
            }
        )
    return objeciones


def escribir_formato_interno(
    objeciones: list[dict], salida: Path, fecha: datetime, usuario: str
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        sys.stderr.write("ERROR: falta openpyxl. Instalalo con: py -m pip install openpyxl\n")
        sys.exit(2)

    wb = Workbook()
    ws = wb.active
    ws.title = "OBJECIONES"

    for col, etiq in enumerate(COLUMNAS_SALIDA, start=1):
        ws.cell(row=1, column=col, value=etiq).font = Font(bold=True)

    for r, o in enumerate(objeciones, start=2):
        crdobserv = f"{o['codigo']} {o['obs']}".strip()
        valores = {
            "CDCONSEC": 1,
            "CDFECDOC": fecha,
            "CRNCXC": factura_larga(o["factura"]),
            "CROFECOBJ": fecha,
            "CROREFERE": None,
            "CROOBSERV": None,
            "CROCLAOBJ": 0,
            "CRNCLAOBJ": None,
            "GENUSUARIO4": usuario,
            "CRNCONOBJ": o["codigo"],
            "SLNSERPRO": o["tecnologia"],
            "IDRIPS": None,
            "CTNCENCOS": None,
            "CROVALOBJ": o["valor"],
            "CRDOBSERV": crdobserv,
            "CROTIPOBJ": 0,
        }
        for col, etiq in enumerate(COLUMNAS_SALIDA, start=1):
            c = ws.cell(row=r, column=col, value=valores[etiq])
            if etiq in ("CDFECDOC", "CROFECOBJ"):
                c.number_format = "yyyy-mm-dd hh:mm:ss"

    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(salida))


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Convierte el Excel de objeciones del portal al formato interno (CRO*).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--entrada", type=Path, required=True, help="Excel de objeciones del portal."
    )
    parser.add_argument("--salida", type=Path, required=True, help="Excel .xlsx de salida (CRO*).")
    parser.add_argument(
        "--fecha",
        type=str,
        default=None,
        help="Fecha del documento/objeción (YYYY-MM-DD). Por defecto: hoy.",
    )
    parser.add_argument(
        "--usuario", type=str, default="999", help="Valor de GENUSUARIO4 (por defecto 999)."
    )
    args = parser.parse_args()
    setup_logging()

    if not args.entrada.is_file():
        logger.error(f"No existe la entrada: {args.entrada}")
        return 1
    try:
        fecha = datetime.strptime(args.fecha, "%Y-%m-%d") if args.fecha else datetime.now()
    except ValueError:
        logger.error(f"--fecha inválida: {args.fecha!r} (usá YYYY-MM-DD).")
        return 1

    objeciones = leer_objeciones(args.entrada)
    if not objeciones:
        logger.error("No se leyó ninguna objeción del Excel de entrada.")
        return 1

    escribir_formato_interno(objeciones, args.salida, fecha, args.usuario)

    facturas = sorted({factura_larga(o["factura"]) for o in objeciones})
    total = sum(o["valor"] for o in objeciones)
    logger.info(f"Excel interno generado: {args.salida}")
    logger.info(f"  Objeciones: {len(objeciones)} | Facturas: {', '.join(facturas)}")
    logger.info(f"  Suma Valor glosado (objeciones): ${total:,}".replace(",", "."))
    logger.info(
        f"  Fecha usada (CDFECDOC/CROFECOBJ): {fecha:%Y-%m-%d}  ·  GENUSUARIO4={args.usuario}"
    )
    logger.info(
        "  Revisá: la fecha, y que CRDOBSERV (código + observación) sea como lo espera tu sistema."
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
