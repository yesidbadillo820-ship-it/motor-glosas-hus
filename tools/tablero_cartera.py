#!/usr/bin/env python3
"""Tablero de Radicación y Cartera — Cuentas Médicas y Cartera de la ESE HUS.

Es el paso AGUAS ABAJO del radicador (tools/radicar_facturacion.py): una vez
que las facturas están radicadas ante las entidades, este módulo lleva el
CONTROL de cartera del área:

  1. CONSOLIDA cuánto se ha radicado (cantidad y valor) por entidad y por mes,
     leyendo uno o varios reportes del radicador (CSV/XLSX).
  2. CRUZA ese radicado con el SEGUIMIENTO de cartera (una planilla Excel que el
     equipo alimenta a mano): valor glosado, devuelto, pagado, fechas y notas
     crédito — por número de factura.
  3. CALCULA los indicadores de cartera: radicado, glosado/devuelto, pagado,
     SALDO (lo que queda debiendo), % de recaudo, % de glosa y la MORA por
     tramos (aging 0-30 / 31-60 / 61-90 / +90 días).
  4. GENERA un tablero HTML profesional y autocontenido (se abre con doble clic,
     funciona sin internet) con KPIs, gráficos y una tabla filtrable por factura.

Seguro por defecto: NO toca las carpetas de soportes; solo LEE los reportes y la
planilla de seguimiento, y ESCRIBE el HTML (y, con --crear-plantilla, la
planilla Excel inicial pre-cargada con lo radicado).

Reutiliza el catálogo de entidades (data/perfiles_radicacion.json) y la
normalización de factura del radicador, así que el cruce radicado→pagado es
consistente con el resto del motor.
"""

from __future__ import annotations

import argparse
import csv
import logging
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

# Reutilizamos la normalización de factura y el catálogo de entidades del
# radicador (mismo tools/), para que el cruce por número de factura sea idéntico.
sys.path.insert(0, str(Path(__file__).resolve().parent))
try:
    from radicar_facturacion import _norm, cargar_perfiles, normalizar_factura

    _RAD_OK = True
except Exception:  # pragma: no cover - fallback si se ejecuta aislado
    _RAD_OK = False

    def _norm(s: str | None) -> str:
        return " ".join((s or "").strip().upper().split())

    def normalizar_factura(fac: str) -> str:
        f = (fac or "").strip().upper()
        for p in ("HUS",):
            if f.startswith(p):
                f = f[len(p) :]
        f = f.lstrip("0")
        return f or "0"

    def cargar_perfiles(_ruta):  # type: ignore[misc]
        return None


logger = logging.getLogger("tablero_cartera")

# Tramos de mora (aging) de la cartera, en días desde la radicación.
TRAMOS_MORA: tuple[tuple[str, int, int], ...] = (
    ("0-30", 0, 30),
    ("31-60", 31, 60),
    ("61-90", 61, 90),
    ("+90", 91, 10**9),
)

# Estados de cartera (ciclo de vida posterior a la radicación), de peor a mejor.
ESTADOS_CARTERA = (
    "PENDIENTE_RADICAR",
    "GLOSADA_DEVUELTA",
    "RADICADA_TRAMITE",
    "PAGO_PARCIAL",
    "PAGADA",
)

# Encabezados aceptados en la planilla de seguimiento (tolerante a variantes).
_ALIAS_COLS: dict[str, set[str]] = {
    "factura": {"FACTURA", "FACTURA HUS", "NUMERO FACTURA", "NRO FACTURA", "NO FACTURA"},
    "fecha_radicacion": {"FECHA RADICACION", "FECHA DE RADICACION", "RADICACION", "FECHA RADICADO"},
    "valor_glosado": {"VALOR GLOSADO", "GLOSA", "GLOSADO", "VALOR GLOSA"},
    "valor_devuelto": {"VALOR DEVUELTO", "DEVOLUCION", "DEVUELTO", "VALOR DEVOLUCION"},
    "valor_pagado": {"VALOR PAGADO", "PAGADO", "PAGO", "VALOR PAGO", "VALOR RECAUDADO"},
    "fecha_pago": {"FECHA PAGO", "FECHA DE PAGO", "FECHA RECAUDO"},
    "motivo_glosa": {"MOTIVO GLOSA", "MOTIVO", "CAUSAL", "MOTIVO DE GLOSA", "CONCEPTO GLOSA"},
    "numero_nota_credito": {"NOTA CREDITO", "NUMERO NOTA CREDITO", "NC", "NRO NOTA CREDITO"},
    "estado": {"ESTADO", "ESTADO CARTERA", "ESTADO GESTION"},
    "observaciones": {"OBSERVACIONES", "OBSERVACION", "NOTAS", "COMENTARIO"},
}

# Columnas de la planilla de seguimiento que generamos con --crear-plantilla.
COLS_PLANTILLA = [
    "FACTURA",
    "ENTIDAD",
    "VALOR_RADICADO",
    "FECHA_RADICACION",
    "VALOR_GLOSADO",
    "VALOR_DEVUELTO",
    "VALOR_PAGADO",
    "FECHA_PAGO",
    "MOTIVO_GLOSA",
    "NOTA_CREDITO",
    "ESTADO",
    "OBSERVACIONES",
]


# ─── Modelo ──────────────────────────────────────────────────────────────────


@dataclass
class FacturaCartera:
    factura: str
    factura_norm: str
    entidad: str = "SIN ENTIDAD"
    entidad_id: str = ""
    canal: str = ""
    regimen: str = ""
    valor_radicado: float = 0.0
    estado_radicacion: str = ""  # LISTA / REVISAR_TIPIFICACION / FALTAN_SOPORTES…
    # Seguimiento (cartera, alimentado por el equipo):
    fecha_radicacion: date | None = None
    valor_glosado: float = 0.0
    valor_devuelto: float = 0.0
    valor_pagado: float = 0.0
    fecha_pago: date | None = None
    motivo_glosa: str = ""
    numero_nota_credito: str = ""
    estado_manual: str = ""
    observaciones: str = ""

    # ── Derivados ──
    @property
    def saldo(self) -> float:
        """Lo que la entidad queda debiendo = radicado − pagado."""
        return round(self.valor_radicado - self.valor_pagado, 2)

    @property
    def objetado(self) -> float:
        """Parte del saldo que está objetada (glosa + devolución)."""
        return round(self.valor_glosado + self.valor_devuelto, 2)

    @property
    def en_tramite(self) -> float:
        """Saldo sin objetar (por pagar en trámite normal)."""
        return round(max(0.0, self.saldo - self.objetado), 2)

    @property
    def pct_recaudo(self) -> float:
        return (
            round(100.0 * self.valor_pagado / self.valor_radicado, 1)
            if self.valor_radicado
            else 0.0
        )

    @property
    def pct_glosa(self) -> float:
        return round(100.0 * self.objetado / self.valor_radicado, 1) if self.valor_radicado else 0.0

    def dias_mora(self, corte: date) -> int | None:
        if self.fecha_radicacion is None or self.saldo <= 0:
            return None
        return max(0, (corte - self.fecha_radicacion).days)

    def tramo_mora(self, corte: date) -> str | None:
        d = self.dias_mora(corte)
        if d is None:
            return None
        for nombre, lo, hi in TRAMOS_MORA:
            if lo <= d <= hi:
                return nombre
        return TRAMOS_MORA[-1][0]

    @property
    def estado_cartera(self) -> str:
        if self.estado_manual:
            return _norm(self.estado_manual).replace(" ", "_")
        if self.fecha_radicacion is None:
            return "PENDIENTE_RADICAR"
        if self.valor_radicado > 0 and self.valor_pagado >= self.valor_radicado - 0.5:
            return "PAGADA"
        if self.valor_pagado > 0:
            return "PAGO_PARCIAL"
        if self.objetado > 0:
            return "GLOSADA_DEVUELTA"
        return "RADICADA_TRAMITE"


# ─── Lectura de tablas (CSV / XLSX) ──────────────────────────────────────────


def _leer_tabla(ruta: Path) -> list[list]:
    """Devuelve las filas (incluido el encabezado) de un CSV o XLSX."""
    if ruta.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as e:  # pragma: no cover
            raise SystemExit(
                "ERROR: para leer .xlsx falta openpyxl (py -m pip install openpyxl)."
            ) from e
        wb = load_workbook(filename=str(ruta), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        return [list(r) for r in ws.iter_rows(values_only=True)]
    with ruta.open("r", encoding="utf-8-sig", newline="") as fh:
        return list(csv.reader(fh))


def _parse_valor(v) -> float:
    """Parsea un valor en pesos. Acepta número o texto en formato colombiano
    ('$ 1.234.567,89', '1.234.567', '1234567')."""
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    for ch in ("$", " ", " ", "COP", "cop"):
        s = s.replace(ch, "")
    if not s:
        return 0.0
    if "," in s and "." in s:  # 1.234.567,89 → punto = miles, coma = decimal
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:  # 1234567,89 → coma decimal
        s = s.replace(",", ".")
    elif s.count(".") > 1:  # 1.234.567 → puntos de miles
        s = s.replace(".", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_fecha(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()[:10]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _indice_columnas(headers: list[str], alias: dict[str, set[str]]) -> dict[str, int]:
    norm = [_norm(str(h or "")).replace("_", " ") for h in headers]
    idx: dict[str, int] = {}
    for campo, nombres in alias.items():
        for i, h in enumerate(norm):
            if h in nombres:
                idx[campo] = i
                break
    return idx


def _celda_raw(fila: list, idx: dict[str, int], campo: str):
    """Valor crudo de la columna `campo` en la fila (None si no está)."""
    i = idx.get(campo)
    return fila[i] if i is not None and i < len(fila) else None


def _celda(fila: list, idx: dict[str, int], campo: str) -> str:
    """Texto de la columna `campo` en la fila (cadena vacía si falta)."""
    v = _celda_raw(fila, idx, campo)
    return "" if v is None else str(v)


# ─── Carga de la radicación (reportes del radicador) ─────────────────────────

# Encabezados del reporte del radicador que nos interesan.
_ALIAS_RAD: dict[str, set[str]] = {
    "factura": {"FACTURA"},
    "entidad": {"ENTIDAD"},
    "entidad_id": {"ENTIDAD ID", "ENTIDADID"},
    "canal": {"CANAL"},
    "valor_radicado": {"VALOR TOTAL", "VALOR RADICADO", "VALOR"},
    "estado_radicacion": {"ESTADO"},
}


def cargar_radicacion(rutas: list[Path]) -> dict[str, FacturaCartera]:
    """Lee uno o varios reportes del radicador y arma el inventario de
    facturas radicadas indexado por factura normalizada. Si una factura
    aparece en varios reportes, gana el último (re-radicación)."""
    inventario: dict[str, FacturaCartera] = {}
    for ruta in rutas:
        filas = _leer_tabla(ruta)
        if not filas:
            continue
        idx = _indice_columnas([str(h or "") for h in filas[0]], _ALIAS_RAD)
        if "factura" not in idx:
            logger.warning(f"{ruta.name}: no encontré columna FACTURA; lo salto.")
            continue
        n = 0
        for fila in filas[1:]:
            fac = _celda(fila, idx, "factura").strip()
            if not fac:
                continue
            fn = normalizar_factura(fac)
            if fn == "0":
                continue
            inventario[fn] = FacturaCartera(
                factura=fac,
                factura_norm=fn,
                entidad=_celda(fila, idx, "entidad").strip() or "SIN ENTIDAD",
                entidad_id=_celda(fila, idx, "entidad_id").strip(),
                canal=_celda(fila, idx, "canal").strip(),
                valor_radicado=_parse_valor(_celda_raw(fila, idx, "valor_radicado")),
                estado_radicacion=_celda(fila, idx, "estado_radicacion").strip(),
            )
            n += 1
        logger.info(f"Radicación: {n} factura(s) leídas de {ruta.name}")
    return inventario


# ─── Carga del seguimiento de cartera (planilla del equipo) ──────────────────


def cargar_seguimiento(ruta: Path) -> dict[str, dict]:
    """Lee la planilla de seguimiento de cartera y devuelve, por factura
    normalizada, los datos que el equipo alimentó (glosa, pago, fechas…)."""
    filas = _leer_tabla(ruta)
    if not filas:
        return {}
    idx = _indice_columnas([str(h or "") for h in filas[0]], _ALIAS_COLS)
    if "factura" not in idx:
        raise ValueError(
            f"--seguimiento: no encontré la columna FACTURA en {ruta.name} "
            f"(encabezados: {filas[0]})"
        )
    out: dict[str, dict] = {}
    for fila in filas[1:]:
        fac = _celda_raw(fila, idx, "factura")
        if fac in (None, ""):
            continue
        fn = normalizar_factura(str(fac))
        if fn == "0":
            continue
        out[fn] = {
            "fecha_radicacion": _parse_fecha(_celda_raw(fila, idx, "fecha_radicacion")),
            "valor_glosado": _parse_valor(_celda_raw(fila, idx, "valor_glosado")),
            "valor_devuelto": _parse_valor(_celda_raw(fila, idx, "valor_devuelto")),
            "valor_pagado": _parse_valor(_celda_raw(fila, idx, "valor_pagado")),
            "fecha_pago": _parse_fecha(_celda_raw(fila, idx, "fecha_pago")),
            "motivo_glosa": str(_celda_raw(fila, idx, "motivo_glosa") or "").strip(),
            "numero_nota_credito": str(_celda_raw(fila, idx, "numero_nota_credito") or "").strip(),
            "estado_manual": str(_celda_raw(fila, idx, "estado") or "").strip(),
            "observaciones": str(_celda_raw(fila, idx, "observaciones") or "").strip(),
        }
    logger.info(f"Seguimiento: {len(out)} factura(s) con datos de cartera en {ruta.name}")
    return out


def fusionar(
    inventario: dict[str, FacturaCartera], seguimiento: dict[str, dict]
) -> list[FacturaCartera]:
    """Aplica el seguimiento sobre el inventario radicado (por factura)."""
    for fn, datos in seguimiento.items():
        f = inventario.get(fn)
        if f is None:
            continue  # factura del seguimiento que no está en la radicación: se ignora
        for k, v in datos.items():
            if v not in (None, "", 0.0):
                setattr(f, k, v)
    return list(inventario.values())


# ─── Plantilla de seguimiento (pre-cargada con lo radicado) ──────────────────


def crear_plantilla(facturas: list[FacturaCartera], ruta: Path) -> None:
    """Genera la planilla Excel de seguimiento, una fila por factura radicada,
    con FACTURA/ENTIDAD/VALOR_RADICADO ya rellenos y el resto en blanco para que
    cartera complete glosas, pagos y fechas."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:  # pragma: no cover
        raise SystemExit("ERROR: --crear-plantilla necesita openpyxl.") from e

    wb = Workbook()
    ws = wb.active
    ws.title = "SEGUIMIENTO"
    head_fill = PatternFill("solid", fgColor="1E3A8A")
    head_font = Font(bold=True, color="FFFFFF")
    edit_fill = PatternFill("solid", fgColor="FFF7E6")  # columnas a diligenciar
    for c, h in enumerate(COLS_PLANTILLA, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    anchos = {"FACTURA": 16, "ENTIDAD": 30, "MOTIVO_GLOSA": 28, "OBSERVACIONES": 30}
    editables = {
        "FECHA_RADICACION",
        "VALOR_GLOSADO",
        "VALOR_DEVUELTO",
        "VALOR_PAGADO",
        "FECHA_PAGO",
        "MOTIVO_GLOSA",
        "NOTA_CREDITO",
        "ESTADO",
        "OBSERVACIONES",
    }
    for c, h in enumerate(COLS_PLANTILLA, 1):
        ws.column_dimensions[get_column_letter(c)].width = anchos.get(h, 16)
    for r, f in enumerate(sorted(facturas, key=lambda x: x.entidad), start=2):
        ws.cell(row=r, column=1, value=f.factura)
        ws.cell(row=r, column=2, value=f.entidad)
        ws.cell(row=r, column=3, value=round(f.valor_radicado, 2))
        for c, h in enumerate(COLS_PLANTILLA, 1):
            if h in editables:
                ws.cell(row=r, column=c).fill = edit_fill
    ws.freeze_panes = "A2"
    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ruta))
    logger.info(f"Plantilla de seguimiento: {ruta} ({len(facturas)} facturas pre-cargadas)")


# ─── Agregaciones para el tablero ────────────────────────────────────────────


def _mes_label(d: date | None) -> str:
    return f"{d.year:04d}-{d.month:02d}" if d else "Sin fecha"


def agregar(facturas: list[FacturaCartera], corte: date) -> dict:
    """Calcula todos los indicadores que consume el HTML."""
    tot_rad = sum(f.valor_radicado for f in facturas)
    tot_glo = sum(f.valor_glosado for f in facturas)
    tot_dev = sum(f.valor_devuelto for f in facturas)
    tot_pag = sum(f.valor_pagado for f in facturas)
    tot_saldo = sum(f.saldo for f in facturas)

    # Por entidad
    por_entidad: dict[str, dict] = {}
    for f in facturas:
        e = por_entidad.setdefault(
            f.entidad,
            {
                "entidad": f.entidad,
                "n": 0,
                "radicado": 0.0,
                "glosado": 0.0,
                "devuelto": 0.0,
                "pagado": 0.0,
                "saldo": 0.0,
                "_dias": [],
                "listas": 0,
            },
        )
        e["n"] += 1
        e["radicado"] += f.valor_radicado
        e["glosado"] += f.valor_glosado
        e["devuelto"] += f.valor_devuelto
        e["pagado"] += f.valor_pagado
        e["saldo"] += f.saldo
        if f.estado_radicacion == "LISTA":
            e["listas"] += 1
        dm = f.dias_mora(corte)
        if dm is not None:
            e["_dias"].append(dm)
    entidades = []
    for e in por_entidad.values():
        e["pct_recaudo"] = round(100.0 * e["pagado"] / e["radicado"], 1) if e["radicado"] else 0.0
        e["pct_glosa"] = (
            round(100.0 * (e["glosado"] + e["devuelto"]) / e["radicado"], 1)
            if e["radicado"]
            else 0.0
        )
        e["dias_prom"] = round(sum(e["_dias"]) / len(e["_dias"])) if e["_dias"] else 0
        e.pop("_dias", None)
        entidades.append(e)
    entidades.sort(key=lambda x: x["radicado"], reverse=True)

    # Por mes (radicación) — tendencia radicado vs pagado
    por_mes: dict[str, dict] = {}
    for f in facturas:
        m = por_mes.setdefault(
            _mes_label(f.fecha_radicacion), {"mes": "", "radicado": 0.0, "pagado": 0.0, "n": 0}
        )
        m["mes"] = _mes_label(f.fecha_radicacion)
        m["radicado"] += f.valor_radicado
        m["pagado"] += f.valor_pagado
        m["n"] += 1
    meses = [m for k, m in sorted(por_mes.items()) if k != "Sin fecha"]

    # Aging (mora por tramos, solo saldo > 0)
    aging = {nombre: 0.0 for nombre, _, _ in TRAMOS_MORA}
    for f in facturas:
        t = f.tramo_mora(corte)
        if t:
            aging[t] += f.saldo

    # Estados de radicación (módulo "ayuda a radicar")
    estados_rad: dict[str, int] = {}
    for f in facturas:
        estados_rad[f.estado_radicacion or "—"] = estados_rad.get(f.estado_radicacion or "—", 0) + 1

    # Estados de cartera
    estados_car: dict[str, int] = {}
    for f in facturas:
        estados_car[f.estado_cartera] = estados_car.get(f.estado_cartera, 0) + 1

    # Top motivos de glosa (por valor)
    motivos: dict[str, dict] = {}
    for f in facturas:
        if f.objetado > 0 and f.motivo_glosa:
            mo = motivos.setdefault(
                f.motivo_glosa, {"motivo": f.motivo_glosa, "valor": 0.0, "n": 0}
            )
            mo["valor"] += f.objetado
            mo["n"] += 1
    top_motivos = sorted(motivos.values(), key=lambda x: x["valor"], reverse=True)[:10]

    return {
        "generado": corte.isoformat(),
        "kpis": {
            "n_facturas": len(facturas),
            "radicado": round(tot_rad, 2),
            "glosado": round(tot_glo, 2),
            "devuelto": round(tot_dev, 2),
            "pagado": round(tot_pag, 2),
            "saldo": round(tot_saldo, 2),
            "pct_recaudo": round(100.0 * tot_pag / tot_rad, 1) if tot_rad else 0.0,
            "pct_glosa": round(100.0 * (tot_glo + tot_dev) / tot_rad, 1) if tot_rad else 0.0,
            "n_pagadas": sum(1 for f in facturas if f.estado_cartera == "PAGADA"),
            "n_glosadas": sum(1 for f in facturas if f.objetado > 0),
            "n_listas": sum(1 for f in facturas if f.estado_radicacion == "LISTA"),
        },
        "entidades": entidades,
        "meses": meses,
        "aging": aging,
        "estados_radicacion": estados_rad,
        "estados_cartera": estados_car,
        "motivos": top_motivos,
        "facturas": [
            {
                "factura": f.factura,
                "entidad": f.entidad,
                "canal": f.canal,
                "radicado": round(f.valor_radicado, 2),
                "glosado": round(f.valor_glosado, 2),
                "devuelto": round(f.valor_devuelto, 2),
                "pagado": round(f.valor_pagado, 2),
                "saldo": round(f.saldo, 2),
                "pct_recaudo": f.pct_recaudo,
                "estado_radicacion": f.estado_radicacion,
                "estado_cartera": f.estado_cartera,
                "dias_mora": f.dias_mora(corte),
                "fecha_radicacion": f.fecha_radicacion.isoformat() if f.fecha_radicacion else "",
                "fecha_pago": f.fecha_pago.isoformat() if f.fecha_pago else "",
                "motivo_glosa": f.motivo_glosa,
                "nota_credito": f.numero_nota_credito,
                "observaciones": f.observaciones,
            }
            for f in sorted(facturas, key=lambda x: x.saldo, reverse=True)
        ],
    }


# ─── CLI ─────────────────────────────────────────────────────────────────────


def _fecha(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="tablero_cartera.py",
        description="Tablero de Radicación y Cartera — Cuentas Médicas ESE HUS.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument(
        "--radicacion",
        type=Path,
        nargs="+",
        required=True,
        help="Uno o varios reportes del radicador (CSV/XLSX). Fuente de lo RADICADO.",
    )
    p.add_argument(
        "--seguimiento",
        type=Path,
        default=None,
        help="Planilla Excel de cartera (glosas/pagos/fechas por factura). Opcional.",
    )
    p.add_argument(
        "--salida",
        type=Path,
        default=Path("tablero_cartera.html"),
        help="Archivo HTML del tablero (def.: tablero_cartera.html).",
    )
    p.add_argument(
        "--crear-plantilla",
        type=Path,
        default=None,
        dest="crear_plantilla",
        help="Genera la planilla de seguimiento pre-cargada con lo radicado y termina.",
    )
    p.add_argument(
        "--perfiles",
        type=Path,
        default=None,
        help="Catálogo de entidades (def.: data/perfiles_radicacion.json).",
    )
    p.add_argument(
        "--fecha-corte",
        type=_fecha,
        default=None,
        dest="corte",
        help="Fecha de corte AAAA-MM-DD para la mora (def.: hoy).",
    )
    p.add_argument(
        "--titulo",
        type=str,
        default="Tablero de Radicación y Cartera — ESE HUS",
        help="Título que se muestra en el tablero.",
    )
    p.add_argument("--log", type=Path, default=None, help="Archivo de log opcional.")
    return p


def _setup_logging(ruta: Path | None) -> None:
    handlers: list[logging.Handler] = [logging.StreamHandler(sys.stdout)]
    if ruta is not None:
        handlers.append(logging.FileHandler(ruta, encoding="utf-8"))
    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", handlers=handlers
    )


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    _setup_logging(args.log)

    faltan = [r for r in args.radicacion if not r.is_file()]
    if faltan:
        logger.error(f"No existe(n) el/los reporte(s): {', '.join(str(x) for x in faltan)}")
        return 1

    inventario = cargar_radicacion(args.radicacion)
    if not inventario:
        logger.error("No se leyó ninguna factura de los reportes de radicación.")
        return 1

    # Enriquecer entidad/canal/régimen desde el catálogo (si está disponible).
    cfg = cargar_perfiles(args.perfiles) if _RAD_OK else None
    if cfg is not None:
        por_id = {e.id: e for e in getattr(cfg, "entidades", [])}
        for f in inventario.values():
            perfil = por_id.get(f.entidad_id)
            if perfil is not None:
                f.canal = f.canal or perfil.canal
                f.regimen = perfil.regimen

    if args.crear_plantilla is not None:
        crear_plantilla(list(inventario.values()), args.crear_plantilla)
        logger.info("Plantilla creada. Complétala y vuelve a correr con --seguimiento.")
        return 0

    seguimiento = cargar_seguimiento(args.seguimiento) if args.seguimiento else {}
    if args.seguimiento and not seguimiento:
        logger.warning(
            "La planilla de seguimiento no aportó datos; el tablero saldrá solo con radicación."
        )

    facturas = fusionar(inventario, seguimiento)
    corte = args.corte or date.today()
    datos = agregar(facturas, corte)
    datos["titulo"] = args.titulo

    from _tablero_html import render_html

    html = render_html(datos)
    args.salida.parent.mkdir(parents=True, exist_ok=True)
    args.salida.write_text(html, encoding="utf-8")
    logger.info(f"Tablero HTML: {args.salida}")

    k = datos["kpis"]
    logger.info("=" * 60)
    logger.info("RESUMEN DE CARTERA")
    logger.info(f"  Facturas:   {k['n_facturas']:,}")
    logger.info(f"  Radicado:   $ {k['radicado']:,.0f}")
    logger.info(f"  Pagado:     $ {k['pagado']:,.0f}  ({k['pct_recaudo']}% recaudo)")
    logger.info(f"  Glosa/Dev.: $ {k['glosado'] + k['devuelto']:,.0f}  ({k['pct_glosa']}%)")
    logger.info(f"  Saldo:      $ {k['saldo']:,.0f}")
    logger.info("=" * 60)
    return 0


if __name__ == "__main__":
    sys.exit(main())
