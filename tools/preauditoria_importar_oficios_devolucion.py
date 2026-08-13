"""Importa los oficios de devolución históricos (DEV-PRE-AUD) desde el Excel.

PARA QUÉ. Tras importar el consolidado histórico, las devoluciones quedaron
en la base como eventos con su motivo y sus valores, pero sin el OFICIO DE
DEVOLUCIÓN que las agrupa — y sin oficio no hay PDF. El equipo llevó sus
oficios en OFICIOS_DEVOLUCIONES_CONSECUTIVOS.xlsx (hoja PREAUDITORIA): un
bloque por consecutivo con «Consecutivo: DEV-PRE-AUD-####-AAAA», «Fecha:» y
la tabla de facturas. Este bot lee esos bloques, crea los oficios y les
amarra los eventos de devolución que ya están en la base; con eso el botón
PDF de la pestaña «Oficios de devolución» funciona de una.

CÓMO DECIDE:
  - Solo importa los bloques con fecha >= --desde (los viejos no hacen
    falta como PDF; siguen contados en el historial de cada factura).
  - Un bloque repetido (mismo consecutivo dos veces en el Excel) se une.
  - Para cada factura del bloque busca su ÚLTIMO evento de devolución sin
    oficio asignado y se lo amarra; si la factura no está en la base o no
    tiene evento de devolución, se reporta y se salta (no se inventa nada).
  - Si el consecutivo ya existe en la base, el bloque completo se salta.
  - SOLO MIRAR por defecto; con "aplicar" escribe en una transacción.

USO EN EL PC DE CARTERA (desde C:\\motor-glosas\\repo):
  venv\\Scripts\\python.exe tools\\preauditoria_importar_oficios_devolucion.py "C:\\ruta\\OFICIOS_DEVOLUCIONES_CONSECUTIVOS.xlsx" --desde 30/07/2026
  venv\\Scripts\\python.exe tools\\preauditoria_importar_oficios_devolucion.py "C:\\ruta\\OFICIOS_DEVOLUCIONES_CONSECUTIVOS.xlsx" --desde 30/07/2026 aplicar
"""

from __future__ import annotations

import re
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from _dinero import a_numero  # noqa: E402  (lector único de pesos)

try:
    import openpyxl
except ImportError:  # pragma: no cover - autoinstalación para el PC del auditor
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

MARCA_IMPORT = "importacion-historica-excel"
RE_CONSEC = re.compile(r"Consecutivo:\s*(DEV-PRE-AUD-(\d+)-(\d{4}))")
RE_FECHA = re.compile(r"Fecha:\s*([0-9/ ]+)")
RE_FACTURA = re.compile(r"^HUS\d{7,}$")
RE_OFICIO = re.compile(r"^FHUS-[A-Z]+-I?\d+-\d+$")


def _fecha(texto: str) -> datetime | None:
    t = (texto or "").strip().replace(" ", "")
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(t, fmt)
        except ValueError:
            continue
    return None


def leer_bloques(ruta: str) -> list[dict]:
    """Devuelve un bloque por consecutivo (los repetidos se unen)."""
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    ws = wb["PREAUDITORIA"] if "PREAUDITORIA" in wb.sheetnames else wb.worksheets[0]
    bloques: dict[str, dict] = {}
    actual: dict | None = None
    for row in ws.iter_rows(min_row=1, max_col=12, values_only=True):
        celdas = [str(c).strip() if c is not None else "" for c in (row or [])]
        linea = " ".join(c for c in celdas if c)
        m = RE_CONSEC.search(linea)
        if m:
            clave = m.group(1)
            if clave not in bloques:
                bloques[clave] = {
                    "consecutivo": clave,
                    "numero": int(m.group(2)),
                    "anio": int(m.group(3)),
                    "fecha": None,
                    "facturas": {},  # factura -> {envio, valor, oficio, motivo}
                }
            actual = bloques[clave]
            continue
        if actual is None:
            continue
        mf = RE_FECHA.search(linea)
        if mf and actual["fecha"] is None:
            actual["fecha"] = _fecha(mf.group(1))
            continue
        # Fila de factura del bloque: (…, No., ENVIO, FACTURA, F_FACTURA,
        # VALOR, NIT, ENTIDAD, OFICIO, MOTIVO, …) — se ubica por contenido.
        fact = next((c for c in celdas if RE_FACTURA.fullmatch(c)), None)
        if fact:
            idx = celdas.index(fact)
            oficio = next((c for c in celdas if RE_OFICIO.fullmatch(c)), "")
            valor = a_numero(celdas[idx + 2]) if idx + 2 < len(celdas) else 0.0
            envio = celdas[idx - 1] if idx >= 1 else ""
            motivo = celdas[-1] if celdas and len(celdas[-1]) > 10 else ""
            if fact not in actual["facturas"]:
                actual["facturas"][fact] = {
                    "envio": envio.split(".")[0],
                    "valor": valor,
                    "oficio": oficio,
                    "motivo": motivo,
                }
    wb.close()
    # Fecha faltante: se hereda del bloque anterior (van en orden en el Excel)
    lista = sorted(bloques.values(), key=lambda b: (b["anio"], b["numero"]))
    previa = None
    for b in lista:
        if b["fecha"] is None:
            b["fecha"] = previa
        else:
            previa = b["fecha"]
    return lista


def planear(bloques: list[dict], desde: datetime, con: sqlite3.Connection) -> dict:
    cur = con.cursor()
    existentes = {r[0] for r in cur.execute("SELECT consecutivo FROM preaud_oficios_devolucion")}
    plan = {"crear": [], "saltados_existen": [], "fuera_de_rango": 0}
    for b in bloques:
        if b["fecha"] is None or b["fecha"] < desde:
            plan["fuera_de_rango"] += 1
            continue
        if b["consecutivo"] in existentes:
            plan["saltados_existen"].append(b["consecutivo"])
            continue
        detalle = {"bloque": b, "amarra": [], "sin_evento": [], "sin_factura": []}
        for fact, info in b["facturas"].items():
            fila = cur.execute("SELECT id FROM preaud_facturas WHERE factura=?", (fact,)).fetchone()
            if not fila:
                detalle["sin_factura"].append(fact)
                continue
            ev = cur.execute(
                "SELECT id, valor_snapshot FROM preaud_factura_eventos "
                "WHERE factura=? AND tipo_evento IN ('DEVUELTA','NUEVAMENTE_DEVUELTA') "
                "AND oficio_devolucion_id IS NULL ORDER BY id DESC LIMIT 1",
                (fact,),
            ).fetchone()
            if not ev:
                detalle["sin_evento"].append(fact)
                continue
            detalle["amarra"].append(
                {
                    "factura": fact,
                    "factura_id": fila[0],
                    "evento_id": ev[0],
                    "valor": ev[1] or info["valor"],
                    "oficio": info["oficio"],
                }
            )
        plan["crear"].append(detalle)
    return plan


def aplicar(plan: dict, con: sqlite3.Connection) -> dict:
    cur = con.cursor()
    cur.execute("BEGIN IMMEDIATE")
    hechos = {"oficios": 0, "eventos_amarrados": 0, "facturas_marcadas": 0}
    try:
        for det in plan["crear"]:
            b, amarra = det["bloque"], det["amarra"]
            if not amarra:
                continue
            # El oficio de recepción del bloque: el FHUS más repetido entre
            # sus facturas (si está en la base).
            rec_id = None
            fhus = [a["oficio"] for a in amarra if a["oficio"]]
            if fhus:
                mas_comun = max(set(fhus), key=fhus.count)
                fila = cur.execute(
                    "SELECT id FROM preaud_oficios_recepcion WHERE numero_radicado=?",
                    (mas_comun,),
                ).fetchone()
                rec_id = fila[0] if fila else None
            cur.execute(
                "INSERT INTO preaud_oficios_devolucion "
                "(consecutivo, anio, numero, oficio_recepcion_id, fecha_generado, "
                " generado_por, total_facturas, total_valor) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    b["consecutivo"],
                    b["anio"],
                    b["numero"],
                    rec_id,
                    b["fecha"].strftime("%Y-%m-%d %H:%M:%S"),
                    MARCA_IMPORT,
                    len(amarra),
                    sum(a["valor"] for a in amarra),
                ),
            )
            dev_id = cur.lastrowid
            hechos["oficios"] += 1
            for a in amarra:
                cur.execute(
                    "UPDATE preaud_factura_eventos SET oficio_devolucion_id=? WHERE id=?",
                    (dev_id, a["evento_id"]),
                )
                hechos["eventos_amarrados"] += 1
                cur.execute(
                    "UPDATE preaud_facturas SET oficio_devolucion_id=? "
                    "WHERE id=? AND resultado_actual='DEVUELTA'",
                    (dev_id, a["factura_id"]),
                )
                hechos["facturas_marcadas"] += cur.rowcount if cur.rowcount > 0 else 0
        con.commit()
    except Exception:
        if con.in_transaction:
            con.rollback()
        raise
    return hechos


def main() -> int:
    args = sys.argv[1:]
    if not args:
        print(__doc__)
        return 1
    ruta_excel = args[0]
    modo_aplicar = "aplicar" in [a.lower() for a in args[1:]]
    desde = datetime(2026, 7, 30)
    for i, a in enumerate(args):
        if a == "--desde" and i + 1 < len(args):
            f = _fecha(args[i + 1])
            if not f:
                print(f"Fecha --desde no entendida: {args[i + 1]} (use dd/mm/aaaa)")
                return 1
            desde = f
    ruta_db = next(
        (a for a in args[1:] if a.lower().endswith(".db")),
        str(Path(__file__).resolve().parent.parent / "data" / "motorglosas.db"),
    )
    if not Path(ruta_excel).exists():
        print(f"No se encontró el Excel: {ruta_excel}")
        return 1
    if not Path(ruta_db).exists():
        print(f"No se encontró la base de datos: {ruta_db}")
        return 1

    print("=" * 70)
    print("  IMPORTAR OFICIOS DE DEVOLUCIÓN HISTÓRICOS (DEV-PRE-AUD)")
    print(f"  Excel: {ruta_excel}")
    print(f"  Base:  {ruta_db}")
    print(f"  Desde: {desde.strftime('%d/%m/%Y')}")
    print(f"  Modo:  {'APLICAR (escribe)' if modo_aplicar else 'SOLO MIRAR (no escribe nada)'}")
    print("=" * 70)

    bloques = leer_bloques(ruta_excel)
    print(
        f"\nBloques leídos del Excel: {len(bloques)} "
        f"(del {bloques[0]['consecutivo']} al {bloques[-1]['consecutivo']})"
    )

    if modo_aplicar:
        con = sqlite3.connect(ruta_db, timeout=30)
    else:
        con = sqlite3.connect(f"file:{ruta_db}?mode=ro", uri=True, timeout=30)
    con.execute("PRAGMA busy_timeout=15000")

    plan = planear(bloques, desde, con)
    print(f"\nFuera del rango de fechas (se ignoran): {plan['fuera_de_rango']}")
    if plan["saltados_existen"]:
        print(f"Ya existían en la base (se saltan): {', '.join(plan['saltados_existen'])}")
    print(f"Oficios de devolución a crear: {len(plan['crear'])}\n")
    for det in plan["crear"]:
        b = det["bloque"]
        print(
            f"  {b['consecutivo']}  fecha {b['fecha'].strftime('%d/%m/%Y')}  "
            f"→ {len(det['amarra'])} facturas con su evento listo"
        )
        if det["sin_factura"]:
            print(f"      OJO no están en la base (se saltan): {', '.join(det['sin_factura'])}")
        if det["sin_evento"]:
            print(
                f"      OJO sin evento de devolución libre (se saltan): "
                f"{', '.join(det['sin_evento'])}"
            )

    if not modo_aplicar:
        print("\nNada se escribió (SOLO MIRAR). Para aplicar de verdad, agregue: aplicar")
        con.close()
        return 0

    hechos = aplicar(plan, con)
    con.close()
    print("\nLISTO — escrito en la base:")
    for k, v in hechos.items():
        print(f"   {k}: {v}")
    print("\nRecargue la pestaña «Oficios de devolución»: cada uno trae su botón PDF.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
