"""Importa el consolidado histórico de pre-auditoría (Excel del equipo) a la base.

POR QUÉ EXISTE. El 04-08-2026 el sistema se mudó al PC de cartera con una base
provisional vacía (la historia quedó encerrada en la VM de Google). El equipo
llevó SIEMPRE su consolidado a mano en Excel — CONSOLIDADO_PRE_AUDITORIA_2026:
una fila por pasada de factura, desde el 13-04 — así que ese Excel ES la
historia completa del módulo, más completa incluso que la base vieja.

CÓMO LEE EL EXCEL (Hoja1, una fila por PASADA de la factura):
  ITEM | F_RECIBIDO | ENVIO | AUDITOR | FACTURA | F_FACTURA | VALOR | NIT |
  ENTIDAD | CORREO F.E. | PREAUDITORIA RADICACION SINAC (obs) | RADICAR_1 |
  OBSERVACIONES ADICIONALES | F_DEV | F_REINSPECCIÓN | OBSERVACIONES |
  RADICAR_2 | NºOFICIO FACTURACION

  - La misma factura aparece en varias filas cuando facturación la reenvió
    (otra pasada, casi siempre con otro envío).
  - Dentro de una fila: si tiene F_DEV hubo devolución; si tiene
    F_REINSPECCIÓN la factura volvió; RADICAR_1 queda con el resultado FINAL
    de esa pasada (SI/NO). RADICAR_2 casi siempre vacío; cuando trae fecha,
    esa es la radicación de la segunda vuelta (vale como SI).

QUÉ ESCRIBE (mismas tablas que usa la página):
  - preaud_oficios_recepcion: los oficios FHUS reales de la columna final;
    para filas sin oficio, uno sintético por fecha: HIST-AAAA-MM-DD.
  - preaud_fuente_radicacion / preaud_fuente_dgreport: para que el consolidado
    muestre VALOR, NIT, ENTIDAD y CORREO F.E. (el módulo los lee por JOIN).
  - preaud_envios_cargados: el ledger de envíos por oficio.
  - preaud_facturas + preaud_factura_eventos: la factura canónica con su
    estado final y el historial de eventos (ESCRITA → DEVUELTA → REINGRESO →
    SUBSANADA/RADICADA/NUEVAMENTE_DEVUELTA) con fechas y snapshots reales.

REGLAS DE SEGURIDAD:
  - NUNCA borra ni modifica lo que la PÁGINA creó o tocó: si una factura fue
    creada por la página, o alguno de sus eventos vino de la página, se salta
    completa — el sistema manda. Fuentes/oficios/ledger: solo INSERT si falta.
  - Las facturas que ESTE MISMO importador creó sí se pueden PONER AL DÍA
    cuando el Excel avanza (el equipo la radicó/devolvió/subsanó después de
    la última corrida): se agregan SOLO los eventos que faltan al final y se
    actualiza el estado. La historia ya guardada jamás se modifica; si la
    historia nueva no encaja como continuación de la guardada, la factura se
    reporta como conflicto y no se toca.
  - Idempotente: correrlo dos veces no duplica nada.
  - Sin argumento de modo: SOLO MIRAR (no escribe; muestra qué haría).
    Con "aplicar": escribe, en una sola transacción.

USO EN EL PC DE CARTERA (PowerShell o cmd, desde C:\\motor-glosas\\repo):
  venv\\Scripts\\python.exe tools\\preauditoria_importar_consolidado.py "C:\\ruta\\CONSOLIDADO_PRE_AUDITORIA_2026.xlsx"
  venv\\Scripts\\python.exe tools\\preauditoria_importar_consolidado.py "C:\\ruta\\CONSOLIDADO_PRE_AUDITORIA_2026.xlsx" aplicar
"""

from __future__ import annotations

import re
import sqlite3
import sys
from collections import defaultdict
from datetime import datetime, time
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover - autoinstalación para el PC del auditor
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

# Igual que el módulo web (app/services/preauditoria_service.py)
RESULTADO_PENDIENTE = "PENDIENTE"
RESULTADO_RADICAR = "RADICAR"
RESULTADO_DEVUELTA = "DEVUELTA"
ESTADO_NUEVA = "NUEVA"
ESTADO_RADICADA = "RADICADA"
ESTADO_DEVUELTA_PEND = "DEVUELTA_PEND_SUBSANACION"
ESTADO_EN_SUBSANACION = "EN_SUBSANACION"
ESTADO_SUBSANADA = "SUBSANADA"
ESTADO_NUEV_DEVUELTA = "NUEVAMENTE_DEVUELTA"
ESTADO_BLOQUEADA = "BLOQUEADA_LIMITE"

MARCA_IMPORT = "importacion-historica-excel"

# El Excel trae el primer nombre (o las iniciales, en el consolidado ADRES);
# la página usa el nombre completo del equipo.
AUDITORES = {
    "OSCAR": "OSCAR VILLAMIZAR",
    "CAROLINA": "CAROLINA CIFUENTES",
    "CLAUDIA": "CLAUDIA SUAREZ",
    "DIANEYDA": "DIANEYDA QUINTERO",
    "CAMILO": "CAMILO CASTILLO",
    "ELIAS": "ELIAS CARVAJAL",
    "EDGAR": "EDGAR SILVA",
    "VANESSA": "VANESSA OSPINA",
    "DI": "DIANEYDA QUINTERO",
    "EC": "ELIAS CARVAJAL",
    "ES": "EDGAR SILVA",
}


def _txt(v) -> str:
    return str(v).strip() if v is not None and str(v).strip() else ""


def _fecha(v) -> datetime | None:
    if isinstance(v, datetime):
        return v
    t = _txt(v)
    if not t:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(t[:19], fmt)
        except ValueError:
            continue
    return None


def _iso(dt: datetime | None) -> str | None:
    return dt.strftime("%Y-%m-%d %H:%M:%S") if dt else None


class Pasada:
    """Una fila del Excel = una pasada de la factura por pre-auditoría."""

    __slots__ = (
        "item",
        "f_recibido",
        "envio",
        "auditor",
        "factura",
        "f_factura",
        "valor",
        "nit",
        "entidad",
        "correo_fe",
        "obs_preaud",
        "radicar_1",
        "obs_adic",
        "f_dev",
        "f_reinsp",
        "obs_2",
        "radicar_2",
        "oficio",
    )

    def __init__(self, row):
        self.item = row[0]
        self.f_recibido = _fecha(row[1]) or datetime(2026, 1, 1)
        self.envio = _txt(row[2]).split(".")[0]  # 225417.0 → 225417
        self.auditor = AUDITORES.get(_txt(row[3]).upper(), _txt(row[3]).upper() or None)
        self.factura = _txt(row[4])
        self.f_factura = _fecha(row[5])
        try:
            self.valor = float(row[6] or 0)
        except (TypeError, ValueError):
            self.valor = 0.0
        self.nit = _txt(row[7])
        self.entidad = _txt(row[8])
        self.correo_fe = "SI" if _txt(row[9]).upper().startswith("S") else "NO"
        self.obs_preaud = _txt(row[10])
        self.radicar_1 = _txt(row[11]).upper()
        self.obs_adic = _txt(row[12])
        self.f_dev = _fecha(row[13])
        self.f_reinsp = _fecha(row[14])
        self.obs_2 = _txt(row[15])
        # RADICAR_2: "SI"/"NO", o una fecha (la práctica del equipo:
        # anotar la fecha en que por fin se radicó = SI).
        r2 = _txt(row[16]).upper()
        if _fecha(row[16]) is not None:
            r2 = "SI"
        self.radicar_2 = r2
        self.oficio = _txt(row[17])

    @property
    def resultado_final(self) -> str:
        """El desenlace de ESTA pasada: RADICAR, DEVUELTA o PENDIENTE."""
        if self.radicar_2 == "SI" or self.radicar_1 == "SI":
            return RESULTADO_RADICAR
        if self.radicar_2 == "NO" or self.radicar_1 == "NO":
            return RESULTADO_DEVUELTA
        # Sin decisión escrita; si hubo reinspección quedó esperando decidir.
        return RESULTADO_PENDIENTE

    @property
    def hubo_devolucion(self) -> bool:
        """Si esta pasada dejó una devolución ANTES de su desenlace final.

        Con F_DEV escrita la devolución es explícita. Sin F_DEV, un NO en
        RADICAR_1/RADICAR_2 también registra que la factura fue devuelta —
        pero solo mientras no haya reinspección: cuando la hay, la única
        prueba de la devolución inicial es la F_DEV (así se cargó toda la
        historia del 05-08 y así se mantiene la compatibilidad). Esto cubre
        la práctica del equipo de anotar después la radicación final en
        RADICAR_2 sin que se borre la devolución del pasado.
        """
        if self.f_dev is not None:
            return True
        return self.f_reinsp is None and (
            self.radicar_1 == "NO"
            or self.radicar_2 == "NO"
            or self.resultado_final == RESULTADO_DEVUELTA
        )

    @property
    def motivo(self) -> str:
        partes = [p for p in (self.obs_preaud, self.obs_adic, self.obs_2) if p]
        return " · ".join(partes)


def _oficio_limpio(texto: str) -> str:
    """Normaliza el número de oficio: sin espacios y con la I del formato.

    El consolidado ADRES trae variantes como 'FHUS- AS-101139-26' donde la
    'I' quedó digitada como '1'. La forma canónica es FHUS-AS-I01139-26.
    """
    t = re.sub(r"\s+", "", texto or "")
    return re.sub(r"^(FHUS-[A-Z]+-)1(\d{5}-\d{2})$", r"\g<1>I\g<2>", t)


def _fila_desde_adres(row: tuple) -> tuple:
    """Traduce una fila del consolidado ADRES (26 columnas) al orden clásico.

    Columnas ADRES: Oficio | Item | Fecha_Recibido | Envío | AUD | HUS |
    Fecha_Factura | Valor | NIT | Entidad | Correo F.E. | Obs Preauditoría |
    Radicar_1 | Obs Adicionales | Fecha_Entrega_Fact | Obs_FACTURACIÓN |
    Fecha_Dev_CARTERA | Fecha_Segunda_Revisión | Segunda_Obs_SINAC |
    Fecha_Dev_FACTURACIÓN | Segunda_Obs_FACTURACIÓN | Fecha_Dev_CARTERA |
    Radicar_2 | Fecha_Radicación | Número_Radicado | INFOPOL
    """
    f_dev = next((row[i] for i in (16, 19, 21) if _fecha(row[i]) is not None), None)
    obs_2 = " · ".join(p for p in (_txt(row[15]), _txt(row[18]), _txt(row[20])) if p)
    radicado = _txt(row[24])
    if radicado:
        obs_2 = f"{obs_2} · Radicado ADRES: {radicado}" if obs_2 else f"Radicado ADRES: {radicado}"
    radicar_2 = row[22] if _txt(row[22]) else row[23]  # la fecha de radicación vale como SI
    return (
        row[1],  # ITEM
        row[2],  # F_RECIBIDO
        row[3],  # ENVIO
        row[4],  # AUDITOR (iniciales)
        row[5],  # FACTURA
        row[6],  # F_FACTURA
        row[7],  # VALOR
        row[8],  # NIT
        row[9],  # ENTIDAD
        row[10],  # CORREO F.E.
        row[11],  # obs preauditoría
        row[12],  # RADICAR_1
        row[13],  # obs adicionales
        f_dev,  # F_DEV (la primera fecha de devolución que aparezca)
        row[17],  # F_REINSPECCIÓN (Fecha_Segunda_Revisión)
        obs_2,  # OBSERVACIONES segunda vuelta
        radicar_2,  # RADICAR_2
        _oficio_limpio(_txt(row[0])),  # OFICIO
    )


def leer_excel(ruta: str) -> tuple[list[Pasada], list[str]]:
    """Lee el Excel (formato clásico del equipo o consolidado ADRES).

    Devuelve (pasadas, avisos). Los avisos son problemas de datos que el
    auditor debe conocer: fechas dañadas, filas repetidas idénticas.
    """
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    if "Hoja1" in wb.sheetnames:
        ws = wb["Hoja1"]
    elif "CONSOLIDADO" in wb.sheetnames:
        ws = wb["CONSOLIDADO"]
    else:
        ws = wb.worksheets[0]
    filas = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()
    encabezado = [_txt(c).upper() for c in (filas[0] if filas else [])]
    es_adres = bool(encabezado) and encabezado[0] == "OFICIO"

    avisos: list[str] = []
    pasadas: list[Pasada] = []
    vistas: set[tuple] = set()
    ancho = 26 if es_adres else 18
    col_factura = 5 if es_adres else 4
    for numero, row in enumerate(filas[1:], 2):
        if row is None:
            continue
        # En modo read_only openpyxl recorta las columnas vacías del final:
        # se rellena para no perder filas válidas.
        if len(row) < ancho:
            row = tuple(row) + (None,) * (ancho - len(row))
        if not _txt(row[col_factura]).upper().startswith("HUS"):
            continue
        if es_adres:
            row = _fila_desde_adres(row)
        llave = tuple(str(v) for v in row)
        if llave in vistas:
            avisos.append(f"Fila {numero}: repetida idéntica a otra ya leída — se ignora")
            continue
        vistas.add(llave)
        if isinstance(row[13], time):
            avisos.append(
                f"Fila {numero} ({_txt(row[4])}): F_DEV trae una hora suelta "
                f"('{row[13]}') en vez de una fecha — se lee como sin fecha; "
                f"corrija esa celda en el Excel"
            )
        pasadas.append(Pasada(row))
    # Orden cronológico por factura: fecha de recibido y, de empate, ITEM.
    pasadas.sort(key=lambda p: (p.factura, p.f_recibido, p.item or 0))
    return pasadas, avisos


def _radicado_oficio(p: Pasada) -> str:
    """El FHUS real si la fila lo trae; si no, uno sintético por fecha."""
    return p.oficio or f"HIST-{p.f_recibido.strftime('%Y-%m-%d')}"


def _estado_en_base(cur, factura: str) -> dict | None:
    """Lo que la base ya sabe de la factura, para clasificar y revalidar."""
    fila = cur.execute(
        "SELECT id, creado_por, estado, resultado_actual, ronda_actual, "
        "num_devoluciones, envio_actual, oficio_fhus FROM preaud_facturas WHERE factura=?",
        (factura,),
    ).fetchone()
    if fila is None:
        return None
    eventos = cur.execute(
        "SELECT tipo_evento, creado_por FROM preaud_factura_eventos WHERE factura_id=? ORDER BY id",
        (fila[0],),
    ).fetchall()
    return {
        "id": fila[0],
        "creado_por": fila[1],
        "estado": fila[2],
        "resultado": fila[3],
        "ronda": fila[4],
        "num_dev": fila[5],
        "envio": fila[6],
        "oficio": fila[7],
        "tipos": [e[0] for e in eventos],
        "ajeno": any(e[1] != MARCA_IMPORT for e in eventos),
    }


def planear(pasadas: list[Pasada], con: sqlite3.Connection) -> dict:
    """Arma el plan completo SIN escribir: qué se crea y qué se salta."""
    cur = con.cursor()

    existentes = {r[0] for r in cur.execute("SELECT factura FROM preaud_facturas")}
    oficios_db = {
        r[1]: r[0] for r in cur.execute("SELECT id, numero_radicado FROM preaud_oficios_recepcion")
    }
    fuentes_rad = {r[0] for r in cur.execute("SELECT factura FROM preaud_fuente_radicacion")}
    fuentes_dg = {r[0] for r in cur.execute("SELECT factura FROM preaud_fuente_dgreport")}
    ledger = {
        (r[0], r[1]) for r in cur.execute("SELECT envio, oficio_id FROM preaud_envios_cargados")
    }

    por_factura: dict[str, list[Pasada]] = defaultdict(list)
    for p in pasadas:
        por_factura[p.factura].append(p)

    plan = {
        "facturas_nuevas": {},  # factura -> (estado, resultado, eventos[...])
        "facturas_saltadas": [],  # de la página (creadas o tocadas por ella)
        "facturas_al_dia": [],  # del import, sin nada nuevo en el Excel
        "facturas_actualizar": {},  # factura -> info con cola de eventos nueva
        "facturas_encabezado": {},  # historia igual, encabezado por sanar
        "facturas_conflicto": {},  # factura -> (historia en base, historia Excel)
        "oficios_nuevos": {},  # radicado -> fecha_recibido mínima
        "oficios_existentes": set(),
        "fuentes_rad_nuevas": {},  # factura -> última pasada (datos descriptivos)
        "fuentes_dg_nuevas": {},  # factura -> última pasada con correo SI
        "ledger_nuevo": set(),  # (envio, radicado)
        "envios_con_mas_de_3_oficios": [],
        "resumen_estados": defaultdict(int),
    }

    envio_oficios: dict[str, set] = defaultdict(set)

    for factura, filas in por_factura.items():
        ultima = filas[-1]

        # Oficios y ledger se planean SIEMPRE (aun si la factura ya existe,
        # otra factura del mismo envío puede necesitarlos).
        for p in filas:
            rad = _radicado_oficio(p)
            if rad in oficios_db:
                plan["oficios_existentes"].add(rad)
            else:
                previa = plan["oficios_nuevos"].get(rad)
                if previa is None or p.f_recibido < previa:
                    plan["oficios_nuevos"][rad] = p.f_recibido
            if p.envio:
                envio_oficios[p.envio].add(rad)
                clave = (p.envio, rad)
                if rad in oficios_db and (p.envio, oficios_db[rad]) in ledger:
                    continue
                plan["ledger_nuevo"].add(clave)

        # Estado final + historial de eventos (se calcula SIEMPRE: sirve tanto
        # para crear la factura como para ponerla al día si ya existe)
        eventos = []
        num_dev = 0
        ronda = 0
        for p in filas:
            ronda += 1
            rad = _radicado_oficio(p)
            base = {
                "envio": p.envio,
                "radicado": rad,
                "auditor": p.auditor,
                "valor": p.valor,
                "nit": p.nit,
                "entidad": p.entidad,
                "correo": p.correo_fe,
                "f_factura": p.f_factura,
            }
            eventos.append(
                {
                    **base,
                    "tipo": "REINGRESO" if ronda > 1 else "ESCRITA",
                    "fecha": p.f_recibido,
                    "ronda": ronda,
                    "motivo": "",
                    "resultado": None,
                    "estado": None,
                }
            )
            if p.hubo_devolucion:
                num_dev += 1
                eventos.append(
                    {
                        **base,
                        "tipo": "DEVUELTA" if num_dev == 1 else "NUEVAMENTE_DEVUELTA",
                        "fecha": p.f_dev or p.f_recibido,
                        "ronda": ronda,
                        "motivo": p.motivo,
                        "resultado": RESULTADO_DEVUELTA,
                        "estado": ESTADO_DEVUELTA_PEND if num_dev == 1 else ESTADO_NUEV_DEVUELTA,
                    }
                )
            if p.f_reinsp is not None:
                ronda += 1
                eventos.append(
                    {
                        **base,
                        "tipo": "REINGRESO",
                        "fecha": p.f_reinsp,
                        "ronda": ronda,
                        "motivo": "",
                        "resultado": None,
                        "estado": None,
                    }
                )
                if p.resultado_final == RESULTADO_DEVUELTA:
                    num_dev += 1
                    eventos.append(
                        {
                            **base,
                            "tipo": "NUEVAMENTE_DEVUELTA",
                            "fecha": p.f_reinsp,
                            "ronda": ronda,
                            "motivo": p.motivo,
                            "resultado": RESULTADO_DEVUELTA,
                            "estado": ESTADO_NUEV_DEVUELTA,
                        }
                    )
            if p.resultado_final == RESULTADO_RADICAR:
                tipo = "SUBSANADA" if num_dev else "RADICADA"
                eventos.append(
                    {
                        **base,
                        "tipo": tipo,
                        "fecha": p.f_reinsp or p.f_dev or p.f_recibido,
                        "ronda": ronda,
                        "motivo": p.motivo if num_dev else "",
                        "resultado": RESULTADO_RADICAR,
                        "estado": ESTADO_SUBSANADA if num_dev else ESTADO_RADICADA,
                    }
                )

        resultado = ultima.resultado_final
        if resultado == RESULTADO_RADICAR:
            estado = ESTADO_SUBSANADA if num_dev else ESTADO_RADICADA
        elif resultado == RESULTADO_DEVUELTA:
            if num_dev >= 3:
                estado = ESTADO_BLOQUEADA
            elif num_dev >= 2:
                estado = ESTADO_NUEV_DEVUELTA
            else:
                estado = ESTADO_DEVUELTA_PEND
        else:
            # Pendiente de decidir: si ya hubo reingreso (por reinspección en
            # la misma fila o por una fila nueva de reenvío), la factura está
            # EN_SUBSANACION — NUEVA solo existe en la primera ronda.
            hubo_reingreso = ultima.f_reinsp is not None or ronda > 1
            estado = ESTADO_EN_SUBSANACION if hubo_reingreso else ESTADO_NUEVA

        info = {
            "ultima": ultima,
            "eventos": eventos,
            "estado": estado,
            "resultado": resultado,
            "num_dev": num_dev,
            "ronda": max(e["ronda"] for e in eventos),
        }

        if factura not in existentes:
            plan["resumen_estados"][f"{estado} / {resultado}"] += 1
            plan["facturas_nuevas"][factura] = info
            if factura not in fuentes_rad:
                plan["fuentes_rad_nuevas"][factura] = ultima
            if factura not in fuentes_dg and ultima.correo_fe == "SI":
                plan["fuentes_dg_nuevas"][factura] = ultima
            continue

        # La factura ya está en la base: decidir si es de la página (no se
        # toca), si ya está al día, si solo hay que agregarle la cola nueva
        # de eventos, si el encabezado quedó desalineado, o si las historias
        # no encajan (conflicto).
        db = _estado_en_base(cur, factura)
        tipos_excel = [e["tipo"] for e in eventos]
        rad_final = _radicado_oficio(ultima)
        encabezado_igual = (
            db["estado"] == estado
            and db["resultado"] == resultado
            and db["ronda"] == info["ronda"]
            and db["num_dev"] == num_dev
            and (db["envio"] == (ultima.envio or db["envio"]))
            and db["oficio"] == rad_final
        )
        if db["creado_por"] != MARCA_IMPORT or db["ajeno"]:
            plan["facturas_saltadas"].append(factura)
        elif db["tipos"] == tipos_excel:
            if encabezado_igual:
                plan["facturas_al_dia"].append(factura)
            else:
                info.update(id=db["id"], cola=[], estado_antes=db["estado"], tipos_db=db["tipos"])
                plan["facturas_encabezado"][factura] = info
        elif db["tipos"] == tipos_excel[: len(db["tipos"])]:
            info.update(
                id=db["id"],
                cola=eventos[len(db["tipos"]) :],
                estado_antes=db["estado"],
                tipos_db=db["tipos"],
            )
            plan["facturas_actualizar"][factura] = info
            if factura not in fuentes_rad:
                plan["fuentes_rad_nuevas"][factura] = ultima
            if factura not in fuentes_dg and ultima.correo_fe == "SI":
                plan["fuentes_dg_nuevas"][factura] = ultima
        else:
            plan["facturas_conflicto"][factura] = (db["tipos"], tipos_excel)

    plan["facturas_saltadas"].sort()
    plan["facturas_al_dia"].sort()

    plan["envios_con_mas_de_3_oficios"] = sorted(
        e for e, ofs in envio_oficios.items() if len(ofs) > 3
    )
    return plan


def _insertar_evento(cur, factura_id: int, factura: str, ev: dict, oficios_db: dict) -> None:
    cur.execute(
        "INSERT INTO preaud_factura_eventos "
        "(factura_id, factura, tipo_evento, subsanacion_num, ronda, "
        " estado_resultante, envio, oficio_id, oficio_fhus, f_recibido, "
        " resultado, auditor, motivo, observaciones, valor_snapshot, "
        " nit_snapshot, entidad_snapshot, correo_fe_snapshot, "
        " fecha_factura_snapshot, creado_en, creado_por) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (
            factura_id,
            factura,
            ev["tipo"],
            max(ev["ronda"] - 1, 0),
            ev["ronda"],
            ev["estado"],
            ev["envio"] or None,
            oficios_db[ev["radicado"]],
            ev["radicado"],
            _iso(ev["fecha"]),
            ev["resultado"],
            ev["auditor"],
            ev["motivo"] or None,
            ev["motivo"] or None,
            ev["valor"],
            ev["nit"] or None,
            ev["entidad"] or None,
            ev["correo"],
            _iso(ev["f_factura"]),
            _iso(ev["fecha"]),
            MARCA_IMPORT,
        ),
    )


def aplicar(plan: dict, con: sqlite3.Connection) -> dict:
    cur = con.cursor()
    cur.execute("BEGIN IMMEDIATE")
    hechos = {
        "oficios": 0,
        "fuentes": 0,
        "dgreport": 0,
        "ledger": 0,
        "facturas": 0,
        "facturas_actualizadas": 0,
        "encabezados_sanados": 0,
        "cambiaron_mientras_tanto": 0,
        "eventos": 0,
    }
    try:
        oficios_db = {
            r[1]: r[0]
            for r in cur.execute("SELECT id, numero_radicado FROM preaud_oficios_recepcion")
        }
        for rad, fecha in sorted(plan["oficios_nuevos"].items(), key=lambda kv: kv[1]):
            cur.execute(
                "INSERT INTO preaud_oficios_recepcion "
                "(numero_radicado, fecha_recibido, observaciones, creado_en, creado_por) "
                "VALUES (?, ?, ?, ?, ?)",
                (
                    rad,
                    _iso(fecha.replace(hour=8) if fecha.hour == 0 else fecha),
                    "Importado del consolidado histórico en Excel",
                    _iso(datetime.now()),
                    MARCA_IMPORT,
                ),
            )
            oficios_db[rad] = cur.lastrowid
            hechos["oficios"] += 1

        for envio, rad in sorted(plan["ledger_nuevo"]):
            cur.execute(
                "INSERT OR IGNORE INTO preaud_envios_cargados "
                "(envio, oficio_id, total_facturas, nuevas, reingresos, cargado_por, cargado_en) "
                "VALUES (?, ?, 0, 0, 0, ?, ?)",
                (envio, oficios_db[rad], MARCA_IMPORT, _iso(datetime.now())),
            )
            hechos["ledger"] += cur.rowcount if cur.rowcount > 0 else 0

        for factura, p in plan["fuentes_rad_nuevas"].items():
            cur.execute(
                "INSERT OR IGNORE INTO preaud_fuente_radicacion "
                "(factura, envio, f_recibido, f_factura, valor, nit, entidad, "
                " fuente_archivo, importado_por, importado_en, actualizado_en) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    factura,
                    p.envio or "-",
                    _iso(p.f_recibido),
                    _iso(p.f_factura),
                    p.valor,
                    p.nit or None,
                    p.entidad or None,
                    plan.get("fuente_archivo", "CONSOLIDADO_PRE_AUDITORIA_2026.xlsx"),
                    MARCA_IMPORT,
                    _iso(datetime.now()),
                    _iso(datetime.now()),
                ),
            )
            hechos["fuentes"] += cur.rowcount if cur.rowcount > 0 else 0

        for factura, p in plan["fuentes_dg_nuevas"].items():
            cur.execute(
                "INSERT OR IGNORE INTO preaud_fuente_dgreport "
                "(factura, correo_fe, fuente_archivo, importado_por, importado_en, actualizado_en) "
                "VALUES (?, 'SI', ?, ?, ?, ?)",
                (
                    factura,
                    plan.get("fuente_archivo", "CONSOLIDADO_PRE_AUDITORIA_2026.xlsx"),
                    MARCA_IMPORT,
                    _iso(datetime.now()),
                    _iso(datetime.now()),
                ),
            )
            hechos["dgreport"] += cur.rowcount if cur.rowcount > 0 else 0

        for factura, info in plan["facturas_nuevas"].items():
            u: Pasada = info["ultima"]
            rad_final = _radicado_oficio(u)
            ult_evento = info["eventos"][-1]
            cur.execute(
                "INSERT OR IGNORE INTO preaud_facturas "
                "(factura, envio_actual, oficio_actual_id, oficio_fhus, f_recibido, "
                " estado, resultado_actual, ronda_actual, num_subsanacion, "
                " num_devoluciones, pendiente_subsanacion, auditor, fecha_auditoria, "
                " motivo_ultima_devolucion, observaciones, creado_en, creado_por, actualizado_en) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    factura,
                    u.envio or None,
                    oficios_db[rad_final],
                    rad_final,
                    _iso(u.f_recibido),
                    info["estado"],
                    info["resultado"],
                    info["ronda"],
                    max(info["ronda"] - 1, 0),
                    info["num_dev"],
                    1 if info["resultado"] == RESULTADO_DEVUELTA else 0,
                    u.auditor,
                    _iso(ult_evento["fecha"]) if info["resultado"] != RESULTADO_PENDIENTE else None,
                    u.motivo if info["resultado"] == RESULTADO_DEVUELTA else None,
                    u.motivo or None,
                    _iso(datetime.now()),
                    MARCA_IMPORT,
                    _iso(datetime.now()),
                ),
            )
            if cur.rowcount <= 0:
                continue
            factura_id = cur.lastrowid
            hechos["facturas"] += 1
            for ev in info["eventos"]:
                _insertar_evento(cur, factura_id, factura, ev, oficios_db)
                hechos["eventos"] += 1

        # Facturas del propio import que avanzaron (o cuyo encabezado quedó
        # desalineado): solo se AGREGA la cola de eventos nueva y se pone al
        # día el encabezado. Nada existente se toca. Antes de escribir cada
        # una se REVALIDA contra la base dentro de la transacción, por si la
        # página escribió algo entre el plan y este momento.
        pendientes = list(plan["facturas_actualizar"].items()) + list(
            plan["facturas_encabezado"].items()
        )
        for factura, info in pendientes:
            db = _estado_en_base(cur, factura)
            if (
                db is None
                or db["id"] != info["id"]
                or db["creado_por"] != MARCA_IMPORT
                or db["ajeno"]
                or db["tipos"] != info["tipos_db"]
            ):
                hechos["cambiaron_mientras_tanto"] += 1
                continue
            u: Pasada = info["ultima"]
            rad_final = _radicado_oficio(u)
            ult_evento = info["eventos"][-1]
            for ev in info["cola"]:
                _insertar_evento(cur, info["id"], factura, ev, oficios_db)
                hechos["eventos"] += 1
            # Al reingresar, la página limpia el amarre al oficio de
            # devolución del ciclo anterior — aquí se imita ese comportamiento.
            hubo_reingreso = any(ev["tipo"] == "REINGRESO" for ev in info["cola"])
            cur.execute(
                "UPDATE preaud_facturas SET envio_actual=COALESCE(?, envio_actual), "
                "oficio_actual_id=?, oficio_fhus=?, f_recibido=?, estado=?, "
                "resultado_actual=?, ronda_actual=?, num_subsanacion=?, "
                "num_devoluciones=?, pendiente_subsanacion=?, "
                "auditor=COALESCE(?, auditor), fecha_auditoria=?, "
                "motivo_ultima_devolucion=?, observaciones=?, "
                "oficio_devolucion_id=CASE WHEN ? THEN NULL ELSE oficio_devolucion_id END, "
                "actualizado_en=? WHERE id=?",
                (
                    u.envio or None,
                    oficios_db[rad_final],
                    rad_final,
                    _iso(u.f_recibido),
                    info["estado"],
                    info["resultado"],
                    info["ronda"],
                    max(info["ronda"] - 1, 0),
                    info["num_dev"],
                    1 if info["resultado"] == RESULTADO_DEVUELTA else 0,
                    u.auditor,
                    _iso(ult_evento["fecha"]) if info["resultado"] != RESULTADO_PENDIENTE else None,
                    u.motivo if info["resultado"] == RESULTADO_DEVUELTA else None,
                    u.motivo or None,
                    1 if hubo_reingreso else 0,
                    _iso(datetime.now()),
                    info["id"],
                ),
            )
            if info["cola"]:
                hechos["facturas_actualizadas"] += 1
            else:
                hechos["encabezados_sanados"] += 1

        con.commit()
    except Exception:
        if con.in_transaction:
            con.rollback()
        raise
    return hechos


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 1
    ruta_excel = sys.argv[1]
    modo_aplicar = len(sys.argv) > 2 and sys.argv[2].lower() == "aplicar"
    ruta_db = None
    for arg in sys.argv[2:]:
        if arg.lower().endswith(".db"):
            ruta_db = arg
    if not ruta_db:
        base = Path(__file__).resolve().parent.parent / "data" / "motorglosas.db"
        ruta_db = str(base)
    if not Path(ruta_excel).exists():
        print(f"No se encontró el Excel: {ruta_excel}")
        return 1
    if not Path(ruta_db).exists():
        print(f"No se encontró la base de datos: {ruta_db}")
        return 1

    print("=" * 70)
    print("  IMPORTAR EL CONSOLIDADO HISTÓRICO DE PRE-AUDITORÍA")
    print(f"  Excel: {ruta_excel}")
    print(f"  Base:  {ruta_db}")
    print(f"  Modo:  {'APLICAR (escribe)' if modo_aplicar else 'SOLO MIRAR (no escribe nada)'}")
    print("=" * 70)

    pasadas, avisos = leer_excel(ruta_excel)
    print(f"\nFilas leídas del Excel (pasadas): {len(pasadas)}")
    for aviso in avisos:
        print(f"   OJO: {aviso}")

    if modo_aplicar:
        con = sqlite3.connect(ruta_db, timeout=30)
    else:
        con = sqlite3.connect(f"file:{ruta_db}?mode=ro", uri=True, timeout=30)
    con.execute("PRAGMA busy_timeout=15000")

    plan = planear(pasadas, con)
    plan["fuente_archivo"] = Path(ruta_excel).name

    print(f"\nFacturas nuevas a crear:      {len(plan['facturas_nuevas'])}")
    print(f"Facturas del import ya al día (sin cambios): {len(plan['facturas_al_dia'])}")
    n_eventos_cola = sum(len(i["cola"]) for i in plan["facturas_actualizar"].values())
    print(
        f"Facturas del import que AVANZARON y se ponen al día: "
        f"{len(plan['facturas_actualizar'])} (+{n_eventos_cola} eventos nuevos)"
    )
    for f, i in sorted(plan["facturas_actualizar"].items())[:12]:
        print(f"   {f}: {i['estado_antes']} → {i['estado']} (+{len(i['cola'])} eventos)")
    if len(plan["facturas_actualizar"]) > 12:
        print(f"   ... y {len(plan['facturas_actualizar']) - 12} más")
    if plan["facturas_encabezado"]:
        print(
            f"Facturas con la historia igual pero el encabezado por sanar: "
            f"{len(plan['facturas_encabezado'])}"
        )
        for f, i in sorted(plan["facturas_encabezado"].items())[:8]:
            print(f"   {f}: {i['estado_antes']} → {i['estado']}")
    print(
        f"Facturas de la página (no se tocan, el sistema manda): {len(plan['facturas_saltadas'])}"
    )
    if plan["facturas_saltadas"][:8]:
        print(f"   p. ej.: {', '.join(plan['facturas_saltadas'][:8])}")
    if plan["facturas_conflicto"]:
        print(
            f"\nOJO — {len(plan['facturas_conflicto'])} facturas con historia que NO encaja "
            f"(no se tocan; revisar esas filas del Excel a mano — si a la fila le "
            f"falta la F_DEV de una devolución que sí pasó, escribirla suele destrabarla):"
        )
        for f, (en_base, en_excel) in sorted(plan["facturas_conflicto"].items())[:10]:
            print(f"   {f}: base={'→'.join(en_base)}  excel={'→'.join(en_excel)}")
    print(
        f"Oficios nuevos:               {len(plan['oficios_nuevos'])} "
        f"(reales FHUS: {sum(1 for r in plan['oficios_nuevos'] if not r.startswith('HIST-'))}, "
        f"sintéticos HIST-: {sum(1 for r in plan['oficios_nuevos'] if r.startswith('HIST-'))})"
    )
    print(f"Oficios que ya existían:      {len(plan['oficios_existentes'])}")
    print(f"Renglones de ledger de envíos: {len(plan['ledger_nuevo'])}")
    print(f"Fuentes de radicación nuevas: {len(plan['fuentes_rad_nuevas'])}")
    print(f"Correos F.E. (SI) nuevos:     {len(plan['fuentes_dg_nuevas'])}")

    print("\nEstados finales que quedarían:")
    for clave, n in sorted(plan["resumen_estados"].items(), key=lambda kv: -kv[1]):
        print(f"   {n:5d}  {clave}")

    if plan["envios_con_mas_de_3_oficios"]:
        print(
            f"\nOJO: {len(plan['envios_con_mas_de_3_oficios'])} envíos históricos pasaron "
            f"por MÁS de 3 oficios (la regla del tope aplica solo a cargas nuevas): "
            f"{', '.join(plan['envios_con_mas_de_3_oficios'][:6])}"
        )

    if not modo_aplicar:
        print("\nNada se escribió (SOLO MIRAR). Para aplicar de verdad, agregue: aplicar")
        con.close()
        return 0

    hechos = aplicar(plan, con)
    con.close()
    print("\nLISTO — escrito en la base:")
    for k, v in hechos.items():
        print(f"   {k}: {v}")
    print("\nRecargue la página de Pre-auditoría para ver el consolidado lleno.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
