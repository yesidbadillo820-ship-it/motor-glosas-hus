"""preparar_lote_dgh_dispensario.py — Convierte el export de glosas de Dinamica
Gerencial (DGH) al formato estandar que consume el motor de conciliacion.

El area recibe el lote de glosas iniciales como un export DGH (columnas
`FacturaCartera.*` / `ListadoConceptos.*`, hojas 'INICIAL' + detalle). El motor
(asistente / expediente / piloto) espera el layout tipo HUS.xlsx (FACTURA,
CODIGO CONCEPTO DE GLOSA, VALOR OBJETADO, SERVICIO OBJETADO, ...). Esta
herramienta hace la conversion SIN transcripcion manual:

- Detecta sola la hoja de detalle (la que trae `FacturaCartera.Factura` y
  `ListadoConceptos.ConceptoObjecion.Codigo`).
- CODIGO CONCEPTO DE GLOSA = codigo + nombre del concepto (p. ej.
  "TA0201 El cargo por consulta...") -> familia_de/codigo_corto funcionan igual.
- SERVICIO OBJETADO = descripcion + codigo del servicio (p. ej.
  "CONSULTA DE CONTROL... 890380H") -> el motor de evidencia extrae el codigo.
- FECHA ATENCION: el export DGH NO trae la fecha de atencion; se usa
  `FacturaCartera.Fecha` (fecha de la factura) como aproximacion DOCUMENTADA
  (columna ORIGEN FECHA = "FECHA FACTURA (aprox)"). La fecha real de atencion
  queda en los RIPS y el expediente la puede contrastar despues.
- PACIENTE/IDENTIFICACION quedan vacios: el export no los trae; salen del RIPS
  cuando el piloto cruza con el indice de soportes (nunca se inventan).
- Si el libro trae la hoja 'INICIAL', cruza por factura la fecha de VENCE (para
  priorizar) y el radicado.

USO
---
    py tools\\preparar_lote_dgh_dispensario.py ^
        --dgh    "D:\\...\\GLOSAS_23_JULIO.xlsx" ^
        --salida "D:\\...\\HUS_LOTE_23JUL.xlsx"
"""

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path

logger = logging.getLogger("preparar_lote_dgh")

# Encabezados del export DGH (se mapea por NOMBRE, no por posicion).
COL_FACTURA = "FacturaCartera.Factura"
COL_VALOR_FACTURA = "FacturaCartera.Valor"
COL_FECHA_FACTURA = "FacturaCartera.Fecha"
COL_CONSECUTIVO = "Consecutivo"
COL_COD_CONCEPTO = "ListadoConceptos.ConceptoObjecion.Codigo"
COL_NOM_CONCEPTO = "ListadoConceptos.ConceptoObjecion.Nombre"
COL_COD_SERVICIO = "ListadoConceptos.ServicioProductoFactura.Codigo"
COL_DESC_SERVICIO = "ListadoConceptos.ServicioProductoFactura.Descripcion"
COL_VALOR_OBJ = "ListadoConceptos.ValorObjecion"
COL_MOTIVO = "ListadoConceptos.Observaciones"

REQUERIDAS = (COL_FACTURA, COL_COD_CONCEPTO, COL_VALOR_OBJ)

ENCABEZADOS_SALIDA = [
    "FACTURA",
    "RADICADO",
    "FECHA ATENCION",
    "NOMBRES Y APELLIDOS",
    "IDENTIFICACION",
    "VALOR FACTURA",
    "# GLOSA",
    "CODIGO CONCEPTO DE GLOSA",
    "MOTIVO DE GLOSA",
    "VALOR OBJETADO",
    "SERVICIO OBJETADO",
    # Extras informativos (el motor los ignora; ayudan al auditor):
    "VENCE",
    "CONSECUTIVO DGH",
    "ORIGEN FECHA",
]


def setup_logging() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")


def _num(x) -> int:
    try:
        return int(float(x))
    except (TypeError, ValueError):
        return 0


def detectar_hoja_detalle(wb) -> tuple[str, list, list[tuple]]:
    """Devuelve (nombre_hoja, encabezados, filas_datos) de la hoja de detalle DGH."""
    for nombre in wb.sheetnames:
        ws = wb[nombre]
        filas = list(ws.iter_rows(values_only=True))
        if not filas:
            continue
        headers = [str(h) if h is not None else "" for h in filas[0]]
        if all(req in headers for req in REQUERIDAS):
            return nombre, headers, filas[1:]
    raise ValueError(
        "Ninguna hoja trae las columnas DGH esperadas "
        f"({', '.join(REQUERIDAS)}). ¿Es el export de Dinamica Gerencial?"
    )


def leer_inicial(wb) -> dict[str, dict]:
    """Si existe la hoja 'INICIAL', devuelve factura -> {vence, radicado}."""
    if "INICIAL" not in wb.sheetnames:
        return {}
    filas = list(wb["INICIAL"].iter_rows(values_only=True))
    if not filas:
        return {}
    H = {str(h): i for i, h in enumerate(filas[0]) if h is not None}
    i_fac, i_ven, i_rad = H.get("FACTURA"), H.get("VENCE"), H.get("RADICADO")
    out: dict[str, dict] = {}
    for r in filas[1:]:
        if i_fac is None or not r[i_fac]:
            continue
        out[str(r[i_fac]).strip()] = {
            "vence": r[i_ven] if i_ven is not None else None,
            "radicado": str(r[i_rad]).strip() if i_rad is not None and r[i_rad] else "",
        }
    return out


def convertir(dgh: Path, salida: Path) -> dict:
    """Convierte el export DGH al layout estandar. Devuelve el resumen."""
    import openpyxl

    wb = openpyxl.load_workbook(str(dgh), data_only=True, read_only=True)
    hoja, headers, datos = detectar_hoja_detalle(wb)
    inicial = leer_inicial(wb)
    wb.close()
    H = {h: i for i, h in enumerate(headers)}  # primera ocurrencia de cada nombre

    def g(row, col):
        i = H.get(col)
        return row[i] if i is not None and i < len(row) else None

    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = "Hoja1"
    ws.append(ENCABEZADOS_SALIDA)

    contador: dict[str, int] = {}
    total = filas_out = 0
    facturas: set[str] = set()
    for r in datos:
        factura = str(g(r, COL_FACTURA) or "").strip()
        cod = str(g(r, COL_COD_CONCEPTO) or "").strip()
        if not factura or not cod:
            continue
        contador[factura] = contador.get(factura, 0) + 1
        nombre_con = str(g(r, COL_NOM_CONCEPTO) or "").strip()
        desc_srv = str(g(r, COL_DESC_SERVICIO) or "").strip()
        cod_srv = str(g(r, COL_COD_SERVICIO) or "").strip()
        valor = _num(g(r, COL_VALOR_OBJ))
        extra = inicial.get(factura, {})
        ws.append(
            [
                factura,
                extra.get("radicado", "") or str(g(r, COL_CONSECUTIVO) or ""),
                g(r, COL_FECHA_FACTURA),  # aproximacion documentada (ORIGEN FECHA)
                "",  # paciente: sale del RIPS via indice, no se inventa
                "",  # identificacion: idem
                _num(g(r, COL_VALOR_FACTURA)),
                str(contador[factura]),
                (cod + " " + nombre_con).strip(),
                str(g(r, COL_MOTIVO) or "").strip(),
                valor,
                (desc_srv + " " + cod_srv).strip(),
                extra.get("vence"),
                str(g(r, COL_CONSECUTIVO) or ""),
                "FECHA FACTURA (aprox)",
            ]
        )
        total += valor
        filas_out += 1
        facturas.add(factura)

    salida.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(str(salida))
    return {
        "hoja_detalle": hoja,
        "glosas": filas_out,
        "facturas": len(facturas),
        "valor_objetado_total": total,
        "con_vence": sum(1 for f in facturas if inicial.get(f, {}).get("vence")),
    }


def main(argv: list[str] | None = None) -> int:
    setup_logging()
    ap = argparse.ArgumentParser(
        description="Convierte el export DGH de glosas al formato del motor (HUS.xlsx)."
    )
    ap.add_argument("--dgh", type=Path, required=True, help="Export de Dinamica Gerencial (xlsx)")
    ap.add_argument("--salida", type=Path, required=True, help="Excel convertido (layout HUS)")
    args = ap.parse_args(argv)

    try:
        r = convertir(args.dgh, args.salida)
    except (OSError, ValueError) as e:
        logger.error("No pude convertir %s: %s", args.dgh, e)
        return 1
    logger.info(
        "Convertido (hoja '%s'): %d glosas de %d facturas, $%s objetados  ->  %s",
        r["hoja_detalle"],
        r["glosas"],
        r["facturas"],
        f"{r['valor_objetado_total']:,}",
        args.salida,
    )
    logger.info("  FECHA ATENCION = fecha de FACTURA (aprox; la real esta en el RIPS).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
