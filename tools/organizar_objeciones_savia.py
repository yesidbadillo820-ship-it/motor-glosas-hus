"""organizar_objeciones_savia.py — Arma las objeciones de SAVIA SALUD en el
formato del Dispensario (OBJECIONES).

Toma el Excel de glosas de SAVIA SALUD (8 columnas: Numero_factura,
Cod_Servicio, Servicio, Cantidad_Servicio, Valor_Unitario, Valor_Glosa,
Motivo_Esp_Glosa_Valor_A, Observacion_Glosa_A — como `SAVIA_SALUD_8.03.xlsx`) y
lo convierte al **formato del Dispensario** — el mismo layout de 16 columnas del
archivo `OBJECIONES_DISPENSARIO_HUS*.xlsx`, que es la GUÍA/plantilla de salida:

    CDCONSEC | CDFECDOC | CRNCXC | CROFECOBJ | CROREFERE | CROOBSERV | CROCLAOBJ |
    CRNCLAOBJ | GENUSUARIO4 | CRNCONOBJ | SLNSERPRO | IDRIPS | CTNCENCOS |
    CROVALOBJ | CRDOBSERV | CROTIPOBJ

Genera **un archivo por factura** identificado como de SAVIA
(`OBJECIONES_SAVIA_<CRNCXC>.xlsx` por defecto; se cambia con `--prefijo`). La
estructura interna es la de la guía, pero el archivo es de SAVIA, no del
Dispensario.

MAPEO DE CAMPOS (SAVIA → Dispensario):
    CRNCXC       ← Numero_factura            (HUS443697 → HUS0000443697, 10 dígitos)
    CRNCONOBJ    ← Motivo_Esp_Glosa_Valor_A  (TA08 → TA0801, se completa el consecutivo)
    SLNSERPRO    ← Cod_Servicio
    CROVALOBJ    ← Valor_Glosa
    CRDOBSERV    ← "<CRNCONOBJ> <Observacion_Glosa_A>$<Valor_Glosa>" (formato Dispensario)
    CDFECDOC / CROFECOBJ ← --fecha (default: hoy)
    CDCONSEC     ← consecutivo POR FACTURA (1-1-1 la 1ª, 2-2-2 la 2ª, …)
    CROTIPOBJ    ← por factura: solo TA/FA/SO/AU→0, solo CL→1, mezcla con CL→2
    CROCLAOBJ=0, GENUSUARIO4=999  (constantes de la guía)
    CROREFERE, CROOBSERV, CRNCLAOBJ, IDRIPS, CTNCENCOS ← vacíos

CÓDIGO DE OBJECIÓN (--codigo-sufijo / --mapa-codigos):
    SAVIA usa códigos de 4 caracteres (grupo+concepto): TA08, CL07, SO61. El
    Dispensario usa 6 (grupo+concepto+consecutivo): TA0801, CL0701, SO6101. Por
    defecto se completa con el sufijo "01" (el que traía la guía). Se cambia con
    --codigo-sufijo, o se fuerza código por código con --mapa-codigos
    {"TA08": "TA0805", ...}.

USO:
    py organizar_objeciones_savia.py \
        --entrada "SAVIA_SALUD_8.03.xlsx" \
        --salida  "OBJECIONES_SAVIA"           # carpeta destino

INSTALACIÓN (una vez):
    py -m pip install openpyxl
"""

from __future__ import annotations

import argparse
import datetime as _dt
import json
import logging
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

# Un solo lector de pesos para todos los bots: la copia que vivía aquí
# multiplicaba por cien los valores con centavos.
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _dinero import a_entero  # noqa: E402

logger = logging.getLogger("organizar_savia")


# ─── Formato de salida: layout del Dispensario (archivo OBJECIONES) ───────────

# Orden EXACTO de columnas del export del Dispensario (OBJECIONES_DISPENSARIO_*).
COLUMNAS_DISPENSARIO: tuple[str, ...] = (
    "CDCONSEC",
    "CDFECDOC",
    "CRNCXC",
    "CROFECOBJ",
    "CROREFERE",
    "CROOBSERV",
    "CROCLAOBJ",
    "CRNCLAOBJ",
    "GENUSUARIO4",
    "CRNCONOBJ",
    "SLNSERPRO",
    "IDRIPS",
    "CTNCENCOS",
    "CROVALOBJ",
    "CRDOBSERV",
    "CROTIPOBJ",
)

# Valores constantes tomados tal cual de la guía OBJECIONES_DISPENSARIO_*.xlsx.
# OJO con los tipos: en los archivos reales CDCONSEC y GENUSUARIO4 se guardan
# como TEXTO ('1', '999'); CROCLAOBJ/CROVALOBJ/CROTIPOBJ como NÚMERO.
# CROTIPOBJ NO es constante: se calcula por factura (ver crotipobj_factura).
CDCONSEC_DEFAULT = 1
CROCLAOBJ_CONST = 0
GENUSUARIO4_CONST = "999"
CODIGO_SUFIJO_DEFAULT = "01"


# ─── Lectura del Excel de SAVIA SALUD (8 columnas) ───────────────────────────


def _norm_header(h: object) -> str:
    """Normaliza un encabezado: mayúsculas, sin tildes, sin espacios de más."""
    s = unicodedata.normalize("NFKD", str(h or "").strip().upper())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.split())


# Alias de las columnas de SAVIA (tolerante a tildes/espacios/mayúsculas).
COLUMNAS_SAVIA = {
    "factura": {"NUMERO_FACTURA", "NUMERO FACTURA", "FACTURA"},
    "cod_servicio": {"COD_SERVICIO", "COD SERVICIO", "CODIGO SERVICIO", "CODIGO_SERVICIO"},
    "servicio": {"SERVICIO", "DESCRIPCION SERVICIO"},
    "cantidad": {"CANTIDAD_SERVICIO", "CANTIDAD SERVICIO", "CANTIDAD"},
    "valor_unitario": {"VALOR_UNITARIO", "VALOR UNITARIO"},
    "valor_glosa": {"VALOR_GLOSA", "VALOR GLOSA", "VALOR OBJETADO"},
    "motivo": {"MOTIVO_ESP_GLOSA_VALOR_A", "MOTIVO ESP GLOSA VALOR A", "MOTIVO", "CODIGO GLOSA"},
    "observacion": {
        "OBSERVACION_GLOSA_A",
        "OBSERVACION GLOSA A",
        "OBSERVACION",
        "OBSERVACION GLOSA",
    },
}
# Índices fijos de respaldo (layout de SAVIA_SALUD_8.03.xlsx).
IDX_FALLBACK = {
    "factura": 0,
    "cod_servicio": 1,
    "servicio": 2,
    "cantidad": 3,
    "valor_unitario": 4,
    "valor_glosa": 5,
    "motivo": 6,
    "observacion": 7,
}


def _resolver_columnas(headers: list[str]) -> dict[str, int]:
    """Devuelve {clave: indice} mapeando por nombre de encabezado; si una
    columna no se reconoce por nombre, usa el índice fijo de respaldo."""
    norm = [_norm_header(h) for h in headers]
    idx: dict[str, int] = {}
    for clave, alias in COLUMNAS_SAVIA.items():
        encontrado = next((i for i, h in enumerate(norm) if h in alias), None)
        if encontrado is None:
            encontrado = IDX_FALLBACK[clave]
            logger.debug(
                f"  columna '{clave}' no reconocida por nombre; uso índice fijo {encontrado}"
            )
        idx[clave] = encontrado
    return idx


def _cell(row: tuple, idx: dict[str, int], clave: str) -> object:
    """Valor de la columna `clave` en `row`, o None si el índice se sale de rango."""
    i = idx[clave]
    return row[i] if i < len(row) else None


def _num(v: object) -> int:
    """Pesos enteros, con el lector único de `tools/_dinero.py`.

    Lo que hacía antes: `re.sub(r"[^\\d\\-]", "", str(v))` — borraba **todo**
    lo que no fuera dígito, incluida la coma decimal. Así `"1.365,50"` se
    volvía `136550`: **cien veces el valor real**, y ese número salía en el
    Excel que se carga al ERP. La validación de suma no lo veía porque el
    encabezado y los renglones se inflaban por igual.

    El lector compartido aplica la regla colombiana: un separador de miles
    lleva tres dígitos detrás; uno o dos son decimales.
    """
    return a_entero(v)


# ─── Normalización de factura (corta → larga del Dispensario) ────────────────

_RE_FACTURA = re.compile(r"^([A-Za-z]+)0*(\d+)$")


def factura_larga(fac: object, ancho: int = 10) -> str:
    """HUS443697 → HUS0000443697 (rellena con ceros a `ancho` dígitos, el formato
    largo que usa el Dispensario). Idempotente: HUS0000443697 → HUS0000443697.
    Si no matchea el patrón, devuelve el texto tal cual."""
    s = str(fac or "").strip()
    m = _RE_FACTURA.match(s)
    if not m:
        return s
    return m.group(1).upper() + m.group(2).zfill(ancho)


# ─── Código de objeción (SAVIA 4 chars → Dispensario 6 chars) ────────────────


def codigo_dispensario(
    motivo: str,
    sufijo: str = CODIGO_SUFIJO_DEFAULT,
    mapa: dict[str, str] | None = None,
) -> str:
    """Completa el código de SAVIA al de 6 caracteres del Dispensario.

    - Si `mapa` trae el código exacto, se usa ese (override manual).
    - Si el código es de 4 caracteres (grupo+concepto, p.ej. TA08), se le agrega
      el `sufijo` del consecutivo: TA08 → TA0801.
    - Si ya viene con 6 o más caracteres, se deja tal cual.
    """
    cod = (motivo or "").strip().upper()
    if mapa and cod in mapa:
        return mapa[cod]
    if re.fullmatch(r"[A-Z]{2}\d{2}", cod):
        return cod + sufijo
    return cod


# ─── CROTIPOBJ: tipo de objeción por factura ─────────────────────────────────

# Grupos ADMINISTRATIVOS del catálogo (tarifas, facturación, soportes,
# autorizaciones…). CL (calidad/pertinencia) es el grupo clínico.
GRUPO_CLINICO = "CL"


def crotipobj_factura(grupos: set[str]) -> int:
    """Tipo de objeción de la factura según sus grupos de concepto (2 letras):

    - solo administrativos (TA/FA/SO/AU…)  → 0
    - solo CL (calidad)                    → 1
    - administrativos + CL (mezclada)      → 2
    """
    tiene_cl = GRUPO_CLINICO in grupos
    tiene_admin = any(g != GRUPO_CLINICO for g in grupos)
    if tiene_cl and tiene_admin:
        return 2
    if tiene_cl:
        return 1
    return 0


# ─── Texto de la observación en el formato del Dispensario ───────────────────

_RE_VALOR_FINAL = re.compile(r"\$\s*[\d.,]+\s*$")


def construir_crdobserv(codigo: str, observacion: str, valor: int) -> str:
    """Arma el CRDOBSERV con el formato del Dispensario: ``<código> <texto>$<valor>``.

    Verificado contra los archivos reales (OBJECIONES_DISPENSARIO / EMSSANAR):
    el texto SIEMPRE empieza con el código de objeción y termina pegado a
    ``$<valor objetado>``. Evita duplicar el código o el ``$valor`` si el texto
    de SAVIA ya los trajera."""
    t = (observacion or "").strip()
    cod = (codigo or "").strip()
    if cod and t.upper().startswith(cod.upper()):
        t = t[len(cod) :].strip()
    t = _RE_VALOR_FINAL.sub("", t).strip()
    prefijo = f"{cod} " if cod else ""
    return f"{prefijo}{t}${valor}"


# ─── Transformación SAVIA → Dispensario ──────────────────────────────────────


def construir_registros(
    ruta: Path,
    fecha: _dt.datetime,
    consecutivo: int,
    codigo_sufijo: str,
    mapa_codigos: dict[str, str] | None,
) -> list[dict]:
    """Lee el Excel de SAVIA y devuelve una lista de dicts, uno por objeción, ya
    con las 16 columnas del formato del Dispensario."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.stderr.write("ERROR: falta openpyxl. py -m pip install openpyxl\n")
        sys.exit(2)

    wb = load_workbook(filename=str(ruta), data_only=True, read_only=True)
    ws = wb.active
    filas = ws.iter_rows(values_only=True)
    try:
        headers = list(next(filas))
    except StopIteration:
        logger.warning(f"  {ruta.name}: hoja vacía")
        return []
    idx = _resolver_columnas([str(h) for h in headers])

    # CDCONSEC es el consecutivo POR FACTURA: todas las filas de la 1ª factura
    # llevan `consecutivo`, las de la 2ª `consecutivo+1`, y así (1-1-1, 2-2-2…),
    # numerando por orden de aparición.
    consec_por_factura: dict[str, int] = {}

    registros: list[dict] = []
    for r in filas:
        if r is None:
            continue
        factura_raw = str(_cell(r, idx, "factura") or "").strip()
        observacion = str(_cell(r, idx, "observacion") or "").strip()
        motivo = str(_cell(r, idx, "motivo") or "").strip()
        # Fila sin factura ni contenido: se salta.
        if not factura_raw and not observacion and not motivo:
            continue

        crncxc = factura_larga(factura_raw)
        if crncxc not in consec_por_factura:
            consec_por_factura[crncxc] = consecutivo + len(consec_por_factura)
        codigo = codigo_dispensario(motivo, codigo_sufijo, mapa_codigos)
        valor = _num(_cell(r, idx, "valor_glosa"))
        registros.append(
            {
                # CDCONSEC como TEXTO, igual que los archivos reales.
                "CDCONSEC": str(consec_por_factura[crncxc]),
                "CDFECDOC": fecha,
                "CRNCXC": crncxc,
                "CROFECOBJ": fecha,
                "CROREFERE": None,
                "CROOBSERV": None,
                "CROCLAOBJ": CROCLAOBJ_CONST,
                "CRNCLAOBJ": None,
                "GENUSUARIO4": GENUSUARIO4_CONST,
                "CRNCONOBJ": codigo,
                "SLNSERPRO": str(_cell(r, idx, "cod_servicio") or "").strip() or None,
                "IDRIPS": None,
                "CTNCENCOS": None,
                "CROVALOBJ": valor,
                "CRDOBSERV": construir_crdobserv(codigo, observacion, valor),
                "CROTIPOBJ": 0,  # placeholder: se calcula por factura abajo
            }
        )
    wb.close()

    # CROTIPOBJ se define POR FACTURA según la mezcla de conceptos que lleve:
    #   solo administrativos (TA/FA/SO/AU…) → 0
    #   solo CL (calidad/pertinencia)       → 1
    #   administrativos + CL                → 2
    # Todas las filas de la factura llevan el mismo valor.
    grupos_por_factura: dict[str, set[str]] = defaultdict(set)
    for reg in registros:
        grupos_por_factura[reg["CRNCXC"]].add(str(reg["CRNCONOBJ"])[:2].upper())
    for reg in registros:
        reg["CROTIPOBJ"] = crotipobj_factura(grupos_por_factura[reg["CRNCXC"]])
    return registros


# ─── Escritura del/los Excel del Dispensario ─────────────────────────────────


# Formato de celda POR COLUMNA, copiado 1:1 del archivo real
# (OBJECIONES_EMSSANAR_HUS0000515948.xlsx). 'mm-dd-yy' es el formato builtin
# de FECHA CORTA de Excel (se muestra dd/mm/yyyy según la config regional, sin
# horas); '@' es texto; CROVALOBJ lleva el formato contable de miles.
FORMATOS_DISPENSARIO: dict[str, str] = {
    "CDCONSEC": "@",
    "CDFECDOC": "mm-dd-yy",
    "CRNCXC": "@",
    "CROFECOBJ": "mm-dd-yy",
    "CROREFERE": "@",
    "CROOBSERV": "@",
    "CROCLAOBJ": "General",
    "CRNCLAOBJ": "@",
    "GENUSUARIO4": "@",
    "CRNCONOBJ": "@",
    "SLNSERPRO": "@",
    "IDRIPS": "@",
    "CTNCENCOS": "@",
    "CROVALOBJ": '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-',
    "CRDOBSERV": "@",
    "CROTIPOBJ": "0",
}


def _escribir_hoja(registros: list[dict], salida: Path) -> None:
    """Escribe UN archivo del Dispensario (hoja OBJECIONES) con las 16 columnas."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "OBJECIONES"

    fill_hdr = PatternFill("solid", fgColor="1F4E78")
    font_hdr = Font(bold=True, color="FFFFFF")
    for col, nombre in enumerate(COLUMNAS_DISPENSARIO, start=1):
        c = ws.cell(row=1, column=col, value=nombre)
        c.fill = fill_hdr
        c.font = font_hdr

    fila = 2
    for reg in registros:
        for col, nombre in enumerate(COLUMNAS_DISPENSARIO, start=1):
            celda = ws.cell(row=fila, column=col, value=reg.get(nombre))
            celda.number_format = FORMATOS_DISPENSARIO[nombre]
        fila += 1

    ws.freeze_panes = "A2"
    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(salida))


PREFIJO_DEFAULT = "OBJECIONES_SAVIA"


def escribir_por_factura(
    registros: list[dict],
    carpeta: Path,
    prefijo: str = PREFIJO_DEFAULT,
    consecutivo: int = CDCONSEC_DEFAULT,
) -> list[Path]:
    """Escribe un archivo por factura, nombrado <prefijo>_<CRNCXC>.xlsx (por
    defecto OBJECIONES_SAVIA_<CRNCXC>.xlsx). La estructura interna es la de la
    guía del Dispensario, pero el archivo se identifica como de SAVIA.

    Cada archivo lleva UNA factura, así que su CDCONSEC arranca en `consecutivo`
    (1 por defecto), igual que la guía de una sola factura. Devuelve las rutas
    generadas."""
    por_factura: dict[str, list[dict]] = defaultdict(list)
    for reg in registros:
        por_factura[reg["CRNCXC"]].append(reg)

    generados: list[Path] = []
    for crncxc, regs in sorted(por_factura.items()):
        # Archivo standalone: su única factura es el consecutivo 1 (como TEXTO).
        for reg in regs:
            reg["CDCONSEC"] = str(consecutivo)
        destino = carpeta / f"{prefijo}_{crncxc}.xlsx"
        _escribir_hoja(regs, destino)
        generados.append(destino)
        logger.info(f"  {destino.name}: {len(regs)} objeciones")
    return generados


def escribir_consolidado(registros: list[dict], salida: Path) -> None:
    """Escribe un único archivo con todas las facturas juntas."""
    _escribir_hoja(registros, salida)


# ─── CLI ─────────────────────────────────────────────────────────────────────


def _cargar_mapa(ruta: Path | None) -> dict[str, str]:
    if ruta is None:
        return {}
    try:
        data = json.loads(ruta.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        logger.error(f"No pude leer el mapa de códigos {ruta}: {e}")
        sys.exit(2)
    if not isinstance(data, dict):
        logger.error('El mapa de códigos debe ser un objeto JSON {"TA08": "TA0801", ...}')
        sys.exit(2)
    return {str(k).strip().upper(): str(v).strip() for k, v in data.items()}


def _parse_fecha(texto: str | None) -> _dt.datetime:
    if not texto:
        hoy = _dt.date.today()
        return _dt.datetime(hoy.year, hoy.month, hoy.day)
    try:
        d = _dt.datetime.strptime(texto.strip(), "%Y-%m-%d")
        return d
    except ValueError:
        logger.error(f"--fecha inválida: {texto!r} (usá YYYY-MM-DD, p.ej. 2026-07-16)")
        sys.exit(2)


def _resumen(registros: list[dict]) -> None:
    facturas = defaultdict(int)
    codigos = defaultdict(int)
    total = 0
    for reg in registros:
        facturas[reg["CRNCXC"]] += 1
        codigos[reg["CRNCONOBJ"]] += 1
        total += reg["CROVALOBJ"]
    logger.info(f"  Facturas: {len(facturas)}  |  Objeciones: {len(registros)}")
    logger.info(f"  Valor glosado total: ${total:,.0f}")
    logger.info("  Códigos de objeción (CRNCONOBJ):")
    for cod, n in sorted(codigos.items(), key=lambda kv: (-kv[1], kv[0])):
        logger.info(f"    {cod or '(sin código)'}: {n}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        "--entrada",
        type=Path,
        required=True,
        help="Excel de glosas de SAVIA SALUD (8 columnas, como SAVIA_SALUD_8.03.xlsx).",
    )
    parser.add_argument(
        "--salida",
        type=Path,
        required=True,
        help="Carpeta destino (un archivo <prefijo>_<factura>.xlsx por factura), "
        "o archivo .xlsx si se usa --consolidado.",
    )
    parser.add_argument(
        "--prefijo",
        default=PREFIJO_DEFAULT,
        help=f"Prefijo del nombre de cada archivo por factura. Default: {PREFIJO_DEFAULT} "
        f"(→ {PREFIJO_DEFAULT}_HUS0000443697.xlsx).",
    )
    parser.add_argument(
        "--consolidado",
        action="store_true",
        help="En vez de un archivo por factura, junta todo en un solo Excel (--salida es el .xlsx).",
    )
    parser.add_argument(
        "--fecha",
        default=None,
        help="Fecha para CDFECDOC/CROFECOBJ (YYYY-MM-DD). Default: hoy.",
    )
    parser.add_argument(
        "--codigo-sufijo",
        default=CODIGO_SUFIJO_DEFAULT,
        help=f"Consecutivo con que se completa el código de 4→6 (TA08 → TA08XX). Default: {CODIGO_SUFIJO_DEFAULT}.",
    )
    parser.add_argument(
        "--mapa-codigos",
        type=Path,
        default=None,
        help='JSON opcional para forzar códigos: {"TA08": "TA0805", ...}.',
    )
    parser.add_argument(
        "--consecutivo",
        type=int,
        default=CDCONSEC_DEFAULT,
        help=f"Número inicial del consecutivo por factura (CDCONSEC). Default: {CDCONSEC_DEFAULT}. "
        "En el consolidado la 1ª factura lleva este valor, la 2ª +1, etc. (1-1-1, 2-2-2…).",
    )
    parser.add_argument("--log", type=Path, default=None, help="Guarda un log adicional a archivo.")
    args = parser.parse_args(argv)

    handlers: list[logging.Handler] = [logging.StreamHandler(sys.stdout)]
    if args.log is not None:
        args.log.parent.mkdir(parents=True, exist_ok=True)
        handlers.append(logging.FileHandler(args.log, encoding="utf-8"))
    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", handlers=handlers
    )

    if not args.entrada.is_file():
        logger.error(f"No existe el archivo de entrada: {args.entrada}")
        return 1

    fecha = _parse_fecha(args.fecha)
    mapa = _cargar_mapa(args.mapa_codigos)

    logger.info(f"Leyendo glosas de SAVIA: {args.entrada.name}")
    registros = construir_registros(
        args.entrada,
        fecha=fecha,
        consecutivo=args.consecutivo,
        codigo_sufijo=args.codigo_sufijo,
        mapa_codigos=mapa,
    )
    if not registros:
        logger.error("No se encontró ninguna objeción en el archivo de entrada.")
        return 1

    if args.consolidado:
        escribir_consolidado(registros, args.salida)
        logger.info(f"\nExcel de SAVIA (consolidado, formato de la guía): {args.salida}")
    else:
        generados = escribir_por_factura(
            registros, args.salida, prefijo=args.prefijo, consecutivo=args.consecutivo
        )
        logger.info(f"\n{len(generados)} archivo(s) de SAVIA en: {args.salida}")

    _resumen(registros)
    return 0


if __name__ == "__main__":
    sys.exit(main())
