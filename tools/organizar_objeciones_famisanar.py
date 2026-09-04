"""organizar_objeciones_famisanar.py — Arma las objeciones de FAMISANAR en el
formato de trabajo de 16 columnas (hoja OBJECIONES).

Toma el Excel de devoluciones/glosas que entrega FAMISANAR (export
"DEVYGLOSAS…": 4 columnas — NRO_FACTURA, CODIGO_DEVOLUCION, VALOR DEVOLUCION,
OBSERVACION) y lo convierte al mismo layout de 16 columnas que se usa para
SAVIA y el Dispensario:

    CDCONSEC | CDFECDOC | CRNCXC | CROFECOBJ | CROREFERE | CROOBSERV | CROCLAOBJ |
    CRNCLAOBJ | GENUSUARIO4 | CRNCONOBJ | SLNSERPRO | IDRIPS | CTNCENCOS |
    CROVALOBJ | CRDOBSERV | CROTIPOBJ

La particularidad de FAMISANAR: el Excel NO trae columna de código de servicio.
El código viene EMBEBIDO en el texto de la observación ("… CÓDIGO   903867 …",
"… CÓDIGO   U19965499-11 …"). Este bot lo EXTRAE del texto para llenar
SLNSERPRO. Las filas cortas tipo "AUD EXTRA - …" no traen código y quedan con
SLNSERPRO vacío (igual que las estancias en los archivos del Dispensario).

MAPEO DE CAMPOS (FAMISANAR → 16 columnas):
    CRNCXC       ← NRO_FACTURA          (HUS532670 → HUS0000532670, 10 dígitos)
    CRNCONOBJ    ← CODIGO_DEVOLUCION    (ya viene de 6: CL0801, CO0701, TA0801…)
    SLNSERPRO    ← extraído del texto ("CÓDIGO <x>") y HOMOLOGADO al código HUS:
                   CUPS tal cual; medicamentos U/P → se quita la letra
                   (U20162259-04 → 20162259-04); dispositivos 9101xxxx → código
                   FMQ del HUS (equivalencias fijas confirmadas, ver
                   MAPA_SERVICIOS_DEFAULT); los no mapeados quedan tal cual +
                   aviso (completar con --mapa-servicios)
    CROVALOBJ    ← VALOR DEVOLUCION
    CRDOBSERV    ← "<CRNCONOBJ> <OBSERVACION>$<valor>" (formato de trabajo)
    CDFECDOC / CROFECOBJ ← --fecha (default: hoy), en FECHA CORTA
    CDCONSEC     ← consecutivo POR FACTURA (1-1-1 la 1ª, 2-2-2 la 2ª, …), texto
    CROTIPOBJ    ← por factura: solo TA/FA/SO/AU/CO→0, solo CL→1, mezcla con CL→2
    CROCLAOBJ=0, GENUSUARIO4='999' (texto)  |  resto de columnas vacías

Las 7 reglas del formato son las mismas del bot de SAVIA
(`organizar_objeciones_savia.py`) — verificadas contra archivos reales
(OBJECIONES_DISPENSARIO_* y OBJECIONES_EMSSANAR_*). Este tool es autocontenido
a propósito (patrón del repo: un archivo por bot, ejecutable suelto).

USO:
    py organizar_objeciones_famisanar.py \
        --entrada "FAMISANAR_11.35.1.xlsx" \
        --salida  "OBJECIONES_FAMISANAR"      # carpeta destino (o .xlsx con --consolidado)

INSTALACIÓN (una vez):
    py -m pip install openpyxl
"""

from __future__ import annotations

import argparse
import datetime as _dt
import difflib
import json
import logging
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

# Un solo lector de pesos para todos los bots (tools/_dinero.py): la copia
# local multiplicaba por cien los valores con centavos.
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _dinero import a_entero  # noqa: E402

logger = logging.getLogger("organizar_famisanar")


# ─── Formato de salida: layout de 16 columnas (hoja OBJECIONES) ──────────────

COLUMNAS_DISPENSARIO: tuple[str, ...] = (
    "CDCONSEC",
    "CDFECDOC",
    "CRNCXC",
    "CROFECOBJ",
    "CROREFERE",
    "CROOBSERV",
    "CROCLAOBJ",
    "CRNCLAOBJ",
    "GENUSUARIO4",
    "CRNCONOBJ",
    "SLNSERPRO",
    "IDRIPS",
    "CTNCENCOS",
    "CROVALOBJ",
    "CRDOBSERV",
    "CROTIPOBJ",
)

# Constantes de la guía. CDCONSEC y GENUSUARIO4 van como TEXTO en los archivos
# reales; CROCLAOBJ/CROVALOBJ/CROTIPOBJ como número.
CDCONSEC_DEFAULT = 1
CROCLAOBJ_CONST = 0
GENUSUARIO4_CONST = "999"
CODIGO_SUFIJO_DEFAULT = "01"

# number_format por columna, copiado 1:1 de los archivos reales (EMSSANAR).
# 'mm-dd-yy' = fecha corta builtin de Excel (se ve dd/mm/yyyy según la config
# regional, sin horas); '@' = texto; CROVALOBJ con formato contable de miles.
FORMATOS_DISPENSARIO: dict[str, str] = {
    "CDCONSEC": "@",
    "CDFECDOC": "mm-dd-yy",
    "CRNCXC": "@",
    "CROFECOBJ": "mm-dd-yy",
    "CROREFERE": "@",
    "CROOBSERV": "@",
    "CROCLAOBJ": "General",
    "CRNCLAOBJ": "@",
    "GENUSUARIO4": "@",
    "CRNCONOBJ": "@",
    "SLNSERPRO": "@",
    "IDRIPS": "@",
    "CTNCENCOS": "@",
    "CROVALOBJ": '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-',
    "CRDOBSERV": "@",
    "CROTIPOBJ": "0",
}


# ─── Lectura del Excel de FAMISANAR (4 columnas) ─────────────────────────────


def _norm_header(h: object) -> str:
    """Normaliza un encabezado: mayúsculas, sin tildes, sin espacios de más."""
    s = unicodedata.normalize("NFKD", str(h or "").strip().upper())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.split())


# Alias de las columnas del export de FAMISANAR (tolerante a variaciones).
COLUMNAS_FAMISANAR = {
    "factura": {"NRO_FACTURA", "NRO FACTURA", "NUMERO_FACTURA", "NUMERO FACTURA", "FACTURA"},
    "codigo_devolucion": {
        "CODIGO_DEVOLUCION",
        "CODIGO DEVOLUCION",
        "CODIGO_GLOSA",
        "CODIGO GLOSA",
        "COD DEVOLUCION",
    },
    "valor": {"VALOR DEVOLUCION", "VALOR_DEVOLUCION", "VALOR GLOSA", "VALOR_GLOSA", "VALOR"},
    "observacion": {"OBSERVACION", "OBSERVACIONES", "OBSERVACION GLOSA", "DETALLE"},
}
# Índices fijos de respaldo (layout de FAMISANAR_11.35.1.xlsx).
IDX_FALLBACK = {
    "factura": 0,
    "codigo_devolucion": 1,
    "valor": 2,
    "observacion": 3,
}


def _resolver_columnas(headers: list[str]) -> dict[str, int]:
    """Devuelve {clave: indice} mapeando por nombre de encabezado; si una
    columna no se reconoce por nombre, usa el índice fijo de respaldo."""
    norm = [_norm_header(h) for h in headers]
    idx: dict[str, int] = {}
    for clave, alias in COLUMNAS_FAMISANAR.items():
        encontrado = next((i for i, h in enumerate(norm) if h in alias), None)
        if encontrado is None:
            encontrado = IDX_FALLBACK[clave]
            logger.debug(
                f"  columna '{clave}' no reconocida por nombre; uso índice fijo {encontrado}"
            )
        idx[clave] = encontrado
    return idx


def _cell(row: tuple, idx: dict[str, int], clave: str) -> object:
    """Valor de la columna `clave` en `row`, o None si el índice se sale de rango."""
    i = idx[clave]
    return row[i] if i < len(row) else None


def _num(v: object) -> int:
    """Pesos enteros, con el lector único de `tools/_dinero.py` (regla
    colombiana: un separador de miles lleva tres dígitos detrás; uno o dos son
    decimales — así "1.365,50" es 1365 y no 136550)."""
    return a_entero(v)


# ─── Normalización de factura (corta → larga) ────────────────────────────────

_RE_FACTURA = re.compile(r"^([A-Za-z]+)0*(\d+)$")


def factura_larga(fac: object, ancho: int = 10) -> str:
    """HUS532670 → HUS0000532670 (rellena con ceros a `ancho` dígitos).
    Idempotente. Si no matchea el patrón, devuelve el texto tal cual."""
    s = str(fac or "").strip()
    m = _RE_FACTURA.match(s)
    if not m:
        return s
    return m.group(1).upper() + m.group(2).zfill(ancho)


# ─── Código de objeción ──────────────────────────────────────────────────────


def codigo_objecion(codigo: str, sufijo: str = CODIGO_SUFIJO_DEFAULT, mapa=None) -> str:
    """FAMISANAR ya entrega el código de 6 caracteres (CL0801, CO0701…): se
    deja tal cual. Red de seguridad: si viniera de 4 (grupo+concepto) se
    completa con el `sufijo`, y `mapa` (--mapa-codigos) fuerza equivalencias
    puntuales — misma semántica que el bot de SAVIA."""
    cod = (codigo or "").strip().upper()
    if mapa and cod in mapa:
        return mapa[cod]
    if re.fullmatch(r"[A-Z]{2}\d{2}", cod):
        return cod + sufijo
    return cod


# ─── Extracción del código de servicio desde la observación ──────────────────

# En los textos de FAMISANAR el código del servicio/insumo viene como
# "… CÓDIGO   903867 SE RECONOCE …" / "… CÓDIGO   U19965499-11 VALOR …".
# Acepta dígitos, letras y guiones (91017235, U19965499-11, P32606-02).
_RE_COD_SERVICIO = re.compile(r"C[OÓ]DIGO\s+([A-Za-z0-9][A-Za-z0-9.\-]*)", re.IGNORECASE)


def extraer_cod_servicio(observacion: str) -> str:
    """Extrae el código de servicio embebido en el texto de la objeción.
    Devuelve '' si el texto no trae código (p. ej. filas 'AUD EXTRA - …').

    Cuando FAMISANAR deja la etiqueta CÓDIGO vacía —"… CATETER INTRAVENOSO 20
    CÓDIGO    VALOR UNITARIO FACTURADO…"— lo que sigue es la frase siguiente,
    no un código: todo código de servicio trae al menos un dígito, así que sin
    dígitos se descarta (antes entraba la palabra VALOR a SLNSERPRO)."""
    m = _RE_COD_SERVICIO.search(observacion or "")
    if not m:
        return ""
    cod = m.group(1).strip(".-")
    return cod if any(c.isdigit() for c in cod) else ""


# ─── Homologación: código FAMISANAR → código HUS ─────────────────────────────

# Los códigos que FAMISANAR escribe en sus textos NO son (todos) los del HUS:
#   - CUPS de 6 dígitos (890202, 903867, 735301…): son el estándar nacional,
#     los mismos que usa el HUS → tal cual.
#   - Medicamentos con LETRA adelante (U20162259-04, P32606-02…): la letra es
#     de FAMISANAR; sin ella queda el código del HUS. Verificado contra los
#     archivos de trabajo: U20162259-04 → 20162259-04 (METOCLOPRAMIDA, match
#     exacto en OBJECIONES_EMSSANAR) y P32606-02 → raíz 32606 (SODIO LACTATO).
#   - Dispositivos 9101xxxx (catéteres, llaves, electrodos…): catálogo propio
#     de FAMISANAR sin equivalencia conocida → se dejan tal cual y se reportan,
#     hasta tener el maestro (cargarlo vía --mapa-servicios).
_RE_MED_CON_LETRA = re.compile(r"^[A-Za-z](\d[\dA-Za-z.\-]*)$")

# Equivalencias FIJAS dispositivo FAMISANAR → código FMQ del HUS, confirmadas
# por la auditora contra el archivo de trabajo OBJECIONES_LOTE_02 (los insumos
# del HUS van con código FMQ). Evidencia de cada una:
#   91017235 → FMQ0112   CATETER INTRAVENOSO 18 — valor $5.800 idéntico
#   91012136 → FMQ0182-1 LLAVES DE TRES VIAS — nombre exacto (único)
#   91017424 → FMQ0952   ELECTRODO ECG ADULTO — 3×$800 = $2.400 idéntico
#   91017278 → FMQ0159   BOLSA RECOLECTORA DE ORINA ADULTO — $18.100 idéntico
# Un --mapa-servicios de la CLI puede AGREGAR equivalencias o PISAR estas.
MAPA_SERVICIOS_DEFAULT: dict[str, str] = {
    "91017235": "FMQ0112",
    "91012136": "FMQ0182-1",
    "91017424": "FMQ0952",
    "91017278": "FMQ0159",
}


def homologar_cod_servicio(cod: str, mapa: dict[str, str] | None = None) -> tuple[str, str]:
    """Devuelve (codigo_homologado, regla_aplicada). Reglas, en orden:
    'mapa' (equivalencias fijas + las de --mapa-servicios, que pisan a las
    fijas) → 'letra' (quita la letra inicial de medicamentos) → 'igual'
    (CUPS y demás, tal cual)."""
    c = (cod or "").strip()
    if not c:
        return "", "vacio"
    combinado = {**MAPA_SERVICIOS_DEFAULT, **(mapa or {})}
    if c in combinado:
        return combinado[c], "mapa"
    m = _RE_MED_CON_LETRA.match(c)
    if m:
        return m.group(1), "letra"
    return c, "igual"


# ─── Cruce contra los servicios facturados del DGH ───────────────────────────
#
# El código que FAMISANAR escribe en el texto NO siempre es el del hospital:
# los dispositivos van con nomenclatura IUM (91022534) y los medicamentos con
# el sufijo rellenado con ceros (P32606-02 donde el HUS tiene 32606-2). Sin el
# código del HUS, DGH no reconoce el renglón.
#
# Con el export de servicios facturados del DGH (--servicios-dgh) se busca, en
# ESA factura, de qué servicio habla la objeción: código, nombre y valor. Lo
# que se escribe en SLNSERPRO es entonces el código real del hospital, y de
# paso se llena CTNCENCOS (centro de costo), que hasta ahora salía vacío.
#
# Regla que no se rompe: si el cruce no es confiable NO se inventa un servicio
# — queda lo que se sabía por el texto y el renglón se reporta para revisión.

COLUMNAS_DGH_SERVICIOS = {
    "codigo": {"SERVICIOS DGH", "SERVICIOS_DGH", "SLNSERPRO SERVICIO", "SLNSERPRO_SERVICIO"},
    "desc_institucional": {"DESCRIPCION INSTITUCIONAL", "DESCRIPCION_INSTITUCIONAL"},
    "cups": {"SLNSERPRO CUPS", "SLNSERPRO_CUPS"},
    "desc_cups": {"DESCRIPCION CUPS", "DESCRIPCION_CUPS"},
    "cod_medicamento": {"CODIGO MEDICAMENTO", "CODIGO_MEDICAMENTO"},
    "nombre_medicamento": {"NOMBRE MEDICAMENTO", "NOMBRE_MEDICAMENTO"},
    "centro_costo": {"NOM CENTRO COSTO", "NOM_CENTRO_COSTO", "CENTRO COSTO", "CENTRO_COSTO"},
    "factura": {"FACTURA", "NRO_FACTURA", "NRO FACTURA", "NUMERO FACTURA"},
    "cantidad": {"CAT SERVICIOS", "CAT_SERVICIOS", "CANTIDAD", "CANT"},
    "valor": {"VR SERVICIO", "VR_SERVICIO", "VALOR SERVICIO"},
}

# Palabras que no distinguen un servicio de otro.
_VACIAS_DESC = frozenset(
    {
        "DE",
        "DEL",
        "LA",
        "EL",
        "LOS",
        "LAS",
        "POR",
        "CON",
        "SIN",
        "PARA",
        "EN",
        "A",
        "Y",
        "O",
        "X",
        "AL",
        "UN",
        "UNA",
        "SU",
        "MAS",
        "REF",
        "TIPO",
    }
)

_RE_NO_ALFA = re.compile(r"[^A-Z0-9 ]+")
_RE_NUM_LETRA = re.compile(r"(?<=\d)(?=[A-Z])")
_RE_LETRA_NUM = re.compile(r"(?<=[A-Z])(?=\d)")


def norm_desc(texto: object) -> str:
    """MAYÚSCULAS, sin tildes, sin puntuación, espacios colapsados."""
    if texto is None:
        return ""
    t = _norm_header(texto)
    return " ".join(_RE_NO_ALFA.sub(" ", t).split())


def palabras_desc(desc: str) -> set[str]:
    """Palabras significativas, separando número y unidad pegados (1ML → 1 ML),
    para que 'JERINGA DESECHABLE 1ML' del DGH y 'JERINGA 1 ML CON AGUJA' de
    FAMISANAR compartan las mismas palabras."""
    if not desc:
        return set()
    t = _RE_LETRA_NUM.sub(" ", _RE_NUM_LETRA.sub(" ", desc))
    return {w for w in t.split() if w and w not in _VACIAS_DESC}


def parecido_desc(a: str, b: str) -> float:
    """0..1 — parecido entre dos nombres de servicio ya normalizados."""
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    ratio = difflib.SequenceMatcher(None, a, b).ratio()
    if a in b or b in a:
        ratio = max(ratio, 0.90)
    ta, tb = palabras_desc(a), palabras_desc(b)
    if ta and tb:
        comunes = ta & tb
        # Sin ninguna palabra de fondo en común, el parecido letra a letra es
        # ruido ('VITAMINA D3 CAP X 1 000' vs. 'JERINGA 1ML 25G'): se topa.
        if not {w for w in comunes if len(w) >= 3 and not w.isdigit()}:
            ratio = min(ratio, 0.30)
        ratio = max(ratio, len(comunes) / len(ta | tb) * 0.95)
        # Cobertura: qué tanto del nombre más corto está dentro del más largo.
        # FAMISANAR antepone la categoría ('LINEA INFUSION E INYECCION - …') y
        # el DGH usa el nombre pelado.
        corto = min(len(ta), len(tb))
        cobertura = len(comunes) / corto if corto else 0.0
        if corto < 2 or len(comunes) < 2:
            cobertura = min(cobertura, 0.55)
        ratio = max(ratio, cobertura * 0.85)
    return ratio


def variantes_codigo(cod: object) -> set[str]:
    """Todas las formas en que la misma referencia puede venir escrita:
    la letra que antepone FAMISANAR (P32606-02), la H que agrega el DGH al
    CUPS (903437H) y el sufijo rellenado con ceros (32606-02 = 32606-2)."""
    crudo = _norm_header(cod)
    base = re.sub(r"[^A-Z0-9]", "", crudo)
    if not base:
        return set()
    v = {base}
    if len(base) > 1 and base[0] in ("P", "U") and base[1].isdigit():
        v.add(base[1:])
    if base.endswith("H") and base[:-1].isdigit():
        v.add(base[:-1])
    m = re.match(r"^([PU]?)([A-Z0-9]+)-0*(\d+)$", crudo)
    if m:
        v.add(re.sub(r"[^A-Z0-9]", "", f"{m.group(2)}{m.group(3)}"))
        v.add(re.sub(r"[^A-Z0-9]", "", f"{m.group(1)}{m.group(2)}{m.group(3)}"))
    return {x for x in v if x}


class LineaDgh:
    """Un renglón del export de servicios facturados del DGH."""

    __slots__ = (
        "codigo",
        "descripcion",
        "descripciones",
        "codigos",
        "centro_costo",
        "cantidad",
        "valor",
        "unitario",
        "usos",
    )

    def __init__(
        self, codigo, descripcion, desc_cups, nombre_med, cups, cod_med, centro, cant, valor
    ):
        self.codigo = codigo
        self.descripcion = descripcion or desc_cups or nombre_med
        self.descripciones = {norm_desc(d) for d in (descripcion, desc_cups, nombre_med) if d}
        self.codigos: set[str] = set()
        for c in (codigo, cups, cod_med):
            self.codigos |= variantes_codigo(c)
        self.centro_costo = centro
        self.cantidad = cant
        self.valor = valor
        self.unitario = (valor / cant) if cant else valor
        self.usos = 0


def leer_servicios_dgh(ruta: Path) -> dict[str, list[LineaDgh]]:
    """{factura: [servicios facturados]} desde el export del DGH."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(ruta), data_only=True, read_only=True)
    try:
        ws = wb.active
        filas = ws.iter_rows(values_only=True)
        headers = list(next(filas, ()) or ())
        norm = [_norm_header(h) for h in headers]
        idx: dict[str, int] = {}
        for clave, alias in COLUMNAS_DGH_SERVICIOS.items():
            i = next((n for n, h in enumerate(norm) if h in alias), None)
            if i is not None:
                idx[clave] = i
        if "factura" not in idx or "valor" not in idx:
            raise ValueError(
                f"{ruta.name} no parece el export de servicios del DGH: no encontré "
                "las columnas FACTURA y Vr_SERVICIO."
            )

        def dato(fila, clave):
            i = idx.get(clave)
            return fila[i] if i is not None and i < len(fila) else None

        fuera: dict[str, list[LineaDgh]] = defaultdict(list)
        for fila in filas:
            if fila is None:
                continue
            factura = str(dato(fila, "factura") or "").strip()
            if not factura:
                continue
            fuera[factura_larga(factura)].append(
                LineaDgh(
                    codigo=str(dato(fila, "codigo") or "").strip(),
                    descripcion=str(dato(fila, "desc_institucional") or "").strip(),
                    desc_cups=str(dato(fila, "desc_cups") or "").strip(),
                    nombre_med=str(dato(fila, "nombre_medicamento") or "").strip(),
                    cups=str(dato(fila, "cups") or "").strip(),
                    cod_med=str(dato(fila, "cod_medicamento") or "").strip(),
                    centro=str(dato(fila, "centro_costo") or "").strip(),
                    cant=_num(dato(fila, "cantidad")),
                    valor=_num(dato(fila, "valor")),
                )
            )
        return dict(fuera)
    finally:
        wb.close()


# Lo que se puede leer del texto de la objeción además del código: el nombre
# del servicio y el valor unitario que FAMISANAR dice haber facturado.
_RE_DESC_TARIFA = re.compile(
    r"PARA\s+EL\s+SERVICIO\s+(?P<desc>.+?)\s+C[ÓO]DIGO", re.IGNORECASE | re.DOTALL
)
_RE_DESC_COBERTURA = re.compile(
    r"SERVICIO\s+SIN\s+COBERTURA\s+(?P<desc>.+?)\s*C[ÓO]DIGO", re.IGNORECASE | re.DOTALL
)
_RE_DESC_AUTORIZACION = re.compile(
    r"(?:\bPOS\b|NO\s+POS)\s+(?P<desc>.+?)\s*C[ÓO]DIGO", re.IGNORECASE | re.DOTALL
)
_RE_DESC_CANTIDAD = re.compile(
    r"CANT\.?\s+DE\s+(?P<desc>.+?)\s+C[ÓO]DIGO", re.IGNORECASE | re.DOTALL
)
_RE_DESC_AUD_EXTRA = re.compile(
    r"SE\s+(?:OBJETAN?|RECONOCE)\s+(?P<desc>.+?)"
    r"(?=\s+(?:CANT|CNT|CAN)\b|\s+C[ÓO]DIGO\b|\s{3,}|[,\.]\s|$)",
    re.IGNORECASE | re.DOTALL,
)
_RE_UNITARIO = re.compile(
    r"VALOR\s+(?:UNITARIO\s+)?FACTURADO\s+POR\s+(?:LA\s+)?IPS\s*(?:DE)?\s*\$?\s*([\d\.,]+)",
    re.IGNORECASE,
)
# Código pegado adelante del nombre ("FMQ0113 CATETER INTRAVENOSO 20") y
# código IUM largo pegado adelante ("1O1044511000101 OXIGENO").
_RE_COD_PEGADO = re.compile(r"^(?P<cod>[A-Z]{2,4}\d[A-Z0-9\-\.]*)\s+(?P<resto>.{4,})$")
_RE_COD_LARGO_PEGADO = re.compile(r"^[0-9A-Z]{7,}\s+(?=[A-Z])")


def descripcion_del_texto(observacion: str) -> str:
    """El nombre del servicio tal como lo escribió FAMISANAR, sacado del texto."""
    t = " ".join((observacion or "").split())
    if not t:
        return ""
    for rx in (
        _RE_DESC_TARIFA,
        _RE_DESC_COBERTURA,
        _RE_DESC_AUTORIZACION,
        _RE_DESC_CANTIDAD,
        _RE_DESC_AUD_EXTRA,
    ):
        m = rx.search(t)
        if not m:
            continue
        desc = " ".join(m.group("desc").split()).strip(" .-")
        pegado = _RE_COD_PEGADO.match(desc)
        if pegado:
            desc = pegado.group("resto").strip()
        return desc[:300]
    return ""


def valor_unitario_del_texto(observacion: str) -> int:
    """El valor unitario que FAMISANAR declara haber facturado (0 si no viene)."""
    m = _RE_UNITARIO.search(" ".join((observacion or "").split()))
    return _num(m.group(1)) if m else 0


def codigo_servicio_del_texto(observacion: str) -> str:
    """El código del servicio, venga donde venga: detrás de la etiqueta CÓDIGO
    o pegado adelante del nombre ("… SERVICIO SIN COBERTURA  FMQ0113 CATETER
    INTRAVENOSO 20  CÓDIGO   VALOR UNITARIO…", donde la etiqueta va vacía)."""
    cod = extraer_cod_servicio(observacion)
    if cod:
        return cod
    t = " ".join((observacion or "").split())
    for rx in (_RE_DESC_COBERTURA, _RE_DESC_TARIFA, _RE_DESC_AUTORIZACION, _RE_DESC_CANTIDAD):
        m = rx.search(t)
        if not m:
            continue
        cuerpo = " ".join(m.group("desc").split()).strip(" .-")
        pegado = _RE_COD_PEGADO.match(cuerpo)
        if pegado:
            return pegado.group("cod")
        # "DUN0018" solo: el nombre ES el código.
        if re.fullmatch(r"[A-Z]{2,5}\d[A-Z0-9\-\.]*", cuerpo.upper()):
            return cuerpo.upper()
        break
    return ""


_RE_CANTIDAD_TEXTO = re.compile(r"DE\s+(\d+)\s*UNIDAD\(ES\)", re.IGNORECASE)
_RE_CANTIDAD_FACT = re.compile(r"C(?:ANT|NT)\s*F\w*\s*(\d+)", re.IGNORECASE)


def cantidad_del_texto(observacion: str) -> int:
    """Las unidades que FAMISANAR dice haber objetado (0 si no vienen)."""
    t = " ".join((observacion or "").split())
    m = _RE_CANTIDAD_TEXTO.search(t) or _RE_CANTIDAD_FACT.search(t)
    if not m:
        return 0
    try:
        return int(m.group(1))
    except ValueError:
        return 0


def formas_descripcion(desc: str) -> list[str]:
    """Formas alternativas de leer el nombre que escribió FAMISANAR: completo,
    sin el código IUM pegado adelante y sólo lo que va después del guion
    ('LINEA INFUSION E INYECCION - JERINGA 5 ML' → 'JERINGA 5 ML')."""
    base = norm_desc(desc)
    if not base:
        return []
    formas = [base]
    sin_cod = _RE_COD_LARGO_PEGADO.sub("", base)
    if sin_cod and sin_cod != base:
        formas.append(sin_cod)
    if " - " in (desc or ""):
        cola = norm_desc(desc.split(" - ")[-1])
        if cola and cola not in formas:
            formas.append(cola)
    return formas


TOLERANCIA_PESOS = 0.5
PUNTAJE_MINIMO = 3.0
UMBRAL_ALTA = 6.0
UMBRAL_MEDIA = 4.5


class Cruce:
    """Resultado de buscar el servicio de una objeción en el export del DGH."""

    __slots__ = ("linea", "confianza", "motivos", "puntaje", "aviso")

    def __init__(self, linea=None, confianza="SIN CRUCE", motivos="", puntaje=0.0, aviso=""):
        self.linea = linea
        self.confianza = confianza
        self.motivos = motivos
        self.puntaje = puntaje
        self.aviso = aviso


def _mismo_valor(a: float, b: float) -> bool:
    return abs(a - b) < TOLERANCIA_PESOS


def _puntuar(
    linea: LineaDgh,
    cods: set[str],
    formas: list[str],
    unitario: int,
    valor: int,
    cantidad: int,
    ctx: dict,
):
    """Qué tanto esta línea del DGH se parece a lo que dice la objeción."""
    puntos = 0.0
    motivos: list[str] = []

    if cods and linea.codigos:
        if cods & linea.codigos:
            puntos += 3.0
            motivos.append("código")
        elif any(a.startswith(b) or b.startswith(a) for a in cods for b in linea.codigos):
            puntos += 1.5
            motivos.append("código parcial")

    mejor_desc = 0.0
    for forma in formas:
        for d in linea.descripciones:
            mejor_desc = max(mejor_desc, parecido_desc(forma, d))
    if mejor_desc >= 0.99:
        puntos += 3.0
        motivos.append("nombre exacto")
    elif mejor_desc >= 0.55:
        puntos += 3.0 * mejor_desc
        motivos.append(f"nombre {mejor_desc:.0%}")

    puntos_valor = 0.0
    if unitario and _mismo_valor(linea.unitario, unitario):
        puntos_valor = 2.0
        motivos.append("valor unitario")
    if valor and _mismo_valor(linea.valor, valor):
        puntos_valor = max(puntos_valor, 2.0)
        motivos.append("valor del renglón")
    if unitario and _mismo_valor(linea.valor, unitario):
        puntos_valor = max(puntos_valor, 1.5)
        motivos.append("valor")
    puntos += puntos_valor

    if cantidad and linea.cantidad and _mismo_valor(linea.cantidad, cantidad):
        puntos += 0.5
        motivos.append("cantidad")

    # Un valor que en toda la factura sólo lo tiene UN servicio identifica el
    # renglón aunque FAMISANAR use otro código y otro nombre.
    if puntos_valor:
        if ctx.get("codigos_con_ese_valor") == 1:
            puntos += 1.5
            motivos.append("valor único en la factura")
        elif ctx.get("codigo_preferido") and linea.codigo == ctx["codigo_preferido"]:
            puntos += 1.5
            motivos.append("valor + nombre más parecido")

    # Código igual y nombre distinto: si el valor coincide es la misma línea y
    # FAMISANAR escribió mal el nombre (busca el código en CUPS y no en el
    # catálogo del hospital); si el valor no coincide, no es el mismo servicio.
    if "código" in motivos and formas and mejor_desc < 0.35:
        if puntos_valor:
            motivos.append("¡el nombre no concuerda!")
        else:
            puntos -= 2.5
            motivos.append("¡el nombre no concuerda!")
    return puntos, motivos


def resolver_servicio(
    cod_texto: str, observacion: str, valor: int, lineas: list[LineaDgh]
) -> Cruce:
    """Busca en las líneas del DGH de ESA factura el servicio del que habla la
    objeción. Si no hay un candidato con puntaje suficiente devuelve un cruce
    vacío: nunca escribe un servicio del que no está seguro."""
    if not lineas:
        return Cruce(aviso="la factura no está en el export del DGH")

    cods = variantes_codigo(cod_texto)
    unitario = valor_unitario_del_texto(observacion)
    cantidad = cantidad_del_texto(observacion)
    formas = formas_descripcion(descripcion_del_texto(observacion))

    def coincide_valor(linea: LineaDgh) -> bool:
        return bool(
            (unitario and _mismo_valor(linea.unitario, unitario))
            or (valor and _mismo_valor(linea.valor, valor))
        )

    con_ese_valor = [x for x in lineas if coincide_valor(x)]
    ctx = {"codigos_con_ese_valor": len({x.codigo for x in con_ese_valor})}
    # Si varias líneas comparten el valor, el nombre desempata.
    ctx["codigo_preferido"] = ""
    if ctx["codigos_con_ese_valor"] > 1 and formas:
        por_codigo: dict[str, float] = {}
        for linea in con_ese_valor:
            for forma in formas:
                for d in linea.descripciones:
                    r = parecido_desc(forma, d)
                    if r > por_codigo.get(linea.codigo, 0.0):
                        por_codigo[linea.codigo] = r
        orden = sorted(por_codigo.items(), key=lambda kv: kv[1], reverse=True)
        if orden and orden[0][1] >= 0.30 and (len(orden) == 1 or orden[0][1] - orden[1][1] >= 0.15):
            ctx["codigo_preferido"] = orden[0][0]

    mejor, mejor_pts, mejor_motivos = None, 0.0, []
    for linea in lineas:
        pts, motivos = _puntuar(linea, cods, formas, unitario, valor, cantidad, ctx)
        # A igualdad de puntaje, preferir una línea todavía no usada: FAMISANAR
        # manda una objeción por unidad y el DGH tiene un renglón por unidad.
        pts_desempate = pts - min(linea.usos, 3) * 0.15
        if pts_desempate > mejor_pts:
            mejor, mejor_pts, mejor_motivos = linea, pts_desempate, motivos

    if mejor is None or mejor_pts < PUNTAJE_MINIMO:
        return Cruce(
            motivos=", ".join(mejor_motivos),
            puntaje=round(mejor_pts, 2),
            aviso="no se identificó el servicio: completar a mano",
        )

    mejor.usos += 1
    confianza = (
        "ALTA" if mejor_pts >= UMBRAL_ALTA else "MEDIA" if mejor_pts >= UMBRAL_MEDIA else "BAJA"
    )
    aviso = ""
    if "¡el nombre no concuerda!" in mejor_motivos:
        confianza = "MEDIA" if confianza == "ALTA" else confianza
        aviso = "el nombre del servicio en el archivo de FAMISANAR no coincide con el del DGH"
    elif confianza == "BAJA":
        aviso = "cruce débil: verificar antes de subir"
    return Cruce(mejor, confianza, ", ".join(mejor_motivos), round(mejor_pts, 2), aviso)


# ─── CROTIPOBJ: tipo de objeción por factura ─────────────────────────────────

GRUPO_CLINICO = "CL"


def crotipobj_factura(grupos: set[str]) -> int:
    """Tipo de objeción de la factura según sus grupos de concepto (2 letras):
    solo administrativos (TA/FA/SO/AU/CO…) → 0; solo CL → 1; mezcla con CL → 2.
    """
    tiene_cl = GRUPO_CLINICO in grupos
    tiene_admin = any(g != GRUPO_CLINICO for g in grupos)
    if tiene_cl and tiene_admin:
        return 2
    if tiene_cl:
        return 1
    return 0


# ─── Texto de la observación en el formato de trabajo ────────────────────────

_RE_VALOR_FINAL = re.compile(r"\$\s*([\d.,]+)\s*$")


def construir_crdobserv(codigo: str, observacion: str, valor: int) -> str:
    """CRDOBSERV = ``<código> <texto>$<valor>`` (formato de los archivos
    reales). Normaliza las corridas largas de espacios que traen los exports
    de FAMISANAR. Anti-duplicado con cuidado: quita el código del inicio si ya
    viene, y un ``$<monto>`` final SOLO si es el MISMO valor de la objeción.
    Si el monto final es otro (p. ej. "VALOR UNITARIO FACTURADO POR IPS
    $ 244,800" cuando la objeción es 207100), se CONSERVA — es información
    real del texto, no un duplicado (hallazgo de la revisión adversarial:
    18/37 filas del archivo real perdían el valor unitario)."""
    t = " ".join((observacion or "").split())
    cod = (codigo or "").strip()
    if cod and t.upper().startswith(cod.upper()):
        t = t[len(cod) :].strip()
    m = _RE_VALOR_FINAL.search(t)
    if m and _num(m.group(1)) == valor:
        t = t[: m.start()].strip()
    prefijo = f"{cod} " if cod else ""
    return f"{prefijo}{t}${valor}"


# ─── Transformación FAMISANAR → 16 columnas ──────────────────────────────────


def construir_registros(
    ruta: Path,
    fecha: _dt.datetime,
    consecutivo: int,
    codigo_sufijo: str,
    mapa_codigos: dict[str, str] | None,
    mapa_servicios: dict[str, str] | None = None,
    homologar: bool = True,
    servicios_dgh: dict[str, list[LineaDgh]] | None = None,
    trazas: list[dict] | None = None,
) -> list[dict]:
    """Lee el Excel de FAMISANAR y devuelve una lista de dicts, uno por
    objeción, con las 16 columnas del formato de trabajo. Con `homologar`
    (default), el código de servicio extraído del texto se convierte al del
    HUS (ver homologar_cod_servicio); `mapa_servicios` agrega equivalencias
    explícitas (p. ej. los dispositivos 9101xxxx).

    Con `servicios_dgh` (el export de servicios facturados del DGH) se busca
    además, factura por factura, de qué servicio habla cada objeción: cuando
    el cruce es confiable, SLNSERPRO queda con el código REAL del hospital y
    CTNCENCOS con su centro de costo. Si no lo es, se deja lo que se sabía por
    el texto y el renglón queda marcado para revisión.

    `trazas`, si se pasa, recibe un dict por objeción con el detalle del cruce
    (qué se leyó del texto, con qué renglón del DGH cruzó y por qué). Va aparte
    para que cada registro conserve exactamente las 16 columnas del formato."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.stderr.write("ERROR: falta openpyxl. py -m pip install openpyxl\n")
        sys.exit(2)

    wb = load_workbook(filename=str(ruta), data_only=True, read_only=True)
    ws = wb.active
    filas = ws.iter_rows(values_only=True)
    try:
        headers = list(next(filas))
    except StopIteration:
        logger.warning(f"  {ruta.name}: hoja vacía")
        return []
    idx = _resolver_columnas([str(h) for h in headers])

    consec_por_factura: dict[str, int] = {}
    sin_codigo = 0
    cruces: dict[str, int] = defaultdict(int)
    homologados: dict[str, int] = defaultdict(int)
    sin_regla: dict[str, str] = {}

    registros: list[dict] = []
    for r in filas:
        if r is None:
            continue
        factura_raw = str(_cell(r, idx, "factura") or "").strip()
        observacion = str(_cell(r, idx, "observacion") or "").strip()
        cod_dev = str(_cell(r, idx, "codigo_devolucion") or "").strip()
        if not factura_raw and not observacion and not cod_dev:
            continue

        crncxc = factura_larga(factura_raw)
        if crncxc not in consec_por_factura:
            consec_por_factura[crncxc] = consecutivo + len(consec_por_factura)
        codigo = codigo_objecion(cod_dev, codigo_sufijo, mapa_codigos)
        valor = _num(_cell(r, idx, "valor"))
        cod_servicio = codigo_servicio_del_texto(observacion)
        cod_famisanar, regla = "", ""
        if not cod_servicio:
            sin_codigo += 1
        elif homologar:
            cod_famisanar = cod_servicio
            cod_servicio, regla = homologar_cod_servicio(cod_famisanar, mapa_servicios)
            homologados[regla] += 1

        # Cruce contra los servicios que el DGH tiene facturados en ESA factura.
        cruce = Cruce()
        centro_costo = None
        if servicios_dgh is not None:
            cruce = resolver_servicio(
                cod_famisanar or cod_servicio, observacion, valor, servicios_dgh.get(crncxc, [])
            )
            cruces[cruce.confianza] += 1
            if cruce.linea is not None:
                # Manda el código del hospital: es el que DGH reconoce.
                cod_servicio = cruce.linea.codigo or cod_servicio
                centro_costo = cruce.linea.centro_costo or None
        # Dispositivos FAMISANAR (9101xxxx…) sin equivalencia HUS: se reportan
        # para completar el mapa con el maestro, pero sólo los que el cruce
        # contra el DGH tampoco pudo ubicar — con el export a la mano la
        # mayoría se resuelve sola y no hay nada que cargar al mapa.
        if (
            cod_famisanar
            and regla == "igual"
            and not re.fullmatch(r"\d{6}", cod_famisanar)
            and cruce.linea is None
        ):
            sin_regla[cod_famisanar] = observacion[:60]

        registros.append(
            {
                "CDCONSEC": str(consec_por_factura[crncxc]),
                "CDFECDOC": fecha,
                "CRNCXC": crncxc,
                "CROFECOBJ": fecha,
                "CROREFERE": None,
                "CROOBSERV": None,
                "CROCLAOBJ": CROCLAOBJ_CONST,
                "CRNCLAOBJ": None,
                "GENUSUARIO4": GENUSUARIO4_CONST,
                "CRNCONOBJ": codigo,
                "SLNSERPRO": cod_servicio or None,
                "IDRIPS": None,
                "CTNCENCOS": centro_costo,
                "CROVALOBJ": valor,
                "CRDOBSERV": construir_crdobserv(codigo, observacion, valor),
                "CROTIPOBJ": 0,  # placeholder: se calcula por factura abajo
            }
        )
        if trazas is not None:
            trazas.append(
                {
                    "factura": crncxc,
                    "codigo_objecion": codigo,
                    "valor": valor,
                    "cod_famisanar": codigo_servicio_del_texto(observacion),
                    "desc_famisanar": descripcion_del_texto(observacion),
                    "unitario_famisanar": valor_unitario_del_texto(observacion),
                    "cod_dgh": cruce.linea.codigo if cruce.linea else "",
                    "servicio_dgh": cruce.linea.descripcion if cruce.linea else "",
                    "unitario_dgh": cruce.linea.unitario if cruce.linea else "",
                    "centro_costo": centro_costo or "",
                    "confianza": cruce.confianza,
                    "motivos": cruce.motivos,
                    "puntaje": cruce.puntaje,
                    "aviso": cruce.aviso,
                    "observacion": observacion,
                }
            )
    wb.close()

    # CROTIPOBJ por factura según la mezcla de conceptos.
    grupos_por_factura: dict[str, set[str]] = defaultdict(set)
    for reg in registros:
        grupos_por_factura[reg["CRNCXC"]].add(str(reg["CRNCONOBJ"])[:2].upper())
    for reg in registros:
        reg["CROTIPOBJ"] = crotipobj_factura(grupos_por_factura[reg["CRNCXC"]])

    if sin_codigo:
        logger.info(
            f"  ⚠ {sin_codigo} objeciones sin código de servicio en el texto "
            "(filas 'AUD EXTRA'): SLNSERPRO queda vacío."
        )
    if homologar and homologados:
        partes = []
        if homologados.get("igual"):
            partes.append(f"{homologados['igual']} tal cual (CUPS/otros)")
        if homologados.get("letra"):
            partes.append(f"{homologados['letra']} con letra FAMISANAR quitada (U/P…)")
        if homologados.get("mapa"):
            partes.append(f"{homologados['mapa']} por mapa de equivalencias")
        logger.info(f"  Homologación de códigos de servicio: {', '.join(partes)}.")
    if servicios_dgh is not None:
        total = sum(cruces.values()) or 1
        ubicados = cruces["ALTA"] + cruces["MEDIA"]
        logger.info(
            f"  Cruce contra los servicios del DGH: {ubicados} de {total} servicios "
            f"ubicados con confianza alta/media ({ubicados / total:.0%})."
        )
        logger.info(
            f"    ALTA={cruces['ALTA']}  MEDIA={cruces['MEDIA']}  BAJA={cruces['BAJA']}  "
            f"SIN CRUCE={cruces['SIN CRUCE']}  → revisar {cruces['BAJA'] + cruces['SIN CRUCE']}."
        )
    if sin_regla:
        logger.warning(
            f"  ⚠ {len(sin_regla)} códigos de FAMISANAR SIN equivalencia HUS conocida "
            "(dispositivos): quedan tal cual. Completalos con --mapa-servicios:"
        )
        for cod, desc in sorted(sin_regla.items()):
            logger.warning(f"      {cod}  ← {desc}…")
    return registros


# ─── Escritura ───────────────────────────────────────────────────────────────


def _escribir_hoja(registros: list[dict], salida: Path) -> None:
    """Escribe UN .xlsx con hoja OBJECIONES y las 16 columnas del formato."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "OBJECIONES"

    fill_hdr = PatternFill("solid", fgColor="1F4E78")
    font_hdr = Font(bold=True, color="FFFFFF")
    for col, nombre in enumerate(COLUMNAS_DISPENSARIO, start=1):
        c = ws.cell(row=1, column=col, value=nombre)
        c.fill = fill_hdr
        c.font = font_hdr

    fila = 2
    for reg in registros:
        for col, nombre in enumerate(COLUMNAS_DISPENSARIO, start=1):
            celda = ws.cell(row=fila, column=col, value=reg.get(nombre))
            celda.number_format = FORMATOS_DISPENSARIO[nombre]
        fila += 1

    ws.freeze_panes = "A2"
    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(salida))


PREFIJO_DEFAULT = "OBJECIONES_FAMISANAR"


def escribir_por_factura(
    registros: list[dict],
    carpeta: Path,
    prefijo: str = PREFIJO_DEFAULT,
    consecutivo: int = CDCONSEC_DEFAULT,
) -> list[Path]:
    """Un archivo <prefijo>_<CRNCXC>.xlsx por factura. Cada archivo standalone
    lleva UNA factura, así que su CDCONSEC se reinicia a `consecutivo` (texto)."""
    por_factura: dict[str, list[dict]] = defaultdict(list)
    for reg in registros:
        por_factura[reg["CRNCXC"]].append(reg)

    generados: list[Path] = []
    for crncxc, regs in sorted(por_factura.items()):
        # Copia por registro: NO se muta la lista original — un consolidado
        # posterior sobre los mismos registros perdería el 1,2,3… por factura
        # (hallazgo de la revisión adversarial).
        regs_out = [{**reg, "CDCONSEC": str(consecutivo)} for reg in regs]
        destino = carpeta / f"{prefijo}_{crncxc}.xlsx"
        _escribir_hoja(regs_out, destino)
        generados.append(destino)
        logger.info(f"  {destino.name}: {len(regs)} objeciones")
    return generados


COLUMNAS_REPORTE: tuple[tuple[str, int], ...] = (
    ("Factura", 18),
    ("Cód. glosa", 11),
    ("Valor objetado", 14),
    ("Servicio según FAMISANAR", 42),
    ("Cód. según FAMISANAR", 18),
    ("Vr. unitario FAMISANAR", 16),
    ("Servicio en el DGH", 42),
    ("Cód. en el DGH (SLNSERPRO)", 20),
    ("Vr. unitario DGH", 15),
    ("Centro de costo", 28),
    ("Confianza", 11),
    ("Por qué cruzó", 34),
    ("Puntaje", 9),
    ("Qué revisar", 46),
    ("Observación de FAMISANAR", 70),
)

COLORES_CONFIANZA = {
    "ALTA": "E2EFDA",
    "MEDIA": "FFF2CC",
    "BAJA": "FCE4D6",
    "SIN CRUCE": "F8CBAD",
}


def escribir_reporte_cruce(trazas: list[dict], salida: Path) -> None:
    """Excel de trabajo del auditor: CRUCE (todo), REVISAR (lo que hay que
    confirmar a mano) y RESUMEN por factura. No es el archivo que se sube:
    es el respaldo de por qué cada objeción quedó con ese servicio."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    fill_hdr = PatternFill("solid", fgColor="1F4E78")
    font_hdr = Font(bold=True, color="FFFFFF")

    def encabezar(ws, columnas) -> None:
        for col, (nombre, ancho) in enumerate(columnas, start=1):
            c = ws.cell(row=1, column=col, value=nombre)
            c.fill = fill_hdr
            c.font = font_hdr
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col)].width = ancho
        ws.freeze_panes = "A2"

    def fila(t: dict) -> list:
        return [
            t["factura"],
            t["codigo_objecion"],
            t["valor"],
            t["desc_famisanar"],
            t["cod_famisanar"],
            t["unitario_famisanar"] or "",
            t["servicio_dgh"],
            t["cod_dgh"],
            t["unitario_dgh"],
            t["centro_costo"],
            t["confianza"],
            t["motivos"],
            t["puntaje"],
            t["aviso"],
            t["observacion"],
        ]

    ws = wb.active
    ws.title = "CRUCE"
    encabezar(ws, COLUMNAS_REPORTE)
    for i, t in enumerate(trazas, start=2):
        for col, valor in enumerate(fila(t), start=1):
            celda = ws.cell(row=i, column=col, value=valor)
            if col == 11:
                celda.fill = PatternFill(
                    "solid", fgColor=COLORES_CONFIANZA.get(t["confianza"], "F8CBAD")
                )

    ws2 = wb.create_sheet("REVISAR")
    encabezar(ws2, COLUMNAS_REPORTE)
    pendientes = [t for t in trazas if t["aviso"] or t["confianza"] in ("BAJA", "SIN CRUCE")]
    for i, t in enumerate(pendientes, start=2):
        for col, valor in enumerate(fila(t), start=1):
            ws2.cell(row=i, column=col, value=valor)

    ws3 = wb.create_sheet("RESUMEN")
    encabezar(
        ws3,
        (
            ("Factura", 18),
            ("Objeciones", 11),
            ("Valor objetado", 16),
            ("ALTA", 8),
            ("MEDIA", 8),
            ("BAJA", 8),
            ("Sin cruce", 10),
            ("Estado", 32),
        ),
    )
    por_factura: dict[str, list[dict]] = defaultdict(list)
    for t in trazas:
        por_factura[t["factura"]].append(t)
    linea = 2
    for factura in sorted(por_factura):
        grupo = por_factura[factura]
        cuenta = {k: sum(1 for t in grupo if t["confianza"] == k) for k in COLORES_CONFIANZA}
        pendiente = cuenta["BAJA"] + cuenta["SIN CRUCE"]
        estado = "Listo para subir" if not pendiente else f"Revisar {pendiente} objeción(es)"
        valores = [
            factura,
            len(grupo),
            sum(t["valor"] for t in grupo),
            cuenta["ALTA"],
            cuenta["MEDIA"],
            cuenta["BAJA"],
            cuenta["SIN CRUCE"],
            estado,
        ]
        for col, valor in enumerate(valores, start=1):
            ws3.cell(row=linea, column=col, value=valor)
        linea += 1
    totales = [
        "TOTAL",
        len(trazas),
        sum(t["valor"] for t in trazas),
        *[sum(1 for t in trazas if t["confianza"] == k) for k in COLORES_CONFIANZA],
        "",
    ]
    for col, valor in enumerate(totales, start=1):
        ws3.cell(row=linea, column=col, value=valor).font = Font(bold=True)

    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(salida))


def escribir_consolidado(registros: list[dict], salida: Path) -> None:
    """Un único archivo con todas las facturas (CDCONSEC 1,2,3… por factura)."""
    _escribir_hoja(registros, salida)


# ─── CLI ─────────────────────────────────────────────────────────────────────


def _cargar_mapa(ruta: Path | None) -> dict[str, str]:
    if ruta is None:
        return {}
    try:
        data = json.loads(ruta.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        logger.error(f"No pude leer el mapa de códigos {ruta}: {e}")
        sys.exit(2)
    if not isinstance(data, dict):
        logger.error('El mapa de códigos debe ser un objeto JSON {"CL08": "CL0801", ...}')
        sys.exit(2)
    return {str(k).strip().upper(): str(v).strip() for k, v in data.items()}


def _parse_fecha(texto: str | None) -> _dt.datetime:
    if not texto:
        hoy = _dt.date.today()
        return _dt.datetime(hoy.year, hoy.month, hoy.day)
    try:
        return _dt.datetime.strptime(texto.strip(), "%Y-%m-%d")
    except ValueError:
        logger.error(f"--fecha inválida: {texto!r} (usá YYYY-MM-DD, p.ej. 2026-08-13)")
        sys.exit(2)


def _resumen(registros: list[dict]) -> None:
    facturas = defaultdict(int)
    codigos = defaultdict(int)
    total = 0
    for reg in registros:
        facturas[reg["CRNCXC"]] += 1
        codigos[reg["CRNCONOBJ"]] += 1
        total += reg["CROVALOBJ"]
    logger.info(f"  Facturas: {len(facturas)}  |  Objeciones: {len(registros)}")
    logger.info(f"  Valor glosado total: ${total:,.0f}")
    logger.info("  Códigos de objeción (CRNCONOBJ):")
    for cod, n in sorted(codigos.items(), key=lambda kv: (-kv[1], kv[0])):
        logger.info(f"    {cod or '(sin código)'}: {n}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        "--entrada",
        type=Path,
        required=True,
        help="Excel de FAMISANAR (4 columnas: NRO_FACTURA, CODIGO_DEVOLUCION, "
        "VALOR DEVOLUCION, OBSERVACION).",
    )
    parser.add_argument(
        "--salida",
        type=Path,
        required=True,
        help="Carpeta destino (un archivo <prefijo>_<factura>.xlsx por factura), "
        "o archivo .xlsx si se usa --consolidado.",
    )
    parser.add_argument(
        "--prefijo",
        default=PREFIJO_DEFAULT,
        help=f"Prefijo del nombre de cada archivo por factura. Default: {PREFIJO_DEFAULT}.",
    )
    parser.add_argument(
        "--consolidado",
        action="store_true",
        help="En vez de un archivo por factura, junta todo en un solo Excel (--salida es el .xlsx).",
    )
    parser.add_argument(
        "--fecha",
        default=None,
        help="Fecha para CDFECDOC/CROFECOBJ (YYYY-MM-DD). Default: hoy.",
    )
    parser.add_argument(
        "--codigo-sufijo",
        default=CODIGO_SUFIJO_DEFAULT,
        help="Consecutivo con que se completaría un código de 4 chars (red de seguridad). "
        f"Default: {CODIGO_SUFIJO_DEFAULT}.",
    )
    parser.add_argument(
        "--mapa-codigos",
        type=Path,
        default=None,
        help='JSON opcional para forzar códigos de objeción: {"CO0701": "CO0702", ...}.',
    )
    parser.add_argument(
        "--mapa-servicios",
        type=Path,
        default=None,
        help="JSON con equivalencias FAMISANAR→HUS para códigos de SERVICIO sin regla "
        '(los dispositivos 9101xxxx): {"91017235": "<código HUS>", ...}.',
    )
    parser.add_argument(
        "--servicios-dgh",
        type=Path,
        default=None,
        help="Export de servicios facturados del DGH (columnas SERVICIOS DGH, "
        "DESCRIPCION INSTITUCIONAL, NOM_CENTRO_COSTO, FACTURA, CAT_SERVICIOS, "
        "Vr_SERVICIO). Con él, SLNSERPRO queda con el código real del hospital y "
        "CTNCENCOS con el centro de costo.",
    )
    parser.add_argument(
        "--reporte-cruce",
        type=Path,
        default=None,
        help="Excel de trabajo con el detalle del cruce (hojas CRUCE, REVISAR y "
        "RESUMEN). Requiere --servicios-dgh.",
    )
    parser.add_argument(
        "--sin-homologar",
        action="store_true",
        help="Deja en SLNSERPRO el código tal cual viene de FAMISANAR (sin quitar la "
        "letra de medicamentos ni aplicar el mapa de servicios).",
    )
    parser.add_argument(
        "--consecutivo",
        type=int,
        default=CDCONSEC_DEFAULT,
        help=f"Número inicial del consecutivo por factura (CDCONSEC). Default: {CDCONSEC_DEFAULT}.",
    )
    parser.add_argument("--log", type=Path, default=None, help="Guarda un log adicional a archivo.")
    args = parser.parse_args(argv)

    handlers: list[logging.Handler] = [logging.StreamHandler(sys.stdout)]
    if args.log is not None:
        args.log.parent.mkdir(parents=True, exist_ok=True)
        handlers.append(logging.FileHandler(args.log, encoding="utf-8"))
    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", handlers=handlers
    )

    if not args.entrada.is_file():
        logger.error(f"No existe el archivo de entrada: {args.entrada}")
        return 1

    if args.reporte_cruce is not None and args.servicios_dgh is None:
        logger.error("--reporte-cruce necesita --servicios-dgh (es el detalle de ese cruce).")
        return 1

    fecha = _parse_fecha(args.fecha)
    mapa = _cargar_mapa(args.mapa_codigos)
    mapa_servicios = _cargar_mapa(args.mapa_servicios)

    servicios_dgh = None
    if args.servicios_dgh is not None:
        if not args.servicios_dgh.is_file():
            logger.error(f"No existe el export del DGH: {args.servicios_dgh}")
            return 1
        logger.info(f"Leyendo servicios facturados del DGH: {args.servicios_dgh.name}")
        try:
            servicios_dgh = leer_servicios_dgh(args.servicios_dgh)
        except ValueError as e:
            logger.error(str(e))
            return 1
        logger.info(
            f"  {sum(len(v) for v in servicios_dgh.values())} renglones de servicio "
            f"en {len(servicios_dgh)} facturas."
        )

    trazas: list[dict] = []
    logger.info(f"Leyendo glosas de FAMISANAR: {args.entrada.name}")
    registros = construir_registros(
        args.entrada,
        fecha=fecha,
        consecutivo=args.consecutivo,
        codigo_sufijo=args.codigo_sufijo,
        mapa_codigos=mapa,
        mapa_servicios=mapa_servicios,
        homologar=not args.sin_homologar,
        servicios_dgh=servicios_dgh,
        trazas=trazas,
    )
    if not registros:
        logger.error("No se encontró ninguna objeción en el archivo de entrada.")
        return 1

    if args.consolidado:
        escribir_consolidado(registros, args.salida)
        logger.info(f"\nExcel de FAMISANAR (consolidado, formato de trabajo): {args.salida}")
    else:
        generados = escribir_por_factura(
            registros, args.salida, prefijo=args.prefijo, consecutivo=args.consecutivo
        )
        logger.info(f"\n{len(generados)} archivo(s) de FAMISANAR en: {args.salida}")

    if args.reporte_cruce is not None:
        escribir_reporte_cruce(trazas, args.reporte_cruce)
        pendientes = sum(1 for t in trazas if t["aviso"] or t["confianza"] in ("BAJA", "SIN CRUCE"))
        logger.info(
            f"Detalle del cruce: {args.reporte_cruce} "
            f"({pendientes} renglón(es) en la hoja REVISAR)."
        )

    _resumen(registros)
    return 0


if __name__ == "__main__":
    sys.exit(main())
