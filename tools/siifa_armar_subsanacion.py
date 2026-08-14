#!/usr/bin/env python3
"""siifa_armar_subsanacion.py — El archivo para subsanar lo que la EPS reiteró.

CUÁNDO SE USA. El hospital respondió la glosa, la EPS la miró y NO la levantó:
la reiteró. Ahí arranca la etapa 4 del trámite y el hospital tiene SIETE días
hábiles para subsanar. Vencido ese plazo sin subsanar, la glosa queda en firme
y el valor se pierde.

QUÉ HACE. Toma el informe de seguimientos, busca lo que está reiterado y
todavía sin subsanar, y arma el Excel que carga `responder_glosas_siifa.py`
con `--accion reiteracion-respuesta`.

EN QUÉ SE DIFERENCIA DE LA PRIMERA RESPUESTA. La subsanación no repite el
mismo texto: la EPS ya lo leyó y no lo aceptó. Este escrito (a) deja constancia
de que el hospital ya había respondido y en qué fecha, (b) señala que la
reiteración no desvirtúa lo sustentado, y (c) insiste en la posición
institucional. El argumento de fondo de la causal se conserva, porque sigue
siendo el que corresponde.

USO:
    py siifa_armar_subsanacion.py ^
        --informe "D:\\...\\SIIFA\\informe_13AGO.xlsx" ^
        --salida  "D:\\...\\SIIFA\\SUBSANACION.xlsx"

    REM Después, el piloto de 1 (regla del repo) y el cargue:
    py responder_glosas_siifa.py --excel "D:\\...\\SUBSANACION.xlsx" ^
        --accion reiteracion-respuesta --piloto 1 --reporte "piloto_sub.csv"

OJO CON LAS DEVOLUCIONES REITERADAS. La subsanación por API está confirmada
sólo para GLOSAS (`PUT /api/SeguimientoFacturaGlosa/ReiteracionRespuesta`).
Para una DEVOLUCIÓN reiterada no se conoce la puerta, y mandarla por la de
glosas escribiría sobre otro registro. Por eso salen en una hoja aparte,
marcadas, y el bot de cargue las bloquea. Para saber si esa puerta existe:
`py siifa_sondear_endpoints.py` (sólo consulta, no escribe nada).

INSTALACIÓN (una vez):
    py -m pip install openpyxl
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from _dinero import a_entero  # noqa: E402
import siifa_perfiles as perfiles  # noqa: E402
from siifa_estado_tramite import POR_SUBSANAR, clasificar  # noqa: E402
from siifa_redactar_respuestas import LIMITE_OBSERVACION, argumento_para  # noqa: E402

# El mismo código con el que se respondió que no se acepta la glosa. La
# subsanación sostiene la posición: si el hospital fuera a aceptar ahora, el
# camino no es este sino la nota crédito (ver docs/CONTEXTO_SIIFA.md).
CODIGO_SUBSANACION = "RE9901"


def cierre_de(ips=None) -> str:
    quien = getattr(ips, "nombre_legal", None) or "ESE HOSPITAL UNIVERSITARIO DE SANTANDER"
    return (
        f"POR LO ANTERIOR, {quien} SUBSANA LA GLOSA REITERADA Y SOLICITA SU "
        "LEVANTAMIENTO Y EL PAGO DEL VALOR OBJETADO"
    )


COLS = [
    "ID_SEGUIMIENTO_FACTURA_GLOSA",
    "ID_SEGUIMIENTO_FACTURA_DEVOLUCION",
    "NUMERO_FACTURA",
    "TIPO",
    "CODIGO_GLOSA",
    "DESCRIPCION_GLOSA",
    "VALOR_GLOSA",
    "QUE_DECIDIO_LA_EPS",
    "FECHA_RESPUESTA_ANTERIOR",
    "REVISAR",
    "CODIGO_RESPUESTA",
    "FECHA_RESPUESTA",
    "OBSERVACION_RESPUESTA",
]


def _limpiar(valor: object) -> str:
    texto = "" if valor is None else str(valor)
    return " ".join(texto.split()).strip()


def _pesos(valor: object) -> str:
    return f"${a_entero(valor):,}".replace(",", ".")


def _fecha_corta(valor: object) -> str:
    texto = _limpiar(valor)[:10]
    try:
        return datetime.strptime(texto, "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        return texto


def redactar_subsanacion(linea: dict, limite: int = LIMITE_OBSERVACION, ips=None) -> dict[str, str]:
    """El escrito de subsanación de UNA glosa reiterada."""
    codigo = _limpiar(linea.get("codigo_glosa")).upper()
    banco = argumento_para(codigo)
    fecha_anterior = _fecha_corta(linea.get("fecha_respuesta"))
    quien = getattr(ips, "nombre_legal", None) or "ESE HOSPITAL UNIVERSITARIO DE SANTANDER"
    CIERRE = cierre_de(ips)

    encabezado = (
        f"{quien} SUBSANA LA GLOSA REITERADA SOBRE LA "
        f"FACTURA {_limpiar(linea.get('numero_factura')).upper()} POR VALOR DE "
        f"{_pesos(linea.get('valor_glosa'))} BAJO EL CÓDIGO {codigo}"
    )
    if fecha_anterior:
        encabezado += f", RESPONDIDA POR ESTA INSTITUCIÓN EL {fecha_anterior}"

    reiteracion = _limpiar(linea.get("descripcion_reiteracion"))
    if reiteracion:
        constancia = (
            f". LA ENTIDAD RESPONSABLE DE PAGO REGISTRÓ COMO DECISIÓN INICIAL: "
            f"{reiteracion.upper().rstrip('.')}"
        )
    else:
        constancia = ". LA ENTIDAD REITERÓ LA GLOSA SIN REGISTRAR DECISIÓN MOTIVADA"

    insistencia = (
        ". LA REITERACIÓN NO APORTA ELEMENTO NUEVO ALGUNO QUE DESVIRTÚE LO YA SUSTENTADO "
        "POR ESTA INSTITUCIÓN, NI CONTROVIERTE LOS SOPORTES DE LA ATENCIÓN, POR LO QUE "
        "LA OBJECIÓN CARECE DE FUNDAMENTO Y SE MANTIENE ÍNTEGRAMENTE LA POSICIÓN "
        "INSTITUCIONAL"
    )

    argumento = banco["argumento"].replace("{IPS}", quien)
    texto = f"{encabezado}{constancia}{insistencia}. {argumento}. {CIERRE}"
    if len(texto) > limite:
        # Se recorta la cita de la EPS: el sustento propio y el cierre no se
        # pueden perder, que es lo mismo que hace la primera respuesta.
        sobra = len(texto) - limite
        recorte = max(0, len(constancia) - sobra - 5)
        texto = (
            f"{encabezado}{constancia[:recorte].rstrip()}...{insistencia}. {argumento}. {CIERRE}"
        )
    return {
        "CODIGO_RESPUESTA": CODIGO_SUBSANACION,
        "OBSERVACION_RESPUESTA": texto,
        "REVISAR": banco["revisar"],
    }


def es_devolucion(fila: dict) -> bool:
    return _limpiar(fila.get("tipo_seguimiento")).upper().startswith("DEVOL")


def armar(filas: list[dict], hoy: str, ips=None) -> tuple[list[dict], list[dict]]:
    """Separa lo que se puede subsanar por API de lo que no.

    Las devoluciones reiteradas van aparte a propósito: su puerta de
    subsanación no está confirmada y mandarlas por la de glosas escribiría
    sobre otro registro de la plataforma, con un OK falso en el reporte.
    """
    glosas, devoluciones = [], []
    for f in filas:
        if clasificar(f) != POR_SUBSANAR:
            continue
        fila = {
            "ID_SEGUIMIENTO_FACTURA_GLOSA": f.get("id_seguimiento_factura_glosa", ""),
            "ID_SEGUIMIENTO_FACTURA_DEVOLUCION": f.get("id_seguimiento_factura_devolucion", ""),
            "NUMERO_FACTURA": f.get("numero_factura", ""),
            "TIPO": f.get("tipo_seguimiento", ""),
            "CODIGO_GLOSA": f.get("codigo_glosa", ""),
            "DESCRIPCION_GLOSA": f.get("descripcion_glosa", ""),
            "VALOR_GLOSA": a_entero(f.get("valor_glosa")),
            "QUE_DECIDIO_LA_EPS": f.get("descripcion_reiteracion", ""),
            "FECHA_RESPUESTA_ANTERIOR": _fecha_corta(f.get("fecha_respuesta")),
            "FECHA_RESPUESTA": hoy,
        }
        if es_devolucion(f):
            fila["CODIGO_RESPUESTA"] = ""
            fila["OBSERVACION_RESPUESTA"] = ""
            fila["REVISAR"] = (
                "DEVOLUCIÓN REITERADA: la puerta de subsanación de devoluciones no está "
                "confirmada. NO cargar por la de glosas. Sondear primero."
            )
            devoluciones.append(fila)
        else:
            fila.update(redactar_subsanacion(f, ips=ips))
            glosas.append(fila)
    return glosas, devoluciones


def escribir(glosas: list[dict], devoluciones: list[dict], ruta: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise SystemExit("ERROR: falta openpyxl. py -m pip install openpyxl") from None

    azul = PatternFill("solid", fgColor="1F4E78")
    blanco = Font(bold=True, color="FFFFFF")
    ambar = PatternFill("solid", fgColor="FFF2CC")

    wb = Workbook()

    def hoja(nombre: str, filas: list[dict], marcar: bool):
        ws = wb.create_sheet(nombre) if wb.sheetnames != ["Sheet"] else wb.active
        ws.title = nombre
        for i, h in enumerate(COLS, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.fill = azul
            c.font = blanco
        for r, f in enumerate(filas, start=2):
            for i, h in enumerate(COLS, 1):
                c = ws.cell(row=r, column=i, value=f.get(h, ""))
                c.alignment = Alignment(
                    vertical="top",
                    wrap_text=h in ("OBSERVACION_RESPUESTA", "DESCRIPCION_GLOSA", "REVISAR"),
                )
                if h == "VALOR_GLOSA":
                    c.number_format = '"$"#,##0'
            if marcar:
                ws.cell(row=r, column=COLS.index("REVISAR") + 1).fill = ambar
        anchos = {
            "OBSERVACION_RESPUESTA": 90,
            "DESCRIPCION_GLOSA": 40,
            "REVISAR": 40,
            "QUE_DECIDIO_LA_EPS": 24,
            "NUMERO_FACTURA": 14,
        }
        for i, h in enumerate(COLS, 1):
            ws.column_dimensions[get_column_letter(i)].width = anchos.get(h, 16)
        ws.freeze_panes = "A2"

    hoja("RESPUESTAS", glosas, marcar=False)
    if devoluciones:
        hoja("DEVOLUCIONES_NO_CARGAR", devoluciones, marcar=True)

    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ruta))


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--informe", required=True, help="Informe de siifa_reporte_seguimientos.py.")
    ap.add_argument("--salida", required=True, help="Excel de subsanación a generar.")
    ap.add_argument("--fecha", help="Fecha de la subsanación (AAAA-MM-DD). Por defecto, hoy.")
    perfiles.agregar_argumento(ap)
    args = ap.parse_args()
    ips = perfiles.buscar(args.ips)
    perfiles.anunciar(ips)

    from siifa_novedades import leer_informe

    filas = leer_informe(Path(args.informe))
    hoy = args.fecha or datetime.now().strftime("%Y-%m-%d")
    glosas, devoluciones = armar(filas, hoy, ips=ips)

    if not glosas and not devoluciones:
        print("\nNo hay nada reiterado pendiente de subsanar. Nada que hacer.\n")
        return

    salida = Path(args.salida)
    escribir(glosas, devoluciones, salida)

    print()
    print("=" * 74)
    print("SUBSANACIÓN DE GLOSAS REITERADAS")
    print("=" * 74)
    print(
        f"  Glosas listas para subsanar : {len(glosas):>4}   {_pesos(sum(g['VALOR_GLOSA'] for g in glosas))}"
    )
    if devoluciones:
        valor_dev = sum(
            max(
                (d["VALOR_GLOSA"] for d in devoluciones if d["NUMERO_FACTURA"] == f),
                default=0,
            )
            for f in {d["NUMERO_FACTURA"] for d in devoluciones}
        )
        print(f"  Devoluciones reiteradas     : {len(devoluciones):>4}   {_pesos(valor_dev)}")
        print()
        print("  Las devoluciones NO se cargan: su puerta de subsanación no está")
        print("  confirmada y mandarlas por la de glosas escribiría sobre otro")
        print("  registro. Quedan en la hoja DEVOLUCIONES_NO_CARGAR.")
    print()
    print(f"  Archivo: {salida}")
    print()
    print("  Revisar la columna REVISAR y después, el PILOTO DE 1 (regla del repo):")
    print(f'     py tools\\responder_glosas_siifa.py --excel "{salida}" \\')
    print('        --accion reiteracion-respuesta --piloto 1 --reporte "piloto_subsanacion.csv"')
    print()


if __name__ == "__main__":
    main()
