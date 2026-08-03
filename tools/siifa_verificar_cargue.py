#!/usr/bin/env python3
"""siifa_verificar_cargue.py — Le pregunta a SIIFA qué quedó registrado de
verdad después de un cargue, y lo deja probado por escrito.

POR QUÉ EXISTE
Que el bot diga OK significa que la API contestó bien. Lo que vale en una
conciliación es lo que SIIFA tiene guardado. Con 1.082 respuestas no se
pueden tomar 1.082 pantallazos: este script consulta la plataforma glosa
por glosa y arma la hoja de verificación —y, si se pide, una constancia en
PDF por factura para anexar a soportes.

QUÉ COMPARA
De cada glosa que se subió: si SIIFA la muestra con respuesta, si el código
es el que se mandó y si la FECHA es la que se mandó (la del día en que el
hospital respondió, no la de hoy: de eso depende que la respuesta cuente
como dada dentro del término del artículo 57 de la Ley 1438 de 2011).

USO:
    REM Verificar todo lo que trae el archivo de cargue
    py siifa_verificar_cargue.py ^
        --excel  "D:\\...\\SIIFA\\respuestas_GLOSAS.xlsx" ^
        --salida "D:\\...\\SIIFA\\verificacion_GLOSAS.xlsx"

    REM Verificar solo lo que el cargue dio por bueno, y sacar constancias
    py siifa_verificar_cargue.py ^
        --excel   "D:\\...\\SIIFA\\respuestas_GLOSAS.xlsx" ^
        --reporte "D:\\...\\SIIFA\\reporte_GLOSAS.csv" ^
        --salida  "D:\\...\\SIIFA\\verificacion_GLOSAS.xlsx" ^
        --constancias "D:\\...\\SIIFA\\EVIDENCIAS"

INSTALACIÓN (una vez):
    py -m pip install httpx openpyxl
    py -m pip install reportlab     (solo si se van a sacar constancias)
"""

from __future__ import annotations

import argparse
import csv
import logging
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from siifa_client import (  # noqa: E402
    SiifaApiError,
    SiifaClient,
    credenciales_desde_env,
)
from siifa_reporte_seguimientos import _fila_reporte  # noqa: E402

logger = logging.getLogger("siifa_verificar")

REGISTRADA = "REGISTRADA"
SIN_RESPUESTA = "NO APARECE RESPONDIDA"
NO_ESTA = "NO SE ENCONTRO EN SIIFA"
CON_DIFERENCIAS = "REGISTRADA CON DIFERENCIAS"

COLUMNAS = [
    "ID_SEGUIMIENTO_FACTURA_GLOSA",
    "NUMERO_FACTURA",
    "CODIGO_GLOSA",
    "VALOR_GLOSA",
    "RESULTADO",
    "DIFERENCIAS",
    "CODIGO_ENVIADO",
    "CODIGO_EN_SIIFA",
    "FECHA_ENVIADA",
    "FECHA_EN_SIIFA",
    "OBSERVACION_EN_SIIFA",
]


def _texto(valor) -> str:
    return "" if valor is None else str(valor).strip()


def _dia(valor) -> str:
    """La fecha sin la hora: SIIFA la devuelve como 2026-05-11T00:00:00."""
    texto = _texto(valor)
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")
    return texto[:10]


def leer_cargado(ruta: Path, hoja: str | None = None) -> list[dict]:
    """Lo que se subió, leído del mismo Excel que usó el bot de cargue."""
    from openpyxl import load_workbook

    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[hoja] if hoja else wb[wb.sheetnames[0]]
    filas = ws.iter_rows(values_only=True)
    encabezados = [_texto(c).upper().replace(" ", "_") for c in next(filas)]
    if "ID_SEGUIMIENTO_FACTURA_GLOSA" not in encabezados:
        raise SystemExit(
            f"\nERROR: a {ruta.name} le falta la columna ID_SEGUIMIENTO_FACTURA_GLOSA.\n"
            "Tiene que ser el mismo Excel que se usó para cargar.\n"
        )

    cargado = []
    for f in filas:
        fila = dict(zip(encabezados, f))
        ident = fila.get("ID_SEGUIMIENTO_FACTURA_GLOSA")
        if ident in (None, ""):
            continue
        cargado.append(
            {
                "id": int(ident),
                "factura": _texto(fila.get("NUMERO_FACTURA")),
                "codigo_glosa": _texto(fila.get("CODIGO_GLOSA")),
                "valor_glosa": fila.get("VALOR_GLOSA"),
                "codigo": _texto(fila.get("CODIGO_RESPUESTA")),
                "fecha": _dia(fila.get("FECHA_RESPUESTA")),
            }
        )
    return cargado


def ids_del_reporte(ruta: Path) -> set[int]:
    """Los ids que el cargue dio por buenos (estado OK en el CSV)."""
    ids: set[int] = set()
    with open(ruta, newline="", encoding="utf-8-sig") as f:
        for fila in csv.DictReader(f):
            if _texto(fila.get("estado")).upper() != "OK":
                continue
            try:
                ids.add(int(fila["id_seguimiento_factura_glosa"]))
            except (KeyError, TypeError, ValueError):
                continue
    return ids


def consultar(cliente: SiifaClient, facturas: list[str]) -> dict[int, dict]:
    """Lo que SIIFA tiene hoy, indexado por id de seguimiento.

    Se consulta FACTURA POR FACTURA y no glosa por glosa: 17 consultas en
    vez de 1.082, que es la diferencia entre un minuto y una tarde.
    """
    en_siifa: dict[int, dict] = {}
    for i, factura in enumerate(facturas, 1):
        logger.info("Consultando la factura %s en SIIFA (%d de %d)...", factura, i, len(facturas))
        try:
            for crudo in cliente.listar_seguimientos(numero_factura=factura):
                fila = _fila_reporte(crudo)
                try:
                    en_siifa[int(fila["id_seguimiento_factura_glosa"])] = fila
                except (TypeError, ValueError):
                    continue
        except SiifaApiError as exc:
            logger.error("No se pudo consultar la factura %s: %s", factura, exc)
    return en_siifa


def comparar(enviada: dict, en_siifa: dict | None) -> dict:
    """Una fila de la hoja de verificación: lo enviado contra lo registrado."""
    fila = {
        "ID_SEGUIMIENTO_FACTURA_GLOSA": enviada["id"],
        "NUMERO_FACTURA": enviada["factura"],
        "CODIGO_GLOSA": enviada["codigo_glosa"],
        "VALOR_GLOSA": enviada["valor_glosa"],
        "CODIGO_ENVIADO": enviada["codigo"],
        "FECHA_ENVIADA": enviada["fecha"],
        "CODIGO_EN_SIIFA": "",
        "FECHA_EN_SIIFA": "",
        "OBSERVACION_EN_SIIFA": "",
        "DIFERENCIAS": "",
    }
    if en_siifa is None:
        fila["RESULTADO"] = NO_ESTA
        fila["DIFERENCIAS"] = "SIIFA no devuelve esta glosa: revisar el id y volver a consultar."
        return fila

    fila["CODIGO_EN_SIIFA"] = _texto(en_siifa.get("codigo_respuesta"))
    fila["FECHA_EN_SIIFA"] = _dia(en_siifa.get("fecha_respuesta"))
    fila["OBSERVACION_EN_SIIFA"] = _texto(en_siifa.get("observacion_respuesta"))

    if _texto(en_siifa.get("tiene_respuesta")).upper() != "SI":
        fila["RESULTADO"] = SIN_RESPUESTA
        fila["DIFERENCIAS"] = "La glosa sigue sin respuesta en SIIFA: hay que volver a subirla."
        return fila

    avisos = []
    if enviada["codigo"] and fila["CODIGO_EN_SIIFA"] != enviada["codigo"]:
        avisos.append(f"el código quedó {fila['CODIGO_EN_SIIFA']} y se mandó {enviada['codigo']}")
    if enviada["fecha"] and fila["FECHA_EN_SIIFA"] != enviada["fecha"]:
        avisos.append(
            f"la fecha quedó {fila['FECHA_EN_SIIFA']} y se mandó {enviada['fecha']} "
            "— con la fecha corrida, la EPS puede alegar respuesta fuera de término"
        )
    fila["RESULTADO"] = CON_DIFERENCIAS if avisos else REGISTRADA
    fila["DIFERENCIAS"] = "; ".join(avisos)
    return fila


def escribir(filas: list[dict], ruta: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "VERIFICACION"
    for i, h in enumerate(COLUMNAS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.font = Font(bold=True, color="FFFFFF")

    bien = PatternFill("solid", fgColor="C6EFCE")
    mal = PatternFill("solid", fgColor="FFC7CE")
    ojo = PatternFill("solid", fgColor="FFEB9C")
    for r, fila in enumerate(filas, start=2):
        for i, h in enumerate(COLUMNAS, 1):
            c = ws.cell(row=r, column=i, value=fila.get(h))
            c.alignment = Alignment(
                vertical="top", wrap_text=h in ("DIFERENCIAS", "OBSERVACION_EN_SIIFA")
            )
            if h == "VALOR_GLOSA":
                c.number_format = '"$"#,##0'
            if h == "RESULTADO":
                c.fill = (
                    bien
                    if fila["RESULTADO"] == REGISTRADA
                    else ojo
                    if fila["RESULTADO"] == CON_DIFERENCIAS
                    else mal
                )
    for i, ancho in enumerate([16, 14, 12, 14, 26, 50, 14, 14, 14, 14, 60], 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNAS))}{ws.max_row}"
    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta)
    logger.info("Hoja de verificación: %s (%d glosas)", ruta, len(filas))


def constancias(filas: list[dict], carpeta: Path, cuando: str) -> int:
    """Una constancia PDF por factura, para anexar a los soportes.

    Reemplaza al pantallazo del portal: dice lo mismo —qué respondió el
    hospital, con qué código y con qué fecha— pero consultado a la API
    oficial del Ministerio y con la fecha y hora de la consulta.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        logger.error(
            "Para las constancias falta reportlab. Instalarlo con:  py -m pip install reportlab"
        )
        return 0

    carpeta.mkdir(parents=True, exist_ok=True)
    estilos = getSampleStyleSheet()
    titulo = ParagraphStyle("titulo", parent=estilos["Title"], fontSize=13, spaceAfter=6)
    chico = ParagraphStyle("chico", parent=estilos["Normal"], fontSize=7, leading=9)
    normal = ParagraphStyle("normal", parent=estilos["Normal"], fontSize=9, leading=12)

    por_factura: dict[str, list[dict]] = {}
    for f in filas:
        por_factura.setdefault(f["NUMERO_FACTURA"] or "SIN_FACTURA", []).append(f)

    hechas = 0
    for factura, glosas in sorted(por_factura.items()):
        ruta = carpeta / f"CONSTANCIA_SIIFA_{factura}.pdf"
        doc = SimpleDocTemplate(
            str(ruta),
            pagesize=letter,
            leftMargin=2 * cm,
            rightMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
            title=f"Constancia SIIFA {factura}",
        )
        cuerpo = [
            Paragraph("E.S.E. HOSPITAL UNIVERSITARIO DE SANTANDER", titulo),
            Paragraph(
                "CONSTANCIA DE RESPUESTA REGISTRADA EN SIIFA<br/>"
                "(Sistema de Información de Facturación — Ministerio de Salud y "
                "Protección Social)",
                normal,
            ),
            Spacer(1, 8),
            Paragraph(f"<b>Factura:</b> {factura}", normal),
            Paragraph(f"<b>Fecha y hora de la consulta:</b> {cuando}", normal),
            Spacer(1, 10),
        ]
        datos = [
            [
                Paragraph("<b>Id seguimiento</b>", chico),
                Paragraph("<b>Glosa</b>", chico),
                Paragraph("<b>Valor</b>", chico),
                Paragraph("<b>Respuesta registrada</b>", chico),
                Paragraph("<b>Fecha</b>", chico),
                Paragraph("<b>Estado</b>", chico),
            ]
        ]
        for g in sorted(glosas, key=lambda x: str(x["ID_SEGUIMIENTO_FACTURA_GLOSA"])):
            valor = g.get("VALOR_GLOSA")
            datos.append(
                [
                    Paragraph(str(g["ID_SEGUIMIENTO_FACTURA_GLOSA"]), chico),
                    Paragraph(str(g["CODIGO_GLOSA"] or ""), chico),
                    Paragraph(f"${int(valor):,}".replace(",", ".") if valor else "", chico),
                    Paragraph(str(g["CODIGO_EN_SIIFA"] or ""), chico),
                    Paragraph(str(g["FECHA_EN_SIIFA"] or ""), chico),
                    Paragraph(str(g["RESULTADO"]), chico),
                ]
            )
        tabla = Table(datos, colWidths=[2.6 * cm, 2 * cm, 2.6 * cm, 3.4 * cm, 2.4 * cm, 4 * cm])
        tabla.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        cuerpo.append(tabla)
        cuerpo.append(Spacer(1, 12))

        for g in glosas:
            if not g.get("OBSERVACION_EN_SIIFA"):
                continue
            cuerpo.append(
                Paragraph(
                    f"<b>Texto registrado — seguimiento {g['ID_SEGUIMIENTO_FACTURA_GLOSA']}:</b>",
                    normal,
                )
            )
            cuerpo.append(Paragraph(str(g["OBSERVACION_EN_SIIFA"]), chico))
            cuerpo.append(Spacer(1, 8))

        cuerpo.append(
            Paragraph(
                "Los datos de esta constancia fueron consultados directamente a la API "
                "oficial de interoperabilidad de SIIFA en la fecha y hora indicadas.",
                chico,
            )
        )
        doc.build(cuerpo)
        hechas += 1
    logger.info("Constancias en PDF: %d archivo(s) en %s", hechas, carpeta)
    return hechas


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--excel", required=True, help="El mismo Excel que se usó para cargar.")
    ap.add_argument("--hoja", help="Nombre de la hoja (default: la primera).")
    ap.add_argument("--reporte", help="CSV del cargue: verifica solo las filas que dieron OK.")
    ap.add_argument("--salida", required=True, help="Excel de verificación que se va a escribir.")
    ap.add_argument("--constancias", help="Carpeta donde dejar una constancia PDF por factura.")
    ap.add_argument("--verbose", action="store_true")
    args = ap.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s  %(message)s",
        datefmt="%H:%M:%S",
        stream=sys.stdout,
    )
    if not args.verbose:
        logging.getLogger("httpx").setLevel(logging.WARNING)
        logging.getLogger("httpcore").setLevel(logging.WARNING)

    cargado = leer_cargado(Path(args.excel), args.hoja)
    if args.reporte:
        buenos = ids_del_reporte(Path(args.reporte))
        cargado = [c for c in cargado if c["id"] in buenos]
        logger.info("Del reporte de cargue salieron %d glosas con estado OK.", len(buenos))
    if not cargado:
        raise SystemExit("\nNo hay glosas que verificar con esos archivos.\n")

    facturas = sorted({c["factura"] for c in cargado if c["factura"]})
    logger.info("A verificar: %d glosas de %d factura(s).", len(cargado), len(facturas))

    usuario, password = credenciales_desde_env()
    with SiifaClient() as cliente:
        try:
            cliente.login(usuario, password)
        except SiifaApiError as exc:
            raise SystemExit(f"\nNo se pudo entrar a SIIFA: {exc}\n")
        en_siifa = consultar(cliente, facturas)

    filas = [comparar(c, en_siifa.get(c["id"])) for c in cargado]
    escribir(filas, Path(args.salida))

    cuando = datetime.now().strftime("%d/%m/%Y %H:%M")
    if args.constancias:
        constancias(filas, Path(args.constancias), cuando)

    cuenta = Counter(f["RESULTADO"] for f in filas)
    print(
        f"\nVERIFICACIÓN ({cuando})\n"
        f"  registradas en SIIFA            : {cuenta[REGISTRADA]:5}\n"
        f"  registradas CON DIFERENCIAS     : {cuenta[CON_DIFERENCIAS]:5}\n"
        f"  siguen SIN respuesta            : {cuenta[SIN_RESPUESTA]:5}\n"
        f"  no aparecen en SIIFA            : {cuenta[NO_ESTA]:5}\n"
    )
    if cuenta[REGISTRADA] == len(filas):
        print("Todo lo que se subió está registrado en SIIFA.\n")
    else:
        print(
            "Revisar en la hoja las filas que NO dicen REGISTRADA (van en rojo o amarillo):\n"
            "las que siguen sin respuesta hay que volver a subirlas.\n"
        )


if __name__ == "__main__":
    main()
