"""organizar_objeciones_vco.py — Organiza las objeciones VCO de una entidad.

Convierte entre los dos formatos que maneja el equipo de cartera del HUS para
las objeciones (glosas) que llegan por el portal VCO (vco.ctamedicas.com) de
COOSALUD, FIDUPREVISORA, SAVIA SALUD y demás entidades:

1) CONSOLIDADO VCO (una fila por servicio objetado — formato del acta):

    CROOBSERV | NUMERO FACTURA | VALOR GLOSA | CODIGO GLOSA ESPECIFICA |
    OBSERVACION | CODIGO SERVICIO | DESCRIPCION SERVICIO | CANTIDAD |
    VALOR UNITARIO SERVICIO | VALOR TOTAL SERVICIO

2) PLANTILLA OBJECIONES (cargue masivo al ERP — 16 columnas técnicas):

    CDCONSEC | CDFECDOC | CRNCXC | CROFECOBJ | CROREFERE | CROOBSERV |
    CROCLAOBJ | CRNCLAOBJ | GENUSUARIO4 | CRNCONOBJ | SLNSERPRO | IDRIPS |
    CTNCENCOS | CROVALOBJ | CRDOBSERV | CROTIPOBJ

El bot AUTODETECTA el formato del Excel de entrada por sus encabezados
(match tolerante a mayúsculas, tildes y espacios) y genera el otro formato:

- consolidado  → cargue OBJECIONES (para registrar la glosa en el ERP)
- cargue/export→ CONSOLIDADO VCO   (para trabajar el acta como con las
                                     demás entidades)

USO (SAVIA SALUD):
    py organizar_objeciones_vco.py ^
        --entrada "D:\\GLOSAS 2026\\CONSOLIDADO_VCO_SAVIA.xlsx" ^
        --entidad "SAVIA SALUD"

    py organizar_objeciones_vco.py ^
        --entrada "D:\\GLOSAS 2026\\OBJECIONES_SAVIA.xlsx" ^
        --entidad "SAVIA SALUD" ^
        --salida  "D:\\GLOSAS 2026\\CONSOLIDADO_VCO_SAVIA_SALUD.xlsx"

Flags útiles del cargue: --fecha-documento, --fecha-objecion, --usuario,
--centro-costos, --tipo-objecion, --consecutivo-inicial, --sin-prefijo,
--detalle-servicio. Ver --help.

Supuestos del mapeo (ajustables por flag; validar contra un cargue previo):
- CDCONSEC: mismo consecutivo para todas las filas de una misma factura
  (un documento por factura), arrancando en --consecutivo-inicial.
- CROCLAOBJ/CRNCLAOBJ: derivados del código de glosa Res. 3047
  (TA2901 → clase "TA", concepto general 29).
- CRNCONOBJ: código específico completo (TA2901).
- IDRIPS y CTNCENCOS: vacíos salvo que se pasen (--centro-costos).
"""

from __future__ import annotations

import argparse
import datetime as dt
import logging
import re
import sys
import unicodedata
from collections import OrderedDict
from pathlib import Path

logger = logging.getLogger("organizar_objeciones_vco")

# ---------------------------------------------------------------------------
# Encabezados de los dos formatos
# ---------------------------------------------------------------------------

COLUMNAS_CARGUE = [
    "CDCONSEC",
    "CDFECDOC",
    "CRNCXC",
    "CROFECOBJ",
    "CROREFERE",
    "CROOBSERV",
    "CROCLAOBJ",
    "CRNCLAOBJ",
    "GENUSUARIO4",
    "CRNCONOBJ",
    "SLNSERPRO",
    "IDRIPS",
    "CTNCENCOS",
    "CROVALOBJ",
    "CRDOBSERV",
    "CROTIPOBJ",
]

COLUMNAS_CONSOLIDADO = [
    "CROOBSERV",
    "NUMERO FACTURA",
    "VALOR GLOSA",
    "CODIGO GLOSA ESPECIFICA",
    "OBSERVACION",
    "CODIGO SERVICIO",
    "DESCRIPCION SERVICIO",
    "CANTIDAD",
    "VALOR UNITARIO SERVICIO",
    "VALOR TOTAL SERVICIO",
]

# Alias tolerantes por campo del consolidado (ya normalizados con _norm).
_ALIAS_CONSOLIDADO = {
    "acta": ["CROOBSERV", "ACTA", "REFERENCIA", "NUMERO ACTA", "ACTA VCO"],
    "factura": [
        "NUMERO FACTURA",
        "NUMERO DE FACTURA",
        "NRO FACTURA",
        "FACTURA",
        "NUM FACTURA",
    ],
    "valor_glosa": [
        "VALOR GLOSA",
        "VLR GLOSA",
        "VALOR OBJETADO",
        "VALOR OBJECION",
        "VALOR",
    ],
    "codigo_glosa": [
        "CODIGO GLOSA ESPECIFICA",
        "CODIGO GLOSA",
        "COD GLOSA",
        "CODIGO GLOSA ESPECIFICO",
        "CODIGO",
    ],
    "observacion": ["OBSERVACION", "OBSERVACIONES", "DETALLE", "DETALLE GLOSA"],
    "codigo_servicio": [
        "CODIGO SERVICIO",
        "COD SERVICIO",
        "CODIGO DEL SERVICIO",
        "CUPS",
    ],
    "descripcion_servicio": [
        "DESCRIPCION SERVICIO",
        "DESCRIPCION DEL SERVICIO",
        "SERVICIO",
        "DESCRIPCION",
    ],
    "cantidad": ["CANTIDAD", "CANT", "CANTIDAD SERVICIO"],
    "valor_unitario": [
        "VALOR UNITARIO SERVICIO",
        "VALOR UNITARIO",
        "VLR UNITARIO",
    ],
    "valor_total": [
        "VALOR TOTAL SERVICIO",
        "VALOR TOTAL",
        "VLR TOTAL",
    ],
}

# Alias tolerantes por campo del cargue ERP.
_ALIAS_CARGUE = {campo: [campo] for campo in COLUMNAS_CARGUE}

# Patrón de código de glosa Res. 3047: 2 letras + concepto general + específico.
_RE_CODIGO_GLOSA = re.compile(r"^([A-Z]{2})\s*(\d{2})\s*(\d{2})$")

# "(CEFRADINA AMP X 1 GR CANTIDAD 5)" al final de la observación del acta.
_RE_DETALLE_OBS = re.compile(r"\(([^()]*?)\s+CANTIDAD\s+(\d+)\s*\)\s*$")


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------


def _norm(texto) -> str:
    """MAYÚSCULAS, sin tildes ni espacios repetidos (match tolerante)."""
    if texto is None:
        return ""
    plano = unicodedata.normalize("NFKD", str(texto))
    plano = "".join(ch for ch in plano if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", plano).strip().upper()


def _texto(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def _numero(valor) -> float | int | None:
    """Valor numérico tolerante ('$ 1.234,56', '1,234.56', 42800)."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return valor
    crudo = re.sub(r"[^\d,.\-]", "", str(valor))
    if not crudo:
        return None
    # '1.234,56' (formato es-CO) → '1234.56'; '1,234.56' → '1234.56'
    if "," in crudo and "." in crudo:
        if crudo.rfind(",") > crudo.rfind("."):
            crudo = crudo.replace(".", "").replace(",", ".")
        else:
            crudo = crudo.replace(",", "")
    elif "," in crudo:
        entero, _, dec = crudo.rpartition(",")
        es_decimal = entero and len(dec) != 3
        crudo = crudo.replace(",", "." if es_decimal else "")
    elif crudo.count(".") > 1:
        crudo = crudo.replace(".", "")  # '1.234.567' es-CO → 1234567
    elif "." in crudo:
        entero, _, dec = crudo.rpartition(".")
        if entero.lstrip("-") and len(dec) == 3:
            crudo = crudo.replace(".", "")  # '1.234' es-CO → 1234
    try:
        num = float(crudo)
    except ValueError:
        return None
    return int(num) if num.is_integer() else num


def _fecha(valor: str | None) -> dt.date:
    """Parsea DD/MM/YYYY o YYYY-MM-DD; sin valor → hoy."""
    if not valor:
        return dt.date.today()
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return dt.datetime.strptime(valor.strip(), formato).date()
        except ValueError:
            continue
    raise SystemExit(f"ERROR: fecha inválida {valor!r}. Usar DD/MM/YYYY.")


def _partir_codigo_glosa(codigo: str) -> tuple[str, int | str]:
    """'TA2901' → ('TA', 29). Si no calza el patrón devuelve ('', '')."""
    m = _RE_CODIGO_GLOSA.match(_norm(codigo))
    if not m:
        return "", ""
    return m.group(1), int(m.group(2))


def _mapear_encabezados(fila_encabezado, alias: dict[str, list[str]]) -> dict[str, int]:
    """{campo: índice de columna} para los encabezados presentes."""
    indices: dict[str, int] = {}
    normalizados = [_norm(c) for c in fila_encabezado]
    for campo, nombres in alias.items():
        for nombre in nombres:
            if nombre in normalizados:
                indices[campo] = normalizados.index(nombre)
                break
    return indices


# ---------------------------------------------------------------------------
# Lectura del Excel de entrada
# ---------------------------------------------------------------------------


def _abrir_hoja(ruta: Path, nombre_hoja: str | None):
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(ruta), data_only=True, read_only=True)
    if nombre_hoja:
        buscada = _norm(nombre_hoja)
        for ws in wb.worksheets:
            if _norm(ws.title) == buscada:
                return wb, ws
        disponibles = ", ".join(ws.title for ws in wb.worksheets)
        raise SystemExit(f"ERROR: no existe la hoja {nombre_hoja!r}. Hojas: {disponibles}")
    return wb, wb.worksheets[0]


def leer_entrada(ruta: Path, nombre_hoja: str | None = None):
    """Lee el Excel y devuelve ('consolidado'|'cargue', filas, indices).

    `filas` son tuplas crudas (sin el encabezado); `indices` mapea campo →
    índice de columna según el formato detectado.
    """
    wb, ws = _abrir_hoja(ruta, nombre_hoja)
    try:
        if hasattr(ws, "reset_dimensions"):
            ws.reset_dimensions()  # exports con <dimension> mentiroso (A1:P1)
        todas = [
            fila
            for fila in ws.iter_rows(values_only=True)
            if any(c is not None and str(c).strip() != "" for c in fila)
        ]
    finally:
        wb.close()

    if not todas:
        raise SystemExit(f"ERROR: {ruta} no tiene filas con datos.")

    encabezado, filas = todas[0], todas[1:]

    idx_cargue = _mapear_encabezados(encabezado, _ALIAS_CARGUE)
    if "CRNCXC" in idx_cargue and "CROVALOBJ" in idx_cargue:
        logger.info("Formato detectado: cargue OBJECIONES (%d filas)", len(filas))
        return "cargue", filas, idx_cargue

    idx_cons = _mapear_encabezados(encabezado, _ALIAS_CONSOLIDADO)
    if "factura" in idx_cons and "valor_glosa" in idx_cons:
        logger.info("Formato detectado: CONSOLIDADO VCO (%d filas)", len(filas))
        return "consolidado", filas, idx_cons

    raise SystemExit(
        "ERROR: no reconozco el formato del Excel. Se esperaba un CONSOLIDADO "
        "VCO (columnas NUMERO FACTURA + VALOR GLOSA) o una plantilla "
        f"OBJECIONES (columnas CRNCXC + CROVALOBJ). Encabezados: {encabezado}"
    )


# ---------------------------------------------------------------------------
# consolidado → cargue OBJECIONES
# ---------------------------------------------------------------------------


def consolidado_a_cargue(filas, idx: dict[str, int], cfg) -> list[list]:
    """Arma las filas del cargue ERP a partir del consolidado del acta."""

    def celda(fila, campo):
        pos = idx.get(campo)
        return fila[pos] if pos is not None and pos < len(fila) else None

    consecutivos: OrderedDict[str, int] = OrderedDict()
    salida: list[list] = []
    for fila in filas:
        factura = _texto(celda(fila, "factura"))
        if not factura:
            logger.warning("Fila sin factura, se omite: %s", fila)
            continue
        if factura not in consecutivos:
            consecutivos[factura] = cfg.consecutivo_inicial + len(consecutivos)

        acta = _texto(celda(fila, "acta")) or cfg.referencia
        codigo_glosa = _texto(celda(fila, "codigo_glosa"))
        clase, concepto_general = _partir_codigo_glosa(codigo_glosa)
        observacion = _texto(celda(fila, "observacion"))

        if cfg.detalle_servicio:
            desc = _texto(celda(fila, "descripcion_servicio"))
            cant = _texto(celda(fila, "cantidad"))
            vunit = _numero(celda(fila, "valor_unitario"))
            vtotal = _numero(celda(fila, "valor_total"))
            partes = [
                p
                for p in (
                    f"SERVICIO {desc}" if desc else "",
                    f"CANT {cant}" if cant else "",
                    f"VLR UNIT {vunit}" if vunit is not None else "",
                    f"VLR TOTAL {vtotal}" if vtotal is not None else "",
                )
                if p
            ]
            if partes:
                observacion = f"{observacion} — {' '.join(partes)}".strip(" —")

        factura_erp = factura
        if cfg.sin_prefijo:
            factura_erp = re.sub(r"^[A-Za-z]+", "", factura)

        salida.append(
            [
                consecutivos[factura],  # CDCONSEC
                cfg.fecha_documento,  # CDFECDOC
                factura_erp,  # CRNCXC
                cfg.fecha_objecion,  # CROFECOBJ
                acta,  # CROREFERE
                acta,  # CROOBSERV
                clase,  # CROCLAOBJ
                concepto_general,  # CRNCLAOBJ
                cfg.usuario,  # GENUSUARIO4
                codigo_glosa,  # CRNCONOBJ
                _texto(celda(fila, "codigo_servicio")),  # SLNSERPRO
                "",  # IDRIPS (no viene en el consolidado)
                cfg.centro_costos,  # CTNCENCOS
                _numero(celda(fila, "valor_glosa")),  # CROVALOBJ
                observacion,  # CRDOBSERV
                cfg.tipo_objecion,  # CROTIPOBJ
            ]
        )
    return salida


def escribir_cargue(filas_cargue: list[list], ruta_salida: Path) -> None:
    """Cargue con el formato exacto de la plantilla OBJECIONES del ERP."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "OBJECIONES"
    ws.append(COLUMNAS_CARGUE)
    for fila in filas_cargue:
        ws.append(fila)

    for col_idx, nombre in enumerate(COLUMNAS_CARGUE, start=1):
        letra = ws.cell(row=1, column=col_idx).column_letter
        ancho = (
            max(
                len(nombre),
                *(len(_texto(f[col_idx - 1])) for f in filas_cargue),
                8,
            )
            if filas_cargue
            else max(len(nombre), 8)
        )
        ws.column_dimensions[letra].width = min(ancho + 2, 80)

    idx_fechas = [COLUMNAS_CARGUE.index(c) + 1 for c in ("CDFECDOC", "CROFECOBJ")]
    idx_valor = COLUMNAS_CARGUE.index("CROVALOBJ") + 1
    for num_fila in range(2, len(filas_cargue) + 2):
        for c in idx_fechas:
            ws.cell(row=num_fila, column=c).number_format = "DD/MM/YYYY"
        ws.cell(row=num_fila, column=idx_valor).number_format = "#,##0"

    ws.freeze_panes = "A2"  # como la plantilla original
    wb.save(str(ruta_salida))


# ---------------------------------------------------------------------------
# cargue/export → CONSOLIDADO VCO
# ---------------------------------------------------------------------------


def cargue_a_consolidado(filas, idx: dict[str, int]) -> list[list]:
    """Arma el CONSOLIDADO VCO a partir de un export con columnas del ERP."""

    def celda(fila, campo):
        pos = idx.get(campo)
        return fila[pos] if pos is not None and pos < len(fila) else None

    salida: list[list] = []
    for fila in filas:
        factura = _texto(celda(fila, "CRNCXC"))
        if not factura:
            logger.warning("Fila sin factura (CRNCXC), se omite: %s", fila)
            continue

        codigo = _texto(celda(fila, "CRNCONOBJ"))
        if not codigo:
            clase = _texto(celda(fila, "CROCLAOBJ"))
            general = _texto(celda(fila, "CRNCLAOBJ"))
            codigo = f"{clase}{int(general):02d}" if clase and general.isdigit() else clase

        observacion = _texto(celda(fila, "CRDOBSERV"))
        acta = _texto(celda(fila, "CROOBSERV")) or _texto(celda(fila, "CROREFERE"))

        # Mejor esfuerzo: "(DESC ... CANTIDAD n)" al final de la observación.
        descripcion, cantidad = "", None
        m = _RE_DETALLE_OBS.search(_norm(observacion))
        if m:
            descripcion, cantidad = m.group(1).strip(), int(m.group(2))

        salida.append(
            [
                acta,  # CROOBSERV (referencia del acta VCO)
                factura,  # NUMERO FACTURA
                _numero(celda(fila, "CROVALOBJ")),  # VALOR GLOSA
                codigo,  # CODIGO GLOSA ESPECIFICA
                observacion,  # OBSERVACION
                _texto(celda(fila, "SLNSERPRO")),  # CODIGO SERVICIO
                descripcion,  # DESCRIPCION SERVICIO
                cantidad,  # CANTIDAD
                None,  # VALOR UNITARIO SERVICIO (no viene en el export)
                None,  # VALOR TOTAL SERVICIO (no viene en el export)
            ]
        )
    return salida


def escribir_consolidado(filas_cons: list[list], ruta_salida: Path) -> None:
    """CONSOLIDADO VCO con el mismo formato del ejemplo de las entidades."""
    from openpyxl import Workbook

    # Formato contable del ejemplo (miles con punto, sin decimales).
    fmt_contable = '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-'

    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    ws.append(COLUMNAS_CONSOLIDADO)
    for fila in filas_cons:
        ws.append(fila)

    idx_valores = [
        COLUMNAS_CONSOLIDADO.index(c) + 1
        for c in ("VALOR GLOSA", "VALOR UNITARIO SERVICIO", "VALOR TOTAL SERVICIO")
    ]
    for num_fila in range(2, len(filas_cons) + 2):
        for c in idx_valores:
            ws.cell(row=num_fila, column=c).number_format = fmt_contable

    anchos = {
        "A": 32,
        "B": 16,
        "C": 13,
        "D": 12,
        "E": 19,
        "F": 14,
        "G": 45,
        "H": 10,
        "I": 14,
        "J": 14,
    }
    for letra, ancho in anchos.items():
        ws.column_dimensions[letra].width = ancho

    wb.save(str(ruta_salida))


# ---------------------------------------------------------------------------
# Resumen de control
# ---------------------------------------------------------------------------


def _resumen(filas: list[list], idx_factura: int, idx_valor: int, idx_acta: int) -> str:
    por_acta: OrderedDict[str, dict] = OrderedDict()
    for fila in filas:
        acta = _texto(fila[idx_acta]) or "(sin acta)"
        info = por_acta.setdefault(acta, {"facturas": set(), "filas": 0, "valor": 0})
        info["facturas"].add(_texto(fila[idx_factura]))
        info["filas"] += 1
        info["valor"] += _numero(fila[idx_valor]) or 0
    lineas = ["", "RESUMEN DE CONTROL (validar contra el acta):"]
    total_filas = total_valor = 0
    for acta, info in por_acta.items():
        lineas.append(
            f"  {acta}: {len(info['facturas'])} facturas, "
            f"{info['filas']} objeciones, valor glosado $ {info['valor']:,.0f}"
        )
        total_filas += info["filas"]
        total_valor += info["valor"]
    lineas.append(f"  TOTAL: {total_filas} objeciones, valor glosado $ {total_valor:,.0f}")
    return "\n".join(lineas)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Organiza las objeciones VCO de una entidad: consolidado del acta "
            "↔ plantilla OBJECIONES de cargue al ERP (autodetecta el formato)."
        )
    )
    parser.add_argument("--entrada", required=True, help="Excel de entrada")
    parser.add_argument("--salida", help="Excel de salida (default: junto a la entrada)")
    parser.add_argument("--hoja", help="Nombre de la hoja (default: la primera)")
    parser.add_argument(
        "--entidad",
        default="SAVIA SALUD",
        help="Entidad para el nombre del archivo de salida (default: SAVIA SALUD)",
    )
    parser.add_argument(
        "--referencia",
        default="",
        help="Referencia del acta VCO si el consolidado no trae columna CROOBSERV/ACTA",
    )
    parser.add_argument(
        "--fecha-documento",
        help="CDFECDOC del cargue, DD/MM/YYYY (default: hoy)",
    )
    parser.add_argument(
        "--fecha-objecion",
        help="CROFECOBJ del cargue, DD/MM/YYYY (default: igual a --fecha-documento)",
    )
    parser.add_argument(
        "--usuario", default="CARTERA", help="GENUSUARIO4 del cargue (default: CARTERA)"
    )
    parser.add_argument("--centro-costos", default="", help="CTNCENCOS del cargue (default: vacío)")
    parser.add_argument("--tipo-objecion", default="", help="CROTIPOBJ del cargue (default: vacío)")
    parser.add_argument(
        "--consecutivo-inicial",
        type=int,
        default=1,
        help="Primer CDCONSEC; se asigna uno por factura (default: 1)",
    )
    parser.add_argument(
        "--sin-prefijo",
        action="store_true",
        help="Quita el prefijo alfabético de la factura en CRNCXC (HUS521454 → 521454)",
    )
    parser.add_argument(
        "--detalle-servicio",
        action="store_true",
        help="Anexa servicio/cantidad/valores a CRDOBSERV en el cargue",
    )
    args = parser.parse_args(argv)

    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

    ruta_entrada = Path(args.entrada)
    if not ruta_entrada.exists():
        sys.stderr.write(f"ERROR: no existe {ruta_entrada}\n")
        return 2

    args.fecha_documento = _fecha(args.fecha_documento)
    args.fecha_objecion = (
        _fecha(args.fecha_objecion) if args.fecha_objecion else args.fecha_documento
    )

    formato, filas, idx = leer_entrada(ruta_entrada, args.hoja)
    entidad_archivo = re.sub(r"\s+", "_", _norm(args.entidad)) or "ENTIDAD"

    if formato == "consolidado":
        filas_salida = consolidado_a_cargue(filas, idx, args)
        if not filas_salida:
            sys.stderr.write("ERROR: el consolidado no tiene filas con factura.\n")
            return 2
        ruta_salida = Path(
            args.salida or ruta_entrada.with_name(f"OBJECIONES_{entidad_archivo}.xlsx")
        )
        escribir_cargue(filas_salida, ruta_salida)
        print(f"OK: cargue OBJECIONES con {len(filas_salida)} filas → {ruta_salida}")
        print(
            _resumen(
                filas_salida,
                COLUMNAS_CARGUE.index("CRNCXC"),
                COLUMNAS_CARGUE.index("CROVALOBJ"),
                COLUMNAS_CARGUE.index("CROOBSERV"),
            )
        )
    else:
        filas_salida = cargue_a_consolidado(filas, idx)
        if not filas_salida:
            sys.stderr.write(
                "ERROR: el archivo OBJECIONES no tiene filas de datos (solo "
                "encabezados). Exportá/pegá las objeciones de la entidad y "
                "volvé a correr el bot.\n"
            )
            return 2
        ruta_salida = Path(
            args.salida or ruta_entrada.with_name(f"CONSOLIDADO_VCO_{entidad_archivo}.xlsx")
        )
        escribir_consolidado(filas_salida, ruta_salida)
        print(f"OK: CONSOLIDADO VCO con {len(filas_salida)} filas → {ruta_salida}")
        print(
            _resumen(
                filas_salida,
                COLUMNAS_CONSOLIDADO.index("NUMERO FACTURA"),
                COLUMNAS_CONSOLIDADO.index("VALOR GLOSA"),
                COLUMNAS_CONSOLIDADO.index("CROOBSERV"),
            )
        )
    return 0


if __name__ == "__main__":
    sys.exit(main())
