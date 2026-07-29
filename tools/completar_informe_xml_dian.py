"""completar_informe_xml_dian.py — Completa el informe de revisión XML (DE4401).

Para las facturas devueltas por la EPS con código DE4401 (inconsistencias de
la factura electrónica), este bot toma el Excel del informe, busca el XML
DIAN (y el JSON RIPS/CUV si existe) de cada factura en la carpeta indicada,
analiza los campos y completa las columnas:

- VALOR (XML): total a pagar según el XML (y aviso si difiere del Excel).
- NUMERO_CONTRATO y COBERTURA_PLAN_BENEFICIOS: del sector salud del XML
  (bloque Interoperabilidad, anexo técnico FEV salud).
- VALIDACIÓN DIAN: CUFE, firma digital, acuse de validación de la DIAN
  (ApplicationResponse código 02) y autorización de numeración.
- CONCLUSIÓN (ARGUMENTO GLOSA): dictamen listo para sustentar si la
  devolución procede o no, con lo encontrado.
- ARCHIVO XML: nombre del archivo hallado.
- RESPUESTA DGH: texto corto listo para pegar en el portal.

Además agrega una hoja RESUMEN con los totales y deja el Excel con formato
profesional (semáforo, filtros, paneles congelados).

USO
---
    py tools\\completar_informe_xml_dian.py ^
        --excel "INFORME_REVISION_XML_NUEVA_EPS.xlsx" ^
        --facturas "\\\\172.16.32.83\\factura_electronica_net22\\202607\\FACTURAS_SALUD" ^
        [--salida "INFORME_..._COMPLETO.xlsx"] [--hoja DE4401]

Normalmente se lanza con doble clic desde COMPLETAR_INFORME_XML.cmd.
Solo LEE los XML/JSON; nunca los modifica. Requiere openpyxl.
"""

from __future__ import annotations

import argparse
import html
import json
import logging
import re
import sys
import unicodedata
import zipfile
from collections import Counter
from pathlib import Path, PurePosixPath

logger = logging.getLogger("informe_xml_dian")

VERSION = "2.1 (23 de julio de 2026)"

# ─────────────────────────────────────────────────────────────────────────────
# Utilidades
# ─────────────────────────────────────────────────────────────────────────────


def normalizar(texto: str) -> str:
    if not texto:
        return ""
    t = unicodedata.normalize("NFKD", str(texto))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", t).strip().upper()


def norm_factura(valor: str) -> str:
    """HUS0000533650 → HUS533650 (tolerante a ceros, guiones y espacios)."""
    v = normalizar(valor).replace(" ", "").replace("-", "")
    m = re.match(r"^([A-Z]*)0*(\d+)$", v)
    if m and m.group(2):
        return m.group(1) + m.group(2)
    return v


_RE_FACTURA_TOKEN = re.compile(r"([A-Z]{2,}0*\d{4,})")


def facturas_en_nombre(nombre: str) -> set[str]:
    """Todos los identificadores de factura reconocibles en un nombre."""
    return {norm_factura(t) for t in _RE_FACTURA_TOKEN.findall(normalizar(nombre))}


def _texto_xml(ruta: Path) -> str:
    """Lee el XML y des-escapa el Invoice embebido (CDATA o entidades)."""
    try:
        contenido = ruta.read_text(encoding="utf-8", errors="ignore")
    except OSError:
        return ""
    contenido = contenido.replace("<![CDATA[", "").replace("]]>", "")
    if "&lt;" in contenido:
        contenido = html.unescape(contenido)
    return contenido


def _buscar(patron: str, texto: str, grupo: int = 1) -> str:
    m = re.search(patron, texto, re.DOTALL | re.IGNORECASE)
    return m.group(grupo).strip() if m else ""


def _pesos(n: int | None) -> str:
    return "" if n is None else "$" + f"{n:,}".replace(",", ".")


def _monto(valor: str) -> int | None:
    try:
        return int(round(float(valor)))
    except (ValueError, TypeError):
        return None


# ─────────────────────────────────────────────────────────────────────────────
# Análisis del XML DIAN (FEV sector salud)
# ─────────────────────────────────────────────────────────────────────────────


def _campo_sector_salud(texto: str, nombre: str) -> tuple[str, str]:
    """Devuelve (valor, esquema) del AdditionalInformation con ese Name."""
    m = re.search(
        rf"<AdditionalInformation>\s*<Name>{nombre}</Name>\s*"
        rf"<Value(?P<attrs>[^>]*)>(?P<valor>.*?)</Value>",
        texto,
        re.DOTALL | re.IGNORECASE,
    )
    if not m:
        return "", ""
    valor = re.sub(r"\s+", " ", m.group("valor")).strip()
    esquema = _buscar(r'schemeID="([^"]+)"', m.group("attrs") or "")
    return valor, esquema


def analizar_xml(ruta: Path) -> dict:
    """Extrae todos los datos relevantes del XML de la factura electrónica."""
    texto = _texto_xml(ruta)
    d: dict = {"archivo": ruta.name, "legible": bool(texto)}
    if not texto:
        return d

    d["num_factura"] = _buscar(
        r"<(?:\w+:)?Invoice[\s>].*?<(?:\w+:)?ID(?=[\s>])[^>]*>([^<]+)<", texto
    )
    d["fecha"] = _buscar(r"<(?:\w+:)?IssueDate>([^<]+)<", texto)
    d["cufe"] = _buscar(r'<(?:\w+:)?UUID[^>]*schemeName="CUFE[^"]*"[^>]*>([^<]+)<', texto)
    d["firmada"] = "<ds:Signature" in texto
    d["autorizacion"] = _buscar(r"<sts:InvoiceAuthorization>([^<]+)<", texto)
    d["nit_emisor"] = _buscar(r"AccountingSupplierParty.*?<(?:\w+:)?CompanyID[^>]*>([^<]+)<", texto)
    d["adquiriente"] = _buscar(
        r"AccountingCustomerParty.*?<(?:\w+:)?RegistrationName>([^<]+)<", texto
    )
    d["nit_adquiriente"] = _buscar(
        r"AccountingCustomerParty.*?<(?:\w+:)?CompanyID[^>]*>([^<]+)<", texto
    )
    d["total"] = _monto(
        _buscar(r"LegalMonetaryTotal.*?<(?:\w+:)?PayableAmount[^>]*>([^<]+)<", texto)
    )

    # Acuse de la DIAN embebido (ApplicationResponse): 02 = documento validado
    d["acuse_dian"] = _buscar(r"<ApplicationResponse.*?<(?:\w+:)?ResponseCode>\s*([0-9]+)", texto)
    d["acuse_descripcion"] = _buscar(
        r"<ApplicationResponse.*?<(?:\w+:)?Description>([^<]+)<", texto
    )

    # Sector salud (Interoperabilidad)
    d["tiene_sector_salud"] = "<Interoperabilidad>" in texto
    for campo in (
        "NUMERO_CONTRATO",
        "COBERTURA_PLAN_BENEFICIOS",
        "MODALIDAD_PAGO",
        "NUMERO_POLIZA",
        "COPAGO",
        "CUOTA_MODERADORA",
    ):
        valor, esquema = _campo_sector_salud(texto, campo)
        d[campo.lower()] = valor
        if esquema:
            d[campo.lower() + "_esquema"] = esquema
    return d


def validar_dian(x: dict) -> tuple[str, bool]:
    """Texto de la columna VALIDACIÓN DIAN y si quedó todo en regla."""
    if not x.get("legible"):
        return ("XML ilegible o dañado: no se pudo analizar.", False)
    partes, ok = [], True
    if x.get("cufe"):
        partes.append(f"CUFE presente ({x['cufe'][:16]}…)")
    else:
        partes.append("SIN CUFE")
        ok = False
    partes.append("firmada digitalmente" if x.get("firmada") else "SIN FIRMA DIGITAL")
    ok = ok and x.get("firmada", False)
    if x.get("acuse_dian") == "02":
        partes.append("validada por la DIAN (acuse 02)")
    elif x.get("acuse_dian"):
        partes.append(f"acuse DIAN código {x['acuse_dian']} — revisar")
        ok = False
    else:
        partes.append("sin acuse de validación DIAN en el archivo")
        ok = False
    if x.get("autorizacion"):
        partes.append(f"autorización de numeración {x['autorizacion']}")
    if x.get("tiene_sector_salud"):
        partes.append("con sector salud (anexo FEV salud)")
    else:
        partes.append("SIN el bloque de sector salud")
        ok = False
    return (("VÁLIDA — " if ok else "CON OBSERVACIONES — ") + "; ".join(partes) + ".", ok)


def construir_conclusion(
    factura: str, x: dict | None, dian_ok: bool, valor_excel: int | None, rips_ok: str
) -> tuple[str, str, str]:
    """Devuelve (conclusión, respuesta DGH, estado) para la fila."""
    if x is None:
        return (
            "NO se encontró el XML de la factura en la ruta indicada. Verificar el "
            "número, la carpeta del período o si la factura fue reexpedida con otra "
            "numeración. Sin el XML no es posible sustentar la respuesta a la "
            "devolución DE4401.",
            "No fue posible ubicar el XML en el repositorio del período; se "
            "solicita a facturación confirmar la ruta o la reexpedición.",
            "SIN XML",
        )

    contrato = x.get("numero_contrato", "")
    cobertura = x.get("cobertura_plan_beneficios", "")
    problemas = []
    num_xml = norm_factura(x.get("num_factura", ""))
    if num_xml and num_xml != factura:
        problemas.append(f"el archivo corresponde a OTRA factura ({x.get('num_factura')})")
    if not dian_ok:
        problemas.append("la validación DIAN presenta observaciones")
    if not contrato or normalizar(contrato) in ("SIN CONTRATO", "N/A", "NA", "0"):
        problemas.append(f"el campo NUMERO_CONTRATO del XML está '{contrato or 'VACÍO'}'")
    if not cobertura:
        problemas.append("el campo COBERTURA_PLAN_BENEFICIOS del XML está vacío")
    difiere_valor = (
        valor_excel is not None and x.get("total") is not None and valor_excel != x["total"]
    )
    nota_valor = (
        f" (valor XML {_pesos(x.get('total'))} vs Excel {_pesos(valor_excel)} — conciliar)"
        if difiere_valor
        else ""
    )

    if not problemas:
        conclusion = (
            f"La factura electrónica {x.get('num_factura') or factura} CUMPLE los "
            "requisitos de la FEV del sector salud: está validada por la DIAN "
            "(CUFE y firma digital), registra la cobertura "
            f"'{cobertura}' y el contrato '{contrato}' en el bloque de "
            "interoperabilidad (Res. 506/2021 MinSalud y Res. 2275/2023), por valor "
            f"de {_pesos(x.get('total'))}{nota_valor}"
            + (f". {rips_ok}" if rips_ok else "")
            + ". En consecuencia, la devolución DE4401 NO PROCEDE y se solicita "
            "levantar la devolución y dar trámite de auditoría a la factura."
        )
        respuesta = (
            f"Se anexa la representación XML de la factura {x.get('num_factura') or factura}, "
            f"validada por la DIAN (CUFE {x.get('cufe', '')[:24]}…), con cobertura "
            f"'{cobertura}' y contrato '{contrato}' debidamente diligenciados en el "
            "sector salud del XML. La factura cumple los requisitos de la FEV; se "
            "solicita levantar la devolución DE4401 y continuar el trámite."
        )
        return conclusion, respuesta, "NO PROCEDE DEVOLUCIÓN"

    conclusion = (
        f"REVISAR: en el XML hallado para la factura {factura} "
        + "; ".join(problemas)
        + nota_valor
        + (f". {rips_ok}" if rips_ok else "")
        + ". Si la inconsistencia es real, la devolución DE4401 sería procedente: "
        "corregir el XML con facturación y reexpedir; si no, sustentar con lo "
        "anterior."
    )
    respuesta = (
        "Factura en revisión con el área de facturación por los campos del XML "
        "señalados en la devolución; se dará respuesta con el soporte corregido "
        "o la sustentación correspondiente."
    )
    return conclusion, respuesta, "REVISAR"


# ─────────────────────────────────────────────────────────────────────────────
# Índice de archivos por factura
# ─────────────────────────────────────────────────────────────────────────────


class ArchivoEnZip:
    """Un .xml/.json guardado dentro de un .zip (paquete de la DIAN).

    Imita lo mínimo de Path (name, stem, suffix, read_text, stat) para que
    el resto del bot lo trate igual que un archivo suelto.
    """

    def __init__(self, zip_path: Path, miembro: str):
        self._zip = zip_path
        self._miembro = miembro
        pm = PurePosixPath(miembro.replace("\\", "/"))
        self.name = f"{zip_path.name} → {pm.name}"
        self.stem = pm.stem
        self.suffix = pm.suffix.lower()

    def read_text(self, encoding: str = "utf-8", errors: str = "ignore") -> str:
        try:
            with zipfile.ZipFile(self._zip) as z:
                return z.read(self._miembro).decode(encoding, errors=errors)
        except (zipfile.BadZipFile, KeyError, RuntimeError) as e:
            raise OSError(str(e)) from e

    def stat(self):
        return self._zip.stat()


def indexar_archivos(raiz: Path) -> tuple[dict, dict, dict]:
    """Índices {factura: [xml]} y {factura: [json]} + diagnóstico de la ruta.

    El número de factura se busca en el nombre del archivo Y en el nombre de
    las carpetas que lo contienen: los repositorios de facturación suelen
    guardar cada factura en su propia subcarpeta (FACTURAS_SALUD\\HUS533650\\…)
    con archivos de nombre genérico de la DIAN (ad0901…xml) que no traen el
    número por ninguna parte. Si el XML/JSON viene comprimido en un .zip
    (paquete DIAN), también se lee adentro del .zip.
    """
    xmls: dict[str, list] = {}
    jsons: dict[str, list] = {}
    diag: dict = {
        "total": 0,
        "por_ext": Counter(),
        "subcarpetas": 0,
        "muestra_carpetas": [],
        "muestra_archivos": [],
        "zip_miembros": 0,
        "zip_danados": 0,
    }

    def claves_de(stem: str, p: Path) -> set[str]:
        claves = set(facturas_en_nombre(stem))
        for carpeta in p.parents:
            if carpeta == raiz or carpeta == carpeta.parent:
                break
            claves |= facturas_en_nombre(carpeta.name)
        return claves

    try:
        for hijo in sorted(raiz.iterdir()):
            if hijo.is_dir():
                diag["subcarpetas"] += 1
                if len(diag["muestra_carpetas"]) < 15:
                    diag["muestra_carpetas"].append(hijo.name)
    except OSError as e:
        logger.error(f"  No se pudo listar la ruta: {e}")

    for p in raiz.rglob("*"):
        if not p.is_file():
            continue
        ext = p.suffix.lower()
        diag["total"] += 1
        diag["por_ext"][ext or "(sin extensión)"] += 1
        if len(diag["muestra_archivos"]) < 25:
            try:
                diag["muestra_archivos"].append(str(p.relative_to(raiz)))
            except ValueError:
                diag["muestra_archivos"].append(str(p))
        if ext in (".xml", ".json"):
            destino = xmls if ext == ".xml" else jsons
            for fact in claves_de(p.stem, p):
                destino.setdefault(fact, []).append(p)
        elif ext == ".zip":
            try:
                with zipfile.ZipFile(p) as z:
                    miembros = z.namelist()
            except (zipfile.BadZipFile, OSError):
                diag["zip_danados"] += 1
                continue
            for miembro in miembros:
                a = ArchivoEnZip(p, miembro)
                if a.suffix not in (".xml", ".json"):
                    continue
                diag["zip_miembros"] += 1
                destino = xmls if a.suffix == ".xml" else jsons
                for fact in claves_de(a.stem, p) | facturas_en_nombre(p.stem):
                    destino.setdefault(fact, []).append(a)
    logger.info(
        f"  Archivos vistos: {diag['total']} | XML/JSON dentro de ZIP: {diag['zip_miembros']} | "
        f"claves con XML: {len(xmls)} | con JSON: {len(jsons)}"
    )
    return xmls, jsons, diag


def ordenar_xmls(candidatos: list[Path], factura: str = "") -> list[Path]:
    """Mejor candidato primero: el que trae el número de esta factura en el
    nombre, luego el de FACTURA (no FACOSTE/nota) y el más reciente."""

    def puntaje(p: Path):
        nu = p.name.upper()
        return (
            bool(factura) and factura not in facturas_en_nombre(p.stem),
            "FACOSTE" in nu,
            "FACTURA" not in nu,
            -p.stat().st_mtime,
        )

    return sorted(candidatos, key=puntaje)


def elegir_xml(candidatos: list[Path], factura: str = "") -> Path:
    return ordenar_xmls(candidatos, factura)[0]


def revisar_json(candidatos: list[Path], factura: str) -> str:
    """Nota corta sobre el RIPS/CUV JSON de la factura (si existe)."""
    for p in candidatos:
        try:
            data = json.loads(p.read_text(encoding="utf-8-sig", errors="ignore"))
        except (json.JSONDecodeError, OSError):
            continue
        if not isinstance(data, dict):
            continue
        num = norm_factura(str(data.get("numFactura") or data.get("NumFactura") or ""))
        if data.get("CodigoUnicoValidacion"):
            estado = "válido" if data.get("ResultState") else "CON RECHAZO"
            return f"El CUV JSON está presente y {estado}"
        if num:
            coincide = "coincide" if num == factura else f"trae {num} (NO coincide)"
            return f"El RIPS JSON está presente y su número de factura {coincide}"
    return ""


# ─────────────────────────────────────────────────────────────────────────────
# Excel
# ─────────────────────────────────────────────────────────────────────────────

COLUMNAS_SALIDA = [
    "VALOR (XML)",
    "NUMERO_CONTRATO",
    "COBERTURA_PLAN_BENEFICIOS",
    "VALIDACIÓN DIAN",
    "CONCLUSIÓN (ARGUMENTO GLOSA)",
    "ARCHIVO XML",
    "RESPUESTA DGH",
]


def completar(excel: Path, raiz: Path, salida: Path, hoja: str | None) -> int:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(excel)
    ws = wb[hoja] if hoja else wb[wb.sheetnames[0]]
    encabezados = {str(c.value).strip(): c.column for c in ws[1] if c.value}

    col_factura = encabezados.get("FACTURA") or encabezados.get("FAC")
    if not col_factura:
        logger.error("El Excel no tiene columna FACTURA/FAC en la primera fila.")
        return 2
    col_valor_excel = encabezados.get("VALOR FACTURA")
    faltantes = [c for c in COLUMNAS_SALIDA if c not in encabezados]
    if faltantes:
        logger.error(f"Faltan columnas en el Excel: {', '.join(faltantes)}")
        return 2
    col = {nombre: encabezados[nombre] for nombre in COLUMNAS_SALIDA}

    logger.info(f"  Hoja '{ws.title}': {ws.max_row - 1} facturas por completar")
    xmls, jsons, diag = indexar_archivos(raiz)

    FILL_OK = PatternFill("solid", fgColor="C6EFCE")
    FONT_OK = Font(color="006100", size=10)
    FILL_REV = PatternFill("solid", fgColor="FFEB9C")
    FONT_REV = Font(color="9C6500", size=10)
    FILL_MAL = PatternFill("solid", fgColor="FFC7CE")
    FONT_MAL = Font(color="9C0006", bold=True, size=10)
    BORDE = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

    conteo = {"NO PROCEDE DEVOLUCIÓN": 0, "REVISAR": 0, "SIN XML": 0}
    for i, fila in enumerate(ws.iter_rows(min_row=2), 2):
        crudo = fila[col_factura - 1].value
        if crudo is None or not str(crudo).strip():
            continue
        factura = norm_factura(str(crudo))
        valor_excel = _monto(str(fila[col_valor_excel - 1].value)) if col_valor_excel else None

        candidatos = xmls.get(factura, [])
        x = None
        if candidatos:
            orden = ordenar_xmls(candidatos, factura)
            x = analizar_xml(orden[0])
            # La carpeta de la factura puede traer varios XML (factura, notas,
            # acuses). Si el elegido no corresponde, se prueban los demás.
            if norm_factura(x.get("num_factura", "")) != factura:
                for otro in orden[1:]:
                    alt = analizar_xml(otro)
                    if norm_factura(alt.get("num_factura", "")) == factura:
                        x = alt
                        break
        rips_nota = revisar_json(jsons.get(factura, []), factura)
        texto_dian, dian_ok = validar_dian(x) if x else ("", False)
        conclusion, respuesta, estado = construir_conclusion(
            factura, x, dian_ok, valor_excel, rips_nota
        )
        conteo[estado] += 1

        valores = {
            "VALOR (XML)": (x or {}).get("total"),
            "NUMERO_CONTRATO": (x or {}).get("numero_contrato", ""),
            "COBERTURA_PLAN_BENEFICIOS": (x or {}).get("cobertura_plan_beneficios", ""),
            "VALIDACIÓN DIAN": texto_dian if x else "SIN XML PARA VALIDAR",
            "CONCLUSIÓN (ARGUMENTO GLOSA)": conclusion,
            "ARCHIVO XML": (x or {}).get("archivo", "NO ENCONTRADO"),
            "RESPUESTA DGH": respuesta,
        }
        if estado == "NO PROCEDE DEVOLUCIÓN":
            fill, font = FILL_OK, FONT_OK
        elif estado == "SIN XML":
            fill, font = FILL_MAL, FONT_MAL
        else:
            fill, font = FILL_REV, FONT_REV
        for nombre, valor in valores.items():
            cel = ws.cell(row=i, column=col[nombre], value=valor)
            cel.border = BORDE
            cel.alignment = Alignment(vertical="top", wrap_text=True)
            if nombre == "VALOR (XML)" and isinstance(valor, int):
                cel.number_format = "#,##0"
            if nombre == "CONCLUSIÓN (ARGUMENTO GLOSA)":
                cel.fill, cel.font = fill, font

        if (i - 1) % 25 == 0:
            logger.info(f"  Procesadas {i - 1} facturas…")

    # Formato general de la hoja
    anchos = {
        "VALOR (XML)": 13,
        "NUMERO_CONTRATO": 18,
        "COBERTURA_PLAN_BENEFICIOS": 22,
        "VALIDACIÓN DIAN": 46,
        "CONCLUSIÓN (ARGUMENTO GLOSA)": 70,
        "ARCHIVO XML": 34,
        "RESPUESTA DGH": 60,
    }
    for nombre, ancho in anchos.items():
        ws.column_dimensions[get_column_letter(col[nombre])].width = ancho
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Hoja RESUMEN
    if "RESUMEN" in wb.sheetnames:
        del wb["RESUMEN"]
    res = wb.create_sheet("RESUMEN")
    AZUL = PatternFill("solid", fgColor="1F4E79")
    BLANCA = Font(color="FFFFFF", bold=True)
    filas_resumen = [
        ("INFORME REVISIÓN XML — DEVOLUCIÓN DE4401", ""),
        ("Facturas en el informe", sum(conteo.values())),
        ("NO PROCEDE la devolución (XML en regla)", conteo["NO PROCEDE DEVOLUCIÓN"]),
        ("REVISAR con facturación (XML con observaciones)", conteo["REVISAR"]),
        ("SIN XML en la ruta indicada", conteo["SIN XML"]),
        ("", ""),
        ("Ruta de facturas analizada", str(raiz)),
        (
            "Criterios",
            "CUFE + firma digital + acuse DIAN 02 + sector salud "
            "(NUMERO_CONTRATO y COBERTURA_PLAN_BENEFICIOS) + cruce de valor y RIPS/CUV",
        ),
        ("Versión del bot", VERSION),
    ]
    for j, (a, b) in enumerate(filas_resumen, 1):
        res.cell(row=j, column=1, value=a)
        res.cell(row=j, column=2, value=b)
    res.cell(row=1, column=1).fill = AZUL
    res.cell(row=1, column=1).font = BLANCA
    res.column_dimensions["A"].width = 48
    res.column_dimensions["B"].width = 90

    # Hoja DIAGNOSTICO: qué vio el bot en la ruta. Es la clave para entender
    # a distancia por qué una corrida sale "SIN XML".
    if "DIAGNOSTICO" in wb.sheetnames:
        del wb["DIAGNOSTICO"]
    dg = wb.create_sheet("DIAGNOSTICO")
    ext_txt = " | ".join(f"{e}: {n}" for e, n in diag["por_ext"].most_common()) or "NINGUNO"
    filas_diag: list[tuple[str, object]] = [
        ("DIAGNÓSTICO DE LA RUTA DE FACTURAS", ""),
        (
            "Para qué sirve esta hoja",
            "Muestra qué encontró el bot en la ruta. Si el informe sale con muchas "
            "facturas SIN XML, envíe este mismo Excel para el diagnóstico.",
        ),
        ("Versión del bot", VERSION),
        ("Ruta analizada", str(raiz)),
        ("Subcarpetas de primer nivel", diag["subcarpetas"]),
        ("Archivos vistos en total", diag["total"]),
        ("Archivos por tipo", ext_txt),
        ("XML/JSON hallados dentro de ZIP", diag["zip_miembros"]),
        ("ZIP dañados o ilegibles", diag["zip_danados"]),
        ("Facturas del informe CON XML hallado", sum(conteo.values()) - conteo["SIN XML"]),
        ("Facturas del informe SIN XML", conteo["SIN XML"]),
        ("", ""),
        ("Primeras subcarpetas vistas", " | ".join(diag["muestra_carpetas"]) or "NINGUNA"),
        ("Primeros archivos vistos (ruta relativa)", ""),
    ]
    for m in diag["muestra_archivos"]:
        filas_diag.append(("", m))
    if not diag["muestra_archivos"]:
        filas_diag.append(("", "NINGÚN ARCHIVO VISTO — revisar la ruta o los permisos de red"))
    for j, (a, b) in enumerate(filas_diag, 1):
        dg.cell(row=j, column=1, value=a)
        dg.cell(row=j, column=2, value=b)
    dg.cell(row=1, column=1).fill = AZUL
    dg.cell(row=1, column=1).font = BLANCA
    dg.column_dimensions["A"].width = 42
    dg.column_dimensions["B"].width = 100

    wb.save(salida)
    logger.info("")
    logger.info("=" * 70)
    logger.info(
        f"RESULTADO: {sum(conteo.values())} facturas — "
        f"{conteo['NO PROCEDE DEVOLUCIÓN']} en regla | "
        f"{conteo['REVISAR']} por revisar | {conteo['SIN XML']} sin XML"
    )
    logger.info(f"EXCEL:     {salida}")
    logger.info("=" * 70)
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--excel", type=Path, required=True, help="Informe a completar")
    parser.add_argument(
        "--facturas",
        type=Path,
        required=True,
        help="Carpeta raíz con los XML (y JSON) de las facturas",
    )
    parser.add_argument(
        "--salida",
        type=Path,
        default=None,
        help="Excel de salida (por defecto <informe>_COMPLETO.xlsx)",
    )
    parser.add_argument("--hoja", default=None, help="Nombre de la hoja (por defecto la primera)")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO, format="%(message)s", handlers=[logging.StreamHandler(sys.stdout)]
    )

    logger.info("=" * 70)
    logger.info(f"COMPLETAR INFORME XML DIAN (DE4401) — Motor Glosas HUS — v{VERSION}")
    logger.info("=" * 70)
    if not args.excel.is_file():
        logger.error(f"No existe el Excel: {args.excel}")
        return 1
    if not args.facturas.is_dir():
        logger.error(f"No existe la carpeta de facturas: {args.facturas}")
        return 1
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        logger.error("Falta openpyxl. Instalar con: py -m pip install openpyxl")
        return 1

    salida = args.salida or args.excel.with_name(args.excel.stem + "_COMPLETO.xlsx")
    return completar(args.excel, args.facturas, salida, args.hoja)


if __name__ == "__main__":
    raise SystemExit(main())
