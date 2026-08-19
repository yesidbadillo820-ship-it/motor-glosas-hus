"""ajustar_detallado_glosas.py — deja el detallado de factura SOLO con lo que
la entidad sigue glosando.

Automatiza el trabajo manual de depurar los "detallados de factura" antes de
armar la respuesta a la entidad. En una sola corrida:

  1. Lee el **consolidado** de facturas a trabajar y le **quita los duplicados**.
  2. Abre el **Excel del detallado** (el que baja el sistema con todas las
     facturas del lote apiladas una tras otra, muchas más de las que se van a
     trabajar) y **borra las facturas que no están en el consolidado**.
  3. **Quita el encabezado institucional** (logo, dirección, NIT, ciudad, QR,
     CUFE, "Página 1/1") y cambia el título "FACTURA ELECTRONICA DE VENTA" por
     **"DETALLADO DE FACTURA"** en cada factura que queda.
  4. Cruza cada ítem contra el **ReporteGlosasReclamPAQUETE NNNNN.xlsx** (lo que
     la entidad aprobó vs. lo que sigue glosando) y:
       - **quita** los ítems que la entidad ya aprobó (valor glosado = 0),
       - **ajusta** cantidad y valor de los ítems aprobados a medias (queda solo
         la parte que sigue glosada),
       - **deja igual** los ítems glosados en su totalidad.
  5. Borra los títulos de grupo que quedaron sin ítems (CONSULTAS MEDICAS,
     MEDICAMENTOS POS, DERECHOS DE SALA, …), **recalcula el subtotal y el total**
     y escribe el **total en letras**.
  6. Escribe el Excel corregido y una **bitácora CSV** ítem por ítem con lo que
     hizo (y con lo que NO pudo cruzar, para revisión del auditor).

IMPORTANTE — el reporte de glosas trae el MISMO ítem repartido en varias filas
(ej. VENDA DE GASA en una fila por 4 unidades y otra por 2). El bot **suma todas
las filas del ítem** antes de decidir. Revisar solo una fila subestima lo que
sigue glosado, que es el error típico del proceso manual.

USO (Windows, desde C:\\temp-notas):

    REM 1) Ver qué encontró, sin escribir nada (recomendado la primera vez):
    py tools\\ajustar_detallado_glosas.py ^
        --consolidado "D:\\USUARIO CARTERA\\Downloads\\CONSOLIDADO.xlsx" ^
        --detallado   "D:\\USUARIO CARTERA\\Downloads\\DETALLADOS PAQUETE 31068.xlsx" ^
        --reporte-glosas "D:\\USUARIO CARTERA\\Downloads\\ReporteGlosasReclamPAQUETE 31068.xlsx" ^
        --diagnostico

    REM 2) Generar el Excel corregido + la bitácora:
    py tools\\ajustar_detallado_glosas.py ^
        --consolidado "D:\\USUARIO CARTERA\\Downloads\\CONSOLIDADO.xlsx" ^
        --detallado   "D:\\USUARIO CARTERA\\Downloads\\DETALLADOS PAQUETE 31068.xlsx" ^
        --reporte-glosas "D:\\USUARIO CARTERA\\Downloads\\ReporteGlosasReclamPAQUETE 31068.xlsx" ^
        --salida      "D:\\USUARIO CARTERA\\Documents\\COOSALUD\\DETALLADOS_31068_AJUSTADO.xlsx" ^
        --reporte-csv "D:\\USUARIO CARTERA\\Documents\\COOSALUD\\bitacora_31068.csv"

Seguro por defecto: NO toca los archivos originales. Solo los lee y escribe el
Excel de salida y la bitácora CSV.

Requiere: py -m pip install openpyxl   (opcional: Pillow, para conservar
imágenes que estén FUERA del encabezado; las del encabezado se eliminan igual).
"""

from __future__ import annotations

import argparse
import csv
import logging
import re
import sys
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

# Reutilizamos la normalización de factura del radicador (mismo tools/), para
# que el cruce por número de factura sea idéntico en todo el proyecto.
sys.path.insert(0, str(Path(__file__).resolve().parent))

from _dinero import a_numero  # noqa: E402  (lector único de pesos)

try:  # pragma: no cover - camino normal
    from radicar_facturacion import normalizar_factura
except Exception:  # pragma: no cover - fallback si se ejecuta aislado

    def normalizar_factura(fac: str) -> str:
        f = (fac or "").strip().upper()
        f = re.sub(r"^HUS", "", f)
        return f.lstrip("0") or "0"


logger = logging.getLogger("ajustar_detallado")

# ─── Marcas del formato del detallado ────────────────────────────────────────

TITULO_ORIGINAL = "FACTURA ELECTRONICA DE"
TITULO_NUEVO = "DETALLADO DE FACTURA"

MARCA_SUBTOTAL = "VALOR SUBTOTAL DE SERVICIOS PRESTADOS"
MARCA_COPAGO = "VALOR CUOTA COPAGO"
MARCA_ANTICIPO = "VALOR ANTICIPO PAGADO POR EL USUARIO"
MARCA_COPAGO_USUARIO = "VALOR CUOTA DE COPAGO Y/O CUOTA MODERADORA ASUMIDA POR EL USUARIO"
MARCA_TOTAL_ORDEN = "VALOR TOTAL ORDEN DE SERVICIO"
MARCA_TOTAL_LETRAS = "TOTAL:"

# Pie que el sistema imprime UNA vez al final del archivo (texto legal, nombre
# del reporte, licencia). No pertenece a ninguna factura y no se toca.
MARCAS_PIE_ARCHIVO = (
    "AUTORIZACION FACTURA ELECTRONICA DE VENTA RESOLUCION",
    "LA PRESENTE FACTURA SE ASIMILA",
    "UNA VEZ VENCIDO EL PLAZO",
    "POR LA SUPERINTENDENCIA BANCARIA",
    "NOMBRE REPORTE",
    "LICENCIADO A:",
)

# Membrete del hospital: lo que hay que quitar del detallado.
_PATRONES_MEMBRETE = (
    r"CARR+ERA 33 ?# ?28",
    r"\bNIT 900\.?006\.?037",
    r"^BUCARAMANGA$",
    r"\bPAGINA \d+ ?/ ?\d+",
    r"\bCUFE\b",
    r"^[0-9A-F]{60,}$",  # la línea del CUFE cuando viene sin el rótulo
)

# Encabezados de la tabla de ítems del detallado (tolerante a variantes).
_ALIAS_COL_ITEM: dict[str, tuple[str, ...]] = {
    "codigo": ("CODIGO", "COD", "COD.", "CODIGO SERVICIO"),
    "nombre": ("NOMBRE", "DESCRIPCION", "SERVICIO", "CONCEPTO"),
    "cantidad": ("CANT", "CANT.", "CANTIDAD"),
    "vr_unit": ("VR UNIT", "VR. UNIT", "VALOR UNITARIO", "VR UNITARIO"),
    "vr_pac": ("VR PAC", "VR. PAC", "VALOR PACIENTE"),
    "vr_ent": ("VR ENT", "VR. ENT", "VALOR ENTIDAD", "VR ENTIDAD"),
}

# Encabezados del ReporteGlosasReclamPAQUETE (tolerante a variantes/tildes).
_ALIAS_COL_GLOSA: dict[str, tuple[str, ...]] = {
    "factura": ("NUMERO FACTURA", "NUM FACTURA", "NRO FACTURA", "FACTURA", "NO FACTURA"),
    "radicacion": ("NUMERO RADICACION", "NUM RADICACION", "RADICACION"),
    "paquete": ("NUMERO PAQUETE", "NUM PAQUETE", "PAQUETE"),
    "consecutivo": ("CONSECUTIVO ITEM", "CONSECUTIVO"),
    "tipo_elemento": ("TIPO ELEMENTO", "TIPO DE ELEMENTO"),
    "codigo": ("COD ELEMENTO", "CODIGO ELEMENTO", "COD. ELEMENTO"),
    "descripcion": ("DESCRIPCION ELEMENTO", "DESCRIPCION DEL ELEMENTO"),
    "cant_reclamada": ("CANTIDAD RECLAMADO", "CANTIDAD RECLAMADA", "CANT RECLAMADO"),
    "valor_reclamado": ("VALOR RECLAMADO", "VLR RECLAMADO"),
    "cant_aprobada": ("CANTIDAD APROBADA", "CANTIDAD APROBADO", "CANT APROBADA"),
    "valor_aprobado": ("VALOR APROBADO", "VLR APROBADO"),
    "valor_glosado": ("VALOR GLOSADO", "VLR GLOSADO"),
    "descripcion_glosa": ("DESCRIPCION GLOSA", "DESCRIPCION DE GLOSA"),
    "anotacion": ("DESCRIPCION ANOTACION", "ANOTACION", "OBSERVACION"),
}

# Encabezados aceptados para la columna de factura del consolidado.
_ALIAS_FACTURA_CONSOLIDADO = (
    "FACTURA",
    "NUMERO FACTURA",
    "NUM FACTURA",
    "NRO FACTURA",
    "NO FACTURA",
    "FACTURA HUS",
    "NUMERO DE FACTURA",
    "NO. FACTURA",
)

_RE_FACTURA = re.compile(r"HUS\s?0*\d{4,12}", re.IGNORECASE)

# Tolerancia en pesos al comparar valores (redondeos del portal).
TOLERANCIA_PESOS = 1.0


# ─── Normalización ───────────────────────────────────────────────────────────


def _norm(s) -> str:
    """Mayúsculas, sin tildes, espacios colapsados."""
    s = "" if s is None else str(s)
    s = s.strip().upper()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    return " ".join(s.split())


def _norm_desc(s) -> str:
    """Clave para comparar descripciones: sin tildes, sin puntuación."""
    t = _norm(s)
    t = re.sub(r"[^A-Z0-9 ]+", " ", t)
    return " ".join(t.split())


def _norm_codigo(s) -> str:
    """Clave para comparar códigos entre la factura y el reporte del ADRES.

    Los dos sistemas escriben el mismo código con distinto relleno de ceros:
    la factura dice ``19935303-4`` y el reporte ``19935303-04``. Se quitan
    espacios y puntos y se normalizan los ceros a la izquierda de cada tramo.
    """
    t = re.sub(r"[\s.]+", "", _norm(s))
    if not t:
        return ""
    tramos = []
    for tramo in t.split("-"):
        sin_ceros = tramo.lstrip("0")
        tramos.append(sin_ceros if sin_ceros else "0")
    return "-".join(tramos)


def _parse_valor(v) -> float:
    """'$ 85.800,00' → 85800.0. Vacío o basura → 0.0.

    Delegado al lector único de pesos (`tools/_dinero.py`): la regla de
    leer "$ 950.000" vive en un solo sitio en este repositorio.
    """
    return a_numero(v)


# ─── Total en letras ─────────────────────────────────────────────────────────

_UNI = (
    "",
    "UNO",
    "DOS",
    "TRES",
    "CUATRO",
    "CINCO",
    "SEIS",
    "SIETE",
    "OCHO",
    "NUEVE",
    "DIEZ",
    "ONCE",
    "DOCE",
    "TRECE",
    "CATORCE",
    "QUINCE",
    "DIECISEIS",
    "DIECISIETE",
    "DIECIOCHO",
    "DIECINUEVE",
    "VEINTE",
    "VEINTIUNO",
    "VEINTIDOS",
    "VEINTITRES",
    "VEINTICUATRO",
    "VEINTICINCO",
    "VEINTISEIS",
    "VEINTISIETE",
    "VEINTIOCHO",
    "VEINTINUEVE",
)
_DEC = ("", "", "", "TREINTA", "CUARENTA", "CINCUENTA", "SESENTA", "SETENTA", "OCHENTA", "NOVENTA")
_CEN = (
    "",
    "CIENTO",
    "DOSCIENTOS",
    "TRESCIENTOS",
    "CUATROCIENTOS",
    "QUINIENTOS",
    "SEISCIENTOS",
    "SETECIENTOS",
    "OCHOCIENTOS",
    "NOVECIENTOS",
)


def _letras_999(n: int, apocope: bool = False) -> str:
    if n <= 0:
        return ""
    if n == 100:
        return "CIEN"
    c, r = divmod(n, 100)
    partes = [_CEN[c]] if c else []
    if r:
        if r < 30:
            partes.append(_UNI[r])
        else:
            d, u = divmod(r, 10)
            partes.append(_DEC[d] + (" Y " + _UNI[u] if u else ""))
    out = " ".join(p for p in partes if p)
    if apocope:
        out = re.sub(r"\bVEINTIUNO$", "VEINTIUN", out)
        out = re.sub(r"\bUNO$", "UN", out)
    return out


def _letras_miles(n: int, apocope: bool = False) -> str:
    miles, uni = divmod(n, 1000)
    partes = []
    if miles == 1:
        partes.append("MIL")
    elif miles:
        partes.append(_letras_999(miles, apocope=True) + " MIL")
    if uni:
        partes.append(_letras_999(uni, apocope=apocope))
    return " ".join(p for p in partes if p)


def _letras_entero(n: int) -> str:
    if n == 0:
        return "CERO"
    millones, resto = divmod(n, 1_000_000)
    partes = []
    if millones == 1:
        partes.append("UN MILLON")
    elif millones:
        partes.append(_letras_miles(millones, apocope=True) + " MILLONES")
    if resto:
        partes.append(_letras_miles(resto))
    return " ".join(p for p in partes if p)


def numero_a_letras(valor: float) -> str:
    """2096754 → 'DOS MILLONES ... CUATRO PESOS CON CERO CTVS M/Cte.'

    Reproduce el mismo texto que imprime el sistema en la fila 'TOTAL:'.
    """
    total = abs(_parse_valor(valor))
    pesos = int(total)
    ctvs = int(round((total - pesos) * 100))
    if ctvs == 100:  # 1234,999 → 1235,00
        pesos, ctvs = pesos + 1, 0
    return f"{_letras_entero(pesos)} PESOS CON {_letras_entero(ctvs)} CTVS M/Cte."


# ─── Modelo ──────────────────────────────────────────────────────────────────


@dataclass
class FilaGlosa:
    """Una fila del ReporteGlosasReclamPAQUETE."""

    factura: str
    codigo: str
    descripcion: str
    tipo_elemento: str = ""
    cant_reclamada: float = 0.0
    valor_reclamado: float = 0.0
    cant_aprobada: float = 0.0
    valor_aprobado: float = 0.0
    valor_glosado: float = 0.0
    descripcion_glosa: str = ""
    anotacion: str = ""
    usada: bool = False

    @property
    def vr_unit(self) -> float:
        return self.valor_reclamado / self.cant_reclamada if self.cant_reclamada else 0.0


@dataclass
class ItemDetallado:
    """Un ítem (servicio/insumo) dentro de la hoja del detallado."""

    fila: int
    filas: list[int]  # todas las filas que ocupa (los merges son verticales)
    grupo: str
    codigo: str
    nombre: str
    cantidad: float
    vr_unit: float
    vr_pac: float
    vr_ent: float
    # Renglón de desglose (sin consecutivo): su valor ya viene sumado en el
    # renglón de arriba, así que NO cuenta para el total de la factura.
    desglose: bool = False
    # Celdas reales donde escribir cuando el ítem se ajusta (resueltas por
    # solapamiento de rangos combinados, no por índice de columna).
    celda_cantidad: Celda | None = None
    celda_vr_ent: Celda | None = None


@dataclass
class Bloque:
    """Un grupo del detallado (CONSULTAS MEDICAS, MATERIALES E INSUMOS, …)."""

    titulo: str
    fila_titulo: int | None
    fila_fin: int
    items: list[ItemDetallado] = field(default_factory=list)


@dataclass
class ResultadoItem:
    factura: str
    hoja: str
    grupo: str
    codigo: str
    nombre: str
    cant_orig: float
    vr_ent_orig: float
    valor_reclamado: float = 0.0
    valor_aprobado: float = 0.0
    valor_glosado: float = 0.0
    accion: str = ""  # QUITADO | AJUSTADO | CONSERVADO | SIN_CRUCE
    cant_nueva: float = 0.0
    vr_ent_nuevo: float = 0.0
    cruce: str = ""  # codigo | descripcion | descripcion~ | cantidad+valor | valor
    filas_reporte: int = 0
    causales: str = ""
    desglose: bool = False
    observacion: str = ""


@dataclass
class ResultadoHoja:
    hoja: str
    factura: str
    estado: str  # AJUSTADA | ELIMINADA | SIN_GLOSAS | SIN_ESTRUCTURA
    items_total: int = 0
    items_quitados: int = 0
    items_ajustados: int = 0
    items_conservados: int = 0
    items_sin_cruce: int = 0
    subtotal_antes: float = 0.0
    subtotal_despues: float = 0.0
    detalle: list[ResultadoItem] = field(default_factory=list)
    # Glosas del reporte que no encontraron su ítem en el detallado, y glosas
    # a la reclamación completa (que no corresponden a ningún ítem).
    glosas_sin_item: list = field(default_factory=list)
    reclamacion: list = field(default_factory=list)
    observacion: str = ""


# ─── Lectura de tablas ───────────────────────────────────────────────────────


def _abrir_libro(ruta: Path, *, solo_datos: bool = True, solo_lectura: bool = False):
    try:
        import openpyxl
    except ImportError:  # pragma: no cover - depende del entorno
        raise SystemExit("Falta openpyxl. Instalalo con:  py -m pip install openpyxl") from None
    return openpyxl.load_workbook(ruta, data_only=solo_datos, read_only=solo_lectura)


def _leer_filas(ruta: Path, hoja: str | None = None) -> list[list]:
    """Devuelve las filas de un xlsx/csv/txt como listas de valores."""
    suf = ruta.suffix.lower()
    if suf in (".csv", ".txt"):
        filas: list[list] = []
        with ruta.open("r", encoding="utf-8-sig", newline="") as fh:
            muestra = fh.read(4096)
            fh.seek(0)
            try:
                dialecto = csv.Sniffer().sniff(muestra, delimiters=",;\t|")
            except csv.Error:
                dialecto = csv.excel
            for fila in csv.reader(fh, dialecto):
                filas.append(list(fila))
        return filas
    # read_only: el ReporteGlosasReclamPAQUETE trae decenas de miles de filas y
    # solo lo leemos (sin read_only tarda ~30 s en vez de ~3 s).
    wb = _abrir_libro(ruta, solo_lectura=True)
    try:
        hojas = [wb[hoja]] if hoja else list(wb.worksheets)
        filas = []
        for ws in hojas:
            filas.extend([list(f) for f in ws.iter_rows(values_only=True)])
        return filas
    finally:
        wb.close()


def _fila_encabezado(filas: list[list], alias: dict[str, tuple[str, ...]], minimo: int = 3) -> int:
    """Índice de la fila que sirve de encabezado (la que más alias reconoce)."""
    mejor, mejor_puntaje = -1, 0
    for i, fila in enumerate(filas[:30]):
        celdas = {_norm(c) for c in fila if c is not None}
        puntaje = sum(1 for opciones in alias.values() if celdas & set(opciones))
        if puntaje > mejor_puntaje:
            mejor, mejor_puntaje = i, puntaje
    return mejor if mejor_puntaje >= minimo else -1


def _mapear_columnas(fila: list, alias: dict[str, tuple[str, ...]]) -> dict[str, int]:
    cols: dict[str, int] = {}
    for j, celda in enumerate(fila):
        etiqueta = _norm(celda)
        if not etiqueta:
            continue
        for campo, opciones in alias.items():
            if campo in cols:
                continue
            if etiqueta in opciones or any(etiqueta.startswith(o) for o in opciones):
                cols[campo] = j
                break
    return cols


# ─── 1) Consolidado de facturas ──────────────────────────────────────────────


def leer_consolidado(ruta: Path) -> tuple[list[str], list[str]]:
    """Facturas a trabajar, sin duplicados.

    Devuelve (facturas_unicas_en_orden, duplicadas). Busca primero una columna
    con encabezado de factura; si no la encuentra, rastrea cualquier celda con
    forma de factura HUS.
    """
    filas = _leer_filas(ruta)
    crudas: list[str] = []

    idx_hdr, col = -1, -1
    for i, fila in enumerate(filas[:30]):
        for j, celda in enumerate(fila):
            if _norm(celda) in _ALIAS_FACTURA_CONSOLIDADO:
                idx_hdr, col = i, j
                break
        if col >= 0:
            break

    if col >= 0:
        for fila in filas[idx_hdr + 1 :]:
            if col < len(fila):
                texto = str(fila[col] or "").strip()
                if texto:
                    crudas.append(texto)
    else:
        for fila in filas:
            for celda in fila:
                m = _RE_FACTURA.search(str(celda or ""))
                if m:
                    crudas.append(m.group(0))

    vistas: dict[str, str] = {}
    unicas: list[str] = []
    duplicadas: list[str] = []
    for cruda in crudas:
        clave = normalizar_factura(cruda)
        if clave in ("", "0"):
            continue
        if clave in vistas:
            duplicadas.append(cruda)
            continue
        vistas[clave] = cruda
        unicas.append(cruda)
    return unicas, duplicadas


# ─── 2) Reporte de glosas ────────────────────────────────────────────────────


def leer_reporte_glosas(ruta: Path, paquete: str | None = None) -> dict[str, list[FilaGlosa]]:
    """Agrupa las filas del ReporteGlosasReclamPAQUETE por factura normalizada."""
    filas = _leer_filas(ruta)
    idx = _fila_encabezado(filas, _ALIAS_COL_GLOSA, minimo=4)
    if idx < 0:
        raise SystemExit(
            f"No reconocí el encabezado del reporte de glosas en {ruta.name}. "
            "Se esperan columnas como 'Número Factura', 'Valor Aprobado' y 'Valor Glosado'."
        )
    cols = _mapear_columnas(filas[idx], _ALIAS_COL_GLOSA)
    faltan = [c for c in ("factura", "valor_glosado") if c not in cols]
    if faltan:
        raise SystemExit(f"Al reporte de glosas le faltan columnas: {', '.join(faltan)}")

    def val(fila: list, campo: str):
        j = cols.get(campo, -1)
        return fila[j] if 0 <= j < len(fila) else None

    por_factura: dict[str, list[FilaGlosa]] = {}
    for fila in filas[idx + 1 :]:
        factura = str(val(fila, "factura") or "").strip()
        if not factura:
            continue
        clave = normalizar_factura(factura)
        if clave in ("", "0"):
            continue
        if paquete and _norm(val(fila, "paquete")) != _norm(paquete):
            continue
        por_factura.setdefault(clave, []).append(
            FilaGlosa(
                factura=factura,
                codigo=str(val(fila, "codigo") or "").strip(),
                descripcion=str(val(fila, "descripcion") or "").strip(),
                tipo_elemento=str(val(fila, "tipo_elemento") or "").strip(),
                cant_reclamada=_parse_valor(val(fila, "cant_reclamada")),
                valor_reclamado=_parse_valor(val(fila, "valor_reclamado")),
                cant_aprobada=_parse_valor(val(fila, "cant_aprobada")),
                valor_aprobado=_parse_valor(val(fila, "valor_aprobado")),
                valor_glosado=_parse_valor(val(fila, "valor_glosado")),
                descripcion_glosa=str(val(fila, "descripcion_glosa") or "").strip(),
                anotacion=str(val(fila, "anotacion") or "").strip(),
            )
        )
    return por_factura


# ─── 3) Lectura de la hoja del detallado ─────────────────────────────────────
#
# El detallado que baja el sistema NO trae una hoja por factura: trae UNA hoja
# con todas las facturas apiladas, y cada celda visible es en realidad un
# *rango combinado* (merge). Peor: los merges del encabezado de la tabla y los
# de los ítems NO están alineados —leer "VR ENT" en la columna del rótulo
# devuelve el valor de "VR PAC"—, así que las columnas se resuelven por
# solapamiento de rangos, no por índice de columna.


@dataclass
class Celda:
    """Un valor de la hoja con el rango de filas/columnas que ocupa."""

    fila: int
    fila_fin: int
    col: int
    col_fin: int
    valor: object

    @property
    def texto(self) -> str:
        return "" if self.valor is None else str(self.valor).strip()


class IndiceHoja:
    """Índice de la hoja: celdas por fila y rangos combinados.

    Se arma una sola vez por hoja (O(celdas)); sin él, resolver cada celda
    combinada costaría un barrido de los ~22.000 merges que trae cada archivo.
    """

    def __init__(self, ws):
        self.ws = ws
        self.anclas: dict[tuple[int, int], tuple[int, int]] = {}
        for r in ws.merged_cells.ranges:
            self.anclas[(r.min_row, r.min_col)] = (r.max_row, r.max_col)
        por_fila: dict[int, list[Celda]] = {}
        for (fila, col), celda in ws._cells.items():
            if celda.value is None:
                continue
            fila_fin, col_fin = self.anclas.get((fila, col), (fila, col))
            por_fila.setdefault(fila, []).append(Celda(fila, fila_fin, col, col_fin, celda.value))
        for celdas in por_fila.values():
            celdas.sort(key=lambda c: c.col)
        self.por_fila = por_fila
        self.max_fila = max(por_fila) if por_fila else 0

    def celdas(self, fila: int) -> list[Celda]:
        return self.por_fila.get(fila, [])

    def texto_fila(self, fila: int) -> str:
        return " ".join(c.texto for c in self.celdas(fila) if c.texto)

    def escribir(self, celda: Celda, valor) -> None:
        self.ws.cell(row=celda.fila, column=celda.col).value = valor
        celda.valor = valor


def _solapamiento(a: tuple[int, int], b: tuple[int, int]) -> int:
    return max(0, min(a[1], b[1]) - max(a[0], b[0]) + 1)


def mapear_por_solapamiento(
    cabeceras: list[tuple[str, tuple[int, int]]], celdas: list[Celda]
) -> dict[str, Celda]:
    """Asigna cada rótulo del encabezado a la celda del ítem que le corresponde.

    Criterio: mayor solapamiento de columnas, cada celda se usa una sola vez y
    los empates se resuelven de izquierda a derecha. Con eso 'VR ENT' [53-64]
    no se lleva el bloque de 'VR PAC' [44-58] cuando ambos solapan 6 columnas:
    el de pac ya quedó tomado por el rótulo anterior.
    """
    usadas: set[int] = set()
    asignado: dict[str, Celda] = {}
    for campo, span in cabeceras:
        mejor, mejor_ov = None, 0
        for i, celda in enumerate(celdas):
            if i in usadas:
                continue
            ov = _solapamiento(span, (celda.col, celda.col_fin))
            if ov > mejor_ov:
                mejor, mejor_ov = i, ov
        if mejor is not None:
            usadas.add(mejor)
            asignado[campo] = celdas[mejor]

    # Rótulos sin solapamiento (formatos sin celdas combinadas): se les asigna
    # la celda libre más cercana que caiga entre el rótulo anterior y el siguiente.
    for pos, (campo, span) in enumerate(cabeceras):
        if campo in asignado:
            continue
        antes = max((asignado[c].col_fin for c, _ in cabeceras[:pos] if c in asignado), default=0)
        despues = min(
            (asignado[c].col for c, _ in cabeceras[pos + 1 :] if c in asignado),
            default=10**6,
        )
        libres = [
            (abs(c.col - span[0]), i)
            for i, c in enumerate(celdas)
            if i not in usadas and antes < c.col < despues
        ]
        if libres:
            _, i = min(libres)
            usadas.add(i)
            asignado[campo] = celdas[i]
    return asignado


def _cabeceras_tabla(celdas: list[Celda]) -> list[tuple[str, tuple[int, int]]] | None:
    """Reconoce la fila 'CÓDIGO / NOMBRE / CANT / VR UNIT / VR PAC / VR ENT'."""
    encontrados: list[tuple[str, tuple[int, int]]] = []
    vistos: set[str] = set()
    for celda in celdas:
        etiqueta = _norm(celda.texto)
        if not etiqueta:
            continue
        for campo, opciones in _ALIAS_COL_ITEM.items():
            if campo in vistos:
                continue
            if etiqueta in opciones or any(etiqueta.startswith(o) for o in opciones):
                vistos.add(campo)
                encontrados.append((campo, (celda.col, celda.col_fin)))
                break
    if {"nombre", "cantidad"} <= vistos and ("vr_ent" in vistos or "vr_unit" in vistos):
        return sorted(encontrados, key=lambda x: x[1][0])
    return None


@dataclass
class FacturaEnHoja:
    """Una factura dentro de la hoja: dónde empieza, dónde termina y su número."""

    factura: str
    fila_ini: int  # fila del título
    fila_fin: int
    celda_titulo: Celda


def _es_titulo(texto: str) -> bool:
    t = _norm(texto)
    return t.startswith(TITULO_ORIGINAL) or t.startswith(TITULO_NUEVO)


def ultima_fila_util(idx: IndiceHoja) -> int:
    """La última fila de la factura, sin el pie legal del archivo.

    Al final de la hoja el sistema imprime el pie de página —la autorización de
    la DIAN, el aviso de la letra de cambio, «Nombre reporte» y «LICENCIADO A»—
    que no es parte de ninguna factura. Se busca de abajo hacia arriba.
    """
    fin_util = idx.max_fila
    for fila in range(idx.max_fila, 0, -1):
        texto = _norm(idx.texto_fila(fila))
        if not texto:
            continue
        if any(m in texto for m in MARCAS_PIE_ARCHIVO):
            fin_util = fila - 1
        else:
            break
    return fin_util


def segmentar_facturas(idx: IndiceHoja) -> list[FacturaEnHoja]:
    """Parte la hoja en facturas. Sirve igual si la hoja trae una sola."""
    inicios: list[tuple[int, Celda]] = []
    for fila in sorted(idx.por_fila):
        for celda in idx.celdas(fila):
            if _es_titulo(celda.texto):
                inicios.append((fila, celda))
                break

    fin_util = ultima_fila_util(idx)

    facturas = []
    for i, (fila, celda) in enumerate(inicios):
        fin = inicios[i + 1][0] - 1 if i + 1 < len(inicios) else fin_util
        texto = idx.texto_fila(fila)
        m = _RE_FACTURA.search(texto)
        facturas.append(
            FacturaEnHoja(
                factura=m.group(0).replace(" ", "") if m else "",
                fila_ini=fila,
                fila_fin=max(fin, fila),
                celda_titulo=celda,
            )
        )
    return facturas


def filas_encabezado_institucional(idx: IndiceHoja, hasta: int) -> set[int]:
    """Filas del membrete del hospital (logo, dirección, NIT, ciudad, QR, CUFE,
    'Página 1/1'). Van antes de la primera factura y no se repiten por factura,
    pero se buscan también en el resto por si el formato cambia."""
    filas = set(range(1, hasta))  # incluye los renglones en blanco del membrete
    for fila in idx.por_fila:
        if fila in filas:
            continue
        texto = _norm(idx.texto_fila(fila))
        if any(re.search(p, texto) for p in _PATRONES_MEMBRETE):
            filas.add(fila)
    return filas


@dataclass
class EstructuraFactura:
    fila_hdr: int
    cabeceras: list[tuple[str, tuple[int, int]]]
    fila_subtotal: int


def detectar_estructura(idx: IndiceHoja, fac: FacturaEnHoja) -> EstructuraFactura | None:
    """Ubica la tabla de ítems y la fila de subtotal dentro de la factura."""
    fila_hdr, cabeceras = -1, None
    for fila in range(fac.fila_ini, fac.fila_fin + 1):
        cab = _cabeceras_tabla(idx.celdas(fila))
        if cab:
            fila_hdr, cabeceras = fila, cab
            break
    if fila_hdr < 0 or cabeceras is None:
        return None
    fila_sub = -1
    for fila in range(fila_hdr + 1, fac.fila_fin + 1):
        if _norm(idx.texto_fila(fila)).startswith(MARCA_SUBTOTAL):
            fila_sub = fila
            break
    if fila_sub < 0:
        return None
    return EstructuraFactura(fila_hdr=fila_hdr, cabeceras=cabeceras, fila_subtotal=fila_sub)


def leer_bloques(idx: IndiceHoja, est: EstructuraFactura) -> list[Bloque]:
    """Recorre la tabla y la parte en grupos, ítems y continuaciones de nombre.

    - **Ítem**: la fila trae cantidad y valor. Puede ocupar varias filas (los
      merges son verticales cuando el nombre es largo).
    - **Título de grupo**: texto que arranca a la izquierda de la columna NOMBRE.
    - **Continuación**: texto en la columna NOMBRE, sin cantidad; es del ítem
      de arriba.
    """
    col_nombre = dict(est.cabeceras).get("nombre", (2, 2))[0]
    bloques: list[Bloque] = [Bloque(titulo="", fila_titulo=None, fila_fin=est.fila_subtotal - 1)]

    fila = est.fila_hdr + 1
    while fila < est.fila_subtotal:
        celdas = idx.celdas(fila)
        if not celdas:
            fila += 1
            continue

        campos = mapear_por_solapamiento(est.cabeceras, celdas)
        cantidad = _parse_valor(campos["cantidad"].valor) if "cantidad" in campos else 0.0
        vr_unit = _parse_valor(campos["vr_unit"].valor) if "vr_unit" in campos else 0.0
        vr_ent = _parse_valor(campos["vr_ent"].valor) if "vr_ent" in campos else 0.0
        vr_pac = _parse_valor(campos["vr_pac"].valor) if "vr_pac" in campos else 0.0

        if cantidad > 0 and (vr_unit > 0 or vr_ent > 0):
            fila_fin = max(c.fila_fin for c in celdas)
            codigo = campos["codigo"].texto if "codigo" in campos else ""
            nombre = campos["nombre"].texto if "nombre" in campos else ""
            if not codigo:
                izquierda = [c.texto for c in celdas if c.col_fin < col_nombre and c.texto]
                codigo = izquierda[-1] if izquierda else ""
            # Nombres largos: unas veces el sistema los mete en una celda
            # combinada de dos filas (ya viene entero) y otras los parte en un
            # segundo renglón dentro del alto del ítem. Ese resto se pega acá.
            for siguiente in range(fila + 1, fila_fin + 1):
                resto = " ".join(
                    c.texto for c in idx.celdas(siguiente) if c.texto and c.col >= col_nombre
                )
                if resto:
                    nombre = f"{nombre} {resto}".strip()
            # Los renglones de DESGLOSE (los que abren un procedimiento
            # quirúrgico en honorarios, derechos de sala, materiales…) vienen
            # SIN número consecutivo. Su valor ya está incluido en el del
            # renglón de arriba, así que no suman al total de la factura.
            a_la_izquierda = [c for c in celdas if c.col_fin < col_nombre and c.texto]
            bloques[-1].items.append(
                ItemDetallado(
                    fila=fila,
                    filas=list(range(fila, fila_fin + 1)),
                    grupo=bloques[-1].titulo,
                    codigo=codigo,
                    nombre=nombre,
                    cantidad=cantidad,
                    vr_unit=vr_unit,
                    vr_pac=vr_pac,
                    vr_ent=vr_ent or cantidad * vr_unit,
                    desglose=len(a_la_izquierda) < 2,
                    celda_cantidad=campos.get("cantidad"),
                    celda_vr_ent=campos.get("vr_ent"),
                )
            )
            fila = fila_fin + 1
            continue

        col_ini = min(c.col for c in celdas)
        if col_ini < col_nombre:
            bloques[-1].fila_fin = fila - 1
            titulo = " ".join(c.texto for c in celdas if c.texto)
            bloques.append(Bloque(titulo=titulo, fila_titulo=fila, fila_fin=est.fila_subtotal - 1))
        elif bloques[-1].items:
            item = bloques[-1].items[-1]
            extra = " ".join(c.texto for c in celdas if c.texto)
            item.filas.append(fila)
            item.nombre = f"{item.nombre} {extra}".strip()
        fila += 1

    return [b for b in bloques if b.items or b.fila_titulo is not None]


# ─── 4) Cruce contra el reporte de glosas ────────────────────────────────────
#
# El reporte del ADRES y el detallado del hospital NO hablan el mismo idioma:
#
#   - los códigos llevan distinto relleno de ceros (19935303-4 / 19935303-04);
#   - a los materiales de osteosíntesis el reporte les agrega un sufijo
#     ("… MATERIAL DE OSTEOSINTESIS UNIDAD 01");
#   - los dispositivos médicos vienen con código INVIMA en el reporte y con
#     código interno en la factura;
#   - el reporte parte un mismo ítem en varias filas.
#
# Por eso el emparejamiento se hace **por rondas**, de la evidencia más fuerte
# a la más débil, y en cada ronda solo se acepta un par cuando es **mutuamente
# único**: si un ítem del detallado empata con dos del reporte (o al revés), se
# deja para una ronda posterior que pueda desempatar por valor. Emparejar de a
# uno "al primero que aparezca" es justo lo que hacía que la PIPERACILINA se
# llevara la glosa del NITRÓGENO URÉICO por tener el mismo valor unitario.

TIPO_RECLAMACION = "RECLAMACION"


@dataclass
class ItemReporte:
    """Todas las filas del reporte que corresponden a un mismo ítem."""

    codigo: str
    descripcion: str
    tipo: str
    filas: list[FilaGlosa] = field(default_factory=list)
    emparejado: bool = False

    @property
    def cant_reclamada(self) -> float:
        return sum(f.cant_reclamada for f in self.filas)

    @property
    def valor_reclamado(self) -> float:
        return sum(f.valor_reclamado for f in self.filas)

    @property
    def valor_aprobado(self) -> float:
        return sum(f.valor_aprobado for f in self.filas)

    @property
    def valor_glosado(self) -> float:
        return sum(f.valor_glosado for f in self.filas)

    @property
    def causales(self) -> str:
        vistas = []
        for f in self.filas:
            if f.descripcion_glosa and f.descripcion_glosa not in vistas:
                vistas.append(f.descripcion_glosa)
        return " | ".join(vistas)


def agrupar_reporte(glosas: list[FilaGlosa]) -> tuple[list[ItemReporte], list[FilaGlosa]]:
    """Agrupa las filas del reporte por ítem.

    Devuelve (ítems, filas de reclamación). Las filas con Tipo Elemento
    'Reclamacion' NO son ítems: son glosas a la reclamación completa (falta un
    anexo, formulario incompleto…), vienen sin código ni descripción y su valor
    repite el de la factura. Si se cruzaran como ítems, duplicarían la glosa.
    """
    items: dict[tuple[str, str], ItemReporte] = {}
    reclamacion: list[FilaGlosa] = []
    for g in glosas:
        if _norm(g.tipo_elemento) == TIPO_RECLAMACION or (
            not g.codigo.strip() and not g.descripcion.strip()
        ):
            reclamacion.append(g)
            continue
        clave = (_norm_codigo(g.codigo), _norm_desc(g.descripcion))
        item = items.get(clave)
        if item is None:
            item = ItemReporte(codigo=g.codigo, descripcion=g.descripcion, tipo=g.tipo_elemento)
            items[clave] = item
        item.filas.append(g)
    return list(items.values()), reclamacion


def _prefijo(a: str, b: str, minimo: int = 12) -> bool:
    """¿Una descripción es el principio de la otra? (sufijos del reporte)."""
    if len(a) < minimo or len(b) < minimo:
        return False
    return a.startswith(b) or b.startswith(a)


def _clave_valor(item) -> str:
    return f"{round(_valor_item(item), 2):.2f}"


def _valor_item(item) -> float:
    return item.vr_ent if isinstance(item, ItemDetallado) else item.valor_reclamado


def _cantidad_item(item) -> float:
    return item.cantidad if isinstance(item, ItemDetallado) else item.cant_reclamada


# Cada ronda es (nombre, función que devuelve la clave del ítem del detallado,
# función que devuelve la clave del ítem del reporte). Clave vacía = no aplica.
RONDAS_CRUCE: tuple[tuple[str, object, object], ...] = (
    ("codigo", lambda d: _norm_codigo(d.codigo), lambda r: _norm_codigo(r.codigo)),
    ("descripcion", lambda d: _norm_desc(d.nombre), lambda r: _norm_desc(r.descripcion)),
    (
        "cantidad+valor",
        lambda d: f"{_cantidad_item(d):g}|{_clave_valor(d)}",
        lambda r: f"{_cantidad_item(r):g}|{_clave_valor(r)}",
    ),
    ("valor", _clave_valor, _clave_valor),
)


def _emparejar_por_clave(
    pendientes_det: list[ItemDetallado],
    pendientes_rep: list[ItemReporte],
    clave_det,
    clave_rep,
) -> list[tuple[ItemDetallado, ItemReporte]]:
    """Empareja solo cuando la clave es única de los dos lados."""
    por_det: dict[str, list[ItemDetallado]] = {}
    for d in pendientes_det:
        k = clave_det(d)
        if k:
            por_det.setdefault(k, []).append(d)
    por_rep: dict[str, list[ItemReporte]] = {}
    for r in pendientes_rep:
        k = clave_rep(r)
        if k:
            por_rep.setdefault(k, []).append(r)
    return [
        (ds[0], por_rep[k][0])
        for k, ds in por_det.items()
        if len(ds) == 1 and len(por_rep.get(k, [])) == 1
    ]


def emparejar(
    items: list[ItemDetallado], reporte: list[ItemReporte]
) -> tuple[dict[int, tuple[ItemReporte, str]], list[ItemReporte]]:
    """Empareja los ítems del detallado con los del reporte.

    Devuelve ({índice del ítem: (ítem del reporte, criterio)}, ítems del reporte
    que no encontraron pareja).
    """
    pares: dict[int, tuple[ItemReporte, str]] = {}
    pos = {id(d): i for i, d in enumerate(items)}
    pendientes_det = list(items)
    pendientes_rep = list(reporte)

    def registrar(nuevos, criterio):
        for d, r in nuevos:
            pares[pos[id(d)]] = (r, criterio)
            r.emparejado = True
        emparejados_det = {id(d) for d, _ in nuevos}
        emparejados_rep = {id(r) for _, r in nuevos}
        return (
            [d for d in pendientes_det if id(d) not in emparejados_det],
            [r for r in pendientes_rep if id(r) not in emparejados_rep],
        )

    for criterio, clave_det, clave_rep in RONDAS_CRUCE:
        if not pendientes_det or not pendientes_rep:
            break
        nuevos = _emparejar_por_clave(pendientes_det, pendientes_rep, clave_det, clave_rep)
        pendientes_det, pendientes_rep = registrar(nuevos, criterio)

        if criterio == "descripcion" and pendientes_det and pendientes_rep:
            # El reporte le agrega sufijos a la descripción ("… MATERIAL DE
            # OSTEOSINTESIS UNIDAD 01"): se acepta el prefijo si es único.
            nuevos = []
            for d in pendientes_det:
                nd = _norm_desc(d.nombre)
                cand = [r for r in pendientes_rep if _prefijo(nd, _norm_desc(r.descripcion))]
                if (
                    len(cand) == 1
                    and sum(
                        1
                        for otro in pendientes_det
                        if _prefijo(_norm_desc(otro.nombre), _norm_desc(cand[0].descripcion))
                    )
                    == 1
                ):
                    nuevos.append((d, cand[0]))
            pendientes_det, pendientes_rep = registrar(nuevos, "descripcion~")

    # Última ronda: VARIOS renglones del detallado contra UNO del reporte.
    # Pasa seguido con los materiales (dos renglones de "TORNILLO BLOQUEADO DE
    # 3.5 MM", uno de 1 y otro de 4 unidades, que el reporte trae como uno solo
    # de 5). Solo se acepta si las cantidades y los valores cuadran exacto, y a
    # cada renglón se le asigna su parte proporcional de la glosa.
    if pendientes_det and pendientes_rep:
        for grupo, rep in _agrupar_varios_a_uno(pendientes_det, pendientes_rep):
            total = sum(d.vr_ent for d in grupo)
            for d in grupo:
                porcion = _porcion(rep, d.vr_ent / total)
                pares[pos[id(d)]] = (porcion, "varios-a-uno")
            rep.emparejado = True

    return pares, [r for r in reporte if not r.emparejado]


def _agrupar_varios_a_uno(
    pendientes_det: list[ItemDetallado], pendientes_rep: list[ItemReporte]
) -> list[tuple[list[ItemDetallado], ItemReporte]]:
    """Grupos de renglones del detallado que son UN solo ítem del reporte."""
    salida: list[tuple[list[ItemDetallado], ItemReporte]] = []
    usados_det: set[int] = set()
    usados_rep: set[int] = set()

    def intentar(clave_det, clave_rep, prefijo=False):
        por_det: dict[str, list[ItemDetallado]] = {}
        for d in pendientes_det:
            if id(d) in usados_det:
                continue
            k = clave_det(d)
            if k:
                por_det.setdefault(k, []).append(d)
        for k, grupo in por_det.items():
            if len(grupo) < 2:
                continue
            if prefijo:
                cand = [
                    r
                    for r in pendientes_rep
                    if id(r) not in usados_rep and _prefijo(k, _norm_desc(r.descripcion))
                ]
            else:
                cand = [r for r in pendientes_rep if id(r) not in usados_rep and clave_rep(r) == k]
            if len(cand) != 1:
                continue
            rep = cand[0]
            cuadra_valor = abs(sum(d.vr_ent for d in grupo) - rep.valor_reclamado) <= len(grupo)
            cuadra_cant = abs(sum(d.cantidad for d in grupo) - rep.cant_reclamada) <= 0.01
            # Cuando el CÓDIGO o la DESCRIPCIÓN COMPLETA coinciden, que el valor
            # cuadre exacto ya es señal firme: no hace falta que además cuadre la
            # cantidad. El caso que lo enseñó: el detallado trae los honorarios
            # del cirujano partidos en dos renglones de $320.600 —porque una
            # cirugía se hizo dos veces— y suman exactamente los $641.200 que
            # reclama la única fila del reporte; exigir 2 renglones = 1 unidad
            # dejaba $654.075 sin descontar en el paquete 31068.
            # Por PREFIJO sí se sigue exigiendo la cantidad: ahí a la descripción
            # del reporte le basta con ser el comienzo de la del detallado, y sin
            # ese segundo candado un nombre genérico corto podría llevarse el
            # grupo equivocado.
            if cuadra_valor and (cuadra_cant or not prefijo):
                salida.append((grupo, rep))
                usados_det.update(id(d) for d in grupo)
                usados_rep.add(id(rep))

    intentar(lambda d: _norm_codigo(d.codigo), lambda r: _norm_codigo(r.codigo))
    intentar(lambda d: _norm_desc(d.nombre), lambda r: _norm_desc(r.descripcion))
    intentar(lambda d: _norm_desc(d.nombre), None, prefijo=True)
    return salida


def _porcion(rep: ItemReporte, fraccion: float) -> ItemReporte:
    """Copia del ítem del reporte con los valores prorrateados."""
    filas = []
    for f in rep.filas:
        datos = dict(f.__dict__)
        for campo in (
            "cant_reclamada",
            "valor_reclamado",
            "cant_aprobada",
            "valor_aprobado",
            "valor_glosado",
        ):
            datos[campo] = datos[campo] * fraccion
        datos["usada"] = False
        filas.append(FilaGlosa(**datos))
    return ItemReporte(codigo=rep.codigo, descripcion=rep.descripcion, tipo=rep.tipo, filas=filas)


def decidir(
    item: ItemDetallado, rep: ItemReporte | None, modo_parcial: str, sin_cruce: str
) -> tuple[str, float, float, str]:
    """(accion, cantidad_nueva, vr_ent_nuevo, observacion)."""
    if rep is None:
        if sin_cruce == "quitar":
            return "QUITADO", 0.0, 0.0, "No aparece en el reporte de glosas"
        return (
            "SIN_CRUCE",
            item.cantidad,
            item.vr_ent,
            "No aparece en el reporte de glosas: se conserva, REVISAR a mano",
        )

    glosado = rep.valor_glosado
    if glosado <= TOLERANCIA_PESOS:
        nota = "La entidad ya lo aprobó (valor glosado 0)"
        if rep.valor_reclamado <= TOLERANCIA_PESOS:
            nota = "El reporte no le reconoce valor (reclamado 0): va dentro de otro servicio"
        return "QUITADO", 0.0, 0.0, nota

    if glosado >= item.vr_ent - TOLERANCIA_PESOS:
        nota = "Sigue glosado en su totalidad"
        if glosado > item.vr_ent + TOLERANCIA_PESOS:
            # El reporte del ADRES glosa más de lo que dice la factura impresa.
            # Se conserva el valor de la factura (no se puede cobrar de más),
            # pero hay que revisar de dónde sale la diferencia.
            nota = (
                f"REVISAR: el reporte glosa ${glosado:,.0f} y el detallado dice "
                f"${item.vr_ent:,.0f} (dif ${glosado - item.vr_ent:,.0f}); "
                "se deja el valor del detallado"
            )
        return "CONSERVADO", item.cantidad, item.vr_ent, nota

    # Aprobado a medias.
    if modo_parcial == "conservar":
        return (
            "CONSERVADO",
            item.cantidad,
            item.vr_ent,
            f"Aprobado parcial (glosado ${glosado:,.0f}), se conserva completo por --modo-parcial",
        )
    if modo_parcial == "quitar":
        return (
            "QUITADO",
            0.0,
            0.0,
            f"Aprobado parcial (glosado ${glosado:,.0f}), se quita por --modo-parcial",
        )

    # La cantidad se recalcula con el valor unitario QUE PAGA LA ENTIDAD
    # (vr_ent / cantidad): cuando hay cuota del paciente, vr_unit es el precio
    # completo y no sirve para prorratear lo que sigue glosado.
    unitario = item.vr_ent / item.cantidad if item.cantidad else 0.0
    unidades = glosado / unitario if unitario > 0 else 0.0
    enteras = round(unidades)

    if enteras >= 1 and abs(unidades - enteras) <= 0.01:
        # Glosa POR UNIDADES: la entidad pagó unas y objetó otras.
        cantidad = float(enteras)
        nota = f"Aprobado parcial: de {item.cantidad:g} queda(n) {cantidad:g} sin pagar"
    else:
        # Glosa POR VALOR (diferencia de tarifa, no de unidades): la cantidad
        # NO se toca —redondearla daría 0 y el renglón quedaría inservible—,
        # solo baja el valor a lo que sigue glosado.
        cantidad = item.cantidad
        nota = (
            f"Aprobado parcial por VALOR (no por unidades): quedan las "
            f"{cantidad:g} unidad(es) por ${glosado:,.0f} de ${item.vr_ent:,.0f}"
        )
    if len(rep.filas) > 1:
        nota += f" (sumadas {len(rep.filas)} filas del reporte)"
    return "AJUSTADO", cantidad, glosado, nota


# ─── 5) Edición de la hoja ───────────────────────────────────────────────────


def eliminar_filas(ws, filas: set[int]) -> int:
    """Borra filas de una hoja re-indexándolas en sitio.

    `ws.delete_rows()` mueve las celdas una por una y no corrige ni las celdas
    combinadas ni los altos de fila: en estos archivos (30.000+ filas, 22.000+
    merges, de los que hay que borrar el 60-90 %) tarda minutos. Re-mapear los
    índices de golpe hace lo mismo en menos de un segundo, y de paso corrige
    merges, altos e imágenes.
    """
    from openpyxl.worksheet.cell_range import CellRange

    if not filas:
        return ws.max_row
    max_fila = ws.max_row
    nuevo: dict[int, int | None] = {}
    corridas = 0
    for f in range(1, max_fila + 1):
        if f in filas:
            corridas += 1
            nuevo[f] = None
        else:
            nuevo[f] = f - corridas

    celdas = {}
    for (f, c), celda in ws._cells.items():
        destino = nuevo.get(f)
        if destino is None:
            continue
        celda.row = destino
        celdas[(destino, c)] = celda
    ws._cells = celdas

    rangos = []
    for r in ws.merged_cells.ranges:
        vivas = [nuevo.get(x) for x in range(r.min_row, r.max_row + 1)]
        vivas = [x for x in vivas if x is not None]
        if not vivas or (len(vivas) == 1 and r.min_col == r.max_col):
            continue
        rangos.append(
            CellRange(min_col=r.min_col, min_row=min(vivas), max_col=r.max_col, max_row=max(vivas))
        )
    ws.merged_cells.ranges = rangos

    dims = dict(ws.row_dimensions)
    ws.row_dimensions.clear()
    for f, dim in sorted(dims.items()):
        destino = nuevo.get(f)
        if destino is None:
            continue
        dim.index = destino
        ws.row_dimensions[destino] = dim

    imagenes = getattr(ws, "_images", None)
    if imagenes:
        quedan = []
        for img in imagenes:
            desde = getattr(getattr(img, "anchor", None), "_from", None)
            if desde is None:
                quedan.append(img)
                continue
            destino = nuevo.get(desde.row + 1)
            if destino is None:
                continue  # la imagen vivía en una fila borrada (logo, QR)
            corrido = desde.row + 1 - destino
            desde.row -= corrido
            hasta = getattr(img.anchor, "to", None)
            if hasta is not None:
                hasta.row -= corrido
            quedan.append(img)
        ws._images = quedan

    ws._current_row = max_fila - corridas
    return ws._current_row


def _celdas_numericas(celdas: list[Celda]) -> list[Celda]:
    return [c for c in celdas if isinstance(c.valor, (int, float))]


def recalcular_totales(
    idx: IndiceHoja, fac: FacturaEnHoja, est: EstructuraFactura, subtotal: float, n_items: int
) -> None:
    """Reescribe subtotal, total de la orden y el total en letras.

    En la fila de subtotal hay DOS números: la **cantidad de ítems** y el
    **importe**. El importe es siempre el último de la fila.
    """
    fila_sub = est.fila_subtotal
    numeros = _celdas_numericas(idx.celdas(fila_sub))
    if numeros:
        idx.escribir(numeros[-1], subtotal)
    if len(numeros) > 1:
        idx.escribir(numeros[-2], n_items)

    descuentos = 0.0
    fila_total = -1
    for fila in range(fila_sub + 1, fac.fila_fin + 1):
        texto = _norm(idx.texto_fila(fila))
        if texto.startswith(MARCA_TOTAL_ORDEN):
            fila_total = fila
            break
        if any(texto.startswith(m) for m in (MARCA_COPAGO, MARCA_ANTICIPO, MARCA_COPAGO_USUARIO)):
            numeros = _celdas_numericas(idx.celdas(fila))
            if numeros:
                descuentos += _parse_valor(numeros[-1].valor)

    total = subtotal - descuentos
    if fila_total > 0:
        numeros = _celdas_numericas(idx.celdas(fila_total))
        if numeros:
            idx.escribir(numeros[-1], total)

    for fila in range(fila_total + 1 if fila_total > 0 else fila_sub + 1, fac.fila_fin + 1):
        celdas = idx.celdas(fila)
        if not celdas or not _norm(celdas[0].texto).startswith(MARCA_TOTAL_LETRAS):
            continue
        for celda in celdas[1:]:
            if isinstance(celda.valor, str):
                idx.escribir(celda, numero_a_letras(total))
                break
        break


def procesar_factura(
    idx: IndiceHoja,
    fac: FacturaEnHoja,
    glosas: list[FilaGlosa],
    *,
    hoja: str,
    modo_parcial: str = "ajustar",
    sin_cruce: str = "conservar",
    aplicar: bool = True,
) -> tuple[ResultadoHoja, set[int]]:
    """Poda los ítems ya aprobados de UNA factura y recalcula sus totales.

    Devuelve el resultado y las filas que hay que borrar de la hoja (el borrado
    se hace de una sola vez al final, para no re-indexar la hoja 300 veces).
    """
    res = ResultadoHoja(hoja=hoja, factura=fac.factura, estado="AJUSTADA")
    est = detectar_estructura(idx, fac)
    if est is None:
        res.estado = "SIN_ESTRUCTURA"
        res.observacion = (
            "No encontré la tabla de ítems (CÓDIGO/NOMBRE/CANT) o la fila "
            f"'{MARCA_SUBTOTAL}'. La factura queda sin tocar."
        )
        return res, set()

    bloques = leer_bloques(idx, est)
    items = [i for b in bloques for i in b.items]
    res.items_total = len(items)
    # El valor de la factura sale de su propia fila de subtotal, NO de sumar los
    # renglones: los de desglose repiten el valor del renglón que los encabeza.
    numeros = _celdas_numericas(idx.celdas(est.fila_subtotal))
    res.subtotal_antes = (
        _parse_valor(numeros[-1].valor)
        if numeros
        else sum(i.vr_ent for i in items if not i.desglose)
    )

    items_reporte, reclamacion = agrupar_reporte(glosas)
    pares, sin_pareja = emparejar(items, items_reporte)
    res.reclamacion = reclamacion
    res.glosas_sin_item = [r for r in sin_pareja if r.valor_glosado > TOLERANCIA_PESOS]

    posicion = {id(i): n for n, i in enumerate(items)}
    # Cada renglón de desglose "cuelga" del último renglón numerado anterior.
    padre_de: dict[int, ItemDetallado | None] = {}
    ultimo: ItemDetallado | None = None
    for i in items:
        if i.desglose:
            padre_de[id(i)] = ultimo
        else:
            padre_de[id(i)] = None
            ultimo = i
    sobrevive: dict[int, bool] = {}

    a_borrar: set[int] = set()
    subtotal = 0.0
    vivos_factura = 0

    for bloque in bloques:
        vivos = 0
        for item in bloque.items:
            rep, criterio = pares.get(posicion[id(item)], (None, ""))
            accion, cant, valor, nota = decidir(item, rep, modo_parcial, sin_cruce)
            res.detalle.append(
                ResultadoItem(
                    factura=fac.factura,
                    hoja=hoja,
                    grupo=bloque.titulo,
                    codigo=item.codigo,
                    nombre=item.nombre,
                    cant_orig=item.cantidad,
                    vr_ent_orig=item.vr_ent,
                    valor_reclamado=rep.valor_reclamado if rep else 0.0,
                    valor_aprobado=rep.valor_aprobado if rep else 0.0,
                    valor_glosado=rep.valor_glosado if rep else 0.0,
                    accion=accion,
                    cant_nueva=cant,
                    vr_ent_nuevo=valor,
                    cruce=criterio,
                    filas_reporte=len(rep.filas) if rep else 0,
                    causales=rep.causales if rep else "",
                    desglose=item.desglose,
                    observacion=nota,
                )
            )
            sobrevive[id(item)] = accion != "QUITADO"

            if accion == "QUITADO":
                res.items_quitados += 1
                a_borrar.update(item.filas)
                continue

            vivos += 1
            # Un renglón de desglose solo suma si el renglón que lo encabeza ya
            # NO está: si quedaran los dos, el valor se contaría dos veces.
            padre = padre_de.get(id(item))
            if item.desglose and padre is not None and sobrevive.get(id(padre)):
                res.detalle[
                    -1
                ].observacion += " | no suma al total: su renglón principal sigue en la factura"
            else:
                vivos_factura += 1
                subtotal += valor
            if accion == "AJUSTADO":
                res.items_ajustados += 1
                if aplicar:
                    if item.celda_cantidad is not None:
                        idx.escribir(item.celda_cantidad, cant)
                    if item.celda_vr_ent is not None:
                        idx.escribir(item.celda_vr_ent, valor)
            elif accion == "SIN_CRUCE":
                res.items_sin_cruce += 1
            else:
                res.items_conservados += 1

        # Grupo que se quedó sin ítems: se borra el bloque completo (título,
        # renglones en blanco y todo), igual que se hace a mano.
        if bloque.fila_titulo is not None and bloque.items and vivos == 0:
            a_borrar.update(range(bloque.fila_titulo, bloque.fila_fin + 1))

    res.subtotal_despues = subtotal
    if not aplicar:
        return res, set()

    celda = fac.celda_titulo
    if _norm(celda.texto).startswith(TITULO_ORIGINAL):
        idx.escribir(celda, TITULO_NUEVO)
    recalcular_totales(idx, fac, est, subtotal, vivos_factura)
    return res, a_borrar


# ─── Orquestación ────────────────────────────────────────────────────────────


def procesar_libro(
    ruta_detallado: Path,
    facturas: list[str],
    glosas_por_factura: dict[str, list[FilaGlosa]],
    *,
    modo_parcial: str = "ajustar",
    sin_cruce: str = "conservar",
    aplicar: bool = True,
    salida: Path | None = None,
) -> list[ResultadoHoja]:
    wb = _abrir_libro(ruta_detallado)
    quiero = {normalizar_factura(f) for f in facturas}
    resultados: list[ResultadoHoja] = []

    for ws in list(wb.worksheets):
        idx = IndiceHoja(ws)
        enHoja = segmentar_facturas(idx)
        if not enHoja:
            logger.warning(
                "%s / hoja '%s': no encontré ninguna factura", ruta_detallado.name, ws.title
            )
            continue
        logger.info(
            "%s / hoja '%s': %d factura(s), %d filas",
            ruta_detallado.name,
            ws.title,
            len(enHoja),
            idx.max_fila,
        )

        a_borrar = filas_encabezado_institucional(idx, enHoja[0].fila_ini) if aplicar else set()
        for fac in enHoja:
            clave = normalizar_factura(fac.factura)
            if quiero and clave not in quiero:
                resultados.append(
                    ResultadoHoja(
                        hoja=ws.title,
                        factura=fac.factura,
                        estado="ELIMINADA",
                        observacion="No está en el consolidado de facturas a trabajar",
                    )
                )
                if aplicar:
                    a_borrar.update(range(fac.fila_ini, fac.fila_fin + 1))
                continue

            glosas = glosas_por_factura.get(clave, [])
            if not glosas:
                resultados.append(
                    ResultadoHoja(
                        hoja=ws.title,
                        factura=fac.factura,
                        estado="SIN_GLOSAS",
                        observacion=(
                            "No tiene filas en el reporte de glosas: la factura queda "
                            "tal cual. REVISAR (¿otro paquete?)"
                        ),
                    )
                )
                continue

            res, filas = procesar_factura(
                idx,
                fac,
                [FilaGlosa(**{**g.__dict__, "usada": False}) for g in glosas],
                hoja=ws.title,
                modo_parcial=modo_parcial,
                sin_cruce=sin_cruce,
                aplicar=aplicar,
            )
            resultados.append(res)
            a_borrar.update(filas)

        if aplicar and a_borrar:
            quedan = eliminar_filas(ws, a_borrar)
            logger.info("   se borran %d filas; la hoja queda con %d", len(a_borrar), quedan)

    trabajadas = sum(1 for r in resultados if r.estado not in ("ELIMINADA",))
    if aplicar and salida is not None:
        if not trabajadas:
            # Este archivo no trae ninguna factura del consolidado: escribir la
            # salida solo dejaría una hoja vacía.
            logger.warning(
                "%s: ninguna de sus %d facturas está en el consolidado; no se escribe salida",
                ruta_detallado.name,
                len(resultados),
            )
        else:
            salida.parent.mkdir(parents=True, exist_ok=True)
            wb.save(salida)
    wb.close()
    return resultados


CAMPOS_CSV = [
    "FACTURA",
    "HOJA",
    "GRUPO",
    "CODIGO",
    "NOMBRE",
    "CANT_ORIGINAL",
    "VR_ENT_ORIGINAL",
    "VALOR_RECLAMADO",
    "VALOR_APROBADO",
    "VALOR_GLOSADO",
    "ACCION",
    "CANT_NUEVA",
    "VR_ENT_NUEVO",
    "CRUCE_POR",
    "FILAS_REPORTE",
    "CAUSALES_GLOSA",
    "TIPO_RENGLON",
    "OBSERVACION",
]


def escribir_bitacora(ruta: Path, resultados: list[ResultadoHoja]) -> None:
    ruta.parent.mkdir(parents=True, exist_ok=True)
    with ruta.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh, delimiter=";")
        w.writerow(CAMPOS_CSV)
        for hoja in resultados:
            if not hoja.detalle:
                w.writerow(
                    [
                        hoja.factura,
                        hoja.hoja,
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        hoja.estado,
                        "",
                        "",
                        "",
                        "",
                        hoja.observacion,
                    ]
                )
                continue
            for d in hoja.detalle:
                w.writerow(
                    [
                        d.factura,
                        d.hoja,
                        d.grupo,
                        d.codigo,
                        d.nombre,
                        f"{d.cant_orig:g}",
                        f"{d.vr_ent_orig:.0f}",
                        f"{d.valor_reclamado:.0f}",
                        f"{d.valor_aprobado:.0f}",
                        f"{d.valor_glosado:.0f}",
                        d.accion,
                        f"{d.cant_nueva:g}",
                        f"{d.vr_ent_nuevo:.0f}",
                        d.cruce,
                        d.filas_reporte,
                        d.causales,
                        "DESGLOSE" if d.desglose else "ITEM",
                        d.observacion,
                    ]
                )
            # Glosas del reporte que no encontraron su item en el detallado y
            # glosas a la reclamacion completa: el auditor tiene que verlas.
            for r in hoja.glosas_sin_item:
                w.writerow(
                    [
                        hoja.factura,
                        hoja.hoja,
                        "",
                        r.codigo,
                        r.descripcion,
                        f"{r.cant_reclamada:g}",
                        "",
                        f"{r.valor_reclamado:.0f}",
                        f"{r.valor_aprobado:.0f}",
                        f"{r.valor_glosado:.0f}",
                        "GLOSA_SIN_ITEM",
                        "",
                        "",
                        "",
                        len(r.filas),
                        r.causales,
                        "El reporte glosa este item pero no aparece en el detallado. REVISAR",
                    ]
                )
            for g in hoja.reclamacion:
                w.writerow(
                    [
                        hoja.factura,
                        hoja.hoja,
                        "",
                        "",
                        "(toda la reclamación)",
                        "",
                        "",
                        f"{g.valor_reclamado:.0f}",
                        f"{g.valor_aprobado:.0f}",
                        f"{g.valor_glosado:.0f}",
                        "GLOSA_RECLAMACION",
                        "",
                        "",
                        "",
                        1,
                        g.descripcion_glosa,
                        "Glosa a la reclamación completa, no a un ítem: no se descuenta del detallado",
                    ]
                )


def _resumir(resultados: list[ResultadoHoja]) -> None:
    ajustadas = [r for r in resultados if r.estado == "AJUSTADA"]
    eliminadas = [r for r in resultados if r.estado == "ELIMINADA"]
    sin_glosas = [r for r in resultados if r.estado == "SIN_GLOSAS"]
    sin_estruct = [r for r in resultados if r.estado == "SIN_ESTRUCTURA"]

    logger.info("─" * 68)
    logger.info("Hojas ajustadas .......... %d", len(ajustadas))
    logger.info("Hojas eliminadas ......... %d (no estaban en el consolidado)", len(eliminadas))
    if sin_glosas:
        logger.warning(
            "Hojas sin glosas ......... %d  → %s",
            len(sin_glosas),
            ", ".join(r.factura for r in sin_glosas[:10]),
        )
    if sin_estruct:
        logger.warning(
            "Hojas sin estructura ..... %d  → %s",
            len(sin_estruct),
            ", ".join(r.factura for r in sin_estruct[:10]),
        )
    if not ajustadas:
        return
    logger.info(
        "Ítems: %d quitados, %d ajustados, %d conservados, %d sin cruce",
        sum(r.items_quitados for r in ajustadas),
        sum(r.items_ajustados for r in ajustadas),
        sum(r.items_conservados for r in ajustadas),
        sum(r.items_sin_cruce for r in ajustadas),
    )
    logger.info(
        "Valor: antes $%s → sigue glosado $%s",
        f"{sum(r.subtotal_antes for r in ajustadas):,.0f}",
        f"{sum(r.subtotal_despues for r in ajustadas):,.0f}",
    )
    revisar = [d for r in ajustadas for d in r.detalle if d.accion == "SIN_CRUCE"]
    if revisar:
        logger.warning(
            "REVISAR: %d ítem(s) no cruzaron con el reporte de glosas "
            "(quedaron en la factura). Ver la bitácora CSV.",
            len(revisar),
        )


def _diagnosticar(resultados: list[ResultadoHoja]) -> None:
    for r in resultados:
        if r.estado == "ELIMINADA":
            logger.info("%-14s hoja '%s' → SE ELIMINA (%s)", r.factura, r.hoja, r.observacion)
            continue
        logger.info(
            "%-14s hoja '%s' → %s | ítems %d (quitar %d, ajustar %d, dejar %d, revisar %d)"
            " | $%s → $%s",
            r.factura,
            r.hoja,
            r.estado,
            r.items_total,
            r.items_quitados,
            r.items_ajustados,
            r.items_conservados,
            r.items_sin_cruce,
            f"{r.subtotal_antes:,.0f}",
            f"{r.subtotal_despues:,.0f}",
        )
        if r.observacion:
            logger.info("               %s", r.observacion)
        for d in r.detalle:
            logger.info(
                "               [%-10s] %-12s %-42s %g → %g  ($%s → $%s) %s",
                d.accion,
                d.codigo,
                d.nombre[:42],
                d.cant_orig,
                d.cant_nueva,
                f"{d.vr_ent_orig:,.0f}",
                f"{d.vr_ent_nuevo:,.0f}",
                f"[{d.cruce}]" if d.cruce else "[sin cruce]",
            )


def construir_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Deja el detallado de factura solo con lo que sigue glosado.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument(
        "--consolidado",
        type=Path,
        help="Excel/CSV con las facturas a trabajar (se le quitan los duplicados). "
        "Si se omite, se trabajan TODAS las hojas del detallado.",
    )
    p.add_argument(
        "--detallado",
        type=Path,
        nargs="+",
        required=True,
        help="Excel del detallado (una hoja por factura). Admite varios archivos.",
    )
    p.add_argument(
        "--reporte-glosas", type=Path, required=True, help="ReporteGlosasReclamPAQUETE NNNNN.xlsx"
    )
    p.add_argument(
        "--salida", type=Path, help="Excel corregido (o carpeta, si se pasan varios --detallado)."
    )
    p.add_argument("--reporte-csv", type=Path, help="Bitácora CSV ítem por ítem.")
    p.add_argument("--paquete", help="Filtrar el reporte de glosas por número de paquete.")
    p.add_argument(
        "--modo-parcial",
        choices=("ajustar", "conservar", "quitar"),
        default="ajustar",
        help="Qué hacer con los ítems que la entidad aprobó a medias. "
        "ajustar (por defecto): deja solo la parte que sigue glosada.",
    )
    p.add_argument(
        "--sin-cruce",
        choices=("conservar", "quitar"),
        default="conservar",
        help="Qué hacer con los ítems de la factura que NO aparecen en el reporte "
        "de glosas. Por defecto se conservan y se marcan para revisión.",
    )
    p.add_argument(
        "--diagnostico",
        action="store_true",
        help="Solo analiza y muestra qué haría; no escribe el Excel de salida.",
    )
    p.add_argument("--verbose", action="store_true", help="Log detallado.")
    return p


def main(argv: list[str] | None = None) -> int:
    args = construir_parser().parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)-8s %(message)s",
    )

    if not args.diagnostico and not args.salida:
        logger.error("Falta --salida (o usá --diagnostico para solo ver qué haría).")
        return 2

    facturas: list[str] = []
    if args.consolidado:
        if not args.consolidado.exists():
            logger.error("No existe el consolidado: %s", args.consolidado)
            return 2
        facturas, duplicadas = leer_consolidado(args.consolidado)
        logger.info(
            "Consolidado: %d factura(s) únicas, %d duplicada(s) descartada(s)",
            len(facturas),
            len(duplicadas),
        )
        if duplicadas:
            logger.info("Duplicadas: %s", ", ".join(duplicadas[:20]))
        if not facturas:
            logger.error("El consolidado no trajo ninguna factura reconocible.")
            return 2
    else:
        logger.warning("Sin --consolidado: se trabajan TODAS las hojas del detallado.")

    if not args.reporte_glosas.exists():
        logger.error("No existe el reporte de glosas: %s", args.reporte_glosas)
        return 2
    glosas = leer_reporte_glosas(args.reporte_glosas, paquete=args.paquete)
    logger.info(
        "Reporte de glosas: %d factura(s), %d fila(s)",
        len(glosas),
        sum(len(v) for v in glosas.values()),
    )

    if facturas:
        faltan = [f for f in facturas if normalizar_factura(f) not in glosas]
        if faltan:
            logger.warning(
                "%d factura(s) del consolidado no están en el reporte de glosas: %s",
                len(faltan),
                ", ".join(faltan[:20]),
            )

    varios = len(args.detallado) > 1
    resultados: list[ResultadoHoja] = []
    for ruta in args.detallado:
        if not ruta.exists():
            logger.error("No existe el detallado: %s", ruta)
            return 2
        salida = None
        if args.salida and not args.diagnostico:
            salida = (args.salida / f"{ruta.stem}_AJUSTADO.xlsx") if varios else args.salida
        logger.info("Procesando %s …", ruta.name)
        resultados.extend(
            procesar_libro(
                ruta,
                facturas,
                glosas,
                modo_parcial=args.modo_parcial,
                sin_cruce=args.sin_cruce,
                aplicar=not args.diagnostico,
                salida=salida,
            )
        )
        if salida and salida.exists():
            logger.info("Escrito: %s", salida)

    if args.diagnostico:
        _diagnosticar(resultados)
    _resumir(resultados)

    if args.reporte_csv:
        escribir_bitacora(args.reporte_csv, resultados)
        logger.info("Bitácora: %s", args.reporte_csv)
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
