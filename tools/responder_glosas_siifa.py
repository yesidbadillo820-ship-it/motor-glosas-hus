#!/usr/bin/env python3
"""responder_glosas_siifa.py — Bot (sin navegador) para cargar en SIIFA las
respuestas de glosas del HUS, leyendo un Excel tipificado igual que el bot de
COOSALUD (tools/responder_glosas_coosalud.py).

SIIFA sí tiene API oficial (a diferencia de COOSALUD/SIMED), así que este
script no abre Playwright: llama directo a
PUT /api/SeguimientoFacturaGlosa/Respuesta (o .../ReiteracionRespuesta para
la subsanación) por cada glosa del Excel. Ver docs/CONTEXTO_SIIFA.md.

EXCEL DE ENTRADA — columnas (acepta con o sin guion bajo, mayus/minus):
    ID_SEGUIMIENTO_FACTURA_GLOSA  → el id que trae tools/siifa_reporte_seguimientos.py
    NUMERO_FACTURA                → solo informativo, para el reporte (opcional)
    CODIGO_RESPUESTA              → código de la Tabla 1.4 Res. 2284/2023 (ej. RESP01)
    OBSERVACION_RESPUESTA         → texto de sustento (opcional pero recomendado)
    FECHA_RESPUESTA                → AAAA-MM-DD (opcional, default: hoy)

PASO A PASO RECOMENDADO:
    1. Correr tools/siifa_reporte_seguimientos.py --sin-respuesta para saber
       qué glosas faltan por responder y conseguir el id_seguimiento_factura_glosa
       de cada una.
    2. Tipificar en ese mismo Excel (o en uno aparte) las columnas
       CODIGO_RESPUESTA y OBSERVACION_RESPUESTA.
    3. Revisar los códigos válidos: --listar-catalogo (no necesita Excel).
    4. PILOTO de 1 sola glosa antes del cargue masivo (regla del repo):
           --solo-id <id>
    5. Cargue completo, con reporte para poder reintentar solo lo que falle:
           --reporte "D:\\...\\reporte_siifa.csv"
    6. Si el cargue se corta, relanzar con --saltar-csv <reporte anterior>.

CREDENCIALES (variables de entorno, NUNCA en el código):
    setx SIIFA_USER <usuario_sispro>
    setx SIIFA_PASSWORD <password>
    setx SIIFA_AUTH_URL <url del servicio de Auth>   (ver docs/CONTEXTO_SIIFA.md — no confirmada)

USO:
    REM Ver los códigos de respuesta válidos
    py responder_glosas_siifa.py --listar-catalogo

    REM Piloto: 1 sola glosa
    py responder_glosas_siifa.py --excel "D:\\...\\respuestas_siifa.xlsx" ^
        --solo-id 123456 --reporte "D:\\...\\piloto_siifa.csv"

    REM Cargue completo
    py responder_glosas_siifa.py --excel "D:\\...\\respuestas_siifa.xlsx" ^
        --reporte "D:\\...\\reporte_siifa.csv"

    REM Reintento de lo que quedó pendiente
    py responder_glosas_siifa.py --excel "D:\\...\\respuestas_siifa.xlsx" ^
        --saltar-csv "D:\\...\\reporte_siifa.csv" ^
        --reporte "D:\\...\\reporte_siifa_pass2.csv"

    REM Paso de subsanación (glosa reiterada por la EPS, no levantada)
    py responder_glosas_siifa.py --excel "D:\\...\\subsanaciones.xlsx" ^
        --accion reiteracion-respuesta --reporte "D:\\...\\reporte_subsanacion.csv"

INSTALACIÓN (una vez):
    py -m pip install httpx openpyxl
"""

from __future__ import annotations

import argparse
import csv
import logging
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from siifa_client import (  # noqa: E402
    SiifaApiError,
    SiifaClient,
    buscar_clave,
    credenciales_desde_env,
)

logger = logging.getLogger("responder_glosas_siifa")

ESTADOS_TERMINALES = {"OK"}

COLUMNAS = {
    "id": {"ID_SEGUIMIENTO_FACTURA_GLOSA", "ID SEGUIMIENTO FACTURA GLOSA", "ID_GLOSA", "ID GLOSA"},
    "factura": {"NUMERO_FACTURA", "NUMERO FACTURA"},
    "codigo": {"CODIGO_RESPUESTA", "CODIGO RESPUESTA", "COD_RESPUESTA", "COD RESPUESTA GLOSA"},
    "observacion": {"OBSERVACION_RESPUESTA", "OBSERVACION RESPUESTA", "OBSERVACION RTA GLOSA"},
    "fecha": {"FECHA_RESPUESTA", "FECHA RESPUESTA"},
    "tipo": {"TIPO", "TIPO_SEGUIMIENTO", "TIPO SEGUIMIENTO"},
    "id_devolucion": {
        "ID_SEGUIMIENTO_FACTURA_DEVOLUCION",
        "ID SEGUIMIENTO FACTURA DEVOLUCION",
    },
}


def _norm_header(h: str) -> str:
    s = unicodedata.normalize("NFKD", (h or "").strip().upper())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.split())


def leer_excel(ruta: Path, hoja: str | None) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("ERROR: falta openpyxl. py -m pip install openpyxl")

    wb = load_workbook(filename=str(ruta), data_only=True, read_only=True)
    nombre_hoja = hoja or wb.sheetnames[0]
    if nombre_hoja not in wb.sheetnames:
        objetivo = nombre_hoja.strip().casefold()
        real = next((s for s in wb.sheetnames if s.strip().casefold() == objetivo), None)
        if real is None:
            raise ValueError(f"El Excel no tiene la hoja '{nombre_hoja}'. Hojas: {wb.sheetnames}")
        nombre_hoja = real
    ws = wb[nombre_hoja]
    rows = ws.iter_rows(values_only=True)
    headers = [_norm_header(str(h or "")) for h in next(rows)]

    idx: dict[str, int] = {}
    for clave, opciones in COLUMNAS.items():
        for i, h in enumerate(headers):
            if h in opciones:
                idx[clave] = i
                break
        if clave not in idx and clave in ("id", "codigo"):
            raise ValueError(
                f"Falta la columna '{clave}' (acepto {sorted(opciones)}). Headers: {headers}"
            )

    filas = []
    for r in rows:
        if r is None:
            continue
        id_raw = r[idx["id"]] if idx.get("id") is not None else None
        if id_raw in (None, ""):
            continue
        try:
            id_seguimiento = int(id_raw)
        except (TypeError, ValueError):
            logger.warning(
                "Fila con id_seguimiento_factura_glosa no numérico, se omite: %r", id_raw
            )
            continue
        codigo = str(r[idx["codigo"]] or "").strip() if idx.get("codigo") is not None else ""
        if not codigo:
            logger.warning("Glosa %s sin CODIGO_RESPUESTA, se omite.", id_seguimiento)
            continue
        filas.append(
            {
                "id": id_seguimiento,
                "factura": str(r[idx["factura"]] or "").strip()
                if idx.get("factura") is not None
                else "",
                "codigo": codigo,
                "observacion": str(r[idx["observacion"]] or "").strip()
                if idx.get("observacion") is not None
                else "",
                "fecha": _fecha_iso(r[idx["fecha"]]) if idx.get("fecha") is not None else None,
                "tipo": str(r[idx["tipo"]] or "").strip().upper()
                if idx.get("tipo") is not None
                else "",
                "id_devolucion": _entero_o_none(r[idx["id_devolucion"]])
                if idx.get("id_devolucion") is not None
                else None,
            }
        )
    return filas


def _entero_o_none(valor) -> int | None:
    try:
        return int(valor)
    except (TypeError, ValueError):
        return None


def _fecha_iso(valor) -> str | None:
    if valor in (None, ""):
        return None
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%dT12:00:00Z")
    texto = str(valor).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(texto, fmt).strftime("%Y-%m-%dT12:00:00Z")
        except ValueError:
            continue
    logger.warning("Fecha de respuesta '%s' no reconocida, se usa la fecha de hoy.", texto)
    return None


def ids_ya_procesados(rutas_csv: list[str]) -> set[int]:
    vistos: set[int] = set()
    for ruta in rutas_csv:
        with open(ruta, newline="", encoding="utf-8-sig") as f:
            for fila in csv.DictReader(f):
                if fila.get("estado") in ESTADOS_TERMINALES:
                    try:
                        vistos.add(int(fila["id_seguimiento_factura_glosa"]))
                    except (KeyError, ValueError):
                        continue
    return vistos


# Nombres de grupo que puede estar usando SIIFA. Con «RESPUESTA» a secas el
# catálogo vino vacío el 05-08-2026, y sin la lista de códigos toca adivinar
# cuál corresponde a cada frase del desplegable del portal —que muestra el
# texto pero no el código—. Se prueban todos y se muestra lo que responda.
GRUPOS_CANDIDATOS = [
    # Los nombres los dijo la propia API en sus mensajes de validación:
    # «no pertenece al grupo RESPUESTA» (glosas) y «no pertenece al grupo
    # RESPUESTA_DEV_PTS_PSS» (devoluciones, 05-08-2026).
    "RESPUESTA_DEV_PTS_PSS",
    "RESPUESTA_GLOSA_PTS_PSS",
    "RESPUESTA",
    "RESPUESTA_GLOSA",
    "RESPUESTAGLOSA",
    "RESPUESTA_DEVOLUCION",
    "RESPUESTADEVOLUCION",
    "RESPUESTA_REITERACION",
    "GLOSA",
    "DEVOLUCION",
    "REITERACION",
]


def _codigo_y_descripcion(fila: dict) -> tuple[str, str]:
    """La API no siempre nombra igual las columnas del catálogo."""
    codigo = buscar_clave(
        fila, "idSeguimientoTipoCodigo", "codigo", "id", "IdSeguimientoTipoCodigo", default="?"
    )
    desc = buscar_clave(fila, "descripcion", "Descripcion", "nombre", "detalle", default="")
    return str(codigo), str(desc)


def listar_catalogo(cliente: SiifaClient, grupo: str, guardar: str | None = None) -> None:
    """Muestra los códigos que acepta SIIFA, con su significado.

    Si el grupo que se pide viene vacío, se prueban los demás nombres
    conocidos: lo que importa es salir con la tabla de códigos, porque el
    portal muestra la frase («la devolución ha sido aceptada al 100%») pero
    no el código que hay que mandarle a la API.
    """
    grupos = [grupo] + [g for g in GRUPOS_CANDIDATOS if g != grupo]
    lineas: list[str] = []
    encontrados = 0
    for g in grupos:
        try:
            codigos = cliente.catalogo_tipo_codigo(g)
        except SiifaApiError as exc:
            lineas.append(f"\nGrupo {g}: no responde ({exc})")
            continue
        if not codigos:
            lineas.append(f"\nGrupo {g}: vacío")
            continue
        encontrados += len(codigos)
        lineas.append(f"\nGrupo {g} ({len(codigos)} códigos):")
        for c in codigos:
            cod, desc = _codigo_y_descripcion(c)
            lineas.append(f"  {cod:<12} {desc}")
        if g == grupo:
            break  # el que se pidió sirvió: no hace falta probar los demás

    texto = "\nCatálogo de códigos de SIIFA" + "\n".join(lineas) + "\n"
    print(texto)
    if not encontrados:
        print(
            "SIIFA no devolvió ningún código en ninguno de los grupos conocidos.\n"
            "Para saber qué código corresponde a cada frase del desplegable:\n"
            "  1. responder UNA glosa (o devolución) a mano en el portal,\n"
            "  2. abrir los tres puntos → Ver Histórico,\n"
            "  3. ahí aparece el código que quedó registrado.\n"
        )
    if guardar:
        ruta = Path(guardar)
        ruta.parent.mkdir(parents=True, exist_ok=True)
        ruta.write_text(texto, encoding="utf-8")
        print(f"Guardado en: {ruta}\n")


def procesar(
    cliente: SiifaClient, filas: list[dict], accion: str, ya_ok: set[int], reporte_writer
) -> tuple[int, int]:
    ok = 0
    err = 0
    for fila in filas:
        id_seg = fila["id"]
        if id_seg in ya_ok:
            reporte_writer.writerow(
                {
                    "id_seguimiento_factura_glosa": id_seg,
                    "numero_factura": fila["factura"],
                    "estado": "YA_OK_PREVIO",
                    "detalle": "saltada por --saltar-csv",
                    "fecha_hora": datetime.now().isoformat(timespec="seconds"),
                }
            )
            continue
        es_devolucion = fila.get("tipo") == "DEVOLUCION"
        # El id de una devolución es EL MISMO id de seguimiento. Se confirmó
        # con el volcado crudo de la API (05-08-2026, factura HUS494196): un
        # seguimiento de tipo DEVOLUCION trae idSeguimientoFactura y nada más
        # —no existe un idSeguimientoFacturaDevolucion aparte—. Lo que cambia
        # entre glosa y devolución no es el id: es la puerta por la que se
        # manda y la lista de códigos contra la que se valida.
        id_a_mandar = fila.get("id_devolucion") or id_seg if es_devolucion else id_seg
        try:
            if accion != "respuesta":
                cliente.responder_reiteracion_glosa(
                    id_seg, fila["codigo"], fila["observacion"], fila["fecha"]
                )
            elif es_devolucion:
                # Una devolución es otra entidad en SIIFA: otra puerta, otro id
                # y otra lista de códigos de respuesta.
                cliente.responder_devolucion(
                    id_a_mandar, fila["codigo"], fila["observacion"], fila["fecha"]
                )
            else:
                cliente.responder_glosa(id_seg, fila["codigo"], fila["observacion"], fila["fecha"])
            estado, detalle = "OK", ""
            ok += 1
            logger.info(
                "%s %s (factura %s): OK",
                "Devolución" if es_devolucion else "Glosa",
                id_a_mandar,
                fila["factura"],
            )
        except SiifaApiError as exc:
            estado, detalle = "ERROR", str(exc)
            err += 1
            logger.error(
                "%s %s (factura %s): ERROR — %s",
                "Devolución" if es_devolucion else "Glosa",
                id_a_mandar,
                fila["factura"],
                exc,
            )
        reporte_writer.writerow(
            {
                "id_seguimiento_factura_glosa": id_seg,
                "numero_factura": fila["factura"],
                "estado": estado,
                "detalle": detalle,
                "fecha_hora": datetime.now().isoformat(timespec="seconds"),
            }
        )
    return ok, err


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--excel", help="Excel con las respuestas tipificadas.")
    ap.add_argument("--hoja", help="Nombre de la hoja (default: la primera).")
    ap.add_argument(
        "--accion",
        choices=["respuesta", "reiteracion-respuesta"],
        default="respuesta",
        help="'respuesta' = responder la glosa formulada (default). "
        "'reiteracion-respuesta' = subsanación de una glosa que la EPS reiteró.",
    )
    ap.add_argument(
        "--solo-id",
        type=int,
        help="Piloto: procesa solo esta glosa (id_seguimiento_factura_glosa).",
    )
    ap.add_argument(
        "--piloto", type=int, metavar="N", help="Procesa solo las primeras N filas del Excel."
    )
    ap.add_argument("--reporte", help="CSV de salida con el resultado de cada fila.")
    ap.add_argument(
        "--saltar-csv", action="append", default=[], help="Reporte(s) previos: se saltan los ya OK."
    )
    ap.add_argument(
        "--listar-catalogo",
        metavar="GRUPO",
        nargs="?",
        const="RESPUESTA",
        help="Solo imprime el catálogo oficial de códigos (default: RESPUESTA) y termina. No necesita --excel.",
    )
    ap.add_argument(
        "--guardar-catalogo",
        help="Archivo de texto donde dejar el catálogo, para poder mandarlo.",
    )
    ap.add_argument("--verbose", action="store_true")
    args = ap.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s  %(message)s",
        datefmt="%H:%M:%S",
        stream=sys.stdout,
    )
    if not args.verbose:
        # httpx habla en inglés y con detalle técnico: sólo con --verbose.
        logging.getLogger("httpx").setLevel(logging.WARNING)
        logging.getLogger("httpcore").setLevel(logging.WARNING)

    usuario, password = credenciales_desde_env()

    with SiifaClient() as cliente:
        try:
            cliente.login(usuario, password)
        except SiifaApiError as exc:
            raise SystemExit(f"ERROR de autenticación: {exc}")

        if args.listar_catalogo is not None:
            listar_catalogo(cliente, args.listar_catalogo, args.guardar_catalogo)
            return

        if not args.excel:
            raise SystemExit(
                "ERROR: falta --excel (o usar --listar-catalogo para solo consultar códigos)."
            )
        if not args.reporte:
            raise SystemExit("ERROR: falta --reporte (ruta del CSV de salida).")

        filas = leer_excel(Path(args.excel), args.hoja)
        if args.solo_id:
            filas = [f for f in filas if f["id"] == args.solo_id]
        if args.piloto:
            filas = filas[: args.piloto]
        if not filas:
            logger.warning("No hay filas para procesar con esos filtros.")
            return

        ya_ok = ids_ya_procesados(args.saltar_csv)
        logger.info(
            "%d fila(s) a procesar (acción: %s, %d ya OK en corridas previas).",
            len(filas),
            args.accion,
            len(ya_ok),
        )

        ruta_reporte = Path(args.reporte)
        ruta_reporte.parent.mkdir(parents=True, exist_ok=True)
        with open(ruta_reporte, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=[
                    "id_seguimiento_factura_glosa",
                    "numero_factura",
                    "estado",
                    "detalle",
                    "fecha_hora",
                ],
            )
            writer.writeheader()
            ok, err = procesar(cliente, filas, args.accion, ya_ok, writer)

    logger.info("Terminado: %d OK, %d con error. Reporte: %s", ok, err, ruta_reporte)
    if err:
        logger.warning(
            "Hubo errores — revisar el CSV y reintentar con --saltar-csv %s", ruta_reporte
        )


if __name__ == "__main__":
    main()
