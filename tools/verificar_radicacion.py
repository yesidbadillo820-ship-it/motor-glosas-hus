"""verificar_radicacion.py — Verifica un lote ANTES de radicarlo ante la EPS,
para PREVENIR glosas en vez de responderlas. Usado por VERIFICAR_RADICACION.cmd.

Revisa la carpeta del lote y saca un Excel semáforo con:

    1. XML de cada factura: que traiga NUMERO_CONTRATO diligenciado, la causa
       de cobertura (FACTURA_SIN_CONTRATO), y la validación de la DIAN (02).
       Sin eso, la glosa "sin contrato" o "documento inválido" es casi segura.
    2. FURIPS (.txt): que todas las líneas tengan el MISMO número de campos
       (una línea trunca o corrida hace rechazar el paquete completo).
    3. Soportes: que exista al menos un archivo de soporte (PDF u otro) con el
       número de cada factura en el nombre.

Puede filtrar por lista de facturas (una por línea) o revisar todo el lote.

USO:
    py tools\\verificar_radicacion.py CARPETA_DEL_LOTE
    py tools\\verificar_radicacion.py CARPETA --lista facturas.txt --salida VERIFICACION.xlsx

Requiere openpyxl (para el Excel). El repo ya lo fija.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from xml.etree import ElementTree as ET

CAMPOS_SALUD = [
    "NUMERO_CONTRATO",
    "NUMERO_POLIZA",
    "FACTURA_SIN_CONTRATO",
    "MODALIDAD_PAGO",
    "COBERTURA_PLAN_BENEFICIOS",
    "CODIGO_PRESTADOR",
]

_EXT_SOPORTE = {".pdf", ".jpg", ".jpeg", ".png", ".tif", ".tiff", ".doc", ".docx", ".zip"}


def _local(tag: str) -> str:
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def _texto(elem, nombre: str) -> str:
    for e in elem.iter():
        if _local(e.tag) == nombre and e.text and e.text.strip():
            return e.text.strip()
    return ""


def _cdata_con(raiz, marcador: str) -> str | None:
    for e in raiz.iter():
        if _local(e.tag) == "Description" and e.text and marcador in e.text:
            return e.text
    return None


def _parse_xml(texto: str):
    limpio = re.sub(r"^\s*<\?xml[^>]*\?>", "", texto).strip()
    return ET.fromstring(limpio)


def normalizar_factura(valor: str) -> str:
    digitos = re.sub(r"\D", "", valor or "")
    return digitos.lstrip("0") or digitos


def revisar_xml(ruta: Path) -> dict:
    """Extrae del XML lo necesario para el chequeo pre-radicación."""
    datos = {
        "archivo": ruta.name,
        "factura": "",
        "eps": "",
        "fecha": "",
        "valor": "",
        "hallazgos": [],
    }
    for c in CAMPOS_SALUD:
        datos[c] = ""
    try:
        raiz = ET.parse(str(ruta)).getroot()
    except Exception as exc:
        datos["hallazgos"].append(f"XML ilegible ({type(exc).__name__})")
        return datos

    datos["factura"] = _texto(raiz, "ParentDocumentID")
    datos["eps"] = _texto(raiz, "RegistrationName")
    validacion = _texto(raiz, "ValidationResultCode")

    cdata = _cdata_con(raiz, "<Invoice")
    if not cdata:
        datos["hallazgos"].append("sin Invoice embebido (no parece factura)")
        return datos
    try:
        inv = _parse_xml(cdata)
    except Exception as exc:
        datos["hallazgos"].append(f"Invoice embebido ilegible ({type(exc).__name__})")
        return datos

    if not datos["factura"]:
        datos["factura"] = _texto(inv, "ID")
    datos["fecha"] = _texto(inv, "IssueDate")
    datos["valor"] = _texto(inv, "PayableAmount")

    for ai in inv.iter():
        if _local(ai.tag) != "AdditionalInformation":
            continue
        nombre = valor = ""
        for hijo in ai:
            if _local(hijo.tag) == "Name":
                nombre = (hijo.text or "").strip()
            elif _local(hijo.tag) == "Value":
                valor = (hijo.text or "").strip()
        if nombre in CAMPOS_SALUD:
            datos[nombre] = valor

    # Los chequeos que previenen las glosas más comunes.
    if not datos["NUMERO_CONTRATO"]:
        datos["hallazgos"].append("NUMERO_CONTRATO vacio (riesgo de glosa 'sin contrato')")
    if validacion != "02":
        datos["hallazgos"].append(f"validacion DIAN '{validacion or 'ausente'}' (debe ser 02)")
    if not datos["valor"]:
        datos["hallazgos"].append("sin valor (PayableAmount)")
    return datos


def revisar_furips(ruta: Path) -> dict:
    """Todas las líneas del FURIPS deben tener el mismo número de campos."""
    res = {"archivo": ruta.name, "lineas": 0, "campos": 0, "hallazgos": []}
    try:
        crudo = ruta.read_bytes()
    except OSError as exc:
        res["hallazgos"].append(f"no se pudo leer ({type(exc).__name__})")
        return res
    try:
        texto = crudo.decode("utf-8-sig")
    except UnicodeDecodeError:
        texto = crudo.decode("cp1252", errors="replace")
    lineas = [ln for ln in texto.splitlines() if ln.strip()]
    res["lineas"] = len(lineas)
    if not lineas:
        res["hallazgos"].append("archivo vacio")
        return res
    cuentas = [ln.count(",") + 1 for ln in lineas]
    moda = max(set(cuentas), key=cuentas.count)
    res["campos"] = moda
    for i, n in enumerate(cuentas, start=1):
        if n != moda:
            res["hallazgos"].append(f"linea {i}: {n} campos (las demas tienen {moda})")
    if len(res["hallazgos"]) > 8:
        extra = len(res["hallazgos"]) - 8
        res["hallazgos"] = res["hallazgos"][:8] + [f"... y {extra} linea(s) mas con problemas"]
    return res


def contar_soportes(raiz: Path, factura: str) -> int:
    """Archivos de soporte cuyo nombre trae el número de la factura."""
    norm = normalizar_factura(factura)
    if not norm:
        return 0
    patron = re.compile(rf"(?<!\d)0*{re.escape(norm)}(?!\d)")
    n = 0
    for p in raiz.rglob("*"):
        if p.is_file() and p.suffix.lower() in _EXT_SOPORTE and patron.search(p.name):
            n += 1
    return n


def cargar_lista(ruta: Path) -> set[str]:
    pedidas = set()
    for ln in ruta.read_text(encoding="utf-8-sig").splitlines():
        s = ln.strip()
        if s and not s.startswith("#"):
            pedidas.add(normalizar_factura(s))
    return pedidas


def escribir_excel(filas_xml: list[dict], filas_furips: list[dict], salida: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    azul = PatternFill("solid", fgColor="1F4E79")
    verde = PatternFill("solid", fgColor="E2EFDA")
    amarillo = PatternFill("solid", fgColor="FFF2CC")

    ws = wb.active
    ws.title = "Facturas"
    ws.append(
        [
            "Estado",
            "Factura",
            "Fecha",
            "Valor (XML)",
            "EPS",
            "NUMERO_CONTRATO",
            "FACTURA_SIN_CONTRATO",
            "Soportes hallados",
            "Hallazgos (corregir ANTES de radicar)",
            "Archivo XML",
        ]
    )
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = azul
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for d in filas_xml:
        listo = not d["hallazgos"]
        ws.append(
            [
                "LISTA" if listo else "REVISAR",
                d.get("factura", ""),
                d.get("fecha", ""),
                d.get("valor", ""),
                d.get("eps", ""),
                d.get("NUMERO_CONTRATO", ""),
                d.get("FACTURA_SIN_CONTRATO", ""),
                d.get("soportes", ""),
                " · ".join(d["hallazgos"]),
                d.get("archivo", ""),
            ]
        )
        for c in ws[ws.max_row]:
            c.fill = verde if listo else amarillo
    anchos = [10, 16, 12, 14, 30, 22, 22, 12, 52, 26]
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("FURIPS")
    ws2.append(["Estado", "Archivo", "Lineas", "Campos por linea", "Hallazgos"])
    for celda in ws2[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = azul
    for f in filas_furips:
        listo = not f["hallazgos"]
        ws2.append(
            [
                "LISTO" if listo else "REVISAR",
                f["archivo"],
                f["lineas"],
                f["campos"],
                " · ".join(f["hallazgos"]),
            ]
        )
        for c in ws2[ws2.max_row]:
            c.fill = verde if listo else amarillo
    for i, w in enumerate([10, 36, 10, 16, 60], start=1):
        ws2.column_dimensions[chr(64 + i)].width = w

    wb.save(str(salida))


def procesar(raiz: Path, lista: Path | None, salida: Path) -> int:
    xmls = sorted(raiz.rglob("*.xml"), key=lambda p: p.name.lower())
    furips = sorted(
        (p for p in raiz.rglob("*.txt") if p.name.upper().startswith("FURIPS")),
        key=lambda p: p.name.lower(),
    )
    pedidas = cargar_lista(lista) if lista else None

    print("=" * 66)
    print("  VERIFICAR RADICACION — chequeo del lote ANTES de radicar")
    print("=" * 66)
    print(f"  Carpeta: {raiz}")
    print(
        f"  XML: {len(xmls)}   FURIPS: {len(furips)}"
        + (f"   Lista: {len(pedidas)}" if pedidas else "")
    )
    print("-" * 66)

    filas_xml = []
    vistas = set()
    for x in xmls:
        d = revisar_xml(x)
        norm = normalizar_factura(d.get("factura", ""))
        if pedidas is not None and norm not in pedidas:
            continue
        d["soportes"] = contar_soportes(raiz, d.get("factura", ""))
        if d["soportes"] == 0:
            d["hallazgos"].append("sin soportes con el numero de factura en el nombre")
        filas_xml.append(d)
        if norm:
            vistas.add(norm)
        marca = "✓" if not d["hallazgos"] else "!"
        print(
            f"  {marca} {d.get('factura', '?'):16s} {(' · '.join(d['hallazgos']) or 'lista para radicar')[:60]}"
        )

    filas_furips = [revisar_furips(f) for f in furips]
    for f in filas_furips:
        marca = "✓" if not f["hallazgos"] else "!"
        print(f"  {marca} {f['archivo'][:40]:40s} {f['lineas']} lineas x {f['campos']} campos")

    escribir_excel(filas_xml, filas_furips, salida)

    listas = sum(1 for d in filas_xml if not d["hallazgos"])
    print("-" * 66)
    print(
        f"  Facturas revisadas: {len(filas_xml)}   Listas: {listas}   Con hallazgos: {len(filas_xml) - listas}"
    )
    if pedidas is not None:
        faltan = sorted(pedidas - vistas)
        if faltan:
            print(f"  OJO: {len(faltan)} factura(s) de la lista SIN XML en la carpeta:")
            print("       " + ", ".join(faltan[:20]) + (" ..." if len(faltan) > 20 else ""))
    print(f"  Informe: {salida}")
    print("=" * 66)
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Verifica el lote antes de radicar (previene glosas)."
    )
    parser.add_argument("carpeta", nargs="?", default=".", help="Carpeta del lote a radicar.")
    parser.add_argument("--lista", type=Path, help="Archivo .txt con las facturas a revisar.")
    parser.add_argument("--salida", type=Path, help="Ruta del Excel de salida.")
    args = parser.parse_args(argv)

    raiz = Path(args.carpeta).expanduser()
    if not raiz.is_dir():
        sys.stderr.write(f"ERROR: no existe la carpeta: {raiz}\n")
        return 2
    if args.lista and not args.lista.is_file():
        sys.stderr.write(f"ERROR: no existe la lista: {args.lista}\n")
        return 2
    salida = args.salida or (raiz / "VERIFICACION_RADICACION.xlsx")
    return procesar(raiz.resolve(), args.lista, salida)


if __name__ == "__main__":
    raise SystemExit(main())
