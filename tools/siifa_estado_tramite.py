#!/usr/bin/env python3
"""siifa_estado_tramite.py — En qué etapa va cada glosa, y a quién le toca mover.

El panel «Avance de auditoría» del portal muestra las cinco etapas del trámite
para UNA factura a la vez. Con 2.600 seguimientos, saber cuáles están
esperando algo del hospital son cientos de clics. Este informe lo responde de
una para todas: qué se ganó, qué se perdió, qué está esperando a la EPS y
—lo que de verdad importa— QUÉ TIENE QUE HACER EL HOSPITAL Y CUÁNDO SE VENCE.

LAS CINCO ETAPAS Y SUS PLAZOS (Res. 2284/2023, Ley 1438/2011 art. 57):

    1. Formulación de la glosa            EPS   20 días hábiles
    2. Respuesta del hospital             HUS   15 días hábiles
    3. Decisión inicial (levanta/reitera) EPS   10 días hábiles
    4. Subsanación de lo no levantado     HUS    7 días hábiles   <-- OJO
    5. Decisión final y pago              EPS    5 días hábiles

La etapa 4 es la trampa: la EPS reitera y el reloj del hospital arranca sin
que nadie avise. Son SIETE días hábiles, menos de la mitad que en la primera
respuesta. Una glosa reiterada que no se subsana queda en firme.

USO:
    py siifa_estado_tramite.py --informe "D:\\...\\informe_13AGO.xlsx"

    REM Con Excel de salida, para trabajar sobre él
    py siifa_estado_tramite.py --informe "D:\\...\\informe_13AGO.xlsx" ^
        --salida "D:\\...\\ESTADO_TRAMITE.xlsx"

    REM Sólo lo que le toca al hospital (lo urgente)
    py siifa_estado_tramite.py --informe "D:\\...\\informe.xlsx" --solo-accion-hus

SOBRE LOS DÍAS HÁBILES: se cuentan de lunes a viernes, SIN descontar los
festivos de Colombia. Un festivo alarga el plazo, así que este informe avisa
un poco antes de lo estricto —nunca después—, que es como debe ser una alerta.

INSTALACIÓN (una vez):
    py -m pip install openpyxl
"""

from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from _dinero import a_entero  # noqa: E402
from siifa_novedades import valor_total  # noqa: E402

# Plazos en días hábiles. Cambiarlos acá si cambia la norma.
PLAZO_RESPUESTA_HUS = 15
PLAZO_DECISION_EPS = 10
PLAZO_SUBSANACION_HUS = 7
PLAZO_DECISION_FINAL_EPS = 5

# Las etapas, en el orden del panel del portal. El orden importa: manda cómo
# se ordena el informe y en qué grupo cae cada glosa.
SIN_RESPONDER = "1. Sin responder"
ESPERANDO_EPS = "2. Respondida, esperando decisión de la EPS"
LEVANTADA = "3. Levantada — ganada"
POR_SUBSANAR = "4. Reiterada — falta subsanar"
SUBSANADA = "5. Subsanada, esperando decisión final"
CERRADA = "6. Con decisión final"

# A quién le toca mover en cada etapa.
LE_TOCA = {
    SIN_RESPONDER: "HOSPITAL",
    ESPERANDO_EPS: "EPS",
    LEVANTADA: "—",
    POR_SUBSANAR: "HOSPITAL",
    SUBSANADA: "EPS",
    CERRADA: "—",
}

PLAZO_DE = {
    SIN_RESPONDER: PLAZO_RESPUESTA_HUS,
    ESPERANDO_EPS: PLAZO_DECISION_EPS,
    POR_SUBSANAR: PLAZO_SUBSANACION_HUS,
    SUBSANADA: PLAZO_DECISION_FINAL_EPS,
}

COLUMNAS_SALIDA = [
    "etapa",
    "le_toca_a",
    "dias_habiles_corridos",
    "vence",
    "situacion",
    "numero_factura",
    "tipo_seguimiento",
    "razon_social_eps",
    "valor_glosa",
    "codigo_glosa",
    "codigo_respuesta",
    "descripcion_reiteracion",
    "fecha_formulacion",
    "fecha_respuesta",
    "id_seguimiento_factura_glosa",
]


def habiles(desde: date, hasta: date) -> int:
    """Días hábiles (lunes a viernes) entre dos fechas, sin contar festivos."""
    if hasta <= desde:
        return 0
    n, d = 0, desde
    while d < hasta:
        d += timedelta(days=1)
        if d.weekday() < 5:
            n += 1
    return n


def sumar_habiles(desde: date, dias: int) -> date:
    d, n = desde, 0
    while n < dias:
        d += timedelta(days=1)
        if d.weekday() < 5:
            n += 1
    return d


def _fecha(valor: object) -> date | None:
    if valor in (None, ""):
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip()[:10]
    try:
        return date.fromisoformat(texto)
    except ValueError:
        return None


def clasificar(fila: dict) -> str:
    """En qué etapa está esta glosa.

    Se mira la DESCRIPCIÓN de la decisión, no el código: SIIFA usa GTL002 para
    una glosa levantada, GRE003 para una reiterada y DRE003 para una
    devolución reiterada, y esa lista puede crecer. La palabra «levantada» o
    «reiterada» en el texto es lo estable.
    """
    respondida = str(fila.get("tiene_respuesta", "")).strip().upper() == "SI"
    if not respondida:
        return SIN_RESPONDER

    decision = str(fila.get("descripcion_reiteracion") or "").strip().lower()
    codigo_decision = str(fila.get("codigo_reiteracion") or "").strip()
    if not decision and not codigo_decision:
        return ESPERANDO_EPS

    # Una glosa levantada PARCIALMENTE deja un resto vivo que el hospital
    # todavía puede subsanar: no se puede dar por ganada.
    if "levantada" in decision and "parcial" not in decision:
        return LEVANTADA

    subsanada = bool(str(fila.get("codigo_reiteracion_respuesta") or "").strip())
    if not subsanada:
        return POR_SUBSANAR
    return SUBSANADA


def fecha_base(fila: dict, etapa: str) -> date | None:
    """Desde cuándo corre el plazo de la etapa en que está la glosa."""
    if etapa == SIN_RESPONDER:
        return _fecha(fila.get("fecha_formulacion"))
    if etapa == SUBSANADA:
        # Los 5 días de la decisión final corren desde la SUBSANACIÓN, y esa
        # fecha no viene en el informe. Usar la de la primera respuesta daba
        # «VENCIDA» sobre una subsanación cargada esa misma mañana —y una mora
        # inventada de la EPS es un reclamo que no se puede sostener—.
        return None
    # De la decisión inicial de la EPS tampoco viene fecha propia, pero ahí sí
    # se puede contar desde la respuesta del hospital: es lo más tarde que
    # pudo haber ocurrido, y por tanto la cuenta más conservadora.
    return _fecha(fila.get("fecha_respuesta"))


def evaluar(filas: list[dict], hoy: date) -> list[dict]:
    """Le pone a cada glosa su etapa, su plazo y su situación."""
    salida = []
    for f in filas:
        etapa = clasificar(f)
        plazo = PLAZO_DE.get(etapa)
        base = fecha_base(f, etapa)
        corridos = vence = None
        if plazo is not None and base is not None:
            corridos = habiles(base, hoy)
            vence = sumar_habiles(base, plazo)
        if plazo is None:
            situacion = "Cerrada"
        elif etapa == SUBSANADA:
            situacion = "Esperando a la EPS (5 días hábiles desde la subsanación)"
        elif corridos is None:
            situacion = "Sin fecha para calcular"
        elif corridos > plazo:
            situacion = f"VENCIDA hace {corridos - plazo} día(s) hábil(es)"
        else:
            situacion = f"Quedan {plazo - corridos} día(s) hábil(es)"
        salida.append(
            {
                **f,
                "etapa": etapa,
                "le_toca_a": LE_TOCA[etapa],
                "dias_habiles_corridos": corridos if corridos is not None else "",
                "vence": vence.isoformat() if vence else "",
                "situacion": situacion,
            }
        )
    return salida


def resumen(evaluadas: list[dict]) -> list[tuple]:
    """Cuántas y cuánto vale cada etapa, en el orden del trámite.

    El valor se calcula con valor_total(), que sabe que una devolución vale
    una vez por factura aunque tenga mil líneas. Sumar `valor_glosa` a secas
    daba $25.221 millones en la etapa 2 —donde hay $310 millones—.
    """
    agg: dict[str, list] = defaultdict(list)
    for f in evaluadas:
        agg[f["etapa"]].append(f)
    orden = [SIN_RESPONDER, ESPERANDO_EPS, LEVANTADA, POR_SUBSANAR, SUBSANADA, CERRADA]
    return [
        (
            e,
            LE_TOCA[e],
            len(agg[e]),
            len({str(f.get("numero_factura") or "") for f in agg[e]}),
            valor_total(agg[e]),
            sum(1 for f in agg[e] if str(f["situacion"]).startswith("VENCIDA")),
        )
        for e in orden
        if e in agg
    ]


def _pesos(valor: object) -> str:
    return f"${a_entero(valor):,}".replace(",", ".")


def mostrar(evaluadas: list[dict], hoy: date) -> None:
    print()
    print("=" * 88)
    print(f"ESTADO DEL TRÁMITE DE GLOSAS EN SIIFA — al {hoy.strftime('%d/%m/%Y')}")
    print("=" * 88)
    print()
    print(f"{'Etapa':<45}{'Le toca a':<11}{'Ítems':>7}{'Valor':>17}{'Vencidas':>9}")
    print("-" * 88)
    for etapa, toca, n, _f, v, venc in resumen(evaluadas):
        marca = str(venc) if venc else "-"
        print(f"{etapa:<45}{toca:<11}{n:>7}{_pesos(v):>17}{marca:>9}")
    print("-" * 88)

    # Lo que le toca al hospital va aparte y de primero: es lo accionable.
    mias = [f for f in evaluadas if f["le_toca_a"] == "HOSPITAL"]
    if mias:
        print()
        print("LO QUE LE TOCA AL HOSPITAL")
        print("-" * 88)
        vencidas = [f for f in mias if str(f["situacion"]).startswith("VENCIDA")]
        if vencidas:
            print(f"  {len(vencidas)} con el plazo VENCIDO — {_pesos(valor_total(vencidas))}")
        proximas = [f for f in mias if not str(f["situacion"]).startswith("VENCIDA")]
        if proximas:
            print(f"  {len(proximas)} dentro del plazo — {_pesos(valor_total(proximas))}")
        print()
        print(f"  {'Factura':<13}{'Etapa':<32}{'Valor':>14}  Situación")
        for f in sorted(mias, key=lambda x: -a_entero(x.get("valor_glosa")))[:25]:
            print(
                f"  {str(f.get('numero_factura') or ''):<13}"
                f"{f['etapa'][:31]:<32}"
                f"{_pesos(f.get('valor_glosa')):>14}  {f['situacion']}"
            )
        if len(mias) > 25:
            print(f"  ... y {len(mias) - 25} más (están en el Excel).")

    ganadas = [f for f in evaluadas if f["etapa"] == LEVANTADA]
    if ganadas:
        print()
        print("GLOSAS LEVANTADAS (la EPS le dio la razón al hospital)")
        print("-" * 88)
        for f in sorted(ganadas, key=lambda x: -a_entero(x.get("valor_glosa"))):
            print(
                f"  {str(f.get('numero_factura') or ''):<13}"
                f"{str(f.get('codigo_glosa') or ''):<9}"
                f"{_pesos(f.get('valor_glosa')):>14}   {str(f.get('razon_social_eps'))[:34]}"
            )
        print(f"  {'TOTAL RECUPERADO':<22}{_pesos(valor_total(ganadas)):>14}")

    # La EPS también tiene plazos. Si se pasa, es incumplimiento reclamable.
    morosas = [
        f
        for f in evaluadas
        if f["le_toca_a"] == "EPS" and str(f["situacion"]).startswith("VENCIDA")
    ]
    if morosas:
        por_eps: dict[str, list] = defaultdict(list)
        for f in morosas:
            por_eps[str(f.get("razon_social_eps") or "?")[:40]].append(f)
        print()
        print("LA EPS ESTÁ EN MORA (respondimos y no ha decidido dentro de su plazo)")
        print("-" * 88)
        for k, grupo in sorted(por_eps.items(), key=lambda kv: -valor_total(kv[1])):
            print(f"  {k:<44}{len(grupo):>7}{_pesos(valor_total(grupo)):>17}")
    print()


def escribir_xlsx(evaluadas: list[dict], ruta: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise SystemExit("ERROR: falta openpyxl. py -m pip install openpyxl") from None

    azul = PatternFill("solid", fgColor="1F4E78")
    blanco = Font(bold=True, color="FFFFFF")
    rojo = PatternFill("solid", fgColor="FFC7CE")
    verde = PatternFill("solid", fgColor="C6EFCE")
    ambar = PatternFill("solid", fgColor="FFF2CC")

    wb = Workbook()

    def hoja(nombre: str, filas: list[dict]):
        ws = wb.create_sheet(nombre)
        for c, h in enumerate(COLUMNAS_SALIDA, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = azul
            cell.font = blanco
        for r, f in enumerate(filas, start=2):
            for c, h in enumerate(COLUMNAS_SALIDA, 1):
                valor = a_entero(f.get(h)) if h == "valor_glosa" else f.get(h, "")
                cell = ws.cell(row=r, column=c, value=valor)
                cell.alignment = Alignment(vertical="top")
                if h == "valor_glosa":
                    cell.number_format = '"$"#,##0'
            if str(f["situacion"]).startswith("VENCIDA"):
                ws.cell(row=r, column=COLUMNAS_SALIDA.index("situacion") + 1).fill = rojo
            elif f["etapa"] == LEVANTADA:
                ws.cell(row=r, column=COLUMNAS_SALIDA.index("etapa") + 1).fill = verde
            elif f["le_toca_a"] == "HOSPITAL":
                ws.cell(row=r, column=COLUMNAS_SALIDA.index("etapa") + 1).fill = ambar
        anchos = {"etapa": 42, "situacion": 30, "razon_social_eps": 32, "numero_factura": 14}
        for c, h in enumerate(COLUMNAS_SALIDA, 1):
            ws.column_dimensions[get_column_letter(c)].width = anchos.get(h, 16)
        ws.freeze_panes = "A2"
        return ws

    ws0 = wb.active
    ws0.title = "RESUMEN"
    encabezado = ["Etapa", "Le toca a", "Ítems", "Facturas", "Valor", "Con plazo vencido"]
    ws0.append(encabezado)
    for c in range(1, len(encabezado) + 1):
        ws0.cell(row=1, column=c).fill = azul
        ws0.cell(row=1, column=c).font = blanco
    for etapa, toca, n, nf, v, venc in resumen(evaluadas):
        ws0.append([etapa, toca, n, nf, v, venc])
        ws0.cell(row=ws0.max_row, column=5).number_format = '"$"#,##0'
        if venc:
            ws0.cell(row=ws0.max_row, column=6).fill = rojo
    for col, ancho in (("A", 44), ("B", 12), ("C", 9), ("D", 10), ("E", 18), ("F", 18)):
        ws0.column_dimensions[col].width = ancho

    mias = [f for f in evaluadas if f["le_toca_a"] == "HOSPITAL"]
    if mias:
        hoja("ACCION_HOSPITAL", sorted(mias, key=lambda x: -a_entero(x.get("valor_glosa"))))
    ganadas = [f for f in evaluadas if f["etapa"] == LEVANTADA]
    if ganadas:
        hoja("LEVANTADAS", sorted(ganadas, key=lambda x: -a_entero(x.get("valor_glosa"))))
    morosas = [
        f
        for f in evaluadas
        if f["le_toca_a"] == "EPS" and str(f["situacion"]).startswith("VENCIDA")
    ]
    if morosas:
        hoja("EPS_EN_MORA", sorted(morosas, key=lambda x: -a_entero(x.get("valor_glosa"))))
    hoja("TODO", evaluadas)

    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ruta))
    print(f"Informe guardado: {ruta}\n")


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--informe", required=True, help="Informe de siifa_reporte_seguimientos.py.")
    ap.add_argument("--salida", help="Excel de salida (opcional).")
    ap.add_argument(
        "--solo-accion-hus",
        action="store_true",
        help="Deja sólo lo que está esperando al hospital.",
    )
    ap.add_argument("--hoy", help="Fecha de corte (2026-08-13). Por defecto, hoy.")
    args = ap.parse_args()

    # Se reutiliza el lector del comparador: un solo sitio que sabe leer el
    # informe de seguimientos, para que no se dupliquen las columnas.
    from siifa_novedades import leer_informe

    filas = leer_informe(Path(args.informe))
    hoy = date.fromisoformat(args.hoy) if args.hoy else date.today()
    evaluadas = evaluar(filas, hoy)
    if args.solo_accion_hus:
        evaluadas = [f for f in evaluadas if f["le_toca_a"] == "HOSPITAL"]

    mostrar(evaluadas, hoy)
    if args.salida:
        escribir_xlsx(evaluadas, Path(args.salida))


if __name__ == "__main__":
    main()
