#!/usr/bin/env python3
"""siifa_redactar_respuestas.py — Redacta la respuesta de las glosas y
devoluciones de SIIFA que el hospital todavía no ha contestado.

QUÉ HACE Y QUÉ NO HACE
Escribe el sustento JURÍDICO y CONTRACTUAL de la no aceptación, que es el
mismo para todas las glosas del mismo código, y le pega el detalle que la
propia EPS registró en su objeción (el ítem, el paciente, la fecha, el CUPS,
el valor). NO inventa hechos clínicos: no dice qué encontró la historia
clínica ni qué hizo el cirujano, porque eso sale del soporte y no de esta
base. Cada respuesta sale con una nota en la columna REVISAR diciendo qué
tiene que verificar el auditor antes de subirla.

La postura es la institucional: NO ACEPTA, se defiende el 100% del valor
(código de respuesta RE9901), que es lo que el HUS viene respondiendo en
Dinámica Gerencial para estos mismos códigos.

USO:
    REM Todo lo que falte por responder, separado en dos archivos
    py siifa_redactar_respuestas.py ^
        --informe  "D:\\...\\SIIFA\\informe_seguimientos.xlsx" ^
        --tramites "D:\\...\\TRAMITES_ENE_A_JUL_2026.xlsx" ^
        --salida-glosas      "D:\\...\\SIIFA\\respuestas_GLOSAS.xlsx" ^
        --salida-devoluciones "D:\\...\\SIIFA\\respuestas_DEVOLUCIONES.xlsx"

`--tramites` es opcional: si se pasa, las glosas que el hospital YA respondió
en DGH salen con ESA respuesta (la real, no una redactada), y sólo se redacta
lo que no tiene antecedente.

`--solo-lo-ya-respondido` deja en los archivos de cargue ÚNICAMENTE las
respuestas reales del hospital (las verdes). Las redactadas por el motor no
se botan: salen en archivos aparte con el sufijo `_REDACTADAS`, para
revisarlas y subirlas después. OJO con el término del artículo 57 de la Ley
1438 de 2011: la glosa que no se contesta a tiempo se entiende ACEPTADA.

Los dos archivos salen listos para tools/responder_glosas_siifa.py, con una
fila por glosa y su id de SIIFA. Antes del cargue masivo: piloto de 1 glosa.

La columna FECHA_RESPUESTA lleva la fecha en que el hospital respondió en DGH
—que es la que se digita en SIIFA— y va vacía en lo redactado, que se está
respondiendo hoy. Subir una respuesta vieja con la fecha de hoy la deja
registrada fuera del término del artículo 57 de la Ley 1438 de 2011.

INSTALACIÓN (una vez):
    py -m pip install openpyxl
"""

from __future__ import annotations

import argparse
import logging
import sys
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from _dinero import a_entero  # noqa: E402
import siifa_perfiles as perfiles  # noqa: E402
from siifa_preparar_respuestas import agrupar, leer_informe  # noqa: E402

logger = logging.getLogger("siifa_redactar")

CODIGO_NO_ACEPTA = "RE9901"

_CIERRE_COMUN = (
    "SE SOLICITA EL LEVANTAMIENTO DE LA OBJECIÓN Y, EN CASO DE PERSISTIR, LA "
    "PROGRAMACIÓN DE MESA DE CONCILIACIÓN DE AUDITORÍA MÉDICA Y/O TÉCNICA ENTRE "
    "LAS PARTES, CONFORME AL ARTÍCULO 57 DE LA LEY 1438 DE 2011 Y A LA RESOLUCIÓN "
    "2284 DE 2023"
)


def cierre_de(ips=None) -> str:
    """El cierre de la respuesta, con los datos de contacto de ESA IPS.

    Sin IPS se usa el HUS, que es como funcionó siempre: hay código que llama
    a redactar() sin decir la entidad y no puede cambiar de comportamiento.

    Si la IPS no tiene contacto cargado, la frase se omite. Poner el correo y
    la dirección del HUS en la respuesta de otra clínica sería mandarle a la
    EPS un dato falso en un escrito con efectos jurídicos.
    """
    if ips is None:
        ips = perfiles.IPS["HUS"]
    contacto = getattr(ips, "contacto", "")
    if contacto:
        return f"{_CIERRE_COMUN}. CUALQUIER INFORMACIÓN AL {contacto}."
    return f"{_CIERRE_COMUN}."


# Compatibilidad con lo ya escrito: el cierre del HUS, que es el de siempre.
CIERRE = cierre_de(None)

# Banco de argumentos. La llave es el código de glosa de la Res. 2284/2023
# (Anexo Técnico 2); si el código exacto no está, se usa el de su familia
# —las dos primeras letras—, porque la causal es la misma.
#
# Cada entrada trae:
#   argumento : el sustento de la no aceptación.
#   revisar   : qué debe confirmar el auditor ANTES de subirla. No es
#               decorativo: es lo que separa una respuesta sostenible de
#               una que la EPS tumba en la conciliación.
ARGUMENTOS: dict[str, dict[str, str]] = {
    # --- TARIFAS: la EPS dice que se cobró por encima de lo pactado --------
    "TA": {
        "argumento": (
            "EL VALOR FACTURADO CORRESPONDE A LA TARIFA INSTITUCIONAL VIGENTE DE LA "
            "{IPS}, ADOPTADA MEDIANTE ACTO "
            "ADMINISTRATIVO Y APLICABLE AL ACUERDO DE VOLUNTADES SUSCRITO CON LA "
            "ENTIDAD RESPONSABLE DE PAGO. LA OBJECIÓN POR MAYOR VALOR NO INDICA LA "
            "CLÁUSULA CONTRACTUAL NI EL MANUAL TARIFARIO CONCRETO DEL CUAL SE "
            "DERIVARÍA LA DIFERENCIA, POR LO QUE NO CUMPLE EL DEBER DE MOTIVACIÓN "
            "EXIGIDO POR EL ARTÍCULO 6 DE LA RESOLUCIÓN 2284 DE 2023, QUE OBLIGA A "
            "QUE LA GLOSA SEA EXPRESA, CONCRETA Y SOPORTADA. EN CONSECUENCIA, EL "
            "ÍTEM SE ENCUENTRA CORRECTAMENTE LIQUIDADO Y SE DEFIENDE LA TOTALIDAD "
            "DEL VALOR"
        ),
        "revisar": (
            "Confirmar la tarifa del ítem contra el contrato con la EPS y la "
            "resolución de tarifas vigente a la fecha de atención."
        ),
    },
    # --- AUTORIZACIÓN: falta volante, o cantidad mayor a la autorizada -----
    "AU": {
        "argumento": (
            "EL SERVICIO OBJETADO FUE EFECTIVAMENTE PRESTADO Y RESULTABA MÉDICAMENTE "
            "NECESARIO PARA LA ATENCIÓN DEL USUARIO. LA AUSENCIA O DIFERENCIA DE UN "
            "TRÁMITE ADMINISTRATIVO DE AUTORIZACIÓN NO DESVIRTÚA LA PRESTACIÓN NI "
            "HABILITA EL NO PAGO: EL ARTÍCULO 17 DE LA LEY 1751 DE 2015 PROTEGE LA "
            "AUTONOMÍA DEL PROFESIONAL TRATANTE, Y EL DECRETO 4747 DE 2007 PROHÍBE "
            "SUPEDITAR LA ATENCIÓN A REQUISITOS ADMINISTRATIVOS CUANDO ESTÁ DE POR "
            "MEDIO LA SALUD DEL PACIENTE. LA {IPS} PONE A DISPOSICIÓN DE LA "
            "ENTIDAD RESPONSABLE DE PAGO LOS SOPORTES DEL TRÁMITE DE NOTIFICACIÓN "
            "ADELANTADO PARA SU VERIFICACIÓN. POR LO ANTERIOR NO SE ACEPTA LA "
            "OBJECIÓN"
        ),
        "revisar": (
            "Adjuntar el soporte de la autorización o del trámite de notificación "
            "(anexos técnicos y correos) antes de subir esta respuesta."
        ),
    },
    # --- SOPORTES: la EPS dice que falta el soporte del servicio ----------
    "SO": {
        "argumento": (
            "EL SERVICIO OBJETADO SE ENCUENTRA DEBIDAMENTE REGISTRADO EN LA HISTORIA "
            "CLÍNICA DEL USUARIO, QUE ES EL SOPORTE IDÓNEO CONFORME AL ANEXO TÉCNICO "
            "1 DE LA RESOLUCIÓN 2284 DE 2023 Y A LA RESOLUCIÓN 1995 DE 1999. LA ESE "
            "HUS PONE A DISPOSICIÓN DE LA ENTIDAD RESPONSABLE DE PAGO EL SOPORTE "
            "CORRESPONDIENTE PARA SU VERIFICACIÓN. LA GLOSA POR AUSENCIA DE SOPORTE "
            "NO PUEDE MANTENERSE CUANDO EL PRESTADOR LO APORTA DENTRO DEL TRÁMITE DE "
            "RESPUESTA, POR LO QUE SE SOLICITA SU LEVANTAMIENTO"
        ),
        "revisar": (
            "IMPRESCINDIBLE: ubicar y anexar el soporte. Si el soporte no existe, "
            "esta respuesta NO se puede sostener y la glosa debería aceptarse."
        ),
    },
    # --- FACTURACIÓN: cobro incluido en otro ítem, o cantidades -----------
    "FA": {
        "argumento": (
            "EL ÍTEM OBJETADO CORRESPONDE A UN SERVICIO O INSUMO CON CÓDIGO CUPS "
            "PROPIO Y DESAGREGACIÓN TARIFARIA INDEPENDIENTE, POR LO QUE NO SE "
            "ENCUENTRA SUBSUMIDO EN LA ATENCIÓN AGRUPADA, EN EL DERECHO DE SALA NI "
            "EN LA ESTANCIA A LA QUE ALUDE LA ENTIDAD. LA ENTIDAD NO INDICA LA NORMA "
            "O CLÁUSULA QUE DISPONGA ESA INCLUSIÓN, COMO LO EXIGE EL ARTÍCULO 6 DE "
            "LA RESOLUCIÓN 2284 DE 2023. EL CARGO CORRESPONDE A UN CONSUMO REAL, "
            "REGISTRADO Y NECESARIO PARA LA ATENCIÓN, POR LO QUE SE DEFIENDE LA "
            "TOTALIDAD DEL VALOR FACTURADO"
        ),
        "revisar": (
            "Verificar en el manual tarifario aplicable que el ítem no esté incluido "
            "en el procedimiento o estancia que menciona la EPS."
        ),
    },
    # --- COBERTURA: la EPS dice que no está cubierto ----------------------
    "CO": {
        "argumento": (
            "EL SERVICIO OBJETADO CORRESPONDE A UNA TECNOLOGÍA EN SALUD FINANCIADA "
            "CON CARGO A LA UPC Y NO SE ENCUENTRA DENTRO DE LAS EXCLUSIONES "
            "EXPRESAMENTE DEFINIDAS POR LA NORMATIVIDAD VIGENTE. LA ENTIDAD NO "
            "ACREDITA EL ACTO ADMINISTRATIVO NI LA CLÁUSULA QUE EXCLUYA LA "
            "PRESTACIÓN. ADEMÁS, EL SERVICIO FUE ORDENADO POR EL PROFESIONAL "
            "TRATANTE EN EJERCICIO DE LA AUTONOMÍA MÉDICA RECONOCIDA EN EL ARTÍCULO "
            "17 DE LA LEY 1751 DE 2015, Y SE PRESTÓ EFECTIVAMENTE AL USUARIO. NO "
            "SIENDO PROCEDENTE LA CAUSAL INVOCADA, SE DEFIENDE LA TOTALIDAD DEL VALOR"
        ),
        "revisar": (
            "Confirmar que la prestación esté contratada con la EPS; si la glosa dice "
            "PRESTACIÓN NO CONTRATADA, revisar el acuerdo de voluntades vigente."
        ),
    },
    # --- PERTINENCIA: la EPS discute el criterio clínico ------------------
    "CL": {
        "argumento": (
            "LA PERTINENCIA DEL SERVICIO PRESTADO ES COMPETENCIA EXCLUSIVA DEL "
            "PROFESIONAL TRATANTE, QUIEN LO ORDENÓ CON FUNDAMENTO EN EL CUADRO "
            "CLÍNICO DEL PACIENTE Y EN SU CRITERIO PROFESIONAL. EL ARTÍCULO 17 DE LA "
            "LEY 1751 DE 2015 GARANTIZA LA AUTONOMÍA MÉDICA Y PROSCRIBE QUE LA "
            "AUDITORÍA DE LA ENTIDAD RESPONSABLE DE PAGO SUSTITUYA EL JUICIO CLÍNICO "
            "DEL MÉDICO QUE ATENDIÓ AL USUARIO. LA VALORACIÓN RETROSPECTIVA DE LA "
            "ENTIDAD NO CONSTITUYE CAUSAL VÁLIDA DE GLOSA, POR LO QUE SE SOLICITA SU "
            "LEVANTAMIENTO"
        ),
        "revisar": (
            "Si es viable, reforzar con la justificación del médico tratante que "
            "consta en la historia clínica."
        ),
    },
    # --- DEVOLUCIÓN por radicación fuera de los 22 días hábiles -----------
    "DE5601": {
        "argumento": (
            "LA {IPS} RADICÓ LA FACTURA "
            "ELECTRÓNICA DE VENTA CON VALIDACIÓN PREVIA Y SUS SOPORTES POR EL CANAL "
            "ACORDADO CON LA ENTIDAD RESPONSABLE DE PAGO, Y PONE A DISPOSICIÓN DE LA "
            "ENTIDAD LA CONSTANCIA DE RADICACIÓN Y EL ACUSE CORRESPONDIENTE PARA "
            "VERIFICAR EL CUMPLIMIENTO DEL TÉRMINO PREVISTO EN EL ARTÍCULO 17 DE LA "
            "RESOLUCIÓN 2284 DE 2023. LA DEVOLUCIÓN NO PROCEDE CUANDO "
            "EL PRESTADOR ACREDITA LA RADICACIÓN EN TÉRMINO, Y UNA FALLA EN EL CANAL "
            "DE RECEPCIÓN DE LA ENTIDAD NO PUEDE TRASLADARSE AL PRESTADOR NI "
            "AFECTAR EL RECONOCIMIENTO DE LOS SERVICIOS EFECTIVAMENTE PRESTADOS. SE "
            "SOLICITA DEJAR SIN EFECTO LA DEVOLUCIÓN Y REPROCESAR LA CUENTA"
        ),
        "revisar": (
            "CRÍTICO: verificar el acuse de radicación y la fecha real de envío del "
            "XML. Si la radicación sí fue extemporánea, esta respuesta no se sostiene "
            "y corresponde volver a radicar la factura."
        ),
    },
    # --- DEVOLUCIÓN por afiliación / responsable de pago ------------------
    "DE1601": {
        "argumento": (
            "LA ATENCIÓN SE PRESTÓ AL USUARIO EN CONDICIONES QUE NO ADMITÍAN "
            "DIFERIMIENTO, Y LA {IPS} PONE A DISPOSICIÓN DE LA ENTIDAD LA CONSULTA "
            "DE DERECHOS EN LA BASE DE DATOS ÚNICA DE AFILIADOS EFECTUADA PARA ESTA "
            "ATENCIÓN. SE SOLICITA A LA ENTIDAD "
            "VALIDAR NUEVAMENTE LA AFILIACIÓN A LA FECHA DE PRESTACIÓN DEL SERVICIO "
            "Y NO A LA FECHA DE RADICACIÓN. EN TODO CASO, TRATÁNDOSE DE UNA ATENCIÓN "
            "YA PRESTADA, EL TRASLADO O LA NOVEDAD DE AFILIACIÓN NO EXTINGUE LA "
            "OBLIGACIÓN DE PAGO A CARGO DE QUIEN OSTENTABA LA CALIDAD DE RESPONSABLE "
            "AL MOMENTO DE LA ATENCIÓN"
        ),
        "revisar": (
            "Adjuntar el pantallazo de la consulta BDUA a la fecha de atención. Si el "
            "usuario efectivamente pertenecía a otro pagador, redirigir el cobro."
        ),
    },
    # --- Cualquier otra devolución ---------------------------------------
    "DE": {
        "argumento": (
            "LA {IPS} NO ACEPTA LA CAUSAL DE "
            "DEVOLUCIÓN INVOCADA. LOS SERVICIOS FACTURADOS FUERON EFECTIVAMENTE "
            "PRESTADOS AL USUARIO Y LA CUENTA SE RADICÓ CON EL LLENO DE LOS "
            "REQUISITOS EXIGIDOS EN EL ANEXO TÉCNICO 1 DE LA RESOLUCIÓN 2284 DE "
            "2023. LA DEVOLUCIÓN DEBE FUNDARSE EN UNA CAUSAL OBJETIVA Y DEBIDAMENTE "
            "MOTIVADA, LO QUE NO SE ACREDITA EN ESTE CASO. SE SOLICITA DEJARLA SIN "
            "EFECTO Y REPROCESAR LA CUENTA"
        ),
        "revisar": "Verificar la causal concreta que invoca la EPS antes de subirla.",
    },
}

ARGUMENTO_GENERICO = {
    "argumento": (
        "LOS SERVICIOS OBJETADOS FUERON EFECTIVAMENTE PRESTADOS AL USUARIO Y SE "
        "ENCUENTRAN DEBIDAMENTE REGISTRADOS EN LA HISTORIA CLÍNICA Y EN LOS "
        "SOPORTES DE COBRO. LA OBJECIÓN NO CUMPLE EL DEBER DE MOTIVACIÓN EXPRESA, "
        "CONCRETA Y SOPORTADA QUE EXIGE EL ARTÍCULO 6 DE LA RESOLUCIÓN 2284 DE "
        "2023, POR LO QUE NO SE ACEPTA Y SE DEFIENDE LA TOTALIDAD DEL VALOR "
        "FACTURADO"
    ),
    "revisar": "Código sin argumento propio: revisar el texto antes de subirlo.",
}


def argumento_para(codigo: str) -> dict[str, str]:
    """Busca el argumento del código exacto; si no, el de su familia."""
    codigo = (codigo or "").strip().upper()
    if codigo in ARGUMENTOS:
        return ARGUMENTOS[codigo]
    return ARGUMENTOS.get(codigo[:2], ARGUMENTO_GENERICO)


def _pesos(valor) -> str:
    """Escribe el valor en pesos para el texto de la respuesta.

    La lectura la hace a_entero, que es el único lector de pesos del repo:
    acá sólo se le pone el formato. Un valor mal leído acabaría escrito en
    una respuesta que se le manda a la EPS.
    """
    if valor in (None, ""):
        return "SIN VALOR REGISTRADO"
    return f"${a_entero(valor):,}".replace(",", ".")


def _limpiar(texto) -> str:
    """El texto de la EPS viene con saltos de línea y basura de exportación."""
    s = str(texto or "").replace("_x000D_", " ").replace("\r", " ").replace("\n", " ")
    return " ".join(s.split()).strip()


# Lo que aguanta el campo Observación de SIIFA. Está medido contra la
# plataforma, no supuesto: en el cargue del 05-08-2026 entraron todas las
# respuestas de hasta 1.499 caracteres y SIIFA rechazó TODAS las de 1.501 en
# adelante —927 en total— con un error que además engaña, porque llega como
# «HTTP 500: error saving the entity changes», que parece falla del servidor.
LIMITE_OBSERVACION = 1500


def redactar(linea: dict, limite: int = LIMITE_OBSERVACION, ips=None) -> dict[str, str]:
    """Arma la respuesta de UNA glosa o devolución.

    El texto sale con cuatro partes: el encabezado (qué se objeta y por
    cuánto), lo que la propia entidad escribió, el sustento normativo del
    código, y el cierre institucional.
    """
    es_devolucion = str(linea.get("tipo_seguimiento") or "").upper() == "DEVOLUCION"
    codigo = str(linea.get("codigo_glosa") or "").strip().upper()
    banco = argumento_para(codigo)
    quien = getattr(ips, "nombre_legal", None) or "ESE HOSPITAL UNIVERSITARIO DE SANTANDER"
    cierre = cierre_de(ips)

    encabezado = (
        f"{quien} NO ACEPTA LA "
        f"{'DEVOLUCIÓN APLICADA' if es_devolucion else 'GLOSA APLICADA'} A LA FACTURA "
        f"{_limpiar(linea.get('numero_factura')).upper()} POR VALOR DE "
        f"{_pesos(linea.get('valor_glosa'))} BAJO EL CÓDIGO {codigo}, "
        f"CUYA CAUSAL ES: {_limpiar(linea.get('descripcion_glosa')).upper().rstrip('.')}"
    )

    dijo = _limpiar(linea.get("observacion_glosa"))
    if dijo and dijo.lower() != "null":
        detalle = f". LA ENTIDAD SUSTENTA SU OBJECIÓN EN LO SIGUIENTE: {dijo.upper().rstrip('.')}"
    else:
        detalle = (
            ". LA ENTIDAD NO REGISTRÓ OBSERVACIÓN ALGUNA QUE SUSTENTE LA OBJECIÓN, "
            "LO QUE POR SÍ SOLO IMPIDE EJERCER EL DERECHO DE CONTRADICCIÓN"
        )

    # El banco de argumentos escribe {IPS} donde va el nombre del prestador:
    # el mismo sustento sirve para las cuatro, cambiando quién lo firma.
    argumento = banco["argumento"].replace("{IPS}", quien)
    texto = f"{encabezado}{detalle}. {argumento}. {cierre}"
    if len(texto) > limite:
        # Lo primero que se recorta es la cita de la EPS: el sustento propio
        # y el cierre no se pueden perder.
        sobra = len(texto) - limite
        recorte = max(0, len(detalle) - sobra - 5)
        texto = f"{encabezado}{detalle[:recorte].rstrip()}... {argumento}. {cierre}"
    return {
        "CODIGO_RESPUESTA": CODIGO_NO_ACEPTA,
        "OBSERVACION_RESPUESTA": texto,
        "REVISAR": banco["revisar"],
    }


def escribir(filas: list[dict], ruta: Path, titulo: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    COLS = [
        "ID_SEGUIMIENTO_FACTURA_GLOSA",
        "ID_SEGUIMIENTO_FACTURA_DEVOLUCION",
        "NUMERO_FACTURA",
        "TIPO",
        "CODIGO_GLOSA",
        "DESCRIPCION_GLOSA",
        "OBSERVACION_GLOSA",
        "VALOR_GLOSA",
        "DIAS_DESDE_EL_REPORTE",
        "ORIGEN_RESPUESTA",
        "REVISAR",
        "CODIGO_RESPUESTA",
        "FECHA_RESPUESTA",
        "OBSERVACION_RESPUESTA",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "RESPUESTAS"
    for i, h in enumerate(COLS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.font = Font(bold=True, color="FFFFFF")

    del_hus = PatternFill("solid", fgColor="C6EFCE")
    redactada = PatternFill("solid", fgColor="FFF2CC")
    for r, fila in enumerate(filas, start=2):
        for i, h in enumerate(COLS, 1):
            c = ws.cell(row=r, column=i, value=fila.get(h))
            c.alignment = Alignment(
                vertical="top",
                wrap_text=h
                in ("DESCRIPCION_GLOSA", "OBSERVACION_GLOSA", "OBSERVACION_RESPUESTA", "REVISAR"),
            )
            if h == "ORIGEN_RESPUESTA":
                c.fill = del_hus if fila.get(h) != "REDACTADA" else redactada
            if h == "VALOR_GLOSA":
                c.number_format = '"$"#,##0'
    anchos = [16, 18, 14, 12, 12, 45, 45, 14, 10, 16, 45, 14, 14, 120]
    for i, a in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = a
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{ws.max_row}"
    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta)
    logger.info("%s: %s (%d respuestas)", titulo, ruta, len(filas))


def armar(informe: Path, tramites: Path | None, ips=None) -> list[dict]:
    """Devuelve una fila por glosa pendiente, con su respuesta puesta."""
    datos = leer_informe(informe)
    grupos = agrupar(datos)

    previas: dict = {}
    if tramites:
        import siifa_cruzar_tramites_dgh as cru

        facturas = {cru.normalizar_factura(v[0].get("numero_factura")) for v in grupos.values()}
        fino, grueso = cru.indexar(cru.leer_tramites(tramites, facturas))
        previas = cru.cruzar(grupos, fino, grueso)

    filas = []
    for clave, lineas in grupos.items():
        previa = previas.get(clave)
        for linea in lineas:
            if previa:
                # La fecha es la del día en que el HUS respondió en DGH, no la
                # de hoy: si sube con la de hoy, en SIIFA la respuesta queda
                # registrada fuera del término.
                resp = {
                    "CODIGO_RESPUESTA": previa["CODIGO_RESPUESTA"],
                    "OBSERVACION_RESPUESTA": previa["OBSERVACION_RESPUESTA"],
                    "FECHA_RESPUESTA": previa.get("FECHA_RESPUESTA"),
                    "REVISAR": previa.get("REVISAR"),
                }
                origen = previa["ORIGEN"]
            else:
                # Lo redactado se está respondiendo hoy: sin fecha, el bot
                # pone la de hoy, que es la que corresponde.
                resp = {**redactar(linea, ips=ips), "FECHA_RESPUESTA": None}
                origen = "REDACTADA"
            filas.append(
                {
                    "ID_SEGUIMIENTO_FACTURA_GLOSA": linea.get("id_seguimiento_factura_glosa"),
                    # Una devolución se responde por otra puerta y con este
                    # id; sin él, el bot la deja sin mandar a propósito.
                    "ID_SEGUIMIENTO_FACTURA_DEVOLUCION": linea.get(
                        "id_seguimiento_factura_devolucion"
                    ),
                    "NUMERO_FACTURA": linea.get("numero_factura"),
                    "TIPO": linea.get("tipo_seguimiento"),
                    "CODIGO_GLOSA": linea.get("codigo_glosa"),
                    "DESCRIPCION_GLOSA": linea.get("descripcion_glosa"),
                    "OBSERVACION_GLOSA": _limpiar(linea.get("observacion_glosa")),
                    "VALOR_GLOSA": linea.get("valor_glosa"),
                    "DIAS_DESDE_EL_REPORTE": None,
                    "ORIGEN_RESPUESTA": origen,
                    **resp,
                }
            )
    return filas


def es_respuesta_real(fila: dict) -> bool:
    """La respuesta la dio el hospital en DGH (verde), no la redactó el motor."""
    return str(fila.get("ORIGEN_RESPUESTA") or "").upper() != "REDACTADA"


def con_sufijo(ruta: Path, sufijo: str) -> Path:
    """respuestas_GLOSAS.xlsx -> respuestas_GLOSAS_REDACTADAS.xlsx"""
    return ruta.with_name(f"{ruta.stem}_{sufijo}{ruta.suffix}")


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--informe", required=True, help="Informe de siifa_reporte_seguimientos.py.")
    ap.add_argument("--tramites", help="Export de DGH, para reusar lo ya respondido (opcional).")
    ap.add_argument("--salida-glosas", required=True, help="Excel de salida de las GLOSAS.")
    ap.add_argument(
        "--salida-devoluciones", required=True, help="Excel de salida de las DEVOLUCIONES."
    )
    ap.add_argument(
        "--solo-lo-ya-respondido",
        action="store_true",
        help=(
            "En los archivos de cargue deja SÓLO las respuestas reales del hospital "
            "(las verdes). Las redactadas por el motor salen aparte, con el sufijo "
            "_REDACTADAS, para revisarlas y subirlas después."
        ),
    )
    ap.add_argument("--verbose", action="store_true")
    perfiles.agregar_argumento(ap)
    args = ap.parse_args()
    ips = perfiles.buscar(args.ips)
    perfiles.anunciar(ips)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s  %(message)s",
        datefmt="%H:%M:%S",
        stream=sys.stdout,
    )

    filas = armar(Path(args.informe), Path(args.tramites) if args.tramites else None, ips=ips)
    glosas = [f for f in filas if str(f["TIPO"]).upper() == "GLOSA"]
    devoluciones = [f for f in filas if str(f["TIPO"]).upper() == "DEVOLUCION"]

    def orden(f):
        return (str(f["NUMERO_FACTURA"]), str(f["CODIGO_GLOSA"]))

    ruta_glosas = Path(args.salida_glosas)
    ruta_devoluciones = Path(args.salida_devoluciones)
    aparte = ""

    if args.solo_lo_ya_respondido:
        for nombre, todas, ruta in (
            ("Glosas", glosas, ruta_glosas),
            ("Devoluciones", devoluciones, ruta_devoluciones),
        ):
            reales = [f for f in todas if es_respuesta_real(f)]
            redactadas = [f for f in todas if not es_respuesta_real(f)]
            escribir(sorted(reales, key=orden), ruta, nombre)
            escribir(
                sorted(redactadas, key=orden),
                con_sufijo(ruta, "REDACTADAS"),
                f"{nombre} redactadas (aparte)",
            )
        aparte = (
            "\nSe cargan SÓLO las respuestas reales del hospital. Las redactadas\n"
            "quedaron en los archivos _REDACTADAS: no se botaron, quedan para\n"
            "revisar y subir después. RECUERDE el artículo 57 de la Ley 1438 de\n"
            "2011: la glosa que no se contesta dentro del término se entiende\n"
            "ACEPTADA, así que esas no pueden quedarse guardadas indefinidamente.\n"
        )
    else:
        escribir(sorted(glosas, key=orden), ruta_glosas, "Glosas")
        escribir(sorted(devoluciones, key=orden), ruta_devoluciones, "Devoluciones")

    origen = Counter(f["ORIGEN_RESPUESTA"] for f in filas)
    print(
        f"\nRESUMEN\n"
        f"  glosas       : {len(glosas):5} respuestas\n"
        f"  devoluciones : {len(devoluciones):5} respuestas\n"
        f"  de esas, respuesta REAL que el hospital ya había dado en DGH: "
        f"{origen['EXACTO'] + origen['POR_CODIGO']}\n"
        f"  redactadas por el motor (revisar antes de subir)            : "
        f"{origen['REDACTADA']}\n"
        f"{aparte}"
        f"\nCada fila trae en REVISAR qué hay que verificar antes de subirla.\n"
        f"Antes del cargue masivo, PILOTO DE 1 GLOSA (regla del repo):\n"
        f'   py tools\\responder_glosas_siifa.py --excel "{args.salida_glosas}" '
        f'--solo-id <id> --reporte "piloto_siifa.csv"\n'
    )


if __name__ == "__main__":
    main()
