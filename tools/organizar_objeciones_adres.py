"""organizar_objeciones_adres.py — pasa las glosas del ADRES al formato OBJECIONES.

Convierte el Excel de glosas del ADRES (el «ADRES DANIEL», el que arma el
auditor con la clasificación y el valor aceptado de cada renglón) al **mismo
formato OBJECIONES de 16 columnas** que se viene cargando en Dinámica
Gerencial (DGH) para COOSALUD:

    CDCONSEC | CDFECDOC | CRNCXC | CROFECOBJ | CROREFERE | CROOBSERV |
    CROCLAOBJ | CRNCLAOBJ | GENUSUARIO4 | CRNCONOBJ | SLNSERPRO | IDRIPS |
    CTNCENCOS | CROVALOBJ | CRDOBSERV | CROTIPOBJ

EL PROBLEMA QUE RESUELVE. El ADRES nombra los servicios con **códigos SOAT**
(29117, 39145, 38134…) y con el registro INVIMA de medicamentos y materiales
(``2016DM-0000315-R2``). DGH los tiene con **su propio código interno**
(``873420``, ``FMQ0041``, ``105M01``). Son dos idiomas distintos para el mismo
servicio, y el archivo no se puede cargar hasta que estén homologados.

CÓMO HOMOLOGA (en este orden, siempre **dentro de la misma factura**):

    1. código directo   — el código del ADRES ya existe en DGH.
    2. SOAT → CUPS      — con el Homologador Gold Standard CUPS/SOAT.
    3. descripción igual
    4. descripción que empieza igual (una es el principio de la otra)
    5. valor exacto + palabras en común (≥ la mitad de las palabras)
    6. descripción muy parecida (85 % o más)

Lo que no se pueda homologar **no se inventa**: el renglón igual sale en el
archivo (con SLNSERPRO vacío) y además queda listado en la hoja REVISAR con el
mejor candidato encontrado, para que el auditor decida.

TODOS LOS SERVICIOS QUEDAN. No se pierde ni un renglón por no haber podido
homologarlo: el archivo de salida tiene tantas filas como glosas trae el ADRES.

GUARDIÁN DE VALORES. Igual que el cruce de DGH de la Suite Cartera: la objeción
nunca puede superar el valor del servicio en DGH ni el saldo de la factura. Si
se pasa, se ajusta al tope y queda anotado en REVISAR.

LOTES DE 300 FACTURAS. DGH no recibe más de 300 facturas por archivo, así que
la salida se parte en ``OBJECIONES_ADRES_LOTE_01.xlsx``, ``_02.xlsx``…

OJO CON EL CÓDIGO DE GLOSA (CRNCONOBJ). El ADRES usa códigos numéricos de
cuatro dígitos (3106, 3209, 4506…) y DGH usa los de seis del Manual Único
(SO3401, CL0101…). **No hay una tabla oficial que los equipare**, así que el
bot escribe el código del ADRES tal cual y le deja al auditor la traducción:
la hoja CODIGOS del archivo REVISAR trae la lista de códigos con el grupo que
les corresponde, y ``--mapa-codigos`` recibe el JSON con la equivalencia
definitiva. Haga siempre el piloto de una factura antes del cargue masivo.

USO:

    py tools\\organizar_objeciones_adres.py ^
        --adres        "ADRES_DANIEL_31068.xlsx" ^
        --dgh          "DGReport_1.xlsx" ^
        --homologador  "01._Homologador_Gold_Standard_CUPS_a_SOAT.xlsx" ^
        --salida       "OBJECIONES_ADRES"

INSTALACIÓN (una vez):

    py -m pip install openpyxl
"""

from __future__ import annotations

import argparse
import datetime as _dt
import difflib
import json
import logging
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

# Un solo lector de pesos para todos los bots (ver tools/_dinero.py).
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _dinero import a_numero, a_texto  # noqa: E402
from ajustar_detallado_glosas import _norm, _norm_codigo, _norm_desc  # noqa: E402

logger = logging.getLogger("objeciones_adres")


# ─── Formato de salida: el layout de OBJECIONES que recibe DGH ───────────────

COLUMNAS_OBJECIONES: tuple[str, ...] = (
    "CDCONSEC",
    "CDFECDOC",
    "CRNCXC",
    "CROFECOBJ",
    "CROREFERE",
    "CROOBSERV",
    "CROCLAOBJ",
    "CRNCLAOBJ",
    "GENUSUARIO4",
    "CRNCONOBJ",
    "SLNSERPRO",
    "IDRIPS",
    "CTNCENCOS",
    "CROVALOBJ",
    "CRDOBSERV",
    "CROTIPOBJ",
)

# Formato de celda por columna, copiado del archivo real de COOSALUD.
# CDCONSEC y GENUSUARIO4 van como TEXTO; los valores, como número.
FORMATOS_OBJECIONES: dict[str, str] = {
    "CDCONSEC": "@",
    "CDFECDOC": "mm-dd-yy",
    "CRNCXC": "@",
    "CROFECOBJ": "mm-dd-yy",
    "CROREFERE": "@",
    "CROOBSERV": "@",
    "CROCLAOBJ": "General",
    "CRNCLAOBJ": "@",
    "GENUSUARIO4": "@",
    "CRNCONOBJ": "@",
    "SLNSERPRO": "@",
    "IDRIPS": "@",
    "CTNCENCOS": "@",
    "CROVALOBJ": '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-',
    "CRDOBSERV": "@",
    "CROTIPOBJ": "0",
}

CROCLAOBJ_CONST = 0
GENUSUARIO4_CONST = "999"
MAX_FACTURAS_POR_LOTE = 300  # tope de DGH (el mismo del cruce de la Suite)
ANCHO_FACTURA = 10  # HUS + 10 dígitos, como los escribe DGH


# ─── Motivos de revisión ─────────────────────────────────────────────────────

REV_SIN_SLNSERPRO = "SIN CODIGO DE SERVICIO EN DGH"
REV_SIN_ELEMENTO = "GLOSA DE TODA LA RECLAMACION (sin servicio)"
REV_SIN_CODIGO_GLOSA = "SIN CODIGO DE GLOSA (glosa total por FURIPS)"
REV_FACTURA_SIN_DGH = "LA FACTURA NO ESTA EN EL REPORTE DE DGH"
REV_VALOR_AJUSTADO = "VALOR AJUSTADO AL TOPE DE DGH"
REV_TODO_ACEPTADO = "GLOSA ACEPTADA COMPLETA (no habria que objetarla)"
REV_REPITE_TOTAL = "RENGLON QUITADO: repetia el glosado de toda la reclamacion"
REV_DUPLICADO = "RENGLON QUITADO: duplicado del ADRES"
REV_AJUSTE_REPORTE = "VALOR AJUSTADO PARA CUADRAR CON EL REPORTE DEL ADRES"
REV_FACTURA_SIN_REPORTE = "LA FACTURA NO ESTA EN EL REPORTE DE RECLAMACIONES"
REV_SERVICIO_ASIGNADO = "CODIGO DE SERVICIO ASIGNADO (no salio del cruce)"


# ─── Datos de entrada ────────────────────────────────────────────────────────


@dataclass
class LineaDgh:
    """Un renglón del reporte de DGH (un servicio facturado)."""

    slnserpro: str = ""
    cups: str = ""
    desc_institucional: str = ""
    desc_cups: str = ""
    cod_medicamento: str = ""
    cod_med_factura: str = ""
    nombre_medicamento: str = ""
    centro_costo: str = ""
    valor: float = 0.0
    saldo: float = 0.0

    def descripciones(self) -> tuple[str, ...]:
        return (
            _norm_desc(self.desc_institucional),
            _norm_desc(self.desc_cups),
            _norm_desc(self.nombre_medicamento),
        )

    def codigos(self) -> tuple[str, ...]:
        return (
            _norm_codigo(self.slnserpro),
            _norm_codigo(self.cups),
            _norm_codigo(self.cod_medicamento),
            _norm_codigo(self.cod_med_factura),
        )

    def rotulo(self) -> str:
        return self.desc_institucional or self.desc_cups or self.nombre_medicamento


@dataclass
class FilaAdres:
    """Una glosa del ADRES: un servicio de una factura con su causal."""

    factura: str = ""
    cod_elemento: str = ""
    descripcion: str = ""
    tipo_elemento: str = ""
    codigo_glosa: str = ""
    texto_glosa: str = ""
    clasificacion: str = ""
    cantidad: float = 0.0
    valor_reclamado: float = 0.0
    valor_glosado: float = 0.0
    valor_aceptado: float = 0.0
    fila_excel: int = 0


@dataclass
class Resolucion:
    """Con qué código de DGH se quedó el renglón y cómo se llegó a él."""

    slnserpro: str = ""
    metodo: str = ""
    candidato: str = ""
    candidato_desc: str = ""
    tope_servicio: float = 0.0
    saldo: float = 0.0


@dataclass
class Revision:
    """Un renglón que el auditor tiene que mirar antes de cargar."""

    factura: str
    motivo: str
    cod_elemento: str = ""
    descripcion: str = ""
    codigo_glosa: str = ""
    valor_glosado: float = 0.0
    detalle: str = ""


# ─── Lectura del Excel del ADRES ─────────────────────────────────────────────


def _norm_header(h: object) -> str:
    s = unicodedata.normalize("NFKD", str(h or "").strip().upper())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.split())


# Encabezados de la hoja de glosas (la que trae la clasificación del auditor).
COLUMNAS_GLOSAS = {
    "factura": {"NUMERO FACTURA", "NRO FACTURA", "FACTURA"},
    "paquete": {"NUMERO PAQUETE", "PAQUETE"},
    "cantidad": {"CANTIDAD RECLAMADO", "CANTIDAD RECLAMADA", "CANTIDAD"},
    "valor_reclamado": {"VALOR RECLAMADO"},
    "valor_glosado": {"VALOR GLOSADO"},
    "cod_elemento": {"COD ELEMENTO", "CODIGO ELEMENTO"},
    "texto_glosa": {"DESCRIPCION GLOSA"},
    "codigo_glosa": {"CODIGO NUMERICO", "CODIGO GLOSA"},
    "clasificacion": {"CLASIFICACION DE LA GLOSA", "CLASIFICACION"},
    "valor_aceptado": {"VALOR ACEPTADO"},
}

# La hoja de glosas se reconoce por traer TODAS estas columnas. Hace falta pedir
# las cuatro: el mismo libro trae una tabla dinámica («Hoja3») con el código y
# el elemento pero con los valores sumados («Suma de Valor Glosado»), y si solo
# se piden dos columnas el bot se queda con la dinámica y saca todo en cero.
HOJA_GLOSAS_OBLIGATORIAS = ("codigo_glosa", "cod_elemento", "valor_glosado", "valor_aceptado")


# Encabezados de la hoja del reporte crudo (la que trae la descripción limpia).
COLUMNAS_REPORTE = {
    "factura": {"NUMERO FACTURA", "NRO FACTURA", "FACTURA"},
    "paquete": {"NUMERO PAQUETE", "PAQUETE"},
    "tipo_elemento": {"TIPO ELEMENTO"},
    "cod_elemento": {"COD ELEMENTO", "CODIGO ELEMENTO"},
    "descripcion": {"DESCRIPCION ELEMENTO"},
}


def _indices(headers: list, alias: dict[str, set[str]]) -> dict[str, int]:
    """{clave: índice de columna}, mapeando por nombre de encabezado."""
    norm = [_norm_header(h) for h in headers]
    idx: dict[str, int] = {}
    for clave, nombres in alias.items():
        pos = next((i for i, h in enumerate(norm) if h in nombres), None)
        if pos is not None:
            idx[clave] = pos
    return idx


def _celda(fila: tuple, idx: dict[str, int], clave: str) -> object:
    pos = idx.get(clave)
    if pos is None or pos >= len(fila):
        return None
    return fila[pos]


def _texto(v: object) -> str:
    return " ".join(str(v).split()) if v is not None else ""


def limpiar_texto_glosa(texto: object, codigo_glosa: str, clasificacion: str = "") -> str:
    """Deja solo el texto de la causal.

    El ADRES concatena todo en una celda:

        ``39145-Procedimientos-3202- La consulta no esta justificada -3202- La
        consulta no esta justificada``

    Como la causal se repite detrás de su código, se corta en la **última**
    aparición de ``<código>-`` y se devuelve lo que sigue.

    Los renglones de **glosa total por FURIPS** no traen código de causal: ahí
    la celda solo repite el código, el tipo y el nombre del servicio, que ya van
    aparte en la objeción. En ese caso se devuelve la clasificación que escribió
    el auditor, que es lo único que dice de qué se trata la glosa.
    """
    t = _texto(texto)
    if not codigo_glosa:
        return _texto(clasificacion)
    marca = f"{codigo_glosa}-"
    corte = t.rfind(marca)
    if corte < 0:
        return t
    return t[corte + len(marca) :].strip()


def leer_adres(
    ruta: Path, paquete: str = "", hoja_glosas: str = "", hoja_reporte: str = ""
) -> list[FilaAdres]:
    """Lee el Excel del ADRES y devuelve una glosa por fila.

    Busca sola la hoja de glosas (la que trae CODIGO NUMERICO y VALOR ACEPTADO)
    y la del reporte crudo (la que trae DESCRIPCION ELEMENTO), y le pega a cada
    glosa la descripción limpia del servicio.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(ruta), data_only=True, read_only=True)
    try:
        nombre_glosas = hoja_glosas or _hoja_con(wb, COLUMNAS_GLOSAS, HOJA_GLOSAS_OBLIGATORIAS)
        if not nombre_glosas:
            raise ValueError(
                "No encontré en el Excel del ADRES una hoja con las columnas "
                + ", ".join(sorted(HOJA_GLOSAS_OBLIGATORIAS))
                + ". Indíquela con --hoja-glosas."
            )
        nombre_reporte = hoja_reporte or _hoja_con(wb, COLUMNAS_REPORTE, ("descripcion",))
        descripciones = _leer_descripciones(wb[nombre_reporte]) if nombre_reporte else {}
        logger.info(
            "  Hoja de glosas: %s%s",
            nombre_glosas,
            f"  |  descripciones desde: {nombre_reporte}" if nombre_reporte else "",
        )
        return _leer_glosas(wb[nombre_glosas], descripciones, paquete)
    finally:
        wb.close()


def _hoja_con(
    wb, alias: dict[str, set[str]], obligatorias: tuple[str, ...], max_filas: int = 1
) -> str:
    """Nombre de la primera hoja cuyo encabezado tenga todas esas columnas.

    `max_filas` sube cuando el encabezado no está en la primera fila (el reporte
    del ADRES trae encima una fila con los totales del paquete).
    """
    for nombre in wb.sheetnames:
        for headers in wb[nombre].iter_rows(min_row=1, max_row=max_filas, values_only=True):
            if not headers:
                continue
            idx = _indices(list(headers), alias)
            if all(c in idx for c in obligatorias):
                return nombre
    return ""


def _leer_descripciones(ws) -> dict[tuple[str, str], tuple[str, str]]:
    """{(factura, código normalizado): (descripción, tipo)} del reporte crudo."""
    filas = ws.iter_rows(values_only=True)
    headers = next(filas, None)
    if not headers:
        return {}
    idx = _indices(list(headers), COLUMNAS_REPORTE)
    fuera: dict[tuple[str, str], tuple[str, str]] = {}
    for fila in filas:
        if fila is None:
            continue
        factura = _texto(_celda(fila, idx, "factura"))
        descripcion = _texto(_celda(fila, idx, "descripcion"))
        if not factura or not descripcion:
            continue
        clave = (factura, _norm_codigo(_celda(fila, idx, "cod_elemento")))
        fuera.setdefault(clave, (descripcion, _texto(_celda(fila, idx, "tipo_elemento"))))
    return fuera


def _leer_glosas(ws, descripciones: dict, paquete: str) -> list[FilaAdres]:
    filas = ws.iter_rows(values_only=True)
    headers = next(filas, None)
    idx = _indices(list(headers or []), COLUMNAS_GLOSAS)
    salida: list[FilaAdres] = []
    for n, fila in enumerate(filas, start=2):
        if fila is None:
            continue
        factura = _texto(_celda(fila, idx, "factura"))
        if not factura:
            continue
        if paquete and _texto(_celda(fila, idx, "paquete")) != paquete:
            continue
        cod = _texto(_celda(fila, idx, "cod_elemento"))
        codigo_glosa = _texto(_celda(fila, idx, "codigo_glosa"))
        descripcion, tipo = descripciones.get((factura, _norm_codigo(cod)), ("", ""))
        salida.append(
            FilaAdres(
                factura=factura,
                cod_elemento=cod,
                descripcion=descripcion,
                tipo_elemento=tipo,
                codigo_glosa=codigo_glosa,
                texto_glosa=limpiar_texto_glosa(
                    _celda(fila, idx, "texto_glosa"),
                    codigo_glosa,
                    _texto(_celda(fila, idx, "clasificacion")),
                ),
                clasificacion=_texto(_celda(fila, idx, "clasificacion")),
                cantidad=a_numero(_celda(fila, idx, "cantidad")),
                valor_reclamado=a_numero(_celda(fila, idx, "valor_reclamado")),
                valor_glosado=a_numero(_celda(fila, idx, "valor_glosado")),
                valor_aceptado=a_numero(_celda(fila, idx, "valor_aceptado")),
                fila_excel=n,
            )
        )
    return salida


# ─── Lectura del reporte de DGH ──────────────────────────────────────────────

COLUMNAS_DGH = {
    "slnserpro": {"SLNSERPRO SERVICIO", "SLNSERPRO_SERVICIO"},
    "desc_institucional": {"DESCRIPCION INSTITUCIONAL"},
    "cups": {"SLNSERPRO CUPS", "SLNSERPRO_CUPS"},
    "desc_cups": {"DESCRIPCION CUPS"},
    "cod_medicamento": {"CODIGO MEDICAMENTO", "CODIGO_MEDICAMENTO"},
    "cod_med_factura": {"COD MED FACTURA", "COD_MED_FACTURA"},
    "nombre_medicamento": {"NOMBRE MEDICAMENTO", "NOMBRE_MEDICAMENTO"},
    "centro_costo": {"CENTRO COSTO", "CENTRO_COSTO"},
    "factura": {"FACTURA", "NUMERO FACTURA"},
    "valor": {"VR SERVICIO", "VR_SERVICIO", "VALOR SERVICIO"},
    "saldo": {"SALDO FACT", "SALDO_FACT", "SALDO"},
}


def leer_dgh(ruta: Path) -> dict[str, list[LineaDgh]]:
    """{factura: [servicios facturados]} desde el DGReport."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(ruta), data_only=True, read_only=True)
    try:
        nombre = _hoja_con(wb, COLUMNAS_DGH, ("factura", "valor"))
        if not nombre:
            raise ValueError(
                f"{ruta.name} no parece el reporte de DGH: no encontré las columnas "
                "FACTURA y Vr_SERVICIO."
            )
        ws = wb[nombre]
        filas = ws.iter_rows(values_only=True)
        idx = _indices(list(next(filas, ()) or ()), COLUMNAS_DGH)
        fuera: dict[str, list[LineaDgh]] = defaultdict(list)
        for fila in filas:
            if fila is None:
                continue
            factura = _texto(_celda(fila, idx, "factura"))
            if not factura:
                continue
            servicio = _texto(_celda(fila, idx, "slnserpro"))
            medicamento = _texto(_celda(fila, idx, "cod_medicamento"))
            fuera[factura].append(
                LineaDgh(
                    slnserpro=servicio or medicamento,
                    cups=_texto(_celda(fila, idx, "cups")),
                    desc_institucional=_texto(_celda(fila, idx, "desc_institucional")),
                    desc_cups=_texto(_celda(fila, idx, "desc_cups")),
                    cod_medicamento=medicamento,
                    cod_med_factura=_texto(_celda(fila, idx, "cod_med_factura")),
                    nombre_medicamento=_texto(_celda(fila, idx, "nombre_medicamento")),
                    centro_costo=_texto(_celda(fila, idx, "centro_costo")),
                    valor=a_numero(_celda(fila, idx, "valor")),
                    saldo=a_numero(_celda(fila, idx, "saldo")),
                )
            )
        return dict(fuera)
    finally:
        wb.close()


# ─── Lectura del reporte de reclamaciones del ADRES ──────────────────────────

# El ReporteReclamPAQUETE_*.xlsx: una fila por reclamación, con lo que el ADRES
# dice oficialmente que reclamó, aprobó y glosó de cada factura. Es la cifra que
# manda: contra ella tiene que cuadrar el archivo que se carga a DGH.
COLUMNAS_RECLAMACION = {
    "factura": {"NUMERO DE FACTURA", "NUMERO FACTURA", "FACTURA"},
    "radicacion": {"NUMERO DE RADICACION", "NUMERO RADICACION"},
    "reclamado": {"VALOR RECLAMADO"},
    "aprobado": {"VALOR APROBADO"},
    "glosado": {"VALOR GLOSADO"},
    "paquete": {"PAQUETE", "NUMERO PAQUETE"},
    "estado": {"ESTADO RECLAMACION", "ESTADO DE RECLAMACION"},
    "extemporaneidad": {"TIPO DE EXTEMPORANEIDAD", "TIPO EXTEMPORANEIDAD"},
}


@dataclass
class Reclamacion:
    """Lo que el ADRES reporta de una factura, a nivel de reclamación."""

    factura: str = ""
    radicacion: str = ""
    reclamado: float = 0.0
    aprobado: float = 0.0
    glosado: float = 0.0
    estado: str = ""
    extemporaneidad: str = ""


def leer_reporte_reclamaciones(ruta: Path, paquete: str = "") -> dict[str, Reclamacion]:
    """{factura larga: Reclamacion} del ReporteReclamPAQUETE del ADRES."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(ruta), data_only=True, read_only=True)
    try:
        # El encabezado NO está en la primera fila: el archivo trae encima una
        # fila con los totales del paquete. Por eso se busca en las primeras.
        nombre = _hoja_con(wb, COLUMNAS_RECLAMACION, ("factura", "glosado"), max_filas=10)
        if not nombre:
            raise ValueError(
                f"{ruta.name} no parece el reporte de reclamaciones del ADRES: no encontré "
                "las columnas 'Número de Factura' y 'Valor Glosado'."
            )
        ws = wb[nombre]
        idx: dict[str, int] = {}
        fuera: dict[str, Reclamacion] = {}
        for fila in ws.iter_rows(values_only=True):
            if fila is None:
                continue
            if not idx:
                # El archivo trae una fila de totales por encima del encabezado.
                candidato = _indices(list(fila), COLUMNAS_RECLAMACION)
                if "factura" in candidato and "glosado" in candidato:
                    idx = candidato
                continue
            factura = _texto(_celda(fila, idx, "factura"))
            if not factura:
                continue
            if paquete and _texto(_celda(fila, idx, "paquete")) != paquete:
                continue
            fuera[factura_larga(factura)] = Reclamacion(
                factura=factura,
                radicacion=_texto(_celda(fila, idx, "radicacion")),
                reclamado=a_numero(_celda(fila, idx, "reclamado")),
                aprobado=a_numero(_celda(fila, idx, "aprobado")),
                glosado=a_numero(_celda(fila, idx, "glosado")),
                estado=_texto(_celda(fila, idx, "estado")),
                extemporaneidad=_texto(_celda(fila, idx, "extemporaneidad")),
            )
        return fuera
    finally:
        wb.close()


# ─── Lectura del Homologador Gold Standard (SOAT → CUPS) ─────────────────────

COLUMNAS_HOMOLOGADOR = {
    "cups": {"CUPS VIGENTE", "CODIGO CUPS", "CUPS"},
    "soat": {"CODIGO SOAT", "COD SOAT", "SOAT"},
}


def leer_homologador(ruta: Path, hoja: str = "CUPS") -> dict[str, set[str]]:
    """{código SOAT normalizado: {códigos CUPS}} del Homologador Gold Standard.

    El encabezado no está en la primera fila (el archivo trae una portada), así
    que se busca la fila que tenga a la vez «CUPS VIGENTE» y «Código SOAT».
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(ruta), data_only=True, read_only=True)
    try:
        ws = wb[hoja] if hoja in wb.sheetnames else wb[wb.sheetnames[0]]
        mapa: dict[str, set[str]] = defaultdict(set)
        idx: dict[str, int] = {}
        for fila in ws.iter_rows(values_only=True):
            if fila is None:
                continue
            if not idx:
                candidato = _indices(list(fila), COLUMNAS_HOMOLOGADOR)
                if "cups" in candidato and "soat" in candidato:
                    idx = candidato
                continue
            soat = _norm_codigo(_celda(fila, idx, "soat"))
            cups = _norm_codigo(_celda(fila, idx, "cups"))
            if soat and cups:
                mapa[soat].add(cups)
        if not idx:
            logger.warning(
                "  %s: no encontré las columnas 'CUPS VIGENTE' y 'Código SOAT'; "
                "sigo sin homologación SOAT→CUPS.",
                ruta.name,
            )
        return dict(mapa)
    finally:
        wb.close()


# ─── Homologación: de qué servicio de DGH habla esta glosa del ADRES ─────────

# Palabras que no distinguen nada al comparar descripciones.
_VACIAS = frozenset(
    {
        "DE",
        "DEL",
        "LA",
        "EL",
        "LOS",
        "LAS",
        "Y",
        "O",
        "EN",
        "POR",
        "CON",
        "A",
        "PARA",
        "UN",
        "UNA",
        "AL",
    }
)
MINIMO_PREFIJO = 12  # caracteres mínimos para comparar «empieza igual»
MINIMO_PALABRAS = 0.5  # fracción de palabras en común que exige el paso 5
MINIMO_SIMILITUD = 0.85  # parecido mínimo del paso 6
TOLERANCIA_PESOS = 0.5

METODO_CODIGO = "codigo directo"
METODO_SOAT = "homologado SOAT→CUPS"
METODO_DESC = "descripcion igual"
METODO_PREFIJO = "descripcion empieza igual"
METODO_VALOR = "valor + palabras en comun"
METODO_SIMILITUD = "descripcion parecida"
METODO_SIN_CRUCE = ""


def _palabras(desc: str) -> set[str]:
    return {p for p in desc.split() if len(p) > 1 and p not in _VACIAS}


def _mismo_valor(a: float, b: float) -> bool:
    return abs(a - b) < TOLERANCIA_PESOS


def _mejor_parecido(objetivo: str, lineas: list[LineaDgh]) -> tuple[float, LineaDgh | None]:
    """La línea de DGH cuya descripción más se parece, con su puntaje 0..1."""
    mejor: tuple[float, LineaDgh | None] = (0.0, None)
    if not objetivo:
        return mejor
    for linea in lineas:
        for desc in linea.descripciones():
            if not desc:
                continue
            puntaje = difflib.SequenceMatcher(None, objetivo, desc).ratio()
            if puntaje > mejor[0]:
                mejor = (puntaje, linea)
    return mejor


def resolver_slnserpro(
    fila: FilaAdres, lineas: list[LineaDgh], soat_a_cups: dict[str, set[str]]
) -> Resolucion:
    """Busca en las líneas de DGH de ESA factura el servicio del que habla la glosa.

    Devuelve el código interno de DGH y por qué camino se llegó a él. Si ningún
    paso da resultado, devuelve el código vacío y el mejor candidato encontrado
    como pista para el auditor — nunca lo escribe como si fuera cierto.
    """
    codigo = _norm_codigo(fila.cod_elemento)
    descripcion = _norm_desc(fila.descripcion)
    if not lineas:
        return Resolucion()

    def resultado(linea: LineaDgh, metodo: str) -> Resolucion:
        mismos = [x for x in lineas if x.slnserpro == linea.slnserpro]
        return Resolucion(
            slnserpro=linea.slnserpro,
            metodo=metodo,
            candidato=linea.slnserpro,
            candidato_desc=linea.rotulo(),
            tope_servicio=round(sum(x.valor for x in mismos), 2),
            saldo=linea.saldo,
        )

    # 1) el código del ADRES ya existe tal cual en DGH.
    if codigo:
        for linea in lineas:
            if codigo in linea.codigos():
                return resultado(linea, METODO_CODIGO)

    # 2) el código SOAT del ADRES traducido a CUPS con el Homologador.
    cups = soat_a_cups.get(codigo, set()) if codigo else set()
    if cups:
        for linea in lineas:
            if _norm_codigo(linea.slnserpro) in cups or _norm_codigo(linea.cups) in cups:
                return resultado(linea, METODO_SOAT)

    if descripcion:
        # 3) la misma descripción.
        for linea in lineas:
            if descripcion in linea.descripciones():
                return resultado(linea, METODO_DESC)
        # 4) una descripción es el principio de la otra (el ADRES la recorta).
        for linea in lineas:
            for otra in linea.descripciones():
                if not otra or min(len(otra), len(descripcion)) < MINIMO_PREFIJO:
                    continue
                if otra.startswith(descripcion) or descripcion.startswith(otra):
                    return resultado(linea, METODO_PREFIJO)

    # 5) el valor coincide exacto y comparten al menos la mitad de las palabras.
    #    (El ADRES dice «Habitación de cuatro ó mas camas» y DGH «Internación
    #    complejidad alta cuatro o mas camas»: distinto nombre, mismo servicio
    #    y mismo valor.)
    mias = _palabras(descripcion)
    unitario = fila.valor_reclamado / fila.cantidad if fila.cantidad else 0.0
    candidatos = [
        linea
        for linea in lineas
        if (
            _mismo_valor(linea.valor, fila.valor_reclamado)
            or (unitario and _mismo_valor(linea.valor, unitario))
        )
        and mias
        and max((len(mias & _palabras(d)) / len(mias)) for d in linea.descripciones())
        >= MINIMO_PALABRAS
    ]
    if candidatos and len({c.slnserpro for c in candidatos}) == 1:
        return resultado(candidatos[0], METODO_VALOR)

    # 6) descripciones muy parecidas (el ADRES trae tildes rotas y abreviaturas).
    puntaje, linea = _mejor_parecido(descripcion, lineas)
    if linea is not None and puntaje >= MINIMO_SIMILITUD:
        return resultado(linea, METODO_SIMILITUD)

    return Resolucion(
        candidato=linea.slnserpro if linea is not None else "",
        candidato_desc=(f"{linea.rotulo()} (parecido {puntaje:.0%})" if linea is not None else ""),
    )


# ─── Reglas del formato OBJECIONES ───────────────────────────────────────────

PREFIJOS_FACTURA = ("HUS",)


def factura_larga(fac: object, ancho: int = ANCHO_FACTURA) -> str:
    """HUS311371 → HUS0000311371. Idempotente; si no reconoce el patrón, la deja igual."""
    s = _texto(fac).replace(" ", "").upper()
    for prefijo in PREFIJOS_FACTURA:
        if s.startswith(prefijo) and s[len(prefijo) :].isdigit():
            return prefijo + s[len(prefijo) :].lstrip("0").zfill(ancho)
    return s


# El grupo del Manual Único al que corresponde cada clasificación del auditor.
# Sirve de guía para traducir el código numérico del ADRES: no lo reemplaza.
GRUPOS_POR_CLASIFICACION: tuple[tuple[str, str], ...] = (
    ("PERTINENCIA", "CL"),
    ("CALIDAD", "CL"),
    ("SOPORTE", "SO"),
    ("HABILITACION", "SO"),
    ("TARIFA", "TA"),
    ("FACTURACION", "FA"),
    ("FACOSTE", "FA"),
    ("CUPS", "FA"),
    ("COBERTURA", "CO"),
    ("AUTORIZACION", "AU"),
)
GRUPO_CLINICO = "CL"


def grupo_dgh(clasificacion: str) -> str:
    """El grupo del Manual Único (CL/SO/TA/FA/CO/AU) que sugiere la clasificación
    que escribió el auditor. Cadena vacía si no se reconoce."""
    texto = _norm(clasificacion)
    for palabra, grupo in GRUPOS_POR_CLASIFICACION:
        if palabra in texto:
            return grupo
    return ""


def crotipobj_factura(clasificaciones: set[str]) -> int:
    """Tipo de objeción de la factura, por la mezcla de causales que trae:

    - solo administrativas (soportes, tarifas, facturación…) → 0
    - solo clínicas (pertinencia/calidad)                    → 1
    - administrativas + clínicas                             → 2
    """
    grupos = {grupo_dgh(c) for c in clasificaciones}
    clinicas = GRUPO_CLINICO in grupos
    administrativas = any(g != GRUPO_CLINICO for g in grupos)
    if clinicas and administrativas:
        return 2
    if clinicas:
        return 1
    return 0


def construir_crdobserv(codigo: str, fila: FilaAdres, valor: int) -> str:
    """El texto de la objeción, con la forma que usa DGH:

    ``<código> <causal> (<código servicio>-<servicio>)$<valor objetado>``
    """
    partes: list[str] = []
    if codigo:
        partes.append(codigo)
    if fila.texto_glosa:
        partes.append(fila.texto_glosa)
    if fila.cod_elemento and fila.descripcion:
        partes.append(f"({fila.cod_elemento}-{fila.descripcion})")
    elif fila.cod_elemento:
        partes.append(f"({fila.cod_elemento})")
    return " ".join(partes).strip() + f"${valor}"


def aplicar_tope(valor: float, tope_servicio: float, saldo: float) -> tuple[int, str]:
    """La objeción nunca supera el valor del servicio en DGH ni el saldo de la
    factura (la misma guarda del cruce de DGH de la Suite Cartera).

    Devuelve (valor final en pesos enteros, motivo del ajuste o cadena vacía)."""
    final = float(valor)
    motivo = ""
    for tope, nombre in (
        (tope_servicio, "valor del servicio en DGH"),
        (saldo, "saldo de la factura"),
    ):
        if tope and tope > 0 and final > tope + TOLERANCIA_PESOS:
            motivo = f"{a_texto(final)} → {a_texto(tope)} ({nombre})"
            final = tope
    return int(round(final)), motivo


# ─── Conciliación con el reporte del ADRES ───────────────────────────────────

# El detalle del ADRES cuenta la misma plata más de una vez. Dos formas:
#
#   1. Renglones que repiten el glosado de TODA la reclamación. Cuando el ADRES
#      glosa la reclamación entera por el FURIPS, además de listar los servicios
#      mete una fila por cada causal de reclamación (2102, 2103…) con el valor
#      COMPLETO. La factura HUS0000311371 aparece así por $39.722.100 cuando el
#      ADRES reporta $13.240.700: el detalle ($13.240.700) más dos renglones de
#      causal, cada uno por el total.
#   2. Renglones repetidos: el mismo servicio, misma cantidad y mismo valor,
#      listado otra vez porque le cayó otra causal encima.
#
# Sumar todo eso y cargarlo a DGH sería objetar tres veces la misma plata. Aquí
# se quita el sobrante, de menos invasivo a más, y todo lo que se quita queda
# anotado en REVISAR.


@dataclass
class Conciliacion:
    """Cómo quedó una factura frente a lo que reporta el ADRES."""

    filas: list[FilaAdres] = field(default_factory=list)
    quitadas: list[tuple[FilaAdres, str]] = field(default_factory=list)
    objetivo: float = 0.0
    ajuste: float = 0.0  # lo que faltó para cuadrar exacto (+ sobra, − falta)


def _clave_repetido(fila: FilaAdres) -> tuple:
    """Dos renglones son el mismo si coinciden servicio, cantidad, valores y causal."""
    return (
        _norm_codigo(fila.cod_elemento),
        fila.cantidad,
        round(fila.valor_reclamado, 2),
        round(fila.valor_glosado, 2),
        fila.codigo_glosa,
    )


def conciliar_factura(filas: list[FilaAdres], glosado_reportado: float) -> Conciliacion:
    """Deja de la factura solo lo que suma el glosado que reporta el ADRES.

    Quita primero los renglones que repiten el total de la reclamación y después
    las repeticiones, siempre la más grande primero y **sin bajarse del
    objetivo** — nunca se quita de más. Lo que quede de diferencia se anota en
    `ajuste` para corregirlo en el renglón mayor, a la vista.
    """
    resultado = Conciliacion(filas=list(filas), objetivo=glosado_reportado)
    if glosado_reportado <= 0 or not filas:
        return resultado
    suma = round(sum(f.valor_glosado for f in resultado.filas), 2)

    # 1) los renglones que repiten el glosado de toda la reclamación.
    if suma > glosado_reportado + TOLERANCIA_PESOS:
        for fila in [
            f
            for f in resultado.filas
            if abs(f.valor_glosado - glosado_reportado) <= TOLERANCIA_PESOS
        ]:
            if len(resultado.filas) <= 1:
                break
            if round(suma - fila.valor_glosado, 2) < glosado_reportado - TOLERANCIA_PESOS:
                continue
            resultado.filas.remove(fila)
            resultado.quitadas.append((fila, REV_REPITE_TOTAL))
            suma = round(suma - fila.valor_glosado, 2)

    # 2) las repeticiones, la más grande primero, sin bajarse del objetivo.
    while suma > glosado_reportado + TOLERANCIA_PESOS:
        vistas: set[tuple] = set()
        repetidas: list[FilaAdres] = []
        for fila in resultado.filas:
            clave = _clave_repetido(fila)
            if clave in vistas:
                repetidas.append(fila)
            else:
                vistas.add(clave)
        repetidas = [
            f
            for f in repetidas
            if round(suma - f.valor_glosado, 2) >= glosado_reportado - TOLERANCIA_PESOS
        ]
        if not repetidas:
            break
        fila = max(repetidas, key=lambda f: f.valor_glosado)
        resultado.filas.remove(fila)
        resultado.quitadas.append((fila, REV_DUPLICADO))
        suma = round(suma - fila.valor_glosado, 2)

    resultado.ajuste = round(suma - glosado_reportado, 2)
    return resultado


# ─── Código de servicio para los renglones que no cruzaron ───────────────────


def servicio_principal(lineas: list[LineaDgh]) -> LineaDgh | None:
    """El servicio de más peso de la factura en DGH: el que más plata suma.

    Es el que se usa cuando el auditor pide que **ningún renglón quede sin
    código de servicio** y el cruce no encontró ninguno. No es una homologación:
    es un destino por defecto, y cada renglón así queda marcado en REVISAR.
    """
    if not lineas:
        return None
    suma: dict[str, float] = defaultdict(float)
    for linea in lineas:
        if linea.slnserpro:
            suma[linea.slnserpro] += linea.valor
    if not suma:
        return None
    codigo = max(suma.items(), key=lambda kv: (kv[1], kv[0]))[0]
    return next(x for x in lineas if x.slnserpro == codigo)


# ─── Armado de los registros ─────────────────────────────────────────────────


@dataclass
class Conversion:
    """Lo que salió de convertir el ADRES: los registros y lo que hay que mirar."""

    registros: list[dict] = field(default_factory=list)
    revisiones: list[Revision] = field(default_factory=list)
    metodos: dict[str, int] = field(default_factory=dict)
    recorte: float = 0.0  # lo que le quitó el guardián de valores de DGH
    quitadas: int = 0  # renglones repetidos que sacó la conciliación
    ajustadas: int = 0  # facturas que hubo que ajustar para cuadrar
    asignados: int = 0  # renglones con código de servicio puesto por defecto


def _conciliar(
    filas: list[FilaAdres],
    reclamaciones: dict[str, Reclamacion] | None,
    salida: Conversion,
) -> list[FilaAdres]:
    """Quita del detalle del ADRES los renglones que repiten plata ya contada.

    Devuelve los renglones que quedan. La diferencia que sobre después de esto la
    cuadra `cuadrar_con_reporte`, ya sobre los valores finales.
    """
    if not reclamaciones:
        return filas

    por_factura: dict[str, list[FilaAdres]] = defaultdict(list)
    for fila in filas:
        por_factura[factura_larga(fila.factura)].append(fila)

    quedan: list[FilaAdres] = []
    sin_reporte: set[str] = set()
    for crncxc, del_factura in por_factura.items():
        reclamacion = reclamaciones.get(crncxc)
        if reclamacion is None:
            if crncxc not in sin_reporte:
                sin_reporte.add(crncxc)
                salida.revisiones.append(
                    Revision(
                        factura=crncxc,
                        motivo=REV_FACTURA_SIN_REPORTE,
                        valor_glosado=sum(f.valor_glosado for f in del_factura),
                        detalle="No se pudo cuadrar: esa factura no está en el reporte del ADRES.",
                    )
                )
            quedan.extend(del_factura)
            continue
        conciliacion = conciliar_factura(del_factura, reclamacion.glosado)
        for fila, motivo in conciliacion.quitadas:
            salida.quitadas += 1
            salida.revisiones.append(
                _revision(
                    fila,
                    crncxc,
                    motivo,
                    f"La factura sumaba de más; el ADRES reporta {a_texto(reclamacion.glosado)} "
                    "glosados.",
                )
            )
        quedan.extend(conciliacion.filas)
    return quedan


def construir_registros(
    filas: list[FilaAdres],
    dgh: dict[str, list[LineaDgh]],
    soat_a_cups: dict[str, set[str]],
    fecha: _dt.datetime,
    mapa_codigos: dict[str, str] | None = None,
    incluir_glosa_total: bool = True,
    reclamaciones: dict[str, Reclamacion] | None = None,
    completar_servicios: bool = False,
) -> Conversion:
    """Convierte las glosas del ADRES en filas del formato OBJECIONES.

    Sale una fila por glosa —ninguna se pierde— y una anotación en REVISAR por
    cada renglón que necesite ojo humano.

    Con `reclamaciones` (el ReporteReclamPAQUETE del ADRES) el total de cada
    factura queda **cuadrado contra lo que el ADRES reporta como glosado**: se
    quitan los renglones que el detalle repite y, si aún queda diferencia, se
    ajusta el renglón mayor. Todo lo quitado y lo ajustado va a REVISAR.

    Con `completar_servicios` ningún renglón sale sin código de servicio: el que
    no cruzó se lleva el candidato más parecido y, si no hay ninguno, el servicio
    de más peso de la factura en DGH. Cada uno queda marcado en REVISAR.
    """
    mapa_codigos = {k.strip(): v.strip() for k, v in (mapa_codigos or {}).items()}
    salida = Conversion()
    metodos: dict[str, int] = defaultdict(int)
    consecutivos: dict[str, int] = {}

    # CROTIPOBJ es de la factura entera, así que primero se ve qué trae cada una.
    clasificaciones: dict[str, set[str]] = defaultdict(set)
    for fila in filas:
        clasificaciones[factura_larga(fila.factura)].add(fila.clasificacion)

    filas = _conciliar(filas, reclamaciones, salida)

    facturas_sin_dgh: set[str] = set()
    for fila in filas:
        es_glosa_total = not fila.codigo_glosa
        if es_glosa_total and not incluir_glosa_total:
            continue
        crncxc = factura_larga(fila.factura)
        if crncxc not in consecutivos:
            consecutivos[crncxc] = len(consecutivos) + 1
        lineas = dgh.get(crncxc) or dgh.get(fila.factura) or []
        if not lineas and crncxc not in facturas_sin_dgh:
            facturas_sin_dgh.add(crncxc)
            salida.revisiones.append(
                Revision(
                    factura=crncxc,
                    motivo=REV_FACTURA_SIN_DGH,
                    detalle="Sin el reporte de DGH no se puede poner el código de servicio.",
                )
            )

        if fila.cod_elemento:
            resolucion = resolver_slnserpro(fila, lineas, soat_a_cups)
        else:
            resolucion = Resolucion()
        metodos[resolucion.metodo or METODO_SIN_CRUCE] += 1

        asignado = ""
        if completar_servicios and not resolucion.slnserpro:
            principal = servicio_principal(lineas)
            if resolucion.candidato:
                asignado = (
                    f"por parecido con {resolucion.candidato} {resolucion.candidato_desc}".strip()
                )
                resolucion.slnserpro = resolucion.candidato
            elif principal is not None:
                asignado = f"servicio de más peso de la factura: {principal.slnserpro} {principal.rotulo()}"
                resolucion.slnserpro = principal.slnserpro
            if asignado:
                salida.asignados += 1

        codigo = mapa_codigos.get(fila.codigo_glosa, fila.codigo_glosa)
        valor, ajuste = aplicar_tope(fila.valor_glosado, resolucion.tope_servicio, resolucion.saldo)
        # El ajuste que cuadra la factura con el reporte del ADRES manda sobre el
        # tope de DGH: el archivo tiene que sumar lo que el ADRES dice que glosó.
        salida.registros.append(
            {
                "CDCONSEC": str(consecutivos[crncxc]),
                "CDFECDOC": fecha,
                "CRNCXC": crncxc,
                "CROFECOBJ": fecha,
                "CROREFERE": None,
                "CROOBSERV": None,
                "CROCLAOBJ": CROCLAOBJ_CONST,
                "CRNCLAOBJ": None,
                "GENUSUARIO4": GENUSUARIO4_CONST,
                "CRNCONOBJ": codigo or None,
                "SLNSERPRO": resolucion.slnserpro or None,
                "IDRIPS": None,
                "CTNCENCOS": None,
                "CROVALOBJ": valor,
                "CRDOBSERV": construir_crdobserv(codigo, fila, valor),
                "CROTIPOBJ": crotipobj_factura(clasificaciones[crncxc]),
            }
        )

        if asignado:
            salida.revisiones.append(
                _revision(
                    fila, crncxc, REV_SERVICIO_ASIGNADO, f"{resolucion.slnserpro}: {asignado}"
                )
            )
        if ajuste:
            salida.recorte += fila.valor_glosado - valor
            salida.revisiones.append(_revision(fila, crncxc, REV_VALOR_AJUSTADO, ajuste))
        if not fila.cod_elemento:
            salida.revisiones.append(
                _revision(
                    fila,
                    crncxc,
                    REV_SIN_ELEMENTO,
                    "El ADRES glosó la reclamación entera; no señala un servicio.",
                )
            )
        elif not resolucion.slnserpro and lineas:
            salida.revisiones.append(
                _revision(
                    fila,
                    crncxc,
                    REV_SIN_SLNSERPRO,
                    f"Candidato más parecido: {resolucion.candidato} {resolucion.candidato_desc}".strip()
                    if resolucion.candidato
                    else "No hay ningún servicio parecido en esa factura.",
                )
            )
        if es_glosa_total:
            salida.revisiones.append(
                _revision(
                    fila,
                    crncxc,
                    REV_SIN_CODIGO_GLOSA,
                    f"Clasificación del auditor: {fila.clasificacion}"
                    if fila.clasificacion
                    else "",
                )
            )
        if fila.valor_aceptado and fila.valor_aceptado >= fila.valor_glosado - TOLERANCIA_PESOS:
            salida.revisiones.append(
                _revision(
                    fila,
                    crncxc,
                    REV_TODO_ACEPTADO,
                    f"Aceptado {a_texto(fila.valor_aceptado)} de {a_texto(fila.valor_glosado)}",
                )
            )

    salida.metodos = dict(metodos)
    cuadrar_con_reporte(salida, reclamaciones)
    return salida


def cuadrar_con_reporte(salida: Conversion, reclamaciones: dict[str, Reclamacion] | None) -> None:
    """Deja cada factura sumando EXACTAMENTE el glosado que reporta el ADRES.

    Se hace de último, sobre los valores ya definitivos, porque el guardián de
    valores de DGH pudo haber recortado algún renglón: si el cuadre se calculara
    antes, ese recorte volvería a descuadrar la factura.

    La diferencia se carga al renglón **mayor**, que es el que menos se deforma
    en proporción y el más fácil de ubicar para revisarlo. Todo ajuste queda
    anotado en REVISAR con el antes y el después.
    """
    if not reclamaciones:
        return
    por_factura: dict[str, list[dict]] = defaultdict(list)
    for registro in salida.registros:
        por_factura[str(registro["CRNCXC"])].append(registro)

    for crncxc, registros in por_factura.items():
        reclamacion = reclamaciones.get(crncxc)
        if reclamacion is None or reclamacion.glosado <= 0:
            continue
        suma = sum(int(r["CROVALOBJ"] or 0) for r in registros)
        sobra = round(suma - reclamacion.glosado)
        if abs(sobra) <= TOLERANCIA_PESOS:
            continue
        # Se empieza por el renglón mayor: es el que menos se deforma en
        # proporción. Si el sobrante no cabe en él, se sigue con el siguiente —
        # nunca se deja un valor negativo.
        ajustados = 0
        for registro in sorted(registros, key=lambda r: -int(r["CROVALOBJ"] or 0)):
            if abs(sobra) <= TOLERANCIA_PESOS:
                break
            antes = int(registro["CROVALOBJ"] or 0)
            despues = max(0, antes - sobra)
            if antes == despues:
                continue
            registro["CROVALOBJ"] = despues
            registro["CRDOBSERV"] = _cambiar_valor_final(
                str(registro["CRDOBSERV"] or ""), antes, despues
            )
            sobra -= antes - despues
            ajustados += 1
            salida.revisiones.append(
                Revision(
                    factura=crncxc,
                    motivo=REV_AJUSTE_REPORTE,
                    cod_elemento=str(registro["SLNSERPRO"] or ""),
                    codigo_glosa=str(registro["CRNCONOBJ"] or ""),
                    valor_glosado=float(despues),
                    detalle=f"{a_texto(antes)} → {a_texto(despues)}, para que la factura sume "
                    f"los {a_texto(reclamacion.glosado)} que reporta glosados el ADRES.",
                )
            )
        if ajustados:
            salida.ajustadas += 1


def _cambiar_valor_final(texto: str, antes: int, despues: int) -> str:
    """Cambia el ``$<valor>`` del final del CRDOBSERV cuando se ajusta el valor."""
    cola = f"${antes}"
    if texto.endswith(cola):
        return texto[: -len(cola)] + f"${despues}"
    return texto


def _revision(fila: FilaAdres, crncxc: str, motivo: str, detalle: str) -> Revision:
    return Revision(
        factura=crncxc,
        motivo=motivo,
        cod_elemento=fila.cod_elemento,
        descripcion=fila.descripcion,
        codigo_glosa=fila.codigo_glosa,
        valor_glosado=fila.valor_glosado,
        detalle=detalle,
    )


# ─── Escritura ───────────────────────────────────────────────────────────────


def _cabecera(ws, nombres) -> None:
    from openpyxl.styles import Font, PatternFill

    relleno = PatternFill("solid", fgColor="1F4E78")
    letra = Font(bold=True, color="FFFFFF")
    for col, nombre in enumerate(nombres, start=1):
        celda = ws.cell(row=1, column=col, value=nombre)
        celda.fill = relleno
        celda.font = letra
    ws.freeze_panes = "A2"


def escribir_objeciones(registros: list[dict], salida: Path) -> None:
    """Un archivo OBJECIONES con las 16 columnas, como lo espera DGH."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "OBJECIONES"
    _cabecera(ws, COLUMNAS_OBJECIONES)
    for n, registro in enumerate(registros, start=2):
        for col, nombre in enumerate(COLUMNAS_OBJECIONES, start=1):
            celda = ws.cell(row=n, column=col, value=registro.get(nombre))
            celda.number_format = FORMATOS_OBJECIONES[nombre]
    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(salida))


def lotes(registros: list[dict], max_facturas: int = MAX_FACTURAS_POR_LOTE) -> list[list[dict]]:
    """Parte los registros en grupos de a lo sumo `max_facturas` facturas.

    DGH no recibe más de 300 facturas por archivo, y una factura nunca se
    puede partir entre dos lotes.
    """
    if max_facturas <= 0:
        return [registros] if registros else []
    por_factura: dict[str, list[dict]] = defaultdict(list)
    for registro in registros:
        por_factura[str(registro["CRNCXC"])].append(registro)
    grupos: list[list[dict]] = []
    actual: list[dict] = []
    cuenta = 0
    for factura in sorted(por_factura):
        if cuenta >= max_facturas:
            grupos.append(actual)
            actual, cuenta = [], 0
        actual.extend(por_factura[factura])
        cuenta += 1
    if actual:
        grupos.append(actual)
    return grupos


COLUMNAS_REVISAR: tuple[str, ...] = (
    "FACTURA",
    "MOTIVO",
    "COD ELEMENTO",
    "DESCRIPCION",
    "CODIGO GLOSA",
    "VALOR GLOSADO",
    "DETALLE",
)


def escribir_revisar(
    conversion: Conversion, filas: list[FilaAdres], salida: Path, mapa_codigos: dict[str, str]
) -> None:
    """El archivo de control: resumen, renglones a revisar y códigos de glosa."""
    from openpyxl import Workbook

    wb = Workbook()

    ws = wb.active
    ws.title = "RESUMEN"
    _cabecera(ws, ("CONCEPTO", "CANTIDAD", "VALOR"))
    facturas = {r["CRNCXC"] for r in conversion.registros}
    total = sum(int(r["CROVALOBJ"] or 0) for r in conversion.registros)
    resumen: list[tuple[str, object, object]] = [
        ("Glosas leídas del ADRES", len(filas), round(sum(f.valor_glosado for f in filas))),
        ("Objeciones escritas", len(conversion.registros), total),
        ("Facturas", len(facturas), None),
    ]
    if conversion.recorte:
        resumen.append(
            ("Recortado por el guardián de valores de DGH", None, round(conversion.recorte))
        )
    for metodo, cuantas in sorted(conversion.metodos.items(), key=lambda kv: -kv[1]):
        etiqueta = metodo or "SIN HOMOLOGAR (queda en REVISAR)"
        resumen.append((f"Código de servicio por: {etiqueta}", cuantas, None))
    motivos: dict[str, list[Revision]] = defaultdict(list)
    for revision in conversion.revisiones:
        motivos[revision.motivo].append(revision)
    for motivo, items in sorted(motivos.items(), key=lambda kv: -len(kv[1])):
        resumen.append(
            (f"A revisar: {motivo}", len(items), round(sum(i.valor_glosado for i in items)))
        )
    for n, (concepto, cuantas, valor) in enumerate(resumen, start=2):
        ws.cell(row=n, column=1, value=concepto)
        ws.cell(row=n, column=2, value=cuantas)
        celda = ws.cell(row=n, column=3, value=valor)
        celda.number_format = FORMATOS_OBJECIONES["CROVALOBJ"]
    ws.column_dimensions["A"].width = 52

    ws = wb.create_sheet("REVISAR")
    _cabecera(ws, COLUMNAS_REVISAR)
    for n, revision in enumerate(conversion.revisiones, start=2):
        for col, valor in enumerate(
            (
                revision.factura,
                revision.motivo,
                revision.cod_elemento,
                revision.descripcion,
                revision.codigo_glosa,
                round(revision.valor_glosado),
                revision.detalle,
            ),
            start=1,
        ):
            ws.cell(row=n, column=col, value=valor)
    for letra, ancho in (("A", 16), ("B", 40), ("C", 20), ("D", 45), ("E", 14), ("G", 60)):
        ws.column_dimensions[letra].width = ancho

    ws = wb.create_sheet("CODIGOS")
    _cabecera(
        ws,
        (
            "CODIGO ADRES",
            "CLASIFICACION DEL AUDITOR",
            "GRUPO DEL MANUAL UNICO",
            "CODIGO DGH (--mapa-codigos)",
            "RENGLONES",
            "VALOR GLOSADO",
            "CAUSAL",
        ),
    )
    cuenta: dict[tuple[str, str], list[FilaAdres]] = defaultdict(list)
    for fila in filas:
        cuenta[(fila.codigo_glosa, fila.clasificacion)].append(fila)
    for n, ((codigo, clasificacion), items) in enumerate(
        sorted(cuenta.items(), key=lambda kv: -len(kv[1])), start=2
    ):
        ws.cell(row=n, column=1, value=codigo or "(sin código)")
        ws.cell(row=n, column=2, value=clasificacion)
        ws.cell(row=n, column=3, value=grupo_dgh(clasificacion))
        ws.cell(row=n, column=4, value=mapa_codigos.get(codigo, ""))
        ws.cell(row=n, column=5, value=len(items))
        celda = ws.cell(row=n, column=6, value=round(sum(i.valor_glosado for i in items)))
        celda.number_format = FORMATOS_OBJECIONES["CROVALOBJ"]
        ws.cell(row=n, column=7, value=next((i.texto_glosa for i in items if i.texto_glosa), ""))
    for letra, ancho in (("A", 14), ("B", 34), ("C", 22), ("D", 26), ("G", 70)):
        ws.column_dimensions[letra].width = ancho

    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(salida))


# ─── CLI ─────────────────────────────────────────────────────────────────────

PREFIJO_DEFAULT = "OBJECIONES_ADRES"


def _parse_fecha(texto: str | None) -> _dt.datetime:
    if not texto:
        hoy = _dt.date.today()
        return _dt.datetime(hoy.year, hoy.month, hoy.day)
    try:
        return _dt.datetime.strptime(texto.strip(), "%Y-%m-%d")
    except ValueError:
        logger.error("--fecha inválida: %r (use YYYY-MM-DD, p.ej. 2026-08-21)", texto)
        raise SystemExit(2) from None


def _cargar_mapa(ruta: Path | None) -> dict[str, str]:
    if ruta is None:
        return {}
    try:
        datos = json.loads(ruta.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        logger.error("No pude leer el mapa de códigos %s: %s", ruta, e)
        raise SystemExit(2) from None
    if not isinstance(datos, dict):
        logger.error('El mapa de códigos debe ser un objeto JSON {"3106": "SO3401", ...}')
        raise SystemExit(2)
    return {str(k).strip(): str(v).strip() for k, v in datos.items()}


def construir_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Pasa las glosas del ADRES al formato OBJECIONES de DGH.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("--adres", type=Path, required=True, help="Excel de glosas del ADRES.")
    p.add_argument("--dgh", type=Path, help="DGReport con los servicios facturados (DGDATATABLE).")
    p.add_argument("--homologador", type=Path, help="Homologador Gold Standard CUPS ↔ SOAT.")
    p.add_argument(
        "--reporte-reclamaciones",
        type=Path,
        help="ReporteReclamPAQUETE del ADRES. Con él, el total de cada factura queda "
        "cuadrado contra el Valor Glosado que reporta el ADRES.",
    )
    p.add_argument(
        "--completar-servicios",
        action="store_true",
        help="Que ningún renglón quede sin código de servicio: el que no cruce se lleva "
        "el candidato más parecido o el servicio de más peso de la factura. Cada uno "
        "queda marcado en REVISAR.",
    )
    p.add_argument("--salida", type=Path, required=True, help="Carpeta destino.")
    p.add_argument(
        "--paquete", default="", help="Deja solo las glosas de ese paquete (p.ej. 31068)."
    )
    p.add_argument(
        "--hoja-glosas", default="", help="Nombre de la hoja de glosas (si no, se detecta)."
    )
    p.add_argument("--hoja-reporte", default="", help="Hoja con la descripción de cada servicio.")
    p.add_argument("--fecha", default=None, help="Fecha de la objeción (YYYY-MM-DD). Default: hoy.")
    p.add_argument(
        "--prefijo",
        default=PREFIJO_DEFAULT,
        help=f"Prefijo de los archivos. Default: {PREFIJO_DEFAULT}.",
    )
    p.add_argument(
        "--max-facturas",
        type=int,
        default=MAX_FACTURAS_POR_LOTE,
        help=f"Facturas por archivo. Default: {MAX_FACTURAS_POR_LOTE} (el tope de DGH).",
    )
    p.add_argument(
        "--mapa-codigos", type=Path, help='JSON {"3106": "SO3401", ...} para el código de DGH.'
    )
    p.add_argument(
        "--excluir-glosa-total",
        action="store_true",
        help="Deja fuera los renglones sin código de glosa (glosa total por FURIPS).",
    )
    p.add_argument("--log", type=Path, help="Guarda además el log en un archivo.")
    return p


def main(argv: list[str] | None = None) -> int:
    args = construir_parser().parse_args(argv)
    handlers: list[logging.Handler] = [logging.StreamHandler(sys.stdout)]
    if args.log is not None:
        args.log.parent.mkdir(parents=True, exist_ok=True)
        handlers.append(logging.FileHandler(args.log, encoding="utf-8"))
    logging.basicConfig(level=logging.INFO, format="%(message)s", handlers=handlers)

    if not args.adres.is_file():
        logger.error("No existe el Excel del ADRES: %s", args.adres)
        return 1

    logger.info("Leyendo las glosas del ADRES: %s", args.adres.name)
    filas = leer_adres(args.adres, args.paquete, args.hoja_glosas, args.hoja_reporte)
    if not filas:
        logger.error("No encontré ninguna glosa en %s.", args.adres.name)
        return 1
    logger.info("  %d glosas · %d facturas", len(filas), len({f.factura for f in filas}))

    dgh: dict[str, list[LineaDgh]] = {}
    if args.dgh:
        if not args.dgh.is_file():
            logger.error("No existe el reporte de DGH: %s", args.dgh)
            return 1
        logger.info("Leyendo los servicios de DGH: %s", args.dgh.name)
        dgh = leer_dgh(args.dgh)
        logger.info("  %d facturas en DGH", len(dgh))
    else:
        logger.warning("Sin --dgh no se puede poner el código de servicio (SLNSERPRO).")

    soat_a_cups: dict[str, set[str]] = {}
    if args.homologador:
        if not args.homologador.is_file():
            logger.error("No existe el homologador: %s", args.homologador)
            return 1
        logger.info("Leyendo el homologador: %s", args.homologador.name)
        soat_a_cups = leer_homologador(args.homologador)
        logger.info("  %d códigos SOAT con CUPS", len(soat_a_cups))

    reclamaciones: dict[str, Reclamacion] = {}
    if args.reporte_reclamaciones:
        if not args.reporte_reclamaciones.is_file():
            logger.error("No existe el reporte de reclamaciones: %s", args.reporte_reclamaciones)
            return 1
        logger.info("Leyendo el reporte del ADRES: %s", args.reporte_reclamaciones.name)
        reclamaciones = leer_reporte_reclamaciones(args.reporte_reclamaciones, args.paquete)
        con_glosa = [r for r in reclamaciones.values() if r.glosado > 0]
        logger.info(
            "  %d reclamaciones · %d con glosa · glosado reportado %s",
            len(reclamaciones),
            len(con_glosa),
            a_texto(sum(r.glosado for r in con_glosa)),
        )
        extemporaneas = {r.extemporaneidad for r in reclamaciones.values() if r.extemporaneidad}
        if extemporaneas:
            logger.info("  Tipo de extemporaneidad: %s", ", ".join(sorted(extemporaneas)))

    mapa = _cargar_mapa(args.mapa_codigos)
    conversion = construir_registros(
        filas,
        dgh,
        soat_a_cups,
        fecha=_parse_fecha(args.fecha),
        mapa_codigos=mapa,
        incluir_glosa_total=not args.excluir_glosa_total,
        reclamaciones=reclamaciones,
        completar_servicios=args.completar_servicios,
    )

    grupos = lotes(conversion.registros, args.max_facturas)
    for n, grupo in enumerate(grupos, start=1):
        destino = args.salida / f"{args.prefijo}_LOTE_{n:02d}.xlsx"
        escribir_objeciones(grupo, destino)
        valor = sum(int(r["CROVALOBJ"] or 0) for r in grupo)
        logger.info(
            "  %s → %d objeciones · %d facturas · %s",
            destino.name,
            len(grupo),
            len({r["CRNCXC"] for r in grupo}),
            a_texto(valor),
        )
    control = args.salida / f"REVISAR_{args.prefijo}.xlsx"
    escribir_revisar(conversion, filas, control, mapa)

    if reclamaciones:
        objetado: dict[str, float] = defaultdict(float)
        for registro in conversion.registros:
            objetado[str(registro["CRNCXC"])] += float(registro["CROVALOBJ"] or 0)
        descuadran = [
            (f, reclamaciones[f].glosado, v)
            for f, v in objetado.items()
            if f in reclamaciones and abs(v - reclamaciones[f].glosado) > TOLERANCIA_PESOS
        ]
        logger.info("\nCuadre contra el reporte del ADRES:")
        logger.info(
            "   glosado que reporta el ADRES : %s",
            a_texto(sum(r.glosado for r in reclamaciones.values())),
        )
        logger.info("   objetado en estos archivos   : %s", a_texto(sum(objetado.values())))
        logger.info(
            "   facturas que cuadran         : %d de %d",
            len(objetado) - len(descuadran),
            len(objetado),
        )
        logger.info("   renglones repetidos quitados : %d", conversion.quitadas)
        logger.info("   facturas ajustadas al reporte: %d", conversion.ajustadas)
        for f, esperado, v in sorted(descuadran, key=lambda x: -abs(x[2] - x[1]))[:10]:
            logger.warning("   [!] %s: reporte %s, objetado %s", f, a_texto(esperado), a_texto(v))

    vacios = sum(1 for r in conversion.registros if not r["SLNSERPRO"])
    logger.info("\nRenglones sin código de servicio: %d", vacios)
    if conversion.asignados:
        logger.info(
            "   %d llevan un código ASIGNADO (no salió del cruce): están en REVISAR.",
            conversion.asignados,
        )

    homologados = sum(v for k, v in conversion.metodos.items() if k)
    con_servicio = sum(1 for f in filas if f.cod_elemento)
    logger.info("\nCódigo de servicio de DGH (SLNSERPRO):")
    for metodo, cuantas in sorted(conversion.metodos.items(), key=lambda kv: -kv[1]):
        logger.info("   %-28s %6d", metodo or "SIN HOMOLOGAR", cuantas)
    if con_servicio:
        logger.info(
            "   homologados: %d de %d renglones con servicio (%.1f%%)",
            homologados,
            con_servicio,
            homologados / con_servicio * 100,
        )
    logger.info("\nArchivo de control: %s", control)
    logger.info(
        "REVISE la hoja CODIGOS antes de cargar: el ADRES usa códigos numéricos y DGH los\n"
        "de seis caracteres del Manual Único. Y haga el piloto de UNA factura primero."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
