"""filtrar_base_dgh.py — Recorta la base DGH a las facturas de un lote.

La base "SERVICIOS FACTURADOS COOSALUD DGH.xlsx" pesa 70 MB y no se puede
mover fácilmente (correo, chat, etc.). Este script deja solo las filas de las
facturas que se están trabajando, y el archivo resultante pesa unos pocos KB.

Se pueden pasar VARIAS bases (tandas). Un Excel no puede tener más de
1.048.576 filas, así que cuando la base completa no cabe hay que bajarla de
DGH por partes (por rango de fechas) y pasarlas todas aquí: el script las lee
en orden y junta lo que encuentra en cada una.

Las facturas se pueden indicar de tres formas (elige una):
  --carpeta   la carpeta "CARGUE MASIVO COOSALUD" del lote (lee las cabeceras
              HUS*.xlsx de la subcarpeta FACTURAS). Es lo más cómodo.
  --lista     un TXT con una factura por línea (HUS522791 / 522791 / HUS0000522791).
  --facturas  la lista separada por comas.

USO:
    py filtrar_base_dgh.py "D:\\...\\SERVICIOS FACTURADOS COOSALUD DGH.xlsx" ^
        --carpeta "D:\\USUARIO CARTERA\\Desktop\\CARGUE MASIVO COOSALUD"

    py filtrar_base_dgh.py "D:\\...\\BASE ENERO-JUNIO.xlsx" "D:\\...\\BASE JULIO-AGOSTO.xlsx" ^
        --lista "D:\\...\\FACTURAS LOTE.txt"

Genera "BASE_DGH_FILTRADA.xlsx" (o lo que diga --salida).

Al final avisa si una base viene recortada (llegó al tope de filas de Excel) y
en qué rango de facturas va cada base, que es la causa típica de que "no
aparezcan" facturas nuevas: la base es vieja y no las trae todavía.
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import Counter
from collections.abc import Iterator
from contextlib import contextmanager
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.stderr.write("Falta openpyxl. Corre: py -m pip install openpyxl\n")
    sys.exit(2)

# DGH exporta la base en .xlsb (Excel binario), que openpyxl no sabe leer.
# pyxlsb es opcional: solo hace falta si la base viene en ese formato.
try:
    from pyxlsb import open_workbook as _abrir_xlsb
except ImportError:
    _abrir_xlsb = None

COLS_FACTURA = ("FACTURA", "NUMERO_FACTURA", "NUMERO FACTURA", "CUENTA")
# Cuántas filas mirar buscando el encabezado (algunos exports traen título).
FILAS_BUSCA_CABECERA = 10
# Columnas del detalle de servicios sin las cuales el consolidador NO puede
# armar el OBJECIONES: sirven para avisar si nos pasaron otro reporte.
COLS_SERVICIOS_MINIMAS = ("SLNSERPRO_SERVICIO", "VR_SERVICIO", "SALDO_FACT")
# Tope duro de filas de una hoja de Excel. Si una base se le acerca, es que
# el export de DGH salió recortado y faltan facturas.
TOPE_EXCEL = 1_048_575
CERCA_DEL_TOPE = 1_000_000


@contextmanager
def filas_de(base: Path) -> Iterator[Iterator[list]]:
    """Filas de la primera hoja, venga el archivo en .xlsx o en .xlsb.

    DGH exporta la base de servicios facturados en .xlsb (Excel binario) y
    openpyxl no lo lee; para ese formato se usa pyxlsb, que solo hace falta
    instalar si de verdad llega un .xlsb.
    """
    if base.suffix.lower() == ".xlsb":
        if _abrir_xlsb is None:
            raise RuntimeError(
                "Esta base viene en formato .xlsb (Excel binario) y falta la "
                "librería para leerlo. Corre:  py -m pip install pyxlsb\n"
                "  (o guárdala como .xlsx desde Excel: Archivo > Guardar como)"
            )
        with _abrir_xlsb(str(base)) as wb:
            with wb.get_sheet(wb.sheets[0]) as sh:
                yield ([c.v for c in fila] for fila in sh.rows())
        return
    wb = openpyxl.load_workbook(base, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        yield (list(fila) for fila in ws.iter_rows(values_only=True))
    finally:
        wb.close()


def nfact(v: object) -> str | None:
    """HUS0000522511 / 0000522511 / 522511 / 522511-1 -> 522511."""
    d = re.sub(r"\D", "", str(v or "")).lstrip("0")
    return d or None


def norm(h: object) -> str:
    return re.sub(r"\s+", " ", str(h or "")).strip().upper()


def facturas_de_carpeta(carpeta: Path) -> set[str]:
    """Facturas del lote: nombres HUS######.xlsx de la subcarpeta FACTURAS."""
    facturas: set[str] = set()
    raiz = carpeta / "FACTURAS" if (carpeta / "FACTURAS").is_dir() else carpeta
    for arch in raiz.rglob("*.xlsx"):
        nombre = arch.name.upper()
        if nombre.startswith("~$") or nombre.startswith(("DETALLE", "GLOSAS", "CONSOLIDADO")):
            continue
        f = nfact(arch.stem)
        if f:
            facturas.add(f)
    return facturas


def leer_base(base: Path, objetivo: set[str]) -> dict:
    """Lee una base DGH y devuelve las filas de las facturas buscadas.

    Además informa cuántas filas trae y entre qué facturas va, para poder
    detectar bases viejas o recortadas.
    """
    print(f"\nLeyendo base: {base.name} (pesa bastante, tarda un poco) ...")
    with filas_de(base) as it:
        # El encabezado no siempre está en la primera fila: hay exports que
        # arrancan con el título del reporte y un par de filas en blanco. Se
        # busca la primera fila que traiga la columna de factura.
        headers: list = []
        idx = None
        primeras: list[list] = []
        for _ in range(FILAS_BUSCA_CABECERA):
            fila = list(next(it, []) or [])
            if not fila:
                continue
            primeras.append(fila)
            pos = next((i for i, h in enumerate(fila) if norm(h) in COLS_FACTURA), None)
            if pos is not None:
                headers, idx = fila, pos
                break
        if idx is None:
            vistas = " | ".join(str([str(c)[:20] for c in f[:8]]) for f in primeras[:4])
            return {
                "error": (
                    f"no hallé la columna FACTURA en las primeras "
                    f"{FILAS_BUSCA_CABECERA} filas. Lo que se leyó: {vistas}"
                )
            }
        if len(primeras) > 1:
            print(f"  (el encabezado estaba en la fila {len(primeras)}, no en la primera)")
        print(f"  Columna de factura: '{headers[idx]}' (posición {idx + 1})")

        faltan_cols = [c for c in COLS_SERVICIOS_MINIMAS if c not in {norm(h) for h in headers}]
        if faltan_cols:
            print(
                "  *** OJO: a este archivo le faltan columnas del detalle de "
                f"servicios ({', '.join(faltan_cols)}). Parece otro reporte, no la "
                "base SERVICIOS FACTURADOS: sirve para consultar, pero el "
                "consolidador no puede armar el OBJECIONES con él."
            )

        filas: list[list] = []
        encontradas: set[str] = set()
        total = 0
        menor = mayor = None
        for row in it:
            total += 1
            f = nfact(row[idx]) if idx < len(row) else None
            if not f:
                continue
            n = int(f)
            if menor is None or n < menor:
                menor = n
            if mayor is None or n > mayor:
                mayor = n
            if f in objetivo:
                filas.append(list(row))
                encontradas.add(f)
        return {
            "headers": headers,
            "filas": filas,
            "encontradas": encontradas,
            "total": total,
            "menor": menor,
            "mayor": mayor,
        }


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Recorta la base DGH a las facturas de un lote.")
    p.add_argument(
        "bases",
        nargs="+",
        help="Ruta(s) de SERVICIOS FACTURADOS COOSALUD DGH.xlsx (varias si bajaste por tandas).",
    )
    p.add_argument("--carpeta", default=None, help='Carpeta "CARGUE MASIVO COOSALUD" del lote.')
    p.add_argument("--lista", default=None, help="TXT con una factura por línea.")
    p.add_argument("--facturas", default=None, help="Facturas separadas por coma.")
    p.add_argument("--salida", default=None, help="Xlsx de salida. Def: BASE_DGH_FILTRADA.xlsx")
    args = p.parse_args(argv)

    rutas = [Path(b.strip().strip('"')) for b in args.bases]
    faltantes = [r for r in rutas if not r.is_file()]
    if faltantes:
        for r in faltantes:
            sys.stderr.write(f"No existe la base: {r}\n")
        # Casi siempre es que el archivo todavía no se ha bajado de DGH, o que
        # el nombre no es el que se escribió. Mostramos qué Excel SÍ hay en esa
        # carpeta para no tener que ir a mirar.
        for carpeta in dict.fromkeys(r.parent for r in faltantes):
            sys.stderr.write(f"\nExcel que SÍ hay en {carpeta}:\n")
            if not carpeta.is_dir():
                sys.stderr.write("  (esa carpeta tampoco existe)\n")
                continue
            hallados = sorted(
                (x for x in carpeta.glob("*.xls*") if not x.name.startswith("~$")),
                key=lambda x: x.stat().st_mtime,
                reverse=True,
            )[:15]
            if not hallados:
                sys.stderr.write("  (ningún Excel en esa carpeta)\n")
            for x in hallados:
                mb = x.stat().st_size / (1024 * 1024)
                sys.stderr.write(f"  {x.name}   ({mb:,.1f} MB)\n")
        sys.stderr.write(
            "\nCopia el nombre EXACTO de la base que vas a usar. Si la base nueva no "
            "aparece en la lista, todavía no se ha bajado de DGH.\n"
        )
        return 2

    objetivo: set[str] = set()
    if args.carpeta:
        carpeta = Path(args.carpeta.strip().strip('"'))
        if not carpeta.is_dir():
            sys.stderr.write(f"No existe la carpeta: {carpeta}\n")
            return 2
        objetivo = facturas_de_carpeta(carpeta)
        print(f"Facturas leídas de la carpeta: {len(objetivo)}")
    elif args.lista:
        lista = Path(args.lista.strip().strip('"'))
        if not lista.is_file():
            sys.stderr.write(f"No existe la lista: {lista}\n")
            return 2
        objetivo = {f for f in (nfact(x) for x in lista.read_text(errors="replace").split()) if f}
        print(f"Facturas leídas del TXT: {len(objetivo)}")
    elif args.facturas:
        objetivo = {f for f in (nfact(x) for x in args.facturas.split(",")) if f}
        print(f"Facturas indicadas: {len(objetivo)}")
    else:
        sys.stderr.write("Indica --carpeta, --lista o --facturas.\n")
        return 2
    if not objetivo:
        sys.stderr.write("No encontré ninguna factura que buscar.\n")
        return 2

    headers: list = []
    # Cuántas copias de cada fila llevamos guardadas. OJO: una base DGH SÍ trae
    # filas idénticas de verdad (el mismo medicamento dispensado varias veces en
    # la misma factura), y cada una es un servicio distinto que se puede objetar.
    # Por eso NO se quitan las repetidas dentro de un mismo archivo: solo se
    # descarta lo que ya venía en una tanda anterior, comparando cuántas veces
    # aparece en cada una. (Quitarlas todas dejó 249 servicios sin cruzar en el
    # lote de 1.573, contra 8 conservándolas.)
    llevadas: Counter[tuple] = Counter()
    todas_filas: list[list] = []
    encontradas: set[str] = set()
    resumen: list[tuple[str, int, int | None, int | None, int]] = []

    for ruta in rutas:
        info = leer_base(ruta, objetivo)
        if info.get("error"):
            sys.stderr.write(f"{ruta.name}: {info['error']}\n")
            return 2
        if not headers:
            headers = info["headers"]
        nuevas = 0
        de_esta = Counter(tuple(f) for f in info["filas"])
        for fila in info["filas"]:
            clave = tuple(fila)
            if llevadas[clave] >= de_esta[clave]:
                continue  # esta tanda no aporta más copias que las que ya hay
            llevadas[clave] += 1
            todas_filas.append(fila)
            nuevas += 1
        encontradas |= info["encontradas"]
        resumen.append((ruta.name, info["total"], info["menor"], info["mayor"], nuevas))
        print(
            f"  Filas leídas: {info['total']:,} | facturas del lote halladas: "
            f"{len(info['encontradas'])} | filas nuevas guardadas: {nuevas:,}"
        )

    out = openpyxl.Workbook()
    wo = out.active
    wo.title = "BASE"
    wo.append(list(headers))
    for fila in todas_filas:
        wo.append(fila)

    salida = (
        Path(args.salida.strip().strip('"'))
        if args.salida
        else rutas[0].parent / "BASE_DGH_FILTRADA.xlsx"
    )
    try:
        out.save(salida)
    except Exception:
        salida = Path.cwd() / "BASE_DGH_FILTRADA.xlsx"
        out.save(salida)

    print("\n===== RESUMEN =====")
    for nombre, total, menor, mayor, nuevas in resumen:
        rango = f"HUS{menor} a HUS{mayor}" if menor and mayor else "sin facturas legibles"
        print(f"  {nombre}")
        print(f"     filas: {total:,} | va de {rango} | aportó {nuevas:,} filas")
        if total >= CERCA_DEL_TOPE:
            print(
                f"     *** OJO: esta base trae {total:,} filas y el tope de Excel es "
                f"{TOPE_EXCEL:,}. Es casi seguro que el export salió RECORTADO y le "
                "faltan facturas. Bájala de DGH por partes (por rango de fechas)."
            )
    print(f"\n  Filas guardadas:      {len(todas_filas):,}")
    print(f"  Facturas encontradas: {len(encontradas)} de {len(objetivo)}")

    faltan = objetivo - encontradas
    if faltan:
        print(f"\n  NO están en la(s) base(s) ({len(faltan)}):")
        ordenadas = sorted(faltan, key=int)
        print("  " + ", ".join("HUS" + f for f in ordenadas[:40]))
        if len(ordenadas) > 40:
            print(f"  ... y {len(ordenadas) - 40} más.")
        mayor_base = max((m for _, _, _, m, _ in resumen if m), default=None)
        menor_falta = int(ordenadas[0])
        if mayor_base and menor_falta > mayor_base:
            print(
                f"\n  *** La base más nueva llega hasta HUS{mayor_base} y la factura que "
                f"falta más antigua es HUS{menor_falta}: la base está DESACTUALIZADA. "
                "Baja de DGH un export nuevo que cubra esas fechas."
            )
    print(f"\nListo. Sube este archivo: {salida.resolve()}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
