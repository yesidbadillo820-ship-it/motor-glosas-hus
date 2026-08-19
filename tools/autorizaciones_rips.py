"""
autorizaciones_rips — Encuentra los números de AUTORIZACIÓN dentro de RIPS
en formato JSON (Res. 2275/2023) y saca un informe Excel del hallazgo.

Qué revisa: en cada RIPS, TODOS los grupos de servicios (consultas,
procedimientos, hospitalización, urgencias, medicamentos, otros servicios…)
y el campo `numAutorizacion` de cada registro. Informa:
  · el número (o números) de autorización encontrados por factura,
  · cada servicio SIN autorización — sea null, vacío o el texto "null",
  · cada servicio con un número DISTINTO al dominante de la factura
    ("otro número"), que suele ser el error que hace rechazar el RIPS.

Entradas: archivos .json sueltos, CARPETAS (se recorren con TODO y sus
subcarpetas) y/o .zip que los contengan — se puede mezclar.

Uso por consola (no necesita ventanas):
  python autorizaciones_rips.py <ruta1> [ruta2 ...] [-o INFORME.xlsx]
"""

import glob
import json
import os
import sys
import tempfile
import zipfile
from collections import Counter

GRUPOS_CONOCIDOS = [
    "consultas",
    "procedimientos",
    "urgencias",
    "hospitalizacion",
    "recienNacidos",
    "medicamentos",
    "otrosServicios",
]

CAMPOS_CODIGO = [
    "codConsulta",
    "codProcedimiento",
    "codTecnologiaSalud",
    "codServicio",
    "codDiagnosticoPrincipal",
]
CAMPOS_FECHA = ["fechaInicioAtencion", "fechaDispensAdmon", "fechaAtencion", "fechaIngreso"]


def _resolver_entradas(entradas):
    """Acepta .json sueltos, carpetas (recursivas) y .zip; devuelve rutas .json."""
    rutas = []
    for e in entradas:
        if os.path.isdir(e):
            for raiz, _dirs, archivos in os.walk(e):
                for a in sorted(archivos):
                    if a.lower().endswith(".json"):
                        rutas.append(os.path.join(raiz, a))
        elif e.lower().endswith(".zip"):
            carpeta_tmp = tempfile.mkdtemp(prefix="ripsaut_")
            with zipfile.ZipFile(e) as z:
                for nombre in z.namelist():
                    if nombre.lower().endswith(".json"):
                        z.extract(nombre, carpeta_tmp)
                        rutas.append(os.path.join(carpeta_tmp, nombre))
        elif e.lower().endswith(".json"):
            rutas.append(e)
        else:
            rutas.extend(sorted(glob.glob(e)))  # comodines tipo *.json
    return rutas


def _clasificar(valor):
    """Devuelve (estado, texto_mostrable). Estados: OK / NULL / VACIO / TEXTO_NULL."""
    if valor is None:
        return "NULL", "(null)"
    texto = str(valor).strip()
    if texto == "":
        return "VACIO", "(vacío)"
    if texto.lower() in ("null", "none"):
        return "TEXTO_NULL", "(texto '%s')" % texto
    return "OK", texto


def _primero(registro, campos):
    for c in campos:
        v = registro.get(c)
        if v not in (None, ""):
            return str(v)
    return ""


def analizar_archivo(ruta, log=print):
    """Lee un RIPS JSON. Devuelve (resumen_dict, [fila_detalle, ...])."""
    try:
        with open(ruta, encoding="utf-8-sig") as f:
            datos = json.load(f)
    except Exception:
        with open(ruta, encoding="latin-1") as f:
            datos = json.load(f)
    if not isinstance(datos, dict) or "usuarios" not in datos:
        raise ValueError("no tiene la estructura de un RIPS JSON (falta 'usuarios')")

    factura = str(datos.get("numFactura") or "").strip() or os.path.basename(ruta)
    nit = str(datos.get("numDocumentoIdObligado") or "").strip()
    filas = []

    for usuario in datos.get("usuarios") or []:
        doc = "%s %s" % (
            usuario.get("tipoDocumentoIdentificacion") or "",
            usuario.get("numDocumentoIdentificacion") or "",
        )
        servicios = usuario.get("servicios") or {}
        grupos = list(servicios.keys()) or []
        for grupo in grupos:
            registros = servicios.get(grupo) or []
            if not isinstance(registros, list):
                continue
            for n, registro in enumerate(registros, 1):
                if not isinstance(registro, dict):
                    continue
                crudo = registro.get("numAutorizacion")
                estado, mostrable = _clasificar(crudo)
                filas.append(
                    dict(
                        archivo=os.path.basename(ruta),
                        factura=factura,
                        usuario=doc.strip(),
                        grupo=grupo,
                        n=n,
                        codigo=_primero(registro, CAMPOS_CODIGO),
                        fecha=_primero(registro, CAMPOS_FECHA)[:16],
                        valor=registro.get("vrServicio"),
                        autorizacion=mostrable,
                        estado=estado,
                    )
                )

    numeros = Counter(f["autorizacion"] for f in filas if f["estado"] == "OK")
    dominante = numeros.most_common(1)[0][0] if numeros else None
    con_num = sum(numeros.values())
    sin_num = len(filas) - con_num
    otros = 0
    for f in filas:
        if f["estado"] == "OK" and f["autorizacion"] != dominante:
            f["estado"] = "OTRO_NUMERO"
            otros += 1
        f["dominante"] = dominante or "-"

    if not filas:
        veredicto = "RIPS SIN SERVICIOS"
    elif not numeros:
        veredicto = "TODA LA FACTURA SIN AUTORIZACIÓN"
    elif len(numeros) > 1:
        veredicto = "¡VARIOS NÚMEROS!: " + " · ".join(
            "%s (%d)" % (num, cant) for num, cant in numeros.most_common()
        )
    elif sin_num:
        veredicto = "AUTORIZACIÓN %s — pero %d servicio(s) SIN autorización" % (dominante, sin_num)
    else:
        veredicto = "AUTORIZACIÓN ÚNICA %s en los %d servicios" % (dominante, len(filas))

    resumen = dict(
        archivo=os.path.basename(ruta),
        ruta=os.path.abspath(ruta),
        factura=factura,
        nit=nit,
        servicios=len(filas),
        con_autorizacion=con_num,
        sin_autorizacion=sin_num,
        otros_numeros=otros,
        numeros_distintos=len(numeros),
        veredicto=veredicto,
    )
    return resumen, filas


def procesar(entradas, salida, log=print):
    rutas = _resolver_entradas(entradas)
    if not rutas:
        raise ValueError("No se encontraron archivos .json en la entrada.")

    resumenes, detalle, ilegibles = [], [], []
    for ruta in rutas:
        log("Leyendo %s…" % os.path.basename(ruta))
        try:
            resumen, filas = analizar_archivo(ruta, log=log)
        except Exception as e:
            log("  [!] %s: %s" % (os.path.basename(ruta), e))
            ilegibles.append((os.path.basename(ruta), str(e)))
            continue
        log("  %s" % resumen["veredicto"])
        resumenes.append(resumen)
        detalle.extend(filas)

    if not resumenes and not ilegibles:
        raise ValueError("Ningún archivo se pudo leer.")

    _escribir_informe(resumenes, detalle, ilegibles, salida)
    problemas = sum(
        1
        for r in resumenes
        if r["sin_autorizacion"] or r["otros_numeros"] or not r["con_autorizacion"]
    )
    log("Informe: %d RIPS revisados, %d con hallazgos → %s" % (len(resumenes), problemas, salida))
    return salida


ESTADO_TEXTO = {
    "OK": "OK",
    "NULL": "SIN AUTORIZACIÓN (null)",
    "VACIO": "SIN AUTORIZACIÓN (vacío)",
    "TEXTO_NULL": "SIN AUTORIZACIÓN (texto null)",
    "OTRO_NUMERO": "¡OTRO NÚMERO! (distinto al dominante)",
}


def _escribir_informe(resumenes, detalle, ilegibles, salida):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    FILL_HDR = PatternFill("solid", fgColor="1F4E79")
    FONT_HDR = Font(bold=True, color="FFFFFF")
    FINO = Side(style="thin", color="B7C6D6")
    BORDE = Border(left=FINO, right=FINO, top=FINO, bottom=FINO)
    ROJO = Font(color="C00000", bold=True)
    NARANJA = Font(color="B45309", bold=True)

    def hoja(ws, encabezados, anchos):
        for j, h in enumerate(encabezados, 1):
            c = ws.cell(1, j, h)
            c.font = FONT_HDR
            c.fill = FILL_HDR
            c.border = BORDE
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for j, a in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(j)].width = a
        ws.freeze_panes = "A2"

    wb = Workbook()

    ws = wb.active
    ws.title = "RESUMEN"
    hoja(
        ws,
        [
            "ARCHIVO",
            "FACTURA",
            "NIT",
            "SERVICIOS",
            "CON AUTORIZACIÓN",
            "SIN AUTORIZACIÓN",
            "CON OTRO NÚMERO",
            "N° DISTINTOS",
            "VEREDICTO",
            "RUTA",
        ],
        [28, 14, 12, 11, 14, 14, 13, 11, 70, 60],
    )
    for i, r in enumerate(resumenes, 2):
        vals = [
            r["archivo"],
            r["factura"],
            r["nit"],
            r["servicios"],
            r["con_autorizacion"],
            r["sin_autorizacion"],
            r["otros_numeros"],
            r["numeros_distintos"],
            r["veredicto"],
            r["ruta"],
        ]
        for j, v in enumerate(vals, 1):
            c = ws.cell(i, j, v)
            c.border = BORDE
        if r["otros_numeros"] or r["numeros_distintos"] > 1:
            ws.cell(i, 9).font = ROJO
        elif r["sin_autorizacion"] or not r["con_autorizacion"]:
            ws.cell(i, 9).font = NARANJA
    ws.auto_filter.ref = "A1:J%d" % max(2, len(resumenes) + 1)
    if ilegibles:
        fila = len(resumenes) + 3
        ws.cell(fila, 1, "Archivos que NO se pudieron leer:").font = ROJO
        for nombre, motivo in ilegibles:
            fila += 1
            ws.cell(fila, 1, nombre)
            ws.cell(fila, 2, motivo)

    ws2 = wb.create_sheet("DETALLE")
    hoja(
        ws2,
        [
            "ARCHIVO",
            "FACTURA",
            "USUARIO",
            "GRUPO",
            "N°",
            "CÓDIGO",
            "FECHA",
            "VALOR",
            "AUTORIZACIÓN",
            "N° DOMINANTE",
            "ESTADO",
        ],
        [28, 14, 16, 16, 6, 14, 17, 12, 18, 16, 32],
    )
    for i, f in enumerate(detalle, 2):
        vals = [
            f["archivo"],
            f["factura"],
            f["usuario"],
            f["grupo"],
            f["n"],
            f["codigo"],
            f["fecha"],
            f["valor"],
            f["autorizacion"],
            f["dominante"],
            ESTADO_TEXTO[f["estado"]],
        ]
        for j, v in enumerate(vals, 1):
            c = ws2.cell(i, j, v)
            c.border = BORDE
        if f["estado"] == "OTRO_NUMERO":
            ws2.cell(i, 11).font = ROJO
        elif f["estado"] != "OK":
            ws2.cell(i, 11).font = NARANJA
    ws2.auto_filter.ref = "A1:K%d" % max(2, len(detalle) + 1)

    ws3 = wb.create_sheet("ALERTAS")
    hoja(
        ws3,
        [
            "ARCHIVO",
            "FACTURA",
            "USUARIO",
            "GRUPO",
            "N°",
            "CÓDIGO",
            "FECHA",
            "VALOR",
            "AUTORIZACIÓN",
            "N° DOMINANTE",
            "HALLAZGO",
        ],
        [28, 14, 16, 16, 6, 14, 17, 12, 18, 16, 32],
    )
    i = 2
    for f in detalle:
        if f["estado"] == "OK":
            continue
        vals = [
            f["archivo"],
            f["factura"],
            f["usuario"],
            f["grupo"],
            f["n"],
            f["codigo"],
            f["fecha"],
            f["valor"],
            f["autorizacion"],
            f["dominante"],
            ESTADO_TEXTO[f["estado"]],
        ]
        for j, v in enumerate(vals, 1):
            c = ws3.cell(i, j, v)
            c.border = BORDE
        ws3.cell(i, 11).font = ROJO if f["estado"] == "OTRO_NUMERO" else NARANJA
        i += 1
    if i == 2:
        ws3.cell(2, 1, "Sin hallazgos: todos los servicios traen su autorización y coinciden.")
    ws3.auto_filter.ref = "A1:K%d" % max(2, i - 1)

    os.makedirs(os.path.dirname(os.path.abspath(salida)) or ".", exist_ok=True)
    wb.save(salida)
    return salida


if __name__ == "__main__":
    args = sys.argv[1:]
    salida = "INFORME_AUTORIZACIONES_RIPS.xlsx"
    if "-o" in args:
        i = args.index("-o")
        salida = args[i + 1]
        args = args[:i] + args[i + 2 :]
    if not args:
        print("Uso: python autorizaciones_rips.py <json|carpeta|zip> [...] [-o INFORME.xlsx]")
        sys.exit(2)
    try:
        procesar(args, salida)
    except Exception as e:
        print("")
        print("[ERROR] %s" % e)
        sys.exit(1)
