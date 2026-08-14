"""dividir_detallado_por_factura.py — un archivo Excel por cada factura.

El detallado que baja el sistema (y el que deja `ajustar_detallado_glosas.py`)
trae **todas las facturas apiladas en una sola hoja**. Para radicar, anexar o
mandar a PDF hace falta **un archivo por factura**. Este script los separa
conservando el formato tal cual: celdas combinadas, anchos de columna, altos de
fila, fuentes, bordes y formato de moneda.

Cada archivo sale con el número de la factura como nombre (`HUS352890.xlsx`) y
listo para imprimir: se le fija el área de impresión y el ajuste a lo ancho de
la página, para que el PDF no salga partido en dos.

USO (Windows, desde C:\\temp-notas):

    REM Un archivo por factura, todos en la misma carpeta:
    py tools\\dividir_detallado_por_factura.py ^
        --detallado "D:\\USUARIO CARTERA\\Documents\\COOSALUD\\*_AJUSTADO.xlsx" ^
        --salida    "D:\\USUARIO CARTERA\\Documents\\COOSALUD\\POR_FACTURA"

    REM Cada factura en su propia carpeta (para meterle los soportes al lado):
    py tools\\dividir_detallado_por_factura.py ^
        --detallado "D:\\...\\DETALLADOS_31068_AJUSTADO.xlsx" ^
        --salida    "D:\\...\\POR_FACTURA" ^
        --carpeta-por-factura

Después, para pasarlos todos a PDF:  `py tools\\excel_a_pdf.py --help`

Seguro por defecto: NO toca los archivos de entrada. Solo los lee.
Requiere: py -m pip install openpyxl
"""

from __future__ import annotations

import argparse
import csv
import logging
import sys
from dataclasses import dataclass
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from ajustar_detallado_glosas import (  # noqa: E402
    MARCAS_PIE_ARCHIVO,
    IndiceHoja,
    _abrir_libro,
    _norm,
    leer_consolidado,
    normalizar_factura,
    segmentar_facturas,
)

logger = logging.getLogger("dividir_detallado")

# Caracteres que Windows no admite en un nombre de archivo.
_INVALIDOS = '<>:"/\\|?*'


@dataclass
class FacturaExtraida:
    factura: str
    archivo: Path
    origen: str
    filas: int


def _nombre_archivo(factura: str, nombre: str) -> str:
    """HUS0000352890 → 'HUS352890' (corto) o 'HUS0000352890' (largo)."""
    limpio = "".join("_" if c in _INVALIDOS else c for c in factura).strip() or "SIN_NUMERO"
    if nombre == "corto":
        clave = normalizar_factura(limpio)
        return f"HUS{clave}" if clave != "0" else limpio
    return limpio


def _filas_del_pie(idx: IndiceHoja) -> list[int]:
    """Filas del pie legal del archivo (van al final de cada factura suelta)."""
    filas: list[int] = []
    for fila in range(idx.max_fila, 0, -1):
        texto = _norm(idx.texto_fila(fila))
        if not texto:
            if filas:
                filas.append(fila)
            continue
        if any(m in texto for m in MARCAS_PIE_ARCHIVO):
            filas.append(fila)
        else:
            break
    return sorted(filas)


def copiar_bloque(ws_origen, destino, filas: list[int], cache: dict | None = None) -> None:
    """Copia esas filas (con su formato) al principio de la hoja destino.

    El estilo se copia **objeto por objeto** (fuente, borde, relleno, alineación,
    formato de número). Copiar el `_style` crudo no sirve: son índices a las
    tablas de estilos del libro de origen y en el libro nuevo apuntan a la nada
    —openpyxl revienta al combinar celdas—. Se cachean por índice, que son unas
    pocas decenas de combinaciones distintas en todo el archivo.
    """
    from copy import copy

    from openpyxl.utils import get_column_letter

    cache = {} if cache is None else cache
    nueva_de = {vieja: i + 1 for i, vieja in enumerate(filas)}

    for letra, dim in ws_origen.column_dimensions.items():
        destino.column_dimensions[letra] = copy(dim)
        destino.column_dimensions[letra].worksheet = destino

    for vieja, nueva in nueva_de.items():
        dim = ws_origen.row_dimensions.get(vieja)
        if dim is not None:
            copiada = copy(dim)
            copiada.index = nueva
            copiada.worksheet = destino
            destino.row_dimensions[nueva] = copiada
        for celda in ws_origen[vieja]:
            if celda.value is None and not celda.has_style:
                continue
            nueva_celda = destino.cell(row=nueva, column=celda.column)
            nueva_celda.value = celda.value
            if not celda.has_style:
                continue
            clave = tuple(celda._style)
            estilo = cache.get(clave)
            if estilo is None:
                estilo = {
                    "font": copy(celda.font),
                    "border": copy(celda.border),
                    "fill": copy(celda.fill),
                    "alignment": copy(celda.alignment),
                    "number_format": celda.number_format,
                    "protection": copy(celda.protection),
                }
                cache[clave] = estilo
            nueva_celda.font = estilo["font"]
            nueva_celda.border = estilo["border"]
            nueva_celda.fill = estilo["fill"]
            nueva_celda.alignment = estilo["alignment"]
            nueva_celda.number_format = estilo["number_format"]
            nueva_celda.protection = estilo["protection"]

    for rango in ws_origen.merged_cells.ranges:
        vivas = [nueva_de[f] for f in range(rango.min_row, rango.max_row + 1) if f in nueva_de]
        if not vivas:
            continue
        destino.merge_cells(
            start_row=min(vivas),
            start_column=rango.min_col,
            end_row=max(vivas),
            end_column=rango.max_col,
        )

    # Formato de página: se copia el del original y se fuerza el ajuste a lo
    # ancho, que es lo que hace que el PDF no salga partido en dos hojas.
    destino.page_setup = copy(ws_origen.page_setup)
    destino.page_margins = copy(ws_origen.page_margins)
    destino.sheet_properties.pageSetUpPr = copy(ws_origen.sheet_properties.pageSetUpPr)
    destino.print_options = copy(ws_origen.print_options)
    # El ancho útil hay que medirlo también sobre los rangos combinados: el
    # valor de una celda combinada vive en su esquina izquierda, así que mirar
    # solo las celdas con valor deja el área de impresión corta y el PDF sale
    # recortado por la derecha.
    ancho = max(
        max((c.column for f in filas for c in ws_origen[f] if c.value is not None), default=1),
        max((r.max_col for r in destino.merged_cells.ranges), default=1),
    )
    destino.print_area = f"A1:{get_column_letter(ancho)}{len(filas)}"
    destino.page_setup.scale = None  # 'escala' y 'ajustar a' se pisan entre sí
    destino.page_setup.fitToWidth = 1
    destino.page_setup.fitToHeight = 0
    destino.sheet_properties.pageSetUpPr.fitToPage = True


def dividir_libro(
    ruta: Path,
    salida: Path,
    *,
    facturas_pedidas: set[str] | None = None,
    nombre: str = "corto",
    carpeta_por_factura: bool = False,
    incluir_pie: bool = True,
    sobrescribir: bool = True,
) -> list[FacturaExtraida]:
    """Escribe un .xlsx por cada factura de `ruta`."""
    import openpyxl

    wb = _abrir_libro(ruta)
    extraidas: list[FacturaExtraida] = []
    cache_estilos: dict = {}
    try:
        for ws in wb.worksheets:
            idx = IndiceHoja(ws)
            enHoja = segmentar_facturas(idx)
            if not enHoja:
                logger.warning("%s / hoja '%s': no encontré facturas", ruta.name, ws.title)
                continue
            pie = _filas_del_pie(idx) if incluir_pie else []

            for fac in enHoja:
                clave = normalizar_factura(fac.factura)
                if facturas_pedidas and clave not in facturas_pedidas:
                    continue
                base = _nombre_archivo(fac.factura, nombre)
                carpeta = salida / base if carpeta_por_factura else salida
                destino_ruta = carpeta / f"{base}.xlsx"
                if destino_ruta.exists() and not sobrescribir:
                    logger.info("%s ya existe, se salta", destino_ruta.name)
                    continue

                filas = list(range(fac.fila_ini, fac.fila_fin + 1))
                filas += [f for f in pie if f > fac.fila_fin]
                nuevo = openpyxl.Workbook()
                hoja = nuevo.active
                hoja.title = base[:31]
                copiar_bloque(ws, hoja, filas, cache_estilos)
                carpeta.mkdir(parents=True, exist_ok=True)
                nuevo.save(destino_ruta)
                nuevo.close()
                extraidas.append(
                    FacturaExtraida(
                        factura=fac.factura,
                        archivo=destino_ruta,
                        origen=ruta.name,
                        filas=len(filas),
                    )
                )
    finally:
        wb.close()
    return extraidas


def construir_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Separa el detallado en un archivo Excel por factura.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument(
        "--detallado", type=Path, nargs="+", required=True, help="Uno o varios Excel del detallado."
    )
    p.add_argument("--salida", type=Path, required=True, help="Carpeta donde dejar los archivos.")
    p.add_argument(
        "--consolidado", type=Path, help="Solo separar estas facturas (Excel/CSV/TXT con la lista)."
    )
    p.add_argument(
        "--nombre",
        choices=("corto", "largo"),
        default="corto",
        help="corto: HUS352890 (por defecto). largo: como venga en el archivo.",
    )
    p.add_argument(
        "--carpeta-por-factura",
        action="store_true",
        help="Deja cada factura en su propia carpeta (salida\\HUS352890\\HUS352890.xlsx).",
    )
    p.add_argument(
        "--sin-pie", action="store_true", help="No copiar el texto legal del final del archivo."
    )
    p.add_argument("--no-sobrescribir", action="store_true", help="Saltar los que ya existan.")
    p.add_argument("--reporte-csv", type=Path, help="Listado de lo generado.")
    p.add_argument("--verbose", action="store_true")
    return p


def main(argv: list[str] | None = None) -> int:
    args = construir_parser().parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)-8s %(message)s",
    )

    pedidas: set[str] | None = None
    if args.consolidado:
        if not args.consolidado.exists():
            logger.error("No existe el consolidado: %s", args.consolidado)
            return 2
        lista, duplicadas = leer_consolidado(args.consolidado)
        pedidas = {normalizar_factura(f) for f in lista}
        logger.info(
            "Consolidado: %d factura(s), %d duplicada(s) descartada(s)", len(lista), len(duplicadas)
        )

    rutas = []
    for patron in args.detallado:
        if patron.exists():
            rutas.append(patron)
            continue
        encontrados = sorted(patron.parent.glob(patron.name))  # comodines de Windows
        if not encontrados:
            logger.error("No existe el detallado: %s", patron)
            return 2
        rutas.extend(encontrados)

    todas: list[FacturaExtraida] = []
    for ruta in rutas:
        logger.info("Separando %s …", ruta.name)
        extraidas = dividir_libro(
            ruta,
            args.salida,
            facturas_pedidas=pedidas,
            nombre=args.nombre,
            carpeta_por_factura=args.carpeta_por_factura,
            incluir_pie=not args.sin_pie,
            sobrescribir=not args.no_sobrescribir,
        )
        logger.info("   %d archivo(s)", len(extraidas))
        todas.extend(extraidas)

    logger.info("─" * 68)
    logger.info("Archivos generados: %d en %s", len(todas), args.salida)
    if pedidas:
        faltan = pedidas - {normalizar_factura(e.factura) for e in todas}
        if faltan:
            logger.warning(
                "%d factura(s) del consolidado no estaban en los detallados: %s",
                len(faltan),
                ", ".join(sorted(faltan)[:20]),
            )

    if args.reporte_csv:
        args.reporte_csv.parent.mkdir(parents=True, exist_ok=True)
        with args.reporte_csv.open("w", encoding="utf-8-sig", newline="") as fh:
            w = csv.writer(fh, delimiter=";")
            w.writerow(["FACTURA", "ARCHIVO", "ORIGEN", "FILAS"])
            for e in todas:
                w.writerow([e.factura, str(e.archivo), e.origen, e.filas])
        logger.info("Reporte: %s", args.reporte_csv)
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
