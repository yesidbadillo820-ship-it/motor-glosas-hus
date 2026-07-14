"""respuesta_tramites_dgh.py — Llena el archivo de TRÁMITES de DGH con la respuesta dada a cada glosa.

ÚLTIMO PASO del ciclo COOSALUD: después de objetar en DGH y responder las
glosas en el portal de COOSALUD, DGH pide registrar la RESPUESTA del trámite.
Se descarga de DGH el export de seguimiento (hoja CRRPSEGUIMIENTORECEPCIONOBJECIO,
una fila por concepto/servicio objetado) y este bot le agrega las 4 columnas
del cargue, con la MISMA lógica de las respuestas del portal:

    FECHA DE CARGUE     la fecha en que se radica la respuesta (--fecha-cargue)
    CODIGO RESPUESTA    RE9502 si la glosa fue EXTEMPORÁNEA (la EPS la formuló
                        pasados los 20 días hábiles desde la radicación de la
                        factura, art. 57 Ley 1438/2011) · RE9901 si fue a tiempo
    VALOR ACEPTADO      0 (el HUS no acepta la glosa)
    OBSERVACION         texto de extemporaneidad, o el texto del área según el
                        tipo (TARIFAS/AUTORIZACION/FACTURACION/SOPORTES)

REGLA DE FACTURA COMPLETA (definida por el área):
  - Factura EXTEMPORÁNEA (RE9502): TODOS sus ítems — incluidos los CL — llevan
    la misma respuesta de extemporaneidad y la factura va completa al archivo.
  - Factura A TIEMPO (RE9901) con ítems de CALIDAD (CL) o COBERTURA (CO): esos
    ítems los responde auditoría médica, así que la factura se QUITA COMPLETA
    del archivo (no solo el concepto CL) — se sube después, cuando las doctoras
    respondan su parte. La lista de facturas excluidas queda en un TXT junto al
    archivo final.

La fecha de radicación de cada factura se lee de las cabeceras del portal
(carpeta FACTURAS del "CARGUE MASIVO COOSALUD"); la fecha de la glosa sale de
la misma fila del export (FechaObjecion). Días hábiles con festivos de Colombia.

USO (sin argumentos pregunta todo; acepta rutas arrastradas):
    py respuesta_tramites_dgh.py
    py respuesta_tramites_dgh.py --plantilla "PARA CARGUE DE TRAMITES COOSALUD.xlsx" ^
        --carpeta "D:\\USUARIO CARTERA\\Desktop\\CARGUE MASIVO COOSALUD" ^
        --fecha-cargue 14/07/2026

Requiere consolidar_coosalud.py en la misma carpeta (comparten textos y motor
de días hábiles).
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

AQUI = Path(__file__).resolve().parent
sys.path.insert(0, str(AQUI))

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    print("Falta openpyxl. Instalar con:  py -m pip install openpyxl")
    sys.exit(1)

try:
    from consolidar_coosalud import (
        CARPETA_FACTURAS,
        COD_RTA_EXTEMPORANEA,
        COD_RTA_NORMAL,
        DIAS_HABILES_EPS,
        OBS_EXTEMPORANEA,
        OBS_POR_TIPO,
        TIPO_POR_PREFIJO,
        a_fecha,
        buscar_archivos,
        consolidar_facturas,
        dias_habiles_entre,
        norm_factura,
        norm_header,
        norm_texto,
    )
except ImportError:
    print("ERROR: falta consolidar_coosalud.py en la misma carpeta que este script.")
    sys.exit(1)

# Nombres EXACTOS del archivo de ejemplo del área (incluye espacios finales
# en las dos primeras — así vienen en la plantilla real y así se respetan).
COLS_NUEVAS = ["FECHA DE CARGUE ", "CODIGO RESPUESTA ", "VALOR ACEPTADO", "OBSERVACION"]


def limpiar_ruta(texto: str) -> str:
    return texto.strip().strip('"').strip("'").strip()


def pedir(mensaje: str) -> str:
    try:
        return input(mensaje)
    except EOFError:
        return ""


def cargar_radicaciones(carpeta: Path) -> dict[str, object]:
    """factura_norm -> fecha_radicacion, desde las cabeceras HUS*.xlsx del portal."""
    archivos = buscar_archivos(carpeta, CARPETA_FACTURAS, "*.xlsx")
    if not archivos:
        raise ValueError(
            f"No encontré cabeceras de factura en {carpeta / CARPETA_FACTURAS}. "
            'Debe ser la carpeta "CARGUE MASIVO COOSALUD" del masivo.'
        )
    h_fact, f_fact = consolidar_facturas(archivos)
    mapa = {norm_header(h): i for i, h in enumerate(h_fact)}
    idx_fact = mapa.get(norm_header("numero_factura"))
    idx_rad = mapa.get(norm_header("fecha_radicacion"))
    if idx_fact is None or idx_rad is None:
        raise ValueError("Las cabeceras de factura no traen numero_factura/fecha_radicacion.")
    radicacion: dict[str, object] = {}
    for r in f_fact:
        f = norm_factura(r[idx_fact])
        fec = a_fecha(r[idx_rad])
        if fec and f not in radicacion:
            radicacion[f] = fec
    return radicacion


def procesar(
    plantilla: Path, carpeta: Path, fecha_cargue, salida: Path | None, max_facturas: int = 499
) -> int:
    print(f"\nLeyendo radicaciones del masivo: {carpeta} ...")
    radicacion = cargar_radicaciones(carpeta)
    print(f"  {len(radicacion)} facturas con fecha de radicación.")

    print(f"Leyendo el export de trámites: {plantilla.name} ...")
    wb = openpyxl.load_workbook(plantilla, read_only=True, data_only=True)
    nombre_hoja = wb.sheetnames[0]
    ws = wb[nombre_hoja]
    filas = list(ws.iter_rows(values_only=True))
    wb.close()
    if not filas:
        raise ValueError(f"'{plantilla.name}' está vacío.")
    hdr = [norm_texto(c) for c in filas[0]]
    try:
        col_fact = hdr.index("FacturaCartera.Factura")
        col_fobj = hdr.index("FechaObjecion")
        col_cod = hdr.index("ListadoConceptos.ConceptoObjecion.Codigo")
    except ValueError:
        raise ValueError(
            f"'{plantilla.name}' no parece el export de trámites de DGH "
            "(faltan FacturaCartera.Factura / FechaObjecion / "
            "ListadoConceptos.ConceptoObjecion.Codigo)."
        ) from None
    # Si la plantilla ya trae las columnas nuevas (re-corrida), no duplicarlas.
    hdr_base = [h for h in hdr if h.upper() not in {c.upper() for c in COLS_NUEVAS}]
    n_base = len(hdr_base)

    # PASADA 1: calcular la respuesta de cada fila y marcar las facturas
    # INCOMPLETAS (a tiempo con ítems CL/CO de auditoría médica, o sin fecha
    # de radicación). Esas facturas se quitan COMPLETAS del archivo.
    sin_rad: set[str] = set()
    blanco: dict[str, int] = {}
    facturas_incompletas: set[str] = set()
    dia_cache: dict[tuple[str, str], int] = {}
    calculadas: list[tuple[str, list]] = []  # (factura_norm, fila completa)
    for r in filas[1:]:
        base = list(r[:n_base]) + [None] * (n_base - len(r))
        fkey = norm_factura(r[col_fact])
        rad = radicacion.get(fkey)
        fobj = a_fecha(r[col_fobj])
        codigo = norm_texto(r[col_cod])
        tipo = TIPO_POR_PREFIJO.get(codigo[:2].upper(), "")

        if rad is None or fobj is None:
            sin_rad.add(fkey)
            facturas_incompletas.add(fkey)
            calculadas.append((fkey, base + [fecha_cargue, "", 0, ""]))
            continue
        clave = (fkey, str(fobj))
        if clave not in dia_cache:
            dia_cache[clave] = dias_habiles_entre(rad, fobj)
        if dia_cache[clave] > DIAS_HABILES_EPS:
            # Extemporánea: TODOS los ítems (incluidos CL) van con RE9502.
            calculadas.append(
                (fkey, base + [fecha_cargue, COD_RTA_EXTEMPORANEA, 0, OBS_EXTEMPORANEA])
            )
        else:
            obs = OBS_POR_TIPO.get(tipo, "")
            if not obs:
                # CALIDAD/COBERTURA a tiempo: responde auditoría médica ->
                # la factura ENTERA se queda por fuera de este cargue.
                blanco[tipo or codigo[:2].upper()] = blanco.get(tipo or "?", 0) + 1
                facturas_incompletas.add(fkey)
                calculadas.append((fkey, base + [fecha_cargue, "", 0, ""]))
            else:
                calculadas.append((fkey, base + [fecha_cargue, COD_RTA_NORMAL, 0, obs]))

    # PASADA 2: solo las facturas COMPLETAS van al archivo.
    salida_filas = [fila for fkey, fila in calculadas if fkey not in facturas_incompletas]
    n_ext = sum(1 for f in salida_filas if f[n_base + 1] == COD_RTA_EXTEMPORANEA)
    n_ok = sum(1 for f in salida_filas if f[n_base + 1] == COD_RTA_NORMAL)

    # El cargue de trámites de DGH acepta máximo ~500 facturas: se parte en
    # LOTES de max_facturas facturas (una factura NUNCA se parte entre lotes).
    orden_facturas: list[str] = []
    filas_por_factura: dict[str, list[list]] = {}
    for fkey, fila in calculadas:
        if fkey in facturas_incompletas:
            continue
        if fkey not in filas_por_factura:
            orden_facturas.append(fkey)
            filas_por_factura[fkey] = []
        filas_por_factura[fkey].append(fila)

    lotes_facturas = [
        orden_facturas[i : i + max_facturas] for i in range(0, len(orden_facturas), max_facturas)
    ]
    base_nombre = salida or plantilla.with_name(
        f"MASIVO COOSALUD {fecha_cargue.strftime('%d%m%Y')}.xlsx"
    )
    archivos_salida: list[tuple[Path, int, int]] = []
    for n_lote, facturas_lote in enumerate(lotes_facturas, 1):
        if len(lotes_facturas) == 1:
            out = base_nombre
        else:
            out = base_nombre.with_name(f"{base_nombre.stem} LOTE {n_lote:02d}.xlsx")
        filas_lote = [fila for fkey in facturas_lote for fila in filas_por_factura[fkey]]
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.title = nombre_hoja
        ws2.append(hdr_base + COLS_NUEVAS)
        for f in filas_lote:
            ws2.append(f)
        col_fc = n_base + 1
        for row_i in range(2, len(filas_lote) + 2):
            ws2.cell(row=row_i, column=col_fc).number_format = "DD/MM/YYYY"
        for i, h in enumerate(hdr_base + COLS_NUEVAS, start=1):
            ws2.column_dimensions[get_column_letter(i)].width = max(12, min(42, len(str(h)) + 4))
        wb2.save(out)
        archivos_salida.append((out, len(facturas_lote), len(filas_lote)))
    out = base_nombre  # para el TXT de pendientes

    total_facturas = len({fkey for fkey, _ in calculadas})
    facturas_salida = total_facturas - len(facturas_incompletas)
    print("\n===== LISTO =====")
    for ruta, n_fact, n_filas in archivos_salida:
        print(f"  {ruta.name:46} {n_fact:>4} facturas · {n_filas:>6} conceptos")
    print(
        f"  TOTAL: {len(archivos_salida)} archivo(s) · {len(salida_filas)} conceptos · "
        f"{facturas_salida} facturas (de {total_facturas}) · máx {max_facturas} facturas por archivo"
    )
    print(f"  Extemporáneas ({COD_RTA_EXTEMPORANEA}): {n_ext}")
    print(f"  A tiempo ({COD_RTA_NORMAL}) con texto del área: {n_ok}")
    if facturas_incompletas:
        pend = out.with_name(out.stem + " - FACTURAS PENDIENTES AUDITORIA MEDICA.txt")
        pend.write_text("\n".join(sorted(facturas_incompletas)) + "\n", encoding="utf-8")
        print(
            f"  QUITADAS del archivo: {len(facturas_incompletas)} facturas COMPLETAS "
            f"(a tiempo con ítems CL/CO de auditoría médica"
            + (f"; {len(sin_rad)} sin radicación" if sin_rad else "")
            + ")."
        )
        print(f"    Lista guardada en: {pend.name}")
        print("    Se suben después, cuando las doctoras respondan su parte.")
        if blanco:
            print(
                "    Ítems que esperan a auditoría médica: "
                + " · ".join(f"{t}: {n}" for t, n in blanco.items())
            )
    print(f"  FECHA DE CARGUE: {fecha_cargue.strftime('%d/%m/%Y')} · VALOR ACEPTADO: 0 en todas")
    return 0


def preguntar_archivo(titulo: str, patron: str) -> Path | None:
    lugares = [
        AQUI,
        AQUI.parent,
        Path.home() / "Downloads",
        Path.home() / "Desktop",
        Path(r"D:\USUARIO CARTERA\Desktop"),
        Path(r"D:\USUARIO CARTERA\Downloads"),
    ]
    cands: dict[Path, float] = {}
    for lugar in lugares:
        if lugar.is_dir():
            for f in lugar.glob(patron):
                if not f.name.startswith("~$"):
                    cands[f] = f.stat().st_mtime
    lista = sorted(cands, key=cands.get, reverse=True)[:5]
    print(f"\n--- {titulo} ---")
    if lista:
        for i, c in enumerate(lista, 1):
            print(f"    {i}) {c}")
        r = limpiar_ruta(pedir("  Escribe el numero, o arrastra aqui el archivo y Enter: "))
    else:
        r = limpiar_ruta(pedir("  Arrastra aqui el archivo y presiona Enter: "))
    if r.isdigit() and lista and 1 <= int(r) <= len(lista):
        return lista[int(r) - 1]
    if r:
        p = Path(r)
        if p.exists():
            return p
        print(f"  ERROR: no existe: {p}")
        return None
    return lista[0] if lista else None


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(
        description="Llena el export de trámites de DGH con la respuesta de cada glosa."
    )
    p.add_argument("--plantilla", default=None, help="Export de DGH (PARA CARGUE DE TRAMITES...).")
    p.add_argument("--carpeta", default=None, help='Carpeta "CARGUE MASIVO COOSALUD" del masivo.')
    p.add_argument("--fecha-cargue", default=None, help="Fecha del cargue DD/MM/AAAA (def: hoy).")
    p.add_argument(
        "--salida", default=None, help="Ruta del archivo final (def: MASIVO COOSALUD <fecha>.xlsx)."
    )
    p.add_argument(
        "--max-facturas-lote",
        type=int,
        default=499,
        help="Máximo de facturas por archivo (el cargue de trámites de DGH "
        "acepta hasta ~500). Def: 499. Si hay más, salen LOTE 01, LOTE 02, ...",
    )
    args = p.parse_args(argv)

    print("=" * 72)
    print("   RESPUESTA DE TRAMITES DGH - agrega las 4 columnas de respuesta")
    print("=" * 72)

    plantilla = (
        Path(args.plantilla)
        if args.plantilla
        else preguntar_archivo("PASO A: el export de tramites de DGH", "*TRAMITES*.xlsx")
    )
    if not plantilla or not plantilla.is_file():
        print("  Sin el export de DGH no puedo continuar.")
        return 1
    carpeta = (
        Path(args.carpeta)
        if args.carpeta
        else preguntar_archivo(
            'PASO B: la carpeta "CARGUE MASIVO COOSALUD" del masivo', "CARGUE MASIVO*"
        )
    )
    if not carpeta or not carpeta.is_dir():
        print("  Sin la carpeta del masivo no puedo continuar.")
        return 1
    if args.fecha_cargue:
        try:
            fecha_cargue = datetime.strptime(args.fecha_cargue, "%d/%m/%Y").date()
        except ValueError:
            print(f"  ERROR: --fecha-cargue debe ser DD/MM/AAAA. Recibido: {args.fecha_cargue}")
            return 1
    else:
        hoy = datetime.now().strftime("%d/%m/%Y")
        r = limpiar_ruta(pedir(f"\n  Fecha del cargue DD/MM/AAAA (Enter = {hoy}): "))
        try:
            fecha_cargue = datetime.strptime(r or hoy, "%d/%m/%Y").date()
        except ValueError:
            print("  Fecha invalida.")
            return 1

    try:
        return procesar(
            plantilla,
            carpeta,
            fecha_cargue,
            Path(args.salida) if args.salida else None,
            max_facturas=max(1, args.max_facturas_lote),
        )
    except ValueError as exc:
        print(f"\n  ERROR: {exc}")
        return 1


if __name__ == "__main__":
    codigo = main()
    pedir("\n  Presiona Enter para salir...")
    raise SystemExit(codigo)
