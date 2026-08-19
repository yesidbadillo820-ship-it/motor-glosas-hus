"""glosas_adres_por_factura.py — un Excel por factura con lo que sigue glosado.

A diferencia de `dividir_detallado_por_factura.py`, este NO necesita el
detallado impreso del hospital: arma cada archivo **directamente del reporte
del ADRES**, que ya trae por ítem cuánto se reclamó, cuánto aprobaron y cuánto
quedó glosado.

Qué hace con cada factura:

  1. Deja **solo lo que sigue glosado**. Lo que el ADRES ya aprobó (glosado en
     cero) no va: eso ya está pago y no hay nada que responder.
  2. Junta los renglones repetidos. El reporte abre **una fila por cada causal**
     del mismo ítem, así que un ítem con tres causales sale tres veces. Se
     colapsa en una sola fila con las causales juntas — si no, la suma sale
     inflada al doble o al triple.
  3. **Verifica contra la cifra oficial del ADRES.** Con `--facturas` se le pasa
     el archivo de facturas del paquete (`FACTURAS PAQUETE NNNNN_NN
     FACTURAS.xlsx`) y cada archivo sale sellado con «CUADRA CON EL ADRES» o con
     el aviso de la diferencia exacta.

Lo tercero es lo importante. En el paquete 31078 el propio reporte del ADRES
**no cuadra consigo mismo en 27 de las 81 facturas** (repite renglones sin que
haya causales distintas que lo expliquen), y son justo las de más plata. Este
bot no inventa una regla para taparlo: lo dice en el archivo y en el resumen,
para que esas facturas se trabajen con el detalle bajado del portal
(Reclamaciones → Reportes Lupa al giro).

USO (Windows, desde C:\\temp-notas):

    py tools\\glosas_adres_por_factura.py ^
        --reporte  "D:\\...\\ReporteGlosasReclamPAQUETE 31078.xlsx" ^
        --facturas "D:\\...\\FACTURAS PAQUETE 31078_81 FACTURAS.xlsx" ^
        --salida   "D:\\...\\POR_FACTURA_31078"

    REM Cada factura en su carpeta, para meterle los soportes al lado:
    ... --carpeta-por-factura

    REM Solo las que no cuadran, para revisarlas aparte:
    ... --solo-descuadradas

Después se pasan a PDF con `tools/excel_a_pdf.py`.

Requiere: py -m pip install openpyxl
"""

from __future__ import annotations

import argparse
import csv
import logging
import sys
from dataclasses import dataclass, field
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _dinero import a_numero  # noqa: E402
from ajustar_detallado_glosas import normalizar_factura  # noqa: E402
from preauditar_glosas_adres import (  # noqa: E402
    FilaMacro,
    _moneda,
    leer_reporte,
)

logger = logging.getLogger("glosas_adres_por_factura")

# Cuánto se permite de diferencia contra la cifra oficial antes de gritar. El
# ADRES maneja centavos, así que un peso es ruido de redondeo.
TOLERANCIA = 1.0

ENCABEZADOS = [
    "CONSECUTIVO",
    "TIPO",
    "CÓDIGO",
    "DESCRIPCIÓN DEL SERVICIO",
    "CANT. RECLAMADA",
    "VR. RECLAMADO",
    "CANT. APROBADA",
    "VR. APROBADO",
    "VR. GLOSADO",
    "CAUSAL(ES)",
    "OBSERVACIÓN DEL ADRES",
]


@dataclass
class ItemGlosado:
    """Un ítem de la factura que sigue glosado, con todas sus causales."""

    fila: FilaMacro
    causales: list[str] = field(default_factory=list)
    anotaciones: list[str] = field(default_factory=list)

    @property
    def valor_glosado(self) -> float:
        return self.fila.valor_glosado or 0.0


@dataclass
class ResultadoFactura:
    factura: str
    clave: str
    items: list[ItemGlosado]
    filas_reporte: int
    oficial: float | None = None

    @property
    def suma(self) -> float:
        return round(sum(i.valor_glosado for i in self.items), 2)

    @property
    def cuadra(self) -> bool | None:
        """None = no se pudo verificar (no se pasó el archivo oficial)."""
        if self.oficial is None:
            return None
        return abs(self.suma - self.oficial) < TOLERANCIA

    @property
    def diferencia(self) -> float:
        return round((self.oficial or 0) - self.suma, 2)


def _contenido(f: FilaMacro) -> tuple:
    """Lo que identifica al ítem, sin la causal ni la anotación."""
    return (
        str(f.consecutivo or ""),
        (f.tipo_elemento or "").strip(),
        (f.codigo or "").strip(),
        (f.descripcion or "").strip(),
        round(f.cant_reclamada or 0, 4),
        round(f.valor_reclamado or 0, 2),
        round(f.cant_aprobada or 0, 4),
        round(f.valor_aprobado or 0, 2),
        round(f.valor_glosado or 0, 2),
    )


def bloques_de_item(filas: list[FilaMacro]):
    """Recorre el reporte agrupando los renglones que son el MISMO ítem.

    El reporte del ADRES abre **una fila por cada causal**. Renglones seguidos
    con el mismo contenido pero causal distinta son un solo ítem; si repiten la
    misma causal son ítems de verdad (una factura puede traer el mismo
    medicamento cargado varias veces) y se respetan.

    Devuelve, por cada bloque: `(filas del bloque, cuántos ítems reales son,
    causales distintas, anotaciones distintas)`.
    """
    i = 0
    while i < len(filas):
        j = i
        while j < len(filas) and _contenido(filas[j]) == _contenido(filas[i]):
            j += 1
        bloque = filas[i:j]
        causales: list[str] = []
        anotaciones: list[str] = []
        for f in bloque:
            texto = (f.descripcion_glosa or "").strip()
            if texto and texto not in causales:
                causales.append(texto)
            nota = (f.anotacion or "").strip()
            if nota and nota not in anotaciones:
                anotaciones.append(nota)
        distintas = len(causales) or 1
        # Tantos ítems reales como veces se repite el ciclo completo de causales.
        reales = len(bloque) // distintas if len(bloque) % distintas == 0 else len(bloque)
        yield bloque, max(reales, 1), causales, anotaciones
        i = j


def marcar_conteo(filas: list[FilaMacro]) -> list[bool]:
    """Para cada renglón, si su valor **cuenta** o ya lo contó otro.

    Sirve para quien necesita conservar todos los renglones (la pantalla los
    muestra uno por causal, porque el gestor decide causal por causal) pero
    **no puede sumar dos veces la misma plata**.
    """
    marcas: list[bool] = []
    for bloque, reales, _causales, _anotaciones in bloques_de_item(filas):
        marcas.extend([True] * reales + [False] * (len(bloque) - reales))
    return marcas


def agrupar_items(filas: list[FilaMacro]) -> list[ItemGlosado]:
    """Los ítems de la factura, ya sin los renglones repetidos por causal."""
    salida: list[ItemGlosado] = []
    for bloque, reales, causales, anotaciones in bloques_de_item(filas):
        for _ in range(reales):
            salida.append(
                ItemGlosado(fila=bloque[0], causales=list(causales), anotaciones=list(anotaciones))
            )
    return salida


def leer_oficiales(ruta: Path) -> dict[str, float]:
    """Valor glosado oficial por factura, del archivo de facturas del paquete.

    Es la única cifra que manda: el reporte por ítem repite renglones y suma de
    más. Se busca la columna del glosado por su nombre, no por su posición.
    """
    import openpyxl

    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    try:
        salida: dict[str, float] = {}
        for ws in wb.worksheets:
            col_factura = col_glosado = None
            for fila in ws.iter_rows(min_row=1, max_row=12, values_only=True):
                for n, celda in enumerate(fila):
                    texto = str(celda or "").strip().upper()
                    if "GLOSADO" in texto:
                        col_glosado = n
                    elif texto in (
                        "ETIQUETAS DE FILA",
                        "NÚMERO FACTURA",
                        "NUMERO FACTURA",
                        "FACTURA",
                    ):
                        col_factura = n
                if col_factura is not None and col_glosado is not None:
                    break
            if col_factura is None or col_glosado is None:
                continue
            for fila in ws.iter_rows(values_only=True):
                if col_factura >= len(fila) or col_glosado >= len(fila):
                    continue
                nombre = str(fila[col_factura] or "").strip()
                if not nombre.upper().startswith("HUS") or nombre.upper().startswith("TOTAL"):
                    continue
                salida[normalizar_factura(nombre)] = round(a_numero(fila[col_glosado]), 2)
            if salida:
                return salida
        return salida
    finally:
        wb.close()


def escribir_factura(res: ResultadoFactura, destino: Path, paquete: str) -> None:
    """El Excel de una factura, listo para imprimir y anexar."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GLOSAS"

    borde = Border(*[Side(style="thin")] * 4)
    negrilla = Font(bold=True)

    ws["A1"] = "DETALLADO DE FACTURA — GLOSAS DEL ADRES"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"Factura: {res.factura}"
    ws["A3"] = f"Paquete: {paquete}"
    primera = res.items[0].fila if res.items else None
    if primera is not None:
        ws["A4"] = f"Radicación: {primera.radicacion or '—'}"
        ws["A5"] = f"Documento del paciente: {primera.doc_victima or '—'}"
    ws["A6"] = "Solo aparece lo que SIGUE GLOSADO. Lo que el ADRES ya aprobó no se incluye."
    ws["A6"].font = Font(italic=True, size=9)
    for celda in ("A2", "A3", "A4", "A5"):
        ws[celda].font = negrilla

    inicio = 8
    for n, titulo in enumerate(ENCABEZADOS, start=1):
        c = ws.cell(row=inicio, column=n, value=titulo)
        c.font = Font(bold=True, size=9)
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.border = borde
        c.alignment = Alignment(wrap_text=True, vertical="center")

    fila_actual = inicio
    for item in res.items:
        fila_actual += 1
        f = item.fila
        valores = [
            f.consecutivo or "",
            f.tipo_elemento or "",
            f.codigo or "",
            f.descripcion or "",
            f.cant_reclamada or 0,
            f.valor_reclamado or 0,
            f.cant_aprobada or 0,
            f.valor_aprobado or 0,
            f.valor_glosado or 0,
            " | ".join(item.causales) or "(el reporte no dice la causal)",
            " | ".join(item.anotaciones),
        ]
        for n, valor in enumerate(valores, start=1):
            c = ws.cell(row=fila_actual, column=n, value=valor)
            c.border = borde
            c.alignment = Alignment(wrap_text=n in (4, 10, 11), vertical="top")
            if n in (6, 7, 8, 9) or (n == 5):
                c.number_format = "#,##0" if n != 5 else "0.##"

    # Total
    fila_actual += 1
    ws.cell(row=fila_actual, column=8, value="TOTAL GLOSADO").font = negrilla
    total = ws.cell(row=fila_actual, column=9, value=res.suma)
    total.font = negrilla
    total.number_format = "#,##0"
    total.border = borde

    # El sello de verificación: qué tan confiable es este archivo.
    fila_actual += 2
    if res.cuadra is True:
        ws.cell(
            row=fila_actual,
            column=1,
            value=f"VERIFICADO: cuadra con la cifra oficial del ADRES ({_moneda(res.oficial)}).",
        ).font = Font(bold=True, color="15803D")
    elif res.cuadra is False:
        ws.cell(
            row=fila_actual,
            column=1,
            value=(
                f"OJO — NO CUADRA. El ADRES dice que esta factura tiene glosado "
                f"{_moneda(res.oficial)}, pero el detalle del reporte suma {_moneda(res.suma)} "
                f"(diferencia {_moneda(abs(res.diferencia))}). El reporte repite renglones. "
                f"Baje el detalle del portal (Reclamaciones → Reportes Lupa al giro) antes de "
                f"responder esta factura."
            ),
        ).font = Font(bold=True, color="B45309")
    else:
        ws.cell(
            row=fila_actual,
            column=1,
            value="Sin verificar: no se pasó el archivo de facturas del paquete (--facturas).",
        ).font = Font(italic=True)

    sin_causal = sum(1 for i in res.items if not i.causales)
    if sin_causal:
        fila_actual += 1
        ws.cell(
            row=fila_actual,
            column=1,
            value=(
                f"{sin_causal} de {len(res.items)} renglones vienen SIN causal escrita en el "
                f"reporte. Sin la causal no se puede objetar ítem por ítem: hay que bajar el "
                f"detalle del portal."
            ),
        ).font = Font(bold=True, color="B45309")

    anchos = [11, 16, 14, 46, 9, 14, 9, 14, 14, 42, 46]
    for n, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(n)].width = ancho

    ws.print_area = f"A1:{get_column_letter(len(ENCABEZADOS))}{fila_actual}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{inicio}:{inicio}"

    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)


def procesar(
    reporte: Path,
    salida: Path,
    *,
    facturas: Path | None = None,
    paquete: str | None = None,
    carpeta_por_factura: bool = False,
    solo_descuadradas: bool = False,
) -> list[ResultadoFactura]:
    filas = leer_reporte(reporte, paquete=paquete)
    if not filas:
        raise ValueError("El reporte no trajo ninguna fila glosada.")
    numero = paquete or (filas[0].paquete or "").strip()

    oficiales = leer_oficiales(facturas) if facturas else {}
    if facturas and not oficiales:
        logger.warning("No se pudo leer ninguna cifra oficial de %s", facturas.name)

    por_factura: dict[str, list[FilaMacro]] = {}
    for f in filas:
        por_factura.setdefault(normalizar_factura(f.factura), []).append(f)

    resultados: list[ResultadoFactura] = []
    for clave, grupo in sorted(por_factura.items()):
        items = agrupar_items(grupo)
        res = ResultadoFactura(
            factura=grupo[0].factura,
            clave=clave,
            items=items,
            filas_reporte=len(grupo),
            oficial=oficiales.get(clave),
        )
        resultados.append(res)

    escritos = 0
    for res in resultados:
        if solo_descuadradas and res.cuadra is not False:
            continue
        nombre = f"{res.factura.strip().replace('/', '-')}.xlsx"
        destino = salida / res.factura.strip() / nombre if carpeta_por_factura else salida / nombre
        escribir_factura(res, destino, numero)
        escritos += 1

    _escribir_resumen(resultados, salida / f"RESUMEN_{numero or 'PAQUETE'}.csv")
    logger.info("%d archivo(s) escrito(s) en %s", escritos, salida)
    return resultados


def _escribir_resumen(resultados: list[ResultadoFactura], destino: Path) -> None:
    destino.parent.mkdir(parents=True, exist_ok=True)
    with open(destino, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh, delimiter=";")
        w.writerow(
            [
                "FACTURA",
                "FILAS_REPORTE",
                "ITEMS_GLOSADOS",
                "SUMA_DETALLE",
                "OFICIAL_ADRES",
                "DIFERENCIA",
                "CUADRA",
                "RENGLONES_SIN_CAUSAL",
            ]
        )
        for r in sorted(resultados, key=lambda x: x.factura):
            sin_causal = sum(1 for i in r.items if not i.causales)
            w.writerow(
                [
                    r.factura,
                    r.filas_reporte,
                    len(r.items),
                    f"{r.suma:.2f}",
                    "" if r.oficial is None else f"{r.oficial:.2f}",
                    "" if r.oficial is None else f"{r.diferencia:.2f}",
                    {True: "SI", False: "NO", None: "SIN VERIFICAR"}[r.cuadra],
                    sin_causal,
                ]
            )


def construir_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Un Excel por factura con lo que sigue glosado, sacado del reporte del ADRES.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("--reporte", required=True, type=Path, help="ReporteGlosasReclamPAQUETE N.xlsx")
    p.add_argument(
        "--facturas",
        type=Path,
        help="FACTURAS PAQUETE N_NN FACTURAS.xlsx — para verificar contra la cifra oficial",
    )
    p.add_argument("--salida", required=True, type=Path, help="Carpeta donde dejar los archivos")
    p.add_argument("--paquete", help="Número de paquete, si el reporte trae varios")
    p.add_argument(
        "--carpeta-por-factura",
        action="store_true",
        help="Cada factura en su propia carpeta",
    )
    p.add_argument(
        "--solo-descuadradas",
        action="store_true",
        help="Escribir solo las facturas que NO cuadran con el ADRES",
    )
    return p


def main(argv: list[str] | None = None) -> int:
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    args = construir_parser().parse_args(argv)
    try:
        resultados = procesar(
            args.reporte,
            args.salida,
            facturas=args.facturas,
            paquete=args.paquete,
            carpeta_por_factura=args.carpeta_por_factura,
            solo_descuadradas=args.solo_descuadradas,
        )
    except Exception as e:  # noqa: BLE001 - mensaje claro para el auditor
        logger.error("No se pudo procesar: %s", e)
        return 1

    cuadran = [r for r in resultados if r.cuadra is True]
    no_cuadran = [r for r in resultados if r.cuadra is False]
    sin_verificar = [r for r in resultados if r.cuadra is None]
    total_oficial = sum(r.oficial or 0 for r in resultados)

    print()
    print(f"Facturas procesadas : {len(resultados)}")
    print(f"Ítems glosados      : {sum(len(r.items) for r in resultados)}")
    if total_oficial:
        print(f"Glosado oficial     : {_moneda(total_oficial)}")
    print(f"  cuadran con el ADRES : {len(cuadran)}")
    if no_cuadran:
        print(f"  NO cuadran           : {len(no_cuadran)}  <- revisar con el detalle del portal")
        for r in sorted(no_cuadran, key=lambda x: -abs(x.diferencia))[:10]:
            print(
                f"     {r.factura:<12} oficial {_moneda(r.oficial):>14}  "
                f"detalle {_moneda(r.suma):>14}"
            )
    if sin_verificar:
        print(f"  sin verificar        : {len(sin_verificar)} (falta --facturas)")
    sin_causal = sum(1 for r in resultados for i in r.items if not i.causales)
    if sin_causal:
        print(f"Renglones sin causal escrita: {sin_causal}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
