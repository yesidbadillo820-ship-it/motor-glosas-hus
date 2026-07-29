"""organizar_objeciones_emssanar.py — Arma el Excel OBJECIONES (cargue DGH) desde los
PDFs de objeción de EMSSANAR ("Objeción a Factura N° HUS..." generados por ripslink.app).

Es el equivalente EMSSANAR del consolidado que ya se arma para COOSALUD: toma uno o
varios PDFs "DETALLES_DE_SERVICIOS_FACTURA_HUS*.pdf" y produce un Excel con la hoja
OBJECIONES en el MISMO formato de 16 columnas del lote de cargue
(OBJECIONES_LOTE_*.xlsx) que se sube al sistema de cartera del HUS:

  CDCONSEC   consecutivo del documento (uno por factura, en orden de factura)
  CDFECDOC   fecha del documento (fecha de objeción del PDF, o --fecha)
  CRNCXC     factura normalizada HUS + 10 dígitos (HUS515948 → HUS0000515948)
  CROFECOBJ  fecha de la objeción (igual a CDFECDOC)
  CROREFERE / CROOBSERV / CRNCLAOBJ / IDRIPS / CTNCENCOS   vacías (igual que el lote)
  CROCLAOBJ  0
  GENUSUARIO4  usuario que carga (default 999, --usuario)
  CRNCONOBJ  código de objeción del Manual Único (TA0801, FA0205, CL0801, ...)
  SLNSERPRO  código del servicio/producto en DGH (CUPS homologado, ver abajo)
  CROVALOBJ  valor objetado en pesos (entero)
  CRDOBSERV  observación completa: código + texto estándar + (servicio) + nota EPS + $valor
  CROTIPOBJ  0 = Glosa, 2 = Devolución (del encabezado "Tipo objeción" del PDF, --tipobj)

REGLAS DE ARMADO (verificadas contra el PDF real y el lote COOSALUD de ejemplo):

  * Una fila por renglón de la tabla "Detalle glosas" del PDF (tecnología × código).

  * FUSIÓN DE DOBLE GLOSA: cuando la EPS objeta el MISMO servicio con una glosa de
    valor total (Cantidad Objetada numérica) Y una de diferencia tarifaria
    (Cantidad Objetada "--"), el total del PDF solo cuenta la mayor. Esas parejas se
    fusionan en UNA fila: CRNCONOBJ/CROVALOBJ = componente de mayor valor y el
    CRDOBSERV lista ambos códigos con su $valor (igual que las filas multi-código
    del lote COOSALUD). Sin esta fusión la suma no cuadra con el "Valor Objetado"
    del encabezado del PDF.

  * Códigos secundarios sin valor propio (bloques apilados bajo el mismo renglón,
    p.ej. SO0601 + FA0603 sobre una misma cánula) van en la misma fila, concatenados
    en CRDOBSERV.

  * HOMOLOGACIÓN DGH: algunos servicios existen en DGH con sufijo "H"
    (876802 → 876802H). La tabla CUPS_A_DGH se derivó del lote real
    OBJECIONES_LOTE_03 (145 códigos); los códigos no listados van tal cual.
    Desactivable con --sin-sufijo-h.

  * VALIDACIÓN: por cada PDF se compara la suma de CROVALOBJ contra el
    "Valor Objetado" del encabezado. Si no cuadra se avisa fuerte (y con
    --estricto el PDF se rechaza).

USO:
    # Un PDF
    py organizar_objeciones_emssanar.py --pdf "DETALLES_DE_SERVICIOS_FACTURA_HUS0000515948.pdf"

    # Carpeta completa (busca DETALLES_DE_SERVICIOS_FACTURA_*.pdf recursivo)
    py organizar_objeciones_emssanar.py --pdf "D:\\OBJECIONES\\EMSSANAR\\JULIO" ^
        --salida "OBJECIONES_EMSSANAR_LOTE_01.xlsx" --consec-inicial 1

INSTALACIÓN (una vez):
    py -m pip install pdfplumber openpyxl
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

# Lector de pesos compartido por los bots de pagador (ver `_dinero.py`).
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _dinero import a_entero  # noqa: E402

logger = logging.getLogger("objeciones_emssanar")

# ─── Layout de la tabla del PDF (bandas X por columna, calibradas del PDF real) ──────
#
# Columnas: Tecnología | Cantidad | Valor Tecnología | Cantidad Objetada |
#           Valor Objetado | Código Objeción | Observación
# El corte se decide por el CENTRO de cada palabra, así una palabra larga que
# arranca pegada al borde no se cae a la columna siguiente.
BANDAS_X = (140, 180, 272, 320, 395, 588)

# IGNORECASE: un código en minúsculas también debe abrir registro/componente —
# si no, sus valores se funden silenciosamente en el registro anterior.
RE_CODIGO_TEC = re.compile(r"^[A-Z]{0,3}\d[\dA-Z.\-]*$", re.IGNORECASE)  # 890701, FMQ0923, 129A02
RE_CODIGO_OBJ = re.compile(r"^[A-Z]{2}\d{4}$", re.IGNORECASE)  # TA0801, FA0205, CL0801...
RE_DINERO = re.compile(r"^\$[\d.,]+$")
RE_PIE_PAGINA = re.compile(r"^Página \d+ de \d+|Reporte generado por", re.IGNORECASE)

# Campos del encabezado de la página 1
RE_FACTURA = re.compile(r"Objeci[oó]n a Factura N[°º]?\s*([A-Z]{2,4}\s?\d+)", re.IGNORECASE)
RE_TIPO_OBJECION = re.compile(r"Tipo objeci[oó]n:\s*(\S+)", re.IGNORECASE)
RE_FECHA_OBJECION = re.compile(r"Fecha de Objeci[oó]n:\s*(\d{2})-(\d{2})-(\d{4})", re.IGNORECASE)
RE_VALOR_FACTURA = re.compile(r"Valor Factura:\s*\$([\d.,]+)", re.IGNORECASE)
RE_VALOR_OBJETADO = re.compile(r"Valor Objetado:\s*\$([\d.,]+)", re.IGNORECASE)

COLUMNAS_OBJECIONES = [
    "CDCONSEC",
    "CDFECDOC",
    "CRNCXC",
    "CROFECOBJ",
    "CROREFERE",
    "CROOBSERV",
    "CROCLAOBJ",
    "CRNCLAOBJ",
    "GENUSUARIO4",
    "CRNCONOBJ",
    "SLNSERPRO",
    "IDRIPS",
    "CTNCENCOS",
    "CROVALOBJ",
    "CRDOBSERV",
    "CROTIPOBJ",
]

# ─── Homologación CUPS → código de servicio DGH ──────────────────────────────────────
#
# En DGH varios servicios existen con sufijo "H" (código institucional del HUS).
# Esta tabla se derivó del lote real OBJECIONES_LOTE_03_LISTO (7.869 filas cargadas):
# todo CUPS que en ese lote apareció como SLNSERPRO con "H" está acá. En los 6 códigos
# que aparecieron en AMBAS formas se tomó la mayoritaria (890283H, 906841H, 881701H,
# 895001 → 879301 y 340401 quedaron en la forma plana por mayoría/empate).
# Los códigos que no están en la tabla se emiten tal cual vienen en el PDF.
CUPS_A_DGH: dict[str, str] = {
    "017202": "017202H",
    "11484": "11484H",
    "129A02": "129A02H",
    "130A02": "130A02H",
    "161101": "161101H",
    "168402": "168402H",
    "19124": "19124H",
    "19297": "19297H",
    "20212": "20212H",
    "20216": "20216H",
    "30109": "30109H",
    "311302": "311302H",
    "324102": "324102H",
    "324104": "324104H",
    "332203": "332203H",
    "342101": "342101H",
    "345102": "345102H",
    "359501": "359501H",
    "371202": "371202H",
    "385320": "385320H",
    "389102": "389102H",
    "511001": "511001H",
    "514301": "514301H",
    "518902": "518902H",
    "542802": "542802H",
    "545001": "545001H",
    "652302": "652302H",
    "751101": "751101H",
    "753601": "753601H",
    "861801": "861801H",
    "864103": "864103H",
    "867001": "867001H",
    "867003": "867003H",
    "869501": "869501H",
    "876241": "876241H",
    "876802": "876802H",
    "878101": "878101H",
    "878301": "878301H",
    "881232": "881232H",
    "881332": "881332H",
    "881401": "881401H",
    "881434": "881434H",
    "881435": "881435H",
    "881443": "881443H",
    "881701": "881701H",
    "882101": "882101H",
    "882221": "882221H",
    "882302": "882302H",
    "882305": "882305H",
    "883101": "883101H",
    "883102": "883102H",
    "883210": "883210H",
    "883531": "883531H",
    "890183": "890183H",
    "890201": "890201H",
    "890202": "890202H",
    "890203": "890203H",
    "890205": "890205H",
    "890261": "890261H",
    "890264": "890264H",
    "890266": "890266H",
    "890268": "890268H",
    "890274": "890274H",
    "890276": "890276H",
    "890278": "890278H",
    "890283": "890283H",
    "890301": "890301H",
    "890302": "890302H",
    "890303": "890303H",
    "890305": "890305H",
    "890361": "890361H",
    "890364": "890364H",
    "890366": "890366H",
    "890368": "890368H",
    "890374": "890374H",
    "890376": "890376H",
    "890378": "890378H",
    "890383": "890383H",
    "890401": "890401H",
    "890402": "890402H",
    "890403": "890403H",
    "890405": "890405H",
    "890461": "890461H",
    "890464": "890464H",
    "890466": "890466H",
    "890468": "890468H",
    "890474": "890474H",
    "890476": "890476H",
    "890478": "890478H",
    "890483": "890483H",
    "890601": "890601H",
    "890703": "890703H",
    "893815": "893815H",
    "895100": "895100H",
    "898001": "898001H",
    "898101": "898101H",
    "901123": "901123H",
    "902208": "902208H",
    "902209": "902209H",
    "903423": "903423H",
    "903806": "903806H",
    "903809": "903809H",
    "903818": "903818H",
    "903825": "903825H",
    "903828": "903828H",
    "903833": "903833H",
    "903841": "903841H",
    "903846": "903846H",
    "903856": "903856H",
    "903861": "903861H",
    "903866": "903866H",
    "903868": "903868H",
    "903895": "903895H",
    "904508": "904508H",
    "904509": "904509H",
    "904516": "904516H",
    "904520": "904520H",
    "904603": "904603H",
    "904712": "904712H",
    "904902": "904902H",
    "904915": "904915H",
    "905102": "905102H",
    "905203": "905203H",
    "906249": "906249H",
    "906317": "906317H",
    "906323": "906323H",
    "906423": "906423H",
    "906625": "906625H",
    "906841": "906841H",
    "906915": "906915H",
    "906917": "906917H",
    "907002": "907002H",
    "907106": "907106H",
    "908455": "908455H",
    "908856": "908856H",
    "908857": "908857H",
    "911004": "911004H",
    "916013": "916013H",
    "923201": "923201H",
    "930860": "930860H",
    "931000": "931000H",
    "932001": "932001H",
    "932002": "932002H",
    "933101": "933101H",
    "933302": "933302H",
    "933401": "933401H",
    "935200": "935200H",
    "935300": "935300H",
    "937000": "937000H",
    "938201": "938201H",
    "938302": "938302H",
    "997102": "997102H",
}


def _exigir_pdfplumber():
    try:
        import pdfplumber
    except ImportError:
        sys.stderr.write(
            "ERROR: falta pdfplumber.\n    py -m pip install pdfplumber\n",
        )
        sys.exit(2)
    return pdfplumber


def _exigir_openpyxl():
    try:
        import openpyxl
    except ImportError:
        sys.stderr.write(
            "ERROR: falta openpyxl.\n    py -m pip install openpyxl\n",
        )
        sys.exit(2)
    return openpyxl


# ─── Helpers de parseo ────────────────────────────────────────────────────────────────


def _a_entero(texto: str) -> int | None:
    """'1.365' → 1365, '2,177,341' → 2177341, '1.365,50' → 1365 (pesos enteros).

    La regla vive ahora en `tools/_dinero.py`, compartida con los bots de SAVIA
    y VCO: un separador de miles lleva tres dígitos detrás; uno o dos son
    decimales. Este bot ya la aplicaba bien; el de SAVIA no, y por eso sacaba
    los valores con centavos multiplicados por cien.

    Se conserva el `None` para "no había número": el resto del bot lo usa para
    distinguir una celda vacía de un cero declarado."""
    s = (texto or "").strip()
    if not re.search(r"\d", s):
        return None
    return a_entero(s)


def parsear_dinero(texto: str) -> int | None:
    """'$114.900' → 114900 (los puntos/comas son separadores de miles)."""
    if not texto or not RE_DINERO.match(texto.strip()):
        return None
    return _a_entero(texto)


def normalizar_factura(cruda: str) -> str:
    """'HUS515948' / 'HUS 515948' → 'HUS0000515948' (prefijo + 10 dígitos).

    Es el formato de CRNCXC del lote de cargue (igual que el lote COOSALUD)."""
    s = re.sub(r"\s+", "", (cruda or "").upper())
    m = re.match(r"^([A-Z]+)0*(\d+)$", s)
    if not m:
        return s
    return f"{m.group(1)}{int(m.group(2)):010d}"


def separar_codigo_descripcion(texto: str) -> tuple[str, str]:
    """'890701 - CONSULTA DE URGENCIAS...' → ('890701', 'CONSULTA DE URGENCIAS...')."""
    partes = (texto or "").split(" - ", 1)
    if len(partes) == 2:
        return partes[0].strip(), partes[1].strip()
    return (texto or "").strip(), ""


def homologar_servicio(codigo: str, aplicar_sufijo_h: bool = True) -> str:
    if not aplicar_sufijo_h:
        return codigo
    return CUPS_A_DGH.get(codigo, codigo)


def _limpiar_texto(palabras: list[str]) -> str:
    return re.sub(r"\s+", " ", " ".join(palabras)).strip()


# ─── Parseo del encabezado (página 1) ────────────────────────────────────────────────


def parsear_encabezado(texto_pagina1: str) -> dict:
    """Extrae factura, tipo, fecha de objeción y valores del bloque superior del PDF."""
    enc: dict = {}
    m = RE_FACTURA.search(texto_pagina1)
    if m:
        enc["factura"] = normalizar_factura(m.group(1))
    m = RE_TIPO_OBJECION.search(texto_pagina1)
    if m:
        enc["tipo_objecion"] = m.group(1).strip()
    m = RE_FECHA_OBJECION.search(texto_pagina1)
    if m:
        dia, mes, anio = (int(g) for g in m.groups())
        enc["fecha_objecion"] = datetime(anio, mes, dia)
    m = RE_VALOR_FACTURA.search(texto_pagina1)
    if m:
        enc["valor_factura"] = _a_entero(m.group(1))
    m = RE_VALOR_OBJETADO.search(texto_pagina1)
    if m:
        enc["valor_objetado"] = _a_entero(m.group(1))
    return enc


def tipobj_desde_encabezado(tipo_objecion: str) -> int:
    """'Glosa' → 0, 'Devolución' → 2 (CROTIPOBJ del lote de cargue)."""
    s = unicodedata.normalize("NFKD", (tipo_objecion or "").lower())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return 2 if "devolu" in s else 0


# ─── Parseo de la tabla "Detalle glosas" ─────────────────────────────────────────────
#
# La tabla se reconstruye desde las PALABRAS del PDF (posición X → columna), porque
# las celdas envuelven en varias líneas, los registros cruzan páginas y pdfplumber
# no detecta el borde entre "Código Objeción" y "Observación".
#
# Cada REGISTRO (renglón lógico) arranca en la línea donde la columna Tecnología
# trae "CODIGO -". Un registro puede traer VARIOS componentes de objeción: bloques
# "XX9999 - ..." apilados en la columna Código Objeción (el segundo bloque no repite
# tecnología ni valores → mismo renglón, código adicional).


# Vocabulario del encabezado de la tabla. Una línea compuesta SOLO por estas
# palabras es encabezado (o su segunda línea envuelta, "Tecnología Objetada") y se
# descarta; las líneas de datos siempre traen además códigos, montos o texto libre.
_PALABRAS_ENCABEZADO = {
    "Tecnología",
    "Cantidad",
    "Valor",
    "Objetada",
    "Objetado",
    "Código",
    "Objeción",
    "Observación",
}


def _es_linea_encabezado(linea: list[dict]) -> bool:
    return bool(linea) and {w["text"] for w in linea} <= _PALABRAS_ENCABEZADO


def _columna_de(palabra: dict) -> int:
    centro = (palabra["x0"] + palabra["x1"]) / 2
    for i, borde in enumerate(BANDAS_X):
        if centro < borde:
            return i
    return len(BANDAS_X)


def _lineas_de_pagina(page) -> list[list[dict]]:
    """Agrupa las palabras de la página en líneas visuales (por coordenada top)."""
    por_top: dict[int, list[dict]] = defaultdict(list)
    for w in page.extract_words():
        por_top[round(w["top"])].append(w)
    return [sorted(por_top[t], key=lambda w: w["x0"]) for t in sorted(por_top)]


def _nuevo_registro() -> dict:
    return {
        "tecnologia": [],
        "cantidad": None,
        "valor_tecnologia": None,
        "cantidad_objetada": None,
        "valor_objetado": None,
        "componentes": [],  # [{"palabras_cod": [...], "palabras_obs": [...]}]
    }


def extraer_registros(pdf) -> list[dict]:
    """Recorre todas las páginas y devuelve los registros crudos de la tabla."""
    registros: list[dict] = []
    actual: dict | None = None

    for page in pdf.pages:
        lineas = _lineas_de_pagina(page)

        # Tope del PRIMER encabezado de la tabla (si la página lo trae): se salta
        # todo lo que esté a esa altura o por encima (título, NIT, valores, etc.).
        # Se usa el primero a propósito: si una plantilla repitiera el encabezado a
        # mitad de página, con el último se perderían los registros de arriba. Las
        # repeticiones se filtran igual porque toda línea que sea puro vocabulario
        # de encabezado se descarta (_es_linea_encabezado).
        tope_encabezado = None
        for linea in lineas:
            if _es_linea_encabezado(linea) and any(w["text"] == "Código" for w in linea):
                tope_encabezado = max(w["top"] for w in linea)
                break

        for linea in lineas:
            texto_linea = " ".join(w["text"] for w in linea)
            if RE_PIE_PAGINA.search(texto_linea):
                continue
            if tope_encabezado is not None and linea[0]["top"] <= tope_encabezado + 13:
                continue
            if _es_linea_encabezado(linea):
                continue
            # Bloque de firmas al final ("Auditor Principal:", correos): sin columnas
            # de tecnología ni código, se corta ahí.
            if texto_linea.startswith(("Auditor Principal", "Notificado por", "Correo:")):
                if actual:
                    registros.append(actual)
                    actual = None
                continue

            por_col: dict[int, list[str]] = defaultdict(list)
            for w in linea:
                por_col[_columna_de(w)].append(w["text"])

            tec = por_col.get(0, [])
            es_inicio = len(tec) >= 2 and tec[1] == "-" and RE_CODIGO_TEC.match(tec[0])
            if es_inicio:
                if actual:
                    registros.append(actual)
                actual = _nuevo_registro()
            if actual is None:
                continue  # texto suelto antes del primer registro

            actual["tecnologia"] += tec
            if por_col.get(1) and actual["cantidad"] is None and por_col[1][0].isdigit():
                actual["cantidad"] = int(por_col[1][0])
            if por_col.get(2) and actual["valor_tecnologia"] is None:
                actual["valor_tecnologia"] = parsear_dinero(por_col[2][0])
            if por_col.get(3) and actual["cantidad_objetada"] is None:
                actual["cantidad_objetada"] = por_col[3][0]
            if por_col.get(4) and actual["valor_objetado"] is None:
                actual["valor_objetado"] = parsear_dinero(por_col[4][0])

            cod = por_col.get(5, [])
            obs = por_col.get(6, [])
            # ¿Arranca un componente nuevo? ("TA0801 -" al inicio de la columna código)
            if len(cod) >= 2 and cod[1] == "-" and RE_CODIGO_OBJ.match(cod[0]):
                actual["componentes"].append({"palabras_cod": [], "palabras_obs": []})
            if (cod or obs) and not actual["componentes"]:
                actual["componentes"].append({"palabras_cod": [], "palabras_obs": []})
            if actual["componentes"]:
                actual["componentes"][-1]["palabras_cod"] += cod
                actual["componentes"][-1]["palabras_obs"] += obs

    if actual:
        registros.append(actual)
    return registros


def consolidar_registro(reg: dict) -> dict:
    """Registro crudo → renglón limpio con tecnología y componentes separados."""
    tec_codigo, tec_desc = separar_codigo_descripcion(_limpiar_texto(reg["tecnologia"]))
    componentes = []
    for comp in reg["componentes"]:
        cod, texto = separar_codigo_descripcion(_limpiar_texto(comp["palabras_cod"]))
        componentes.append(
            {
                # upper(): DGH y la tabla CUPS_A_DGH trabajan en mayúsculas
                "codigo": cod.upper(),
                "texto": texto,
                "observacion": _limpiar_texto(comp["palabras_obs"]),
            }
        )
    return {
        "tec_codigo": tec_codigo.upper(),
        "tec_desc": tec_desc,
        "cantidad": reg["cantidad"],
        "valor_tecnologia": reg["valor_tecnologia"],
        "cantidad_objetada": (reg["cantidad_objetada"] or "").strip(),
        "valor_objetado": reg["valor_objetado"],
        "componentes": componentes,
    }


# ─── Fusión de doble glosa (valor total + diferencia tarifaria) ──────────────────────


def fusionar_dobles_glosas(renglones: list[dict]) -> list[dict]:
    """Cuando el mismo servicio trae una glosa de valor TOTAL (Cantidad Objetada
    numérica) y otra de DIFERENCIA tarifaria ("--"), la EPS solo cuenta la mayor en
    su "Valor Objetado". Se fusionan de a pares (misma tecnología) y el renglón
    resultante toma CRNCONOBJ/CROVALOBJ del componente de MAYOR valor — normalmente
    el total, pero puede ser la diferencia (p.ej. total de 1 unidad vs diferencia
    sobre todas las unidades). Ambos códigos quedan en CRDOBSERV con su $valor."""
    por_tec: dict[str, dict[str, list[int]]] = defaultdict(lambda: {"total": [], "dif": []})
    for i, r in enumerate(renglones):
        clave = "total" if r["cantidad_objetada"].isdigit() else "dif"
        por_tec[r["tec_codigo"]][clave].append(i)

    def _valor(idx: int) -> int:
        return renglones[idx]["valor_objetado"] or 0

    # Emparejar por afinidad de valor (mayor con mayor), no por posición: con varios
    # totales de valores distintos, el orden del documento cruzaría los pares.
    pares: list[tuple[int, int]] = []
    for grupos in por_tec.values():
        totales = sorted(grupos["total"], key=_valor, reverse=True)
        difs = sorted(grupos["dif"], key=_valor, reverse=True)
        # strict=False a propósito: se emparejan tantos como haya de cada lado
        pares += list(zip(totales, difs, strict=False))

    absorbido_por: dict[int, int] = {}  # idx secundario → idx principal que lo absorbe
    for idx_total, idx_dif in pares:
        if _valor(idx_total) >= _valor(idx_dif):
            absorbido_por[idx_dif] = idx_total
        else:
            absorbido_por[idx_total] = idx_dif

    fusionados: list[dict] = []
    extras: dict[int, list[dict]] = defaultdict(list)
    for sec, prin in absorbido_por.items():
        extras[prin].append(renglones[sec])
    for i, r in enumerate(renglones):
        if i in absorbido_por:
            continue
        if i in extras:
            r = dict(r)
            r["fusionados"] = extras[i]
        fusionados.append(r)
    return fusionados


# ─── Armado de filas OBJECIONES ──────────────────────────────────────────────────────


def _texto_componente(comp: dict, tec_codigo: str, tec_desc: str, valor: int | None) -> str:
    """Un componente → 'COD TEXTO ESTANDAR (SERVICIO): NOTA EPS$VALOR' (estilo del lote)."""
    partes = [comp["codigo"], comp["texto"]]
    txt = " ".join(p for p in partes if p)
    txt += f" ({tec_codigo}-{tec_desc})" if tec_desc else f" ({tec_codigo})"
    if comp["observacion"]:
        txt += f": {comp['observacion']}"
    if valor is not None:
        txt += f"${valor}"
    return txt.upper()


def armar_crdobserv(renglon: dict) -> str:
    """CRDOBSERV: componente principal + secundarios sin valor + renglones fusionados."""
    lineas = []
    comps = renglon["componentes"]
    for i, comp in enumerate(comps):
        # Solo el primer componente del renglón lleva el valor (los apilados no traen)
        valor = renglon["valor_objetado"] if i == 0 else None
        lineas.append(_texto_componente(comp, renglon["tec_codigo"], renglon["tec_desc"], valor))
    for extra in renglon.get("fusionados", []):
        for i, comp in enumerate(extra["componentes"]):
            valor = extra["valor_objetado"] if i == 0 else None
            lineas.append(_texto_componente(comp, extra["tec_codigo"], extra["tec_desc"], valor))
    return "\n".join(lineas)


def renglon_a_fila(
    renglon: dict,
    consec: int,
    factura: str,
    fecha: datetime,
    usuario: str,
    tipobj: int,
    aplicar_sufijo_h: bool = True,
) -> dict:
    """Renglón consolidado → dict con las 16 columnas del lote OBJECIONES."""
    comps = renglon["componentes"]
    principal = comps[0] if comps else {"codigo": "", "texto": "", "observacion": ""}
    return {
        "CDCONSEC": str(consec),
        "CDFECDOC": fecha,
        "CRNCXC": factura,
        "CROFECOBJ": fecha,
        "CROREFERE": None,
        "CROOBSERV": None,
        "CROCLAOBJ": 0,
        "CRNCLAOBJ": None,
        "GENUSUARIO4": usuario,
        "CRNCONOBJ": principal["codigo"],
        "SLNSERPRO": homologar_servicio(renglon["tec_codigo"], aplicar_sufijo_h),
        "IDRIPS": None,
        "CTNCENCOS": None,
        "CROVALOBJ": renglon["valor_objetado"] or 0,
        "CRDOBSERV": armar_crdobserv(renglon),
        "CROTIPOBJ": tipobj,
    }


# ─── Procesamiento de un PDF completo ────────────────────────────────────────────────


def procesar_pdf(ruta: Path, aplicar_sufijo_h: bool = True) -> dict:
    """Devuelve {"encabezado": {...}, "renglones": [...], "cuadra": bool}."""
    pdfplumber = _exigir_pdfplumber()
    with pdfplumber.open(str(ruta)) as pdf:
        texto_p1 = pdf.pages[0].extract_text() or ""
        encabezado = parsear_encabezado(texto_p1)
        registros = extraer_registros(pdf)

    renglones = [consolidar_registro(r) for r in registros]
    incompletos = [r for r in renglones if r["valor_objetado"] is None or not r["componentes"]]
    if incompletos:
        for r in incompletos:
            logger.warning(
                f"  renglón incompleto en {ruta.name}: {r['tec_codigo']} {r['tec_desc'][:40]}"
            )
    renglones = fusionar_dobles_glosas(renglones)

    suma = sum(r["valor_objetado"] or 0 for r in renglones)
    esperado = encabezado.get("valor_objetado")
    cuadra = esperado is None or suma == esperado
    return {
        "encabezado": encabezado,
        "renglones": renglones,
        "suma": suma,
        "cuadra": cuadra,
    }


def buscar_pdfs(rutas: list[str]) -> list[Path]:
    """Expande archivos y carpetas (recursivo, DETALLES_DE_SERVICIOS_FACTURA_*.pdf)."""
    encontrados: list[Path] = []
    for r in rutas:
        p = Path(r)
        if p.is_dir():
            encontrados += sorted(p.rglob("DETALLES_DE_SERVICIOS_FACTURA_*.pdf"))
            encontrados += sorted(p.rglob("DETALLES_DE_SERVICIOS_FACTURA_*.PDF"))
        elif p.is_file():
            encontrados.append(p)
        else:
            logger.warning(f"ruta inexistente: {r}")
    # dedup preservando orden
    vistos: set[Path] = set()
    unicos = []
    for p in encontrados:
        if p not in vistos:
            vistos.add(p)
            unicos.append(p)
    return unicos


# ─── Escritura del Excel ─────────────────────────────────────────────────────────────

_FORMATO_TEXTO = "@"
_FORMATO_FECHA = "mm-dd-yy"
_FORMATO_MILES = '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-'


def escribir_excel(filas: list[dict], salida: Path) -> None:
    """Hoja OBJECIONES con las 16 columnas y los mismos formatos del lote de ejemplo."""
    openpyxl = _exigir_openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OBJECIONES"
    ws.append(COLUMNAS_OBJECIONES)

    formatos = {
        "CDFECDOC": _FORMATO_FECHA,
        "CROFECOBJ": _FORMATO_FECHA,
        "CROCLAOBJ": "General",
        "CROVALOBJ": _FORMATO_MILES,
        "CROTIPOBJ": "0",
    }
    for fila in filas:
        ws.append([fila[c] for c in COLUMNAS_OBJECIONES])
        for idx, col in enumerate(COLUMNAS_OBJECIONES, start=1):
            celda = ws.cell(row=ws.max_row, column=idx)
            celda.number_format = formatos.get(col, _FORMATO_TEXTO)
    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(salida))


# ─── Main ────────────────────────────────────────────────────────────────────────────


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(
        description="Arma el Excel OBJECIONES (cargue DGH) desde PDFs de objeción EMSSANAR."
    )
    ap.add_argument(
        "--pdf",
        nargs="+",
        required=True,
        help="PDF(s) o carpeta(s) con DETALLES_DE_SERVICIOS_FACTURA_*.pdf",
    )
    ap.add_argument(
        "--salida",
        default="OBJECIONES_EMSSANAR.xlsx",
        help="Excel de salida (default OBJECIONES_EMSSANAR.xlsx)",
    )
    ap.add_argument(
        "--consec-inicial",
        type=int,
        default=1,
        help="CDCONSEC de la primera factura (default 1)",
    )
    ap.add_argument(
        "--fecha",
        default=None,
        help="Forzar CDFECDOC/CROFECOBJ (DD-MM-AAAA). Default: la fecha de objeción de cada PDF",
    )
    ap.add_argument("--usuario", default="999", help="GENUSUARIO4 (default 999)")
    ap.add_argument(
        "--tipobj",
        type=int,
        choices=(0, 2),
        default=None,
        help="Forzar CROTIPOBJ. Default: 0 si el PDF dice Glosa, 2 si dice Devolución",
    )
    ap.add_argument(
        "--sin-sufijo-h",
        action="store_true",
        help="No homologar CUPS al código DGH con sufijo H (emitir tal cual el PDF)",
    )
    ap.add_argument(
        "--estricto",
        action="store_true",
        help="Rechazar los PDFs cuya suma no cuadre con el Valor Objetado del encabezado",
    )
    args = ap.parse_args(argv)

    logging.basicConfig(level=logging.INFO, format="%(message)s", stream=sys.stdout)

    fecha_forzada = None
    if args.fecha:
        try:
            dia, mes, anio = (int(x) for x in args.fecha.split("-"))
            fecha_forzada = datetime(anio, mes, dia)
        except ValueError:
            sys.stderr.write(f"ERROR: --fecha inválida: {args.fecha} (esperaba DD-MM-AAAA)\n")
            return 2

    pdfs = buscar_pdfs(args.pdf)
    if not pdfs:
        sys.stderr.write("ERROR: no encontré ningún PDF para procesar.\n")
        return 2
    logger.info(f"PDFs a procesar: {len(pdfs)}")

    resultados = []
    con_error = 0
    sin_cuadrar = 0
    for ruta in pdfs:
        try:
            res = procesar_pdf(ruta, aplicar_sufijo_h=not args.sin_sufijo_h)
        except Exception as e:
            logger.error(f"✗ {ruta.name}: no se pudo parsear ({e})")
            con_error += 1
            continue
        enc = res["encabezado"]
        factura = enc.get("factura") or normalizar_factura(ruta.stem.split("_")[-1])
        estado = "OK" if res["cuadra"] else "NO CUADRA"
        logger.info(
            f"{'✓' if res['cuadra'] else '⚠'} {factura}: {len(res['renglones'])} renglones, "
            f"suma ${res['suma']:,} vs encabezado "
            f"${enc.get('valor_objetado', 0):,} → {estado}"
        )
        if not res["cuadra"]:
            sin_cuadrar += 1
            if args.estricto:
                logger.error(f"  descartado por --estricto: {ruta.name}")
                con_error += 1
                continue
        res["factura"] = factura
        resultados.append(res)

    if not resultados:
        sys.stderr.write("ERROR: ningún PDF quedó procesable.\n")
        return 1

    # Si dos PDFs traen la MISMA factura (copias en subcarpetas), se queda el primero
    vistas: set[str] = set()
    unicos = []
    for res in resultados:
        if res["factura"] in vistas:
            logger.warning(f"⚠ factura repetida {res['factura']}: se ignora el PDF duplicado")
            continue
        vistas.add(res["factura"])
        unicos.append(res)
    resultados = unicos

    # Orden por número de factura (igual que el lote de ejemplo) y consecutivo por factura
    def _clave_factura(res):
        m = re.search(r"(\d+)$", res["factura"])
        return (int(m.group(1)) if m else 0, res["factura"])

    resultados.sort(key=_clave_factura)

    filas: list[dict] = []
    for i, res in enumerate(resultados):
        consec = args.consec_inicial + i
        enc = res["encabezado"]
        fecha = fecha_forzada or enc.get("fecha_objecion") or datetime.now()
        tipobj = args.tipobj
        if tipobj is None:
            tipobj = tipobj_desde_encabezado(enc.get("tipo_objecion", ""))
        for renglon in res["renglones"]:
            filas.append(
                renglon_a_fila(
                    renglon,
                    consec=consec,
                    factura=res["factura"],
                    fecha=fecha,
                    usuario=args.usuario,
                    tipobj=tipobj,
                    aplicar_sufijo_h=not args.sin_sufijo_h,
                )
            )

    salida = Path(args.salida)
    escribir_excel(filas, salida)
    total = sum(f["CROVALOBJ"] for f in filas)
    logger.info(
        f"\nListo: {salida} — {len(filas)} filas, {len(resultados)} facturas, "
        f"total objetado ${total:,}"
    )
    if con_error:
        logger.warning(f"({con_error} PDF(s) con problemas — revisar arriba)")
    if sin_cuadrar and not args.estricto:
        logger.warning(
            f"(⚠ {sin_cuadrar} factura(s) NO CUADRAN con el encabezado del PDF: "
            "se incluyeron igual — revisar antes de cargar)"
        )
    # exit != 0 también cuando alguna factura no cuadró: que el aviso no pase
    # desapercibido en corridas automatizadas (el Excel igual queda escrito).
    return 0 if not con_error and not sin_cuadrar else 1


if __name__ == "__main__":
    sys.exit(main())
