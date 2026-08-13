#!/usr/bin/env python3
"""siifa_novedades.py — Qué llegó nuevo a SIIFA desde la última revisión.

El portal dice cuántos registros hay en total («Mostrando 2611 a 2620 de 2620
registros») pero no cuáles son nuevos: hay que ir página por página buscando
lo que no estaba. Este script compara el informe recién bajado contra el
anterior y responde, para cada novedad: qué ENTIDAD glosó o devolvió, qué
FACTURA, si es GLOSA o DEVOLUCIÓN, por cuánto y con qué causal.

CÓMO SE USA (dos pasos):

    REM 1) Bajar el informe de hoy (el de siempre, opción [1] del bot)
    py siifa_reporte_seguimientos.py --salida "D:\\SIIFA\\informe_HOY.xlsx"

    REM 2) Comparar contra el que ya se tenía
    py siifa_novedades.py --nuevo "D:\\SIIFA\\informe_HOY.xlsx" ^
                          --anterior "D:\\SIIFA\\informe_seguimientos.xlsx"

Si no se tiene a mano el informe anterior, se puede correr sólo con --nuevo:
en ese caso lista todo lo que está SIN RESPUESTA, que —estando ya cargadas
todas las respuestas del corte pasado— es justamente lo que acaba de entrar.

    py siifa_novedades.py --nuevo "D:\\SIIFA\\informe_HOY.xlsx"

Deja un Excel (NOVEDADES_SIIFA.xlsx) con una hoja por novedad y un resumen
por entidad, y muestra el detalle en pantalla para poder leerlo de una vez.
"""

from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from _dinero import a_entero  # noqa: E402

# Las que el auditor necesita ver de una novedad, en el orden en que las lee.
COLUMNAS_VISTA = [
    "razon_social_eps",
    "nit_eps",
    "numero_factura",
    "tipo_seguimiento",
    "valor_glosa",
    "codigo_glosa",
    "descripcion_glosa",
    "fecha_formulacion",
    "tiene_respuesta",
    "id_seguimiento_factura_glosa",
]

TITULOS = {
    "razon_social_eps": "Entidad (EPS)",
    "nit_eps": "NIT",
    "numero_factura": "Factura",
    "tipo_seguimiento": "Tipo",
    "valor_glosa": "Valor",
    "codigo_glosa": "Causal",
    "descripcion_glosa": "Descripción de la causal",
    "fecha_formulacion": "Fecha",
    "tiene_respuesta": "¿Respondida?",
    "id_seguimiento_factura_glosa": "Id seguimiento",
}


def leer_informe(ruta: Path) -> list[dict]:
    """Lee un informe de seguimientos (el que produce siifa_reporte_seguimientos)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("ERROR: falta openpyxl. py -m pip install openpyxl") from None

    if not ruta.exists():
        raise SystemExit(f"ERROR: no encuentro el archivo:\n  {ruta}\n")

    wb = load_workbook(str(ruta), read_only=True, data_only=True)
    ws = wb["SEGUIMIENTOS"] if "SEGUIMIENTOS" in wb.sheetnames else wb.worksheets[0]
    filas_crudas = ws.iter_rows(values_only=True)
    try:
        encabezado = [str(c or "").strip() for c in next(filas_crudas)]
    except StopIteration:
        raise SystemExit(f"ERROR: el archivo está vacío:\n  {ruta}\n") from None

    if "id_seguimiento_factura_glosa" not in encabezado:
        raise SystemExit(
            f"ERROR: {ruta.name} no parece un informe de seguimientos de SIIFA.\n"
            "Se espera el Excel que produce la opción [1] del bot "
            "(siifa_reporte_seguimientos.py).\n"
        )

    filas = []
    for valores in filas_crudas:
        # strict=False a propósito: una fila más corta que el encabezado se
        # completa con lo que haya, en vez de tumbar todo el informe.
        fila = {k: ("" if v is None else v) for k, v in zip(encabezado, valores, strict=False)}
        if str(fila.get("id_seguimiento_factura_glosa", "")).strip():
            filas.append(fila)
    wb.close()
    return filas


def _id(fila: dict) -> str:
    return str(fila.get("id_seguimiento_factura_glosa", "")).strip()


def comparar(nuevas: list[dict], anteriores: list[dict] | None) -> dict:
    """Separa lo que llegó nuevo de lo que ya estaba.

    Sin informe anterior no se puede saber qué es «nuevo» de verdad, así que
    se devuelve como novedad lo que está sin responder — que es lo accionable
    y, con el corte anterior ya cargado al 100%, viene a ser lo mismo.
    """
    sin_respuesta = [f for f in nuevas if str(f.get("tiene_respuesta", "")).strip().upper() != "SI"]
    if anteriores is None:
        return {
            "modo": "sin-respuesta",
            "novedades": sin_respuesta,
            "sin_respuesta": sin_respuesta,
            "desaparecidos": [],
            "total_nuevo": len(nuevas),
            "total_anterior": None,
        }

    ids_antes = {_id(f) for f in anteriores}
    ids_ahora = {_id(f) for f in nuevas}
    novedades = [f for f in nuevas if _id(f) not in ids_antes]
    # Un seguimiento que estaba y ya no está no debería pasar nunca; si pasa,
    # el auditor tiene que saberlo (puede ser un informe anterior incompleto).
    desaparecidos = [f for f in anteriores if _id(f) not in ids_ahora]
    return {
        "modo": "comparacion",
        "novedades": novedades,
        "sin_respuesta": sin_respuesta,
        "desaparecidos": desaparecidos,
        "total_nuevo": len(nuevas),
        "total_anterior": len(anteriores),
    }


def es_devolucion(fila: dict) -> bool:
    return str(fila.get("tipo_seguimiento", "")).strip().upper().startswith("DEVOL")


def valor_total(filas: list[dict]) -> int:
    """Cuánto vale un grupo de seguimientos, contando bien las devoluciones.

    ÚSESE ESTA FUNCIÓN SIEMPRE que haya que sumar valores de seguimientos de
    SIIFA. Sumar `valor_glosa` a secas está mal en cuanto haya una devolución
    de por medio, y el error no se nota: da una cifra grande y creíble. Ver la
    explicación en resumen_por_entidad().
    """
    total = 0
    facturas_devueltas: dict[str, int] = {}
    for f in filas:
        valor = a_entero(f.get("valor_glosa"))
        if es_devolucion(f):
            factura = str(f.get("numero_factura") or "")
            facturas_devueltas[factura] = max(facturas_devueltas.get(factura, 0), valor)
        else:
            total += valor
    return total + sum(facturas_devueltas.values())


def resumen_por_entidad(filas: list[dict]) -> list[tuple]:
    """Agrupa por entidad y tipo: cuántos ítems, cuántas facturas y cuánto vale.

    OJO CON EL VALOR DE LAS DEVOLUCIONES. Una glosa objeta un ítem y trae el
    valor de ese ítem: sumar los ítems da el valor glosado. Una devolución
    rechaza la factura ENTERA, y SIIFA repite el valor de la factura en cada
    una de sus líneas. Sumarlas multiplica la plata: las 1.340 líneas de
    SANITAS daban $24.917 millones cuando el valor real de esas 9 facturas es
    $111 millones —224 veces más—. Por eso en devoluciones el valor se cuenta
    una sola vez por factura.
    """
    agg: dict[tuple, list] = defaultdict(list)
    for f in filas:
        clave = (
            str(f.get("razon_social_eps") or f.get("nit_eps") or "SIN ENTIDAD IDENTIFICADA"),
            str(f.get("tipo_seguimiento") or "?"),
        )
        agg[clave].append(f)
    resumen = [
        (
            entidad,
            tipo,
            len(grupo),
            len({str(f.get("numero_factura") or "") for f in grupo}),
            valor_total(grupo),
        )
        for (entidad, tipo), grupo in agg.items()
    ]
    return sorted(resumen, key=lambda r: -r[4])


def _pesos(valor: object) -> str:
    return f"${a_entero(valor):,}".replace(",", ".")


def mostrar(res: dict, limite: int) -> None:
    novedades = res["novedades"]
    print()
    print("=" * 78)
    if res["modo"] == "comparacion":
        print(
            f"SIIFA tiene ahora {res['total_nuevo']} seguimientos; "
            f"el informe anterior tenía {res['total_anterior']}."
        )
    else:
        print(f"SIIFA tiene {res['total_nuevo']} seguimientos en total.")
        print("(Sin informe anterior para comparar: muestro lo que está SIN RESPONDER.)")
    print("=" * 78)

    if not novedades:
        print("\nNo hay novedades: no llegó nada nuevo y no queda nada sin responder.\n")
        return

    print(f"\nNOVEDADES: {len(novedades)} seguimiento(s).\n")
    print("RESUMEN POR ENTIDAD")
    print("-" * 78)
    print(f"{'Entidad':<32}{'Tipo':<12}{'Ítems':>7}{'Facturas':>9}{'Valor':>18}")
    print("-" * 78)
    total_items = total_valor = 0
    for entidad, tipo, items, facturas, valor in resumen_por_entidad(novedades):
        print(f"{entidad[:31]:<32}{tipo:<12}{items:>7}{facturas:>9}{_pesos(valor):>18}")
        total_items += items
        total_valor += valor
    print("-" * 78)
    print(f"{'TOTAL':<44}{total_items:>7}{'':>9}{_pesos(total_valor):>18}")
    if any(es_devolucion(f) for f in novedades):
        print(
            "\nNota: en las devoluciones el valor es el de la FACTURA COMPLETA y se\n"
            "cuenta una sola vez, aunque la devolución tenga muchas líneas."
        )

    print("\nDETALLE")
    print("-" * 78)
    print(f"{'Factura':<14}{'Tipo':<12}{'Entidad':<26}{'Valor':>13}  Causal")
    print("-" * 78)
    for f in novedades[:limite]:
        entidad = str(f.get("razon_social_eps") or f.get("nit_eps") or "")[:25]
        print(
            f"{str(f.get('numero_factura') or ''):<14}"
            f"{str(f.get('tipo_seguimiento') or ''):<12}"
            f"{entidad:<26}"
            f"{_pesos(f.get('valor_glosa')):>13}  "
            f"{str(f.get('codigo_glosa') or '')}"
        )
    if len(novedades) > limite:
        print(f"... y {len(novedades) - limite} más. El detalle completo queda en el Excel.")

    if res["desaparecidos"]:
        print(
            f"\nATENCIÓN: {len(res['desaparecidos'])} seguimiento(s) que estaban en el "
            "informe anterior ya no aparecen. Puede que el informe anterior haya\n"
            "quedado incompleto. Están en la hoja YA_NO_ESTAN del Excel."
        )
    print()


def escribir_xlsx(res: dict, ruta: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise SystemExit("ERROR: falta openpyxl. py -m pip install openpyxl") from None

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    dev_fill = PatternFill("solid", fgColor="FFF2CC")

    wb = Workbook()
    ws = wb.active
    ws.title = "NOVEDADES"
    for c, h in enumerate(COLUMNAS_VISTA, 1):
        cell = ws.cell(row=1, column=c, value=TITULOS.get(h, h))
        cell.fill = header_fill
        cell.font = header_font
    for r, fila in enumerate(res["novedades"], start=2):
        for c, h in enumerate(COLUMNAS_VISTA, 1):
            valor = a_entero(fila.get(h)) if h == "valor_glosa" else fila.get(h, "")
            cell = ws.cell(row=r, column=c, value=valor)
            cell.alignment = Alignment(vertical="top", wrap_text=(h == "descripcion_glosa"))
            if h == "valor_glosa":
                cell.number_format = '"$"#,##0'
        # Las devoluciones se responden por otra puerta: conviene verlas aparte.
        if str(fila.get("tipo_seguimiento", "")).upper().startswith("DEVOL"):
            for c in range(1, len(COLUMNAS_VISTA) + 1):
                ws.cell(row=r, column=c).fill = dev_fill
    anchos = {"razon_social_eps": 32, "descripcion_glosa": 45, "numero_factura": 14}
    for c, h in enumerate(COLUMNAS_VISTA, 1):
        ws.column_dimensions[get_column_letter(c)].width = anchos.get(h, 16)
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("RESUMEN")
    encabezado = ["Entidad (EPS)", "Tipo", "Ítems", "Facturas", "Valor"]
    ws2.append(encabezado)
    for c in range(1, len(encabezado) + 1):
        ws2.cell(row=1, column=c).fill = header_fill
        ws2.cell(row=1, column=c).font = header_font
    for entidad, tipo, items, facturas, valor in resumen_por_entidad(res["novedades"]):
        ws2.append([entidad, tipo, items, facturas, valor])
        ws2.cell(row=ws2.max_row, column=5).number_format = '"$"#,##0'
    for col, ancho in (("A", 34), ("B", 14), ("C", 10), ("D", 12), ("E", 18)):
        ws2.column_dimensions[col].width = ancho

    if res["desaparecidos"]:
        ws3 = wb.create_sheet("YA_NO_ESTAN")
        for c, h in enumerate(COLUMNAS_VISTA, 1):
            cell = ws3.cell(row=1, column=c, value=TITULOS.get(h, h))
            cell.fill = header_fill
            cell.font = header_font
        for r, fila in enumerate(res["desaparecidos"], start=2):
            for c, h in enumerate(COLUMNAS_VISTA, 1):
                ws3.cell(row=r, column=c, value=fila.get(h, ""))
        for col, ancho in (("A", 34), ("C", 14)):
            ws3.column_dimensions[col].width = ancho

    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ruta))
    print(f"Archivo con las novedades: {ruta}\n")


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--nuevo", required=True, help="Informe de seguimientos recién bajado.")
    ap.add_argument("--anterior", help="Informe de la revisión pasada, para comparar.")
    ap.add_argument("--salida", help="Excel de salida (default: NOVEDADES_SIIFA.xlsx al lado).")
    ap.add_argument(
        "--limite", type=int, default=40, help="Cuántas novedades mostrar en pantalla (default 40)."
    )
    args = ap.parse_args()

    nuevo = Path(args.nuevo)
    nuevas = leer_informe(nuevo)
    anteriores = leer_informe(Path(args.anterior)) if args.anterior else None

    res = comparar(nuevas, anteriores)
    mostrar(res, args.limite)

    if res["novedades"] or res["desaparecidos"]:
        salida = Path(args.salida) if args.salida else nuevo.parent / "NOVEDADES_SIIFA.xlsx"
        escribir_xlsx(res, salida)


if __name__ == "__main__":
    main()
