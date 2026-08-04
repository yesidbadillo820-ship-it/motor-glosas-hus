"""verificar_detallado_ajustado.py — revisa el Excel que dejó el ajustador.

Vuelve a **leer el archivo de salida** de `ajustar_detallado_glosas.py` y lo
contrasta contra el archivo original, el consolidado y el reporte del ADRES.
Es una segunda opinión: no confía en lo que el bot dijo que hizo, sino en lo
que quedó escrito en el archivo.

Comprueba, factura por factura:

 1. que no quede ni rastro del membrete ni del título viejo;
 2. que las facturas que quedan sean **exactamente** las del consolidado que
    estaban en el archivo original (ni una de más, ni una de menos);
 3. que la suma de los renglones == subtotal == total de la orden;
 4. que ese subtotal coincida con lo que el **reporte del ADRES** dice que
    sigue glosado, descontando lo que la bitácora ya declara como excepción;
 5. que el total en letras corresponda al número.

USO (Windows, desde C:\\temp-notas):

    py tools\\verificar_detallado_ajustado.py ^
        --ajustado  "D:\\...\\DETALLADOS_31068_AJUSTADO.xlsx" ^
        --original  "D:\\...\\DETALLADOS PAQUETE 31068.xlsx" ^
        --consolidado "D:\\...\\CONSOLIDADO.xlsx" ^
        --reporte-glosas "D:\\...\\ReporteGlosasReclamPAQUETE 31068.xlsx" ^
        --bitacora "D:\\...\\bitacora_31068.csv"

Devuelve 0 si no encontró fallas y 1 si encontró alguna.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
import ajustar_detallado_glosas as aj  # noqa: E402

RE_FAC = re.compile(r"HUS\s?0*\d{4,12}", re.I)


def bloques_de_fila(ws, anclas, fila):
    out = []
    for c in ws[fila]:
        if c.value is None:
            continue
        fmax, cmax = anclas.get((fila, c.column), (fila, c.column))
        out.append((c.column, cmax, c.value, fmax))
    out.sort()
    return out


def ajustes_conocidos(bitacora: Path) -> dict[str, float]:
    """Diferencias que la bitácora ya declara (no son errores del bot)."""
    import csv as _csv

    out: dict[str, float] = {}
    with bitacora.open(encoding="utf-8-sig") as fh:
        for row in _csv.DictReader(fh, delimiter=";"):
            fac = aj.normalizar_factura(row["FACTURA"])
            ent = float(row["VR_ENT_ORIGINAL"] or 0)
            glo = float(row["VALOR_GLOSADO"] or 0)
            if row["ACCION"] == "GLOSA_SIN_ITEM":
                out[fac] = out.get(fac, 0.0) + glo
            elif row["ACCION"] == "SIN_CRUCE":
                out[fac] = out.get(fac, 0.0) - ent
            elif row["ACCION"] in ("CONSERVADO", "AJUSTADO") and glo > ent + 1:
                out[fac] = out.get(fac, 0.0) + (glo - ent)
            if "no suma al total" in (row.get("OBSERVACION") or ""):
                # El reporte glosa el renglón principal Y su desglose; la
                # factura solo cobra el principal.
                out[fac] = out.get(fac, 0.0) + float(row["VR_ENT_NUEVO"] or 0)
    return out


def renglones_que_no_suman(bitacora: Path) -> dict[str, tuple[float, int]]:
    """Renglones de desglose que la bitácora declara que no suman al total."""
    import csv as _csv

    out: dict[str, tuple[float, int]] = {}
    with bitacora.open(encoding="utf-8-sig") as fh:
        for row in _csv.DictReader(fh, delimiter=";"):
            if "no suma al total" not in (row.get("OBSERVACION") or ""):
                continue
            fac = aj.normalizar_factura(row["FACTURA"])
            valor, n = out.get(fac, (0.0, 0))
            out[fac] = (valor + float(row["VR_ENT_NUEVO"] or 0), n + 1)
    return out


def verificar(
    salida: Path, original: Path, glosas: dict, quiero: set, ajustes=None, no_suman=None
) -> list[str]:
    ajustes = ajustes or {}
    no_suman = no_suman or {}
    fallas: list[str] = []
    wb = openpyxl.load_workbook(salida, data_only=True)
    ws = wb.worksheets[0]
    anclas = {(r.min_row, r.min_col): (r.max_row, r.max_col) for r in ws.merged_cells.ranges}

    filas = {}
    for f in range(1, ws.max_row + 1):
        b = bloques_de_fila(ws, anclas, f)
        if b:
            filas[f] = b
    texto_todo = " || ".join(" ".join(str(v) for _, _, v, _ in b) for b in filas.values()).upper()

    # 1) membrete y título viejo
    for prohibido in ("CUFE", "CARRRERA 33", "NIT 900.006.037", "PÁGINA 1/1", "PAGINA 1/1"):
        if prohibido in texto_todo:
            fallas.append(f"quedó rastro del membrete: {prohibido!r}")
    for b in filas.values():
        for _, _, v, _ in b:
            if aj._norm(v).startswith("FACTURA ELECTRONICA DE"):
                fallas.append("quedó el título viejo 'FACTURA ELECTRONICA DE'")
                break

    # 2) facturas presentes
    inicios = []
    for f, b in sorted(filas.items()):
        if any(aj._norm(str(v)).startswith(aj.TITULO_NUEVO) for _, _, v, _ in b):
            m = RE_FAC.search(" ".join(str(v) for _, _, v, _ in b))
            inicios.append((f, aj.normalizar_factura(m.group(0)) if m else "??"))
    presentes = {k for _, k in inicios}

    wb0 = openpyxl.load_workbook(original, read_only=True, data_only=True)
    ws0 = wb0.worksheets[0]
    orig = set()
    for row in ws0.iter_rows(values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue
        t = " ".join(str(v) for v in vals)
        if "FACTURA ELECTRONICA" in t.upper():
            m = RE_FAC.search(t)
            if m:
                orig.add(aj.normalizar_factura(m.group(0)))
    wb0.close()

    esperadas = orig & quiero
    if presentes != esperadas:
        fallas.append(
            f"facturas: sobran {sorted(presentes - esperadas)[:5]} / "
            f"faltan {sorted(esperadas - presentes)[:5]}"
        )

    # 3-6) por factura
    fin = ws.max_row
    for i, (f0, fac) in enumerate(inicios):
        f1 = inicios[i + 1][0] - 1 if i + 1 < len(inicios) else fin
        hdr = sub = -1
        cabeceras = None
        for f in range(f0, f1 + 1):
            b = filas.get(f, [])
            if hdr < 0:
                cab = aj._cabeceras_tabla([aj.Celda(f, fm, c0, c1, v) for c0, c1, v, fm in b])
                if cab:
                    hdr, cabeceras = f, cab
                    continue
            if sub < 0 and any(aj._norm(str(v)).startswith(aj.MARCA_SUBTOTAL) for _, _, v, _ in b):
                sub = f
        if hdr < 0 or sub < 0:
            fallas.append(f"{fac}: no se reconoce la estructura tras el ajuste")
            continue

        suma = 0.0
        n_items = 0
        # Renglones que la bitácora declara que NO suman (desglose cuyo renglón
        # principal sigue en la factura): se descuentan de lo que debe cuadrar.
        no_suma, n_no_suma = no_suman.get(fac, (0.0, 0))
        for f in range(hdr + 1, sub):
            b = filas.get(f, [])
            if not b:
                continue
            campos = aj.mapear_por_solapamiento(
                cabeceras, [aj.Celda(f, fm, c0, c1, v) for c0, c1, v, fm in b]
            )
            cant = aj._parse_valor(campos["cantidad"].valor) if "cantidad" in campos else 0
            ent = aj._parse_valor(campos["vr_ent"].valor) if "vr_ent" in campos else 0
            if cant > 0 and ent > 0:
                # Los renglones de desglose (sin consecutivo) no vuelven a
                # sumar cuando el renglón que los encabeza sigue en la factura.
                suma += ent
                n_items += 1

        suma -= no_suma
        n_items -= n_no_suma
        nums = [v for _, _, v, _ in filas.get(sub, []) if isinstance(v, (int, float))]
        if not nums or abs(nums[-1] - suma) > 1:
            fallas.append(
                f"{fac}: subtotal {nums[-1] if nums else '?'} != suma de ítems {suma:.0f}"
            )
        if len(nums) > 1 and nums[-2] != n_items:
            fallas.append(f"{fac}: la fila de subtotal dice {nums[-2]} ítems y quedaron {n_items}")

        # Lo que DEBE quedar = glosado de los ítems del reporte,
        #   - menos las glosas que no tienen ítem en el detallado,
        #   + los ítems del detallado que no cruzaron (se conservan),
        #   - el exceso de las glosas que superan el valor del detallado.
        items_rep, _reclam = aj.agrupar_reporte(glosas.get(fac, []))
        glosado_rep = sum(r.valor_glosado for r in items_rep) - ajustes.get(fac, 0.0)
        if glosado_rep and abs(glosado_rep - suma) > 1:
            fallas.append(
                f"{fac}: subtotal {suma:.0f} != glosado de ítems del reporte {glosado_rep:.0f} "
                f"(dif {suma - glosado_rep:+.0f})"
            )

        for f in range(sub, f1 + 1):
            b = filas.get(f, [])
            if b and aj._norm(str(b[0][2])).startswith(aj.MARCA_TOTAL_ORDEN):
                nums = [v for _, _, v, _ in b if isinstance(v, (int, float))]
                if nums and abs(nums[-1] - suma) > 1:
                    fallas.append(f"{fac}: total de orden {nums[-1]} != subtotal {suma:.0f}")
            if b and aj._norm(str(b[0][2])).startswith(aj.MARCA_TOTAL_LETRAS):
                letras = next((v for _, _, v, _ in b[1:] if isinstance(v, str)), "")
                if letras != aj.numero_a_letras(suma):
                    fallas.append(f"{fac}: letras {letras!r} != {aj.numero_a_letras(suma)!r}")
    wb.close()
    return fallas


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    p.add_argument("--ajustado", type=Path, required=True, help="Excel que dejó el ajustador.")
    p.add_argument("--original", type=Path, required=True, help="El detallado sin tocar.")
    p.add_argument("--consolidado", type=Path, required=True, help="Facturas a trabajar.")
    p.add_argument("--reporte-glosas", type=Path, required=True)
    p.add_argument("--bitacora", type=Path, help="Bitácora CSV de la corrida (recomendado).")
    p.add_argument("--limite", type=int, default=40, help="Cuántas fallas listar.")
    args = p.parse_args(argv)

    glosas = aj.leer_reporte_glosas(args.reporte_glosas)
    lista, _ = aj.leer_consolidado(args.consolidado)
    quiero = {aj.normalizar_factura(f) for f in lista}
    ajustes = ajustes_conocidos(args.bitacora) if args.bitacora else {}
    no_suman = renglones_que_no_suman(args.bitacora) if args.bitacora else {}
    fallas = verificar(args.ajustado, args.original, glosas, quiero, ajustes, no_suman)

    print(f"\n{'SIN FALLAS' if not fallas else f'{len(fallas)} FALLA(S)'} en {args.ajustado.name}")
    for f in fallas[: args.limite]:
        print("  -", f)
    if len(fallas) > args.limite:
        print(f"  ... y {len(fallas) - args.limite} más")
    return 1 if fallas else 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
