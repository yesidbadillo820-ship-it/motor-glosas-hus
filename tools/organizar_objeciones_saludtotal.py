"""organizar_objeciones_saludtotal.py — Arma las objeciones de SALUD TOTAL en
el formato de trabajo de 16 columnas (hoja OBJECIONES), listo para el cargue
masivo de recepción.

Toma el Excel de notificación de glosas de SALUD TOTAL (export
"NotificacionGLS_…": 6 columnas — NumeroFac_, NombreServicio, CantidadFac,
ValorGlosaTotalxServ, Observaciones, CodMotvGlosaEspc) y lo convierte al mismo
layout de 16 columnas que usan SAVIA, FAMISANAR y el Dispensario:

    CDCONSEC | CDFECDOC | CRNCXC | CROFECOBJ | CROREFERE | CROOBSERV | CROCLAOBJ |
    CRNCLAOBJ | GENUSUARIO4 | CRNCONOBJ | SLNSERPRO | IDRIPS | CTNCENCOS |
    CROVALOBJ | CRDOBSERV | CROTIPOBJ

Particularidades de SALUD TOTAL:
  - La factura viene SIN prefijo (464306): se completa a HUS0000464306
    (`--prefijo-factura` para cambiar el prefijo).
  - NO trae código de servicio, solo el NOMBRE. SLNSERPRO queda vacío, salvo
    que se pase `--maestro` (un Excel de referencia — p. ej. un OBJECIONES ya
    trabajado tipo LOTE_02, o un listado código|nombre — con el que se
    homologa por NOMBRE normalizado, sin inventar: lo que no matchea exacto
    queda vacío y se reporta).
  - Algunos textos vienen con el encoding dañado («Ã“» en vez de «Ó»):
    se reparan automáticamente.

MAPEO (SALUD TOTAL → 16 columnas):
    CRNCXC       ← NumeroFac_           (464306 → HUS0000464306)
    CRNCONOBJ    ← CodMotvGlosaEspc     (ya viene de 6: CL0801, TA0601…)
    SLNSERPRO    ← vacío | homologado por nombre con --maestro
    CROVALOBJ    ← ValorGlosaTotalxServ (lector único tools/_dinero.py)
    CRDOBSERV    ← "<CRNCONOBJ> <Observaciones>$<valor>" (regla del formato)
    CDFECDOC / CROFECOBJ ← --fecha (default: hoy), FECHA CORTA
    CDCONSEC     ← consecutivo POR FACTURA (1-1-1, 2-2-2…), como texto
    CROTIPOBJ    ← por factura: solo TA/FA/SO/AU/CO→0, solo CL→1, mezcla→2
    CROCLAOBJ=0, GENUSUARIO4='999' (texto)  |  resto de columnas vacías

USO:
    py organizar_objeciones_saludtotal.py \
        --entrada "PARA_MASIVO.xlsx" \
        --salida  "OBJECIONES_SALUDTOTAL"      # carpeta (o .xlsx con --consolidado)

INSTALACIÓN (una vez):
    py -m pip install openpyxl
"""

from __future__ import annotations

import argparse
import datetime as _dt
import json
import logging
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

# Un solo lector de pesos para todos los bots (tools/_dinero.py).
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _dinero import a_entero  # noqa: E402

logger = logging.getLogger("organizar_saludtotal")


# ─── Formato de salida: layout de 16 columnas (hoja OBJECIONES) ──────────────

COLUMNAS_DISPENSARIO: tuple[str, ...] = (
    "CDCONSEC",
    "CDFECDOC",
    "CRNCXC",
    "CROFECOBJ",
    "CROREFERE",
    "CROOBSERV",
    "CROCLAOBJ",
    "CRNCLAOBJ",
    "GENUSUARIO4",
    "CRNCONOBJ",
    "SLNSERPRO",
    "IDRIPS",
    "CTNCENCOS",
    "CROVALOBJ",
    "CRDOBSERV",
    "CROTIPOBJ",
)

CDCONSEC_DEFAULT = 1
CROCLAOBJ_CONST = 0
GENUSUARIO4_CONST = "999"
PREFIJO_FACTURA_DEFAULT = "HUS"

# number_format por columna, copiado 1:1 de los archivos reales (EMSSANAR).
FORMATOS_DISPENSARIO: dict[str, str] = {
    "CDCONSEC": "@",
    "CDFECDOC": "mm-dd-yy",
    "CRNCXC": "@",
    "CROFECOBJ": "mm-dd-yy",
    "CROREFERE": "@",
    "CROOBSERV": "@",
    "CROCLAOBJ": "General",
    "CRNCLAOBJ": "@",
    "GENUSUARIO4": "@",
    "CRNCONOBJ": "@",
    "SLNSERPRO": "@",
    "IDRIPS": "@",
    "CTNCENCOS": "@",
    "CROVALOBJ": '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-',
    "CRDOBSERV": "@",
    "CROTIPOBJ": "0",
}


# ─── Lectura del Excel de SALUD TOTAL (6 columnas) ───────────────────────────


def _norm_header(h: object) -> str:
    """Normaliza un encabezado: mayúsculas, sin tildes, sin espacios de más."""
    s = unicodedata.normalize("NFKD", str(h or "").strip().upper())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.split())


# Alias del export de SALUD TOTAL (tolerante a variaciones).
COLUMNAS_SALUDTOTAL = {
    "factura": {
        "NUMEROFAC_",
        "NUMEROFAC",
        "NUMERO FAC",
        "NRO_FACTURA",
        "NUMERO_FACTURA",
        "FACTURA",
    },
    "servicio": {"NOMBRESERVICIO", "NOMBRE SERVICIO", "SERVICIO", "DESCRIPCION SERVICIO"},
    "cantidad": {"CANTIDADFAC", "CANTIDAD FAC", "CANTIDAD"},
    "valor": {
        "VALORGLOSATOTALXSERV",
        "VALOR GLOSA TOTAL X SERV",
        "VALOR GLOSA",
        "VALOR_GLOSA",
        "VALOR",
    },
    "observacion": {"OBSERVACIONES", "OBSERVACION", "OBSERVACION GLOSA"},
    "motivo": {"CODMOTVGLOSAESPC", "COD MOTV GLOSA ESPC", "CODIGO GLOSA", "CODIGO_GLOSA", "MOTIVO"},
}
IDX_FALLBACK = {
    "factura": 0,
    "servicio": 1,
    "cantidad": 2,
    "valor": 3,
    "observacion": 4,
    "motivo": 5,
}


def _resolver_columnas(headers: list[str]) -> dict[str, int]:
    """Devuelve {clave: indice} mapeando por nombre de encabezado; si una
    columna no se reconoce por nombre, usa el índice fijo de respaldo."""
    norm = [_norm_header(h) for h in headers]
    idx: dict[str, int] = {}
    for clave, alias in COLUMNAS_SALUDTOTAL.items():
        encontrado = next((i for i, h in enumerate(norm) if h in alias), None)
        if encontrado is None:
            encontrado = IDX_FALLBACK[clave]
            logger.debug(
                f"  columna '{clave}' no reconocida por nombre; uso índice fijo {encontrado}"
            )
        idx[clave] = encontrado
    return idx


def _cell(row: tuple, idx: dict[str, int], clave: str) -> object:
    """Valor de la columna `clave` en `row`, o None si el índice se sale de rango."""
    i = idx[clave]
    return row[i] if i < len(row) else None


def _num(v: object) -> int:
    """Pesos enteros, con el lector único de `tools/_dinero.py` (regla
    colombiana: separador de miles lleva tres dígitos detrás; uno o dos son
    decimales — así "1.365,50" es 1365 y no 136550)."""
    return a_entero(v)


# ─── Reparación de textos con encoding dañado ────────────────────────────────

_MARCAS_MOJIBAKE = ("Ã", "Â", "â€")


def _bytes_originales(s: str) -> bytes | None:
    """Reconstruye los bytes originales de un texto mal decodificado. El export
    de SALUD TOTAL mezcla las dos variantes clásicas: caracteres latin-1
    directos (U+0081) y caracteres cp1252 (la comilla « " » que aparece dentro
    de «Ã“»), así que se convierte carácter por carácter."""
    out = bytearray()
    for ch in s:
        cp = ord(ch)
        if cp <= 0xFF:
            out.append(cp)
            continue
        try:
            out += ch.encode("cp1252")
        except UnicodeEncodeError:
            return None
    return bytes(out)


def reparar_texto(s: str) -> str:
    """Repara el mojibake típico de los exports de SALUD TOTAL («Ã“» → «Ó»,
    «CÃPSULA» → «CÁPSULA»). Solo repara si el texto trae las marcas Y la
    reconstrucción a UTF-8 funciona y reduce las marcas; si no, tal cual."""
    if not s or not any(m in s for m in _MARCAS_MOJIBAKE):
        return s
    crudo = _bytes_originales(s)
    if crudo is None:
        return s
    try:
        arreglado = crudo.decode("utf-8")
    except UnicodeDecodeError:
        return s
    # Solo aceptar si de verdad quitó marcas (no empeorar).
    if sum(arreglado.count(m) for m in _MARCAS_MOJIBAKE) < sum(
        s.count(m) for m in _MARCAS_MOJIBAKE
    ):
        return arreglado
    return s


# ─── Normalización de factura (número pelado → larga con prefijo) ────────────

_RE_FACTURA_LETRAS = re.compile(r"^([A-Za-z]+)0*(\d+)$")
_RE_FACTURA_NUM = re.compile(r"^0*(\d+)$")


def factura_larga(fac: object, prefijo: str = PREFIJO_FACTURA_DEFAULT, ancho: int = 10) -> str:
    """SALUD TOTAL entrega la factura SIN prefijo: 464306 → HUS0000464306.
    Si ya viene con letras (HUS464306 / HUS0000464306) se respeta el prefijo
    que traiga. Si no matchea ningún patrón, se devuelve tal cual."""
    s = str(fac or "").strip()
    m = _RE_FACTURA_NUM.match(s)
    if m:
        return prefijo.upper() + m.group(1).zfill(ancho)
    m = _RE_FACTURA_LETRAS.match(s)
    if m:
        return m.group(1).upper() + m.group(2).zfill(ancho)
    return s


# ─── Homologación opcional del servicio por NOMBRE (--maestro) ───────────────


def _norm_nombre(s: str) -> str:
    """Normaliza un nombre de servicio para el match: mayúsculas, sin tildes,
    sin signos, espacios colapsados."""
    s = unicodedata.normalize("NFKD", (s or "").upper())
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    return " ".join(s.split())


_RE_COD_NOMBRE = re.compile(r"\(([A-Z0-9.\-]{3,20})-(.{4,90}?)\)")


def cargar_maestro(ruta: Path) -> dict[str, str]:
    """Construye {nombre_normalizado: codigo} desde un Excel de referencia.

    Acepta dos formas:
      1. Un archivo OBJECIONES ya trabajado (p. ej. LOTE_02): saca los pares
         de los textos "(CODIGO-NOMBRE)" del CRDOBSERV cuyo código coincide
         con el SLNSERPRO de la fila.
      2. Un listado simple con columnas código | nombre (las dos primeras).
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(ruta), data_only=True, read_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    wb.close()
    if not filas:
        return {}

    headers = [_norm_header(h) for h in filas[0]]
    mapa: dict[str, str] = {}

    if "SLNSERPRO" in headers and "CRDOBSERV" in headers:
        i_cod = headers.index("SLNSERPRO")
        i_txt = headers.index("CRDOBSERV")
        for r in filas[1:]:
            cod = str(r[i_cod] or "").strip() if i_cod < len(r) else ""
            txt = str(r[i_txt] or "") if i_txt < len(r) else ""
            if not cod:
                continue
            for m in _RE_COD_NOMBRE.finditer(txt):
                if m.group(1).strip() == cod:
                    mapa.setdefault(_norm_nombre(m.group(2)), cod)
    else:
        # Listado simple: columna 0 = código, columna 1 = nombre.
        for r in filas[1:]:
            if len(r) < 2 or not r[0] or not r[1]:
                continue
            mapa.setdefault(_norm_nombre(str(r[1])), str(r[0]).strip())

    logger.info(f"  Maestro de servicios: {len(mapa)} nombres con código ({ruta.name})")
    return mapa


# ─── CROTIPOBJ: tipo de objeción por factura ─────────────────────────────────

GRUPO_CLINICO = "CL"


def crotipobj_factura(grupos: set[str]) -> int:
    """Solo administrativos (TA/FA/SO/AU/CO…) → 0; solo CL → 1; mezcla → 2."""
    tiene_cl = GRUPO_CLINICO in grupos
    tiene_admin = any(g != GRUPO_CLINICO for g in grupos)
    if tiene_cl and tiene_admin:
        return 2
    if tiene_cl:
        return 1
    return 0


# ─── Texto de la observación en el formato de trabajo ────────────────────────

_RE_VALOR_FINAL = re.compile(r"\$\s*([\d.,]+)\s*$")


def construir_crdobserv(codigo: str, observacion: str, valor: int) -> str:
    """CRDOBSERV = ``<código> <texto>$<valor>``. Normaliza espacios, repara el
    encoding dañado y evita duplicar: quita el código del inicio si ya viene y
    un ``$<monto>`` final SOLO si es el mismo valor de la objeción (un monto
    distinto es información real y se conserva)."""
    t = " ".join(reparar_texto(observacion or "").split())
    cod = (codigo or "").strip()
    if cod and t.upper().startswith(cod.upper()):
        t = t[len(cod) :].strip()
    m = _RE_VALOR_FINAL.search(t)
    if m and _num(m.group(1)) == valor:
        t = t[: m.start()].strip()
    prefijo = f"{cod} " if cod else ""
    return f"{prefijo}{t}${valor}"


# ─── Transformación SALUD TOTAL → 16 columnas ────────────────────────────────


def construir_registros(
    ruta: Path,
    fecha: _dt.datetime,
    consecutivo: int,
    prefijo_factura: str = PREFIJO_FACTURA_DEFAULT,
    maestro: dict[str, str] | None = None,
) -> list[dict]:
    """Lee el Excel de SALUD TOTAL y devuelve una lista de dicts, uno por
    objeción, con las 16 columnas del formato de trabajo. Si `maestro` viene,
    llena SLNSERPRO homologando el NOMBRE del servicio (match exacto
    normalizado — sin inventar); sin maestro, SLNSERPRO queda vacío."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.stderr.write("ERROR: falta openpyxl. py -m pip install openpyxl\n")
        sys.exit(2)

    wb = load_workbook(filename=str(ruta), data_only=True, read_only=True)
    ws = wb.active
    filas = ws.iter_rows(values_only=True)
    try:
        headers = list(next(filas))
    except StopIteration:
        logger.warning(f"  {ruta.name}: hoja vacía")
        return []
    idx = _resolver_columnas([str(h) for h in headers])

    consec_por_factura: dict[str, int] = {}
    con_codigo = 0
    sin_codigo: dict[str, int] = defaultdict(int)

    registros: list[dict] = []
    for r in filas:
        if r is None:
            continue
        factura_raw = str(_cell(r, idx, "factura") or "").strip()
        observacion = str(_cell(r, idx, "observacion") or "").strip()
        motivo = str(_cell(r, idx, "motivo") or "").strip().upper()
        if not factura_raw and not observacion and not motivo:
            continue

        crncxc = factura_larga(factura_raw, prefijo_factura)
        if crncxc not in consec_por_factura:
            consec_por_factura[crncxc] = consecutivo + len(consec_por_factura)
        valor = _num(_cell(r, idx, "valor"))
        nombre_serv = reparar_texto(str(_cell(r, idx, "servicio") or "").strip())

        cod_servicio = ""
        if maestro:
            cod_servicio = maestro.get(_norm_nombre(nombre_serv), "")
            if cod_servicio:
                con_codigo += 1
            else:
                sin_codigo[nombre_serv] += 1

        registros.append(
            {
                "CDCONSEC": str(consec_por_factura[crncxc]),
                "CDFECDOC": fecha,
                "CRNCXC": crncxc,
                "CROFECOBJ": fecha,
                "CROREFERE": None,
                "CROOBSERV": None,
                "CROCLAOBJ": CROCLAOBJ_CONST,
                "CRNCLAOBJ": None,
                "GENUSUARIO4": GENUSUARIO4_CONST,
                "CRNCONOBJ": motivo,
                "SLNSERPRO": cod_servicio or None,
                "IDRIPS": None,
                "CTNCENCOS": None,
                "CROVALOBJ": valor,
                "CRDOBSERV": construir_crdobserv(motivo, observacion, valor),
                "CROTIPOBJ": 0,  # placeholder: se calcula por factura abajo
            }
        )
    wb.close()

    # CROTIPOBJ por factura según la mezcla de conceptos.
    grupos_por_factura: dict[str, set[str]] = defaultdict(set)
    for reg in registros:
        grupos_por_factura[reg["CRNCXC"]].add(str(reg["CRNCONOBJ"])[:2].upper())
    for reg in registros:
        reg["CROTIPOBJ"] = crotipobj_factura(grupos_por_factura[reg["CRNCXC"]])

    if maestro:
        logger.info(
            f"  Homologación por nombre: {con_codigo} con código, "
            f"{sum(sin_codigo.values())} sin match (quedan vacíos)."
        )
        for nombre in sorted(sin_codigo)[:15]:
            logger.warning(f"      sin código: {nombre[:70]}")
    else:
        logger.info(
            "  SALUD TOTAL no trae código de servicio: SLNSERPRO queda vacío "
            "(usa --maestro para homologar por nombre)."
        )
    return registros


# ─── Escritura ───────────────────────────────────────────────────────────────


def _escribir_hoja(registros: list[dict], salida: Path) -> None:
    """Escribe UN .xlsx con hoja OBJECIONES y las 16 columnas del formato."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "OBJECIONES"

    fill_hdr = PatternFill("solid", fgColor="1F4E78")
    font_hdr = Font(bold=True, color="FFFFFF")
    for col, nombre in enumerate(COLUMNAS_DISPENSARIO, start=1):
        c = ws.cell(row=1, column=col, value=nombre)
        c.fill = fill_hdr
        c.font = font_hdr

    fila = 2
    for reg in registros:
        for col, nombre in enumerate(COLUMNAS_DISPENSARIO, start=1):
            celda = ws.cell(row=fila, column=col, value=reg.get(nombre))
            celda.number_format = FORMATOS_DISPENSARIO[nombre]
        fila += 1

    ws.freeze_panes = "A2"
    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(salida))


PREFIJO_DEFAULT = "OBJECIONES_SALUDTOTAL"


def escribir_por_factura(
    registros: list[dict],
    carpeta: Path,
    prefijo: str = PREFIJO_DEFAULT,
    consecutivo: int = CDCONSEC_DEFAULT,
) -> list[Path]:
    """Un archivo <prefijo>_<CRNCXC>.xlsx por factura. Standalone → CDCONSEC
    reinicia en `consecutivo` (texto), escribiendo COPIAS (no muta la lista)."""
    por_factura: dict[str, list[dict]] = defaultdict(list)
    for reg in registros:
        por_factura[reg["CRNCXC"]].append(reg)

    generados: list[Path] = []
    for crncxc, regs in sorted(por_factura.items()):
        regs_out = [{**reg, "CDCONSEC": str(consecutivo)} for reg in regs]
        destino = carpeta / f"{prefijo}_{crncxc}.xlsx"
        _escribir_hoja(regs_out, destino)
        generados.append(destino)
        logger.info(f"  {destino.name}: {len(regs)} objeciones")
    return generados


def escribir_consolidado(registros: list[dict], salida: Path) -> None:
    """Un único archivo con todas las facturas (CDCONSEC 1,2,3… por factura)."""
    _escribir_hoja(registros, salida)


# ─── CLI ─────────────────────────────────────────────────────────────────────


def _parse_fecha(texto: str | None) -> _dt.datetime:
    if not texto:
        hoy = _dt.date.today()
        return _dt.datetime(hoy.year, hoy.month, hoy.day)
    try:
        return _dt.datetime.strptime(texto.strip(), "%Y-%m-%d")
    except ValueError:
        logger.error(f"--fecha inválida: {texto!r} (usá YYYY-MM-DD, p.ej. 2026-08-20)")
        sys.exit(2)


def _cargar_mapa_json(ruta: Path | None) -> dict[str, str]:
    if ruta is None:
        return {}
    try:
        data = json.loads(ruta.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        logger.error(f"No pude leer el mapa {ruta}: {e}")
        sys.exit(2)
    if not isinstance(data, dict):
        logger.error('El mapa debe ser un objeto JSON {"NOMBRE": "CODIGO", ...}')
        sys.exit(2)
    return {_norm_nombre(str(k)): str(v).strip() for k, v in data.items()}


def _resumen(registros: list[dict]) -> None:
    facturas = defaultdict(int)
    codigos = defaultdict(int)
    total = 0
    for reg in registros:
        facturas[reg["CRNCXC"]] += 1
        codigos[reg["CRNCONOBJ"]] += 1
        total += reg["CROVALOBJ"]
    logger.info(f"  Facturas: {len(facturas)}  |  Objeciones: {len(registros)}")
    logger.info(f"  Valor glosado total: ${total:,.0f}")
    logger.info("  Códigos de objeción (CRNCONOBJ):")
    for cod, n in sorted(codigos.items(), key=lambda kv: (-kv[1], kv[0])):
        logger.info(f"    {cod or '(sin código)'}: {n}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        "--entrada",
        type=Path,
        required=True,
        help="Excel de SALUD TOTAL (6 columnas: NumeroFac_, NombreServicio, "
        "CantidadFac, ValorGlosaTotalxServ, Observaciones, CodMotvGlosaEspc).",
    )
    parser.add_argument(
        "--salida",
        type=Path,
        required=True,
        help="Carpeta destino (un archivo por factura), o .xlsx con --consolidado.",
    )
    parser.add_argument(
        "--prefijo",
        default=PREFIJO_DEFAULT,
        help=f"Prefijo del nombre de cada archivo por factura. Default: {PREFIJO_DEFAULT}.",
    )
    parser.add_argument(
        "--consolidado",
        action="store_true",
        help="Junta todo en un solo Excel (--salida es el .xlsx).",
    )
    parser.add_argument(
        "--fecha", default=None, help="Fecha para CDFECDOC/CROFECOBJ (YYYY-MM-DD). Default: hoy."
    )
    parser.add_argument(
        "--prefijo-factura",
        default=PREFIJO_FACTURA_DEFAULT,
        help="Prefijo que se antepone a la factura pelada de SALUD TOTAL "
        f"(464306 → {PREFIJO_FACTURA_DEFAULT}0000464306). Default: {PREFIJO_FACTURA_DEFAULT}.",
    )
    parser.add_argument(
        "--maestro",
        type=Path,
        default=None,
        help="Excel de referencia para homologar el código de servicio por NOMBRE "
        "(un OBJECIONES trabajado tipo LOTE_02, o un listado código|nombre).",
    )
    parser.add_argument(
        "--mapa-servicios",
        type=Path,
        default=None,
        help='JSON {"NOMBRE DEL SERVICIO": "CODIGO", ...} que se suma al maestro '
        "(y le gana en caso de choque).",
    )
    parser.add_argument(
        "--consecutivo",
        type=int,
        default=CDCONSEC_DEFAULT,
        help=f"Número inicial del consecutivo por factura (CDCONSEC). Default: {CDCONSEC_DEFAULT}.",
    )
    parser.add_argument("--log", type=Path, default=None, help="Guarda un log adicional a archivo.")
    args = parser.parse_args(argv)

    handlers: list[logging.Handler] = [logging.StreamHandler(sys.stdout)]
    if args.log is not None:
        args.log.parent.mkdir(parents=True, exist_ok=True)
        handlers.append(logging.FileHandler(args.log, encoding="utf-8"))
    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", handlers=handlers
    )

    if not args.entrada.is_file():
        logger.error(f"No existe el archivo de entrada: {args.entrada}")
        return 1

    fecha = _parse_fecha(args.fecha)
    maestro: dict[str, str] = {}
    if args.maestro is not None:
        if not args.maestro.is_file():
            logger.error(f"No existe el maestro: {args.maestro}")
            return 1
        maestro = cargar_maestro(args.maestro)
    maestro.update(_cargar_mapa_json(args.mapa_servicios))

    logger.info(f"Leyendo glosas de SALUD TOTAL: {args.entrada.name}")
    registros = construir_registros(
        args.entrada,
        fecha=fecha,
        consecutivo=args.consecutivo,
        prefijo_factura=args.prefijo_factura,
        maestro=maestro or None,
    )
    if not registros:
        logger.error("No se encontró ninguna objeción en el archivo de entrada.")
        return 1

    if args.consolidado:
        escribir_consolidado(registros, args.salida)
        logger.info(f"\nExcel de SALUD TOTAL (consolidado, formato de trabajo): {args.salida}")
    else:
        generados = escribir_por_factura(
            registros, args.salida, prefijo=args.prefijo, consecutivo=args.consecutivo
        )
        logger.info(f"\n{len(generados)} archivo(s) de SALUD TOTAL en: {args.salida}")

    _resumen(registros)
    return 0


if __name__ == "__main__":
    sys.exit(main())
