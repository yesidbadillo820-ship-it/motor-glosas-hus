"""organizar_objeciones_dispensario.py — PDF de auditoría del Dispensario → Excel OBJECIONES (DGH).

Lee uno o varios PDF "DETALLE DE AUDITORIA Y GLOSAS" que envía el auditor del
Dispensario Médico Bucaramanga (CONSORCIO AUDITOOL) y produce, por cada
factura glosada, un Excel con UNA FILA POR OBJECIÓN en el formato OBJECIONES
que se importa en Dinámica Gerencial (mismo layout del ejemplo
OBJECIONES_EMSSANAR_HUS<n>.xlsx: columnas CDCONSEC..CROTIPOBJ).

El PDF es una tabla multicolumna (PREFIJO | FACTURA | FECHA ATENCION |
NOMBRES | VALOR FACTURA | VALOR OBJETADO | CODIGO+CONCEPTO | MOTIVO). Las
columnas CONCEPTO y MOTIVO vienen pegadas en el texto plano, así que la
extracción se hace por coordenadas de carácter (pdfplumber) y no por regex
sobre el texto corrido.

Validación integrada: la suma de los valores objetados extraídos de cada
factura se compara contra el "Total Factura" que imprime el propio PDF; si no
cuadra, el script lo marca y termina con código de salida 1.

USO RÁPIDO
----------

    # Un PDF (genera un Excel por factura glosada, en la carpeta del PDF)
    py tools\\organizar_objeciones_dispensario.py --pdf "D:\\...\\AUDITORIA_GLOSA_...pdf"

    # Carpeta con varios PDFs + salida en otra carpeta
    py tools\\organizar_objeciones_dispensario.py ^
        --carpeta "D:\\...\\GLOSAS_2026\\DISPENSARIO_<fecha>" ^
        --salida-dir "D:\\...\\GLOSAS_2026\\DISPENSARIO_<fecha>\\OBJECIONES"

    # Todo consolidado en un solo Excel
    py tools\\organizar_objeciones_dispensario.py --pdf <pdf> --consolidado objeciones.xlsx

DEPENDENCIAS
------------
    py -m pip install pdfplumber openpyxl
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

logger = logging.getLogger("organizar_objeciones")

# ---------------------------------------------------------------------------
# Layout del PDF (coordenadas X de cada columna, página de 1029 pt de ancho).
# Un carácter pertenece a la columna donde cae el centro de su caja.
# ---------------------------------------------------------------------------
COLUMNAS = {
    "prefijo": (28.0, 68.0),
    "factura": (68.0, 113.0),
    "fecha": (113.0, 162.0),
    "nombre": (162.0, 330.0),
    "vfactura": (330.0, 382.0),
    "vobjetado": (382.0, 440.0),
    "concepto": (440.0, 659.5),
    "motivo": (659.5, 10_000.0),
}

# Banda vertical con datos (excluye cabecera de la tabla y pie de página).
Y_DATOS_MIN = 165.0
Y_DATOS_MAX = 592.0

# Código de glosa al inicio de la columna CONCEPTO: "CL03 01", "SO08 01", ...
RE_CODIGO_GLOSA = re.compile(r"^([A-Z]{2}\d{2})\s+(\d{2})\b\s*(.*)$", re.DOTALL)

# Prefijo de factura en la primera columna: "HUS" (2 a 5 letras).
RE_PREFIJO = re.compile(r"^[A-ZÑ]{2,5}$")

RE_FECHA = re.compile(r"\d{2}/\d{2}/\d{4}")
RE_VALOR = re.compile(r"[\d.]+,\d{2}")
RE_FECHA_IMPRESION = re.compile(r"Fecha\s+Impresi[oó]n:\s*(\d{2}/\d{2}/\d{2,4})")

# Servicio glosado dentro del motivo: "SE GLOSA CODIGO 21519H ..." o
# "SE GLOSA EL CODIGO 908337H ...".
RE_SERVICIO = re.compile(r"SE\s+GLOSA\s+(?:EL\s+)?CODIGO\s+([A-Z0-9]+)", re.IGNORECASE)

# Artefactos de extracción tipo "(cid:9)" (tabs embebidos en la fuente).
RE_CID = re.compile(r"\(cid:\d+\)")

ENCABEZADOS = [
    "CDCONSEC",
    "CDFECDOC",
    "CRNCXC",
    "CROFECOBJ",
    "CROREFERE",
    "CROOBSERV",
    "CROCLAOBJ",
    "CRNCLAOBJ",
    "GENUSUARIO4",
    "CRNCONOBJ",
    "SLNSERPRO",
    "IDRIPS",
    "CTNCENCOS",
    "CROVALOBJ",
    "CRDOBSERV",
    "CROTIPOBJ",
]

FMT_CONTABLE = '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-'
FMT_FECHA = "mm-dd-yy"


@dataclass
class Glosa:
    codigo: str  # CL0301, SO0801, ...
    concepto: str  # texto de la columna CONCEPTO DE GLOSA
    motivo: str  # texto de la columna MOTIVO DE GLOSA
    valor: int  # valor objetado en pesos


@dataclass
class Factura:
    prefijo: str
    numero: str
    fecha_atencion: str = ""
    paciente: str = ""
    valor_factura: int = 0
    total_objetado_pdf: int | None = None  # "Total Factura" del propio PDF
    glosas: list[Glosa] = field(default_factory=list)

    @property
    def cxc(self) -> str:
        """N° de CxC en DGH: prefijo + número a 10 dígitos (HUS0000530265)."""
        return f"{self.prefijo}{int(self.numero):010d}"

    @property
    def total_extraido(self) -> int:
        return sum(g.valor for g in self.glosas)


def setup_logging() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")


def _to_int(s: str) -> int:
    """'4.677.656,00' (formato Colombia) → 4677656. La parte decimal se ignora."""
    parte_entera = s.split(",")[0]
    return int(re.sub(r"\.", "", parte_entera) or "0")


def _limpiar(texto: str) -> str:
    """Colapsa espacios y remueve artefactos (cid:N) de la extracción."""
    return re.sub(r"\s+", " ", RE_CID.sub(" ", texto)).strip()


# ---------------------------------------------------------------------------
# Extracción por coordenadas
# ---------------------------------------------------------------------------


def _filas_de_pagina(page) -> list[dict[str, str]]:
    """Convierte una página en filas visuales: dict columna → texto.

    Agrupa los caracteres en renglones por su coordenada `top` (clusters con
    tolerancia, porque celdas de la misma fila pueden variar 1-2 pt) y dentro
    de cada renglón reparte los caracteres a su columna por el centro X.
    """
    chars = [c for c in page.chars if Y_DATOS_MIN <= c["top"] <= Y_DATOS_MAX]
    if not chars:
        return []

    chars.sort(key=lambda c: c["top"])
    renglones: list[list[dict]] = [[chars[0]]]
    for c in chars[1:]:
        if c["top"] - renglones[-1][-1]["top"] <= 3.0:
            renglones[-1].append(c)
        else:
            renglones.append([c])

    filas: list[dict[str, str]] = []
    for renglon in renglones:
        fila: dict[str, str] = {}
        for nombre, (x0, x1) in COLUMNAS.items():
            celda = [c for c in renglon if x0 <= (c["x0"] + c["x1"]) / 2 < x1]
            celda.sort(key=lambda c: c["x0"])
            texto, x_prev = "", None
            for c in celda:
                if x_prev is not None and c["x0"] - x_prev > 1.2:
                    texto += " "
                texto += c["text"]
                x_prev = c["x1"]
            fila[nombre] = _limpiar(texto)
        if any(fila.values()):
            filas.append(fila)
    return filas


def parsear_pdf(ruta: Path) -> tuple[list[Factura], str]:
    """Extrae las facturas con sus glosas de un PDF DETALLE DE AUDITORIA Y GLOSAS.

    Devuelve (facturas, fecha_impresion "dd/mm/yy").
    """
    try:
        import pdfplumber
    except ImportError:
        sys.stderr.write("ERROR: falta pdfplumber.\nInstalalo con:  py -m pip install pdfplumber\n")
        sys.exit(2)

    facturas: list[Factura] = []
    factura: Factura | None = None
    glosa: Glosa | None = None
    fecha_impresion = ""

    with pdfplumber.open(str(ruta)) as pdf:
        for page in pdf.pages:
            texto_pagina = page.extract_text() or ""
            m_imp = RE_FECHA_IMPRESION.search(texto_pagina)
            if m_imp and not fecha_impresion:
                fecha_impresion = m_imp.group(1)

            for fila in _filas_de_pagina(page):
                # --- Cierre de factura: "Total Factura: HUS - <n>  <objetado> ... <a pagar>"
                if "Total" in fila["prefijo"]:
                    if factura is not None:
                        m_val = RE_VALOR.search(fila["vobjetado"])
                        if m_val:
                            factura.total_objetado_pdf = _to_int(m_val.group(0))
                        factura, glosa = None, None
                    continue

                # --- Encabezado de factura: prefijo (HUS) + número
                if RE_PREFIJO.match(fila["prefijo"]) and fila["factura"].replace(" ", "").isdigit():
                    glosa = None
                    factura = Factura(
                        prefijo=fila["prefijo"].strip(),
                        numero=fila["factura"].replace(" ", ""),
                        paciente=fila["nombre"],
                    )
                    m_fec = RE_FECHA.search(fila["fecha"])
                    if m_fec:
                        factura.fecha_atencion = m_fec.group(0)
                    m_vf = RE_VALOR.search(fila["vfactura"])
                    if m_vf:
                        factura.valor_factura = _to_int(m_vf.group(0))
                    facturas.append(factura)
                    # ojo: la primera glosa viene en el MISMO renglón (cae al código de abajo)

                if factura is None:
                    continue

                # --- Inicio de glosa: código "XX99 99" al comienzo de CONCEPTO
                m_cod = RE_CODIGO_GLOSA.match(fila["concepto"])
                if m_cod:
                    m_val = RE_VALOR.search(fila["vobjetado"])
                    glosa = Glosa(
                        codigo=m_cod.group(1) + m_cod.group(2),
                        concepto=m_cod.group(3).strip(),
                        motivo=fila["motivo"],
                        valor=_to_int(m_val.group(0)) if m_val else 0,
                    )
                    factura.glosas.append(glosa)
                    continue

                # --- Continuación de la glosa en curso
                if glosa is not None:
                    if fila["concepto"]:
                        glosa.concepto = _limpiar(glosa.concepto + " " + fila["concepto"])
                    if fila["motivo"]:
                        glosa.motivo = _limpiar(glosa.motivo + " " + fila["motivo"])

    return facturas, fecha_impresion


# ---------------------------------------------------------------------------
# Armado del Excel OBJECIONES
# ---------------------------------------------------------------------------


def _servicio_de(glosa: Glosa) -> str:
    """Código del servicio/producto glosado, si el motivo lo menciona."""
    m = RE_SERVICIO.search(glosa.motivo)
    return m.group(1) if m else ""


def filas_objeciones(factura: Factura, fecha_doc: datetime) -> list[list]:
    """Filas (16 columnas) del formato OBJECIONES para una factura."""
    filas = []
    for g in factura.glosas:
        observacion = f"{g.codigo} {g.concepto}: {g.motivo}${g.valor}"
        filas.append(
            [
                "1",  # CDCONSEC
                fecha_doc,  # CDFECDOC
                factura.cxc,  # CRNCXC
                fecha_doc,  # CROFECOBJ
                None,  # CROREFERE
                None,  # CROOBSERV
                0,  # CROCLAOBJ
                None,  # CRNCLAOBJ
                "999",  # GENUSUARIO4
                g.codigo,  # CRNCONOBJ
                _servicio_de(g) or None,  # SLNSERPRO
                None,  # IDRIPS
                None,  # CTNCENCOS
                g.valor,  # CROVALOBJ
                observacion,  # CRDOBSERV
                0,  # CROTIPOBJ
            ]
        )
    return filas


def escribir_excel(filas: list[list], salida: Path) -> None:
    """Escribe el Excel con los mismos tipos y formatos del ejemplo de EMSSANAR."""
    try:
        from openpyxl import Workbook
    except ImportError:
        sys.stderr.write("ERROR: falta openpyxl. Instalalo con: py -m pip install openpyxl\n")
        sys.exit(2)

    formatos = {2: FMT_FECHA, 4: FMT_FECHA, 7: "General", 14: FMT_CONTABLE, 16: "0"}
    wb = Workbook()
    ws = wb.active
    ws.title = "OBJECIONES"
    ws.append(ENCABEZADOS)
    for fila in filas:
        ws.append(fila)
        r = ws.max_row
        for col in range(1, 17):
            ws.cell(row=r, column=col).number_format = formatos.get(col, "@")
    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(salida))


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def _parse_fecha(s: str) -> datetime:
    """'10/07/26' o '10/07/2026' → datetime."""
    dd, mm, yy = s.split("/")
    anio = int(yy) + 2000 if len(yy) == 2 else int(yy)
    return datetime(anio, int(mm), int(dd))


def procesar(
    pdfs: list[Path],
    salida_dir: Path,
    entidad: str,
    fecha_cli: str | None,
    consolidado: Path | None,
) -> int:
    """Corre el pipeline completo. Devuelve exit code (0 ok / 1 con descuadres)."""
    todas_filas: list[list] = []
    descuadres = 0
    generados = 0

    for pdf in pdfs:
        facturas, fecha_imp = parsear_pdf(pdf)
        if not facturas:
            logger.warning(
                "%s: no se encontraron facturas — ¿es un DETALLE DE AUDITORIA Y GLOSAS?", pdf.name
            )
            continue
        fecha_doc = _parse_fecha(fecha_cli or fecha_imp or datetime.now().strftime("%d/%m/%Y"))

        for fac in facturas:
            if fac.total_objetado_pdf is not None and fac.total_extraido != fac.total_objetado_pdf:
                logger.error(
                    "%s: DESCUADRE — glosas extraídas suman $%s pero el PDF dice $%s",
                    fac.cxc,
                    f"{fac.total_extraido:,}",
                    f"{fac.total_objetado_pdf:,}",
                )
                descuadres += 1
            if not fac.glosas:
                logger.info("%s (%s): sin glosas — no se genera Excel", fac.cxc, fac.paciente)
                continue

            filas = filas_objeciones(fac, fecha_doc)
            todas_filas.extend(filas)
            generados += 1
            if consolidado is None:
                destino = salida_dir / f"OBJECIONES_{entidad}_{fac.cxc}.xlsx"
                escribir_excel(filas, destino)
                logger.info(
                    "%s (%s): %d objeciones, total $%s %s → %s",
                    fac.cxc,
                    fac.paciente,
                    len(fac.glosas),
                    f"{fac.total_extraido:,}",
                    "✓ cuadra con el PDF"
                    if fac.total_extraido == fac.total_objetado_pdf
                    else "(sin total en PDF)"
                    if fac.total_objetado_pdf is None
                    else "✗ NO CUADRA",
                    destino,
                )
            else:
                logger.info(
                    "%s (%s): %d objeciones, total $%s",
                    fac.cxc,
                    fac.paciente,
                    len(fac.glosas),
                    f"{fac.total_extraido:,}",
                )

    if consolidado is not None and todas_filas:
        escribir_excel(todas_filas, consolidado)
        logger.info(
            "Consolidado: %d objeciones de %d facturas → %s",
            len(todas_filas),
            generados,
            consolidado,
        )

    if not todas_filas:
        logger.warning("No se extrajo ninguna objeción.")
        return 1
    return 1 if descuadres else 0


def main(argv: list[str] | None = None) -> int:
    setup_logging()
    ap = argparse.ArgumentParser(
        description="Convierte el PDF de auditoría del Dispensario (AUDITOOL) al Excel OBJECIONES de DGH.",
    )
    fuente = ap.add_mutually_exclusive_group(required=True)
    fuente.add_argument("--pdf", type=Path, help="Un PDF DETALLE DE AUDITORIA Y GLOSAS")
    fuente.add_argument("--carpeta", type=Path, help="Carpeta con varios PDFs")
    ap.add_argument(
        "--salida-dir", type=Path, default=None, help="Carpeta de salida (default: la del PDF)"
    )
    ap.add_argument(
        "--consolidado",
        type=Path,
        default=None,
        help="Generar UN SOLO xlsx con todas las facturas en esta ruta",
    )
    ap.add_argument(
        "--entidad",
        default="DISPENSARIO",
        help="Nombre de la entidad para el archivo de salida (default: DISPENSARIO)",
    )
    ap.add_argument(
        "--fecha",
        default=None,
        help="Fecha de la objeción dd/mm/aaaa (default: Fecha Impresión del PDF)",
    )
    args = ap.parse_args(argv)

    if args.pdf:
        pdfs = [args.pdf]
    else:
        pdfs = sorted(args.carpeta.glob("*.pdf")) + sorted(args.carpeta.glob("*.PDF"))
        if not pdfs:
            logger.error("No hay PDFs en %s", args.carpeta)
            return 2

    salida_dir = args.salida_dir or (pdfs[0].parent if args.pdf else args.carpeta)
    return procesar(pdfs, salida_dir, args.entidad.upper(), args.fecha, args.consolidado)


if __name__ == "__main__":
    sys.exit(main())
