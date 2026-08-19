#!/usr/bin/env python3
"""siifa_balance.py — El balance de un corte: qué había, qué se respondió,
qué falta y qué llegó nuevo.

Responde de una vez las cuatro preguntas que se hacen después de un cargue:

    1. ¿Qué estaba glosado AL CORTE (la fecha del archivo con el que se
       trabajaron las respuestas)?
    2. De eso, ¿qué quedó RESPONDIDO? (separando lo que ya venía respondido
       de antes y lo que se respondió en este trabajo)
    3. ¿Qué quedó SIN RESPONDER? (lo urgente: el plazo sigue corriendo)
    4. ¿Qué NUEVO ha glosado la EPS desde el corte? (el próximo trabajo)

CÓMO SE USA (una IPS por corrida, con sus dos archivos):

    REM el CORTE es el informe con el que se armaron las respuestas;
    REM HOY es un informe recién bajado (opción [1] del bot).
    py siifa_balance.py --ips SOCORRO ^
        --corte "D:\\...\\SIIFA\\SOCORRO\\SIIFA_informe_seguimientos_SOCORRO.xlsx" ^
        --hoy   "D:\\...\\SIIFA\\SOCORRO\\informe_DESPUES.xlsx"

Deja un Excel (BALANCE_SIIFA_<IPS>.xlsx) con el resumen y el detalle de lo
pendiente, y muestra las cuentas en pantalla para leerlas de una vez.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import siifa_perfiles as perfiles  # noqa: E402
from siifa_novedades import (  # noqa: E402
    COLUMNAS_VISTA,
    TITULOS,
    _pesos,
    leer_informe,
    resumen_por_entidad,
    valor_total,
)

# Los seis cajones del balance. Los nombres salen tal cual en el Excel.
VENIA_RESPONDIDA = "DEL CORTE - ya venía respondida"
RESPONDIDA = "DEL CORTE - respondida en este trabajo"
SIN_RESPONDER = "DEL CORTE - SIGUE SIN RESPONDER"
YA_NO_ESTA = "DEL CORTE - ya no aparece en SIIFA"
NUEVA_RESPONDIDA = "NUEVA - ya respondida"
NUEVA_SIN_RESPONDER = "NUEVA - SIN RESPONDER"


def _id(fila: dict) -> str:
    return str(fila.get("id_seguimiento_factura_glosa", "")).strip()


def _respondida(fila: dict) -> bool:
    return str(fila.get("tiene_respuesta", "")).strip().upper() == "SI"


def balancear(corte: list[dict], hoy: list[dict]) -> dict[str, list[dict]]:
    """Cruza el informe del corte contra el de hoy y reparte en los cajones.

    El detalle que se entrega es siempre el de HOY (trae el estado real de la
    plataforma); del corte sólo se usa para saber qué existía y qué ya venía
    respondido antes de este trabajo.
    """
    del_corte = {_id(f): f for f in corte}
    cajones: dict[str, list[dict]] = {
        VENIA_RESPONDIDA: [],
        RESPONDIDA: [],
        SIN_RESPONDER: [],
        YA_NO_ESTA: [],
        NUEVA_RESPONDIDA: [],
        NUEVA_SIN_RESPONDER: [],
    }

    ids_hoy = set()
    for fila in hoy:
        ids_hoy.add(_id(fila))
        antes = del_corte.get(_id(fila))
        if antes is None:
            cajones[NUEVA_RESPONDIDA if _respondida(fila) else NUEVA_SIN_RESPONDER].append(fila)
        elif not _respondida(fila):
            cajones[SIN_RESPONDER].append(fila)
        elif _respondida(antes):
            cajones[VENIA_RESPONDIDA].append(fila)
        else:
            cajones[RESPONDIDA].append(fila)

    cajones[YA_NO_ESTA] = [f for f in corte if _id(f) not in ids_hoy]
    return cajones


def _cuenta(filas: list[dict]) -> tuple[int, int, int]:
    """(ítems, facturas, valor) de un cajón, con la regla de devoluciones."""
    return (
        len(filas),
        len({str(f.get("numero_factura") or "") for f in filas}),
        valor_total(filas),
    )


def mostrar(cajones: dict[str, list[dict]], ips, limite: int) -> None:
    n_corte = sum(
        len(cajones[c]) for c in (VENIA_RESPONDIDA, RESPONDIDA, SIN_RESPONDER, YA_NO_ESTA)
    )
    print()
    print("=" * 78)
    print(f"BALANCE DE {ips.nombre.upper()}")
    print("=" * 78)

    print(f"\nLO QUE ESTABA GLOSADO AL CORTE: {n_corte} seguimiento(s).")
    print("-" * 78)
    for cajon in (VENIA_RESPONDIDA, RESPONDIDA, SIN_RESPONDER, YA_NO_ESTA):
        items, facturas, valor = _cuenta(cajones[cajon])
        nombre = cajon.split(" - ", 1)[1]
        print(f"  {nombre:<38}{items:>7} en {facturas:>6} factura(s) {_pesos(valor):>16}")

    nuevas = cajones[NUEVA_RESPONDIDA] + cajones[NUEVA_SIN_RESPONDER]
    print(f"\nLO NUEVO DESDE EL CORTE: {len(nuevas)} seguimiento(s).")
    print("-" * 78)
    for cajon in (NUEVA_RESPONDIDA, NUEVA_SIN_RESPONDER):
        items, facturas, valor = _cuenta(cajones[cajon])
        nombre = cajon.split(" - ", 1)[1]
        print(f"  {nombre:<38}{items:>7} en {facturas:>6} factura(s) {_pesos(valor):>16}")

    pendientes = cajones[SIN_RESPONDER] + cajones[NUEVA_SIN_RESPONDER]
    if not pendientes:
        print("\nNADA PENDIENTE: todo lo que SIIFA tiene está respondido.\n")
        return

    items, facturas, valor = _cuenta(pendientes)
    print(f"\nPENDIENTE DE RESPONDER (viejo + nuevo): {items} por {_pesos(valor)}.")
    print("-" * 78)
    print(f"{'Entidad':<32}{'Tipo':<12}{'Ítems':>7}{'Facturas':>9}{'Valor':>18}")
    print("-" * 78)
    for entidad, tipo, n, nf, v in resumen_por_entidad(pendientes)[:limite]:
        print(f"{entidad[:31]:<32}{tipo:<12}{n:>7}{nf:>9}{_pesos(v):>18}")
    print("\nEl detalle factura por factura queda en el Excel.\n")

    if cajones[YA_NO_ESTA]:
        print(
            f"ATENCIÓN: {len(cajones[YA_NO_ESTA])} seguimiento(s) del corte ya no aparecen\n"
            "en SIIFA (la EPS pudo reformularlos o retirarlos). Quedan en la hoja\n"
            "YA_NO_ESTAN del Excel para revisarlos.\n"
        )


def escribir(cajones: dict[str, list[dict]], ruta: Path, ips) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise SystemExit("ERROR: falta openpyxl. py -m pip install openpyxl") from None

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    rojo = Font(bold=True, color="9C0006")

    wb = Workbook()
    ws = wb.active
    ws.title = "RESUMEN"
    ws.append([f"Balance SIIFA — {ips.nombre}"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Grupo", "Ítems", "Facturas", "Valor"])
    for c in range(1, 5):
        ws.cell(row=3, column=c).fill = header_fill
        ws.cell(row=3, column=c).font = header_font
    for cajon in (
        VENIA_RESPONDIDA,
        RESPONDIDA,
        SIN_RESPONDER,
        YA_NO_ESTA,
        NUEVA_RESPONDIDA,
        NUEVA_SIN_RESPONDER,
    ):
        items, facturas, valor = _cuenta(cajones[cajon])
        ws.append([cajon, items, facturas, valor])
        ws.cell(row=ws.max_row, column=4).number_format = '"$"#,##0'
        if "SIN RESPONDER" in cajon and items:
            ws.cell(row=ws.max_row, column=1).font = rojo
    for col, ancho in (("A", 44), ("B", 10), ("C", 12), ("D", 18)):
        ws.column_dimensions[col].width = ancho

    # Sólo lo accionable lleva detalle: lo respondido ya no necesita revisión.
    hojas = [
        ("PENDIENTES", cajones[SIN_RESPONDER] + cajones[NUEVA_SIN_RESPONDER]),
        ("NUEVAS", cajones[NUEVA_RESPONDIDA] + cajones[NUEVA_SIN_RESPONDER]),
        ("YA_NO_ESTAN", cajones[YA_NO_ESTA]),
    ]
    from _dinero import a_entero

    for nombre, filas in hojas:
        if not filas:
            continue
        hoja = wb.create_sheet(nombre)
        for c, h in enumerate(COLUMNAS_VISTA, 1):
            cell = hoja.cell(row=1, column=c, value=TITULOS.get(h, h))
            cell.fill = header_fill
            cell.font = header_font
        for r, fila in enumerate(filas, start=2):
            for c, h in enumerate(COLUMNAS_VISTA, 1):
                valor = a_entero(fila.get(h)) if h == "valor_glosa" else fila.get(h, "")
                cell = hoja.cell(row=r, column=c, value=valor)
                cell.alignment = Alignment(vertical="top", wrap_text=(h == "descripcion_glosa"))
                if h == "valor_glosa":
                    cell.number_format = '"$"#,##0'
        anchos = {"razon_social_eps": 32, "descripcion_glosa": 45, "numero_factura": 14}
        for c, h in enumerate(COLUMNAS_VISTA, 1):
            hoja.column_dimensions[get_column_letter(c)].width = anchos.get(h, 16)
        hoja.freeze_panes = "A2"

    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ruta))
    print(f"Archivo con el balance: {ruta}\n")


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument(
        "--corte",
        required=True,
        help="Informe con el que se armaron las respuestas (el archivo del corte).",
    )
    ap.add_argument("--hoy", required=True, help="Informe recién bajado de SIIFA.")
    ap.add_argument("--salida", help="Excel de salida (default: BALANCE_SIIFA_<IPS>.xlsx al lado).")
    ap.add_argument(
        "--limite", type=int, default=25, help="Cuántas entidades mostrar en pantalla (default 25)."
    )
    perfiles.agregar_argumento(ap)
    args = ap.parse_args()
    ips = perfiles.buscar(args.ips)
    perfiles.anunciar(ips)

    corte = leer_informe(Path(args.corte))
    perfiles.verificar_informe(ips, corte, "el informe del corte")
    hoy = leer_informe(Path(args.hoy))
    perfiles.verificar_informe(ips, hoy, "el informe de hoy")

    cajones = balancear(corte, hoy)
    mostrar(cajones, ips, args.limite)

    salida = (
        Path(args.salida)
        if args.salida
        else Path(args.hoy).parent / f"BALANCE_SIIFA_{ips.clave}.xlsx"
    )
    escribir(cajones, salida, ips)


if __name__ == "__main__":
    main()
