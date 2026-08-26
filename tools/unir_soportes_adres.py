"""unir_soportes_adres.py — los dos folios de cada factura del ADRES.

En la carpeta de cada factura los soportes están sueltos y con el nombre con
que salieron del sistema. Este bot los **numera** y los **une en los dos folios**
que pide el ADRES, con el nombre con que hay que subirlos:

    📁 HUS352904
        1 RESPUESTA A GLOSA.pdf   (era RTA_ADRES_HUS352904.pdf)
        2 EPICRISIS.pdf           (era 680010079201_HUS352904_EPICRIS.pdf)
        3 HISTORIA CLINICA.pdf    (era HC.pdf)
        4 AYUDAS DIAGNOSTICAS.pdf (era DX.pdf)
        5 OTROS.pdf
        ────────────────────────► 680010079201_HUS352904_EPICRIS.pdf

        1 FACTURA.pdf             (era 680010079201_HUS352904_FACTURA.pdf)
        2 DETALLADO.pdf           (el detallado en Excel, pasado a PDF)
        3 REPRESENTACION GRAFICA DIAN.pdf
        4 NOTAS CREDITO.pdf       (PENDIENTE: todavía no las han sacado)
        ────────────────────────► 680010079201_HUS352904_FACTURA.pdf

Numerar primero no es adorno: es lo que **deja libre el nombre del folio**,
porque ese nombre es justo el que traían la epicrisis y la factura antes de
renombrarlas.

EL ORDEN DEL FOLIO CLÍNICO es el de la hoja del área:

    1. RESPUESTA A GLOSA          6. NOTAS DE ENFERMERÍA
    2. EPICRISIS                  7. INSUMOS
    3. HISTORIA CLÍNICA           8. OTROS
       (consulta de urgencias,
        terapias, curaciones,
        evoluciones, procedimientos)
    4. AYUDAS DIAGNÓSTICAS
    5. MEDICAMENTOS

LAS NOTAS CRÉDITO quedan pendientes a propósito: son las de los valores
aceptados y todavía no existen. El bot avisa cuántas faltan; cuando lleguen se
vuelve a correr y entran solas de cuartas, sin rehacer nada.

CÓMO SABE DE QUÉ ES CADA PDF. Por el **nombre del archivo**: busca las palabras
con que el equipo los nombra (EPICRISIS, URGENCIAS, TERAPIA, CURACION,
EVOLUCION, MEDICAMENTOS, ENFERMERIA, INSUMOS, FACTURA, DETALLADO…) y las
abreviaturas que usa el auditor (EPI, HC, DX, MED, NTE, INS). Lo que no reconoce
**no se pierde**: va al grupo OTROS y sale listado en el reporte para que el
auditor lo revise. Las palabras se pueden cambiar sin tocar el código, con
`--mapa-nombres`.

SIMULA POR DEFECTO. Armar folios no se deshace de un clic, así que el bot
primero **muestra qué haría** y no escribe nada mientras no se le pase
`--aplicar`. Nunca se toma a sí mismo como entrada, así que se puede correr las
veces que haga falta.

USO (Windows):

    REM 1) PRIMERO en simulación: muestra los dos folios y no toca nada.
    py tools\\unir_soportes_adres.py --folio ^
        --carpeta "Z:\\...\\TECNICOS\\CAROLINA" ^
        --carpeta-facturas "Z:\\...\\4.FACTURAS CON XML\\XML"

    REM 2) Si el listado se ve bien, con --aplicar sí arma los folios.
    py tools\\unir_soportes_adres.py --folio ^
        --carpeta "Z:\\...\\TECNICOS\\CAROLINA" ^
        --carpeta-facturas "Z:\\...\\4.FACTURAS CON XML\\XML" ^
        --convertir-detallado --aplicar --reporte-csv "C:\\temp-rtas\\folios.csv"

    REM Solo numerar los soportes, sin unir nada:
    py tools\\unir_soportes_adres.py --carpeta "Z:\\..." --renombrar --aplicar

    REM El consolidado viejo de un solo PDF (<FACTURA>_SOPORTES.pdf):
    py tools\\unir_soportes_adres.py --carpeta "Z:\\..." --aplicar

INSTALACIÓN (una vez):

    py -m pip install pypdf openpyxl
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import re
import shutil
import sys
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
# El motor de unión ya existe y está probado: aquí solo se le da el ORDEN.
from organizar_soportes_por_factura import factura_del_nombre  # noqa: E402
from unir_pdfs_carpetas import _cargar_lector_escritor, clave_natural, unir_pdfs  # noqa: E402

logger = logging.getLogger("unir_soportes_adres")

SUFIJO_UNIDO = "_SOPORTES"

# Los DOS folios de cada factura, con el nombre que les pone el ADRES:
#   <NIT>_<FACTURA>_EPICRIS.pdf → la respuesta a glosa y la historia clínica
#   <NIT>_<FACTURA>_FACTURA.pdf → la factura y su soporte contable
SUFIJO_EPICRIS = "_EPICRIS"
SUFIJO_FACTURA = "_FACTURA"

FOLIO_EPICRIS = "EPICRIS"
FOLIO_FACTURA = "FACTURA"


# ─── Los grupos del folio clínico, en el orden que pide el área ──────────────


@dataclass(frozen=True)
class Grupo:
    """Un renglón de la lista de soportes."""

    orden: int
    clave: str
    titulo: str
    palabras: tuple[str, ...]
    folio: str = FOLIO_EPICRIS


# El orden de esta tupla ES el orden del PDF. Las palabras se comparan sin
# tildes y en mayúsculas contra el nombre del archivo.
GRUPOS: tuple[Grupo, ...] = (
    Grupo(
        1,
        "RESPUESTA",
        "RESPUESTA A GLOSA",
        (
            "RESPUESTA A GLOSA",
            "RESPUESTA GLOSA",
            "RTA GLOSA",
            # Así se llama el PDF que arma respuestas_adres_por_factura.py:
            # RTA_ADRES_HUS311371.pdf. Es la respuesta a glosa del paquete.
            "RTA ADRES",
            "RESPUESTA",
        ),
    ),
    Grupo(2, "EPICRISIS", "EPICRISIS", ("EPICRISIS", "EPICRIS", "EPI")),
    Grupo(
        3,
        "URGENCIAS",
        "HISTORIA CLINICA - CONSULTA DE URGENCIAS",
        ("CONSULTA DE URGENCIAS", "URGENCIAS", "TRIAGE"),
    ),
    Grupo(
        4,
        "TERAPIAS",
        "HISTORIA CLINICA - TERAPIAS",
        ("TERAPIAS", "TERAPIA", "FISIOTERAPIA", "RESPIRATORIA"),
    ),
    Grupo(5, "CURACIONES", "HISTORIA CLINICA - CURACIONES", ("CURACIONES", "CURACION")),
    Grupo(
        6,
        "EVOLUCIONES",
        "HISTORIA CLINICA - EVOLUCIONES",
        ("EVOLUCIONES", "EVOLUCION", "INTERCONSULTA"),
    ),
    Grupo(
        7,
        "PROCEDIMIENTOS",
        "HISTORIA CLINICA - PROCEDIMIENTOS",
        ("PROCEDIMIENTOS", "PROCEDIMIENTO", "DESCRIPCION QUIRURGICA", "QUIRURGICA"),
    ),
    Grupo(8, "HISTORIA", "HISTORIA CLINICA", ("HISTORIA CLINICA", "HISTORIA", "HC")),
    Grupo(
        9,
        "AYUDAS",
        "AYUDAS DIAGNOSTICAS",
        (
            "AYUDAS DIAGNOSTICAS",
            "AYUDA DIAGNOSTICA",
            "LABORATORIO",
            "IMAGENOLOGIA",
            "RADIOGRAFIA",
            "TOMOGRAFIA",
            "ANGIOTOMOGRAFIA",
            "ECOGRAFIA",
            "RESONANCIA",
            "ELECTROMIOGRAFIA",
            "NEUROCONDUCCION",
            # Así vienen nombrados los PDF de exámenes: con el nombre con que
            # salen en el detallado (bloques LABORATORIOS e IMAGENOLOGIA).
            "GLUCOMETRIA",
            "GASES ARTERIALES",
            "LACTATO",
            "HEMOGRAMA",
            "CUADRO HEMATICO",
            "PARCIAL DE ORINA",
            "CREATININA",
            "IONOGRAMA",
            "PORTATILES",
            "ECOCARDIOGRAMA",
            "DOPPLER",
            "ENDOSCOPIA",
            "DX",
        ),
    ),
    Grupo(
        10,
        "MEDICAMENTOS",
        "MEDICAMENTOS",
        ("MEDICAMENTOS", "MEDICAMENTO", "INGESTA", "PLAN DE MANEJO", "MED"),
    ),
    Grupo(
        11,
        "ENFERMERIA",
        "NOTAS DE ENFERMERIA",
        ("NOTAS DE ENFERMERIA", "NOTAS ENFERMERIA", "ENFERMERIA", "NTE"),
    ),
    Grupo(
        12,
        "INSUMOS",
        "INSUMOS",
        (
            "INSUMOS",
            "INSUMO",
            "GASTOS QUIROFANO",
            "GASTOS DE QUIROFANO",
            "QUIROFANO",
            "SOLICITUDES DE ENFERMERIA",
            "INS",
        ),
    ),
    # OTROS es a la vez un grupo con nombre propio —el equipo nombra archivos
    # «OTROS.pdf»— y el cajón donde cae lo que no se reconoce. Con sus palabras,
    # un OTROS.pdf sale RECONOCIDO y no como algo que el auditor deba mirar.
    Grupo(13, "OTROS", "OTROS", ("OTROS", "OTRO", "SOAT", "CERTIFICACION", "REPS")),
)
GRUPO_OTROS = next(g for g in GRUPOS if g.clave == "OTROS")


# ─── El folio de la FACTURA, con su propio orden adentro ─────────────────────

# El segundo folio de cada factura. El área lo pidió en este orden:
#   1. la FACTURA (la que viene con el XML: 680010079201_HUS######_FACTURA.pdf)
#   2. el DETALLADO (sale en Excel y se pasa a PDF)
#   3. la REPRESENTACIÓN GRÁFICA que sale de la DIAN
#   4. las NOTAS CRÉDITO de los valores aceptados
# Las notas crédito TODAVÍA NO EXISTEN: el bot deja ese renglón como PENDIENTE
# y lo avisa. Cuando lleguen se vuelve a correr y entran solas, sin rehacer nada.
GRUPOS_FACTURA: tuple[Grupo, ...] = (
    Grupo(1, "FACTURA", "FACTURA", ("FACTURA", "FACTURA DE VENTA"), FOLIO_FACTURA),
    Grupo(
        2,
        "DETALLADO",
        "DETALLADO",
        ("DETALLADO DE FACTURA", "DETALLE DE FACTURA", "DETALLADO", "DETALLE"),
        FOLIO_FACTURA,
    ),
    Grupo(
        3,
        "DIAN",
        "REPRESENTACION GRAFICA DIAN",
        ("REPRESENTACION GRAFICA DIAN", "REPRESENTACION GRAFICA", "DIAN"),
        FOLIO_FACTURA,
    ),
    Grupo(
        4,
        "NOTAS",
        "NOTAS CREDITO",
        ("NOTAS CREDITO", "NOTA CREDITO", "NOTA DE CREDITO", "NOTAS DE CREDITO"),
        FOLIO_FACTURA,
    ),
)
GRUPO_FACTURA = next(g for g in GRUPOS_FACTURA if g.clave == "FACTURA")
GRUPO_DETALLADO = next(g for g in GRUPOS_FACTURA if g.clave == "DETALLADO")

# Los renglones del folio de la factura que SÍ tienen que estar hoy. El 4
# (notas crédito) no se cuenta como falta: todavía no las han sacado.
CLAVES_FACTURA_OBLIGATORIAS = ("FACTURA", "DETALLADO", "DIAN")
CLAVE_NOTAS = "NOTAS"

# Todos los grupos, los del folio clínico y los del de la factura.
TODOS_LOS_GRUPOS: tuple[Grupo, ...] = GRUPOS + GRUPOS_FACTURA

# El detallado NO entra al folio clínico: va de segundo en el de la factura.
PALABRAS_DETALLADO = ("DETALLADO", "DETALLE DE FACTURA")


# `RegistroEnfermeria.pdf` no trae separador entre las dos palabras. Al pasarlo
# a mayúsculas quedaba «REGISTROENFERMERIA», donde ENFERMERIA ya no es una
# palabra suelta y el archivo se iba a OTROS. Se corta donde cambia de minúscula
# a mayúscula, ANTES de subirlo todo a mayúsculas.
_RE_PEGADAS = re.compile(r"(?<=[a-z0-9])(?=[A-Z])")


def _norm(texto: object) -> str:
    """Mayúsculas, sin tildes, sin signos: para comparar nombres de archivo."""
    s = unicodedata.normalize("NFKD", _RE_PEGADAS.sub(" ", str(texto or "")).upper())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(re.sub(r"[^A-Z0-9]+", " ", s).split())


def clasificar(nombre: str, mapa: dict[str, str] | None = None) -> Grupo:
    """En qué grupo va un archivo, por su nombre.

    Gana la palabra **más larga** que aparezca, no la primera: así
    «NOTAS DE ENFERMERIA» no se lo lleva «NOTAS», y «CONSULTA DE URGENCIAS» no
    se confunde con «CONSULTA». Lo que no reconoce va a OTROS.
    """
    limpio = _norm(nombre)
    if mapa:
        for palabra, clave in mapa.items():
            if _norm(palabra) and _norm(palabra) in limpio:
                grupo = next(
                    (g for g in TODOS_LOS_GRUPOS if g.clave == clave.strip().upper()), None
                )
                if grupo is not None:
                    return grupo
    mejor: tuple[int, Grupo] = (0, GRUPO_OTROS)
    for grupo in TODOS_LOS_GRUPOS:
        for palabra in grupo.palabras:
            clave = _norm(palabra)
            if clave and _es_palabra(clave, limpio) and len(clave) > mejor[0]:
                mejor = (len(clave), grupo)
    return mejor[1]


def clasificar_con_marca(nombre: str, mapa: dict[str, str] | None = None) -> tuple[Grupo, bool]:
    """(grupo, ¿lo reconoció de verdad?).

    OTROS es a la vez un grupo con nombre propio y el cajón de lo desconocido:
    un archivo llamado «OTROS.pdf» está bien clasificado y NO debe salir en la
    lista de «revisar»; uno llamado «papel raro.pdf» sí.
    """
    grupo = clasificar(nombre, mapa)
    if grupo is not GRUPO_OTROS:
        return grupo, True
    limpio = _norm(nombre)
    caso = any(_es_palabra(_norm(p), limpio) for p in GRUPO_OTROS.palabras if _norm(p))
    if mapa and not caso:
        caso = any(_norm(p) and _norm(p) in limpio for p in mapa)
    return grupo, caso


def _es_palabra(aguja: str, pajar: str) -> bool:
    """¿Aparece `aguja` como palabra completa dentro de `pajar`?

    Con `in` a secas, «INS» (insumos) casaba dentro de «INSTITUCIONAL» y «DX»
    dentro de cualquier código. Las abreviaturas de la lista son cortas: tienen
    que ir sueltas.
    """
    return re.search(rf"(?<![A-Z0-9]){re.escape(aguja)}(?![A-Z0-9])", pajar) is not None


def es_detallado(nombre: str) -> bool:
    limpio = _norm(nombre)
    return any(_norm(p) in limpio for p in PALABRAS_DETALLADO)


# ─── Lo que se encontró en cada carpeta ──────────────────────────────────────


@dataclass
class Soporte:
    """Un PDF de la carpeta, ya clasificado."""

    ruta: Path
    grupo: Grupo
    reconocido: bool = True
    paginas: int = 0


@dataclass
class Factura:
    """Una carpeta de factura y sus soportes en orden.

    Los campos `soportes` / `destino` / `estado` / `paginas` son los del folio
    clínico (el `_EPICRIS.pdf`); los que terminan en `_factura` son los del
    folio de la factura (el `_FACTURA.pdf`).
    """

    factura: str
    carpeta: Path
    soportes: list[Soporte] = field(default_factory=list)
    detallados: list[Path] = field(default_factory=list)
    destino: Path | None = None
    paginas: int = 0
    omitidos: list[str] = field(default_factory=list)
    estado: str = ""
    # El folio de la factura.
    soportes_factura: list[Soporte] = field(default_factory=list)
    destino_epicris: Path | None = None
    destino_factura: Path | None = None
    paginas_factura: int = 0
    estado_factura: str = ""
    # El NIT con que el ADRES nombra los archivos (680010079201). Sale de los
    # propios archivos de la carpeta: no se inventa.
    prefijo: str = ""
    # Detallados que todavía están en Excel y hay que pasar a PDF.
    detallados_sin_pdf: list[Path] = field(default_factory=list)
    # Folios de una corrida anterior, que no se vuelven a meter adentro.
    folios_previos: list[Path] = field(default_factory=list)
    # Los que además hay que mirar: podrían ser un soporte recién agregado.
    folios_dudosos: list[Path] = field(default_factory=list)
    # Renglones que la propia factura ya trae pegados adentro (DETALLADO, DIAN,
    # NOTAS): no se le vuelven a agregar encima.
    trae_la_factura: set[str] = field(default_factory=set)

    def faltantes(self) -> list[str]:
        """Los grupos obligatorios que no aparecieron."""
        presentes = {s.grupo.clave for s in self.soportes}
        return [g.titulo for g in (GRUPOS[0], GRUPOS[1]) if g.clave not in presentes]

    def faltantes_factura(self) -> list[str]:
        """Los renglones del folio de la factura que hoy deberían estar y no están.

        Lo que la propia factura ya trae pegado adentro (`trae_la_factura`) NO
        falta: está, solo que dentro del mismo PDF.
        """
        presentes = {s.grupo.clave for s in self.soportes_factura} | self.trae_la_factura
        return [
            g.titulo
            for g in GRUPOS_FACTURA
            if g.clave in CLAVES_FACTURA_OBLIGATORIAS and g.clave not in presentes
        ]

    def notas_pendientes(self) -> bool:
        """¿Falta el renglón 4 (notas crédito)? No es un error: aún no existen."""
        presentes = {s.grupo.clave for s in self.soportes_factura} | self.trae_la_factura
        return CLAVE_NOTAS not in presentes


ESTADO_UNIDO = "UNIDO"
ESTADO_SIMULADO = "SE UNIRIA"
ESTADO_RENOMBRADO = "RENOMBRADO"
ESTADO_SIN_PDF = "SIN PDF QUE UNIR"
ESTADO_ERROR = "ERROR"


# Caracteres que Windows no admite en un nombre de archivo.
_RE_PROHIBIDOS = re.compile(r'[<>:"/\\|?*]')


def nombre_numerado(orden: int, grupo: Grupo, extension: str = ".pdf") -> str:
    """`1 RESPUESTA A GLOSA.pdf`, `2 HISTORIA CLINICA.pdf`…

    Es como el área nombra los soportes dentro del folio de la factura: el
    número dice en qué orden van y el nombre dice de qué es cada uno, sin tener
    que abrirlos.
    """
    limpio = _RE_PROHIBIDOS.sub(" ", grupo.titulo).strip()
    return f"{orden} {limpio}{extension}"


def nombre_folio(prefijo: str, factura: str, sufijo: str) -> str:
    """`680010079201_HUS352904_EPICRIS.pdf`.

    Es el nombre que ya traían la epicrisis y la factura antes de renombrarlas:
    el mismo con que el ADRES espera el folio armado.
    """
    tallo = f"{prefijo}_{factura}" if prefijo else factura
    return f"{tallo}{sufijo}.pdf"


def prefijo_del_nombre(nombre: str, factura: str) -> str:
    """El NIT con que vienen nombrados los archivos, o «» si el nombre no lo trae.

    `680010079201_HUS352904_EPICRIS` → `680010079201`. Solo lo toma si lo que
    sigue es de verdad esta factura: así no se cuela cualquier número suelto.
    """
    partes = _norm(Path(nombre).stem).split()
    if len(partes) < 2 or not partes[0].isdigit() or len(partes[0]) < 6:
        return ""
    if factura_del_nombre(" ".join(partes[1:])) != factura:
        return ""
    return partes[0]


# Un soporte ya numerado dentro del folio: «1 RESPUESTA A GLOSA.pdf».
_RE_NUMERADO = re.compile(r"^\s*\d+\s+\S")


def _grupos_ya_numerados(archivos: list[Path], mapa: dict[str, str] | None = None) -> set[str]:
    """Qué grupos ya tienen su archivo numerado en la carpeta."""
    return {
        clasificar(r.name, mapa).clave
        for r in archivos
        if r.suffix.lower() == ".pdf" and _RE_NUMERADO.match(r.stem)
    }


def es_folio_previo(tallo: str, factura: str, numerados: set[str]) -> bool:
    """¿Este PDF es el folio que dejó una corrida anterior?

    El folio se llama igual que el archivo del que salió (`..._EPICRIS.pdf`,
    `..._FACTURA.pdf`), así que por el nombre solo no se distinguen. Lo que sí
    los distingue es la carpeta: después de armar los folios, TODOS los soportes
    quedaron numerados («1 …», «2 …»), y con el nombre original ya no queda
    ninguno. Entonces, si en la carpeta hay archivos numerados, el
    `..._EPICRIS.pdf` que quede es el folio, y no se vuelve a meter dentro de
    sí mismo.
    """
    if not numerados:
        return False  # primera corrida: todavía son los archivos de origen
    limpio = _norm(tallo)
    return any(
        limpio.endswith(f"{factura} {s.strip('_')}") for s in (SUFIJO_EPICRIS, SUFIJO_FACTURA)
    )


def _folio_dudoso(tallo: str, factura: str, numerados: set[str]) -> bool:
    """¿Se tomó por folio algo que podría ser un soporte recién agregado?

    Pasa cuando la carpeta ya está armada pero al renglón le falta su archivo
    numerado: por ejemplo un `..._EPICRIS.pdf` sin «2 EPICRISIS.pdf» al lado.
    No se adivina: se toma como folio y se avisa para que el auditor lo mire.
    """
    limpio = _norm(tallo)
    for sufijo, clave in ((SUFIJO_EPICRIS, "EPICRISIS"), (SUFIJO_FACTURA, "FACTURA")):
        if limpio.endswith(f"{factura} {sufijo.strip('_')}") and clave not in numerados:
            return True
    return False


def renombrar_lista(soportes: list[Soporte], aplicar: bool = False) -> list[tuple[Path, str]]:
    """Renombra una lista de soportes a `<n> <GRUPO>.pdf`, en su orden.

    Devuelve [(ruta actual, nombre nuevo)]. Se hace en dos vueltas —primero a un
    nombre temporal— porque el nombre que le toca a un archivo puede ser el que
    todavía tiene otro: renombrando de una, uno pisaría al otro.
    """
    plan = [
        (s.ruta, nombre_numerado(n, s.grupo, s.ruta.suffix))
        for n, s in enumerate(soportes, start=1)
    ]
    if not aplicar:
        return plan
    temporales: list[tuple[Soporte, Path, str]] = []
    for soporte, (ruta, nuevo) in zip(soportes, plan):
        if ruta.name == nuevo or not ruta.exists():
            # Si el archivo ya no está (se lo llevaron del share a mitad de la
            # corrida), no se cae el lote: se sigue con los demás.
            continue
        temporal = ruta.with_name(f"~renombrando~{ruta.name}")
        ruta.rename(temporal)
        temporales.append((soporte, temporal, nuevo))
    for soporte, temporal, nuevo in temporales:
        destino, _ = nombre_libre(temporal.parent, nuevo)
        temporal.rename(destino)
        soporte.ruta = destino
    return plan


def renombrar_en_orden(fac: Factura, aplicar: bool = False) -> list[tuple[Path, str]]:
    """Los soportes del folio clínico, numerados en su orden."""
    return renombrar_lista(fac.soportes, aplicar)


def nombre_libre(carpeta: Path, nombre: str) -> tuple[Path, bool]:
    """(ruta libre en la carpeta, ¿hubo que renombrar?). No se pisa nada."""
    destino = carpeta / nombre
    if not destino.exists():
        return destino, False
    tallo, sufijo = Path(nombre).stem, Path(nombre).suffix
    for n in range(2, 1000):
        candidato = carpeta / f"{tallo} ({n}){sufijo}"
        if not candidato.exists():
            return candidato, True
    raise ValueError(f"No hay un nombre libre para {nombre!r} en {carpeta}")


def _factura_de_carpeta(carpeta: Path) -> str:
    """`HUS379477_PEND. CARTA CORONEL` → `HUS379477`; si no hay número, el nombre.

    Usa el mismo lector que el bot que archiva los soportes: la regla de cómo se
    saca el número de factura de un nombre vive en un solo sitio.
    """
    return factura_del_nombre(carpeta.name) or carpeta.name.strip()


def planificar(
    carpeta: Path,
    facturas: set[str] | None = None,
    mapa: dict[str, str] | None = None,
) -> list[Factura]:
    """Qué se uniría en cada carpeta de factura, sin escribir nada."""
    salida: list[Factura] = []
    for sub in sorted(p for p in carpeta.iterdir() if p.is_dir()):
        numero = _factura_de_carpeta(sub)
        if facturas is not None and _norm(numero) not in facturas:
            continue
        salida.append(_planificar_carpeta(sub, numero, mapa))
    return salida


def _planificar_carpeta(sub: Path, numero: str, mapa: dict[str, str] | None) -> Factura:
    """Lo que hay en la carpeta de UNA factura, repartido en sus dos folios."""
    fac = Factura(factura=numero, carpeta=sub)
    archivos = sorted(
        r for r in sub.rglob("*") if r.is_file() and not r.name.startswith(("~$", "."))
    )
    # El NIT sale del nombre de los propios archivos (la epicrisis y la factura
    # vienen como 680010079201_HUS######_...). Si no lo traen, queda vacío.
    for r in archivos:
        if r.suffix.lower() == ".pdf" and (encontrado := prefijo_del_nombre(r.stem, numero)):
            fac.prefijo = encontrado
            break
    numerados = _grupos_ya_numerados(archivos, mapa)

    pdfs: list[Path] = []
    for r in archivos:
        if r.suffix.lower() in (".xlsx", ".xls") and es_detallado(r.name):
            fac.detallados.append(r)
            continue
        if r.suffix.lower() != ".pdf":
            continue
        if r.stem.upper().endswith(SUFIJO_UNIDO):
            continue  # el consolidado de una corrida anterior
        if es_folio_previo(r.stem, numero, numerados):
            fac.folios_previos.append(r)
            if _folio_dudoso(r.stem, numero, numerados):
                fac.folios_dudosos.append(r)
            continue
        if es_detallado(r.name):
            fac.detallados.append(r)
        pdfs.append(r)

    # Dentro de cada grupo, orden natural por el nombre del archivo.
    soportes = [
        Soporte(ruta=p, grupo=g, reconocido=ok)
        for p in pdfs
        for g, ok in [clasificar_con_marca(p.name, mapa)]
    ]
    soportes.sort(key=lambda s: (s.grupo.orden, clave_natural(s.ruta.name)))
    fac.soportes = [s for s in soportes if s.grupo.folio == FOLIO_EPICRIS]
    fac.soportes_factura = [s for s in soportes if s.grupo.folio == FOLIO_FACTURA]
    # El detallado que solo está en Excel: al folio únicamente entran PDF.
    if not any(s.grupo.clave == "DETALLADO" for s in fac.soportes_factura):
        fac.detallados_sin_pdf = [d for d in fac.detallados if d.suffix.lower() != ".pdf"]
    fac.destino = sub / f"{numero}{SUFIJO_UNIDO}.pdf"
    fac.destino_epicris = sub / nombre_folio(fac.prefijo, numero, SUFIJO_EPICRIS)
    fac.destino_factura = sub / nombre_folio(fac.prefijo, numero, SUFIJO_FACTURA)
    fac.estado = ESTADO_SIMULADO if fac.soportes else ESTADO_SIN_PDF
    fac.estado_factura = ESTADO_SIMULADO if fac.soportes_factura else ESTADO_SIN_PDF
    return fac


def unir(
    carpeta: Path,
    facturas: set[str] | None = None,
    mapa: dict[str, str] | None = None,
    aplicar: bool = False,
) -> list[Factura]:
    """Une los soportes de cada factura. Sin `aplicar`, solo dice qué haría."""
    plan = planificar(carpeta, facturas, mapa)
    if not aplicar:
        return plan
    PdfReader, PdfWriter = _cargar_lector_escritor()
    for fac in plan:
        if fac.estado != ESTADO_SIMULADO or fac.destino is None:
            continue
        try:
            paginas, omitidos = unir_pdfs(
                [s.ruta for s in fac.soportes], fac.destino, PdfReader, PdfWriter
            )
            fac.paginas, fac.omitidos = paginas, omitidos
            fac.estado = ESTADO_UNIDO if paginas else ESTADO_ERROR
            if not paginas:
                fac.omitidos.append("ningún PDF tenía páginas legibles")
        except Exception as e:  # noqa: BLE001 - una factura mala no tumba el lote
            logger.warning("  %s: %s", fac.factura, e)
            fac.estado = ESTADO_ERROR
            fac.omitidos.append(str(e))
    return plan


# ─── Los dos folios de cada factura ──────────────────────────────────────────


def aplicar_folios(plan: list[Factura], aplicar: bool = False, prefijo: str = "") -> list[Factura]:
    """Deja cada factura con sus DOS folios, como los pide el ADRES.

    En la carpeta de la factura:
      1. numera los soportes clínicos («1 RESPUESTA A GLOSA.pdf»,
         «2 EPICRISIS.pdf», «3 HISTORIA CLINICA.pdf»…) y los une en
         `<NIT>_<FACTURA>_EPICRIS.pdf`;
      2. numera la factura y su soporte contable («1 FACTURA.pdf»,
         «2 DETALLADO.pdf», «3 REPRESENTACION GRAFICA DIAN.pdf»,
         «4 NOTAS CREDITO.pdf») y los une en `<NIT>_<FACTURA>_FACTURA.pdf`.

    Numerar primero no es adorno: es lo que **deja libre el nombre del folio**,
    porque ese nombre es justo el que traían la epicrisis (`..._EPICRIS.pdf`) y
    la factura (`..._FACTURA.pdf`) antes de renombrarlas.
    """
    for fac in plan:
        if prefijo and not fac.prefijo:
            fac.prefijo = prefijo
            fac.destino_epicris = fac.carpeta / nombre_folio(prefijo, fac.factura, SUFIJO_EPICRIS)
            fac.destino_factura = fac.carpeta / nombre_folio(prefijo, fac.factura, SUFIJO_FACTURA)
    if not aplicar:
        return plan
    PdfReader, PdfWriter = _cargar_lector_escritor()
    for fac in plan:
        try:
            renombrar_lista(fac.soportes, aplicar=True)
            renombrar_lista(fac.soportes_factura, aplicar=True)
        except OSError as e:
            # Un archivo abierto en Acrobat, sin permisos, o el share que se cae:
            # esa factura se salta y las otras 323 siguen. Queda anotado.
            logger.warning("  %s: no pude renombrar (%s)", fac.factura, e)
            fac.omitidos.append(f"no pude renombrar: {e}")
            fac.estado = fac.estado_factura = ESTADO_ERROR
            continue
        _unir_folio(fac, FOLIO_EPICRIS, PdfReader, PdfWriter)
        _unir_folio(fac, FOLIO_FACTURA, PdfReader, PdfWriter)
    return plan


def _unir_folio(fac: Factura, cual: str, PdfReader, PdfWriter) -> None:
    """Une uno de los dos folios. Una factura mala no tumba el lote."""
    clinico = cual == FOLIO_EPICRIS
    soportes = fac.soportes if clinico else fac.soportes_factura
    destino = fac.destino_epicris if clinico else fac.destino_factura
    if not soportes or destino is None:
        return
    try:
        paginas, omitidos = unir_pdfs([s.ruta for s in soportes], destino, PdfReader, PdfWriter)
        estado = ESTADO_UNIDO if paginas else ESTADO_ERROR
        if not paginas:
            omitidos.append("ningún PDF tenía páginas legibles")
    except Exception as e:  # noqa: BLE001 - una factura mala no tumba el lote
        logger.warning("  %s: %s", fac.factura, e)
        paginas, omitidos, estado = 0, [str(e)], ESTADO_ERROR
    if clinico:
        fac.paginas, fac.estado = paginas, estado
    else:
        fac.paginas_factura, fac.estado_factura = paginas, estado
    fac.omitidos.extend(omitidos)


def armar_folios(
    carpeta: Path,
    facturas: set[str] | None = None,
    mapa: dict[str, str] | None = None,
    aplicar: bool = False,
    prefijo: str = "",
) -> list[Factura]:
    """Planifica y arma los dos folios. Sin `aplicar`, solo dice qué haría."""
    return aplicar_folios(planificar(carpeta, facturas, mapa), aplicar, prefijo)


# ─── La factura, que vive en otra carpeta (la del XML) ───────────────────────


def indice_facturas(carpeta: Path) -> dict[str, Path]:
    """Número de factura → su PDF dentro de «4.FACTURAS CON XML\\XML».

    Ahí los archivos vienen como `680010079201_HUS311736_FACTURA.pdf`.
    """
    indice: dict[str, Path] = {}
    for r in sorted(carpeta.rglob("*.pdf")):
        numero = factura_del_nombre(r.name)
        if numero:
            indice.setdefault(numero, r)
    return indice


# La factura que viene con el XML no siempre es solo la factura. En el paquete
# 31068 el `680010079201_HUS######_FACTURA.pdf` trae los cuatro renglones ya
# pegados: la factura con CUFE, el detallado, la representación gráfica de la
# DIAN y la nota crédito. Si el bot le agregara encima el detallado del Excel,
# el folio subiría al ADRES con el detallado DOS VECES. Por eso se mira qué trae
# adentro antes de tocarlo.
MARCAS_RENGLON: tuple[tuple[str, str], ...] = (
    ("DETALLADO", "DETALLADO FACTURA"),
    ("DIAN", "REPRESENTACION GRAFICA"),
    ("NOTAS", "NOTA CREDITO"),
)

# Cuántas páginas se leen buscando esas marcas. Las traen al principio; leer el
# PDF entero de 324 facturas sobre el share sería lento sin ganar nada.
PAGINAS_A_MIRAR = 25


def renglones_que_trae(pdf: Path, PdfReader=None) -> set[str]:
    """Qué renglones del folio ya vienen pegados dentro de este PDF.

    Devuelve las claves (`DETALLADO`, `DIAN`, `NOTAS`) que encontró. Si el PDF
    no se puede leer, devuelve vacío: se prefiere no afirmar nada a adivinar.
    """
    if PdfReader is None:
        PdfReader, _ = _cargar_lector_escritor()
    faltan = {clave: _norm(marca) for clave, marca in MARCAS_RENGLON}
    trae: set[str] = set()
    try:
        lector = PdfReader(str(pdf))
        for pagina in list(lector.pages)[:PAGINAS_A_MIRAR]:
            if not faltan:
                break
            limpio = _norm(pagina.extract_text() or "")
            for clave in [c for c, marca in faltan.items() if marca in limpio]:
                trae.add(clave)
                faltan.pop(clave)
    except Exception as e:  # noqa: BLE001 - un PDF ilegible no tumba el lote
        logger.debug("No pude mirar dentro de %s: %s", pdf.name, e)
    return trae


ESTADO_COPIADA = "COPIADA"
ESTADO_SE_COPIARIA = "SE COPIARIA"
ESTADO_YA_ESTABA = "YA ESTABA"
ESTADO_SIN_FACTURA = "NO ESTA LA FACTURA"


@dataclass
class Copia:
    """La factura que se trae de la carpeta del XML a la carpeta de la factura."""

    factura: str
    origen: Path | None
    destino: Path | None
    estado: str


def copiar_facturas(
    plan: list[Factura], indice: dict[str, Path], aplicar: bool = False
) -> list[Copia]:
    """Trae a cada carpeta el PDF de su factura. No pisa lo que ya esté.

    Sin `aplicar` no copia nada, pero **sí deja la factura anotada en el plan**:
    así la simulación muestra el folio como va a quedar de verdad, y no uno
    incompleto que asustaría al auditor.
    """
    copias: list[Copia] = []
    for fac in plan:
        if any(s.grupo.clave == "FACTURA" for s in fac.soportes_factura):
            copias.append(Copia(fac.factura, None, None, ESTADO_YA_ESTABA))
            continue
        origen = indice.get(fac.factura)
        if origen is None:
            copias.append(Copia(fac.factura, None, None, ESTADO_SIN_FACTURA))
            continue
        destino = fac.carpeta / origen.name
        if aplicar:
            shutil.copy2(origen, destino)
        _sumar_al_folio(fac, destino, GRUPO_FACTURA)
        # La factura viene nombrada con el NIT: si la carpeta no lo traía, el
        # folio ya puede llevarlo.
        if not fac.prefijo and (nit := prefijo_del_nombre(origen.stem, fac.factura)):
            fac.prefijo = nit
            fac.destino_epicris = fac.carpeta / nombre_folio(nit, fac.factura, SUFIJO_EPICRIS)
            fac.destino_factura = fac.carpeta / nombre_folio(nit, fac.factura, SUFIJO_FACTURA)
        copias.append(
            Copia(fac.factura, origen, destino, ESTADO_COPIADA if aplicar else ESTADO_SE_COPIARIA)
        )
    return copias


def revisar_facturas(plan: list[Factura], indice: dict[str, Path] | None = None) -> None:
    """Mira dentro del PDF de cada factura qué renglones ya trae pegados.

    Hace falta porque en el paquete 31068 el `..._FACTURA.pdf` que viene con el
    XML **ya es el folio completo**: factura + detallado + representación
    gráfica DIAN + nota crédito. Sin esta revisión el bot le pegaría encima el
    detallado del Excel y el folio subiría con el detallado dos veces.
    """
    PdfReader, _ = _cargar_lector_escritor()
    for fac in plan:
        soporte = next((s for s in fac.soportes_factura if s.grupo.clave == "FACTURA"), None)
        origen: Path | None = None
        if soporte is not None and soporte.ruta.exists():
            origen = soporte.ruta
        elif indice is not None:
            origen = indice.get(fac.factura)
        if origen is not None and origen.exists():
            fac.trae_la_factura = renglones_que_trae(origen, PdfReader)


def _sumar_al_folio(fac: Factura, ruta: Path, grupo: Grupo) -> None:
    """Mete un archivo en el folio de la factura, en el renglón que le toca."""
    fac.soportes_factura.append(Soporte(ruta=ruta, grupo=grupo))
    fac.soportes_factura.sort(key=lambda s: (s.grupo.orden, clave_natural(s.ruta.name)))
    fac.estado_factura = ESTADO_SIMULADO


# ─── El detallado, que sale en Excel y al folio solo entran PDF ──────────────


ESTADO_SE_PASARIA = "SE PASARIA A PDF"


def convertir_detallados(plan: list[Factura], aplicar: bool = False) -> list[tuple[str, Path, str]]:
    """Pasa a PDF el detallado que todavía está en Excel (renglón 2 del folio).

    Usa el mismo motor del bot de conversión masiva: el Excel del equipo si
    está, y si no LibreOffice. Si no hay ninguno de los dos, no revienta: lo
    deja anotado para que el auditor lo convierta y vuelva a correr.
    """
    # Si la propia factura ya trae el detallado pegado adentro, no se le agrega
    # otro encima: el folio quedaría con el detallado dos veces.
    pendientes = [
        (fac, d)
        for fac in plan
        for d in fac.detallados_sin_pdf
        if "DETALLADO" not in fac.trae_la_factura
    ]
    if not pendientes:
        return []
    if not aplicar:
        # Igual que con la factura: la simulación muestra el folio como va a
        # quedar, con el detallado ya adentro.
        for fac, d in pendientes:
            _sumar_al_folio(fac, d.with_suffix(".pdf"), GRUPO_DETALLADO)
        return [(fac.factura, d, ESTADO_SE_PASARIA) for fac, d in pendientes]

    import excel_a_pdf

    try:
        motor, binario = excel_a_pdf.elegir_motor("auto", None)
    except SystemExit as e:  # sin Excel ni LibreOffice en el equipo
        return [(fac.factura, d, f"{ESTADO_ERROR}: {e}") for fac, d in pendientes]

    convs = [excel_a_pdf.Conversion(origen=d, destino=d.with_suffix(".pdf")) for _, d in pendientes]
    if motor == "excel":
        excel_a_pdf.convertir_con_excel(convs)
    else:
        excel_a_pdf.convertir_con_libreoffice(convs, str(binario))

    hechos: list[tuple[str, Path, str]] = []
    for (fac, origen), conv in zip(pendientes, convs):
        if conv.estado == "OK":
            _sumar_al_folio(fac, conv.destino, GRUPO_DETALLADO)
            fac.detallados.append(conv.destino)
            fac.detallados_sin_pdf = [d for d in fac.detallados_sin_pdf if d != origen]
            hechos.append((fac.factura, origen, "PASADO A PDF"))
        else:
            hechos.append((fac.factura, origen, f"{ESTADO_ERROR}: {conv.detalle}"))
    return hechos


# ─── Reporte ─────────────────────────────────────────────────────────────────

COLUMNAS_REPORTE = (
    "FACTURA",
    "ORDEN",
    "GRUPO",
    "ARCHIVO",
    "RECONOCIDO",
    "CARPETA",
    "ESTADO",
    "OBSERVACION",
    "FOLIO",
)


def escribir_reporte(ruta: Path, plan: list[Factura]) -> None:
    """Un renglón por soporte: en qué folio y en qué orden va."""
    ruta.parent.mkdir(parents=True, exist_ok=True)
    with open(ruta, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh, delimiter=";")
        w.writerow(COLUMNAS_REPORTE)
        for fac in plan:
            for n, s in enumerate(fac.soportes, start=1):
                w.writerow(
                    [
                        fac.factura,
                        n,
                        s.grupo.titulo,
                        s.ruta.name,
                        "SI" if s.reconocido else "NO - revisar",
                        str(s.ruta.parent),
                        fac.estado,
                        "",
                        FOLIO_EPICRIS,
                    ]
                )
            for n, s in enumerate(fac.soportes_factura, start=1):
                w.writerow(
                    [
                        fac.factura,
                        n,
                        s.grupo.titulo,
                        s.ruta.name,
                        "SI" if s.reconocido else "NO - revisar",
                        str(s.ruta.parent),
                        fac.estado_factura,
                        "",
                        FOLIO_FACTURA,
                    ]
                )
            for det in fac.detallados_sin_pdf:
                w.writerow(
                    [
                        fac.factura,
                        "",
                        "DETALLADO",
                        det.name,
                        "SI",
                        str(det.parent),
                        fac.estado_factura,
                        "está en Excel: hay que pasarlo a PDF para que entre al folio",
                        FOLIO_FACTURA,
                    ]
                )
            if not fac.soportes:
                w.writerow(
                    [
                        fac.factura,
                        "",
                        "",
                        "",
                        "",
                        str(fac.carpeta),
                        fac.estado,
                        "no hay PDF que unir",
                        FOLIO_EPICRIS,
                    ]
                )
            for falta in fac.faltantes():
                w.writerow(
                    [
                        fac.factura,
                        "",
                        falta,
                        "",
                        "NO",
                        str(fac.carpeta),
                        fac.estado,
                        "FALTA este soporte",
                        FOLIO_EPICRIS,
                    ]
                )
            for falta in fac.faltantes_factura():
                w.writerow(
                    [
                        fac.factura,
                        "",
                        falta,
                        "",
                        "NO",
                        str(fac.carpeta),
                        fac.estado_factura,
                        "FALTA este soporte",
                        FOLIO_FACTURA,
                    ]
                )
            if fac.notas_pendientes():
                w.writerow(
                    [
                        fac.factura,
                        "",
                        "NOTAS CREDITO",
                        "",
                        "",
                        str(fac.carpeta),
                        fac.estado_factura,
                        "PENDIENTE: las notas crédito de los valores aceptados aún no existen",
                        FOLIO_FACTURA,
                    ]
                )
            for previo in fac.folios_previos:
                dudoso = previo in fac.folios_dudosos
                w.writerow(
                    [
                        fac.factura,
                        "",
                        "",
                        previo.name,
                        "NO - revisar" if dudoso else "SI",
                        str(previo.parent),
                        "",
                        (
                            "REVISAR: se tomó como el folio de la corrida anterior, pero al "
                            "renglón le falta su archivo numerado. Si es un soporte nuevo, "
                            "renómbrelo con su número y vuelva a correr el bot."
                            if dudoso
                            else "es el folio de una corrida anterior: se vuelve a armar"
                        ),
                        "",
                    ]
                )
            for om in fac.omitidos:
                w.writerow([fac.factura, "", "", "", "", str(fac.carpeta), ESTADO_ERROR, om, ""])


# ─── Lista de facturas a trabajar ────────────────────────────────────────────


def leer_facturas(ruta: Path) -> set[str]:
    """Los números de factura de un Excel de una sola columna (o la que los traiga)."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(ruta), data_only=True, read_only=True)
    try:
        numeros: set[str] = set()
        for hoja in wb.worksheets:
            for fila in hoja.iter_rows(values_only=True):
                for celda in fila or ():
                    texto = str(celda or "").strip()
                    if re.fullmatch(r"HUS\s*0*\d+", texto, re.IGNORECASE):
                        numeros.add(_norm(f"HUS{re.sub(r'[^0-9]', '', texto).lstrip('0')}"))
        return numeros
    finally:
        wb.close()


# ─── CLI ─────────────────────────────────────────────────────────────────────


def _cargar_mapa(ruta: Path | None) -> dict[str, str]:
    if ruta is None:
        return {}
    try:
        datos = json.loads(ruta.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        logger.error("No pude leer el mapa de nombres %s: %s", ruta, e)
        raise SystemExit(2) from None
    if not isinstance(datos, dict):
        logger.error('El mapa debe ser un objeto JSON {"ANGIOTAC": "AYUDAS", ...}')
        raise SystemExit(2)
    return {str(k): str(v) for k, v in datos.items()}


def construir_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Une los soportes de cada factura en un solo PDF, en el orden del ADRES.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument(
        "--carpeta",
        type=Path,
        required=True,
        help="Carpeta del gestor (CAROLINA, CLAUDIA, OSCAR…).",
    )
    p.add_argument(
        "--facturas",
        type=Path,
        help="Excel con las facturas a trabajar. Sin esto, todas las carpetas.",
    )
    p.add_argument(
        "--mapa-nombres", type=Path, help='JSON para agregar palabras: {"ANGIOTAC": "AYUDAS"}.'
    )
    p.add_argument(
        "--renombrar",
        action="store_true",
        help="Solo dejar cada soporte con su nombre numerado dentro del folio: "
        "«1 RESPUESTA A GLOSA.pdf», «2 HISTORIA CLINICA.pdf»…",
    )
    p.add_argument(
        "--folio",
        action="store_true",
        help="El trabajo completo: numera los soportes y arma los DOS folios de cada "
        "factura, «<NIT>_<FACTURA>_EPICRIS.pdf» y «<NIT>_<FACTURA>_FACTURA.pdf».",
    )
    p.add_argument(
        "--carpeta-facturas",
        type=Path,
        help="Carpeta «4.FACTURAS CON XML\\XML», de donde se trae el PDF de cada factura.",
    )
    p.add_argument(
        "--prefijo",
        default="",
        help="NIT con que se nombran los folios (680010079201), por si los archivos "
        "de una carpeta no lo traen. Sin esto se usa el que traigan los archivos.",
    )
    p.add_argument(
        "--convertir-detallado",
        action="store_true",
        help="Pasar a PDF el detallado que esté en Excel, para que entre al folio de la factura.",
    )
    p.add_argument(
        "--aplicar", action="store_true", help="Hacerlo de verdad. Sin esto solo muestra qué haría."
    )
    p.add_argument("--reporte-csv", type=Path, help="Listado de qué archivo quedó en qué grupo.")
    p.add_argument("--log", type=Path, help="Guarda además el log en un archivo.")
    return p


def _mostrar_folio(titulo: str, destino: Path | None, soportes: list[Soporte]) -> None:
    """Un folio en pantalla: el nombre que queda y qué entra adentro, en orden."""
    if not soportes:
        logger.info("     %s: no hay nada todavía", titulo)
        return
    logger.info("     %s → %s", titulo, (destino or Path()).name)
    for n, s in enumerate(soportes, start=1):
        marca = "" if s.reconocido else "   <-- no reconocido, va en OTROS"
        logger.info(
            "        %-34s (era %s)%s",
            nombre_numerado(n, s.grupo, s.ruta.suffix),
            s.ruta.name,
            marca,
        )


def _resumen_folios(
    plan: list[Factura],
    copias: list[Copia],
    conversiones: list[tuple[str, Path, str]],
    aplicar: bool,
) -> None:
    """Lo que quedó (o quedaría) en cada carpeta: los dos folios de la factura."""
    verbo = "Se armaron" if aplicar else "Se armarían"
    con_algo = [f for f in plan if f.soportes or f.soportes_factura]
    logger.info("\n%s los folios de %d factura(s).", verbo, len(con_algo))
    if aplicar:
        logger.info(
            "Páginas escritas: %d en los EPICRIS y %d en las FACTURAS.",
            sum(f.paginas for f in plan),
            sum(f.paginas_factura for f in plan),
        )

    for fac in con_algo[:3]:
        logger.info("\n  %s\\", fac.carpeta.name)
        _mostrar_folio("FOLIO CLINICO", fac.destino_epicris, fac.soportes)
        _mostrar_folio("FOLIO FACTURA", fac.destino_factura, fac.soportes_factura)
    if len(con_algo) > 3:
        logger.info(
            "\n  … y %d factura(s) más (el detalle completo va en el reporte CSV).",
            len(con_algo) - 3,
        )

    if copias:
        cuenta: dict[str, int] = {}
        for c in copias:
            cuenta[c.estado] = cuenta.get(c.estado, 0) + 1
        logger.info(
            "\nFactura traída de la carpeta del XML: %s",
            ", ".join(f"{n} {estado}" for estado, n in sorted(cuenta.items())),
        )
        sin_factura = [c.factura for c in copias if c.estado == ESTADO_SIN_FACTURA]
        if sin_factura:
            logger.info(
                "   Sin PDF de factura (%d): %s", len(sin_factura), ", ".join(sin_factura[:12])
            )

    if conversiones:
        logger.info("\nDetallados que estaban en Excel (%d):", len(conversiones))
        for factura, origen, estado in conversiones[:10]:
            logger.info("   %s: %s — %s", factura, origen.name, estado)
        if len(conversiones) > 10:
            logger.info("   … y %d más, en el reporte CSV.", len(conversiones) - 10)

    faltan_factura = [(f, f.faltantes_factura()) for f in plan if f.faltantes_factura()]
    if faltan_factura:
        logger.info(
            "\nAl folio de la FACTURA le falta un renglón en %d factura(s):", len(faltan_factura)
        )
        for fac, cuales in faltan_factura[:10]:
            logger.info("   %s: falta %s", fac.factura, ", ".join(cuales))
        if len(faltan_factura) > 10:
            logger.info("   … y %d más, en el reporte CSV.", len(faltan_factura) - 10)

    pendientes = [f for f in plan if f.notas_pendientes()]
    if pendientes:
        logger.info(
            "\nNOTAS CRÉDITO: pendientes en %d factura(s). No es un error — todavía no las "
            "han sacado. Cuando lleguen, se vuelve a correr el bot y entran solas de cuartas.",
            len(pendientes),
        )

    completas = [f for f in plan if f.trae_la_factura]
    if completas:
        cuenta: dict[str, int] = {}
        for f in completas:
            for clave in f.trae_la_factura:
                cuenta[clave] = cuenta.get(clave, 0) + 1
        logger.info(
            "\nLa factura ya trae renglones pegados adentro en %d factura(s): %s.",
            len(completas),
            ", ".join(f"{n} con {clave}" for clave, n in sorted(cuenta.items())),
        )
        logger.info("   No se les agrega otro encima: el folio quedaría con el renglón dos veces.")

    dudosos = [(f, p) for f in plan for p in f.folios_dudosos]
    if dudosos:
        logger.info(
            "\nOJO, revise estos archivos (%d): se tomaron como el folio de la corrida "
            "anterior, pero al renglón le falta su archivo numerado. Si alguno es un "
            "soporte nuevo, renómbrelo con su número y vuelva a correr el bot:",
            len(dudosos),
        )
        for fac, p in dudosos[:10]:
            logger.info("   %s: %s", fac.factura, p.name)
        if len(dudosos) > 10:
            logger.info("   … y %d más, en el reporte CSV.", len(dudosos) - 10)

    sin_nit = [f.factura for f in plan if not f.prefijo and (f.soportes or f.soportes_factura)]
    if sin_nit:
        logger.info(
            "\nSin el NIT en el nombre (%d): el folio queda como «HUS######_EPICRIS.pdf». "
            "Si hace falta el NIT, agregue --prefijo 680010079201. Facturas: %s",
            len(sin_nit),
            ", ".join(sin_nit[:12]),
        )


def main(argv: list[str] | None = None) -> int:
    args = construir_parser().parse_args(argv)
    handlers: list[logging.Handler] = [logging.StreamHandler(sys.stdout)]
    if args.log is not None:
        args.log.parent.mkdir(parents=True, exist_ok=True)
        handlers.append(logging.FileHandler(args.log, encoding="utf-8"))
    logging.basicConfig(level=logging.INFO, format="%(message)s", handlers=handlers)

    if not args.carpeta.is_dir():
        logger.error("No existe la carpeta: %s", args.carpeta)
        return 1

    facturas: set[str] | None = None
    if args.facturas:
        if not args.facturas.is_file():
            logger.error("No existe el Excel de facturas: %s", args.facturas)
            return 1
        facturas = leer_facturas(args.facturas)
        logger.info("Lista de trabajo: %d facturas", len(facturas))

    mapa = _cargar_mapa(args.mapa_nombres)
    copias: list[Copia] = []
    conversiones: list[tuple[str, Path, str]] = []
    if args.folio:
        # El trabajo completo: los dos folios de cada factura.
        plan = planificar(args.carpeta, facturas, mapa)
        indice: dict[str, Path] | None = None
        if args.carpeta_facturas:
            if not args.carpeta_facturas.is_dir():
                logger.error("No existe la carpeta de facturas: %s", args.carpeta_facturas)
                return 1
            indice = indice_facturas(args.carpeta_facturas)
            copias = copiar_facturas(plan, indice, args.aplicar)
        # Antes de agregarle nada, mirar qué trae la factura pegado adentro.
        revisar_facturas(plan, indice)
        if args.convertir_detallado:
            conversiones = convertir_detallados(plan, args.aplicar)
        aplicar_folios(plan, args.aplicar, args.prefijo)
    elif args.renombrar:
        # Solo numerar: cada soporte queda como «1 RESPUESTA A GLOSA.pdf», sin
        # pegar nada en un solo PDF.
        plan = planificar(args.carpeta, facturas, mapa)
        for fac in plan:
            if not fac.soportes:
                continue
            renombrar_en_orden(fac, args.aplicar)
            if args.aplicar:
                fac.estado = ESTADO_RENOMBRADO
    else:
        plan = unir(args.carpeta, facturas, mapa, args.aplicar)
    if args.reporte_csv:
        escribir_reporte(args.reporte_csv, plan)

    if not plan:
        logger.info("\nNo encontré carpetas de factura en %s", args.carpeta)
        return 0

    con_pdf = [f for f in plan if f.soportes]
    sin_pdf = [f for f in plan if not f.soportes]
    errores = [f for f in plan if ESTADO_ERROR in (f.estado, f.estado_factura)]
    sin_reconocer = [(f, s) for f in plan for s in f.soportes if not s.reconocido]

    if args.folio:
        _resumen_folios(plan, copias, conversiones, args.aplicar)
    else:
        if args.renombrar:
            verbo = "Se renombraron" if args.aplicar else "Se renombrarían"
        else:
            verbo = "Se unieron" if args.aplicar else "Se unirían"
        logger.info(
            "\n%s los soportes de %d factura(s): %d archivo(s).",
            verbo,
            len(con_pdf),
            sum(len(f.soportes) for f in con_pdf),
        )
        if args.aplicar and not args.renombrar:
            logger.info("Páginas escritas: %d", sum(f.paginas for f in plan))

        for fac in con_pdf[:3]:
            if args.renombrar:
                logger.info("\n  %s\\", fac.carpeta.name)
            else:
                logger.info("\n  %s → %s", fac.factura, (fac.destino or Path()).name)
            for n, s in enumerate(fac.soportes, start=1):
                marca = "" if s.reconocido else "   <-- no reconocido, va en OTROS"
                if args.renombrar:
                    logger.info(
                        "     %-36s (era %s)%s",
                        nombre_numerado(n, s.grupo, s.ruta.suffix),
                        s.ruta.name,
                        marca,
                    )
                else:
                    logger.info("     %2d. [%s] %s%s", n, s.grupo.titulo, s.ruta.name, marca)
        if len(con_pdf) > 3:
            logger.info(
                "\n  … y %d factura(s) más (el detalle completo va en el reporte CSV).",
                len(con_pdf) - 3,
            )

    if sin_reconocer:
        logger.info(
            "\nArchivos que NO se reconocieron (%d) — quedaron en OTROS:", len(sin_reconocer)
        )
        for fac, s in sin_reconocer[:10]:
            logger.info("   %s: %s", fac.factura, s.ruta.name)
        if len(sin_reconocer) > 10:
            logger.info("   … y %d más, en el reporte CSV.", len(sin_reconocer) - 10)

    faltan = [(f, f.faltantes()) for f in plan if f.faltantes()]
    if faltan:
        logger.info("\nFacturas a las que les falta un soporte obligatorio (%d):", len(faltan))
        for fac, cuales in faltan[:10]:
            logger.info("   %s: falta %s", fac.factura, ", ".join(cuales))
        if len(faltan) > 10:
            logger.info("   … y %d más, en el reporte CSV.", len(faltan) - 10)

    if sin_pdf:
        logger.info(
            "\nCarpetas sin PDF que unir (%d): %s",
            len(sin_pdf),
            ", ".join(f.factura for f in sin_pdf[:12]),
        )
    if errores:
        logger.info("\nCon error (%d):", len(errores))
        for fac in errores:
            logger.info("   %s: %s", fac.factura, "; ".join(fac.omitidos))

    if not args.aplicar:
        logger.info(
            "\n*** SIMULACIÓN: no se escribió ningún PDF. Agregue --aplicar para hacerlo. ***"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
