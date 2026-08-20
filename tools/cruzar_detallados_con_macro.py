"""cruzar_detallados_con_macro.py — ¿están todas las facturas del paquete?

Compara la **carpeta de detallados por factura** contra la **macro de respuesta**
del paquete y arma un consolidado que responde dos preguntas:

1. ¿Es la misma cantidad de facturas de un lado y del otro?
2. Si no, ¿cuáles faltan y **cuánta plata** representan?

Las que están en la macro y no tienen detallado son las importantes: el hospital
las está reclamando pero no tiene el papel que lo sustenta.

USO (Windows, desde C:\\temp-notas):

    py tools\\cruzar_detallados_con_macro.py ^
        --detallados "D:\\...\\EXCEL_POR_FACTURA" ^
        --macro      "D:\\...\\RTA GLOSA ADRES PAQ 31068.xlsx" ^
        --salida     "D:\\...\\CONSOLIDADO_31068.xlsx"

Requiere una sola vez: `py -m pip install openpyxl`

Deja un Excel con tres hojas: **RESUMEN**, **FACTURAS** (todas, con su estado y
sus cifras) y **SIN DETALLADO** (solo las que faltan, para trabajarlas).
"""

from __future__ import annotations

import argparse
import logging
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from _dinero import a_numero  # noqa: E402
from ajustar_detallado_glosas import (  # noqa: E402
    IndiceHoja,
    _abrir_libro,
    normalizar_factura,
    segmentar_facturas,
)

logger = logging.getLogger("cruzar_detallados_con_macro")

EXTENSIONES = (".xlsx", ".xlsm")

# Columnas de la macro (1 = A). El encabezado está en la fila 2.
COL_RADICACION = 2
COL_FACTURA = 3
COL_PAQUETE = 4
COL_VALOR_RECLAMADO = 6
COL_VALOR_GLOSADO = 9
COL_DESC_GLOSA = 15
COL_OBSERVACION = 19
COL_VALOR_ACEPTADO = 23

ESTADO_COMPLETA = "CON DETALLADO"
ESTADO_SIN_DETALLADO = "SIN DETALLADO"
ESTADO_SOBRA = "SIN GLOSAS EN LA MACRO"


@dataclass
class FacturaMacro:
    """Lo que la macro dice de una factura."""

    factura: str = ""
    filas: int = 0
    glosas_por_responder: int = 0  # sin las glosas totales, que no se responden
    radicados: set[str] = field(default_factory=set)
    paquetes: set[str] = field(default_factory=set)
    observaciones: dict[str, int] = field(default_factory=lambda: defaultdict(int))
    valor_reclamado: float = 0.0
    valor_glosado: float = 0.0
    valor_aceptado: float = 0.0


@dataclass
class Cruce:
    """Una factura, esté donde esté."""

    factura: str
    estado: str
    archivo: str = ""
    macro: FacturaMacro | None = None

    @property
    def motivo(self) -> str:
        if self.estado != ESTADO_SIN_DETALLADO:
            return ""
        return (
            "El sistema del hospital no exportó el detallado de esta factura. "
            "Hay que pedir esa impresión para poder responder."
        )


def leer_macro(ruta: Path, paquete: str | None = None) -> dict[str, FacturaMacro]:
    """Las facturas de la macro, con sus cifras, agrupadas por factura.

    El encabezado de la macro va en la **segunda** fila: la primera trae el
    título con el total. Se toma toda fila cuyo «Número Factura» empiece por
    HUS, así que sirve igual si el archivo cambia de tamaño.
    """
    wb = _abrir_libro(ruta, solo_datos=True, solo_lectura=True)
    try:
        salida: dict[str, FacturaMacro] = {}
        for fila in wb.worksheets[0].iter_rows(min_row=2, values_only=True):
            if len(fila) < COL_VALOR_ACEPTADO:
                continue
            numero = str(fila[COL_FACTURA - 1] or "").strip()
            if not numero.upper().startswith("HUS"):
                continue
            paq = str(fila[COL_PAQUETE - 1] or "").strip()
            if paquete and paq != str(paquete):
                continue
            dato = salida.setdefault(normalizar_factura(numero), FacturaMacro(factura=numero))
            dato.filas += 1
            # Las filas con «Descripción Glosa» vacía son el desglose de una
            # reclamación glosada entera: no se responden una por una.
            if str(fila[COL_DESC_GLOSA - 1] or "").strip():
                dato.glosas_por_responder += 1
            dato.radicados.add(str(fila[COL_RADICACION - 1] or "").strip())
            dato.paquetes.add(paq)
            dato.observaciones[
                str(fila[COL_OBSERVACION - 1] or "").strip().upper() or "(vacía)"
            ] += 1
            dato.valor_reclamado += a_numero(fila[COL_VALOR_RECLAMADO - 1])
            dato.valor_glosado += a_numero(fila[COL_VALOR_GLOSADO - 1])
            dato.valor_aceptado += a_numero(fila[COL_VALOR_ACEPTADO - 1])
        return salida
    finally:
        wb.close()


def leer_detallados(carpeta: Path, patron: str = "*.xlsx") -> dict[str, str]:
    """{factura: nombre del archivo} de la carpeta de detallados.

    La factura se toma del **contenido** del archivo, no del nombre: es lo que
    de verdad se va a radicar. Si el archivo no dice qué factura es, se usa el
    nombre como respaldo.
    """
    salida: dict[str, str] = {}
    for ruta in sorted(carpeta.glob(patron)):
        if ruta.suffix.lower() not in EXTENSIONES or ruta.name.startswith(("~$", ".")):
            continue
        numero = ""
        try:
            wb = _abrir_libro(ruta, solo_datos=True)
            try:
                for fac in segmentar_facturas(IndiceHoja(wb.worksheets[0])):
                    if fac.factura:
                        numero = fac.factura
                        break
            finally:
                wb.close()
        except Exception as e:  # noqa: BLE001 - un archivo malo no tumba el cruce
            logger.warning("  %s: %s", ruta.name, e)
        salida[normalizar_factura(numero or ruta.stem)] = ruta.name
    return salida


def cruzar(detallados: dict[str, str], macro: dict[str, FacturaMacro]) -> list[Cruce]:
    """Une los dos lados. Devuelve una línea por factura, ordenada."""
    salida = []
    for clave in sorted(set(detallados) | set(macro), key=lambda k: (k not in macro, k)):
        en_macro, en_carpeta = macro.get(clave), detallados.get(clave)
        if en_macro and en_carpeta:
            estado = ESTADO_COMPLETA
        elif en_macro:
            estado = ESTADO_SIN_DETALLADO
        else:
            estado = ESTADO_SOBRA
        salida.append(
            Cruce(
                factura=en_macro.factura if en_macro else (en_carpeta or clave),
                estado=estado,
                archivo=en_carpeta or "",
                macro=en_macro,
            )
        )
    return salida


def _hoja(wb, titulo: str, encabezados: list[str]):
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet(titulo)
    ws.append(encabezados)
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="1F4E79")
        celda.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    return ws


def _anchos(ws, anchos: dict[str, int]) -> None:
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho


def escribir_consolidado(ruta: Path, cruces: list[Cruce]) -> None:
    """El Excel para el auditor: RESUMEN, FACTURAS y SIN DETALLADO."""
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    con = [c for c in cruces if c.estado == ESTADO_COMPLETA]
    sin = [c for c in cruces if c.estado == ESTADO_SIN_DETALLADO]
    sobra = [c for c in cruces if c.estado == ESTADO_SOBRA]
    plata_sin = sum(c.macro.valor_glosado for c in sin if c.macro)
    aceptado_sin = sum(c.macro.valor_aceptado for c in sin if c.macro)

    ws = _hoja(wb, "RESUMEN", ["", ""])
    ws.delete_rows(1)
    filas = [
        ("Facturas en la macro", len(con) + len(sin)),
        ("Facturas con detallado en la carpeta", len(con) + len(sobra)),
        ("", ""),
        ("Están en los dos lados", len(con)),
        ("En la macro pero SIN detallado", len(sin)),
        ("Con detallado pero sin glosas en la macro", len(sobra)),
        ("", ""),
        ("Valor glosado de las que no tienen detallado", plata_sin),
        ("De eso, ya aceptado por el hospital", aceptado_sin),
    ]
    for etiqueta, valor in filas:
        ws.append([etiqueta, valor])
        if isinstance(valor, float):
            ws.cell(row=ws.max_row, column=2).number_format = '"$"#,##0'
        ws.cell(row=ws.max_row, column=1).font = Font(bold=not etiqueta.startswith(" "))
    _anchos(ws, {"A": 46, "B": 18})

    ws = _hoja(
        wb,
        "FACTURAS",
        [
            "FACTURA",
            "ESTADO",
            "ARCHIVO DEL DETALLADO",
            "RADICADO",
            "PAQUETE",
            "FILAS EN LA MACRO",
            "GLOSAS POR RESPONDER",
            "VALOR RECLAMADO",
            "VALOR GLOSADO",
            "VALOR ACEPTADO",
            "QUÉ PASÓ",
        ],
    )
    for c in cruces:
        m = c.macro
        ws.append(
            [
                c.factura,
                c.estado,
                c.archivo,
                " / ".join(sorted(x for x in m.radicados if x)) if m else "",
                " / ".join(sorted(x for x in m.paquetes if x)) if m else "",
                m.filas if m else 0,
                m.glosas_por_responder if m else 0,
                m.valor_reclamado if m else 0.0,
                m.valor_glosado if m else 0.0,
                m.valor_aceptado if m else 0.0,
                c.motivo,
            ]
        )
        for col in (8, 9, 10):
            ws.cell(row=ws.max_row, column=col).number_format = '"$"#,##0'
    _anchos(
        ws,
        {
            "A": 14,
            "B": 22,
            "C": 24,
            "D": 14,
            "E": 10,
            "F": 10,
            "G": 12,
            "H": 16,
            "I": 16,
            "J": 16,
            "K": 60,
        },
    )

    ws = _hoja(
        wb,
        "SIN DETALLADO",
        [
            "FACTURA",
            "RADICADO",
            "FILAS EN LA MACRO",
            "GLOSAS POR RESPONDER",
            "VALOR RECLAMADO",
            "VALOR GLOSADO",
            "VALOR ACEPTADO",
            "CÓMO QUEDÓ RESPONDIDA",
            "QUÉ HAY QUE HACER",
        ],
    )
    for c in sin:
        m = c.macro
        assert m is not None
        obs = " · ".join(f"{k}: {v}" for k, v in sorted(m.observaciones.items()))
        ws.append(
            [
                c.factura,
                " / ".join(sorted(x for x in m.radicados if x)),
                m.filas,
                m.glosas_por_responder,
                m.valor_reclamado,
                m.valor_glosado,
                m.valor_aceptado,
                obs,
                c.motivo,
            ]
        )
        for col in (5, 6, 7):
            ws.cell(row=ws.max_row, column=col).number_format = '"$"#,##0'
    _anchos(ws, {"A": 14, "B": 14, "C": 10, "D": 12, "E": 16, "F": 16, "G": 16, "H": 40, "I": 60})

    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta)


def construir_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Cruza la carpeta de detallados contra la macro del paquete.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("--detallados", type=Path, required=True, help="Carpeta de detallados.")
    p.add_argument("--macro", type=Path, required=True, help="Excel de la macro de respuesta.")
    p.add_argument("--salida", type=Path, required=True, help="Excel del consolidado.")
    p.add_argument("--patron", default="*.xlsx", help="Qué archivos tomar de la carpeta.")
    p.add_argument("--paquete", help="Filtrar por número de paquete, si la macro trae varios.")
    return p


def main(argv: list[str] | None = None) -> int:
    args = construir_parser().parse_args(argv)
    logging.basicConfig(level=logging.INFO, format="%(message)s")

    if not args.detallados.is_dir():
        logger.error("No existe la carpeta de detallados: %s", args.detallados)
        return 1
    detallados = leer_detallados(args.detallados, args.patron)
    macro = leer_macro(args.macro, args.paquete)
    if not macro:
        logger.error("La macro no trae ninguna factura (¿es el archivo correcto?).")
        return 1

    cruces = cruzar(detallados, macro)
    escribir_consolidado(args.salida, cruces)

    con = [c for c in cruces if c.estado == ESTADO_COMPLETA]
    sin = [c for c in cruces if c.estado == ESTADO_SIN_DETALLADO]
    sobra = [c for c in cruces if c.estado == ESTADO_SOBRA]

    def pesos(v: float) -> str:
        return "$" + f"{int(round(v)):,}".replace(",", ".")

    print(f"\nFacturas en la macro     : {len(macro)}")
    print(f"Detallados en la carpeta : {len(detallados)}")
    print(f"\n  están en los dos lados : {len(con)}")
    if sin:
        plata = sum(c.macro.valor_glosado for c in sin if c.macro)
        print(f"  SIN detallado          : {len(sin)}  (glosado {pesos(plata)})")
        for c in sin:
            m = c.macro
            assert m is not None
            print(
                f"     {c.factura:<12} radicado {' / '.join(sorted(m.radicados)):<10} "
                f"{m.filas:>3} glosas   glosado {pesos(m.valor_glosado):>14}   "
                f"aceptado {pesos(m.valor_aceptado):>12}"
            )
    if sobra:
        print(f"  con detallado pero sin glosas en la macro : {len(sobra)}")
        for c in sobra[:10]:
            print(f"     {c.factura}  ({c.archivo})")
    print(f"\nConsolidado en: {args.salida}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
