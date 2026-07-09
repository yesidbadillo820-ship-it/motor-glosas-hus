"""corregir_errores_dgh.py — Lee los ERRORES que devuelve DGH y arma el OBJECIONES de reintento.

Cuando se sube un OBJECIONES a DGH, el sistema devuelve un Excel de validación
(hoja VALIDACIONIMPORTACION). Este bot cruza ese archivo con el OBJECIONES que
se subió y produce "OBJECIONES REINTENTO.xlsx" listo para volver a cargar:

  1. Facturas "ya se encuentra (Objetada) / (Glosa_Total)": se QUITAN.
     Es lo normal cuando la factura ya se había subido en un cargue anterior —
     ya quedó objetada en DGH y no hay nada que reintentar.
  2. "El VALOR OBJECION no puede ser mayor al valor del servicio (V)": se busca
     la fila responsable y se CAPA su valor a V (lo máximo que DGH acepta).
     La fila se identifica en este orden:
       a) es la única fila de la factura con valor mayor a V;
       b) V es el 90% exacto del valor objetado (tarifa pactada COOSALUD);
       c) otra fila del mismo código ya vale exactamente V (valor unitario,
          p. ej. días de estancia 129A02H).
  3. "El TOTAL de conceptos objetados al servicio X (T) es mayor que el valor
     del servicio (V)": se capan las filas más grandes del código X hasta que
     la suma quede en V.
  4. Facturas con errores que no se pueden corregir automáticamente se dejan
     FUERA del reintento y se listan en pantalla para revisión manual.

OJO: DGH NO guarda ningún registro cuando el cargue trae errores — se vuelve a
cargar el Excel COMPLETO. Por eso el reintento lleva TODAS las facturas del
archivo original (con los valores ya corregidos), menos las que ya estaban
objetadas y las de revisión manual, con el consecutivo renumerado desde 1.

USO (sin argumentos pregunta todo, acepta rutas arrastradas):
    py corregir_errores_dgh.py
    py corregir_errores_dgh.py --errores "ERRORES.xlsx" --objeciones "OBJECIONES LOTE 01.xlsx"

Requiere tener consolidar_coosalud.py en la misma carpeta (comparten formato).
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

AQUI = Path(__file__).resolve().parent
sys.path.insert(0, str(AQUI))

try:
    import openpyxl
except ImportError:  # pragma: no cover
    print("Falta openpyxl. Instalar con:  py -m pip install openpyxl")
    sys.exit(1)

try:
    from consolidar_coosalud import COLS_OBJECIONES, a_numero, escribir_hoja, norm_texto
except ImportError:
    print("ERROR: falta consolidar_coosalud.py en la misma carpeta que este script.")
    sys.exit(1)

RE_YA = re.compile(
    r"La FACTURA se encuentra \(([^)]+)\) y por lo tanto no puede ser OBJETADA", re.I
)
RE_VALOR = re.compile(
    r"El VALOR OBJECION no puede ser mayor al valor del servicio\s*\(([\d.,]+)\)", re.I
)
RE_TOTAL = re.compile(
    r"El TOTAL de conceptos objetados al servicio\s+(\S+)\s*\(([\d.,]+)\)\s*es mayor que el valor del servicio\s*\(([\d.,]+)\)",
    re.I,
)


def limpiar_ruta(texto: str) -> str:
    return texto.strip().strip('"').strip("'").strip()


def pedir(mensaje: str) -> str:
    try:
        return input(mensaje)
    except EOFError:
        return ""


def leer_hoja(path: Path) -> tuple[list[str], list[list]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws.max_row == 1 and ws.max_column == 1:
        ws.reset_dimensions()
    filas = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not filas:
        raise ValueError(f"'{path.name}' está vacío.")
    return [norm_texto(c) for c in filas[0]], filas[1:]


def leer_errores(path: Path) -> dict[str, dict]:
    """Agrupa los errores de DGH por consecutivo (columna Documento).

    Devuelve {consec: {"ya": estado|None, "valores": Counter({V: n}),
                       "totales": [(codigo, V)], "otros": [mensajes]}}.
    """
    headers, rows = leer_hoja(path)
    mapa = {h.upper(): i for i, h in enumerate(headers)}
    idx_doc = mapa.get("DOCUMENTO")
    idx_msg = mapa.get("MENSAJEVALIDACION")
    if idx_doc is None or idx_msg is None:
        raise ValueError(
            f"'{path.name}' no parece el archivo de errores de DGH "
            "(faltan las columnas Documento / MensajeValidacion)."
        )
    errores: dict[str, dict] = {}
    for r in rows:
        doc = norm_texto(r[idx_doc])
        msg = norm_texto(r[idx_msg]).replace("_x000d_", " ")
        if not doc or not msg:
            continue
        e = errores.setdefault(doc, {"ya": None, "valores": Counter(), "totales": [], "otros": []})
        m = RE_YA.search(msg)
        if m:
            e["ya"] = m.group(1)
            continue
        m = RE_TOTAL.search(msg)
        if m:
            e["totales"].append((m.group(1).upper(), a_numero(m.group(3)) or 0.0))
            continue
        m = RE_VALOR.search(msg)
        if m:
            v = a_numero(m.group(1))
            if v is not None:
                e["valores"][round(v)] += 1
            continue
        e["otros"].append(msg)
    return errores


def capar_valores(filas: list[list], valores: Counter, idx_sln: int, idx_val: int) -> int | None:
    """Aplica los errores 'VALOR OBJECION > valor del servicio (V)' a las filas
    de UNA factura. Devuelve cuántas filas capó, o None si algún V quedó sin
    identificar (la factura se debe revisar a mano)."""
    capadas = 0
    asignadas: set[int] = set()
    # De mayor a menor V para que los grandes no se roben candidatas chicas.
    for v, cuantas in sorted(valores.items(), reverse=True):
        cand = [
            i
            for i, f in enumerate(filas)
            if i not in asignadas and (a_numero(f[idx_val]) or 0) > v + 0.5
        ]
        elegidas: list[int] = []
        if len(cand) == cuantas:
            elegidas = cand  # (a) las únicas posibles
        else:
            # (b) V = 90% del objetado (tarifa pactada)
            al_90 = [i for i in cand if abs((a_numero(filas[i][idx_val]) or 0) * 0.9 - v) <= 1]
            if len(al_90) == cuantas:
                elegidas = al_90
            else:
                # (c) otra fila del mismo código vale exactamente V (valor unitario)
                unit = {
                    norm_texto(f[idx_sln]).upper()
                    for f in filas
                    if abs((a_numero(f[idx_val]) or 0) - v) <= 0.5
                }
                por_unidad = [i for i in cand if norm_texto(filas[i][idx_sln]).upper() in unit]
                if len(por_unidad) == cuantas:
                    elegidas = por_unidad
        if len(elegidas) != cuantas:
            return None
        for i in elegidas:
            filas[i][idx_val] = int(v)
            asignadas.add(i)
            capadas += 1
    return capadas


def capar_totales(filas: list[list], totales: list, idx_sln: int, idx_val: int) -> int:
    """'El TOTAL objetado al servicio X es mayor que (V)': recorta las filas más
    grandes del código X hasta que la suma quede en V."""
    capadas = 0
    for codigo, v in totales:
        idxs = [i for i, f in enumerate(filas) if norm_texto(f[idx_sln]).upper() == codigo]
        suma = sum(a_numero(filas[i][idx_val]) or 0 for i in idxs)
        for i in sorted(idxs, key=lambda i: -(a_numero(filas[i][idx_val]) or 0)):
            if suma <= v + 0.5:
                break
            val = a_numero(filas[i][idx_val]) or 0
            recorte = min(val, suma - v)
            filas[i][idx_val] = int(round(val - recorte))
            suma -= recorte
            capadas += 1
    return capadas


def corregir(path_err: Path, path_obj: Path, path_salida: Path | None = None) -> int:
    errores = leer_errores(path_err)
    headers, rows = leer_hoja(path_obj)
    cols = [c for c, _ in COLS_OBJECIONES]
    mapa = {h.upper(): i for i, h in enumerate(headers)}
    if "CDCONSEC" not in mapa or "CROVALOBJ" not in mapa:
        raise ValueError(f"'{path_obj.name}' no parece un archivo de OBJECIONES (falta CDCONSEC).")
    idx_con = mapa["CDCONSEC"]
    idx_sln = mapa["SLNSERPRO"]
    idx_val = mapa["CROVALOBJ"]

    por_consec: dict[str, list[list]] = defaultdict(list)
    orden: list[str] = []
    for r in rows:
        c = norm_texto(r[idx_con])
        if c not in por_consec:
            orden.append(c)
        por_consec[c].append(list(r))

    ya_estaban: list[str] = []
    corregidas: list[str] = []
    manuales: list[tuple[str, str]] = []
    total_capadas = 0

    for consec, e in errores.items():
        filas = por_consec.get(consec)
        factura = norm_texto(filas[0][mapa["CRNCXC"]]) if filas else f"consec {consec}"
        if filas is None:
            manuales.append((factura, "el consecutivo del error no está en el OBJECIONES"))
            continue
        if e["ya"]:
            ya_estaban.append(factura)
            continue
        if e["otros"]:
            manuales.append((factura, e["otros"][0][:120]))
            continue
        n = capar_valores(filas, e["valores"], idx_sln, idx_val) if e["valores"] else 0
        if n is None:
            manuales.append((factura, "no pude identificar la fila del error de valor"))
            continue
        n += capar_totales(filas, e["totales"], idx_sln, idx_val)
        if n:
            total_capadas += n
            corregidas.append(factura)

    # DGH no guarda NADA cuando el cargue trae errores: el reintento lleva el
    # archivo COMPLETO (todas las facturas, ya corregidas), menos las que ya
    # estaban objetadas y las de revisión manual. Consecutivo renumerado.
    excluir = {f for f in ya_estaban} | {f for f, _ in manuales}
    consec_salida = [
        c for c in orden if norm_texto(por_consec[c][0][mapa["CRNCXC"]]) not in excluir
    ]
    filas_salida: list[list] = []
    for nuevo, consec in enumerate(consec_salida, 1):
        for f in por_consec[consec]:
            f2 = [f[mapa[c.upper()]] if c.upper() in mapa else None for c in cols]
            f2[0] = str(nuevo)
            filas_salida.append(f2)

    n_err = sum(
        sum(v["valores"].values()) + len(v["totales"]) + len(v["otros"]) + (1 if v["ya"] else 0)
        for v in errores.values()
    )
    print(f"\n  Errores de DGH leídos     : {n_err} en {len(errores)} facturas")
    print(
        f"  Ya estaban objetadas      : {len(ya_estaban)} facturas -> se QUITAN (normal si ya se subieron en un cargue anterior)"
    )
    print(
        f"  Valores corregidos        : {total_capadas} en {len(corregidas)} facturas, capados a lo que DGH acepta"
    )
    print(
        f"  Facturas sin problema     : {len(consec_salida) - len(corregidas)} -> van igual (DGH no guardó nada, se recarga todo)"
    )
    if manuales:
        print(f"  Para revisar A MANO       : {len(manuales)} facturas (no van en el reintento):")
        for fact, motivo in manuales:
            print(f"      {fact}: {motivo}")

    if not filas_salida:
        print("\n  No quedó ninguna factura para reintentar (todas ya estaban objetadas).")
        return 0

    salida = path_salida or path_obj.with_name(path_obj.stem + " REINTENTO.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "OBJECIONES"
    escribir_hoja(wb.active, cols, filas_salida, formatos=[f for _, f in COLS_OBJECIONES])
    wb.save(salida)
    print(f"\n  LISTO: {salida}")
    print(
        f"  {len(filas_salida)} filas · {len(consec_salida)} facturas · consecutivos 1..{len(consec_salida)}"
    )
    print("  Sube ese archivo a DGH igual que el original.")
    return 0


def buscar_candidatos(patron: str) -> list[Path]:
    lugares = [
        AQUI,
        AQUI.parent,
        Path.home() / "Downloads",
        Path.home() / "Desktop",
        Path(r"D:\USUARIO CARTERA\Desktop"),
    ]
    encontrados: dict[Path, float] = {}
    for lugar in lugares:
        if lugar.is_dir():
            for f in lugar.glob(patron):
                if not f.name.startswith("~$"):
                    encontrados[f] = f.stat().st_mtime
    return sorted(encontrados, key=encontrados.get, reverse=True)[:5]


def preguntar_archivo(titulo: str, patron: str) -> Path | None:
    print(f"\n--- {titulo} ---")
    cands = buscar_candidatos(patron)
    if cands:
        for i, c in enumerate(cands, 1):
            print(f"    {i}) {c}")
        r = limpiar_ruta(pedir("  Escribe el numero, o arrastra aqui el archivo y Enter: "))
    else:
        r = limpiar_ruta(pedir("  Arrastra aqui el archivo y presiona Enter: "))
    if r.isdigit() and cands and 1 <= int(r) <= len(cands):
        return cands[int(r) - 1]
    if r:
        p = Path(r)
        if p.is_file():
            return p
        print(f"  ERROR: no existe: {p}")
        return None
    return cands[0] if cands else None


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(
        description="Corrige un OBJECIONES con los errores que devolvió DGH."
    )
    p.add_argument(
        "--errores", default=None, help="Excel de errores de DGH (VALIDACIONIMPORTACION)."
    )
    p.add_argument("--objeciones", default=None, help="El OBJECIONES.xlsx que se subió a DGH.")
    p.add_argument(
        "--salida", default=None, help="Ruta del reintento. Def: <objeciones> REINTENTO.xlsx"
    )
    args = p.parse_args(argv)

    print("=" * 72)
    print("   CORREGIR ERRORES DGH - arma el OBJECIONES de reintento")
    print("=" * 72)

    path_err = (
        Path(args.errores)
        if args.errores
        else preguntar_archivo("PASO A: el Excel de ERRORES que descargaste de DGH", "*ERROR*.xlsx")
    )
    if not path_err or not path_err.is_file():
        print("  Sin el archivo de errores no puedo continuar.")
        return 1
    path_obj = (
        Path(args.objeciones)
        if args.objeciones
        else preguntar_archivo(
            "PASO B: el OBJECIONES que subiste a DGH (el del mismo lote)", "OBJECIONES*.xlsx"
        )
    )
    if not path_obj or not path_obj.is_file():
        print("  Sin el OBJECIONES no puedo continuar.")
        return 1

    try:
        return corregir(path_err, path_obj, Path(args.salida) if args.salida else None)
    except ValueError as exc:
        print(f"\n  ERROR: {exc}")
        return 1


if __name__ == "__main__":
    codigo = main()
    pedir("\n  Presiona Enter para salir...")
    raise SystemExit(codigo)
