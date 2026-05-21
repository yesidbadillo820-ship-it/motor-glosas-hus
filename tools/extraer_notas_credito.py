"""extraer_notas_credito.py — Extrae documentos de Notas Crédito desde un share de red.

Lee un Excel/CSV/TSV con una columna `NOTA CREDITO` y, para cada nota, busca la
carpeta correspondiente en el share de facturación electrónica:

    <base>\\<PERIODO>\\<subcarpeta>\\<nota>

Por ejemplo:

    \\\\172.16.32.83\\factura_electronica_net22\\202605\\FACTURAS_NOTA\\309306

Copia todos los archivos contenidos a:

    <salida>\\<nota>\\...

Al terminar, escribe un reporte CSV con el resultado de cada nota (encontrada,
no encontrada, archivos copiados, etc.).

USO RÁPIDO
----------

    # Desde un Excel
    python extraer_notas_credito.py --excel notas.xlsx

    # Desde el TSV de ejemplo incluido en este repo
    python extraer_notas_credito.py --tsv notas_credito_ejemplo.tsv

    # Personalizar base, subcarpeta y destino
    python extraer_notas_credito.py --excel notas.xlsx \\
        --base "\\\\172.16.32.83\\factura_electronica_net22" \\
        --subcarpeta FACTURAS_NOTA \\
        --salida ./documentos \\
        --reporte ./reporte_extraccion_notas.csv

    # Dry run (no copia, solo reporta qué encontraría)
    python extraer_notas_credito.py --excel notas.xlsx --dry-run

DEPENDENCIAS
------------

- Python 3.9+
- `openpyxl` solo si se usa --excel (ya viene en requirements.txt del proyecto;
  si no, instalar con `pip install openpyxl`). Para --tsv/--csv no se necesita
  nada extra: stdlib pura.

NOTAS
-----

- En Windows, las rutas UNC `\\\\172.16.32.83\\...` funcionan tal cual.
- En Linux/Mac, el share debe estar montado primero (ej. /mnt/factura_electronica)
  y pasarlo como `--base /mnt/factura_electronica`.
- Si un archivo destino ya existe con el mismo tamaño, se omite (reintento seguro).
- Si la misma nota aparece en varios periodos, se prefiere el más reciente
  (orden descendente del nombre de carpeta).
"""

from __future__ import annotations

import argparse
import csv
import logging
import re
import shutil
import sys
from pathlib import Path
from typing import Iterable

DEFAULT_BASE = r"\\172.16.32.83\factura_electronica_net22"
DEFAULT_SUBCARPETA = "FACTURAS_NOTA"
DEFAULT_SALIDA = "documentos"
DEFAULT_REPORTE = "reporte_extraccion_notas.csv"

logger = logging.getLogger("extraer_notas")


def setup_logging(log_file: Path | None = None) -> None:
    fmt = "%(asctime)s [%(levelname)s] %(message)s"
    handlers: list[logging.Handler] = [logging.StreamHandler(sys.stdout)]
    if log_file is not None:
        log_file.parent.mkdir(parents=True, exist_ok=True)
        handlers.append(logging.FileHandler(log_file, encoding="utf-8"))
    logging.basicConfig(level=logging.INFO, format=fmt, handlers=handlers)


# ─── Lectura de la fuente ──────────────────────────────────────────────────

COLUMNAS_REQUERIDAS = ("NOTA CREDITO",)
COLUMNAS_OPCIONALES = ("Radicado", "Acta Conciliacion", "Prefijo Factura",
                       "# Factura", "CONCATENAR")


def _normalizar_headers(headers: Iterable[str]) -> list[str]:
    return [(h or "").strip() for h in headers]


def _filas_desde_iter(headers: list[str], rows_iter: Iterable) -> list[dict]:
    headers_norm = {h.upper(): i for i, h in enumerate(headers)}
    for req in COLUMNAS_REQUERIDAS:
        if req.upper() not in headers_norm:
            raise ValueError(
                f"No encontré la columna requerida '{req}'. "
                f"Headers detectados: {headers}"
            )

    idx_nota = headers_norm["NOTA CREDITO"]
    idx_opc = {col: headers_norm.get(col.upper()) for col in COLUMNAS_OPCIONALES}

    filas: list[dict] = []
    for row in rows_iter:
        if row is None:
            continue
        row = list(row)
        if idx_nota >= len(row) or row[idx_nota] in (None, ""):
            continue
        nota = str(row[idx_nota]).strip()
        if not nota:
            continue
        registro = {"nota": nota}
        for col, idx in idx_opc.items():
            if idx is None or idx >= len(row) or row[idx] in (None, ""):
                registro[col] = ""
            else:
                registro[col] = str(row[idx]).strip()
        filas.append(registro)
    return filas


def leer_notas_excel(ruta: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.stderr.write(
            "ERROR: para leer .xlsx necesitás openpyxl.\n"
            "Instalalo con: pip install openpyxl\n"
            "O exportá el Excel a TSV/CSV y usá --tsv/--csv.\n"
        )
        sys.exit(2)

    wb = load_workbook(filename=str(ruta), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        headers_raw = next(rows)
    except StopIteration:
        raise ValueError(f"El Excel {ruta} está vacío.")
    headers = _normalizar_headers(headers_raw)
    return _filas_desde_iter(headers, rows)


def leer_notas_delimitado(ruta: Path, sep: str) -> list[dict]:
    with ruta.open(encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh, delimiter=sep)
        try:
            headers_raw = next(reader)
        except StopIteration:
            raise ValueError(f"El archivo {ruta} está vacío.")
        headers = _normalizar_headers(headers_raw)
        return _filas_desde_iter(headers, reader)


# ─── Indexación del share ──────────────────────────────────────────────────

def indexar_notas_disponibles(base: Path, subcarpeta: str) -> dict[str, Path]:
    """Indexa carpetas de notas crédito bajo base/<PERIODO>/<subcarpeta>/<nota>.

    Si la misma nota aparece en varios periodos, conserva el más reciente
    (asumiendo orden lexicográfico descendente sobre nombres tipo YYYYMM).
    """
    indice: dict[str, Path] = {}
    if not base.exists():
        logger.error(f"La base no es accesible: {base}")
        return indice

    try:
        candidatos = [d for d in base.iterdir() if d.is_dir()]
    except OSError as e:
        logger.error(f"No puedo listar la base {base}: {e}")
        return indice

    periodos = sorted(
        [d for d in candidatos if (d / subcarpeta).is_dir()],
        key=lambda p: p.name,
        reverse=True,
    )

    if not periodos:
        logger.warning(
            f"No encontré ninguna subcarpeta '{subcarpeta}' bajo {base}. "
            f"Revisá la ruta o usá --subcarpeta para ajustar."
        )
        return indice

    muestra = ", ".join(p.name for p in periodos[:6])
    sufijo = "..." if len(periodos) > 6 else ""
    logger.info(
        f"Indexando {len(periodos)} periodos con '{subcarpeta}' adentro: {muestra}{sufijo}"
    )

    for periodo in periodos:
        carpeta_facturas = periodo / subcarpeta
        try:
            subs = list(carpeta_facturas.iterdir())
        except OSError as e:
            logger.warning(f"No pude listar {carpeta_facturas}: {e}")
            continue
        for sub in subs:
            if sub.is_dir() and sub.name not in indice:
                indice[sub.name] = sub

    logger.info(f"Indexadas {len(indice)} carpetas de notas crédito en total")
    return indice


# ─── Copia de archivos ─────────────────────────────────────────────────────

def copiar_archivos(origen: Path, destino: Path) -> tuple[int, int, int]:
    """Copia archivos de origen a destino (recursivo).

    Devuelve (copiados, omitidos_iguales, errores).
    """
    destino.mkdir(parents=True, exist_ok=True)
    copiados = omitidos = errores = 0
    try:
        items = list(origen.iterdir())
    except OSError as e:
        logger.error(f"No pude listar {origen}: {e}")
        return 0, 0, 1

    for item in items:
        if item.is_file():
            destino_archivo = destino / item.name
            try:
                if (destino_archivo.exists()
                        and destino_archivo.stat().st_size == item.stat().st_size):
                    omitidos += 1
                    continue
                shutil.copy2(item, destino_archivo)
                copiados += 1
            except OSError as e:
                logger.error(f"  ! {item.name}: {e}")
                errores += 1
        elif item.is_dir():
            c, o, e_ = copiar_archivos(item, destino / item.name)
            copiados += c
            omitidos += o
            errores += e_
    return copiados, omitidos, errores


# ─── Orquestación ──────────────────────────────────────────────────────────

def procesar(
    notas: list[dict],
    indice: dict[str, Path],
    base: Path,
    salida: Path,
    fallback_recursivo: bool,
) -> list[dict]:
    notas_unicas: dict[str, dict] = {}
    for n in notas:
        notas_unicas.setdefault(n["nota"], n)

    total = len(notas_unicas)
    logger.info(f"Procesando {total} notas únicas (de {len(notas)} filas)")

    resultados: list[dict] = []
    for i, (nota, info) in enumerate(notas_unicas.items(), start=1):
        prefix = f"[{i}/{total}] nota {nota}"
        origen = indice.get(nota)

        if origen is None and fallback_recursivo and base.exists():
            try:
                for ruta in base.rglob(nota):
                    if ruta.is_dir():
                        origen = ruta
                        logger.info(f"{prefix}: encontrada via rglob en {ruta}")
                        break
            except OSError as e:
                logger.warning(f"{prefix}: error en búsqueda recursiva — {e}")

        if origen is None:
            logger.warning(f"{prefix}: NO ENCONTRADA")
            resultados.append({
                **info,
                "estado": "NO_ENCONTRADA",
                "ruta_origen": "",
                "archivos_copiados": 0,
                "archivos_omitidos": 0,
                "errores": 0,
            })
            continue

        destino = salida / nota
        copiados, omitidos, errores = copiar_archivos(origen, destino)

        if copiados == 0 and omitidos == 0 and errores == 0:
            estado = "VACIA"
        elif errores > 0 and copiados == 0 and omitidos == 0:
            estado = "ERROR"
        elif errores > 0:
            estado = "PARCIAL"
        else:
            estado = "OK"

        logger.info(
            f"{prefix}: {estado} — {copiados} copiados, "
            f"{omitidos} ya existían, {errores} errores"
        )
        resultados.append({
            **info,
            "estado": estado,
            "ruta_origen": str(origen),
            "archivos_copiados": copiados,
            "archivos_omitidos": omitidos,
            "errores": errores,
        })
    return resultados


def escribir_reporte(resultados: list[dict], ruta_reporte: Path) -> None:
    ruta_reporte.parent.mkdir(parents=True, exist_ok=True)
    campos = [
        "nota", "Radicado", "Acta Conciliacion", "Prefijo Factura",
        "# Factura", "CONCATENAR",
        "estado", "ruta_origen",
        "archivos_copiados", "archivos_omitidos", "errores",
    ]
    with ruta_reporte.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=campos)
        writer.writeheader()
        for r in resultados:
            writer.writerow({k: r.get(k, "") for k in campos})
    logger.info(f"Reporte guardado en: {ruta_reporte}")


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    grupo_input = parser.add_mutually_exclusive_group(required=True)
    grupo_input.add_argument("--excel", type=Path, help="Ruta a un archivo .xlsx")
    grupo_input.add_argument("--csv", type=Path, help="Ruta a un archivo .csv (separador coma)")
    grupo_input.add_argument("--tsv", type=Path, help="Ruta a un archivo .tsv (separador tab)")

    parser.add_argument(
        "--base", default=DEFAULT_BASE,
        help=f"Ruta base del share (default: {DEFAULT_BASE})",
    )
    parser.add_argument(
        "--subcarpeta", default=DEFAULT_SUBCARPETA,
        help=f"Subcarpeta dentro de cada periodo (default: {DEFAULT_SUBCARPETA})",
    )
    parser.add_argument(
        "--salida", type=Path, default=Path(DEFAULT_SALIDA),
        help=f"Carpeta destino local (default: {DEFAULT_SALIDA})",
    )
    parser.add_argument(
        "--reporte", type=Path, default=Path(DEFAULT_REPORTE),
        help=f"Archivo CSV de reporte (default: {DEFAULT_REPORTE})",
    )
    parser.add_argument("--log", type=Path, default=None, help="Archivo de log opcional")
    parser.add_argument(
        "--no-fallback-recursivo", action="store_true",
        help="No usar rglob como fallback cuando la nota no está en el índice",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="No copia nada; solo reporta qué encontraría",
    )

    args = parser.parse_args()
    setup_logging(args.log)

    if args.excel:
        notas = leer_notas_excel(args.excel)
    elif args.csv:
        notas = leer_notas_delimitado(args.csv, sep=",")
    else:
        notas = leer_notas_delimitado(args.tsv, sep="\t")

    if not notas:
        logger.error("No hay notas para procesar (¿columna 'NOTA CREDITO' vacía?)")
        return 1
    logger.info(f"Leídas {len(notas)} filas con notas crédito")

    base = Path(args.base)
    indice = indexar_notas_disponibles(base, args.subcarpeta)

    if args.dry_run:
        notas_unicas = {n["nota"] for n in notas}
        encontradas = sorted(n for n in notas_unicas if n in indice)
        faltantes = sorted(notas_unicas - set(indice))
        logger.info(
            f"DRY RUN: {len(encontradas)}/{len(notas_unicas)} notas localizadas "
            f"en el índice. No se copia nada."
        )
        if faltantes:
            logger.info(f"Faltantes ({len(faltantes)}): {', '.join(faltantes[:20])}"
                        f"{'...' if len(faltantes) > 20 else ''}")
        return 0

    resultados = procesar(
        notas, indice, base, args.salida,
        fallback_recursivo=not args.no_fallback_recursivo,
    )
    escribir_reporte(resultados, args.reporte)

    resumen = {"OK": 0, "PARCIAL": 0, "VACIA": 0, "NO_ENCONTRADA": 0, "ERROR": 0}
    total_copiados = total_omitidos = total_errores = 0
    for r in resultados:
        resumen[r["estado"]] = resumen.get(r["estado"], 0) + 1
        total_copiados += r["archivos_copiados"]
        total_omitidos += r["archivos_omitidos"]
        total_errores += r["errores"]

    logger.info("=" * 60)
    logger.info("RESUMEN FINAL")
    logger.info(f"  Notas únicas procesadas: {len(resultados)}")
    for estado, n in resumen.items():
        if n > 0:
            logger.info(f"    {estado}: {n}")
    logger.info(f"  Archivos copiados: {total_copiados}")
    logger.info(f"  Archivos omitidos (ya existían): {total_omitidos}")
    logger.info(f"  Errores: {total_errores}")
    logger.info(f"  Reporte: {args.reporte}")
    logger.info(f"  Carpeta destino: {args.salida}")

    return 0 if total_errores == 0 and resumen["NO_ENCONTRADA"] == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
