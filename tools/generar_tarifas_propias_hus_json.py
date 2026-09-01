"""Convierte el Excel de TARIFAS PROPIAS del HUS a JSON comprimido.

Genera ``app/data/tarifas_propias_hus.json.gz`` con el catálogo institucional
de la ESE Hospital Universitario de Santander: los procedimientos, paquetes,
exámenes ambulatorios y órtesis/prótesis con su valor propio en pesos 2026.

Por qué importa (18-08-2026): el liquidador solo conocía 84 tarifas propias
transcritas a mano. Pero estas tarifas NO son de FAMISANAR: son del HUS, y
casi todos los contratos las usan —COOSALUD, COMPENSAR, SALUD MÍA, AURORA,
PPL, FAMISANAR— cuando pactan «SOAT −X% + TARIFAS PROPIAS/INSTITUCIONALES».
Sin el catálogo completo, el liquidador no podía dar el valor propio de una
colecistectomía ($6.296.900) para ninguna EPS.

Uso:
    python tools/generar_tarifas_propias_hus_json.py /ruta/TARIFAS_HUS.xlsx

Lee cuatro hojas de tarifas PROPIAS (no la hoja UVB, que son valores SOAT por
grupo que el motor ya calcula de la Circular 047/2025):
    • "TARIFAS PROPIAS"    — procedimientos con código institucional
    • "PAQUETES"           — paquetes quirúrgicos por especialidad
    • "AMBULATORIO"        — exámenes ambulatorios (Res. 2706/2025)
    • "ORTESIS Y PROTESIS" — dispositivos

Salida JSON:
    {
      "_meta": {...},
      "tarifas": {
        "512101": {"valor": 6296900, "descripcion": "COLECISTECTOMIA VIA ABIERTA",
                   "codigo_ips": "512101H", "tipo": "PROPIA"},
        ...
      }
    }
"""

from __future__ import annotations

import gzip
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _dinero import a_numero  # noqa: E402 — el único lector de pesos del repo

_AQUI = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_OUT = os.path.join(_AQUI, "app", "data", "tarifas_propias_hus.json.gz")

# (nombre de hoja, tipo). El orden fija la prioridad: si un CUPS aparece en dos
# hojas, gana la primera (la tarifa propia manda sobre el paquete, etc.).
_HOJAS = (
    ("TARIFAS PROPIAS", "PROPIA"),
    ("PAQUETES", "PAQUETE"),
    ("AMBULATORIO", "AMBULATORIO"),
    ("ORTESIS Y PROTESIS", "ORTESIS_PROTESIS"),
)


def _norm(s) -> str:
    return str(s).strip() if s is not None else ""


def _a_pesos(v) -> int | None:
    """Un valor de celda a pesos enteros usando el lector único del repo.
    Devuelve None si no es un valor válido (>0) — no se inventa una tarifa
    para rellenar un código sin precio."""
    n = a_numero(v)
    return int(round(n)) if n and n > 0 else None


def _hoja_por_nombre(wb, nombre: str):
    """Encuentra la hoja aunque el título tenga espacios de sobra."""
    objetivo = nombre.strip().upper()
    for ws in wb.worksheets:
        if str(ws.title).strip().upper() == objetivo:
            return ws
    return None


def _indices(encabezado: list[str]) -> tuple[int, int, int, int]:
    """Ubica las columnas por su nombre: CUPS, descripción, código IPS, valor."""
    up = [_norm(h).upper() for h in encabezado]
    i_cups = i_desc = i_ips = i_val = -1
    for i, h in enumerate(up):
        if i_cups < 0 and ("CUPS" in h or h.startswith("RES ") or "2341" in h or "2706" in h):
            i_cups = i
        elif i_cups >= 0 and i_desc < 0 and "DESCRIPC" in h:
            i_desc = i
        elif i_ips < 0 and ("CODIGO IPS" in h or h == "CÓDIGO" or h == "CODIGO"):
            i_ips = i
        elif i_val < 0 and ("TARIFA" in h or "VALOR" in h):
            i_val = i
    if i_cups < 0:
        i_cups = 0
    if i_val < 0:
        i_val = 4
    return i_cups, i_desc, i_ips, i_val


def generar(src_xlsx: str, out_path: str = _OUT, version: str = "") -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(src_xlsx, read_only=True, data_only=True)
    tarifas: dict[str, dict] = {}
    por_tipo: dict[str, int] = {}
    descartadas = 0

    for nombre, tipo in _HOJAS:
        ws = _hoja_por_nombre(wb, nombre)
        if ws is None:
            continue
        filas = list(ws.iter_rows(values_only=True))
        if not filas:
            continue
        i_cups, i_desc, i_ips, i_val = _indices(list(filas[0]))
        for fila in filas[1:]:
            if not fila:
                continue
            cups = _norm(fila[i_cups]).upper() if i_cups < len(fila) else ""
            if not cups or cups.upper() in ("CUPS", "NONE"):
                continue
            valor = _a_pesos(fila[i_val]) if i_val < len(fila) else None
            if valor is None:
                descartadas += 1
                continue
            if cups in tarifas:  # ya la trae una hoja de mayor prioridad
                continue
            tarifas[cups] = {
                "valor": valor,
                "descripcion": _norm(fila[i_desc]) if 0 <= i_desc < len(fila) else "",
                "codigo_ips": _norm(fila[i_ips]) if 0 <= i_ips < len(fila) else "",
                "tipo": tipo,
            }
            por_tipo[tipo] = por_tipo.get(tipo, 0) + 1

    data = {
        "_meta": {
            "fuente": (
                "Catálogo de TARIFAS PROPIAS de la ESE Hospital Universitario de "
                "Santander (Res. 054/2026 + 124/2026 y anexos), vigencia 2026."
            ),
            "descripcion": (
                "Valor institucional en pesos por código CUPS. Aplica a los "
                "contratos que pactan 'tarifas propias/institucionales del HUS' "
                "(COOSALUD, COMPENSAR, SALUD MÍA, AURORA, PPL, FAMISANAR, etc.)."
            ),
            "norma": "TARIFAS_PROPIAS_HUS_2026",
            "codigos": len(tarifas),
            "por_tipo": por_tipo,
            "descartadas_sin_valor": descartadas,
            "version": version or "2026-TARIFAS-HUS",
        },
        "tarifas": dict(sorted(tarifas.items())),
    }

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with gzip.open(out_path, "wt", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
    return data["_meta"]


# ─────────────────────────────────────────────────────────────────────────
#  RESOLUCIÓN 283 DE 2026 — 20-08-2026
#
#  Yesid mandó "TARIFAS_INSTITUCIONALES_RES_283.xlsx", una resolución nueva
#  del HUS con una sola hoja ("ENCABEZADO NO UVB") de 695 tarifas. Comparada
#  con lo que ya estaba cargado (Res. 054/2026 + 124/2026, 1.932 códigos):
#
#      · 673 códigos NUEVOS que el motor no tenía — sin ellos, el dictamen no
#        podía dar el valor propio de esos procedimientos;
#      · 22 códigos que ya existían y CAMBIARON de valor (p. ej. HIERRO TOTAL
#        $50.000 → $66.500; TROPONINA T $90.900 → $109.100). El hospital
#        estaba defendiendo tarifas viejas;
#      · ninguno igual.
#
#  Decisión del auditor (20-08-2026): la 283 SE SUMA a las anteriores, no las
#  reemplaza. Donde pisan, manda la 283 por ser más reciente. Así no se da de
#  baja ninguna de las 1.910 tarifas que hoy se usan para defender —entre
#  ellas casi todo el laboratorio—.
#
#  Esta función NO regenera el catálogo desde cero: parte del .json.gz que ya
#  existe y le suma. El Excel original de las resoluciones anteriores vive en
#  el PC de cartera, no en el repositorio.
# ─────────────────────────────────────────────────────────────────────────

_HOJA_RES283 = "ENCABEZADO NO UVB"


def _filas_res283(src_xlsx: str) -> list[tuple[str, str, str, int]]:
    """(CUPS, descripción, código IPS, valor) de la hoja de la Res. 283."""
    import openpyxl

    wb = openpyxl.load_workbook(src_xlsx, read_only=True, data_only=True)
    ws = _hoja_por_nombre(wb, _HOJA_RES283)
    if ws is None:
        raise SystemExit(f"El archivo no trae la hoja «{_HOJA_RES283}».")
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return []
    # Encabezado: GRUPO | CUPS | DESCRIPCION CUPS | CODIGO IPS | DESCRIPCION IPS
    #             | FACTOR SMDLV | TARIFA 2026
    i_cups, i_desc, i_ips, i_val = _indices(list(filas[0]))
    salida: list[tuple[str, str, str, int]] = []
    for fila in filas[1:]:
        if not fila:
            continue
        cups = _norm(fila[i_cups]).upper() if i_cups < len(fila) else ""
        if not cups or cups in ("CUPS", "NONE"):
            continue
        valor = _a_pesos(fila[i_val]) if i_val < len(fila) else None
        if valor is None:
            continue
        salida.append(
            (
                cups,
                _norm(fila[i_desc]) if 0 <= i_desc < len(fila) else "",
                _norm(fila[i_ips]) if 0 <= i_ips < len(fila) else "",
                valor,
            )
        )
    return salida


def agregar_res283(src_xlsx: str, out_path: str = _OUT) -> dict:
    """Suma la Res. 283/2026 al catálogo ya generado, sin borrar nada.

    Los CUPS que aparecen varias veces en la resolución CON VALORES DISTINTOS
    quedan FUERA del catálogo a propósito: son tarifas por niveles (p. ej.
    399802 HEMOFILTRACIÓN a $5.450.000 / $7.260.000 / $9.075.000) y el archivo
    no dice cuál aplica a cada caso. Elegir una sería inventarle al dictamen
    una cifra que puede no ser la del paciente. Se devuelven aparte para que
    el motor avise en vez de afirmar.
    """
    with gzip.open(out_path, "rt", encoding="utf-8") as f:
        data = json.load(f)
    tarifas: dict[str, dict] = data.get("tarifas", {})
    antes = len(tarifas)

    filas = _filas_res283(src_xlsx)

    # Un CUPS con más de un valor distinto en la misma resolución = por niveles.
    valores_por_cups: dict[str, set] = {}
    for cups, _d, _i, valor in filas:
        valores_por_cups.setdefault(cups, set()).add(valor)
    por_niveles = {c: sorted(v) for c, v in valores_por_cups.items() if len(v) > 1}

    nuevos = actualizados = 0
    cambios: list[dict] = []
    for cups, desc, ips, valor in filas:
        if cups in por_niveles:
            continue
        anterior = tarifas.get(cups)
        if anterior is None:
            nuevos += 1
        elif anterior.get("valor") != valor:
            actualizados += 1
            cambios.append(
                {
                    "cups": cups,
                    "descripcion": desc,
                    "antes": anterior.get("valor"),
                    "ahora": valor,
                }
            )
        else:
            continue
        tarifas[cups] = {
            "valor": valor,
            "descripcion": desc,
            "codigo_ips": ips,
            "tipo": "PROPIA",
            "norma": "RES_283_2026",
        }

    meta = data.setdefault("_meta", {})
    meta["fuente"] = (
        "Catálogo de TARIFAS PROPIAS de la ESE Hospital Universitario de "
        "Santander (Res. 054/2026 + 124/2026 + Res. 283/2026 y anexos), "
        "vigencia 2026."
    )
    meta["codigos"] = len(tarifas)
    meta["res_283_2026"] = {
        "filas_leidas": len(filas),
        "nuevos": nuevos,
        "actualizados": actualizados,
        "por_niveles_excluidos": por_niveles,
        "cambios_de_valor": cambios,
    }
    data["tarifas"] = dict(sorted(tarifas.items()))

    with gzip.open(out_path, "wt", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
    return {
        "antes": antes,
        "despues": len(tarifas),
        "nuevos": nuevos,
        "actualizados": actualizados,
        "por_niveles_excluidos": por_niveles,
        "cambios_de_valor": cambios,
    }


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python tools/generar_tarifas_propias_hus_json.py /ruta/TARIFAS_HUS.xlsx")
        print("     python tools/generar_tarifas_propias_hus_json.py --res283 /ruta/RES_283.xlsx")
        sys.exit(1)
    if sys.argv[1] == "--res283":
        if len(sys.argv) < 3:
            print("Falta la ruta del Excel de la Res. 283.")
            sys.exit(1)
        resumen = agregar_res283(sys.argv[2])
        print(f"Res. 283/2026 sumada al catálogo en {_OUT}")
        print(json.dumps(resumen, ensure_ascii=False, indent=2))
        sys.exit(0)
    meta = generar(sys.argv[1], version=sys.argv[2] if len(sys.argv) > 2 else "")
    print(f"JSON.gz generado en {_OUT}")
    print(json.dumps(meta, ensure_ascii=False, indent=2))
