"""responder_glosas_simed.py — Carga masiva de respuestas a glosas en SIMED.

Para cada factura del Excel (generado por `extraer_respuestas_glosa.py` y
revisado por vos), recorre las objeciones una por una en el portal SIMED:

1. Login (una sola vez).
2. Por cada factura:
   - Va a "Respuesta Glosa Rad. WEB" (glosasfacturaww.aspx).
   - Filtra por # Factura.
   - Click en el lápiz → abre "Glosas Factura".
   - PARA CADA OBJECIÓN (1..N) del Excel:
     a) Click en el ícono gris de la fila → abre modal "Respuesta Glosa Ips Web".
     b) Tab "Información Glosa": escribe Aceptado, deja NC vacío, escribe
        el Detalle Respuesta sanitizado (sin tildes/ñ).
     c) Tab "Soportes": sube el Trámite + todos los archivos de la carpeta
        de soportes de la factura (las dos fuentes que pasaste).
     d) Click "Confirmar" del modal.
   - Click "Confirmar" del form principal (debajo de la grilla).
   - Click botón verde de la grilla (Enviar/Finalizar) y espera el OK.
3. Reporte CSV con el estado de cada factura.

CREDENCIALES (en variables de entorno):
    setx SIMED_USER 900006037
    setx SIMED_PASSWORD <tu_password>

USO:
    REM Piloto con una factura, browser visible para verificar
    py responder_glosas_simed.py ^
        --excel          "D:\\GLOSAS_2026\\respuestas_glosa.xlsx" ^
        --soportes-glosa "D:\\USUARIO CARTERA\\Documents\\GLOSAS 2026\\DISPENSARIO MEDICO\\Nueva carpeta\\SOPORTES" ^
        --indice         "D:\\indice_facturas_HUS.txt" ^
        --solo HUS0000452150 ^
        --con-cabeza

    REM Lote completo
    py responder_glosas_simed.py ^
        --excel ... --soportes-glosa ... --indice ... ^
        --todas ^
        --reporte "D:\\GLOSAS_2026\\reporte_glosa.csv"

INSTALACIÓN (una vez):
    py -m pip install playwright openpyxl
    py -m playwright install chromium
"""

from __future__ import annotations

import argparse
import csv
import logging
import os
import re
import sys
import time
import unicodedata
from datetime import date
from pathlib import Path

try:
    from playwright.sync_api import (
        Page,
        TimeoutError as PlaywrightTimeout,
        sync_playwright,
    )
except ImportError:
    # Permitir importar helpers (sanitizar, leer_excel_respuestas, etc.) sin
    # playwright instalado. Las funciones de browser fallarán en runtime con
    # un mensaje claro si se ejecuta sin la dependencia.
    Page = object  # type: ignore[assignment,misc]
    PlaywrightTimeout = Exception  # type: ignore[assignment,misc]
    sync_playwright = None  # type: ignore[assignment]


def _exigir_playwright() -> None:
    if sync_playwright is None:
        sys.stderr.write(
            "ERROR: falta playwright.\n"
            "Instalalo con:\n"
            "    py -m pip install playwright\n"
            "    py -m playwright install chromium\n"
        )
        sys.exit(2)

PORTAL_LOGIN = "https://auditool25.tool.com.co/gamexamplelogin.aspx"
PORTAL_GLOSAS = "https://auditool25.tool.com.co/glosasfacturaww.aspx"

logger = logging.getLogger("responder_glosas")


# ─── Setup ──────────────────────────────────────────────────────────────────


def setup_logging(log_file: Path | None = None) -> None:
    fmt = "%(asctime)s [%(levelname)s] %(message)s"
    handlers: list[logging.Handler] = [logging.StreamHandler(sys.stdout)]
    if log_file is not None:
        log_file.parent.mkdir(parents=True, exist_ok=True)
        handlers.append(logging.FileHandler(log_file, encoding="utf-8"))
    logging.basicConfig(level=logging.INFO, format=fmt, handlers=handlers)


def cargar_credenciales() -> tuple[str, str]:
    user = os.environ.get("SIMED_USER", "").strip()
    password = os.environ.get("SIMED_PASSWORD", "").strip()
    if not user or not password:
        sys.stderr.write(
            "ERROR: faltan credenciales. Setealas con:\n"
            "    setx SIMED_USER 900006037\n"
            "    setx SIMED_PASSWORD <tu_password>\n"
            "Despues cerra y reabri PowerShell.\n"
        )
        sys.exit(2)
    return user, password


# ─── Sanitización de texto (sin tildes/ñ/caracteres especiales) ─────────────


def sanitizar(s: str) -> str:
    """Deja SÓLO letras ASCII, números y espacios.

    El portal SIMED valida 'El texto debe iniciar y contener solo letras y
    numeros. NO debe contener caracteres especiales'. Por eso:
      - las tildes se transliteran (Ú→U, Ñ→N) vía NFKD,
      - y CUALQUIER otro signo ( [ ] , ; . - / ' " ( ) : etc.) se reemplaza
        por un espacio (colapsando espacios múltiples).
    Resultado: sólo [A-Za-z0-9] y espacios simples, empezando por letra/número.
    """
    if not s:
        return ""
    base = unicodedata.normalize("NFKD", s)
    sin_tildes = "".join(c for c in base if not unicodedata.combining(c))
    solo_alfanum = re.sub(r"[^A-Za-z0-9]+", " ", sin_tildes)
    return solo_alfanum.strip()


# ─── Lectura del Excel de respuestas ────────────────────────────────────────


def normalizar_factura(s: str) -> str:
    """HUS0000452150 → 452150 (sin prefijo y sin ceros a la izquierda)."""
    s = (s or "").strip().upper()
    m = re.match(r"HUS0*(\d+)$", s)
    return m.group(1) if m else s.lstrip("0") or s


def leer_excel_respuestas(ruta: Path) -> list[dict]:
    """Devuelve una lista de dicts {factura, factura_corta, num, aceptado, detalle}.

    Acepta cabeceras con variantes (tolerante mayúsculas/acentos/espacios).
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.stderr.write("ERROR: falta openpyxl. Instalalo con: py -m pip install openpyxl\n")
        sys.exit(2)

    def norm(h: str) -> str:
        s = unicodedata.normalize("NFKD", (h or "").strip().upper())
        s = "".join(c for c in s if not unicodedata.combining(c))
        return re.sub(r"\s+", " ", s)

    alias = {
        "factura": {"FACTURA", "# FACTURA", "NUM FACTURA", "NUMERO FACTURA"},
        "num": {"# OBJECION", "NUM OBJECION", "OBJECION", "NUMERO OBJECION", "ITEM"},
        "aceptado": {"VALOR ACEPTADO", "ACEPTADO", "VI ACEPTADO"},
        "detalle": {"DETALLE RESPUESTA", "RESPUESTA", "OBSERVACIONES", "DETALLE"},
    }

    wb = load_workbook(filename=str(ruta), data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = ["" if h is None else str(h) for h in next(rows)]
    headers_norm = [norm(h) for h in headers]

    idx: dict[str, int] = {}
    for clave, opciones in alias.items():
        for i, h in enumerate(headers_norm):
            if h in opciones:
                idx[clave] = i
                break
        if clave not in idx:
            raise ValueError(
                f"No encontre la columna '{clave}'. Acepto cualquiera de {sorted(opciones)}. "
                f"Headers detectados: {headers}"
            )

    filas: list[dict] = []
    for row in rows:
        if row is None or all(v in (None, "") for v in row):
            continue
        factura = str(row[idx["factura"]] or "").strip()
        if not factura:
            continue
        try:
            num = int(row[idx["num"]] or 0)
        except (TypeError, ValueError):
            continue
        if num <= 0:
            continue
        aceptado_raw = row[idx["aceptado"]]
        try:
            aceptado = int(float(str(aceptado_raw).replace(",", ".") or 0))
        except (TypeError, ValueError):
            aceptado = 0
        detalle = sanitizar(str(row[idx["detalle"]] or ""))
        filas.append(
            {
                "factura": factura,
                "factura_corta": normalizar_factura(factura),
                "num": num,
                "aceptado": aceptado,
                "detalle": detalle,
            }
        )

    # Agrupar por factura, ordenado por # objeción.
    facturas: dict[str, list[dict]] = {}
    for f in filas:
        facturas.setdefault(f["factura_corta"], []).append(f)
    for fac in facturas.values():
        fac.sort(key=lambda r: r["num"])
    return facturas


# ─── Índice de soportes por factura ─────────────────────────────────────────


def cargar_indice(ruta: Path) -> dict[str, Path]:
    """Lee el TXT indice_facturas_HUS.txt y devuelve {factura_corta: Path}.

    Cada línea es una ruta tipo:
        X:\\SERVIDOR RADICACION\\... \\HUS452150
    El último segmento es HUS<numero>; lo usamos como clave.
    """
    if not ruta.is_file():
        raise FileNotFoundError(f"No existe el indice: {ruta}")
    mapa: dict[str, Path] = {}
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            texto = ruta.read_text(encoding=enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise RuntimeError(f"No pude decodificar el indice: {ruta}")

    for linea in texto.splitlines():
        linea = linea.strip()
        if not linea:
            continue
        m = re.search(r"\\(HUS\d{6,12})\s*$", linea, re.IGNORECASE)
        if not m:
            continue
        factura_corta = normalizar_factura(m.group(1))
        mapa[factura_corta] = Path(linea)
    return mapa


def buscar_soportes(
    factura_corta: str,
    factura_larga: str,
    carpeta_glosa: Path,
    indice: dict[str, Path] | None,
    carpeta_share_override: Path | None = None,
) -> tuple[list[Path], list[str]]:
    """Devuelve (lista_de_archivos_a_subir, advertencias).

    Fuentes (en orden):
      1. PDF de respuesta a la glosa en `carpeta_glosa\\HUS<factura>.pdf`.
      2. Todos los archivos de la carpeta del share — viene de
         `carpeta_share_override` (override directo), o del `indice` si se dio.
         Si no hay ninguna, se omite la fuente 2 (sólo se sube el Trámite).
    """
    archivos: list[Path] = []
    avisos: list[str] = []

    # 1) PDF Trámite/respuesta de la glosa. Buscamos por nombre tolerante.
    candidatos_glosa = [
        carpeta_glosa / f"{factura_larga}.pdf",
        carpeta_glosa / f"HUS{factura_corta}.pdf",
        carpeta_glosa / f"HUS0000{factura_corta}.pdf",
    ]
    encontrado = None
    for c in candidatos_glosa:
        if c.is_file():
            encontrado = c
            break
    if not encontrado and carpeta_glosa.is_dir():
        # Búsqueda relajada: cualquier *.pdf que contenga el número.
        for p in carpeta_glosa.glob("*.pdf"):
            if factura_corta in p.stem:
                encontrado = p
                break
    if encontrado:
        archivos.append(encontrado)
    else:
        avisos.append(f"No hallé el PDF de respuesta para HUS{factura_corta} en {carpeta_glosa}")

    # 2) Carpeta de soportes de la factura (HC, HEV, HAM, etc.).
    carpeta_soportes: Path | None = None
    if carpeta_share_override is not None:
        carpeta_soportes = carpeta_share_override
    elif indice is not None:
        carpeta_soportes = indice.get(factura_corta)
        if carpeta_soportes is None:
            avisos.append(f"No hallé HUS{factura_corta} en el indice (sin soportes del share).")
    else:
        avisos.append("Sin --indice ni --carpeta-share: subo sólo el PDF de Trámite (no los HC/HEV/HAM).")

    if carpeta_soportes is not None:
        if not carpeta_soportes.is_dir():
            avisos.append(f"La carpeta de soportes no es accesible: {carpeta_soportes}")
        else:
            for p in sorted(carpeta_soportes.iterdir()):
                if p.is_file():
                    archivos.append(p)
    return archivos, avisos


# ─── Helpers de browser (reusados/adaptados de cargar_soportes_simed) ───────


def _screenshot_debug(page: Page, etiqueta: str) -> Path:
    out_dir = Path("debug_screenshots")
    out_dir.mkdir(exist_ok=True)
    ts = time.strftime("%H%M%S")
    ruta = out_dir / f"{ts}_{etiqueta}.png"
    try:
        page.screenshot(path=str(ruta), full_page=True)
        logger.info(f"  Screenshot de diagnostico: {ruta}")
    except Exception as e:
        logger.warning(f"  No pude tomar screenshot: {e}")
    return ruta


def login(page: Page, user: str, password: str) -> None:
    logger.info("Login al portal...")
    page.goto(PORTAL_LOGIN, wait_until="networkidle")
    u = page.locator(
        "input[type='text']:visible, input[name*='USR']:visible, input[id*='vUSR']:visible"
    ).first
    p = page.locator("input[type='password']:visible").first
    # fill() atómico: el portal acepta fill en usuario/password (sólo el campo
    # Fecha del modal exige tipeo). Ahorra ~2s vs type(delay=30) char a char.
    u.click(); u.fill(user); u.press("Tab")
    p.click(); p.fill(password); p.press("Enter")
    try:
        page.wait_for_url(lambda url: "gamexamplelogin" not in url, timeout=3000)
    except PlaywrightTimeout:
        boton = page.locator(
            "button:has-text('Iniciar'):not([disabled]), "
            "button:has-text('Ingresar'):not([disabled]), "
            "button:has-text('Aceptar'):not([disabled]), "
            "input[type='submit']:not([disabled])"
        ).first
        try:
            boton.wait_for(state="visible", timeout=5000); boton.click(timeout=10000)
        except PlaywrightTimeout:
            page.evaluate("document.forms[0] && document.forms[0].submit()")
    try:
        page.wait_for_url("**/index.aspx", timeout=15000)
    except PlaywrightTimeout:
        page.wait_for_selector("text=Procesos", timeout=10000)
    logger.info("Login OK")


def ir_a_respuesta_glosa_web(page: Page) -> None:
    """Va a la grilla 'Respuesta Glosa Rad. WEB' (glosasfacturaww.aspx)."""
    page.goto(PORTAL_GLOSAS, wait_until="networkidle")
    page.wait_for_selector("text=Respuesta a Glosas", timeout=15000)


# Memo de los selectores que funcionaron la última vez en el filtro DDO.
# El portal es estable dentro de una sesión: una vez que un selector ganó,
# va a seguir ganando — probarlo primero evita timeouts de 2s por intento.
_memo_filtro: dict[str, str] = {}


def filtrar_por_factura(page: Page, factura_corta: str) -> None:
    """Aplica el filtro de columna '# Factura' (DropDownOptions de GeneXus).

    Idéntica lógica a cargar_soportes_simed.py — el portal es el mismo.
    Los selectores que SIEMPRE ganan en producción van primero (trigger
    `img[src*='DDO']`, input `.dropdown-menu input`), y se memorizan.
    """
    header = page.locator("th:has-text('# Factura')").first
    if header.count() == 0:
        header = page.locator("th:has-text('Factura')").first

    triggers = [
        "img[src*='DDO']",  # ← el que gana siempre en producción
        "[data-toggle='dropdown']",
        "a.dropdown-toggle",
        "a[onclick*='DropDownOptions']",
        "a",
    ]
    if _memo_filtro.get("trigger") in triggers:
        triggers.remove(_memo_filtro["trigger"])
        triggers.insert(0, _memo_filtro["trigger"])
    for sel in triggers:
        try:
            header.locator(sel).first.click(timeout=2000)
            _memo_filtro["trigger"] = sel
            break
        except PlaywrightTimeout:
            continue
    else:
        header.click()

    # Sin sleep fijo: el wait_for(visible) del input ya espera a que el
    # dropdown termine de abrir (incluye el XHR que carga las opciones).
    input_buscar = None
    selectores_input = [
        ".dropdown-menu input[type='text']:visible",  # ← gana siempre
        ".dropdown-menu.show input[type='text']:visible",
        ".DDOOptionFilteringDataContainer input:visible",
        ".DDOOptionFilter input:visible",
        "div[class*='DDO'] input[type='text']:visible",
        "div[class*='dropdown'] input[type='text']:visible",
        "input[type='text']:visible:not([readonly]):not([disabled])",
    ]
    if _memo_filtro.get("input") in selectores_input:
        selectores_input.remove(_memo_filtro["input"])
        selectores_input.insert(0, _memo_filtro["input"])
    for sel in selectores_input:
        try:
            elem = page.locator(sel).first
            elem.wait_for(state="visible", timeout=2500)
            input_buscar = elem
            _memo_filtro["input"] = sel
            break
        except PlaywrightTimeout:
            continue
    if input_buscar is None:
        _screenshot_debug(page, f"filtro_no_encontrado_{factura_corta}")
        raise PlaywrightTimeout("No encontre el input del filtro # Factura.")

    input_buscar.fill(factura_corta)
    page.wait_for_timeout(200)  # mini-debounce del DDO (antes 800ms)
    try:
        page.locator(
            "img[src*='ApplyFilter']:visible, img[title*='Apply' i]:visible, "
            "img[title*='Aplicar' i]:visible, img[src*='Search']:visible"
        ).first.click(timeout=2000)
    except PlaywrightTimeout:
        input_buscar.press("Enter")
    page.wait_for_load_state("networkidle")
    # Sin sleep extra: abrir_factura() espera de forma dirigida a que la fila
    # (o el 'No se encontraron registros') aparezca tras aplicar el filtro.


class FacturaNoPendiente(Exception):
    """La factura no está en la grilla de pendientes (ya finalizada / nada que hacer)."""


class ObjecionNoEnGrilla(Exception):
    """La objeción del Excel no existe en la grilla del portal (no reintentar)."""


def abrir_factura(page: Page, factura_corta: str) -> None:
    """Click en el lápiz de la fila filtrada → abre glosasfactura.aspx.

    Espera DIRIGIDA: en vez de un sleep fijo, pollea barato (checks inmediatos
    cada 150ms) hasta que aparezca la fila de la factura o el mensaje de
    'No se encontraron registros' — lo que llegue primero.
    """
    no_resultados = page.locator("text=No se encontraron registros").first
    fila = page.locator(f"tr:has-text('{factura_corta}')").first
    deadline = time.time() + 7
    while True:
        try:
            if no_resultados.is_visible():
                raise FacturaNoPendiente(
                    f"Factura {factura_corta} no está en pendientes (¿ya finalizada?)."
                )
            if fila.is_visible():
                break
        except FacturaNoPendiente:
            raise
        except Exception:
            pass  # DOM re-renderizando; reintento en el próximo ciclo
        if time.time() >= deadline:
            raise FacturaNoPendiente(
                f"Factura {factura_corta} no aparece en la grilla (¿ya finalizada?)."
            )
        page.wait_for_timeout(150)
    boton_editar = fila.locator(
        "a[title*='Actualizar' i], a[title*='Editar' i], a[title*='Edit' i], "
        "a:has(img[src*='ActionUpdate']), a:has(img[title*='Actualizar' i]), "
        "img[title*='Actualizar' i], img[title*='Editar' i], "
        "a[href*='glosasfactura.aspx']"
    ).first
    try:
        boton_editar.wait_for(state="visible", timeout=5000); boton_editar.click()
    except PlaywrightTimeout:
        _screenshot_debug(page, f"editar_no_encontrado_{factura_corta}"); raise
    page.wait_for_url("**/glosasfactura.aspx*", timeout=15000)
    page.wait_for_selector("text=Información General", timeout=10000)


# ─── Lógica nueva: iterar objeciones y rellenar el modal ────────────────────


def _localizar_fila_objecion(page: Page, num_objecion: int):
    """Devuelve la fila <tr> de la objeción N en la grilla visible, o None.

    En la grilla 'Respuesta Glosa Ips' el '# Objeción' es un span con id que
    contiene 'FCTOBJSEC' (verificado en el DOM real). Matcheamos por su texto
    exacto para no confundir con el radicado u otros números.
    """
    target = str(num_objecion)
    # CSS :has() directo: el engine poda filas sin materializarlas todas,
    # mucho más barato que locator("tr").filter(has=...).
    selectores = (
        f"tr:has(span[id*='FCTOBJSEC']:text-is('{target}'))",
        f"tr:has(td:text-is('{target}'))",  # fallback: celda con sólo el número
    )
    for sel in selectores:
        fila = page.locator(sel).first
        try:
            if fila.count() == 0:
                continue
            fila.wait_for(state="visible", timeout=1500)
            return fila
        except Exception:
            continue
    return None


def _fila_contestada(fila) -> bool:
    """True si la objeción de esa fila YA está contestada (no hay que rehacerla).

    Señales en el DOM real: span FCTOBJEST == 'Contestado' y/o badge verde con
    title 'Contestada/Contentada'.
    """
    try:
        est = fila.locator("span[id*='FCTOBJEST']").first
        if est.count() > 0:
            txt = (est.text_content() or "").strip().lower()
            if "contestad" in txt:
                return True
    except Exception:
        pass
    try:
        badge = fila.locator("span[id*='vGRIDBADGE']").first
        if badge.count() > 0:
            title = (
                badge.get_attribute("data-original-title")
                or badge.get_attribute("title")
                or ""
            ).lower()
            if "contest" in title or "content" in title:
                return True
    except Exception:
        pass
    return False


_RE_PAGINA = re.compile(r"P[áa]gina:?\s*(\d+)\s+de\s+(\d+)", re.IGNORECASE)


def _pagina_actual_grilla(page: Page) -> tuple[int, int] | None:
    """(página_actual, total) si la pantalla muestra UN único 'Página X de Y'.

    Lectura instantánea que evita el ciclo de clicks 'Ant' cuando la grilla ya
    está en página 1. Si el label no existe o es ambiguo (más de un match,
    p.ej. dos grillas paginadas), devuelve None y se usa el camino clásico.
    """
    try:
        loc = page.locator(r"text=/P[áa]gina:?\s*\d+\s+de\s+\d+/i")
        if loc.count() != 1:
            return None
        m = _RE_PAGINA.search(loc.first.text_content(timeout=600) or "")
        if m:
            return int(m.group(1)), int(m.group(2))
    except Exception:
        pass
    return None


def _link_paginacion_deshabilitado(link) -> bool:
    """True si el link 'Ant'/'Sig' está deshabilitado (class/aria/li padre).

    Detectarlo ANTES de clickear ahorra el ciclo click+espera completo cuando
    la grilla ya está en el borde (página 1 para 'Ant', última para 'Sig').
    """
    try:
        return bool(link.evaluate(
            """el => {
                const cls = (el.className || '') + ' ' +
                            (el.closest('li') ? el.closest('li').className : '');
                return /disabled|inactive/i.test(cls)
                    || el.getAttribute('aria-disabled') === 'true'
                    || el.hasAttribute('disabled');
            }""",
            timeout=1000,
        ))
    except Exception:
        return False


def _primera_objecion_visible(page: Page) -> str | None:
    """Texto del primer '# Objeción' visible en la grilla (para detectar página 1)."""
    try:
        loc = page.locator("span[id*='FCTOBJSEC']").first
        if loc.count() > 0:
            return (loc.text_content(timeout=600) or "").strip()
    except Exception:
        pass
    return None


def _esperar_cambio_primera_fila(page: Page, antes: str | None, timeout_ms: int = 4000) -> bool:
    """Espera a que cambie el primer '# Objeción' visible (señal de que la
    paginación AJAX re-renderizó la grilla). True si cambió.

    OJO: acá networkidle NO sirve — el estado de carga es sticky por navegación
    y la paginación de GeneXus es un XHR, así que wait_for_load_state devuelve
    al instante. La señal confiable es el cambio del contenido de la grilla.
    """
    deadline = time.time() + timeout_ms / 1000
    while time.time() < deadline:
        ahora = _primera_objecion_visible(page)
        if ahora is not None and ahora != antes:
            return True
        page.wait_for_timeout(120)
    return False


def _siguiente_pagina_grilla(page: Page) -> bool:
    """Click en 'Sig' de la paginación. Devuelve True si avanzó, False si no hay más."""
    # Salida temprana sin click si el label dice que ya estamos en la última.
    pag = _pagina_actual_grilla(page)
    if pag is not None and pag[0] >= pag[1]:
        return False
    # En GeneXus la paginación suele ser "Ant | 1 2 3 ... | Sig" con links.
    boton_sig = page.locator(
        "a:has-text('Sig'):not(.disabled), a:has-text('Siguiente'):not(.disabled), "
        "img[title*='Sig' i]:visible, img[title*='Siguiente' i]:visible"
    ).first
    try:
        if boton_sig.count() == 0 or not boton_sig.is_visible():
            return False
        if _link_paginacion_deshabilitado(boton_sig):
            return False
        antes = _primera_objecion_visible(page)
        boton_sig.click(timeout=3000)
        return _esperar_cambio_primera_fila(page, antes)
    except PlaywrightTimeout:
        return False


def _ir_primera_pagina_grilla(page: Page) -> None:
    """Lleva la grilla de objeciones a la primera página.

    La grilla puede estar ordenada por otra columna (ej. 'VI Aceptado'), así que
    las objeciones NO siguen el orden 1,2,3 por página; para encontrar una hay
    que arrancar en la página 1 y escanear hacia adelante.

    Fast-paths (sin ningún click):
      1. El label 'Página X de Y' dice X == 1.
      2. El link 'Ant' no existe, no es visible o está deshabilitado.
    """
    try:
        page.wait_for_selector("text=Respuesta Glosa Ips", timeout=8000)
    except Exception:
        return
    pag = _pagina_actual_grilla(page)
    if pag is not None and pag[0] <= 1:
        return
    for _ in range(12):
        ant = page.locator("a:has-text('Ant'), a:has-text('Anterior')").first
        try:
            if ant.count() == 0 or not ant.is_visible():
                return
            if _link_paginacion_deshabilitado(ant):
                return
            antes = _primera_objecion_visible(page)
            ant.click(timeout=2000)
            cambio = _esperar_cambio_primera_fila(page, antes)
        except Exception:
            return
        # Si el primer renglón no cambió, el clic 'Ant' no movió nada → página 1.
        if not cambio:
            return
        pag = _pagina_actual_grilla(page)
        if pag is not None and pag[0] <= 1:
            return


_MODAL_URL_FRAGMENT = "respuestaglosaipsweb"  # respuestaglosaipsweb.aspx


def _ctx_modal_por_texto(page: Page):
    """Fallback legado: UN pase buscando marcadores de texto en todos los frames."""
    marcadores = ["Detalle Respuesta", "Glosa Estado", "Detalle Glosa"]
    for ctx in page.frames:
        for marc in marcadores:
            try:
                loc = ctx.locator(f"text={marc}")
                if loc.count() > 0 and loc.first.is_visible():
                    return ctx
            except Exception:
                continue
    return None


def _ctx_modal(page: Page, timeout_ms: int = 7000):
    """Devuelve el contexto (Page/Frame) donde vive el modal de respuesta, o None.

    El modal SIEMPRE carga en un iframe con URL respuestaglosaipsweb.aspx, así
    que el camino rápido es matchear el frame por URL: `page.frames` es local
    (sin roundtrip al browser), o sea que el chequeo es instantáneo. Sólo si la
    URL nunca aparece se cae al scan de marcadores de texto (costoso) que
    barría todos los frames en versiones anteriores.
    """
    deadline = time.time() + timeout_ms / 1000
    proximo_scan_texto = time.time() + 1.5  # darle ventaja al fast-path por URL
    while True:
        fr = next(
            (f for f in page.frames if _MODAL_URL_FRAGMENT in (f.url or "").lower()),
            None,
        )
        if fr is not None:
            # Frame correcto: esperar (dirigido) a que el form esté renderizado.
            restante = max(800, int((deadline - time.time()) * 1000))
            try:
                fr.locator("text=Detalle Respuesta").first.wait_for(
                    state="visible", timeout=restante
                )
            except Exception:
                logger.warning(
                    "  (frame del modal hallado por URL pero el marcador tardó; sigo igual)"
                )
            return fr
        if time.time() >= proximo_scan_texto:
            ctx = _ctx_modal_por_texto(page)
            if ctx is not None:
                return ctx
            proximo_scan_texto = time.time() + 0.6
        if time.time() >= deadline:
            return None
        page.wait_for_timeout(100)


def _primer_visible(scope, selector: str):
    """Primer elemento VISIBLE que matchea `selector` dentro de `scope`, o None."""
    loc = scope.locator(selector)
    try:
        n = loc.count()
    except Exception:
        return None
    for i in range(n):
        cand = loc.nth(i)
        try:
            if cand.is_visible():
                return cand
        except Exception:
            continue
    return None


def _dump_diagnostico(page: Page, fila, num_objecion: int) -> None:
    """Guarda los frames y el HTML de la fila para diagnosticar selectores."""
    out_dir = Path("debug_screenshots")
    out_dir.mkdir(exist_ok=True)
    ruta = out_dir / f"{time.strftime('%H%M%S')}_dom_objecion_{num_objecion}.txt"
    try:
        partes = ["=== FRAMES (url) ==="]
        for fr in page.frames:
            partes.append(f"  {fr.url}")
        partes.append("\n=== outerHTML de la fila de la objeción ===")
        try:
            partes.append(fila.evaluate("el => el.outerHTML"))
        except Exception as e:
            partes.append(f"(no pude obtener outerHTML: {e})")
        ruta.write_text("\n".join(partes), encoding="utf-8")
        logger.info(f"  Volcado DOM para diagnóstico: {ruta}")
    except Exception as e:
        logger.warning(f"  No pude volcar el DOM: {e}")


def abrir_modal_objecion(page: Page, num_objecion: int, rehacer: bool = False):
    """Abre el modal 'Respuesta Glosa Ips Web' de la objeción N y devuelve el
    contexto (Page o Frame) donde está el formulario.

    Si la objeción ya está contestada y `rehacer` es False, devuelve el
    sentinel 'YA_CONTESTADA' sin abrir nada (para no re-subir soportes ni
    pisar respuestas existentes).
    """
    page.wait_for_selector("text=Respuesta Glosa Ips", timeout=10000)

    # Guarda: si quedó abierto el popup de una objeción anterior, intercepta el
    # click de la fila siguiente. Esperá a que cierre; si no, ciérralo a la fuerza.
    try:
        popup = page.locator("iframe#gxp0_ifrm").first
        if popup.count() > 0 and popup.is_visible():
            try:
                popup.wait_for(state="hidden", timeout=4000)
            except Exception:
                logger.warning("  (popup anterior seguía abierto; lo cierro a la fuerza)")
                _cerrar_popup_forzado(page)
                page.wait_for_timeout(500)
    except Exception:
        pass

    # La grilla puede estar ordenada por otra columna (VI Aceptado), así que la
    # objeción puede estar en cualquier página. Arrancamos en la página 1 y
    # escaneamos hacia adelante TODAS las páginas.
    _ir_primera_pagina_grilla(page)
    fila = None
    for _ in range(12):  # hasta 12 páginas
        fila = _localizar_fila_objecion(page, num_objecion)
        if fila is not None:
            break
        if not _siguiente_pagina_grilla(page):
            break
    if fila is None:
        _screenshot_debug(page, f"objecion_no_hallada_{num_objecion}")
        raise ObjecionNoEnGrilla(
            f"Objeción #{num_objecion} no está en la grilla del portal "
            "(el Excel tiene más objeciones que el portal para esta factura)."
        )

    # Si ya está contestada y no pedimos rehacer, omitir (no re-subir soportes).
    if not rehacer and _fila_contestada(fila):
        return "YA_CONTESTADA"

    # Estrategias de click, en orden — la PRIMERA que haga aparecer el modal gana.
    # El botón real es img[id*='vBTNRESPUESTA'][title='Dar Respuesta']: matchear
    # por id es lo que SIEMPRE gana en producción (la vieja estrategia por
    # title='Dar Respuesta' era el mismo elemento con otro selector — eliminada
    # para no pagar latencia extra por intento fallido).
    estrategias = [
        # 1) Botón exacto del DOM: img id W0104vBTNRESPUESTA_NNNN ("Dar Respuesta").
        ("btn_respuesta_id",
         "img[id*='vBTNRESPUESTA'], a:has(img[id*='vBTNRESPUESTA'])"),
        # 2) Último recurso: cualquier <img> clickable en las primeras 2 celdas.
        ("img_inicio", "xpath=.//td[position()<=2]//img"),
    ]

    ultimo_error = None
    frames_logueados = False
    for nombre, sel in estrategias:
        candidato = _primer_visible(fila, sel)
        if candidato is None:
            continue
        try:
            candidato.scroll_into_view_if_needed(timeout=2000)
        except Exception:
            pass
        try:
            candidato.click(timeout=4000)
        except PlaywrightTimeout as e:
            ultimo_error = e
            continue

        if not frames_logueados:
            urls = [(fr.url or "")[-70:] for fr in page.frames]
            logger.info(f"  click '{nombre}' OK. frames: {urls}")
            frames_logueados = True

        ctx = _ctx_modal(page, timeout_ms=6000)
        if ctx is not None:
            tipo = "página" if ctx == page.main_frame else "iframe"
            logger.info(f"  ✓ modal abierto (estrategia '{nombre}', en {tipo})")
            return ctx

        ultimo_error = "modal no detectado tras el click"
        try:
            page.keyboard.press("Escape")
        except Exception:
            pass
        page.wait_for_timeout(400)

    _dump_diagnostico(page, fila, num_objecion)
    _screenshot_debug(page, f"modal_no_abrio_{num_objecion}")
    raise RuntimeError(
        f"No pude abrir el modal de la objecion #{num_objecion} "
        f"(último: {ultimo_error}). Mirá el volcado DOM en debug_screenshots/."
    )


def _clic_tab(ctx, page: Page, nombre_tab: str, marcador: str | None = None) -> None:
    """Click en un tab del modal por su texto, tolerante.

    Si se pasa `marcador` (selector de un elemento propio de ese tab), se
    espera de forma DIRIGIDA a que sea visible — si el tab ya estaba activo
    (caso típico: 'Información Glosa' es el tab inicial del modal) la espera
    es instantánea. Sin marcador, settle corto como antes.
    """
    try:
        tab = ctx.locator(
            f"a:has-text('{nombre_tab}'), li:has-text('{nombre_tab}') a, "
            f"span:has-text('{nombre_tab}')"
        ).first
        if tab.count() > 0:
            tab.click(timeout=2500)
            if marcador:
                try:
                    ctx.locator(marcador).first.wait_for(state="visible", timeout=2500)
                    return
                except Exception:
                    pass  # marcador no apareció: caer al settle corto
            page.wait_for_timeout(350)
    except PlaywrightTimeout:
        pass


def _setear_fecha_respuesta(ctx, page: Page) -> None:
    """Setea 'Fecha Respuesta IPS' (obligatoria) con la fecha de hoy si está
    vacía. Atómico (fill); si el campo está enmascarado y fill no prende,
    reintenta tipeando sólo dígitos."""
    hoy = date.today().strftime("%d/%m/%Y")
    fecha_input = _primer_visible(ctx, "input[id*='FECRTA' i], input[id*='FECRES' i]")
    if fecha_input is None:
        fecha_input = _primer_visible(
            ctx, "xpath=//*[contains(text(),'Fecha Respuesta')]/following::input[1]"
        )
    if fecha_input is None:
        return
    try:
        val = fecha_input.input_value(timeout=800) or ""
        if any(c.isdigit() for c in val):
            return  # ya trae fecha
        fecha_input.fill(hoy)
        fecha_input.press("Tab")
        page.wait_for_timeout(200)
        nv = fecha_input.input_value(timeout=800) or ""
        if not any(c.isdigit() for c in nv):
            fecha_input.click()
            fecha_input.fill("")
            fecha_input.type(hoy.replace("/", ""), delay=80)  # campo enmascarado
            fecha_input.press("Tab")
        logger.info(f"  Fecha Respuesta seteada: {hoy}")
    except Exception as e:
        logger.warning(f"  No pude setear 'Fecha Respuesta' ({e}); sigo.")


def llenar_informacion_glosa(ctx, page: Page, aceptado: int, detalle: str) -> None:
    """Tab 'Información Glosa' del modal: Aceptado, NC (vacío), Detalle Respuesta.

    `ctx` es el contexto del modal (Page o Frame) devuelto por abrir_modal_objecion.
    """
    # El textarea de Detalle es el marcador del tab: como 'Información Glosa'
    # es el tab inicial del modal, la espera resuelve al instante.
    _clic_tab(ctx, page, "Información Glosa", marcador="textarea")

    # Campo Aceptado.
    aceptado_input = _primer_visible(
        ctx, "input[name*='ACEPTADO' i], input[id*='ACEPTADO' i]"
    )
    if aceptado_input is None:
        aceptado_input = ctx.locator(
            "xpath=//label[contains(., 'Aceptado')]/following::input[1]"
        ).first
    try:
        aceptado_input.click()
        aceptado_input.fill("")
        aceptado_input.type(str(aceptado), delay=20)
        aceptado_input.press("Tab")
    except Exception as e:
        logger.warning(f"  No pude escribir 'Aceptado' ({e}); sigo.")

    # Campo Nota Credito: dejar vacío (ya no van NCs).
    nc_input = ctx.locator(
        "xpath=//label[contains(., 'Nota Credito')]/following::input[1]"
    ).first
    try:
        if nc_input.count() > 0:
            nc_input.fill("")
    except Exception:
        pass

    # Campo Detalle Respuesta (textarea). fill() ATÓMICO, no type char-a-char:
    # el tipeo lento se mezclaba con el campo de fecha y partía el texto.
    detalle_input = _primer_visible(
        ctx, "textarea[name*='DETALLE' i], textarea[id*='DETALLE' i], textarea"
    )
    if detalle_input is None:
        _screenshot_debug(page, "sin_textarea_detalle")
        raise RuntimeError("No encontré el textarea de 'Detalle Respuesta' en el modal.")
    detalle_input.fill(detalle[:4000])

    # Fecha Respuesta IPS (OBLIGATORIO — rojo si queda vacío). Se setea AL FINAL
    # y de forma atómica, sólo si está vacío.
    _setear_fecha_respuesta(ctx, page)


def _buscar_en_frames(page: Page, ctx, selector: str):
    """Primer locator VISIBLE que matchea `selector`, probando primero `ctx`
    y luego todos los frames (por si hay un iframe anidado)."""
    contextos = [ctx] + [f for f in page.frames if f is not ctx]
    for c in contextos:
        cand = _primer_visible(c, selector)
        if cand is not None:
            return cand
    return None


def _ya_tiene_soportes(ctx) -> bool:
    """True si la grilla 'Soportes' del modal ya tiene archivos cargados.

    Los soportes existentes aparecen con estado 'Activo' en la tabla; en una
    objeción pendiente esa tabla está vacía.
    """
    try:
        if ctx.locator("text=Activo").count() > 0:
            return True
    except Exception:
        pass
    return False


def _cerrar_popup_forzado(page: Page) -> None:
    """Cierra a la fuerza el popup de respuesta si quedó colgado abierto
    (con la X del popup o Escape), para no bloquear la objeción siguiente."""
    for sel in (
        "#gxp0_b .close", "#gxp0_b .gx-popup-close", "div.gx-popup .close",
        "div.gx-popup img[onclick*='close' i]", "[id^='gxp'] .close",
    ):
        try:
            b = page.locator(sel).first
            if b.count() > 0 and b.is_visible():
                b.click(timeout=1500)
                page.wait_for_timeout(500)
                break
        except Exception:
            continue
    try:
        page.keyboard.press("Escape")
        page.wait_for_timeout(400)
    except Exception:
        pass


def subir_soportes_modal(ctx, page: Page, archivos: list[Path]) -> None:
    """Tab 'Soportes' del modal: sube los archivos SÓLO si la objeción aún no
    tiene soportes (no re-sube ni duplica los ya cargados)."""
    if not archivos:
        return
    _clic_tab(ctx, page, "Soportes")
    page.wait_for_timeout(600)

    if _ya_tiene_soportes(ctx):
        logger.info("  esta objeción ya tiene soportes cargados → no re-subo")
        return

    # El <input type=file> puede estar en ctx o en un iframe anidado.
    file_input = None
    for c in [ctx] + [f for f in page.frames if f is not ctx]:
        loc = c.locator("input[type='file']")
        try:
            if loc.count() > 0:
                file_input = loc.first
                break
        except Exception:
            continue
    if file_input is None:
        _screenshot_debug(page, "sin_input_file_soportes")
        logger.warning("  No encontré el input de archivos en el tab Soportes; sigo sin soportes.")
        return
    file_input.set_input_files([str(p) for p in archivos])
    page.wait_for_timeout(1000)

    # 'Iniciar subida'.
    btn_iniciar = _buscar_en_frames(
        page, ctx, "button:has-text('Iniciar subida'), a:has-text('Iniciar subida')"
    )
    if btn_iniciar is not None:
        try:
            btn_iniciar.click(timeout=5000)
        except PlaywrightTimeout:
            pass
    else:
        logger.warning("  Sin botón 'Iniciar subida' visible; intento seguir.")

    # Esperar a que terminen las barras de progreso.
    deadline = time.time() + (5 + 4 * len(archivos))
    while time.time() < deadline:
        en_progreso = 0
        for c in [ctx] + [f for f in page.frames if f is not ctx]:
            try:
                en_progreso += c.locator(".progress-bar:not([aria-valuenow='100'])").count()
            except Exception:
                pass
        if en_progreso == 0:
            break
        page.wait_for_timeout(500)

    # 'Agregar' (rojo) — texto EXACTO para no agarrar 'Agregar archivos...'.
    btn_agregar = _buscar_en_frames(
        page, ctx, "button:text-is('Agregar'), input[value='Agregar']"
    )
    if btn_agregar is not None:
        try:
            btn_agregar.click(timeout=5000)
            page.wait_for_timeout(500)
        except PlaywrightTimeout:
            logger.warning("  No pude clickear 'Agregar' del modal de soportes.")


def _hay_caracteres_especiales(page: Page, ctx) -> bool:
    for c in [ctx] + [f for f in page.frames if f is not ctx]:
        try:
            err = c.locator("text=caracteres especiales").first
            if err.count() > 0 and err.is_visible():
                return True
        except Exception:
            continue
    return False


def _settle_post_confirmacion(page: Page) -> None:
    """Settle corto tras el cierre del popup: GeneXus dispara un XHR que
    refresca la grilla (badge 'Contestado'). 400ms cubren ese repaint —
    antes había 400ms acá + 800ms más en el loop de procesar_factura.
    (networkidle no sirve de señal: es sticky por navegación y la grilla se
    refresca por XHR, devolvería al instante sin esperar nada.)"""
    page.wait_for_timeout(400)


def _dump_botones(page: Page, etiqueta: str) -> None:
    """Lista los botones visibles de todos los frames (para descubrir un 2do
    confirm u otros diálogos) en un .txt de diagnóstico."""
    out_dir = Path("debug_screenshots")
    out_dir.mkdir(exist_ok=True)
    ruta = out_dir / f"{time.strftime('%H%M%S')}_botones_{etiqueta}.txt"
    lineas: list[str] = []
    for fr in page.frames:
        try:
            els = fr.locator("button, input[type='button'], input[type='submit'], a.btn")
            for i in range(min(els.count(), 50)):
                e = els.nth(i)
                try:
                    if not e.is_visible():
                        continue
                    txt = (e.inner_text(timeout=400) or "").strip() or (e.get_attribute("value") or "")
                    eid = e.get_attribute("id") or ""
                    lineas.append(f"[{(fr.url or '')[-45:]}] '{txt[:40]}' id={eid}")
                except Exception:
                    continue
        except Exception:
            continue
    try:
        ruta.write_text("\n".join(lineas), encoding="utf-8")
        logger.info(f"  Botones visibles → {ruta}")
    except Exception:
        pass


def confirmar_modal(ctx, page: Page) -> None:
    """Click 'Confirmar' del modal y ESPERA a que el portal guarde y cierre.

    Clave: NO cerramos el popup a la fuerza apenas vence un timeout corto —
    eso abortaba el guardado y dejaba la glosa pendiente. Esperamos con
    wait_for(state='hidden') sobre el iframe, que despierta en el MISMO
    instante en que el portal lo cierra (el polling anterior de 400ms perdía
    hasta 400ms por redondeo, más 1200ms de sleep fijo antes de empezar).

    Presupuesto total: 25s como siempre, en dos fases (5s + 20s) con un
    chequeo de la validación de 'caracteres especiales' entre medio para
    fallar rápido si el portal rechazó el texto.
    """
    # Volver a 'Información Glosa' para asegurar que el Confirmar guarda el form.
    # (Si ya estamos en ese tab — siempre en --sin-soportes — es instantáneo.)
    _clic_tab(ctx, page, "Información Glosa", marcador="textarea")
    btn = _buscar_en_frames(
        page, ctx,
        "input#BTNTRN_ENTER, input[type='button'][value='Confirmar'], "
        "button:text-is('Confirmar')",
    )
    if btn is None:
        _screenshot_debug(page, "modal_sin_confirmar")
        raise RuntimeError("No encontré el botón 'Confirmar' del modal.")
    btn.click(timeout=6000)

    popup = page.locator("iframe#gxp0_ifrm").first
    # Fase 1: caso feliz — el portal cierra en 200-500ms.
    try:
        popup.wait_for(state="hidden", timeout=5000)
        _settle_post_confirmacion(page)
        return
    except PlaywrightTimeout:
        pass

    # Sigue abierto a los 5s: ¿la validación del portal rechazó el texto?
    if _hay_caracteres_especiales(page, ctx):
        _screenshot_debug(page, "validacion_texto_rechazada")
        raise RuntimeError(
            "El portal rechazó el texto por 'caracteres especiales' (revisar sanitización)."
        )

    # Fase 2: resto del presupuesto para guardados lentos del portal.
    try:
        popup.wait_for(state="hidden", timeout=20000)
        _settle_post_confirmacion(page)
        return
    except PlaywrightTimeout:
        pass

    # No cerró en 25s → diagnóstico + cierre forzado; la objeción NO se guardó.
    _dump_botones(page, "modal_no_cierra")
    _screenshot_debug(page, "modal_no_cierra")
    _cerrar_popup_forzado(page)
    raise RuntimeError(
        "El modal no cerró tras Confirmar en 25s — la objeción NO se guardó "
        "(ver el dump de botones por si hay un 2do confirm)."
    )


def confirmar_form_principal(page: Page) -> None:
    """Click 'Confirmar' de la pantalla 'Glosas Factura' (debajo de la grilla)."""
    page.locator(
        "button:has-text('Confirmar'):visible, input[type='button'][value='Confirmar']:visible"
    ).first.click(timeout=8000)
    page.wait_for_load_state("networkidle")


def enviar_finalizar(page: Page, factura_corta: str) -> str:
    """Click botón verde de la grilla (igual que las NCs) y valida el OK del portal."""
    ir_a_respuesta_glosa_web(page)
    filtrar_por_factura(page, factura_corta)
    fila = page.locator(f"tr:has-text('{factura_corta}')").first
    fila.wait_for(state="visible", timeout=8000)
    boton_verde = fila.locator(
        "a:has(img[src*='ActionExportFile2']), img[src*='ActionExportFile2'], "
        "a:has(img[title*='Enviar' i]), img[title*='Enviar' i], "
        "a:has(img[title*='Finalizar' i]), img[title*='Finalizar' i]"
    ).first
    boton_verde.wait_for(state="visible", timeout=5000)
    boton_verde.click()

    # Esperar el diálogo de mensajes del portal y leer su texto. El chequeo de
    # URL sobre page.frames es local (sin roundtrip), así que el poll es barato
    # y arranca de inmediato (antes: 2500ms de sleep fijo + poll de 500ms).
    texto_msg = ""
    deadline = time.time() + 14
    while time.time() < deadline:
        fr = next((f for f in page.frames if "mensaje" in (f.url or "")), None)
        if fr is not None:
            try:
                texto_msg = (fr.locator("body").inner_text(timeout=1500) or "").strip()
            except Exception:
                texto_msg = ""
            if texto_msg:
                break
        page.wait_for_timeout(250)

    if not texto_msg:
        _screenshot_debug(page, f"sin_dialogo_final_{factura_corta}")
        return "OK_SIN_DIALOGO"

    # Confirmar/cerrar el diálogo de mensaje.
    for fr in page.frames:
        try:
            b = fr.locator(
                "button:has-text('Confirmar'), button:has-text('Aceptar'), "
                "button:has-text('OK'), input[value='Confirmar']"
            ).first
            if b.count() > 0 and b.is_visible():
                b.click(timeout=2000)
                page.wait_for_timeout(500)
                break
        except Exception:
            continue

    up = texto_msg.upper()
    if any(k in up for k in ("RESPUESTA CARGADA", "COMPLETAD", "EXITOS")):
        return "OK"
    if "PENDIENTE" in up:
        return f"PENDIENTES: {texto_msg[:140]}"
    return f"MENSAJE: {texto_msg[:160]}"


# ─── Driver principal ───────────────────────────────────────────────────────


def procesar_factura(
    page: Page,
    factura_corta: str,
    factura_larga: str,
    objeciones: list[dict],
    archivos_soportes: list[Path],
    rehacer: bool = False,
    max_obj: int = 0,
    subir_soportes: bool = True,
) -> dict:
    """Procesa UNA factura: filtra, abre, loop objeciones, confirma, envia."""
    a_procesar = objeciones[:max_obj] if max_obj and max_obj > 0 else objeciones
    reg = {"factura": factura_larga, "objeciones": len(a_procesar), "estado": "", "detalle": ""}
    respondidas = 0
    omitidas = 0
    fallidas: list[int] = []
    try:
        ir_a_respuesta_glosa_web(page)
        filtrar_por_factura(page, factura_corta)
        abrir_factura(page, factura_corta)

        for ob in a_procesar:
            # Hasta 2 intentos por objeción: el portal a veces tarda/cuelga de
            # forma intermitente (no abre el modal o no cierra en 25s). Un
            # reintento recupera la mayoría sin dejar cleanup manual.
            ok = False
            no_en_portal = False
            for intento in range(2):
                try:
                    res = abrir_modal_objecion(page, ob["num"], rehacer)
                    if res == "YA_CONTESTADA":
                        omitidas += 1
                        logger.info(f"  objecion #{ob['num']}: ya contestada → omito (usá --rehacer para pisarla)")
                        ok = True
                        break
                    sufijo = "" if intento == 0 else f" (reintento {intento})"
                    logger.info(f"  → objecion #{ob['num']}{sufijo} (aceptado={ob['aceptado']}, detalle {len(ob['detalle'])} chars)")
                    ctx = res
                    llenar_informacion_glosa(ctx, page, ob["aceptado"], ob["detalle"])
                    if subir_soportes:
                        subir_soportes_modal(ctx, page, archivos_soportes)
                    confirmar_modal(ctx, page)
                    respondidas += 1
                    ok = True
                    # Sin sleep extra: confirmar_modal ya hace el settle del
                    # refresh de la grilla (_settle_post_confirmacion).
                    break
                except ObjecionNoEnGrilla as e:
                    # No existe en el portal: no tiene sentido reintentar.
                    logger.info(f"  objecion #{ob['num']}: {e}")
                    no_en_portal = True
                    break
                except Exception as e:
                    logger.warning(f"  intento {intento + 1} de objecion #{ob['num']} falló: {type(e).__name__}: {e}")
                    _cerrar_popup_forzado(page)
                    page.wait_for_timeout(800)
            if not ok and not no_en_portal:
                fallidas.append(ob["num"])
                logger.error(f"  ✗ objecion #{ob['num']} falló tras 2 intentos.")
                _screenshot_debug(page, f"error_obj_{ob['num']}")
                _cerrar_popup_forzado(page)
                page.wait_for_timeout(600)

        if respondidas == 0:
            reg["estado"] = "SIN_PENDIENTES" if not fallidas else "ERROR"
            reg["detalle"] = (
                f"{omitidas} ya contestadas"
                + (f", {len(fallidas)} fallaron: {fallidas[:10]}" if fallidas else "")
            )
            logger.info(f"  {factura_larga}: respondidas=0, omitidas={omitidas}, fallidas={len(fallidas)}.")
            return reg

        # En piloto parcial (--max-obj) no finalizamos: faltarían objeciones y
        # el botón verde sólo avisaría 'glosas pendientes'.
        if max_obj and max_obj > 0:
            reg["estado"] = "PILOTO_PARCIAL"
            reg["detalle"] = (
                f"{respondidas} respondidas de {max_obj} (sin finalizar)"
                + (f", {len(fallidas)} fallaron: {fallidas}" if fallidas else "")
            )
            logger.info(f"  ✓ piloto parcial: {respondidas} respondidas, {len(fallidas)} fallaron.")
            return reg

        # Finalizar SÓLO si respondimos algo nuevo.
        try:
            confirmar_form_principal(page)
        except Exception as e:
            logger.warning(f"  (Confirmar del form principal no aplicó: {e})")
        estado = enviar_finalizar(page, factura_corta)
        reg["estado"] = estado
        reg["detalle"] = (
            f"{respondidas} respondidas, {omitidas} omitidas"
            + (f", {len(fallidas)} fallaron: {fallidas[:10]}" if fallidas else "")
        )
    except FacturaNoPendiente as e:
        reg["estado"] = "NO_PENDIENTE"
        reg["detalle"] = str(e)
        logger.info(f"  ⏭ {factura_larga}: {e}")
    except Exception as e:
        reg["estado"] = "ERROR"
        reg["detalle"] = f"{type(e).__name__}: {e} (respondidas={respondidas}, omitidas={omitidas})"
        logger.error(f"  ✗ {factura_larga}: {reg['detalle']}")
        _screenshot_debug(page, f"error_{factura_corta}")
    return reg


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Carga masiva de respuestas a glosas en SIMED (item por item).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--excel", type=Path, required=True, help="Excel con respuestas (de extraer_respuestas_glosa.py).")
    parser.add_argument("--soportes-glosa", type=Path, default=None, help="Carpeta con HUS<factura>.pdf (Trámite/Respuesta). No requerida con --sin-soportes.")
    parser.add_argument(
        "--indice",
        type=Path,
        default=None,
        help="TXT con rutas <ruta_share>\\HUS<factura>. Si se omite, se usa --carpeta-share o se sube sólo el Trámite.",
    )
    parser.add_argument(
        "--carpeta-share",
        type=Path,
        default=None,
        help="Carpeta del share con los soportes de la factura (HC/HEV/HAM). "
        "Útil para piloto sin tener el indice TXT armado.",
    )
    grupo = parser.add_mutually_exclusive_group(required=True)
    grupo.add_argument("--solo", type=str, help="Procesar solo esta factura (HUS... o número corto).")
    grupo.add_argument("--todas", action="store_true", help="Procesar todas las facturas del Excel.")
    parser.add_argument("--con-cabeza", action="store_true", help="Mostrar el browser (no headless).")
    parser.add_argument("--rehacer", action="store_true",
                        help="Rehacer objeciones YA contestadas (por defecto se omiten para no re-subir soportes).")
    parser.add_argument("--max-obj", type=int, default=0,
                        help="Procesar como mucho N objeciones por factura (0 = todas). Útil para pilotos rápidos.")
    parser.add_argument("--sin-soportes", action="store_true",
                        help="NO subir soportes: sólo carga la respuesta (texto + fecha) y confirma. "
                        "Mucho más rápido y evita que la subida trabe el cierre del modal.")
    parser.add_argument("--lento", action="store_true", help="Slow-motion 300ms (debug).")
    parser.add_argument("--reporte", type=Path, default=Path("reporte_glosa.csv"))
    parser.add_argument("--log", type=Path, default=None)
    args = parser.parse_args()
    setup_logging(args.log)

    if not args.sin_soportes and args.soportes_glosa is None:
        parser.error("--soportes-glosa es obligatorio (o usá --sin-soportes para no subir soportes).")

    _exigir_playwright()
    user, password = cargar_credenciales()
    logger.info(f"Usuario SIMED: {user}")

    facturas = leer_excel_respuestas(args.excel)
    if args.solo:
        objetivo = normalizar_factura(args.solo)
        facturas = {k: v for k, v in facturas.items() if k == objetivo}
        if not facturas:
            logger.error(f"No hallé la factura {args.solo} en el Excel.")
            return 1

    indice: dict[str, Path] | None = None
    if args.indice is not None:
        indice = cargar_indice(args.indice)
        logger.info(f"Indice cargado: {len(indice):,} facturas mapeadas a soportes del share.")
    elif args.carpeta_share is not None:
        if not args.carpeta_share.is_dir():
            logger.error(f"--carpeta-share no es una carpeta accesible: {args.carpeta_share}")
            return 1
        logger.info(f"Carpeta share fija: {args.carpeta_share}")
    else:
        logger.warning("Sin --indice ni --carpeta-share: subiré sólo el PDF de Trámite por objeción.")
    if args.carpeta_share is not None and args.todas:
        logger.error("--carpeta-share sólo tiene sentido con --solo (1 factura). Para masivo usá --indice.")
        return 1
    logger.info(f"Facturas a procesar: {len(facturas)}")

    resultados: list[dict] = []

    # Reporte CSV append-as-you-go: cada factura se escribe apenas termina y
    # se flushea a disco cada 5, así un corte a mitad del masivo (Ctrl+C,
    # caída de red, kill) no pierde el progreso del registro. Antes el CSV
    # se escribía ENTERO recién al final.
    args.reporte.parent.mkdir(parents=True, exist_ok=True)
    f_rep = args.reporte.open("w", newline="", encoding="utf-8-sig")
    w_rep = csv.DictWriter(f_rep, fieldnames=["factura", "objeciones", "estado", "detalle"])
    w_rep.writeheader()
    f_rep.flush()

    def registrar(reg: dict) -> None:
        resultados.append(reg)
        w_rep.writerow(reg)
        if len(resultados) % 5 == 0:
            f_rep.flush()

    t0 = time.time()
    def _sesion_muerta(exc: Exception) -> bool:
        """True si el error indica que el browser/page/context ya no es usable."""
        msg = str(exc).lower()
        return (
            "target page, context or browser has been closed" in msg
            or "browser has been closed" in msg
            or "target closed" in msg
            or "connection closed" in msg
            or exc.__class__.__name__ == "TargetClosedError"
        )

    def _abrir_sesion(p):
        """Crea browser + context + page + login. Devuelve (browser, ctx, page)."""
        b = p.chromium.launch(headless=not args.con_cabeza, slow_mo=300 if args.lento else 0)
        c = b.new_context(accept_downloads=True)
        pg = c.new_page()
        pg.on("dialog", lambda d: d.accept())
        login(pg, user, password)
        return b, c, pg

    with sync_playwright() as p:
        browser, ctx, page = _abrir_sesion(p)
        relogins = 0
        MAX_RELOGINS = 5
        try:
            for i, (factura_corta, objeciones) in enumerate(facturas.items(), start=1):
                factura_larga = objeciones[0]["factura"]
                if args.sin_soportes:
                    archivos, avisos = [], []
                else:
                    archivos, avisos = buscar_soportes(
                        factura_corta, factura_larga, args.soportes_glosa, indice,
                        carpeta_share_override=args.carpeta_share,
                    )
                    for a in avisos:
                        logger.warning(f"  ⚠ {a}")
                extra = " (sin soportes)" if args.sin_soportes else f", {len(archivos)} archivos soporte"
                logger.info(f"[{i}/{len(facturas)}] {factura_larga} — {len(objeciones)} objeciones{extra}")
                if not args.sin_soportes and not archivos:
                    registrar(
                        {"factura": factura_larga, "objeciones": len(objeciones),
                         "estado": "SIN_SOPORTES", "detalle": "; ".join(avisos)}
                    )
                    continue

                # Hasta 2 intentos por factura: si la sesión muere a mitad de
                # camino, reabrimos browser + re-logueamos y reintentamos la
                # MISMA factura una vez antes de pasar a la siguiente.
                for sesion_intento in range(2):
                    reg = procesar_factura(
                        page, factura_corta, factura_larga, objeciones, archivos,
                        rehacer=args.rehacer, max_obj=args.max_obj,
                        subir_soportes=not args.sin_soportes,
                    )
                    if reg["estado"] == "ERROR" and _sesion_muerta(Exception(reg["detalle"])):
                        if relogins >= MAX_RELOGINS:
                            logger.error(
                                f"  ✗ Sesión perdida y ya hice {MAX_RELOGINS} re-logins; "
                                "aborto el masivo. Re-corré el comando para continuar."
                            )
                            registrar(reg)
                            raise RuntimeError("Sesión irrecuperable")
                        relogins += 1
                        logger.warning(
                            f"  ⚠ Sesión cerrada por el portal. Re-login {relogins}/{MAX_RELOGINS} "
                            f"y reintento {factura_larga}…"
                        )
                        try:
                            browser.close()
                        except Exception:
                            pass
                        browser, ctx, page = _abrir_sesion(p)
                        continue  # reintenta la misma factura
                    registrar(reg)
                    break  # éxito o fallo no-sesión: a la siguiente factura
        except RuntimeError:
            pass  # sesión irrecuperable: salida ordenada
        finally:
            try:
                browser.close()
            except Exception:
                pass
            f_rep.close()  # flush final: lo ya procesado queda en disco

    dur = (time.time() - t0) / 60
    logger.info(f"\nReporte: {args.reporte}")
    logger.info(f"Facturas procesadas: {len(resultados)} en {dur:.1f} min")
    from collections import Counter
    c = Counter(r["estado"] for r in resultados)
    for estado, n in c.items():
        logger.info(f"  {estado}: {n}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
