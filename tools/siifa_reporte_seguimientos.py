#!/usr/bin/env python3
"""siifa_reporte_seguimientos.py — Informe masivo de todos los seguimientos
(glosas y devoluciones) del HUS registrados en SIIFA.

El portal SIIFA (Seguimiento → Listar seguimientos) pagina de a 25 registros
y no tiene botón de exportar. Este script llama a la API oficial
(GET /api/SeguimientoFactura/List, ver docs/CONTEXTO_SIIFA.md) y trae TODO
paginando solo, sin importar si son 2.579 o 20.000, y arma un Excel con una
fila por seguimiento más una hoja resumen.

SI ALGO FALLA: correr primero con --diagnostico. Revisa, paso a paso, si hay
conexión, si el usuario y la contraseña sirven, y si la consulta responde.

CREDENCIALES (variables de entorno, NUNCA en el código):
    setx SIIFA_USER <usuario_sispro>
    setx SIIFA_PASSWORD <password>

USO:
    REM Revisar que todo esté en orden (no baja nada)
    py siifa_reporte_seguimientos.py --diagnostico

    REM Todo lo que haya (glosas + devoluciones)
    py siifa_reporte_seguimientos.py --salida "D:\\...\\SIIFA\\informe_seguimientos.xlsx"

    REM Solo glosas sin responder (para priorizar el trabajo del día)
    py siifa_reporte_seguimientos.py --tipo GLOSA --sin-respuesta ^
        --salida "D:\\...\\SIIFA\\glosas_pendientes.xlsx"

    REM Solo una factura puntual
    py siifa_reporte_seguimientos.py --factura HUS532426 --salida "D:\\...\\HUS532426.xlsx"

INSTALACIÓN (una vez):
    py -m pip install httpx openpyxl
"""

from __future__ import annotations

import argparse
import logging
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from siifa_client import (  # noqa: E402
    REGISTROS_POR_PAGINA_DEFECTO,
    SiifaApiError,
    SiifaClient,
    buscar_clave,
    credenciales_desde_env,
)

logger = logging.getLogger("siifa_reporte")

COLUMNAS = [
    "id_seguimiento_factura_glosa",
    "tipo_seguimiento",
    "numero_factura",
    "nit_emisor",
    "razon_social_emisor",
    "nit_eps",
    "razon_social_eps",
    "valor_bruto_factura",
    "valor_glosa",
    "codigo_glosa",
    "descripcion_glosa",
    "observacion_glosa",
    "fecha_formulacion",
    "fecha_reporte",
    "tiene_respuesta",
    "codigo_respuesta",
    "descripcion_respuesta",
    "observacion_respuesta",
    "fecha_respuesta",
    "codigo_reiteracion",
    "descripcion_reiteracion",
    "codigo_reiteracion_respuesta",
    "descripcion_reiteracion_respuesta",
    "anexo",
]


def _fila_reporte(fila: dict) -> dict:
    """Arma la fila del Excel. Usa buscar_clave para que dé igual si la API
    devuelve los campos en camelCase o en PascalCase."""
    factura = buscar_clave(fila, "factura", "Factura", default={}) or {}
    # OJO: el EMISOR de la factura es el HUS (siempre el mismo). La EPS que
    # glosa es el ADQUIRIENTE — es la que sirve para agrupar y priorizar.
    emisor = buscar_clave(factura, "emisor", "Emisor", default={}) or {}
    adquiriente = buscar_clave(factura, "adquiriente", "Adquiriente", default={}) or {}
    tiene_respuesta = bool(buscar_clave(fila, "idSeguimientoTipoCodigoRespuesta"))
    return {
        "id_seguimiento_factura_glosa": buscar_clave(
            fila, "idSeguimientoFactura", "idSeguimientoFacturaGlosa", default=""
        ),
        "tipo_seguimiento": buscar_clave(fila, "tipoSeguimiento", default=""),
        "numero_factura": buscar_clave(factura, "numeroFactura", default=""),
        "nit_emisor": buscar_clave(emisor, "nitEmisor", default=""),
        "razon_social_emisor": buscar_clave(emisor, "razonSocial", default=""),
        "nit_eps": buscar_clave(adquiriente, "nitAdquiriente", default=""),
        "razon_social_eps": buscar_clave(adquiriente, "razonSocial", default=""),
        "valor_bruto_factura": buscar_clave(factura, "valorBruto"),
        "valor_glosa": buscar_clave(fila, "valor", "valorGlosa"),
        "codigo_glosa": buscar_clave(
            fila, "idSeguimientoTipoCodigo", "idSeguimientoTipoCodigoGlosa", default=""
        ),
        "descripcion_glosa": buscar_clave(
            fila,
            "descripcionSeguimientoTipoCodigo",
            "descripcionSeguimientoTipoCodigoGlosa",
            default="",
        ),
        "observacion_glosa": buscar_clave(fila, "observacion", default=""),
        "fecha_formulacion": buscar_clave(fila, "fechaFormulacion", default=""),
        "fecha_reporte": buscar_clave(fila, "fechaReporte", default=""),
        "tiene_respuesta": "SI" if tiene_respuesta else "NO",
        "codigo_respuesta": buscar_clave(fila, "idSeguimientoTipoCodigoRespuesta", default=""),
        "descripcion_respuesta": buscar_clave(
            fila, "descripcionSeguimientoTipoCodigoRespuesta", default=""
        ),
        "observacion_respuesta": buscar_clave(fila, "observacionRespuesta", default=""),
        "fecha_respuesta": buscar_clave(fila, "fechaRespuesta", default=""),
        "codigo_reiteracion": buscar_clave(
            fila,
            "idSeguimientoTipoCodigoReiteracion",
            "idSeguimientoTipoCodigoGlosaReiteracion",
            default="",
        ),
        "descripcion_reiteracion": buscar_clave(
            fila,
            "descripcionSeguimientoTipoCodigoReiteracion",
            "descripcionSeguimientoTipoCodigoGlosaReiteracion",
            default="",
        ),
        "codigo_reiteracion_respuesta": buscar_clave(
            fila,
            "idSeguimientoTipoCodigoReiteracionRespuesta",
            "idSeguimientoTipoCodigoGlosaReiteracionRespuesta",
            default="",
        ),
        "descripcion_reiteracion_respuesta": buscar_clave(
            fila,
            "descripcionSeguimientoTipoCodigoReiteracionRespuesta",
            "descripcionSeguimientoTipoCodigoGlosaReiteracionRespuesta",
            default="",
        ),
        "anexo": buscar_clave(fila, "anexo", default=""),
    }


def _resumen(filas: list[dict]) -> list[tuple]:
    agg: dict[tuple, dict] = defaultdict(lambda: {"cant": 0, "valor": 0.0, "sin_respuesta": 0})
    for f in filas:
        # Agrupado por EPS (adquiriente): es lo que el auditor necesita para
        # saber a quién reclamarle y por cuánto.
        clave = (
            f["razon_social_eps"] or f["nit_eps"] or "SIN EPS IDENTIFICADA",
            f["tipo_seguimiento"] or "?",
        )
        agg[clave]["cant"] += 1
        try:
            agg[clave]["valor"] += float(f["valor_glosa"] or 0)
        except (TypeError, ValueError):
            pass
        if f["tiene_respuesta"] == "NO":
            agg[clave]["sin_respuesta"] += 1
    return [
        (emisor, tipo, d["cant"], d["sin_respuesta"], d["valor"])
        for (emisor, tipo), d in sorted(agg.items(), key=lambda kv: -kv[1]["valor"])
    ]


def escribir_xlsx(filas: list[dict], ruta: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise SystemExit("ERROR: falta openpyxl. py -m pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "SEGUIMIENTOS"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(COLUMNAS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font

    sin_respuesta_fill = PatternFill("solid", fgColor="FFC7CE")
    i_tiene_respuesta = COLUMNAS.index("tiene_respuesta")
    for r, fila in enumerate(filas, start=2):
        for c, h in enumerate(COLUMNAS, 1):
            cell = ws.cell(row=r, column=c, value=fila.get(h))
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=h in ("observacion_glosa", "observacion_respuesta", "descripcion_glosa"),
            )
        if fila["tiene_respuesta"] == "NO":
            ws.cell(row=r, column=i_tiene_respuesta + 1).fill = sin_respuesta_fill

    anchos = {
        "razon_social_emisor": 30,
        "razon_social_eps": 30,
        "descripcion_glosa": 30,
        "observacion_glosa": 45,
        "descripcion_respuesta": 30,
        "observacion_respuesta": 45,
        "numero_factura": 14,
    }
    for c, h in enumerate(COLUMNAS, 1):
        ws.column_dimensions[get_column_letter(c)].width = anchos.get(h, 16)
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("RESUMEN")
    encabezado = ["EPS (adquiriente)", "Tipo", "Cantidad", "Sin respuesta", "Valor total glosado"]
    ws2.append(encabezado)
    for c in range(1, len(encabezado) + 1):
        ws2.cell(row=1, column=c).fill = header_fill
        ws2.cell(row=1, column=c).font = header_font
    for emisor, tipo, cant, sin_resp, valor in _resumen(filas):
        ws2.append([emisor, tipo, cant, sin_resp, round(valor, 0)])
    for col, ancho in (("A", 34), ("B", 12), ("C", 10), ("D", 14), ("E", 20)):
        ws2.column_dimensions[col].width = ancho

    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ruta))
    logger.info("Informe guardado: %s (%d filas)", ruta, len(filas))


def _fecha(texto: str | None) -> date | None:
    if not texto:
        return None
    try:
        return datetime.strptime(texto.strip()[:10], "%Y-%m-%d").date()
    except ValueError:
        raise SystemExit(f"Fecha inválida: '{texto}'. Se escribe así: 2026-07-01") from None


def _meses(ini: date, fin: date) -> list[tuple[date, date]]:
    """Parte el rango en tramos de un mes calendario."""
    tramos = []
    actual = ini.replace(day=1)
    while actual <= fin:
        if actual.month == 12:
            siguiente = actual.replace(year=actual.year + 1, month=1)
        else:
            siguiente = actual.replace(month=actual.month + 1)
        tramos.append((max(actual, ini), min(siguiente - timedelta(days=1), fin)))
        actual = siguiente
    return tramos


def _bajar_rango(
    cliente: SiifaClient,
    filtros: dict,
    ini: date,
    fin: date,
    filas: list[dict],
    fallidos: list[str],
    profundidad: int = 0,
) -> None:
    """Baja un período. Si el servidor no aguanta, lo parte en dos y reintenta.

    Lo bajado se acumula aparte y sólo se suma al informe si el período salió
    completo: así, al partir y reintentar, ningún registro queda duplicado.
    """
    etiqueta = f"{ini} a {fin}"
    parcial: list[dict] = []
    try:
        for cruda in cliente.listar_seguimientos(
            fecha_creacion_inicio=ini.isoformat(),
            fecha_creacion_final=fin.isoformat(),
            **filtros,
        ):
            parcial.append(_fila_reporte(cruda))
    except SiifaApiError as exc:
        # Igual que arriba: quedarse sin respuesta (status 0) significa lo mismo
        # que un 500 — el tramo pedido es demasiado pesado y hay que partirlo.
        pesado = exc.status_code >= 500 or exc.status_code == 0
        if pesado and profundidad < 4 and (fin - ini).days >= 1:
            medio = ini + (fin - ini) / 2
            logger.warning("El período %s es muy pesado para SIIFA; lo parto en dos.", etiqueta)
            _bajar_rango(cliente, filtros, ini, medio, filas, fallidos, profundidad + 1)
            _bajar_rango(
                cliente,
                filtros,
                medio + timedelta(days=1),
                fin,
                filas,
                fallidos,
                profundidad + 1,
            )
            return
        logger.error("No se pudo traer el período %s: %s", etiqueta, exc)
        fallidos.append(etiqueta)
        return

    filas.extend(parcial)
    if parcial:
        logger.info(
            "Período %s: %d registros (total acumulado: %d).", etiqueta, len(parcial), len(filas)
        )


def _sin_repetidos(filas: list[dict]) -> list[dict]:
    """Quita repetidos por id de seguimiento, por si algún período se solapó."""
    vistos: set = set()
    unicas = []
    for f in filas:
        clave = f.get("id_seguimiento_factura_glosa")
        if clave in (None, ""):
            unicas.append(f)
            continue
        if clave not in vistos:
            vistos.add(clave)
            unicas.append(f)
    quitadas = len(filas) - len(unicas)
    if quitadas:
        logger.info("Se quitaron %d registro(s) repetido(s).", quitadas)
    return unicas


def ruta_para_informe_incompleto(destino: Path) -> Path:
    """Devuelve un nombre alterno, con «_PARCIAL», para un informe a medias.

    Un informe incompleto NUNCA debe pisar al bueno. Si el archivo pedido ya
    existe —por ejemplo, el informe completo de una corrida anterior— se
    perdería sin aviso. Se guarda al lado, con el nombre marcado, y el
    anterior queda intacto.
    """
    alterna = destino.with_name(f"{destino.stem}_PARCIAL{destino.suffix}")
    for n in range(2, 100):
        if not alterna.exists():
            break
        alterna = destino.with_name(f"{destino.stem}_PARCIAL_{n}{destino.suffix}")
    return alterna


def correr_diagnostico(cliente: SiifaClient) -> int:
    print("\nRevisión del sistema SIIFA — paso a paso\n" + "=" * 46)
    pasos = cliente.diagnostico()
    for nombre, ok, detalle in pasos:
        marca = "OK  " if ok else "FALLA"
        print(f"  [{marca}] {nombre}: {detalle}")
    fallo = [p for p in pasos if not p[1]]
    if fallo:
        print("\nHay que resolver lo marcado como FALLA antes de bajar el informe.\n")
        return 1
    print("\nTodo en orden: ya se puede correr el informe.\n")
    return 0


def verificar_se_puede_guardar(ruta: Path) -> None:
    """Comprueba que el informe se va a poder guardar ANTES de bajarlo.

    Bajar los 2.600 seguimientos toma varios minutos. Descubrir al final que
    la carpeta no existe, que la unidad de red no está conectada o que se
    escribió cualquier cosa en vez de una ruta es perder todo ese trabajo
    —pasó el 03-08-2026— y encima con el servidor del Ministerio ya cargado.
    """
    try:
        ruta.parent.mkdir(parents=True, exist_ok=True)
        prueba = ruta.parent / ".prueba_escritura_hus.tmp"
        prueba.write_text("ok", encoding="utf-8")
        prueba.unlink()
    except OSError as exc:
        raise SystemExit(
            f"\nERROR: el informe no se va a poder guardar en:\n"
            f"  {ruta}\n"
            f"  ({getattr(exc, 'strerror', None) or exc})\n\n"
            "Revisá que sea la ruta de una CARPETA válida y que la unidad esté\n"
            "conectada. No bajo nada hasta que la ruta sirva.\n"
        )


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--salida", help="Ruta del Excel de salida.")
    ap.add_argument(
        "--diagnostico",
        action="store_true",
        help="Revisa conexión, credenciales y consulta. No baja información.",
    )
    ap.add_argument(
        "--tipo", choices=["GLOSA", "DEVOLUCION"], help="Filtrar por tipo (default: ambos)."
    )
    ap.add_argument(
        "--sin-respuesta", action="store_true", help="Solo lo que todavía no respondió el HUS."
    )
    ap.add_argument("--factura", help="Filtrar por número de factura (ej. HUS532426).")
    ap.add_argument("--desde", help="Fecha de creación desde (AAAA-MM-DD).")
    ap.add_argument("--hasta", help="Fecha de creación hasta (AAAA-MM-DD).")
    ap.add_argument(
        "--pagina-tam",
        type=int,
        default=REGISTROS_POR_PAGINA_DEFECTO,
        help=f"Registros por consulta (default: {REGISTROS_POR_PAGINA_DEFECTO}). "
        "Bajarlo si el servidor va lento.",
    )
    ap.add_argument(
        "--por-meses",
        action="store_true",
        help="Baja mes por mes. Cada consulta es mucho más liviana para el servidor "
        "del Ministerio. Se activa solo si la consulta completa no responde.",
    )
    ap.add_argument("--verbose", action="store_true")
    args = ap.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s  %(message)s",
        datefmt="%H:%M:%S",
        stream=sys.stdout,
    )
    if not args.verbose:
        # httpx habla en inglés y con detalle técnico: sólo con --verbose.
        logging.getLogger("httpx").setLevel(logging.WARNING)
        logging.getLogger("httpcore").setLevel(logging.WARNING)

    if not args.diagnostico and not args.salida:
        raise SystemExit("ERROR: falta --salida (o usar --diagnostico para revisar el sistema).")

    if args.salida:
        verificar_se_puede_guardar(Path(args.salida))

    usuario, password = credenciales_desde_env()
    filas: list[dict] = []
    interrumpido = False
    corte_parcial = False

    with SiifaClient() as cliente:
        if args.diagnostico:
            cliente.guardar_credenciales(usuario, password)
            raise SystemExit(correr_diagnostico(cliente))

        try:
            cliente.login(usuario, password)
        except SiifaApiError as exc:
            raise SystemExit(f"\nNo se pudo entrar a SIIFA: {exc}\n")

        filtros = {
            "tipo_seguimiento": args.tipo,
            "tiene_respuesta": False if args.sin_respuesta else None,
            "numero_factura": args.factura,
            "registros_por_pagina": args.pagina_tam,
        }
        periodos_fallidos: list[str] = []

        try:
            if args.por_meses:
                ini = _fecha(args.desde) or date(2025, 1, 1)
                fin = _fecha(args.hasta) or (date.today() + timedelta(days=1))
                # +1 día de margen: si el servidor del Ministerio va con otro
                # horario, un registro de hoy podría quedar fechado 'mañana'
                # y se perdería del informe.
                logger.info(
                    "Bajando por meses, de %s a %s (cada mes es una consulta liviana).", ini, fin
                )
                for mes_ini, mes_fin in _meses(ini, fin):
                    _bajar_rango(cliente, filtros, mes_ini, mes_fin, filas, periodos_fallidos)
            else:
                for cruda in cliente.listar_seguimientos(
                    fecha_creacion_inicio=args.desde,
                    fecha_creacion_final=args.hasta,
                    **filtros,
                ):
                    filas.append(_fila_reporte(cruda))
        except KeyboardInterrupt:
            interrumpido = True
            logger.warning("Cancelado por el usuario — guardo lo que alcancé a bajar.")
        except SiifaApiError as exc:
            # El servidor del Ministerio no aguanta la consulta completa.
            # En vez de rendirse, se reintenta mes por mes: cada consulta es
            # mucho más liviana y sí responde.
            # status_code 0 = la consulta se quedó sin respuesta (timeout de
            # lectura). Para el caso que nos ocupa significa lo mismo que un 500:
            # el servidor no puede con la consulta completa. Si sólo se
            # contemplara el 500, una corrida donde los tres intentos se agotan
            # por tiempo terminaría sin informe en vez de pasar a meses.
            if (exc.status_code >= 500 or exc.status_code == 0) and not args.por_meses:
                if filas:
                    # Haber alcanzado a bajar un pedazo NO es razón para
                    # rendirse: un informe con un pedazo no le sirve a nadie.
                    # Se sigue mes por mes y al final se quitan los repetidos
                    # por id de seguimiento, así que lo ya bajado no estorba.
                    logger.warning(
                        "La consulta completa se cortó con %d registros bajados. "
                        "Sigo mes por mes para completar el informe...",
                        len(filas),
                    )
                else:
                    logger.warning(
                        "SIIFA no aguanta la consulta completa. Lo intento mes por mes, "
                        "que es mucho más liviano para su servidor..."
                    )
                ini = _fecha(args.desde) or date(2025, 1, 1)
                fin = _fecha(args.hasta) or (date.today() + timedelta(days=1))
                try:
                    for mes_ini, mes_fin in _meses(ini, fin):
                        _bajar_rango(cliente, filtros, mes_ini, mes_fin, filas, periodos_fallidos)
                except KeyboardInterrupt:
                    interrumpido = True
                    logger.warning("Cancelado por el usuario — guardo lo bajado.")
            elif filas:
                # Guardar lo bajado igual: mejor un informe parcial que nada.
                # Pero queda marcado como incompleto, para que se guarde con
                # otro nombre y no pise un informe bueno anterior.
                corte_parcial = True
                logger.error(
                    "Se cortó la consulta (%s). Guardo las %d filas bajadas.", exc, len(filas)
                )
            else:
                raise SystemExit(f"\nNo se pudo consultar SIIFA: {exc}\n")

    filas = _sin_repetidos(filas)
    if periodos_fallidos:
        logger.warning(
            "OJO: no se pudo traer %d período(s): %s. "
            "El informe está incompleto en esas fechas; volvé a intentarlo más tarde.",
            len(periodos_fallidos),
            ", ".join(periodos_fallidos),
        )

    if not filas:
        logger.warning(
            "SIIFA no devolvió ningún registro con esos filtros. "
            "Probá sin filtros, o corré con --diagnostico para verificar el acceso."
        )
        return

    pedido = Path(args.salida)
    incompleto = interrumpido or corte_parcial or bool(periodos_fallidos)
    destino = ruta_para_informe_incompleto(pedido) if incompleto else pedido
    if incompleto and destino != pedido:
        logger.warning(
            "El informe quedó INCOMPLETO. Lo guardo aparte, como «%s», y NO toco «%s»: "
            "así un informe a medias nunca pisa uno bueno.",
            destino.name,
            pedido.name,
        )

    escribir_xlsx(filas, destino)
    sin_respuesta = sum(1 for f in filas if f["tiene_respuesta"] == "NO")
    logger.info(
        "Listo: %d seguimientos (%d sin respuesta del HUS).%s",
        len(filas),
        sin_respuesta,
        " ATENCIÓN: informe PARCIAL — está incompleto, hay que repetirlo." if incompleto else "",
    )


if __name__ == "__main__":
    main()
