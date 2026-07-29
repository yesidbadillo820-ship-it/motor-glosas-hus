"""excel_a_csv.py — El reverso de TXT_A_EXCEL: por cada Excel deja un
NOMBRE.csv delimitado por comas, listo para las plataformas que exigen ese
formato. Usado por EXCEL_A_CSV.cmd.

Sirve para una carpeta o masivo con subcarpetas. El Excel original nunca se
toca. Reglas:

    - Convierte la hoja ACTIVA; con --todas convierte cada hoja a
      NOMBRE_HOJA.csv. Si el libro tiene fórmulas, se usa el último valor
      calculado por Excel.
    - Fechas salen como dd/mm/aaaa (y hora si la trae); los números sin
      ".0" sobrando; comillas solo si el dato trae coma o salto de línea.
    - En ANSI (cp1252), que es lo que esperan las plataformas; si un dato
      no cabe en ANSI, ese archivo sale en UTF-8 y se avisa.
    - Si el NOMBRE.csv ya existe se OMITE (no pisa nada); con --forzar se
      regenera.

USO:
    py tools\\excel_a_csv.py CARPETA
    py tools\\excel_a_csv.py CARPETA --sin-recursion --todas --forzar

Requiere openpyxl. El repo ya lo fija.
"""

from __future__ import annotations

import argparse
import contextlib
import re
import sys
from datetime import datetime, time
from pathlib import Path

_RE_NOMBRE_HOJA = re.compile(r"[^\w\-]+")


def celda_a_texto(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        if v.hour or v.minute or v.second:
            return v.strftime("%d/%m/%Y %H:%M")
        return v.strftime("%d/%m/%Y")
    if isinstance(v, time):
        return v.strftime("%H:%M")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, bool):
        return "SI" if v else "NO"
    return str(v)


def _campo_csv(campo: str) -> str:
    if "," in campo or "\n" in campo or "\r" in campo:
        return '"' + campo.replace('"', '""') + '"'
    return campo


def hoja_a_csv(ws, destino: Path) -> str:
    """Escribe una hoja como csv por comas. Devuelve nota si salió en UTF-8."""
    lineas = []
    for fila in ws.iter_rows(values_only=True):
        campos = [_campo_csv(celda_a_texto(v)) for v in fila]
        lineas.append(",".join(campos))
    # recorta filas vacías del final
    while lineas and not lineas[-1].strip(","):
        lineas.pop()
    contenido = "\r\n".join(lineas) + "\r\n"
    nota = ""
    try:
        crudo = contenido.encode("cp1252")
    except UnicodeEncodeError:
        crudo = contenido.encode("utf-8-sig")
        nota = "en UTF-8 (trae caracteres fuera de ANSI)"
    tmp = destino.with_name(destino.name + ".tmp")
    try:
        tmp.write_bytes(crudo)
        tmp.replace(destino)
    except Exception:
        with contextlib.suppress(OSError):
            tmp.unlink()
        raise
    return nota


def convertir_uno(ruta: Path, todas: bool = False, forzar: bool = False) -> dict:
    res = {"archivo": ruta.name, "estado": "ok", "motivo": "", "csvs": 0}
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(ruta), read_only=True, data_only=True)
    except Exception as exc:
        res.update(estado="error", motivo=f"Excel ilegible ({type(exc).__name__})")
        return res
    try:
        hojas = wb.worksheets if todas else [wb.active]
        notas = []
        for ws in hojas:
            if todas and len(wb.worksheets) > 1:
                sufijo = _RE_NOMBRE_HOJA.sub("_", ws.title).strip("_") or "Hoja"
                destino = ruta.with_name(f"{ruta.stem}_{sufijo}.csv")
            else:
                destino = ruta.with_suffix(".csv")
            if destino.exists() and not forzar:
                notas.append(f"{destino.name} ya existe (usa --forzar para regenerar)")
                continue
            nota = hoja_a_csv(ws, destino)
            res["csvs"] += 1
            if nota:
                notas.append(nota)
        if not todas and len(wb.worksheets) > 1:
            notas.append(f"el libro tiene {len(wb.worksheets)} hojas; se convirtio la activa")
        res["motivo"] = " · ".join(notas)
        if res["csvs"] == 0 and notas:
            res["estado"] = "omitido"
    except Exception as exc:
        res.update(estado="error", motivo=f"no se pudo escribir ({type(exc).__name__})")
    finally:
        wb.close()
    return res


def procesar(
    raiz: Path, sin_recursion: bool = False, todas: bool = False, forzar: bool = False
) -> int:
    buscador = raiz.glob if sin_recursion else raiz.rglob
    archivos = []
    for patron in ("*.xlsx", "*.xlsm"):
        archivos += [a for a in buscador(patron) if a.is_file() and not a.name.startswith("~$")]
    archivos = sorted(set(archivos), key=lambda p: str(p).lower())

    print("=" * 66)
    print("  EXCEL A CSV — deja cada Excel delimitado por comas")
    print("=" * 66)
    print(
        f"  Carpeta: {raiz}" + ("  (sin subcarpetas)" if sin_recursion else "  (con subcarpetas)")
    )
    print(f"  Excel encontrados: {len(archivos)}")
    print("-" * 66)
    if not archivos:
        print("  No hay Excel para convertir aqui.")
        print("=" * 66)
        return 0

    ok = omitidos = errores = 0
    for a in archivos:
        r = convertir_uno(a, todas=todas, forzar=forzar)
        rel = a.relative_to(raiz) if a.is_relative_to(raiz) else a
        extra = f"  [{r['motivo']}]" if r["motivo"] else ""
        if r["estado"] == "ok":
            ok += 1
            print(f"  ✓ {rel}  ->  {r['csvs']} csv{extra}")
        elif r["estado"] == "omitido":
            omitidos += 1
            print(f"  · {rel}  omitido{extra}")
        else:
            errores += 1
            print(f"  ✗ {rel}  ERROR: {r['motivo']}")

    print("-" * 66)
    print(f"  Convertidos: {ok}   Omitidos: {omitidos}   Errores: {errores}")
    print("=" * 66)
    return 0 if errores == 0 else 1


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Convierte cada Excel a .csv delimitado por comas."
    )
    parser.add_argument("carpeta", nargs="?", default=".", help="Carpeta con los Excel.")
    parser.add_argument("--sin-recursion", action="store_true", help="Solo la carpeta raíz.")
    parser.add_argument(
        "--todas", action="store_true", help="Convierte todas las hojas, no solo la activa."
    )
    parser.add_argument("--forzar", action="store_true", help="Regenera aunque el .csv exista.")
    args = parser.parse_args(argv)

    raiz = Path(args.carpeta).expanduser()
    if not raiz.is_dir():
        sys.stderr.write(f"ERROR: no existe la carpeta: {raiz}\n")
        return 2
    return procesar(raiz.resolve(), args.sin_recursion, args.todas, args.forzar)


if __name__ == "__main__":
    raise SystemExit(main())
