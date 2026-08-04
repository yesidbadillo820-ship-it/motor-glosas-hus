#!/usr/bin/env python3
"""Explorador de Radicación — buscador interactivo de facturas de la ESE HUS.

Toma el reporte del radicador (tools/radicar_facturacion.py) y genera un HTML
INTERACTIVO para que los radicadores encuentren al instante:

  - cuáles facturas están LISTAS y cuáles tienen problemas,
  - QUÉ le falta a cada una (RIPS, CUV, FEV, soportes, tipificación…),
  - DÓNDE está cada una (la carpeta del share), con botón para copiar la ruta.

Se filtra por estado con un clic (chips con conteo), se busca por número o
entidad, se abre el detalle de una factura y se exporta la lista filtrada.
Autocontenido: un solo .html, se abre con doble clic y funciona sin internet.

Seguro: solo LEE el reporte y ESCRIBE el HTML; no toca las carpetas de soportes.
"""

from __future__ import annotations

import argparse
import csv
import logging
import re
import sys
from pathlib import Path

# Lector de pesos compartido (ver `tools/_dinero.py`).
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _dinero import a_numero as a_numero_compartido  # noqa: E402

logger = logging.getLogger("explorador_radicacion")

# El radicador escribe en el detalle de ENTIDAD_NO_RESUELTA: "…(pista: 'XXX')…".
# De ahí sacamos el nombre del pagador que no está en el catálogo.
_RE_PISTA = re.compile(r"pista:\s*'([^']*)'")


def _pagador_no_resuelto(detalle: str) -> str:
    """Nombre del pagador leído del FEV que no resolvió contra el catálogo."""
    m = _RE_PISTA.search(detalle or "")
    nombre = m.group(1).strip() if m else ""
    return "" if nombre in ("", "—") else nombre


# Estado del radicador → (etiqueta legible, color). Orden = prioridad de gestión.
ESTADOS_INFO: dict[str, tuple[str, str]] = {
    "SIN_RIPS": ("Sin RIPS", "#b91c1c"),
    "SIN_FEV": ("Sin FEV", "#9f1239"),
    "SIN_CUV": ("Sin CUV", "#dc2626"),
    "FALTAN_SOPORTES": ("Faltan soportes", "#e11d48"),
    "FACTURA_INCONSISTENTE": ("Factura inconsistente", "#c2410c"),
    "ENTIDAD_NO_RESUELTA": ("Entidad no resuelta", "#ea580c"),
    "REVISAR_TIPIFICACION": ("Revisar tipificación", "#d97706"),
    "PARTICULAR": ("Particulares", "#64748b"),
    "LISTA": ("Listas", "#16a34a"),
}

# Columnas del reporte del radicador → nombre interno. Tolerante a mayúsculas.
_CAMPOS = {
    "factura": "factura",
    "entidad": "entidad",
    "canal": "canal",
    "valor_total": "valor",
    "servicios": "servicios",
    "soportes_presentes": "soportes_presentes",
    "soportes_faltantes": "soportes_faltantes",
    "soportes_esperados_faltantes": "soportes_esperados_faltantes",
    "archivos_sin_clasificar": "archivos_sin_clasificar",
    "soportes_clinicos": "soportes_clinicos",
    "estado": "estado",
    "detalle": "detalle",
    "ruta_paquete": "ruta_paquete",
    "carpeta": "carpeta",
}


def _leer_tabla(ruta: Path) -> list[list]:
    """Filas (incluido encabezado) de un CSV o XLSX."""
    if ruta.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as e:  # pragma: no cover
            raise SystemExit("ERROR: para leer .xlsx falta openpyxl.") from e
        wb = load_workbook(filename=str(ruta), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        return [list(r) for r in ws.iter_rows(values_only=True)]
    with ruta.open("r", encoding="utf-8-sig", newline="") as fh:
        return list(csv.reader(fh))


def _parse_valor(v) -> float:
    """Valor monetario → float, con el lector único de `tools/_dinero.py`.

    Traía la misma condición incompleta que el acta y la hoja maestra
    —`s.count(".") > 1`—, así que `"$ 950.000"` se leía como **950,00**:
    dividido por mil. Novena copia de la misma regla en `tools/`.
    """
    return a_numero_compartido(v)


def _falta(estado: str, r: dict) -> str:
    """Frase corta de QUÉ le falta a la factura, para la columna y el filtro."""
    if estado == "LISTA":
        return ""
    if estado == "SIN_RIPS":
        return "Falta el RIPS (.json)"
    if estado == "SIN_FEV":
        return "Falta la factura electrónica (FEV)"
    if estado == "SIN_CUV":
        return "Falta el CUV (validación RIPS)"
    if estado == "FALTAN_SOPORTES":
        f = r.get("soportes_faltantes", "")
        return "Faltan soportes obligatorios: " + f if f else "Faltan soportes obligatorios"
    if estado == "REVISAR_TIPIFICACION":
        if r.get("archivos_sin_clasificar"):
            return "Archivos sin tipificar: " + r["archivos_sin_clasificar"]
        if r.get("soportes_esperados_faltantes"):
            return "Faltan soportes clínicos: " + r["soportes_esperados_faltantes"]
        return "Revisar tipificación"
    if estado == "ENTIDAD_NO_RESUELTA":
        pag = _pagador_no_resuelto(r.get("detalle", ""))
        return (
            f"Pagador sin perfil en el catálogo: {pag}"
            if pag
            else "No se identificó la EPS pagadora"
        )
    if estado == "PARTICULAR":
        return "Paciente particular (no es EPS)"
    if estado == "FACTURA_INCONSISTENTE":
        return "El número no coincide entre RIPS y FEV"
    return r.get("detalle", "")


def cargar_reporte(ruta: Path) -> list[dict]:
    """Lee el reporte del radicador y devuelve una lista de registros."""
    filas = _leer_tabla(ruta)
    if not filas:
        return []
    headers = [str(h or "").strip().lower() for h in filas[0]]
    idx = {campo: headers.index(h) for h, campo in _CAMPOS.items() if h in headers}
    if "factura" not in idx or "estado" not in idx:
        raise ValueError(
            f"{ruta.name}: no parece un reporte del radicador (faltan columnas factura/estado). "
            f"Encabezados: {filas[0]}"
        )

    def cel(fila, campo):
        i = idx.get(campo)
        return (
            str(fila[i]).strip() if i is not None and i < len(fila) and fila[i] is not None else ""
        )

    registros: list[dict] = []
    for fila in filas[1:]:
        if not cel(fila, "factura"):
            continue
        r = {campo: cel(fila, campo) for campo in idx}
        estado = r.get("estado", "") or "—"
        entidad = r.get("entidad", "") or "—"
        # Para las no resueltas, mostrar el PAGADOR detectado (no "SIN_PERFIL"):
        # así se ve qué entidad falta agregar al catálogo y se puede exportar.
        if estado == "ENTIDAD_NO_RESUELTA":
            pag = _pagador_no_resuelto(r.get("detalle", ""))
            if pag:
                entidad = pag
        registros.append(
            {
                "factura": r.get("factura", ""),
                "entidad": entidad,
                "canal": r.get("canal", ""),
                "valor": _parse_valor(r.get("valor", 0)),
                "servicios": r.get("servicios", ""),
                "estado": estado,
                "falta": _falta(estado, r),
                "soportes_presentes": r.get("soportes_presentes", ""),
                "soportes_faltantes": r.get("soportes_faltantes", ""),
                "soportes_esperados_faltantes": r.get("soportes_esperados_faltantes", ""),
                "soportes_clinicos": r.get("soportes_clinicos", ""),
                "archivos_sin_clasificar": r.get("archivos_sin_clasificar", ""),
                "detalle": r.get("detalle", ""),
                "carpeta": r.get("carpeta", ""),
                "ruta_paquete": r.get("ruta_paquete", ""),
            }
        )
    logger.info(f"Reporte: {len(registros)} factura(s) leídas de {ruta.name}")
    return registros


def agregar(registros: list[dict]) -> dict:
    """Conteos por estado (para los chips) y por entidad."""
    por_estado: dict[str, int] = {}
    por_entidad: dict[str, int] = {}
    for r in registros:
        por_estado[r["estado"]] = por_estado.get(r["estado"], 0) + 1
        por_entidad[r["entidad"]] = por_entidad.get(r["entidad"], 0) + 1
    # Ordena los estados por prioridad de gestión (problemas primero, LISTA al final).
    orden = list(ESTADOS_INFO)
    estados = sorted(
        por_estado.items(),
        key=lambda kv: orden.index(kv[0]) if kv[0] in orden else 99,
    )
    return {
        "total": len(registros),
        "estados": [
            {
                "estado": e,
                "n": n,
                "label": ESTADOS_INFO.get(e, (e, "#64748b"))[0],
                "color": ESTADOS_INFO.get(e, (e, "#64748b"))[1],
            }
            for e, n in estados
        ],
        "entidades": sorted(por_entidad),
        "facturas": registros,
    }


# ─── CLI ─────────────────────────────────────────────────────────────────────


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="explorador_radicacion.py",
        description="Explorador interactivo de las facturas del radicador (HTML).",
    )
    p.add_argument(
        "--reporte", type=Path, required=True, help="Reporte del radicador (CSV o XLSX)."
    )
    p.add_argument(
        "--salida",
        type=Path,
        default=Path("explorador_radicacion.html"),
        help="HTML de salida (def.: explorador_radicacion.html).",
    )
    p.add_argument(
        "--titulo",
        type=str,
        default="Explorador de Radicación — ESE HUS",
        help="Título del explorador.",
    )
    p.add_argument("--log", type=Path, default=None, help="Archivo de log opcional.")
    return p


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    handlers: list[logging.Handler] = [logging.StreamHandler(sys.stdout)]
    if args.log is not None:
        handlers.append(logging.FileHandler(args.log, encoding="utf-8"))
    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", handlers=handlers
    )

    if not args.reporte.is_file():
        logger.error(f"No existe el reporte: {args.reporte}")
        return 1

    registros = cargar_reporte(args.reporte)
    if not registros:
        logger.error("El reporte no tiene facturas.")
        return 1

    datos = agregar(registros)
    datos["titulo"] = args.titulo

    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from _explorador_html import render_html

    args.salida.parent.mkdir(parents=True, exist_ok=True)
    args.salida.write_text(render_html(datos), encoding="utf-8")
    logger.info(f"Explorador HTML: {args.salida}")

    logger.info("=" * 60)
    logger.info(f"  {datos['total']:,} facturas · por estado:")
    for e in datos["estados"]:
        logger.info(f"    {e['estado']:24} {e['n']:>6,}")
    logger.info("=" * 60)
    return 0


if __name__ == "__main__":
    sys.exit(main())
