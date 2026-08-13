"""revisar_xml_facturas.py — Revisa los XML de factura electrónica en salud y
extrae la información de CONTRATO para responder la glosa "factura sin contrato".

Por cada XML (AttachedDocument de la DIAN) entra a la factura embebida (el
Invoice que va dentro del CDATA) y lee los campos de interoperabilidad del
sector salud:

    NUMERO_CONTRATO, NUMERO_POLIZA, FACTURA_SIN_CONTRATO, MODALIDAD_PAGO,
    COBERTURA_PLAN_BENEFICIOS, CODIGO_PRESTADOR

...además del número de factura, fecha, valor, la EPS adquirente y el estado
de validación de la DIAN. Con eso arma un Excel (una fila por factura) que
sirve como anexo de prueba: muestra que el XML SÍ trae número de contrato y/o
la causa de atención (p. ej. ATENCION DE URGENCIAS).

Puede filtrar por una lista de facturas (archivo de texto, una por línea) o
procesar todos los XML de la carpeta.

USO:
    py tools\\revisar_xml_facturas.py "\\\\172.16.32.83\\factura_electronica_net22\\202607\\FACTURAS_SALUD"
    py tools\\revisar_xml_facturas.py CARPETA --lista facturas.txt --salida INFORME.xlsx
    py tools\\revisar_xml_facturas.py .        # carpeta actual, todos los XML

Requiere openpyxl (para el Excel). El repo ya lo fija.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from xml.etree import ElementTree as ET

# Campos de salud que nos interesan (nombre tal cual en el XML).
CAMPOS_SALUD = [
    "NUMERO_CONTRATO",
    "NUMERO_POLIZA",
    "FACTURA_SIN_CONTRATO",
    "MODALIDAD_PAGO",
    "COBERTURA_PLAN_BENEFICIOS",
    "CODIGO_PRESTADOR",
]


def _local(tag: str) -> str:
    """Nombre de la etiqueta sin el namespace ({...}Algo -> Algo)."""
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def _texto(elem, nombre: str) -> str:
    """Primer texto de un descendiente cuyo nombre local sea `nombre`."""
    for e in elem.iter():
        if _local(e.tag) == nombre and e.text and e.text.strip():
            return e.text.strip()
    return ""


def _cdata_con(raiz, marcador: str) -> str | None:
    """Devuelve el texto (CDATA) de la primera Description que contenga
    `marcador` (p. ej. '<Invoice' o 'ApplicationResponse')."""
    for e in raiz.iter():
        if _local(e.tag) == "Description" and e.text and marcador in e.text:
            return e.text
    return None


def _parse_xml(texto: str):
    """Parsea un string XML tolerando el prologo <?xml ...?>."""
    limpio = re.sub(r"^\s*<\?xml[^>]*\?>", "", texto).strip()
    return ET.fromstring(limpio)


def normalizar_factura(valor: str) -> str:
    """HUS0000533470 / HUS533470 / 533470 -> '533470' (para comparar)."""
    digitos = re.sub(r"\D", "", valor or "")
    return digitos.lstrip("0") or digitos


def revisar_uno(ruta: Path) -> dict | None:
    """Extrae los datos de un XML de factura. None si no es una factura válida."""
    try:
        raiz = ET.parse(str(ruta)).getroot()
    except Exception as exc:
        return {"archivo": ruta.name, "error": f"XML ilegible ({type(exc).__name__})"}

    factura = _texto(raiz, "ParentDocumentID")
    eps = _texto(raiz, "RegistrationName")  # ReceiverParty va primero en el AttachedDocument
    validacion = _texto(raiz, "ValidationResultCode")

    datos = {
        "archivo": ruta.name,
        "factura": factura,
        "eps": eps,
        "fecha": "",
        "valor": "",
        "validacion_dian": "02 (validado)" if validacion == "02" else (validacion or ""),
        "error": "",
    }
    for c in CAMPOS_SALUD:
        datos[c] = ""

    # Entrar a la factura embebida (Invoice dentro del CDATA).
    cdata = _cdata_con(raiz, "<Invoice")
    if not cdata:
        datos["error"] = "sin Invoice embebido (¿no es factura?)"
        return datos
    try:
        inv = _parse_xml(cdata)
    except Exception as exc:
        datos["error"] = f"Invoice embebido ilegible ({type(exc).__name__})"
        return datos

    if not datos["factura"]:
        datos["factura"] = _texto(inv, "ID")
    datos["fecha"] = _texto(inv, "IssueDate")
    datos["valor"] = _texto(inv, "PayableAmount")

    # Campos de interoperabilidad en salud (pares Name/Value).
    for ai in inv.iter():
        if _local(ai.tag) != "AdditionalInformation":
            continue
        nombre = ""
        valor = ""
        for hijo in ai:
            if _local(hijo.tag) == "Name":
                nombre = (hijo.text or "").strip()
            elif _local(hijo.tag) == "Value":
                valor = (hijo.text or "").strip()
        if nombre in CAMPOS_SALUD:
            datos[nombre] = valor

    return datos


def conclusion(d: dict) -> str:
    """Frase de argumento para la glosa 'sin contrato'."""
    if d.get("error"):
        return f"REVISAR: {d['error']}"
    partes = []
    if d.get("NUMERO_CONTRATO"):
        partes.append(f"SÍ trae contrato: {d['NUMERO_CONTRATO']}")
    else:
        partes.append("Sin NUMERO_CONTRATO en el XML")
    if d.get("FACTURA_SIN_CONTRATO"):
        partes.append(f"causa de cobertura: {d['FACTURA_SIN_CONTRATO']}")
    if d.get("validacion_dian", "").startswith("02"):
        partes.append("validada por DIAN")
    return " · ".join(partes)


def cargar_lista(ruta: Path) -> set[str]:
    pedidas = set()
    for ln in ruta.read_text(encoding="utf-8-sig").splitlines():
        s = ln.strip()
        if s and not s.startswith("#"):
            pedidas.add(normalizar_factura(s))
    return pedidas


def escribir_excel(filas: list[dict], salida: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        sys.stderr.write("ERROR: falta openpyxl.  py -m pip install openpyxl\n")
        sys.exit(2)

    wb = Workbook()
    ws = wb.active
    ws.title = "Revision XML"
    encabezados = [
        "Factura",
        "Fecha",
        "Valor (XML)",
        "EPS adquirente",
        "NUMERO_CONTRATO",
        "FACTURA_SIN_CONTRATO",
        "MODALIDAD_PAGO",
        "COBERTURA_PLAN_BENEFICIOS",
        "NUMERO_POLIZA",
        "CODIGO_PRESTADOR",
        "Validación DIAN",
        "Conclusión (argumento glosa)",
        "Archivo XML",
    ]
    ws.append(encabezados)
    azul = PatternFill("solid", fgColor="1F4E79")
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = azul
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    verde = PatternFill("solid", fgColor="E2EFDA")
    amarillo = PatternFill("solid", fgColor="FFF2CC")
    for d in filas:
        ws.append(
            [
                d.get("factura", ""),
                d.get("fecha", ""),
                d.get("valor", ""),
                d.get("eps", ""),
                d.get("NUMERO_CONTRATO", ""),
                d.get("FACTURA_SIN_CONTRATO", ""),
                d.get("MODALIDAD_PAGO", ""),
                d.get("COBERTURA_PLAN_BENEFICIOS", ""),
                d.get("NUMERO_POLIZA", ""),
                d.get("CODIGO_PRESTADOR", ""),
                d.get("validacion_dian", ""),
                conclusion(d),
                d.get("archivo", ""),
            ]
        )
        fila = ws[ws.max_row]
        if d.get("error"):
            for c in fila:
                c.fill = amarillo
        elif d.get("NUMERO_CONTRATO") or d.get("FACTURA_SIN_CONTRATO"):
            for c in fila:
                c.fill = verde

    anchos = [16, 12, 14, 34, 22, 24, 16, 28, 14, 16, 16, 46, 26]
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    wb.save(str(salida))


def procesar(raiz: Path, lista: Path | None, salida: Path) -> int:
    xmls = sorted(raiz.rglob("*.xml"), key=lambda p: p.name.lower())
    if not xmls:
        print(f"No se encontraron archivos .xml en {raiz}")
        return 0

    pedidas = cargar_lista(lista) if lista else None

    print("=" * 66)
    print("  REVISION DE XML — informacion de contrato para la glosa")
    print("=" * 66)
    print(f"  Carpeta: {raiz}")
    print(
        f"  XML encontrados: {len(xmls)}"
        + (f"  |  Lista pedida: {len(pedidas)}" if pedidas else "")
    )
    print("-" * 66)

    filas = []
    vistas = set()
    con_contrato = 0
    for x in xmls:
        d = revisar_uno(x)
        if not d:
            continue
        norm = normalizar_factura(d.get("factura", ""))
        if pedidas is not None and norm not in pedidas:
            continue
        filas.append(d)
        if norm:
            vistas.add(norm)
        if d.get("NUMERO_CONTRATO"):
            con_contrato += 1
        marca = "✓" if (d.get("NUMERO_CONTRATO") or d.get("FACTURA_SIN_CONTRATO")) else "·"
        if d.get("error"):
            marca = "✗"
        print(
            f"  {marca} {d.get('factura', '?'):16s}  contrato={d.get('NUMERO_CONTRATO', ''):20s} "
            f"sin_contrato={d.get('FACTURA_SIN_CONTRATO', '')}"
        )

    # ordenar por factura
    filas.sort(key=lambda d: normalizar_factura(d.get("factura", "")))
    escribir_excel(filas, salida)

    print("-" * 66)
    print(f"  Facturas en el informe: {len(filas)}")
    print(f"  Con NUMERO_CONTRATO en el XML: {con_contrato}")
    print(
        f"  Con marca FACTURA_SIN_CONTRATO: {sum(1 for d in filas if d.get('FACTURA_SIN_CONTRATO'))}"
    )
    if pedidas is not None:
        faltan = sorted(pedidas - vistas)
        if faltan:
            print(f"  OJO: {len(faltan)} factura(s) de la lista NO se encontraron como XML:")
            print("       " + ", ".join(faltan[:30]) + (" ..." if len(faltan) > 30 else ""))
    print(f"  Informe Excel: {salida}")
    print("=" * 66)
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Revisa los XML de factura en salud y saca la info de contrato.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("carpeta", nargs="?", default=".", help="Carpeta con los XML.")
    parser.add_argument(
        "--lista", type=Path, help="Archivo .txt con las facturas a filtrar (una por línea)."
    )
    parser.add_argument("--salida", type=Path, help="Ruta del Excel de salida.")
    args = parser.parse_args(argv)

    raiz = Path(args.carpeta).expanduser()
    if not raiz.is_dir():
        sys.stderr.write(f"ERROR: no existe la carpeta: {raiz}\n")
        return 2
    salida = args.salida or (raiz / "INFORME_REVISION_XML.xlsx")
    if args.lista and not args.lista.is_file():
        sys.stderr.write(f"ERROR: no existe la lista: {args.lista}\n")
        return 2

    return procesar(raiz.resolve(), args.lista, salida)


if __name__ == "__main__":
    raise SystemExit(main())
