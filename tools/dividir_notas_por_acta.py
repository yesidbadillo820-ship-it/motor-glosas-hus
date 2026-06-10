"""dividir_notas_por_acta.py — Separa las notas crédito por ACTA (correo vs SIMED).

Lee un Excel de notas (ACTA, FACTURA, NOTA ELECTRONICA, ...) y genera dos CSV
con la columna `NOTA CREDITO` (= NOTA ELECTRONICA) que come
`extraer_notas_credito.py`:

  - <salida>\\notas_correo.csv : las del ACTA que va por correo (AC000456 por
    defecto), una NOTA ELECTRONICA por fila.
  - <salida>\\notas_simed.csv  : las del resto de actas (van por el aplicativo).

Ambos grupos necesitan el mismo armado de carpetas; sólo cambia el destino y el
canal de envío.

USO:
    py dividir_notas_por_acta.py ^
        --excel "D:\\USUARIO CARTERA\\Documents\\NOTAS ANTIGUAS\\NOTAS VIEJAS.xlsx" ^
        --salida "D:\\USUARIO CARTERA\\Documents\\NOTAS ANTIGUAS"

    REM Si el acta de correo es otra, pasala con --acta-correo:
    py dividir_notas_por_acta.py --excel ... --salida ... --acta-correo AC000456
"""

from __future__ import annotations

import argparse
import csv
import logging
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

logger = logging.getLogger("dividir_notas")


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", (s or "").strip().upper())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.split())


# Alias por columna, EN ORDEN DE PREFERENCIA (ya normalizados). Para 'factura'
# preferimos la columna con el HUS completo antes que el número pelado.
ALIAS = {
    "acta": ["ACTA", "NUMERO ACTA", "ACTA CONCILIACION", "N ACTA"],
    "factura": ["FACTURA HUS", "FACTURA", "NUMERO FACTURA", "NUM FACTURA"],
    "nota": ["NOTA ELECTRONICA", "NE", "NOTA CREDITO", "NOTAS CREDITO",
             "NOTA ELECTRONICA NUMERO"],
}


def _idx_columnas(headers: list[str]) -> dict[str, int]:
    norm = [_norm(h) for h in headers]
    idx: dict[str, int] = {}
    for clave, opciones in ALIAS.items():
        for opt in opciones:  # respeta el orden de preferencia
            if opt in norm:
                idx[clave] = norm.index(opt)
                break
        if clave not in idx and clave != "acta":
            raise ValueError(
                f"No encontré la columna '{clave}'. Acepté {opciones}. "
                f"Headers: {headers}"
            )
    return idx


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--excel", type=Path, required=True, help="Excel de notas (ACTA, FACTURA, NOTA ELECTRONICA).")
    parser.add_argument("--salida", type=Path, required=True, help="Carpeta donde escribir los CSV.")
    parser.add_argument("--acta-correo", type=str, default="AC000456", help="ACTA cuyas notas van por correo (default AC000456).")
    args = parser.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.stderr.write("ERROR: falta openpyxl. py -m pip install openpyxl\n")
        return 2

    if not args.excel.is_file():
        logger.error(f"No existe el Excel: {args.excel}")
        return 1

    wb = load_workbook(filename=str(args.excel), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        logger.error("Excel vacío.")
        return 1
    idx = _idx_columnas([str(h or "") for h in rows[0]])
    idx_acta = idx.get("acta")  # puede no existir; las sin ACTA van a SIMED

    acta_correo = _norm(args.acta_correo)
    correo: list[tuple[str, str]] = []   # (nota, factura)
    simed: list[tuple[str, str, str]] = []  # (acta, nota, factura)
    sin_nota: list[str] = []  # facturas sin NOTA ELECTRONICA (no procesables)
    por_acta: dict[str, int] = defaultdict(int)

    for r in rows[1:]:
        if r is None:
            continue
        acta = str(r[idx_acta] or "").strip() if idx_acta is not None else ""
        nota = str(r[idx["nota"]] or "").strip()
        factura = str(r[idx["factura"]] or "").strip()
        # Saltar fila de total/resumen (sin factura y sin nota).
        if not factura and not nota:
            continue
        if not nota:
            sin_nota.append(factura or "(sin factura)")
            continue
        por_acta[acta or "(SIN ACTA)"] += 1
        if acta and _norm(acta) == acta_correo:
            correo.append((nota, factura))
        else:
            # ACTA en blanco u otra acta → van por SIMED.
            simed.append((acta or "SIN_ACTA", nota, factura))

    args.salida.mkdir(parents=True, exist_ok=True)

    # CSV de correo: sólo la columna NOTA CREDITO (lo que come el extractor).
    ruta_correo = args.salida / "notas_correo.csv"
    with ruta_correo.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["NOTA CREDITO", "FACTURA HUS"])
        for nota, factura in correo:
            w.writerow([nota, factura])

    # CSV de SIMED.
    ruta_simed = args.salida / "notas_simed.csv"
    with ruta_simed.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["NOTA CREDITO", "FACTURA HUS", "ACTA"])
        for acta, nota, factura in simed:
            w.writerow([nota, factura, acta])

    # Reporte de las que NO se pueden procesar (sin NOTA ELECTRONICA).
    ruta_sin_ne = args.salida / "notas_sin_nota_electronica.csv"
    if sin_nota:
        with ruta_sin_ne.open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["FACTURA HUS (sin nota electronica)"])
            for fac in sin_nota:
                w.writerow([fac])

    logger.info(f"\nNotas por ACTA: {dict(por_acta)}")
    logger.info(f"\n📧 CORREO (ACTA {args.acta_correo}): {len(correo)} notas → {ruta_correo}")
    logger.info(f"🖥️  SIMED (blanco + otras actas): {len(simed)} notas → {ruta_simed}")
    if sin_nota:
        logger.warning(f"\n⚠ {len(sin_nota)} factura(s) SIN nota electrónica (no procesables hasta tener la NE):")
        for fac in sin_nota:
            logger.warning(f"     {fac}")
        logger.warning(f"   Listado en: {ruta_sin_ne}")
    logger.info("\nAmbos CSV usan la columna 'NOTA CREDITO' que lee extraer_notas_credito.py.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
