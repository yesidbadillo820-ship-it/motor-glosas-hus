"""consolidar_coosalud.py — Consolida GLOSAS/DETALLES/FACTURAS de COOSALUD y arma el archivo de OBJECIONES (DGH).

Automatiza los pasos manuales que siguen después de organizar el ZIP con
`organizar_cargue_masivo_coosalud.py` (puntos 2 a 5 de la guía de cartera):

  1. CONSOLIDADO GLOSAS.xlsx
       - Hoja GLOSAS: todas las filas de GLOSAS HUS*.xlsx + columna
         OBSERVACION FINAL = codigo_glosa + " " + justificacion_glosa + "$" + valor_total_glosa
       - Hoja AGRUPADO: una fila por id_detalle con las observaciones combinadas
         (reemplaza el "Agrupar por" + Text.Combine de Power Query).
  2. CONSOLIDADO DETALLE.xlsx
       - Todas las filas de DETALLE HUS*.xlsx + columna OBSERVACION FINAL
         buscada por id_detalle (reemplaza el BUSCARV).
  3. CONSOLIDADO FACTURAS.xlsx
       - Todas las filas de HUS*.xlsx (cabeceras de factura).
  4. SERVICIOS FACTURADOS COOSALUD.xlsx  (solo si se pasa --servicios)
       - La base DGH filtrada: SOLO las facturas trabajadas y SOLO 10 columnas.
       - "Arreglada": en los medicamentos (que vienen con SLNSERPRO_SERVICIO
         vacío) se rellenan SERVICIO/CUPS y descripciones con el código y
         nombre del medicamento, igual que en el proceso manual.
       - Se lee en streaming: la base puede pesar 80MB+ sin agotar la memoria.
  5. OBJECIONES.xlsx  (formato de cargue DGH, igual al archivo de ejemplo)
       - POR LOTE: como el cargue de DGH solo soporta 300 facturas, si la
         carpeta viene en LOTE 01/LOTE 02/... cada lote produce SU PROPIO juego
         de archivos en CONSOLIDADOS\\LOTE XX\\ ("OBJECIONES LOTE 01.xlsx", etc.)
         con el consecutivo arrancando en 1. Sin lotes, sale todo junto como antes.
       - Una fila por servicio glosado (id_detalle con glosas):
         CDCONSEC    consecutivo por factura, como texto (1,1,1... 2,2,2...)
         CDFECDOC    --fecha como fecha Excel (día de la carpeta que se revisa)
         CRNCXC      factura con ceros: HUS0000496207
         CROFECOBJ   --fecha
         CROREFERE / CROOBSERV / CRNCLAOBJ / IDRIPS / CTNCENCOS   vacíos
         CROCLAOBJ   0
         GENUSUARIO4 999 (texto)
         CRNCONOBJ   código completo de la glosa del servicio. Una glosa CL
                     (médica) tiene PRIORIDAD sobre las demás: si el servicio
                     tiene alguna CL, va la CL (la de mayor valor entre las CL);
                     si no hay CL, va la de mayor valor (TA2901, ...)
         SLNSERPRO   codigo_servicio del DETALLE; si se pasa la base DGH y la
                     fila cruza, manda el código DGH (con sufijo H si lo trae).
                     El OBJECIONES.xlsx queda con UNA sola hoja, ya limpio para
                     el cargue; lo que DGH no acepta (servicio ausente, o valor
                     por encima del servicio/saldo) se deja FUERA y se lista en
                     un archivo aparte "REVISAR (no van en el cargue).xlsx".
         CROVALOBJ   valor_glosado del servicio (número)
         CRDOBSERV   OBSERVACION FINAL combinada (una línea por glosa)
         CROTIPOBJ   por FACTURA, según el CRNCONOBJ escrito (no las glosas
                     crudas): 0=administrativa (ningún CRNCONOBJ es CL) ·
                     1=médica (todos los CRNCONOBJ son CL) · 2=mixta (unos CL y
                     otros no). TODAS las filas de la factura llevan el mismo tipo.
  6. CONSOLIDADO RESPUESTAS GLOSAS.xlsx  (respuestas predeterminadas del área)
       - Una fila por glosa con: numero_factura2 (HUS0000...), FECHA RADICACION
         (de la cabecera de factura), FECHA GLOSA (= --fecha), FECHA DE
         VENCIMIENTO (+15 días hábiles), DIA (días hábiles radicación->glosa,
         festivos de Colombia incluidos), EXTEMPORANEA (>20 días hábiles,
         art. 57 Ley 1438/2011), COD RESPUESTA GLOSA / COD (RE9502 extemporánea
         · RE9901 a tiempo) y OBSERVACION RTA GLOSA (texto del área según tipo:
         TARIFAS, AUTORIZACION, FACTURACION, SOPORTES — editables arriba).
       - Hoja FACTURAS con la lista de facturas del lote.

DEFENSAS (validado con revisión adversarial):
  - Valores en formato colombiano ("1.234.567,89", "26,140") se parsean bien.
  - Archivos ~$ de Excel se ignoran; un xlsx corrupto aborta con nombre claro.
  - Si un lote trae los encabezados en otro orden, las filas se reordenan por nombre.
  - id_detalle/códigos tipados como número (123456.0) cruzan igual.
  - Archivos o filas duplicadas (mismo id_glosa / id_detalle) se saltan con aviso
    para no duplicar el valor objetado.
  - xlsx del portal con "dimension" mentirosa se leen completos (reset_dimensions).
  - Facturas que no siguen el patrón HUS se reportan.

USO
---

    REM Después de organizar el ZIP (los consolidados salen en <carpeta>\\CONSOLIDADOS)
    py consolidar_coosalud.py ^
        --carpeta "D:\\USUARIO CARTERA\\Desktop\\CARGUE MASIVO COOSALUD" ^
        --fecha 04/07/2026

    REM Con la base DGH para el punto 4 y el cruce SLNSERPRO del punto 5
    py consolidar_coosalud.py ^
        --carpeta   "D:\\USUARIO CARTERA\\Desktop\\CARGUE MASIVO COOSALUD" ^
        --fecha     04/07/2026 ^
        --servicios "D:\\...\\SERVICIOS FACTURADOS COOSALUD DGH.xlsx"

INSTALACIÓN (una vez):
    py -m pip install openpyxl
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Iterator

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    print("Falta openpyxl. Instalar con:  py -m pip install openpyxl")
    sys.exit(1)

logger = logging.getLogger("consolidar_coosalud")

# --- Estructura de la carpeta organizada (ver organizar_cargue_masivo_coosalud) ---
CARPETA_GLOSAS = "GLOSAS"
CARPETA_DETALLES = "DETALLES"
CARPETA_FACTURAS = "FACTURAS"
SALIDA_DEFECTO = "CONSOLIDADOS"

# Columnas mínimas que se esperan en los archivos del portal.
COL_GLOSAS_REQ = [
    "id_detalle",
    "numero_factura",
    "codigo_glosa",
    "justificacion_glosa",
    "valor_total_glosa",
]
COL_DETALLE_REQ = ["id_detalle", "numero_factura", "codigo_servicio", "valor_glosado"]

# Columnas que se conservan de la base DGH, ya "arreglada" (punto 4 de la guía).
COLS_SERVICIOS = [
    "FACTURA",
    "SLNSERPRO_SERVICIO",
    "DESCRIPCION INSTITUCIONAL",
    "SLNSERPRO_CUPS",
    "DESCRIPCION CUPS",
    "CODIGO_MEDICAMENTO",
    "NOMBRE_MEDICAMENTO",
    "CAT_SERVICIOS",
    "Vr_SERVICIO",
    "SALDO_FACT",
]
# Columnas de COLS_SERVICIOS que llevan códigos (se normalizan, ej. 890466.0 -> 890466).
COLS_SERVICIOS_CODIGO = {"SLNSERPRO_SERVICIO", "SLNSERPRO_CUPS", "CODIGO_MEDICAMENTO"}

# Encabezados y formatos de celda del archivo de OBJECIONES (tomados de la
# plantilla DGH real: fechas mm-dd-yy, valores con miles, el resto texto).
COLS_OBJECIONES = [
    ("CDCONSEC", "@"),
    ("CDFECDOC", "mm-dd-yy"),
    ("CRNCXC", "@"),
    ("CROFECOBJ", "mm-dd-yy"),
    ("CROREFERE", "@"),
    ("CROOBSERV", "@"),
    ("CROCLAOBJ", "General"),
    ("CRNCLAOBJ", "@"),
    ("GENUSUARIO4", "@"),
    ("CRNCONOBJ", "@"),
    ("SLNSERPRO", "@"),
    ("IDRIPS", "@"),
    ("CTNCENCOS", "@"),
    ("CROVALOBJ", '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-'),
    ("CRDOBSERV", "@"),
    ("CROTIPOBJ", "0"),
]

RE_NUM_FACTURA = re.compile(r"HUS0*(\d+)", re.IGNORECASE)

# ============================================================================
# CONSOLIDADO DE RESPUESTAS A GLOSAS (respuestas predeterminadas del área)
# ----------------------------------------------------------------------------
# La Ley 1438 de 2011 (art. 57) da a la EPS 20 días HÁBILES desde la
# radicación de la factura para formular la glosa. Si la glosa llegó después,
# es EXTEMPORÁNEA y se responde con RE9502; si llegó a tiempo, se responde con
# RE9901 y el texto predeterminado según el tipo de glosa.
# Estos textos son EDITABLES: cambiarlos aquí cambia el archivo de respuestas.

DIAS_HABILES_EPS = 20  # plazo de la EPS para glosar (art. 57, Ley 1438/2011)
DIAS_HABILES_RESPUESTA = 15  # plazo del prestador para responder la glosa

COD_RTA_EXTEMPORANEA = "RE9502"
DESC_RTA_EXTEMPORANEA = (
    "RE9502 - La glosa no procede por haber sido generada fuera de los términos "
    "establecidos por la Ley configurándose la aceptación tácita de la factura "
    "de venta en salud"
)
COD_RTA_NORMAL = "RE9901"
DESC_RTA_NORMAL = "RE9901 - Respuesta a glosa dentro de los términos establecidos por la Ley"

OBS_EXTEMPORANEA = (
    "ESE HUS NO ACEPTA GLOSA, SE LE INFORMA A LA ENTIDAD QUE LA OBJECION ES "
    "EXTEMPORANEA, NO FUE NOTIFICADA EN LOS TIEMPOS PERENTORIOS DE LA NORMA ESTO "
    "EN FUNCION DE LA LEY 1438 DE 2011 EN SU ARTÍCULO 57º. TRÁMITE DE GLOSAS. "
    "LAS ENTIDADES RESPONSABLES DEL PAGO DE SERVICIOS DE SALUD DENTRO DE LOS "
    "VEINTE (20) DÍAS HÁBILES SIGUIENTES A LA PRESENTACIÓN DE LA FACTURA CON "
    "TODOS SUS SOPORTES, FORMULARÁN Y COMUNICARÁN A LOS PRESTADORES DE SERVICIOS "
    "DE SALUD LAS GLOSAS A CADA FACTURA, CON BASE EN LA CODIFICACIÓN Y ALCANCE "
    "DEFINIDOS EN LA NORMATIVIDAD VIGENTE. UNA VEZ FORMULADAS LAS GLOSAS A UNA "
    "FACTURA NO SE PODRÁN FORMULAR NUEVAS GLOSAS A LA MISMA FACTURA, SALVO LAS "
    "QUE SURJAN DE HECHOS NUEVOS DETECTADOS EN LA RESPUESTA DADA A LA GLOSA "
    "INICIAL."
)

OBS_POR_TIPO = {
    "TARIFAS": (
        "ESE HUS NO ACEPTA GLOSA POR TARIFAS, TENIENDO EN CUENTA CONTRATO "
        "68001S00060339-24 Y 68001C00060340-24 VIGENTE ENTRE LAS PARTES PARA LA "
        "PRESTACIÓN DEL SERVICIO, POR LO TANTO, SE FACTURA A TARIFA A TARIFAS "
        "ESTABLECIDAS ENTRE LAS PARTES SOAT SMLV-15% Y TARIFAS INSTITUCIONALES "
        "MEDIANTE RESOLUCIÓN DE LA ESE HUS. NOTA: DE ACUERDO AL ARTÍCULO 57 DE "
        "LA LEY 1438 DE 2011, DE NO OBTENERSE RATIFICACIÓN DE LA RESPUESTA A LA "
        "GLOSA EN LOS TÉRMINOS ESTABLECIDOS, SE DARÁ POR LEVANTADA LA RESPECTIVA "
        "OBJECIÓN"
    ),
    "AUTORIZACION": (
        "ESE HUS NO ACEPTA GLOSA POR AUTORIZACIÓN SE EVIDENCIA LA NOTIFICACIÓN A "
        "LA EPS ANEXO AT 02 Y AT 03 Y ENVÍOS CORRESPONDIENTES (1-2-3 Y 4) ANEXOS "
        "PRORROGAS Y EGRESO CON LOS ENVÍOS CORRESPONDIENTES PARA LA SOLICITUD "
        "AUTORIZACIÓN SEGÚN RESOLUCIÓN 3047/2008 Y DECRETO 4747/2007, PACIENTE "
        "CON ATENTACIÓN INTEGRAL. NOTA: DE ACUERDO AL ARTÍCULO 57 DE LA LEY 1438 "
        "DE 2011, DE NO OBTENERSE RATIFICACIÓN DE LA RESPUESTA A LA GLOSA EN LOS "
        "TÉRMINOS ESTABLECIDOS, SE DARÁ POR LEVANTADA LA RESPECTIVA OBJECIÓN"
    ),
    "FACTURACION": (
        "ESE HUS NO ACEPTA GLOSA POR FACTURACION EN SERVICIOS MEDICOS PRESTADOS "
        "AL PACIENTE DURANTE SU ESTANCIA EN LA ESE HUS, SERVICIOS LOS CUALES SE "
        "FACTURAN EN CONCORDANCIA CON EL DECRETO 2423 DEL 96, SERVICIOS PACTADOS "
        "ENTRE LAS PARTES, LOS SERVICISO FACTURADOS ESTAN ORDENADOS SOPORTADOS Y "
        "JUSTIFICADOS EN HISTORIA CLINICA ADJUNTA EN FACTURA ENVIADA A LA ENTIDAD"
    ),
    "SOPORTES": (
        "ESE HUS NO ACEPTA GLOSA SE ANEXA SOPORTE CORRESPONDIENTE, NOTA: SEGÚN "
        "NORMATIVIDAD VIGENTE DE NO OBTENERSE RATIFICACIÓN DE LA GLOSA EN LOS "
        "TÉRMINOS LEGALES, SE DARÁ POR LEVANTADA LA OBJECIÓN DE ACUERDO ARTÍCULO "
        "57 DE LA LEY 1438 DE 2011"
    ),
}

# Respaldo por si alguna glosa viene sin tipo_glosa: prefijo del código -> tipo.
TIPO_POR_PREFIJO = {
    "TA": "TARIFAS",
    "AU": "AUTORIZACION",
    "FA": "FACTURACION",
    "SO": "SOPORTES",
    "CL": "CALIDAD",
    "DE": "DEVOLUCION",
}

COLS_RESPUESTAS_EXTRA = [
    "FECHA RADICACION",
    "FECHA GLOSA",
    "FECHA DE VENCIMIENTO",
    "DEVOLUCIONES",
    "DIA",
    "EXTEMPORANEA",
    "COD RESPUESTA GLOSA",
    "COD",
    "OBSERVACION RTA GLOSA",
]


def setup_logging(verbose: bool = False) -> None:
    logging.basicConfig(
        level=logging.DEBUG if verbose else logging.INFO,
        format="%(message)s",
        stream=sys.stdout,
    )


# --------------------------------------------------------------------------- utilidades
def norm_texto(s: object) -> str:
    return str(s).strip() if s is not None else ""


def norm_header(s: object) -> str:
    """Normaliza un encabezado para compararlo (mayúsculas, sin espacios dobles)."""
    return re.sub(r"\s+", " ", norm_texto(s)).upper()


def _num_factura(s: object) -> int | None:
    """Extrae el número de la factura: HUS0000512396 / HUS512396 / 512396 -> 512396."""
    txt = norm_texto(s).upper()
    m = RE_NUM_FACTURA.search(txt)
    if m:
        return int(m.group(1))
    if txt.isdigit():
        return int(txt)
    return None


def norm_factura(s: object) -> str:
    """Clave de cruce: HUS0000512396 / HUS512396 / 512396 -> HUS512396."""
    n = _num_factura(s)
    return f"HUS{n}" if n is not None else norm_texto(s).upper()


def factura_dgh(s: object) -> str:
    """Factura en el formato del archivo de objeciones: HUS + 10 dígitos (HUS0000496207)."""
    n = _num_factura(s)
    return f"HUS{n:010d}" if n is not None else norm_texto(s).upper()


def norm_codigo(s: object) -> str:
    """Código/ID para cruce: texto plano en mayúsculas, sin el '.0' que Excel
    agrega cuando la celda viene tipada como número (890466.0 -> 890466)."""
    txt = norm_texto(s).upper()
    if re.fullmatch(r"\d+\.0", txt):
        txt = txt[:-2]
    return txt


def norm_desc(s: object) -> str:
    """Descripción de servicio para cruce de respaldo: mayúsculas, sin acentos,
    sin espacios/puntuación duplicados (para comparar portal vs DGH)."""
    txt = norm_texto(s)
    txt = "".join(c for c in unicodedata.normalize("NFKD", txt) if not unicodedata.combining(c))
    txt = re.sub(r"[^A-Z0-9]+", " ", txt.upper())
    return txt.strip()


# Los medicamentos/insumos llevan un sufijo de presentación tras el último "-"
# (19993036-01). El portal y DGH usan el mismo código base pero distinto sufijo:
#   portal 19931216-05  ->  DGH 19931216-5   (mismo, con ceros de más)
#   portal 20055559-1   ->  DGH 20055559-14  (misma base, otra presentación)
def _cod_sufijo_norm(cod: str) -> str:
    """Quita los ceros a la izquierda del sufijo: 19993036-01 -> 19993036-1."""
    return re.sub(r"-0*(\d+)$", r"-\1", cod)


def _cod_base(cod: str) -> str:
    """Base del código, sin el sufijo de presentación: 20055559-14 -> 20055559."""
    return re.sub(r"-\d+$", "", cod)


def _cod_sufijo_num(cod: str) -> int:
    """Número del sufijo de presentación (para elegir el más cercano)."""
    m = re.search(r"-(\d+)$", cod)
    return int(m.group(1)) if m else -1


def _cruzar_por_nombre_med(cruces: dict, fact: str, descripcion: object) -> str | None:
    """Cruce por NOMBRE del medicamento cuando es inequívoco dentro de la factura.

    Evita que dos medicamentos distintos que comparten base + cero a la izquierda
    (portal 20048691-1 KETAMINA vs 20048691-01 SALBUTAMOL) se lleven el mismo
    código DGH: DGH tiene cada uno con su propio código (KETAMINA 20048691-1,
    SALBUTAMOL 20083667-1), y el nombre los separa. Solo cruza si UN único código
    DGH de la factura coincide por nombre (prefijo en cualquier dirección).
    """
    port = norm_desc(descripcion)
    if len(port) < 4:  # nombres muy cortos son ambiguos, se dejan al código
        return None
    candidatos = set()
    for nombre, cod in cruces.get("nom_med", {}).get(fact, ()):  # (nombre_dgh, código)
        if nombre == port or nombre.startswith(port + " ") or port.startswith(nombre + " "):
            candidatos.add(cod)
    return next(iter(candidatos)) if len(candidatos) == 1 else None


def cruzar_codigo(
    cruces: dict, fact: str, codigo_portal: object, descripcion: object = ""
) -> str | None:
    """Busca el código DGH del servicio, por factura, en varios niveles:
    1) código propio del servicio en DGH (SLNSERPRO_SERVICIO) idéntico
    2) código idéntico en cualquier columna
    3) nombre de medicamento inequívoco (separa KETAMINA de SALBUTAMOL cuando el
       código base coincide por un cero a la izquierda)
    4) sufijo sin ceros / misma base (presentación distinta): elige el sufijo
       más cercano. El guardián de valor (en generar_objeciones) impide objetar
       más de lo que DGH tiene, así que un cruce de más nunca genera error.
    Devuelve None si ninguno cruza (el respaldo por descripción va aparte).
    """
    cod = norm_codigo(codigo_portal)
    if not cod:
        return None
    r = cruces["srv_exact"].get((fact, cod)) or cruces["exact"].get((fact, cod))
    if r:
        return r
    r = _cruzar_por_nombre_med(cruces, fact, descripcion)
    if r:
        return r
    candidatos = cruces["sufijo"].get((fact, _cod_sufijo_norm(cod)), set())
    candidatos = candidatos | cruces["base"].get((fact, _cod_base(cod)), set())
    if candidatos:
        objetivo = _cod_sufijo_num(cod)
        return min(candidatos, key=lambda c: (abs(_cod_sufijo_num(c) - objetivo), len(c), c))
    return None


def a_numero(v: object) -> float | None:
    """Convierte un valor monetario a número. Maneja formato colombiano
    ("1.234.567,89"), US ("1,234,567.89"), miles sueltos ("26,140" / "1.234.567")
    y negativos contables. Devuelve None si no es un número.
    (Misma convención que app/services/recepcion_service._a_float.)
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = norm_texto(v)
    if not s:
        return None
    negativo = s.startswith("-") or (s.startswith("(") and s.endswith(")"))
    s = re.sub(r"[^\d.,]", "", s)
    if not s or not any(c.isdigit() for c in s):
        return None

    tiene_punto, tiene_coma = "." in s, "," in s
    if tiene_punto and tiene_coma:
        # El separador decimal es el que aparece más a la derecha.
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")  # 1.234.567,89
        else:
            s = s.replace(",", "")  # 1,234,567.89
    elif tiene_coma:
        # Una sola coma con 1-2 decimales -> decimal; si no, miles.
        if s.count(",") == 1 and len(s.split(",")[1]) in (1, 2):
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    elif tiene_punto:
        # Un solo punto con 1-2 decimales -> decimal; si no, miles.
        if not (s.count(".") == 1 and len(s.split(".")[1]) in (1, 2)):
            s = s.replace(".", "")

    try:
        n = float(s)
    except ValueError:
        return None
    return -n if negativo else n


def fmt_valor(v: object) -> str:
    """Valor para concatenar en OBSERVACION FINAL: entero sin decimales
    (26140) o con hasta 2 decimales (26140.5), nunca notación científica."""
    n = a_numero(v)
    if n is None:
        return norm_texto(v)
    if n == int(n):
        return str(int(n))
    return f"{n:.2f}".rstrip("0").rstrip(".")


def valor_o_numero(v: object) -> object:
    """Para celdas de valores: número real si se puede, si no el texto original."""
    n = a_numero(v)
    if n is None:
        return norm_texto(v)
    return int(n) if n == int(n) else n


def leer_xlsx(path: Path) -> tuple[list[str], list[list]]:
    """Lee la primera hoja: (encabezados, filas). Filas vacías se descartan.

    - reset_dimensions(): algunos xlsx de portales declaran un rango menor al
      real y openpyxl en read_only los leería vacíos/truncados.
    - Un archivo corrupto o bloqueado aborta con el nombre del archivo.
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(
            f"No se pudo leer '{path.name}': {exc}. "
            "¿Archivo corrupto, incompleto o abierto en Excel?"
        ) from exc
    try:
        ws = wb.active
        ws.reset_dimensions()
        it = ws.iter_rows(values_only=True)
        headers = [norm_texto(h) for h in next(it, [])]
        filas = []
        for row in it:
            if row is None or all(c is None or norm_texto(c) == "" for c in row):
                continue
            filas.append(list(row))
        return headers, filas
    finally:
        wb.close()


def buscar_archivos(carpeta: Path, subcarpeta: str, patron: str) -> list[Path]:
    """Busca recursivo (cubre LOTE 01, LOTE 02... o estructura plana).

    Ignora los archivos de bloqueo de Excel (~$...) y ocultos.
    """
    base = carpeta / subcarpeta
    if not base.is_dir():
        return []
    return sorted(
        (p for p in base.rglob(patron) if not p.name.startswith(("~$", "."))),
        key=lambda p: p.name.upper(),
    )


def descubrir_lotes(carpeta: Path) -> dict[str, dict[str, list[Path]]]:
    """Agrupa los archivos por LOTE (la subcarpeta donde viven) en cada categoría.

    DGH solo acepta cargues de máximo 300 facturas, así que cada LOTE produce su
    propio juego de consolidados + OBJECIONES. El organizador deja los lotes
    ALINEADOS: el LOTE 01 de GLOSAS/DETALLES/FACTURAS cubre las mismas facturas.

    Devuelve {lote: {"GLOSAS": [...], "DETALLES": [...], "FACTURAS": [...]}},
    ordenado por nombre de lote. Estructura plana (archivos sueltos, sin
    subcarpetas) -> un único lote "" y la salida queda como siempre.
    """
    lotes: dict[str, dict[str, list[Path]]] = {}
    for cat in (CARPETA_GLOSAS, CARPETA_DETALLES, CARPETA_FACTURAS):
        raiz = carpeta / cat
        for p in buscar_archivos(carpeta, cat, "*.xlsx"):
            lote = "" if p.parent == raiz else p.parent.name
            lotes.setdefault(lote, {}).setdefault(cat, []).append(p)
    return dict(sorted(lotes.items()))


def indice_columnas(headers: list[str], requeridas: list[str], contexto: str) -> dict[str, int]:
    """Mapa nombre_columna -> posición. Falla claro si falta una requerida."""
    mapa = {norm_header(h): i for i, h in enumerate(headers)}
    idx = {}
    for col in requeridas:
        pos = mapa.get(norm_header(col))
        if pos is None:
            raise ValueError(
                f"{contexto}: no se encontró la columna '{col}'. Encabezados: {headers}"
            )
        idx[col] = pos
    return idx


def alinear_filas(
    headers: list[str], filas: list[list], headers_base: list[str], contexto: str
) -> list[list]:
    """Deja las filas en el orden de columnas de headers_base.

    Si los encabezados coinciden solo rellena filas cortas; si vienen en otro
    orden (o con columnas extra/faltantes) las reordena POR NOMBRE y avisa,
    para que el consolidado nunca quede desalineado en silencio.
    """
    norm_b = [norm_header(h) for h in headers_base]
    norm_f = [norm_header(h) for h in headers]
    if norm_f == norm_b:
        n = len(headers_base)
        return [f + [None] * (n - len(f)) if len(f) < n else f for f in filas]

    logger.warning(
        "  %s: encabezados distintos al primer archivo; se reordenan por nombre.", contexto
    )
    extras = [h for h in norm_f if h not in set(norm_b)]
    if extras:
        logger.warning("    columnas extra ignoradas: %s", ", ".join(extras[:6]))
    pos = {h: i for i, h in enumerate(norm_f)}
    return [[f[pos[h]] if h in pos and pos[h] < len(f) else None for h in norm_b] for f in filas]


def escribir_hoja(
    ws, headers: list[str], filas: list[list], formatos: list[str] | None = None
) -> None:
    ws.append(headers)
    for f in filas:
        ws.append(f)
    if formatos:
        for col_i, fmt in enumerate(formatos, start=1):
            if fmt in ("General", None):
                continue
            for row_i in range(2, len(filas) + 2):
                ws.cell(row=row_i, column=col_i).number_format = fmt
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(40, len(str(h)) + 4))


# ------------------------------------------------------------- días hábiles CO
def _pascua(anio: int) -> "datetime.date":
    """Domingo de Pascua (algoritmo gregoriano anónimo)."""
    from datetime import date as _date

    a = anio % 19
    b, c = divmod(anio, 100)
    d, e = divmod(b, 4)
    g = (8 * b + 13) // 25
    h = (19 * a + b - d - g + 15) % 30
    i, k = divmod(c, 4)
    l = (32 + 2 * e + 2 * i - h - k) % 7  # noqa: E741
    m = (a + 11 * h + 19 * l) // 433
    mes = (h + l - 7 * m + 90) // 25
    dia = (h + l - 7 * m + 33 * mes + 19) % 32
    return _date(anio, mes, dia)


def festivos_colombia(anio: int) -> set:
    """Festivos de Colombia (Ley 51 de 1983 'Emiliani'). Verificado contra los
    consolidados reales del área (mayo 18, junio 8/15/29 y julio 20 de 2026)."""
    from datetime import date as _date
    from datetime import timedelta

    def lunes_siguiente(d):
        return d if d.weekday() == 0 else d + timedelta(days=7 - d.weekday())

    p = _pascua(anio)
    fijos = {
        _date(anio, 1, 1),  # Año Nuevo
        _date(anio, 5, 1),  # Día del Trabajo
        _date(anio, 7, 20),  # Independencia
        _date(anio, 8, 7),  # Boyacá
        _date(anio, 12, 8),  # Inmaculada
        _date(anio, 12, 25),  # Navidad
        p - timedelta(days=3),  # Jueves Santo
        p - timedelta(days=2),  # Viernes Santo
    }
    emiliani = {
        _date(anio, 1, 6),  # Reyes
        _date(anio, 3, 19),  # San José
        _date(anio, 6, 29),  # San Pedro y San Pablo
        _date(anio, 8, 15),  # Asunción
        _date(anio, 10, 12),  # Raza
        _date(anio, 11, 1),  # Todos los Santos
        _date(anio, 11, 11),  # Independencia de Cartagena
        p + timedelta(days=39),  # Ascensión
        p + timedelta(days=60),  # Corpus Christi
        p + timedelta(days=68),  # Sagrado Corazón
    }
    return fijos | {lunes_siguiente(d) for d in emiliani}


def _es_habil(d, festivos: set) -> bool:
    return d.weekday() < 5 and d not in festivos


def dias_habiles_entre(desde, hasta) -> int:
    """Días hábiles en (desde, hasta]: sin contar el día de inicio, contando el
    final. Con este conteo los consolidados reales del área cuadran exacto."""
    from datetime import timedelta

    if desde is None or hasta is None or hasta <= desde:
        return 0
    festivos = set()
    for anio in range(desde.year, hasta.year + 1):
        festivos |= festivos_colombia(anio)
    n, d = 0, desde
    while d < hasta:
        d += timedelta(days=1)
        if _es_habil(d, festivos):
            n += 1
    return n


def sumar_dias_habiles(desde, n: int):
    """Fecha del n-ésimo día hábil después de `desde`."""
    from datetime import timedelta

    festivos = festivos_colombia(desde.year) | festivos_colombia(desde.year + 1)
    d = desde
    while n > 0:
        d += timedelta(days=1)
        if _es_habil(d, festivos):
            n -= 1
    return d


def a_fecha(v):
    """Fecha desde celda Excel: datetime/date o texto dd/mm/aaaa · aaaa-mm-dd."""
    from datetime import date as _date
    from datetime import datetime as _dt

    if isinstance(v, _dt):
        return v.date()
    if isinstance(v, _date):
        return v
    s = norm_texto(v)[:19]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
        try:
            return _dt.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def generar_respuestas_glosas(
    h_glosas: list[str],
    f_glosas: list[list],
    h_fact: list[str],
    f_fact: list[list],
    fecha_masivo,
) -> tuple[list[str], list[list], dict]:
    """Arma el consolidado de RESPUESTAS a las glosas (respuestas predeterminadas).

    Por cada glosa: si la EPS la formuló pasados los 20 días hábiles desde la
    radicación de la factura (art. 57, Ley 1438/2011) va con RE9502 y el texto
    de extemporaneidad; si fue a tiempo, va con RE9901 y el texto del área
    según el tipo de glosa (TARIFAS, AUTORIZACION, FACTURACION, SOPORTES).
    """
    mapa_g = {norm_header(h): i for i, h in enumerate(h_glosas)}
    idx_fact_g = mapa_g.get(norm_header("numero_factura"))
    idx_cod = mapa_g.get(norm_header("codigo_glosa"))
    idx_tipo = mapa_g.get(norm_header("tipo_glosa"))
    idx_obs_final = mapa_g.get(norm_header("OBSERVACION FINAL"))

    # radicación por factura (viene en la cabecera HUS*.xlsx del portal)
    mapa_f = {norm_header(h): i for i, h in enumerate(h_fact)}
    idx_fact_f = mapa_f.get(norm_header("numero_factura"))
    idx_rad = mapa_f.get(norm_header("fecha_radicacion"))
    radicacion: dict[str, object] = {}
    if idx_fact_f is not None and idx_rad is not None:
        for r in f_fact:
            f = norm_factura(r[idx_fact_f])
            fec = a_fecha(r[idx_rad])
            if fec and f not in radicacion:
                radicacion[f] = fec

    fecha_glosa_masivo = fecha_masivo.date() if hasattr(fecha_masivo, "date") else fecha_masivo
    vencimiento = sumar_dias_habiles(fecha_glosa_masivo, DIAS_HABILES_RESPUESTA)

    # encabezados: los del portal + numero_factura2 tras numero_factura,
    # sin OBSERVACION FINAL (esa es del archivo de OBJECIONES), + las nuevas.
    headers: list[str] = []
    for i, h in enumerate(h_glosas):
        if i == idx_obs_final:
            continue
        headers.append(h)
        if i == idx_fact_g:
            headers.append("numero_factura2")
    headers += COLS_RESPUESTAS_EXTRA

    filas: list[list] = []
    sin_radicacion: set[str] = set()
    filas_sin_rad = 0
    sin_texto: dict[str, int] = {}
    n_ext = 0
    pos2 = headers.index("numero_factura2")
    for r in f_glosas:
        fact = r[idx_fact_g] if idx_fact_g is not None else ""
        fkey = norm_factura(fact)
        rad = radicacion.get(fkey)
        codigo = norm_texto(r[idx_cod]) if idx_cod is not None else ""
        tipo = norm_texto(r[idx_tipo]).upper() if idx_tipo is not None else ""
        if not tipo:
            tipo = TIPO_POR_PREFIJO.get(codigo[:2].upper(), "")

        base = [c for i, c in enumerate(r) if i != idx_obs_final]
        # numero_factura2 en la posición siguiente a numero_factura
        base = base[:pos2] + [factura_dgh(fact)] + base[pos2:]

        if rad is None:
            sin_radicacion.add(fkey)
            filas_sin_rad += 1
            filas.append(base + [None, fecha_glosa_masivo, vencimiento, "NO", None, "", "", "", ""])
            continue

        dia = dias_habiles_entre(rad, fecha_glosa_masivo)
        extemporanea = dia > DIAS_HABILES_EPS
        if extemporanea:
            n_ext += 1
            cod, desc, obs = COD_RTA_EXTEMPORANEA, DESC_RTA_EXTEMPORANEA, OBS_EXTEMPORANEA
        else:
            cod, desc = COD_RTA_NORMAL, DESC_RTA_NORMAL
            obs = OBS_POR_TIPO.get(tipo, "")
            if not obs:
                # CALIDAD (y cualquier tipo sin texto del área): la respuesta es
                # de auditoría médica — se deja TODO vacío (código y observación)
                # para que las doctoras diligencien su propia respuesta.
                cod, desc = "", ""
                sin_texto[tipo or codigo[:2].upper()] = sin_texto.get(tipo or "?", 0) + 1
        devol = "SI" if codigo[:2].upper() == "DE" or tipo == "DEVOLUCION" else "NO"
        filas.append(
            base
            + [
                rad,
                fecha_glosa_masivo,
                vencimiento,
                devol,
                dia,
                "SI" if extemporanea else "NO",
                desc,
                cod,
                obs,
            ]
        )

    resumen = {
        "extemporaneas": n_ext,
        "a_tiempo": len(filas) - n_ext - filas_sin_rad,
        "sin_radicacion": sin_radicacion,
        "sin_texto": sin_texto,
    }
    return headers, filas, resumen


# --------------------------------------------------------------------------- pasos
def consolidar_glosas(archivos: list[Path]) -> tuple[list[str], list[list], dict[str, dict]]:
    """Une los GLOSAS *.xlsx, agrega OBSERVACION FINAL y agrupa por id_detalle.

    Deduplica por id_glosa (archivos o filas repetidas se saltan con aviso).

    Devuelve (headers, filas, agrupado) donde agrupado[id_detalle_norm] = {
        'observacion':  texto combinado con saltos de línea,
        'conceptos':    set de prefijos de 2 letras (TA, CL, ...),
        'codigo_mayor': codigo_glosa del servicio; una glosa CL (médica) tiene
                        prioridad; si no hay CL, la de mayor valor,
        'max_valor':    valor de esa glosa,
    }.
    """
    headers_base: list[str] | None = None
    filas: list[list] = []
    agrupado: dict[str, dict] = {}
    vistas: set = set()
    duplicadas = 0

    for arch in archivos:
        headers, rows = leer_xlsx(arch)
        if not rows:
            logger.warning("  (vacío) %s", arch.name)
            continue
        if headers_base is None:
            headers_base = headers
        rows = alinear_filas(headers, rows, headers_base, arch.name)
        idx = indice_columnas(headers_base, COL_GLOSAS_REQ, arch.name)
        mapa_h = {norm_header(h): i for i, h in enumerate(headers_base)}
        idx_id_glosa = mapa_h.get("ID_GLOSA")

        for row in rows:
            # Clave de duplicado: id_glosa si existe; si no, la fila completa.
            if idx_id_glosa is not None and norm_texto(row[idx_id_glosa]):
                clave = norm_codigo(row[idx_id_glosa])
            else:
                clave = tuple(norm_texto(c) for c in row)
            if clave in vistas:
                duplicadas += 1
                continue
            vistas.add(clave)

            codigo = norm_texto(row[idx["codigo_glosa"]])
            justif = norm_texto(row[idx["justificacion_glosa"]])
            valor = row[idx["valor_total_glosa"]]
            obs = f"{codigo} {justif}${fmt_valor(valor)}"
            filas.append(list(row) + [obs])

            id_det = norm_codigo(row[idx["id_detalle"]])
            g = agrupado.setdefault(
                id_det,
                {
                    "observacion": [],
                    "conceptos": set(),
                    "codigo_mayor": "",
                    "max_valor": None,
                    "mayor_es_cl": False,
                },
            )
            g["observacion"].append(obs)
            g["conceptos"].add(codigo[:2].upper())
            v = a_numero(valor)
            es_cl = codigo[:2].upper() == "CL"
            if not g["codigo_mayor"]:
                g["codigo_mayor"] = codigo
                g["max_valor"] = v
                g["mayor_es_cl"] = es_cl
            else:
                # Una glosa CL (médica) manda por encima de cualquier otra, sin
                # importar el valor. Dentro del mismo tipo (CL vs CL, u otra vs
                # otra) gana la de mayor valor; los empates conservan la primera.
                mayor_por_valor = v is not None and (g["max_valor"] is None or v > g["max_valor"])
                if (es_cl and not g["mayor_es_cl"]) or (
                    es_cl == g["mayor_es_cl"] and mayor_por_valor
                ):
                    g["codigo_mayor"] = codigo
                    g["max_valor"] = v
                    g["mayor_es_cl"] = es_cl

    if duplicadas:
        logger.warning("  OJO: %d glosas duplicadas se saltaron (archivos repetidos).", duplicadas)

    for g in agrupado.values():
        g["observacion"] = "\n".join(g["observacion"])

    if headers_base is None:
        raise ValueError("No se encontró ningún archivo GLOSAS con datos.")
    return headers_base + ["OBSERVACION FINAL"], filas, agrupado


def consolidar_detalles(
    archivos: list[Path], agrupado: dict[str, dict]
) -> tuple[list[str], list[list], list[dict]]:
    """Une los DETALLE *.xlsx + OBSERVACION FINAL por id_detalle.

    Deduplica por id_detalle (un servicio repetido duplicaría el valor objetado).
    Devuelve además la lista de servicios glosados (insumo de OBJECIONES).
    """
    headers_base: list[str] | None = None
    filas: list[list] = []
    glosados: list[dict] = []
    vistos: set = set()
    duplicados = 0

    for arch in archivos:
        headers, rows = leer_xlsx(arch)
        if not rows:
            logger.warning("  (vacío) %s", arch.name)
            continue
        if headers_base is None:
            headers_base = headers
        rows = alinear_filas(headers, rows, headers_base, arch.name)
        idx = indice_columnas(headers_base, COL_DETALLE_REQ, arch.name)
        mapa_h = {norm_header(h): i for i, h in enumerate(headers_base)}
        idx_desc = mapa_h.get("DESCRIPCION")
        idx_vt = mapa_h.get(norm_header("valor_total"))
        idx_cm = mapa_h.get(norm_header("valor_cuota_moderadora"))

        for row in rows:
            id_det = norm_codigo(row[idx["id_detalle"]])
            if id_det in vistos:
                duplicados += 1
                continue
            vistos.add(id_det)

            g = agrupado.get(id_det)
            obs = g["observacion"] if g else ""
            filas.append(list(row) + [obs])

            if g:
                glosados.append(
                    {
                        "factura": norm_texto(row[idx["numero_factura"]]),
                        "id_detalle": id_det,
                        "codigo_servicio": norm_texto(row[idx["codigo_servicio"]]),
                        "descripcion": norm_texto(row[idx_desc]) if idx_desc is not None else "",
                        "valor_glosado": valor_o_numero(row[idx["valor_glosado"]]),
                        "valor_servicio": row[idx_vt] if idx_vt is not None else None,
                        "copago": row[idx_cm] if idx_cm is not None else None,
                        "observacion": obs,
                        "conceptos": g["conceptos"],
                        "codigo_mayor": g["codigo_mayor"],
                    }
                )

    if duplicados:
        logger.warning(
            "  OJO: %d líneas de detalle duplicadas se saltaron (archivos repetidos).", duplicados
        )

    if headers_base is None:
        raise ValueError("No se encontró ningún archivo DETALLE con datos.")
    return headers_base + ["OBSERVACION FINAL"], filas, glosados


def consolidar_facturas(archivos: list[Path]) -> tuple[list[str], list[list]]:
    headers_base: list[str] | None = None
    filas: list[list] = []
    vistas: set = set()
    for arch in archivos:
        headers, rows = leer_xlsx(arch)
        if not rows:
            continue
        if headers_base is None:
            headers_base = headers
        rows = alinear_filas(headers, rows, headers_base, arch.name)
        for row in rows:
            clave = tuple(norm_texto(c) for c in row)
            if clave in vistas:
                continue
            vistas.add(clave)
            filas.append(list(row))
    if headers_base is None:
        raise ValueError("No se encontró ningún archivo de FACTURAS con datos.")
    return headers_base, filas


def _iterar_base(path: Path) -> Iterator[list]:
    """Itera las filas de la base DGH (xlsx/csv/txt) en streaming, sin
    materializarla: la base real pesa 80MB+ y cargarla entera agota la RAM.
    La primera fila que produce es el encabezado."""
    if path.suffix.lower() in {".csv", ".txt"}:
        import csv as _csv

        with open(path, newline="", encoding="utf-8-sig", errors="replace") as fh:
            muestra = fh.read(4096)
            fh.seek(0)
            try:
                dialecto = _csv.Sniffer().sniff(muestra, delimiters=";,\t|")
            except _csv.Error:
                dialecto = _csv.excel
            for row in _csv.reader(fh, dialecto):
                yield list(row)
    else:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError(f"No se pudo leer la base DGH '{path.name}': {exc}") from exc
        try:
            ws = wb.active
            ws.reset_dimensions()
            for row in ws.iter_rows(values_only=True):
                yield list(row)
        finally:
            wb.close()


def cargar_base_dgh(
    path: Path, facturas_trabajadas: set[str]
) -> tuple[list[list], dict[tuple[str, str], str]]:
    """Lee la base DGH en streaming, filtra las facturas trabajadas y la "arregla".

    El arreglo (igual al proceso manual): en las filas de medicamentos, que
    vienen con SLNSERPRO_SERVICIO vacío, se rellenan SERVICIO/CUPS y sus
    descripciones con CODIGO_MEDICAMENTO / NOMBRE_MEDICAMENTO.

    Devuelve:
      - filas para SERVICIOS FACTURADOS (solo COLS_SERVICIOS, ya arregladas)
      - cruce[(factura_norm, codigo_norm)] = código DGH del servicio (SLNSERPRO)
    """
    logger.info("Leyendo base DGH: %s (puede tardar si pesa mucho)...", path)

    it = _iterar_base(path)
    headers = [norm_texto(h) for h in next(it, [])]
    mapa = {norm_header(h): i for i, h in enumerate(headers)}

    def col(nombre: str) -> int | None:
        return mapa.get(norm_header(nombre))

    idx_fact = col("FACTURA")
    if idx_fact is None:
        raise ValueError(f"Base DGH: no se encontró la columna FACTURA. Encabezados: {headers}")
    idx_srv = col("SLNSERPRO_SERVICIO")
    idx_cups = col("SLNSERPRO_CUPS")
    idx_cod_med = col("CODIGO_MEDICAMENTO")
    idx_cod_med_fact = col("COD_MED_FACTURA")
    idx_nom_med = col("NOMBRE_MEDICAMENTO")
    idx_desc_inst = col("DESCRIPCION INSTITUCIONAL")
    idx_desc_cups = col("DESCRIPCION CUPS")
    idx_vr = col("Vr_SERVICIO")
    idx_saldo = col("SALDO_FACT")

    idx_salida = [col(c) for c in COLS_SERVICIOS]
    faltantes = [c for c, pos in zip(COLS_SERVICIOS, idx_salida) if pos is None]
    if len(faltantes) >= 5:
        # A este archivo le faltan casi todas las columnas: NO es la base DGH.
        raise ValueError(
            f"'{path.name}' NO parece la base DGH de SERVICIOS FACTURADOS "
            f"(le faltan {len(faltantes)} columnas, ej: {', '.join(faltantes[:4])}...). "
            "El archivo correcto se llama 'SERVICIOS FACTURADOS COOSALUD DGH.xlsx' "
            "y trae SLNSERPRO_SERVICIO, SLNSERPRO_CUPS, CODIGO_MEDICAMENTO, etc."
        )
    for c in faltantes:
        logger.warning("Base DGH: falta la columna '%s' (saldrá vacía).", c)

    def celda(row: list, i: int | None) -> str:
        return norm_texto(row[i]) if i is not None and i < len(row) else ""

    filas_filtradas: list[list] = []
    # Índices de cruce (todos por factura):
    #   srv_exact = código == SLNSERPRO_SERVICIO propio (máxima prioridad)
    #   exact     = código idéntico en cualquier columna
    #   sufijo    = código con el sufijo sin ceros (conjunto de códigos DGH)
    #   base      = misma base, presentación distinta (conjunto de códigos DGH)
    #   desc      = por descripción, solo si es inequívoca
    #   valor     = suma de Vr_SERVICIO por código   ·   saldo = SALDO_FACT por factura
    cruce_srv_exact: dict[tuple[str, str], str] = {}
    cruce_exact: dict[tuple[str, str], str] = {}
    cruce_sufijo: dict[tuple[str, str], set] = {}
    cruce_base: dict[tuple[str, str], set] = {}
    desc_a_cod: dict[tuple[str, str], set] = {}
    nom_med_idx: dict[str, set] = {}  # factura -> {(nombre_medicamento_norm, código DGH)}
    valor: dict[tuple[str, str], float] = {}
    saldo: dict[str, float] = {}
    total = 0
    for row in it:
        total += 1
        if row is None or all(c is None or norm_texto(c) == "" for c in row):
            continue
        fact = norm_factura(celda(row, idx_fact))
        if fact not in facturas_trabajadas:
            continue

        servicio = norm_codigo(celda(row, idx_srv))
        cod_med = norm_codigo(celda(row, idx_cod_med))
        nom_med = celda(row, idx_nom_med)

        # --- Arreglo: medicamentos rellenan las columnas de servicio ---
        valores = {}
        for c, i in zip(COLS_SERVICIOS, idx_salida):
            v = celda(row, i)
            valores[c] = norm_codigo(v) if c in COLS_SERVICIOS_CODIGO else v
        if not servicio and cod_med:
            valores["SLNSERPRO_SERVICIO"] = cod_med
            valores["DESCRIPCION INSTITUCIONAL"] = nom_med
            valores["SLNSERPRO_CUPS"] = cod_med
            valores["DESCRIPCION CUPS"] = nom_med
        filas_filtradas.append([valores[c] for c in COLS_SERVICIOS])

        # --- Cruce: cualquier código conocido de la fila -> código DGH efectivo ---
        slnserpro = valores["SLNSERPRO_SERVICIO"]
        if slnserpro:
            cruce_srv_exact.setdefault((fact, slnserpro), slnserpro)
            for i in (idx_srv, idx_cups, idx_cod_med, idx_cod_med_fact):
                cod = norm_codigo(celda(row, i))
                if cod:
                    cruce_exact.setdefault((fact, cod), slnserpro)
                    cruce_sufijo.setdefault((fact, _cod_sufijo_norm(cod)), set()).add(slnserpro)
                    cruce_base.setdefault((fact, _cod_base(cod)), set()).add(slnserpro)
            # descripción del servicio DGH (respaldo de cruce por descripción)
            for i in (idx_desc_inst, idx_desc_cups):
                desc = norm_desc(celda(row, i))
                if desc:
                    desc_a_cod.setdefault((fact, desc), set()).add(slnserpro)
            # nombre del medicamento (para separar meds con base+cero a la
            # izquierda). Se indexa por cod_med, que identifica a un medicamento
            # tanto en la base cruda (SLNSERPRO vacío) como en la ya "arreglada".
            if cod_med:
                nom_med_norm = norm_desc(nom_med)
                if nom_med_norm:
                    nom_med_idx.setdefault(fact, set()).add((nom_med_norm, slnserpro))
            # valor del servicio (para no objetar más de lo que DGH tiene)
            v = a_numero(celda(row, idx_vr))
            if v is not None:
                valor[(fact, slnserpro)] = valor.get((fact, slnserpro), 0.0) + v
        s = a_numero(celda(row, idx_saldo))
        if s is not None:
            saldo[fact] = s

    # La descripción solo sirve de cruce si es inequívoca (un único código DGH).
    cruce_desc = {k: next(iter(v)) for k, v in desc_a_cod.items() if len(v) == 1}
    cruces = {
        "srv_exact": cruce_srv_exact,
        "exact": cruce_exact,
        "sufijo": cruce_sufijo,
        "base": cruce_base,
        "desc": cruce_desc,
        "nom_med": {f: tuple(s) for f, s in nom_med_idx.items()},
        "valor": valor,
        "saldo": saldo,
    }

    logger.info(
        "Base DGH: %d filas leídas · %d de las %d facturas trabajadas.",
        total,
        len(filas_filtradas),
        len(facturas_trabajadas),
    )
    if not filas_filtradas:
        raise ValueError(
            f"la base '{path.name}' no contiene NINGUNA de las "
            f"{len(facturas_trabajadas)} facturas trabajadas. ¿Es la base "
            "correcta y está actualizada con las fechas de estas facturas?"
        )
    return filas_filtradas, cruces


def generar_objeciones(
    glosados: list[dict],
    fecha: datetime,
    cruces: dict | None,
    excluir_no_cruzados: bool = True,
) -> tuple[list[list], list[list]]:
    """Arma las filas del archivo de OBJECIONES (una por servicio glosado).

    Si hay base DGH y `excluir_no_cruzados` (por defecto), los servicios que DGH
    no tiene en la factura NO se agregan al OBJECIONES (irían con error de
    "servicio no asociado"); quedan solo en la lista de NO_CRUZADOS.

    Devuelve (filas_objeciones, filas_no_cruzadas, filas_ajustadas_por_valor).
    """
    filas: list[list] = []
    no_cruzados: list[list] = []
    ajustados: list[list] = []
    malformadas: set[str] = set()
    por_respaldo = 0
    capados_copago = 0
    excluidos = 0
    # Guardián de valor: lo ya objetado por (factura, código DGH) y por factura,
    # para no pasarse del valor del servicio ni del saldo de la cuenta.
    acum_cod: dict[tuple[str, str], float] = {}
    acum_fac: dict[str, float] = {}
    consecutivo = 0
    factura_actual: str | None = None

    # CROTIPOBJ se fija por FACTURA más abajo, según el CRNCONOBJ realmente
    # escrito en cada fila (no según las glosas crudas). Ver el bloque al final.
    for srv in glosados:
        fact = srv["factura"]
        if _num_factura(fact) is None:
            malformadas.add(fact)

        # SLNSERPRO: el codigo_servicio del DETALLE; si hay base DGH y la fila
        # cruza (código exacto, sufijo sin ceros, misma base o descripción),
        # manda el código DGH. Si NO cruza, o el valor se pasa, va a NO_CRUZADOS.
        cod_portal = norm_codigo(srv["codigo_servicio"])
        slnserpro = cod_portal or None
        valor_final = srv["valor_glosado"]
        motivo = None
        if cruces is not None:
            fkey = norm_factura(fact)
            del_dgh = cruzar_codigo(
                cruces, fkey, srv["codigo_servicio"], srv.get("descripcion", "")
            )
            if not del_dgh:
                # Último respaldo: por descripción (rescata códigos totalmente
                # distintos, p. ej. DERECHOS DE SALA PARA CURACIONES).
                del_dgh = cruces["desc"].get((fkey, norm_desc(srv.get("descripcion", ""))))
            if del_dgh:
                if del_dgh != cod_portal:
                    por_respaldo += 1
                slnserpro = del_dgh
            else:
                motivo = "no está en DGH"

            # --- Guardián de valor/saldo (solo si el código cruzó) ---
            # DGH no acepta objetar más que el valor del servicio ni más que el
            # saldo de la cuenta. Se CAPA la objeción a lo que quede disponible
            # (para no perder la objeción entera) y se reporta el ajuste. Si ya
            # no queda cupo, se manda a NO_CRUZADOS.
            if motivo is None:
                valobj = a_numero(srv["valor_glosado"])
                if valobj is not None:
                    lim_cod = cruces["valor"].get((fkey, slnserpro))
                    lim_sal = cruces["saldo"].get(fkey)
                    cupo = float("inf")
                    if lim_cod is not None:
                        cupo = min(cupo, lim_cod - acum_cod.get((fkey, slnserpro), 0.0))
                    if lim_sal is not None:
                        cupo = min(cupo, lim_sal - acum_fac.get(fkey, 0.0))
                    # Copago (cuota moderadora): DGH descuenta el copago del
                    # valor del servicio, así que el máximo objetable de esta
                    # línea es (valor del servicio - copago). La cuota la paga
                    # el paciente, no la EPS; si se objeta de más, DGH la rechaza
                    # con "El VALOR OBJECION no puede ser mayor al valor del
                    # servicio". Se capa la línea a lo que DGH sí acepta.
                    vserv = a_numero(srv.get("valor_servicio"))
                    copago = a_numero(srv.get("copago")) or 0.0
                    if vserv is not None and copago > 0.5:
                        tope_copago = vserv - copago
                        if tope_copago < cupo:
                            cupo = tope_copago
                            capados_copago += 1
                    if cupo <= 0.5:
                        motivo = "sin cupo en DGH (servicio o saldo ya cubierto por otras glosas)"
                    elif valobj > cupo + 0.5:
                        # Capar al máximo que DGH acepta y registrar el ajuste.
                        capado = int(round(cupo))
                        ajustados.append(
                            [
                                factura_dgh(fact),
                                srv["id_detalle"],
                                slnserpro,
                                srv.get("descripcion", ""),
                                srv["valor_glosado"],
                                capado,
                                srv["observacion"][:90],
                            ]
                        )
                        valor_final = capado
                        valobj = float(capado)
                    acum_cod[(fkey, slnserpro)] = acum_cod.get((fkey, slnserpro), 0.0) + (
                        valobj or 0.0
                    )
                    acum_fac[fkey] = acum_fac.get(fkey, 0.0) + (valobj or 0.0)

            if motivo:
                no_cruzados.append(
                    [
                        factura_dgh(fact),
                        srv["id_detalle"],
                        srv["codigo_servicio"],
                        srv.get("descripcion", ""),
                        srv["valor_glosado"],
                        motivo,
                        srv["observacion"][:90],
                    ]
                )
                if excluir_no_cruzados:
                    # No se agrega al OBJECIONES (DGH lo marcaría con error).
                    excluidos += 1
                    continue

        # El consecutivo solo avanza para las filas que SÍ quedan en el archivo,
        # para no dejar huecos cuando se excluyen los no cruzados.
        if fact != factura_actual:
            consecutivo += 1
            factura_actual = fact

        filas.append(
            [
                str(consecutivo),  # CDCONSEC (texto, como en la plantilla)
                fecha,  # CDFECDOC (fecha Excel)
                factura_dgh(fact),  # CRNCXC   (HUS0000496207)
                fecha,  # CROFECOBJ
                None,  # CROREFERE
                None,  # CROOBSERV
                0,  # CROCLAOBJ
                None,  # CRNCLAOBJ
                "999",  # GENUSUARIO4 (texto)
                srv["codigo_mayor"],  # CRNCONOBJ (CL manda; si no, la de mayor valor)
                slnserpro,  # SLNSERPRO
                None,  # IDRIPS
                None,  # CTNCENCOS
                valor_final,  # CROVALOBJ (número; capado al tope DGH si aplicaba)
                srv["observacion"],  # CRDOBSERV
                0,  # CROTIPOBJ (se fija por factura más abajo, según CRNCONOBJ)
            ]
        )

    # --- CROTIPOBJ definitivo: según el CRNCONOBJ realmente escrito ---
    # DGH clasifica cada objeción por su CONCEPTO PRINCIPAL (CRNCONOBJ), no por
    # las glosas que solo quedan en la observación. Como una glosa CL hace que el
    # CRNCONOBJ de ese servicio salga CL, el tipo de la factura sigue esos códigos
    # escritos: si TODOS son CL -> médica (1); si NINGUNO es CL -> administrativa
    # (0); si de verdad hay mezcla de CL y no-CL -> mixta (2).
    # (Antes se miraban las glosas crudas: una factura toda-CL salía mixta por las
    #  TA que solo van en la observación, y una factura cuyas CL no cruzaron a DGH
    #  salía mixta aunque en el archivo solo queden filas administrativas.)
    cl_por_factura: dict[str, set] = {}
    for fila in filas:
        es_cl = str(fila[9] or "").upper()[:2] == "CL"
        cl_por_factura.setdefault(fila[2], set()).add(es_cl)
    tipo_por_factura = {
        f: (1 if flags == {True} else 0 if flags == {False} else 2)
        for f, flags in cl_por_factura.items()
    }
    for fila in filas:
        fila[15] = tipo_por_factura[fila[2]]

    if por_respaldo:
        logger.info(
            "  %d servicios tomaron el código DGH por respaldo (sufijo/base/descripción).",
            por_respaldo,
        )
    if excluidos:
        logger.info(
            "  %d servicios NO están en DGH: se dejaron FUERA del OBJECIONES "
            "(quedan en la hoja NO_CRUZADOS).",
            excluidos,
        )
    if ajustados:
        logger.info(
            "  %d objeciones se CAPARON al valor que DGH acepta (hoja VALOR_AJUSTADO).",
            len(ajustados),
        )
    if capados_copago:
        logger.info(
            "  %d de esas objeciones se caparon por COPAGO (cuota moderadora): "
            "DGH no deja objetar la parte que paga el paciente.",
            capados_copago,
        )
    if malformadas:
        logger.warning(
            "OJO: %d factura(s) no siguen el patrón HUS y salen tal cual en CRNCXC: %s",
            len(malformadas),
            ", ".join(sorted(malformadas)[:5]),
        )
    return filas, no_cruzados, ajustados


# --------------------------------------------------------------------------- main
def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Consolida GLOSAS/DETALLES/FACTURAS de COOSALUD y genera el archivo de OBJECIONES (DGH).",
    )
    p.add_argument(
        "--carpeta",
        required=True,
        help='Carpeta organizada (la "CARGUE MASIVO COOSALUD" que crea el organizador).',
    )
    p.add_argument(
        "--fecha",
        required=True,
        help="Fecha del día que se revisa, DD/MM/AAAA (ej: 04/07/2026). Va en CDFECDOC y CROFECOBJ.",
    )
    p.add_argument(
        "--servicios",
        default=None,
        help="Base DGH de servicios facturados (xlsx/csv/txt) para el punto 4 y el cruce SLNSERPRO.",
    )
    p.add_argument(
        "--incluir-no-cruzados",
        action="store_true",
        help="Deja en el OBJECIONES los servicios que DGH no tiene (marcarán error). "
        "Por defecto se dejan FUERA (solo en la hoja NO_CRUZADOS) para que el cargue pase.",
    )
    p.add_argument(
        "--omitir-facturas",
        default=None,
        help="Archivo de texto con facturas que YA están objetadas en DGH (una por "
        "línea, ej. HUS509457). Se dejan fuera del OBJECIONES — si van, DGH "
        "rechaza el cargue COMPLETO. Los consolidados sí las conservan.",
    )
    p.add_argument(
        "--salida",
        default=None,
        help=f"Carpeta de salida. Def: <carpeta>\\{SALIDA_DEFECTO}",
    )
    p.add_argument("-v", "--verbose", action="store_true", help="Log detallado.")
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    setup_logging(args.verbose)

    try:
        fecha = datetime.strptime(args.fecha, "%d/%m/%Y")
    except ValueError:
        logger.error(
            "ERROR: --fecha debe ser DD/MM/AAAA (ej: 04/07/2026). Recibido: %s", args.fecha
        )
        return 1

    carpeta = Path(args.carpeta)
    if not carpeta.is_dir():
        logger.error("ERROR: no existe la carpeta %s", carpeta)
        return 1
    salida = Path(args.salida) if args.salida else carpeta / SALIDA_DEFECTO
    salida.mkdir(parents=True, exist_ok=True)

    # Facturas que ya están objetadas en DGH: se dejan fuera del OBJECIONES
    # (DGH rechaza el cargue COMPLETO si va una repetida y no guarda nada).
    omitidas: set[str] = set()
    if args.omitir_facturas:
        p_omit = Path(args.omitir_facturas)
        if not p_omit.is_file():
            logger.error("ERROR: no existe el archivo de --omitir-facturas: %s", p_omit)
            return 1
        with p_omit.open(encoding="utf-8", errors="replace") as fh:
            omitidas = {norm_factura(lin) for lin in fh if lin.strip()}
        logger.info(
            "Se omitirán del OBJECIONES %d facturas ya objetadas en DGH (%s).",
            len(omitidas),
            p_omit.name,
        )

    lotes = descubrir_lotes(carpeta)
    n_glo = sum(len(v.get(CARPETA_GLOSAS, [])) for v in lotes.values())
    n_det = sum(len(v.get(CARPETA_DETALLES, [])) for v in lotes.values())
    n_fac = sum(len(v.get(CARPETA_FACTURAS, [])) for v in lotes.values())
    con_lotes = any(nombre for nombre in lotes)
    logger.info(
        "Archivos: %d GLOSAS · %d DETALLES · %d FACTURAS · %s",
        n_glo,
        n_det,
        n_fac,
        f"{len(lotes)} lote(s): un OBJECIONES por lote" if con_lotes else "sin lotes",
    )
    if not n_glo or not n_det:
        logger.error("ERROR: faltan archivos en GLOSAS o DETALLES dentro de %s", carpeta)
        return 1
    for nombre, archivos in lotes.items():
        if not archivos.get(CARPETA_GLOSAS) or not archivos.get(CARPETA_DETALLES):
            logger.error(
                "ERROR: el lote '%s' no tiene archivos en GLOSAS o DETALLES. "
                "Los lotes deben estar alineados (mismo LOTE en las 3 carpetas).",
                nombre or "(raíz)",
            )
            return 1

    try:
        # --- Puntos 2 y 3: consolidados + observación final, POR LOTE ----------
        # (DGH solo acepta cargues de máx 300 facturas: cada lote sale aparte.)
        datos: dict[str, dict] = {}
        for nombre, archivos in lotes.items():
            logger.info("\n===== %s =====", nombre or "CONSOLIDANDO (sin lotes)")
            logger.info("[1/4] Consolidando GLOSAS...")
            h_glosas, f_glosas, agrupado = consolidar_glosas(archivos[CARPETA_GLOSAS])
            logger.info("  %d glosas · %d id_detalle distintos", len(f_glosas), len(agrupado))

            logger.info("[2/4] Consolidando DETALLES (+ OBSERVACION FINAL)...")
            h_det, f_det, glosados = consolidar_detalles(archivos[CARPETA_DETALLES], agrupado)
            logger.info("  %d líneas de detalle · %d servicios glosados", len(f_det), len(glosados))

            logger.info("[3/4] Consolidando FACTURAS...")
            arch_fact = archivos.get(CARPETA_FACTURAS, [])
            h_fact, f_fact = consolidar_facturas(arch_fact) if arch_fact else ([], [])
            logger.info("  %d facturas", len(f_fact))

            # Facturas con COPAGO/cuota moderadora: DGH les descuenta el copago
            # al valor de los servicios, así que la objeción puede salir mayor
            # y el cargue marcaría error de valor. Se avisa desde ya.
            mapa_det = {norm_header(h): i for i, h in enumerate(h_det)}
            idx_cm = mapa_det.get(norm_header("valor_cuota_moderadora"))
            idx_fac_det = mapa_det.get(norm_header("numero_factura"))
            if idx_cm is not None and idx_fac_det is not None:
                copagos: dict[str, float] = {}
                for fila_det in f_det:
                    cm = a_numero(fila_det[idx_cm])
                    if cm and cm > 0:
                        fkey = norm_factura(fila_det[idx_fac_det])
                        copagos[fkey] = copagos.get(fkey, 0.0) + cm
                if copagos:
                    top = sorted(copagos.items(), key=lambda x: -x[1])
                    logger.warning(
                        "  OJO: %d factura(s) con COPAGO. Si DGH devuelve errores de "
                        "valor en este lote, casi seguro son estas (las de copago "
                        "grande suelen tocar A MANO en DGH): %s",
                        len(copagos),
                        " · ".join(f"{f} (${v:,.0f})" for f, v in top[:5]).replace(",", "."),
                    )

            # Glosas cuyo id_detalle no apareció en ningún DETALLE (avisar, no perder).
            ids_detalle = {g["id_detalle"] for g in glosados}
            huerfanas = [k for k in agrupado if k not in ids_detalle]
            if huerfanas:
                logger.warning(
                    "OJO: %d id_detalle con glosas NO aparecen en los DETALLE (ej: %s)",
                    len(huerfanas),
                    ", ".join(huerfanas[:5]),
                )
            datos[nombre] = {
                "h_glosas": h_glosas,
                "f_glosas": f_glosas,
                "agrupado": agrupado,
                "h_det": h_det,
                "f_det": f_det,
                "glosados": glosados,
                "h_fact": h_fact,
                "f_fact": f_fact,
            }

        # --- Punto 4: base DGH filtrada y arreglada ----------------------------
        # Si la base falla (ruta mala, archivo abierto/corrupto) NO se aborta:
        # se avisa y se sigue sin cruce (SLNSERPRO sale del codigo del detalle),
        # para no perder los consolidados ni el OBJECIONES por un problema de ruta.
        cruces = None
        f_serv: list[list] = []
        base_ok = False
        if args.servicios:
            # La base pesa 80MB+: se lee UNA sola vez con las facturas de TODOS
            # los lotes; luego cada lote usa el cruce y filtra sus propias filas.
            facturas_trabajadas = {
                norm_factura(s["factura"]) for d in datos.values() for s in d["glosados"]
            }
            try:
                f_serv, cruces = cargar_base_dgh(Path(args.servicios), facturas_trabajadas)
                base_ok = True
            except ValueError as exc:
                logger.warning(
                    "\nAVISO: no se pudo usar la base DGH -> %s\n"
                    "Se genera TODO igual, pero SLNSERPRO saldra del codigo_servicio "
                    "del detalle (sin cruce DGH). Revisa la ruta del archivo.\n",
                    exc,
                )

        # --- Punto 5 + escritura: un juego de archivos POR LOTE -----------------
        # Cada lote queda en CONSOLIDADOS\LOTE XX\ con su propio OBJECIONES
        # (máx 300 facturas, que es lo que soporta el cargue de DGH). El
        # consecutivo CDCONSEC arranca en 1 dentro de cada archivo.
        resumen: list[tuple[str, int, int, int, int]] = []
        for nombre, d in datos.items():
            salida_lote = salida / nombre if nombre else salida
            salida_lote.mkdir(parents=True, exist_ok=True)
            sufijo = f" {nombre}" if nombre else ""

            logger.info("\n[4/4] %s: generando OBJECIONES...", nombre or "FINAL")
            glosados_lote = d["glosados"]
            if omitidas:
                antes = {norm_factura(s["factura"]) for s in glosados_lote}
                glosados_lote = [
                    s for s in glosados_lote if norm_factura(s["factura"]) not in omitidas
                ]
                quitadas = len(antes & omitidas)
                if quitadas:
                    logger.info(
                        "  %d facturas omitidas (ya objetadas en DGH, no van al OBJECIONES).",
                        quitadas,
                    )
            f_obj, no_cruzados, ajustados = generar_objeciones(
                glosados_lote, fecha, cruces, excluir_no_cruzados=not args.incluir_no_cruzados
            )
            n_facturas_obj = int(f_obj[-1][0]) if f_obj else 0
            logger.info("  %d filas de objeciones · %d facturas", len(f_obj), n_facturas_obj)
            if no_cruzados:
                logger.warning(
                    "  %d servicios NO cruzaron con la base DGH (van al archivo REVISAR)",
                    len(no_cruzados),
                )

            logger.info("  Escribiendo en %s ...", salida_lote)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "GLOSAS"
            escribir_hoja(ws, d["h_glosas"], d["f_glosas"])
            ws2 = wb.create_sheet("AGRUPADO")
            escribir_hoja(
                ws2,
                ["id_detalle", "OBSERVACION FINAL"],
                [[k, v["observacion"]] for k, v in d["agrupado"].items()],
            )
            wb.save(salida_lote / f"CONSOLIDADO GLOSAS{sufijo}.xlsx")

            wb = openpyxl.Workbook()
            wb.active.title = "DETALLE"
            escribir_hoja(wb.active, d["h_det"], d["f_det"])
            wb.save(salida_lote / f"CONSOLIDADO DETALLE{sufijo}.xlsx")

            if d["f_fact"]:
                wb = openpyxl.Workbook()
                wb.active.title = "FACTURAS"
                escribir_hoja(wb.active, d["h_fact"], d["f_fact"])
                wb.save(salida_lote / f"CONSOLIDADO FACTURAS{sufijo}.xlsx")

            # CONSOLIDADO RESPUESTAS GLOSAS: las respuestas predeterminadas del
            # área para TODAS las glosas del lote (RE9502 extemporánea / RE9901
            # a tiempo con el texto según tipo de glosa).
            h_rta, f_rta, res_rta = generar_respuestas_glosas(
                d["h_glosas"], d["f_glosas"], d["h_fact"], d["f_fact"], fecha
            )
            formatos_rta: list[str | None] = [None] * len(h_rta)
            for col_fecha in ("FECHA RADICACION", "FECHA GLOSA", "FECHA DE VENCIMIENTO"):
                formatos_rta[h_rta.index(col_fecha)] = "DD/MM/YYYY"
            wb = openpyxl.Workbook()
            wb.active.title = "BASE"
            escribir_hoja(wb.active, h_rta, f_rta, formatos=formatos_rta)
            pos_f2 = h_rta.index("numero_factura2")
            vistas_f2: list[str] = []
            for fila_rta in f_rta:
                f2 = norm_texto(fila_rta[pos_f2])
                if f2 and f2 not in vistas_f2:
                    vistas_f2.append(f2)
            ws_fact2 = wb.create_sheet("FACTURAS")
            escribir_hoja(ws_fact2, ["FACTURAS"], [[f2] for f2 in vistas_f2])
            wb.save(salida_lote / f"CONSOLIDADO RESPUESTAS GLOSAS{sufijo}.xlsx")
            logger.info(
                "  Respuestas a glosas: %d a tiempo (%s) · %d extemporáneas (%s).",
                res_rta["a_tiempo"],
                COD_RTA_NORMAL,
                res_rta["extemporaneas"],
                COD_RTA_EXTEMPORANEA,
            )
            if res_rta["sin_radicacion"]:
                logger.warning(
                    "  OJO: %d factura(s) sin fecha de radicación (la respuesta queda "
                    "sin código, revisar): %s",
                    len(res_rta["sin_radicacion"]),
                    ", ".join(sorted(res_rta["sin_radicacion"])[:5]),
                )
            if res_rta["sin_texto"]:
                logger.warning(
                    "  Glosas A TIEMPO que responde AUDITORÍA MÉDICA (quedan sin "
                    "código ni observación, las diligencian las doctoras): %s",
                    " · ".join(f"{t}: {n}" for t, n in res_rta["sin_texto"].items()),
                )

            if base_ok:
                # Solo las filas DGH de las facturas de ESTE lote.
                facts_lote = {norm_factura(s["factura"]) for s in d["glosados"]}
                f_serv_lote = [r for r in f_serv if norm_factura(r[0]) in facts_lote]
                wb = openpyxl.Workbook()
                wb.active.title = "SERVICIOS"
                escribir_hoja(wb.active, COLS_SERVICIOS, f_serv_lote)
                wb.save(salida_lote / f"SERVICIOS FACTURADOS COOSALUD{sufijo}.xlsx")

            # OBJECIONES: una sola hoja (como la guía, lista para el cargue).
            wb = openpyxl.Workbook()
            wb.active.title = "OBJECIONES"
            escribir_hoja(
                wb.active,
                [c for c, _ in COLS_OBJECIONES],
                f_obj,
                formatos=[f for _, f in COLS_OBJECIONES],
            )
            wb.save(salida_lote / f"OBJECIONES{sufijo}.xlsx")

            # Lo que quedó fuera / ajustado va a un archivo APARTE (no se sube
            # a DGH), solo para tu control. El OBJECIONES no lo lleva.
            if no_cruzados or ajustados:
                wb_rev = openpyxl.Workbook()
                ws_nc = wb_rev.active
                ws_nc.title = "NO_INCLUIDOS"
                escribir_hoja(
                    ws_nc,
                    [
                        "factura",
                        "id_detalle",
                        "codigo_servicio",
                        "descripcion",
                        "valor_glosado",
                        "motivo",
                        "observacion (inicio)",
                    ],
                    no_cruzados,
                )
                if ajustados:
                    ws_aj = wb_rev.create_sheet("VALOR_AJUSTADO")
                    escribir_hoja(
                        ws_aj,
                        [
                            "factura",
                            "id_detalle",
                            "SLNSERPRO",
                            "descripcion",
                            "valor_glosado_portal",
                            "valor_objetado_capado",
                            "observacion (inicio)",
                        ],
                        ajustados,
                    )
                wb_rev.save(salida_lote / f"REVISAR (no van en el cargue){sufijo}.xlsx")

            resumen.append(
                (nombre, len(f_obj), n_facturas_obj, len(no_cruzados), len(d["f_glosas"]))
            )

    except ValueError as exc:
        logger.error("ERROR: %s", exc)
        return 1

    logger.info("\n===== LISTO =====")
    if not base_ok and args.servicios:
        logger.info("  (SIN base DGH: SLNSERPRO salió del codigo_servicio del detalle)")
    for nombre, n_obj, n_fact_obj, n_nc, n_glosas in resumen:
        etiqueta = f"{nombre}\\OBJECIONES {nombre}.xlsx" if nombre else "OBJECIONES.xlsx"
        logger.info(
            "  %-38s %d filas · %d facturas · %d glosas%s",
            etiqueta,
            n_obj,
            n_fact_obj,
            n_glosas,
            f" · {n_nc} sin cruce DGH" if n_nc else "",
        )
    total_obj = sum(r[1] for r in resumen)
    total_fact = sum(r[2] for r in resumen)
    if len(resumen) > 1:
        logger.info(
            "  TOTAL: %d lotes · %d filas · %d facturas (cada OBJECIONES se sube por separado)",
            len(resumen),
            total_obj,
            total_fact,
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
