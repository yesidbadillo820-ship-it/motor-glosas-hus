#!/usr/bin/env python3
"""siifa_cruzar_tramites_dgh.py — Rellena la hoja de trabajo de SIIFA con las
respuestas que el hospital YA dio, sacándolas de la base de trámites de
Dinámica Gerencial (DGH).

EL PROBLEMA QUE RESUELVE
Muchas de las glosas que SIIFA muestra como "sin respuesta" ya fueron
contestadas por el HUS: la respuesta está registrada en DGH, pero nunca se
cargó a la plataforma del Ministerio. Volver a redactarlas sería trabajo
perdido, y además dos respuestas distintas para la misma glosa es un
problema en una conciliación.

Este script cruza las dos bases y deja la hoja de trabajo con la respuesta
ya escrita donde exista, marcando de dónde salió cada una:

    EXACTO      la misma factura, el mismo código y el mismo valor objetado.
    POR_CODIGO  la misma factura y el mismo código, pero el valor no calza
                (la respuesta sirve igual, pero conviene mirarla).
    ESCRIBIR    no hay respuesta previa: la escribe el auditor.

Las filas donde DGH tiene MÁS DE UNA respuesta distinta quedan con la más
reciente y marcadas en la columna REVISAR: ahí el auditor decide.

USO:
    py siifa_cruzar_tramites_dgh.py ^
        --informe  "D:\\...\\SIIFA\\informe_seguimientos.xlsx" ^
        --tramites "D:\\...\\TRAMITES_ENE_A_JUL_2026.xlsx" ^
        --salida   "D:\\...\\SIIFA\\respuestas_PARA_REVISAR.xlsx"

Después: revisar la hoja, completar lo que diga ESCRIBIR, y seguir con
tools/siifa_preparar_respuestas.py --expandir para dejarla lista para el bot.

La base de trámites sale de DGH (hoja «BD TRAMITE») y debe traer, por lo
menos, las columnas FACTURA, el código del concepto objetado, el valor
objetado, codigo_respuesta y observacion_respuesta.

INSTALACIÓN (una vez):
    py -m pip install openpyxl
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from siifa_preparar_respuestas import (  # noqa: E402
    agrupar,
    escribir_plantilla,
    leer_informe,
)

logger = logging.getLogger("siifa_cruce")

# Nombres posibles de cada columna en el export de DGH. El export trae rutas
# larguísimas del estilo «ListadoConceptos.DetalleRecepcionObjecion.…», y no
# siempre salen iguales; se busca por el final del nombre.
COLUMNAS_DGH = {
    "factura": ("factura",),
    "codigo": ("conceptoobjecion.codigo", "concepto.codigo", "codigo_glosa"),
    "valor": ("detallerecepcionobjecion.valorobjecion", "valorobjecion"),
    "codigo_respuesta": ("codigo_respuesta",),
    "observacion_respuesta": ("observacion_respuesta",),
    "fecha_respuesta": ("fecha_respuesta",),
    "estado": ("estadoactual",),
}


def normalizar_factura(valor) -> str:
    """HUS0000437197 (DGH) y HUS437197 (SIIFA) son la misma factura.

    DGH rellena con ceros a la izquierda y SIIFA no. Sin esto, el cruce da
    cero coincidencias aunque las dos bases hablen de lo mismo.
    """
    return re.sub(r"\D", "", str(valor or "")).lstrip("0")


def _entero(valor) -> int | None:
    try:
        return int(float(valor))
    except (TypeError, ValueError):
        return None


def fecha_dia(valor) -> str:
    """La fecha en que el hospital respondió, como AAAA-MM-DD.

    Es la fecha que hay que digitar en SIIFA: la del día en que el HUS
    contestó de verdad, no la de hoy. Si se sube con la fecha de hoy, en el
    histórico de SIIFA la respuesta queda registrada meses después de la
    glosa —es decir, fuera del término— y eso es lo primero que mira la EPS
    en una conciliación.
    """
    if valor in (None, ""):
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")
    texto = str(valor).strip()
    for trozo in (texto, texto[:10]):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
            try:
                return datetime.strptime(trozo, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    logger.warning("Fecha de respuesta de DGH no reconocida: %r", texto)
    return ""


def _ubicar_columnas(columnas: list[str]) -> dict[str, int]:
    limpias = [str(c or "").strip().lower() for c in columnas]
    posiciones: dict[str, int] = {}
    for campo, sufijos in COLUMNAS_DGH.items():
        for i, nombre in enumerate(limpias):
            if any(nombre.endswith(s) or nombre == s for s in sufijos):
                posiciones[campo] = i
                break
    faltan = {"factura", "codigo", "codigo_respuesta"} - set(posiciones)
    if faltan:
        raise SystemExit(
            f"\nERROR: al archivo de trámites le faltan columnas: {', '.join(sorted(faltan))}.\n"
            "Debe ser el export de DGH con la hoja «BD TRAMITE».\n"
        )
    return posiciones


def leer_tramites(ruta: Path, facturas: set[str], hoja: str = "BD TRAMITE") -> list[dict]:
    """Lee el export de DGH quedándose SOLO con las facturas que interesan.

    El export real pesa 74 MB y trae medio millón de líneas de todas las EPS
    y todos los meses. Filtrar mientras se lee es lo que permite trabajarlo
    en el equipo del auditor sin que se quede sin memoria.
    """
    from openpyxl import load_workbook

    wb = load_workbook(ruta, read_only=True, data_only=True)
    if hoja not in wb.sheetnames:
        if len(wb.sheetnames) != 1:
            raise SystemExit(
                f"\nERROR: el archivo {ruta.name} no tiene la hoja «{hoja}».\n"
                f"Hojas que trae: {', '.join(wb.sheetnames)}\n"
            )
        hoja = wb.sheetnames[0]

    filas = wb[hoja].iter_rows(values_only=True)
    pos = _ubicar_columnas(list(next(filas)))
    tramites = []
    leidas = 0
    for f in filas:
        leidas += 1
        factura = normalizar_factura(f[pos["factura"]])
        if factura not in facturas:
            continue
        tramites.append(
            {
                "factura": factura,
                "codigo": str(f[pos["codigo"]] or "").strip(),
                "valor": _entero(f[pos["valor"]]) if "valor" in pos else None,
                "codigo_respuesta": str(f[pos["codigo_respuesta"]] or "").strip(),
                "observacion_respuesta": str(f[pos["observacion_respuesta"]] or "").strip()
                if "observacion_respuesta" in pos
                else "",
                "fecha_respuesta": fecha_dia(f[pos["fecha_respuesta"]])
                if "fecha_respuesta" in pos
                else "",
                "estado": str(f[pos["estado"]] or "") if "estado" in pos else "",
            }
        )
    logger.info(
        "Trámites de DGH: %d líneas leídas, %d son de las facturas que están en SIIFA.",
        leidas,
        len(tramites),
    )
    return [t for t in tramites if t["codigo_respuesta"]]


def indexar(tramites: list[dict]) -> tuple[dict, dict]:
    """Dos índices: el fino (con valor) y el grueso (solo factura+código)."""
    fino: dict = defaultdict(list)
    grueso: dict = defaultdict(list)
    for t in tramites:
        fino[(t["factura"], t["codigo"], t["valor"])].append(t)
        grueso[(t["factura"], t["codigo"])].append(t)
    return fino, grueso


def _elegir(candidatos: list[dict]) -> tuple[dict, bool]:
    """De varias respuestas previas, la más reciente. Avisa si hubo empate.

    Que DGH tenga dos respuestas distintas para la misma glosa no es raro
    (se corrigió, o la respondieron dos personas). Se toma la última, pero
    el auditor tiene que verlo: por eso el aviso.
    """
    distintas = {(c["codigo_respuesta"], c["observacion_respuesta"]) for c in candidatos}
    if len(distintas) == 1:
        return candidatos[0], False
    return max(candidatos, key=lambda c: c["fecha_respuesta"]), True


def respuesta_para_grupo(lineas: list[dict], fino: dict, grueso: dict) -> dict | None:
    """Busca en DGH la respuesta que ya se dio para estas líneas de SIIFA."""
    factura = normalizar_factura(lineas[0].get("numero_factura"))
    codigo = str(lineas[0].get("codigo_glosa") or "").strip()

    candidatos: list[dict] = []
    origen = "EXACTO"
    for linea in lineas:
        hallados = fino.get((factura, codigo, _entero(linea.get("valor_glosa"))))
        if hallados:
            candidatos.extend(hallados)
    if not candidatos:
        candidatos = grueso.get((factura, codigo), [])
        origen = "POR_CODIGO"
    if not candidatos:
        return None

    elegida, hubo_empate = _elegir(candidatos)
    avisos = []
    if hubo_empate:
        avisos.append("DGH tiene más de una respuesta para esta glosa: quedó la más reciente")
    if origen == "POR_CODIGO":
        avisos.append("el valor objetado no calza con el de DGH")
    if not elegida["fecha_respuesta"]:
        avisos.append("DGH no trae la fecha en que se respondió: revisar antes de subirla")
    return {
        "CODIGO_RESPUESTA": elegida["codigo_respuesta"],
        "OBSERVACION_RESPUESTA": elegida["observacion_respuesta"],
        "FECHA_RESPUESTA": elegida["fecha_respuesta"],
        "ORIGEN": origen,
        "REVISAR": "; ".join(avisos) or None,
    }


def cruzar(grupos: dict[str, list[dict]], fino: dict, grueso: dict) -> dict[str, dict]:
    previas: dict[str, dict] = {}
    for clave, lineas in grupos.items():
        encontrada = respuesta_para_grupo(lineas, fino, grueso)
        if encontrada:
            previas[clave] = encontrada
    return previas


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--informe", required=True, help="Informe de siifa_reporte_seguimientos.py.")
    ap.add_argument("--tramites", required=True, help="Export de DGH (hoja BD TRAMITE).")
    ap.add_argument("--salida", required=True, help="Hoja de trabajo que queda pre-llenada.")
    ap.add_argument("--tipo", choices=["GLOSA", "DEVOLUCION"], help="Trabajar solo un tipo.")
    ap.add_argument("--factura", help="Trabajar solo una factura.")
    ap.add_argument("--verbose", action="store_true")
    args = ap.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s  %(message)s",
        datefmt="%H:%M:%S",
        stream=sys.stdout,
    )

    datos = leer_informe(Path(args.informe))
    grupos = agrupar(datos, tipo=args.tipo, factura=args.factura)
    if not grupos:
        raise SystemExit("\nNo quedó nada pendiente con esos filtros. Nada que hacer.\n")
    lineas = sum(len(v) for v in grupos.values())
    logger.info("SIIFA: %d líneas pendientes, agrupadas en %d respuestas.", lineas, len(grupos))

    facturas = {normalizar_factura(x[0].get("numero_factura")) for x in grupos.values()}
    logger.info("Buscando en DGH las respuestas de %d facturas (esto tarda)...", len(facturas))
    fino, grueso = indexar(leer_tramites(Path(args.tramites), facturas))

    previas = cruzar(grupos, fino, grueso)
    escribir_plantilla(grupos, Path(args.salida), ya_respondidas=previas)

    cubiertas = sum(len(grupos[c]) for c in previas)
    por_origen = Counter(v["ORIGEN"] for v in previas.values())
    a_revisar = sum(1 for v in previas.values() if v["REVISAR"])
    print(
        f"\nRESULTADO DEL CRUCE\n"
        f"  respuestas que ya existían en DGH : {len(previas):4} de {len(grupos)}"
        f"  ({cubiertas} de {lineas} líneas)\n"
        f"      cruce exacto (factura+código+valor) : {por_origen['EXACTO']:4}\n"
        f"      solo por factura+código            : {por_origen['POR_CODIGO']:4}\n"
        f"  hay que escribirlas de cero       : {len(grupos) - len(previas):4}\n"
        f"  marcadas para revisar             : {a_revisar:4}\n"
        f"\nHoja de trabajo: {args.salida}\n"
        f"Revise primero las filas con algo escrito en REVISAR, después complete "
        f"las que dicen ESCRIBIR, y siga con:\n"
        f'   py tools\\siifa_preparar_respuestas.py --informe "{args.informe}" '
        f'--expandir "{args.salida}" --salida "respuestas_siifa.xlsx"\n'
    )


if __name__ == "__main__":
    main()
