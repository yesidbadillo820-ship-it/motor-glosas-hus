"""responder_glosas_coosalud.py — Respuesta masiva de glosas en el portal COOSALUD (vco.ctamedicas.com).

Para cada factura del Excel consolidado (hoja BASE), agrupa las glosas por
(código de respuesta + justificación) y usa el "Responder Masivamente" del
portal: marca los checkboxes del grupo, abre el modal, selecciona el código,
pega la justificación, adjunta el PDF PDX si el grupo es de SOPORTES, confirma,
y al final le da "Terminar Respuesta" capturando el pantallazo de evidencia
("¡Usted ha cerrado una cuenta!") en la carpeta EVIDENCIA.

FLUJO POR FACTURA:
  1. Bolsa de Respuestas → Buscar la factura → botón azul ▶.
  2. Sección GLOSAS → Mostrar: Todos.
  3. Por cada grupo de respuesta del Excel (saltando las ya RESPONDIDA):
     checkboxes → Responder Masivamente → modal (código + justificación
     [+ PDF PDX para SOPORTES]) → Responder Glosa → Continuar.
  4. Cuando no quedan SIN RESPUESTA → Terminar Respuesta → "Sí, Terminar!"
     → pantallazo del cartel "¡Usted ha cerrado una cuenta!" → Continuar.

REGLA DE SOPORTES (verificada contra el Excel, 0 excepciones):
  grupo con tipo_glosa == SOPORTES ⟺ justificación contiene "ANEXA SOPORTE"
  → se adjunta el PDF tipificado PDX_*.pdf de la carpeta de la factura
  (ubicada vía el índice TXT). Si no se encuentra el PDX, el grupo se SALTA y
  la factura queda sin terminar (reportada como PENDIENTE_PDX) para no
  prometer un soporte que no se adjuntó.

CONCEPTOS DE CALIDAD (PERTINENCIA): NO se responden — la hoja CALIDAD del
Excel se ignora (sólo se lee la hoja BASE).

CREDENCIALES (variables de entorno, NO en el código):
    setx COOSALUD_USER <usuario>
    setx COOSALUD_PASSWORD <contraseña>
    (cerrar y reabrir la terminal para que tomen efecto)

USO:
    REM Piloto: una factura, browser visible
    py responder_glosas_coosalud.py ^
        --excel  "D:\\...\\CONSOLIDADO COOSALUD DIA 28.xlsx" ^
        --indice "D:\\USUARIO CARTERA\\Desktop\\BUSCADOR_HUS\\indice_facturas_HUS.txt" ^
        --solo HUS496197 --con-cabeza

    REM Masivo (140 facturas de la hoja BASE)
    py responder_glosas_coosalud.py ^
        --excel ... --indice ... --todas ^
        --reporte "D:\\...\\reporte_coosalud.csv"

INSTALACIÓN (una vez):
    py -m pip install playwright openpyxl
    py -m playwright install chromium
"""

from __future__ import annotations

import argparse
import csv
import logging
import os
import re
import sys
import time
import unicodedata
from collections import defaultdict
from pathlib import Path

try:
    from playwright.sync_api import (
        Page,
        TimeoutError as PlaywrightTimeout,
        sync_playwright,
    )
except ImportError:
    Page = object  # type: ignore[assignment,misc]
    PlaywrightTimeout = Exception  # type: ignore[assignment,misc]
    sync_playwright = None  # type: ignore[assignment]


def _exigir_playwright() -> None:
    if sync_playwright is None:
        sys.stderr.write(
            "ERROR: falta playwright.\n"
            "    py -m pip install playwright\n"
            "    py -m playwright install chromium\n"
        )
        sys.exit(2)


PORTAL_BASE = "https://vco.ctamedicas.com"
PORTAL_LOGIN = f"{PORTAL_BASE}/app/login"
PORTAL_HOME = f"{PORTAL_BASE}/app/inicio"
PORTAL_BOLSA = f"{PORTAL_BASE}/app/respuestaGlosaSearch"

# Texto estándar para cerrar glosas que están en el portal pero no en el
# Excel del HUS (residuales). Se usa SOLO con --cerrar-residuales y se
# adjunta con código RE9901 (el mismo que el bot usa por default para
# PERTINENCIA del Excel). Cubre el motivo más común: el HUS no acepta la
# glosa por pertinencia y solicita el levantamiento integro.
TEXTO_GENERICO_RESIDUALES = (
    "ESE HUS NO ACEPTA LA GLOSA POR CONCEPTO DE PERTINENCIA INTERPUESTA POR "
    "COOSALUD SOBRE LOS SERVICIOS EN MENCION. LOS SERVICIOS PRESTADOS A ESTE "
    "USUARIO POR LA ESE HUS FUERON PERTINENTES SEGUN EL DIAGNOSTICO Y LA "
    "EVOLUCION CLINICA DEL PACIENTE, SOPORTADOS EN LA HISTORIA CLINICA QUE "
    "REPOSA EN LA ESE HUS. SE SOLICITA EL LEVANTAMIENTO DE LA GLOSA Y EL PAGO "
    "DE LO FACTURADO."
)
COD_GENERICO_RESIDUALES = "RE9901"
PORTAL_PAUSA = f"{PORTAL_BASE}/app/respuestaGlosaPause"

MAX_PDF_MB = 10

logger = logging.getLogger("responder_coosalud")


# ─── Setup ──────────────────────────────────────────────────────────────────


def setup_logging(log_file: Path | None = None) -> None:
    fmt = "%(asctime)s [%(levelname)s] %(message)s"
    handlers: list[logging.Handler] = [logging.StreamHandler(sys.stdout)]
    if log_file is not None:
        log_file.parent.mkdir(parents=True, exist_ok=True)
        handlers.append(logging.FileHandler(log_file, encoding="utf-8"))
    logging.basicConfig(level=logging.INFO, format=fmt, handlers=handlers)


def cargar_credenciales() -> tuple[str, str]:
    user = os.environ.get("COOSALUD_USER", "").strip()
    password = os.environ.get("COOSALUD_PASSWORD", "").strip()
    if not user or not password:
        sys.stderr.write(
            "ERROR: faltan credenciales. Setealas con:\n"
            "    setx COOSALUD_USER <usuario>\n"
            "    setx COOSALUD_PASSWORD <contraseña>\n"
            "Despues cerra y reabri la terminal.\n"
        )
        sys.exit(2)
    return user, password


# ─── Lectura del Excel consolidado (hoja BASE) ──────────────────────────────


def _norm_header(h: str) -> str:
    s = unicodedata.normalize("NFKD", (h or "").strip().upper())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.split())


COLUMNAS = {
    "id_glosa": {"ID_GLOSA", "ID GLOSA"},
    "factura": {"NUMERO_FACTURA", "NUMERO FACTURA"},
    "tipo": {"TIPO_GLOSA", "TIPO GLOSA"},
    "cod_rta": {"COD RESPUESTA GLOSA", "CODIGO RESPUESTA GLOSA"},
    "obs_rta": {"OBSERVACION RTA GLOSA", "OBSERVACION RESPUESTA GLOSA"},
}

# Códigos de respuesta que NO requieren adjuntar soporte, aunque la glosa
# original sea de tipo_glosa == SOPORTES. RE9502 = glosa extemporánea
# (aceptación tácita por radicación fuera de términos): la glosa se rechaza
# por vencimiento del plazo legal, no por falta de soporte clínico, así que
# el portal no espera ningún PDF. Sin esto, esas glosas quedaban PENDIENTE_PDX
# buscando un soporte que no corresponde.
CODIGOS_SIN_SOPORTE = {"RE9502"}


def leer_excel(ruta: Path, hoja: str, incluir_calidad: bool = False) -> dict[str, dict]:
    """Devuelve {factura: {"grupos": [...], "calidad": N}} donde cada grupo es
    {cod, cod_corto, obs, ids: [id_glosa...], es_soporte: bool}.

    Por defecto, las glosas con tipo_glosa == CALIDAD (pertinencia) NO se
    responden: se excluyen de los grupos y sólo se cuentan en "calidad" — esas
    facturas quedan abiertas (sin Terminar) para el equipo médico.

    Con incluir_calidad=True las glosas CALIDAD SÍ se incluyen en los grupos
    a responder (usando su COD_RTA y OBSERVACION_RTA del Excel, igual que las
    otras). Útil cuando la planilla ya viene con la respuesta tipificada para
    pertinencia y se quiere cerrar la factura completa en una sola corrida.

    Los grupos preservan el orden de aparición en el Excel."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.stderr.write("ERROR: falta openpyxl. py -m pip install openpyxl\n")
        sys.exit(2)

    wb = load_workbook(filename=str(ruta), data_only=True, read_only=True)
    # Match del nombre de hoja tolerante a mayus/minus y espacios: los
    # Excel exportados por Power Query suelen traer nombres inconsistentes
    # ('BASE ' con espacio al final, 'CALIDAD', etc.). Asi '--hoja BASE'
    # encuentra la hoja 'BASE ' sin que el usuario tenga que adivinar el
    # nombre exacto con el espacio.
    if hoja not in wb.sheetnames:
        objetivo = hoja.strip().casefold()
        real = next((s for s in wb.sheetnames if s.strip().casefold() == objetivo), None)
        if real is None:
            raise ValueError(f"El Excel no tiene la hoja '{hoja}'. Hojas: {wb.sheetnames}")
        if real != hoja:
            logger.info(f"  Hoja '{hoja}' → uso '{real}' (match tolerante a espacios/mayusculas).")
        hoja = real
    ws = wb[hoja]
    rows = ws.iter_rows(values_only=True)
    headers = [_norm_header(str(h or "")) for h in next(rows)]

    idx: dict[str, int] = {}
    for clave, opciones in COLUMNAS.items():
        for i, h in enumerate(headers):
            if h in opciones:
                idx[clave] = i
                break
        if clave not in idx:
            raise ValueError(
                f"Falta la columna '{clave}' (acepto {sorted(opciones)}). Headers: {headers}"
            )

    # factura → lista ordenada de (clave_grupo → grupo)
    grupos_por_fac: dict[str, dict] = defaultdict(dict)
    calidad_por_fac: dict[str, int] = defaultdict(int)
    for r in rows:
        if r is None:
            continue
        fac = str(r[idx["factura"]] or "").strip()
        id_glosa = str(r[idx["id_glosa"]] or "").strip()
        if not fac or not id_glosa:
            continue
        tipo = str(r[idx["tipo"]] or "").strip().upper()
        if tipo == "CALIDAD" and not incluir_calidad:
            calidad_por_fac[fac] += 1
            continue  # pertinencia: NO se responde
        cod = str(r[idx["cod_rta"]] or "").strip()
        obs = str(r[idx["obs_rta"]] or "").strip()
        if not cod or not obs:
            # Sin respuesta definida: si era CALIDAD se cuenta como pendiente
            # de equipo médico; las otras se ignoran.
            if tipo == "CALIDAD":
                calidad_por_fac[fac] += 1
            continue
        key = (cod, obs)
        g = grupos_por_fac[fac].get(key)
        if g is None:
            m = re.match(r"([A-Z]{2}\d{4})", cod)
            g = {
                "cod": cod,
                "cod_corto": m.group(1) if m else cod[:6],
                "obs": obs,
                "ids": [],
                "es_soporte": False,
            }
            grupos_por_fac[fac][key] = g
        g["ids"].append(id_glosa)
        # Solo exigimos soporte si la glosa es de tipo SOPORTES Y el código de
        # respuesta efectivamente lo requiere. Para RE9502 (extemporánea) no se
        # sube soporte, así que no marcamos es_soporte y la glosa se responde y
        # cierra normal en vez de quedar PENDIENTE_PDX.
        if tipo == "SOPORTES" and g["cod_corto"] not in CODIGOS_SIN_SOPORTE:
            g["es_soporte"] = True

    todas = set(grupos_por_fac) | set(calidad_por_fac)
    return {
        fac: {
            "grupos": list(grupos_por_fac.get(fac, {}).values()),
            "calidad": calidad_por_fac.get(fac, 0),
        }
        for fac in todas
    }


# ─── Índice de soportes (factura → carpeta del share) ───────────────────────


def cargar_indice(ruta: Path) -> dict[str, Path]:
    mapa: dict[str, Path] = {}
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            texto = ruta.read_text(encoding=enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise RuntimeError(f"No pude decodificar el indice: {ruta}")
    for linea in texto.splitlines():
        linea = linea.strip()
        m = re.search(r"\\(HUS\d+)\s*$", linea, re.IGNORECASE)
        if m:
            mapa[m.group(1).upper()] = Path(linea)
    return mapa


# Prefijos de archivo aceptados como soporte, en orden de preferencia.
# COOSALUD acepta cualquier PDF tipificado del envío del HUS — preferimos PDX
# (el que el HUS arma específicamente como respuesta de soportes), pero si no
# está usamos HAM o PDE que también son válidos como soporte clínico.
PREFIJOS_SOPORTE = ("PDX", "HAM", "PDE")


def _archivos_con_prefijo(carpeta: Path, prefijos: tuple[str, ...]) -> list[tuple[str, Path]]:
    """Devuelve [(prefijo, path)] para los archivos que matcheen <PREFIJO>*.pdf
    en la carpeta, respetando el ORDEN de `prefijos` (primero los PDX, después
    los HAM, etc.)."""
    out: list[tuple[str, Path]] = []
    for prefijo in prefijos:
        encontrados: set[Path] = set()
        for pat in (
            f"{prefijo}*.pdf",
            f"{prefijo}*.PDF",
            f"{prefijo.lower()}*.pdf",
            f"{prefijo.capitalize()}*.pdf",
        ):
            for p in carpeta.glob(pat):
                encontrados.add(p)
        for p in sorted(encontrados):
            out.append((prefijo, p))
    return out


def _pdx_en(carpeta: Path) -> list[Path]:
    """Compat: devuelve solo paths PDX (sin tupla). Mantengo por si lo usa
    código viejo, pero `buscar_pdx` ya usa `_archivos_con_prefijo`."""
    return [p for _, p in _archivos_con_prefijo(carpeta, ("PDX",))]


RE_ENV_NUM = re.compile(r"^(ENV-\d+)-", re.IGNORECASE)


def _carpetas_alternativas(carpeta: Path) -> list[Path]:
    """En el share del HUS los PDFs no siempre están en la misma carpeta del
    RIPS (donde viven los JSONs). Patrones observados:
        ...\\KARIN\\RIPS\\ENV-227237-OK\\HUS508259      ← índice (JSONs)
        ...\\KARIN\\ENV-227237-OK-C-DGH\\IMG\\HUS...   ← hermana (caso A)
        ...\\VANESSA\\RIPS\\ENV-226686-OK\\HUS500258   ← índice (JSONs)
        ...\\VANESSA\\ENV-226686-R-C\\IMG\\HUS...       ← hermana (caso B)
    En el caso B la hermana NO arranca con `ENV-XXX-OK` sino directo con
    `ENV-XXX-R-C` (sufijo distinto al ENV original). Por eso matcheamos
    por el NÚMERO del envío (`ENV-NNNN-`), no por el nombre completo.
    Devuelve las carpetas hermanas que coinciden con `ENV-NNNN-*` y que
    tengan adentro `/IMG/<factura>/` o `/<factura>/`."""
    factura = carpeta.name
    env_dir = carpeta.parent  # ENV-XXX-OK
    if not env_dir.name.upper().startswith("ENV"):
        return []
    # Extraer el "ENV-<numero>-" para matchear hermanas con cualquier sufijo
    # (no solo el sufijo del env_dir original).
    m = RE_ENV_NUM.match(env_dir.name)
    if not m:
        return []
    env_base = m.group(1).upper() + "-"  # "ENV-226686-"
    alternativas: list[Path] = []
    raices: list[Path] = [env_dir.parent]
    abuelo = env_dir.parent.parent
    if abuelo and abuelo != env_dir.parent:
        raices.append(abuelo)
    vistos: set[Path] = set()
    for raiz in raices:
        if not raiz.is_dir():
            continue
        try:
            vecinas = list(raiz.iterdir())
        except OSError:
            continue
        for vecina in vecinas:
            try:
                if not vecina.is_dir() or vecina == env_dir:
                    continue
                # Hermana del envío: arranca con `ENV-<numero>-`
                if not vecina.name.upper().startswith(env_base):
                    continue
                for sub in (vecina / "IMG" / factura, vecina / factura):
                    try:
                        if sub.is_dir() and sub not in vistos:
                            vistos.add(sub)
                            alternativas.append(sub)
                    except OSError:
                        continue
            except OSError:
                continue
    return alternativas


def buscar_pdx(factura: str, indice: dict[str, Path]) -> tuple[Path | None, str]:
    """Busca un PDF de soporte para la factura, probando los prefijos en
    orden: PDX (preferido), HAM, PDE. Si la carpeta del índice no tiene PDFs
    (solo RIPS/JSONs), prueba la carpeta hermana de IMG/.
    Devuelve (ruta o None, detalle)."""
    carpeta = indice.get(factura.upper())
    if carpeta is None:
        return None, f"{factura} no está en el índice"
    if not carpeta.is_dir():
        return None, f"carpeta no accesible: {carpeta}"

    candidatos = _archivos_con_prefijo(carpeta, PREFIJOS_SOPORTE)
    if not candidatos:
        # rglob como red de seguridad por si el PDF cuelga de un subdir
        # dentro de la propia carpeta del índice.
        for prefijo in PREFIJOS_SOPORTE:
            for p in sorted(carpeta.rglob(f"{prefijo}*.pdf")):
                candidatos.append((prefijo, p))
            if candidatos:
                break
    if not candidatos:
        # Fallback: carpeta hermana (RIPS → IMG).
        for alt in _carpetas_alternativas(carpeta):
            cand_alt = _archivos_con_prefijo(alt, PREFIJOS_SOPORTE)
            if cand_alt:
                candidatos = cand_alt
                prefijo_usado = cand_alt[0][0]
                if prefijo_usado == "PDX":
                    logger.info(f"  PDX no estaba en {carpeta.name}; lo encontré en hermana IMG/")
                else:
                    logger.info(f"  PDX no estaba; uso {prefijo_usado} de hermana IMG/")
                break
    if not candidatos:
        nombres_aceptados = "/".join(PREFIJOS_SOPORTE)
        pdfs = sorted(carpeta.glob("*.pdf")) + sorted(carpeta.glob("*.PDF"))
        if pdfs:
            muestras = ", ".join(p.name for p in pdfs[:6])
            extra = "" if len(pdfs) <= 6 else f" (+{len(pdfs) - 6} más)"
            return None, (
                f"sin {nombres_aceptados}*.pdf en {carpeta} | PDFs presentes: {muestras}{extra}"
            )
        return None, f"sin {nombres_aceptados}*.pdf en {carpeta} (ni en carpeta hermana IMG/)"

    prefijo_usado, pdf = candidatos[0]
    mb = pdf.stat().st_size / (1024 * 1024)
    if mb > MAX_PDF_MB:
        return None, f"{pdf.name} pesa {mb:.1f}MB (límite {MAX_PDF_MB}MB)"
    # Avisar cuando NO es el prefijo preferido (PDX) para que quede en el
    # log y vos puedas auditar después qué facturas se cerraron con HAM/PDE.
    if prefijo_usado != "PDX":
        logger.info(f"  ⚠ usando {prefijo_usado} como soporte (no había PDX): {pdf.name}")
    elif len(candidatos) > 1:
        logger.warning(f"  ({factura}: {len(candidatos)} candidatos, uso {pdf.name})")
    return pdf, pdf.name


# ─── Browser: sesión y navegación ───────────────────────────────────────────


def _screenshot_debug(page: Page, etiqueta: str) -> None:
    out = Path("debug_screenshots")
    out.mkdir(exist_ok=True)
    ruta = out / f"{time.strftime('%H%M%S')}_{etiqueta}.png"
    try:
        page.screenshot(path=str(ruta), full_page=True)
        logger.info(f"  Screenshot de diagnóstico: {ruta}")
    except Exception as e:
        logger.warning(f"  No pude tomar screenshot: {e}")


def login(page: Page, user: str, password: str) -> None:
    logger.info("Login al portal COOSALUD...")
    try:
        page.goto(PORTAL_LOGIN, wait_until="domcontentloaded", timeout=90000)
    except PlaywrightTimeout:
        # El portal a veces tarda en disparar DOMContentLoaded; con que la
        # navegación haya arrancado alcanza — los waits posteriores esperan
        # los elementos concretos.
        page.goto(PORTAL_LOGIN, wait_until="commit", timeout=90000)
    page.wait_for_timeout(1500)
    # Si ya hay sesión (el portal redirige al home con el menú), no hay login.
    if page.locator("text=Respuesta Glosas").count() > 0:
        logger.info("Sesión ya activa.")
        return
    user_input = page.locator(
        "input[type='text']:visible, input[type='email']:visible, "
        "input[name*='user' i]:visible, input[name*='usuario' i]:visible"
    ).first
    pwd_input = page.locator("input[type='password']:visible").first
    user_input.fill(user)
    pwd_input.fill(password)
    boton = page.locator(
        "button[type='submit']:visible, input[type='submit']:visible, "
        "button:has-text('Ingresar'):visible, button:has-text('Iniciar'):visible, "
        "button:has-text('Entrar'):visible, button:has-text('Login'):visible"
    ).first
    try:
        boton.click(timeout=4000)
    except PlaywrightTimeout:
        pwd_input.press("Enter")
    # Señal de adentro: el menú lateral con "Respuesta Glosas".
    page.wait_for_selector("text=Respuesta Glosas", timeout=20000)
    logger.info("Login OK")


class FacturaNoEnBolsa(Exception):
    """La factura no aparece en la Bolsa de Respuestas (ya cerrada / no glosada)."""


# Latch global: una vez que cargamos la Bolsa por primera vez (paciente, por
# el menú), las siguientes navegaciones usan el goto directo, que es mucho
# más rápido en la sesión ya "calentada".
_BOLSA_VISITADA = False


def _goto_paciente(page: Page, url: str, intentos: int = 3) -> None:
    """goto con reintentos: el portal tiene rachas de minutos enteros sin
    responder (visto en producción: 2 min sin contestar /app/inicio). Cada
    intento prueba domcontentloaded 60s y, si no, commit 60s más."""
    for n in range(1, intentos + 1):
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=60000)
            return
        except PlaywrightTimeout:
            pass
        try:
            page.goto(url, wait_until="commit", timeout=60000)
            return
        except PlaywrightTimeout:
            if n == intentos:
                raise
            logger.info(f"  el portal no responde; reintento navegación ({n}/{intentos})…")
            page.wait_for_timeout(3000)


def _esta_en_bolsa(page: Page) -> bool:
    try:
        return (
            "respuestaGlosaSearch" in (page.url or "")
            and page.locator("text=FILTROS BOLSA RESPUESTA").count() > 0
        )
    except Exception:
        return False


def ir_a_bolsa(page: Page) -> None:
    """Navega a Respuesta Glosas → Bolsa Respuesta.

    PRIMERA vez en la sesión: paciente. Vamos al home y clickeamos el menú
    lateral, esperando hasta 90s a que el datatable termine de cargar (el
    usuario confirmó que esa carga inicial es lenta).
    SIGUIENTES veces: goto directo a /respuestaGlosaSearch, que sí responde
    rápido una vez que la página ya corrió la primera vez en la sesión.
    """
    global _BOLSA_VISITADA
    if _esta_en_bolsa(page):
        return

    if _BOLSA_VISITADA:
        # Camino rápido (sesión ya calentada).
        _goto_paciente(page, PORTAL_BOLSA)
        page.wait_for_selector("text=FILTROS BOLSA RESPUESTA", timeout=60000)
        page.wait_for_timeout(500)
        return

    # Primera vez: por el menú, con paciencia.
    _goto_paciente(page, PORTAL_HOME)
    page.wait_for_selector("text=Respuesta Glosas", timeout=30000)
    try:
        page.locator(
            "xpath=//a[normalize-space()='Respuesta Glosas'] | "
            "//span[normalize-space()='Respuesta Glosas']/ancestor::a[1]"
        ).first.click(timeout=5000)
    except PlaywrightTimeout:
        page.get_by_text("Respuesta Glosas").first.click()
    page.wait_for_timeout(500)
    sub = page.locator(
        "xpath=//a[contains(normalize-space(), 'Bolsa')] | "
        "//span[contains(normalize-space(), 'Bolsa')]/ancestor::a[1]"
    ).first
    # no_wait_after: el click dispara una navegación LENTA (la carga inicial
    # del datatable); si el click la espera, revienta su propio timeout.
    sub.click(timeout=10000, no_wait_after=True)
    # La PRIMERA carga del datatable puede tardar más de un minuto.
    page.wait_for_selector("text=FILTROS BOLSA RESPUESTA", timeout=120000)
    page.wait_for_timeout(1500)
    _BOLSA_VISITADA = True


def ir_a_en_pausa(page: Page) -> None:
    """Navega a Respuesta Glosas → En Pausa (las respuestas abiertas a medias).

    Cuando una respuesta se abre y no se Termina, la factura sale de la Bolsa
    y aparece acá. Para cuando llegamos a esta grilla la app ya está caliente
    (siempre pasamos antes por la Bolsa), así que el goto directo alcanza."""
    _goto_paciente(page, PORTAL_PAUSA)
    # OJO: 'EN PAUSA' a secas matchea el ítem del menú lateral (oculto). El
    # título real de la página es 'RESPUESTA GLOSA EN PAUSA'; si tarda, la
    # caja Buscar (gate real) lo cubre después.
    try:
        page.wait_for_selector("text=RESPUESTA GLOSA EN PAUSA", timeout=20000)
    except PlaywrightTimeout:
        pass
    page.wait_for_timeout(800)


def _esperar_caja_buscar(page: Page, nombre_grilla: str, timeout_s: int = 180):
    """Espera (con paciencia) la caja 'Buscar:' del datatable — aparece recién
    cuando la grilla terminó su carga lenta. Devuelve el locator."""
    deadline = time.time() + timeout_s
    aviso = False
    while time.time() < deadline:
        for sel in (
            "input[type='search']",
            "xpath=//label[contains(., 'Buscar')]//input",
            "input[placeholder*='Buscar' i]",
        ):
            loc = page.locator(sel)
            try:
                if loc.count() > 0 and loc.first.is_visible():
                    return loc.first
            except Exception:
                continue
        if not aviso:
            logger.info(f"  esperando que cargue la grilla de {nombre_grilla} (es lenta)…")
            aviso = True
        page.wait_for_timeout(1500)
    _screenshot_debug(page, f"sin_caja_buscar_{nombre_grilla}")
    raise RuntimeError(
        f"La grilla de {nombre_grilla} no cargó en {timeout_s}s (sin caja 'Buscar:')."
    )


def _cerrar_cartel_continuar(page: Page, espera_s: float = 0) -> bool:
    """Cierra el cartel emergente con botón 'Continuar' si está en pantalla.
    Al abrir una cuenta desde la Bolsa el portal la ASIGNA y bloquea todo con
    '¡Se ha asignado la cuenta!' — hasta no darle Continuar la sección GLOSAS
    no se puede tocar. espera_s=0 → un solo chequeo sin esperar."""
    fin = time.time() + espera_s
    while True:
        try:
            btn = _primer_visible(page.locator("button:has-text('Continuar')"))
            if btn is not None:
                btn.click(timeout=3000)
                logger.info("  cartel de asignación → Continuar")
                page.wait_for_timeout(400)
                return True
        except Exception:
            pass
        if time.time() >= fin:
            return False
        page.wait_for_timeout(300)


def _buscar_y_entrar(page: Page, factura: str, nombre_grilla: str, timeout_s: int = 60) -> bool:
    """Busca la factura en la grilla actual y entra con el botón ▶.
    True = entró al detalle. False = la grilla dio señal de vacío o venció el
    plazo sin fila (la factura no está en ESTA grilla)."""
    buscar = _esperar_caja_buscar(page, nombre_grilla)
    # El datatable filtra con eventos de TECLADO (keyup): fill() setea el valor
    # sin teclas y la grilla nunca filtra (por eso a mano sí funcionaba). Hay
    # que tipear de verdad, y de yapa disparamos keyup por JS.
    buscar.click()
    buscar.fill("")
    buscar.type(factura, delay=40)
    try:
        buscar.evaluate("el => { el.dispatchEvent(new KeyboardEvent('keyup', {bubbles: true})); }")
    except Exception:
        pass
    page.wait_for_timeout(1200)  # debounce del datatable

    # Localización por celda EXACTA de NUMERO FACTURA (tr:has-text se confunde
    # con plantillas ocultas y con los otros números de la fila).
    def _fila_visible():
        for sel in (
            f"tbody tr:has(td:text-is('{factura}'))",
            f"tbody tr:has(td:has-text('{factura}'))",
        ):
            cand = _primer_visible(page.locator(sel))
            if cand is not None:
                return cand
        return None

    inicio = time.time()
    aviso = False
    fila = None
    while True:
        fila = _fila_visible()
        if fila is not None:
            break
        vacio = False
        try:
            vacio = page.evaluate(
                """() => {
                    const t = document.body.innerText || '';
                    return /No se encontraron|No hay (datos|registros)|ning[uú]n registro|de 0 Entradas/i.test(t);
                }"""
            )
        except Exception:
            pass
        if vacio:
            return False
        if time.time() - inicio > timeout_s:
            _screenshot_debug(page, f"{nombre_grilla}_sin_fila_{factura}")
            return False
        if not aviso and time.time() - inicio > 5:
            logger.info("  esperando resultados de la búsqueda…")
            aviso = True
        page.wait_for_timeout(400)

    # Click en el botón azul ▶ de OPCIONES (último botón/anchor de la fila).
    accion = _primer_visible(fila.locator("button")) or _primer_visible(fila.locator("a"))
    if accion is None:
        accion = _primer_visible(fila.locator("td").last.locator("button, a, [onclick]"))
    if accion is None:
        _screenshot_debug(page, f"sin_boton_play_{factura}")
        raise RuntimeError(f"{factura}: no hallé el botón ▶ en la fila de {nombre_grilla}.")
    accion.click()
    # Primera apertura desde la Bolsa → cartel "¡Se ha asignado la cuenta!".
    _cerrar_cartel_continuar(page, espera_s=3)
    # Página de detalle: aparece la sección GLOSAS.
    page.wait_for_selector("text=GLOSAS", timeout=60000)
    page.wait_for_timeout(800)
    return True


def abrir_factura(page: Page, factura: str) -> None:
    """Abre la factura buscándola primero en la Bolsa y, si no está, en
    En Pausa (las respuestas que quedaron abiertas a medias van ahí)."""
    ir_a_bolsa(page)
    if _buscar_y_entrar(page, factura, "Bolsa"):
        return
    logger.info("  no está en la Bolsa; pruebo en 'En Pausa'…")
    ir_a_en_pausa(page)
    if _buscar_y_entrar(page, factura, "En Pausa"):
        logger.info("  ▶ reanudada desde 'En Pausa'")
        return
    raise FacturaNoEnBolsa(f"{factura} no está ni en la Bolsa ni En Pausa (¿ya cerrada?).")


def preparar_grilla(page: Page) -> None:
    """Pone 'Mostrar: Todos' en la sección GLOSAS (hay facturas de 800+ glosas)."""
    candidatos = (
        "xpath=//*[contains(text(),'GLOSAS')]/following::select[1]",
        "select[name*='length']",
        "select",
    )
    for sel in candidatos:
        loc = page.locator(sel)
        try:
            n = loc.count()
        except Exception:
            continue
        for i in range(min(n, 6)):
            s = loc.nth(i)
            try:
                if not s.is_visible():
                    continue
                opciones = s.locator("option").all_inner_texts()
                if any("Todos" in o for o in opciones):
                    s.select_option(label=next(o for o in opciones if "Todos" in o))
                    page.wait_for_timeout(800)
                    return
            except Exception:
                continue
    # Algunos datatables ya vienen en 'Todos' (como en los pantallazos).
    logger.info("  (no pude setear Mostrar=Todos; sigo con lo visible)")


def leer_estados(page: Page) -> dict[str, str]:
    """Devuelve {id_glosa: estado} de la grilla GLOSAS ('SIN RESPUESTA'/'RESPONDIDA')."""
    estados: dict[str, str] = {}
    datos = page.evaluate(
        """() => {
            const out = [];
            for (const tr of document.querySelectorAll('tr')) {
                const tds = tr.querySelectorAll('td');
                if (tds.length < 5) continue;
                const id = (tds[0].innerText || '').trim();
                if (!/^\\d{6,}$/.test(id)) continue;
                const texto = (tr.innerText || '').toUpperCase();
                let estado = '';
                if (texto.includes('SIN RESPUESTA')) estado = 'SIN RESPUESTA';
                else if (texto.includes('RESPONDIDA')) estado = 'RESPONDIDA';
                if (estado) out.push([id, estado]);
            }
            return out;
        }"""
    )
    for id_glosa, estado in datos:
        estados[id_glosa] = estado
    return estados


def esperar_glosas(page: Page, timeout_s: int = 120) -> dict[str, str]:
    """Espera a que la grilla GLOSAS tenga filas y se ESTABILICE (la página de
    la cuenta carga por partes: primero GESTION CUENTA, después GLOSAS; y tras
    'Mostrar: Todos' el datatable redibuja). Mientras tanto puede saltar el
    cartel de asignación → lo cierra. Si nunca aparecen filas, aborta SIN
    tocar nada (jamás dar Terminar con la grilla vacía)."""
    inicio = time.time()
    aviso = False
    previo = -1
    estables = 0
    while True:
        _cerrar_cartel_continuar(page)
        estados = leer_estados(page)
        n = len(estados)
        if n > 0:
            if n == previo:
                estables += 1
                if estables >= 2:  # 3 lecturas iguales seguidas → grilla quieta
                    return estados
            else:
                estables = 0
            previo = n
            if time.time() - inicio > timeout_s:
                return estados  # mejor esfuerzo: hay filas aunque siga moviéndose
        else:
            if time.time() - inicio > timeout_s:
                _screenshot_debug(page, "glosas_no_cargo")
                raise RuntimeError(
                    f"la grilla GLOSAS no mostró filas en {timeout_s}s — no toco "
                    "nada (la cuenta quedó asignada: reintentá y la va a "
                    "encontrar En Pausa)"
                )
            if not aviso and time.time() - inicio > 8:
                logger.info("  esperando que cargue la sección GLOSAS…")
                aviso = True
        page.wait_for_timeout(700)


def _contar_marcadas(page: Page) -> int:
    try:
        return page.evaluate(
            "() => document.querySelectorAll(\"tbody input[type='checkbox']:checked\").length"
        )
    except Exception:
        return 0


def marcar_checkboxes(page: Page, ids: list[str], todas_pendientes: bool) -> int:
    """Marca los checkboxes de las glosas `ids` y devuelve cuántas quedaron
    marcadas (verificado contando en el DOM).

    El JS del portal maneja el estado de los checkboxes a su manera, así que
    NO usamos check()/uncheck() (su verificación estricta revienta con
    'Clicking the checkbox did not change its state'): clickeamos sin
    verificación y validamos nosotros contando los marcados."""

    def _marcadas_de(ids_buscar: list[str]) -> set[str]:
        try:
            res = page.evaluate(
                """(ids) => {
                    const want = new Set(ids); const out = [];
                    for (const tr of document.querySelectorAll('tbody tr')) {
                        const tds = tr.querySelectorAll('td');
                        if (!tds.length) continue;
                        const id = (tds[0].innerText || '').trim();
                        if (!want.has(id)) continue;
                        const cb = tr.querySelector("input[type='checkbox']");
                        if (cb && cb.checked) out.push(id);
                    }
                    return out;
                }""",
                ids_buscar,
            )
            return set(res)
        except Exception:
            return set()

    if todas_pendientes:
        maestro = _primer_visible(page.locator("thead input[type='checkbox']"))
        if maestro is not None:
            maestro.click(force=True)
            page.wait_for_timeout(700)
            if len(_marcadas_de(ids)) >= len(ids):
                return len(ids)

    # Camino rápido: UN evaluate clickea todos los checkboxes objetivo.
    # (Fila por fila con locators tarda ~0.7s por glosa: una tanda de 200
    # se iba a 2½ minutos; así baja a segundos.)
    try:
        page.evaluate(
            """(ids) => {
                const want = new Set(ids);
                for (const tr of document.querySelectorAll('tbody tr')) {
                    const tds = tr.querySelectorAll('td');
                    if (!tds.length) continue;
                    const id = (tds[0].innerText || '').trim();
                    if (!want.has(id)) continue;
                    const cb = tr.querySelector("input[type='checkbox']");
                    if (cb && !cb.checked) cb.click();
                }
            }""",
            ids,
        )
        page.wait_for_timeout(500)
    except Exception:
        pass
    hechas = _marcadas_de(ids)
    faltan = [i for i in ids if i not in hechas]
    if faltan:
        logger.info(f"  ({len(faltan)} checkboxes no entraron en bloque; voy fila por fila)")

    # Lento (fila por fila con eventos), solo para las que falten.
    for id_glosa in faltan:
        fila = page.locator(f"tr:has(td:text-is('{id_glosa}'))").first
        try:
            cb = fila.locator("input[type='checkbox']").last
            if cb.evaluate("el => el.checked"):
                continue
            cb.click(force=True, timeout=3000)
            page.wait_for_timeout(60)
            if cb.evaluate("el => el.checked"):
                continue
            # Último recurso: setear el estado y disparar los eventos que
            # escucha el framework del portal.
            cb.evaluate(
                "el => { el.checked = true;"
                " el.dispatchEvent(new Event('click', {bubbles: true}));"
                " el.dispatchEvent(new Event('change', {bubbles: true})); }"
            )
        except Exception as e:
            logger.warning(f"  no pude marcar id_glosa {id_glosa}: {e}")
    page.wait_for_timeout(300)
    return len(_marcadas_de(ids))


def _desmarcar_todo(page: Page) -> None:
    """Desmarca todos los checkboxes de la grilla clickeando los marcados
    (sin uncheck(): misma incompatibilidad que check())."""
    try:
        for _ in range(3):
            if _contar_marcadas(page) == 0:
                return
            page.evaluate(
                """() => { for (const cb of document.querySelectorAll("tbody input[type='checkbox']:checked")) cb.click(); }"""
            )
            page.wait_for_timeout(300)
    except Exception:
        pass


def _primer_visible(loc):
    """Primer elemento VISIBLE de un locator, o None (los .first a secas
    pueden agarrar plantillas ocultas del DOM)."""
    try:
        n = loc.count()
    except Exception:
        return None
    for i in range(n):
        e = loc.nth(i)
        try:
            if e.is_visible():
                return e
        except Exception:
            continue
    return None


def _esperar_responder_masivamente_habilitado(page: Page, timeout_s: int = 90) -> bool:
    """El portal deja el botón 'Responder Masivamente' DISABLED mientras
    está aplicando la respuesta del grupo anterior (aunque la grilla ya
    muestre RESPONDIDA en pantalla, el backend sigue trabajando). Esto
    bloquea el siguiente click. Espera hasta que se HABILITE de nuevo."""
    deadline = time.time() + timeout_s
    while time.time() < deadline:
        btn = _primer_visible(page.locator("button:has-text('Responder Masivamente')"))
        if btn is not None:
            try:
                if not btn.evaluate("el => !!el.disabled || el.getAttribute('disabled')!==null"):
                    return True
            except Exception:
                return True
        page.wait_for_timeout(1000)
    return False


def responder_grupo(page: Page, grupo: dict, pdf: Path | None) -> None:
    """Abre el modal 'Responder Masivamente' y carga código + justificación
    (+ PDF si aplica). Lanza excepción si el portal no confirma."""
    # Antes del click: esperar a que el portal NO esté procesando algo del
    # grupo anterior (con esto evitamos abrir el modal en estado raro).
    if not _esperar_responder_masivamente_habilitado(page, timeout_s=90):
        _screenshot_debug(page, "responder_masivamente_disabled")
        raise RuntimeError(
            "El botón 'Responder Masivamente' siguió DISABLED 90s "
            "(el portal seguía procesando el grupo anterior)."
        )
    btn_rm = _primer_visible(page.locator("button:has-text('Responder Masivamente')"))
    if btn_rm is None:
        raise RuntimeError("No veo el botón 'Responder Masivamente'.")
    btn_rm.click()
    page.wait_for_selector("text=Respondiendo Masivamente", timeout=15000)
    page.wait_for_timeout(600)

    # Contenedor del modal visible: scope para todos los campos (en el DOM
    # puede haber plantillas/modales ocultos con los mismos textos e IDs).
    scope = page
    for sel in (".modal-content", ".modal-dialog", ".modal", "[role='dialog']"):
        cand = _primer_visible(page.locator(sel).filter(has_text="Respondiendo Masivamente"))
        if cand is not None:
            scope = cand
            break

    # ── Código de respuesta (dropdown, posiblemente select2) ──
    cod = grupo["cod_corto"]
    seleccionado = False
    # 1) <select> VISIBLE dentro del modal (nativo estilizado).
    for sel in scope.locator("select").all():
        try:
            if not sel.is_visible():
                continue
            opciones = sel.locator("option").all_inner_texts()
            match = next((o for o in opciones if cod in o), None)
            if match:
                sel.select_option(label=match)
                seleccionado = True
                break
        except Exception:
            continue
    # 2) Select OCULTO detrás de un select2: setear por JS + change (jQuery
    # incluido, que es lo que escucha select2).
    if not seleccionado:
        try:
            seleccionado = bool(
                page.evaluate(
                    """(cod) => {
                    for (const s of document.querySelectorAll('select')) {
                        for (const o of s.options) {
                            if ((o.text || '').includes(cod)) {
                                s.value = o.value;
                                s.dispatchEvent(new Event('change', {bubbles: true}));
                                if (window.jQuery) window.jQuery(s).trigger('change');
                                return true;
                            }
                        }
                    }
                    return false;
                }""",
                    cod,
                )
            )
        except Exception:
            pass
    # 3) Combo select2 por UI: click + tipear + elegir.
    if not seleccionado:
        combo = _primer_visible(scope.locator(".select2-selection")) or _primer_visible(
            scope.locator("[role='combobox']")
        )
        if combo is None:
            raise RuntimeError("No encontré el dropdown de RESPUESTA en el modal.")
        combo.click()
        page.wait_for_timeout(400)
        caja = _primer_visible(page.locator("input.select2-search__field")) or _primer_visible(
            page.locator("input[type='search']")
        )
        caja.fill(cod)
        page.wait_for_timeout(800)
        opcion = _primer_visible(page.locator(f"li.select2-results__option:has-text('{cod}')"))
        if opcion is None:
            # Fallback: el filtro de select2 a veces no devuelve la opcion
            # aunque exista (timing del ajax que filtra). Vaciamos la caja
            # para listar TODAS las opciones del dropdown y, si la lista
            # contiene `cod`, la clickeamos directamente. El mismo barrido
            # sirve de diagnostico cuando el portal realmente no ofrece el
            # codigo (ej. tipo de glosa restringido).
            ofrecidos: list[str] = []
            try:
                caja.fill("")
                # Esperar la re-hidratacion del listado (mas robusto que un
                # sleep fijo: si el ajax tarda >800ms igual lo captamos).
                opciones = page.locator(
                    "li.select2-results__option:not(.select2-results__message):visible"
                )
                opciones.first.wait_for(timeout=4000)
                for it in opciones.all():
                    try:
                        txt = (it.inner_text() or "").strip()
                    except Exception:
                        continue
                    if not txt:
                        continue
                    ofrecidos.append(txt[:80])
                    if cod in txt:
                        # Re-resolvemos por texto para el click: si el ajax
                        # re-renderizo la lista entre el escaneo y el click,
                        # el nth(K) podria haber cambiado de posicion.
                        opcion = opciones.filter(has_text=cod).first
                        break
            except Exception:
                pass
            if opcion is None:
                if ofrecidos:
                    logger.error(
                        f"  Códigos que SÍ ofrece el dropdown ({len(ofrecidos)}): "
                        + " | ".join(ofrecidos[:25])
                    )
                raise RuntimeError(f"El dropdown no ofrece el código {cod}.")
            logger.info(f"  ↻ filtro select2 fallo con '{cod}'; lo eligi de la lista completa")
        opcion.click()
        seleccionado = True
    page.wait_for_timeout(500)

    # ── Justificación (textarea) — el portal ACEPTA tildes, va tal cual ──
    textarea = _primer_visible(scope.locator("textarea"))
    if textarea is None:
        raise RuntimeError("No encontré el textarea de JUSTIFICACION en el modal.")
    textarea.fill(grupo["obs"])
    # Forzar blur + change para que la validación JS del portal habilite
    # el botón 'Responder Glosa' (fill() setea el valor pero no siempre
    # dispara los eventos que escucha el framework del modal).
    try:
        textarea.evaluate(
            "el => { el.dispatchEvent(new Event('input', {bubbles: true}));"
            " el.dispatchEvent(new Event('change', {bubbles: true}));"
            " el.dispatchEvent(new Event('blur', {bubbles: true})); }"
        )
    except Exception:
        pass
    page.wait_for_timeout(300)

    # ── PDF (solo grupos de soporte) ──
    if pdf is not None:
        file_input = scope.locator("input[type='file']").first
        file_input.set_input_files(str(pdf))
        # Los uploads del portal son asíncronos; con 800ms se vio que el
        # click siguiente disparaba pero el form aún no tenía el adjunto y
        # el submit pasaba en falso. Bumpeamos a 2.5s y, además, esperamos
        # al feedback típico del portal (nombre del archivo a la vista) si
        # aparece.
        page.wait_for_timeout(2500)
        try:
            scope.locator(f"text={pdf.name}").first.wait_for(state="visible", timeout=4000)
        except PlaywrightTimeout:
            pass
        logger.info(f"    adjunto: {pdf.name}")

    # ── Responder Glosa ──
    btn = (
        _primer_visible(scope.locator("#btnAnswerMasivoGlosa"))
        or _primer_visible(scope.locator("#btnAnswerGlosa"))
        or _primer_visible(scope.locator("button:has-text('Responder Glosa')"))
        or _primer_visible(page.locator("button:has-text('Responder Glosa')"))
    )
    if btn is None:
        _screenshot_debug(page, "sin_boton_responder_glosa")
        raise RuntimeError("No veo el botón 'Responder Glosa' del modal.")
    # El portal deja el botón disabled mientras está procesando la respuesta
    # del grupo anterior (o validando que tengamos cod + obs + PDF). Esperamos
    # hasta 45s a que se HABILITE, en vez de clickear ciego y agotar el
    # timeout del click reintentando la factura entera 4 veces.
    try:
        btn.wait_for(state="visible", timeout=5000)
        habilitado = False
        for _ in range(90):
            try:
                if not btn.evaluate("el => !!el.disabled || el.getAttribute('disabled')!==null"):
                    habilitado = True
                    break
            except Exception:
                habilitado = True
                break
            page.wait_for_timeout(1000)
        if not habilitado:
            _screenshot_debug(page, "responder_glosa_disabled")
            raise RuntimeError(
                "El botón 'Responder Glosa' siguió deshabilitado 90s "
                "(¿faltó código, justificación o PDF? ¿portal lento?)."
            )
    except PlaywrightTimeout:
        pass
    btn.click()

    # SEÑAL FUERTE de submit OK: el modal 'Respondiendo Masivamente' se cierra.
    # (wait_for_selector "Se ha dado Respuesta" a secas puede matchear un texto
    # residual del DOM y devolver inmediato sin que la respuesta se guarde.)
    try:
        page.wait_for_selector("text=Respondiendo Masivamente", state="hidden", timeout=60000)
    except PlaywrightTimeout:
        _screenshot_debug(page, "modal_no_cierra")
        raise RuntimeError(
            "El modal 'Respondiendo Masivamente' no se cerró tras click en "
            "'Responder Glosa' — el portal no aceptó la respuesta."
        )
    # Sweet alert de confirmación.
    try:
        page.wait_for_selector("text=Se ha dado Respuesta", timeout=15000)
    except PlaywrightTimeout:
        # Si el modal se cerró igual debe estar guardado; lo informamos pero
        # no rompemos: el chequeo final por leer_estados decide.
        logger.info("    (no vi el cartel 'Se ha dado Respuesta'; sigo)")
    cont = _primer_visible(page.locator("button:has-text('Continuar')"))
    if cont is not None:
        cont.click()
    page.wait_for_timeout(1500)
    # La grilla GLOSAS se redibuja tras la respuesta; le damos chance.
    try:
        page.wait_for_selector("text=Cargando", state="hidden", timeout=15000)
    except PlaywrightTimeout:
        pass


def terminar_respuesta(page: Page, factura: str, evidencias: Path) -> str:
    """Click 'Terminar Respuesta' → 'Sí, Terminar!' → pantallazo del cartel
    '¡Usted ha cerrado una cuenta!' en EVIDENCIA → Continuar."""
    evidencias.mkdir(parents=True, exist_ok=True)
    btn_t = _primer_visible(page.locator("button:has-text('Terminar Respuesta')"))
    if btn_t is None:
        raise RuntimeError("No veo el botón 'Terminar Respuesta'.")
    btn_t.click()
    page.wait_for_selector("text=Desea Terminar", timeout=15000)
    btn_si = _primer_visible(page.locator("button:has-text('Terminar!')")) or _primer_visible(
        page.locator("button:has-text('Si, Terminar')")
    )
    if btn_si is None:
        raise RuntimeError("No veo el botón 'Sí, Terminar!'.")
    btn_si.click()

    # Cartel final de evidencia.
    try:
        page.wait_for_selector("text=ha cerrado una cuenta", timeout=30000)
    except PlaywrightTimeout:
        _screenshot_debug(page, f"sin_cartel_cierre_{factura}")
        return "TERMINADA_SIN_CARTEL"
    ruta = evidencias / f"{factura}_cierre.png"
    page.screenshot(path=str(ruta), full_page=True)
    logger.info(f"  📸 Evidencia: {ruta}")
    try:
        cont = _primer_visible(page.locator("button:has-text('Continuar')"))
        if cont is not None:
            cont.click(timeout=5000)
    except PlaywrightTimeout:
        pass
    return "OK"


# ─── Driver por factura ─────────────────────────────────────────────────────


def procesar_factura(
    page: Page,
    factura: str,
    grupos: list[dict],
    calidad: int,
    indice: dict[str, Path],
    evidencias: Path,
    max_grupos: int = 0,
    cerrar_residuales: bool = False,
) -> dict:
    total_glosas = sum(len(g["ids"]) for g in grupos)
    reg = {
        "factura": factura,
        "grupos": len(grupos),
        "glosas": total_glosas,
        "estado": "",
        "detalle": "",
    }
    try:
        abrir_factura(page, factura)
        esperar_glosas(page)  # la sección GLOSAS carga después que GESTION CUENTA
        preparar_grilla(page)
        estados = esperar_glosas(page)  # releer ya con 'Mostrar: Todos' aplicado

        respondidas_previas = sum(1 for e in estados.values() if e == "RESPONDIDA")
        pendientes_portal = {i for i, e in estados.items() if e == "SIN RESPUESTA"}
        logger.info(
            f"  portal: {len(estados)} glosas en grilla, "
            f"{respondidas_previas} ya respondidas, {len(pendientes_portal)} pendientes"
        )

        grupos_hechos = 0
        tandas_hechas = 0  # respuestas enviadas en ESTA carga de página
        grupos_saltados_pdx: list[str] = []
        a_procesar = grupos[:max_grupos] if max_grupos > 0 else grupos
        # El modal con cientos de glosas marcadas se rompe (el dropdown no
        # carga los códigos); partimos los grupos grandes en tandas.
        LOTE_MAX = 200

        def _recargar_pagina() -> None:
            """El modal del portal es de UN SOLO USO por carga de página: el
            primer 'Responder Glosa' funciona, el segundo queda disabled para
            siempre. Recargar (F5) lo resetea — verificado: al reintentar la
            factura (página fresca) el grupo que fallaba salía al primer
            intento."""
            logger.info("  ↻ recargo la página (el modal es de un solo uso)")
            try:
                page.reload(wait_until="domcontentloaded", timeout=90000)
            except PlaywrightTimeout:
                page.reload(wait_until="commit", timeout=90000)
            esperar_glosas(page)
            preparar_grilla(page)
            esperar_glosas(page)

        for n, g in enumerate(a_procesar, start=1):
            ids_pend = [i for i in g["ids"] if i in pendientes_portal]
            if not ids_pend:
                logger.info(f"  grupo {n}/{len(grupos)}: ya respondido → omito")
                continue

            pdf = None
            if g["es_soporte"]:
                pdf, detalle_pdx = buscar_pdx(factura, indice)
                if pdf is None:
                    logger.warning(f"  grupo {n} es SOPORTES pero {detalle_pdx} → SALTO el grupo")
                    grupos_saltados_pdx.append(detalle_pdx)
                    continue

            n_tandas = (len(ids_pend) + LOTE_MAX - 1) // LOTE_MAX
            logger.info(
                f"  → grupo {n}/{len(grupos)}: {len(ids_pend)} glosas, cod {g['cod_corto']}"
                + (" + PDF" if pdf else "")
                + (f" (en {n_tandas} tandas de ≤{LOTE_MAX})" if n_tandas > 1 else "")
            )
            restantes = list(ids_pend)
            k = 0
            while restantes:
                tanda = restantes[:LOTE_MAX]
                restantes = restantes[LOTE_MAX:]
                k += 1
                if tandas_hechas > 0:
                    _recargar_pagina()
                if n_tandas > 1:
                    logger.info(f"    tanda {k}/{n_tandas}: {len(tanda)} glosas")
                todas = set(tanda) == pendientes_portal
                marcadas = marcar_checkboxes(page, tanda, todas_pendientes=todas)
                if marcadas == 0:
                    raise RuntimeError(f"no pude marcar ningún checkbox del grupo {n}")
                responder_grupo(page, g, pdf)
                tandas_hechas += 1
                pendientes_portal -= set(tanda)
                _desmarcar_todo(page)
                # Esperar a que la grilla confirme lo recién respondido antes
                # de seguir (si no, el reload puede agarrar al backend a
                # mitad del guardado).
                confirmadas: set[str] = set()
                t_dl = time.time() + 90
                while time.time() < t_dl:
                    try:
                        estados_act = leer_estados(page)
                    except Exception:
                        estados_act = {}
                    confirmadas = {i for i in tanda if estados_act.get(i) == "RESPONDIDA"}
                    if len(confirmadas) >= len(tanda):
                        break
                    page.wait_for_timeout(1500)
                if len(confirmadas) < len(tanda):
                    logger.warning(
                        f"  ⚠ El portal confirmó {len(confirmadas)}/{len(tanda)} "
                        f"glosas del grupo {n} en 90s; sigo igual"
                    )
            grupos_hechos += 1

        if max_grupos > 0:
            reg["estado"] = "PILOTO_PARCIAL"
            reg["detalle"] = f"{grupos_hechos} grupos respondidos (sin Terminar)"
            return reg

        # ¿Quedó algo sin responder? Damos paciencia: tras el último grupo
        # la grilla puede tardar en reflejar los cambios; releemos hasta 20s.
        deadline = time.time() + 20
        while True:
            estados = leer_estados(page)
            aun_pendientes = sum(1 for e in estados.values() if e == "SIN RESPUESTA")
            if aun_pendientes == 0 or time.time() > deadline:
                break
            page.wait_for_timeout(1000)
        if grupos_saltados_pdx:
            reg["estado"] = "PENDIENTE_PDX"
            reg["detalle"] = f"{grupos_hechos} grupos ok; sin PDX: {'; '.join(grupos_saltados_pdx)}"
            return reg

        # --cerrar-residuales: si el portal tiene glosas sin respuesta que el
        # Excel no cubría, marcarlas con código RE9901 y texto genérico de
        # pertinencia. Esto permite Terminar la factura completa en una pasada
        # sin tener que abrir cada glosa residual a mano.
        if cerrar_residuales and aun_pendientes > 0:
            ids_residuales = [i for i, e in estados.items() if e == "SIN RESPUESTA"]
            logger.info(
                f"  --cerrar-residuales: respondiendo {len(ids_residuales)} glosa(s) "
                f"residual(es) con cód {COD_GENERICO_RESIDUALES}"
            )
            grupo_residual = {
                "cod": COD_GENERICO_RESIDUALES,
                "cod_corto": COD_GENERICO_RESIDUALES,
                "obs": TEXTO_GENERICO_RESIDUALES,
                "ids": ids_residuales,
                "es_soporte": False,
            }
            restantes = list(ids_residuales)
            k_res = 0
            n_tandas_res = (len(restantes) + LOTE_MAX - 1) // LOTE_MAX
            while restantes:
                tanda = restantes[:LOTE_MAX]
                restantes = restantes[LOTE_MAX:]
                k_res += 1
                if tandas_hechas > 0:
                    _recargar_pagina()
                if n_tandas_res > 1:
                    logger.info(f"    residuales tanda {k_res}/{n_tandas_res}: {len(tanda)} glosas")
                marcadas = marcar_checkboxes(page, tanda, todas_pendientes=False)
                if marcadas == 0:
                    raise RuntimeError("no pude marcar checkboxes de las glosas residuales")
                responder_grupo(page, grupo_residual, None)
                tandas_hechas += 1
                _desmarcar_todo(page)
                # Esperar a que la grilla confirme antes de pasar a la siguiente tanda
                t_dl_res = time.time() + 60
                confirmadas_res: set[str] = set()
                while time.time() < t_dl_res:
                    try:
                        estados_act = leer_estados(page)
                    except Exception:
                        estados_act = {}
                    confirmadas_res = {i for i in tanda if estados_act.get(i) == "RESPONDIDA"}
                    if len(confirmadas_res) >= len(tanda):
                        break
                    page.wait_for_timeout(1500)
            # Releer estado final
            estados = leer_estados(page)
            aun_pendientes = sum(1 for e in estados.values() if e == "SIN RESPUESTA")

        if calidad > 0:
            # Los conceptos CALIDAD no se responden: la factura queda ABIERTA
            # (sin Terminar) para que el equipo médico maneje la pertinencia.
            reg["estado"] = "OK_CALIDAD_ABIERTA"
            reg["detalle"] = (
                f"{grupos_hechos} grupos respondidos; {calidad} glosas CALIDAD no se "
                f"responden (portal muestra {aun_pendientes} sin respuesta); queda SIN Terminar"
            )
            return reg
        if aun_pendientes:
            reg["estado"] = "PENDIENTES"
            reg["detalle"] = (
                f"{grupos_hechos} grupos ok, pero el portal aún muestra {aun_pendientes} "
                "SIN RESPUESTA (glosas del portal que no están en el Excel)"
            )
            return reg

        estado_fin = terminar_respuesta(page, factura, evidencias)
        reg["estado"] = estado_fin
        reg["detalle"] = f"{grupos_hechos} grupos respondidos, evidencia capturada"
    except FacturaNoEnBolsa as e:
        reg["estado"] = "NO_EN_BOLSA"
        reg["detalle"] = str(e)
        logger.info(f"  ⏭ {e}")
    except Exception as e:
        reg["estado"] = "ERROR"
        reg["detalle"] = f"{type(e).__name__}: {e}"
        logger.error(f"  ✗ {factura}: {reg['detalle']}")
        _screenshot_debug(page, f"error_{factura}")
    return reg


# ─── Main ───────────────────────────────────────────────────────────────────


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Respuesta masiva de glosas en COOSALUD (vco.ctamedicas.com).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--excel", type=Path, required=True, help="Excel consolidado de glosas COOSALUD."
    )
    parser.add_argument("--hoja", type=str, default="BASE", help="Hoja a procesar (default BASE).")
    parser.add_argument(
        "--incluir-calidad",
        action="store_true",
        help=(
            "Responder TAMBIÉN las glosas tipo_glosa==CALIDAD usando el "
            "COD_RTA y OBSERVACION del Excel. Por defecto se excluyen y la "
            "factura queda en pausa para el equipo médico. Activá este flag "
            "cuando la planilla ya trae respuesta tipificada para pertinencia "
            "y querés cerrar la factura completa en una sola corrida."
        ),
    )
    parser.add_argument(
        "--indice",
        type=Path,
        default=None,
        help="TXT índice factura→carpeta (para el PDX de SOPORTES).",
    )
    grupo = parser.add_mutually_exclusive_group(required=True)
    grupo.add_argument("--solo", type=str, help="Procesar solo esta factura (HUS...).")
    grupo.add_argument(
        "--facturas", type=str, help="Lista de facturas separadas por coma (HUS...,HUS...)."
    )
    grupo.add_argument("--lista", type=Path, help="TXT con una factura por línea (HUS...).")
    grupo.add_argument(
        "--todas", action="store_true", help="Procesar todas las facturas de la hoja."
    )
    parser.add_argument(
        "--max-grupos",
        type=int,
        default=0,
        help="Responder como mucho N grupos (piloto; no Termina la factura).",
    )
    parser.add_argument(
        "--max-facturas",
        type=int,
        default=0,
        help="Procesar como mucho N facturas (piloto; corta la lista de --todas).",
    )
    parser.add_argument(
        "--evidencias",
        type=Path,
        default=Path("EVIDENCIA"),
        help="Carpeta para los pantallazos de cierre (default: EVIDENCIA).",
    )
    parser.add_argument("--con-cabeza", action="store_true", help="Mostrar el browser.")
    parser.add_argument("--lento", action="store_true", help="Slow-motion 300ms (debug).")
    parser.add_argument("--reporte", type=Path, default=Path("reporte_coosalud.csv"))
    parser.add_argument(
        "--saltar-csv",
        type=Path,
        action="append",
        default=[],
        help="CSV(s) de reportes previos: las facturas con estado terminal "
        "(OK, OK_CALIDAD_ABIERTA, SOLO_CALIDAD, NO_EN_BOLSA, TERMINADA_SIN_CARTEL) "
        "se omiten. Podés pasarlo varias veces.",
    )
    parser.add_argument(
        "--cerrar-residuales",
        action="store_true",
        help=(
            "Si tras responder todos los grupos del Excel el portal sigue "
            "mostrando glosas SIN RESPUESTA (residuales que no están en el "
            "Excel), responderlas con código RE9901 y texto genérico de "
            "pertinencia, y Terminar la factura. Sin este flag, el bot deja "
            "la factura en estado PENDIENTES."
        ),
    )
    parser.add_argument("--log", type=Path, default=None)
    args = parser.parse_args()
    setup_logging(args.log)

    _exigir_playwright()
    user, password = cargar_credenciales()
    logger.info(f"Usuario COOSALUD: {user}")

    facturas = leer_excel(args.excel, args.hoja, incluir_calidad=args.incluir_calidad)
    if args.incluir_calidad:
        logger.info(
            "  --incluir-calidad: las glosas CALIDAD también se responden con su texto del Excel."
        )
    tot_glosas = sum(len(g["ids"]) for f in facturas.values() for g in f["grupos"])
    tot_calidad = sum(f["calidad"] for f in facturas.values())
    logger.info(
        f"Hoja {args.hoja}: {len(facturas)} facturas, {tot_glosas:,} glosas a responder"
        + (f" (+{tot_calidad} CALIDAD excluidas)" if tot_calidad else "")
        + "."
    )

    if args.solo:
        objetivo = args.solo.strip().upper()
        facturas = {k: v for k, v in facturas.items() if k.upper() == objetivo}
        if not facturas:
            logger.error(f"No hallé la factura {args.solo} en la hoja {args.hoja}.")
            return 1

    if args.facturas or args.lista:
        if args.lista:
            if not args.lista.is_file():
                logger.error(f"No existe el archivo de lista: {args.lista}")
                return 1
            texto = args.lista.read_text(encoding="utf-8-sig", errors="replace")
            objetivos = {
                x.strip().upper() for x in texto.replace(",", "\n").splitlines() if x.strip()
            }
        else:
            objetivos = {f.strip().upper() for f in args.facturas.split(",") if f.strip()}
        encontradas = {k: v for k, v in facturas.items() if k.upper() in objetivos}
        faltantes = objetivos - {k.upper() for k in encontradas}
        if not encontradas:
            logger.error(f"Ninguna de las facturas pedidas está en la hoja {args.hoja}.")
            return 1
        if faltantes:
            logger.warning(f"⚠ No están en la hoja {args.hoja}: {', '.join(sorted(faltantes))}")
        logger.info(f"Procesando {len(encontradas)} de {len(objetivos)} facturas pedidas.")
        facturas = encontradas

    if args.saltar_csv:
        ESTADOS_TERMINALES = {
            "OK",
            "OK_CALIDAD_ABIERTA",
            "SOLO_CALIDAD",
            "NO_EN_BOLSA",
            "TERMINADA_SIN_CARTEL",
        }
        ya_listas: set[str] = set()
        for ruta in args.saltar_csv:
            if not ruta.is_file():
                logger.warning(f"⚠ --saltar-csv: no existe {ruta}, lo ignoro.")
                continue
            with ruta.open("r", encoding="utf-8-sig", newline="") as fh:
                reader = csv.DictReader(fh)
                for fila in reader:
                    if (fila.get("estado") or "").strip().upper() in ESTADOS_TERMINALES:
                        f = (fila.get("factura") or "").strip().upper()
                        if f:
                            ya_listas.add(f)
            logger.info(
                f"  --saltar-csv {ruta.name}: acumulado {len(ya_listas)} facturas terminales."
            )
        if ya_listas:
            antes = len(facturas)
            facturas = {k: v for k, v in facturas.items() if k.upper() not in ya_listas}
            logger.info(
                f"Salto {antes - len(facturas)} facturas ya cerradas según reportes previos; quedan {len(facturas)}."
            )

    if args.max_facturas > 0:
        recorte = dict(list(facturas.items())[: args.max_facturas])
        logger.info(f"Piloto: proceso solo {len(recorte)} de {len(facturas)} facturas.")
        facturas = recorte

    indice: dict[str, Path] = {}
    if args.indice is not None:
        indice = cargar_indice(args.indice)
        logger.info(f"Índice cargado: {len(indice):,} facturas mapeadas.")
    else:
        n_sop = sum(1 for f in facturas.values() for g in f["grupos"] if g["es_soporte"])
        if n_sop:
            logger.warning(f"⚠ Sin --indice: {n_sop} grupos de SOPORTES quedarán PENDIENTE_PDX.")

    # Reporte CSV incremental (append-as-you-go).
    args.reporte.parent.mkdir(parents=True, exist_ok=True)
    f_rep = args.reporte.open("w", newline="", encoding="utf-8-sig")
    w_rep = csv.DictWriter(f_rep, fieldnames=["factura", "grupos", "glosas", "estado", "detalle"])
    w_rep.writeheader()
    resultados: list[dict] = []

    def registrar(reg: dict) -> None:
        resultados.append(reg)
        w_rep.writerow(reg)
        if len(resultados) % 5 == 0:
            f_rep.flush()

    def _sesion_muerta(detalle: str) -> bool:
        d = detalle.lower()
        return (
            "target page, context or browser has been closed" in d
            or "target closed" in d
            or "browser has been closed" in d
            or "connection closed" in d
            or "targetclosederror" in d
        )

    def _es_lentitud(detalle: str) -> bool:
        """Errores que huelen a racha de lentitud del portal (no a bug):
        vale la pena reintentar la misma factura desde cero."""
        d = detalle.lower()
        return (
            "timeout" in d
            or "no cargó" in d
            or "no mostró filas" in d
            or "no responde" in d
            or "net::err" in d
        )

    def _abrir_sesion(p):
        b = p.chromium.launch(headless=not args.con_cabeza, slow_mo=300 if args.lento else 0)
        c = b.new_context(accept_downloads=True)
        pg = c.new_page()
        # El portal es lento: timeouts generosos por defecto para TODA
        # navegación y acción (los waits puntuales ya tienen los suyos).
        pg.set_default_navigation_timeout(120000)
        pg.set_default_timeout(30000)
        pg.on("dialog", lambda d: d.accept())
        login(pg, user, password)
        return b, c, pg

    def _abrir_sesion_con_reintentos(p, intentos: int = 3):
        ultimo = None
        for n in range(1, intentos + 1):
            try:
                return _abrir_sesion(p)
            except Exception as e:
                ultimo = e
                logger.warning(f"Apertura de sesión falló (intento {n}/{intentos}): {e}")
                time.sleep(5)
        raise RuntimeError(f"No pude abrir sesión tras {intentos} intentos: {ultimo}")

    t0 = time.time()
    with sync_playwright() as p:
        browser, ctx, page = _abrir_sesion_con_reintentos(p)
        relogins = 0
        MAX_RELOGINS = 5
        try:
            for i, (factura, datos) in enumerate(facturas.items(), start=1):
                grupos = datos["grupos"]
                calidad = datos["calidad"]
                extra = f" (+{calidad} CALIDAD excluidas)" if calidad else ""
                logger.info(
                    f"[{i}/{len(facturas)}] {factura} — {len(grupos)} grupo(s), "
                    f"{sum(len(g['ids']) for g in grupos)} glosas{extra}"
                )
                if not grupos:
                    registrar(
                        {
                            "factura": factura,
                            "grupos": 0,
                            "glosas": 0,
                            "estado": "SOLO_CALIDAD",
                            "detalle": f"{calidad} glosas CALIDAD; nada que responder",
                        }
                    )
                    continue
                INTENTOS_FACTURA = 3
                intento = 0
                while True:
                    intento += 1
                    reg = procesar_factura(
                        page,
                        factura,
                        grupos,
                        calidad,
                        indice,
                        args.evidencias,
                        max_grupos=args.max_grupos,
                        cerrar_residuales=args.cerrar_residuales,
                    )
                    if reg["estado"] == "ERROR" and intento < INTENTOS_FACTURA:
                        det = reg["detalle"]
                        if _sesion_muerta(det):
                            if relogins >= MAX_RELOGINS:
                                logger.error(
                                    "Sesión irrecuperable; re-corré el comando para continuar."
                                )
                                registrar(reg)
                                raise RuntimeError("Sesión irrecuperable")
                            relogins += 1
                            logger.warning(f"  ⚠ Sesión caída. Re-login {relogins}/{MAX_RELOGINS}…")
                            try:
                                browser.close()
                            except Exception:
                                pass
                            browser, ctx, page = _abrir_sesion_con_reintentos(p)
                            continue
                        if _es_lentitud(det):
                            logger.warning(
                                f"  ⚠ Racha de lentitud del portal; reintento la factura "
                                f"({intento}/{INTENTOS_FACTURA - 1})…"
                            )
                            time.sleep(8)
                            continue
                    registrar(reg)
                    break
        except RuntimeError:
            pass
        finally:
            try:
                browser.close()
            except Exception:
                pass
            f_rep.close()

    dur = (time.time() - t0) / 60
    logger.info(f"\nReporte: {args.reporte}")
    logger.info(f"Evidencias: {args.evidencias}")
    logger.info(f"Facturas procesadas: {len(resultados)} en {dur:.1f} min")
    from collections import Counter

    for estado, n in Counter(r["estado"] for r in resultados).items():
        logger.info(f"  {estado}: {n}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
