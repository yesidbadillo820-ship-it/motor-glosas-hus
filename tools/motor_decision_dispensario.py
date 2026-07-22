"""motor_decision_dispensario.py — Modulo de Decision (el "cerebro").

No es jurídico, no es IA: son REGLAS DE NEGOCIO. Toma el expediente probatorio
(hechos probados + contradicciones + contrato + cartera) y, por cada glosa,
decide:

  - Defendibilidad (0-100%): que tan sostenible es el cobro con lo que hay.
  - Riesgos multidimensionales: probatorio, documental, contractual, tarifario,
    financiero, jurídico (cada uno ALTO/MEDIO/BAJO con su razon).
  - Documentos requeridos (matriz de precedencia) y los que faltan.
  - ACCION recomendada: solicitar levantamiento / aceptar parcial / solicitar
    soporte / escalar / conciliar / aceptar / cobro jurídico.

Asi el Modulo Jurídico solo FUNDAMENTA la decision con la norma, y el de
Argumentacion solo la REDACTA. Cada recomendacion es explicable y trazable.

Entrada: expedientes_probatorios.json (salida del Modulo 3.5).
Salida:  expedientes_decision.json (+ reporte opcional).

USO
---
    py tools\\motor_decision_dispensario.py ^
        --expedientes "D:\\...\\expedientes_probatorios.json" ^
        --salida      "D:\\...\\expedientes_decision.json" ^
        --reporte     "D:\\...\\DECISION.xlsx"
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from pathlib import Path

_TOOLS = Path(__file__).resolve().parent
if str(_TOOLS) not in sys.path:
    sys.path.insert(0, str(_TOOLS))
from motor_verificacion_dispensario import HECHOS_POR_FAMILIA  # noqa: E402

logger = logging.getLogger("motor_decision")

UMBRAL_ALTO = 5_000_000
UMBRAL_MEDIO = 1_000_000

# Palabras que marcan una contradiccion GRAVE (obliga a escalar antes de responder).
ALERTA_GRAVE = ("paciente", "cartera", "otra factura")


def setup_logging() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")


def _riesgo_valor(v: float) -> str:
    if v >= UMBRAL_ALTO:
        return "ALTO"
    if v >= UMBRAL_MEDIO:
        return "MEDIO"
    return "BAJO"


def documentos_requeridos(familia: str) -> list[str]:
    """Matriz de precedencia: que documentos necesita esta familia de glosa.

    Union de los grupos 'requiere_alguno' de los hechos de la familia. Asi el
    sistema sabe QUE buscar (no busca todo, solo lo necesario).
    """
    req: list[str] = []
    for spec in HECHOS_POR_FAMILIA.get(familia, []):
        for t in spec.get("requiere_alguno", []):
            if t not in req:
                req.append(t)
    return req


def calificar_glosa(glosa: dict, exp: dict) -> dict:
    """Decision determinista para una glosa. Devuelve el bloque 'decision'."""
    hechos = glosa.get("hechos_probados", [])
    valor = float(glosa.get("valor_objetado", 0) or 0)
    confs = [(h["nivel_confianza"] if h.get("probado") else 0.0) for h in hechos]
    defendibilidad = round(100 * sum(confs) / len(confs)) if confs else 0
    faltantes = sorted({f for h in hechos for f in h.get("faltantes", [])})

    alertas = exp.get("alertas", [])
    graves = [a for a in alertas if any(k in a.lower() for k in ALERTA_GRAVE)]

    # --- Riesgos multidimensionales ---
    riesgos = {
        "probatorio": (
            "ALTO"
            if any(not h.get("probado") for h in hechos)
            else "MEDIO"
            if any(h.get("nivel_confianza", 0) < 0.85 for h in hechos)
            else "BAJO"
            if hechos
            else "ALTO"
        ),
        "documental": "ALTO" if len(faltantes) >= 2 else ("MEDIO" if faltantes else "BAJO"),
        "contractual": (
            "ALTO"
            if exp.get("contrato", {}).get("codigo") == "?"
            else "MEDIO"
            if exp.get("contrato", {}).get("pendiente_confirmar")
            else "BAJO"
        ),
        "tarifario": "MEDIO" if glosa.get("familia") == "TARIFAS" else "BAJO",
        "financiero": _riesgo_valor(valor),
        "juridico": "MEDIO",  # se afina en el Modulo Juridico
    }
    razones: dict[str, str] = {}
    no_prob = [h["hecho"] for h in hechos if not h.get("probado")]
    if no_prob:
        razones["probatorio"] = "No se probó: " + "; ".join(no_prob)
    if faltantes:
        razones["documental"] = "Falta: " + "; ".join(faltantes[:4])
    if riesgos["contractual"] != "BAJO":
        razones["contractual"] = "Contrato por confirmar o sin resolver a la fecha."
    if riesgos["tarifario"] == "MEDIO":
        razones["tarifario"] = "Requiere cotejar el valor facturado contra la tarifa pactada."

    # --- Decision (regla de negocio) ---
    if graves:
        decision, motivo = "ESCALAR", "Contradiccion critica: " + graves[0]
    elif defendibilidad >= 85:
        decision, motivo = "SOLICITAR_LEVANTAMIENTO", "Hechos probados con evidencia alta."
    elif defendibilidad >= 60:
        if faltantes:
            decision = "SOLICITAR_SOPORTE"
            motivo = "Defensa viable; falta soporte: " + "; ".join(faltantes[:2])
        else:
            decision, motivo = "ACEPTAR_PARCIAL", "Evidencia parcial; defender lo probado."
    elif defendibilidad >= 40:
        decision = "SOLICITAR_SOPORTE"
        motivo = "Evidencia insuficiente; recolectar soporte antes de responder."
    else:
        decision, motivo = "CONCILIAR", "Sin evidencia suficiente para sostener el cobro."

    prioridad = "ALTA" if valor >= UMBRAL_ALTO else ("MEDIA" if valor >= UMBRAL_MEDIO else "BAJA")
    return {
        "defendibilidad": defendibilidad,
        "riesgos": riesgos,
        "riesgo_razones": razones,
        "documentos_requeridos": documentos_requeridos(glosa.get("familia", "")),
        "documentos_faltantes": faltantes,
        "decision": decision,
        "motivo_decision": motivo,
        "prioridad": prioridad,
    }


# Orden de severidad para elegir la decision dominante de un expediente.
_SEVERIDAD = [
    "SOLICITAR_LEVANTAMIENTO",
    "ACEPTAR_PARCIAL",
    "SOLICITAR_SOPORTE",
    "CONCILIAR",
    "ESCALAR",
]


def decidir_expedientes(payload: dict) -> dict:
    from collections import Counter

    conteo: Counter = Counter()
    for exp in payload.get("expedientes", []):
        defs = []
        decisiones = []
        for g in exp.get("glosas", []):
            dec = calificar_glosa(g, exp)
            g["decision"] = dec
            defs.append(dec["defendibilidad"])
            decisiones.append(dec["decision"])
            conteo[dec["decision"]] += 1
        exp["defendibilidad_promedio"] = round(sum(defs) / len(defs)) if defs else 0
        # decision dominante = la mas severa presente
        exp["decision_dominante"] = next((d for d in reversed(_SEVERIDAD) if d in decisiones), "")
    payload["resumen_decision"] = dict(conteo)
    return payload


def escribir_reporte(payload: dict, salida: Path) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    color = {
        "SOLICITAR_LEVANTAMIENTO": "D9EAD3",
        "ACEPTAR_PARCIAL": "FFF2CC",
        "SOLICITAR_SOPORTE": "FCE5CD",
        "CONCILIAR": "F4CCCC",
        "ESCALAR": "EAD1DC",
    }
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Decision"
    ws.append(
        [
            "Factura",
            "Glosa",
            "Familia",
            "Valor objetado",
            "Defendibilidad %",
            "Decision",
            "Prioridad",
            "R. probatorio",
            "R. documental",
            "R. financiero",
            "Docs faltantes",
            "Motivo",
        ]
    )
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E5F")
    for exp in payload.get("expedientes", []):
        for g in exp.get("glosas", []):
            d = g.get("decision", {})
            r = d.get("riesgos", {})
            ws.append(
                [
                    exp.get("factura"),
                    g.get("codigo_corto", ""),
                    g.get("familia", ""),
                    g.get("valor_objetado", 0),
                    d.get("defendibilidad", 0),
                    d.get("decision", ""),
                    d.get("prioridad", ""),
                    r.get("probatorio", ""),
                    r.get("documental", ""),
                    r.get("financiero", ""),
                    "; ".join(d.get("documentos_faltantes", [])),
                    d.get("motivo_decision", ""),
                ]
            )
            fila = ws.max_row
            ws.cell(row=fila, column=4).number_format = "#,##0"
            cc = ws.cell(row=fila, column=6)
            if d.get("decision") in color:
                cc.fill = PatternFill("solid", fgColor=color[d["decision"]])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{ws.max_row}"
    for w, col in zip((16, 12, 12, 15, 14, 22, 10, 12, 12, 12, 34, 50), "ABCDEFGHIJKL"):
        ws.column_dimensions[col].width = w
    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(salida))


def main(argv: list[str] | None = None) -> int:
    setup_logging()
    ap = argparse.ArgumentParser(
        description="Motor de decision (reglas de negocio) de conciliacion."
    )
    ap.add_argument(
        "--expedientes",
        type=Path,
        required=True,
        help="expedientes_probatorios.json (salida del Modulo 3.5)",
    )
    ap.add_argument("--salida", type=Path, required=True, help="expedientes_decision.json")
    ap.add_argument("--reporte", type=Path, default=None, help="Excel legible (opcional)")
    args = ap.parse_args(argv)

    try:
        payload = json.loads(args.expedientes.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        logger.error("No pude leer %s: %s", args.expedientes, e)
        return 1

    payload = decidir_expedientes(payload)
    args.salida.parent.mkdir(parents=True, exist_ok=True)
    args.salida.write_text(json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")
    logger.info("Decisiones por glosa: %s  -> %s", payload.get("resumen_decision", {}), args.salida)
    if args.reporte:
        escribir_reporte(payload, args.reporte)
        logger.info("Reporte legible: %s", args.reporte)
    return 0


if __name__ == "__main__":
    sys.exit(main())
