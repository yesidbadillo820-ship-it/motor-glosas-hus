"""asistente_conciliacion_dispensario.py — Asistente de conciliación de glosas del Dispensario.

Construye, POR CADA GLOSA, la mejor defensa técnica/jurídica/documental a partir
de TODOS los soportes de la factura. Recorre la carpeta de soportes (p. ej. la
unidad de red `Y:\\...\\DISPENSARIO\\...`), lee cada documento (RIPS, XML/CUFE,
CUV, PDF de historia clínica/epicrisis/orden/etc., con OCR opcional para
escaneados), arma la MATRIZ DE EVIDENCIA, verifica coherencia y redacta la
respuesta técnica lista para el pagador.

REGLA PRINCIPAL: nunca inventa evidencia. Si un soporte no existe, lo dice; si la
evidencia es insuficiente, marca "Información insuficiente" y qué documento falta.
Toda conclusión indica documento, ruta, página (cuando aplica) y norma.

CORRE EN LA MÁQUINA DONDE ESTÁ MONTADA LA CARPETA DE SOPORTES (Windows con `Y:\\`).

USO
---
    py tools\\asistente_conciliacion_dispensario.py ^
        --excel   "D:\\...\\HUS.xlsx" ^
        --soportes-raiz "Y:\\7. JULIO 2026 - SOPORTES RADICACION\\DISPENSARIO" ^
        --salida  "D:\\...\\MATRIZ_EVIDENCIA_Y_RESPUESTAS.xlsx"

    # Piloto de una factura (recomendado antes del masivo):
    ...  --solo HUS524435 --con-oficios "D:\\...\\OFICIOS"

DEPENDENCIAS
------------
    py -m pip install openpyxl pdfplumber
    # OCR opcional (escaneados):  py -m pip install pytesseract pdf2image  + Tesseract instalado
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

# Reutiliza los lectores del repo (tools/adres/).
_ADRES = Path(__file__).resolve().parent / "adres"
if str(_ADRES) not in sys.path:
    sys.path.insert(0, str(_ADRES))

logger = logging.getLogger("asistente_conciliacion")

# ---------------------------------------------------------------------------
# Clasificación de soportes por token del nombre (convención del Dispensario)
#   Ej.: RIPS_900006037_HUS524435.json, HEV_..._HUS524435.pdf
# ---------------------------------------------------------------------------
TIPOS_DOC = {
    "RIPS": "RIPS",
    "XML": "Factura electronica (XML/CUFE)",
    "FEV": "Factura electronica (representacion PDF)",
    "CUV": "Validacion MinSalud (CUV)",
    "CRC": "Constancia de radicacion",
    "HEV": "Historia clinica / Evolucion",
    "HC": "Historia clinica",
    "EPICRISIS": "Epicrisis",
    "EPIC": "Epicrisis",
    "OPF": "Orden medica / Formula",
    "OM": "Orden medica",
    "ORDEN": "Orden medica",
    "HAM": "Hoja de administracion de medicamentos",
    "PDE": "Descripcion / detalle de prestacion",
    "PDX": "Descripcion quirurgica",
    "AUT": "Autorizacion",
    "AUTORIZACION": "Autorizacion",
    "MIPRES": "MIPRES",
    "LAB": "Resultados de laboratorio",
    "IMG": "Imagenologia / ayudas diagnosticas",
    "NC": "Nota credito",
    "ND": "Nota debito",
    "PAGO": "Comprobante de pago",
    "ADM": "Documento administrativo",
}

# Qué tipo(s) de documento sustentan cada familia de glosa
SOPORTE_POR_FAMILIA = {
    "TARIFAS": ["Factura electronica (XML/CUFE)", "RIPS"],
    "SOPORTES": [
        "Historia clinica / Evolucion",
        "Orden medica / Formula",
        "Hoja de administracion de medicamentos",
        "Resultados de laboratorio",
    ],
    "CALIDAD": ["Historia clinica / Evolucion", "Epicrisis", "Descripcion quirurgica"],
    "FACTURACION": [
        "Factura electronica (XML/CUFE)",
        "RIPS",
        "Descripcion / detalle de prestacion",
    ],
    "COBERTURA": ["Autorizacion", "Historia clinica / Evolucion"],
    "AUTORIZACION": ["Autorizacion", "MIPRES"],
    "PERTINENCIA": ["Historia clinica / Evolucion", "Epicrisis"],
}

NORMA_FAMILIA = {
    "TARIFAS": "Contrato aplicable (tarifa pactada) + Decreto 780/2016",
    "SOPORTES": "Res. 2284/2023 + Res. 2335/2023 (RIPS)",
    "CALIDAD": "Res. 2284/2023 (Manual Unico de Glosas)",
    "FACTURACION": "Res. 2284/2023 + normativa de factura electronica en salud",
    "COBERTURA": "Ley 1438/2011 Art. 57 + normativa SOAT",
    "AUTORIZACION": "Res. 2284/2023 (soporte de autorizacion)",
    "PERTINENCIA": "Res. 2284/2023 + guia de practica clinica",
}

FAMILIAS = {
    "TA": "TARIFAS",
    "SO": "SOPORTES",
    "CL": "CALIDAD",
    "FA": "FACTURACION",
    "CO": "COBERTURA",
    "AU": "AUTORIZACION",
    "PT": "PERTINENCIA",
}

CORTE_287 = datetime(2025, 11, 30)
FIN_440 = datetime(2026, 7, 30)


# ---------------------------------------------------------------------------
# Estructuras
# ---------------------------------------------------------------------------
@dataclass
class Documento:
    tipo: str
    ruta: Path
    nombre: str
    paginas: int = 0
    texto: str = ""
    requiere_ocr: bool = False
    datos: dict = field(default_factory=dict)  # p.ej. RIPS: numFactura, usuarios


@dataclass
class Glosa:
    factura: str
    radicado: str
    fecha_atencion: datetime | None
    paciente: str
    identificacion: str
    valor_factura: float
    num_glosa: str
    codigo: str
    familia: str
    motivo: str
    valor_objetado: float
    servicio: str


@dataclass
class Resultado:
    glosa: Glosa
    contrato: str
    matriz: list[dict]  # filas de la matriz de evidencia
    coherencia: list[str]  # checks (OK/ALERTA ...)
    conclusion: str  # Procede / No procede / Parcial / Insuficiente
    confianza: int  # 0..100
    norma: str
    oficio: str  # respuesta técnica redactada


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------
def setup_logging() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")


def _num(x) -> float:
    try:
        return float(str(x).replace(",", "").replace("$", "") or 0)
    except (TypeError, ValueError):
        return 0.0


def norm_factura(f) -> str:
    """426013 / HUS426013 / HUS0000426013 -> HUS0000426013."""
    s = str(f).strip().upper().replace("HUS", "").lstrip("0")
    return "HUS" + s.zfill(10)


def factura_corta(f) -> str:
    """Devuelve solo los digitos significativos, para matching de carpetas."""
    return str(f).strip().upper().replace("HUS", "").lstrip("0")


def familia_de(codigo: str) -> str:
    return FAMILIAS.get(str(codigo).strip().upper()[:2], "OTRA")


def codigo_corto(codigo: str) -> str:
    """'CO46 01 COBERTURA-...' -> 'CO4601'. Si no hay patron, primeros 6 alfanum."""
    m = re.search(r"([A-Z]{2}\d{2})\s*(\d{2})?", str(codigo).upper())
    if m:
        return m.group(1) + (m.group(2) or "")
    return re.sub(r"[^A-Z0-9]", "", str(codigo).upper())[:6]


def contrato_por_fecha(fatencion: datetime | None) -> str:
    if not isinstance(fatencion, datetime):
        return "?"
    if fatencion <= CORTE_287:
        return "287"
    if fatencion <= FIN_440:
        return "440"
    return "?"


def clasificar(nombre: str) -> str:
    base = nombre.rsplit(".", 1)[0]
    tokens = re.split(r"[ _\-.]", base.upper())
    for t in tokens:  # el token de tipo suele ir primero (HEV_..., RIPS_...)
        if t in TIPOS_DOC:
            return TIPOS_DOC[t]
    # fallback por extensión
    ext = nombre.rsplit(".", 1)[-1].lower() if "." in nombre else ""
    if ext == "xml":
        return TIPOS_DOC["XML"]
    if ext == "json":
        return "JSON sin clasificar"
    return "Documento sin clasificar"


# ---------------------------------------------------------------------------
# Motor documental + OCR: lee un archivo y devuelve un Documento
# ---------------------------------------------------------------------------
def leer_pdf(ruta: Path, ocr: bool) -> tuple[str, int, bool]:
    """Devuelve (texto, num_paginas, requiere_ocr). OCR solo si ocr=True y hay libs."""
    try:
        import pdfplumber
    except ImportError:
        return ("", 0, True)
    texto, npag = "", 0
    with pdfplumber.open(str(ruta)) as pdf:
        npag = len(pdf.pages)
        for p in pdf.pages:
            texto += (p.extract_text() or "") + "\n"
    if len(texto.strip()) >= 30:
        return (texto, npag, False)
    # Sin texto: probablemente escaneado
    if not ocr:
        return (texto, npag, True)
    try:
        import pytesseract  # noqa: F401
        from pdf2image import convert_from_path

        imgs = convert_from_path(str(ruta))
        texto = "\n".join(pytesseract.image_to_string(im, lang="spa") for im in imgs)
        return (texto, npag, False)
    except Exception as e:  # noqa: BLE001
        logger.warning("OCR no disponible/fallo en %s: %s", ruta.name, e)
        return (texto, npag, True)


def leer_documento(ruta: Path, ocr: bool) -> Documento:
    tipo = clasificar(ruta.name)
    doc = Documento(tipo=tipo, ruta=ruta, nombre=ruta.name)
    ext = ruta.suffix.lower()
    try:
        if ext == ".json":
            from inspeccionar_soportes import cargar_json  # lector del repo

            data = cargar_json(ruta) or {}
            doc.datos = data
            if "numFactura" in data or "usuarios" in data:
                doc.tipo = TIPOS_DOC["RIPS"]
        elif ext == ".xml":
            try:
                from factura_lectura import cargar_factura_xml

                doc.datos = cargar_factura_xml(ruta)
            except Exception:  # noqa: BLE001
                doc.datos = {}
            doc.tipo = TIPOS_DOC["XML"]
        elif ext == ".pdf":
            doc.texto, doc.paginas, doc.requiere_ocr = leer_pdf(ruta, ocr)
    except Exception as e:  # noqa: BLE001
        logger.warning("No pude leer %s: %s", ruta.name, e)
    return doc


# ---------------------------------------------------------------------------
# Motor de indexación: mapea cada factura a su carpeta y sus documentos
# ---------------------------------------------------------------------------
def indexar_soportes(raiz: Path, facturas: set[str], ocr: bool) -> dict[str, list[Documento]]:
    """Recorre TODA la carpeta y agrupa documentos por factura (match por HUS<num>).

    facturas: set de facturas cortas (solo digitos significativos).
    """
    idx: dict[str, list[Documento]] = {f: [] for f in facturas}
    if not raiz.exists():
        logger.error("La carpeta de soportes no existe: %s (¿está montada la unidad?)", raiz)
        return idx
    total = 0
    for ruta in raiz.rglob("*"):
        if not ruta.is_file():
            continue
        # ¿a qué factura pertenece? buscamos HUS<num> en el nombre o en la ruta
        texto_ruta = (ruta.name + " " + str(ruta.parent)).upper()
        m = re.findall(r"HUS0*([0-9]{4,10})", texto_ruta)
        if not m:
            continue
        for corta in {c.lstrip("0") for c in m}:
            if corta in idx:
                idx[corta].append(leer_documento(ruta, ocr))
                total += 1
                break
    logger.info("Indexados %d documentos para %d facturas del lote", total, len(facturas))
    return idx


def docs_desde_indice(
    indice_ruta: Path, facturas: set[str], ocr: bool
) -> dict[str, list[Documento]]:
    """Arma los documentos del lote consultando un INDICE ya construido.

    No recorre `Y:\\`: lee del JSON las rutas de los documentos de cada factura y
    abre SOLO esos archivos. Mucho mas rapido que escanear la unidad de red.
    """
    idx: dict[str, list[Documento]] = {f: [] for f in facturas}
    try:
        payload = json.loads(Path(indice_ruta).read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        logger.error("No pude leer el indice %s: %s", indice_ruta, e)
        return idx
    total = 0
    for reg in payload.get("documentos", []):
        corta = str(reg.get("factura", ""))
        if corta not in idx:
            continue
        ruta = Path(reg.get("ruta", ""))
        if not ruta.exists():
            logger.warning("El indice apunta a un archivo que ya no existe: %s", ruta)
            continue
        idx[corta].append(leer_documento(ruta, ocr))
        total += 1
    logger.info("Del indice: %d documentos para %d facturas del lote", total, len(facturas))
    return idx


# ---------------------------------------------------------------------------
# Motor de cruces / coherencia
# ---------------------------------------------------------------------------
def verificar_coherencia(g: Glosa, docs: list[Documento]) -> list[str]:
    checks: list[str] = []
    rips = next((d for d in docs if d.tipo == TIPOS_DOC["RIPS"]), None)
    if rips and rips.datos:
        numf = str(rips.datos.get("numFactura", "")).upper()
        if numf:
            ok = factura_corta(numf) == factura_corta(g.factura)
            checks.append(
                ("OK" if ok else "ALERTA") + f": RIPS numFactura={numf} vs glosa {g.factura}"
            )
        docs_pac = {
            str(u.get("numDocumentoIdentificacion", "")) for u in rips.datos.get("usuarios", [])
        }
        if docs_pac and g.identificacion:
            ok = str(g.identificacion) in docs_pac
            checks.append(
                ("OK" if ok else "ALERTA")
                + f": documento paciente {g.identificacion} en RIPS={sorted(docs_pac)}"
            )
    else:
        checks.append("ALERTA: no se encontro RIPS para cruzar factura/paciente")
    # coherencia de servicio glosado vs RIPS
    if rips and rips.datos and g.servicio:
        cod_serv = re.findall(r"\d{4,7}", str(g.servicio))
        if cod_serv:
            todos = json.dumps(rips.datos)
            hay = any(c in todos for c in cod_serv)
            checks.append(
                ("OK" if hay else "ALERTA")
                + f": codigo de servicio {cod_serv[0]} {'presente' if hay else 'NO hallado'} en RIPS"
            )
    return checks


# ---------------------------------------------------------------------------
# Motor jurídico + auditoría + confianza: concluye por glosa
# ---------------------------------------------------------------------------
def concluir(
    g: Glosa, docs: list[Documento], matriz: list[dict], checks: list[str]
) -> tuple[str, int]:
    """Devuelve (conclusion, confianza%). Reglas deterministas y explicables."""
    requeridos = SOPORTE_POR_FAMILIA.get(g.familia, [])
    hallados = {d.tipo for d in docs}
    faltan = [t for t in requeridos if t not in hallados]
    alertas = [c for c in checks if c.startswith("ALERTA")]
    hay_ocr_pendiente = any(d.requiere_ocr for d in docs)

    if not docs:
        return ("Informacion insuficiente", 30)
    if g.familia == "TARIFAS":
        # Se defiende con contrato+factura; si están, alta confianza
        base_ok = (
            TIPOS_DOC["XML"] in hallados
            or TIPOS_DOC["FEV"] in hallados
            or TIPOS_DOC["RIPS"] in hallados
        )
        if base_ok and not alertas:
            return ("No procede", 90 if not hay_ocr_pendiente else 80)
        return ("Procede parcialmente", 70)
    # Familias que dependen de soporte clínico
    if not faltan and not alertas:
        return ("No procede", 95 if not hay_ocr_pendiente else 82)
    if not faltan and alertas:
        return ("Procede parcialmente", 70)
    if faltan and (hallados & set(requeridos)):
        return ("Procede parcialmente", 65)
    return ("Informacion insuficiente", 45)


# ---------------------------------------------------------------------------
# Motor de redacción: arma la respuesta técnica
# ---------------------------------------------------------------------------
def redactar_oficio(r: "Resultado") -> str:
    g = r.glosa
    ev = "\n".join(
        f"   - {m['tipo']}: {'SI' if m['encontrado'] else 'NO'}"
        + (f"  (pag {m['pagina']})" if m.get("pagina") else "")
        + (f"  ruta: {m['ruta']}" if m.get("ruta") else "")
        for m in r.matriz
    )
    coher = "\n".join(f"   - {c}" for c in r.coherencia) or "   - (sin cruces automaticos)"
    fatt = g.fecha_atencion.date().isoformat() if isinstance(g.fecha_atencion, datetime) else "?"
    return f"""RESPUESTA TECNICA A GLOSA — {g.factura}

1. IDENTIFICACION
   Factura: {g.factura}   Radicado: {g.radicado}
   Paciente: {g.paciente}   Documento: {g.identificacion}
   Contrato aplicable (por fecha de atencion {fatt}): {r.contrato}
   Glosa: {g.codigo} ({g.familia})   Valor objetado: ${g.valor_objetado:,.0f}
   Servicio objetado: {g.servicio}

2. HALLAZGO / MOTIVO DE LA GLOSA
   {g.motivo}

3. EVIDENCIA ENCONTRADA (matriz)
{ev}

4. CRUCES DE COHERENCIA
{coher}

5. ANALISIS JURIDICO
   Norma aplicable: {r.norma}

6. CONCLUSION: {r.conclusion.upper()}   (nivel de confianza: {r.confianza}%)

7. ACCION RECOMENDADA
   {
        "Solicitar LEVANTAMIENTO de la glosa con la evidencia relacionada."
        if r.conclusion == "No procede"
        else "Aportar el/los documento(s) faltante(s) y solicitar levantamiento parcial."
        if r.conclusion == "Procede parcialmente"
        else "Recolectar el soporte faltante antes de responder."
        if r.conclusion == "Informacion insuficiente"
        else "Mantener la glosa."
    }

NOTA: respuesta generada por el asistente a partir de los soportes hallados.
Debe ser revisada y avalada por el auditor antes de su envio. No se anexa
evidencia inexistente; los documentos 'NO' deben conseguirse."""


# ---------------------------------------------------------------------------
# Carga de glosas desde el Excel (layout HUS.xlsx)
# ---------------------------------------------------------------------------
def cargar_glosas(ruta_excel: Path, solo: str | None) -> list[Glosa]:
    import openpyxl

    wb = openpyxl.load_workbook(str(ruta_excel), data_only=True, read_only=True)
    ws = wb["Hoja1"] if "Hoja1" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    H = {h: i for i, h in enumerate(rows[0])}

    def g(row, col):
        return row[H[col]] if col in H else None

    solo_corta = factura_corta(solo) if solo else None
    glosas = []
    for r in rows[1:]:
        if not g(r, "FACTURA"):
            continue
        if solo_corta and factura_corta(g(r, "FACTURA")) != solo_corta:
            continue
        cod = str(g(r, "CODIGO CONCEPTO DE GLOSA") or "")
        glosas.append(
            Glosa(
                factura=norm_factura(g(r, "FACTURA")),
                radicado=str(g(r, "RADICADO") or ""),
                fecha_atencion=g(r, "FECHA ATENCION")
                if isinstance(g(r, "FECHA ATENCION"), datetime)
                else None,
                paciente=str(g(r, "NOMBRES Y APELLIDOS") or ""),
                identificacion=str(g(r, "IDENTIFICACION") or ""),
                valor_factura=_num(g(r, "VALOR FACTURA")),
                num_glosa=str(g(r, "# GLOSA") or ""),
                codigo=cod.strip(),
                familia=familia_de(cod),
                motivo=str(g(r, "MOTIVO DE GLOSA") or ""),
                valor_objetado=_num(g(r, "VALOR OBJETADO")),
                servicio=str(g(r, "SERVICIO OBJETADO") or ""),
            )
        )
    return glosas


# ---------------------------------------------------------------------------
# Orquestación por glosa
# ---------------------------------------------------------------------------
def construir_matriz(g: Glosa, docs: list[Documento]) -> list[dict]:
    """Matriz de evidencia: por cada tipo requerido/hallado, encontrado + ruta + pagina."""
    requeridos = SOPORTE_POR_FAMILIA.get(g.familia, [])
    tipos = list(dict.fromkeys(requeridos + [d.tipo for d in docs]))
    filas = []
    for t in tipos:
        d = next((x for x in docs if x.tipo == t), None)
        filas.append(
            {
                "tipo": t,
                "encontrado": d is not None,
                "ruta": str(d.ruta) if d else "",
                "pagina": (d.paginas if d and d.paginas else ""),
                "requiere_ocr": bool(d and d.requiere_ocr),
            }
        )
    return filas


def procesar(glosas: list[Glosa], idx: dict[str, list[Documento]], kb: dict) -> list[Resultado]:
    resultados = []
    for g in glosas:
        docs = idx.get(factura_corta(g.factura), [])
        matriz = construir_matriz(g, docs)
        checks = verificar_coherencia(g, docs)
        contrato = contrato_por_fecha(g.fecha_atencion)
        norma = NORMA_FAMILIA.get(g.familia, "Res. 2284/2023")
        # Precedente en base de conocimiento (caso análogo)
        clave = f"{codigo_corto(g.codigo)}|{re.findall(r'[0-9]{4,7}', g.servicio)[:1]}"
        if clave in kb:
            checks.append(
                f"PRECEDENTE: caso analogo resuelto antes -> {kb[clave]['conclusion']} (verificar hechos/norma)"
            )
        conclusion, confianza = concluir(g, docs, matriz, checks)
        r = Resultado(
            glosa=g,
            contrato=contrato,
            matriz=matriz,
            coherencia=checks,
            conclusion=conclusion,
            confianza=confianza,
            norma=norma,
            oficio="",
        )
        r.oficio = redactar_oficio(r)
        resultados.append(r)
    return resultados


# ---------------------------------------------------------------------------
# Salida: Excel (matriz + respuestas) y oficios opcionales
# ---------------------------------------------------------------------------
def escribir_excel(resultados: list[Resultado], salida: Path) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    # Hoja Respuestas
    ws = wb.active
    ws.title = "Respuestas"
    cols = [
        "Factura",
        "Radicado",
        "Paciente",
        "Glosa",
        "Familia",
        "Valor objetado",
        "Contrato",
        "Conclusion",
        "Confianza %",
        "Norma",
        "Docs hallados",
        "Docs faltantes",
        "Alertas coherencia",
    ]
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E5F")
    fill_con = {
        "No procede": "D9EAD3",
        "Procede parcialmente": "FFF2CC",
        "Informacion insuficiente": "F4CCCC",
        "Procede": "F4CCCC",
    }
    for r in resultados:
        hall = [m["tipo"] for m in r.matriz if m["encontrado"]]
        falt = [m["tipo"] for m in r.matriz if not m["encontrado"]]
        alertas = [c for c in r.coherencia if c.startswith("ALERTA")]
        ws.append(
            [
                r.glosa.factura,
                r.glosa.radicado,
                r.glosa.paciente,
                r.glosa.codigo,
                r.glosa.familia,
                r.glosa.valor_objetado,
                r.contrato,
                r.conclusion,
                r.confianza,
                r.norma,
                "; ".join(hall),
                "; ".join(falt),
                " | ".join(alertas),
            ]
        )
        row = ws.max_row
        ws.cell(row=row, column=6).number_format = "#,##0"
        cc = ws.cell(row=row, column=8)
        if r.conclusion in fill_con:
            cc.fill = PatternFill("solid", fgColor=fill_con[r.conclusion])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{ws.max_row}"
    # Hoja Matriz_Evidencia (una fila por documento)
    ws2 = wb.create_sheet("Matriz_Evidencia")
    ws2.append(
        ["Factura", "Glosa", "Tipo documento", "Encontrado", "Pagina/Pags", "Requiere OCR", "Ruta"]
    )
    for c in ws2[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E6E86")
    for r in resultados:
        for m in r.matriz:
            ws2.append(
                [
                    r.glosa.factura,
                    r.glosa.codigo,
                    m["tipo"],
                    "SI" if m["encontrado"] else "NO",
                    m["pagina"],
                    "SI" if m["requiere_ocr"] else "",
                    m["ruta"],
                ]
            )
    ws2.freeze_panes = "A2"
    for w, col in zip((16, 12, 34, 11, 12, 12, 70), "ABCDEFG"):
        ws2.column_dimensions[col].width = w
    for w, col in zip((16, 12, 26, 12, 12, 44, 55, 12, 40, 26, 30, 30, 50), "ABCDEFGHIJKLM"):
        ws.column_dimensions[col].width = w
    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(salida))


def escribir_oficios(resultados: list[Resultado], carpeta: Path) -> None:
    carpeta.mkdir(parents=True, exist_ok=True)
    for r in resultados:
        nombre = f"OFICIO_{r.glosa.factura}_{codigo_corto(r.glosa.codigo)}_{r.glosa.num_glosa}.txt"
        (carpeta / nombre).write_text(r.oficio, encoding="utf-8")


# ---------------------------------------------------------------------------
# Base de conocimiento (precedentes)
# ---------------------------------------------------------------------------
def cargar_kb(ruta: Path) -> dict:
    if ruta and ruta.exists():
        try:
            return json.loads(ruta.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return {}
    return {}


def actualizar_kb(kb: dict, resultados: list[Resultado], ruta: Path) -> None:
    for r in resultados:
        if r.conclusion in ("No procede", "Procede parcialmente") and r.confianza >= 80:
            cs = re.findall(r"[0-9]{4,7}", r.glosa.servicio)[:1]
            clave = f"{codigo_corto(r.glosa.codigo)}|{cs}"
            kb[clave] = {
                "conclusion": r.conclusion,
                "norma": r.norma,
                "familia": r.glosa.familia,
                "confianza": r.confianza,
            }
    if ruta:
        ruta.write_text(json.dumps(kb, ensure_ascii=False, indent=2), encoding="utf-8")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main(argv: list[str] | None = None) -> int:
    setup_logging()
    ap = argparse.ArgumentParser(description="Asistente de conciliacion de glosas del Dispensario.")
    ap.add_argument("--excel", type=Path, required=True, help="Excel de glosas (layout HUS.xlsx)")
    ap.add_argument(
        "--soportes-raiz",
        type=Path,
        default=None,
        help="Carpeta raiz de soportes (Y:\\...). Recorre la unidad. Alternativa: --indice.",
    )
    ap.add_argument(
        "--indice",
        type=Path,
        default=None,
        help="Indice JSON de soportes (indexar_soportes_dispensario.py). Evita recorrer Y:\\.",
    )
    ap.add_argument(
        "--salida", type=Path, required=True, help="Excel de salida (matriz + respuestas)"
    )
    ap.add_argument(
        "--solo", default=None, help="Procesar una sola factura (piloto), ej. HUS524435"
    )
    ap.add_argument(
        "--con-oficios",
        type=Path,
        default=None,
        help="Carpeta para escribir un oficio .txt por glosa",
    )
    ap.add_argument(
        "--kb", type=Path, default=None, help="Ruta del JSON de base de conocimiento (precedentes)"
    )
    ap.add_argument(
        "--ocr",
        action="store_true",
        help="Activar OCR para PDFs escaneados (requiere pytesseract+pdf2image)",
    )
    args = ap.parse_args(argv)

    if not args.soportes_raiz and not args.indice:
        ap.error("indica --indice <json> (recomendado) o --soportes-raiz <carpeta>")

    glosas = cargar_glosas(args.excel, args.solo)
    if not glosas:
        logger.error("No se cargaron glosas del Excel (¿--solo con factura inexistente?)")
        return 1
    logger.info("Glosas a procesar: %d", len(glosas))

    facturas = {factura_corta(g.factura) for g in glosas}
    if args.indice:
        idx = docs_desde_indice(args.indice, facturas, args.ocr)
    else:
        idx = indexar_soportes(args.soportes_raiz, facturas, args.ocr)
    kb = cargar_kb(args.kb) if args.kb else {}
    resultados = procesar(glosas, idx, kb)

    escribir_excel(resultados, args.salida)
    logger.info("Excel de resultados: %s", args.salida)
    if args.con_oficios:
        escribir_oficios(resultados, args.con_oficios)
        logger.info("Oficios en: %s", args.con_oficios)
    if args.kb:
        actualizar_kb(kb, resultados, args.kb)
        logger.info("Base de conocimiento actualizada: %s", args.kb)

    # Resumen
    from collections import Counter

    resumen = Counter(r.conclusion for r in resultados)
    for k, v in resumen.most_common():
        logger.info("  %-26s %d", k, v)
    return 0


if __name__ == "__main__":
    sys.exit(main())
