#!/usr/bin/env python3
"""siifa_informe_del_cargue.py — Qué quedó cargado, probado contra SIIFA.

POR QUÉ EXISTE, Y EN QUÉ SE DIFERENCIA DEL OTRO VERIFICADOR. Que el bot diga
OK significa que la API contestó bien; lo que vale en una conciliación es lo
que SIIFA tiene guardado. `siifa_verificar_cargue.py` lo comprueba consultando
FACTURA POR FACTURA, que sirve para un cargue de 17 facturas pero no para uno
de 12.255: serían 12.255 consultas. Este parte del informe masivo —una sola
bajada— y cruza contra los reportes del cargue.

QUÉ RESPONDE:
  · Cuántas respuestas se mandaron, cuántas quedaron REGISTRADAS en SIIFA y
    cuántas fallaron.
  · Las que el bot dio por buenas pero SIIFA no tiene — el caso peligroso: se
    daría por respondido algo que sigue vencido.
  · El desglose por entidad pagadora, por tipo y por causal, con valores.
  · Los errores agrupados por causa, para saber qué reintentar.

USO (después del cargue, con el informe recién bajado):

    py siifa_reporte_seguimientos.py --ips SOCORRO ^
        --salida "D:\\...\\SOCORRO\\informe_DESPUES.xlsx"

    py siifa_informe_del_cargue.py --ips SOCORRO ^
        --informe "D:\\...\\SOCORRO\\informe_DESPUES.xlsx" ^
        --reporte "D:\\...\\SOCORRO\\reporte_cargue.csv" ^
        --salida  "D:\\...\\SOCORRO\\INFORME_DEL_CARGUE.xlsx"

`--reporte` se puede repetir, para juntar varias tandas en un solo informe.

INSTALACIÓN (una vez):
    py -m pip install openpyxl
"""

from __future__ import annotations

import argparse
import csv
import sys
from collections import Counter, defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from _dinero import a_entero  # noqa: E402
import siifa_perfiles as perfiles  # noqa: E402
from siifa_novedades import leer_informe, valor_total  # noqa: E402

# Los tres estados en que puede quedar una respuesta mandada.
REGISTRADA = "REGISTRADA EN SIIFA"
NO_QUEDO = "EL BOT DIJO OK PERO SIIFA NO LA TIENE"
CON_ERROR = "NO SE PUDO CARGAR"

COLUMNAS = [
    "estado",
    "numero_factura",
    "tipo_seguimiento",
    "razon_social_eps",
    "valor_glosa",
    "codigo_glosa",
    "codigo_respuesta",
    "fecha_respuesta",
    "detalle",
    "id_seguimiento_factura_glosa",
]


def leer_reportes(rutas: list[Path]) -> dict[int, dict]:
    """Lo que dijo el bot de cada respuesta. La última corrida manda.

    Si una glosa falló en la primera tanda y salió bien en el reintento, lo
    que cuenta es el reintento: por eso no se acumula, se reemplaza.
    """
    por_id: dict[int, dict] = {}
    for ruta in rutas:
        if not ruta.exists():
            raise SystemExit(f"\nNo encuentro el reporte del cargue:\n  {ruta}\n")
        with open(ruta, newline="", encoding="utf-8") as f:
            for fila in csv.DictReader(f):
                try:
                    ident = int(fila.get("id_seguimiento_factura_glosa", ""))
                except (TypeError, ValueError):
                    continue
                estado = str(fila.get("estado", "")).strip().upper()
                # YA_OK_PREVIO es una fila saltada por --saltar-csv: se cargó
                # en una corrida anterior, así que cuenta como buena.
                por_id[ident] = {
                    "ok": estado in ("OK", "YA_OK_PREVIO"),
                    "detalle": str(fila.get("detalle", "") or "").strip(),
                    "cuando": str(fila.get("fecha_hora", "") or "").strip(),
                }
    return por_id


def cruzar(reportes: dict[int, dict], informe: list[dict]) -> list[dict]:
    """Junta lo que dijo el bot con lo que SIIFA tiene de verdad."""
    en_siifa: dict[int, dict] = {}
    for fila in informe:
        try:
            en_siifa[int(fila.get("id_seguimiento_factura_glosa", ""))] = fila
        except (TypeError, ValueError):
            continue

    salida = []
    for ident, rep in reportes.items():
        siifa = en_siifa.get(ident, {})
        respondida = str(siifa.get("tiene_respuesta", "")).strip().upper() == "SI"
        if not rep["ok"] and respondida:
            # El intento dio error pero SIIFA SÍ la tiene. Pasa de dos formas:
            # un timeout donde la escritura entró y la confirmación se perdió
            # (el reintento recibe «ya tiene un registro previo»), o una fila
            # repetida. Lo que manda es el estado de la plataforma.
            estado = REGISTRADA
            detalle = "El intento dio error, pero SIIFA la tiene registrada"
        elif not rep["ok"]:
            estado, detalle = CON_ERROR, rep["detalle"]
        elif respondida:
            estado, detalle = REGISTRADA, ""
        elif siifa:
            estado, detalle = NO_QUEDO, "El bot reportó OK pero el informe no la trae respondida"
        else:
            # No está en el informe: puede que la EPS la haya reformulado, o
            # que el informe se bajara antes del cargue.
            estado, detalle = (
                NO_QUEDO,
                "No aparece en el informe (¿reformulada por la EPS, o informe viejo?)",
            )
        salida.append(
            {
                "estado": estado,
                "detalle": detalle,
                "id_seguimiento_factura_glosa": ident,
                "numero_factura": siifa.get("numero_factura", ""),
                "tipo_seguimiento": siifa.get("tipo_seguimiento", ""),
                "razon_social_eps": siifa.get("razon_social_eps", ""),
                "valor_glosa": a_entero(siifa.get("valor_glosa")),
                "codigo_glosa": siifa.get("codigo_glosa", ""),
                "codigo_respuesta": siifa.get("codigo_respuesta", ""),
                "fecha_respuesta": str(siifa.get("fecha_respuesta", ""))[:10],
            }
        )
    return salida


def _pesos(valor: object) -> str:
    return f"${a_entero(valor):,}".replace(",", ".")


def _por(filas: list[dict], clave) -> list[tuple]:
    agg: dict[str, list] = defaultdict(list)
    for f in filas:
        agg[str(clave(f) or "(sin dato)")].append(f)
    return sorted(
        (
            (k, len(g), len({str(x["numero_factura"]) for x in g}), valor_total(g))
            for k, g in agg.items()
        ),
        key=lambda t: -t[3],
    )


def mostrar(cruzadas: list[dict], ips) -> None:
    por_estado: dict[str, list] = defaultdict(list)
    for f in cruzadas:
        por_estado[f["estado"]].append(f)

    print()
    print("=" * 82)
    print(f"INFORME DEL CARGUE — {ips.nombre}")
    print("=" * 82)
    print(f"\n  Respuestas mandadas: {len(cruzadas)}\n")
    print(f"  {'Estado':<42}{'Ítems':>9}{'Valor':>18}")
    print("-" * 82)
    for estado in (REGISTRADA, NO_QUEDO, CON_ERROR):
        grupo = por_estado.get(estado, [])
        if grupo:
            print(f"  {estado:<42}{len(grupo):>9}{_pesos(valor_total(grupo)):>18}")
    print("-" * 82)

    buenas = por_estado.get(REGISTRADA, [])
    if buenas:
        print(f"\n  LO QUE QUEDÓ REGISTRADO ({len(buenas)} respuestas)")
        print("-" * 82)
        print(f"  {'Entidad responsable de pago':<44}{'Ítems':>8}{'Facturas':>10}{'Valor':>18}")
        for nombre, n, nf, valor in _por(buenas, lambda f: f["razon_social_eps"])[:8]:
            print(f"  {nombre[:43]:<44}{n:>8}{nf:>10}{_pesos(valor):>18}")
        print(f"\n  {'Tipo':<44}{'Ítems':>8}{'Facturas':>10}{'Valor':>18}")
        for nombre, n, nf, valor in _por(buenas, lambda f: f["tipo_seguimiento"]):
            print(f"  {nombre[:43]:<44}{n:>8}{nf:>10}{_pesos(valor):>18}")

    perdidas = por_estado.get(NO_QUEDO, [])
    if perdidas:
        print(
            f"\n  [!] ATENCIÓN: {len(perdidas)} respuestas que el bot dio por buenas NO están\n"
            f"      registradas en SIIFA ({_pesos(valor_total(perdidas))}). Son las más\n"
            f"      peligrosas: se darían por respondidas y siguen corriendo el plazo."
        )

    errores = por_estado.get(CON_ERROR, [])
    if errores:
        print(f"\n  LO QUE NO SE PUDO CARGAR ({len(errores)})")
        print("-" * 82)
        for detalle, n in Counter(f["detalle"][:70] for f in errores).most_common(6):
            print(f"  {n:>7}  {detalle}")
    print()


def escribir(cruzadas: list[dict], ruta: Path, ips) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise SystemExit("ERROR: falta openpyxl. py -m pip install openpyxl") from None

    azul, blanco = PatternFill("solid", fgColor="1F4E78"), Font(bold=True, color="FFFFFF")
    verde = PatternFill("solid", fgColor="C6EFCE")
    rojo = PatternFill("solid", fgColor="FFC7CE")
    ambar = PatternFill("solid", fgColor="FFF2CC")
    colores = {REGISTRADA: verde, CON_ERROR: rojo, NO_QUEDO: ambar}

    wb = Workbook()
    ws = wb.active
    ws.title = "RESUMEN"
    ws.append([f"INFORME DEL CARGUE — {ips.nombre} (NIT {ips.nit})"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws.append([])
    encabezado = ["Estado", "Ítems", "Facturas", "Valor"]
    ws.append(encabezado)
    for c in range(1, len(encabezado) + 1):
        ws.cell(row=3, column=c).fill = azul
        ws.cell(row=3, column=c).font = blanco
    por_estado: dict[str, list] = defaultdict(list)
    for f in cruzadas:
        por_estado[f["estado"]].append(f)
    for estado in (REGISTRADA, NO_QUEDO, CON_ERROR):
        grupo = por_estado.get(estado, [])
        if grupo:
            ws.append(
                [
                    estado,
                    len(grupo),
                    len({str(x["numero_factura"]) for x in grupo}),
                    valor_total(grupo),
                ]
            )
            ws.cell(row=ws.max_row, column=1).fill = colores[estado]
            ws.cell(row=ws.max_row, column=4).number_format = '"$"#,##0'

    buenas = por_estado.get(REGISTRADA, [])
    if buenas:
        for titulo, clave in (
            ("POR ENTIDAD RESPONSABLE DE PAGO", lambda f: f["razon_social_eps"]),
            ("POR TIPO", lambda f: f["tipo_seguimiento"]),
            ("POR CAUSAL", lambda f: f["codigo_glosa"]),
        ):
            ws.append([])
            ws.append([titulo])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
            ws.append(["", "Ítems", "Facturas", "Valor"])
            for c in range(1, 5):
                ws.cell(row=ws.max_row, column=c).fill = azul
                ws.cell(row=ws.max_row, column=c).font = blanco
            for nombre, n, nf, valor in _por(buenas, clave):
                ws.append([nombre, n, nf, valor])
                ws.cell(row=ws.max_row, column=4).number_format = '"$"#,##0'
    for col, ancho in (("A", 52), ("B", 12), ("C", 12), ("D", 20)):
        ws.column_dimensions[col].width = ancho

    ws2 = wb.create_sheet("DETALLE")
    for i, h in enumerate(COLUMNAS, 1):
        c = ws2.cell(row=1, column=i, value=h)
        c.fill, c.font = azul, blanco
    orden = {REGISTRADA: 2, NO_QUEDO: 0, CON_ERROR: 1}
    for r, f in enumerate(
        sorted(cruzadas, key=lambda x: (orden[x["estado"]], -x["valor_glosa"])), 2
    ):
        for i, h in enumerate(COLUMNAS, 1):
            c = ws2.cell(row=r, column=i, value=f.get(h, ""))
            c.alignment = Alignment(vertical="top", wrap_text=(h == "detalle"))
            if h == "valor_glosa":
                c.number_format = '"$"#,##0'
        ws2.cell(row=r, column=1).fill = colores[f["estado"]]
    anchos = {"estado": 38, "detalle": 50, "razon_social_eps": 32, "numero_factura": 16}
    for i, h in enumerate(COLUMNAS, 1):
        ws2.column_dimensions[get_column_letter(i)].width = anchos.get(h, 16)
    ws2.freeze_panes = "A2"

    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ruta))


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--informe", required=True, help="Informe de SIIFA bajado DESPUÉS del cargue.")
    ap.add_argument(
        "--reporte", action="append", required=True, help="CSV del cargue (se puede repetir)."
    )
    ap.add_argument("--salida", required=True, help="Excel del informe a generar.")
    perfiles.agregar_argumento(ap)
    args = ap.parse_args()
    ips = perfiles.buscar(args.ips)
    perfiles.anunciar(ips)

    informe = leer_informe(Path(args.informe))
    perfiles.verificar_informe(ips, informe)
    reportes = leer_reportes([Path(r) for r in args.reporte])
    if not reportes:
        raise SystemExit("\nLos reportes del cargue no traen ninguna fila con id.\n")

    cruzadas = cruzar(reportes, informe)
    mostrar(cruzadas, ips)
    escribir(cruzadas, Path(args.salida), ips)
    print(f"  Informe guardado: {args.salida}\n")


if __name__ == "__main__":
    main()
