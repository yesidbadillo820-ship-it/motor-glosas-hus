"""
nucleo.msg_tools — Correos de relación de pagos (.msg) → Excel consolidado.

Las entidades pagadoras (Fiduprevisora/PPL, EPS, ADRES...) envían la
"relación de pagos del día" como un correo de Outlook (.msg) con el detalle
en un Excel ADJUNTO. Este módulo abre cada correo, saca ese adjunto y
preserva los valores TAL CUAL vienen (fechas como fechas, montos como
números — nunca como texto), para no perder calidad del dato al consolidar
varios correos en un solo archivo.

Por qué existe: el analista recibe un pago por día en un correo aparte;
para conciliar cartera necesita verlos todos juntos, sumables y filtrables.

Dependencia difícil (decisión documentada):
    extract_msg declara como dependencia obligatoria 'red-black-tree-mod',
    cuyo instalador (setup.py heredado) es INCOMPATIBLE con el setuptools
    moderno: falla con "AttributeError: install_layout" al compilar,
    reproducido con setuptools nuevo tanto del sistema como recién bajado
    de PyPI (no es un problema de un solo sistema operativo). Esa librería
    SOLO la usa extract_msg para RE-ESCRIBIR/guardar un .msg (método
    `.save()`); este módulo NUNCA guarda .msg, solo los LEE. Por eso, en
    vez de instalar ese paquete roto, se inyecta un stub de
    'red_black_dict_mod' en sys.modules ANTES de importar extract_msg: el
    import queda satisfecho, y si algún día se intentara usar esa función
    de guardado (que este módulo no expone), fallaría con un mensaje claro
    en vez de silenciosamente.
"""

import glob
import io
import os
import sys
import tempfile
import types
import zipfile


def _parchear_red_black_stub():
    if "red_black_dict_mod" in sys.modules:
        return
    stub = types.ModuleType("red_black_dict_mod")

    class RedBlackTree:
        def __init__(self, *a, **k):
            raise NotImplementedError(
                "red-black-tree-mod no está instalado (solo hace falta para "
                "RE-ESCRIBIR archivos .msg; esta herramienta solo los lee)."
            )

    stub.RedBlackTree = RedBlackTree
    sys.modules["red_black_dict_mod"] = stub


try:
    _parchear_red_black_stub()
    import extract_msg
except ImportError:
    extract_msg = None

MENSAJE_SIN_EXTRACT_MSG = (
    "Falta el lector de correos Outlook (.msg). Ejecute INICIAR_SUITE.bat "
    "con internet una vez, o instale manualmente en dos pasos:\n"
    "  pip install --no-deps extract-msg\n"
    "  pip install olefile beautifulsoup4 compressed-rtf ebcdic RTFDE tzlocal"
)


def _exigir_extract_msg():
    if extract_msg is None:
        raise RuntimeError(MENSAJE_SIN_EXTRACT_MSG)


# ------------------------------------------------------------------ lectura


def _limpiar_nombre(nombre):
    """Los nombres de adjunto de extract_msg a veces traen un NUL final."""
    return (nombre or "").rstrip("\x00").strip()


def leer_adjuntos_pago(ruta_msg, log=print):
    """Abre un .msg y devuelve (metadata_correo, [(nombre_adjunto, bytes), ...]).

    Solo se devuelven adjuntos .xlsx/.xls (los pagos siempre vienen así).
    """
    _exigir_extract_msg()
    m = extract_msg.openMsg(ruta_msg)
    try:
        meta = dict(
            archivo=os.path.basename(ruta_msg),
            asunto=(m.subject or "").strip(),
            remitente=str(m.sender or "").strip(),
            fecha_envio=m.date,
        )
        adjuntos = []
        for a in m.attachments:
            try:
                nombre = _limpiar_nombre(a.getFilename())
            except Exception:
                continue
            if nombre.lower().endswith((".xlsx", ".xls")):
                adjuntos.append((nombre, a.data))
        if not adjuntos:
            log("  [!] %s: sin adjunto Excel (revisar manualmente)" % meta["archivo"])
        return meta, adjuntos
    finally:
        m.close()


def _es_fila_con_datos(fila_dict, col_factura):
    v = fila_dict.get(col_factura)
    return v is not None and str(v).strip() != ""


def filas_desde_adjunto(datos_xlsx, nombre_adjunto):
    """Lee el Excel adjunto y devuelve (encabezados, [fila_dict, ...]).

    Preserva los tipos nativos de openpyxl (datetime, float, int, str): no
    se convierte nada a texto, así no se pierde precisión ni formato.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(datos_xlsx), data_only=True)
    ws = wb[wb.sheetnames[0]]

    fila_hdr = None
    for r in range(1, min(ws.max_row, 10) + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if sum(1 for v in vals if v not in (None, "")) >= 3:
            fila_hdr = r
            break
    if fila_hdr is None:
        raise ValueError("No se encontró fila de encabezado en %s" % nombre_adjunto)

    encabezados = []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(fila_hdr, c).value
        encabezados.append(str(h).strip() if h is not None else "COL_%d" % c)

    col_factura = next((h for h in encabezados if "FACTURA" in h.upper()), encabezados[0])

    filas = []
    for r in range(fila_hdr + 1, ws.max_row + 1):
        fila = {encabezados[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        if _es_fila_con_datos(fila, col_factura):
            filas.append(fila)
    return encabezados, filas


def _resolver_entradas(entradas):
    """Acepta rutas .msg sueltas, una carpeta, o un .zip; devuelve rutas .msg."""
    rutas = []
    for e in entradas:
        if os.path.isdir(e):
            rutas.extend(sorted(glob.glob(os.path.join(e, "*.msg"))))
        elif e.lower().endswith(".zip"):
            carpeta_tmp = tempfile.mkdtemp(prefix="msgtools_")
            with zipfile.ZipFile(e) as z:
                for nombre in z.namelist():
                    if nombre.lower().endswith(".msg"):
                        z.extract(nombre, carpeta_tmp)
                        rutas.append(os.path.join(carpeta_tmp, nombre))
        elif e.lower().endswith(".msg"):
            rutas.append(e)
    return rutas


def _fmt_fecha_correo(dt):
    if dt is None:
        return None
    try:
        return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(dt)


# --------------------------------------------------------------- consolidar


def consolidar(entradas, salida, log=print):
    """Consolida N correos .msg de relación de pagos en UN Excel.

    `entradas`: lista de rutas .msg, carpetas con .msg, y/o .zip que los
    contenga (se puede mezclar). Devuelve la ruta del Excel generado.
    """
    _exigir_extract_msg()
    rutas = _resolver_entradas(entradas)
    if not rutas:
        raise ValueError("No se encontraron archivos .msg en la entrada.")

    encabezados_ref = None
    todas_filas = []
    resumen_por_correo = []
    sin_adjunto = []

    for ruta in rutas:
        log("Leyendo %s…" % os.path.basename(ruta))
        meta, adjuntos = leer_adjuntos_pago(ruta, log=log)
        if not adjuntos:
            sin_adjunto.append(meta["archivo"])
            continue
        filas_correo = 0
        for nombre_adj, datos in adjuntos:
            try:
                encabezados, filas = filas_desde_adjunto(datos, nombre_adj)
            except Exception as e:
                log("  [!] No se pudo leer el adjunto %s: %s" % (nombre_adj, e))
                continue
            if encabezados_ref is None:
                encabezados_ref = encabezados
            elif encabezados != encabezados_ref:
                log(
                    "  [!] %s: columnas distintas al primer archivo (se "
                    "conservan igual; revisar antes de sumar)." % nombre_adj
                )
            for fila in filas:
                fila["_ORIGEN_CORREO"] = meta["archivo"]
                fila["_ORIGEN_ADJUNTO"] = nombre_adj
                fila["_FECHA_ENVIO_CORREO"] = _fmt_fecha_correo(meta["fecha_envio"])
                todas_filas.append(fila)
                filas_correo += 1
        resumen_por_correo.append((meta["archivo"], filas_correo))

    if not todas_filas:
        raise ValueError("Ningún adjunto trajo filas de pago legibles.")

    col_factura = next((h for h in encabezados_ref if "FACTURA" in h.upper()), encabezados_ref[0])
    col_comprobante = next((h for h in encabezados_ref if "COMPROBANTE" in h.upper()), None)
    col_valor_pagado = next((h for h in encabezados_ref if h.upper() == "VALOR PAGADO"), None)
    if col_valor_pagado is None:
        log("  [!] No se encontró columna 'VALOR PAGADO'; los totales de RESUMEN quedarán en 0.")

    # Marcar posibles duplicados (misma factura+comprobante+valor pagado en
    # más de un correo): NO se eliminan, solo se marcan — "sin perder ningún
    # dato" incluye no borrar nada sin que el analista lo decida.
    contador = {}
    for fila in todas_filas:
        clave = (fila.get(col_factura), fila.get(col_comprobante), fila.get(col_valor_pagado))
        contador[clave] = contador.get(clave, 0) + 1
    for fila in todas_filas:
        clave = (fila.get(col_factura), fila.get(col_comprobante), fila.get(col_valor_pagado))
        fila["_POSIBLE_DUPLICADO"] = "SI" if contador[clave] > 1 else "-"

    ruta_salida = _escribir_consolidado(
        encabezados_ref, todas_filas, resumen_por_correo, sin_adjunto, salida, col_valor_pagado
    )
    total_pagado = sum(float(f.get(col_valor_pagado) or 0) for f in todas_filas)
    log(
        "Consolidado: %d correos, %d filas, $%s pagado → %s"
        % (len(rutas), len(todas_filas), format(int(total_pagado), ","), ruta_salida)
    )
    return ruta_salida


def _escribir_consolidado(
    encabezados, filas, resumen_por_correo, sin_adjunto, salida, col_valor_pagado
):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    FMT_CONT = '_-* #,##0_-;\\-* #,##0_-;_-* "-"??_-;_-@_-'
    FILL_HDR = PatternFill("solid", fgColor="1F4E79")
    FONT_HDR = Font(bold=True, color="FFFFFF")
    FINO = Side(style="thin", color="B7C6D6")
    BORDE = Border(left=FINO, right=FINO, top=FINO, bottom=FINO)

    columnas_finales = list(encabezados) + [
        "_ORIGEN_CORREO",
        "_ORIGEN_ADJUNTO",
        "_FECHA_ENVIO_CORREO",
        "_POSIBLE_DUPLICADO",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "CONSOLIDADO"
    for j, h in enumerate(columnas_finales, 1):
        c = ws.cell(1, j, h)
        c.font = FONT_HDR
        c.fill = FILL_HDR
        c.border = BORDE
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cols_fecha = {h for h in encabezados if "FECHA" in h.upper() and "TEXTO" not in h.upper()}
    cols_valor = {
        h
        for h in encabezados
        if h.upper().startswith("VALOR") or "PAGOS EN RUTA" in h.upper() or "IMPUESTO" in h.upper()
    }

    for i, fila in enumerate(filas, 2):
        for j, h in enumerate(columnas_finales, 1):
            v = fila.get(h)
            c = ws.cell(i, j, v)
            c.border = BORDE
            if h in cols_fecha and v is not None:
                c.number_format = "dd/mm/yyyy"
            elif h in cols_valor and isinstance(v, (int, float)):
                c.number_format = FMT_CONT

    for j, h in enumerate(columnas_finales, 1):
        ancho = min(38, max(12, len(h) + 2))
        ws.column_dimensions[get_column_letter(j)].width = ancho
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(columnas_finales)), len(filas) + 1)

    ws2 = wb.create_sheet("RESUMEN")
    ws2.append(["CORREO (archivo .msg)", "FILAS"])
    for j in (1, 2):
        ws2.cell(1, j).font = FONT_HDR
        ws2.cell(1, j).fill = FILL_HDR
    fila_r = 2
    for nombre, n in resumen_por_correo:
        ws2.cell(fila_r, 1, nombre)
        ws2.cell(fila_r, 2, n)
        fila_r += 1
    ws2.cell(fila_r, 1, "TOTAL").font = Font(bold=True)
    ws2.cell(fila_r, 2, "=SUM(B2:B%d)" % (fila_r - 1)).font = Font(bold=True)
    fila_r += 2
    ws2.cell(fila_r, 1, "VALOR PAGADO TOTAL").font = Font(bold=True)
    total_cell = ws2.cell(fila_r, 2, sum(float(f.get(col_valor_pagado) or 0) for f in filas))
    total_cell.number_format = FMT_CONT
    total_cell.font = Font(bold=True)
    if sin_adjunto:
        fila_r += 2
        ws2.cell(fila_r, 1, "Correos SIN adjunto Excel (revisar manualmente):").font = Font(
            bold=True, color="C00000"
        )
        for nombre in sin_adjunto:
            fila_r += 1
            ws2.cell(fila_r, 1, nombre)
    ws2.column_dimensions["A"].width = 55
    ws2.column_dimensions["B"].width = 14

    os.makedirs(os.path.dirname(os.path.abspath(salida)) or ".", exist_ok=True)
    wb.save(salida)
    return salida
