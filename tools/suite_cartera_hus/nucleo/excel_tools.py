"""
nucleo.excel_tools — Unir varios Excel en uno solo, sin dañar el dato.

Dos formas de unir (el analista elige):
  - "apilar": los archivos traen las MISMAS columnas (o casi) y se quiere una
    sola tabla con todas las filas, una debajo de otra — el caso típico de los
    cortes mensuales de DGH o los exportes del portal. Cada fila queda marcada
    con el archivo del que salió, y si un archivo trae columnas nuevas se
    agregan al final (nunca se pierde una columna ni una fila).
  - "hojas": cada archivo queda como una hoja aparte dentro de un solo libro
    (útil para entregar "todo en un archivo" sin mezclar tablas distintas).

Regla de oro (la misma de msg_tools): los valores se copian TAL CUAL vienen
(fechas como fechas, montos como números — nunca como texto), para que el
resultado se pueda sumar y filtrar directo.
"""

import glob
import os
import re
import tempfile
import zipfile

EXTENSIONES = (".xlsx", ".xlsm")


def _resolver_entradas(entradas):
    """Acepta rutas .xlsx/.xlsm sueltas, carpetas, o un .zip; devuelve rutas."""
    rutas = []
    for e in entradas:
        if os.path.isdir(e):
            for ext in EXTENSIONES:
                rutas.extend(sorted(glob.glob(os.path.join(e, "*" + ext))))
        elif e.lower().endswith(".zip"):
            carpeta_tmp = tempfile.mkdtemp(prefix="exceltools_")
            with zipfile.ZipFile(e) as z:
                for nombre in z.namelist():
                    if nombre.lower().endswith(EXTENSIONES):
                        z.extract(nombre, carpeta_tmp)
                        rutas.append(os.path.join(carpeta_tmp, nombre))
        elif e.lower().endswith(EXTENSIONES):
            rutas.append(e)
    # los temporales de Excel abiertos (~$archivo.xlsx) no son archivos reales
    return [r for r in rutas if not os.path.basename(r).startswith("~$")]


def _fila_encabezado(ws):
    """Primera fila con 2+ celdas llenas (salta títulos sueltos); si no hay,
    la primera con algo. Devuelve el número de fila o None si está vacía."""
    primera_con_algo = None
    for r in range(1, min(ws.max_row, 10) + 1):
        llenas = sum(
            1 for c in range(1, ws.max_column + 1) if ws.cell(r, c).value not in (None, "")
        )
        if llenas >= 2:
            return r
        if llenas >= 1 and primera_con_algo is None:
            primera_con_algo = r
    return primera_con_algo


def _leer_tabla(ruta, log=print):
    """Lee la primera hoja y devuelve (encabezados, [fila_dict, ...]).

    Preserva los tipos nativos de openpyxl (datetime, float, int, str).
    """
    import openpyxl

    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        fila_hdr = _fila_encabezado(ws)
        if fila_hdr is None:
            raise ValueError("La primera hoja de %s está vacía." % os.path.basename(ruta))

        encabezados = []
        for c in range(1, ws.max_column + 1):
            h = ws.cell(fila_hdr, c).value
            encabezados.append(str(h).strip() if h not in (None, "") else "COL_%d" % c)

        filas = []
        for r in range(fila_hdr + 1, ws.max_row + 1):
            fila = {encabezados[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
            if any(v not in (None, "") for v in fila.values()):
                filas.append(fila)
        return encabezados, filas
    finally:
        wb.close()


def _nombre_hoja(nombre_archivo, usados):
    """Nombre de hoja válido para Excel (31 chars, sin caracteres prohibidos)."""
    base = os.path.splitext(os.path.basename(nombre_archivo))[0]
    base = re.sub(r"[\[\]:*?/\\]", "_", base).strip() or "HOJA"
    nombre = base[:31]
    n = 2
    while nombre in usados:
        sufijo = "_%d" % n
        nombre = base[: 31 - len(sufijo)] + sufijo
        n += 1
    usados.add(nombre)
    return nombre


def unir(entradas, salida, modo="apilar", log=print):
    """Une N archivos Excel en UNO. `modo`: "apilar" u "hojas".

    `entradas`: lista de rutas .xlsx/.xlsm, carpetas, y/o .zip (se puede
    mezclar). Devuelve la ruta del Excel generado.
    """
    rutas = _resolver_entradas(entradas)
    if not rutas:
        raise ValueError("No se encontraron archivos Excel (.xlsx/.xlsm) en la entrada.")
    if modo not in ("apilar", "hojas"):
        raise ValueError("Modo desconocido: %r (use 'apilar' u 'hojas')." % modo)

    if modo == "hojas":
        return _unir_hojas(rutas, salida, log=log)
    return _unir_apilando(rutas, salida, log=log)


def _unir_apilando(rutas, salida, log=print):
    columnas = []  # orden: las del primer archivo, luego las nuevas según aparecen
    todas_filas = []
    resumen = []

    for ruta in rutas:
        nombre = os.path.basename(ruta)
        log("Leyendo %s…" % nombre)
        try:
            encabezados, filas = _leer_tabla(ruta, log=log)
        except Exception as e:
            log("  [!] No se pudo leer %s: %s" % (nombre, e))
            resumen.append((nombre, 0, "ERROR: %s" % e))
            continue
        nuevas = [h for h in encabezados if h not in columnas]
        if columnas and nuevas:
            log("  [!] %s trae columnas nuevas: %s (se agregan al final)" % (nombre, nuevas))
        columnas.extend(nuevas)
        for fila in filas:
            fila["_ARCHIVO_ORIGEN"] = nombre
            todas_filas.append(fila)
        resumen.append((nombre, len(filas), ""))

    if not todas_filas:
        raise ValueError("Ningún archivo trajo filas legibles.")

    _escribir_apilado(columnas + ["_ARCHIVO_ORIGEN"], todas_filas, resumen, salida)
    log("Unidos: %d archivos, %d filas → %s" % (len(rutas), len(todas_filas), salida))
    return salida


def _escribir_apilado(columnas, filas, resumen, salida):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    FILL_HDR = PatternFill("solid", fgColor="1F4E79")
    FONT_HDR = Font(bold=True, color="FFFFFF")
    FINO = Side(style="thin", color="B7C6D6")
    BORDE = Border(left=FINO, right=FINO, top=FINO, bottom=FINO)

    wb = Workbook()
    ws = wb.active
    ws.title = "UNIDO"
    for j, h in enumerate(columnas, 1):
        c = ws.cell(1, j, h)
        c.font = FONT_HDR
        c.fill = FILL_HDR
        c.border = BORDE
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, fila in enumerate(filas, 2):
        for j, h in enumerate(columnas, 1):
            v = fila.get(h)
            c = ws.cell(i, j, v)
            c.border = BORDE
            if hasattr(v, "year") and hasattr(v, "month"):
                c.number_format = "dd/mm/yyyy"

    for j, h in enumerate(columnas, 1):
        ws.column_dimensions[get_column_letter(j)].width = min(38, max(12, len(h) + 2))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(columnas)), len(filas) + 1)

    ws2 = wb.create_sheet("RESUMEN")
    ws2.append(["ARCHIVO", "FILAS", "NOTA"])
    for j in (1, 2, 3):
        ws2.cell(1, j).font = FONT_HDR
        ws2.cell(1, j).fill = FILL_HDR
    fila_r = 2
    for nombre, n, nota in resumen:
        ws2.cell(fila_r, 1, nombre)
        ws2.cell(fila_r, 2, n)
        if nota:
            ws2.cell(fila_r, 3, nota).font = Font(color="C00000")
        fila_r += 1
    ws2.cell(fila_r, 1, "TOTAL").font = Font(bold=True)
    ws2.cell(fila_r, 2, "=SUM(B2:B%d)" % (fila_r - 1)).font = Font(bold=True)
    ws2.column_dimensions["A"].width = 55
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 45

    os.makedirs(os.path.dirname(os.path.abspath(salida)) or ".", exist_ok=True)
    wb.save(salida)
    return salida


def _unir_hojas(rutas, salida, log=print):
    import openpyxl
    from openpyxl import Workbook

    wb_out = Workbook()
    wb_out.remove(wb_out.active)
    usados = set()
    copiados = 0

    for ruta in rutas:
        nombre = os.path.basename(ruta)
        log("Copiando %s…" % nombre)
        try:
            wb_in = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
        except Exception as e:
            log("  [!] No se pudo leer %s: %s" % (nombre, e))
            continue
        try:
            ws_in = wb_in[wb_in.sheetnames[0]]
            ws_out = wb_out.create_sheet(_nombre_hoja(nombre, usados))
            for fila in ws_in.iter_rows(values_only=True):
                ws_out.append(list(fila))
            copiados += 1
        finally:
            wb_in.close()

    if not copiados:
        raise ValueError("Ningún archivo se pudo copiar como hoja.")

    os.makedirs(os.path.dirname(os.path.abspath(salida)) or ".", exist_ok=True)
    wb_out.save(salida)
    log("Unidos: %d archivos, cada uno en su hoja → %s" % (copiados, salida))
    return salida
