"""exportar_conciliacion_dispensario.py — Genera CONCILIACION.xlsx desde una
carpeta de piloto YA corrida, sin volver a leer los PDF.

Lee las carpetas por expediente que dejo el piloto (expediente.json /
decision.json / hechos.json) y arma la hoja de trabajo consolidada (una fila por
glosa). El valor objetado lo toma del HUS.xlsx (no queda en los JSON del piloto).

USO
---
    py tools\\exportar_conciliacion_dispensario.py ^
        --piloto-dir "D:\\USUARIO CARTERA\\Piloto6" ^
        --excel      "D:\\USUARIO CARTERA\\Downloads\\HUS.xlsx" ^
        --salida     "D:\\USUARIO CARTERA\\Piloto6\\CONCILIACION.xlsx"
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from collections import deque
from pathlib import Path

_TOOLS = Path(__file__).resolve().parent
if str(_TOOLS) not in sys.path:
    sys.path.insert(0, str(_TOOLS))
from asistente_conciliacion_dispensario import (  # noqa: E402
    cargar_glosas,
    codigo_corto,
    factura_corta,
)

logger = logging.getLogger("exportar_conciliacion")

COLOR = {
    "SOLICITAR_LEVANTAMIENTO": "D9EAD3",
    "ACEPTAR_PARCIAL": "FFF2CC",
    "SOLICITAR_SOPORTE": "FCE5CD",
    "CONCILIAR": "F4CCCC",
    "ESCALAR": "EAD1DC",
}


def _load(ruta: Path) -> dict:
    try:
        return json.loads(ruta.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def _valores_por_glosa(excel: Path) -> dict:
    """(factura_corta, codigo_corto, servicio) -> cola de valores objetados.
    Se usa una cola para respetar glosas repetidas (mismo servicio y valor)."""
    valores: dict = {}
    for g in cargar_glosas(excel, None):
        clave = (factura_corta(g.factura), codigo_corto(g.codigo), str(g.servicio))
        try:
            v = int(float(g.valor_objetado))
        except (TypeError, ValueError):
            v = 0
        valores.setdefault(clave, deque()).append(v)
    return valores


def construir_filas(piloto_dir: Path, valores: dict) -> list[dict]:
    filas: list[dict] = []
    for carpeta in sorted(p for p in piloto_dir.iterdir() if p.is_dir()):
        exp = _load(carpeta / "expediente.json")
        dec = _load(carpeta / "decision.json")
        hec = _load(carpeta / "hechos.json")
        if not dec.get("glosas"):
            continue
        factura = exp.get("factura", carpeta.name)
        fc = factura_corta(factura)
        paciente = exp.get("paciente", {}).get("nombre", "")
        contrato = exp.get("contrato", {}).get("codigo", "")
        saldo = exp.get("cartera", {}).get("saldo", 0)
        edad = exp.get("cartera", {}).get("edad_dias", 0)
        alertas = "; ".join(hec.get("alertas", []))
        hglosas = hec.get("glosas", [])
        for i, g in enumerate(dec.get("glosas", [])):
            d = g.get("decision", {})
            servicio = str(g.get("servicio", ""))
            cod = g.get("codigo_corto", "")
            hechos = hglosas[i].get("hechos_probados", []) if i < len(hglosas) else []
            evid = "; ".join(ref for h in hechos for ref in h.get("evidencia", []))
            cola = valores.get((fc, cod, servicio))
            valor = cola.popleft() if cola else 0
            filas.append(
                {
                    "factura": factura,
                    "paciente": paciente,
                    "contrato": contrato,
                    "saldo": saldo,
                    "edad": edad,
                    "codigo": cod,
                    "servicio": servicio[:60],
                    "valor": valor,
                    "defendibilidad": d.get("defendibilidad", 0),
                    "decision": d.get("decision", ""),
                    "prioridad": d.get("prioridad", ""),
                    "evidencia": evid,
                    "faltantes": "; ".join(d.get("documentos_faltantes", [])),
                    "riesgo": d.get("riesgos", {}).get("probatorio", ""),
                    "alertas": alertas,
                    "motivo": d.get("motivo_decision", ""),
                }
            )
    return filas


def escribir_excel(filas: list[dict], salida: Path) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conciliacion"
    ws.append(
        [
            "Factura",
            "Paciente",
            "Contrato",
            "Saldo cartera",
            "Edad (d)",
            "Codigo glosa",
            "Servicio",
            "Valor objetado",
            "Defendibilidad %",
            "Decision",
            "Prioridad",
            "Evidencia (paginas)",
            "Documentos faltantes",
            "Riesgo probatorio",
            "Alertas del expediente",
            "Motivo / accion",
        ]
    )
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E5F")
        c.alignment = Alignment(vertical="center", wrap_text=True)
    for f in filas:
        ws.append(
            [
                f["factura"],
                f["paciente"],
                f["contrato"],
                f["saldo"],
                f["edad"],
                f["codigo"],
                f["servicio"],
                f["valor"],
                f["defendibilidad"],
                f["decision"],
                f["prioridad"],
                f["evidencia"],
                f["faltantes"],
                f["riesgo"],
                f["alertas"],
                f["motivo"],
            ]
        )
        fila = ws.max_row
        ws.cell(row=fila, column=4).number_format = "#,##0"
        ws.cell(row=fila, column=8).number_format = "#,##0"
        cc = ws.cell(row=fila, column=10)
        if f["decision"] in COLOR:
            cc.fill = PatternFill("solid", fgColor=COLOR[f["decision"]])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:P{ws.max_row}"
    anchos = (15, 26, 9, 14, 8, 12, 48, 14, 13, 22, 9, 46, 30, 14, 40, 46)
    for w, col in zip(anchos, "ABCDEFGHIJKLMNOP"):
        ws.column_dimensions[col].width = w
    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(salida))


def main(argv: list[str] | None = None) -> int:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
    ap = argparse.ArgumentParser(
        description="Genera CONCILIACION.xlsx desde una carpeta de piloto ya corrida."
    )
    ap.add_argument("--piloto-dir", type=Path, required=True, help="Carpeta del piloto (Piloto6)")
    ap.add_argument("--excel", type=Path, required=True, help="HUS.xlsx (para el valor objetado)")
    ap.add_argument("--salida", type=Path, required=True, help="CONCILIACION.xlsx de salida")
    args = ap.parse_args(argv)

    valores = _valores_por_glosa(args.excel)
    filas = construir_filas(args.piloto_dir, valores)
    if not filas:
        logger.error("No encontre glosas en %s (¿es la carpeta del piloto?).", args.piloto_dir)
        return 1
    escribir_excel(filas, args.salida)
    logger.info("Hoja de conciliacion: %d glosas  ->  %s", len(filas), args.salida)
    return 0


if __name__ == "__main__":
    sys.exit(main())
