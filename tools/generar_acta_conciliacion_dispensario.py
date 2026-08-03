"""Genera el ACTA DE CONCILIACION de las 147 facturas del Dispensario
usando como base el formato real ACTA SINAC N 720 (mismo libro, mismos estilos).

- Parte del .xlsm original (keep_vba): conserva logos, merges, firmas, macros.
- Hoja ACTA: inserta las 444 lineas de glosa (agrupadas por factura, mayor
  valor primero), actualiza encabezado (147 facturas, periodo, totales) y la
  fila TOTAL con formulas. Anade columnas de apoyo AM..AS (estado respuesta,
  resultado en acta, soporte, diferencia vs cartera, rutas + hipervinculo FE,
  lista para conciliar).
- Hoja DASHBOARD nueva: los 12 indicadores pedidos, con formulas sobre ACTA.
- Hojas GLOSAS/TRAMITES: se rellenan con los exports crudos (RECEPCION/TRAMITE)
  filtrados a las 147 facturas. NOTAS queda solo con encabezado (sin NC).
- NO inventa: lo que no existe queda vacio o PENDIENTE.

ESTADO DE ESTA HERRAMIENTA
--------------------------
Este es el generador VIGENTE (el que produjo el acta que usa el area). Se
preserva aqui tal cual se ejecuto, para no perder el trabajo. Antes de darlo
por productivo quedan tres tareas, documentadas en
`docs/MODULO_CONCILIACION_DISPENSARIO.md` (seccion 16.3):

1. Parametrizar con argparse las rutas de entrada (hoy fijas en las constantes
   U/TPL/HUS/TRA/REC/CAR/EST y la salida OUT), que apuntan a la carpeta de la
   sesion en que se corrio.
2. Cambiar el `assert` fijo de 147 facturas / 444 glosas por un parametro
   opcional de validacion, para poder correr otros lotes.
3. Reusar `num`/`txt`/`normalizar_factura` de `hoja_maestra_conciliacion.py` en
   lugar de las copias locales (`num`/`txt`/`norm`).

USO ACTUAL
----------
    py tools\\generar_acta_conciliacion_dispensario.py "D:\\ruta\\ACTA.xlsm"

Las cifras de control que debe reproducir estan en el documento del modulo:
147 facturas / 444 glosas, facturado $1.267.976.805, glosado $317.640.524,
aceptado en tramite $1.758.956, 471 formulas y 0 errores tras recalcular.
"""

import re
import sys
from collections import defaultdict
from copy import copy
from datetime import datetime
from pathlib import Path

# Lector de pesos compartido (ver `tools/_dinero.py`): el que vivía aquí
# dividía por mil los valores con un solo punto de miles.
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _dinero import a_numero as a_numero_compartido  # noqa: E402

import openpyxl
from openpyxl.utils import get_column_letter as L

U = Path("/root/.claude/uploads/116eb458-ad18-5e2c-a877-732ea7f1ccbe")
TPL = U / "285e701e-ACTA_SINAC_N_720__ESE_HUS__DISPENSARIO_MEDICO.xlsm"
HUS = U / "ed760996-HUS.xlsx"
TRA = U / "1dd26b78-TRAMITE_DE_OBJECCION_DISPENSARIO_MEDICO.xlsx"
REC = U / "b7d039e7-RECEPCION_DE_OBJECIONES_DISPENSARIO_MEDICO.xlsx"
CAR = U / "2ba94eec-DISPENSARIO_MEDICO_CORTE_30062026_CRUCE_ACTAS.xlsx"
OUT = (
    Path(sys.argv[1])
    if len(sys.argv) > 1
    else Path(
        "/tmp/claude-0/-home-user-motor-glosas-hus/116eb458-ad18-5e2c-a877-732ea7f1ccbe/"
        "scratchpad/ACTA_CONCILIACION_147_DISPENSARIO.xlsm"
    )
)

FE_RAIZ = r"\\172.16.32.83\factura_electronica_net22"

FAMILIAS = {
    "TA": ("TARIFAS", "ADMINISTRATIVA"),
    "SO": ("SOPORTES", "ADMINISTRATIVA"),
    "AU": ("AUTORIZACION", "ADMINISTRATIVA"),
    "FA": ("FACTURACION", "ADMINISTRATIVA"),
    "CO": ("COBERTURA", "ADMINISTRATIVA"),
    "CL": ("CALIDAD / PERTINENCIA", "MEDICA"),
    "PE": ("PERTINENCIA", "MEDICA"),
}


def norm(pref, fac):
    m = re.search(r"(\d+)", str(fac) if fac is not None else "")
    if not m:
        return ""
    p = str(pref).strip() if pref else "HUS"
    return p + m.group(1).zfill(10)


def num(x):
    """Valor monetario → float, con el lector único del repo.

    Tenía el mismo agujero que el informe de cartera: un **solo** punto de
    miles no se manejaba. La condición era `s.count(".") > 1`, así que
    `"3.211.891"` salía bien pero `"$ 950.000"` se leía como **950,00** y
    `"2.500"` como **2,50** — dividido por mil.

    Es el bug que el auditor ya había visto en un informe. Ahora este archivo
    usa el lector único de `tools/_dinero.py`.
    """
    return a_numero_compartido(x)


def txt(x):
    if x is None:
        return ""
    s = str(x).strip()
    return "" if s in ("None", "-") else s


def cod_corto(c):
    m = re.match(r"([A-Z]{2}\s?\d{2}\s?\d{2}|[A-Z]{2}\d{2}|[A-Z]{2}\s?\d+)", txt(c).upper())
    return m.group(1).replace(" ", "") if m else txt(c).split(" ")[0]


def familia_de(cod):
    pref = cod_corto(cod)[:2]
    return FAMILIAS.get(pref, ("OTRA", "ADMINISTRATIVA"))


# ---------------------------------------------------------------- fuentes


def main() -> None:
    """El guion completo: lee las fuentes, arma el acta y la guarda.

    Estaba suelto a nivel de módulo, así que **importar el archivo lo
    ejecutaba** — y por eso no se podía probar ninguna de sus funciones sin
    tener los Excel de producción a mano. Metido en `main()`, el módulo se
    importa sin efectos y el guion sigue corriendo igual desde la consola.
    """
    print("Cargando fuentes...")
    wb = openpyxl.load_workbook(HUS, read_only=True, data_only=True)
    hus_rows = list(wb["Hoja1"].iter_rows(values_only=True))[1:]
    wb.close()

    glosas = []  # una por linea del lote
    for r in hus_rows:
        f = norm(r[8], r[9])
        if not f:
            continue
        glosas.append(
            {
                "factura": f,
                "radicado": txt(r[7]),
                "fecha_atencion": r[10],
                "fecha_radicacion": r[11],
                "paciente": txt(r[15]),
                "identificacion": txt(r[14]),
                "valor_factura": num(r[16]),
                "num_glosa": txt(r[17]),
                "cod_full": txt(r[18]),
                "motivo": txt(r[19]),
                "valor_objetado": num(r[20]),
                "servicio": txt(r[21]),
                "cod_respuesta": txt(r[22]),
                "acepta_ips": num(r[23]),
                "levanta_ips": num(r[24]),
                "ratifica_ips": num(r[25]),
                "fecha_acta_rta": r[26],
                "acta_rta": txt(r[27]),
                "soporte_esm": txt(r[30]).upper() or "PENDIENTE",
                "rta_auditor": txt(r[31]),
                "rta_ips": txt(r[32]),
            }
        )
    facs = sorted({g["factura"] for g in glosas})
    assert len(facs) == 147 and len(glosas) == 444, (len(facs), len(glosas))

    # fecha factura real + fecha objecion (por factura) desde TRAMITE / RECEPCION
    wb = openpyxl.load_workbook(TRA, read_only=True, data_only=True)
    tra_rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))[1:]
    wb.close()
    fecha_fac, fecha_obj = {}, {}
    for r in tra_rows:
        f = norm("HUS", r[7])
        if f in (fecha_fac.keys() | set()) or f not in set(facs):
            pass
        if f in set(facs):
            if f not in fecha_fac and isinstance(r[15], datetime):
                fecha_fac[f] = r[15]
            if f not in fecha_obj and isinstance(r[3], datetime):
                fecha_obj[f] = r[3]
    wb = openpyxl.load_workbook(REC, read_only=True, data_only=True)
    rec_rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))[1:]
    wb.close()
    for r in rec_rows:
        f = norm("HUS", r[11])
        if f in set(facs):
            if f not in fecha_fac and isinstance(r[17], datetime):
                fecha_fac[f] = r[17]
            if f not in fecha_obj and isinstance(r[20], datetime):
                fecha_obj[f] = r[20]

    # estado de cartera JUN 2026 (VLR ACEPTADO por factura + abogado asignado)
    EST = U / "e8113d00-Copia_de_Estado_Cartera_JUN_2026_Con_Cruces.xlsx"
    wb = openpyxl.load_workbook(EST, read_only=True, data_only=True)
    wse = wb["ARCH CON CRUCES JUN"]
    acep_cartera, abogado = {}, {}
    hdr_seen = False
    for r in wse.iter_rows(values_only=True):
        if not hdr_seen:
            if r and txt(r[0]) == "NIT":
                hdr_seen = True
            continue
        f = norm("HUS", r[6]) if len(r) > 6 else ""
        if not f:
            continue
        v = num(r[21]) if len(r) > 21 else 0.0
        acep_cartera[f] = max(acep_cartera.get(f, 0.0), v)
        ab = txt(r[25]) if len(r) > 25 else ""
        if ab and ab.upper() != "#N/A":
            abogado[f] = ab
    wb.close()

    # centro de costo por linea (RECEPCION col32), indexado por factura -> valor -> {cc}
    cc_por_fac = defaultdict(lambda: defaultdict(set))
    cc_todos = defaultdict(set)
    for r in rec_rows:
        f = norm("HUS", r[11])
        if f and len(r) > 32 and r[32]:
            cc = txt(r[32])
            cc_por_fac[f][round(num(r[31]))].add(cc)
            cc_todos[f].add(cc)

    def centro_costo(f, valor):
        """Centro de costo real de la linea: match por (factura, valor); si es
        ambiguo se listan los centros reales de la factura (nunca se inventa)."""
        por_valor = cc_por_fac.get(f, {}).get(round(valor), set())
        if len(por_valor) == 1:
            return next(iter(por_valor))
        todos = cc_todos.get(f, set())
        if len(todos) == 1:
            return next(iter(todos))
        if por_valor:
            return " / ".join(sorted(por_valor))
        if todos:
            return " / ".join(sorted(todos))
        return ""

    # cartera (saldo + valor glosa) y resultado en actas (CRUCE ACTAS)
    wb = openpyxl.load_workbook(CAR, read_only=True, data_only=True)
    car_rows = list(wb["CARTERA"].iter_rows(values_only=True))
    cru_rows = list(wb["CRUCE ACTAS"].iter_rows(values_only=True))
    wb.close()
    cartera = {}
    for r in car_rows[2:]:
        f = norm("HUS", r[2])
        if f:
            cartera[f] = {"saldo": num(r[5]), "vglosa": num(r[10]), "estado": txt(r[24])}
    acta_res = defaultdict(list)
    for r in cru_rows[2:]:
        f = norm("HUS", r[0])
        acta = txt(r[1])
        if f and acta and f in set(facs):
            partes = []
            if num(r[4]):
                partes.append(f"ACEPTADA ${num(r[4]):,.0f}")
            if num(r[5]):
                partes.append(f"LEVANTADA ${num(r[5]):,.0f}")
            if num(r[6]):
                partes.append(f"RATIFICADA ${num(r[6]):,.0f}")
            acta_res[f].append(f"{acta}: " + (" / ".join(partes) or "SIN VALORES"))

    # ---------------------------------------------------------------- orden
    suma_fac = defaultdict(float)
    for g in glosas:
        suma_fac[g["factura"]] += g["valor_objetado"]
    orden_facs = sorted(facs, key=lambda f: -suma_fac[f])
    glosas.sort(key=lambda g: (orden_facs.index(g["factura"]), -g["valor_objetado"]))
    N = len(glosas)  # 444

    # ---------------------------------------------------------------- plantilla
    print("Abriendo plantilla (keep_vba)...")
    wb = openpyxl.load_workbook(TPL, keep_vba=True)
    ws = wb["ACTA"]

    FIRST, TPL_LAST = 12, 22  # filas de datos de la plantilla
    extra = N - (TPL_LAST - FIRST + 1)  # 433 filas nuevas
    INS_AT = FIRST + 1  # insertar despues de la primera fila de datos

    # 1) guardar merges y alturas por debajo del punto de insercion
    merges_abajo = [str(m) for m in ws.merged_cells.ranges if m.min_row >= INS_AT]
    for m in merges_abajo:
        ws.unmerge_cells(m)
    alturas_abajo = {r: d.height for r, d in ws.row_dimensions.items() if r >= INS_AT and d.height}

    # 2) insertar filas
    ws.insert_rows(INS_AT, extra)

    # 3) restaurar merges y alturas desplazadas
    for m in merges_abajo:
        mm = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", m)
        c1, r1, c2, r2 = mm.group(1), int(mm.group(2)), mm.group(3), int(mm.group(4))
        ws.merge_cells(f"{c1}{r1 + extra}:{c2}{r2 + extra}")
    for r, h in list(ws.row_dimensions.items()):
        pass  # limpiamos abajo
    for r in sorted(alturas_abajo, reverse=True):
        ws.row_dimensions[r + extra].height = alturas_abajo[r]

    LAST = FIRST + N - 1  # 455
    TOT = LAST + 1  # 456 (fila TOTAL original 23 + 433)

    # 4) copiar estilo de la fila 12 a todas las filas de datos (C..AL => 3..38)
    STYLE_COLS = list(range(3, 39))
    base = {c: ws.cell(row=FIRST, column=c) for c in STYLE_COLS}
    for r in range(FIRST + 1, LAST + 1):
        ws.row_dimensions[r].height = 21.75
        for c in STYLE_COLS:
            cel = ws.cell(row=r, column=c)
            cel._style = copy(base[c]._style)

    # 5) encabezados columnas de apoyo AM..AS (39..45), estilo del encabezado I11
    EXTRA_HDRS = [
        "ESTADO RESPUESTA",
        "RESULTADO EN ACTAS PREVIAS",
        "SOPORTE ENTREGA ESM",
        "DIFERENCIA VS CARTERA (por factura)",
        "RUTA SOPORTES (PENDIENTE raiz Y:/X:)",
        "RUTA FACTURA ELECTRONICA",
        "LISTA PARA CONCILIAR",
    ]
    hdr_style = ws.cell(row=11, column=9)  # I11
    for i, h in enumerate(EXTRA_HDRS):
        c = 39 + i
        cel = ws.cell(row=11, column=c, value=h)
        cel._style = copy(hdr_style._style)
        ws.column_dimensions[L(c)].width = 26
    ws.column_dimensions["AQ"].width = 40
    ws.column_dimensions["AR"].width = 40

    # 6) volcar las 444 filas
    print("Escribiendo 444 filas...")
    # (aquí vivía `data_style_ref`, un dict que se armaba y nadie leía.
    #  Estaba a nivel de módulo, donde ruff no lo veía; al meter el guion
    #  en `main()` quedó a la vista y se borra.)
    prev_fac = None
    for i, g in enumerate(glosas):
        r = FIRST + i
        f = g["factura"]
        fam, tipo = familia_de(g["cod_full"])
        first_of_fac = f != prev_fac
        prev_fac = f
        ff = fecha_fac.get(f)
        aaaamm = ""
        for cand in (ff, g["fecha_radicacion"], g["fecha_atencion"]):
            if isinstance(cand, datetime):
                aaaamm = f"{cand.year:04d}{cand.month:02d}"
                break
        vals = {
            3: i + 1,
            4: g["acta_rta"] or g["radicado"],
            5: f,
            6: ff if ff else g["fecha_atencion"],
            7: tipo,
            8: fam,
            9: cod_corto(g["cod_full"]),
            10: g["motivo"] or g["cod_full"],
            11: g["valor_factura"],
            12: g["valor_objetado"],
            13: g["ratifica_ips"],
            14: g["acepta_ips"],  # VALOR ACEPTA IPS (mesa)
            15: 0,  # VALOR LEVANTA ENTIDAD (mesa)
            16: 0,  # VALOR RATIFICADO (mesa)
            17: "",  # DESCRIPCION DE CONCILIACION (mesa)
            18: "",
            19: abogado.get(f, ""),  # ABOGADO ASIGNADO (estado de cartera JUN)
            # VALOR ACEPTADO EN TRAMITE: en el lote todas las lineas van en 0
            # (RE9901); el estado de cartera registra aceptado a nivel FACTURA en 8
            # casos -> se pone en la PRIMERA fila de la factura para que la SUMA
            # de la columna cuadre sin duplicar.
            20: (acep_cartera.get(f, 0.0) if first_of_fac else 0) or g["acepta_ips"],
            21: centro_costo(f, g["valor_objetado"]),  # CENTRO DE COSTO (recepcion)
            22: "PENDIENTE",  # CUENTA CONTABLE: no viene en SIMED/DGH/cartera
            23: 0,  # VALOR REFACTURAR
            26: f"=L{r}-N{r}",  # Z DIFERENCIA
            27: g["fecha_radicacion"],  # AA
            28: fecha_obj.get(f, ""),  # AB FECHA OBJECION
            29: g["rta_ips"],  # AC RESPUESTA A GLOSA INICIAL EN TRAMITE
        }
        for c, v in vals.items():
            ws.cell(row=r, column=c, value=v)
        # columnas de apoyo
        estado = "RESPONDIDA" if (g["rta_ips"] or g["cod_respuesta"]) else "SIN RESPUESTA"
        car = cartera.get(f)
        dif = ""
        if first_of_fac:
            if car:
                d = car["vglosa"] - suma_fac[f]
                dif = round(d) if abs(d) > 1 else 0
            else:
                dif = "NO ESTA EN CARTERA"
        ruta_sop = f"PENDIENTE raiz -> ...\\{aaaamm}\\{f}" if aaaamm else "PENDIENTE"
        ruta_fe = f"{FE_RAIZ}\\{aaaamm}" if aaaamm else "PENDIENTE"
        lista = ""
        if first_of_fac:
            # Lista = respondida + sin diferencia vs cartera + existe en cartera.
            # (El soporte de entrega ESM va aparte: el lote trae 'NO' en todas —
            # la entidad no ha confirmado recibo; es un punto PARA la mesa, no un
            # bloqueo nuestro: las 444 tienen radicado de entrega de la respuesta.)
            ok = estado == "RESPONDIDA" and dif == 0 and car is not None
            lista = "SI" if ok else "NO"
        extras = {
            39: estado,
            40: "; ".join(acta_res.get(f, [])) if first_of_fac else "",
            41: g["soporte_esm"],
            42: dif,
            43: ruta_sop,
            44: ruta_fe,
            45: lista,
        }
        for c, v in extras.items():
            cel = ws.cell(row=r, column=c, value=v)
            cel._style = copy(base[9]._style)  # estilo celda de datos (I12)
            if c == 42 and isinstance(v, (int, float)):
                cel.number_format = "#,##0"
            if c == 44 and aaaamm:
                cel.hyperlink = ruta_fe
        # numero formato moneda por si el estilo base no lo trae
        for c in (11, 12, 13, 14, 15, 16, 20, 23, 26):
            ws.cell(row=r, column=c).number_format = "#,##0"

    # 7) fila TOTAL (ahora en TOT): reescribir formulas al rango nuevo
    ws.cell(row=TOT, column=9, value="TOTAL ")
    ws.cell(
        row=TOT,
        column=11,
        value=f"=SUMPRODUCT((E{FIRST}:E{LAST}<>E{FIRST - 1}:E{LAST - 1})*K{FIRST}:K{LAST})",
    )
    for c in (12, 13, 14, 15, 16, 20, 23, 26):
        col = L(c)
        ws.cell(row=TOT, column=c, value=f"=SUM({col}{FIRST}:{col}{LAST})")
    ws.cell(row=TOT, column=21, value="-")

    # 8) encabezado del acta
    ws["F9"] = "2025 - 2026"
    ws["K9"] = f"=SUMPRODUCT((E{FIRST}:E{LAST}<>E{FIRST - 1}:E{LAST - 1})*1)"
    ws["M9"] = f"=M{TOT}"
    ws["P9"] = f"=N{TOT}"
    ws["R9"] = f"=O{TOT}"
    ws["T9"] = f"=P{TOT}"
    ws["W9"] = f"=W{TOT}"
    ws["R7"] = "POR DEFINIR"
    ws["V7"] = " N.º POR ASIGNAR"

    # ---------------------------------------------------------------- DASHBOARD
    print("Dashboard...")
    if "DASHBOARD" in wb.sheetnames:
        del wb["DASHBOARD"]
    dash = wb.create_sheet("DASHBOARD", 0)
    from openpyxl.styles import Alignment, Font, PatternFill

    HF = PatternFill("solid", fgColor="1F4E78")
    WF = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    TIT = Font(bold=True, size=14, color="1F4E78", name="Arial")
    SUB = Font(italic=True, size=9, color="595959", name="Arial")
    AR = Font(name="Arial", size=10)
    dash["B2"] = "DASHBOARD - CONCILIACION DISPENSARIO MEDICO (147 FACTURAS)"
    dash["B2"].font = TIT
    dash["B3"] = (
        "Universo: lote de conciliacion (147 facturas / 444 glosas). "
        "Corte cartera 30/06/2026. HUS0000443525 no cruza con cartera (marcada)."
    )
    dash["B3"].font = SUB
    kpis = [
        (
            "Numero total de facturas pendientes por conciliar",
            f"=SUMPRODUCT((ACTA!E{FIRST}:E{LAST}<>ACTA!E{FIRST - 1}:E{LAST - 1})*1)",
            "0",
        ),
        ("Valor total pendiente por conciliar", f"=SUM(ACTA!M{FIRST}:M{LAST})", "#,##0"),
        ("Valor total glosado (glosa inicial)", f"=SUM(ACTA!L{FIRST}:L{LAST})", "#,##0"),
        (
            "Valor recuperable (defendido por el HUS = glosado - aceptado)",
            f"=SUM(ACTA!L{FIRST}:L{LAST})-SUM(ACTA!T{FIRST}:T{LAST})",
            "#,##0",
        ),
        ("Valor ratificado por el HUS en respuesta", f"=SUM(ACTA!M{FIRST}:M{LAST})", "#,##0"),
        (
            "Valor levantado (resultado mesa - se llena en conciliacion)",
            f"=SUM(ACTA!O{FIRST}:O{LAST})",
            "#,##0",
        ),
        (
            "Valor sin respuesta",
            f'=SUMIF(ACTA!AM{FIRST}:AM{LAST},"SIN RESPUESTA",ACTA!L{FIRST}:L{LAST})',
            "#,##0",
        ),
        (
            "Valor conciliado (acepta IPS + levanta entidad - se llena en mesa)",
            f"=SUM(ACTA!N{FIRST}:N{LAST})+SUM(ACTA!O{FIRST}:O{LAST})",
            "#,##0",
        ),
        (
            "Facturas sin soporte de entrega ESM",
            f'=SUMPRODUCT((ACTA!E{FIRST}:E{LAST}<>ACTA!E{FIRST - 1}:E{LAST - 1})*(ACTA!AO{FIRST}:AO{LAST}="NO"))',
            "0",
        ),
        (
            "Facturas con glosas sin respuesta",
            f'=SUMPRODUCT((ACTA!E{FIRST}:E{LAST}<>ACTA!E{FIRST - 1}:E{LAST - 1})*(ACTA!AM{FIRST}:AM{LAST}="SIN RESPUESTA"))',
            "0",
        ),
        (
            "Facturas con diferencias vs cartera",
            f'=SUMPRODUCT((ACTA!AP{FIRST}:AP{LAST}<>"")*(ACTA!AP{FIRST}:AP{LAST}<>0)*1)',
            "0",
        ),
        (
            "Facturas listas para conciliar (respondida + sin dif. + en cartera)",
            f'=COUNTIF(ACTA!AS{FIRST}:AS{LAST},"SI")',
            "0",
        ),
    ]
    dash["B5"] = "INDICADOR"
    dash["E5"] = "VALOR"
    for cc in ("B5", "C5", "D5", "E5"):
        dash[cc].fill = HF
        dash[cc].font = WF
    dash.merge_cells("B5:D5")
    r = 6
    for nombre, formula, fmt in kpis:
        dash.merge_cells(f"B{r}:D{r}")
        dash[f"B{r}"] = nombre
        dash[f"B{r}"].font = AR
        c = dash[f"E{r}"]
        c.value = formula
        c.number_format = fmt
        c.font = Font(name="Arial", size=11, bold=True)
        c.alignment = Alignment(horizontal="right")
        r += 1
    dash.column_dimensions["B"].width = 46
    dash.column_dimensions["C"].width = 16
    dash.column_dimensions["D"].width = 16
    dash.column_dimensions["E"].width = 20
    dash["B19"] = "Nota: N (acepta IPS), O (levanta entidad), P (ratificado) y Q "
    dash["B20"] = "(descripcion) de la hoja ACTA se diligencian EN la mesa de conciliacion."
    dash["B19"].font = SUB
    dash["B20"].font = SUB
    dash.sheet_view.showGridLines = False

    # ---------------------------------------------------------------- GLOSAS / TRAMITES
    print("Hojas crudas filtradas...")
    fset = set(facs)

    def rellenar_cruda(nombre, filas, col_factura, header_keep):
        wsx = wb[nombre]
        # borrar datos viejos (desde header_keep+1)
        wsx.delete_rows(header_keep + 1, wsx.max_row)
        n = 0
        for row in filas:
            f = norm("HUS", row[col_factura])
            if f in fset:
                wsx.append(list(row))
                n += 1
        return n

    n_rec = rellenar_cruda("GLOSAS", rec_rows, 11, 3)
    n_tra = rellenar_cruda("TRAMITES", tra_rows, 7, 3)
    wb["NOTAS"].delete_rows(3, wb["NOTAS"].max_row)

    # NOTA: las hojas GLOSAS/TRAMITES quedan con las columnas del export original
    # (RECEPCION 34 cols / TRAMITE 33 cols). Reemplazo los encabezados de la
    # plantilla por los del export para que columna y dato coincidan SIEMPRE.
    wbh = openpyxl.load_workbook(REC, read_only=True)
    rec_hdr = next(wbh[wbh.sheetnames[0]].iter_rows(values_only=True))
    wbh.close()
    wbh = openpyxl.load_workbook(TRA, read_only=True)
    tra_hdr = next(wbh[wbh.sheetnames[0]].iter_rows(values_only=True))
    wbh.close()
    for nombre, hdr in (("GLOSAS", rec_hdr), ("TRAMITES", tra_hdr)):
        wsx = wb[nombre]
        wsx.delete_rows(1, 3)
        wsx.insert_rows(1)
        for i, h in enumerate(hdr, 1):
            wsx.cell(row=1, column=i, value=h)

    # --------------------------------------------------------------- CF acotado
    # La plantilla trae formato condicional sobre columnas ENTERAS (Z1:Z1048576,
    # NOTAS A71:A1048576): eso cuelga el recalculo. Se acota al rango real.
    from openpyxl.formatting.formatting import ConditionalFormattingList

    viejos = [(str(cf.sqref), cf.rules) for cf in ws.conditional_formatting]
    ws.conditional_formatting = ConditionalFormattingList()
    for sqref, rules in viejos:
        if "Z1:Z1048576" in sqref:
            nuevo = f"Z{FIRST}:Z{TOT}"
        elif sqref.startswith("W12"):
            nuevo = f"W{FIRST}:W{LAST}"
        else:
            nuevo = sqref
        for rule in rules:
            ws.conditional_formatting.add(nuevo, rule)
    for sn in ("TRAMITES", "NOTAS", "GLOSAS"):
        wb[sn].conditional_formatting = ConditionalFormattingList()

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"OK -> {OUT}")
    print(f"  ACTA: filas {FIRST}..{LAST} (444 glosas), TOTAL en fila {TOT}")
    print(f"  GLOSAS (recepcion filtrada): {n_rec} filas | TRAMITES: {n_tra} filas")


if __name__ == "__main__":
    main()
