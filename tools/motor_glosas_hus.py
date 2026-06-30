"""motor_glosas_hus.py — Aprende del histórico TRAMITE y propone respuestas para RECEPCION.

Lee un Excel TRAMITE (glosas ya respondidas por el HUS) y aprende, por cada
(EPS, código de objeción), cuál es el código de respuesta y el texto que el
HUS usa típicamente. Después aplica ese aprendizaje a una RECEPCION (glosas
pendientes) y genera un XLSX enriquecido con columnas nuevas:

    COD_RESPUESTA_SUGERIDO   — RE9901 / RE9502 / RE9601 / ...
    VALOR_ACEPTADO_SUGERIDO  — casi siempre 0 (el HUS rechaza el 99% en histórico)
    TEXTO_RESPUESTA_SUGERIDO — plantilla del HUS adaptada al servicio
    CONFIANZA                — ALTA / MEDIA / BAJA según especificidad del match
    FUENTE_PATRON            — de qué (EPS, código) salió la plantilla

El humano revisa el XLSX y corrige antes de pasar a radicación.

USO:
    py motor_glosas_hus.py ^
        --tramite   "D:\\...\\TRAMITE_DE_GLOSA.xlsx" ^
        --recepcion "D:\\...\\RECEPCION_DE_GLOSA.xlsx" ^
        --salida    "D:\\...\\RECEPCION_CON_RESPUESTAS.xlsx"

    # Filtrar la recepción por una EPS (substring case-insensitive):
    py motor_glosas_hus.py --tramite ... --recepcion ... --salida ... --eps DISPENSARIO

    # Acotar a una factura puntual:
    py motor_glosas_hus.py --tramite ... --recepcion ... --salida ... --factura HUS0000511438

DEPENDENCIAS:
    py -m pip install openpyxl
"""

from __future__ import annotations

import argparse
import logging
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

logger = logging.getLogger("motor_glosas")

# ─── Columnas esperadas en los dos archivos ─────────────────────────────────
# (los nombres son los que exporta DG.NET con todo el prefijo del modelo)

COLS_TRAMITE = {
    "tercero": "RecepcionObjecion.FacturaCartera.Tercero.NombreCompletoAN",
    "factura": "RecepcionObjecion.FacturaCartera.Factura",
    "cod_obj": "ListadoConceptos.DetalleRecepcionObjecion.ConceptoObjecion.Codigo",
    "cod_res": "ListadoConceptos.ConceptoObjecion.Codigo",
    "val_obj": "ListadoConceptos.DetalleRecepcionObjecion.ValorObjecion",
    "val_ace": "ListadoConceptos.ValorAceptado",
    "servicio": "ListadoConceptos.DetalleRecepcionObjecion.ServicioProductoFactura.Descripcion",
    "observ": "ListadoConceptos.Observaciones",
}
# Formato del export "Listado de Recepción de Objeción" (headers sin prefijo).
COLS_RECEPCION = {
    "tercero": "FacturaCartera.Tercero.NombreCompletoAN",
    "factura": "FacturaCartera.Factura",
    "cod_obj": "ListadoConceptos.ConceptoObjecion.Codigo",
    "val_obj": "ListadoConceptos.ValorObjecion",
    "servicio": "ListadoConceptos.ServicioProductoFactura.Descripcion",
    "observ_eps": "ListadoConceptos.Observaciones",
}
# Formato del export "Listado de Trámite de Objeción" usado como recepción
# (mismo archivo TRAMITE → toma las glosas como si fueran pendientes).
COLS_RECEPCION_DESDE_TRAMITE = {
    "tercero": "RecepcionObjecion.FacturaCartera.Tercero.NombreCompletoAN",
    "factura": "RecepcionObjecion.FacturaCartera.Factura",
    "cod_obj": "ListadoConceptos.DetalleRecepcionObjecion.ConceptoObjecion.Codigo",
    "val_obj": "ListadoConceptos.DetalleRecepcionObjecion.ValorObjecion",
    "servicio": "ListadoConceptos.DetalleRecepcionObjecion.ServicioProductoFactura.Descripcion",
    "observ_eps": "ListadoConceptos.DetalleRecepcionObjecion.ConceptoObjecion.Nombre",
}


def _norm(s: str) -> str:
    s = (s or "").strip().upper()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    return " ".join(s.split())


def _idx(headers: list, mapping: dict) -> dict:
    """Para cada clave de mapping devuelve el índice de la columna en headers
    (busca match exacto). Lanza error si falta alguna."""
    h2i = {h: i for i, h in enumerate(headers)}
    out = {}
    for k, col in mapping.items():
        if col not in h2i:
            raise ValueError(
                f"Falta la columna {col!r} en el Excel. "
                f"Encabezados disponibles (parcial): {headers[:6]} …"
            )
        out[k] = h2i[col]
    return out


# ─── Aprendizaje ────────────────────────────────────────────────────────────


class PatronesGlosas:
    """Indexa el histórico TRAMITE por (eps_norm, cod_objecion) y por
    cod_objecion solo, guardando: códigos de respuesta vistos y los textos
    de Observaciones que el HUS usó."""

    def __init__(self):
        # Por (eps, cod_obj) → Counter de cod_res
        self.cod_por_eps_codobj: dict = defaultdict(Counter)
        # Por (eps, cod_obj, cod_res) → lista de textos (en orden de aparición)
        self.textos_por_eps_codobj_codres: dict = defaultdict(list)
        # Aggregate por cod_obj solo (todas las EPS)
        self.cod_por_codobj: dict = defaultdict(Counter)
        self.textos_por_codobj_codres: dict = defaultdict(list)
        # ¿Cuánto valor acepta el HUS en cada caso? (acept_total / acept_parcial / rechazo)
        self.val_por_eps_codobj_codres: dict = defaultdict(list)

    def aprender(self, fila: dict) -> None:
        eps = fila["eps"]
        co = fila["cod_obj"]
        cr = fila["cod_res"]
        if not co or not cr:
            return
        texto = (fila.get("observ") or "").strip()
        v_obj = fila.get("val_obj") or 0
        v_ace = fila.get("val_ace") or 0
        # Categoría de aceptación
        if v_ace == 0:
            cat = "RECHAZA"
        elif v_obj and v_ace == v_obj:
            cat = "ACEPTA_TOTAL"
        else:
            cat = "ACEPTA_PARCIAL"

        self.cod_por_eps_codobj[(eps, co)][cr] += 1
        self.cod_por_codobj[co][cr] += 1
        if texto:
            self.textos_por_eps_codobj_codres[(eps, co, cr)].append(texto)
            self.textos_por_codobj_codres[(co, cr)].append(texto)
        self.val_por_eps_codobj_codres[(eps, co, cr)].append((cat, v_obj, v_ace))

    def proponer(self, eps: str, cod_obj: str) -> dict:
        """Devuelve dict con cod_res sugerido + texto + confianza + fuente."""
        # 1) Match EXACTO: misma EPS, mismo cod_obj.
        contador = self.cod_por_eps_codobj.get((eps, cod_obj))
        if contador:
            cr, n = contador.most_common(1)[0]
            tot = sum(contador.values())
            textos = self.textos_por_eps_codobj_codres.get((eps, cod_obj, cr), [])
            return {
                "cod_res": cr,
                "texto": textos[0] if textos else "",
                "confianza": "ALTA" if n / tot >= 0.7 else "MEDIA",
                "fuente": f"EPS exacta ({n}/{tot} en historial)",
                "valor_aceptado": self._valor_aceptado_tipico(eps, cod_obj, cr),
            }
        # 2) Match por COD_OBJ (cualquier EPS).
        contador = self.cod_por_codobj.get(cod_obj)
        if contador:
            cr, n = contador.most_common(1)[0]
            tot = sum(contador.values())
            textos = self.textos_por_codobj_codres.get((cod_obj, cr), [])
            return {
                "cod_res": cr,
                "texto": textos[0] if textos else "",
                "confianza": "MEDIA" if n / tot >= 0.7 else "BAJA",
                "fuente": f"cualquier EPS ({n}/{tot} en historial)",
                "valor_aceptado": 0,  # por defecto rechazo
            }
        # 3) Fallback por defecto: rechazo con RE9901.
        return {
            "cod_res": "RE9901",
            "texto": "ESE HUS NO ACEPTA GLOSA. [REVISAR — sin histórico para este código de objeción]",
            "confianza": "BAJA",
            "fuente": "default (sin histórico)",
            "valor_aceptado": 0,
        }

    def _valor_aceptado_tipico(self, eps, cod_obj, cod_res) -> int:
        casos = self.val_por_eps_codobj_codres.get((eps, cod_obj, cod_res), [])
        cats = Counter(c[0] for c in casos)
        if not cats:
            return 0
        dom = cats.most_common(1)[0][0]
        return 0 if dom == "RECHAZA" else -1  # -1 sentinela: parcial (humano decide)


# ─── Lectura/escritura Excel ────────────────────────────────────────────────


def cargar_tramite(ruta: Path) -> PatronesGlosas:
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(ruta), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"{ruta} está vacío")
    headers = list(rows[0])
    idx = _idx(headers, COLS_TRAMITE)
    pat = PatronesGlosas()
    for r in rows[1:]:
        pat.aprender(
            {
                "eps": _norm(str(r[idx["tercero"]] or "")),
                "factura": str(r[idx["factura"]] or "").strip(),
                "cod_obj": (str(r[idx["cod_obj"]] or "").strip() or None),
                "cod_res": (str(r[idx["cod_res"]] or "").strip() or None),
                "val_obj": r[idx["val_obj"]] or 0,
                "val_ace": r[idx["val_ace"]] or 0,
                "servicio": str(r[idx["servicio"]] or "").strip(),
                "observ": str(r[idx["observ"]] or "").strip(),
            }
        )
    logger.info(
        f"Patrones aprendidos del histórico: "
        f"{len(pat.cod_por_codobj)} códigos de objeción distintos, "
        f"{len(pat.cod_por_eps_codobj)} combinaciones (EPS, código)"
    )
    return pat


def procesar_recepcion(
    ruta_recep: Path,
    ruta_salida: Path,
    patrones: PatronesGlosas,
    filtro_eps: str | None,
    filtro_factura: str | None,
) -> tuple[int, int]:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb_in = load_workbook(filename=str(ruta_recep), data_only=True, read_only=True)
    ws_in = wb_in[wb_in.sheetnames[0]]
    rows = list(ws_in.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"{ruta_recep} está vacío")
    headers = list(rows[0])
    # Auto-detectar formato: si los headers traen prefijo RecepcionObjecion. son
    # del Listado de Trámite (que también sirve como input — el motor toma cada
    # fila como una glosa a evaluar).
    try:
        idx = _idx(headers, COLS_RECEPCION)
        logger.info("  formato detectado: RECEPCION (export del Listado de Recepción)")
    except ValueError:
        idx = _idx(headers, COLS_RECEPCION_DESDE_TRAMITE)
        logger.info("  formato detectado: TRAMITE (export del Listado de Trámite)")

    fil_eps = _norm(filtro_eps or "")
    fil_fac = (filtro_factura or "").strip().upper()

    # Salida: copia las columnas originales y agrega las 5 nuevas.
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "RECEPCION_CON_RESPUESTAS"
    NUEVAS = [
        "COD_RESPUESTA_SUGERIDO",
        "VALOR_ACEPTADO_SUGERIDO",
        "TEXTO_RESPUESTA_SUGERIDO",
        "CONFIANZA",
        "FUENTE_PATRON",
    ]
    out_headers = headers + NUEVAS
    fill = PatternFill("solid", fgColor="1F4E78")
    font_h = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(out_headers, 1):
        cell = ws_out.cell(row=1, column=c, value=h)
        cell.fill = fill
        cell.font = font_h
    # Resaltar las nuevas con un fondo distinto en cabecera y celdas.
    fill_nuevas_hdr = PatternFill("solid", fgColor="C65911")
    for c in range(len(headers) + 1, len(out_headers) + 1):
        ws_out.cell(row=1, column=c).fill = fill_nuevas_hdr

    fila_dest = 2
    total = 0
    incluidas = 0
    counter_fuentes = Counter()
    counter_codres = Counter()
    for r in rows[1:]:
        total += 1
        eps = _norm(str(r[idx["tercero"]] or ""))
        factura = str(r[idx["factura"]] or "").strip().upper()
        if fil_eps and fil_eps not in eps:
            continue
        if fil_fac and factura != fil_fac:
            continue
        cod_obj = str(r[idx["cod_obj"]] or "").strip()
        if not cod_obj:
            continue
        propuesta = patrones.proponer(eps, cod_obj)
        v_ace = propuesta["valor_aceptado"]
        if v_ace == -1:
            v_ace = "REVISAR (histórico: parcial)"

        # Copiar todas las columnas originales
        for c, val in enumerate(r, 1):
            ws_out.cell(row=fila_dest, column=c, value=val)
        # Agregar las 5 nuevas
        col_off = len(headers) + 1
        ws_out.cell(row=fila_dest, column=col_off, value=propuesta["cod_res"])
        ws_out.cell(row=fila_dest, column=col_off + 1, value=v_ace)
        cell_texto = ws_out.cell(row=fila_dest, column=col_off + 2, value=propuesta["texto"])
        cell_texto.alignment = Alignment(wrap_text=True, vertical="top")
        ws_out.cell(row=fila_dest, column=col_off + 3, value=propuesta["confianza"])
        ws_out.cell(row=fila_dest, column=col_off + 4, value=propuesta["fuente"])
        counter_fuentes[propuesta["fuente"].split(" (")[0]] += 1
        counter_codres[propuesta["cod_res"]] += 1
        incluidas += 1
        fila_dest += 1

    # Anchos cómodos para los nuevos campos
    col_off = len(headers) + 1
    for col, ancho in zip(
        range(col_off, col_off + 5),
        [16, 18, 80, 12, 30],
    ):
        ws_out.column_dimensions[ws_out.cell(row=1, column=col).column_letter].width = ancho
    ws_out.freeze_panes = "A2"
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(str(ruta_salida))

    logger.info("")
    logger.info(f"Salida: {ruta_salida}")
    logger.info(f"  Glosas en RECEPCION: {total:,}")
    logger.info(f"  Glosas procesadas (con cod_obj y filtros aplicados): {incluidas:,}")
    logger.info("  Distribución por código de respuesta sugerido:")
    for cr, n in counter_codres.most_common():
        pct = 100 * n / max(1, incluidas)
        logger.info(f"    {cr}: {n:>4} ({pct:.0f}%)")
    logger.info("  Distribución por fuente del patrón:")
    for f, n in counter_fuentes.most_common():
        pct = 100 * n / max(1, incluidas)
        logger.info(f"    {f}: {n:>4} ({pct:.0f}%)")
    return total, incluidas


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        "--tramite",
        type=Path,
        required=True,
        help="Excel TRAMITE_DE_GLOSA (histórico ya respondido).",
    )
    parser.add_argument(
        "--recepcion",
        type=Path,
        required=True,
        help="Excel RECEPCION_DE_GLOSA (glosas pendientes).",
    )
    parser.add_argument(
        "--salida", type=Path, required=True, help="XLSX de salida (RECEPCION + columnas SUGERIDO)."
    )
    parser.add_argument(
        "--eps",
        type=str,
        default=None,
        help="Filtrar glosas por nombre de EPS (substring, ej. DISPENSARIO).",
    )
    parser.add_argument(
        "--factura", type=str, default=None, help="Filtrar solo esta factura (HUS...)."
    )
    args = parser.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

    try:
        from openpyxl import load_workbook  # noqa: F401
    except ImportError:
        sys.stderr.write("ERROR: falta openpyxl. py -m pip install openpyxl\n")
        return 2

    if not args.tramite.is_file():
        logger.error(f"No existe el Excel TRAMITE: {args.tramite}")
        return 1
    if not args.recepcion.is_file():
        logger.error(f"No existe el Excel RECEPCION: {args.recepcion}")
        return 1

    logger.info(f"Aprendiendo del histórico: {args.tramite.name}")
    patrones = cargar_tramite(args.tramite)
    logger.info(f"Procesando RECEPCION: {args.recepcion.name}")
    if args.eps:
        logger.info(f"  filtro EPS: substring '{args.eps}'")
    if args.factura:
        logger.info(f"  filtro factura: {args.factura}")
    procesar_recepcion(
        args.recepcion, args.salida, patrones, filtro_eps=args.eps, filtro_factura=args.factura
    )
    logger.info(
        "\nAbrí el XLSX, revisá las 5 columnas nuevas (a la derecha) y corregí "
        "lo que sea necesario antes de pasar a radicación."
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
