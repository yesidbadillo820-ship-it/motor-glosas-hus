"""Convierte el Excel del homologador CUPS → SOAT a JSON comprimido.

Genera app/data/cups_soat_homologacion.json.gz a partir del Excel oficial
de homologación CUPS (Manual Único) → SOAT (Manual Tarifario).

Uso:
    python tools/generar_cups_soat_json.py /ruta/al/HOMOLOGADOR.xlsx

Reconoce DOS formatos, y elige solo mirando el encabezado:

**Antiguo (5 columnas).** El que se cargó en junio de 2025:
    0: COD. CUPS RES. 2775   ·  1: COD. SOAT  ·  2: Descripción SOAT
    3: COD. CUPS (repetido)  ·  4: DESCRIPCIÓN SERVICIO O PROCEDIMIENTO

**Gold Standard (44 columnas).** El que subió Yesid en agosto de 2026. Trae
lo mismo y además el artículo y el capítulo del Manual SOAT donde está cada
código, la clase (PUNTOS, SMDLV…), el valor en UVB y la cobertura. Eso es
justamente lo que le hace falta a un dictamen de tarifas: no basta con decir
«el CUPS 039101 corresponde al SOAT 39100», hay que poder decir en qué
artículo del manual está escrito. El encabezado no está en la primera fila
—arriba hay un título— así que se busca.

Salida JSON (las llaves nuevas solo aparecen si el Excel las trae):
    {
      "_meta": {...},
      "cups_a_soat": {
        "012403": [
          {"soat": "1101", "desc_soat": "...", "desc_cups": "...",
           "articulo": "ARTICULO 48: ...", "capitulo": "CAPITULO V",
           "clase": "PUNTOS", "uvb": "7.86", "cobertura": "POS"},
          ...
        ]
      }
    }

Un CUPS puede mapear a varios SOAT (multi-mapeo ~11.7%).
"""

from __future__ import annotations

import gzip
import json
import os
import sys
from collections import OrderedDict

# Ruta de salida relativa a la raíz del repo.
_OUT = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "app",
    "data",
    "cups_soat_homologacion.json.gz",
)


def _txt(v) -> str:
    return str(v).strip() if v is not None else ""


def _buscar_encabezado(rows: list[tuple]) -> tuple[int, dict[str, int]] | tuple[None, None]:
    """Ubica la fila del encabezado y en qué columna quedó cada campo.

    En el Gold Standard el encabezado NO está en la primera fila: arriba hay
    un título («HOMOLOGADOR CUPS A SOAT», «GOLD STANDARD»). Buscarlo por su
    contenido evita tener que saberse el número de fila, que cambia cada vez
    que alguien le agrega una línea al título.
    """
    for i, fila in enumerate(rows[:40]):
        cabecera = {_txt(c).upper(): j for j, c in enumerate(fila) if _txt(c)}
        if "CUPS VIGENTE" in cabecera and "CÓDIGO SOAT" in cabecera:
            return i, cabecera
    return None, None


def _leer_gold_standard(rows: list[tuple]) -> "OrderedDict[str, list[dict]]":
    """Lee el formato de 44 columnas."""
    hdr_idx, cab = _buscar_encabezado(rows)
    if hdr_idx is None or cab is None:
        return OrderedDict()

    def col(*nombres: str) -> int | None:
        for n in nombres:
            if n in cab:
                return cab[n]
        return None

    i_cups = col("CUPS VIGENTE")
    i_soat = col("CÓDIGO SOAT", "CODIGO SOAT")
    i_desc_soat = col("DESCRIPCIÓN SOAT", "DESCRIPCION SOAT")
    i_desc_cups = col("DESCRIPCIÓN", "DESCRIPCION")
    i_art = col("ARTÍCULO SOAT", "ARTICULO SOAT")
    i_cap = col("CAPÍTULO SOAT", "CAPITULO SOAT")
    i_clase = col("CLASE")
    i_uvb = col("VALOR UVB")
    i_cob = col("COBERTURA")

    def celda(fila: tuple, idx: int | None) -> str:
        if idx is None or idx >= len(fila):
            return ""
        return _txt(fila[idx])

    salida: "OrderedDict[str, list[dict]]" = OrderedDict()
    for fila in rows[hdr_idx + 1 :]:
        cups_s = celda(fila, i_cups)
        soat_s = celda(fila, i_soat)
        # El archivo usa ";" como "aquí no hay nada". Si se deja pasar, se
        # cargan homologaciones al código SOAT ";", que no existe.
        if not cups_s or cups_s == ";":
            continue
        if not soat_s or soat_s == ";":
            # El manual no le asigna código SOAT a este procedimiento. Es un
            # hecho útil para el dictamen —la entidad no puede objetar la
            # tarifa citando un SOAT que no existe—, así que se guarda
            # marcado, NUNCA como si fuera un código.
            if cups_s not in salida:
                salida[cups_s] = [
                    {
                        "sin_homologacion": True,
                        "desc_cups": celda(fila, i_desc_cups),
                        **(
                            {"cobertura": celda(fila, i_cob)}
                            if celda(fila, i_cob) and celda(fila, i_cob) != ";"
                            else {}
                        ),
                    }
                ]
            continue
        # Si antes se había marcado como sin homologación y ahora aparece un
        # código de verdad, el código manda.
        if salida.get(cups_s) and all(e.get("sin_homologacion") for e in salida[cups_s]):
            salida[cups_s] = []
        entrada = {
            "soat": soat_s,
            "desc_soat": celda(fila, i_desc_soat),
            "desc_cups": celda(fila, i_desc_cups),
        }
        # Solo se guarda lo que de verdad venga: un campo vacío en el JSON
        # invita a que el dictamen lo cite en blanco.
        for llave, idx in (
            ("articulo", i_art),
            ("capitulo", i_cap),
            ("clase", i_clase),
            ("uvb", i_uvb),
            ("cobertura", i_cob),
        ):
            v = celda(fila, idx)
            if v and v != ";":
                entrada[llave] = v
        salida.setdefault(cups_s, [])
        if not any(e["soat"] == soat_s for e in salida[cups_s]):
            salida[cups_s].append(entrada)
    return salida


def _leer_formato_antiguo(rows: list[tuple]) -> "OrderedDict[str, list[dict]]":
    """Lee el formato de 5 columnas (el cargado en junio de 2025)."""
    salida: "OrderedDict[str, list[dict]]" = OrderedDict()
    for r in rows[1:]:
        if len(r) < 5:
            continue
        cups, soat, desc_soat, _cups2, desc_cups = r[0], r[1], r[2], r[3], r[4]
        if cups is None or soat is None:
            continue
        cups_s = _txt(cups)
        soat_s = _txt(soat)
        entrada = {
            "soat": soat_s,
            "desc_soat": _txt(desc_soat),
            "desc_cups": _txt(desc_cups),
        }
        salida.setdefault(cups_s, [])
        if not any(e["soat"] == soat_s for e in salida[cups_s]):
            salida[cups_s].append(entrada)
    return salida


def generar(src_xlsx: str, out_path: str = _OUT, version: str = "") -> dict:
    import openpyxl  # import local para no exigir openpyxl en producción

    wb = openpyxl.load_workbook(src_xlsx, read_only=True, data_only=True)
    # El Gold Standard trae también TRAZABILIDAD y NOTAS DE AUDITORIA; la de
    # los códigos es la que se llama CUPS.
    nombre_hoja = "CUPS" if "CUPS" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[nombre_hoja]
    rows: list[tuple] = list(ws.iter_rows(values_only=True))
    wb.close()

    hdr_idx, _cab = _buscar_encabezado(rows)
    formato = "GOLD_STANDARD" if hdr_idx is not None else "ANTIGUO_5_COLUMNAS"
    cups_a_soat = _leer_gold_standard(rows) if hdr_idx is not None else _leer_formato_antiguo(rows)

    n = sum(len(v) for v in cups_a_soat.values())
    con_articulo = sum(1 for v in cups_a_soat.values() for e in v if e.get("articulo"))
    sin_homol = sum(
        1 for v in cups_a_soat.values() if v and all(e.get("sin_homologacion") for e in v)
    )

    data = {
        "_meta": {
            "fuente": "Homologación CUPS (Manual Único) → Manual Tarifario SOAT",
            "descripcion": (
                "Mapeo oficial de código CUPS a código SOAT con descripciones. "
                "Un CUPS puede mapear a varios SOAT."
            ),
            "formato": formato,
            "hoja": nombre_hoja,
            "filas_origen": n,
            "cups_unicos": len(cups_a_soat),
            "soat_unicos": len(
                {e["soat"] for v in cups_a_soat.values() if v for e in v if e.get("soat")}
            ),
            "con_articulo_soat": con_articulo,
            "sin_homologacion_directa": sin_homol,
            "version": version or "2025-06-30",
        },
        "cups_a_soat": cups_a_soat,
    }

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with gzip.open(out_path, "wt", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
    return data["_meta"]


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python tools/generar_cups_soat_json.py /ruta/al/HOMOLOGADOR.xlsx [version]")
        sys.exit(1)
    meta = generar(sys.argv[1], version=sys.argv[2] if len(sys.argv) > 2 else "")
    print(f"JSON.gz generado en {_OUT}")
    print(json.dumps(meta, ensure_ascii=False, indent=2))
