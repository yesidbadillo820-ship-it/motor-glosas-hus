"""extraer_respuestas_glosa.py — Pre-rellena un Excel de respuestas a glosas desde el PDF.

Lee uno o varios PDF de "Trámite de Objeción" del HUS y produce un Excel con
UNA fila por objeción (factura, #obj, código, valor objetado, valor aceptado,
servicio y la respuesta del HUS extraída del campo "Observaciones:").

Ese Excel es el INPUT del script `responder_glosas_simed.py`: vos lo revisás
en Excel (corregir aceptado, ajustar texto si querés) y lo guardás. El
responder lo lee, sanitiza el texto (sin tildes/ñ) y lo carga al portal.

USO RÁPIDO
----------

    # Un PDF
    py extraer_respuestas_glosa.py ^
        --pdf    "D:\\...\\SOPORTES\\HUS452150.pdf" ^
        --salida "D:\\...\\respuestas_glosa.xlsx"

    # Una carpeta con varios PDFs (uno por factura)
    py extraer_respuestas_glosa.py ^
        --carpeta "D:\\USUARIO CARTERA\\Documents\\GLOSAS 2026\\DISPENSARIO MEDICO\\Nueva carpeta\\SOPORTES" ^
        --salida  "D:\\...\\respuestas_glosa.xlsx"

DEPENDENCIAS
------------
    py -m pip install pdfplumber openpyxl
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from pathlib import Path

logger = logging.getLogger("extraer_respuestas")

# Códigos de objeción/respuesta (2 letras + 3 dígitos, ej. TA080, SO070, RE990).
RE_CODIGO = re.compile(r"[A-Z]{2}\d{3}")

# Cabecera del PDF: factura HUS.
RE_FACTURA = re.compile(r"FACTURA:\s*(HUS\d{6,12})", re.IGNORECASE)
RE_NUM_OBJECION_DOC = re.compile(r"N[°º]?\s*OBJECION:\s*(\d+)", re.IGNORECASE)

# Bloque de UNA objeción: arranca con código al inicio de línea, agarra los
# dos valores ($ X,XX $ Y,YY), todo lo del medio (incluye servicio) y el texto
# después de "Observaciones:" hasta la siguiente objeción o fin del doc.
RE_BLOQUE = re.compile(
    r"^(?P<cod>[A-Z]{2}\d{3})"
    r".*?\$\s*(?P<v_obj>[\d.,]+)\s*\$\s*(?P<v_ace>[\d.,]+)"
    r"(?P<medio>.*?)Observaciones:\s*(?P<obs>.*?)"
    r"(?=^[A-Z]{2}\d{3}|\Z)",
    re.MULTILINE | re.DOTALL,
)

# Línea de servicio dentro del bloque: "734005 - LABORATORIO - INMUNOLOGIA".
RE_SERVICIO = re.compile(r"(\d{4,7}\s*-\s*[A-ZÁÉÍÓÚÑ].*?)(?=\n|Observaciones)")


def setup_logging() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")


def _to_int(s: str) -> int:
    """Convierte '207.703,00' (formato Colombia: miles con `.`, decimal con `,`)
    al entero 207703. Ignora la parte decimal porque las glosas siempre son enteras."""
    parte_entera = s.split(",")[0]
    return int(re.sub(r"\.", "", parte_entera) or "0")


def extraer_texto_pdf(ruta: Path) -> str:
    """Devuelve todo el texto del PDF concatenado, con saltos de página."""
    try:
        import pdfplumber
    except ImportError:
        sys.stderr.write(
            "ERROR: falta pdfplumber.\n"
            "Instalalo con:  py -m pip install pdfplumber\n"
        )
        sys.exit(2)
    with pdfplumber.open(str(ruta)) as pdf:
        return "\n".join((p.extract_text() or "") for p in pdf.pages)


def extraer_objeciones(ruta: Path) -> list[dict]:
    """Extrae las objeciones de un PDF de Trámite de Objeción.

    Devuelve una lista de dicts (uno por objeción) con:
      factura, num_objecion_doc, num, cod_objecion, valor_objetado,
      valor_aceptado, servicio, observaciones.
    """
    texto = extraer_texto_pdf(ruta)

    m_fac = RE_FACTURA.search(texto)
    factura = m_fac.group(1) if m_fac else ruta.stem.upper()

    m_obj_doc = RE_NUM_OBJECION_DOC.search(texto)
    num_obj_doc = m_obj_doc.group(1) if m_obj_doc else ""

    objeciones: list[dict] = []
    for m in RE_BLOQUE.finditer(texto):
        srv_match = RE_SERVICIO.search(m.group("medio"))
        objeciones.append(
            {
                "factura": factura,
                "num_objecion_doc": num_obj_doc,
                "num": len(objeciones) + 1,  # 1-based en orden de aparición
                "cod_objecion": m.group("cod"),
                "valor_objetado": _to_int(m.group("v_obj")),
                "valor_aceptado": _to_int(m.group("v_ace")),
                "servicio": (srv_match.group(1).strip() if srv_match else ""),
                "observaciones": re.sub(r"\s+", " ", m.group("obs")).strip(),
            }
        )
    return objeciones


def escribir_excel(filas: list[dict], salida: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        sys.stderr.write("ERROR: falta openpyxl. Instalalo con: py -m pip install openpyxl\n")
        sys.exit(2)

    wb = Workbook()
    ws = wb.active
    ws.title = "Respuestas Glosa"

    encabezados = [
        ("Factura", 18),
        ("# Objeción", 11),
        ("Cód.", 8),
        ("Servicio", 38),
        ("Valor Objetado", 16),
        ("Valor Aceptado", 16),
        ("Detalle Respuesta", 100),
    ]
    fill_hdr = PatternFill("solid", fgColor="1F4E78")
    font_hdr = Font(bold=True, color="FFFFFF")
    for col, (etiq, ancho) in enumerate(encabezados, start=1):
        c = ws.cell(row=1, column=col, value=etiq)
        c.fill = fill_hdr
        c.font = font_hdr
        ws.column_dimensions[c.column_letter].width = ancho

    for r, fila in enumerate(filas, start=2):
        ws.cell(row=r, column=1, value=fila["factura"])
        ws.cell(row=r, column=2, value=fila["num"])
        ws.cell(row=r, column=3, value=fila["cod_objecion"])
        ws.cell(row=r, column=4, value=fila["servicio"])
        ws.cell(row=r, column=5, value=fila["valor_objetado"]).number_format = "#,##0"
        ws.cell(row=r, column=6, value=fila["valor_aceptado"]).number_format = "#,##0"
        det = ws.cell(row=r, column=7, value=fila["observaciones"])
        det.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    wb.save(str(salida))


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Pre-rellena Excel de respuestas a glosas desde el PDF Trámite de Objeción.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    grupo = parser.add_mutually_exclusive_group(required=True)
    grupo.add_argument("--pdf", type=Path, help="Un solo PDF de Trámite de Objeción.")
    grupo.add_argument(
        "--carpeta",
        type=Path,
        help="Carpeta con varios PDFs (cada uno = una factura).",
    )
    parser.add_argument("--salida", type=Path, required=True, help="Excel .xlsx de salida.")
    args = parser.parse_args()
    setup_logging()

    pdfs: list[Path]
    if args.pdf:
        if not args.pdf.is_file():
            logger.error(f"No existe el PDF: {args.pdf}")
            return 1
        pdfs = [args.pdf]
    else:
        if not args.carpeta.is_dir():
            logger.error(f"No es una carpeta: {args.carpeta}")
            return 1
        pdfs = sorted(p for p in args.carpeta.iterdir() if p.suffix.lower() == ".pdf")
        if not pdfs:
            logger.error(f"La carpeta no tiene PDFs: {args.carpeta}")
            return 1

    todas: list[dict] = []
    for p in pdfs:
        try:
            obj = extraer_objeciones(p)
        except Exception as e:
            logger.error(f"{p.name}: ERROR — {e}")
            continue
        if not obj:
            logger.warning(f"{p.name}: 0 objeciones (¿formato distinto?)")
            continue
        fac = obj[0]["factura"]
        total_obj = sum(o["valor_objetado"] for o in obj)
        logger.info(f"{p.name} → factura {fac}: {len(obj)} objeciones, total objetado ${total_obj:,}")
        todas.extend(obj)

    if not todas:
        logger.error("No se extrajo ninguna objeción.")
        return 1

    args.salida.parent.mkdir(parents=True, exist_ok=True)
    escribir_excel(todas, args.salida)
    logger.info(f"\nExcel generado: {args.salida}")
    logger.info(f"  Facturas: {len({o['factura'] for o in todas})}")
    logger.info(f"  Objeciones: {len(todas)}")
    logger.info("\nProximo paso:")
    logger.info("  1) Abrí el Excel, revisá 'Detalle Respuesta' (extraído del PDF) y 'Valor Aceptado'.")
    logger.info("  2) Corré responder_glosas_simed.py con --excel <ese.xlsx>.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
