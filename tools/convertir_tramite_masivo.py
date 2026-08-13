"""convertir_tramite_masivo.py — Convierte el export de CRRP a Excel de respuestas.

Lee el Excel masivo del CRRP (`DISPENSARIO TRAMITE DE OBJECIONES MASIVO.xlsx`,
una fila por objeción × concepto) y produce un Excel con el formato exacto que
espera `responder_glosas_simed.py`:

    Factura | # Objeción | Cód. | Servicio | Valor Objetado | Valor Aceptado | Detalle Respuesta

La numeración de # Objeción es 1..N en el orden en que aparecen las filas para
cada factura — coincide con la numeración de la grilla del portal SIMED
(verificado contra HUS0000452150).

USO:
    py convertir_tramite_masivo.py ^
        --tramite "D:\\USUARIO CARTERA\\Downloads\\DISPENSARIO TRAMITE DE OBJECIONES MASIVO.xlsx" ^
        --salida  "D:\\USUARIO CARTERA\\Documents\\GLOSAS 2026\\respuestas_MASIVO.xlsx"
"""

from __future__ import annotations

import argparse
import logging
import sys
from collections import defaultdict
from pathlib import Path

# Lector de pesos compartido (ver `tools/_dinero.py`): el que vivía aquí
# leía cero cuando el valor venía en formato colombiano.
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _dinero import a_entero  # noqa: E402

logger = logging.getLogger("convertir_tramite")

# Índices verificados del Excel de CRRP.
IDX_FACTURA = 7
IDX_VALOR_OBJETADO = 23
IDX_VALOR_ACEPTADO = 24
IDX_CODIGO = 25  # ConceptoObjecion.Codigo (TA0801, etc.)
IDX_SERVICIO = 30  # ServicioProductoFactura.Descripcion
IDX_OBS = 32  # Observaciones del detalle (la RESPUESTA del HUS)


def _num(v) -> int:
    """Pesos enteros, con el lector único de `tools/_dinero.py`.

    Lo que hacía antes: `int(float(str(v).replace(",", ".")))`. Cambiaba la
    coma por punto y nada más, así que el formato colombiano lo destruía:

        "1.234.567"  →  ValueError  →  0      un millón se volvía cero
        "$114.900"   →  ValueError  →  0
        "50.000"     →  50                    dividido por mil
        "1.365,50"   →  ValueError  →  0

    Décima copia de la misma regla en `tools/`, y la única que seguía viva en
    producción cuando se encontró.
    """
    return a_entero(v)


def convertir(ruta_tramite: Path, ruta_salida: Path) -> tuple[int, int]:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        sys.stderr.write("ERROR: falta openpyxl. py -m pip install openpyxl\n")
        sys.exit(2)

    wb = load_workbook(filename=str(ruta_tramite), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))[1:]  # saltar encabezado

    # Agrupar por factura preservando orden de aparición.
    por_factura: dict[str, list[tuple]] = defaultdict(list)
    for r in rows:
        fac = r[IDX_FACTURA]
        if not fac:
            continue
        por_factura[str(fac).strip()].append(r)

    # Escribir el Excel destino con el formato del responder.
    out = Workbook()
    sh = out.active
    sh.title = "Respuestas Glosa"
    encabezados = [
        ("Factura", 18),
        ("# Objeción", 11),
        ("Cód.", 8),
        ("Servicio", 38),
        ("Valor Objetado", 16),
        ("Valor Aceptado", 16),
        ("Detalle Respuesta", 100),
    ]
    fill_hdr = PatternFill("solid", fgColor="1F4E78")
    font_hdr = Font(bold=True, color="FFFFFF")
    for col, (etiq, ancho) in enumerate(encabezados, start=1):
        c = sh.cell(row=1, column=col, value=etiq)
        c.fill = fill_hdr
        c.font = font_hdr
        sh.column_dimensions[c.column_letter].width = ancho

    fila_dest = 2
    total_obj = 0
    for factura, fac_rows in sorted(por_factura.items()):
        # # Objeción = posición 1..N en el orden del Excel.
        for n, r in enumerate(fac_rows, start=1):
            sh.cell(row=fila_dest, column=1, value=factura)
            sh.cell(row=fila_dest, column=2, value=n)
            sh.cell(row=fila_dest, column=3, value=str(r[IDX_CODIGO] or "")[:10])
            sh.cell(row=fila_dest, column=4, value=str(r[IDX_SERVICIO] or "")[:200])
            sh.cell(
                row=fila_dest, column=5, value=_num(r[IDX_VALOR_OBJETADO])
            ).number_format = "#,##0"
            sh.cell(
                row=fila_dest, column=6, value=_num(r[IDX_VALOR_ACEPTADO])
            ).number_format = "#,##0"
            det = sh.cell(row=fila_dest, column=7, value=str(r[IDX_OBS] or ""))
            det.alignment = Alignment(wrap_text=True, vertical="top")
            fila_dest += 1
            total_obj += 1

    sh.freeze_panes = "A2"
    out.save(str(ruta_salida))
    return len(por_factura), total_obj


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        "--tramite", type=Path, required=True, help="Excel del CRRP (TRAMITE DE OBJECIONES MASIVO)."
    )
    parser.add_argument(
        "--salida", type=Path, required=True, help="Excel destino con formato del responder."
    )
    args = parser.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

    if not args.tramite.is_file():
        logger.error(f"No existe: {args.tramite}")
        return 1

    n_fac, n_obj = convertir(args.tramite, args.salida)
    logger.info(f"\nExcel convertido: {args.salida}")
    logger.info(f"  Facturas: {n_fac:,}")
    logger.info(f"  Objeciones: {n_obj:,}")
    logger.info("\nPróximo paso:")
    logger.info(
        '  py tools\\responder_glosas_simed.py --excel "<este.xlsx>" --todas --sin-soportes ...'
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
