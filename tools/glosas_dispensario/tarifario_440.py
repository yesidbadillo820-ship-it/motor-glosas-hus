"""Tarifario del contrato 440-DIGSA-DMBUG-2025 (CD477) para el cruce dinámico.

Carga los anexos que entregó Contratación y arma un índice por código:
  - 6.2 PRECIOS DE REFERENCIA (servicios): hojas SOAT, SERVICIOS DE
    PROCEDIMIENTOS, PAQUETES Y/O PROGRAMAS, LABORATORIO E IMAGENOLOGIA —
    columnas CUPS / DESCRIPCION / CODIGO IPS / PRECIO DE REFERENCIA / TARIFA.
  - TARIFAS DEL CONTRATO (medicamentos y dispositivos): anexos por CODIGO
    CUM / NOMBRE / PRECIO DE VENTA.

La búsqueda va de lo exacto a lo general: código IPS tal cual (039001H,
902210AMB) → CUPS pelado (sin sufijo) → CUM (con o sin el guion de la
presentación). Sin coincidencia devuelve None: la respuesta NO cita tarifas
que no estén pactadas (regla de no invención).
"""

from __future__ import annotations

import re
from pathlib import Path


def _num(v) -> int | None:
    if v in (None, ""):
        return None
    s = str(v).replace("$", "").replace(" ", "").strip()
    if s.count(",") and s.count("."):
        s = s.replace(".", "").replace(",", ".")
    elif s.count(","):
        s = s.replace(",", ".")
    try:
        return round(float(s))
    except ValueError:
        return None


def _norm_cod(c) -> str:
    return re.sub(r"\s", "", str(c or "")).upper()


def cargar_tarifario(ruta_servicios: Path | None, ruta_medicamentos: Path | None) -> dict:
    """Devuelve el índice {codigo_normalizado: {precio, descripcion, fuente}}."""
    from openpyxl import load_workbook

    indice: dict[str, dict] = {}

    def anotar(codigo, precio, descripcion, fuente):
        cod = _norm_cod(codigo)
        if cod and precio and cod not in indice:
            indice[cod] = dict(
                precio=precio,
                descripcion=" ".join(str(descripcion or "").split())[:80],
                fuente=fuente,
            )

    if ruta_servicios and Path(ruta_servicios).is_file():
        wb = load_workbook(ruta_servicios, read_only=True, data_only=True)
        for hoja in wb.sheetnames:
            ws = wb[hoja]
            encabezado = None
            for fila in ws.iter_rows(values_only=True):
                vals = [str(v).strip() if v is not None else "" for v in fila]
                if encabezado is None:
                    if any(v.upper().startswith("CUPS") for v in vals) and any(
                        "PRECIO" in v.upper() for v in vals
                    ):
                        encabezado = {v.upper(): i for i, v in enumerate(vals) if v}
                        i_cups = next(i for v, i in encabezado.items() if v.startswith("CUPS"))
                        i_ips = next((i for v, i in encabezado.items() if "CODIGO IPS" in v), None)
                        i_des = next(
                            (
                                i
                                for v, i in encabezado.items()
                                if "DESCRIPCION" in v or "DESCRIPCIÓN" in v
                            ),
                            None,
                        )
                        i_pre = next(i for v, i in encabezado.items() if "PRECIO" in v)
                    continue
                precio = _num(fila[i_pre]) if i_pre < len(fila) else None
                des = fila[i_des] if i_des is not None and i_des < len(fila) else ""
                if i_ips is not None and i_ips < len(fila):
                    anotar(fila[i_ips], precio, des, f"anexo 6.2 · {hoja.strip()}")
                anotar(fila[i_cups], precio, des, f"anexo 6.2 · {hoja.strip()}")
        wb.close()

    if ruta_medicamentos and Path(ruta_medicamentos).is_file():
        wb = load_workbook(ruta_medicamentos, read_only=True, data_only=True)
        for hoja in wb.sheetnames:
            ws = wb[hoja]
            encabezado = None
            for fila in ws.iter_rows(values_only=True):
                vals = [str(v).strip() if v is not None else "" for v in fila]
                if encabezado is None:
                    if any(v.upper().startswith(("CODIGO CUM", "CODIGO")) for v in vals) and any(
                        "PRECIO" in v.upper() for v in vals
                    ):
                        arriba = {v.upper(): i for i, v in enumerate(vals) if v}
                        i_cod = next(i for v, i in arriba.items() if v.startswith("CODIGO"))
                        i_nom = next(
                            (
                                i
                                for v, i in arriba.items()
                                if "NOMBRE" in v or "DESCRIPCION" in v or "DESCRIPCIÓN" in v
                            ),
                            None,
                        )
                        i_pre = next(i for v, i in arriba.items() if "PRECIO" in v)
                        encabezado = True
                    continue
                precio = _num(fila[i_pre]) if i_pre < len(fila) else None
                nom = fila[i_nom] if i_nom is not None and i_nom < len(fila) else ""
                cod = _norm_cod(fila[i_cod] if i_cod < len(fila) else "")
                anotar(cod, precio, nom, f"tarifas del contrato · {hoja.strip()}")
                if "-" in cod:  # el CUM también sin la presentación (20103720-02 → 20103720)
                    anotar(cod.split("-")[0], precio, nom, f"tarifas del contrato · {hoja.strip()}")
        wb.close()

    return indice


def tarifa_de(indice: dict, codigo_servicio) -> dict | None:
    """Busca el código del servicio glosado: exacto → CUPS pelado → CUM base."""
    cod = _norm_cod(codigo_servicio)
    if not cod:
        return None
    if cod in indice:
        return indice[cod]
    pelado = re.sub(r"[A-Z]+$", "", cod)  # 039001H → 039001
    if pelado and pelado in indice:
        return indice[pelado]
    if "-" in cod and cod.split("-")[0] in indice:
        return indice[cod.split("-")[0]]
    return None


def frase_tarifaria(codigo, tarifa: dict) -> str:
    valor = f"${tarifa['precio']:,.0f}".replace(",", ".")
    return (
        f"LA TARIFA PACTADA EN EL ANEXO TARIFARIO DEL CONTRATO 440-DIGSA-DMBUG-2025 "
        f"PARA EL CODIGO {_norm_cod(codigo)} ({tarifa['descripcion']}) ES DE {valor} "
        f"({tarifa['fuente'].upper()}), FILA QUE SE REMITE CON LA PRESENTE RESPUESTA."
    )
