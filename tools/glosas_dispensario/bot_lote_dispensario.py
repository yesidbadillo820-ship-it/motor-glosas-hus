"""BOT RPA del lote de glosas del Dispensario — flujo completo por paquete GI.

Uso (en el equipo de cartera, desde C:\\motor-glosas\\repo):

    py tools\\glosas_dispensario\\bot_lote_dispensario.py --excel "D:\\...\\GLOSAS_X.xlsx"

Qué hace, en orden:
  1. Pide el código GI (o recíbelo con --gi) y crea la carpeta del paquete
     con su subcarpeta `soportes`: <base>\\<GI>\\soportes.
  2. Genera el Excel de respuestas con gen_lote.py — hereda la DIRECTRIZ CL:
     las objeciones CL de glosas Médica/Mixta NO se responden (van a la hoja
     "PARA GESTION MEDICA" y quedan intactas para la nota crédito del equipo
     médico).
  3. Recorre UNA vez las carpetas de soportes de radicación de la unidad Y:,
     arma el índice de PDF por factura y copia el de cada factura a `soportes`.
  4. COTEJO DE COBRO (cotejo_tarifa.py): lee cada PDF con la cascada
     pdfplumber → PyPDF2 → OCR (extraer_factura_pdf.py), saca el valor por el
     que de verdad se facturó el servicio y lo compara con la tarifa pactada.
     Escribe al lado de cada respuesta si hay mayor valor cobrado, cuánto, y
     la RESPUESTA SUGERIDA. Si la misma diferencia porcentual se repite en el
     lote, es la actualización de tarifas del año y NO se sugiere aceptar. El
     Excel del cargue conserva el Valor Aceptado en 0: aceptar una glosa lo
     decide el auditor, no el bot.
  5. LA REDACCION, por escenario: el argumento de cada respuesta lo dicta el
     veredicto del cotejo, con las mismas cifras verificadas —
       · cobro a tarifa    → el valor facturado ES lo pactado: causal infundada;
       · vigencia 2026     → el mayor valor es el ajuste de los parágrafos 3
                             y 4 del contrato: el cobro es válido;
       · sobrecobro real   → se redacta la aceptación parcial (solo entra al
                             texto que se sube con --redactar-aceptacion).
     Fuera de esos tres manda la redacción prudente, que cita el PDF y el
     anexo pero no proclama cifras sin cotejo. Nada leido ni pactado = nada
     agregado (no se inventa).
  6. Si no se pasa --sin-cargue, corre el robot del portal
     (responder_glosas_simed.py) con el Excel del paquete.
  7. Convierte las evidencias de la corrida (los .png del robot) en el PDF
     del paquete <GI>_EVIDENCIAS.pdf y deja todo dentro de la carpeta GI.

Reglas: piloto antes de masivo (use --piloto HUS...), credenciales del portal
solo por variables de entorno (SIMED_USER / SIMED_PASSWORD), y español claro.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import sys
import time
from collections import Counter
from pathlib import Path

AQUI = Path(__file__).resolve().parent
REPO = AQUI.parents[1]
sys.path.insert(0, str(AQUI))

import cotejo_tarifa  # noqa: E402  (el cotejo de lo facturado contra lo pactado)

RUTAS_SOPORTES_DEFECTO = [
    r"Y:\2. FEBRERO 2026 - SOPORTES RADICACION CARPETA 2",
    r"Y:\3. MARZO 2026 - SOPORTES RADICACION",
    r"Y:\4. ABRIL 2026 - SOPORTES RADICACION",
    r"Y:\5. MAYO 2026 - SOPORTES RADICACION",
    r"Y:\6. JUNIO 2026 - SOPORTES RADICACION",
    r"Y:\7. JULIO 2026 - SOPORTES RADICACION",
    r"Y:\8. AGOSTO 2026 - SOPORTES RADICACION",
    r"Y:\9.SEPTIEMBRE - SOPORTES RADICACION",
]
BASE_DEFECTO = r"D:\USUARIO CARTERA\Documents"


def asegurar_dependencias() -> None:
    """Autoinstala lo que falte (patrón de los bots del repo)."""
    for paquete, modulo in [
        ("openpyxl", "openpyxl"),
        ("pdfplumber", "pdfplumber"),
        ("PyPDF2", "PyPDF2"),
        ("pymupdf", "pymupdf"),
        ("pillow", "PIL"),
    ]:
        try:
            __import__(modulo)
        except ImportError:
            print(f"Instalando {paquete}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", paquete])


def factura_corta(fac: str) -> str:
    return str(fac).replace("HUS", "").lstrip("0") or "0"


RE_FACTURA_EN_NOMBRE = re.compile(r"HUS0*(\d{4,})", re.IGNORECASE)


def indexar_soportes(rutas: list[str]) -> dict[str, list[Path]]:
    """Recorre UNA sola vez las carpetas de radicación y arma el índice
    {número de factura → PDF que le corresponde}.

    Las carpetas de la unidad Y: vienen así:

        Y:\\9.SEPTIEMBRE - SOPORTES RADICACION\\DISPENSARIO\\LILIANA\\
            ENV-233972-OK\\HUS552002\\FEV_900006037_HUS552002.pdf

    o sea: la factura está en el nombre del archivo Y en el de su carpeta, y
    dentro de esa carpeta hay varios soportes (epicrisis, autorización...).
    Por eso el índice ordena los candidatos: primero la factura electrónica
    (los archivos `FEV_...`, que son los que traen el detalle del cobro),
    luego los que nombran la factura, y de último el resto de la carpeta.
    Entre meses manda el más reciente (el último de `rutas`).

    Recorrer una vez y no una por factura es lo que hace viable el lote: son
    ocho carpetas de red con decenas de miles de archivos.
    """
    indice: dict[str, list[tuple[tuple[int, int], Path]]] = {}
    for orden_raiz, raiz in enumerate(reversed(rutas)):  # el mes más reciente primero
        base = Path(raiz)
        if not base.is_dir():
            continue
        try:
            archivos = list(base.rglob("*.pdf"))
        except OSError as e:
            print(f"    (no pude recorrer {raiz}: {e})")
            continue
        for pdf in archivos:
            en_nombre = {m.group(1).lstrip("0") for m in RE_FACTURA_EN_NOMBRE.finditer(pdf.name)}
            en_carpeta = {
                m.group(1).lstrip("0") for m in RE_FACTURA_EN_NOMBRE.finditer(pdf.parent.name)
            }
            for clave in en_nombre | en_carpeta:
                if clave in en_nombre:
                    prioridad = 0 if pdf.name.upper().startswith("FEV") else 1
                else:
                    prioridad = 2  # solo lo nombra la carpeta: es otro soporte
                indice.setdefault(clave, []).append(((orden_raiz, prioridad), pdf))
    return {
        clave: [ruta for _, ruta in sorted(cands, key=lambda c: c[0])]
        for clave, cands in indice.items()
    }


def buscar_en_indice(indice: dict[str, list[Path]], fac: str) -> Path | None:
    """El PDF de la factura dentro del índice ya recorrido."""
    candidatos = indice.get(factura_corta(fac))
    return candidatos[0] if candidatos else None


def buscar_pdf_factura(fac: str, rutas: list[str]) -> Path | None:
    """Busca el PDF de UNA factura (recorre las carpetas). Para un lote entero
    use `indexar_soportes` una vez y luego `buscar_en_indice`."""
    return buscar_en_indice(indexar_soportes(rutas), fac)


def frase_anclaje(fac: str, datos: dict) -> str | None:
    """SOLO con datos extraídos del PDF. Sin paciente ni total no hay frase."""
    partes = [f"LA FACTURA {fac}"]
    if datos.get("paciente"):
        partes.append(f"EXPEDIDA A NOMBRE DEL PACIENTE {datos['paciente']}")
    if datos.get("total"):
        partes.append(f"POR VALOR TOTAL DE ${datos['total']:,.0f}".replace(",", "."))
    if len(partes) == 1:
        return None
    return (
        ", ".join(partes) + ", SE ANEXA COMO SOPORTE DE LA PRESENTE RESPUESTA "
        "JUNTO CON SUS ANEXOS DE RADICACION."
    )


def _plata(v) -> str:
    return f"${v:,.0f}".replace(",", ".")


def hallar_servicio_en_pdf(
    servicios: list[list], cups: str, descripcion: str
) -> tuple[str, int | None, list[int]] | None:
    """Ubica en las filas extraídas del PDF la del servicio glosado: primero
    por el código (con o sin sufijo), luego por palabras de la descripción.
    Devuelve (texto de la fila, último valor de la fila, todos sus importes)
    o None. Los importes van completos porque el cotejo necesita distinguir
    el valor unitario del total de la línea."""
    cod = re.sub(r"\s", "", str(cups or "")).upper()
    pelado = re.sub(r"[A-Z]+$", "", cod)
    palabras = [
        p
        for p in re.findall(r"[A-ZÑ]{4,}", str(descripcion or "").upper())
        if p not in ("PARA", "COMO")
    ]

    def importes_de(fila: list) -> list[int]:
        """Los valores en pesos de la fila, sin confundirlos con los códigos.

        El código del servicio (890275H, HUS0000542497) tiene letras y el
        centro de costo son dígitos sueltos: leerlos como plata fue el error
        que hacía ver un cobro de $890.275 donde había uno de $192.600. Por eso
        mandan los números escritos como dinero —con separador de miles o con
        $—, y solo si la fila no trae ninguno se miran los dígitos pelados.
        """
        sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
        from _dinero import a_entero  # noqa: PLC0415  (el UNICO lector de pesos)

        con_formato, pelados = [], []
        # Se mira token por token: pdfplumber a veces devuelve el renglón
        # entero en una sola celda ("890275H CONSULTA 1 192.600 192.600"), y
        # mirando la celda completa no se leía ningún valor.
        for celda in fila:
            for texto in str(celda or "").split():
                if not re.search(r"\d", texto):
                    continue
                limpio = re.sub(r"[\s$]", "", texto).upper().rstrip(".,")
                if re.search(r"[A-ZÑ]", limpio):  # 890275H, HUS0000542497, "12UND"
                    continue
                if limpio in (cod, pelado):  # el código del servicio, no un valor
                    continue
                n = a_entero(texto)
                if n < 100:  # cantidades y códigos cortos
                    continue
                (con_formato if re.search(r"[.,]", limpio) else pelados).append(n)
        return con_formato or pelados

    def resultado(fila: list) -> tuple[str, int | None, list[int]]:
        importes = importes_de(fila)
        texto = " ".join(" ".join(str(c) for c in fila).split())[:120]
        return texto, (importes[-1] if importes else None), importes

    for fila in servicios or []:
        fila_txt = " ".join(str(c) for c in fila).upper()
        if cod and (cod in fila_txt or (pelado and pelado in fila_txt)):
            return resultado(fila)
    for fila in servicios or []:
        fila_txt = " ".join(str(c) for c in fila).upper()
        if palabras and sum(1 for p in palabras if p in fila_txt) >= max(2, len(palabras) // 2):
            return resultado(fila)
    return None


def _quien(datos_pdf: dict | None) -> str:
    """«DEL USUARIO X» solo si el PDF trajo el nombre. Sin nombre, nada: el
    paciente no se inventa."""
    paciente = (datos_pdf or {}).get("paciente")
    return f" DEL USUARIO {paciente}" if paciente else ""


def parrafo_por_escenario(
    cotejo: dict,
    datos_pdf: dict | None,
    servicio_pdf: tuple | None,
    cups: str = "",
    servicio: str = "",
    factura: str = "",
) -> str | None:
    """Redacta el argumento según el escenario matemático que dictó el cotejo.

    Tres caminos, uno por veredicto, todos con las mismas cifras que el
    cotejo verificó (nada se recalcula acá, para que el texto y las columnas
    del Excel no puedan contradecirse):

      1. COBRO A TARIFA  → el valor facturado ES la tarifa: causal infundada.
      2. MAYOR VALOR POR VIGENCIA → el mayor valor es la actualización 2026
         de los parágrafos 3 y 4: el cobro es contractualmente válido.
      3. MAYOR VALOR VERIFICADO → hay sobrecobro real: se plantea la
         aceptación parcial por la diferencia.

    Fuera de esos tres (sin cotejo posible, o cobro por debajo de lo pactado)
    devuelve None y manda la redacción prudente de `parrafo_evidencia`, que
    no proclama cifras sin cotejo.

    OJO con el escenario 3: este texto acepta una glosa. Quien decide subirlo
    al portal es el auditor (`--redactar-aceptacion`), nunca el bot solo.
    """
    archivo = str((datos_pdf or {}).get("archivo") or "").upper()
    fuente = f" {archivo}" if archivo else ""
    # De dónde salió el valor facturado, dicho con precisión: el soporte que
    # se leyó, o —cuando el PDF no estaba a la mano— la factura en el sistema
    # de cartera del hospital. Nunca se cita un documento que no se abrió.
    if archivo:
        visto = f"AL REVISAR EL SOPORTE{fuente}"
        verificado = f"VERIFICANDO LA FACTURA{fuente}"
        validado = f"VALIDADO EL SOPORTE{fuente}"
    else:
        visto = verificado = validado = (
            f"VERIFICADA LA FACTURA {factura} EN EL SISTEMA DE CARTERA DE LA ESE"
            if factura
            else "VERIFICADO EL VALOR FACTURADO EN EL SISTEMA DE CARTERA DE LA ESE"
        )
    # El servicio se nombra como lo conoce la EPS (código y descripción del
    # export), no con la fila cruda del PDF: esa lleva cantidades y valores
    # sueltos que dentro de la frase se leen como un error. La fila leída
    # queda en la columna FUENTE DEL COTEJO, que es donde sirve.
    etiqueta = " ".join(f"{cups} {servicio}".split()).strip() or (
        servicio_pdf[0] if servicio_pdf else "EL SERVICIO GLOSADO"
    )
    facturado = cotejo.get("valor_facturado")
    pactada = cotejo.get("tarifa")
    veredicto = cotejo.get("veredicto")

    if veredicto == cotejo_tarifa.A_TARIFA:
        return (
            f"{visto}, SE EVIDENCIA LA ATENCION{_quien(datos_pdf)}. "
            f"EL VALOR FACTURADO DE {_plata(facturado)} POR EL SERVICIO {etiqueta} "
            "CORRESPONDE A LA TARIFA PACTADA EXACTA EN EL CONTRATO 440-DIGSA-DMBUG-2025, "
            "POR LO QUE LA CAUSAL DE GLOSA ES INFUNDADA."
        )

    if veredicto == cotejo_tarifa.POR_VIGENCIA:
        return (
            f"{verificado}{_quien(datos_pdf)}, EL VALOR COBRADO DE "
            f"{_plata(facturado)} POR EL SERVICIO {etiqueta} REFLEJA LA ACTUALIZACION DE "
            "TARIFAS DE LA VIGENCIA 2026 CONTEMPLADA EN LOS PARAGRAFOS 3 Y 4 DEL CONTRATO "
            "440-DIGSA-DMBUG-2025 (MODIFICATORIO CONTRACTUAL Y RESOLUCION TARIFARIA DE LA "
            "ESE HOSPITAL UNIVERSITARIO DE SANTANDER), DOCUMENTOS QUE SE REMITEN CON LA "
            "PRESENTE RESPUESTA. EL COBRO ES CONTRACTUALMENTE VALIDO."
        )

    if veredicto == cotejo_tarifa.SOBRECOBRO:
        return (
            f"{validado}{_quien(datos_pdf)}, LA TARIFA PACTADA PARA EL "
            f"CODIGO {cups} ES DE {_plata(pactada)} Y SE FACTURO {_plata(facturado)}. "
            f"SE ACEPTA LA GLOSA POR EL MAYOR VALOR COBRADO DE {_plata(cotejo['aceptar'])}"
            + (
                ", Y SE SOLICITA EL LEVANTAMIENTO DE LOS "
                f"{_plata(cotejo['resto'])} RESTANTES, POR CORRESPONDER A VALOR PACTADO."
                if cotejo.get("resto")
                else "."
            )
        )
    return None


def parrafo_evidencia(
    datos_pdf: dict | None,
    servicio_pdf: tuple[str, int | None] | None,
    tarifa: dict | None,
    valor_objetado: int,
    cups: str = "",
) -> str | None:
    """El constructor de la respuesta con TRAZABILIDAD: abre citando el PDF
    fuente, inyecta lo que de verdad se leyó (paciente, servicio, valor) y
    conecta con el cruce del contrato 440. Reglas duras:
      - sin datos leídos ni tarifa pactada → None (queda el texto del motor);
      - 'COINCIDE EXACTAMENTE' solo se afirma cuando los números coinciden."""
    frases = []
    leyo = bool(datos_pdf and (datos_pdf.get("paciente") or datos_pdf.get("total") or servicio_pdf))
    if leyo:
        frases.append(
            f"AL REVISAR EL DOCUMENTO DE SOPORTE {datos_pdf['archivo'].upper()}"
            + (
                f", SE EVIDENCIA LA ATENCION PRESTADA AL USUARIO {datos_pdf['paciente']}"
                if datos_pdf.get("paciente")
                else ", SE VERIFICA LA FACTURA Y SU RADICACION"
            )
            + "."
        )
        if servicio_pdf:
            texto_serv, valor_serv = servicio_pdf[0], servicio_pdf[1]
            frases.append(
                "EL DOCUMENTO DETALLA EL COBRO DEL SERVICIO "
                + texto_serv
                + (f" POR VALOR DE {_plata(valor_serv)}." if valor_serv else ".")
            )
        elif datos_pdf.get("total"):
            frases.append(
                f"EL DOCUMENTO REGISTRA UN VALOR TOTAL FACTURADO DE {_plata(datos_pdf['total'])}."
            )
    if tarifa:
        valor_leido = servicio_pdf[1] if servicio_pdf else None
        base = (
            "AL CRUZAR ESTA EVIDENCIA CON EL CONTRATO 440-DIGSA-DMBUG-2025, "
            if leyo
            else "AL VERIFICAR EL CONTRATO 440-DIGSA-DMBUG-2025, "
        )
        pactada = f"{_plata(tarifa['precio'])} ({tarifa['fuente'].upper()})"
        if valor_leido and valor_leido == tarifa["precio"]:
            frases.append(
                base + f"SE VERIFICA QUE EL VALOR FACTURADO CORRESPONDE EXACTAMENTE A LA "
                f"TARIFA PACTADA EN EL ANEXO 6.2 ({pactada}), POR LO QUE LA CAUSAL DE "
                "GLOSA ES INFUNDADA."
            )
        elif valor_objetado and valor_objetado == tarifa["precio"]:
            frases.append(
                base + f"SE VERIFICA QUE EL VALOR OBJETADO CORRESPONDE EXACTAMENTE A LA "
                f"TARIFA PACTADA PARA EL CODIGO {cups} ({pactada}), POR LO QUE LA CAUSAL "
                "DE GLOSA ES INFUNDADA."
            )
        else:
            # SIN COTEJO NO SE PROCLAMA CIFRA: citar un valor del anexo que no
            # coincide con lo facturado le sirve a la EPS para ratificar con
            # el propio soporte del hospital (caso HUS0000542497, 04-09-2026).
            frases.append(
                base + f"EL CODIGO {cups} ({tarifa['descripcion']}) SE ENCUENTRA PACTADO EN "
                "EL ANEXO TARIFARIO DEL CONTRATO, FILA QUE SE REMITE CON LA PRESENTE RESPUESTA."
            )
    return " ".join(frases) if frases else None


COLUMNAS_COTEJO = [
    "VALOR FACTURADO (PDF)",
    "TARIFA PACTADA (440)",
    "DIFERENCIA",
    "¿SOBRECOBRO?",
    "VALOR SUGERIDO A ACEPTAR",
    "RESPUESTA SUGERIDA",
    "FUENTE DEL COTEJO",
]
COLORES_VEREDICTO = {
    cotejo_tarifa.SOBRECOBRO: "FFC7CE",  # rojo claro: sí hay que aceptar algo
    cotejo_tarifa.POR_VIGENCIA: "FFEB9C",  # ámbar: revisar antes de aceptar
    cotejo_tarifa.A_TARIFA: "C6EFCE",  # verde: la glosa es infundada
    cotejo_tarifa.POR_DEBAJO: "C6EFCE",
    cotejo_tarifa.SIN_COTEJO: "D9D9D9",  # gris: falta información
}


def agregar_cotejo_excel(ruta_excel: Path, cotejos: dict[tuple[str, int], dict]) -> int:
    """Escribe al lado de cada respuesta el cotejo del cobro: qué se facturó de
    verdad (leído del PDF), qué se pactó en el contrato, la diferencia y la
    RESPUESTA SUGERIDA.

    Las columnas van al final de la hoja "Respuestas Glosa": el robot del
    portal busca sus columnas por nombre, así que no lo estorban. Y el
    "Valor Aceptado" del cargue NO se toca —sigue en 0—: la sugerencia es para
    que el auditor decida, porque aceptar una glosa es decisión del hospital,
    no del bot. Además deja la hoja "COTEJO DE COBRO" con lo que hay que
    revisar (sobrecobros y diferencias por vigencia), que es el paquete de
    trabajo del auditor."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(ruta_excel)
    ws = wb["Respuestas Glosa"]
    encabezados = [str(c.value or "") for c in ws[1]]
    if COLUMNAS_COTEJO[0] in encabezados:
        inicio = encabezados.index(COLUMNAS_COTEJO[0]) + 1
    else:
        inicio = ws.max_column + 1
        for i, nombre in enumerate(COLUMNAS_COTEJO):
            c = ws.cell(row=1, column=inicio + i, value=nombre)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E78")
            c.alignment = Alignment(vertical="center", wrap_text=True)
        for i, ancho in enumerate([18, 18, 14, 26, 16, 90, 46]):
            ws.column_dimensions[get_column_letter(inicio + i)].width = ancho

    escritas = 0
    for fila in ws.iter_rows(min_row=2):
        clave = (str(fila[0].value or "").strip(), int(fila[1].value or 0))
        cot = cotejos.get(clave)
        if not cot:
            continue
        valores = [
            cot.get("valor_facturado"),
            cot.get("tarifa"),
            cot.get("diferencia"),
            cot["veredicto"]
            + (f" (+{cot['porcentaje']}%)" if cot.get("porcentaje") and cot["diferencia"] else ""),
            cot.get("aceptar") or 0,
            cot["respuesta"],
            cot.get("fuente", ""),
        ]
        for i, v in enumerate(valores):
            c = ws.cell(row=fila[0].row, column=inicio + i, value=v)
            c.alignment = Alignment(wrap_text=i >= 3, vertical="top")
            if i < 3 or i == 4:
                c.number_format = '"$"#,##0'
            if i == 3:
                c.fill = PatternFill(
                    "solid", fgColor=COLORES_VEREDICTO.get(cot["veredicto"], "D9D9D9")
                )
        escritas += 1

    revisar = {
        k: v
        for k, v in cotejos.items()
        if v["veredicto"] in (cotejo_tarifa.SOBRECOBRO, cotejo_tarifa.POR_VIGENCIA)
    }
    if "COTEJO DE COBRO" in wb.sheetnames:
        del wb["COTEJO DE COBRO"]
    if revisar:
        hoja = wb.create_sheet("COTEJO DE COBRO")
        cols = ["Factura", "# Objeción", *COLUMNAS_COTEJO[:-1]]
        hoja.append(cols)
        for i, _ in enumerate(cols, 1):
            c = hoja.cell(row=1, column=i)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="9C0006")
            c.alignment = Alignment(vertical="center", wrap_text=True)
        for (fac, num), v in sorted(revisar.items()):
            hoja.append(
                [
                    fac,
                    num,
                    v.get("valor_facturado"),
                    v.get("tarifa"),
                    v.get("diferencia"),
                    v["veredicto"],
                    v.get("aceptar") or 0,
                    v["respuesta"],
                ]
            )
        for i, ancho in enumerate([16, 11, 18, 18, 14, 26, 16, 110], 1):
            hoja.column_dimensions[get_column_letter(i)].width = ancho
        for rr in hoja.iter_rows(min_row=2):
            rr[7].alignment = Alignment(wrap_text=True, vertical="top")
            for c in rr[2:5] + (rr[6],):
                c.number_format = '"$"#,##0'
            rr[5].fill = PatternFill("solid", fgColor=COLORES_VEREDICTO.get(rr[5].value, "D9D9D9"))
        hoja.freeze_panes = "A2"
    wb.save(ruta_excel)
    return escritas


def enriquecer_excel(
    ruta_excel: Path,
    por_factura: dict[str, str] | None = None,
    por_linea: dict[tuple[str, int], str] | None = None,
    apertura_por_linea: dict[tuple[str, int], str] | None = None,
) -> int:
    """Arma la respuesta final de cada línea. `apertura_por_linea` es el
    párrafo de evidencia con trazabilidad: entra JUSTO DESPUES del encabezado
    ("... OBJETADO BAJO EL CONCEPTO XX:"), antes del argumento del motor.
    `por_linea`/`por_factura` se insertan antes del cierre institucional.
    No toca la hoja PARA GESTION MEDICA."""
    import re

    from openpyxl import load_workbook

    MARCA = "POR LO EXPUESTO, SE SOLICITA EL LEVANTAMIENTO"  # arranque del cierre del motor
    RE_ENCABEZADO = re.compile(r"OBJETADO\s+BAJO\s+EL\s+CONCEPTO\s+[A-Z0-9]+\s*:")
    por_factura = por_factura or {}
    por_linea = por_linea or {}
    apertura_por_linea = apertura_por_linea or {}
    wb = load_workbook(ruta_excel)
    ws = wb["Respuestas Glosa"]
    tocadas = 0
    for fila in ws.iter_rows(min_row=2):
        fac = str(fila[0].value or "").strip()
        num = int(fila[1].value or 0)
        detalle = str(fila[7].value or "")
        cambio = False

        apertura = apertura_por_linea.get((fac, num))
        if apertura and apertura not in detalle:
            m = RE_ENCABEZADO.search(detalle)
            if m:
                detalle = detalle[: m.end()] + " " + apertura + " " + detalle[m.end() :].lstrip()
            else:
                detalle = apertura + " " + detalle
            cambio = True

        frases = [
            f for f in (por_linea.get((fac, num)), por_factura.get(fac)) if f and f not in detalle
        ]
        if frases:
            agregado = " ".join(frases)
            k = detalle.find(MARCA)
            detalle = (
                detalle[:k].rstrip() + " " + agregado + " " + detalle[k:]
                if k > 0
                else detalle + " " + agregado
            )
            cambio = True

        if cambio:
            fila[7].value = " ".join(detalle.split())
            tocadas += 1
    wb.save(ruta_excel)
    return tocadas


def evidencias_a_pdf(gi: str, facturas: list[str], carpeta_gi: Path) -> Path | None:
    """Junta los pantallazos de la corrida de HOY (evidencias_glosa/) en el
    PDF del paquete, copiando además los .png a la carpeta GI."""
    from PIL import Image

    origen = REPO / "evidencias_glosa"
    hoy = time.strftime("%Y%m%d")
    elegidos = []
    for fac in sorted(facturas):
        cands = sorted(origen.glob(f"HUS{factura_corta(fac)}_{hoy}_*.png"))
        if cands:
            destino = carpeta_gi / cands[-1].name
            shutil.copy2(cands[-1], destino)
            elegidos.append(destino)
    if not elegidos:
        return None
    pdf = carpeta_gi / f"{gi}_EVIDENCIAS.pdf"
    for i, png in enumerate(elegidos):
        with Image.open(png) as im:
            im.convert("RGB").save(pdf, "PDF", append=(i > 0))
    return pdf


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--excel", type=Path, required=True, help="Export DGH del lote (GLOSAS_X.xlsx)"
    )
    parser.add_argument(
        "--gi", type=str, default=None, help="Código GI (si se omite, se pide en pantalla)"
    )
    parser.add_argument(
        "--base", type=Path, default=Path(BASE_DEFECTO), help="Dónde crear la carpeta GI"
    )
    parser.add_argument("--rutas-soportes", nargs="*", default=RUTAS_SOPORTES_DEFECTO)
    parser.add_argument(
        "--piloto", type=str, default=None, help="Cargar SOLO esta factura (piloto)"
    )
    parser.add_argument(
        "--sin-cargue", action="store_true", help="Preparar todo sin subir al portal"
    )
    parser.add_argument(
        "--redactar-aceptacion",
        action="store_true",
        help=(
            "En las líneas con sobrecobro verificado, escribir la aceptación en el texto "
            "que se sube al portal. Sin esta bandera la aceptación queda solo como "
            "sugerencia en el Excel, para que la decida el auditor."
        ),
    )
    parser.add_argument(
        "--con-cabeza", action="store_true", help="Mostrar el navegador en el cargue"
    )
    parser.add_argument(
        "--tarifario-servicios",
        type=Path,
        default=Path(BASE_DEFECTO) / "TARIFARIO_440" / "6.2 PRECIOS DE REFERENCIA.xlsx",
        help="Anexo 6.2 del contrato 440 (CUPS/servicios)",
    )
    parser.add_argument(
        "--tarifario-medicamentos",
        type=Path,
        default=Path(BASE_DEFECTO) / "TARIFARIO_440" / "TARIFAS DEL CONTRATO.xlsx",
        help="Anexos de medicamentos y dispositivos del contrato 440 (CUM)",
    )
    args = parser.parse_args()
    # gen_lote.py corre desde otra carpeta: las rutas relativas que escriba el
    # auditor deben quedar absolutas antes de pasárselas.
    args.excel = args.excel.expanduser().resolve()
    args.base = args.base.expanduser().resolve()

    asegurar_dependencias()
    from extraer_factura_pdf import extraer_datos_factura
    from tarifario_440 import cargar_tarifario, tarifa_de

    gi = (args.gi or input("Código GI del paquete (ej. GI-33-5400-2026): ")).strip()
    if not gi:
        print("Sin código GI no hay paquete.")
        return 1
    carpeta_gi = args.base / gi
    soportes = carpeta_gi / "soportes"
    soportes.mkdir(parents=True, exist_ok=True)
    print(f"[1/7] Carpeta del paquete: {carpeta_gi}")

    # 2. Generar respuestas (gen_lote hereda la exclusión CL médica/mixta)
    excel_resp = carpeta_gi / f"respuestas_glosa_{gi}.xlsx"
    dump = carpeta_gi / f"dump_{gi}.json"
    print("[2/7] Generando respuestas con gen_lote.py (directriz CL incluida)...")
    r = subprocess.run(
        [sys.executable, str(AQUI / "gen_lote.py"), str(args.excel), str(excel_resp), str(dump)],
        cwd=AQUI,
    )
    if r.returncode != 0 or not excel_resp.is_file():
        print("gen_lote.py falló — revise el export.")
        return 1
    lineas = json.load(open(dump, encoding="utf-8"))
    facturas = sorted({d["factura"] for d in lineas})
    print(f"      Facturas a responder: {len(facturas)}")

    # 2b. Tarifario del contrato 440: la tarifa de cada glosa de tarifas
    # (el párrafo de evidencia la citará; sin pacto no se cita nada)
    indice = cargar_tarifario(args.tarifario_servicios, args.tarifario_medicamentos)
    tarifa_por_linea: dict[tuple[str, int], dict] = {}  # la que se cita en la respuesta
    tarifa_de_cotejo: dict[tuple[str, int], dict] = {}  # la de TODAS las líneas, para el cotejo
    if indice:
        for d in lineas:
            t = tarifa_de(indice, d.get("cups"))
            if not t:
                continue
            tarifa_de_cotejo[(d["factura"], d["num"])] = t
            if str(d.get("tipo", "")).startswith(("TARIFA", "LISTA_PRECIOS")):
                tarifa_por_linea[(d["factura"], d["num"])] = t
        print(
            f"      Tarifario cargado ({len(indice)} códigos): "
            f"{len(tarifa_de_cotejo)} glosas con tarifa pactada hallada "
            f"({len(tarifa_por_linea)} de ellas se citan en la respuesta)"
        )
    else:
        print(
            "      Sin tarifario a mano (revise --tarifario-servicios/--tarifario-medicamentos): "
            "las respuestas van sin la fila tarifaria citada."
        )

    # 3. Soportes de radicación desde la unidad Y: + lectura en cascada
    print("[3/7] Recorriendo las carpetas de radicación (una sola pasada)...")
    indice_soportes = indexar_soportes(args.rutas_soportes)
    print(f"      Soportes indexados: {len(indice_soportes)} facturas con PDF a la vista")
    extraidos, sin_pdf = {}, []
    for fac in facturas:
        pdf = buscar_en_indice(indice_soportes, fac)
        if pdf is None:
            sin_pdf.append(fac)
            continue
        destino = soportes / f"{fac}{pdf.suffix.lower()}"
        try:
            shutil.copy2(pdf, destino)
        except OSError as e:
            print(f"    {fac}: no pude copiar {pdf} ({e})")
            sin_pdf.append(fac)
            continue
        datos = extraer_datos_factura(destino)
        extraidos[fac] = datos
        print(f"    {fac}: soporte OK | lectura: {datos['metodo']}")
    json.dump(
        extraidos,
        open(carpeta_gi / f"lectura_facturas_{gi}.json", "w", encoding="utf-8"),
        ensure_ascii=False,
        indent=1,
    )

    # 4. COTEJO DE COBRO: ¿de verdad se está cobrando de más?
    # Va ANTES de redactar porque el texto de la respuesta depende del
    # escenario que dicte el cotejo. Primero se elige, de la línea leída en el
    # PDF, cuál es el valor unitario; con eso y la tarifa pactada se calculan
    # los factores que se repiten en el lote (esos son actualización de
    # tarifas del año, no sobrecobro) y recién entonces se dicta el veredicto.
    servicios_hallados: dict[tuple[str, int], tuple] = {}
    elegidos: dict[tuple[str, int], tuple[int | None, str]] = {}
    for d in lineas:
        clave = (d["factura"], d["num"])
        datos = extraidos.get(d["factura"])
        serv = (
            hallar_servicio_en_pdf(datos.get("servicios"), d.get("cups"), d.get("serv"))
            if datos
            else None
        )
        if serv:
            servicios_hallados[clave] = serv
        precio = (tarifa_de_cotejo.get(clave) or {}).get("precio")
        elegidos[clave] = cotejo_tarifa.elegir_valor_facturado(serv[2] if serv else [], precio)
    # Cuántas objeciones trae cada factura: el total del DGH solo sirve de
    # valor del servicio cuando la factura tiene una sola.
    objeciones_por_factura = Counter(d["factura"] for d in lineas)

    def _precio(clave):
        return (tarifa_de_cotejo.get(clave) or {}).get("precio")

    # Los factores del lote se calculan con lo mejor que haya de cada línea —el
    # PDF si se leyó, y si no el total del DGH—: son los que dejan ver que un
    # mismo porcentaje de más se repite y es la actualización del año.
    candidatos = []
    for d in lineas:
        clave = (d["factura"], d["num"])
        valor = elegidos[clave][0]
        if not valor and objeciones_por_factura[d["factura"]] == 1:
            valor = int(d.get("valor_factura") or 0) or None
        candidatos.append((valor, _precio(clave)))
    factores = cotejo_tarifa.factores_repetidos(candidatos)

    # Y ahora sí, con los factores conocidos, se decide si el total del DGH
    # puede usarse para cada línea que no tenga lectura del PDF.
    origen: dict[tuple[str, int], str] = {}
    for d in lineas:
        clave = (d["factura"], d["num"])
        if elegidos[clave][0]:
            origen[clave] = "PDF"
            continue
        valor, motivo = cotejo_tarifa.valor_desde_dgh(
            int(d.get("valor_factura") or 0),
            _precio(clave),
            objeciones_por_factura[d["factura"]],
            factores,
        )
        if valor:
            elegidos[clave] = (valor, "")
            origen[clave] = "DGH"
        elif motivo and not elegidos[clave][1]:
            elegidos[clave] = (None, motivo)
    cotejos: dict[tuple[str, int], dict] = {}
    for d in lineas:
        clave = (d["factura"], d["num"])
        valor, motivo = elegidos[clave]
        t = tarifa_de_cotejo.get(clave)
        cot = cotejo_tarifa.cotejar(
            valor, t, int(d.get("valor") or 0), d.get("cups", ""), factores, motivo
        )
        serv = servicios_hallados.get(clave)
        datos = extraidos.get(d["factura"]) or {}
        cot["fuente"] = " | ".join(
            x
            for x in (
                datos.get("archivo", ""),
                f"línea del PDF: {serv[0]}" if serv else "",
                (
                    f"valor total de la factura {d['factura']} según el sistema de cartera (DGH)"
                    if origen.get(clave) == "DGH"
                    else ""
                ),
                (t or {}).get("fuente", ""),
            )
            if x
        )
        cotejos[clave] = cot
    escritas = agregar_cotejo_excel(excel_resp, cotejos)
    json.dump(
        {f"{k[0]}#{k[1]}": v for k, v in cotejos.items()},
        open(carpeta_gi / f"cotejo_cobro_{gi}.json", "w", encoding="utf-8"),
        ensure_ascii=False,
        indent=1,
    )
    resumen = Counter(c["veredicto"] for c in cotejos.values())
    sugerido = sum(c["aceptar"] for c in cotejos.values())
    print(
        f"[4/7] Cotejo de cobro en {escritas} líneas: "
        f"{resumen[cotejo_tarifa.SOBRECOBRO]} con mayor valor VERIFICADO, "
        f"{resumen[cotejo_tarifa.POR_VIGENCIA]} por actualización de vigencia (revisar), "
        f"{resumen[cotejo_tarifa.A_TARIFA] + resumen[cotejo_tarifa.POR_DEBAJO]} cobradas a tarifa "
        f"o por debajo, {resumen[cotejo_tarifa.SIN_COTEJO]} sin cotejo posible."
    )
    if sugerido:
        print(
            f"      Valor sugerido a aceptar en total: {_plata(sugerido)} — el Excel del cargue "
            "SIGUE CON VALOR ACEPTADO EN 0: la aceptación la decide usted en la hoja "
            "'COTEJO DE COBRO'."
        )

    # 5. LA REDACCIÓN, por escenario: el argumento de cada respuesta lo dicta
    # el veredicto del cotejo (a tarifa / vigencia 2026 / sobrecobro real), y
    # usa exactamente las mismas cifras verificadas. Fuera de esos tres casos
    # manda la redacción prudente, que no proclama cifras sin cotejo.
    aperturas: dict[tuple[str, int], str] = {}
    por_escenario = Counter()
    con_pdf = 0
    for d in lineas:
        clave = (d["factura"], d["num"])
        datos = extraidos.get(d["factura"])
        servicio_pdf = servicios_hallados.get(clave)
        cot = cotejos[clave]
        parrafo = parrafo_por_escenario(
            cot, datos, servicio_pdf, d.get("cups", ""), d.get("serv", ""), d["factura"]
        )
        if (
            parrafo
            and cot["veredicto"] == cotejo_tarifa.SOBRECOBRO
            and not args.redactar_aceptacion
        ):
            # El texto acepta una glosa: no se sube sin que el auditor lo pida.
            # Queda en la columna RESPUESTA SUGERIDA y en la hoja de trabajo.
            parrafo = None
        if parrafo:
            por_escenario[cot["veredicto"]] += 1
        else:
            parrafo = parrafo_evidencia(
                datos,
                servicio_pdf,
                tarifa_por_linea.get(clave),
                int(d.get("valor") or 0),
                d.get("cups", ""),
            )
        if parrafo:
            aperturas[clave] = parrafo
            con_pdf += 1 if datos else 0
    if aperturas:
        n = enriquecer_excel(excel_resp, apertura_por_linea=aperturas)
        print(
            f"[5/7] Respuestas redactadas: {n} líneas ({con_pdf} citan el PDF leído). "
            f"Por escenario: {por_escenario[cotejo_tarifa.A_TARIFA]} a tarifa exacta, "
            f"{por_escenario[cotejo_tarifa.POR_VIGENCIA]} defendidas por vigencia 2026, "
            f"{por_escenario[cotejo_tarifa.SOBRECOBRO]} con aceptación redactada."
        )
        if resumen[cotejo_tarifa.SOBRECOBRO] and not args.redactar_aceptacion:
            print(
                f"      {resumen[cotejo_tarifa.SOBRECOBRO]} línea(s) con sobrecobro NO llevan la "
                "aceptación en el texto que se sube: está en la columna RESPUESTA SUGERIDA. "
                "Para que el texto la lleve, corra con --redactar-aceptacion."
            )
    else:
        print("[5/7] Sin datos de PDF ni tarifas: las respuestas van con el texto del motor.")
    if sin_pdf:
        print(
            f"      SIN SOPORTE EN Y: ({len(sin_pdf)}): {', '.join(sin_pdf[:12])}"
            + ("..." if len(sin_pdf) > 12 else "")
        )

    # 6. Cargue al portal
    if args.sin_cargue:
        print("[6/7] --sin-cargue: el Excel queda listo; corra el robot cuando quiera.")
    else:
        reporte = carpeta_gi / f"reporte_{gi}.csv"
        cmd = [
            sys.executable,
            str(REPO / "tools" / "responder_glosas_simed.py"),
            "--excel",
            str(excel_resp),
            "--sin-soportes",
            "--reporte",
            str(reporte),
        ]
        cmd += ["--solo", args.piloto] if args.piloto else ["--todas"]
        if args.con_cabeza:
            cmd.append("--con-cabeza")
        print(
            f"[6/7] Corriendo el robot del portal ({'piloto ' + args.piloto if args.piloto else 'todas'})..."
        )
        subprocess.run(cmd, cwd=REPO)

        # 7. Evidencias del paquete
        pdf_ev = evidencias_a_pdf(gi, [args.piloto] if args.piloto else facturas, carpeta_gi)
        print(
            f"[7/7] Evidencias del paquete: {pdf_ev if pdf_ev else 'sin pantallazos de hoy (¿corrió el robot?)'}"
        )

    print(f"\nPaquete {gi} listo en {carpeta_gi}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
