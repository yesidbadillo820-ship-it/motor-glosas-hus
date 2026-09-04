"""Genera el Excel de respuestas para UN lote (un export DGH), todas las facturas
Dispensario, usando las plantillas endurecidas de glosa_motor.py.

Uso: py gen_lote.py fuente.xlsx salida.xlsx [dump.json]

La salida es el formato exacto que consume tools/responder_glosas_simed.py:
hoja "Respuestas Glosa" con columnas Factura | # Objeción | Cód. | Servicio |
Valor Objetado | Valor Aceptado (0) | Cod Respuesta (RE9901) | Detalle Respuesta.
La numeración 1..N por factura sigue el orden del export, que es el orden de la
grilla del portal (verificado con el lote del 09-07-2026).

Directriz 03-09-2026: las objeciones con causal CL de glosas Médica o Mixta NO
se responden con el bot — van a la hoja "PARA GESTION MEDICA" y las trabaja el
equipo médico a mano (si aceptan la glosa, la nota crédito debe poder cruzar;
una respuesta genérica ya cargada lo impide). Esas objeciones conservan su
número en la grilla, así el robot responde las demás sin correrse de fila.
"""

import json
import sys
from collections import Counter, defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from glosa_motor import disp, money, norm, redactar

src, out = sys.argv[1], sys.argv[2]
jsonout = sys.argv[3] if len(sys.argv) > 3 else None

wb = load_workbook(src, data_only=True, read_only=True)
det = None
for ws in wb.worksheets:
    hdr = [ws.cell(1, cc).value for cc in range(1, ws.max_column + 1)]
    if "ListadoConceptos.ConceptoObjecion.Codigo" in hdr:
        det, H = ws, hdr
        break
if det is None:
    sys.exit("No encontré la hoja de detalle (ListadoConceptos.*)")
c = {
    k: H.index(v)
    for k, v in dict(
        ent="FacturaCartera.PlanBeneficio.Contrato.Entidad.NombreEntidad",
        fac="FacturaCartera.Factura",
        gc="ListadoConceptos.ConceptoObjecion.Codigo",
        sc="ListadoConceptos.ServicioProductoFactura.Codigo",
        sd="ListadoConceptos.ServicioProductoFactura.Descripcion",
        vl="ListadoConceptos.ValorObjecion",
        vh="ValorObjecion",
        ob="ListadoConceptos.Observaciones",
    ).items()
}
i_tt = H.index("TipoObjecionTramite") if "TipoObjecionTramite" in H else None
# El valor total de la factura no viene en todos los export; sin él, el
# cotejo simplemente se queda sin esa segunda fuente.
i_vf = H.index("FacturaCartera.Valor") if "FacturaCartera.Valor" in H else None


def es_cl_para_medicos(tipo_tramite: str, code: str) -> bool:
    """Directriz 03-09-2026: las glosas Médica o Mixta con causal CL no se
    responden con el bot — quedan a gestión manual del equipo médico, para que
    si ellos aceptan la glosa, la nota crédito pueda cruzar en el sistema (una
    respuesta genérica ya cargada impide ese cruce)."""
    t = norm(tipo_tramite).upper()
    return (t.startswith("MEDIC") or t == "MIXTA") and norm(code).upper().startswith("CL")


seq = defaultdict(int)
out_rows, tally, dump, excluidas = [], Counter(), [], []
for r in det.iter_rows(min_row=2, values_only=True):
    if not r[c["fac"]] or not disp(norm(r[c["ent"]])):
        continue
    fac = str(r[c["fac"]]).strip()
    code = norm(r[c["gc"]])
    cups = norm(r[c["sc"]])
    serv = norm(r[c["sd"]])
    valor = int(r[c["vl"]] or r[c["vh"]] or 0)
    # El valor TOTAL de la factura, como lo tiene el sistema de cartera del
    # hospital. Sirve de segunda fuente para el cotejo cuando el PDF no está
    # a la mano: si la factura es de un solo servicio, ese total ES el valor
    # facturado del servicio glosado.
    valor_factura = int(r[i_vf] or 0) if i_vf is not None else 0
    obs = r[c["ob"]]
    # La numeración es POSICIONAL (el orden de la grilla del portal): las
    # excluidas también cuentan, para que el robot responda cada objeción
    # restante con su número real y jamás toque la CL omitida.
    seq[fac] += 1
    if i_tt is not None and es_cl_para_medicos(r[i_tt], code):
        excluidas.append([fac, seq[fac], code, f"{cups} - {serv}", valor, norm(obs)])
        continue
    det_txt, t = redactar(fac, code, cups, serv, valor, obs)
    tally[t] += 1
    out_rows.append([fac, seq[fac], code, f"{cups} - {serv}", valor, 0, "RE9901", det_txt])
    dump.append(
        dict(
            factura=fac,
            num=seq[fac],
            code=code,
            cups=cups,
            serv=serv,
            valor=valor,
            valor_factura=valor_factura,
            obs=norm(obs),
            tipo=t,
            detalle=det_txt,
        )
    )

owb = Workbook()
osh = owb.active
osh.title = "Respuestas Glosa"
cols = [
    "Factura",
    "# Objeción",
    "Cód.",
    "Servicio",
    "Valor Objetado",
    "Valor Aceptado",
    "Cod Respuesta",
    "Detalle Respuesta",
]
osh.append(cols)
fill = PatternFill("solid", fgColor="1F4E78")
for i, _ in enumerate(cols, 1):
    x = osh.cell(row=1, column=i)
    x.font = Font(bold=True, color="FFFFFF")
    x.fill = fill
    x.alignment = Alignment(vertical="center")
for row in out_rows:
    osh.append(row)
for i, w in enumerate([16, 10, 9, 44, 14, 14, 13, 130], 1):
    osh.column_dimensions[get_column_letter(i)].width = w
for rr in osh.iter_rows(min_row=2):
    rr[7].alignment = Alignment(wrap_text=True, vertical="top")
osh.freeze_panes = "A2"

# Las CL médicas/mixtas quedan aisladas en su propia hoja: es el paquete de
# trabajo del equipo médico, y NO lo lee el robot (solo lee "Respuestas Glosa").
if excluidas:
    med = owb.create_sheet("PARA GESTION MEDICA")
    med.append(
        [
            "Factura",
            "# Objeción (grilla portal)",
            "Cód.",
            "Servicio",
            "Valor Objetado",
            "Observación de la EPS",
        ]
    )
    for i, _ in enumerate(med[1], 1):
        x = med.cell(row=1, column=i)
        x.font = Font(bold=True, color="FFFFFF")
        x.fill = PatternFill("solid", fgColor="9C0006")
    for row in excluidas:
        med.append(row)
    for i, w in enumerate([16, 12, 9, 44, 14, 120], 1):
        med.column_dimensions[get_column_letter(i)].width = w
    for rr in med.iter_rows(min_row=2):
        rr[5].alignment = Alignment(wrap_text=True, vertical="top")
    med.freeze_panes = "A2"
owb.save(out)
if jsonout:
    json.dump(dump, open(jsonout, "w", encoding="utf-8"), ensure_ascii=False, indent=1)

print("OK ->", out)
if excluidas:
    print(
        f"EXCLUIDAS PARA GESTION MEDICA (CL medica/mixta): {len(excluidas)} objeciones "
        f"de {len({e[0] for e in excluidas})} facturas por {money(sum(e[4] for e in excluidas))} "
        f"— ver hoja 'PARA GESTION MEDICA'; el bot NO las responde."
    )
print(
    f"Ítems: {len(out_rows)} | Facturas: {len({r[0] for r in out_rows})} "
    f"| Valor obj: {money(sum(r[4] for r in out_rows))}"
)
print("Por tipo:", dict(tally))
prob = []
for r in out_rows:
    d = r[7]
    if (
        d != d.upper()
        or "\n" in d
        or not d.startswith("ESE HUS NO ACEPTA LA GLOSA APLICADA A LA FACTURA")
        or r[2].upper() not in d
        or "CARTERA@HUS.GOV.CO" not in d
    ):
        prob.append((r[0], r[1]))
print("Problemas de formato:", prob if prob else "NINGUNO")
if out_rows:
    print("Largo detalle:", min(len(r[7]) for r in out_rows), "-", max(len(r[7]) for r in out_rows))
