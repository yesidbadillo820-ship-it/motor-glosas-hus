"""Completa RESPUESTA/NO/FECHA DE TRAMITE Y/O ACTA en el consolidado de glosas aceptadas.

Cruza el archivo mensual de glosas aceptadas contra la hoja GENERAL de la
CIRCULARIZACION DE GLOSAS y llena, en las filas que les falte la respuesta:

  - RESPUESTA TRAMITE GLOSA Y/O ACTA  (conceptos de conciliacion unificados)
  - NO DE TRAMITE Y/O ACTA            (numero de acta; solo si viene vacio)
  - FECHA DE TRAMITE Y/O ACTA         (fecha de firma; solo si viene vacio)

El formato del consolidado cambia de un mes a otro (nombre de la hoja, fila de
encabezados y posicion de las columnas), asi que TODO se detecta por el texto
del encabezado, nunca por posicion fija. Los meses vistos hasta hoy:

  jun-2026  hoja "BD"              encabezados fila 4   respuesta en col W
  jul-2026  hoja "BD VLR ACEPTADO" encabezados fila 3   respuesta en col X

Reglas de resolucion, por fila a completar:
  1. Se busca la factura en GENERAL. Si la fila ya trae numero de acta, se usa
     ese; si no, el que se pueda leer en la observacion de la nota credito.
  2. Se escribe UN PARRAFO POR CADA GLOSA ACEPTADA del acta (valor aceptado > 0),
     separados por una linea en blanco. NO se agrupan los que repiten texto:
     dos renglones iguales son dos glosas distintas y agruparlos escondería la
     plata de uno de ellos. Las glosas de valor 0 (las que la entidad levanto,
     las ratificadas y las que la ESE no acepto) NO entran: no pintan en una
     nota credito y su texto dice lo contrario de lo que la nota documenta.
  3. Si el acta no registra nada aceptado para esa factura, o la factura no esta
     en la circularizacion (p.ej. actas de vigencia anterior), la respuesta se
     toma del texto de la propia observacion de la nota credito (la parte
     posterior a "DONDE"/"EN LA CUAL").
  4. Numero y fecha solo se escriben si la celda venia vacia: nunca se pisa un
     dato que ya puso el auditor.
  5. Cuando lo aceptado en el acta no coincide con lo que acredito la nota, se
     escribe el aviso en una columna nueva al final ("NOVEDAD ACEPTADO VS NOTA
     CREDITO") y tambien queda en el reporte CSV.

Uso:
    python tools/completar_tramite_glosas_aceptadas.py ACEPTADAS.xlsx CIRCULARIZACION.xlsx SALIDA.xlsx [REPORTE.csv]
"""

import csv
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import datetime

import openpyxl

# Encabezados que identifican cada columna del consolidado. Se compara por
# "contiene", sobre el texto normalizado (mayusculas y sin tildes), asi que
# basta con un fragmento distintivo. El orden importa: gana el primero que
# aparezca en la fila de encabezados.
COLUMNAS_BD = {
    "factura": ["FACTURA"],
    "obs": ["OBSERVC NOTA CREDITO", "OBSERVC NTA", "OBSERVACION NOTA", "OBSERVC"],
    "valor": ["VALOR ACEPTADO", "VLR NOTA", "VALOR NOTA"],
    "respuesta": ["RESPUESTA TRAMITE"],
    "num": ["NO DE TRAMITE", "N DE TRAMITE", "NRO DE TRAMITE"],
    "fecha": ["FECHA DE TRAMITE"],
    "tipo_tramite": ["TRAMITE Y/O ACTA"],
}
# Columnas cuyo nombre esta contenido en el de otra ("FACTURA" dentro de "VALOR
# FACTURA"; "TRAMITE Y/O ACTA" dentro de "NO DE TRAMITE Y/O ACTA"): para estas
# se exige que el encabezado sea exactamente el alias, no que lo contenga.
EXACTAS = {"factura", "tipo_tramite"}

# Columnas (0-based) de la hoja GENERAL de la circularizacion (estable)
GEN_FACTURA, GEN_VAL_ACEPTADO, GEN_ACTA, GEN_FECHA, GEN_CONCEPTO = 1, 5, 11, 12, 13

RE_ACTA = re.compile(r"ACTA\s*(?:DE\s+CONCILIACI\w+\s*)?(?:N[O°ºRo\.\s]{0,4})?\s*(\d{3,4})", re.I)
RE_SOLO_NUM = re.compile(r"(\d{3,4})")
RE_FECHA_DMA = re.compile(r"(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})")
MESES = {
    "ENERO": 1,
    "FEBRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
    "JULIO": 7,
    "AGOSTO": 8,
    "SEPTIEMBRE": 9,
    "OCTUBRE": 10,
    "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}
RE_FECHA_TEXTO = re.compile(r"(\d{1,2})\s+DE\s+(" + "|".join(MESES) + r")\s+DE\s+(\d{4})", re.I)


def limpiar(texto):
    if texto is None:
        return ""
    return str(texto).replace("_x000D_", "").replace("\r", "").strip()


def normalizar(v):
    """Mayusculas y sin tildes, para comparar encabezados entre meses."""
    s = unicodedata.normalize("NFD", limpiar(v).upper())
    return " ".join("".join(ch for ch in s if not unicodedata.combining(ch)).split())


def a_numero(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return round(float(v))
    try:
        return round(float(str(v).replace(",", "").strip()))
    except ValueError:
        return 0


def a_fecha(v):
    if isinstance(v, datetime):
        return v
    s = limpiar(v)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def pesos(n):
    """Formato colombiano: 1234567 -> $1.234.567"""
    return "$" + f"{round(n):,}".replace(",", ".")


def numero_de_acta(v):
    """Lee el numero de una celda de acta: 862, '862', 'Acta No. 828'."""
    s = limpiar(v)
    if not s:
        return None
    m = RE_ACTA.search(s) or RE_SOLO_NUM.search(s)
    return int(m.group(1)) if m else None


def parsear_obs(obs):
    """Extrae (acta, fecha, respuesta) de la observacion de la nota credito."""
    texto = limpiar(obs)
    m = RE_ACTA.search(texto)
    acta = int(m.group(1)) if m else None
    fecha = None
    resto = texto[m.end() :] if m else texto
    mf = RE_FECHA_DMA.search(resto[:60])
    if mf:
        d, mth, y = (int(x) for x in mf.groups())
        try:
            fecha = datetime(y, mth, d)
        except ValueError:
            fecha = None
    else:
        mt = RE_FECHA_TEXTO.search(resto[:80])
        if mt:
            try:
                fecha = datetime(int(mt.group(3)), MESES[mt.group(2).upper()], int(mt.group(1)))
            except ValueError:
                fecha = None

    respuesta = ""
    ancla = re.search(r"DONDE\s*:?|EN LA CUAL\s*[;:]?", resto, re.I)
    if ancla:
        respuesta = resto[ancla.end() :]
    else:
        inicio = re.search(r"EN CONCILI\w+|ESE HUS", resto, re.I)
        if inicio:
            respuesta = resto[inicio.start() :]
    respuesta = respuesta.strip(" :;,.\n")
    # colapsa el texto cuando la observacion trae la misma respuesta duplicada
    mitad = len(respuesta) // 2
    if mitad > 40:
        a, b = respuesta[:mitad].strip(" .\n"), respuesta[mitad:].strip(" .\n")
        if a and a == b:
            respuesta = a
    return acta, fecha, respuesta.strip()


def buscar_subconjunto(vals, objetivo):
    """Indices cuyo valor suma exactamente `objetivo`; None si no existe.

    Solo se devuelven glosas con valor aceptado > 0. Las de valor 0 son las
    que la entidad levanto o que la ESE no acepto: sumar cero no altera el
    total, asi que si se incluyeran la comprobacion de "cuadra exacto" las
    dejaria pasar y el texto de una nota credito terminaria diciendo "ESE HUS
    NO ACEPTA GLOSA", que es lo contrario de lo que la nota documenta.
    """
    no_cero = [(i, v) for i, v in enumerate(vals) if v > 0]
    if objetivo == sum(v for _, v in no_cero):
        return [i for i, _ in no_cero]
    for i, v in no_cero:
        if v == objetivo:
            return [i]
    alcanzables = {0: []}
    for i, v in no_cero:
        nuevos = {}
        for s, idxs in alcanzables.items():
            ns = s + v
            if ns <= objetivo and ns not in alcanzables and ns not in nuevos:
                nuevos[ns] = idxs + [i]
        alcanzables.update(nuevos)
        if len(alcanzables) > 500000:
            break
    return alcanzables.get(objetivo)


def cargar_general(ruta):
    wb = openpyxl.load_workbook(ruta, read_only=True)
    ws = wb["GENERAL"]
    por_factura = defaultdict(list)
    for fila in ws.iter_rows(min_row=3, values_only=True):
        factura = limpiar(fila[GEN_FACTURA])
        if not factura.upper().startswith("HUS"):
            continue
        por_factura[factura.upper()].append(
            {
                "val": a_numero(fila[GEN_VAL_ACEPTADO]),
                "acta": numero_de_acta(fila[GEN_ACTA]),
                "fecha": a_fecha(fila[GEN_FECHA]),
                "concepto": limpiar(fila[GEN_CONCEPTO]),
            }
        )
    wb.close()
    return por_factura


def localizar_hoja_y_encabezados(wb, hoja_pedida=None):
    """Devuelve (worksheet, fila_encabezados, {campo: indice_0based}).

    Busca en cada hoja la fila (entre las 15 primeras) que contenga el
    encabezado de RESPUESTA TRAMITE, que es la columna que da sentido al
    archivo. Sobre esa fila mapea el resto de columnas por nombre.
    """
    hojas = [wb[hoja_pedida]] if hoja_pedida else wb.worksheets
    for ws in hojas:
        for nfila, fila in enumerate(ws.iter_rows(min_row=1, max_row=15), 1):
            encabezados = [normalizar(c.value) for c in fila]
            if not any("RESPUESTA TRAMITE" in h for h in encabezados):
                continue
            mapa = {}
            for campo, alias in COLUMNAS_BD.items():
                for idx, h in enumerate(encabezados):
                    if not h:
                        continue
                    hit = (
                        any(h == a for a in alias)
                        if campo in EXACTAS
                        else any(a in h for a in alias)
                    )
                    if hit:
                        mapa[campo] = idx
                        break
            faltan = {"factura", "obs", "valor", "respuesta", "num", "fecha"} - set(mapa)
            if faltan:
                raise SystemExit(
                    f"En la hoja '{ws.title}' fila {nfila} no se hallaron las columnas: "
                    f"{sorted(faltan)}. Encabezados leidos: "
                    f"{[h for h in encabezados if h][:30]}"
                )
            return ws, nfila, mapa
    raise SystemExit(
        "No se encontro ninguna hoja con la columna 'RESPUESTA TRAMITE GLOSA Y/O ACTA'. "
        f"Hojas del archivo: {wb.sheetnames}"
    )


def resolver_fila(factura, valor, obs, acta_fila, general):
    """Devuelve (respuesta, acta, fecha, notas_de_revision).

    `acta_fila` es el numero de acta que ya trae la fila (o None). Cuando
    existe manda sobre el que se lea en la observacion: lo puso el auditor.
    """
    acta_obs, fecha_obs, resp_obs = parsear_obs(obs)
    acta_ref = acta_fila or acta_obs
    notas = []
    if acta_fila and acta_obs and acta_fila != acta_obs:
        notas.append(f"la fila dice acta {acta_fila} y la observacion dice acta {acta_obs}")
    candidatas = general.get(factura.upper(), [])

    grupo = []
    if candidatas:
        por_acta = defaultdict(list)
        for c in candidatas:
            por_acta[c["acta"]].append(c)
        if acta_ref in por_acta:
            grupo = por_acta[acta_ref]
        else:
            con_cuadre = [
                g for g in por_acta.values() if buscar_subconjunto([c["val"] for c in g], valor)
            ]
            candidatos = con_cuadre or list(por_acta.values())
            grupo = max(candidatos, key=lambda g: g[0]["fecha"] or datetime.min)
            if acta_ref and grupo:
                notas.append(
                    f"la NC cita acta {acta_ref}; circularizacion la registra en acta {grupo[0]['acta']}"
                )

    if grupo:
        acta, fecha = grupo[0]["acta"], grupo[0]["fecha"]
        # Solo lo ACEPTADO. Las glosas de valor 0 son las que la entidad levanto,
        # las que se ratificaron o las que la ESE no acepto: no pintan en una nota
        # credito. Se escribe un parrafo por cada renglon aceptado, SIN agrupar los
        # que repiten texto, porque dos renglones iguales son dos glosas y agrupar
        # los esconderia la plata de uno de ellos.
        aceptadas = [c for c in grupo if c["val"] > 0]
        if aceptadas:
            conceptos = [c["concepto"] for c in aceptadas if c["concepto"]]
            sin_aceptar = [c for c in aceptadas if "ACEPTA" not in c["concepto"].upper()]
            if sin_aceptar:
                notas.append(
                    f"{len(sin_aceptar)} renglon(es) tienen valor aceptado pero su texto "
                    f"no dice que la ESE acepte: revisar redaccion en la circularizacion"
                )
            total = sum(c["val"] for c in aceptadas)
            novedad = ""
            if total != valor:
                dif = total - valor
                sobra = "de mas en el acta" if dif > 0 else "de mas en la nota"
                cuadra = buscar_subconjunto([c["val"] for c in aceptadas], valor)
                pista = (
                    " (la nota corresponde a algunos renglones del acta, no a todos)"
                    if cuadra
                    else " (ningun grupo de renglones del acta da el valor de la nota)"
                )
                novedad = (
                    f"El acta acepta {pesos(total)} y la nota credito acredita "
                    f"{pesos(valor)}: diferencia {pesos(abs(dif))} {sobra}.{pista}"
                )
                notas.append(novedad)
            return "\n\n".join(conceptos), acta, fecha, novedad, notas
        novedad = (
            f"El acta no registra ningun valor aceptado para esta factura y la nota "
            f"credito acredita {pesos(valor)}. Respuesta tomada de la observacion de la nota."
        )
        notas.append(novedad)
        return resp_obs, acta, fecha, novedad, notas

    novedad = (
        "La factura no aparece en la circularizacion bajo esa acta. "
        "Respuesta y datos tomados de la observacion de la nota credito."
    )
    notas.append(novedad)
    return resp_obs, acta_ref, fecha_obs, novedad, notas


def main():
    if len(sys.argv) < 4:
        print(__doc__)
        sys.exit(1)
    ruta_bd, ruta_circ, ruta_salida = sys.argv[1:4]
    ruta_reporte = sys.argv[4] if len(sys.argv) > 4 else None

    general = cargar_general(ruta_circ)
    wb = openpyxl.load_workbook(ruta_bd)
    ws, fila_enc, col = localizar_hoja_y_encabezados(wb)
    letra = openpyxl.utils.get_column_letter
    print(
        f"Hoja '{ws.title}', encabezados en la fila {fila_enc}. Columnas detectadas: "
        + ", ".join(f"{c}={letra(i + 1)}" for c, i in sorted(col.items(), key=lambda x: x[1]))
    )

    # Columna nueva al final para avisar diferencias entre lo aceptado en el acta
    # y lo que realmente acredito la nota credito.
    col_novedad = max(i for c, i in col.items()) + 1
    while ws.cell(row=fila_enc, column=col_novedad + 1).value not in (None, ""):
        col_novedad += 1
    celda_enc = ws.cell(row=fila_enc, column=col_novedad + 1)
    if not limpiar(celda_enc.value):
        celda_enc.value = "NOVEDAD ACEPTADO VS NOTA CREDITO"
    print(f"Novedades se escriben en la columna {letra(col_novedad + 1)}")

    reporte, llenas, solo_respuesta = [], 0, 0
    for fila in ws.iter_rows(min_row=fila_enc + 1):
        factura = limpiar(fila[col["factura"]].value)
        if not factura:
            continue
        # Se trabaja la fila cuando le falta la respuesta, que es la columna
        # que da sentido al registro. Numero y fecha pueden venir ya puestos.
        if limpiar(fila[col["respuesta"]].value):
            continue
        valor = a_numero(fila[col["valor"]].value)
        acta_fila = numero_de_acta(fila[col["num"]].value)
        respuesta, acta, fecha, novedad, notas = resolver_fila(
            factura, valor, fila[col["obs"]].value, acta_fila, general
        )
        if respuesta:
            celda = fila[col["respuesta"]]
            celda.value = respuesta
            celda.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
        if novedad:
            ws.cell(row=fila[0].row, column=col_novedad + 1).value = novedad
        # Nunca se pisa un numero o una fecha que ya estaban escritos.
        if acta and fila[col["num"]].value is None:
            fila[col["num"]].value = acta
        if fecha and fila[col["fecha"]].value is None:
            fila[col["fecha"]].value = fecha
            fila[col["fecha"]].number_format = "dd/mm/yyyy"
        llenas += 1
        if acta_fila and fila[col["fecha"]].value is not None:
            solo_respuesta += 1
        if not respuesta:
            notas.append("SIN RESPUESTA: revisar manualmente")
        for n in notas:
            reporte.append(
                {
                    "fila": fila[0].row,
                    "factura": factura,
                    "valor_nota": valor,
                    "acta": acta,
                    "nota": n,
                }
            )

    wb.save(ruta_salida)
    print(f"Filas diligenciadas: {llenas} (de ellas {solo_respuesta} ya traian numero y fecha)")
    print(f"Archivo generado: {ruta_salida}")
    if ruta_reporte:
        with open(ruta_reporte, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=["fila", "factura", "valor_nota", "acta", "nota"])
            w.writeheader()
            w.writerows(reporte)
        print(f"Reporte de revision ({len(reporte)} notas): {ruta_reporte}")
    else:
        for r in reporte:
            print(f"  fila {r['fila']} {r['factura']}: {r['nota']}")


if __name__ == "__main__":
    main()
