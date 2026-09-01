#!/usr/bin/env python3
"""siifa_revisar_antes_de_cargar.py — La revisión previa al cargue.

POR QUÉ EXISTE. Un cargue de 60.000 respuestas no se puede corregir sobre la
marcha: lo que entra mal, entra mal para siempre. Esta revisión mira el
archivo ANTES de tocar la plataforma y separa lo que está listo de lo que hay
que arreglar, para que el cargue salga limpio de una.

QUÉ REVISA, y por qué cada cosa:

  1. QUE EL ARCHIVO SEA DE ESA IPS. Se compara el NIT del emisor contra el de
     la IPS elegida. Cargar el archivo de una clínica con las credenciales de
     otra no se puede deshacer.
  2. LO QUE YA ESTÁ RESPONDIDO EN SIIFA. Volver a cargarlo PISA la respuesta
     que ya está registrada —y con ella su fecha, que es la que prueba que se
     respondió a tiempo—. Se deja fuera salvo que se pida lo contrario.
  3. OBSERVACIONES DE MÁS DE 1.500 CARACTERES. SIIFA las rechaza con un
     HTTP 500 que no dice la causa. Se recortan en el último punto y quedan
     marcadas para revisarlas.
  4. RESPUESTAS SIN TEXTO. Un código sin sustento no defiende nada ante la
     EPS. Salen aparte.
  5. QUE EL CÓDIGO CORRESPONDA AL TIPO. Los códigos de devolución (RE95xx,
     RE96xx, RE97xx) y los de glosa (RE99xx) no son intercambiables.
  6. QUE LA FECHA NO SEA ANTERIOR A LA FORMULACIÓN. SIIFA rechaza una
     respuesta fechada antes de que la EPS formulara la glosa.

USO:
    py siifa_revisar_antes_de_cargar.py --ips SOCORRO ^
        --archivo "D:\\...\\SOCORRO\\respuestas.xlsx" ^
        --salida  "D:\\...\\SOCORRO\\LISTO_PARA_CARGAR.xlsx"

Deja dos archivos: el de cargue (sólo lo que puede subir sin problemas) y
`..._REVISAR.xlsx` con lo que quedó fuera y por qué.

INSTALACIÓN (una vez):
    py -m pip install openpyxl
"""

from __future__ import annotations

import argparse
import sys
from collections import Counter
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from _dinero import a_entero  # noqa: E402
import siifa_perfiles as perfiles  # noqa: E402
from siifa_novedades import leer_informe, valor_total  # noqa: E402

LIMITE_OBSERVACION = 1500

# Los códigos de respuesta a una DEVOLUCIÓN (grupo RESPUESTA_DEV_PTS_PSS,
# confirmado contra la API el 05-08-2026). Todo lo demás es de glosa.
CODIGOS_DEVOLUCION = {"RE9501", "RE9601", "RE9701"}

COLS_SALIDA = [
    "ID_SEGUIMIENTO_FACTURA_GLOSA",
    "ID_SEGUIMIENTO_FACTURA_DEVOLUCION",
    "NUMERO_FACTURA",
    "TIPO",
    "CODIGO_GLOSA",
    "VALOR_GLOSA",
    "RAZON_SOCIAL_EPS",
    "MOTIVO",
    "CODIGO_RESPUESTA",
    "FECHA_RESPUESTA",
    "OBSERVACION_RESPUESTA",
]


def _v(fila: dict, col: str) -> str:
    """El valor de una columna, tratando como vacío lo que el Excel trae como
    texto 'None' o 'null' — que es como llega de la API y no es un dato."""
    valor = fila.get(col)
    if valor in (None, ""):
        return ""
    texto = str(valor).strip()
    return "" if texto.lower() in ("none", "null") else texto


def _fecha(valor: object) -> date | None:
    texto = str(valor or "").strip()[:10]
    try:
        return date.fromisoformat(texto)
    except ValueError:
        return None


def recortar(texto: str, limite: int = LIMITE_OBSERVACION) -> str:
    """Corta en el último punto o espacio, para no partir una palabra."""
    if len(texto) <= limite:
        return texto
    corte = texto[: limite - 3]
    for sep in (". ", " "):
        pos = corte.rfind(sep)
        if pos > limite * 0.6:
            return corte[: pos + (1 if sep == ". " else 0)].rstrip() + "..."
    return corte.rstrip() + "..."


def es_devolucion(fila: dict) -> bool:
    return _v(fila, "tipo_seguimiento").upper().startswith("DEVOL")


def revisar(filas: list[dict], ips, incluir_ya_respondidas: bool = False) -> tuple[list, list]:
    """Separa lo que se puede cargar de lo que hay que mirar primero."""
    listas, aparte = [], []
    for f in filas:
        codigo = _v(f, "codigo_respuesta").upper()
        obs = _v(f, "observacion_respuesta")
        dev = es_devolucion(f)
        motivos = []

        if not codigo:
            motivos.append("Sin código de respuesta")
        if not obs:
            motivos.append("Sin texto de respuesta: un código sin sustento no defiende nada")
        if codigo and dev and codigo not in CODIGOS_DEVOLUCION:
            motivos.append(f"Es DEVOLUCIÓN y {codigo} es un código de glosa")
        if codigo and not dev and codigo in CODIGOS_DEVOLUCION:
            motivos.append(f"Es GLOSA y {codigo} es un código de devolución")

        f_resp, f_form = _fecha(_v(f, "fecha_respuesta")), _fecha(_v(f, "fecha_formulacion"))
        if f_resp and f_form and f_resp < f_form:
            motivos.append(f"La respuesta ({f_resp}) es anterior a la glosa ({f_form})")

        ya = _v(f, "tiene_respuesta").upper() == "SI"
        if ya and not incluir_ya_respondidas:
            motivos.append("Ya está respondida en SIIFA: volver a cargarla pisaría esa respuesta")

        recortada = len(obs) > LIMITE_OBSERVACION
        salida = {
            "ID_SEGUIMIENTO_FACTURA_GLOSA": _v(f, "id_seguimiento_factura_glosa"),
            "ID_SEGUIMIENTO_FACTURA_DEVOLUCION": _v(f, "id_seguimiento_factura_devolucion"),
            "NUMERO_FACTURA": _v(f, "numero_factura"),
            "TIPO": _v(f, "tipo_seguimiento"),
            "CODIGO_GLOSA": _v(f, "codigo_glosa"),
            "VALOR_GLOSA": a_entero(f.get("valor_glosa")),
            "RAZON_SOCIAL_EPS": _v(f, "razon_social_eps"),
            "CODIGO_RESPUESTA": codigo,
            "FECHA_RESPUESTA": _v(f, "fecha_respuesta")[:10],
            "OBSERVACION_RESPUESTA": recortar(obs) if recortada else obs,
            # Se guarda el original para poder juzgar el recorte sin abrir el
            # archivo de donde salió.
            "_tipo_seguimiento": _v(f, "tipo_seguimiento"),
            "_valor": a_entero(f.get("valor_glosa")),
            "_numero_factura": _v(f, "numero_factura"),
        }
        if motivos:
            salida["MOTIVO"] = " | ".join(motivos)
            aparte.append(salida)
        else:
            salida["MOTIVO"] = (
                f"Texto recortado de {len(obs)} a {LIMITE_OBSERVACION} caracteres — revisar"
                if recortada
                else ""
            )
            listas.append(salida)
    return listas, aparte


def _pesos(valor: object) -> str:
    return f"${a_entero(valor):,}".replace(",", ".")


def _para_valor(filas: list[dict]) -> list[dict]:
    """Las filas en el formato que entiende valor_total()."""
    return [
        {
            "tipo_seguimiento": f["_tipo_seguimiento"],
            "valor_glosa": f["_valor"],
            "numero_factura": f["_numero_factura"],
        }
        for f in filas
    ]


def escribir(filas: list[dict], ruta: Path, titulo: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise SystemExit("ERROR: falta openpyxl. py -m pip install openpyxl") from None

    wb = Workbook()
    ws = wb.active
    ws.title = titulo
    azul, blanco = PatternFill("solid", fgColor="1F4E78"), Font(bold=True, color="FFFFFF")
    ambar = PatternFill("solid", fgColor="FFF2CC")
    for i, h in enumerate(COLS_SALIDA, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill, c.font = azul, blanco
    for r, f in enumerate(filas, start=2):
        for i, h in enumerate(COLS_SALIDA, 1):
            c = ws.cell(row=r, column=i, value=f.get(h, ""))
            c.alignment = Alignment(
                vertical="top", wrap_text=h in ("OBSERVACION_RESPUESTA", "MOTIVO")
            )
            if h == "VALOR_GLOSA":
                c.number_format = '"$"#,##0'
        if f.get("MOTIVO"):
            ws.cell(row=r, column=COLS_SALIDA.index("MOTIVO") + 1).fill = ambar
    anchos = {
        "OBSERVACION_RESPUESTA": 80,
        "MOTIVO": 45,
        "RAZON_SOCIAL_EPS": 30,
        "NUMERO_FACTURA": 15,
    }
    for i, h in enumerate(COLS_SALIDA, 1):
        ws.column_dimensions[get_column_letter(i)].width = anchos.get(h, 16)
    ws.freeze_panes = "A2"
    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ruta))


def mostrar(listas: list[dict], aparte: list[dict], ips, total: int) -> None:
    print()
    print("=" * 78)
    print(f"REVISIÓN PREVIA AL CARGUE — {ips.nombre}")
    print("=" * 78)
    print(f"\n  Filas del archivo          : {total:>7}")
    print(
        f"  LISTAS PARA CARGAR         : {len(listas):>7}   {_pesos(valor_total(_para_valor(listas)))}"
    )
    print(
        f"  Quedan fuera por revisar   : {len(aparte):>7}   {_pesos(valor_total(_para_valor(aparte)))}"
    )

    if listas:
        tipos = Counter(f["TIPO"] for f in listas)
        codigos = Counter(f["CODIGO_RESPUESTA"] for f in listas)
        recortadas = sum(1 for f in listas if f["MOTIVO"])
        print("\n  De las que se van a cargar:")
        for t, n in tipos.most_common():
            print(f"    {t:<14}{n:>7}")
        print(f"    códigos: {dict(codigos.most_common())}")
        if recortadas:
            print(
                f"\n    [!] {recortadas} tenían más de {LIMITE_OBSERVACION} caracteres y se"
                f" recortaron.\n        SIIFA no acepta más; sin recortar, esas rebotan."
            )

    if aparte:
        print("\n  POR QUÉ QUEDARON FUERA:")
        for motivo, n in Counter(f["MOTIVO"] for f in aparte).most_common():
            print(f"    {n:>7}  {motivo[:64]}")
    print()


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--archivo", required=True, help="Excel con las respuestas a cargar.")
    ap.add_argument("--salida", required=True, help="Excel de cargue a generar.")
    ap.add_argument(
        "--incluir-ya-respondidas",
        action="store_true",
        help="Cargar también las que YA tienen respuesta en SIIFA (las PISA).",
    )
    perfiles.agregar_argumento(ap)
    args = ap.parse_args()
    ips = perfiles.buscar(args.ips)
    perfiles.anunciar(ips)

    filas = leer_informe(Path(args.archivo))
    perfiles.verificar_informe(ips, filas, "el archivo de respuestas")

    listas, aparte = revisar(filas, ips, args.incluir_ya_respondidas)
    mostrar(listas, aparte, ips, len(filas))

    salida = Path(args.salida)
    if listas:
        escribir(listas, salida, "RESPUESTAS")
        print(f"  Archivo de cargue : {salida}")
    if aparte:
        revisar_ruta = salida.with_name(f"{salida.stem}_REVISAR{salida.suffix}")
        escribir(aparte, revisar_ruta, "REVISAR")
        print(f"  Lo que quedó fuera: {revisar_ruta}")

    if listas:
        print("\n  Antes del cargue completo, PILOTO DE 1 (regla del repo):")
        print(f"     py tools\\responder_glosas_siifa.py --ips {ips.clave} \\")
        print(f'        --excel "{salida}" --piloto 1 --reporte "piloto_{ips.clave}.csv"')
    print()


if __name__ == "__main__":
    main()
