"""extraer_detalle_glosas.py — Extrae el detalle por factura de una carpeta de glosas archivadas.

Recorre una carpeta de día del servidor de glosas (la que organiza
organizar_correos_glosas.py o el archivo manual), lee los adjuntos descargados
de cada entidad y produce el EXCEL DE ENTREGA con el formato del consolidado
de gestores: hojas INICIAL / RATIFICADA / DEVOLUCIONES, una fila por factura,
con RESPONSABLE asignado por entidad y FECHA DE VENCIMIENTO en días hábiles
(15 para inicial, 7 para ratificada, festivos colombianos).

Formatos que entiende (detectados por nombre/contenido, también dentro de .zip):

- AUDITOOL/DISPENSARIO:  AUDITORIA_GLOSA_*.csv          (; latin-1, por servicio)
- FAMISANAR:             DEVYGLOSAS*.txt                (| por servicio, códigos DE→devolución)
- SURA:                  *MASIVO*.TXT                   (; por glosa)
- SANITAS:               .xlsx hoja "Glosa"             (por ítem de glosa)
- POLICIA/GENESIS:       .xlsx hoja "OBJECIONES"        (CRNCXC/CROVALOBJ)
- FACTRAMED/NUEVA EPS:   .xlsx "Registro_Inforamcion.." (filtrado por ventana de fechas)
- AXA:                   PDF de liquidación             (No. factura / Glosado)
- SEGUROS BOLIVAR:       PDF de acta                    (Num. Factura / Val. Glosado inicial)
- PDF escaneado (DEV*):  fila mínima con el radicado para completar a mano

Los PDF que son correos impresos ("Correo de Hospital Universitario...") se
ignoran: son la evidencia, no los datos.

USO RÁPIDO
----------

    # La carpeta de un día completo
    py extraer_detalle_glosas.py --carpeta "Z:\\SERVIDOR GLOSAS\\...\\2026\\07 JULIO\\02"

    # Solo ver el resumen sin escribir el Excel
    py extraer_detalle_glosas.py --carpeta "...\\07 JULIO\\02" --dry-run

    # Con mapeo propio de gestores
    py extraer_detalle_glosas.py --carpeta "...\\02" --config gestores_glosas.ejemplo.json

DEPENDENCIAS
------------
- openpyxl (pip install openpyxl) para leer/escribir Excel.
- pdfplumber (pip install pdfplumber) opcional, para los PDF de AXA/Bolívar.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import logging
import re
import sys
import unicodedata
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path

DEFAULT_SALIDA = "ENTREGA_GLOSAS_{fecha}.xlsx"
DEFAULT_VENTANA_FACTRAMED = 7  # días hacia atrás de Fecha_Hora_Glosa que se incluyen

CATEGORIAS = ("INICIAL", "RATIFICADA", "DEVOLUCIONES")
PLAZO_HABILES = {"INICIAL": 15, "RATIFICADA": 7}

# Festivos Colombia (mismos del motor: app/services/glosa_service.py FERIADOS_CO)
FERIADOS_CO = {
    "2025-01-01",
    "2025-01-06",
    "2025-03-24",
    "2025-04-17",
    "2025-04-18",
    "2025-05-01",
    "2025-06-02",
    "2025-06-23",
    "2025-06-30",
    "2025-07-20",
    "2025-08-07",
    "2025-08-18",
    "2025-10-13",
    "2025-11-03",
    "2025-11-17",
    "2025-12-08",
    "2025-12-25",
    "2026-01-01",
    "2026-01-12",
    "2026-03-23",
    "2026-04-02",
    "2026-04-03",
    "2026-05-01",
    "2026-05-18",
    "2026-06-08",
    "2026-06-15",
    "2026-06-29",
    "2026-07-20",
    "2026-08-07",
    "2026-08-17",
    "2026-10-12",
    "2026-11-02",
    "2026-11-16",
    "2026-12-08",
    "2026-12-25",
    "2027-01-01",
    "2027-01-11",
    "2027-03-22",
    "2027-04-01",
    "2027-04-02",
    "2027-05-01",
    "2027-05-17",
    "2027-06-07",
    "2027-06-14",
    "2027-06-28",
    "2027-07-20",
    "2027-08-07",
    "2027-08-16",
    "2027-10-11",
    "2027-11-01",
    "2027-11-15",
    "2027-12-08",
    "2027-12-25",
    "2028-01-01",
    "2028-01-10",
    "2028-03-20",
    "2028-04-13",
    "2028-04-14",
    "2028-05-01",
    "2028-05-15",
    "2028-06-05",
    "2028-06-12",
    "2028-06-26",
    "2028-07-20",
    "2028-08-07",
    "2028-08-14",
    "2028-10-09",
    "2028-10-30",
    "2028-11-06",
    "2028-11-13",
}

RE_FACTURA_HUS = re.compile(r"\b(E?HUS\d{4,})\b")
RE_RADICADO_NOMBRE = re.compile(r"RADICADO\s*N?[O0]?\.?\s*(\d+)", re.IGNORECASE)

# _LIGERO: copias reducidas hechas a mano para compartir; duplicarían los valores
ARCHIVOS_IGNORADOS = re.compile(r"MANUAL-GESTION|DETALLE DE CAMPOS|_LIGERO|^~\$", re.IGNORECASE)

logger = logging.getLogger("extraer_detalle")

# ─── Gestores por entidad (sobreescribible con --config archivo.json) ────────
# Derivado del consolidado real; "POR ASIGNAR" donde no hay histórico claro.

GESTORES_DEFECTO: dict = {
    "por_defecto": "POR ASIGNAR",
    "entidades": {
        "DISPENSARIO": {
            "empresa": "DISPENSARIO MEDICO BUCARAMANGA",
            "INICIAL": "FANNY BARAJAS",
            "RATIFICADA": "POR ASIGNAR",
            "DEVOLUCIONES": "JOHANNA MORENO",
        },
        "AXA": {
            "empresa": "AXA COLPATRIA SEGUROS S.A.",
            "INICIAL": "SEBASTIAN ACUÑA",
            "RATIFICADA": "JORGE GUARIN",
            "DEVOLUCIONES": "SEBASTIAN ACUÑA",
        },
        "SEGUROS BOLIVAR": {
            "empresa": "SEGUROS COMERCIALES BOLIVAR",
            "INICIAL": "SEBASTIAN ACUÑA",
            "RATIFICADA": "JORGE GUARIN",
            "DEVOLUCIONES": "SEBASTIAN ACUÑA",
        },
        "SEGUROS MUNDIAL": {
            "empresa": "COMPAÑIA MUNDIAL DE SEGUROS S.A.",
            "INICIAL": "SEBASTIAN ACUÑA",
            "RATIFICADA": "JORGE GUARIN",
            "DEVOLUCIONES": "SEBASTIAN ACUÑA",
        },
        "SANITAS": {
            "empresa": "ENTIDAD PROMOTORA DE SALUD SANITAS S.A.S",
            "INICIAL": "POR ASIGNAR",
            "RATIFICADA": "JORGE GUARIN",
            "DEVOLUCIONES": "SEBASTIAN ACUÑA",
        },
        "NUEVA EPS": {
            "empresa": "NUEVA EMPRESA PROMOTORA DE SALUD S.A.",
            "INICIAL": "DIANEYDA QUINTERO",
            "RATIFICADA": "JORGE GUARIN",
            "DEVOLUCIONES": "DIANEYDA QUINTERO",
        },
        "COOSALUD": {
            "empresa": "COOSALUD ENTIDAD PROMOTORA DE SALUD S.A.",
            "INICIAL": "DIANEYDA QUINTERO",
            "RATIFICADA": "JORGE GUARIN",
            "DEVOLUCIONES": "ALBA PEREZ",
        },
        "ADRES": {
            "empresa": "ADRES - ACCIDENTES DE TRANSITO",
            "INICIAL": "MARICELA ROJAS",
            "RATIFICADA": "EQUIPO ADRES",
            "DEVOLUCIONES": "MARICELA ROJAS",
        },
        "FAMISANAR": {"empresa": "EPS FAMISANAR"},
        "SURA": {"empresa": "EPS SURA"},
        "POLICIA": {"empresa": "DIRECCION DE SANIDAD POLICIA NACIONAL"},
        "SALUD MIA": {"empresa": "SALUD MIA EPS"},
        "FACTRAMED": {"empresa": "NUEVA EMPRESA PROMOTORA DE SALUD S.A."},
        "LA PREVISORA": {"empresa": "LA PREVISORA S A COMPAÑIA DE SEGUROS"},
        "PREVISORA": {"empresa": "LA PREVISORA S A COMPAÑIA DE SEGUROS"},
        "SAVIA SALUD": {"empresa": "SAVIA SALUD EPS"},
        "POSITIVA": {"empresa": "POSITIVA COMPAÑIA DE SEGUROS"},
        "MUTUAL": {"empresa": "ASOCIACION MUTUAL SER"},
        "FOMAG": {"empresa": "FOMAG - FIDUPREVISORA"},
    },
}

ENCABEZADOS = {
    "INICIAL": [
        "RESPONSABLE",
        "FECHA ENTREGA",
        "FECHA RADICACION",
        "FECHA DE DOCUMENTO (DGH)",
        "FECHA NOTIFICACION OBJECIÓN",
        "EMPRESA",
        "NUMERO DE FACTURA",
        "CONSECUTIVO DGH",
        "VALOR GLOSA",
        "FECHA DE VENCIMIENTO",
        "DEVOLUCION\nSI/NO",
        "DIAS \nRADICACION VS RECEPCIO",
        "NUMERO RADICADO",
        "REFERENCIA",
        "OBSERVACION TECNICO",
        "TECNICO QUE RECEPCIONO",
    ],
    "RATIFICADA": [
        "RESPONSABLE",
        "FECHA ENTREGA",
        "FECHA DE DOCUMENTO (DGH)",
        "FECHA NOTIFICACION OBJECIÓN",
        "EMPRESA",
        "NUMERO DE FACTURA",
        "CONSECUTIVO DGH",
        "VALOR GLOSA",
        "FECHA VENCIMIENTO",
        "OBSERVACION TECNICO ",
        "TECNICO QUE RECEPCIONO ",
    ],
    "DEVOLUCIONES": [
        "RESPONSABLE",
        "FECHA ENTREGA",
        "FECHA DE RADICACION",
        "FECHA DE DOCUMENTO",
        "FECHA DE FACTURA",
        "FECHA NOTIFICACION DEVOLUCION",
        "EMPRESA",
        "NUMERO DE FACTURA",
        "CONSECUTIVO DGH",
        "VALOR FACTURA",
        "FECHA DE VENCIMIENTO",
        "REFERENCIA",
        "OBSERVACIONES TECNICO ",
        "TECNICO QUE RECEPCIONO",
    ],
}


# ─── Utilidades ──────────────────────────────────────────────────────────────


def setup_logging(log_file: Path | None = None) -> None:
    fmt = "%(asctime)s [%(levelname)s] %(message)s"
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(errors="replace")
    handlers: list[logging.Handler] = [logging.StreamHandler(sys.stdout)]
    if log_file is not None:
        log_file.parent.mkdir(parents=True, exist_ok=True)
        handlers.append(logging.FileHandler(log_file, encoding="utf-8"))
    logging.basicConfig(level=logging.INFO, format=fmt, handlers=handlers)


def _normalizar(texto: str) -> str:
    descompuesto = unicodedata.normalize("NFKD", str(texto))
    limpio = "".join(c for c in descompuesto if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", limpio).upper().strip()


def a_fecha(valor) -> date | None:
    """Convierte lo que venga (datetime, '18/06/26', '2026/06/25', '01/07/2026 00:00')."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip().split()[0] if str(valor).strip() else ""
    for formato in ("%d/%m/%Y", "%d/%m/%y", "%Y/%m/%d", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            fecha = datetime.strptime(texto, formato).date()
            if fecha.year > 1990:  # '1900/01/01' = sin fecha
                return fecha
            return None
        except ValueError:
            continue
    return None


def a_numero(valor) -> float | None:
    """'$9.576.693,00' -> 9576693.0 | '2,800' -> 2800.0 | 117300 -> 117300.0"""
    if valor is None or valor == "":
        return None
    if isinstance(valor, int | float):
        return float(valor)
    texto = re.sub(r"[^\d.,-]", "", str(valor))
    if not re.search(r"\d", texto):
        return None
    # el separador decimal es el ÚLTIMO . o , y solo si deja 1-2 dígitos
    entera, decimal = texto, ""
    ultimo = max(texto.rfind("."), texto.rfind(","))
    if ultimo != -1 and len(texto) - ultimo - 1 in (1, 2):
        entera, decimal = texto[:ultimo], texto[ultimo + 1 :]
    entera = re.sub(r"[.,]", "", entera)
    try:
        return float(f"{entera}.{decimal or 0}")
    except ValueError:
        return None


def sumar_habiles(desde: date, habiles: int) -> date:
    """Fecha de vencimiento: `habiles` días hábiles después de `desde`."""
    actual, faltan = desde, habiles
    while faltan > 0:
        actual += timedelta(days=1)
        if actual.weekday() < 5 and actual.isoformat() not in FERIADOS_CO:
            faltan -= 1
    return actual


def cargar_gestores(ruta: Path | None) -> dict:
    config = json.loads(json.dumps(GESTORES_DEFECTO))
    if ruta is None:
        return config
    try:
        propia = json.loads(ruta.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        sys.stderr.write(f"ERROR: no pude leer el config de gestores {ruta}: {exc}\n")
        sys.exit(2)
    config.update(propia)
    return config


def info_entidad(carpeta_entidad: str, gestores: dict) -> tuple[str, dict]:
    """(EMPRESA formal, reglas de gestor) para una carpeta como 'DISPENSARIO SOFIA OK'."""
    clave_carpeta = _normalizar(carpeta_entidad)
    for clave, reglas in gestores.get("entidades", {}).items():
        limpia = _normalizar(clave)
        if clave_carpeta == limpia or clave_carpeta.startswith(limpia + " "):
            return reglas.get("empresa", clave), reglas
    # sin regla: el nombre de la carpeta sin la marca ' OK' final
    empresa = re.sub(r"\s+OK$", "", clave_carpeta)
    return empresa, {}


def responsable_para(reglas: dict, categoria: str, gestores: dict) -> str:
    return reglas.get(categoria) or gestores.get("por_defecto", "POR ASIGNAR")


# ─── Parsers (cada uno devuelve list[dict] por servicio/ítem) ────────────────
# Claves: categoria (None = usar la carpeta), factura, valor, fecha_notificacion,
# fecha_radicacion, fecha_factura, no_glosa, radicado, observacion


def _categoria_por_codigo(codigo: str) -> str | None:
    """Códigos estándar (Res. 3047/2284): DE* = devolución; FA/TA/SO/AU/CO/CL/UN = glosa."""
    codigo = (codigo or "").strip().upper()
    if codigo.startswith("DE"):
        return "DEVOLUCIONES"
    return None


def parser_auditool_csv(contenido: bytes, nombre: str) -> list[dict]:
    texto = contenido.decode("latin-1", errors="replace")
    lineas = texto.splitlines()
    idx_encabezado = next(
        (i for i, ln in enumerate(lineas) if "FACTURA" in ln.upper() and ";" in ln), None
    )
    if idx_encabezado is None:
        return []
    filas = list(csv.reader(lineas[idx_encabezado:], delimiter=";"))
    encabezado = [_normalizar(c) for c in filas[0]]
    categoria = "RATIFICADA" if any("RATIFICADA" in c for c in encabezado) else "INICIAL"

    def col(*nombres: str) -> int | None:
        for n in nombres:
            for i, c in enumerate(encabezado):
                if n in c:
                    return i
        return None

    i_prefijo = col("PREFIJO")
    i_factura = next((i for i, c in enumerate(encabezado) if c == "FACTURA"), col("FACTURA"))
    i_radicado = col("RADICADO FACTURA", "RADICADO")
    i_valor = col("VALOR GLOSA RATIFICADA", "VALOR GLOSA INICIAL", "VALOR GLOSA")
    i_fnotif = col("NOTIFICACION AL PRESTADOR", "NOTIFICACION")
    i_frad = col("FEC. PRESENTACION", "PRESENTACION FACTURA")
    i_femision = col("FECHA EMISION")
    i_codigo = col("CODIGO GLOSA")
    resultados = []
    for fila in filas[1:]:
        if i_factura is None or len(fila) <= i_factura or not fila[i_factura].strip():
            continue

        def _v(i: int | None, fila=fila) -> str:
            return fila[i].strip() if i is not None and len(fila) > i else ""

        prefijo = _v(i_prefijo)
        resultados.append(
            {
                "categoria": categoria,
                "factura": f"{prefijo}{_v(i_factura)}",
                "valor": a_numero(_v(i_valor)),
                "fecha_notificacion": a_fecha(_v(i_fnotif)),
                "fecha_radicacion": a_fecha(_v(i_frad)),
                "fecha_factura": a_fecha(_v(i_femision)),
                "no_glosa": (_v(i_codigo).split() or [""])[0][:12],
                "radicado": _v(i_radicado),
                "observacion": "",
            }
        )
    return resultados


def parser_famisanar_txt(contenido: bytes, nombre: str) -> list[dict]:
    texto = contenido.decode("latin-1", errors="replace")
    filas = list(csv.reader(texto.splitlines(), delimiter="|"))
    if not filas or "NRO_FACTURA" not in "|".join(filas[0]).upper():
        return []
    encabezado = [_normalizar(c) for c in filas[0]]
    idx = {c: i for i, c in enumerate(encabezado)}
    resultados = []
    for fila in filas[1:]:
        if len(fila) < len(encabezado):
            continue
        codigo = fila[idx.get("CODIGO_DEVOLUCION", -1)] if "CODIGO_DEVOLUCION" in idx else ""
        resultados.append(
            {
                "categoria": _categoria_por_codigo(codigo),
                "factura": fila[idx["NRO_FACTURA"]].strip(),
                "valor": a_numero(fila[idx.get("VALOR DEVOLUCION", -1)]),
                "fecha_notificacion": a_fecha(fila[idx.get("FECHA_ENVIO", -1)]),
                "fecha_radicacion": None,
                "fecha_factura": None,
                "no_glosa": codigo.strip(),
                "radicado": fila[idx.get("NRO_RADICACION", -1)].strip(),
                "observacion": "",
            }
        )
    return resultados


def parser_sura_txt(contenido: bytes, nombre: str) -> list[dict]:
    texto = contenido.decode("latin-1", errors="replace")
    filas = list(csv.reader(texto.splitlines(), delimiter=";"))
    if not filas or "NUMEROFACTURA" not in "".join(filas[0]).upper():
        return []
    idx = {_normalizar(c): i for i, c in enumerate(filas[0])}
    resultados = []
    for fila in filas[1:]:
        if len(fila) < 3 or not fila[idx["NUMEROFACTURA"]].strip():
            continue
        prefijo = fila[idx.get("PREFIJOFACTURA", -1)].strip() if "PREFIJOFACTURA" in idx else ""
        resultados.append(
            {
                "categoria": "INICIAL",
                "factura": f"{prefijo}{fila[idx['NUMEROFACTURA']].strip()}",
                "valor": a_numero(fila[idx.get("VALORGLOSA", -1)]),
                "fecha_notificacion": a_fecha(fila[idx.get("FECHAGLOSA", -1)]),
                "fecha_radicacion": a_fecha(fila[idx.get("FECHARADICACION", -1)]),
                "fecha_factura": a_fecha(fila[idx.get("FECHAFACTURA", -1)]),
                "no_glosa": fila[idx.get("NUMEROGLOSA", -1)].strip(),
                "radicado": fila[idx.get("NUMERORADICADO", -1)].strip(),
                "observacion": "",
            }
        )
    return resultados


def _filas_xlsx(contenido: bytes):
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    hoja = wb[wb.sheetnames[0]]
    return hoja.iter_rows(values_only=True)


def parser_sanitas_xlsx(contenido: bytes, nombre: str) -> list[dict]:
    filas = _filas_xlsx(contenido)
    encabezado = [_normalizar(c or "") for c in next(filas)]
    if "NUMERO DE FACTURA" not in encabezado or not any("VALOR GLOSADO" in c for c in encabezado):
        return []
    idx = {c: i for i, c in enumerate(encabezado)}
    i_valor = next(
        (
            idx[c]
            for c in ("VALOR REAL GLOSA", "VALOR GLOSADO PRESTADOR", "VALOR GLOSADO GLOSA")
            if c in idx
        ),
        None,
    )
    resultados = []
    for fila in filas:
        factura = fila[idx["NUMERO DE FACTURA"]]
        if not factura:
            continue
        codigo = str(fila[idx.get("CODIGO CONCEPTO GENERAL", 0)] or "")
        resultados.append(
            {
                "categoria": _categoria_por_codigo(codigo),
                "factura": str(factura).strip(),
                "valor": a_numero(fila[i_valor]) if i_valor is not None else None,
                "fecha_notificacion": None,
                "fecha_radicacion": a_fecha(fila[idx.get("FECHA DE RADICACION", -1)])
                if "FECHA DE RADICACION" in idx
                else None,
                "fecha_factura": a_fecha(fila[idx.get("FECHA DE FACTURA", -1)])
                if "FECHA DE FACTURA" in idx
                else None,
                "no_glosa": str(fila[idx.get("ID GLOSA", 0)] or "").strip(),
                "radicado": str(fila[idx.get("NUMERO DE RADICACION", 0)] or "").strip()
                if "NUMERO DE RADICACION" in idx
                else "",
                "observacion": "",
            }
        )
    return resultados


def parser_objeciones_xlsx(contenido: bytes, nombre: str) -> list[dict]:
    filas = _filas_xlsx(contenido)
    encabezado = [_normalizar(c or "") for c in next(filas)]
    if "CRNCXC" not in encabezado:
        return []
    idx = {c: i for i, c in enumerate(encabezado)}
    resultados = []
    for fila in filas:
        factura = fila[idx["CRNCXC"]]
        if not factura:
            continue
        codigo = str(fila[idx.get("CRNCONOBJ", 0)] or "")
        resultados.append(
            {
                "categoria": _categoria_por_codigo(codigo),
                "factura": str(factura).strip(),
                "valor": a_numero(fila[idx.get("CROVALOBJ", -1)]) if "CROVALOBJ" in idx else None,
                "fecha_notificacion": a_fecha(fila[idx.get("CROFECOBJ", -1)])
                if "CROFECOBJ" in idx
                else None,
                "fecha_radicacion": None,
                "fecha_factura": a_fecha(fila[idx.get("CDFECDOC", -1)])
                if "CDFECDOC" in idx
                else None,
                "no_glosa": codigo.strip(),
                "radicado": "",
                "observacion": "",
            }
        )
    return resultados


def parser_factramed_xlsx(
    contenido: bytes, nombre: str, desde: date | None = None, hasta: date | None = None
) -> list[dict]:
    filas = _filas_xlsx(contenido)
    encabezado = [_normalizar(c or "") for c in next(filas)]
    if "NUMERO_FACTURA" not in encabezado or "CODIGO_GLOSA_DEV" not in encabezado:
        return []
    idx = {c: i for i, c in enumerate(encabezado)}
    resultados, omitidas = [], 0
    for fila in filas:
        factura = fila[idx["NUMERO_FACTURA"]]
        if not factura:
            continue
        fecha_glosa = a_fecha(fila[idx.get("FECHA_HORA_GLOSA", -1)])
        if desde and (
            fecha_glosa is None or fecha_glosa < desde or (hasta and fecha_glosa > hasta)
        ):
            omitidas += 1
            continue
        codigo = str(fila[idx.get("CODIGO_GLOSA_DEV", 0)] or "")
        resultados.append(
            {
                "categoria": _categoria_por_codigo(codigo),
                "factura": str(factura).strip(),
                "valor": a_numero(fila[idx.get("VALOR_GLOSA_DEV", -1)]),
                "valor_factura": a_numero(fila[idx.get("VALOR_TOTAL_FACTURA", -1)]),
                "fecha_notificacion": fecha_glosa,
                "fecha_radicacion": a_fecha(fila[idx.get("FECHA_HORA_RADICACION", -1)]),
                "fecha_factura": None,
                "no_glosa": str(fila[idx.get("NUMERO_GLOSA_DEV", 0)] or "").strip(),
                "radicado": str(fila[idx.get("NUMERO_RADICACION", 0)] or "").strip(),
                "observacion": "",
            }
        )
    if omitidas:
        logger.info(
            f"  {nombre}: {omitidas} filas fuera de la ventana de fechas "
            f"({desde} a {hasta}) omitidas; usa --ventana-factramed para ampliar"
        )
    return resultados


def parser_savia_xlsx(contenido: bytes, nombre: str) -> list[dict]:
    """SAVIA SALUD: una fila por servicio glosado (Numero_factura / Valor_Glosa)."""
    filas = _filas_xlsx(contenido)
    encabezado = [_normalizar(c or "") for c in next(filas, [])]
    if "NUMERO_FACTURA" not in encabezado or "VALOR_GLOSA" not in encabezado:
        return []
    idx = {c: i for i, c in enumerate(encabezado)}
    resultados = []
    for fila in filas:
        factura = fila[idx["NUMERO_FACTURA"]]
        if not factura:
            continue
        resultados.append(
            {
                "categoria": None,
                "factura": str(factura).strip(),
                "valor": a_numero(fila[idx["VALOR_GLOSA"]]),
                "fecha_notificacion": None,
                "fecha_radicacion": a_fecha(fila[idx.get("FECHA_RADICACION", -1)])
                if "FECHA_RADICACION" in idx
                else None,
                "fecha_factura": a_fecha(fila[idx.get("FECHA_FACTURA", -1)])
                if "FECHA_FACTURA" in idx
                else None,
                "no_glosa": str(fila[idx.get("MOTIVO_GLOSA_VALOR_A", 0)] or "").strip()[:12],
                "radicado": str(fila[idx.get("NUMERO_RADICADO", 0)] or "").strip()
                if "NUMERO_RADICADO" in idx
                else "",
                "observacion": "",
            }
        )
    return resultados


def parser_coosalud_xlsx(contenido: bytes, nombre: str) -> list[dict]:
    """COOSALUD: GLOSAS HUS*.xlsx dentro de zips anidados (una fila por ítem)."""
    filas = _filas_xlsx(contenido)
    encabezado = [_normalizar(c or "") for c in next(filas, [])]
    if "NUMERO_FACTURA" not in encabezado or "VALOR_TOTAL_GLOSA" not in encabezado:
        return []
    idx = {c: i for i, c in enumerate(encabezado)}
    resultados = []
    for fila in filas:
        factura = fila[idx["NUMERO_FACTURA"]]
        if not factura:
            continue
        codigo = str(fila[idx.get("CODIGO_GLOSA", 0)] or "")
        resultados.append(
            {
                "categoria": _categoria_por_codigo(codigo),
                "factura": str(factura).strip(),
                "valor": a_numero(fila[idx["VALOR_TOTAL_GLOSA"]]),
                "fecha_notificacion": a_fecha(fila[idx.get("FECHA_GLOSA", -1)])
                if "FECHA_GLOSA" in idx
                else None,
                "fecha_radicacion": None,
                "fecha_factura": None,
                "no_glosa": codigo.strip(),
                "radicado": "",
                "observacion": "",
            }
        )
    return resultados


def parser_policia_rad_xlsx(contenido: bytes, nombre: str) -> list[dict]:
    """POLICÍA 'GLOSA A RAD NNNN.xlsx': certificación con encabezado en la
    fila ~14 (No.FACTURA / VR. GLOSA) y el radicado unas filas arriba."""
    filas = list(_filas_xlsx(contenido))
    idx_encabezado = None
    for i, fila in enumerate(filas[:25]):
        normalizados = [_normalizar(str(c or "")) for c in fila]
        if any("NO.FACTURA" in c or "NO. FACTURA" in c for c in normalizados):
            idx_encabezado = i
            break
    if idx_encabezado is None:
        return []
    encabezado = [_normalizar(str(c or "")) for c in filas[idx_encabezado]]

    def col(nombre_col: str) -> int | None:
        return next((i for i, c in enumerate(encabezado) if nombre_col in c), None)

    i_factura = col("NO.FACTURA") or col("NO. FACTURA")
    i_valor = col("VR. GLOSA")
    i_fglosa = col("FECHA GLOSA")
    i_ffactura = col("FECHA DE FACTURA")
    radicado = ""
    for fila in filas[:idx_encabezado]:
        textos = [str(c) for c in fila if c is not None]
        for j, texto in enumerate(textos):
            if "RADICACION" in _normalizar(texto) and j + 1 < len(textos):
                radicado = re.sub(r"\D", "", textos[j + 1])[:10]
    if not radicado:
        m = re.search(r"RAD\D*(\d+)", _normalizar(nombre))
        radicado = m.group(1) if m else ""
    resultados = []
    for fila in filas[idx_encabezado + 1 :]:
        factura = fila[i_factura] if i_factura is not None and len(fila) > i_factura else None
        if not factura or not RE_FACTURA_HUS.search(str(factura)):
            continue
        resultados.append(
            {
                "categoria": None,
                "factura": str(factura).strip(),
                "valor": a_numero(fila[i_valor]) if i_valor is not None else None,
                "fecha_notificacion": a_fecha(fila[i_fglosa]) if i_fglosa is not None else None,
                "fecha_radicacion": None,
                "fecha_factura": a_fecha(fila[i_ffactura]) if i_ffactura is not None else None,
                "no_glosa": "",
                "radicado": radicado,
                "observacion": "",
            }
        )
    return resultados


def _fila_minima_por_nombre(nombre: str) -> list[dict]:
    """Último recurso: la factura viene en el nombre del archivo (ej. MUTUAL
    'Validación de subsanación Factura N° HUS492542 OK.xlsx')."""
    factura = RE_FACTURA_HUS.search(_normalizar(nombre))
    if not factura:
        return []
    return [
        {
            "categoria": None,
            "factura": factura.group(1),
            "valor": None,
            "fecha_notificacion": None,
            "fecha_radicacion": None,
            "fecha_factura": None,
            "no_glosa": "",
            "radicado": "",
            "observacion": f"formato sin parser ({nombre}): completar a mano",
        }
    ]


def _texto_pdf(contenido: bytes) -> str:
    try:
        import pdfplumber
    except ImportError:
        logger.warning("pdfplumber no instalado: los PDF no se extraen (pip install pdfplumber)")
        return ""
    try:
        with pdfplumber.open(io.BytesIO(contenido)) as pdf:
            return "\n".join(pagina.extract_text() or "" for pagina in pdf.pages)
    except Exception as exc:
        logger.warning(f"PDF ilegible: {exc}")
        return ""


def parser_pdf(contenido: bytes, nombre: str) -> list[dict]:
    """PDFs de AXA (liquidación), Bolívar (acta) o escaneados (fila mínima)."""
    texto = _texto_pdf(contenido)
    if "CORREO DE HOSPITAL UNIVERSITARIO" in _normalizar(texto[:300]):
        return []  # es el correo impreso, no datos
    if texto.strip():
        plano = _normalizar(texto)
        # el acta de Bolívar primero: es la firma más específica
        if re.search(r"ACTA\s*NRO", plano) and "VAL. GLOSADO" in plano:
            return _parser_bolivar(texto)
        # AXA/liquidaciones tabulares: 'No.reclamo' / 'No.factura' (con punto)
        if "GLOSADO" in plano and re.search(r"\bNO\.\s*(RECLAMO|FACTURA)", plano):
            return _parser_axa(texto)
    # escaneado o formato desconocido: fila mínima si hay radicado o factura
    radicado = RE_RADICADO_NOMBRE.search(nombre)
    factura = RE_FACTURA_HUS.search(texto or nombre)
    if radicado or factura:
        return [
            {
                "categoria": None,
                "factura": factura.group(1) if factura else "",
                "valor": None,
                "fecha_notificacion": None,
                "fecha_radicacion": None,
                "fecha_factura": None,
                "no_glosa": "",
                "radicado": radicado.group(1) if radicado else "",
                "observacion": f"PDF sin datos extraíbles ({nombre}): completar a mano",
            }
        ]
    return []


def _parser_axa(texto: str) -> list[dict]:
    """Liquidación tabular: la línea de cada reclamo trae la factura y el
    'Glosado' es el ÚLTIMO número de la línea."""
    fecha_doc = next(
        (a_fecha(f) for f in re.findall(r"\b\d{4}-\d{1,2}-\d{1,2}\b", texto) if a_fecha(f)), None
    )
    resultados = []
    for linea in texto.splitlines():
        m = RE_FACTURA_HUS.search(linea)
        if not m or "FACTURA" in _normalizar(linea):  # salta el encabezado de la tabla
            continue
        numeros = re.findall(r"\d[\d.,]*", linea[m.end() :])
        valor = a_numero(numeros[-1]) if numeros else None
        fecha_linea = next(
            (a_fecha(t) for t in re.findall(r"\d{2}-\d{2}-\d{4}", linea) if a_fecha(t)), None
        )
        resultados.append(
            {
                "categoria": None,
                "factura": m.group(1),
                "valor": valor,
                "fecha_notificacion": fecha_linea or fecha_doc,
                "fecha_radicacion": None,
                "fecha_factura": None,
                "no_glosa": "",
                "radicado": "",
                "observacion": "" if valor is not None else "valor glosado: ver PDF",
            }
        )
    return resultados


def _parser_bolivar(texto: str) -> list[dict]:
    facturas = list(dict.fromkeys(RE_FACTURA_HUS.findall(texto)))
    acta = re.search(r"Acta\s+Nro\.?:?\s*(\d+)", texto, re.IGNORECASE)
    fecha = a_fecha((re.search(r"Fecha:\s*([\d/]+)", texto) or [None, None])[1])
    # 'Val. Glosado inicial' = primer monto de cada ítem '(n.nn) ...'
    montos = [
        a_numero(m)
        for m in re.findall(
            r"^\s*\(\d+\.\d+\).*?([\d.,]{4,})\s+[\d.,]+\s+[\d.,]+\s+[\d.,]+\s*$",
            texto,
            re.MULTILINE,
        )
    ]
    total = sum(m for m in montos if m) if montos else None
    return [
        {
            "categoria": None,
            "factura": factura,
            "valor": total if len(facturas) == 1 else None,
            "fecha_notificacion": fecha,
            "fecha_radicacion": None,
            "fecha_factura": None,
            "no_glosa": acta.group(1) if acta else "",
            "radicado": acta.group(1) if acta else "",
            "observacion": "" if (total and len(facturas) == 1) else "valores: ver acta PDF",
        }
        for factura in facturas
    ]


def elegir_parser(nombre: str):
    plano = _normalizar(Path(nombre).name)
    extension = Path(nombre).suffix.lower()
    if ARCHIVOS_IGNORADOS.search(plano):
        return None
    if plano.startswith("AUDITORIA_GLOSA") and extension == ".csv":
        return parser_auditool_csv
    if plano.startswith("DEVYGLOSAS") and extension == ".txt":
        return parser_famisanar_txt
    if "MASIVO" in plano and extension == ".txt":
        return parser_sura_txt
    if extension == ".xlsx":
        return _parser_xlsx_auto
    if extension == ".pdf":
        return parser_pdf
    return None


def _parser_xlsx_auto(contenido: bytes, nombre: str, **kwargs) -> list[dict]:
    """Lee el encabezado UNA vez y enruta por firma de columnas (un día grande
    trae 20+ zips de COOSALUD: probar todos los parsers reabriendo cada xlsx
    multiplicaba el tiempo por 7)."""
    try:
        encabezado = {_normalizar(str(c or "")) for c in next(_filas_xlsx(contenido), [])}
    except Exception:
        encabezado = set()
    if "NUMERO DE FACTURA" in encabezado and any("VALOR GLOSADO" in c for c in encabezado):
        return parser_sanitas_xlsx(contenido, nombre)
    if "CRNCXC" in encabezado:
        return parser_objeciones_xlsx(contenido, nombre)
    if "NUMERO_FACTURA" in encabezado and "VALOR_TOTAL_GLOSA" in encabezado:
        return parser_coosalud_xlsx(contenido, nombre)
    if "NUMERO_FACTURA" in encabezado and "VALOR_GLOSA" in encabezado:
        return parser_savia_xlsx(contenido, nombre)
    if "NUMERO_FACTURA" in encabezado and "CODIGO_GLOSA_DEV" in encabezado:
        return parser_factramed_xlsx(contenido, nombre, **kwargs)
    filas = parser_policia_rad_xlsx(contenido, nombre)  # encabezado en fila intermedia
    if filas:
        return filas
    # los DETALLE/HUSxxxx.xlsx acompañan al GLOSAS de COOSALUD: mismo contenido,
    # no deben producir filas 'completar a mano' junto a las reales
    if _normalizar(Path(nombre).stem).startswith(("DETALLE", "HUS")):
        return []
    return _fila_minima_por_nombre(nombre)


# ─── Recorrido de la carpeta del día ─────────────────────────────────────────


def _categoria_de_carpeta(nombre: str) -> str | None:
    plano = _normalizar(nombre)
    for categoria in CATEGORIAS:
        if plano == categoria or plano == categoria + "S" or plano.startswith(categoria + " "):
            return categoria
    if plano.startswith("RATIFICADA"):  # 'RATIFICADAS OK'
        return "RATIFICADA"
    return None


def extraer_de_archivo(
    ruta_o_nombre: str, contenido: bytes, ventana: tuple[date | None, date | None]
) -> tuple[list[dict], bool]:
    """(filas, tenia_parser). Abre .zip recursivamente."""
    if ARCHIVOS_IGNORADOS.search(_normalizar(Path(ruta_o_nombre).name)):
        return [], True  # ignorado a propósito (manuales, copias _LIGERO): sin reporte
    if Path(ruta_o_nombre).suffix.lower() == ".zip":
        filas, alguno = [], False
        try:
            with zipfile.ZipFile(io.BytesIO(contenido)) as zf:
                for interno in zf.namelist():
                    if interno.endswith("/"):
                        continue
                    sub, tenia = extraer_de_archivo(interno, zf.read(interno), ventana)
                    filas.extend(sub)
                    alguno = alguno or tenia
        except zipfile.BadZipFile:
            logger.warning(f"  ZIP corrupto: {ruta_o_nombre}")
        return filas, alguno
    parser = elegir_parser(ruta_o_nombre)
    if parser is None:
        return [], False
    if parser is _parser_xlsx_auto:
        return parser(contenido, Path(ruta_o_nombre).name, desde=ventana[0], hasta=ventana[1]), True
    return parser(contenido, Path(ruta_o_nombre).name), True


def fecha_de_carpeta(carpeta: Path) -> date:
    """'...\\2026\\07 JULIO\\02' -> 2026-07-02; si no se puede, hoy."""
    try:
        dia = int(re.match(r"(\d{1,2})", carpeta.name).group(1))
        mes = int(re.match(r"(\d{1,2})", carpeta.parent.name).group(1))
        anio = int(re.match(r"(\d{4})", carpeta.parent.parent.name).group(1))
        return date(anio, mes, dia)
    except (AttributeError, ValueError):
        return date.today()


def recorrer_dia(
    carpeta_dia: Path, gestores: dict, ventana_factramed: int
) -> tuple[list[dict], list[str]]:
    """Devuelve (filas por factura ya agrupadas, archivos sin parser)."""
    fecha_dia = fecha_de_carpeta(carpeta_dia)
    crudas: list[dict] = []
    sin_parser: list[str] = []
    ventana = (fecha_dia - timedelta(days=ventana_factramed), fecha_dia)
    for carpeta_categoria in sorted(carpeta_dia.iterdir()):
        if not carpeta_categoria.is_dir():
            continue
        categoria_carpeta = _categoria_de_carpeta(carpeta_categoria.name)
        if categoria_carpeta is None:
            logger.info(f"Carpeta ignorada (sin categoría): {carpeta_categoria.name}")
            continue
        for carpeta_entidad in sorted(carpeta_categoria.iterdir()):
            if not carpeta_entidad.is_dir():
                continue
            empresa, reglas = info_entidad(carpeta_entidad.name, gestores)
            for archivo in sorted(carpeta_entidad.rglob("*")):
                if not archivo.is_file():
                    continue
                try:
                    contenido = archivo.read_bytes()
                except OSError as exc:
                    logger.warning(f"  No pude leer {archivo.name}: {exc}")
                    continue
                filas, tenia = extraer_de_archivo(archivo.name, contenido, ventana)
                if not tenia and Path(archivo.name).suffix.lower() not in (".pdf",):
                    sin_parser.append(
                        f"{carpeta_categoria.name}/{carpeta_entidad.name}/{archivo.name}"
                    )
                for fila in filas:
                    fila["categoria"] = fila.get("categoria") or categoria_carpeta
                    fila["empresa"] = empresa
                    fila["reglas"] = reglas
                    fila["origen"] = f"{carpeta_entidad.name}/{archivo.name}"
                crudas.extend(filas)
                if filas:
                    logger.info(
                        f"  [{categoria_carpeta}] {carpeta_entidad.name} / {archivo.name}: "
                        f"{len(filas)} ítems"
                    )
    return agrupar_por_factura(crudas), sin_parser


def agrupar_por_factura(crudas: list[dict]) -> list[dict]:
    """Una fila por (categoría, empresa, factura): valores sumados, primeras fechas."""
    grupos: dict[tuple, dict] = {}
    for fila in crudas:
        clave = (fila["categoria"], fila["empresa"], fila.get("factura") or fila.get("radicado"))
        grupo = grupos.get(clave)
        if grupo is None:
            grupos[clave] = fila | {"items": 1}
            continue
        grupo["items"] += 1
        if fila.get("valor") is not None:
            grupo["valor"] = (grupo.get("valor") or 0) + fila["valor"]
            # un valor real supera la advertencia de una fila mínima previa
            if "completar a mano" in (grupo.get("observacion") or ""):
                grupo["observacion"] = fila.get("observacion", "")
        if fila.get("valor_factura") and not grupo.get("valor_factura"):
            grupo["valor_factura"] = fila["valor_factura"]
        for campo in ("fecha_notificacion", "fecha_radicacion", "fecha_factura", "radicado"):
            if not grupo.get(campo) and fila.get(campo):
                grupo[campo] = fila[campo]
        if fila.get("no_glosa") and fila["no_glosa"] not in (grupo.get("no_glosa") or ""):
            unidos = f"{grupo.get('no_glosa', '')} {fila['no_glosa']}".split()
            grupo["no_glosa"] = " ".join(list(dict.fromkeys(unidos))[:4])
    return list(grupos.values())


# ─── Excel de entrega ────────────────────────────────────────────────────────


def _fila_para_hoja(fila: dict, categoria: str, fecha_entrega: date, gestores: dict) -> list:
    responsable = responsable_para(fila.get("reglas", {}), categoria, gestores)
    vence = ""
    if categoria in PLAZO_HABILES and fila.get("fecha_notificacion"):
        vence = sumar_habiles(fila["fecha_notificacion"], PLAZO_HABILES[categoria])
    comun = {
        "responsable": responsable,
        "entrega": fecha_entrega,
        "notificacion": fila.get("fecha_notificacion") or "",
        "empresa": fila.get("empresa", ""),
        "factura": fila.get("factura", ""),
        "valor": fila.get("valor") if fila.get("valor") is not None else "",
        "vence": vence,
        "obs": fila.get("observacion", ""),
        "radicado": fila.get("radicado", ""),
        "no_glosa": fila.get("no_glosa", ""),
    }
    if categoria == "INICIAL":
        return [
            comun["responsable"],
            comun["entrega"],
            fila.get("fecha_radicacion") or "",
            "",
            comun["notificacion"],
            comun["empresa"],
            comun["factura"],
            comun["no_glosa"],
            comun["valor"],
            comun["vence"],
            "N",
            "",
            comun["radicado"],
            "",
            comun["obs"],
            "",
        ]
    if categoria == "RATIFICADA":
        return [
            comun["responsable"],
            comun["entrega"],
            "",
            comun["notificacion"],
            comun["empresa"],
            comun["factura"],
            comun["no_glosa"],
            comun["valor"],
            comun["vence"],
            comun["obs"],
            "",
        ]
    return [
        comun["responsable"],
        comun["entrega"],
        fila.get("fecha_radicacion") or "",
        "",
        fila.get("fecha_factura") or "",
        comun["notificacion"],
        comun["empresa"],
        comun["factura"],
        comun["no_glosa"],
        fila.get("valor_factura") or comun["valor"],
        "",
        comun["radicado"],
        comun["obs"],
        "",
    ]


def escribir_entrega(
    filas: list[dict], sin_parser: list[str], salida: Path, fecha_entrega: date, gestores: dict
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)
    negrita = Font(bold=True)
    for categoria in CATEGORIAS:
        hoja = wb.create_sheet(categoria)
        hoja.append(ENCABEZADOS[categoria])
        for celda in hoja[1]:
            celda.font = negrita
        hoja.freeze_panes = "A2"
        del_grupo = sorted(
            (f for f in filas if f["categoria"] == categoria),
            key=lambda f: (f.get("empresa", ""), str(f.get("factura", ""))),
        )
        for fila in del_grupo:
            hoja.append(_fila_para_hoja(fila, categoria, fecha_entrega, gestores))
    resumen = wb.create_sheet("RESUMEN")
    resumen.append(["CATEGORIA", "EMPRESA", "FACTURAS", "VALOR"])
    for celda in resumen[1]:
        celda.font = negrita
    conteo: dict[tuple, list] = {}
    for fila in filas:
        clave = (fila["categoria"], fila.get("empresa", ""))
        acumulado = conteo.setdefault(clave, [0, 0.0])
        acumulado[0] += 1
        acumulado[1] += fila.get("valor") or 0
    for (categoria, empresa), (n, valor) in sorted(conteo.items()):
        resumen.append([categoria, empresa, n, valor])
    if sin_parser:
        resumen.append([])
        resumen.append(["ARCHIVOS SIN PARSER (revisar a mano)"])
        for ruta in sin_parser:
            resumen.append([ruta])
    wb.save(salida)


# ─── Orquestación ────────────────────────────────────────────────────────────


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--carpeta", type=Path, required=True, help="Carpeta del día a procesar")
    parser.add_argument("--salida", type=Path, default=None, help="Ruta del Excel de entrega")
    parser.add_argument("--config", type=Path, default=None, help="JSON de gestores por entidad")
    parser.add_argument(
        "--ventana-factramed",
        type=int,
        default=DEFAULT_VENTANA_FACTRAMED,
        help=f"Días hacia atrás de Fecha_Hora_Glosa a incluir del Excel FACTRAMED "
        f"(default: {DEFAULT_VENTANA_FACTRAMED})",
    )
    parser.add_argument("--dry-run", action="store_true", help="Solo muestra el resumen")
    parser.add_argument("--log", type=Path, default=None, help="Archivo de log opcional")
    args = parser.parse_args()
    setup_logging(args.log)

    if not args.carpeta.is_dir():
        sys.stderr.write(f"ERROR: no existe la carpeta {args.carpeta}\n")
        return 2
    gestores = cargar_gestores(args.config)
    logger.info(f"Procesando {args.carpeta}")
    filas, sin_parser = recorrer_dia(args.carpeta, gestores, args.ventana_factramed)

    logger.info("=" * 60)
    logger.info("RESUMEN")
    for categoria in CATEGORIAS:
        del_grupo = [f for f in filas if f["categoria"] == categoria]
        valor = sum(f.get("valor") or 0 for f in del_grupo)
        logger.info(f"  {categoria}: {len(del_grupo)} facturas | ${valor:,.0f}")
    for ruta in sin_parser:
        logger.warning(f"  SIN PARSER: {ruta}")
    logger.info("=" * 60)

    if args.dry_run:
        logger.info("Simulacro: no se escribió ningún Excel")
        return 0
    hoy = date.today()
    salida = args.salida or args.carpeta / DEFAULT_SALIDA.format(fecha=hoy.isoformat())
    if salida.exists():
        respaldo = salida.with_name(f"{salida.stem} ({datetime.now():%H.%M}){salida.suffix}")
        logger.warning(f"{salida.name} ya existe; escribo {respaldo.name}")
        salida = respaldo
    escribir_entrega(filas, sin_parser, salida, hoy, gestores)
    logger.info(f"Excel de entrega: {salida}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
