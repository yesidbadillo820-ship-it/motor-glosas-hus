"""extraer_respuestas_glosa_mutualser.py — Excel de respuestas a glosas MUTUAL SER desde el PDF.

Lee uno o varios PDF de "Trámite de Objeción" del HUS (reporte CRRPTramiteObjecion,
variante Glosa_Ratificada / Administrativo / Mixta que se radica ante MUTUAL SER en
el portal Zona Ser — https://portalzonaser.mutualser.com) y produce un Excel con
UNA fila por objeción/ítem con:

    Factura | # Objeción | Código Glosa | Código Respuesta | Servicio |
    Valor Objetado | Valor Aceptado | Detalle Respuesta

Ese Excel es el INPUT de `responder_glosas_mutual_ser.py`.

DIFERENCIAS CON extraer_respuestas_glosa.py (SIMED/Dispensario)
--------------------------------------------------------------
El PDF de MUTUAL SER trae la tabla "CONCEPTO OBJECIÓN | CONCEPTO RESPUESTA | SERVICIO
| VALOR OBJETADO | VALOR ACEPTADO", donde en cada ítem conviven DOS códigos:
  - el código de la GLOSA (p.ej. CL0101, TA0201, FA0802) y
  - el código de la RESPUESTA (RE9901).
pdfplumber los parte por el ancho de columna: el 6º dígito del código cae al inicio
de la línea siguiente ("CL010" + "1" = CL0101; "RE990" + "1" = RE9901). Este
extractor reconstruye el código de 6 caracteres y captura además el código de
respuesta, que el extractor de SIMED no necesitaba.

El texto de la respuesta ("Observaciones:") suele ser idéntico para todos los ítems
en una glosa RATIFICADA (rechazo total con solicitud de conciliación) — el bot lo
agrupa para responder en bloque.

USO RÁPIDO
----------

    # Un PDF
    py extraer_respuestas_glosa_mutualser.py ^
        --pdf    "D:\\...\\HUS0000492542.pdf" ^
        --salida "D:\\...\\respuestas_mutualser.xlsx"

    # Una carpeta con varios PDFs (uno por factura)
    py extraer_respuestas_glosa_mutualser.py ^
        --carpeta "D:\\...\\MUTUALSER\\SOPORTES" ^
        --salida  "D:\\...\\respuestas_mutualser.xlsx"

DEPENDENCIAS
------------
    py -m pip install pdfplumber openpyxl
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from pathlib import Path

logger = logging.getLogger("extraer_respuestas_mutualser")

# Cabecera del PDF.
RE_FACTURA = re.compile(r"FACTURA:\s*(HUS\d{6,12})", re.IGNORECASE)
RE_NUM_OBJECION_DOC = re.compile(r"N[°º]?\s*OBJECION:\s*(\d+)", re.IGNORECASE)
RE_CONTRATO = re.compile(r"CONTRATO:\s*([^\n]+?)\s+PLAN:", re.IGNORECASE)
RE_TIPO = re.compile(
    r"Glosa_?(Ratificada|Inicial)?\s*(Mixta|Administrativo|Asistencial)?", re.IGNORECASE
)

# Bloque de UNA objeción: arranca con el código de glosa (2 letras + 3 dígitos) al
# inicio de línea, agarra el PRIMER par de valores "$ v_obj $ v_ace" (el código de
# respuesta RE990 queda dentro del tramo capturado), todo lo del medio (incluye el
# servicio) y el texto tras "Observaciones:" hasta la próxima objeción o TOTALES/fin.
# NOTA: no se exige RE990 en la cabecera — el layout de pdfplumber lo parte en varias
# líneas y exigirlo descartaba objeciones (se verifica sumando contra el TOTAL oficial).
RE_BLOQUE = re.compile(
    r"^(?P<cod>[A-Z]{2}\d{3})"
    r".*?\$\s*(?P<v_obj>[\d.,]+)\s*\$\s*(?P<v_ace>[\d.,]+)"
    r"(?P<medio>.*?)Observaciones:\s*(?P<obs>.*?)"
    r"(?=^[A-Z]{2}\d{3}|^TOTALES|\Z)",
    re.MULTILINE | re.DOTALL,
)

# Línea de servicio dentro del bloque: "734004 - LABORATORIO - HEMATOLOGIA".
RE_SERVICIO = re.compile(r"(\d{4,7}\s*-\s*[A-ZÁÉÍÓÚÑ].*?)(?=\n|Observaciones)")

# Código de respuesta completo (RE990 + versión).
RE_COD_RESP = re.compile(r"\b(RE\d{3})\b")

# Marcadores de pie/encabezado de página que ensucian el texto de "Observaciones:"
# cuando una objeción cruza un salto de página. Se corta el obs en el primero que
# aparezca para quedarnos con la respuesta canónica limpia.
RE_PIE_PAGINA = re.compile(
    r"\s*(?:"
    r"Nombre\s+reporte"
    r"|LICENCIADO\s+A:"
    r"|Fecha\s+Actual\s*:"
    r"|EMPRESA\s+SOCIAL\s+DEL\s+ESTADO\s+HOSPITAL"
    r"|Glosa_?(?:Ratificada|Inicial)"
    r"|Elabor[oó]\s*:"
    r"|CONCEPTO\s+OBJECI[OÓ]N\s+CONCEPTO\s+RESPUESTA"
    r"|\d{1,3}/\d{1,3}\b"
    r")",
    re.IGNORECASE,
)


def _limpiar_obs(obs: str) -> str:
    """Corta el texto de Observaciones en el primer marcador de pie/encabezado de
    página (aparece cuando la objeción cruza un salto de página) y colapsa espacios,
    para que todos los ítems con la misma respuesta queden con texto idéntico."""
    obs = re.sub(r"\s+", " ", obs).strip()
    m = RE_PIE_PAGINA.search(obs)
    if m:
        obs = obs[: m.start()].strip()
    return obs


def setup_logging() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")


def _to_int(s: str) -> int:
    """'1.706.000,00' (formato Colombia) -> 1706000. Ignora decimales (glosas enteras)."""
    parte_entera = s.split(",")[0]
    return int(re.sub(r"\.", "", parte_entera) or "0")


def extraer_texto_pdf(ruta: Path) -> str:
    try:
        import pdfplumber
    except ImportError:
        sys.stderr.write("ERROR: falta pdfplumber.\nInstalalo con:  py -m pip install pdfplumber\n")
        sys.exit(2)
    with pdfplumber.open(str(ruta)) as pdf:
        return "\n".join((p.extract_text() or "") for p in pdf.pages)


# Línea de continuación de la descripción de la objeción: empieza con el 6º dígito
# del código (versión) seguido de la descripción en texto — p.ej. "1 observación...",
# "2 apoyos diagnósticos...". NO matchea la línea de servicio ("730102 - URGENCIAS":
# no hay espacio tras el primer dígito) ni las líneas de la respuesta (sin dígito
# inicial). El `[(]` cubre casos como "1 (HEMOGLOBINA".
RE_VERSION_OBJ = re.compile(r"^\s*(\d)\s+[A-Za-zÁÉÍÓÚÑ(]", re.MULTILINE)


def _completar_codigo(cod5: str, bloque: str) -> str:
    """Reconstruye el código de 6 caracteres. El 6º dígito (versión) cae al inicio
    de alguna línea de continuación de la descripción de la objeción; pdfplumber a
    veces intercala primero la descripción de la respuesta, así que se busca en todo
    el bloque (saltando la cabecera) la primera línea 'dígito + texto'."""
    nl = bloque.find("\n")
    cuerpo = bloque[nl + 1 :] if nl != -1 else ""
    m = RE_VERSION_OBJ.search(cuerpo)
    if m:
        return cod5 + m.group(1)
    return cod5


def _codigo_respuesta(bloque: str, default: str = "") -> str:
    """Código de respuesta del ítem (RE9901 en glosa ratificada). Si en el bloque no
    quedó capturado 'RE990' (pdfplumber lo partió a otra línea), usa el `default`
    dominante del documento."""
    if "RE990" in bloque:
        return "RE9901"
    return default


def extraer_objeciones(ruta: Path) -> list[dict]:
    texto = extraer_texto_pdf(ruta)

    m_fac = RE_FACTURA.search(texto)
    factura = m_fac.group(1) if m_fac else ruta.stem.upper()

    m_obj_doc = RE_NUM_OBJECION_DOC.search(texto)
    num_obj_doc = m_obj_doc.group(1) if m_obj_doc else ""

    m_contrato = RE_CONTRATO.search(texto)
    contrato = m_contrato.group(1).strip() if m_contrato else ""

    # Código de respuesta dominante del documento (RE9901 en glosa ratificada), para
    # rellenar los bloques donde pdfplumber separó 'RE990' fuera del tramo capturado.
    resp_default = "RE9901" if "RE990" in texto else ""

    objeciones: list[dict] = []
    for m in RE_BLOQUE.finditer(texto):
        bloque = m.group(0)
        srv_match = RE_SERVICIO.search(m.group("medio"))
        objeciones.append(
            {
                "factura": factura,
                "num_objecion_doc": num_obj_doc,
                "contrato": contrato,
                "num": len(objeciones) + 1,
                "cod_glosa": _completar_codigo(m.group("cod"), bloque),
                "cod_respuesta": _codigo_respuesta(bloque, resp_default),
                "valor_objetado": _to_int(m.group("v_obj")),
                "valor_aceptado": _to_int(m.group("v_ace")),
                "servicio": (srv_match.group(1).strip() if srv_match else ""),
                "observaciones": _limpiar_obs(m.group("obs")),
            }
        )
    return objeciones


def escribir_excel(filas: list[dict], salida: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        sys.stderr.write("ERROR: falta openpyxl. Instalalo con: py -m pip install openpyxl\n")
        sys.exit(2)

    wb = Workbook()
    ws = wb.active
    ws.title = "Respuestas Glosa"

    encabezados = [
        ("Factura", 18),
        ("# Objeción", 11),
        ("Código Glosa", 13),
        ("Código Respuesta", 16),
        ("Servicio", 40),
        ("Valor Objetado", 16),
        ("Valor Aceptado", 16),
        ("Detalle Respuesta", 100),
    ]
    fill_hdr = PatternFill("solid", fgColor="1F4E78")
    font_hdr = Font(bold=True, color="FFFFFF")
    for col, (etiq, ancho) in enumerate(encabezados, start=1):
        c = ws.cell(row=1, column=col, value=etiq)
        c.fill = fill_hdr
        c.font = font_hdr
        ws.column_dimensions[c.column_letter].width = ancho

    for r, fila in enumerate(filas, start=2):
        ws.cell(row=r, column=1, value=fila["factura"])
        ws.cell(row=r, column=2, value=fila["num"])
        ws.cell(row=r, column=3, value=fila["cod_glosa"])
        ws.cell(row=r, column=4, value=fila["cod_respuesta"])
        ws.cell(row=r, column=5, value=fila["servicio"])
        ws.cell(row=r, column=6, value=fila["valor_objetado"]).number_format = "#,##0"
        ws.cell(row=r, column=7, value=fila["valor_aceptado"]).number_format = "#,##0"
        det = ws.cell(row=r, column=8, value=fila["observaciones"])
        det.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    wb.save(str(salida))


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Pre-rellena Excel de respuestas a glosas MUTUAL SER desde el PDF Trámite de Objeción.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    grupo = parser.add_mutually_exclusive_group(required=True)
    grupo.add_argument("--pdf", type=Path, help="Un solo PDF de Trámite de Objeción.")
    grupo.add_argument(
        "--carpeta", type=Path, help="Carpeta con varios PDFs (cada uno = una factura)."
    )
    parser.add_argument("--salida", type=Path, required=True, help="Excel .xlsx de salida.")
    args = parser.parse_args()
    setup_logging()

    if args.pdf:
        if not args.pdf.is_file():
            logger.error(f"No existe el PDF: {args.pdf}")
            return 1
        pdfs = [args.pdf]
    else:
        if not args.carpeta.is_dir():
            logger.error(f"No es una carpeta: {args.carpeta}")
            return 1
        pdfs = sorted(p for p in args.carpeta.iterdir() if p.suffix.lower() == ".pdf")
        if not pdfs:
            logger.error(f"La carpeta no tiene PDFs: {args.carpeta}")
            return 1

    todas: list[dict] = []
    for p in pdfs:
        try:
            obj = extraer_objeciones(p)
        except Exception as e:
            logger.error(f"{p.name}: ERROR — {e}")
            continue
        if not obj:
            logger.warning(f"{p.name}: 0 objeciones (¿formato distinto?)")
            continue
        fac = obj[0]["factura"]
        total_obj = sum(o["valor_objetado"] for o in obj)
        total_ace = sum(o["valor_aceptado"] for o in obj)
        logger.info(
            f"{p.name} → {fac}: {len(obj)} objeciones | objetado ${total_obj:,} | aceptado ${total_ace:,}"
        )
        todas.extend(obj)

    if not todas:
        logger.error("No se extrajo ninguna objeción.")
        return 1

    args.salida.parent.mkdir(parents=True, exist_ok=True)
    escribir_excel(todas, args.salida)
    logger.info(f"\nExcel generado: {args.salida}")
    logger.info(f"  Facturas: {len({o['factura'] for o in todas})}")
    logger.info(f"  Objeciones: {len(todas)}")
    logger.info("\nProximo paso:")
    logger.info("  1) Abrí el Excel y revisá 'Valor Aceptado' y 'Detalle Respuesta'.")
    logger.info("  2) Corré responder_glosas_mutual_ser.py --excel <ese.xlsx>.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
