"""expediente_conciliacion.py — Modelo de datos UNICO de la conciliacion.

Este es el nucleo de la plataforma: un EXPEDIENTE por factura, con un
identificador unico (`id_expediente`) que relaciona TODO — factura, paciente,
contrato, radicado, glosas, RIPS, historia clinica, XML, cartera, pagos,
respuestas, oficios y actas. Todos los demas modulos (evidencia, juridico,
argumentacion, dashboard, aprendizaje) LEEN Y ESCRIBEN sobre este expediente,
en vez de repetir busquedas sobre `Y:\\` o reinterpretar cada Excel.

Fuentes que ensambla:
- HUS.xlsx (glosas del Dispensario)         -> via asistente_conciliacion_dispensario.cargar_glosas
- indice de soportes (JSON)                 -> via indexar_soportes_dispensario
- Estado de cartera (xlsx)                  -> saldo / edad / deterioro por factura
- Contrato aplicable por fecha de atencion  -> 287 / 440 + base tarifaria

USO
---
    py tools\\expediente_conciliacion.py ^
        --excel   "D:\\...\\HUS.xlsx" ^
        --indice  "D:\\...\\indice_soportes.json" ^
        --cartera "D:\\...\\Estado_Cartera_JUN_2026.xlsx" ^
        --salida  "D:\\...\\expedientes.json"

El JSON resultante es la "fuente de la verdad" del lote.
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path

_TOOLS = Path(__file__).resolve().parent
if str(_TOOLS) not in sys.path:
    sys.path.insert(0, str(_TOOLS))
from asistente_conciliacion_dispensario import (  # noqa: E402
    cargar_glosas,
    codigo_corto,
    contrato_por_fecha,
    factura_corta,
)

logger = logging.getLogger("expediente")

BASE_TARIFARIA = {
    "287": "SOAT 2025 -15% + tarifas institucionales/propias",
    "440": "SOAT SMLV -20% + tarifas PROPIAS + regulados (Circ.19/2024)",
    "?": "Contrato por confirmar",
}


# ---------------------------------------------------------------------------
# Modelo (dataclasses -> JSON con asdict)
# ---------------------------------------------------------------------------
@dataclass
class Paciente:
    nombre: str = ""
    identificacion: str = ""


@dataclass
class GlosaExp:
    num: str = ""
    codigo: str = ""
    codigo_corto: str = ""
    familia: str = ""
    motivo: str = ""
    valor_objetado: int = 0
    servicio: str = ""
    cod_respuesta: str = ""
    respuesta_ips: str = ""


@dataclass
class ContratoExp:
    codigo: str = "?"  # 287 / 440 / ?
    base_tarifaria: str = ""
    codigo_cartera: str = ""  # U22031, C26001, ... (codigo interno de cartera)
    pendiente_confirmar: bool = False


@dataclass
class CarteraExp:
    en_cartera: bool = False
    saldo: int = 0
    edad_dias: int = 0
    edad_rango: str = ""
    deterioro: int = 0


@dataclass
class SoporteExp:
    tipo: str = ""
    nombre: str = ""
    ruta: str = ""


@dataclass
class RespuestaExp:
    """La llenan los motores (evidencia/juridico/argumentacion)."""

    glosa_codigo: str = ""
    conclusion: str = ""  # Procede / No procede / Parcial / Insuficiente
    confianza: int = 0
    norma: str = ""
    oficio_ruta: str = ""


@dataclass
class Expediente:
    id_expediente: str
    factura: str
    factura_corta: str
    radicado: str = ""
    fecha_atencion: str = ""  # ISO date
    fecha_radicacion: str = ""
    mes_radicacion: str = ""
    paciente: Paciente = field(default_factory=Paciente)
    valor_factura: int = 0
    valor_objetado_total: int = 0
    contrato: ContratoExp = field(default_factory=ContratoExp)
    cartera: CarteraExp = field(default_factory=CarteraExp)
    glosas: list[GlosaExp] = field(default_factory=list)
    soportes: list[SoporteExp] = field(default_factory=list)
    respuestas: list[RespuestaExp] = field(default_factory=list)
    actas: list[dict] = field(default_factory=list)
    pagos: list[dict] = field(default_factory=list)
    estado: str = "PENDIENTE"  # PENDIENTE / EN_ANALISIS / CONCILIADO / RESPONDIDO / PAGADO


def _iso(x) -> str:
    if isinstance(x, (datetime, date)):
        return x.date().isoformat() if isinstance(x, datetime) else x.isoformat()
    return ""


def _int(x) -> int:
    try:
        return int(float(x))
    except (TypeError, ValueError):
        return 0


# ---------------------------------------------------------------------------
# Cartera: extrae saldo/edad/deterioro por factura del Estado de Cartera xlsx
# ---------------------------------------------------------------------------
def cartera_desde_xlsx(ruta: Path, facturas: set[str]) -> dict[str, dict]:
    """Devuelve dict factura_corta -> {saldo, edad, edad_rango, deterioro, contrato}.

    El archivo trae un titulo en la fila 1 y los encabezados en la fila 2.
    Columnas (0-based): 4 CONTRATO, 6 FACTURA, 7 SALDO, 10 VALFAC, 28 EDAD,
    30 AGRUPADO EDAD, 33 DETERIORO.
    """
    import openpyxl

    out: dict[str, dict] = {}
    wb = openpyxl.load_workbook(str(ruta), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    next(it, None)  # titulo
    next(it, None)  # encabezados
    for r in it:
        fac = r[6] if len(r) > 6 else None
        if not fac:
            continue
        corta = factura_corta(fac)
        if corta in facturas and corta not in out:
            out[corta] = {
                "contrato": str(r[4] or ""),
                "saldo": _int(r[7]),
                "edad": _int(r[28]) if len(r) > 28 else 0,
                "edad_rango": str(r[30]).strip() if len(r) > 30 and r[30] else "",
                "deterioro": _int(r[33]) if len(r) > 33 else 0,
            }
    wb.close()
    return out


def soportes_por_factura(indice: dict) -> dict[str, list[dict]]:
    out: dict[str, list[dict]] = {}
    for d in indice.get("documentos", []):
        out.setdefault(str(d.get("factura", "")), []).append(d)
    return out


# ---------------------------------------------------------------------------
# Ensamblado
# ---------------------------------------------------------------------------
def construir_expedientes(glosas, indice: dict, cartera: dict) -> list[Expediente]:
    """Agrupa las glosas por factura y arma un Expediente unificado por cada una."""
    por_factura: dict[str, list] = {}
    for g in glosas:
        por_factura.setdefault(g.factura, []).append(g)

    sop = soportes_por_factura(indice)
    expedientes: list[Expediente] = []
    for factura, gs in por_factura.items():
        corta = factura_corta(factura)
        g0 = gs[0]
        contr = contrato_por_fecha(g0.fecha_atencion)
        cart = cartera.get(corta, {})
        exp = Expediente(
            id_expediente=factura,  # la factura normalizada ES el id unico
            factura=factura,
            factura_corta=corta,
            radicado=str(g0.radicado),
            fecha_atencion=_iso(g0.fecha_atencion),
            paciente=Paciente(nombre=g0.paciente, identificacion=g0.identificacion),
            valor_factura=_int(g0.valor_factura),
            valor_objetado_total=sum(_int(g.valor_objetado) for g in gs),
            contrato=ContratoExp(
                codigo=contr,
                base_tarifaria=BASE_TARIFARIA.get(contr, ""),
                codigo_cartera=cart.get("contrato", ""),
                pendiente_confirmar=(contr == "287"),  # falta acta de inicio del 287
            ),
            cartera=CarteraExp(
                en_cartera=corta in cartera,
                saldo=cart.get("saldo", 0),
                edad_dias=cart.get("edad", 0),
                edad_rango=cart.get("edad_rango", ""),
                deterioro=cart.get("deterioro", 0),
            ),
            glosas=[
                GlosaExp(
                    num=str(g.num_glosa),
                    codigo=g.codigo,
                    codigo_corto=codigo_corto(g.codigo),
                    familia=g.familia,
                    motivo=g.motivo,
                    valor_objetado=_int(g.valor_objetado),
                    servicio=g.servicio,
                    cod_respuesta=str(getattr(g, "codrta", "") or ""),
                    respuesta_ips=str(getattr(g, "rta_ips", "") or ""),
                )
                for g in gs
            ],
            soportes=[
                SoporteExp(
                    tipo=d.get("tipo", ""), nombre=d.get("nombre", ""), ruta=d.get("ruta", "")
                )
                for d in sop.get(corta, [])
            ],
        )
        expedientes.append(exp)
    return expedientes


def guardar_expedientes(expedientes: list[Expediente], salida: Path) -> None:
    salida.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "generado": datetime.now().isoformat(timespec="seconds"),
        "total": len(expedientes),
        "expedientes": [asdict(e) for e in expedientes],
    }
    salida.write_text(json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")


def cargar_expedientes(salida: Path) -> dict:
    if salida and Path(salida).exists():
        try:
            return json.loads(Path(salida).read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return {}
    return {}


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main(argv: list[str] | None = None) -> int:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
    ap = argparse.ArgumentParser(
        description="Arma el expediente unico de conciliacion por factura."
    )
    ap.add_argument("--excel", type=Path, required=True, help="HUS.xlsx (glosas del Dispensario)")
    ap.add_argument("--indice", type=Path, default=None, help="Indice de soportes (JSON)")
    ap.add_argument("--cartera", type=Path, default=None, help="Estado de cartera (xlsx)")
    ap.add_argument("--salida", type=Path, required=True, help="expedientes.json de salida")
    ap.add_argument("--solo", default=None, help="Una sola factura (piloto)")
    args = ap.parse_args(argv)

    glosas = cargar_glosas(args.excel, args.solo)
    if not glosas:
        logger.error("No se cargaron glosas del Excel.")
        return 1
    facturas = {factura_corta(g.factura) for g in glosas}

    indice = {}
    if args.indice:
        indice = cargar_expedientes(args.indice)  # mismo formato JSON generico
    cartera = {}
    if args.cartera:
        logger.info("Leyendo cartera de %s ...", args.cartera)
        cartera = cartera_desde_xlsx(args.cartera, facturas)

    expedientes = construir_expedientes(glosas, indice, cartera)
    guardar_expedientes(expedientes, args.salida)

    from collections import Counter

    logger.info("Expedientes: %d  ->  %s", len(expedientes), args.salida)
    logger.info("  Por contrato: %s", dict(Counter(e.contrato.codigo for e in expedientes)))
    logger.info(
        "  En cartera: %d / %d", sum(e.cartera.en_cartera for e in expedientes), len(expedientes)
    )
    con_sop = sum(1 for e in expedientes if e.soportes)
    logger.info("  Con soportes indexados: %d / %d", con_sop, len(expedientes))
    return 0


if __name__ == "__main__":
    sys.exit(main())
