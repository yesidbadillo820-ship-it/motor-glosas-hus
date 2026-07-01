"""responder_glosas_mutual_ser.py — Respuesta a glosas en el portal MUTUAL SER (Zona Ser).

Portal: https://portalzonaser.mutualser.com
Módulo: AUDITORIA DE CUENTAS MEDICAS → GESTIÓN DE RESPUESTAS DE GLOSAS →
        CONSULTAR CUENTAS MÉDICAS GLOSADAS

Bot Playwright que carga las respuestas a glosas del HUS (una fila por objeción,
generadas por `extraer_respuestas_glosa_mutualser.py`) sin intervención humana,
siguiendo el mismo patrón que responder_glosas_coosalud.py / _simed.py.

⚠ ESTADO: v1 — IMPLEMENTADO (selectores calibrados con el volcado real), PENDIENTE
    de PILOTO. El portal es React/MUI; los selectores usan texto/rol + posición de
    celda (las clases/ids son inestables). Flujo de SUBSANACIÓN implementado:
    abrir factura → SUBSANAR GLOSA → por cada ítem: expandir "+" (col. TECNOLOGÍA),
    poner VALOR RATIFICADO ACEPTADO IPS = 0, escribir OBSERVACIÓN (modal ≤1000) →
    ACEPTAR TOTAL RATIFICADO. El soporte PDF es opcional y v1 lo omite.
    SEGURIDAD: sin --finalizar el bot SOLO llena (no envía); --max-items N limita a
    N ítems para pilotear. Modo --explorar sigue disponible para re-calibrar.

CREDENCIALES (variables de entorno, NO en el código):
    setx MUTUALSER_USER  gerencia@hus.gov.co
    setx MUTUALSER_PASSWORD  <contraseña>
    (cerrar y reabrir la terminal para que tomen efecto)

INSTALACIÓN (una vez):
    py -m pip install playwright openpyxl
    py -m playwright install chromium

reCAPTCHA — IMPORTANTE:
    El login tiene reCAPTCHA y NO valida en un navegador automatizado (Playwright).
    Solución recomendada: abrí TU Chrome con puerto de depuración, logueate a mano
    (ahí el captcha sí pasa) y el bot se CONECTA a ese Chrome con --cdp:

    REM (1 vez) cerrá Chrome y abrilo con depuración + perfil dedicado:
    "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" ^
        --remote-debugging-port=9222 --user-data-dir="C:\\temp-notas\\zonaser-chrome"
    REM en ese Chrome: entrá a https://portalzonaser.mutualser.com, logueate, abrí el módulo.

USO:
    REM 1) Explorar/calibrar conectándose a tu Chrome (recomendado):
    py responder_glosas_mutual_ser.py --explorar --cdp http://localhost:9222

    REM 1b) Explorar lanzando el browser (si el captcha te deja):
    py responder_glosas_mutual_ser.py --explorar --con-cabeza

    REM 2) PILOTO SEGURO: llena SOLO 1 ítem y NO envía (para revisar en pantalla):
    py responder_glosas_mutual_ser.py --excel respuestas_mutualser.xlsx ^
        --solo HUS0000510639 --cdp http://127.0.0.1:9222 --max-items 1

    REM 3) Factura completa, llena todo y ENVÍA (ACEPTAR TOTAL RATIFICADO):
    py responder_glosas_mutual_ser.py --excel respuestas_mutualser.xlsx ^
        --solo HUS0000510639 --cdp http://127.0.0.1:9222 --finalizar

    REM 4) Masivo (todas), enviando:
    py responder_glosas_mutual_ser.py --excel respuestas_mutualser.xlsx --todas ^
        --cdp http://127.0.0.1:9222 --finalizar --reporte reporte_mutualser.csv
"""

from __future__ import annotations

import argparse
import contextlib
import csv
import logging
import os
import re
import sys
import time
import unicodedata
from pathlib import Path

try:
    from playwright.sync_api import (
        Page,
        sync_playwright,
    )
    from playwright.sync_api import (
        TimeoutError as PlaywrightTimeout,
    )
except ImportError:
    Page = object  # type: ignore[assignment,misc]
    PlaywrightTimeout = Exception  # type: ignore[assignment,misc]
    sync_playwright = None  # type: ignore[assignment]


def _exigir_playwright() -> None:
    if sync_playwright is None:
        sys.stderr.write(
            "ERROR: falta playwright.\n"
            "    py -m pip install playwright\n"
            "    py -m playwright install chromium\n"
        )
        sys.exit(2)


# ─── Constantes del portal ───────────────────────────────────────────────────

PORTAL_BASE = "https://portalzonaser.mutualser.com"
PORTAL_LOGIN = f"{PORTAL_BASE}/auth/login"
PORTAL_DASHBOARD = f"{PORTAL_BASE}/dashboard"
PORTAL_MODULO = (
    f"{PORTAL_BASE}/dashboard/applications/auditoria-de-cuentas-medicas/GESTION-RESPUESTAS-GLOSAS"
)

# Señal de login exitoso: textos/URL que sólo aparecen ya autenticado.
LOGIN_OK_TEXTOS = ("Inicio", "AUDITORIA DE CUENTAS MEDICAS", "GESTIÓN DE RESPUESTAS")

STORAGE_STATE_DEFAULT = "mutualser_session.json"
MAX_PDF_MB = 10

logger = logging.getLogger("responder_mutualser")


# ─── Setup ───────────────────────────────────────────────────────────────────


def setup_logging(log_file: Path | None = None) -> None:
    fmt = "%(asctime)s [%(levelname)s] %(message)s"
    handlers: list[logging.Handler] = [logging.StreamHandler(sys.stdout)]
    if log_file is not None:
        log_file.parent.mkdir(parents=True, exist_ok=True)
        handlers.append(logging.FileHandler(log_file, encoding="utf-8"))
    logging.basicConfig(level=logging.INFO, format=fmt, handlers=handlers)


def cargar_credenciales() -> tuple[str, str]:
    user = os.environ.get("MUTUALSER_USER", "").strip()
    password = os.environ.get("MUTUALSER_PASSWORD", "").strip()
    if not user or not password:
        sys.stderr.write(
            "ERROR: faltan credenciales. Setealas con:\n"
            "    setx MUTUALSER_USER <correo>\n"
            "    setx MUTUALSER_PASSWORD <contraseña>\n"
            "Despues cerra y reabri la terminal.\n"
        )
        sys.exit(2)
    return user, password


# ─── Lectura del Excel de respuestas (salida del extractor MUTUAL SER) ────────


def _norm_header(h: str) -> str:
    s = unicodedata.normalize("NFKD", (h or "").strip().upper())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.split())


# Alias tolerantes por columna (acepta la salida del extractor y variantes manuales).
COLUMNAS = {
    "factura": {"FACTURA", "# FACTURA", "NUMERO FACTURA", "NUMERO_FACTURA"},
    "num": {"# OBJECION", "NUM OBJECION", "OBJECION", "ITEM", "# OBJECIÓN"},
    "cod_glosa": {"CODIGO GLOSA", "COD GLOSA", "CODIGO_GLOSA", "COD."},
    "cod_rta": {"CODIGO RESPUESTA", "COD RESPUESTA", "COD RESPUESTA GLOSA"},
    "servicio": {"SERVICIO"},
    "aceptado": {"VALOR ACEPTADO", "ACEPTADO"},
    "objetado": {"VALOR OBJETADO", "OBJETADO"},
    "detalle": {"DETALLE RESPUESTA", "OBSERVACION RTA GLOSA", "OBSERVACIONES", "DETALLE"},
}


def normalizar_factura(f: str) -> str:
    """'HUS0000492542' -> '492542' (sin prefijo HUS ni ceros a la izquierda)."""
    s = re.sub(r"(?i)^hus", "", (f or "").strip())
    s = s.lstrip("0")
    return s or "0"


def _to_int(s) -> int:
    if s is None:
        return 0
    if isinstance(s, (int, float)):
        return int(s)
    txt = str(s).split(",")[0]
    return int(re.sub(r"[^\d]", "", txt) or "0")


def leer_excel_respuestas(ruta: Path) -> dict[str, dict]:
    """Devuelve {factura: {"items": [...], "grupos": [...]}} donde cada item es
    {num, cod_glosa, cod_rta, servicio, aceptado, detalle} y cada grupo agrupa los
    items por (cod_rta, detalle) idénticos (respuesta masiva)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.stderr.write("ERROR: falta openpyxl. Instalalo con: py -m pip install openpyxl\n")
        sys.exit(2)

    wb = load_workbook(ruta, data_only=True, read_only=True)
    ws = wb.active
    filas = ws.iter_rows(values_only=True)
    encabezados = [_norm_header(str(c)) for c in next(filas)]

    idx: dict[str, int] = {}
    for clave, alias in COLUMNAS.items():
        for i, h in enumerate(encabezados):
            if h in alias:
                idx[clave] = i
                break

    faltan = [k for k in ("factura", "detalle") if k not in idx]
    if faltan:
        sys.stderr.write(
            f"ERROR: el Excel no tiene columnas {faltan}. Encabezados: {encabezados}\n"
        )
        sys.exit(2)

    por_factura: dict[str, dict] = {}
    for fila in filas:
        if not fila:
            continue
        fac_raw = fila[idx["factura"]] if idx.get("factura") is not None else None
        detalle = fila[idx["detalle"]] if idx.get("detalle") is not None else None
        if not fac_raw or not detalle:
            continue
        fac = str(fac_raw).strip().upper()
        item = {
            "num": _to_int(fila[idx["num"]]) if "num" in idx else 0,
            "cod_glosa": str(fila[idx["cod_glosa"]]).strip()
            if "cod_glosa" in idx and fila[idx["cod_glosa"]]
            else "",
            "cod_rta": str(fila[idx["cod_rta"]]).strip()
            if "cod_rta" in idx and fila[idx["cod_rta"]]
            else "",
            "servicio": str(fila[idx["servicio"]]).strip()
            if "servicio" in idx and fila[idx["servicio"]]
            else "",
            "aceptado": _to_int(fila[idx["aceptado"]]) if "aceptado" in idx else 0,
            "objetado": _to_int(fila[idx["objetado"]]) if "objetado" in idx else 0,
            "detalle": str(detalle).strip(),
        }
        por_factura.setdefault(fac, {"items": []})["items"].append(item)

    # Agrupar por (cod_rta, detalle) idénticos — en glosa ratificada es un solo grupo.
    for datos in por_factura.values():
        grupos: dict[tuple, dict] = {}
        for it in datos["items"]:
            clave = (it["cod_rta"], it["detalle"])
            g = grupos.setdefault(
                clave, {"cod_rta": it["cod_rta"], "detalle": it["detalle"], "items": []}
            )
            g["items"].append(it)
        datos["grupos"] = list(grupos.values())
    return por_factura


def sanitizar(texto: str) -> str:
    """Deja SOLO [A-Za-z0-9] y espacios, translitera tildes/ñ. Se usa SOLO si el
    portal de MUTUAL SER rechaza caracteres especiales (a confirmar en --explorar;
    por defecto el bot manda el texto tal cual)."""
    s = unicodedata.normalize("NFKD", texto or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Za-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


# ─── Login con reCAPTCHA (sesión persistida) ─────────────────────────────────


def _login_ok(page: Page) -> bool:
    try:
        if "/auth/login" in page.url:
            return False
        for t in LOGIN_OK_TEXTOS:
            if page.get_by_text(t, exact=False).count() > 0:
                return True
        return "/dashboard" in page.url
    except Exception:
        return False


def login_interactivo(page: Page, user: str, password: str, timeout_captcha_s: int = 240) -> None:
    """Login que rellena usuario/clave y ESPERA a que un humano resuelva el
    reCAPTCHA y entre. Al detectar la sesión activa, el llamador guarda el
    storage_state para reusarlo sin captcha en corridas posteriores."""
    logger.info("Abriendo login de MUTUAL SER…")
    page.goto(PORTAL_LOGIN, wait_until="domcontentloaded")
    if _login_ok(page):
        logger.info("Sesión ya activa.")
        return

    # Rellenar credenciales (selectores tolerantes).
    try:
        page.locator(
            "input[type=email], input[name*=correo i], input[name*=email i], input[type=text]"
        ).first.fill(user)
        page.locator("input[type=password]").first.fill(password)
        logger.info("Usuario y contraseña rellenados.")
    except Exception as e:
        logger.warning(f"No pude rellenar credenciales automáticamente: {e}")

    logger.warning(
        "⏳ RESOLVÉ EL reCAPTCHA y hacé clic en INGRESAR en la ventana del browser. "
        f"Espero hasta {timeout_captcha_s}s a que entres…"
    )
    t0 = time.time()
    while time.time() - t0 < timeout_captcha_s:
        if _login_ok(page):
            logger.info("✅ Login exitoso (sesión activa).")
            return
        time.sleep(2)
    raise RuntimeError("No se detectó login tras esperar el captcha (timeout).")


def abrir_sesion(p, args, user: str, password: str):
    """Abre/conecta el browser. Devuelve (browser, ctx, page, es_cdp).

    MODO --cdp URL (recomendado por el reCAPTCHA): se CONECTA a un Chrome real que
    el usuario abrió con --remote-debugging-port y donde YA inició sesión a mano. El
    captcha lo resuelve el usuario en su Chrome normal, así reCAPTCHA nunca ve un
    navegador automatizado (que es lo que lo hace fallar). El bot sólo maneja la
    pestaña ya autenticada. NO cierra ese Chrome al terminar.

    MODO default: lanza Chromium/Chrome controlado por Playwright y reutiliza
    --storage-state si existe; si no, login interactivo. OJO: en este modo el
    reCAPTCHA suele NEGARSE a validar (navegador detectado como automatizado) — usá
    --cdp si el captcha no pasa.
    """
    if args.cdp:
        logger.info(f"Conectando a tu Chrome vía CDP: {args.cdp}")
        # En Windows 'localhost' suele resolver a IPv6 (::1) y Chrome escucha en IPv4
        # (127.0.0.1) -> ECONNREFUSED. Probar variantes IPv4 automáticamente.
        candidatos = [args.cdp]
        for alias in ("localhost", "::1", "0.0.0.0"):
            if alias in args.cdp:
                candidatos.append(args.cdp.replace(alias, "127.0.0.1"))
        browser = None
        ultimo_error: Exception | None = None
        for url in dict.fromkeys(candidatos):  # dedup preservando orden
            try:
                browser = p.chromium.connect_over_cdp(url)
                if url != args.cdp:
                    logger.info(f"Conectado usando {url} (fallback IPv4).")
                break
            except Exception as e:  # noqa: BLE001
                ultimo_error = e
        if browser is None:
            raise RuntimeError(
                f"No pude conectarme a Chrome en {args.cdp} (ni por 127.0.0.1). Abrí "
                "Chrome con --remote-debugging-port=9222 y "
                '--user-data-dir="C:\\temp-notas\\zonaser-chrome", verificá '
                f"http://127.0.0.1:9222/json/version, y reintentá. Detalle: {ultimo_error}"
            ) from ultimo_error
        ctx = browser.contexts[0] if browser.contexts else browser.new_context()
        page = ctx.pages[0] if ctx.pages else ctx.new_page()
        page.set_default_navigation_timeout(120000)
        page.set_default_timeout(30000)
        if not _login_ok(page):
            logger.warning(
                "No detecto sesión activa en ese Chrome. En esa ventana: iniciá "
                "sesión en el portal (resolvé el captcha) y abrí el módulo / la "
                "factura ANTES de seguir. Igual continúo con la pestaña actual."
            )
        return browser, ctx, page, True

    storage = Path(args.storage_state)
    launch_kwargs = {
        "headless": not args.con_cabeza,
        "slow_mo": 300 if args.lento else 0,
        # Reduce la huella de automatización (ayuda algo, pero el reCAPTCHA igual
        # puede fallar — para eso está --cdp).
        "args": ["--disable-blink-features=AutomationControlled"],
    }
    if args.canal:
        launch_kwargs["channel"] = args.canal
    browser = p.chromium.launch(**launch_kwargs)
    ctx_kwargs = {"accept_downloads": True}
    if storage.is_file():
        logger.info(f"Reutilizando sesión guardada: {storage}")
        ctx_kwargs["storage_state"] = str(storage)
    ctx = browser.new_context(**ctx_kwargs)
    page = ctx.new_page()
    page.set_default_navigation_timeout(120000)
    page.set_default_timeout(30000)
    page.on("dialog", lambda d: d.accept())

    # Validar la sesión reutilizada; si no sirve, re-login interactivo.
    page.goto(PORTAL_MODULO, wait_until="domcontentloaded")
    if not _login_ok(page):
        if not args.con_cabeza:
            raise RuntimeError(
                "Sesión inválida/expirada y estás headless. Corré una vez con "
                "--con-cabeza (o mejor --cdp) para autenticarte y regenerar la sesión."
            )
        login_interactivo(page, user, password)
        ctx.storage_state(path=str(storage))
        logger.info(f"Sesión guardada en {storage} (reutilizable sin captcha).")
        page.goto(PORTAL_MODULO, wait_until="domcontentloaded")
    return browser, ctx, page, False


# ─── Exploración/calibración del DOM (equivalente web de dump_dg.py) ──────────


def explorar(page: Page, salida_dir: Path) -> None:
    """Vuelca la estructura del módulo para calibrar los  # TODO(portal):
    screenshot + inventario de inputs, botones, selects, links y cabeceras de tabla."""
    salida_dir.mkdir(parents=True, exist_ok=True)
    ts = time.strftime("%Y%m%d_%H%M%S")
    page.wait_for_timeout(3000)
    shot = salida_dir / f"explorar_{ts}.png"
    try:
        page.screenshot(path=str(shot), full_page=True)
        logger.info(f"Screenshot: {shot}")
    except Exception as e:
        logger.warning(f"No pude sacar screenshot: {e}")

    dump = page.evaluate(
        """() => {
        const txt = el => (el.innerText || el.value || el.getAttribute('aria-label') || '').trim().slice(0,80);
        // data-testid o class del/los SVG interno(s) (MUI: AddIcon, CloudUploadIcon...).
        const svgInfo = el => Array.from(el.querySelectorAll('svg'))
            .map(s => s.getAttribute('data-testid') || (s.getAttribute('class')||'').slice(0,45))
            .filter(Boolean).join(' , ');
        const desc = el => {
            const a = [];
            for (const at of ['id','name','type','placeholder','role','aria-label','data-testid','title','class']) {
                const v = el.getAttribute(at); if (v) a.push(at+'='+v.slice(0,55));
            }
            const svg = svgInfo(el);
            return el.tagName.toLowerCase()+' {'+a.join(' ')+'}'+(svg?' [svg: '+svg+']':'')+' '+txt(el);
        };
        const pick = (sel,n=100) => Array.from(document.querySelectorAll(sel)).slice(0,n).map(desc);
        const clean = h => (h||'').replace(/\\s+/g,' ');
        // HTML crudo de la 1ª tabla y de cualquier modal/dialog abierto — la fuente de
        // verdad para ver cómo se anidan input de valor y botón de observación.
        const tabla = document.querySelector('table');
        const dialog = document.querySelector('[role=dialog], .MuiDialog-root, .MuiModal-root');
        return {
            url: location.href,
            botones: pick('button, [role=button], input[type=button], input[type=submit], a[href]', 160),
            inputs: pick('input, textarea'),
            selects: pick('select, [role=combobox], mat-select'),
            iconos: Array.from(document.querySelectorAll('svg[data-testid]'))
                .slice(0,200).map(s => s.getAttribute('data-testid')),
            th: Array.from(document.querySelectorAll('th, [role=columnheader]')).slice(0,60).map(e=>txt(e)),
            links: Array.from(document.querySelectorAll('a')).slice(0,80).map(a=>({t:txt(a), href:a.getAttribute('href')})),
            tabla_html: clean(tabla ? tabla.outerHTML : '').slice(0, 40000),
            dialog_html: clean(dialog ? dialog.outerHTML : '').slice(0, 8000),
        };
    }"""
    )
    out = salida_dir / f"explorar_{ts}.txt"
    with out.open("w", encoding="utf-8") as f:
        f.write(f"URL: {dump.get('url')}\n\n")
        for seccion in ("th", "botones", "iconos", "selects", "inputs", "links"):
            f.write(f"===== {seccion.upper()} =====\n")
            for row in dump.get(seccion, []):
                f.write(f"  {row}\n")
            f.write("\n")
        for seccion in ("dialog_html", "tabla_html"):
            html = dump.get(seccion) or ""
            f.write(f"===== {seccion.upper()} ({len(html)} chars) =====\n")
            f.write(html + "\n\n")
    logger.info(f"Volcado del DOM: {out}")
    logger.info(
        "Usá este volcado para completar los  # TODO(portal)  de este script "
        "(grilla, apertura de factura, formulario de respuesta, finalizar)."
    )


# ─── Procesamiento por factura (TODO: calibrar contra el portal) ─────────────


# Índices de columna (0-based) de la tabla "Detalle de respuesta de glosa"
# (24 columnas, confirmados con el volcado --explorar 2026-07-01):
COL_TECNOLOGIA = 2  # celda con el botón "+" que expande/activa la fila
COL_VALOR_ACEPTADO = 19  # VALOR RATIFICADO ACEPTADO IPS (input al activar la fila)
COL_OBSERVACION = 21  # OBSERVACIONES DE SUBSANACIÓN (input + botón libro -> modal)
COL_SOPORTE = 22  # SOPORTE DE SUBSANACIÓN (íconos: ver / subir / cancelar)


def _abrir_factura(page: Page, factura: str) -> None:
    """Va a la grilla y abre la factura por su número corto (ej. 'HUS510639').
    Deja la pantalla en 'Detalle de respuesta de glosa'."""
    page.goto(PORTAL_MODULO, wait_until="domcontentloaded")
    corto = "HUS" + normalizar_factura(factura)  # 'HUS0000510639' -> 'HUS510639'
    logger.info(f"  Buscando {corto} en la grilla…")
    objetivo = page.get_by_text(corto, exact=True)
    objetivo.first.wait_for(timeout=90000)
    objetivo.first.click()
    # Señal de que abrió el detalle: el botón SUBSANAR GLOSA.
    page.get_by_role("button", name="SUBSANAR GLOSA", exact=True).wait_for(timeout=60000)
    logger.info(f"  Detalle de {corto} abierto.")


def _dump_tabla(page: Page, evidencias: Path, nombre: str) -> None:
    """Diagnóstico: vuelca el outerHTML de la 1ª tabla a un archivo para inspeccionar
    la estructura real (clases MUI estables) cuando un selector falla."""
    try:
        evidencias.mkdir(parents=True, exist_ok=True)
        html = page.locator("table").first.evaluate("e => e.outerHTML")
        (evidencias / nombre).write_text(html, encoding="utf-8")
        logger.info(f"    [diag] tabla volcada: {evidencias / nombre}")
    except Exception as e:  # noqa: BLE001
        logger.warning(f"    [diag] no pude volcar la tabla: {e}")


def _plus_items(page: Page):
    """Botones "+" de expandir, uno por ítem: viven en la celda TECNOLOGÍA (la 3ª,
    nth-child(3)) de cada fila de ítem. La sub-fila de detalle que se inserta al
    expandir NO tiene botón en esa celda, así que este selector siempre devuelve
    exactamente los N ítems, en orden."""
    return page.locator("table tbody tr td:nth-child(3) button")


def _fila_activa(page: Page):
    """La sub-fila de detalle activa: la única que tiene el input de valor (adorno)."""
    return page.locator("table tbody tr").filter(
        has=page.locator("input.MuiInputBase-inputAdornedEnd")
    )


def _set_valor(page: Page, valor: int, timeout: int = 12000) -> None:
    # El input de VALOR RATIFICADO ACEPTADO IPS es el único con adorno al final
    # (clase MUI estable, no hasheada). Al expandir un ítem sólo hay uno visible.
    inp = page.locator("table input.MuiInputBase-inputAdornedEnd").first
    inp.wait_for(state="visible", timeout=timeout)
    inp.click()
    inp.fill(str(valor))
    inp.press("Tab")  # dispara la validación (check verde)


def _set_observacion(page: Page, texto: str) -> None:
    # Botón de la celda OBSERVACIONES DE SUBSANACIÓN (col 21) de la fila ACTIVA -> modal.
    fila = _fila_activa(page).first
    fila.locator("td").nth(COL_OBSERVACION).locator("button").first.click()
    # El modal ("Observaciones de subsanación") NO expone role=dialog de forma fiable;
    # lo operamos por su textarea (con contador de caracteres) + el botón ACEPTAR.
    ta = page.locator("textarea").last
    ta.wait_for(state="visible", timeout=15000)
    ta.fill(texto)
    page.get_by_role("button", name="ACEPTAR", exact=True).click()
    ta.wait_for(state="hidden", timeout=15000)


def subsanar_items(
    page: Page,
    valor: int,
    texto: str,
    evidencias: Path,
    max_items: int = 0,
    lento: bool = False,
) -> int:
    """Activa el modo SUBSANAR y llena cada ítem con (valor, texto). Devuelve cuántos
    se llenaron. NOTA: procesa la página actual (>1 página aún es un TODO)."""
    page.get_by_role("button", name="SUBSANAR GLOSA", exact=True).click()
    page.wait_for_timeout(1500)
    _dump_tabla(page, evidencias, "dbg_subsanar_inicial.html")
    total = _plus_items(page).count()
    n = min(total, max_items) if max_items else total
    logger.info(f"  Modo subsanar activo. Ítems: {total} — a llenar: {n}")
    hechos = 0
    for i in range(n):
        # Expandir el ítem i (acordeón: colapsa el que estaba abierto). Re-queryeamos
        # por si el DOM se re-renderizó al insertar/quitar la sub-fila de detalle.
        _plus_items(page).nth(i).click()
        page.wait_for_timeout(700)
        if i == 0:
            _dump_tabla(page, evidencias, "dbg_expandido.html")
        try:
            _set_valor(page, valor)
            _set_observacion(page, texto)
        except Exception as e:  # noqa: BLE001
            _dump_tabla(page, evidencias, f"dbg_error_item{i + 1}.html")
            raise RuntimeError(
                f"Falló el llenado del ítem {i + 1}: {str(e)[:150]}. Se volcó el HTML de "
                f"la tabla en {evidencias} (dbg_*.html) para calibrar el selector."
            ) from e
        hechos += 1
        logger.info(f"    ítem {i + 1}/{n} ✓")
        if lento:
            page.wait_for_timeout(300)
    return hechos


def finalizar_factura(page: Page, factura: str, evidencias: Path) -> str:
    """Cierra la subsanación con ACEPTAR TOTAL RATIFICADO y captura evidencia."""
    btn = page.get_by_role("button", name="ACEPTAR TOTAL RATIFICADO", exact=True)
    btn.wait_for(timeout=15000)
    btn.click()
    page.wait_for_timeout(2500)
    evidencias.mkdir(parents=True, exist_ok=True)
    shot = evidencias / f"{factura}_ok.png"
    with contextlib.suppress(Exception):
        page.screenshot(path=str(shot), full_page=True)
    return f"finalizada; evidencia {shot.name}"


def procesar_factura(
    page: Page,
    factura: str,
    datos: dict,
    evidencias: Path,
    max_items: int = 0,
    finalizar: bool = False,
    lento: bool = False,
) -> dict:
    grupos = datos["grupos"]
    items = [it for g in grupos for it in g["items"]]
    # Respuesta uniforme (glosa ratificada): 1 grupo -> mismo (aceptado, detalle) para todos.
    valor = grupos[0]["items"][0]["aceptado"] if grupos and grupos[0]["items"] else 0
    texto = grupos[0]["detalle"] if grupos else ""
    reg = {
        "factura": factura,
        "grupos": len(grupos),
        "items": len(items),
        "estado": "",
        "detalle": "",
    }
    try:
        _abrir_factura(page, factura)
        hechos = subsanar_items(page, valor, texto, evidencias, max_items=max_items, lento=lento)
        if finalizar:
            det = finalizar_factura(page, factura, evidencias)
            reg["estado"] = "OK"
            reg["detalle"] = f"{hechos} ítems; {det}"
        else:
            reg["estado"] = "LLENADO_SIN_ENVIAR"
            reg["detalle"] = f"{hechos} ítems llenados; revisá y re-corré con --finalizar"
    except Exception as e:
        reg["estado"] = "ERROR"
        reg["detalle"] = str(e)[:300]
    return reg


# ─── CLI / main ──────────────────────────────────────────────────────────────


def _seleccionar_facturas(por_factura: dict, args) -> dict:
    if args.solo:
        objetivo = {args.solo.strip().upper()}
    elif args.facturas:
        objetivo = {f.strip().upper() for f in args.facturas.split(",") if f.strip()}
    elif args.lista:
        objetivo = {
            ln.strip().upper()
            for ln in Path(args.lista).read_text(encoding="utf-8").splitlines()
            if ln.strip()
        }
    else:  # --todas
        return por_factura
    return {f: d for f, d in por_factura.items() if f in objetivo}


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Respuesta a glosas en el portal MUTUAL SER (Zona Ser).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--excel", type=Path, help="Excel de respuestas (extraer_respuestas_glosa_mutualser.py)."
    )
    grupo = parser.add_mutually_exclusive_group(required=False)
    grupo.add_argument("--solo", type=str, help="Procesar SOLO esta factura (piloto).")
    grupo.add_argument("--facturas", type=str, help="Lista separada por coma.")
    grupo.add_argument("--lista", type=Path, help="TXT con una factura por línea.")
    grupo.add_argument("--todas", action="store_true", help="Todas las facturas del Excel.")
    parser.add_argument(
        "--explorar",
        action="store_true",
        help="Volcar el DOM del módulo para calibrar (no responde).",
    )
    parser.add_argument(
        "--storage-state",
        type=str,
        default=STORAGE_STATE_DEFAULT,
        help="JSON de sesión (evita el captcha).",
    )
    parser.add_argument(
        "--evidencias",
        type=Path,
        default=Path("EVIDENCIA_MUTUALSER"),
        help="Carpeta de screenshots de cierre.",
    )
    parser.add_argument(
        "--con-cabeza",
        action="store_true",
        help="Browser visible (necesario para resolver el captcha).",
    )
    parser.add_argument(
        "--cdp",
        type=str,
        default=None,
        help=(
            "Conectar a un Chrome REAL ya abierto (ej. http://localhost:9222) donde "
            "vos iniciaste sesión a mano. Evita que el reCAPTCHA detecte automatización."
        ),
    )
    parser.add_argument(
        "--canal",
        type=str,
        default=None,
        help="Canal del browser en modo lanzado (ej. 'chrome' o 'msedge').",
    )
    parser.add_argument("--lento", action="store_true", help="slow_mo 300ms (debug visual).")
    parser.add_argument(
        "--max-items",
        type=int,
        default=0,
        help="Llenar como mucho N ítems por factura (piloto). 0 = todos.",
    )
    parser.add_argument(
        "--finalizar",
        action="store_true",
        help=(
            "Hacer click en ACEPTAR TOTAL RATIFICADO al terminar (envía). SIN este "
            "flag el bot SOLO llena los campos y NO envía (revisás y re-corrés)."
        ),
    )
    parser.add_argument(
        "--reporte", type=Path, default=Path("reporte_mutualser.csv"), help="CSV de salida."
    )
    parser.add_argument("--log", type=Path, default=None, help="Archivo de log adicional.")
    args = parser.parse_args()

    setup_logging(args.log)
    _exigir_playwright()
    # En modo --cdp el login lo hace el usuario en su Chrome: no hacen falta credenciales.
    user, password = ("", "") if args.cdp else cargar_credenciales()

    if not args.explorar and not args.excel:
        parser.error("indicá --excel (o usá --explorar para calibrar el portal).")
    if not args.explorar and not (args.solo or args.facturas or args.lista or args.todas):
        parser.error("elegí qué facturas procesar: --solo / --facturas / --lista / --todas.")

    with sync_playwright() as p:
        browser, ctx, page, es_cdp = abrir_sesion(p, args, user, password)
        try:
            if args.explorar:
                if es_cdp:
                    logger.info(
                        "Modo CDP: vuelco la pestaña ACTUAL de tu Chrome "
                        "(abrí la factura en modo subsanación para capturar esos botones)."
                    )
                explorar(page, args.evidencias)
                return 0

            por_factura = leer_excel_respuestas(args.excel)
            seleccion = _seleccionar_facturas(por_factura, args)
            if not seleccion:
                logger.warning("No hay facturas que coincidan con la selección.")
                return 0
            logger.info(f"Facturas a procesar: {len(seleccion)}")

            args.reporte.parent.mkdir(parents=True, exist_ok=True)
            f_rep = args.reporte.open("w", newline="", encoding="utf-8-sig")
            w_rep = csv.DictWriter(
                f_rep, fieldnames=["factura", "grupos", "items", "estado", "detalle"]
            )
            w_rep.writeheader()
            resultados: list[dict] = []
            t0 = time.time()
            for i, (factura, datos) in enumerate(seleccion.items(), start=1):
                n_items = sum(len(g["items"]) for g in datos["grupos"])
                logger.info(
                    f"[{i}/{len(seleccion)}] {factura} — {len(datos['grupos'])} grupo(s), {n_items} ítems"
                )
                reg = procesar_factura(
                    page,
                    factura,
                    datos,
                    args.evidencias,
                    max_items=args.max_items,
                    finalizar=args.finalizar,
                    lento=args.lento,
                )
                resultados.append(reg)
                w_rep.writerow(reg)
                if i % 5 == 0:
                    f_rep.flush()
                logger.info(f"    → {reg['estado']}: {reg['detalle']}")
            f_rep.close()

            from collections import Counter

            dur = (time.time() - t0) / 60
            logger.info(f"\nReporte: {args.reporte} | {len(resultados)} facturas en {dur:.1f} min")
            for estado, n in Counter(r["estado"] for r in resultados).items():
                logger.info(f"  {estado}: {n}")
            if any(r["estado"] == "LLENADO_SIN_ENVIAR" for r in resultados):
                logger.warning(
                    "Se LLENARON los campos pero NO se envió (sin --finalizar). Revisá en "
                    "el browser y, si está OK, re-corré agregando  --finalizar."
                )
        finally:
            # browser.close() sobre una conexión CDP sólo DESCONECTA (no mata el
            # Chrome del usuario); sobre un browser lanzado, lo cierra.
            with contextlib.suppress(Exception):
                browser.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
