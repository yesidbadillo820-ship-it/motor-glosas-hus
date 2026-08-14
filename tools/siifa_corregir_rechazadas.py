#!/usr/bin/env python3
"""siifa_corregir_rechazadas.py — Arma el archivo para volver a subir SOLO lo
que SIIFA rechazó, ya corregido.

POR QUÉ EXISTE
En el cargue del 05-08-2026 quedaron 1.422 respuestas sin subir, por dos
causas que no se ven a simple vista en el reporte:

1. **El texto pasaba de 1.500 caracteres.** SIIFA no lo dice: contesta
   «HTTP 500: error saving the entity changes», que parece falla del
   servidor. Entraron todas las de hasta 1.499 y ninguna de 1.501 o más.
2. **El código de respuesta no lo acepta la plataforma.** Ese sí lo dice:
   «el código no existe, no está activo o no pertenece al grupo RESPUESTA».
   Pasó con RE9701, que se usa en DGH pero no en SIIFA.

Este script lee el Excel que se cargó y el reporte del cargue, se queda con
las filas que quedaron en ERROR, les aplica las dos correcciones y las deja
en un Excel nuevo listo para responder_glosas_siifa.py. No toca las que ya
quedaron subidas.

USO:
    py siifa_corregir_rechazadas.py ^
        --excel    "D:\\...\\SIIFA\\respuestas_DEVOLUCIONES.xlsx" ^
        --reporte  "D:\\...\\SIIFA\\reporte_DEVOLUCIONES.csv" ^
        --reporte  "D:\\...\\SIIFA\\reporte_DEVOLUCIONES_2.csv" ^
        --cambiar-codigo RE9701=RE9901 ^
        --salida   "D:\\...\\SIIFA\\respuestas_DEVOLUCIONES_CORREGIDAS.xlsx"

Se pueden pasar varios --reporte: una fila cuenta como subida si quedó OK en
CUALQUIERA de ellos.

La columna CORRECCION dice qué se le hizo a cada fila. Las que quedaron con
el texto recortado o con el código cambiado hay que MIRARLAS antes de subir:
el recorte quita palabras del final y el cambio de código puede alterar el
sentido de la respuesta.

INSTALACIÓN (una vez):
    py -m pip install openpyxl
"""

from __future__ import annotations

import argparse
import csv
import logging
import sys
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from siifa_redactar_respuestas import LIMITE_OBSERVACION, redactar  # noqa: E402

logger = logging.getLogger("siifa_corregir")

# Homologación de códigos de DGH que SIIFA no acepta. HOY ESTÁ VACÍA, a
# propósito, y conviene dejar por escrito por qué.
#
# Con las devoluciones que el hospital aceptó (RE9701) SIIFA contesta que el
# código «no existe, no está activo o no pertenece al grupo RESPUESTA». Lo
# primero que se pensó fue cambiarlo por el código de aceptación de las
# glosas. Está mal: el auditor confirmó contra el portal que para una
# devolución el código correcto ES RE9701 —el desplegable de «Responder
# Devolución» lo ofrece—. Lo que falla es la PUERTA: el bot manda glosas y
# devoluciones por PUT /api/SeguimientoFacturaGlosa/Respuesta, que valida
# contra los códigos del grupo de GLOSA.
#
# Mientras no se sepa por dónde se responde una devolución
# (tools/siifa_sondear_endpoints.py lo averigua sin escribir nada), cambiarle
# el código a esas respuestas sería taparle la boca al error: quedarían
# registradas diciendo algo distinto de lo que el hospital decidió.
HOMOLOGACION: dict[str, str] = {}

# Columnas que necesita el bot de cargue, más las que ayudan a revisar.
COLUMNAS = [
    "ID_SEGUIMIENTO_FACTURA_GLOSA",
    "ID_SEGUIMIENTO_FACTURA_DEVOLUCION",
    "NUMERO_FACTURA",
    "TIPO",
    "CODIGO_GLOSA",
    "VALOR_GLOSA",
    "ORIGEN_RESPUESTA",
    "CORRECCION",
    "CODIGO_RESPUESTA",
    "FECHA_RESPUESTA",
    "OBSERVACION_RESPUESTA",
]


def _texto(valor) -> str:
    return "" if valor is None else str(valor).strip()


def recortar(texto: str, limite: int = LIMITE_OBSERVACION) -> tuple[str, bool]:
    """Devuelve (texto que cabe, si hubo que recortar).

    Se corta en el último punto o, si no hay, en el último espacio: partir
    una palabra por la mitad en un documento que se le manda a la EPS queda
    mal. Los puntos suspensivos avisan de que la respuesta sigue.
    """
    texto = _texto(texto)
    if len(texto) <= limite:
        return texto, False
    corte = texto[: limite - 3]
    punto = corte.rfind(". ")
    if punto > limite // 2:
        return corte[: punto + 1], True
    espacio = corte.rfind(" ")
    if espacio > limite // 2:
        corte = corte[:espacio]
    return corte.rstrip(" ,;:") + "...", True


def ids_subidos(rutas: list[str]) -> set[int]:
    """Los ids que quedaron OK en cualquiera de los reportes del cargue."""
    subidos: set[int] = set()
    for ruta in rutas:
        with open(ruta, newline="", encoding="utf-8-sig") as f:
            for fila in csv.DictReader(f):
                if _texto(fila.get("estado")).upper() not in ("OK", "YA_OK_PREVIO"):
                    continue
                try:
                    subidos.add(int(fila["id_seguimiento_factura_glosa"]))
                except (KeyError, TypeError, ValueError):
                    continue
    return subidos


def leer(ruta: Path, hoja: str | None = None) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[hoja] if hoja else wb[wb.sheetnames[0]]
    filas = ws.iter_rows(values_only=True)
    encabezados = [_texto(c).upper().replace(" ", "_") for c in next(filas)]
    if "ID_SEGUIMIENTO_FACTURA_GLOSA" not in encabezados:
        raise SystemExit(
            f"\nERROR: a {ruta.name} le falta la columna ID_SEGUIMIENTO_FACTURA_GLOSA.\n"
            "Tiene que ser el mismo Excel que se usó para cargar.\n"
        )
    return [dict(zip(encabezados, f)) for f in filas if f[0] not in (None, "")]


def lineas_del_informe(ruta: Path) -> dict[int, dict]:
    """La glosa original de cada id, para poder RE-redactar en vez de cortar."""
    from siifa_preparar_respuestas import leer_informe

    lineas = {}
    for linea in leer_informe(ruta):
        try:
            lineas[int(linea["id_seguimiento_factura_glosa"])] = linea
        except (KeyError, TypeError, ValueError):
            continue
    return lineas


def corregir(fila: dict, cambios: dict[str, str], limite: int, informe: dict | None = None) -> dict:
    """Aplica las correcciones y deja dicho qué se le hizo a cada fila."""
    hechas = []
    codigo = _texto(fila.get("CODIGO_RESPUESTA")).upper()
    nuevo = cambios.get(codigo)
    if nuevo:
        hechas.append(f"código {codigo} cambiado por {nuevo} — REVISAR que el texto concuerde")
        codigo = nuevo

    texto = _texto(fila.get("OBSERVACION_RESPUESTA"))
    linea = (informe or {}).get(int(fila["ID_SEGUIMIENTO_FACTURA_GLOSA"]))
    if (
        len(texto) > limite
        and linea
        and _texto(fila.get("ORIGEN_RESPUESTA")).upper() == "REDACTADA"
    ):
        # La escribió el motor: se vuelve a escribir con el límite bueno. Así
        # el recorte se lo lleva la cita de la EPS y no el cierre del hospital
        # (mesa de conciliación y correo), que es parte de la fórmula.
        texto = redactar(linea, limite=limite)["OBSERVACION_RESPUESTA"]
        hechas.append(f"respuesta rehecha para que quepa en {limite} caracteres")
    else:
        texto, se_recorto = recortar(texto, limite)
        if se_recorto:
            hechas.append(
                f"texto recortado a {limite} caracteres — REVISAR que no falte lo esencial"
            )

    salida = {c: fila.get(c) for c in COLUMNAS}
    salida["CODIGO_RESPUESTA"] = codigo
    salida["OBSERVACION_RESPUESTA"] = texto
    salida["CORRECCION"] = "; ".join(hechas) or "sin cambios: reintento tal cual"
    return salida


def escribir(filas: list[dict], ruta: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "RESPUESTAS"
    for i, h in enumerate(COLUMNAS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.font = Font(bold=True, color="FFFFFF")

    ojo = PatternFill("solid", fgColor="FFEB9C")
    for r, fila in enumerate(filas, start=2):
        for i, h in enumerate(COLUMNAS, 1):
            c = ws.cell(row=r, column=i, value=fila.get(h))
            c.alignment = Alignment(
                vertical="top", wrap_text=h in ("CORRECCION", "OBSERVACION_RESPUESTA")
            )
            if h == "VALOR_GLOSA":
                c.number_format = '"$"#,##0'
            if h == "CORRECCION" and "REVISAR" in _texto(fila.get("CORRECCION")):
                c.fill = ojo
    for i, ancho in enumerate([16, 18, 14, 12, 12, 14, 16, 52, 14, 14, 120], 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNAS))}{ws.max_row}"
    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta)
    logger.info("Archivo para volver a subir: %s (%d respuestas)", ruta, len(filas))


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--excel", required=True, help="El Excel que se usó para cargar.")
    ap.add_argument("--hoja", help="Nombre de la hoja (default: la primera).")
    ap.add_argument(
        "--reporte", action="append", required=True, help="CSV del cargue (se puede repetir)."
    )
    ap.add_argument("--salida", required=True, help="Excel nuevo, solo con lo rechazado.")
    ap.add_argument(
        "--cambiar-codigo",
        action="append",
        default=[],
        metavar="VIEJO=NUEVO",
        help="Cambia un código de respuesta que SIIFA no acepta (ej. RE9701=RE9901).",
    )
    ap.add_argument(
        "--limite",
        type=int,
        default=LIMITE_OBSERVACION,
        help=f"Máximo de caracteres de la observación (default: {LIMITE_OBSERVACION}).",
    )
    ap.add_argument(
        "--homologar",
        action="store_true",
        help="Aplica la tabla de homologación de códigos del repo (hoy está vacía: "
        "ver el comentario de HOMOLOGACION en este archivo).",
    )
    ap.add_argument(
        "--solo-codigo",
        metavar="CODIGO",
        help="Se queda SOLO con las respuestas que traen ese código de respuesta.",
    )
    ap.add_argument(
        "--excluir-codigo",
        metavar="CODIGO",
        help="Deja fuera las respuestas que traen ese código (para tratarlas aparte).",
    )
    ap.add_argument(
        "--informe",
        help="Informe de siifa_reporte_seguimientos.py. Con él, las respuestas que "
        "escribió el motor se vuelven a escribir para que quepan, en vez de cortarles "
        "el final.",
    )
    ap.add_argument("--verbose", action="store_true")
    args = ap.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s  %(message)s",
        datefmt="%H:%M:%S",
        stream=sys.stdout,
    )

    cambios = dict(HOMOLOGACION) if args.homologar else {}
    for par in args.cambiar_codigo:
        if "=" not in par:
            raise SystemExit(f"\nERROR: --cambiar-codigo se escribe VIEJO=NUEVO, no '{par}'.\n")
        viejo, nuevo = par.split("=", 1)
        cambios[viejo.strip().upper()] = nuevo.strip().upper()

    filas = leer(Path(args.excel), args.hoja)
    if args.solo_codigo:
        filas = [
            f
            for f in filas
            if _texto(f.get("CODIGO_RESPUESTA")).upper() == args.solo_codigo.upper()
        ]
    if args.excluir_codigo:
        filas = [
            f
            for f in filas
            if _texto(f.get("CODIGO_RESPUESTA")).upper() != args.excluir_codigo.upper()
        ]
    if not filas:
        raise SystemExit("\nNinguna respuesta del archivo cumple ese filtro de código.\n")
    subidos = ids_subidos(args.reporte)
    rechazadas = [f for f in filas if int(f["ID_SEGUIMIENTO_FACTURA_GLOSA"]) not in subidos]
    if not rechazadas:
        raise SystemExit("\nNo quedó nada rechazado: todo el archivo está subido.\n")

    informe = lineas_del_informe(Path(args.informe)) if args.informe else None
    corregidas = [corregir(f, cambios, args.limite, informe) for f in rechazadas]
    escribir(corregidas, Path(args.salida))

    def que_se_hizo(correccion: str) -> str:
        partes = []
        if "rehecha" in correccion:
            partes.append("respuesta rehecha")
        if "recortado" in correccion:
            partes.append("texto recortado")
        if "código" in correccion:
            partes.append("código cambiado")
        return " + ".join(partes) or "sin cambios"

    cuenta = Counter(que_se_hizo(f["CORRECCION"]) for f in corregidas)
    print(
        f"\nDel archivo original ({len(filas)} respuestas) ya estaban subidas {len(subidos)}.\n"
        f"Quedan {len(corregidas)} para volver a subir:\n"
        + "".join(f"  {k:18s}: {v:5d}\n" for k, v in cuenta.most_common())
        + "\nRevisar la columna CORRECCION (lo amarillo) antes de subir.\n"
        "Antes del cargue completo, PILOTO DE 1 (regla del repo):\n"
        f'   py tools\\responder_glosas_siifa.py --excel "{args.salida}" '
        f'--piloto 1 --reporte "piloto_correccion.csv"\n'
    )


if __name__ == "__main__":
    main()
