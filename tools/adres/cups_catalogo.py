"""cups_catalogo.py — Lookup de nombres CUPS (normatividad MSPS).

El RIPS no trae el nombre de consultas/procedimientos (sólo el código CUPS,
p.ej. 890435). Este módulo carga un catálogo CUPS oficial (código → nombre)
desde CSV/TSV/XLSX y resuelve la descripción que falta en el FUR SERVICIOS.

El catálogo lo provee el usuario (la tabla CUPS de la normatividad). Detecta
solo las columnas de código y nombre, sin importar el orden ni encabezados
exactos.

Sólo stdlib (csv); openpyxl únicamente si el catálogo es .xlsx.
"""

from __future__ import annotations

import csv
from pathlib import Path


def _norm_cod(c: str) -> str:
    """Normaliza un código para comparar (sin espacios ni ceros a la izquierda)."""
    return (c or "").strip().lstrip("0") or "0"


def _detectar_columnas(headers: list[str]) -> tuple[int, int]:
    """Devuelve (idx_codigo, idx_nombre) detectando por encabezado."""
    cod_idx = nom_idx = None
    for i, h in enumerate(headers):
        hl = (h or "").strip().lower()
        if cod_idx is None and ("cup" in hl or "codigo" in hl or "código" in hl or hl == "cod"):
            cod_idx = i
        elif nom_idx is None and ("nombre" in hl or "descrip" in hl or "servicio" in hl):
            nom_idx = i
    if cod_idx is None:
        cod_idx = 0
    if nom_idx is None:
        nom_idx = 1 if len(headers) > 1 else 0
    return cod_idx, nom_idx


def _filas(ruta: Path) -> list[list[str]]:
    suf = ruta.suffix.lower()
    if suf in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise ImportError("Para catálogo .xlsx instalá openpyxl (pip install openpyxl)") from e
        wb = load_workbook(filename=str(ruta), read_only=True, data_only=True)
        ws = wb.active
        return [
            ["" if c is None else str(c) for c in row] for row in ws.iter_rows(values_only=True)
        ]
    sep = "\t" if suf == ".tsv" else ","
    with ruta.open(encoding="utf-8-sig", newline="") as fh:
        return [list(r) for r in csv.reader(fh, delimiter=sep)]


def cargar_catalogo(ruta: Path) -> dict[str, str]:
    """Devuelve {codigo_cups: nombre}. Indexa también el código normalizado."""
    filas = _filas(ruta)
    if not filas:
        return {}
    cod_idx, nom_idx = _detectar_columnas(filas[0])
    cat: dict[str, str] = {}
    for row in filas[1:]:
        if max(cod_idx, nom_idx) >= len(row):
            continue
        cod = (row[cod_idx] or "").strip()
        nom = (row[nom_idx] or "").strip()
        if not cod or not nom:
            continue
        cat[cod] = nom
        cat.setdefault(_norm_cod(cod), nom)
    return cat


def buscar(cat: dict[str, str], codigo: str) -> str:
    """Busca el nombre por código exacto o normalizado."""
    if not codigo or not cat:
        return ""
    return cat.get(codigo.strip()) or cat.get(_norm_cod(codigo)) or ""
