"""validar_furips.py — Bot de validación FURIPS vs soportes (Circular 022/2023 ADRES).

Valida de forma MASIVA las reclamaciones FURIPS del HUS contra:

1. La malla de la Circular 022 de 2023 de la ADRES (anexo técnico FURIPS 1 de
   102 campos y FURIPS 2 de 9 campos): obligatoriedad, formatos de fecha/hora,
   valores permitidos, longitudes máximas y reglas condicionales (estado de
   aseguramiento, remisión, transporte primario, UCI, etc.).
2. Los SOPORTES de cada carpeta de factura: RIPS JSON, CUV JSON, Factura
   electrónica DIAN (XML), representación gráfica de la factura (PDF) y
   EPICRISIS (PDF). Cruza dato por dato: paciente, documento, fechas de
   ingreso/egreso, diagnósticos, valores, número de factura, NIT, etc.
3. La coherencia interna FURIPS1 ↔ FURIPS2 (sumatorias de amparos, líneas
   quirúrgicas, consecutivo de reclamación, reclamado ≤ facturado).

Produce un INFORME EXCEL detallado con 7 hojas: RESUMEN, HALLAZGOS,
FURIPS1 CAMPOS, FURIPS2 LINEAS, CRUCE SOPORTES, SOPORTES y LEYENDA.

USO
---

    # Masivo: una raíz con carpetas por factura (los TXT FURIPS pueden estar
    # en la raíz o dentro de cualquier carpeta; se detectan solos)
    py validar_furips.py --raiz "C:\\FACTURAS\\ADRES MAYO"

    # Una sola carpeta de factura
    py validar_furips.py --carpeta "C:\\FACTURAS\\HUS374152"

    # Indicando los TXT explícitamente y la salida
    py validar_furips.py --raiz "C:\\FACTURAS" ^
        --furips1 "C:\\FACTURAS\\FURIPS1680010079201.txt" ^
        --furips2 "C:\\FACTURAS\\FURIPS2680010079201.txt" ^
        --salida  "C:\\FACTURAS\\INFORME_FURIPS.xlsx"

DEPENDENCIAS
------------
- openpyxl (para el Excel). Obligatoria.
- pdfplumber o pypdf (para leer los PDF de factura y epicrisis). Opcional:
  sin ellas el bot corre igual pero marca los PDF como "no legibles".
"""

from __future__ import annotations

import argparse
import contextlib
import json
import logging
import re
import sys
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

logger = logging.getLogger("validar_furips")

_DIR = Path(__file__).resolve().parent
if str(_DIR) not in sys.path:
    sys.path.insert(0, str(_DIR))

# ─────────────────────────────────────────────────────────────────────────────
# Severidades
# ─────────────────────────────────────────────────────────────────────────────
ERROR = "ERROR"
ADVERTENCIA = "ADVERTENCIA"
INFO = "INFO"
OK = "OK"

# Códigos DANE de departamento vigentes
DEPTOS_DANE = {
    "05",
    "08",
    "11",
    "13",
    "15",
    "17",
    "18",
    "19",
    "20",
    "23",
    "25",
    "27",
    "41",
    "44",
    "47",
    "50",
    "52",
    "54",
    "63",
    "66",
    "68",
    "70",
    "73",
    "76",
    "81",
    "85",
    "86",
    "88",
    "91",
    "94",
    "95",
    "97",
    "99",
}

TIPOS_DOC_VICTIMA = {"CC", "CE", "CN", "PA", "TI", "RC", "AS", "MS", "CD", "SC", "PE", "PT", "DE"}
TIPOS_DOC_PROPIETARIO = {"CC", "CE", "CD", "DE", "SC", "PE", "PT", "NI", "TI"}
TIPOS_DOC_CONDUCTOR = {"CC", "CE", "PA", "RC", "TI", "MS", "AS", "CD", "SC", "DE", "PE", "PT"}
TIPOS_DOC_MEDICO = {"CC", "CE", "PE", "PA", "PT"}
NATURALEZAS_EVENTO = {
    "01",
    "02",
    "03",
    "04",
    "05",
    "06",
    "07",
    "08",
    "09",
    "10",
    "11",
    "12",
    "13",
    "14",
    "15",
    "16",
    "17",
    "25",
    "26",
    "27",
}
TIPOS_VEHICULO = {"1", "2", "3", "4", "5", "6", "7", "8", "10", "14", "17", "19", "20", "21", "22"}
ESTADOS_ASEGURAMIENTO = {"1", "2", "3", "4", "6", "7", "8"}
NOMBRE_ESTADO_ASEG = {
    "1": "Asegurado",
    "2": "No asegurado",
    "3": "Vehículo fantasma",
    "4": "Póliza falsa",
    "6": "Asegurado D.2497",
    "7": "No aseg. propietario indeterminado",
    "8": "No asegurado sin placa",
}
TIPOS_SERVICIO_F2 = {
    "1": "Medicamentos",
    "2": "Procedimientos",
    "3": "Transporte primario",
    "4": "Transporte secundario",
    "5": "Insumos",
    "6": "Dispositivos médicos",
    "7": "Material de osteosíntesis",
    "8": "Procedimiento fuera de manual tarifario",
}

_RE_FECHA = re.compile(r"^\d{2}/\d{2}/\d{4}$")
_RE_HORA = re.compile(r"^\d{2}:\d{2}$")
_RE_CIE10 = re.compile(r"^[A-Z]\d{2}[0-9A-Z]?$")
_RE_PLACA = re.compile(r"^[A-Z0-9]{3,10}$")
_RE_NUM = re.compile(r"^\d+$")

# ─────────────────────────────────────────────────────────────────────────────
# Especificación FURIPS 1 — Tabla 1 del anexo técnico, Circular 022/2023 ADRES
# (nro, sección, concepto, longitud máx, formato, valores permitidos, obligatoriedad)
# Obligatoriedad: SI / NO / código de condición que se evalúa en código.
# ─────────────────────────────────────────────────────────────────────────────
E1 = [
    (1, "I. Datos de la reclamación", "Número de radicado anterior", 10, None, None, "C_RGO"),
    (
        2,
        "I. Datos de la reclamación",
        "RGO Respuesta a glosa u objeción",
        1,
        None,
        {"0", "1", "6"},
        "NO",
    ),
    (3, "I. Datos de la reclamación", "Número de factura", 20, None, None, "SI"),
    (
        4,
        "I. Datos de la reclamación",
        "Número consecutivo de la reclamación",
        12,
        "num",
        None,
        "SI",
    ),
    (5, "II. Datos del prestador", "Código de habilitación del prestador", 12, "num", None, "SI"),
    (6, "III. Datos de la víctima", "Primer apellido de la víctima", 20, None, None, "SI"),
    (7, "III. Datos de la víctima", "Segundo apellido de la víctima", 30, None, None, "NO"),
    (8, "III. Datos de la víctima", "Primer nombre de la víctima", 20, None, None, "SI"),
    (9, "III. Datos de la víctima", "Segundo nombre de la víctima", 30, None, None, "NO"),
    (
        10,
        "III. Datos de la víctima",
        "Tipo de documento de la víctima",
        2,
        None,
        TIPOS_DOC_VICTIMA,
        "SI",
    ),
    (11, "III. Datos de la víctima", "Número de documento de la víctima", 16, None, None, "SI"),
    (12, "III. Datos de la víctima", "Fecha de nacimiento de la víctima", 10, "fecha", None, "NO"),
    (13, "III. Datos de la víctima", "Fecha de fallecimiento", 10, "fecha", None, "NO"),
    (14, "III. Datos de la víctima", "Sexo de la víctima", 1, None, {"F", "M", "O"}, "SI"),
    (
        15,
        "III. Datos de la víctima",
        "Dirección de residencia de la víctima",
        100,
        None,
        None,
        "SI",
    ),
    (16, "III. Datos de la víctima", "Código departamento de residencia", 2, "depto", None, "SI"),
    (17, "III. Datos de la víctima", "Código municipio de residencia", 3, "mun", None, "SI"),
    (18, "III. Datos de la víctima", "Teléfono de la víctima", 10, None, None, "SI"),
    (
        19,
        "III. Datos de la víctima",
        "Condición de la víctima",
        1,
        None,
        {"1", "2", "3", "4"},
        "C_NAT01",
    ),
    (20, "IV. Sitio del evento", "Naturaleza del evento", 2, None, NATURALEZAS_EVENTO, "SI"),
    (21, "IV. Sitio del evento", "Descripción del otro evento", 25, None, None, "C_NAT17"),
    (22, "IV. Sitio del evento", "Dirección de ocurrencia del evento", 100, None, None, "SI"),
    (23, "IV. Sitio del evento", "Fecha de ocurrencia del evento", 10, "fecha", None, "SI"),
    (24, "IV. Sitio del evento", "Hora de ocurrencia del evento", 5, "hora", None, "SI"),
    (25, "IV. Sitio del evento", "Código departamento de ocurrencia", 2, "depto", None, "SI"),
    (26, "IV. Sitio del evento", "Código municipio de ocurrencia", 3, "mun", None, "SI"),
    (27, "IV. Sitio del evento", "Zona de ocurrencia del evento", 1, None, {"U", "R"}, "NO"),
    (
        28,
        "V. Vehículo involucrado",
        "Estado de aseguramiento",
        1,
        None,
        ESTADOS_ASEGURAMIENTO,
        "C_NAT01",
    ),
    (29, "V. Vehículo involucrado", "Marca del vehículo", 15, None, None, "C_VEH_NO3"),
    (30, "V. Vehículo involucrado", "Placa del vehículo", 10, "placa", None, "C_VEH_PLACA"),
    (31, "V. Vehículo involucrado", "Tipo de vehículo", 2, None, TIPOS_VEHICULO, "C_VEH_NO3"),
    (32, "V. Vehículo involucrado", "Código de la aseguradora", 6, None, None, "C_VEH_POLIZA"),
    (33, "V. Vehículo involucrado", "Número de póliza SOAT", 20, None, None, "C_VEH_POLIZA"),
    (
        34,
        "V. Vehículo involucrado",
        "Fecha inicio vigencia póliza",
        10,
        "fecha",
        None,
        "C_VEH_POLIZA",
    ),
    (
        35,
        "V. Vehículo involucrado",
        "Fecha final vigencia póliza",
        10,
        "fecha",
        None,
        "C_VEH_POLIZA",
    ),
    (
        36,
        "V. Vehículo involucrado",
        "Número de radicado SIRAS (id_atencion)",
        20,
        None,
        None,
        "C_NAT01",
    ),
    (
        37,
        "V. Vehículo involucrado",
        "Cobro por agotamiento tope aseguradora",
        1,
        None,
        {"0", "1"},
        "C_NAT01",
    ),
    (
        38,
        "VI. Atención de la víctima",
        "CUPS servicio principal hospitalización",
        6,
        None,
        None,
        "C_ESTANCIA",
    ),
    (
        39,
        "VI. Atención de la víctima",
        "Complejidad del procedimiento quirúrgico",
        1,
        None,
        {"1", "2", "3"},
        "C_QX",
    ),
    (
        40,
        "VI. Atención de la víctima",
        "CUPS procedimiento quirúrgico principal",
        6,
        None,
        None,
        "C_QX",
    ),
    (
        41,
        "VI. Atención de la víctima",
        "CUPS procedimiento quirúrgico secundario",
        6,
        None,
        None,
        "NO",
    ),
    (42, "VI. Atención de la víctima", "Se prestó servicio UCI", 1, None, {"0", "1"}, "C_UCI_F2"),
    (43, "VI. Atención de la víctima", "Días de UCI reclamados", 2, "num", None, "C_UCI"),
    (
        44,
        "VII. Propietario del vehículo",
        "Tipo documento del propietario",
        2,
        None,
        TIPOS_DOC_PROPIETARIO,
        "C_PROP",
    ),
    (
        45,
        "VII. Propietario del vehículo",
        "Número documento del propietario",
        16,
        None,
        None,
        "C_PROP",
    ),
    (
        46,
        "VII. Propietario del vehículo",
        "Primer apellido del propietario o razón social",
        40,
        None,
        None,
        "C_PROP",
    ),
    (47, "VII. Propietario del vehículo", "Segundo apellido del propietario", 30, None, None, "NO"),
    (
        48,
        "VII. Propietario del vehículo",
        "Primer nombre del propietario",
        20,
        None,
        None,
        "C_PROP_NAT",
    ),
    (49, "VII. Propietario del vehículo", "Segundo nombre del propietario", 30, None, None, "NO"),
    (
        50,
        "VII. Propietario del vehículo",
        "Dirección de residencia del propietario",
        200,
        None,
        None,
        "C_PROP",
    ),
    (51, "VII. Propietario del vehículo", "Teléfono del propietario", 10, None, None, "NO"),
    (
        52,
        "VII. Propietario del vehículo",
        "Código departamento residencia propietario",
        2,
        "depto",
        None,
        "C_PROP",
    ),
    (
        53,
        "VII. Propietario del vehículo",
        "Código municipio residencia propietario",
        3,
        "mun",
        None,
        "C_PROP",
    ),
    (54, "VIII. Conductor del vehículo", "Primer apellido del conductor", 20, None, None, "C_COND"),
    (55, "VIII. Conductor del vehículo", "Segundo apellido del conductor", 30, None, None, "NO"),
    (56, "VIII. Conductor del vehículo", "Primer nombre del conductor", 20, None, None, "C_COND"),
    (57, "VIII. Conductor del vehículo", "Segundo nombre del conductor", 30, None, None, "NO"),
    (
        58,
        "VIII. Conductor del vehículo",
        "Tipo documento del conductor",
        2,
        None,
        TIPOS_DOC_CONDUCTOR,
        "C_COND",
    ),
    (
        59,
        "VIII. Conductor del vehículo",
        "Número documento del conductor",
        16,
        None,
        None,
        "C_COND",
    ),
    (
        60,
        "VIII. Conductor del vehículo",
        "Dirección de residencia del conductor",
        200,
        None,
        None,
        "C_COND",
    ),
    (
        61,
        "VIII. Conductor del vehículo",
        "Código departamento residencia conductor",
        2,
        "depto",
        None,
        "C_COND",
    ),
    (
        62,
        "VIII. Conductor del vehículo",
        "Código municipio residencia conductor",
        3,
        "mun",
        None,
        "C_COND",
    ),
    (63, "VIII. Conductor del vehículo", "Teléfono del conductor", 10, None, None, "NO"),
    (64, "IX. Datos de remisión", "Tipo de referencia", 1, None, {"1", "2", "3"}, "C_REMISION"),
    (65, "IX. Datos de remisión", "Fecha de remisión", 10, "fecha", None, "C_REMISION"),
    (66, "IX. Datos de remisión", "Hora de salida", 5, "hora", None, "C_REMISION"),
    (
        67,
        "IX. Datos de remisión",
        "Código habilitación prestador remitente",
        12,
        "num",
        None,
        "C_REMISION",
    ),
    (68, "IX. Datos de remisión", "Profesional que remite", 60, None, None, "C_REMISION"),
    (69, "IX. Datos de remisión", "Cargo de la persona que remite", 30, None, None, "C_REMISION"),
    (70, "IX. Datos de remisión", "Fecha de aceptación", 10, "fecha", None, "C_REMISION"),
    (71, "IX. Datos de remisión", "Hora de aceptación", 5, "hora", None, "C_REMISION"),
    (
        72,
        "IX. Datos de remisión",
        "Código habilitación prestador que recibe",
        12,
        "num",
        None,
        "C_REMISION",
    ),
    (73, "IX. Datos de remisión", "Profesional que recibe", 60, None, None, "C_REMISION"),
    (
        74,
        "IX. Datos de remisión",
        "Placa ambulancia traslado interinstitucional",
        6,
        "placa",
        None,
        "C_REMISION",
    ),
    (
        75,
        "X. Transporte de la víctima",
        "Placa ambulancia traslado primario",
        6,
        "placa",
        None,
        "C_TRANSP",
    ),
    (
        76,
        "X. Transporte de la víctima",
        "Transporte de la víctima desde",
        40,
        None,
        None,
        "C_TRANSP",
    ),
    (
        77,
        "X. Transporte de la víctima",
        "Transporte de la víctima hasta",
        40,
        None,
        None,
        "C_TRANSP",
    ),
    (
        78,
        "X. Transporte de la víctima",
        "Tipo de servicio del transporte",
        1,
        None,
        {"1", "2"},
        "C_TRANSP",
    ),
    (79, "X. Transporte de la víctima", "Zona donde recoge la víctima", 1, None, {"U", "R"}, "NO"),
    (80, "XI. Certificación de la atención", "Fecha de ingreso", 10, "fecha", None, "SI"),
    (81, "XI. Certificación de la atención", "Hora de ingreso", 5, "hora", None, "SI"),
    (82, "XI. Certificación de la atención", "Fecha de egreso", 10, "fecha", None, "SI"),
    (83, "XI. Certificación de la atención", "Hora de egreso", 5, "hora", None, "SI"),
    (
        84,
        "XI. Certificación de la atención",
        "Diagnóstico principal de ingreso",
        4,
        "cie10",
        None,
        "SI",
    ),
    (
        85,
        "XI. Certificación de la atención",
        "Diagnóstico de ingreso asociado 1",
        4,
        "cie10",
        None,
        "NO",
    ),
    (
        86,
        "XI. Certificación de la atención",
        "Diagnóstico de ingreso asociado 2",
        4,
        "cie10",
        None,
        "NO",
    ),
    (
        87,
        "XI. Certificación de la atención",
        "Diagnóstico principal de egreso",
        4,
        "cie10",
        None,
        "SI",
    ),
    (
        88,
        "XI. Certificación de la atención",
        "Diagnóstico de egreso asociado 1",
        4,
        "cie10",
        None,
        "NO",
    ),
    (
        89,
        "XI. Certificación de la atención",
        "Diagnóstico de egreso asociado 2",
        4,
        "cie10",
        None,
        "NO",
    ),
    (90, "XII. Médico o profesional tratante", "Primer apellido del médico", 20, None, None, "SI"),
    (91, "XII. Médico o profesional tratante", "Segundo apellido del médico", 30, None, None, "NO"),
    (92, "XII. Médico o profesional tratante", "Primer nombre del médico", 20, None, None, "SI"),
    (93, "XII. Médico o profesional tratante", "Segundo nombre del médico", 30, None, None, "NO"),
    (
        94,
        "XII. Médico o profesional tratante",
        "Tipo documento del médico",
        2,
        None,
        TIPOS_DOC_MEDICO,
        "SI",
    ),
    (95, "XII. Médico o profesional tratante", "Número documento del médico", 16, None, None, "SI"),
    (
        96,
        "XII. Médico o profesional tratante",
        "Número de registro del médico",
        16,
        None,
        None,
        "SI",
    ),
    (
        97,
        "XIII. Amparos que reclama",
        "Total facturado gastos médico-quirúrgicos",
        15,
        "num",
        None,
        "SI",
    ),
    (
        98,
        "XIII. Amparos que reclama",
        "Total reclamado gastos médico-quirúrgicos",
        15,
        "num",
        None,
        "SI",
    ),
    (
        99,
        "XIII. Amparos que reclama",
        "Total facturado transporte y movilización",
        15,
        "num",
        None,
        "SI",
    ),
    (
        100,
        "XIII. Amparos que reclama",
        "Total reclamado transporte y movilización",
        15,
        "num",
        None,
        "SI",
    ),
    (
        101,
        "XIV. Declaración del prestador",
        "Manifestación de servicios habilitados",
        3,
        None,
        {"0", "1"},
        "SI",
    ),
    (102, "XIV. Declaración del prestador", "Descripción del evento", 1000, None, None, "NO"),
]

E2 = [
    (1, "Número de factura", 20, None, "SI"),
    (2, "Número consecutivo de la reclamación", 20, "num", "SI"),
    (3, "Tipo de servicio", 1, None, "SI"),
    (4, "Código del servicio", 15, None, "COND"),
    (5, "Descripción del servicio o elemento reclamado", 100, None, "COND"),
    (6, "Cantidad de servicios", 15, "num", "SI"),
    (7, "Valor unitario", 15, "num", "SI"),
    (8, "Valor total facturado", 15, "num", "SI"),
    (9, "Valor total reclamado", 15, "num", "SI"),
]

NOMBRE_CAMPO_F1 = {n: nombre for (n, _sec, nombre, _l, _f, _v, _o) in E1}

TEXTO_OBLIG = {
    "SI": "Obligatorio",
    "NO": "Opcional",
    "C_RGO": "Obligatorio si hay RGO (campo 2)",
    "C_NAT01": "Obligatorio si naturaleza evento = 01 (tránsito)",
    "C_NAT17": "Obligatorio si naturaleza evento = 17 (otro)",
    "C_VEH_NO3": "Obligatorio salvo vehículo fantasma (28=3)",
    "C_VEH_PLACA": "Obligatorio si estado aseg. = 1,2,4,6,7",
    "C_VEH_POLIZA": "Obligatorio si estado aseg. = 1,4,6",
    "C_PROP": "Obligatorio si estado aseg. = 1,2,4,6 (con 7/8 se revisa)",
    "C_PROP_NAT": "Obligatorio (persona natural) si estado aseg. = 1,2,4,6 (con 7/8 se revisa)",
    "C_COND": "Obligatorio si estado aseg. = 1,2,4,6,8",
    "C_REMISION": "Obligatorio si existe remisión (grupo IX)",
    "C_TRANSP": "Obligatorio si se cobra transporte primario",
    "C_QX": "Obligatorio si se facturan procedimientos quirúrgicos",
    "C_UCI": "Obligatorio si se prestó UCI (campo 42=1 o UCI en FURIPS2)",
    "C_ESTANCIA": "Obligatorio cuando se reclama estancia (líneas de estancia en FURIPS2)",
    "C_UCI_F2": "Obligatorio cuando se factura estancia en UCI (FURIPS2)",
}


# ─────────────────────────────────────────────────────────────────────────────
# Utilidades
# ─────────────────────────────────────────────────────────────────────────────
def normalizar(texto: str) -> str:
    """Mayúsculas, sin tildes, espacios colapsados (para comparar textos)."""
    if not texto:
        return ""
    t = unicodedata.normalize("NFKD", str(texto))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", t).strip().upper()


def norm_factura(valor: str) -> str:
    """HUS0000374152 / FE-00123 / 0001234 → clave tolerante a ceros y guiones."""
    v = normalizar(valor).replace(" ", "").replace("-", "")
    m = re.match(r"^([A-Z]*)0*(\d+)$", v)
    if m and m.group(2):
        return m.group(1) + m.group(2)
    return v


_RE_FACTURA_TOKEN = re.compile(r"^[A-Z]{2,}0*\d{3,}$")


def factura_desde_nombre(nombre: str) -> str:
    """Extrae el número de factura del nombre de un archivo o carpeta.

    Convención HUS: 680010079201_HUS374152_EPICRIS.pdf → HUS374152.
    También acepta 'HUS374152', 'FACTURA HUS999 (2)', etc. Devuelve ""
    si el nombre no contiene un identificador de factura reconocible
    (prefijo alfabético + dígitos).
    """
    base = normalizar(nombre.rsplit(".", 1)[0])
    for token in base.replace("-", "_").replace(" ", "_").split("_"):
        if _RE_FACTURA_TOKEN.match(token):
            return token
    m = re.search(r"\b([A-Z]{2,}0*\d{3,})\b", base)
    return m.group(1) if m else ""


def parse_fecha_ddmmaaaa(valor: str) -> date | None:
    try:
        return datetime.strptime(valor.strip(), "%d/%m/%Y").date()
    except (ValueError, AttributeError):
        return None


_MESES_ES = {
    "ENE": 1,
    "FEB": 2,
    "MAR": 3,
    "ABR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AGO": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DIC": 12,
}


def parse_fecha_flexible(valor: str) -> date | None:
    """Fechas en cualquiera de los formatos vistos en soportes.

    Soporta: DD/MM/AAAA, D/MM/AAAA, AAAA-MM-DD, 'AAAA-MM-DD HH:MM',
    '11 feb. 2025', ISO con hora.
    """
    if not valor:
        return None
    v = str(valor).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(v[:10], fmt).date()
        except ValueError:
            pass
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})", v)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            return None
    m = re.match(r"^(\d{1,2})\s+([A-Za-z]{3})\w*\.?\s+(\d{4})", v)
    if m:
        mes = _MESES_ES.get(normalizar(m.group(2))[:3])
        if mes:
            try:
                return date(int(m.group(3)), mes, int(m.group(1)))
            except ValueError:
                return None
    return None


def parse_hora(valor: str) -> tuple[int, int] | None:
    m = re.match(r"^(\d{1,2}):(\d{2})$", (valor or "").strip())
    if not m:
        return None
    h, mi = int(m.group(1)), int(m.group(2))
    if h > 23 or mi > 59:
        return None
    return (h, mi)


def a_entero(valor: str) -> int | None:
    v = (valor or "").strip()
    if not v:
        return None
    try:
        return int(float(v))
    except ValueError:
        return None


def extraer_montos(texto: str) -> set[int]:
    """Todos los valores monetarios de un texto ('$ 95.400,00' → 95400)."""
    montos: set[int] = set()
    for m in re.finditer(r"\$?\s*([\d]{1,3}(?:[.,][\d]{3})*(?:[.,]\d{2})?|\d{4,})", texto):
        crudo = m.group(1)
        limpio = crudo
        # quitar decimales ',00' / '.00' (un sufijo [.,]dd es siempre decimal:
        # los grupos de miles son de 3 dígitos) y separadores de miles
        if re.search(r"[.,]\d{2}$", limpio):
            limpio = limpio[:-3]
        limpio = limpio.replace(".", "").replace(",", "")
        if limpio.isdigit() and len(limpio) >= 3:
            montos.add(int(limpio))
    return montos


def extraer_fechas(texto: str) -> set[date]:
    """Todas las fechas reconocibles de un texto (varios formatos)."""
    fechas: set[date] = set()
    for m in re.finditer(r"\b(\d{1,2})/(\d{1,2})/(\d{4})\b", texto):
        with contextlib.suppress(ValueError):
            fechas.add(date(int(m.group(3)), int(m.group(2)), int(m.group(1))))
    for m in re.finditer(r"\b(\d{4})-(\d{2})-(\d{2})\b", texto):
        with contextlib.suppress(ValueError):
            fechas.add(date(int(m.group(1)), int(m.group(2)), int(m.group(3))))
    for m in re.finditer(r"\b(\d{1,2})\s+([A-Za-z]{3})\w*\.?\s+(?:de\s+)?(\d{4})", texto):
        mes = _MESES_ES.get(normalizar(m.group(2))[:3])
        if mes:
            with contextlib.suppress(ValueError):
                fechas.add(date(int(m.group(3)), mes, int(m.group(1))))
    return fechas


def cargar_json_tolerante(ruta: Path):
    """Devuelve el JSON parseado (cualquier tipo) o None si no parsea.
    El llamador debe verificar isinstance(x, dict) antes de usar .get()."""
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with ruta.open(encoding=enc) as fh:
                return json.load(fh)
        except (UnicodeDecodeError, json.JSONDecodeError):
            continue
        except OSError:
            return None
    return None


def leer_texto_tolerante(ruta: Path) -> str:
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return ruta.read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
        except OSError as e:
            logger.error(f"No pude leer {ruta}: {e}")
            return ""
    try:
        return ruta.read_text(encoding="utf-8", errors="replace")
    except OSError as e:
        logger.error(f"No pude leer {ruta}: {e}")
        return ""


# ─── Extracción de texto PDF (pdfplumber → pypdf → nada) ─────────────────────
_MOTOR_PDF = ""


def _detectar_motor_pdf() -> str:
    global _MOTOR_PDF
    if _MOTOR_PDF:
        return _MOTOR_PDF
    try:
        import pdfplumber  # noqa: F401

        _MOTOR_PDF = "pdfplumber"
    except ImportError:
        try:
            import pypdf  # noqa: F401

            _MOTOR_PDF = "pypdf"
        except ImportError:
            _MOTOR_PDF = "ninguno"
    return _MOTOR_PDF


def extraer_texto_pdf(ruta: Path) -> tuple[str, int]:
    """Devuelve (texto, nro_paginas). Texto vacío si no hay motor o es escaneado."""
    motor = _detectar_motor_pdf()
    try:
        if motor == "pdfplumber":
            import pdfplumber

            with pdfplumber.open(ruta) as pdf:
                paginas = len(pdf.pages)
                partes = [(p.extract_text() or "") for p in pdf.pages]
                return ("\n".join(partes), paginas)
        if motor == "pypdf":
            from pypdf import PdfReader

            reader = PdfReader(str(ruta))
            partes = [(p.extract_text() or "") for p in reader.pages]
            return ("\n".join(partes), len(reader.pages))
    except Exception as e:  # PDF corrupto/cifrado: se reporta, no se aborta
        logger.warning(f"  No pude leer el PDF {ruta.name}: {e}")
        return ("", 0)
    return ("", 0)


# ─────────────────────────────────────────────────────────────────────────────
# Modelo de hallazgos
# ─────────────────────────────────────────────────────────────────────────────
@dataclass
class Hallazgo:
    factura: str
    origen: str  # MALLA FURIPS1 | FURIPS2 | CRUCE RIPS | CRUCE CUV | CRUCE FEV XML | CRUCE FACTURA PDF | CRUCE EPICRISIS | SOPORTES | ESTRUCTURA
    severidad: str
    campo: str  # "23" o "F2 L4 c7" o "-"
    concepto: str
    valor_furips: str
    valor_soporte: str
    fuente: str
    regla: str
    detalle: str


@dataclass
class ResultadoFactura:
    factura: str
    registro_f1: list[str] | None = None
    lineas_f2: list[list[str]] = field(default_factory=list)
    carpeta: Path | None = None
    archivos: list[Path] = field(default_factory=list)
    hallazgos: list[Hallazgo] = field(default_factory=list)
    estado_campos_f1: dict[int, tuple[str, str]] = field(default_factory=dict)  # n -> (estado, obs)
    estado_lineas_f2: list[tuple[str, str]] = field(default_factory=list)
    soportes: list[dict] = field(default_factory=list)
    cruces: list[dict] = field(default_factory=list)
    datos_soportes: dict = field(default_factory=dict)

    def agregar(
        self,
        origen: str,
        severidad: str,
        campo: str,
        concepto: str,
        valor_furips: str,
        regla: str,
        detalle: str = "",
        valor_soporte: str = "",
        fuente: str = "",
    ) -> None:
        self.hallazgos.append(
            Hallazgo(
                factura=self.factura,
                origen=origen,
                severidad=severidad,
                campo=str(campo),
                concepto=concepto,
                valor_furips=valor_furips,
                valor_soporte=valor_soporte,
                fuente=fuente,
                regla=regla,
                detalle=detalle,
            )
        )

    @property
    def errores(self) -> int:
        return sum(1 for h in self.hallazgos if h.severidad == ERROR)

    @property
    def advertencias(self) -> int:
        return sum(1 for h in self.hallazgos if h.severidad == ADVERTENCIA)


# ─────────────────────────────────────────────────────────────────────────────
# Lectura de archivos FURIPS
# ─────────────────────────────────────────────────────────────────────────────
_RE_ARCHIVO_F1 = re.compile(r"^FURIPS\s*_?1", re.IGNORECASE)
_RE_ARCHIVO_F2 = re.compile(r"^FURIPS\s*_?2", re.IGNORECASE)


def descubrir_archivos_furips(raiz: Path) -> tuple[list[Path], list[Path]]:
    """Encuentra los TXT FURIPS1/FURIPS2 en la raíz y todas sus subcarpetas."""
    f1, f2 = [], []
    for p in sorted(raiz.rglob("*.txt")):
        if _RE_ARCHIVO_F1.match(p.name):
            f1.append(p)
        elif _RE_ARCHIVO_F2.match(p.name):
            f2.append(p)
    return f1, f2


def validar_nombre_archivo(ruta: Path, tipo: str) -> list[str]:
    """Valida FURIPS<t>CODIGOHABILITACIONDDMMAAAA.txt. Devuelve observaciones."""
    obs = []
    base = ruta.stem
    resto = re.sub(rf"^FURIPS\s*_?{tipo}", "", base, flags=re.IGNORECASE)
    if len(resto) < 20:
        obs.append(
            f"Nombre de archivo '{ruta.name}' no cumple la estructura "
            f"FURIPS{tipo}CODIGOHABILITACION(12)DDMMAAAA(8) de la Circular 022/2023."
        )
        return obs
    codigo, fecha = resto[:12], resto[12:20]
    if not codigo.isdigit():
        obs.append(f"'{ruta.name}': el código de habilitación '{codigo}' no es numérico.")
    try:
        datetime.strptime(fecha, "%d%m%Y")
    except ValueError:
        obs.append(f"'{ruta.name}': la fecha de generación '{fecha}' no es DDMMAAAA válida.")
    return obs


def leer_furips1(rutas: list[Path]) -> tuple[dict[str, list[list[str]]], list[str]]:
    """Lee y agrupa registros FURIPS1 por factura normalizada.

    Devuelve ({factura_norm: [registros]}, observaciones_de_archivo).
    """
    registros: dict[str, list[list[str]]] = {}
    obs: list[str] = []
    for ruta in rutas:
        obs.extend(validar_nombre_archivo(ruta, "1"))
        contenido = leer_texto_tolerante(ruta)
        for nro, linea in enumerate(contenido.splitlines(), 1):
            if not linea.strip():
                continue
            campos = linea.split(",")
            if len(campos) > 102:
                # comas de más: casi siempre en la descripción final (campo 102)
                campos = campos[:101] + [",".join(campos[101:])]
                obs.append(
                    f"{ruta.name} línea {nro}: {len(linea.split(','))} campos "
                    f"(se esperan 102). Hay comas dentro de un campo de texto; "
                    f"la malla ADRES rechaza el registro (se unieron en el campo 102)."
                )
            elif len(campos) < 102:
                obs.append(
                    f"{ruta.name} línea {nro}: solo {len(campos)} campos "
                    f"(se esperan 102). Registro incompleto."
                )
                campos = campos + [""] * (102 - len(campos))
            campos = [c.strip() for c in campos]
            clave = norm_factura(campos[2])
            if not clave:
                obs.append(f"{ruta.name} línea {nro}: sin número de factura (campo 3).")
                clave = f"(SIN FACTURA) {ruta.name}:{nro}"
            registros.setdefault(clave, []).append(campos)
    return registros, obs


def leer_furips2(rutas: list[Path]) -> tuple[dict[str, list[list[str]]], list[str]]:
    """Lee y agrupa líneas FURIPS2 por factura normalizada."""
    lineas: dict[str, list[list[str]]] = {}
    obs: list[str] = []
    for ruta in rutas:
        obs.extend(validar_nombre_archivo(ruta, "2"))
        contenido = leer_texto_tolerante(ruta)
        for nro, linea in enumerate(contenido.splitlines(), 1):
            if not linea.strip():
                continue
            campos = linea.split(",")
            if len(campos) > 9:
                # comas de más → dentro de la descripción (campo 5)
                campos = campos[:4] + [",".join(campos[4 : len(campos) - 4])] + campos[-4:]
                obs.append(
                    f"{ruta.name} línea {nro}: comas dentro de la descripción "
                    f"(campo 5); la malla ADRES rechaza el archivo."
                )
            elif len(campos) < 9:
                obs.append(f"{ruta.name} línea {nro}: solo {len(campos)} campos (se esperan 9).")
                campos = campos + [""] * (9 - len(campos))
            campos = [c.strip() for c in campos]
            clave = norm_factura(campos[0])
            if not clave:
                obs.append(f"{ruta.name} línea {nro}: sin número de factura (campo 1).")
                clave = f"(SIN FACTURA) {ruta.name}:{nro}"
            lineas.setdefault(clave, []).append(campos)
    return lineas, obs


# ─────────────────────────────────────────────────────────────────────────────
# Malla FURIPS 1 (Circular 022/2023, Tabla 1)
# ─────────────────────────────────────────────────────────────────────────────
def _hay_lineas_quirurgicas(lineas_f2: list[list[str]]) -> bool:
    """Línea principal quirúrgica: tipo 2 con cantidad 0 y valores 0."""
    return any(ln[2] == "2" and a_entero(ln[5]) == 0 and a_entero(ln[7]) == 0 for ln in lineas_f2)


def _evaluar_condicion(cond: str, c: dict[int, str], ctx: dict) -> tuple[bool, bool]:
    """Devuelve (obligatorio, debe_estar_vacio) para la condición dada."""
    estado = c.get(28, "")
    nat = c.get(20, "")
    if cond == "SI":
        return True, False
    if cond == "NO":
        return False, False
    if cond == "C_RGO":
        return bool(c.get(2, "")), False
    if cond == "C_NAT01":
        return nat == "01", False
    if cond == "C_NAT17":
        return nat == "17", False
    if cond == "C_VEH_NO3":
        # Con vehículo fantasma (3) pasan a ser OPCIONALES (no prohibidos)
        return (nat == "01" and estado not in ("", "3")), False
    if cond == "C_VEH_PLACA":
        return estado in {"1", "2", "4", "6", "7"}, estado == "3"
    if cond == "C_VEH_POLIZA":
        return estado in {"1", "4", "6"}, False
    if cond in ("C_PROP", "C_PROP_NAT"):
        # Persona jurídica: razón social va en 46 y el primer nombre (48) puede ir vacío.
        # Estados 7 (propietario indeterminado) y 8: la Tabla 1 y el instructivo VII
        # se contradicen — se degradan a ADVERTENCIA aparte (ver validar_malla_f1).
        oblig = estado in {"1", "2", "4", "6"}
        if cond == "C_PROP_NAT" and c.get(44, "") in ("NI", "NIT"):
            oblig = False
        return oblig, estado == "3"
    if cond == "C_COND":
        return estado in {"1", "2", "4", "6", "8"}, estado == "3"
    if cond == "C_REMISION":
        return ctx["hay_remision"], False
    if cond == "C_TRANSP":
        return ctx["hay_transporte"], False
    if cond == "C_QX":
        return ctx["hay_qx"], False
    if cond == "C_UCI":
        return (c.get(42, "") == "1" or ctx.get("hay_uci", False)), False
    if cond == "C_ESTANCIA":
        return ctx.get("hay_estancia", False), False
    if cond == "C_UCI_F2":
        return ctx.get("hay_uci", False), False
    return False, False


def validar_malla_f1(res: ResultadoFactura) -> None:
    reg = res.registro_f1
    if reg is None:
        res.agregar(
            "ESTRUCTURA",
            ERROR,
            "-",
            "Registro FURIPS 1",
            "",
            "Circular 022/2023",
            (
                "La factura tiene carpeta de soportes pero NO tiene registro en el FURIPS 1."
                if res.carpeta is not None
                else "La factura NO tiene registro en el FURIPS 1."
            ),
        )
        return
    c = {n: reg[n - 1] for n in range(1, 103)}

    ctx = {
        "hay_remision": any(c.get(n, "") for n in range(64, 75)),
        "hay_transporte": (a_entero(c.get(99, "")) or 0) > 0
        or any(ln[2] == "3" for ln in res.lineas_f2),
        "hay_qx": _hay_lineas_quirurgicas(res.lineas_f2),
        # Heurística sobre el FURIPS2: códigos SOAT de estancia (381xx) y de
        # cuidado intensivo (385xx), o descripciones de estancia/UCI.
        "hay_estancia": any(
            ln[2] == "2"
            and (
                ln[3].startswith(("381", "385"))
                or re.search(
                    r"ESTANCIA|HABITACI|INTERNACI|CUIDADO\s+(INTENSIVO|INTERMEDIO)|\bUCI\b",
                    ln[4],
                    re.IGNORECASE,
                )
            )
            for ln in res.lineas_f2
        ),
        "hay_uci": any(
            ln[2] == "2"
            and (
                ln[3].startswith("385")
                or re.search(r"CUIDADO\s+INTENSIVO|\bUCI\b", ln[4], re.IGNORECASE)
            )
            for ln in res.lineas_f2
        ),
    }

    nat = c.get(20, "")
    estado_aseg = c.get(28, "")
    seccion_v_prohibida = nat in NATURALEZAS_EVENTO and nat != "01"

    for n, _seccion, nombre, longitud, formato, permitidos, oblig in E1:
        valor = c.get(n, "")
        estado, obs = OK, ""
        obligatorio, debe_vacio = _evaluar_condicion(oblig, c, ctx)

        # Instructivo secc. V: solo se diligencia en accidentes de tránsito.
        if seccion_v_prohibida and 28 <= n <= 37:
            obligatorio = False
            if valor:
                debe_vacio = True

        if debe_vacio and valor:
            estado = ERROR
            if seccion_v_prohibida and 28 <= n <= 37:
                obs = (
                    f"La sección V (vehículo) solo se diligencia en accidentes de "
                    f"tránsito (naturaleza 01); con naturaleza {nat} debe ir vacía "
                    f"y tiene '{valor}' (instructivo Circular 022/2023)."
                )
            else:
                obs = f"Debe estar VACÍO cuando el vehículo es fantasma (28=3) y tiene '{valor}'."
        elif obligatorio and not valor:
            estado = ERROR
            obs = f"Campo obligatorio sin diligenciar ({TEXTO_OBLIG.get(oblig, oblig)})."
            # Activación por heurística sobre el FURIPS2 (no por un campo del
            # propio registro): se degrada a ADVERTENCIA para acotar falsos
            # positivos si la IPS usa códigos atípicos.
            if oblig in ("C_ESTANCIA", "C_UCI_F2") or (oblig == "C_UCI" and c.get(42, "") != "1"):
                estado = ADVERTENCIA
                obs += " (Detectado por las líneas de estancia/UCI del FURIPS 2.)"
        elif (
            not valor
            and oblig in ("C_PROP", "C_PROP_NAT")
            and estado_aseg in ("7", "8")
            and not (oblig == "C_PROP_NAT" and c.get(44, "") in ("NI", "NIT"))
        ):
            estado = ADVERTENCIA
            obs = (
                f"Vacío con estado de aseguramiento {estado_aseg}: la Tabla 1 y el "
                "instructivo VII de la Circular se contradicen sobre su exigencia "
                "(el 7 se define como 'se desconocen los datos del propietario'). "
                "Documentar el caso en la reclamación."
            )
        elif valor:
            if len(valor) > longitud:
                estado = ERROR
                obs = f"Longitud {len(valor)} supera el máximo {longitud} de la Circular."
            if permitidos and valor not in permitidos:
                estado = ERROR
                obs = (
                    obs + " " if obs else ""
                ) + f"Valor '{valor}' fuera de los permitidos: {', '.join(sorted(permitidos))}."
            if formato == "fecha" and not (_RE_FECHA.match(valor) and parse_fecha_ddmmaaaa(valor)):
                estado = ERROR
                obs = (obs + " " if obs else "") + f"'{valor}' no es fecha DD/MM/AAAA válida."
            elif formato == "hora" and not (_RE_HORA.match(valor) and parse_hora(valor)):
                estado = ERROR
                obs = (obs + " " if obs else "") + f"'{valor}' no es hora HH:MM (24h) válida."
            elif formato == "num" and not _RE_NUM.match(valor):
                estado = ERROR
                obs = (obs + " " if obs else "") + f"'{valor}' debe ser numérico sin separadores."
            elif formato == "cie10" and not _RE_CIE10.match(valor):
                estado = ERROR
                obs = (obs + " " if obs else "") + f"'{valor}' no tiene formato CIE-10 (ej. S423)."
            elif formato == "depto":
                if not (len(valor) == 2 and valor.isdigit()):
                    estado = ERROR
                    obs = (obs + " " if obs else "") + f"'{valor}' no es código DANE de 2 dígitos."
                elif valor not in DEPTOS_DANE:
                    estado = ERROR
                    obs = (
                        obs + " " if obs else ""
                    ) + f"'{valor}' no es un departamento DANE válido."
            elif formato == "mun" and not (len(valor) == 3 and valor.isdigit()):
                estado = ERROR
                obs = (
                    obs + " " if obs else ""
                ) + f"'{valor}' no es código DANE de municipio (3 dígitos)."
            elif formato == "placa" and not _RE_PLACA.match(normalizar(valor).replace(" ", "")):
                estado = ERROR
                obs = (
                    obs + " " if obs else ""
                ) + f"Placa '{valor}' con caracteres no permitidos (solo letras y números)."
            if "�" in valor:
                if estado == OK:
                    estado = ADVERTENCIA
                obs = (
                    obs + " " if obs else ""
                ) + "Caracteres dañados por codificación (�): la malla puede rechazarlo."
        else:
            obs = "Vacío (permitido)."

        res.estado_campos_f1[n] = (estado, obs)
        if estado in (ERROR, ADVERTENCIA):
            res.agregar(
                "MALLA FURIPS1",
                estado,
                n,
                nombre,
                valor,
                f"Circular 022/2023 Tabla 1 campo {n} — {TEXTO_OBLIG.get(oblig, 'Formato/valores')}",
                obs,
            )

    _validar_coherencia_f1(res, c, ctx)


def _validar_coherencia_f1(res: ResultadoFactura, c: dict[int, str], ctx: dict) -> None:
    """Reglas de coherencia interna del FURIPS 1 (fechas, edades, personas)."""
    f_nac = parse_fecha_ddmmaaaa(c.get(12, ""))
    f_evento = parse_fecha_ddmmaaaa(c.get(23, ""))
    f_ingreso = parse_fecha_ddmmaaaa(c.get(80, ""))
    f_egreso = parse_fecha_ddmmaaaa(c.get(82, ""))
    h_ingreso = parse_hora(c.get(81, ""))
    h_egreso = parse_hora(c.get(83, ""))

    if f_ingreso and f_egreso:
        dt_in = datetime(f_ingreso.year, f_ingreso.month, f_ingreso.day, *(h_ingreso or (0, 0)))
        dt_eg = datetime(f_egreso.year, f_egreso.month, f_egreso.day, *(h_egreso or (23, 59)))
        if dt_eg < dt_in:
            res.agregar(
                "MALLA FURIPS1",
                ERROR,
                "80-83",
                "Fechas de ingreso/egreso",
                f"ingreso {c.get(80)} {c.get(81)} / egreso {c.get(82)} {c.get(83)}",
                "Coherencia XI. Certificación de la atención",
                "La fecha/hora de EGRESO es anterior a la de INGRESO.",
            )
    if f_evento and f_ingreso and f_ingreso < f_evento:
        res.agregar(
            "MALLA FURIPS1",
            ERROR,
            "23/80",
            "Fecha evento vs fecha ingreso",
            f"evento {c.get(23)} / ingreso {c.get(80)}",
            "Coherencia IV vs XI",
            "La fecha de ingreso es ANTERIOR a la fecha del evento/accidente.",
        )
    if f_nac:
        if f_evento and f_nac > f_evento:
            res.agregar(
                "MALLA FURIPS1",
                ERROR,
                "12/23",
                "Fecha nacimiento vs evento",
                f"nacimiento {c.get(12)} / evento {c.get(23)}",
                "Coherencia III vs IV",
                "La fecha de nacimiento es posterior a la fecha del evento.",
            )
        elif f_ingreso:
            edad = (
                f_ingreso.year
                - f_nac.year
                - ((f_ingreso.month, f_ingreso.day) < (f_nac.month, f_nac.day))
            )
            tipo_doc = c.get(10, "")
            if tipo_doc == "TI" and edad >= 18:
                res.agregar(
                    "MALLA FURIPS1",
                    ADVERTENCIA,
                    "10/12",
                    "Tipo doc vs edad",
                    f"TI con {edad} años",
                    "Coherencia identificación",
                    "Víctima con Tarjeta de Identidad pero mayor de edad a la fecha de "
                    "atención: causal frecuente de glosa (debería ser CC).",
                )
            if tipo_doc == "CC" and edad < 18:
                res.agregar(
                    "MALLA FURIPS1",
                    ADVERTENCIA,
                    "10/12",
                    "Tipo doc vs edad",
                    f"CC con {edad} años",
                    "Coherencia identificación",
                    "Víctima con Cédula de Ciudadanía pero menor de edad.",
                )

    # Condición de la víctima = conductor → debe coincidir con los datos del conductor
    if c.get(19, "") == "1" and c.get(59, ""):
        if c.get(59) != c.get(11):
            res.agregar(
                "MALLA FURIPS1",
                ADVERTENCIA,
                "19/59",
                "Víctima conductora vs conductor",
                f"víctima {c.get(10)} {c.get(11)} / conductor {c.get(58)} {c.get(59)}",
                "Coherencia III vs VIII",
                "La víctima figura como CONDUCTOR (19=1) pero el documento del "
                "conductor (59) no coincide con el de la víctima (11).",
            )
        elif c.get(58, "") and c.get(58) != c.get(10):
            res.agregar(
                "MALLA FURIPS1",
                ADVERTENCIA,
                "10/58",
                "Tipo doc víctima vs conductor",
                f"víctima {c.get(10)} / conductor {c.get(58)}",
                "Coherencia III vs VIII",
                "Misma persona (víctima=conductor) con TIPO de documento distinto "
                "en las secciones III y VIII.",
            )

    # Vigencia de póliza vs fecha del evento
    f_v1 = parse_fecha_ddmmaaaa(c.get(34, ""))
    f_v2 = parse_fecha_ddmmaaaa(c.get(35, ""))
    if f_evento and f_v1 and f_v2 and not (f_v1 <= f_evento <= f_v2):
        res.agregar(
            "MALLA FURIPS1",
            ADVERTENCIA,
            "34/35",
            "Vigencia póliza vs evento",
            f"vigencia {c.get(34)}–{c.get(35)} / evento {c.get(23)}",
            "Coherencia V",
            "El evento está FUERA de la vigencia de la póliza reportada.",
        )

    # Días UCI vs estancia
    dias_uci = a_entero(c.get(43, ""))
    if dias_uci and f_ingreso and f_egreso:
        estancia = (f_egreso - f_ingreso).days
        if dias_uci > max(estancia, 1):
            res.agregar(
                "MALLA FURIPS1",
                ERROR,
                "43",
                "Días de UCI vs estancia",
                f"UCI {dias_uci} días / estancia {estancia} días",
                "Circular 022/2023 campo 43",
                "Los días de UCI superan la estancia total (ingreso→egreso).",
            )

    if c.get(32, "") and c.get(28, "") in ("2", "7", "8"):
        res.agregar(
            "MALLA FURIPS1",
            ADVERTENCIA,
            "32",
            "Código de aseguradora con vehículo no asegurado",
            c.get(32, ""),
            "Instructivo V Circular 022/2023",
            "El vehículo figura como no asegurado y trae código de aseguradora: "
            "'la ADRES no es una aseguradora, no debe ir ningún código si el "
            "vehículo es no asegurado'.",
        )

    if c.get(37, "") == "1" and c.get(28, "") != "6":
        res.agregar(
            "MALLA FURIPS1",
            ADVERTENCIA,
            "37",
            "Cobro agotamiento tope",
            f"37=1 con estado aseguramiento {c.get(28)}",
            "Circular 022/2023 campo 37",
            "El cobro por agotamiento de tope solo aplica con estado "
            "de aseguramiento 6 (Asegurado D.2497).",
        )

    if c.get(101, "") == "0":
        res.agregar(
            "MALLA FURIPS1",
            ADVERTENCIA,
            "101",
            "Manifestación de servicios habilitados",
            "0",
            "Circular 022/2023 campo 101",
            "El prestador manifiesta NO tener los servicios habilitados (101=0): "
            "revisar antes de radicar.",
        )

    # Estado 7 (propietario indeterminado): el instructivo pide 'sin información' en conductor
    if c.get(28, "") == "7" and not c.get(54, ""):
        res.agregar(
            "MALLA FURIPS1",
            ADVERTENCIA,
            "54-63",
            "Datos del conductor",
            "",
            "Instructivo VIII Circular 022/2023",
            "Con estado 7, el instructivo indica registrar 'sin información' en los "
            "datos del conductor y anexar el soporte correspondiente.",
        )


# ─────────────────────────────────────────────────────────────────────────────
# FURIPS 2 y coherencia FURIPS1 ↔ FURIPS2
# ─────────────────────────────────────────────────────────────────────────────
def validar_furips2(res: ResultadoFactura) -> None:
    reg = res.registro_f1
    c4 = reg[3] if reg else ""
    c3 = reg[2] if reg else ""

    if not res.lineas_f2:
        if reg:
            res.agregar(
                "FURIPS2",
                ERROR,
                "-",
                "Líneas FURIPS 2",
                "",
                "Circular 022/2023 Tabla 2",
                "La factura tiene registro FURIPS 1 pero NINGUNA línea en el FURIPS 2.",
            )
        return

    if reg is None:
        res.agregar(
            "FURIPS2",
            ERROR,
            "-",
            "Líneas FURIPS 2 sin registro FURIPS 1",
            f"{len(res.lineas_f2)} líneas",
            "Circular 022/2023 Tabla 2",
            "La factura tiene líneas en el FURIPS 2 pero NO aparece en el "
            "FURIPS 1: la malla ADRES rechaza el archivo.",
        )

    # Cada observación lleva su propia severidad: (severidad, texto).
    # (línea, severidad, obs, tipo, valor de la línea)
    hallazgos_linea: list[tuple[int, str, str, str, str]] = []
    vistas: dict[tuple, int] = {}
    for i, ln in enumerate(res.lineas_f2, 1):
        partes: list[tuple[str, str]] = []
        (f2_factura, f2_consec, tipo, codigo, descripcion, cantidad, vr_unit, vr_fact, vr_recl) = ln

        if c3 and norm_factura(f2_factura) != norm_factura(c3):
            partes.append((ERROR, f"Factura '{f2_factura}' no coincide con el FURIPS 1 ('{c3}')."))
        if c4 and f2_consec != c4:
            partes.append(
                (
                    ERROR,
                    f"Consecutivo '{f2_consec}' no coincide con el campo 4 del FURIPS 1 ('{c4}').",
                )
            )
        if tipo not in TIPOS_SERVICIO_F2:
            partes.append((ERROR, f"Tipo de servicio '{tipo}' fuera de los permitidos (1-8)."))
        if tipo in {"3", "4", "5", "7", "8"} and codigo:
            partes.append(
                (
                    ERROR,
                    f"Código '{codigo}' debe ir VACÍO para tipo de servicio {tipo} "
                    f"({TIPOS_SERVICIO_F2.get(tipo, '')}).",
                )
            )
        if tipo in {"1", "2", "6"} and not codigo:
            detalle_cod = {
                "1": "para medicamentos debe ir el CUM/IUM INVIMA",
                "2": "para procedimientos debe ir el código del manual tarifario SOAT",
                "6": "para dispositivos médicos debe ir el registro sanitario INVIMA",
            }[tipo]
            partes.append((ERROR, f"Código del servicio VACÍO: {detalle_cod} (campo 4 Tabla 2)."))
        if tipo == "1" and codigo and not re.match(r"^\d+-\d+$|^\d{4,}$", codigo):
            partes.append(
                (ADVERTENCIA, f"Código '{codigo}' no parece CUM (expediente-consecutivo) ni IUM.")
            )
        if tipo in {"3", "4", "5", "6", "7", "8"} and not descripcion:
            partes.append((ERROR, f"Descripción obligatoria para tipo de servicio {tipo}."))

        n_cant, n_unit, n_fact, n_recl = (
            a_entero(cantidad),
            a_entero(vr_unit),
            a_entero(vr_fact),
            a_entero(vr_recl),
        )
        for etiqueta, crudo, num in (
            ("cantidad", cantidad, n_cant),
            ("valor unitario", vr_unit, n_unit),
            ("valor facturado", vr_fact, n_fact),
            ("valor reclamado", vr_recl, n_recl),
        ):
            if crudo == "" or num is None:
                partes.append(
                    (ERROR, f"{etiqueta.capitalize()} '{crudo}' no es numérico u obligatorio.")
                )
            elif "." in crudo or "," in crudo:
                partes.append(
                    (ERROR, f"{etiqueta.capitalize()} '{crudo}' trae separadores/decimales.")
                )

        es_linea_principal_qx = (
            tipo == "2" and n_cant == 0 and n_unit == 0 and n_fact == 0 and n_recl == 0
        )
        if None not in (n_cant, n_unit, n_fact) and not es_linea_principal_qx:
            if n_cant <= 0:
                partes.append(
                    (ERROR, "Cantidad debe ser mayor a cero (salvo línea principal quirúrgica).")
                )
            for etq_v, num_v in (
                ("Valor unitario", n_unit),
                ("Valor facturado", n_fact),
                ("Valor reclamado", n_recl),
            ):
                if num_v is not None and num_v <= 0:
                    partes.append(
                        (
                            ERROR,
                            f"{etq_v} debe ser mayor a cero (campos 7-9 Tabla 2, "
                            "salvo línea principal quirúrgica).",
                        )
                    )
            if n_cant * n_unit != n_fact:
                partes.append(
                    (
                        ERROR,
                        f"Valor facturado {n_fact:,} ≠ cantidad {n_cant} × unitario {n_unit:,}.",
                    )
                )
        if n_recl is not None and n_fact is not None and n_recl > n_fact:
            partes.append((ERROR, f"Valor reclamado {n_recl:,} SUPERA el facturado {n_fact:,}."))

        clave = tuple(ln)
        if clave in vistas:
            partes.append(
                (
                    ADVERTENCIA,
                    f"Línea idéntica a la línea {vistas[clave]}: la Circular exige agrupar "
                    f"servicios idénticos en una sola línea.",
                )
            )
        else:
            vistas[clave] = i

        for pos, (_n, _nom, lmax, _f, _o) in enumerate(E2):
            if len(ln[pos]) > lmax:
                partes.append((ERROR, f"Campo {pos + 1} supera longitud máxima {lmax}."))

        estado = ERROR if any(s == ERROR for s, _ in partes) else ADVERTENCIA if partes else OK
        obs = " ".join(t for _s, t in partes) if partes else "Línea válida."
        res.estado_lineas_f2.append((estado, obs))
        valor_ln = " | ".join(
            x
            for x in (
                f"cód '{codigo}'" if codigo else "cód vacío",
                descripcion[:30],
                f"cant {cantidad}",
                f"vr.fact {vr_fact}",
                f"vr.recl {vr_recl}",
            )
            if x
        )
        for sev, texto in partes:
            hallazgos_linea.append((i, sev, texto, tipo, valor_ln))

    # Agrupar los hallazgos por PATRÓN para no repetir cientos de filas iguales
    # (el detalle línea a línea queda completo en la hoja FURIPS2 LINEAS).
    grupos: dict[tuple[str, str], list[tuple[int, str, str, str]]] = {}
    for i, sev, texto, tipo, valor_ln in hallazgos_linea:
        patron = re.sub(r"\d[\d.,]*", "N", re.sub(r"'[^']*'", "'…'", texto))
        grupos.setdefault((sev, patron), []).append((i, texto, tipo, valor_ln))
    for (sev, _patron), casos in grupos.items():
        lineas_txt = ", ".join(f"L{i}" for i, _p, _t, _v in casos[:8])
        if len(casos) > 8:
            lineas_txt += f" … (+{len(casos) - 8} más)"
        tipos_grupo = {t for _i, _p, t, _v in casos}
        tipo0 = casos[0][2]
        concepto = (
            f"{len(casos)} línea(s) — tipo {tipo0} {TIPOS_SERVICIO_F2.get(tipo0, '')}"
            if len(tipos_grupo) == 1
            else f"{len(casos)} línea(s)"
        )
        res.agregar(
            "FURIPS2",
            sev,
            lineas_txt,
            concepto,
            casos[0][3][:80] + (" (y otras líneas)" if len(casos) > 1 else ""),
            "Circular 022/2023 Tabla 2",
            casos[0][1]
            + (
                f" (Aplica a {len(casos)} líneas: {lineas_txt}. "
                f"Ver detalle en la hoja FURIPS2 LINEAS.)"
                if len(casos) > 1
                else f" (Línea {casos[0][0]}.)"
            ),
        )

    if reg:
        _validar_sumatorias(res, reg)


def _validar_sumatorias(res: ResultadoFactura, reg: list[str]) -> None:
    """Campos 97-100 del FURIPS1 vs sumatorias del FURIPS2 (Circular 022/2023)."""
    suma = {"fact_med": 0, "recl_med": 0, "fact_tra": 0, "recl_tra": 0}
    for ln in res.lineas_f2:
        fact, recl = a_entero(ln[7]) or 0, a_entero(ln[8]) or 0
        if ln[2] == "3":
            suma["fact_tra"] += fact
            suma["recl_tra"] += recl
        else:
            suma["fact_med"] += fact
            suma["recl_med"] += recl

    comparaciones = [
        (97, "fact_med", "Total facturado gastos médico-quirúrgicos"),
        (98, "recl_med", "Total reclamado gastos médico-quirúrgicos"),
        (99, "fact_tra", "Total facturado transporte primario"),
        (100, "recl_tra", "Total reclamado transporte primario"),
    ]
    for campo, clave, nombre in comparaciones:
        declarado = a_entero(reg[campo - 1])
        if declarado is None:
            continue
        if declarado != suma[clave]:
            res.agregar(
                "FURIPS2",
                ERROR,
                str(campo),
                nombre,
                f"{declarado:,}",
                "Circular 022/2023 campos 97-100",
                f"El FURIPS 1 declara {declarado:,} pero la sumatoria de las líneas "
                f"FURIPS 2 da {suma[clave]:,} (diferencia {declarado - suma[clave]:,}).",
                valor_soporte=f"{suma[clave]:,}",
                fuente="FURIPS 2",
            )

    recl, fact = a_entero(reg[97]), a_entero(reg[96])
    if recl is not None and fact is not None and recl > fact:
        res.agregar(
            "MALLA FURIPS1",
            ERROR,
            "97/98",
            "Reclamado vs facturado",
            f"facturado {fact:,} / reclamado {recl:,}",
            "Circular 022/2023",
            "El total reclamado supera el total facturado.",
        )


# ─────────────────────────────────────────────────────────────────────────────
# Soportes de la carpeta
# ─────────────────────────────────────────────────────────────────────────────
TOKENS_SOPORTE = {
    "RIP": "RIPS JSON",
    "RIPS": "RIPS JSON",
    "CUV": "CUV JSON",
    "FACTURA": "FACTURA",
    "FEV": "FACTURA XML",
    "EPICRIS": "EPICRISIS",
    "EPI": "EPICRISIS",
    "HEV": "HOJA EVOLUCIÓN",
    "HAU": "HOJA URGENCIAS",
    "INFOPOL": "INFORME POLICIAL",
    "IPO": "INFORME POLICIAL",
    "DQX": "DESCRIPCIÓN QUIRÚRGICA",
    "RAN": "REGISTRO ANESTESIA",
    "TAP": "TRASLADO ASISTENCIAL",
    "FMO": "FACT. OSTEOSÍNTESIS",
    "OPF": "ORDEN FACULTATIVA",
    "PDX": "APOYO DIAGNÓSTICO",
    "HAM": "ADMIN. MEDICAMENTOS",
    "CRC": "COMPROBANTE RECIBIDO",
}


def clasificar_soporte(nombre: str) -> str:
    base = nombre.rsplit(".", 1)[0]
    for t in reversed(base.replace("-", "_").split("_")):
        tu = t.upper()
        if tu in TOKENS_SOPORTE:
            tipo = TOKENS_SOPORTE[tu]
            if tipo == "FACTURA":
                return "FACTURA XML" if nombre.lower().endswith(".xml") else "FACTURA PDF"
            return tipo
    return "OTRO"


def analizar_carpeta(res: ResultadoFactura, sin_pdf: bool = False) -> None:
    """Inventaría los soportes de la factura, los lee y valida presencia."""
    if not res.archivos:
        res.agregar(
            "SOPORTES",
            ADVERTENCIA,
            "-",
            "Soportes de la factura",
            "",
            "Res. 2284/2023",
            "No se encontraron soportes para esta factura: ni carpeta propia ni "
            "archivos con el número de factura en el nombre "
            "(<codHabilitacion>_<factura>_<TIPO>.ext).",
        )
        return

    datos = res.datos_soportes
    encontrados: dict[str, Path] = {}
    for p in res.archivos:
        tipo = clasificar_soporte(p.name)
        if tipo not in encontrados:
            encontrados[tipo] = p
        try:
            tam = p.stat().st_size
        except OSError:
            tam = 0
        res.soportes.append(
            {
                "archivo": p.name,
                "tipo": tipo,
                "bytes": tam,
                "legible": "-",
                "obs": "",
            }
        )

    # RIPS
    if "RIPS JSON" in encontrados:
        rips = cargar_json_tolerante(encontrados["RIPS JSON"])
        if isinstance(rips, dict) and rips:
            datos["rips"] = rips
        else:
            res.agregar(
                "SOPORTES",
                ERROR,
                "-",
                "RIPS JSON",
                encontrados["RIPS JSON"].name,
                "Res. 2284/2023",
                "El RIPS no se pudo parsear como JSON."
                if rips is None
                else "El RIPS parsea pero está vacío o no tiene la estructura "
                "esperada (objeto con numFactura/usuarios).",
            )
    else:
        res.agregar(
            "SOPORTES",
            ERROR,
            "-",
            "RIPS JSON",
            "",
            "Res. 2284/2023",
            "FALTA el RIPS JSON de esta factura (soporte obligatorio ADRES).",
        )

    # CUV
    if "CUV JSON" in encontrados:
        cuv = cargar_json_tolerante(encontrados["CUV JSON"])
        if isinstance(cuv, dict) and cuv:
            datos["cuv"] = cuv
        else:
            res.agregar(
                "SOPORTES",
                ERROR,
                "-",
                "CUV JSON",
                encontrados["CUV JSON"].name,
                "Res. 2284/2023",
                "El CUV no se pudo parsear como JSON."
                if cuv is None
                else "El CUV parsea pero está vacío o no tiene la estructura "
                "esperada (objeto con NumFactura/CodigoUnicoValidacion).",
            )
    else:
        res.agregar(
            "SOPORTES",
            ERROR,
            "-",
            "CUV JSON",
            "",
            "Res. 2284/2023",
            "FALTA el CUV (código único de validación RIPS) de esta factura.",
        )

    # FEV XML
    if "FACTURA XML" in encontrados:
        datos["fev"] = _leer_fev(encontrados["FACTURA XML"])
    else:
        res.agregar(
            "SOPORTES",
            ERROR,
            "-",
            "Factura electrónica XML",
            "",
            "Res. 2284/2023",
            "FALTA la factura electrónica DIAN (XML) de esta factura.",
        )

    # PDFs
    motor = _detectar_motor_pdf()
    for tipo, etiqueta in (("FACTURA PDF", "factura_pdf"), ("EPICRISIS", "epicrisis")):
        if tipo not in encontrados:
            severidad = ERROR if tipo == "EPICRISIS" else ADVERTENCIA
            res.agregar(
                "SOPORTES",
                severidad,
                "-",
                tipo,
                "",
                "Res. 2284/2023",
                f"FALTA el soporte {tipo} para esta factura."
                + (
                    " (La epicrisis/resumen de atención es soporte obligatorio.)"
                    if tipo == "EPICRISIS"
                    else ""
                ),
            )
            continue
        if sin_pdf:
            continue
        if motor == "ninguno":
            res.agregar(
                "SOPORTES",
                ADVERTENCIA,
                "-",
                tipo,
                encontrados[tipo].name,
                "Dependencias",
                "No hay librería PDF instalada (pdfplumber o pypdf): "
                "no se pudo leer el contenido del PDF.",
            )
            continue
        texto, paginas = extraer_texto_pdf(encontrados[tipo])
        datos[etiqueta] = {"texto": texto, "paginas": paginas, "archivo": encontrados[tipo].name}
        for s in res.soportes:
            if s["archivo"] == encontrados[tipo].name:
                s["legible"] = "SI" if len(texto.strip()) > 50 else "NO (escaneado)"
        if len(texto.strip()) <= 50:
            res.agregar(
                "SOPORTES",
                ADVERTENCIA,
                "-",
                tipo,
                encontrados[tipo].name,
                "Legibilidad",
                f"El PDF ({paginas} páginas) no tiene capa de texto "
                "(parece escaneado): el cruce automático no es posible, revisar manualmente.",
            )


def _leer_fev(ruta: Path) -> dict:
    """Extrae encabezado y totales del FEV DIAN con regex tolerantes."""
    try:
        from factura_lectura import _texto_normalizado, cargar_factura_xml

        texto = _texto_normalizado(ruta)
        items = cargar_factura_xml(ruta)
    except Exception:
        texto = ruta.read_text(encoding="utf-8", errors="ignore")
        texto = texto.replace("<![CDATA[", "").replace("]]>", "")
        import html as _html

        if "&lt;" in texto:
            texto = _html.unescape(texto)
        items = {}

    info: dict = {"archivo": ruta.name, "items": items}
    m = re.search(r"<(?:\w+:)?Invoice[\s>].*?<(?:\w+:)?ID[^>]*>([^<]+)</", texto, re.DOTALL)
    info["num_factura"] = m.group(1).strip() if m else ""
    m = re.search(r"<(?:\w+:)?IssueDate>([^<]+)<", texto)
    info["fecha_emision"] = m.group(1).strip() if m else ""
    m = re.search(
        r"AccountingSupplierParty.*?<(?:\w+:)?RegistrationName>([^<]+)<", texto, re.DOTALL
    )
    info["emisor"] = m.group(1).strip() if m else ""
    m = re.search(r"AccountingSupplierParty.*?<(?:\w+:)?CompanyID[^>]*>([^<]+)<", texto, re.DOTALL)
    info["nit_emisor"] = m.group(1).strip() if m else ""
    m = re.search(
        r"AccountingCustomerParty.*?<(?:\w+:)?RegistrationName>([^<]+)<", texto, re.DOTALL
    )
    info["adquiriente"] = m.group(1).strip() if m else ""
    m = re.search(r"LegalMonetaryTotal.*?<(?:\w+:)?PayableAmount[^>]*>([^<]+)<", texto, re.DOTALL)
    info["total_pagar"] = m.group(1).strip() if m else ""
    m = re.search(
        r"LegalMonetaryTotal.*?<(?:\w+:)?LineExtensionAmount[^>]*>([^<]+)<", texto, re.DOTALL
    )
    info["total_lineas"] = m.group(1).strip() if m else ""
    return info


# ─────────────────────────────────────────────────────────────────────────────
# Cruces FURIPS ↔ soportes
# ─────────────────────────────────────────────────────────────────────────────
def _monto_xml(valor: str) -> int | None:
    try:
        return int(round(float(valor)))
    except (ValueError, TypeError):
        return None


def _agregar_cruce(
    res: ResultadoFactura,
    dato: str,
    valor_furips: str,
    valores: dict[str, str],
    resultado: str,
    obs: str = "",
) -> None:
    if resultado == "COINCIDE" and not any(v for v in valores.values()):
        # Ninguna fuente aportó valor: no se comparó nada realmente.
        resultado = "SIN DATO PARA CRUZAR"
    res.cruces.append(
        {
            "dato": dato,
            "furips": valor_furips,
            "rips": valores.get("rips", ""),
            "cuv": valores.get("cuv", ""),
            "xml": valores.get("xml", ""),
            "factura_pdf": valores.get("factura_pdf", ""),
            "epicrisis": valores.get("epicrisis", ""),
            "resultado": resultado,
            "obs": obs,
        }
    )


def _nombres_en_texto(nombres: list[str], texto_norm: str) -> bool:
    tokens = [normalizar(n) for n in nombres if n and len(normalizar(n)) >= 2]
    return bool(tokens) and all(t in texto_norm for t in tokens)


def cruzar_soportes(res: ResultadoFactura) -> None:
    reg = res.registro_f1
    datos = res.datos_soportes
    if reg is None or not datos:
        # Sin registro FURIPS o sin ningún soporte extraíble no hay cruce:
        # la ausencia ya queda reportada en SOPORTES/RESUMEN.
        return
    c = {n: reg[n - 1] for n in range(1, 103)}

    rips = datos.get("rips") or {}
    cuv = datos.get("cuv") or {}
    fev = datos.get("fev") or {}
    txt_fac = normalizar((datos.get("factura_pdf") or {}).get("texto", ""))
    txt_epi = normalizar((datos.get("epicrisis") or {}).get("texto", ""))
    fac_legible = len(txt_fac) > 50
    epi_legible = len(txt_epi) > 50

    usuario = (rips.get("usuarios") or [{}])[0] if rips else {}
    procs = ((usuario.get("servicios") or {}).get("procedimientos") or []) if usuario else []
    proc0 = procs[0] if procs else {}

    # ── Número de factura ────────────────────────────────────────────────────
    valores = {
        "rips": rips.get("numFactura", ""),
        "cuv": cuv.get("NumFactura", ""),
        "xml": fev.get("num_factura", ""),
        "factura_pdf": "presente"
        if fac_legible and norm_factura(c[3]) in txt_fac.replace(" ", "")
        else ("" if not fac_legible else "NO ENCONTRADO"),
    }
    difs = [
        f
        for f in ("rips", "cuv", "xml")
        if valores[f] and norm_factura(valores[f]) != norm_factura(c[3])
    ]
    if difs or valores["factura_pdf"] == "NO ENCONTRADO":
        fuentes = ", ".join(
            difs + (["factura PDF"] if valores["factura_pdf"] == "NO ENCONTRADO" else [])
        )
        _agregar_cruce(
            res, "Número de factura", c[3], valores, "DIFERENCIA", f"No coincide en: {fuentes}"
        )
        res.agregar(
            "CRUCE FEV XML" if "xml" in difs else "CRUCE RIPS",
            ERROR,
            "3",
            "Número de factura",
            c[3],
            "Cruce FURIPS vs soportes",
            f"El número de factura difiere en: {fuentes}.",
            valor_soporte="; ".join(f"{k}={v}" for k, v in valores.items() if v),
            fuente=fuentes,
        )
    else:
        _agregar_cruce(res, "Número de factura", c[3], valores, "COINCIDE")

    # ── Documento de la víctima ──────────────────────────────────────────────
    doc = c[11]
    valores = {
        "rips": f"{usuario.get('tipoDocumentoIdentificacion', '')} {usuario.get('numDocumentoIdentificacion', '')}".strip(),
        "factura_pdf": ""
        if not fac_legible
        else ("presente" if doc and doc in txt_fac else "NO ENCONTRADO"),
        "epicrisis": ""
        if not epi_legible
        else ("presente" if doc and doc in txt_epi else "NO ENCONTRADO"),
    }
    problemas = []
    if usuario:
        if usuario.get("numDocumentoIdentificacion", "") != doc:
            problemas.append(f"RIPS trae documento {usuario.get('numDocumentoIdentificacion')}")
        if (
            usuario.get("tipoDocumentoIdentificacion", "")
            and usuario.get("tipoDocumentoIdentificacion") != c[10]
        ):
            problemas.append(
                f"RIPS trae tipo {usuario.get('tipoDocumentoIdentificacion')} y FURIPS {c[10]}"
            )
    if valores["factura_pdf"] == "NO ENCONTRADO":
        problemas.append("documento no aparece en la factura PDF")
    if valores["epicrisis"] == "NO ENCONTRADO":
        problemas.append("documento no aparece en la epicrisis")
    if problemas:
        _agregar_cruce(
            res,
            "Documento de la víctima",
            f"{c[10]} {doc}",
            valores,
            "DIFERENCIA",
            "; ".join(problemas),
        )
        res.agregar(
            "CRUCE RIPS" if usuario else "CRUCE EPICRISIS",
            ERROR,
            "10/11",
            "Documento de la víctima",
            f"{c[10]} {doc}",
            "Cruce FURIPS vs soportes",
            "; ".join(problemas) + ".",
            valor_soporte=valores["rips"],
            fuente="RIPS/PDFs",
        )
    else:
        _agregar_cruce(res, "Documento de la víctima", f"{c[10]} {doc}", valores, "COINCIDE")

    # ── Nombre de la víctima ─────────────────────────────────────────────────
    nombres = [c[8], c[9], c[6], c[7]]  # nombres + apellidos
    nombre_completo = " ".join(n for n in nombres if n)
    valores = {
        "factura_pdf": ""
        if not fac_legible
        else ("presente" if _nombres_en_texto(nombres, txt_fac) else "NO ENCONTRADO"),
        "epicrisis": ""
        if not epi_legible
        else ("presente" if _nombres_en_texto(nombres, txt_epi) else "NO ENCONTRADO"),
    }
    problemas = [
        f"nombre no aparece completo en {k.replace('_', ' ')}"
        for k in ("factura_pdf", "epicrisis")
        if valores[k] == "NO ENCONTRADO"
    ]
    if problemas:
        _agregar_cruce(
            res,
            "Nombre de la víctima",
            nombre_completo,
            valores,
            "DIFERENCIA",
            "; ".join(problemas),
        )
        res.agregar(
            "CRUCE EPICRISIS",
            ADVERTENCIA,
            "6-9",
            "Nombre de la víctima",
            nombre_completo,
            "Cruce FURIPS vs soportes",
            "; ".join(problemas) + ". Verificar apellidos/nombres contra el soporte.",
            fuente="PDFs",
        )
    else:
        _agregar_cruce(res, "Nombre de la víctima", nombre_completo, valores, "COINCIDE")

    # ── Fecha de nacimiento y sexo ───────────────────────────────────────────
    if usuario:
        f_nac_f = parse_fecha_ddmmaaaa(c[12])
        f_nac_r = parse_fecha_flexible(usuario.get("fechaNacimiento", ""))
        if f_nac_f and f_nac_r and f_nac_f != f_nac_r:
            _agregar_cruce(
                res,
                "Fecha de nacimiento",
                c[12],
                {"rips": usuario.get("fechaNacimiento", "")},
                "DIFERENCIA",
            )
            res.agregar(
                "CRUCE RIPS",
                ERROR,
                "12",
                "Fecha de nacimiento",
                c[12],
                "Cruce FURIPS vs RIPS",
                f"FURIPS dice {c[12]} y el RIPS {usuario.get('fechaNacimiento')}.",
                valor_soporte=usuario.get("fechaNacimiento", ""),
                fuente="RIPS",
            )
        elif f_nac_f and f_nac_r:
            _agregar_cruce(
                res,
                "Fecha de nacimiento",
                c[12],
                {"rips": usuario.get("fechaNacimiento", "")},
                "COINCIDE",
            )
        sexo_r = normalizar(usuario.get("codSexo", ""))
        if sexo_r and c[14] and sexo_r != c[14]:
            _agregar_cruce(res, "Sexo", c[14], {"rips": sexo_r}, "DIFERENCIA")
            res.agregar(
                "CRUCE RIPS",
                ERROR,
                "14",
                "Sexo de la víctima",
                c[14],
                "Cruce FURIPS vs RIPS",
                f"FURIPS dice {c[14]} y el RIPS {sexo_r}.",
                valor_soporte=sexo_r,
                fuente="RIPS",
            )
        elif sexo_r:
            _agregar_cruce(res, "Sexo", c[14], {"rips": sexo_r}, "COINCIDE")

    # ── Código de habilitación / NIT ─────────────────────────────────────────
    cod_prest_rips = str(proc0.get("codPrestador", "")) if proc0 else ""
    if cod_prest_rips:
        r = "COINCIDE" if cod_prest_rips == c[5] else "DIFERENCIA"
        _agregar_cruce(res, "Código habilitación prestador", c[5], {"rips": cod_prest_rips}, r)
        if r == "DIFERENCIA":
            res.agregar(
                "CRUCE RIPS",
                ERROR,
                "5",
                "Código de habilitación",
                c[5],
                "Cruce FURIPS vs RIPS",
                f"FURIPS dice {c[5]} y el RIPS {cod_prest_rips}.",
                valor_soporte=cod_prest_rips,
                fuente="RIPS",
            )
    nit_rips = str(rips.get("numDocumentoIdObligado", "")) if rips else ""
    nit_xml = re.sub(r"\D", "", fev.get("nit_emisor", ""))[:9]
    if nit_rips and nit_xml:
        r = "COINCIDE" if nit_rips == nit_xml else "DIFERENCIA"
        _agregar_cruce(res, "NIT del prestador", nit_xml, {"rips": nit_rips, "xml": nit_xml}, r)
        if r == "DIFERENCIA":
            res.agregar(
                "CRUCE FEV XML",
                ADVERTENCIA,
                "-",
                "NIT del prestador",
                nit_xml,
                "Cruce RIPS vs FEV",
                f"RIPS declara NIT {nit_rips} y el XML {nit_xml}.",
                valor_soporte=nit_rips,
                fuente="RIPS/XML",
            )

    # ── Fechas de ingreso/egreso ─────────────────────────────────────────────
    f_ing = parse_fecha_ddmmaaaa(c[80])
    fechas_rips = parse_fecha_flexible(proc0.get("fechaInicioAtencion", "")) if proc0 else None
    fechas_fac = (
        extraer_fechas((datos.get("factura_pdf") or {}).get("texto", "")) if fac_legible else set()
    )
    fechas_epi = (
        extraer_fechas((datos.get("epicrisis") or {}).get("texto", "")) if epi_legible else set()
    )
    valores = {
        "rips": proc0.get("fechaInicioAtencion", "") if proc0 else "",
        "factura_pdf": ""
        if not fac_legible
        else ("presente" if f_ing in fechas_fac else "NO ENCONTRADA"),
        "epicrisis": ""
        if not epi_legible
        else ("presente" if f_ing in fechas_epi else "NO ENCONTRADA"),
    }
    problemas = []
    if f_ing and fechas_rips and f_ing != fechas_rips:
        problemas.append(f"RIPS inicia atención {fechas_rips.strftime('%d/%m/%Y')}")
    if valores["factura_pdf"] == "NO ENCONTRADA":
        problemas.append("la fecha de ingreso no aparece en la factura PDF")
    if valores["epicrisis"] == "NO ENCONTRADA":
        problemas.append("la fecha de ingreso no aparece en la epicrisis")
    if problemas:
        _agregar_cruce(
            res, "Fecha de ingreso", f"{c[80]} {c[81]}", valores, "DIFERENCIA", "; ".join(problemas)
        )
        res.agregar(
            "CRUCE EPICRISIS",
            ADVERTENCIA if not (f_ing and fechas_rips and f_ing != fechas_rips) else ERROR,
            "80",
            "Fecha de ingreso",
            f"{c[80]} {c[81]}",
            "Cruce FURIPS vs soportes (XI: los datos deben coincidir con el soporte)",
            "; ".join(problemas) + ".",
            valor_soporte=valores["rips"],
            fuente="RIPS/PDFs",
        )
    elif f_ing:
        _agregar_cruce(res, "Fecha de ingreso", f"{c[80]} {c[81]}", valores, "COINCIDE")

    f_evento = parse_fecha_ddmmaaaa(c[23])
    if f_evento and epi_legible:
        r = "COINCIDE" if f_evento in fechas_epi else "NO ENCONTRADA"
        _agregar_cruce(
            res,
            "Fecha del evento/accidente",
            c[23],
            {"epicrisis": "presente" if r == "COINCIDE" else "NO ENCONTRADA"},
            r,
        )
        if r != "COINCIDE":
            res.agregar(
                "CRUCE EPICRISIS",
                ADVERTENCIA,
                "23",
                "Fecha del evento",
                c[23],
                "Cruce FURIPS vs epicrisis",
                "La fecha del accidente no se encontró en el texto de la epicrisis.",
                fuente="EPICRISIS",
            )

    # ── Diagnósticos ─────────────────────────────────────────────────────────
    dx_rips = normalizar(proc0.get("codDiagnosticoPrincipal", "")) if proc0 else ""
    for campo, nombre in (
        (84, "Diagnóstico principal de ingreso"),
        (87, "Diagnóstico principal de egreso"),
    ):
        dx = normalizar(c[campo])
        if not dx:
            continue
        valores = {
            "rips": dx_rips,
            "epicrisis": ""
            if not epi_legible
            else (
                "presente"
                if (dx in txt_epi or f"{dx[:3]}.{dx[3:]}" in txt_epi)
                else "NO ENCONTRADO"
            ),
        }
        problemas = []
        if dx_rips and campo == 84 and dx_rips != dx:
            problemas.append(f"RIPS trae {dx_rips}")
        if valores["epicrisis"] == "NO ENCONTRADO":
            problemas.append("el código no aparece en la epicrisis")
        if problemas:
            _agregar_cruce(res, nombre, dx, valores, "DIFERENCIA", "; ".join(problemas))
            res.agregar(
                "CRUCE RIPS" if (dx_rips and campo == 84 and dx_rips != dx) else "CRUCE EPICRISIS",
                ADVERTENCIA,
                str(campo),
                nombre,
                dx,
                "Cruce FURIPS vs soportes",
                "; ".join(problemas) + ".",
                valor_soporte=dx_rips,
                fuente="RIPS/EPICRISIS",
            )
        else:
            _agregar_cruce(res, nombre, dx, valores, "COINCIDE")

    # ── Valores ──────────────────────────────────────────────────────────────
    total_fact = (a_entero(c[97]) or 0) + (a_entero(c[99]) or 0)
    total_xml = _monto_xml(fev.get("total_pagar", "")) if fev else None
    suma_rips = (
        sum(
            (a_entero(str(p.get("vrServicio") or "")) or 0)
            for u in (rips.get("usuarios") or [])
            for tipo in (
                "consultas",
                "procedimientos",
                "medicamentos",
                "otrosServicios",
                "urgencias",
                "hospitalizacion",
            )
            for p in ((u.get("servicios") or {}).get(tipo) or [])
        )
        if rips
        else None
    )
    montos_fac = (
        extraer_montos((datos.get("factura_pdf") or {}).get("texto", "")) if fac_legible else set()
    )
    valores = {
        "rips": f"{suma_rips:,}" if suma_rips else "",
        "xml": f"{total_xml:,}" if total_xml is not None else "",
        "factura_pdf": ""
        if not fac_legible
        else ("presente" if total_fact in montos_fac else "NO ENCONTRADO"),
    }
    problemas = []
    if total_xml is not None and total_xml != total_fact:
        problemas.append(f"el XML DIAN factura {total_xml:,}")
    if suma_rips is not None and suma_rips != total_fact:
        problemas.append(f"el RIPS suma {suma_rips:,}")
    if valores["factura_pdf"] == "NO ENCONTRADO":
        problemas.append("el total no aparece en la factura PDF")
    if problemas:
        _agregar_cruce(
            res,
            "Total facturado (97+99)",
            f"{total_fact:,}",
            valores,
            "DIFERENCIA",
            "; ".join(problemas),
        )
        res.agregar(
            "CRUCE FEV XML",
            ERROR if (total_xml is not None and total_xml != total_fact) else ADVERTENCIA,
            "97/99",
            "Total facturado vs soportes",
            f"{total_fact:,}",
            "Cruce FURIPS vs factura/RIPS",
            f"FURIPS declara {total_fact:,} pero " + "; ".join(problemas) + ".",
            valor_soporte=valores["xml"] or valores["rips"],
            fuente="XML/RIPS/PDF",
        )
    else:
        _agregar_cruce(res, "Total facturado (97+99)", f"{total_fact:,}", valores, "COINCIDE")

    # ── Consecutivo (nro de ingreso) en factura PDF ──────────────────────────
    if fac_legible and c[4]:
        r = "COINCIDE" if c[4] in txt_fac else "NO ENCONTRADO"
        _agregar_cruce(
            res,
            "Consecutivo reclamación (ingreso)",
            c[4],
            {"factura_pdf": "presente" if r == "COINCIDE" else "NO ENCONTRADO"},
            r,
        )
        if r == "NO ENCONTRADO":
            res.agregar(
                "CRUCE FACTURA PDF",
                ADVERTENCIA,
                "4",
                "Consecutivo de la reclamación",
                c[4],
                "Cruce FURIPS vs factura PDF",
                "El consecutivo (número de ingreso) no aparece en la factura PDF.",
                fuente="FACTURA PDF",
            )

    # ── FURIPS2 vs líneas del XML ────────────────────────────────────────────
    # Solo se cruza por código si la factura electrónica usa la MISMA codificación
    # que el FURIPS 2 (muchas IPS facturan con códigos UNSPSC/internos distintos
    # del SOAT: en ese caso el cruce por código sería ruido y se cruza por total).
    items_xml: dict = fev.get("items") or {}
    lineas_con_codigo = [
        (i, ln) for i, ln in enumerate(res.lineas_f2, 1) if ln[2] in {"1", "2", "6"} and ln[3]
    ]
    if items_xml and lineas_con_codigo:
        codigos_xml = {normalizar(k) for k in items_xml}
        coincidencias = sum(1 for _i, ln in lineas_con_codigo if normalizar(ln[3]) in codigos_xml)
        if coincidencias >= max(1, len(lineas_con_codigo) // 3):
            faltantes = [
                (i, ln) for i, ln in lineas_con_codigo if normalizar(ln[3]) not in codigos_xml
            ]
            for i, ln in faltantes[:20]:
                res.agregar(
                    "CRUCE FEV XML",
                    ADVERTENCIA,
                    f"F2 L{i}",
                    "Código de servicio vs factura XML",
                    ln[3],
                    "Cruce FURIPS2 vs FEV",
                    f"El código '{ln[3]}' de la línea {i} del FURIPS 2 no aparece "
                    f"como ítem de la factura electrónica.",
                    valor_soporte=", ".join(sorted(codigos_xml)[:8]),
                    fuente="FEV XML",
                )
        else:
            res.agregar(
                "CRUCE FEV XML",
                INFO,
                "-",
                "Codificación de la factura XML",
                f"{len(items_xml)} ítems XML",
                "Cruce FURIPS2 vs FEV",
                "La factura electrónica usa una codificación distinta a la del "
                "FURIPS 2 (ej. UNSPSC/interna): el cruce por código no aplica; "
                "se validó el TOTAL facturado contra el XML.",
            )

    # ── CUV ──────────────────────────────────────────────────────────────────
    if cuv:
        codigo_cuv = cuv.get("CodigoUnicoValidacion", "") or ""
        if not codigo_cuv:
            res.agregar(
                "CRUCE CUV",
                ERROR,
                "-",
                "Código único de validación",
                "",
                "Res. 2284/2023",
                "El CUV no contiene CodigoUnicoValidacion.",
            )
        if cuv.get("ResultState") is False:
            res.agregar(
                "CRUCE CUV",
                ERROR,
                "-",
                "Estado de validación RIPS",
                "ResultState=false",
                "MSPS",
                "El Ministerio NO validó exitosamente el RIPS (ResultState=false).",
            )
        rechazos = [
            r
            for r in (cuv.get("ResultadosValidacion") or [])
            if normalizar(str(r.get("Clase", ""))) == "RECHAZO"
        ]
        notifs = [
            r
            for r in (cuv.get("ResultadosValidacion") or [])
            if normalizar(str(r.get("Clase", ""))) == "NOTIFICACION"
        ]
        if rechazos:
            res.agregar(
                "CRUCE CUV",
                ERROR,
                "-",
                "Rechazos en validación RIPS",
                f"{len(rechazos)} rechazo(s)",
                "MSPS",
                "; ".join(
                    f"{r.get('Codigo')}: {r.get('Descripcion', '')[:80]}" for r in rechazos[:5]
                ),
            )
        if notifs:
            res.agregar(
                "CRUCE CUV",
                INFO,
                "-",
                "Notificaciones validación RIPS",
                f"{len(notifs)} notificación(es)",
                "MSPS",
                "; ".join(f"{r.get('Codigo')}" for r in notifs[:10])
                + ". Son avisos del validador RIPS, no impiden radicar.",
            )
        _agregar_cruce(
            res,
            "CUV (validación RIPS MSPS)",
            "OK" if cuv.get("ResultState") else "FALLIDO",
            {"cuv": (codigo_cuv[:24] + "…") if codigo_cuv else "SIN CÓDIGO"},
            "COINCIDE" if (cuv.get("ResultState") and codigo_cuv) else "DIFERENCIA",
            f"{len(rechazos)} rechazos, {len(notifs)} notificaciones",
        )


# ─────────────────────────────────────────────────────────────────────────────
# Informe Excel
# ─────────────────────────────────────────────────────────────────────────────
def generar_excel(
    resultados: list[ResultadoFactura], obs_archivos: list[str], salida: Path
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    AZUL = "1F4E79"
    FILL_HEADER = PatternFill("solid", fgColor=AZUL)
    FONT_HEADER = Font(color="FFFFFF", bold=True, size=10)
    FILLS = {
        ERROR: PatternFill("solid", fgColor="FFC7CE"),
        ADVERTENCIA: PatternFill("solid", fgColor="FFEB9C"),
        OK: PatternFill("solid", fgColor="C6EFCE"),
        INFO: PatternFill("solid", fgColor="DDEBF7"),
    }
    FONTS = {
        ERROR: Font(color="9C0006", bold=True, size=10),
        ADVERTENCIA: Font(color="9C6500", size=10),
        OK: Font(color="006100", size=10),
        INFO: Font(color="1F4E79", size=10),
    }
    BORDE = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

    wb = Workbook()

    def hoja(nombre, encabezados, anchos):
        ws = wb.create_sheet(nombre)
        for j, (enc, ancho) in enumerate(zip(encabezados, anchos, strict=False), 1):
            cel = ws.cell(row=1, column=j, value=enc)
            cel.fill, cel.font = FILL_HEADER, FONT_HEADER
            cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(j)].width = ancho
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30
        return ws

    def fila(ws, valores, col_estado=None, estado=None, cols_moneda=()):
        r = ws.max_row + 1
        for j, v in enumerate(valores, 1):
            if isinstance(v, str):
                # openpyxl rechaza caracteres de control (TXT truncados/binarios)
                v = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", v)
            cel = ws.cell(row=r, column=j, value=v)
            cel.border = BORDE
            cel.alignment = Alignment(vertical="top", wrap_text=True)
            cel.font = Font(size=10)
            if j in cols_moneda and isinstance(v, (int, float)):
                cel.number_format = "#,##0"
                cel.alignment = Alignment(vertical="top", horizontal="right")
        if col_estado and estado in FILLS:
            cel = ws.cell(row=r, column=col_estado)
            cel.fill, cel.font = FILLS[estado], FONTS[estado]

    # ── RESUMEN ──────────────────────────────────────────────────────────────
    ws = hoja(
        "RESUMEN",
        [
            "FACTURA",
            "CONSECUTIVO",
            "PACIENTE",
            "DOCUMENTO",
            "FECHA EVENTO",
            "INGRESO",
            "EGRESO",
            "VR FACTURADO",
            "VR RECLAMADO",
            "LÍNEAS F2",
            "FURIPS1",
            "CARPETA",
            "RIPS",
            "CUV",
            "FEV XML",
            "FACT. PDF",
            "EPICRISIS",
            "ERRORES",
            "ADVERT.",
            "ESTADO",
            "PRINCIPALES HALLAZGOS",
        ],
        [14, 11, 26, 13, 11, 11, 11, 13, 13, 8, 8, 9, 6, 6, 8, 9, 10, 8, 8, 13, 60],
    )
    wb.active = ws
    del wb["Sheet"]

    for res in resultados:
        reg = res.registro_f1
        c = {n: reg[n - 1] for n in range(1, 103)} if reg else {}
        tipos = {s["tipo"]: s for s in res.soportes}

        def sop(tipo, res=res, tipos=tipos):
            if not res.archivos:
                return "—"
            s = tipos.get(tipo)
            if not s:
                return "FALTA"
            if s.get("legible") == "NO (escaneado)":
                return "SIN TEXTO"
            return "SI"

        estado = "CON ERRORES" if res.errores else "REVISAR" if res.advertencias else "CUMPLE"
        principales = "; ".join(
            f"[{h.severidad[:3]}] {h.concepto}: {h.detalle[:80]}"
            for h in sorted(
                (h for h in res.hallazgos if h.severidad in (ERROR, ADVERTENCIA)),
                key=lambda h: (h.severidad != ERROR, h.origen),
            )[:3]
        )
        paciente = " ".join(x for x in (c.get(8), c.get(9), c.get(6), c.get(7)) if x)
        fila(
            ws,
            [
                res.factura,
                c.get(4, ""),
                paciente,
                f"{c.get(10, '')} {c.get(11, '')}".strip(),
                c.get(23, ""),
                f"{c.get(80, '')} {c.get(81, '')}".strip(),
                f"{c.get(82, '')} {c.get(83, '')}".strip(),
                (a_entero(c.get(97, "")) or 0) + (a_entero(c.get(99, "")) or 0) if reg else "",
                (a_entero(c.get(98, "")) or 0) + (a_entero(c.get(100, "")) or 0) if reg else "",
                len(res.lineas_f2),
                "SI" if reg else "FALTA",
                "SI" if res.carpeta else "FALTA",
                sop("RIPS JSON"),
                sop("CUV JSON"),
                sop("FACTURA XML"),
                sop("FACTURA PDF"),
                sop("EPICRISIS"),
                res.errores,
                res.advertencias,
                estado,
                principales,
            ],
            col_estado=20,
            estado=(
                ERROR if estado == "CON ERRORES" else ADVERTENCIA if estado == "REVISAR" else OK
            ),
            cols_moneda=(8, 9),
        )
    ws.auto_filter.ref = f"A1:U{ws.max_row}"

    # ── HALLAZGOS ────────────────────────────────────────────────────────────
    ws = hoja(
        "HALLAZGOS",
        [
            "FACTURA",
            "ORIGEN",
            "SEVERIDAD",
            "CAMPO Nº",
            "CONCEPTO",
            "VALOR FURIPS",
            "VALOR SOPORTE",
            "FUENTE",
            "REGLA (referencia)",
            "DETALLE",
        ],
        [14, 16, 13, 9, 30, 24, 24, 14, 30, 65],
    )
    orden_sev = {ERROR: 0, ADVERTENCIA: 1, INFO: 2}
    todos = sorted(
        (h for res in resultados for h in res.hallazgos),
        key=lambda h: (h.factura, orden_sev.get(h.severidad, 3), h.origen),
    )
    for h in todos:
        fila(
            ws,
            [
                h.factura,
                h.origen,
                h.severidad,
                h.campo,
                h.concepto,
                h.valor_furips,
                h.valor_soporte,
                h.fuente,
                h.regla,
                h.detalle,
            ],
            col_estado=3,
            estado=h.severidad,
        )
    for obs in obs_archivos:
        fila(
            ws,
            [
                "(ARCHIVOS TXT)",
                "ESTRUCTURA",
                ADVERTENCIA,
                "-",
                "Archivo FURIPS",
                "",
                "",
                "",
                "Circular 022/2023 nombre/estructura",
                obs,
            ],
            col_estado=3,
            estado=ADVERTENCIA,
        )
    ws.auto_filter.ref = f"A1:J{ws.max_row}"

    # ── FURIPS1 CAMPOS ───────────────────────────────────────────────────────
    ws = hoja(
        "FURIPS1 CAMPOS",
        [
            "FACTURA",
            "Nº",
            "SECCIÓN",
            "CAMPO",
            "VALOR",
            "OBLIGATORIEDAD",
            "ESTADO",
            "OBSERVACIÓN",
        ],
        [14, 6, 30, 38, 30, 34, 13, 55],
    )
    for res in resultados:
        if not res.registro_f1:
            continue
        for n, seccion, nombre, _l, _f, _v, oblig in E1:
            estado, obs = res.estado_campos_f1.get(n, (OK, ""))
            fila(
                ws,
                [
                    res.factura,
                    n,
                    seccion,
                    nombre,
                    res.registro_f1[n - 1],
                    TEXTO_OBLIG.get(oblig, oblig),
                    estado,
                    obs,
                ],
                col_estado=7,
                estado=estado,
            )
    ws.auto_filter.ref = f"A1:H{ws.max_row}"

    # ── FURIPS2 LINEAS ───────────────────────────────────────────────────────
    ws = hoja(
        "FURIPS2 LINEAS",
        [
            "FACTURA",
            "CONSECUTIVO",
            "LÍNEA",
            "TIPO SERVICIO",
            "CÓDIGO",
            "DESCRIPCIÓN",
            "CANTIDAD",
            "VR UNITARIO",
            "VR FACTURADO",
            "VR RECLAMADO",
            "ESTADO",
            "OBSERVACIÓN",
        ],
        [14, 12, 7, 22, 12, 40, 10, 12, 13, 13, 13, 50],
    )
    for res in resultados:
        for i, ln in enumerate(res.lineas_f2, 1):
            estado, obs = (
                res.estado_lineas_f2[i - 1] if i <= len(res.estado_lineas_f2) else (OK, "")
            )
            fila(
                ws,
                [
                    res.factura,
                    ln[1],
                    i,
                    f"{ln[2]} - {TIPOS_SERVICIO_F2.get(ln[2], '?')}",
                    ln[3],
                    ln[4],
                    a_entero(ln[5]),
                    a_entero(ln[6]),
                    a_entero(ln[7]),
                    a_entero(ln[8]),
                    estado,
                    obs,
                ],
                col_estado=11,
                estado=estado,
                cols_moneda=(8, 9, 10),
            )
    ws.auto_filter.ref = f"A1:L{ws.max_row}"

    # ── CRUCE SOPORTES ───────────────────────────────────────────────────────
    ws = hoja(
        "CRUCE SOPORTES",
        [
            "FACTURA",
            "DATO",
            "VALOR FURIPS",
            "RIPS",
            "CUV",
            "FACTURA XML",
            "FACTURA PDF",
            "EPICRISIS",
            "RESULTADO",
            "OBSERVACIÓN",
        ],
        [14, 28, 24, 18, 18, 18, 16, 16, 14, 45],
    )
    for res in resultados:
        for cz in res.cruces:
            estado = (
                OK
                if cz["resultado"] == "COINCIDE"
                else INFO
                if cz["resultado"] == "SIN DATO PARA CRUZAR"
                else ADVERTENCIA
            )
            fila(
                ws,
                [
                    res.factura,
                    cz["dato"],
                    cz["furips"],
                    cz["rips"],
                    cz["cuv"],
                    cz["xml"],
                    cz["factura_pdf"],
                    cz["epicrisis"],
                    cz["resultado"],
                    cz["obs"],
                ],
                col_estado=9,
                estado=estado,
            )
    ws.auto_filter.ref = f"A1:J{ws.max_row}"

    # ── SOPORTES ─────────────────────────────────────────────────────────────
    ws = hoja(
        "SOPORTES",
        [
            "FACTURA",
            "CARPETA",
            "ARCHIVO",
            "TIPO DE SOPORTE",
            "TAMAÑO (KB)",
            "TEXTO LEGIBLE",
            "OBSERVACIÓN",
        ],
        [14, 34, 44, 22, 12, 14, 40],
    )
    for res in resultados:
        for s in res.soportes:
            fila(
                ws,
                [
                    res.factura,
                    str(res.carpeta or ""),
                    s["archivo"],
                    s["tipo"],
                    round(s["bytes"] / 1024, 1),
                    s["legible"],
                    s["obs"],
                ],
            )
        if not res.archivos:
            fila(
                ws,
                [
                    res.factura,
                    "(SIN SOPORTES)",
                    "",
                    "",
                    "",
                    "",
                    "No se encontraron soportes para esta factura.",
                ],
            )
    ws.auto_filter.ref = f"A1:G{ws.max_row}"

    # ── LEYENDA ──────────────────────────────────────────────────────────────
    ws = hoja("LEYENDA", ["SECCIÓN", "EXPLICACIÓN"], [30, 120])
    leyenda = [
        (
            "Qué valida este informe",
            "1) Malla de la Circular 022 de 2023 ADRES: los 102 campos del FURIPS 1 y los 9 del "
            "FURIPS 2 (obligatoriedad, formatos, valores permitidos, longitudes, condicionales). "
            "2) Coherencia FURIPS1↔FURIPS2 (sumatorias campos 97-100, consecutivo, reclamado ≤ facturado). "
            "3) Cruce contra los soportes de cada carpeta: RIPS JSON, CUV, factura electrónica XML "
            "(DIAN), factura PDF y epicrisis PDF.",
        ),
        (
            "ERROR (rojo)",
            "Incumple la malla de la Circular o hay diferencia comprobada entre el FURIPS y un "
            "soporte. La ADRES normalmente lo devuelve o glosa. Corregir antes de radicar.",
        ),
        (
            "ADVERTENCIA (amarillo)",
            "Dato que no se pudo confirmar en los soportes o situación atípica que suele generar "
            "glosa (revisión manual recomendada).",
        ),
        (
            "INFO (azul)",
            "Contexto útil que no impide radicar (ej. notificaciones del validador RIPS).",
        ),
        (
            "CUMPLE / REVISAR / CON ERRORES",
            "Estado por factura en RESUMEN: sin hallazgos / solo advertencias / con al menos un error.",
        ),
        (
            "Cruce 'NO ENCONTRADO'",
            "El dato del FURIPS no aparece textualmente en el PDF. Si el PDF es escaneado "
            "(sin capa de texto) el cruce no es posible y se marca 'SIN TEXTO'.",
        ),
        (
            "Normativa",
            "Circular 022 de 2023 ADRES (formularios y anexos técnicos FURIPS 1 y 2); "
            "Resolución 2284 de 2023 (soportes); Decreto 780/2016 mod. 2466/2022 (manual tarifario "
            "SOAT); Resolución 762 de 2023 (identificación AS/MS).",
        ),
        (
            "Herramienta",
            "tools/adres/validar_furips.py — Motor Glosas HUS. Uso: py validar_furips.py --raiz "
            '"C:\\CARPETA\\FACTURAS" [--salida INFORME.xlsx]. Los TXT FURIPS1/FURIPS2 se detectan '
            "automáticamente en la raíz o subcarpetas.",
        ),
    ]
    for sec, exp in leyenda:
        fila(ws, [sec, exp])

    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(salida)


# ─────────────────────────────────────────────────────────────────────────────
# Orquestación
# ─────────────────────────────────────────────────────────────────────────────
def descubrir_soportes(base: Path) -> dict[str, list[Path]]:
    """Mapa factura_norm → archivos de soporte, agrupados por NOMBRE de archivo.

    Soporta las dos organizaciones reales del HUS:
    - Una carpeta por factura (HUS374152/...): el número de factura viene
      igual en el nombre de cada soporte (<codhab>_<factura>_<TOKEN>.ext).
    - Una carpeta PLANA (p.ej. SOPORTES/) con los archivos de muchas
      facturas mezclados: se agrupan por el número de factura del nombre.
    Si el archivo no trae el número, se intenta con el nombre de su carpeta.
    """
    soportes: dict[str, list[Path]] = {}
    try:
        archivos = [p for p in base.rglob("*") if p.is_file()]
    except OSError:
        return soportes
    for p in archivos:
        if _RE_ARCHIVO_F1.match(p.name) or _RE_ARCHIVO_F2.match(p.name):
            continue
        if clasificar_soporte(p.name) == "OTRO":
            continue
        fact = factura_desde_nombre(p.name) or factura_desde_nombre(p.parent.name)
        if not fact:
            continue
        soportes.setdefault(norm_factura(fact), []).append(p)
    return soportes


def procesar(
    raiz: Path | None,
    carpeta: Path | None,
    furips1: Path | None,
    furips2: Path | None,
    sin_pdf: bool,
) -> tuple[list[ResultadoFactura], list[str]]:
    base = carpeta or raiz
    assert base is not None

    if furips1:
        rutas_f1 = [furips1]
        rutas_f2 = [furips2] if furips2 else []
    else:
        rutas_f1, rutas_f2 = descubrir_archivos_furips(base)

    logger.info(f"Archivos FURIPS 1 encontrados: {len(rutas_f1)}")
    for r in rutas_f1:
        logger.info(f"   - {r}")
    logger.info(f"Archivos FURIPS 2 encontrados: {len(rutas_f2)}")
    for r in rutas_f2:
        logger.info(f"   - {r}")

    registros_f1, obs1 = leer_furips1(rutas_f1)
    lineas_f2, obs2 = leer_furips2(rutas_f2)
    obs_archivos = obs1 + obs2

    soportes = descubrir_soportes(base)
    logger.info(f"Facturas con soportes detectadas: {len(soportes)}")

    if carpeta:
        # Modo una-carpeta: solo validar las facturas con soportes en ella
        claves = set(soportes) or (set(registros_f1) | set(lineas_f2))
    else:
        claves = set(registros_f1) | set(soportes) | set(lineas_f2)

    resultados: list[ResultadoFactura] = []
    for clave in sorted(claves):
        regs = registros_f1.get(clave, [])
        res = ResultadoFactura(factura=clave)
        res.archivos = sorted(soportes.get(clave, []))
        res.carpeta = res.archivos[0].parent if res.archivos else None
        res.lineas_f2 = lineas_f2.get(clave, [])
        if len(regs) > 1:
            res.agregar(
                "ESTRUCTURA",
                ADVERTENCIA,
                "-",
                "Registros duplicados",
                f"{len(regs)} registros",
                "Circular 022/2023",
                f"La factura aparece {len(regs)} veces en el FURIPS 1; "
                f"se valida el primer registro.",
            )
        res.registro_f1 = regs[0] if regs else None

        validar_malla_f1(res)
        validar_furips2(res)
        analizar_carpeta(res, sin_pdf=sin_pdf)
        cruzar_soportes(res)
        resultados.append(res)
        logger.info(
            f"  {clave}: {res.errores} errores, {res.advertencias} advertencias"
            f"{' (sin soportes)' if not res.archivos else ''}"
            f"{' (sin registro FURIPS1)' if res.registro_f1 is None else ''}"
        )
    return resultados, obs_archivos


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        "--raiz", type=Path, default=None, help="Raíz con carpetas de factura (modo masivo)"
    )
    parser.add_argument("--carpeta", type=Path, default=None, help="Carpeta de UNA factura")
    parser.add_argument(
        "--furips1", type=Path, default=None, help="Ruta del TXT FURIPS 1 (si no, se detecta solo)"
    )
    parser.add_argument(
        "--furips2", type=Path, default=None, help="Ruta del TXT FURIPS 2 (si no, se detecta solo)"
    )
    parser.add_argument(
        "--salida", type=Path, default=None, help="Ruta del informe Excel (por defecto, en la raíz)"
    )
    parser.add_argument(
        "--sin-pdf", action="store_true", help="No leer el contenido de los PDF (más rápido)"
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO, format="%(message)s", handlers=[logging.StreamHandler(sys.stdout)]
    )

    if not args.raiz and not args.carpeta:
        parser.error("Indique --raiz (masivo) o --carpeta (una factura).")
    base = args.carpeta or args.raiz
    if not base.is_dir():
        logger.error(f"No existe la carpeta: {base}")
        return 1
    for etiqueta, ruta in (("--furips1", args.furips1), ("--furips2", args.furips2)):
        if ruta is not None and not ruta.is_file():
            logger.error(f"No existe el archivo {etiqueta}: {ruta}")
            return 1

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        logger.error("Falta openpyxl (para el Excel). Instalar con: py -m pip install openpyxl")
        return 1

    motor = _detectar_motor_pdf()
    logger.info("=" * 72)
    logger.info("BOT DE VALIDACIÓN FURIPS — Circular 022/2023 ADRES — Motor Glosas HUS")
    logger.info("=" * 72)
    if motor == "ninguno" and not args.sin_pdf:
        logger.warning("⚠ Sin librería PDF (pdfplumber/pypdf): los PDF no se podrán leer.")
        logger.warning("  Instalar con: py -m pip install pypdf")
    else:
        logger.info(f"Lector de PDF: {motor}")

    resultados, obs_archivos = procesar(
        args.raiz, args.carpeta, args.furips1, args.furips2, args.sin_pdf
    )

    if not resultados:
        logger.error("No se encontraron registros FURIPS ni carpetas de factura para validar.")
        return 1

    salida = args.salida or (base / f"INFORME_VALIDACION_FURIPS_{date.today():%Y%m%d}.xlsx")
    generar_excel(resultados, obs_archivos, salida)

    tot_err = sum(r.errores for r in resultados)
    tot_adv = sum(r.advertencias for r in resultados)
    con_err = sum(1 for r in resultados if r.errores)
    logger.info("")
    logger.info("=" * 72)
    logger.info(
        f"RESULTADO: {len(resultados)} facturas validadas — "
        f"{con_err} con errores | {tot_err} errores | {tot_adv} advertencias"
    )
    logger.info(f"INFORME:   {salida}")
    logger.info("=" * 72)
    return 0


if __name__ == "__main__":
    sys.exit(main())
