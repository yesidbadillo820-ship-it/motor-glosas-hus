"""fur_lectura.py — Lectura del export consolidado del HUS para el FUR (ADRES).

El HUS produce un CSV (una fila por factura) que junta los datos de víctima,
evento, vehículo, propietario, conductor, atención y diagnósticos del SOAT.
Ese CSV NO tiene la estructura de las 62 columnas oficiales del FUR; es un
layout interno del hospital con columnas extra (fecha de nacimiento, sexo,
marca del vehículo, horas, diagnósticos, médico, valores) intercaladas.

Este módulo mapea cada posición del CSV al nombre de campo oficial del FUR
(Diccionario de Campos FUR de la ADRES) y normaliza formatos:
  - fechas DD/MM/AAAA → AAAA-MM-DD
  - zona U/R → 02/01
  - municipio depto+muni (dos columnas) → DIVIPOLA de 5 dígitos

El mapeo posicional se verificó contra dos filas reales de distinta
naturaleza (accidente de tránsito y explosión), confirmando que los campos
condicionales (condición víctima, estado aseguramiento, bloque vehículo,
propietario, conductor) sólo se llenan en accidentes de tránsito.

IMPORTANTE: el layout del CSV puede cambiar si el HUS modifica su export.
Usar `--diagnostico` en generar_fur.py para verificar el mapeo contra el
archivo real antes de confiar en la salida.

Sólo librería estándar (csv).
"""

from __future__ import annotations

import csv
import re
from pathlib import Path

# ── Las 62 columnas oficiales del FUR, en orden (nombre técnico JSON) ────────
CAMPOS_FUR: list[str] = [
    "NIT_PRESTADOR",                                                  # 1
    "NUM_FACTURA",                                                    # 2
    "Tipo_documento_identidad_victima",                              # 3
    "Numero_documento_identidad_victima",                           # 4
    "Tipo_de_poblacion_especial",                                    # 5
    "Primer_nombre_victima",                                          # 6
    "Segundo_nombre_victima",                                         # 7
    "Primer_apellido_victima",                                        # 8
    "Segundo_apellido_victima",                                       # 9
    "Direccion_residencia_victima",                                  # 10
    "Codigo_municipio_residencia_victima",                          # 11
    "Telefono_victima",                                               # 12
    "Naturaleza_del_evento",                                          # 13
    "Descripcion_del_otro_evento",                                    # 14
    "Condicion_victima",                                              # 15
    "Fecha_de_ocurrencia_evento",                                    # 16
    "Zona_de_ocurrencia_evento",                                     # 17
    "Codigo_municipio_ocurrencia_evento",                           # 18
    "Direccion_de_ocurrencia_evento",                               # 19
    "Descripcion_corta_de_lo_ocurrido_en_el_evento",               # 20
    "Estado_de_aseguramiento",                                       # 21
    "Placa_vehiculo",                                                 # 22
    "Tipo_de_Vehiculo",                                              # 23
    "Codigo_de_la_aseguradora",                                      # 24
    "Numero_de_poliza_SOAT",                                         # 25
    "Fecha_de_inicio_de_vigencia_de_la_poliza",                    # 26
    "Fecha_final_de_vigencia_de_la_poliza",                         # 27
    "Numero_de_radicado_SIRAS",                                      # 28
    "Cobro_por_agotamiento_tope_Aseguradora",                       # 29
    "Tipo_de_documento_de_identidad_del_propietario",              # 30
    "Numero_de_documento_de_identidad_del_propietario",            # 31
    "Primer_nombre_del_propietario_o_razon_social",                # 32
    "Segundo_nombre_del_propietario",                               # 33
    "Primer_apellido_del_propietario",                              # 34
    "Segundo_apellido_del_propietario",                             # 35
    "Direccion_de_residencia_del_propietario",                     # 36
    "Telefono_de_residencia_del_propietario",                      # 37
    "Codigo_del_municipio_de_residencia_del_propietario",          # 38
    "Tipo_de_documento_de_identidad_del_conductor",                # 39
    "Numero_de_documento_de_identidad_del_conductor",              # 40
    "Primer_nombre_del_conductor",                                  # 41
    "Segundo_nombre_del_conductor",                                 # 42
    "Primer_apellido_del_conductor",                                # 43
    "Segundo_apellido_del_conductor",                               # 44
    "Codigo_del_municipio_de_residencia_del_conductor",            # 45
    "Direccion_de_residencia_del_conductor",                       # 46
    "Telefono_de_residencia_del_conductor",                        # 47
    "Uso_material_de_osteosintesis_en_la_atencion",                # 48
    "Es_atencion_inicial_paciente_remitido_o_control",             # 49
    "Placa_ambulancia_que_realiza_el_traslado_secundario",         # 50
    "Tipo_de_servicio_del_transporte_secundario",                  # 51
    "Codigo_de_habilitacion_del_prestador_que_remite",             # 52
    "Codigo_de_habilitación_del_prestador_que_recibe",             # 53
    "TIPO_de_documento_Profesional_que_recibe",                    # 54
    "Numero_de_documento_Profesional_que_recibe",                  # 55
    "Fecha_de_aceptacion",                                          # 56
    "Hora_aceptacion",                                             # 57
    "Tipo_de_servicio_del_transporte",                             # 58
    "Placa_ambulancia_que_realiza_el_traslado",                   # 59
    "Codigo_de_habilitacion_del_prestador_que_recibe_transporte_primario",  # 60
    "Transporte_de_la_victima_desde_el_sitio_del_evento_Direccion",         # 61
    "Transporte_de_la_victima_hasta_el_fin_del_recorrido_direccion_IPS",    # 62
]

# Etiquetas legibles (encabezado en español de la plantilla oficial), en orden.
ENCABEZADOS_FUR: list[str] = [
    "NIT del Prestador",
    "Número de Factura",
    "Tipo documento víctima",
    "Número documento víctima",
    "Tipo de población especial",
    "Primer nombre víctima",
    "Segundo nombre víctima",
    "Primer apellido víctima",
    "Segundo apellido víctima",
    "Dirección residencia víctima",
    "Código municipio residencia víctima",
    "Teléfono víctima",
    "Naturaleza del evento",
    "Descripción otro evento",
    "Condición víctima",
    "Fecha de ocurrencia del evento",
    "Zona de ocurrencia del evento",
    "Código municipio de ocurrencia",
    "Dirección de ocurrencia del evento",
    "Descripción corta del evento",
    "Estado de aseguramiento",
    "Placa del vehículo",
    "Tipo del vehículo",
    "Código de la Aseguradora",
    "Número de póliza SOAT",
    "Fecha de inicio de vigencia de la póliza",
    "Fecha final de vigencia de la póliza",
    "Número de radicado SIRAS",
    "Cobro por agotamiento de tope (aseguradora)",
    "Tipo doc. propietario",
    "Número doc. propietario",
    "Primer nombre / razón social propietario",
    "Segundo nombre propietario",
    "Primer apellido propietario",
    "Segundo apellido propietario",
    "Dirección residencia propietario",
    "Teléfono residencia propietario",
    "Código municipio residencia propietario",
    "Tipo doc. conductor",
    "Número doc. conductor",
    "Primer nombre conductor",
    "Segundo nombre conductor",
    "Primer apellido conductor",
    "Segundo apellido conductor",
    "Código municipio residencia conductor",
    "Dirección residencia conductor",
    "Teléfono conductor",
    "Uso material de osteosíntesis",
    "Tipo de atención facturada",
    "Placa ambulancia traslado secundario",
    "Tipo servicio transporte secundario",
    "Código habilitación prestador que remite",
    "Código habilitación prestador que recibe",
    "Tipo doc. profesional que recibe",
    "Número doc. profesional que recibe",
    "Fecha de aceptación",
    "Hora de aceptación",
    "Tipo servicio transporte primario",
    "Placa ambulancia transporte primario",
    "Cód. habilitación prestador que recibe (T. primario)",
    "Dirección sitio del evento (origen T. primario)",
    "Dirección IPS destino (T. primario)",
]

# Campos obligatorios (VERDE) y condicionantes (AMARILLO) según el diccionario.
# Los condicionantes se exigen sólo si se cumple su condición (ver
# generar_fur.validar_fila).
OBLIGATORIOS = {
    "NIT_PRESTADOR", "NUM_FACTURA", "Tipo_documento_identidad_victima",
    "Numero_documento_identidad_victima", "Primer_nombre_victima",
    "Primer_apellido_victima", "Naturaleza_del_evento",
    "Fecha_de_ocurrencia_evento", "Zona_de_ocurrencia_evento",
    "Codigo_municipio_ocurrencia_evento", "Direccion_de_ocurrencia_evento",
    "Descripcion_corta_de_lo_ocurrido_en_el_evento",
    "Es_atencion_inicial_paciente_remitido_o_control",
}


def _fecha(s: str) -> str:
    """DD/MM/AAAA → AAAA-MM-DD. Devuelve '' si no matchea."""
    m = re.match(r"\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*$", s or "")
    if not m:
        return (s or "").strip()
    d, mes, a = m.groups()
    return f"{a}-{int(mes):02d}-{int(d):02d}"


def _muni(depto: str, muni: str) -> str:
    """Combina depto(2)+muni(3) en DIVIPOLA de 5 dígitos."""
    depto, muni = (depto or "").strip(), (muni or "").strip()
    if not depto and not muni:
        return ""
    return f"{depto.zfill(2)}{muni.zfill(3)}"


def _zona(s: str) -> str:
    """U→02 (Urbana), R→01 (Rural)."""
    s = (s or "").strip().upper()
    return {"U": "02", "R": "01"}.get(s, "")


# Índices del CSV interno del HUS (verificados contra filas reales).
# c = lista de campos de la fila ya separada por coma.
def _mapear_fila(c: list[str], nit_prestador: str) -> dict:
    """Convierte una fila del CSV del HUS en un dict con los 62 campos FUR."""
    g = lambda i: c[i].strip() if i < len(c) and c[i] is not None else ""  # noqa: E731

    fila = dict.fromkeys(CAMPOS_FUR, "")
    fila["NIT_PRESTADOR"] = nit_prestador
    fila["NUM_FACTURA"] = g(2)
    # Víctima (orden CSV: apellido1, apellido2, nombre1, nombre2)
    fila["Primer_apellido_victima"] = g(5)
    fila["Segundo_apellido_victima"] = g(6)
    fila["Primer_nombre_victima"] = g(7)
    fila["Segundo_nombre_victima"] = g(8)
    fila["Tipo_documento_identidad_victima"] = g(9)
    fila["Numero_documento_identidad_victima"] = g(10)
    # [11] fecha nac, [13] sexo → no son campos del FUR
    fila["Tipo_de_poblacion_especial"] = g(12)
    fila["Direccion_residencia_victima"] = g(14)
    fila["Codigo_municipio_residencia_victima"] = _muni(g(15), g(16))
    fila["Telefono_victima"] = g(17)
    fila["Condicion_victima"] = g(18)
    fila["Naturaleza_del_evento"] = g(19)
    fila["Descripcion_del_otro_evento"] = g(20)
    fila["Direccion_de_ocurrencia_evento"] = g(21)
    fila["Fecha_de_ocurrencia_evento"] = _fecha(g(22))
    # [23] hora ocurrencia → no es campo del FUR
    fila["Codigo_municipio_ocurrencia_evento"] = _muni(g(24), g(25))
    fila["Zona_de_ocurrencia_evento"] = _zona(g(26))
    fila["Estado_de_aseguramiento"] = g(27)
    # [28] marca → no es campo del FUR
    fila["Placa_vehiculo"] = g(29)
    fila["Tipo_de_Vehiculo"] = g(30)
    fila["Codigo_de_la_aseguradora"] = g(31)
    fila["Numero_de_poliza_SOAT"] = g(32)
    fila["Fecha_de_inicio_de_vigencia_de_la_poliza"] = _fecha(g(33))
    fila["Fecha_final_de_vigencia_de_la_poliza"] = _fecha(g(34))
    fila["Numero_de_radicado_SIRAS"] = g(35)
    fila["Cobro_por_agotamiento_tope_Aseguradora"] = g(36)
    # [37..42] códigos de servicio/CUPS → pertenecen al FUR SERVICIOS, no al FUR
    # Propietario (orden CSV: apellido1, apellido2, nombre1, nombre2)
    fila["Tipo_de_documento_de_identidad_del_propietario"] = g(43)
    fila["Numero_de_documento_de_identidad_del_propietario"] = g(44)
    fila["Primer_apellido_del_propietario"] = g(45)
    fila["Segundo_apellido_del_propietario"] = g(46)
    fila["Primer_nombre_del_propietario_o_razon_social"] = g(47)
    fila["Segundo_nombre_del_propietario"] = g(48)
    fila["Direccion_de_residencia_del_propietario"] = g(49)
    fila["Telefono_de_residencia_del_propietario"] = g(50)
    fila["Codigo_del_municipio_de_residencia_del_propietario"] = _muni(g(51), g(52))
    # Conductor (orden CSV: apellido1, apellido2, nombre1, nombre2)
    fila["Tipo_de_documento_de_identidad_del_conductor"] = g(57)
    fila["Numero_de_documento_de_identidad_del_conductor"] = g(58)
    fila["Primer_apellido_del_conductor"] = g(53)
    fila["Segundo_apellido_del_conductor"] = g(54)
    fila["Primer_nombre_del_conductor"] = g(55)
    fila["Segundo_nombre_del_conductor"] = g(56)
    fila["Direccion_de_residencia_del_conductor"] = g(59)
    fila["Codigo_del_municipio_de_residencia_del_conductor"] = _muni(g(60), g(61))
    fila["Telefono_de_residencia_del_conductor"] = g(62)
    # Descripción corta del evento = última columna (texto largo)
    fila["Descripcion_corta_de_lo_ocurrido_en_el_evento"] = (c[-1].strip() if c else "")
    # Es_atencion_inicial: el HUS exporta atención inicial; por defecto 1.
    # (Si la factura es de transporte/remisión, ajustar a mano.)
    fila["Es_atencion_inicial_paciente_remitido_o_control"] = "1"
    return fila


def leer_fur_hus(ruta: Path, nit_prestador: str) -> list[dict]:
    """Lee el export del HUS (.csv/.txt/.tsv o .xlsx) y devuelve dicts FUR."""
    return [_mapear_fila(c, nit_prestador) for c in filas_crudas(ruta)]


def _rebase_factura(row: list[str]) -> list[str] | None:
    """Re-alinea una fila para que el número de factura quede en el índice 2.

    El layout verificado tiene 2 columnas vacías y la factura (HUS…) en la
    posición 2. Si el origen (p. ej. un Excel) no trae esas dos columnas
    iniciales, buscamos la celda 'HUS<dígitos>' y reconstruimos la fila para
    que caiga en el índice 2. Devuelve None si no hay factura HUS en la fila.
    """
    idx = None
    for i, v in enumerate(row):
        if re.match(r"^HUS\d+$", (v or "").strip(), re.IGNORECASE):
            idx = i
            break
    if idx is None:
        return None
    return ["", ""] + [(v or "").strip() for v in row[idx:]]


def _leer_xlsx(ruta: Path) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("Para leer .xlsx instalá openpyxl (pip install openpyxl)")
    filas: list[list[str]] = []
    wb = load_workbook(filename=str(ruta), read_only=True, data_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            celdas = ["" if c is None else str(c) for c in row]
            reb = _rebase_factura(celdas)
            if reb:
                filas.append(reb)
    return filas


def _leer_texto(ruta: Path) -> list[list[str]]:
    """Lee CSV/TSV/TXT probando coma, tab y punto y coma."""
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            texto = ruta.read_text(encoding=enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        return []
    lineas = texto.splitlines()
    mejor: list[list[str]] = []
    for sep in (",", "\t", ";"):
        filas: list[list[str]] = []
        for ln in lineas:
            reb = _rebase_factura(ln.split(sep))
            if reb:
                filas.append(reb)
        if len(filas) > len(mejor):
            mejor = filas
    return mejor


def filas_crudas(ruta: Path) -> list[list[str]]:
    """Devuelve las filas (re-alineadas, factura en índice 2) de cualquier formato."""
    if ruta.suffix.lower() in (".xlsx", ".xlsm"):
        return _leer_xlsx(ruta)
    return _leer_texto(ruta)
