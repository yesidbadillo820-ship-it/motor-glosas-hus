"""generar_fur.py — Excel FUR (Formato Único de Reclamación) ADRES.

Lee el export consolidado del HUS (CSV con una fila por factura: víctima,
evento, vehículo, propietario, conductor, atención) y produce el Excel con
las 62 columnas oficiales del FUR, en el orden y nombres del Diccionario de
Campos de la ADRES.

MODOS:
  * Por factura:  --fur export.csv --factura HUS428139 --salida HUS428139_FUR.xlsx
  * Masivo:       --fur export.csv --salida lote.xlsx        (todas las facturas)
  * Diagnóstico:  --fur export.csv --factura HUS428139 --diagnostico
                  (imprime posición→valor→campo para verificar el mapeo)

VALIDACIÓN: marca en rojo (y reporta) los campos VERDE (obligatorios) vacíos
y los AMARILLO (condicionantes) que apliquen según naturaleza/estado. No
inventa datos: si el export del HUS no trae un campo, queda vacío y se reporta.

DEPENDENCIA: openpyxl (sólo para escribir el .xlsx).
"""

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from fur_lectura import (  # noqa: E402
    CAMPOS_FUR,
    ENCABEZADOS_FUR,
    OBLIGATORIOS,
    filas_crudas,
    leer_fur_hus,
)
from rips_lectura import cargar_rips  # noqa: E402

logger = logging.getLogger("generar_fur")

NIT_HUS_DEFECTO = "900006037"  # NIT del HUS sin dígito de verificación


def setup_logging() -> None:
    logging.basicConfig(level=logging.INFO, format="%(message)s")


def hallar_fur_csv(carpeta: Path) -> Path | None:
    """Busca un CSV de export FUR en la carpeta (que tenga filas HUS...)."""
    for p in sorted(carpeta.iterdir()):
        if p.is_file() and p.suffix.lower() in (".csv", ".txt"):
            try:
                if filas_crudas(p):
                    return p
            except Exception:
                continue
    return None


def nit_desde_rips(carpeta: Path) -> str | None:
    for p in sorted(carpeta.glob("*RIP*.json")):
        try:
            data = cargar_rips(p)
            nit = (data.get("numDocumentoIdObligado") or "").strip()
            if nit:
                return nit
        except Exception:
            continue
    return None


def validar_fila(fila: dict) -> list[str]:
    """Devuelve la lista de campos faltantes (obligatorios o condicionantes)."""
    faltan: list[str] = []
    for campo in OBLIGATORIOS:
        if not (fila.get(campo) or "").strip():
            faltan.append(campo)

    naturaleza = (fila.get("Naturaleza_del_evento") or "").strip()
    estado = (fila.get("Estado_de_aseguramiento") or "").strip()
    es_transito = naturaleza == "01"

    if es_transito:
        # Condición y tipo de vehículo siempre exigibles en tránsito.
        for campo in ("Condicion_victima", "Tipo_de_Vehiculo"):
            if not (fila.get(campo) or "").strip():
                faltan.append(campo)
        # Placa: obligatoria si estado en (2,4,6,7); vacía si (3,8).
        if estado in ("2", "4", "6", "7") and not (fila.get("Placa_vehiculo") or "").strip():
            faltan.append("Placa_vehiculo")
        # SIRAS: obligatorio si la ocurrencia es posterior al 01/06/2023.
        fecha = (fila.get("Fecha_de_ocurrencia_evento") or "").strip()
        if fecha >= "2023-06-01" and not (fila.get("Numero_de_radicado_SIRAS") or "").strip():
            faltan.append("Numero_de_radicado_SIRAS")
        # Conductor: obligatorio si estado en (2,4,6,7,8).
        if estado in ("2", "4", "6", "7", "8"):
            for campo in ("Tipo_de_documento_de_identidad_del_conductor",
                          "Numero_de_documento_de_identidad_del_conductor",
                          "Primer_nombre_del_conductor", "Primer_apellido_del_conductor"):
                if not (fila.get(campo) or "").strip():
                    faltan.append(campo)
        # Propietario: obligatorio si estado en (2,4,6,8).
        if estado in ("2", "4", "6", "8"):
            for campo in ("Tipo_de_documento_de_identidad_del_propietario",
                          "Numero_de_documento_de_identidad_del_propietario",
                          "Primer_nombre_del_propietario_o_razon_social",
                          "Primer_apellido_del_propietario"):
                if not (fila.get(campo) or "").strip():
                    faltan.append(campo)
    return faltan


def diagnostico(ruta: Path, num_factura: str) -> None:
    """Imprime posición→valor→campo mapeado para una factura."""
    crudas = filas_crudas(ruta)
    objetivo = None
    for c in crudas:
        if len(c) > 2 and c[2].strip().upper() == num_factura.strip().upper():
            objetivo = c
            break
    if objetivo is None:
        logger.error(f"No encontré la factura {num_factura} en {ruta.name}")
        return
    from fur_lectura import _mapear_fila

    fila = _mapear_fila(objetivo, NIT_HUS_DEFECTO)
    # Mapa inverso campo→valor para mostrar junto a cada posición conocida.
    logger.info(f"\n=== Posiciones crudas del CSV para {num_factura} ===")
    for i, v in enumerate(objetivo):
        marca = ""
        if i in (28,):
            marca = "  (marca vehículo — NO va al FUR)"
        elif i in (11,):
            marca = "  (fecha nac. — NO va al FUR)"
        elif i in (13,):
            marca = "  (sexo — NO va al FUR)"
        elif i in (23,):
            marca = "  (hora ocurrencia — NO va al FUR)"
        elif 37 <= i <= 42:
            marca = "  (códigos servicio/CUPS — van al FUR SERVICIOS)"
        logger.info(f"  [{i:>3}] {(v or '')[:40]:<40}{marca}")
    logger.info(f"\n=== Campos FUR resultantes para {num_factura} ===")
    for campo in CAMPOS_FUR:
        val = fila.get(campo, "")
        if val:
            logger.info(f"  {campo:<55} = {val[:60]}")


def escribir_excel(filas: list[dict], salida: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        sys.stderr.write("ERROR: falta openpyxl. Instalá: py -m pip install openpyxl\n")
        sys.exit(2)

    wb = Workbook()
    ws = wb.active
    ws.title = "FUR"
    encabezado_fill = PatternFill("solid", fgColor="1F4E78")
    encabezado_font = Font(bold=True, color="FFFFFF")
    falta_fill = PatternFill("solid", fgColor="FFC7CE")  # rojo claro

    # Fila 1: encabezados en español.
    for col, etiqueta in enumerate(ENCABEZADOS_FUR, start=1):
        celda = ws.cell(row=1, column=col, value=etiqueta)
        celda.fill = encabezado_fill
        celda.font = encabezado_font

    # Filas de datos.
    for r, fila in enumerate(filas, start=2):
        faltan = set(validar_fila(fila))
        for col, campo in enumerate(CAMPOS_FUR, start=1):
            celda = ws.cell(row=r, column=col, value=fila.get(campo, ""))
            if campo in faltan:
                celda.fill = falta_fill

    wb.save(str(salida))


def main() -> None:
    parser = argparse.ArgumentParser(description="Genera el Excel FUR (ADRES) desde el export del HUS.")
    parser.add_argument("--fur", type=Path, help="CSV export del HUS (víctima/evento/vehículo/conductor).")
    parser.add_argument("--carpeta", type=Path, help="Carpeta de la factura (para buscar el CSV y el NIT desde el RIPS).")
    parser.add_argument("--factura", type=str, default=None, help="Número de factura (modo por factura). Si se omite: masivo.")
    parser.add_argument("--nit", type=str, default=None, help="NIT del prestador (si no, se toma del RIPS o el del HUS).")
    parser.add_argument("--salida", type=Path, help="Excel .xlsx de salida.")
    parser.add_argument("--diagnostico", action="store_true", help="Imprime el mapeo posición→campo para verificar.")
    parser.add_argument("--buscar", type=Path, default=None, help="Escanea una carpeta (recursivo) y reporta qué archivos contienen filas de facturas HUS.")
    args = parser.parse_args()
    setup_logging()

    if args.buscar:
        if not args.buscar.is_dir():
            logger.error(f"No es una carpeta: {args.buscar}")
            sys.exit(1)
        logger.info(f"Buscando exports con filas HUS en: {args.buscar}\n")
        encontrados = 0
        for p in sorted(args.buscar.rglob("*")):
            if not p.is_file() or p.suffix.lower() not in (".csv", ".txt", ".tsv", ".xlsx", ".xlsm"):
                continue
            try:
                filas = filas_crudas(p)
            except Exception as e:
                logger.info(f"  (no pude leer {p.name}: {e})")
                continue
            if filas:
                encontrados += 1
                facturas = sorted({f[2] for f in filas if len(f) > 2})
                muestra = ", ".join(facturas[:3]) + (" …" if len(facturas) > 3 else "")
                logger.info(f"  ✓ {p}")
                logger.info(f"      {len(filas)} fila(s), {len(facturas)} factura(s): {muestra}")
        if not encontrados:
            logger.info("  No hallé ningún archivo con filas de facturas HUS.")
        else:
            logger.info(f"\n  Usá el archivo correcto con --fur \"<ruta>\".")
        return

    fur_path = args.fur
    if not fur_path and args.carpeta:
        fur_path = hallar_fur_csv(args.carpeta)
    if not fur_path or not fur_path.is_file():
        if args.fur and not args.fur.is_file():
            logger.error(f"El archivo NO existe en esa ruta: {args.fur}")
            padre = args.fur.parent
            if padre.is_dir():
                cand = [p.name for p in sorted(padre.iterdir())
                        if p.suffix.lower() in (".csv", ".txt", ".tsv", ".xlsx", ".xlsm")]
                if cand:
                    logger.error("  Archivos candidatos en esa carpeta:")
                    for n in cand:
                        logger.error(f"    {n}")
        else:
            logger.error("No encontré el export del HUS. Pasalo con --fur o --carpeta.")
        sys.exit(1)
    logger.info(f"Export FUR del HUS: {fur_path.name}")

    if args.diagnostico:
        if not args.factura:
            logger.error("--diagnostico requiere --factura.")
            sys.exit(1)
        diagnostico(fur_path, args.factura)
        return

    nit = args.nit or (nit_desde_rips(args.carpeta) if args.carpeta else None) or NIT_HUS_DEFECTO
    logger.info(f"NIT prestador: {nit}")

    filas = leer_fur_hus(fur_path, nit)
    logger.info(f"  Facturas en el export: {len(filas)}")

    if args.factura:
        objetivo = args.factura.strip().upper()
        filas = [f for f in filas if (f.get("NUM_FACTURA") or "").strip().upper() == objetivo]
        if not filas:
            logger.error(f"No encontré la factura {args.factura} en el export.")
            sys.exit(1)

    if not args.salida:
        logger.error("Falta --salida.")
        sys.exit(1)

    # Reporte de validación.
    total_faltan = 0
    for fila in filas:
        faltan = validar_fila(fila)
        if faltan:
            total_faltan += 1
            fac = fila.get("NUM_FACTURA", "?")
            logger.info(f"  ⚠ {fac}: {len(faltan)} campo(s) por completar → {', '.join(faltan[:6])}"
                        + (" …" if len(faltan) > 6 else ""))
    if total_faltan:
        logger.info(f"  {total_faltan}/{len(filas)} factura(s) con campos pendientes (resaltados en rojo).")
    else:
        logger.info("  Todas las facturas pasaron la validación de obligatorios.")

    escribir_excel(filas, args.salida)
    logger.info(f"  ✓ Excel FUR generado: {args.salida}")


if __name__ == "__main__":
    main()
