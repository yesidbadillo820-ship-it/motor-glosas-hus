"""generar_fur_servicios.py — Genera el Excel FUR SERVICIOS desde el RIPS.

Lee el RIPS de una factura y arma el Excel del **FUR SERVICIOS** (Diccionario
de 13 campos, Resolución 2284/2023). El Excel sale **limpio**: una sola hoja,
una sola fuente y tamaño, sin colores ni filas de leyenda (igual al ejemplo
oficial). Las columnas que no salen del RIPS quedan vacías para que las
complete el prestador.

Columnas auto-rellenadas desde el RIPS (+ catálogo CUPS para la descripción
de consultas/procedimientos):
    Número factura, NIT, Codificación CUPS, Descripción, Cantidad,
    Valor unitario/total facturado, Código del servicio (medicamentos/otros)
    y Tipo de servicio.

Columnas que completa el prestador (quedan vacías):
    Código general del procedimiento quirúrgico, Consecutivo quirúrgico,
    Código del servicio (SOAT) de procedimientos. Los "reclamado" se pre-cargan
    = facturado como punto de partida.

USO
---

    # Desde la carpeta de la factura (busca el *_RIP.json adentro)
    py generar_fur_servicios.py ^
        --carpeta "C:\\Users\\Usuario\\Downloads\\FACTURAS\\HUS428139" ^
        --salida  "C:\\Users\\Usuario\\Downloads\\FACTURAS\\HUS428139_FURSERVICIOS.xlsx"

    # O apuntando directo al RIPS
    py generar_fur_servicios.py --rips "...\\HUS428139_RIP.json" --salida salida.xlsx

DEPENDENCIA: openpyxl (pip install openpyxl)
"""

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from cups_catalogo import buscar as buscar_cups  # noqa: E402
from cups_catalogo import cargar_catalogo  # noqa: E402
from factura_lectura import hallar_factura_xml  # noqa: E402, F401
from factura_pdf_lectura import (  # noqa: E402
    descripciones_dian_pdf,
    hallar_factura_pdf,
    leer_factura_pdf,
)
from furips_lectura import (  # noqa: E402
    cargar_furips2,
    facturas_en_furips2,
    hallar_furips2,
)
from rips_lectura import (  # noqa: E402
    cargar_rips,
    datos_generales,
    extraer_lineas_servicios,
    inferir_tipo_por_codigo,
)

logger = logging.getLogger("gen_fur_servicios")

# (clave interna, encabezado legible) en el orden exacto del Diccionario FUR
# SERVICIOS. El encabezado visible coincide con el ejemplo oficial.
COLUMNAS = [
    ("NUM_FACTURA", "Número factura"),
    ("NIT_PRESTADOR", "NIT Prestador"),
    ("Tipo_de_servicio", "Tipo de servicio"),
    ("Codigo_general_del_procedimiento_quirurgico", "Código general del procedimiento quirúrgico"),
    ("Consecutivo_procedimiento_quirurgico", "Consecutivo Procedimiento quirúrgico"),
    ("Codigo_del_servicio", "Código del servicio"),
    ("Codificacion_CUPS", "Codificación CUPS"),
    (
        "Descripción_del_servicio_o_elemento_reclamado",
        "Descripción del servicio o elemento reclamado",
    ),
    ("Cantidad_de_servicios", "Cantidad de servicios"),
    ("Valor_unitario_facturado", "Valor unitario facturado"),
    ("Valor_unitario_reclamado", "Valor unitario reclamado"),
    ("Valor_total_facturado", "Valor total facturado"),
    ("Valor_total_reclamado", "Valor total reclamado"),
]


def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
    )


def _vr_eq(a: str, b: str) -> bool:
    """Compara dos importes que pueden venir como '83400', '83400.00' o ''."""
    try:
        return abs(float(a) - float(b)) < 0.5
    except (TypeError, ValueError):
        return False


def enriquecer_con_factura(linea: dict, items_factura: dict, usadas_por_precio: set) -> dict:
    """Completa descripción y Tipo de servicio desde la factura.

    Estrategia:
      1) Match por código directo (codTecnologiaSalud — funciona para
         medicamentos CUM, insumos FMQ/QX y osteosíntesis FMO).
      2) Si no hay match por código (típico de consultas/procedimientos:
         RIPS trae CUPS, factura trae SOAT), buscar UN solo ítem de la
         factura cuyo vr_total coincida y aún no se haya usado. Si hay
         empate, no se decide (queda para revisión manual).
    """
    cod = (linea.get("cod_servicio") or "").strip()
    enriquecida = dict(linea)

    # 1) Match por código directo
    if cod and cod in items_factura:
        item = items_factura[cod]
        if not enriquecida.get("descripcion") and item.get("descripcion"):
            enriquecida["descripcion"] = item["descripcion"]
        return enriquecida

    # 2) Match por valor (consultas/procedimientos donde RIPS=CUPS ≠ factura=SOAT)
    vr_total_rips = linea.get("vr_total") or ""
    if vr_total_rips:
        candidatos = [
            (k, v)
            for k, v in items_factura.items()
            if k not in usadas_por_precio and _vr_eq(v.get("vr_total", ""), vr_total_rips)
        ]
        if len(candidatos) == 1:
            cod_fact, item = candidatos[0]
            usadas_por_precio.add(cod_fact)
            if not enriquecida.get("descripcion") and item.get("descripcion"):
                enriquecida["descripcion"] = item["descripcion"]
            # Para procedimientos, el código del vendedor de la factura es el SOAT
            if linea.get("tipo_rips") in ("consultas", "procedimientos", "urgencias") and not cod:
                enriquecida["cod_servicio"] = cod_fact
        # Si hay múltiples candidatos o ninguno, dejamos como está (sin adivinar).

    return enriquecida


def fila_desde_linea(ln: dict) -> dict:
    """Mapea una línea normalizada a las 13 columnas del FUR SERVICIOS."""
    tipo = ln.get("tipo_servicio_fur") or inferir_tipo_por_codigo(ln.get("cod_servicio", ""))
    return {
        "NUM_FACTURA": ln.get("num_factura", ""),
        "NIT_PRESTADOR": ln.get("nit_prestador", ""),
        "Tipo_de_servicio": tipo,
        "Codigo_general_del_procedimiento_quirurgico": ln.get("cod_general_quirurgico", ""),
        "Consecutivo_procedimiento_quirurgico": ln.get("consecutivo_quirurgico", ""),
        "Codigo_del_servicio": ln.get("cod_servicio", ""),
        "Codificacion_CUPS": ln.get("cups", ""),
        "Descripción_del_servicio_o_elemento_reclamado": ln.get("descripcion", ""),
        "Cantidad_de_servicios": ln.get("cantidad", ""),
        "Valor_unitario_facturado": ln.get("vr_unitario", ""),
        # Punto de partida: reclamado = facturado (el prestador lo puede bajar)
        "Valor_unitario_reclamado": ln.get("vr_unitario", ""),
        "Valor_total_facturado": ln.get("vr_total", ""),
        "Valor_total_reclamado": ln.get("vr_total", ""),
    }


def _tipo_fur_desde_pdf(seccion: str, codigo: str) -> str:
    """Infiere el Tipo de servicio FUR (1-8) según sección del PDF y código."""
    cod_up = (codigo or "").upper()
    if cod_up.startswith("FMO"):
        return "7"  # Material de osteosíntesis
    if cod_up.startswith("FMQ") or cod_up.startswith("QX"):
        return "5"  # Insumos
    if seccion == "medicamentos":
        return "1"
    if seccion == "osteo":
        return "7"
    if seccion == "insumos":
        return "5"
    if seccion in ("consultas", "procedimientos", "quirurgico", "estancias", "derechos_sala"):
        return "2"
    return ""


def _linea_desde_pdf(meta: dict, ln_pdf: dict) -> dict:
    """Convierte una línea parseada del PDF al formato normalizado."""
    seccion = ln_pdf.get("seccion", "")
    codigo = ln_pdf.get("codigo", "")
    return {
        "num_factura": meta.get("num_factura", ""),
        "nit_prestador": meta.get("nit_prestador", ""),
        "tipo_servicio_fur": _tipo_fur_desde_pdf(seccion, codigo),
        "cod_servicio": codigo,
        "cups": "",  # se enlaza luego desde el RIPS por valor
        "descripcion": ln_pdf.get("descripcion", ""),
        "cantidad": ln_pdf.get("cantidad", ""),
        "vr_unitario": ln_pdf.get("vr_unitario", ""),
        "vr_total": ln_pdf.get("vr_total", ""),
        "cod_general_quirurgico": ln_pdf.get("cod_general_quirurgico", ""),
        "consecutivo_quirurgico": ln_pdf.get("consecutivo_quirurgico", ""),
        "_seccion": seccion,
    }


def _cups_por_valor_desde_rips(data: dict) -> dict[str, list[str]]:
    """Indexa CUPS del RIPS por vr_total (string). Sirve para enlazar con el PDF."""
    idx: dict[str, list[str]] = {}
    for u in data.get("usuarios") or []:
        servicios = u.get("servicios") or {}
        for tipo in ("consultas", "procedimientos", "urgencias"):
            for it in servicios.get(tipo) or []:
                cups = it.get("codProcedimiento") or it.get("codConsulta") or ""
                vr = it.get("vrServicio")
                if not cups or vr in (None, "", 0):
                    continue
                key = str(int(vr)) if isinstance(vr, (int, float)) else str(vr)
                idx.setdefault(key, []).append(cups)
    return idx


def _descripciones_por_codigo_desde_rips(data: dict) -> dict[str, str]:
    """Indexa nombre de tecnología por código (codTecnologiaSalud) desde el RIPS.

    Sirve para llenar la descripción de medicamentos (cuyo CUM aparece tanto
    en el RIPS como en el FURIPS2) y de cualquier otro servicio donde el
    código del FURIPS2 coincida con el del RIPS.
    """
    idx: dict[str, str] = {}
    for u in data.get("usuarios") or []:
        servicios = u.get("servicios") or {}
        for tipo in ("consultas", "procedimientos", "medicamentos",
                     "otrosServicios", "urgencias", "recienNacidos"):
            for it in servicios.get(tipo) or []:
                cod = (it.get("codTecnologiaSalud") or "").strip()
                nom = (it.get("nomTecnologiaSalud") or "").strip()
                if cod and nom:
                    idx.setdefault(cod, nom)
    return idx


def escribir_excel(filas: list[dict], ruta: Path, meta: dict) -> None:
    """Escribe el Excel FUR SERVICIOS limpio: una sola hoja, una sola fuente y
    tamaño, sin colores ni filas de leyenda. Igual al formato del ejemplo oficial.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.stderr.write("ERROR: falta openpyxl.\n  Instalalo con:  py -m pip install openpyxl\n")
        sys.exit(2)

    wb = Workbook()
    ws = wb.active
    ws.title = "FUR_SERVICIOS"

    anchos = [14, 12, 9, 16, 14, 18, 12, 42, 11, 14, 14, 14, 14]

    # Fila 1: encabezados (solo negrita, misma fuente y tamaño por defecto)
    for col_idx, (_clave, encabezado) in enumerate(COLUMNAS, 1):
        c = ws.cell(row=1, column=col_idx, value=encabezado)
        c.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = anchos[col_idx - 1]

    # Filas de datos: texto plano, sin formato ni color
    for i, fila in enumerate(filas, start=2):
        for col_idx, (clave, _enc) in enumerate(COLUMNAS, 1):
            ws.cell(row=i, column=col_idx, value=fila.get(clave, ""))

    ws.freeze_panes = "A2"

    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta)


def hallar_rips(carpeta: Path) -> Path | None:
    candidatos = [
        p
        for p in carpeta.iterdir()
        if p.is_file()
        and p.suffix.lower() == ".json"
        and "RIP" in p.name.upper()
        and "CUV" not in p.name.upper()
    ]
    return candidatos[0] if candidatos else None


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    grupo = parser.add_mutually_exclusive_group(required=True)
    grupo.add_argument("--carpeta", type=Path, help="Carpeta de la factura (busca el *_RIP.json)")
    grupo.add_argument("--rips", type=Path, help="Ruta directa al RIPS .json")
    parser.add_argument("--salida", type=Path, required=True, help="Excel .xlsx de salida")
    parser.add_argument(
        "--cups",
        type=Path,
        default=None,
        help="Catálogo CUPS oficial (código→nombre) CSV/TSV/XLSX. Resuelve la "
        "descripción de consultas/procedimientos que el RIPS no trae.",
    )
    parser.add_argument(
        "--soat",
        type=Path,
        default=None,
        help="Manual tarifario ISS/SOAT (código→descripción) CSV/TSV/XLSX. "
        "Resuelve por código del servicio los procedimientos que un mismo "
        "código de tarifa cubre (los que si no quedarían marcados [ELEGIR]).",
    )
    parser.add_argument(
        "--furips2",
        type=Path,
        default=None,
        help="Ruta al FURIPS2*.txt (detalle de servicios). Si se omite y se "
        "pasó --carpeta, se busca en la carpeta y en su padre. Cuando está "
        "presente, es la fuente principal del detalle (en vez del PDF).",
    )
    args = parser.parse_args()
    setup_logging()

    if args.carpeta:
        if not args.carpeta.is_dir():
            logger.error(f"No existe la carpeta: {args.carpeta}")
            return 1
        rips_path = hallar_rips(args.carpeta)
        if not rips_path:
            logger.error(f"No encontré un *_RIP.json en {args.carpeta}")
            return 1
    else:
        rips_path = args.rips
        if not rips_path.is_file():
            logger.error(f"No existe el RIPS: {rips_path}")
            return 1

    logger.info(f"RIPS: {rips_path.name}")
    data = cargar_rips(rips_path)
    meta = datos_generales(data)

    # Fuente del detalle por orden de preferencia:
    #   1) FURIPS2 (archivo plano del HUS): trae código SOAT, descripción,
    #      cantidad, valores y descomposición quirúrgica ya estructurada.
    #      Es la fuente más confiable cuando está disponible.
    #   2) PDF de la factura: parser sobre el "Detalle de Cargos".
    #   3) RIPS: fallback. Pierde la descomposición SOAT/quirúrgica.
    # Aunque FURIPS2 sea la fuente principal, el PDF se carga también para
    # enriquecer descripciones de procedimientos (representación DIAN) y el
    # RIPS para el CUPS por valor.
    furips2_path: Path | None = args.furips2
    if not furips2_path and args.carpeta:
        furips2_path = hallar_furips2(args.carpeta)

    pdf_path: Path | None = None
    if args.carpeta:
        pdf_path = hallar_factura_pdf(args.carpeta)

    if furips2_path:
        logger.info(f"FURIPS2: {furips2_path.name}")
        factura_id = meta["num_factura"]
        lineas = cargar_furips2(
            furips2_path, factura_id, nit_prestador=meta["nit_prestador"]
        )
        logger.info(f"  Líneas leídas del FURIPS2: {len(lineas)}")
        if not lineas:
            disponibles = facturas_en_furips2(furips2_path)[:8]
            logger.warning(
                f"  FURIPS2 no tiene líneas para {factura_id!r}. "
                f"Facturas en el archivo (primeras 8): {disponibles}"
            )
    elif pdf_path:
        logger.info(f"PDF Factura: {pdf_path.name}")
        try:
            lineas_pdf = leer_factura_pdf(pdf_path)
        except SystemExit:
            raise
        except Exception as e:
            logger.error(f"No pude parsear el PDF de la factura: {e}")
            lineas_pdf = []
        logger.info(f"  Líneas leídas del PDF: {len(lineas_pdf)}")
        # Convertimos cada línea del PDF al formato normalizado que usa fila_desde_linea
        lineas = [_linea_desde_pdf(meta, ln) for ln in lineas_pdf]
        pdf_path = None  # ya consumido como fuente principal; no usar para enriquecer
    else:
        logger.warning("Ni FURIPS2 ni PDF encontrados — fallback al RIPS.")
        lineas = extraer_lineas_servicios(data)

    if not lineas:
        logger.error("No hay líneas para escribir (PDF/RIPS vacíos).")
        return 1

    # Descripción desde el RIPS por código directo: cubre medicamentos (CUM)
    # cuyo codTecnologiaSalud existe tanto en RIPS como en FURIPS2.
    desc_por_codigo = _descripciones_por_codigo_desde_rips(data)
    enlazados_desc_rips = 0
    for ln in lineas:
        if ln.get("descripcion"):
            continue
        cod = (ln.get("cod_servicio") or "").strip()
        if cod and cod in desc_por_codigo:
            ln["descripcion"] = desc_por_codigo[cod]
            enlazados_desc_rips += 1
    logger.info(f"  Descripciones desde RIPS (por código): {enlazados_desc_rips}")

    # Descripción desde la representación DIAN del PDF: cubre procedimientos
    # SOAT cuya descripción no viene en el FURIPS2 (21712, 21102, 19304…).
    # Se separa en precios ÚNICOS (un solo servicio → se asigna limpio) y
    # precios AMBIGUOS (un código de tarifa que cubre varios procedimientos →
    # se marcan con [ELEGIR] al final, ver más abajo).
    desc_ambiguas_dian: dict[str, str] = {}
    marcadas = 0
    if pdf_path:
        try:
            desc_unicas_dian, desc_ambiguas_dian = descripciones_dian_pdf(pdf_path)
        except Exception as e:
            logger.warning(f"  No pude extraer descripciones del PDF DIAN: {e}")
            desc_unicas_dian, desc_ambiguas_dian = {}, {}
        logger.info(f"  Precios con descripción única en PDF DIAN: {len(desc_unicas_dian)}")
        logger.info(
            f"  Precios ambiguos en PDF DIAN (un código, varios servicios): "
            f"{len(desc_ambiguas_dian)}"
        )
        enlazados_desc_dian = 0
        for ln in lineas:
            if ln.get("descripcion"):
                continue
            vr_u = ln.get("vr_unitario", "")
            if vr_u and vr_u in desc_unicas_dian:
                ln["descripcion"] = desc_unicas_dian[vr_u]
                enlazados_desc_dian += 1
        logger.info(f"  Descripciones desde PDF DIAN (precio único): {enlazados_desc_dian}")

    # Propagación por código SOAT: si una línea quedó sin descripción pero otra
    # con el MISMO cod_servicio ya la tiene (de cualquier fuente anterior), la
    # copiamos — mismo código = mismo servicio. Cubre el caso en que sólo
    # algunas repeticiones del código lograron enlazar por valor.
    desc_por_cod_servicio: dict[str, str] = {}
    for ln in lineas:
        cod = (ln.get("cod_servicio") or "").strip()
        d = (ln.get("descripcion") or "").strip()
        if cod and d:
            desc_por_cod_servicio.setdefault(cod, d)
    propagados = 0
    for ln in lineas:
        if ln.get("descripcion"):
            continue
        cod = (ln.get("cod_servicio") or "").strip()
        if cod and cod in desc_por_cod_servicio:
            ln["descripcion"] = desc_por_cod_servicio[cod]
            propagados += 1
    logger.info(f"  Descripciones propagadas por código SOAT: {propagados}")

    # Manual tarifario ISS/SOAT (opcional): descripción oficial por código del
    # servicio. Resuelve los códigos de tarifa que cubren varios procedimientos
    # (los que si no quedarían marcados [ELEGIR]) con el nombre canónico del
    # manual. Sólo llena lo que sigue vacío, así las descripciones específicas y
    # limpias del DIAN/RIPS mantienen prioridad.
    if args.soat:
        if not args.soat.is_file():
            logger.warning(
                f"  ⚠ --soat apunta a un archivo que NO EXISTE: {args.soat}. "
                "Los códigos ambiguos quedarán marcados [ELEGIR]."
            )
        else:
            cat_soat = cargar_catalogo(args.soat)
            logger.info(f"Manual ISS/SOAT: {len(cat_soat)} códigos cargados desde {args.soat.name}")
            if not cat_soat:
                logger.warning(
                    "  ⚠ El catálogo --soat se cargó VACÍO. Revisá el formato "
                    "(encabezados código/descripción, separador correcto)."
                )
            enlazados_soat = 0
            for ln in lineas:
                if ln.get("descripcion"):
                    continue
                cod = (ln.get("cod_servicio") or "").strip()
                nom = buscar_cups(cat_soat, cod) if cod else ""
                if nom:
                    ln["descripcion"] = nom
                    enlazados_soat += 1
            logger.info(f"  Descripciones desde manual ISS/SOAT (por código): {enlazados_soat}")

    # Candidatos ambiguos del DIAN: para precios donde un código SOAT cubre
    # varios procedimientos, no hay forma automática de elegir. Se rellenan los
    # renglones que sigan vacíos con los candidatos unidos y marcados [ELEGIR];
    # el coordinador escoge el correcto antes de radicar. Se aplica de ÚLTIMO,
    # para que cualquier fuente limpia anterior tenga prioridad.
    if desc_ambiguas_dian:
        for ln in lineas:
            if ln.get("descripcion"):
                continue
            vr_u = ln.get("vr_unitario", "")
            if vr_u and vr_u in desc_ambiguas_dian:
                ln["descripcion"] = desc_ambiguas_dian[vr_u]
                marcadas += 1
        logger.info(f"  Líneas marcadas [ELEGIR] (resolver antes de radicar): {marcadas}")

    # CUPS desde el RIPS: enlazamos por vr_total (1:1 cuando es único). El RIPS
    # es la única fuente del CUPS para procedimientos/consultas.
    cups_por_valor = _cups_por_valor_desde_rips(data)
    enlazados_cups = 0
    for ln in lineas:
        if ln.get("cups"):
            continue
        if ln.get("tipo_servicio_fur") not in ("", "2"):
            continue
        candidatos = cups_por_valor.get(ln.get("vr_total", ""), [])
        if len(candidatos) == 1:
            ln["cups"] = candidatos[0]
            enlazados_cups += 1
    logger.info(f"  CUPS enlazados desde RIPS (por valor único): {enlazados_cups}")

    # Catálogo CUPS opcional: nombre oficial si quedó algo sin descripción
    if args.cups and args.cups.is_file():
        cat = cargar_catalogo(args.cups)
        logger.info(f"Catálogo CUPS: {len(cat)} códigos")
        rellenadas = 0
        for ln in lineas:
            if not ln.get("descripcion") and ln.get("cups"):
                nom = buscar_cups(cat, ln["cups"])
                if nom:
                    ln["descripcion"] = nom
                    rellenadas += 1
        logger.info(f"  Descripciones rellenadas con catálogo CUPS: {rellenadas}")

    if marcadas:
        logger.info(
            f"  ⚠ {marcadas} línea(s) con [ELEGIR]: elegí el procedimiento correcto "
            f"(el código de tarifa cubre varios) antes de radicar."
        )
    sin_desc = sum(1 for ln in lineas if not ln.get("descripcion"))
    if sin_desc:
        logger.info(f"  Líneas aún sin descripción (revisar manual): {sin_desc}")

    filas = [fila_desde_linea(ln) for ln in lineas]
    escribir_excel(filas, args.salida, meta)

    logger.info(f"  Factura: {meta['num_factura']}   NIT: {meta['nit_prestador']}")
    logger.info(f"  Líneas de servicio: {len(filas)}")
    logger.info(f"  ✓ Excel generado (limpio, sin formato): {args.salida}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
