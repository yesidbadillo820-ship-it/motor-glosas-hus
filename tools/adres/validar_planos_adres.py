"""validar_planos_adres.py — Validador de archivos planos de la Circular 022/2023 ADRES.

Valida los archivos planos de reclamaciones ante la ADRES **antes de
radicarlos**: revisa la NOMENCLATURA del nombre del archivo, las REGLAS
GENERALES del anexo técnico (delimitador coma, fechas DD/MM/AAAA, números
sin separadores, sin comillas, sin relleno, longitudes MÁXIMAS) y, donde el
anexo está cargado, la malla campo a campo. Entrega un INFORME EXCEL
detallado (hojas RESUMEN con semáforo por archivo y errores más repetidos,
HALLAZGOS, POR CAMPO, AVISOS y LEYENDA) y, si se pide, el mismo reporte en
JSON o CSV con: Nombre del Archivo, Línea, Campo, Valor y Descripción.

ARCHIVOS QUE RECONOCE (nomenclatura estricta, validada con expresión regular)
----------------------------------------------------------------------------
    FURIPS1 + código habilitación (12 dígitos) + fecha DDMMAAAA   → Tabla 1
    FURIPS2 + código habilitación (12 dígitos) + fecha DDMMAAAA   → Tabla 2
    FURTRAN + código habilitación (12 dígitos) + fecha DDMMAAAA   → Tabla 3
    FUCTAS2 + código aseguradora (6 dígitos) + período MMAAAA     → Tabla 4
    FURCEN  + código del evento + fecha DDMMAAAA + hora HHMM      → Tabla 5

    (Ejemplo válido: FURIPS112345678901201102023.txt)

MALLAS DE CONTENIDO
-------------------
- FURIPS 1 (102 campos) y FURIPS 2 (9 campos): se reutilizan las tablas E1 y
  E2 de `validar_furips.py` (construidas desde la propia Circular 022/2023).
  Aquí se validan longitudes, formatos, valores permitidos y obligatorios
  incondicionales; las obligatoriedades CONDICIONALES (dependen de otros
  campos o del FURIPS2) las evalúa el validador profundo `validar_furips.py`.
- FURTRAN, FUCTAS y FURCEN: la nomenclatura y las reglas generales se validan
  YA; la malla campo a campo queda como punto de extensión (`TABLAS[...] =
  None`) hasta cargar el anexo técnico correspondiente de la Circular — este
  validador NO inventa especificaciones.

REGLAS GENERALES QUE APLICA A TODA LÍNEA (Circular 022/2023)
------------------------------------------------------------
- Extensión .txt (un .csv se procesa igual pero se advierte que la Circular
  exige .txt). Los JSON no son anexos FURIPS: se avisa y se omiten (los JSON
  de RIPS/CUV los valida `validar_furips.py`).
- Campos separados SOLO por coma; los vacíos se conservan entre comas (,,).
- Prohibido encerrar campos entre comillas dobles y prohibidos los
  caracteres de control / fin de archivo dentro de los registros.
- Una coma dentro de un campo de texto libre descuadra el registro: se
  reporta como "número de campos distinto del esperado".
- Fechas en DD/MM/AAAA (con los slash) y que existan en el calendario.
- Números sin puntos ni comas (ni miles ni decimales).
- Prohibido el relleno: espacios a la izquierda/derecha en cualquier campo;
  ceros a la izquierda en consecutivos, cantidades y valores. (Los códigos
  DANE de departamento/municipio SÍ conservan su cero inicial: no se tocan.)
- Las longitudes del anexo son MÁXIMOS, no tamaños fijos.

USO
---
    py validar_planos_adres.py --ruta "C:\\FACTURAS\\ADRES"        (carpeta)
    py validar_planos_adres.py --ruta FURIPS1...01102023.txt      (un archivo)
    py validar_planos_adres.py --ruta . --formato json            (solo JSON)

Salida: REPORTE_PLANOS_ADRES_<fecha>.xlsx en la carpeta validada (con
--formato también .csv o .json). Solo LEE los archivos; no modifica nada.
Única dependencia: openpyxl para el Excel (sin él, entrega CSV).
Normalmente se lanza con doble clic desde VALIDAR_PLANOS_ADRES.cmd.
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import re
import sys
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path

# Malla FURIPS construida desde la Circular 022/2023 (mismo directorio).
sys.path.insert(0, str(Path(__file__).resolve().parent))
from validar_furips import E1, E2, TEXTO_OBLIG  # noqa: E402

logger = logging.getLogger("validar_planos_adres")

ERROR = "ERROR"
ADVERTENCIA = "ADVERTENCIA"
INFO = "INFO"

# ─────────────────────────────────────────────────────────────────────────────
# 1) NOMENCLATURA de los archivos (expresiones regulares de la Circular)
# ─────────────────────────────────────────────────────────────────────────────

RE_FURIPS1 = re.compile(r"^FURIPS1(?P<habilitacion>\d{12})(?P<fecha>\d{8})$", re.IGNORECASE)
RE_FURIPS2 = re.compile(r"^FURIPS2(?P<habilitacion>\d{12})(?P<fecha>\d{8})$", re.IGNORECASE)
RE_FURTRAN = re.compile(r"^FURTRAN(?P<habilitacion>\d{12})(?P<fecha>\d{8})$", re.IGNORECASE)
RE_FUCTAS = re.compile(r"^FUCTAS2(?P<aseguradora>\d{6})(?P<periodo>\d{6})$", re.IGNORECASE)
# FURCEN: el código del evento no tiene longitud fija en el nombre; el motor
# separa desde el FINAL (últimos 12 dígitos = DDMMAAAA + HHMM).
RE_FURCEN = re.compile(r"^FURCEN(?P<resto>\d{13,})$", re.IGNORECASE)


def _fecha_ddmmaaaa_valida(texto: str) -> bool:
    try:
        datetime.strptime(texto, "%d%m%Y")
        return True
    except ValueError:
        return False


def _periodo_mmaaaa_valido(texto: str) -> bool:
    if len(texto) != 6 or not texto.isdigit():
        return False
    mes, anio = int(texto[:2]), int(texto[2:])
    return 1 <= mes <= 12 and 2000 <= anio <= 2100


def _hora_hhmm_valida(texto: str) -> bool:
    if len(texto) != 4 or not texto.isdigit():
        return False
    return 0 <= int(texto[:2]) <= 23 and 0 <= int(texto[2:]) <= 59


def detectar_tipo(nombre: str) -> str | None:
    """FURIPS1 / FURIPS2 / FURTRAN / FUCTAS / FURCEN según el prefijo."""
    base = nombre.upper()
    for prefijo, tipo in (
        ("FURIPS1", "FURIPS1"),
        ("FURIPS2", "FURIPS2"),
        ("FURTRAN", "FURTRAN"),
        ("FUCTAS", "FUCTAS"),
        ("FURCEN", "FURCEN"),
    ):
        if base.startswith(prefijo):
            return tipo
    return None


def validar_nomenclatura(ruta: Path) -> list[Hallazgo]:
    """Valida el NOMBRE del archivo contra la estructura de la Circular."""
    errores: list[Hallazgo] = []
    nombre, base, ext = ruta.name, ruta.stem, ruta.suffix.lower()
    tipo = detectar_tipo(base)

    def err(descripcion: str, severidad: str = ERROR) -> None:
        errores.append(Hallazgo(nombre, 0, "Nombre del archivo", base, descripcion, severidad))

    if ext == ".csv":
        err("La Circular 022/2023 exige extensión .txt (se validará igual).", ADVERTENCIA)
    elif ext != ".txt":
        err(f"Extensión '{ext}' no permitida: los anexos FURIPS son .txt.")

    if tipo in ("FURIPS1", "FURIPS2", "FURTRAN"):
        rx = {"FURIPS1": RE_FURIPS1, "FURIPS2": RE_FURIPS2, "FURTRAN": RE_FURTRAN}[tipo]
        m = rx.match(base)
        if not m:
            err(
                f"El nombre no cumple {tipo} + código de habilitación (12 dígitos) "
                f"+ fecha DDMMAAAA. Ejemplo: {tipo}12345678901201102023.txt"
            )
        elif not _fecha_ddmmaaaa_valida(m.group("fecha")):
            err(f"La fecha del nombre '{m.group('fecha')}' no es una fecha DDMMAAAA real.")
    elif tipo == "FUCTAS":
        m = RE_FUCTAS.match(base)
        if not m:
            err(
                "El nombre no cumple FUCTAS2 + código de aseguradora (6 dígitos) "
                "+ período MMAAAA. Ejemplo: FUCTAS2123456012026.txt"
            )
        elif not _periodo_mmaaaa_valido(m.group("periodo")):
            err(f"El período del nombre '{m.group('periodo')}' no es un MMAAAA válido.")
    elif tipo == "FURCEN":
        m = RE_FURCEN.match(base)
        if not m:
            err(
                "El nombre no cumple FURCEN + código del evento + fecha DDMMAAAA "
                "+ hora HHMM (todo en dígitos). Ejemplo: FURCEN001011020231430.txt"
            )
        else:
            resto = m.group("resto")
            fecha, hora = resto[-12:-4], resto[-4:]
            if not _fecha_ddmmaaaa_valida(fecha):
                err(f"La fecha del nombre '{fecha}' no es una fecha DDMMAAAA real.")
            if not _hora_hhmm_valida(hora):
                err(f"La hora del nombre '{hora}' no es una hora HHMM válida (00:00-23:59).")
    else:
        err(
            "El nombre no corresponde a ningún archivo de la Circular 022/2023 "
            "(FURIPS1, FURIPS2, FURTRAN, FUCTAS2 o FURCEN)."
        )
    return errores


# ─────────────────────────────────────────────────────────────────────────────
# 2) Modelo del hallazgo y reporte
# ─────────────────────────────────────────────────────────────────────────────


@dataclass
class Hallazgo:
    archivo: str
    linea: int  # 0 = el problema es del nombre o de la estructura del archivo
    campo: str  # "23 - Concepto" o "Nombre del archivo"
    valor: str
    descripcion: str
    severidad: str  # ERROR | ADVERTENCIA | INFO


# ─────────────────────────────────────────────────────────────────────────────
# 3) REGLAS GENERALES (aplican a todos los archivos de la Circular)
# ─────────────────────────────────────────────────────────────────────────────

_RE_FECHA_SLASH = re.compile(r"^\d{2}/\d{2}/\d{4}$")
_RE_SOLO_DIGITOS = re.compile(r"^\d+$")
_RE_CONTROL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
# Ceros a la izquierda = relleno prohibido SOLO donde no hay códigos DANE:
_RE_CONCEPTO_SIN_CEROS = re.compile(r"consecutivo|cantidad|valor", re.IGNORECASE)


def _validar_fecha_campo(valor: str) -> str | None:
    if not _RE_FECHA_SLASH.match(valor):
        return "Fecha no tiene formato DD/MM/AAAA (con los slash)."
    try:
        datetime.strptime(valor, "%d/%m/%Y")
    except ValueError:
        return f"La fecha '{valor}' no existe en el calendario (¿día y mes al revés?)."
    return None


def _validar_numero_campo(valor: str) -> str | None:
    if _RE_SOLO_DIGITOS.match(valor):
        return None
    if re.fullmatch(r"[\d.,]+", valor):
        if "." in valor and valor.count(".") == 1 and valor.split(".")[1].isdigit():
            return "Contiene decimales: los valores van sin decimales."
        return "Contiene separador de miles (punto o coma): los números van corridos."
    return f"Debe ser numérico y contiene caracteres no numéricos ('{valor[:20]}')."


def validar_campo_global(valor: str, etiqueta: str, archivo: str, n_linea: int) -> list[Hallazgo]:
    """Reglas generales que aplican a CUALQUIER campo, con o sin anexo."""
    errores: list[Hallazgo] = []

    def err(descripcion: str, severidad: str = ERROR) -> None:
        errores.append(Hallazgo(archivo, n_linea, etiqueta, valor[:80], descripcion, severidad))

    if '"' in valor:
        err('El campo contiene comillas dobles ("): prohibido encerrar campos en comillas.')
    if _RE_CONTROL.search(valor):
        err("El campo contiene caracteres de control / fin de archivo: prohibidos.")
    if valor != valor.strip():
        err("Relleno con espacios a la izquierda o derecha: prohibido para justificar campos.")
    return errores


def validar_campo_con_anexo(
    valor: str,
    numero: int,
    concepto: str,
    longitud: int | None,
    formato: str | None,
    permitidos: set[str] | None,
    oblig: str,
    archivo: str,
    n_linea: int,
) -> list[Hallazgo]:
    """Reglas del anexo técnico para UN campo: longitud máxima, formato,
    valores permitidos, obligatoriedad incondicional y relleno con ceros."""
    etiqueta = f"{numero} - {concepto}"
    errores = validar_campo_global(valor, etiqueta, archivo, n_linea)

    def err(descripcion: str, severidad: str = ERROR) -> None:
        errores.append(Hallazgo(archivo, n_linea, etiqueta, valor[:80], descripcion, severidad))

    if not valor:
        if oblig == "SI":
            err("Campo obligatorio vacío según la Circular 022/2023.")
        elif oblig not in ("NO", "SI"):
            err(
                f"Campo vacío con obligatoriedad condicional ({TEXTO_OBLIG.get(oblig, oblig)}): "
                "verificar con el validador profundo (validar_furips.py).",
                INFO,
            )
        return errores

    if longitud and len(valor) > longitud:
        err(f"Excede la longitud máxima de {longitud} (tiene {len(valor)} caracteres).")

    if formato == "fecha":
        detalle = _validar_fecha_campo(valor)
        if detalle:
            err(detalle)
    elif formato == "hora":
        if not re.fullmatch(r"\d{2}:\d{2}", valor) or not _hora_hhmm_valida(valor.replace(":", "")):
            err("Hora no tiene formato HH:MM válido (00:00-23:59).")
    elif formato == "num":
        detalle = _validar_numero_campo(valor)
        if detalle:
            err(detalle)
        elif len(valor) > 1 and valor.startswith("0") and _RE_CONCEPTO_SIN_CEROS.search(concepto):
            err("Relleno con ceros a la izquierda: prohibido para justificar campos.")
    elif formato == "cie10":
        if not re.fullmatch(r"[A-Z]\d{2}[0-9A-Z]?", valor.upper()):
            err(f"'{valor}' no es un código CIE-10 válido (letra + 2 dígitos [+ 1).")
    elif formato in ("depto", "mun"):
        # DIVIPOLA partido como lo pide la Circular: departamento (2) y
        # municipio (3, sin el prefijo del departamento).
        esperado = 2 if formato == "depto" else 3
        if not valor.isdigit() or len(valor) != esperado:
            nombre_fmt = "departamento" if formato == "depto" else "municipio"
            err(f"Código DANE de {nombre_fmt} debe tener {esperado} dígitos.")

    if permitidos and valor not in permitidos:
        muestra = ", ".join(sorted(permitidos)[:12])
        err(f"Valor '{valor}' no está en los permitidos de la Circular: {muestra}.")
    return errores


# ─────────────────────────────────────────────────────────────────────────────
# 4) MALLAS por archivo (anexos técnicos). Formato común de campo:
#    (numero, concepto, longitud_max, formato, permitidos, obligatoriedad)
# ─────────────────────────────────────────────────────────────────────────────

_MALLA_FURIPS1 = [(n, c, lon, fmt, perm, ob) for (n, _sec, c, lon, fmt, perm, ob) in E1]
_MALLA_FURIPS2 = [(n, c, lon, fmt, None, ob) for (n, c, lon, fmt, ob) in E2]

# FURTRAN (Tabla 3), FUCTAS (Tabla 4) y FURCEN (Tabla 5): PENDIENTES de cargar
# desde el anexo técnico de la Circular. None = solo nomenclatura + reglas
# generales (este validador NO inventa especificaciones de campos).
TABLAS: dict[str, list | None] = {
    "FURIPS1": _MALLA_FURIPS1,
    "FURIPS2": _MALLA_FURIPS2,
    "FURTRAN": None,
    "FUCTAS": None,
    "FURCEN": None,
}


def validar_linea(linea: str, n_linea: int, tipo: str, archivo: str) -> list[Hallazgo]:
    """Valida UNA línea: estructura (número de campos) + campo a campo."""
    errores: list[Hallazgo] = []
    campos = linea.split(",")
    malla = TABLAS.get(tipo)

    if malla is not None and len(campos) != len(malla):
        errores.append(
            Hallazgo(
                archivo,
                n_linea,
                "Registro completo",
                linea[:80],
                f"La línea tiene {len(campos)} campos y la Tabla de {tipo} exige "
                f"{len(malla)} separados por coma. Si algún texto libre trae una "
                "coma interna, retirarla: descuadra todo el registro.",
                ERROR,
            )
        )

    if malla is None:
        for i, valor in enumerate(campos, 1):
            errores.extend(validar_campo_global(valor, f"Campo {i}", archivo, n_linea))
        return errores

    for (numero, concepto, longitud, formato, permitidos, oblig), valor in zip(
        malla, campos, strict=False
    ):
        errores.extend(
            validar_campo_con_anexo(
                valor, numero, concepto, longitud, formato, permitidos, oblig, archivo, n_linea
            )
        )
    return errores


# ─────────────────────────────────────────────────────────────────────────────
# 5) Procesamiento de archivos y reporte
# ─────────────────────────────────────────────────────────────────────────────


def _leer_texto(ruta: Path) -> str:
    datos = ruta.read_bytes()
    try:
        return datos.decode("utf-8-sig")
    except UnicodeDecodeError:
        return datos.decode("latin-1")


def validar_archivo(ruta: Path) -> list[Hallazgo]:
    """Valida un archivo completo: nomenclatura + contenido línea a línea."""
    nombre = ruta.name
    if ruta.suffix.lower() == ".json":
        return [
            Hallazgo(
                nombre,
                0,
                "Archivo",
                "",
                "Los anexos FURIPS de la Circular 022/2023 son planos .txt; los JSON "
                "(RIPS/CUV) se validan con validar_furips.py. Archivo omitido.",
                ADVERTENCIA,
            )
        ]

    errores = validar_nomenclatura(ruta)
    tipo = detectar_tipo(ruta.stem)
    if tipo is None:
        return errores
    if TABLAS.get(tipo) is None:
        errores.append(
            Hallazgo(
                nombre,
                0,
                "Archivo",
                tipo,
                f"El anexo de campos de {tipo} aún no está cargado en el validador: "
                "se aplican la nomenclatura y las reglas generales de la Circular.",
                INFO,
            )
        )

    try:
        texto = _leer_texto(ruta)
    except OSError as e:
        errores.append(Hallazgo(nombre, 0, "Archivo", "", f"No se pudo leer: {e}.", ERROR))
        return errores

    lineas = texto.splitlines()
    while lineas and not lineas[-1].strip():
        lineas.pop()
    if not lineas:
        errores.append(Hallazgo(nombre, 0, "Archivo", "", "El archivo está vacío.", ERROR))
        return errores

    for n, linea in enumerate(lineas, 1):
        if not linea.strip():
            errores.append(
                Hallazgo(nombre, n, "Registro completo", "", "Línea en blanco intermedia.", ERROR)
            )
            continue
        errores.extend(validar_linea(linea, n, tipo, nombre))
    return errores


def descubrir_archivos(ruta: Path) -> list[Path]:
    """Archivos de la Circular en la ruta (recursivo): FURIPS*/FURTRAN*/FUCTAS*/FURCEN*."""
    if ruta.is_file():
        return [ruta]
    encontrados = [
        p
        for p in sorted(ruta.rglob("*"))
        if p.is_file()
        and p.suffix.lower() in (".txt", ".csv")
        and detectar_tipo(p.stem) is not None
    ]
    return encontrados


def escribir_json(hallazgos: list[Hallazgo], destino: Path) -> None:
    resumen = {
        "generado": datetime.now().isoformat(timespec="seconds"),
        "norma": "Circular 022 de 2023 - ADRES",
        "total_errores": sum(1 for h in hallazgos if h.severidad == ERROR),
        "total_advertencias": sum(1 for h in hallazgos if h.severidad == ADVERTENCIA),
        "hallazgos": [asdict(h) for h in hallazgos],
    }
    destino.write_text(json.dumps(resumen, ensure_ascii=False, indent=2), encoding="utf-8")


def escribir_csv(hallazgos: list[Hallazgo], destino: Path) -> None:
    with destino.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(
            [
                "NOMBRE DEL ARCHIVO",
                "LINEA",
                "CAMPO QUE FALLA",
                "VALOR",
                "DESCRIPCION DEL ERROR",
                "SEVERIDAD",
            ]
        )
        for h in hallazgos:
            w.writerow([h.archivo, h.linea, h.campo, h.valor, h.descripcion, h.severidad])


def escribir_xlsx(hallazgos: list[Hallazgo], archivos_meta: list[dict], destino: Path) -> None:
    """Informe Excel detallado: RESUMEN, HALLAZGOS, POR CAMPO, AVISOS, LEYENDA."""
    from collections import Counter

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    AZUL = "1F4E79"
    FILL_HEADER = PatternFill("solid", fgColor=AZUL)
    FONT_HEADER = Font(color="FFFFFF", bold=True, size=10)
    FILLS = {
        ERROR: PatternFill("solid", fgColor="FFC7CE"),
        ADVERTENCIA: PatternFill("solid", fgColor="FFEB9C"),
        INFO: PatternFill("solid", fgColor="DDEBF7"),
        "OK": PatternFill("solid", fgColor="C6EFCE"),
    }
    FONTS = {
        ERROR: Font(color="9C0006", bold=True, size=10),
        ADVERTENCIA: Font(color="9C6500", size=10),
        INFO: Font(color="1F4E79", size=10),
        "OK": Font(color="006100", size=10),
    }
    BORDE = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

    wb = Workbook()
    wb.remove(wb.active)

    def hoja(nombre, encabezados, anchos):
        ws = wb.create_sheet(nombre)
        for j, (enc, ancho) in enumerate(zip(encabezados, anchos, strict=False), 1):
            cel = ws.cell(row=1, column=j, value=enc)
            cel.fill, cel.font = FILL_HEADER, FONT_HEADER
            cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(j)].width = ancho
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 28
        return ws

    def fila(ws, valores, estado=None, col_estado=None):
        r = ws.max_row + 1
        for j, v in enumerate(valores, 1):
            cel = ws.cell(row=r, column=j, value=v)
            cel.border = BORDE
            cel.alignment = Alignment(vertical="top", wrap_text=True)
            if estado and (col_estado is None or j == col_estado):
                cel.fill, cel.font = FILLS[estado], FONTS[estado]

    errores = [h for h in hallazgos if h.severidad == ERROR]
    advertencias = [h for h in hallazgos if h.severidad == ADVERTENCIA]
    avisos = [h for h in hallazgos if h.severidad == INFO]

    # ── RESUMEN ──────────────────────────────────────────────────────────────
    ws = wb.create_sheet("RESUMEN")
    ws.cell(row=1, column=1, value="VALIDACIÓN DE ARCHIVOS PLANOS — CIRCULAR 022 DE 2023 ADRES")
    ws.cell(row=1, column=1).fill = FILL_HEADER
    ws.cell(row=1, column=1).font = Font(color="FFFFFF", bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    generales = [
        ("Generado", f"{datetime.now():%d/%m/%Y %H:%M}"),
        ("Archivos validados", len(archivos_meta)),
        ("Errores (corregir ANTES de radicar)", len(errores)),
        ("Advertencias (revisar)", len(advertencias)),
        ("Avisos informativos", len(avisos)),
    ]
    for i, (a, b) in enumerate(generales, 3):
        ws.cell(row=i, column=1, value=a).font = Font(bold=True, size=10)
        ws.cell(row=i, column=2, value=b)
    ws.column_dimensions["A"].width = 36

    fila_tabla = len(generales) + 4
    encabezados = ["ARCHIVO", "TIPO", "LÍNEAS", "ERRORES", "ADVERTENCIAS", "AVISOS", "ESTADO"]
    anchos = [36, 10, 9, 10, 14, 9, 14]
    for j, (enc, ancho) in enumerate(zip(encabezados, anchos, strict=False), 1):
        cel = ws.cell(row=fila_tabla, column=j, value=enc)
        cel.fill, cel.font = FILL_HEADER, FONT_HEADER
        ws.column_dimensions[get_column_letter(j)].width = ancho
    r = fila_tabla
    for meta in archivos_meta:
        propios = [h for h in hallazgos if h.archivo == meta["archivo"]]
        ne = sum(1 for h in propios if h.severidad == ERROR)
        na = sum(1 for h in propios if h.severidad == ADVERTENCIA)
        ni = sum(1 for h in propios if h.severidad == INFO)
        estado = "CON ERRORES" if ne else ("REVISAR" if na else "CUMPLE")
        clave = ERROR if ne else (ADVERTENCIA if na else "OK")
        r += 1
        valores = [meta["archivo"], meta["tipo"], meta["lineas"], ne, na, ni, estado]
        for j, v in enumerate(valores, 1):
            cel = ws.cell(row=r, column=j, value=v)
            cel.border = BORDE
            if j == 7:
                cel.fill, cel.font = FILLS[clave], FONTS[clave]

    r += 2
    ws.cell(
        row=r, column=1, value="LOS ERRORES MÁS REPETIDOS (para corregir en bloque)"
    ).font = Font(bold=True, size=11, color=AZUL)
    top = Counter(h.descripcion for h in errores).most_common(10)
    r += 1
    for j, enc in enumerate(("VECES", "DESCRIPCIÓN DEL ERROR"), 1):
        cel = ws.cell(row=r, column=j, value=enc)
        cel.fill, cel.font = FILL_HEADER, FONT_HEADER
    for desc, veces in top:
        r += 1
        ws.cell(row=r, column=1, value=veces).border = BORDE
        cel = ws.cell(row=r, column=2, value=desc)
        cel.border = BORDE
        cel.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)

    # ── HALLAZGOS (errores y advertencias, para trabajar) ────────────────────
    ws = hoja(
        "HALLAZGOS",
        ["ARCHIVO", "LÍNEA", "CAMPO QUE FALLA", "VALOR", "DESCRIPCIÓN DEL ERROR", "SEVERIDAD"],
        [30, 7, 30, 24, 70, 13],
    )
    for h in errores + advertencias:
        fila(
            ws,
            [h.archivo, h.linea, h.campo, h.valor, h.descripcion, h.severidad],
            estado=h.severidad,
            col_estado=6,
        )
    ws.auto_filter.ref = f"A1:F{ws.max_row}"

    # ── POR CAMPO (agrupado: qué campo corregir y dónde) ─────────────────────
    ws = hoja(
        "POR CAMPO",
        ["CAMPO QUE FALLA", "VECES", "ARCHIVO(S)", "LÍNEAS (primeras)", "ERROR TÍPICO"],
        [32, 8, 30, 24, 66],
    )
    grupos: dict[str, list[Hallazgo]] = {}
    for h in errores + advertencias:
        grupos.setdefault(h.campo, []).append(h)
    for campo, lista in sorted(grupos.items(), key=lambda kv: -len(kv[1])):
        archivos_grupo = sorted({h.archivo for h in lista})
        lineas = sorted({h.linea for h in lista})
        tipica = Counter(h.descripcion for h in lista).most_common(1)[0][0]
        fila(
            ws,
            [
                campo,
                len(lista),
                " | ".join(archivos_grupo),
                ", ".join(str(n) for n in lineas[:12]) + ("…" if len(lineas) > 12 else ""),
                tipica,
            ],
        )
    ws.auto_filter.ref = f"A1:E{ws.max_row}"

    # ── AVISOS (INFO: no bloquean, contexto) ─────────────────────────────────
    ws = hoja(
        "AVISOS",
        ["ARCHIVO", "LÍNEA", "CAMPO", "AVISO (no bloquea la radicación)"],
        [30, 7, 30, 100],
    )
    for h in avisos:
        fila(ws, [h.archivo, h.linea, h.campo, h.descripcion], estado=INFO, col_estado=4)
    ws.auto_filter.ref = f"A1:D{ws.max_row}"

    # ── LEYENDA ──────────────────────────────────────────────────────────────
    ws = hoja("LEYENDA", ["TEMA", "EXPLICACIÓN"], [30, 110])
    leyenda = [
        ("ERROR", "Incumple la Circular 022/2023: corregirlo ANTES de radicar ante la ADRES."),
        (
            "ADVERTENCIA",
            "No bloquea, pero conviene revisarlo (p. ej. extensión .csv en vez de .txt).",
        ),
        (
            "AVISO (hoja AVISOS)",
            "Información: campos con obligatoriedad condicional que evalúa el "
            "validador profundo (validar_furips.py), o archivos omitidos.",
        ),
        (
            "Línea descuadrada",
            "Cuando una línea tiene campos de MÁS o de MENOS (casi siempre por una "
            "coma dentro de un texto libre o una coma sobrante), los demás errores de ESA línea "
            "suelen ser consecuencia del corrimiento: corrija primero la coma y vuelva a validar.",
        ),
        (
            "Reglas generales",
            "Separador: solo comas (vacíos como ,,). Sin comillas dobles ni "
            "caracteres de control. Fechas DD/MM/AAAA reales. Números sin puntos ni comas. Sin "
            "relleno de espacios; sin ceros a la izquierda en consecutivos, cantidades y valores. "
            "Las longitudes del anexo son MÁXIMOS.",
        ),
        (
            "Nomenclatura",
            "FURIPS1/FURIPS2/FURTRAN + habilitación (12 dígitos) + DDMMAAAA; "
            "FUCTAS2 + aseguradora (6) + MMAAAA; FURCEN + código del evento + DDMMAAAA + HHMM.",
        ),
        (
            "Mallas cargadas",
            "FURIPS 1 (102 campos) y FURIPS 2 (9). FURTRAN/FUCTAS/FURCEN: por "
            "ahora nomenclatura + reglas generales (falta cargar su anexo técnico).",
        ),
    ]
    for tema, exp in leyenda:
        fila(ws, [tema, exp])

    wb.save(destino)


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--ruta", type=Path, required=True, help="Archivo o carpeta a validar")
    parser.add_argument(
        "--salida", type=Path, default=None, help="Base del reporte (sin extensión)"
    )
    parser.add_argument(
        "--formato",
        choices=("xlsx", "json", "csv", "todos"),
        default="xlsx",
        help="Formato del reporte (por defecto Excel detallado)",
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO, format="%(message)s", handlers=[logging.StreamHandler(sys.stdout)]
    )

    if not args.ruta.exists():
        logger.error(f"No existe la ruta: {args.ruta}")
        return 2

    archivos = descubrir_archivos(args.ruta)
    logger.info("=" * 70)
    logger.info("VALIDADOR DE ARCHIVOS PLANOS — Circular 022/2023 ADRES")
    logger.info("=" * 70)
    if not archivos:
        logger.error(
            "No se encontraron archivos FURIPS1/FURIPS2/FURTRAN/FUCTAS/FURCEN "
            f"(.txt) en: {args.ruta}"
        )
        return 2

    hallazgos: list[Hallazgo] = []
    archivos_meta: list[dict] = []
    for ruta in archivos:
        propios = validar_archivo(ruta)
        hallazgos.extend(propios)
        try:
            lineas = sum(1 for ln in _leer_texto(ruta).splitlines() if ln.strip())
        except OSError:
            lineas = 0
        archivos_meta.append(
            {"archivo": ruta.name, "tipo": detectar_tipo(ruta.stem) or "?", "lineas": lineas}
        )
        ne = sum(1 for h in propios if h.severidad == ERROR)
        na = sum(1 for h in propios if h.severidad == ADVERTENCIA)
        estado = "SIN ERRORES" if ne == 0 else f"{ne} error(es)"
        logger.info(f"  {ruta.name}: {estado}, {na} advertencia(s)")

    carpeta = args.ruta if args.ruta.is_dir() else args.ruta.parent
    base = args.salida or carpeta / f"REPORTE_PLANOS_ADRES_{datetime.now():%Y%m%d_%H%M%S}"
    formato = args.formato
    if formato in ("xlsx", "todos"):
        try:
            escribir_xlsx(hallazgos, archivos_meta, base.with_suffix(".xlsx"))
            logger.info(f"\nREPORTE EXCEL: {base.with_suffix('.xlsx')}")
        except ImportError:
            logger.warning(
                "openpyxl no está instalado: se entrega el reporte en CSV. "
                "Para el Excel detallado: py -m pip install openpyxl"
            )
            formato = "csv" if formato == "xlsx" else formato
    if formato in ("json", "todos"):
        escribir_json(hallazgos, base.with_suffix(".json"))
        logger.info(f"REPORTE JSON:  {base.with_suffix('.json')}")
    if formato in ("csv", "todos"):
        escribir_csv(hallazgos, base.with_suffix(".csv"))
        logger.info(f"REPORTE CSV:   {base.with_suffix('.csv')}")

    errores = sum(1 for h in hallazgos if h.severidad == ERROR)
    logger.info("=" * 70)
    logger.info(
        f"RESULTADO: {len(archivos)} archivo(s) — {errores} error(es), "
        f"{sum(1 for h in hallazgos if h.severidad == ADVERTENCIA)} advertencia(s)"
    )
    logger.info("=" * 70)
    return 1 if errores else 0


if __name__ == "__main__":
    raise SystemExit(main())
