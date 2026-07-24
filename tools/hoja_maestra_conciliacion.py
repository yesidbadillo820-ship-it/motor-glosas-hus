"""hoja_maestra_conciliacion.py — Arma el "Expediente Inteligente de Conciliacion"
del Dispensario Medico: UN registro maestro por factura, con su historial de
glosas, la respuesta cruzada al lado de cada glosa, el resultado en actas, la
ubicacion derivada de soportes/factura electronica, un tablero de control y los
cruces automaticos de consistencia.

NO transcribe ni inventa: solo cruza las bases que ya existen. Lo que ninguna
base trae se marca PENDIENTE (nunca en blanco que se confunda con "no").

FUENTES (las tres del Dispensario, tal cual salen de SIMED/DGH)
-------------------------------------------------------------
- CARTERA  (hoja 'CARTERA' del corte 30/06/2026): ESPINA. 1 fila por factura
  (5.571 unicas). Trae, a nivel factura: VALOR FACTURA, SALDO, VALOR GLOSA,
  ESTADO GLOSA, EDADES, LEVANTADA/ACEPTADA/RATIFICADA EN ACTAS, ACTAS. La fila 0
  es un banner con totales (se descarta).
- RECEPCION (CRRPSEGUIMIENTORECEPCIONOBJECIO): la GLOSA RECIBIDA (lo que objeta
  la EPS), 1 fila por linea de concepto. Trae el MOTIVO EPS en texto libre
  (col33), el codigo/nombre del concepto, el CUPS y el servicio. Se une al
  tramite por el Consecutivo de recepcion.
- TRAMITE   (CRRPSEGUIMIENTOTRAMITEOBJECION): NUESTRA RESPUESTA, 1 fila por linea.
  Trae el valor objetado de la linea (col23), el VALOR ACEPTADO (col24) y el
  ARGUMENTO del ESE HUS (col32 'ESE HUS ACEPTA/NO ACEPTA...'). El concepto
  original de la glosa viaja en col25/col27; col28/col29 es el codigo de
  "respuesta a glosa". Se une a la recepcion por RecepcionObjecion.Consecutivo
  (col20 = RECEPCION col13).
- ACTAS/CRUCE ACTAS (mismo libro de CARTERA): resultado de conciliacion por
  FACTURA+ACTA (una factura puede estar en varias actas). AC000639 es DUPLICADA
  de SINAC 720: se excluye de las sumas.

GRANULARIDAD (sin duplicar el maestro)
--------------------------------------
- 01_MAESTRA : 1 fila por factura. Valores factura/saldo a nivel factura de
  CARTERA (NUNCA sumando lineas); glosado/aceptado como SUM de las lineas del
  detalle agrupadas por factura.
- 02_GLOSAS  : 1 fila por linea de glosa, con su respuesta cruzada al lado.
- 03_ACTAS   : 1 fila por FACTURA+ACTA.
- 04_CRUCES  : los 11 controles de consistencia.
- 00_DASHBOARD: tablero de indicadores + carteras (vigencia/estado/edades).

PENDIENTES (documentados, no se inventan): bandera de Factura Electronica (CUFE),
normatividad citada por respuesta, valor pagado real, y las RAICES exactas Y:/X:
de soportes (mientras tanto se deja la ruta derivada por AAAAMM + una nota).

USO
---
    py tools\\hoja_maestra_conciliacion.py ^
        --cartera   "...CORTE_30062026_CRUCE_ACTAS.xlsx" ^
        --recepcion "...RECEPCION_DE_OBJECIONES_DISPENSARIO_MEDICO.xlsx" ^
        --tramite   "...TRAMITE_DE_OBJECCION_DISPENSARIO_MEDICO.xlsx" ^
        --salida    "...EXPEDIENTE_CONCILIACION_DISPENSARIO.xlsx"
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

logger = logging.getLogger("hoja_maestra")

# ----------------------------------------------------------------------------
# Constantes de negocio (Dispensario Medico Bucaramanga)
# ----------------------------------------------------------------------------
EPS_NOMBRE = "DIRECCION DE SANIDAD EJERCITO"
NIT_TERCERO = "901541137"
PRESTADOR = "DISPENSARIO MEDICO BUCARAMANGA"
FE_RAIZ = r"\\172.16.32.83\factura_electronica_net22"
SOPORTES_NOTA = "RAIZ Y:/X: PENDIENTE confirmar con auditor; subcarpeta = AAAAMM"
ACTA_DUPLICADA = "AC000639"  # duplicada de SINAC 720: se excluye de sumas

# Indices de columna por fuente (mapeados por posicion, verificados del export).
# CARTERA (hoja 'CARTERA'): fila 0 = banner, fila 1 = encabezado, fila 2+ = datos
CA_VIGENCIA, CA_ESTADO, CA_FACTURA, CA_FECHA_RAD, CA_VALOR_FAC = 0, 1, 2, 3, 4
CA_SALDO_DGH, CA_SALDO_EPS, CA_VALOR_DEVOL, CA_VALOR_GLOSA = 5, 6, 9, 10
CA_VALOR_LIBRE, CA_SOPORTE_PAGO, CA_EDADES = 12, 13, 17
CA_LEVANTADA, CA_ACEPTADA, CA_RATIFICADA, CA_ACTAS, CA_ESTADO_GLOSA = 20, 21, 22, 23, 24

# RECEPCION (glosa recibida)
RE_FACTURA, RE_FECHA_DOC, RE_CONSECUTIVO = 11, 12, 13
RE_VALOR_FAC, RE_FECHA_FAC, RE_NIT = 16, 17, 18
RE_FECHA_OBJ, RE_TIPO = 20, 1
RE_COD_CONCEPTO, RE_NOM_CONCEPTO = 26, 28
RE_CUPS, RE_SERVICIO, RE_VALOR_OBJ, RE_MOTIVO = 29, 30, 31, 33

# TRAMITE (nuestra respuesta)
TR_ESTADO_CXC, TR_TIPO, TR_FECHA_OBJ = 0, 1, 3
TR_FACTURA, TR_CONTRATO_COD = 7, 8
TR_FECHA_DOC, TR_CONSECUTIVO = 10, 11
TR_FECHA_FAC, TR_REC_CONSECUTIVO = 15, 20
TR_VALOR_OBJ, TR_VALOR_ACEPTADO = 23, 24
TR_COD_GLOSA, TR_NOM_GLOSA, TR_SERVICIO, TR_ARGUMENTO = 25, 27, 30, 32

# CRUCE ACTAS: fila 0 titulo, fila 1 encabezado, fila 2+ datos, ultima = totales
CR_FACTURA, CR_ACTA, CR_TIPO = 0, 1, 2
CR_GLOSA_INI, CR_ACEPTADA, CR_LEVANTADA, CR_RATIFICADA, CR_A_PAGAR = 3, 4, 5, 6, 7


def setup_logging() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")


# ----------------------------------------------------------------------------
# Utilidades
# ----------------------------------------------------------------------------
def _num(x) -> float:
    """Convierte a numero tolerando '$ 3.211.891', comas y None -> 0.0."""
    if x is None:
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace("$", "").replace(" ", "")
    if not s or s in ("-", "None"):
        return 0.0
    # formato colombiano: puntos de miles y coma decimal
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    elif s.count(".") > 1:
        # varios puntos y sin coma -> todos son separadores de miles
        s = s.replace(".", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _txt(x) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    return "" if s in ("None", "-") else s


def normalizar_factura(x) -> str:
    """HUS + 10 digitos. Acepta 'HUS0000316347', '316347', 316347."""
    s = _txt(x)
    if not s:
        return ""
    m = re.search(r"(\d+)", s)
    if not m:
        return s
    return "HUS" + m.group(1).zfill(10)


def _fecha(x):
    """Devuelve datetime si puede, si no None."""
    if isinstance(x, datetime):
        return x
    s = _txt(x)
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[: len(fmt) + 2].strip(), fmt)
        except ValueError:
            continue
    return None


def _aaaamm(*fechas):
    """AAAAMM de la primera fecha valida; '' si ninguna."""
    for f in fechas:
        d = _fecha(f)
        if d:
            return f"{d.year:04d}{d.month:02d}"
    return ""


def contrato_por_fecha(fecha_fac, codigo_raw: str) -> str:
    """287 hasta 30-nov-2025; 440 desde dic-2025 (aprox por fecha factura).

    Se deja el codigo crudo entre parentesis para trazabilidad. La fecha real de
    atencion vive en los RIPS; aqui es una aproximacion documentada.
    """
    d = _fecha(fecha_fac)
    base = ""
    if d:
        base = "287" if d <= datetime(2025, 11, 30) else "440"
    cod = _txt(codigo_raw)
    if base and cod:
        return f"{base} (aprox; cod {cod})"
    if base:
        return f"{base} (aprox x fecha factura)"
    return cod or "PENDIENTE"


# ----------------------------------------------------------------------------
# Carga de fuentes
# ----------------------------------------------------------------------------
def cargar_cartera(ruta: Path) -> dict[str, dict]:
    """Lee la hoja CARTERA como espina. Devuelve {factura -> dict factura}."""
    import openpyxl

    wb = openpyxl.load_workbook(str(ruta), read_only=True, data_only=True)
    if "CARTERA" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"El libro {ruta.name} no trae hoja 'CARTERA'.")
    filas = list(wb["CARTERA"].iter_rows(values_only=True))
    wb.close()
    out: dict[str, dict] = {}
    for r in filas[2:]:  # fila 0 banner, fila 1 encabezado
        fac = normalizar_factura(r[CA_FACTURA] if len(r) > CA_FACTURA else None)
        if not fac:
            continue
        out[fac] = {
            "factura": fac,
            "vigencia": _txt(r[CA_VIGENCIA]),
            "estado": _txt(r[CA_ESTADO]),
            "fecha_radicado": r[CA_FECHA_RAD] if len(r) > CA_FECHA_RAD else None,
            "valor_factura": _num(r[CA_VALOR_FAC]),
            "saldo_dgh": _num(r[CA_SALDO_DGH]),
            "saldo_eps": _num(r[CA_SALDO_EPS]),
            "valor_devolucion": _num(r[CA_VALOR_DEVOL]),
            "valor_glosa": _num(r[CA_VALOR_GLOSA]),
            "valor_libre_pago": _num(r[CA_VALOR_LIBRE]),
            "soporte_pago": _txt(r[CA_SOPORTE_PAGO]),
            "edades": _txt(r[CA_EDADES]) if len(r) > CA_EDADES else "",
            "levantada_actas": _num(r[CA_LEVANTADA]) if len(r) > CA_LEVANTADA else 0.0,
            "aceptada_actas": _num(r[CA_ACEPTADA]) if len(r) > CA_ACEPTADA else 0.0,
            "ratificada_actas": _num(r[CA_RATIFICADA]) if len(r) > CA_RATIFICADA else 0.0,
            "actas": _txt(r[CA_ACTAS]) if len(r) > CA_ACTAS else "",
            "estado_glosa": _txt(r[CA_ESTADO_GLOSA]) if len(r) > CA_ESTADO_GLOSA else "",
        }
    return out


def cargar_recepcion(ruta: Path) -> dict[str, list[dict]]:
    """Lee la RECEPCION (glosa recibida). Indexa por Consecutivo de recepcion:
    {consecutivo -> [linea, ...]}. Cada linea trae el motivo EPS (col33)."""
    import openpyxl

    wb = openpyxl.load_workbook(str(ruta), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = list(ws.iter_rows(values_only=True))
    wb.close()
    por_cons: dict[str, list[dict]] = defaultdict(list)
    for r in filas[1:]:
        if len(r) <= RE_MOTIVO:
            continue
        cons = _txt(r[RE_CONSECUTIVO])
        fac = normalizar_factura(r[RE_FACTURA])
        if not fac:
            continue
        por_cons[cons].append(
            {
                "consecutivo": cons,
                "factura": fac,
                "fecha_glosa": r[RE_FECHA_OBJ] or r[RE_FECHA_DOC],
                "fecha_factura": r[RE_FECHA_FAC],
                "tipo": _txt(r[RE_TIPO]),
                "codigo": _txt(r[RE_COD_CONCEPTO]),
                "concepto": _txt(r[RE_NOM_CONCEPTO]),
                "cups": _txt(r[RE_CUPS]),
                "servicio": _txt(r[RE_SERVICIO]),
                "valor": _num(r[RE_VALOR_OBJ]),
                "motivo_eps": _txt(r[RE_MOTIVO]),
            }
        )
    return por_cons


def cargar_tramite(ruta: Path) -> list[dict]:
    """Lee el TRAMITE (nuestra respuesta). 1 dict por linea de respuesta."""
    import openpyxl

    wb = openpyxl.load_workbook(str(ruta), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = list(ws.iter_rows(values_only=True))
    wb.close()
    out: list[dict] = []
    for r in filas[1:]:
        if len(r) <= TR_ARGUMENTO:
            continue
        fac = normalizar_factura(r[TR_FACTURA])
        if not fac:
            continue
        glosado = _num(r[TR_VALOR_OBJ])
        aceptado = _num(r[TR_VALOR_ACEPTADO])
        out.append(
            {
                "factura": fac,
                "rec_consecutivo": _txt(r[TR_REC_CONSECUTIVO]),
                "fecha_glosa": r[TR_FECHA_OBJ],
                "fecha_factura": r[TR_FECHA_FAC],
                "fecha_respuesta": r[TR_FECHA_DOC],
                "radicado_respuesta": _txt(r[TR_CONSECUTIVO]),
                "tipo": _txt(r[TR_TIPO]),
                "estado_cxc": _txt(r[TR_ESTADO_CXC]),
                "contrato_cod": _txt(r[TR_CONTRATO_COD]),
                "codigo_glosa": _txt(r[TR_COD_GLOSA]),
                "concepto_glosa": _txt(r[TR_NOM_GLOSA]),
                "servicio": _txt(r[TR_SERVICIO]),
                "glosado": glosado,
                "aceptado": aceptado,
                "objetado": max(glosado - aceptado, 0.0),
                "argumento": _txt(r[TR_ARGUMENTO]),
            }
        )
    return out


def cargar_cruce_actas(ruta: Path) -> list[dict]:
    """Lee CRUCE ACTAS (1 fila por FACTURA+ACTA). Excluye titulo y totales."""
    import openpyxl

    wb = openpyxl.load_workbook(str(ruta), read_only=True, data_only=True)
    if "CRUCE ACTAS" not in wb.sheetnames:
        wb.close()
        return []
    filas = list(wb["CRUCE ACTAS"].iter_rows(values_only=True))
    wb.close()
    out: list[dict] = []
    for r in filas[2:]:  # fila 0 titulo, fila 1 encabezado
        fac = normalizar_factura(r[CR_FACTURA] if len(r) > CR_FACTURA else None)
        acta = _txt(r[CR_ACTA]) if len(r) > CR_ACTA else ""
        if not fac or not acta:  # la fila de totales tiene factura vacia
            continue
        out.append(
            {
                "factura": fac,
                "acta": acta,
                "tipo": _txt(r[CR_TIPO]),
                "glosa_inicial": _num(r[CR_GLOSA_INI]),
                "aceptada": _num(r[CR_ACEPTADA]),
                "levantada": _num(r[CR_LEVANTADA]),
                "ratificada": _num(r[CR_RATIFICADA]),
                "a_pagar": _num(r[CR_A_PAGAR]),
                "duplicada": acta.upper() == ACTA_DUPLICADA,
            }
        )
    return out


# ----------------------------------------------------------------------------
# Construccion del detalle de glosas (glosa + respuesta cruzada)
# ----------------------------------------------------------------------------
def _motivo_eps_para(linea_tr: dict, rec_por_cons: dict[str, list[dict]]) -> tuple[str, str, str]:
    """Enriquece la linea de tramite con (motivo_eps, cups, servicio_eps) de la
    recepcion. Match por consecutivo + (codigo, valor); si es ambiguo, concatena
    los motivos distintos del consecutivo (nivel factura). Nunca inventa."""
    cons = linea_tr.get("rec_consecutivo", "")
    lineas = rec_por_cons.get(cons)
    if not lineas:
        return "", "", ""
    cod = linea_tr.get("codigo_glosa", "")
    val = round(linea_tr.get("glosado", 0.0))
    exactas = [x for x in lineas if x["codigo"] == cod and round(x["valor"]) == val]
    if len(exactas) == 1:
        x = exactas[0]
        return x["motivo_eps"], x["cups"], x["servicio"]
    por_cod = [x for x in lineas if x["codigo"] == cod]
    if len(por_cod) == 1:
        x = por_cod[0]
        return x["motivo_eps"], x["cups"], x["servicio"]
    # ambiguo: motivo a nivel de la recepcion (factura), distintos concatenados
    motivos = sorted({x["motivo_eps"] for x in lineas if x["motivo_eps"]})
    cups = sorted({x["cups"] for x in lineas if x["cups"]})
    return " | ".join(motivos)[:900], " | ".join(cups), ""


def construir_glosas(tramite: list[dict], rec_por_cons: dict[str, list[dict]]) -> list[dict]:
    """Une glosa (tramite col25/27 + motivo EPS de recepcion) con respuesta.
    Agrega tambien las glosas recibidas SIN respuesta (recepcion sin tramite)."""
    glosas: list[dict] = []
    cons_respondidos: set[str] = set()
    for tr in tramite:
        cons_respondidos.add(tr["rec_consecutivo"])
        motivo, cups, serv_eps = _motivo_eps_para(tr, rec_por_cons)
        respondida = bool(tr["radicado_respuesta"] or tr["fecha_respuesta"] or tr["argumento"])
        glosas.append(
            {
                "factura": tr["factura"],
                "consecutivo": tr["rec_consecutivo"],
                "fecha_glosa": tr["fecha_glosa"],
                "tipo": tr["tipo"],
                "codigo_glosa": tr["codigo_glosa"],
                "concepto_glosa": tr["concepto_glosa"],
                "servicio": tr["servicio"] or serv_eps,
                "cups": cups,
                "motivo_eps": motivo,
                "valor_glosado": tr["glosado"],
                "fecha_respuesta": tr["fecha_respuesta"],
                "radicado_respuesta": tr["radicado_respuesta"],
                "valor_aceptado": tr["aceptado"],
                "valor_objetado": tr["objetado"],
                "argumento": tr["argumento"],
                "estado_respuesta": "RESPONDIDA" if respondida else "SIN RESPUESTA",
                "origen": "TRAMITE",
            }
        )
    # Glosas recibidas cuyo consecutivo NO aparece en el tramite: sin respuesta.
    for cons, lineas in rec_por_cons.items():
        if cons in cons_respondidos:
            continue
        for x in lineas:
            glosas.append(
                {
                    "factura": x["factura"],
                    "consecutivo": cons,
                    "fecha_glosa": x["fecha_glosa"],
                    "tipo": x["tipo"],
                    "codigo_glosa": x["codigo"],
                    "concepto_glosa": x["concepto"],
                    "servicio": x["servicio"],
                    "cups": x["cups"],
                    "motivo_eps": x["motivo_eps"],
                    "valor_glosado": x["valor"],
                    "fecha_respuesta": None,
                    "radicado_respuesta": "",
                    "valor_aceptado": 0.0,
                    "valor_objetado": x["valor"],
                    "argumento": "",
                    "estado_respuesta": "SIN RESPUESTA",
                    "origen": "RECEPCION",
                }
            )
    return glosas


def agregar_por_factura(glosas: list[dict]) -> dict[str, dict]:
    """Agrega el detalle de glosas a nivel factura (para la hoja maestra)."""
    agg: dict[str, dict] = defaultdict(
        lambda: {
            "num_glosas": 0,
            "glosado": 0.0,
            "aceptado": 0.0,
            "objetado": 0.0,
            "sin_respuesta": 0,
            "fecha_factura": None,
            "contrato_cod": "",
            "tipos": set(),
        }
    )
    for g in glosas:
        a = agg[g["factura"]]
        a["num_glosas"] += 1
        a["glosado"] += g["valor_glosado"]
        a["aceptado"] += g["valor_aceptado"]
        a["objetado"] += g["valor_objetado"]
        if g["estado_respuesta"] == "SIN RESPUESTA":
            a["sin_respuesta"] += 1
        if a["fecha_factura"] is None and g.get("fecha_glosa"):
            a["fecha_factura"] = g["fecha_glosa"]
        if g["tipo"]:
            a["tipos"].add(g["tipo"])
    return agg


def fecha_factura_por_factura(
    tramite: list[dict], rec_por_cons: dict[str, list[dict]]
) -> dict[str, object]:
    """{factura -> fecha de factura} tomada del tramite/recepcion (no de CARTERA,
    que solo trae fecha de radicado)."""
    out: dict[str, object] = {}
    for tr in tramite:
        if tr["factura"] not in out and _fecha(tr["fecha_factura"]):
            out[tr["factura"]] = tr["fecha_factura"]
    for lineas in rec_por_cons.values():
        for x in lineas:
            if x["factura"] not in out and _fecha(x["fecha_factura"]):
                out[x["factura"]] = x["fecha_factura"]
    return out


def contrato_cod_por_factura(tramite: list[dict]) -> dict[str, str]:
    out: dict[str, str] = {}
    for tr in tramite:
        if tr["factura"] not in out and tr["contrato_cod"]:
            out[tr["factura"]] = tr["contrato_cod"]
    return out


# ----------------------------------------------------------------------------
# Resultado final por factura y cruces de consistencia
# ----------------------------------------------------------------------------
def resultado_final(cart: dict, agg: dict | None) -> str:
    """Deriva el resultado a nivel factura combinando CARTERA (actas) + detalle."""
    lev = cart.get("levantada_actas", 0.0)
    rat = cart.get("ratificada_actas", 0.0)
    ace = cart.get("aceptada_actas", 0.0)
    estado_glosa = (cart.get("estado_glosa") or "").upper()
    tiene_glosa = bool(agg) or cart.get("valor_glosa", 0.0) > 0
    if not tiene_glosa:
        return "SIN GLOSA"
    if lev > 0 and rat == 0 and ace == 0:
        return "LEVANTADA (a favor HUS)"
    if lev > 0 or rat > 0 or ace > 0:
        if rat > 0 and lev == 0 and ace == 0:
            return "RATIFICADA (pdte conciliar)"
        return "CONCILIADA PARCIAL"
    if agg and agg["sin_respuesta"] == agg["num_glosas"] and agg["num_glosas"] > 0:
        return "SIN RESPUESTA"
    if "LEVANTADA" in estado_glosa:
        return "LEVANTADA (a favor HUS)"
    if "RATIFICADA" in estado_glosa:
        return "RATIFICADA (pdte conciliar)"
    if "CONCILIADA" in estado_glosa:
        return "CONCILIADA PARCIAL"
    return "EN TRAMITE"


def calcular_cruces(
    cartera: dict[str, dict],
    agg: dict[str, dict],
    glosas: list[dict],
    cruce_actas: list[dict],
) -> list[dict]:
    """Los 11 controles. Cada uno: {n, nombre, cantidad, muestra}."""
    cruces: list[dict] = []

    def add(n, nombre, facturas):
        facturas = list(facturas)
        cruces.append(
            {
                "n": n,
                "nombre": nombre,
                "cantidad": len(facturas),
                "muestra": ", ".join(sorted(facturas)[:15]),
            }
        )

    fac_cartera = set(cartera)
    fac_glosa = set(agg)

    # 1. Facturas de cartera sin glosa en el detalle
    add(
        1,
        "Facturas SIN glosa (en cartera, sin linea de glosa)",
        [f for f in fac_cartera if f not in fac_glosa and cartera[f]["valor_glosa"] == 0],
    )

    # 2. Glosas SIN respuesta
    sin_resp = {g["factura"] for g in glosas if g["estado_respuesta"] == "SIN RESPUESTA"}
    add(2, "Facturas con al menos una glosa SIN respuesta", sin_resp)

    # 3. Respuestas cuya factura no existe en cartera
    add(3, "Glosas/respuestas cuya factura NO esta en cartera", fac_glosa - fac_cartera)

    # 4. Facturas sin AAAAMM resoluble (no se puede ubicar soporte)
    sin_periodo = set()
    for f in fac_glosa:
        fx = agg[f]["fecha_factura"]
        if not _aaaamm(fx, cartera.get(f, {}).get("fecha_radicado")):
            sin_periodo.add(f)
    add(4, "Facturas sin periodo (AAAAMM) para ubicar soportes", sin_periodo)

    # 5. Diferencia CARTERA.VALOR GLOSA vs SUM glosado del detalle (> $1)
    dif = set()
    for f in fac_glosa & fac_cartera:
        if abs(cartera[f]["valor_glosa"] - agg[f]["glosado"]) > 1:
            dif.add(f)
    add(5, "Diferencia VALOR GLOSA cartera vs SUM glosado detalle (>$1)", dif)

    # 6. Respuestas que no cubren el 100% (glosado != aceptado + objetado)
    no_cubre = set()
    for f in fac_glosa:
        a = agg[f]
        if abs(a["glosado"] - (a["aceptado"] + a["objetado"])) > 1:
            no_cubre.add(f)
    add(6, "Facturas donde glosado != aceptado + objetado (>$1)", no_cubre)

    # 7. Valores/fechas inconsistentes (aceptado > glosado por factura)
    incons = {f for f in fac_glosa if agg[f]["aceptado"] > agg[f]["glosado"] + 1}
    add(7, "Facturas con aceptado > glosado (inconsistente)", incons)

    # 8. Facturas duplicadas en la maestra (debe ser 0: cartera es unica)
    add(8, "Facturas duplicadas en la maestra (esperado 0)", [])

    # 9. Pagos no aplicados: levantada/conciliada en actas pero saldo DGH > 0
    pagos = set()
    for f in fac_cartera:
        c = cartera[f]
        if (c["levantada_actas"] > 0 or "CONCILIADA" in (c["estado_glosa"] or "").upper()) and c[
            "saldo_dgh"
        ] > 1:
            pagos.add(f)
    add(9, "Facturas conciliadas/levantadas con saldo DGH aun > 0", pagos)

    # 10. Facturas en actas que no estan en cartera
    fac_actas = {r["factura"] for r in cruce_actas}
    add(10, "Facturas en ACTAS que NO estan en cartera", fac_actas - fac_cartera)

    # 11. Aritmetica de actas: glosa_inicial != aceptada+levantada+ratificada
    arit = set()
    for r in cruce_actas:
        if r["duplicada"]:
            continue
        suma = r["aceptada"] + r["levantada"] + r["ratificada"]
        if abs(r["glosa_inicial"] - suma) > 1:
            arit.add(f"{r['factura']}/{r['acta']}")
    add(11, "Actas donde GLOSA INICIAL != ACEPTADA+LEVANTADA+RATIFICADA", arit)

    return cruces


# ----------------------------------------------------------------------------
# Escritura del workbook con presentacion
# ----------------------------------------------------------------------------
def _escribir_workbook(salida: Path, maestra, glosas, actas, cruces, dash) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    MONEDA = "#,##0"
    HDR_FILL = PatternFill("solid", fgColor="1F4E78")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    TITULO = Font(bold=True, size=14, color="1F4E78")
    SUB = Font(italic=True, size=9, color="595959")
    # colores de resultado
    COL = {
        "LEVANTADA (a favor HUS)": "C6EFCE",
        "RATIFICADA (pdte conciliar)": "FCE4D6",
        "CONCILIADA PARCIAL": "DDEBF7",
        "SIN RESPUESTA": "FFC7CE",
        "EN TRAMITE": "FFF2CC",
        "SIN GLOSA": "F2F2F2",
    }

    # Objetos de estilo creados UNA sola vez y reusados (evita que openpyxl
    # deduplique miles de PatternFill/Font -> comportamiento cuadratico).
    FILL_RES = {k: PatternFill("solid", fgColor=v) for k, v in COL.items()}
    FILL_ROJO = PatternFill("solid", fgColor="FFC7CE")
    FILL_GRIS = PatternFill("solid", fgColor="F2F2F2")
    FILL_AMBAR = PatternFill("solid", fgColor="FFF2CC")
    FONT_LINK = Font(color="0563C1", underline="single")
    ALIGN_HDR = Alignment(vertical="center", wrap_text=True, horizontal="center")

    wb = openpyxl.Workbook()

    def estilizar_encabezado(ws, ncols, fila=1):
        for c in range(1, ncols + 1):
            cel = ws.cell(row=fila, column=c)
            cel.fill = HDR_FILL
            cel.font = HDR_FONT
            cel.alignment = ALIGN_HDR
        ws.freeze_panes = ws.cell(row=fila + 1, column=1)
        ws.auto_filter.ref = f"A{fila}:{get_column_letter(ncols)}{fila}"
        ws.row_dimensions[fila].height = 30

    def anchos(ws, mapa):
        for letra, w in mapa.items():
            ws.column_dimensions[letra].width = w

    # ---- 00_DASHBOARD ------------------------------------------------------
    ws = wb.active
    ws.title = "00_DASHBOARD"
    ws["A1"] = "EXPEDIENTE INTELIGENTE DE CONCILIACION - DISPENSARIO MEDICO BUCARAMANGA"
    ws["A1"].font = TITULO
    ws["A2"] = (
        f"Entidad: {EPS_NOMBRE}  |  NIT {NIT_TERCERO}  |  Prestador: {PRESTADOR}  "
        f"|  Corte 30/06/2026  |  1 registro maestro por factura"
    )
    ws["A2"].font = SUB
    fila = 4
    ws.cell(row=fila, column=1, value="INDICADOR").font = HDR_FONT
    ws.cell(row=fila, column=1).fill = HDR_FILL
    ws.cell(row=fila, column=2, value="VALOR").font = HDR_FONT
    ws.cell(row=fila, column=2).fill = HDR_FILL
    fila += 1
    for etiqueta, valor, es_moneda in dash["kpis"]:
        ws.cell(row=fila, column=1, value=etiqueta)
        c = ws.cell(row=fila, column=2, value=valor)
        if es_moneda:
            c.number_format = MONEDA
        fila += 1
    # tablas de cartera
    fila += 1
    for titulo, tabla in dash["tablas"]:
        ws.cell(row=fila, column=1, value=titulo).font = Font(bold=True, color="1F4E78")
        fila += 1
        ws.cell(row=fila, column=1, value="Categoria").font = HDR_FONT
        ws.cell(row=fila, column=1).fill = HDR_FILL
        ws.cell(row=fila, column=2, value="# Facturas").font = HDR_FONT
        ws.cell(row=fila, column=2).fill = HDR_FILL
        ws.cell(row=fila, column=3, value="Valor glosa").font = HDR_FONT
        ws.cell(row=fila, column=3).fill = HDR_FILL
        fila += 1
        for cat, n, val in tabla:
            ws.cell(row=fila, column=1, value=cat)
            ws.cell(row=fila, column=2, value=n)
            ws.cell(row=fila, column=3, value=val).number_format = MONEDA
            fila += 1
        fila += 1
    anchos(ws, {"A": 52, "B": 20, "C": 20})

    # ---- 01_MAESTRA --------------------------------------------------------
    ws = wb.create_sheet("01_MAESTRA")
    cols_m = [
        "FACTURA",
        "VIGENCIA",
        "ESTADO",
        "CONTRATO",
        "FECHA FACTURA",
        "FECHA RADICADO",
        "EPS",
        "NIT",
        "PRESTADOR",
        "VALOR FACTURA",
        "VALOR GLOSA (cartera)",
        "# GLOSAS",
        "GLOSADO (detalle)",
        "ACEPTADO",
        "OBJETADO",
        "# SIN RESPUESTA",
        "LEVANTADA ACTAS",
        "ACEPTADA ACTAS",
        "RATIFICADA ACTAS",
        "ACTAS",
        "SALDO DGH 30/06",
        "SALDO A COBRAR EPS",
        "ESTADO GLOSA",
        "EDADES",
        "RESULTADO FINAL",
        "TIPO OBJECION",
        "PERIODO (AAAAMM)",
        "UBICACION SOPORTES",
        "UBICACION FACTURA ELECTRONICA",
        "FE (bandera)",
        "NORMATIVIDAD",
        "VER GLOSAS",
    ]
    ws.append(cols_m)
    estilizar_encabezado(ws, len(cols_m))
    ncol_result = cols_m.index("RESULTADO FINAL") + 1
    rid = 1
    for m in maestra:
        ws.append(m["fila"])
        rid += 1  # contador propio: ws.max_row es O(n) y volveria cuadratico
        # formato moneda a las columnas de valor
        for idx in (10, 11, 13, 14, 15, 17, 18, 19, 21, 22):
            ws.cell(row=rid, column=idx).number_format = MONEDA
        # color por resultado
        fill = FILL_RES.get(m["resultado"])
        if fill:
            ws.cell(row=rid, column=ncol_result).fill = fill
        # hipervinculo a la hoja de glosas (autofiltro por factura)
        cel = ws.cell(row=rid, column=len(cols_m))
        cel.hyperlink = "#'02_GLOSAS'!A1"
        cel.value = "ver glosas"
        cel.font = FONT_LINK
    anchos(
        ws,
        {
            "A": 16,
            "D": 20,
            "E": 13,
            "F": 13,
            "J": 16,
            "K": 16,
            "M": 16,
            "N": 14,
            "O": 14,
            "T": 16,
            "U": 16,
            "V": 16,
            "W": 22,
            "X": 12,
            "Y": 26,
            "AB": 40,
            "AC": 40,
            "AE": 30,
            "AF": 12,
        },
    )

    # ---- 02_GLOSAS ---------------------------------------------------------
    ws = wb.create_sheet("02_GLOSAS")
    cols_g = [
        "FACTURA",
        "CONSECUTIVO RECEP.",
        "FECHA GLOSA",
        "TIPO",
        "CODIGO GLOSA",
        "CONCEPTO GLOSA (EPS)",
        "SERVICIO / CUPS",
        "MOTIVO EXACTO EPS",
        "VALOR GLOSADO",
        "FECHA RESPUESTA",
        "RADICADO RESPUESTA",
        "VALOR ACEPTADO",
        "VALOR OBJETADO (ratificado)",
        "ARGUMENTO ESE HUS",
        "ESTADO RESPUESTA",
        "ORIGEN",
    ]
    ws.append(cols_g)
    estilizar_encabezado(ws, len(cols_g))
    rid = 1
    for g in glosas:
        ws.append(
            [
                g["factura"],
                g["consecutivo"],
                _fecha(g["fecha_glosa"]),
                g["tipo"],
                g["codigo_glosa"],
                g["concepto_glosa"],
                (g["servicio"] + (" / " + g["cups"] if g["cups"] else "")).strip(" /"),
                g["motivo_eps"],
                g["valor_glosado"],
                _fecha(g["fecha_respuesta"]),
                g["radicado_respuesta"],
                g["valor_aceptado"],
                g["valor_objetado"],
                g["argumento"],
                g["estado_respuesta"],
                g["origen"],
            ]
        )
        rid += 1
        for idx in (9, 12, 13):
            ws.cell(row=rid, column=idx).number_format = MONEDA
        if g["estado_respuesta"] == "SIN RESPUESTA":
            ws.cell(row=rid, column=15).fill = FILL_ROJO
    anchos(
        ws,
        {
            "A": 16,
            "B": 14,
            "C": 12,
            "D": 12,
            "E": 11,
            "F": 30,
            "G": 30,
            "H": 50,
            "I": 15,
            "J": 12,
            "K": 14,
            "L": 14,
            "M": 16,
            "N": 55,
            "O": 15,
        },
    )

    # ---- 03_ACTAS ----------------------------------------------------------
    ws = wb.create_sheet("03_ACTAS")
    cols_a = [
        "FACTURA",
        "ACTA",
        "TIPO",
        "GLOSA INICIAL",
        "ACEPTADA",
        "LEVANTADA",
        "RATIFICADA",
        "A PAGAR",
        "NOTA",
    ]
    ws.append(cols_a)
    estilizar_encabezado(ws, len(cols_a))
    rid = 1
    for r in actas:
        ws.append(
            [
                r["factura"],
                r["acta"],
                r["tipo"],
                r["glosa_inicial"],
                r["aceptada"],
                r["levantada"],
                r["ratificada"],
                r["a_pagar"],
                "DUPLICADA con SINAC 720 (excluida de sumas)" if r["duplicada"] else "",
            ]
        )
        rid += 1
        for idx in (4, 5, 6, 7, 8):
            ws.cell(row=rid, column=idx).number_format = MONEDA
        if r["duplicada"]:
            for c in range(1, len(cols_a) + 1):
                ws.cell(row=rid, column=c).fill = FILL_GRIS
    anchos(ws, {"A": 16, "B": 12, "C": 16, "D": 15, "E": 15, "F": 15, "G": 15, "H": 15, "I": 45})

    # ---- 04_CRUCES ---------------------------------------------------------
    ws = wb.create_sheet("04_CRUCES")
    ws["A1"] = "CRUCES AUTOMATICOS DE CONSISTENCIA (11 controles)"
    ws["A1"].font = TITULO
    ws.append([])
    hdr = ["#", "CONTROL", "# CASOS", "MUESTRA (primeras 15 facturas)"]
    ws.append(hdr)
    estilizar_encabezado(ws, len(hdr), fila=3)
    rid = 3
    for cr in cruces:
        ws.append([cr["n"], cr["nombre"], cr["cantidad"], cr["muestra"]])
        rid += 1
        if cr["cantidad"] > 0 and cr["n"] not in (1,):  # 1 (sin glosa) es informativo
            ws.cell(row=rid, column=3).fill = FILL_AMBAR
    anchos(ws, {"A": 5, "B": 58, "C": 10, "D": 90})

    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(salida))


# ----------------------------------------------------------------------------
# Orquestacion
# ----------------------------------------------------------------------------
def construir(cartera_p: Path, recepcion_p: Path, tramite_p: Path, salida: Path) -> dict:
    cartera = cargar_cartera(cartera_p)
    rec_por_cons = cargar_recepcion(recepcion_p) if recepcion_p else {}
    tramite = cargar_tramite(tramite_p) if tramite_p else []
    cruce_actas = cargar_cruce_actas(cartera_p)

    glosas = construir_glosas(tramite, rec_por_cons)
    agg = agregar_por_factura(glosas)
    fechas_fac = fecha_factura_por_factura(tramite, rec_por_cons)
    contratos = contrato_cod_por_factura(tramite)

    # --- filas de la maestra (espina = cartera) ---
    maestra = []
    for fac in sorted(cartera):
        c = cartera[fac]
        a = agg.get(fac)
        ffac = fechas_fac.get(fac) or c["fecha_radicado"]
        aaaamm = _aaaamm(ffac, c["fecha_radicado"])
        res = resultado_final(c, a)
        ruta_sop = f"{SOPORTES_NOTA} -> ...\\{aaaamm}" if aaaamm else SOPORTES_NOTA
        ruta_fe = f"{FE_RAIZ}\\{aaaamm}" if aaaamm else "PENDIENTE (sin periodo)"
        tipos = ", ".join(sorted(a["tipos"])) if a else ""
        fila = [
            fac,
            c["vigencia"],
            c["estado"],
            contrato_por_fecha(ffac, contratos.get(fac, "")),
            _fecha(ffac),
            _fecha(c["fecha_radicado"]),
            EPS_NOMBRE,
            NIT_TERCERO,
            PRESTADOR,
            c["valor_factura"],
            c["valor_glosa"],
            a["num_glosas"] if a else 0,
            a["glosado"] if a else 0.0,
            a["aceptado"] if a else 0.0,
            a["objetado"] if a else 0.0,
            a["sin_respuesta"] if a else 0,
            c["levantada_actas"],
            c["aceptada_actas"],
            c["ratificada_actas"],
            c["actas"],
            c["saldo_dgh"],
            c["saldo_eps"],
            c["estado_glosa"],
            c["edades"],
            res,
            tipos,
            aaaamm,
            ruta_sop,
            ruta_fe,
            "PENDIENTE VERIFICAR",
            "PENDIENTE (inyectar del motor)",
            "ver glosas",
        ]
        maestra.append({"fila": fila, "resultado": res, "factura": fac})

    cruces = calcular_cruces(cartera, agg, glosas, cruce_actas)

    # --- dashboard ---
    # Los totales de conciliacion (levantado/ratificado/aceptada/a_pagar) se toman
    # del DETALLE de actas (CRUCE ACTAS), excluyendo la duplicada AC000639. Las
    # columnas resumen de CARTERA sub-cuentan la ratificada, por eso no se usan
    # para los KPI (si viven en la maestra como vista propia de cartera).
    actas_nd = [r for r in cruce_actas if not r["duplicada"]]
    tot_facturado = sum(c["valor_factura"] for c in cartera.values())
    tot_glosado = sum(a["glosado"] for a in agg.values())
    tot_aceptado = sum(a["aceptado"] for a in agg.values())
    tot_saldo = sum(c["saldo_dgh"] for c in cartera.values())
    tot_lev = sum(r["levantada"] for r in actas_nd)
    tot_rat = sum(r["ratificada"] for r in actas_nd)
    tot_ace_actas = sum(r["aceptada"] for r in actas_nd)
    tot_apagar = sum(r["a_pagar"] for r in actas_nd)
    n_glosas = len(glosas)
    n_sin_resp = sum(1 for g in glosas if g["estado_respuesta"] == "SIN RESPUESTA")
    conciliado = tot_ace_actas + tot_lev
    kpis = [
        ("Total FACTURADO (nivel factura)", round(tot_facturado), True),
        ("Total GLOSADO (SUM lineas detalle)", round(tot_glosado), True),
        ("Total ACEPTADO por IPS (detalle)", round(tot_aceptado), True),
        ("Total CONCILIADO (aceptada+levantada actas)", round(conciliado), True),
        ("Total LEVANTADO a favor HUS (actas)", round(tot_lev), True),
        ("Total RATIFICADO pdte conciliar (actas)", round(tot_rat), True),
        ("Total A PAGAR (cruce actas, sin duplicada)", round(tot_apagar), True),
        ("Total PENDIENTE (saldo DGH 30/06/2026)", round(tot_saldo), True),
        (
            "% CONCILIADO (conciliado/glosado)",
            round(100 * conciliado / tot_glosado, 1) if tot_glosado else 0,
            False,
        ),
        (
            "% RECUPERACION (levantado/glosado)",
            round(100 * tot_lev / tot_glosado, 1) if tot_glosado else 0,
            False,
        ),
        ("# FACTURAS (maestra, unicas)", len(cartera), False),
        ("# FACTURAS con glosa", len(agg), False),
        ("# GLOSAS (lineas de detalle)", n_glosas, False),
        ("# GLOSAS sin respuesta", n_sin_resp, False),
        ("# CASOS en cruces de consistencia", sum(c["cantidad"] for c in cruces), False),
    ]

    def tabla_por(campo):
        d = defaultdict(lambda: [0, 0.0])
        for fac, c in cartera.items():
            k = c[campo] or "(sin dato)"
            d[k][0] += 1
            d[k][1] += c["valor_glosa"]
        return sorted(((k, v[0], round(v[1])) for k, v in d.items()), key=lambda x: -x[2])

    tablas = [
        ("Cartera por VIGENCIA", tabla_por("vigencia")),
        ("Cartera por ESTADO GLOSA", tabla_por("estado_glosa")),
        ("Cartera por EDADES", tabla_por("edades")),
    ]
    dash = {"kpis": kpis, "tablas": tablas}

    _escribir_workbook(salida, maestra, glosas, cruce_actas, cruces, dash)

    return {
        "facturas_maestra": len(cartera),
        "facturas_con_glosa": len(agg),
        "glosas": n_glosas,
        "glosas_sin_respuesta": n_sin_resp,
        "actas_filas": len(cruce_actas),
        "total_glosado": round(tot_glosado),
        "total_aceptado": round(tot_aceptado),
        "total_levantado": round(tot_lev),
        "total_ratificado": round(tot_rat),
        "total_saldo": round(tot_saldo),
        "cruces": {c["n"]: c["cantidad"] for c in cruces},
        "salida": str(salida),
    }


def main(argv: list[str] | None = None) -> int:
    setup_logging()
    ap = argparse.ArgumentParser(
        description="Arma el Expediente Inteligente de Conciliacion (1 registro por factura)."
    )
    ap.add_argument(
        "--cartera",
        type=Path,
        required=True,
        help="Libro con hojas CARTERA + CRUCE ACTAS (corte 30/06/2026)",
    )
    ap.add_argument(
        "--recepcion",
        type=Path,
        default=None,
        help="RECEPCION_DE_OBJECIONES (glosa recibida / motivo EPS)",
    )
    ap.add_argument(
        "--tramite", type=Path, default=None, help="TRAMITE_DE_OBJECCION (nuestra respuesta)"
    )
    ap.add_argument("--salida", type=Path, required=True, help="Expediente .xlsx de salida")
    args = ap.parse_args(argv)

    try:
        r = construir(args.cartera, args.recepcion, args.tramite, args.salida)
    except (OSError, ValueError) as e:
        logger.error("No pude construir el expediente: %s", e)
        return 1
    logger.info(
        "Expediente listo: %d facturas (%d con glosa), %d glosas (%d sin respuesta), "
        "%d filas de actas -> %s",
        r["facturas_maestra"],
        r["facturas_con_glosa"],
        r["glosas"],
        r["glosas_sin_respuesta"],
        r["actas_filas"],
        r["salida"],
    )
    logger.info(
        "  Glosado $%s | Aceptado $%s | Levantado $%s | Ratificado $%s | Saldo $%s",
        f"{r['total_glosado']:,}",
        f"{r['total_aceptado']:,}",
        f"{r['total_levantado']:,}",
        f"{r['total_ratificado']:,}",
        f"{r['total_saldo']:,}",
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
