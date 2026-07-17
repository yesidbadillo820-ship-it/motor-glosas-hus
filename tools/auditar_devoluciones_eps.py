"""auditar_devoluciones_eps.py — Audita las devoluciones de Nueva EPS cruzando,
por cada factura, TRES fuentes: la factura del DGH (soportes), el RIPS (JSON) y
los soportes de autorización (PDF OPF/PDE). Usado por AUDITAR_DEV_EPS.cmd.

Sigue la guía de auditoría del HUS: del Excel de devoluciones toma cada factura
y verifica en el JSON de RIPS el tipo y número de documento del usuario y el
número de autorización, comparándolos con lo que traen los soportes de
autorización. Llena en el Excel los bloques:

    FACTURA DGH          -> TIPO, NUMERO DE DOCUMENTO, NOMBRE, SERVICIO
    RIPS - JSON          -> TIPO, NUMERO DE DOCUMENTO, Nº AUTORIZACION
    SOPORTES AUTORIZACION-> Nº AUTORIZACION, TIPO, NUMERO DE DOCUMENTO
    OBSERVACION          -> hallazgo de la auditoría (diferencias)

El hallazgo típico: en recién nacidos el JSON trae un documento provisional
(RC largo) distinto al de la factura y los soportes -> se marca la diferencia.

De dónde saca los archivos (busca por número de factura, una sola pasada):
    - JSON de RIPS: en la base de facturas (por defecto la ruta de red del
      HUS) como Rips_<FAC>.json dentro de la carpeta de la factura.
    - Soportes OPF/PDE: en la base de soportes (por defecto Y:) como
      OPF_*_<FAC>.pdf y PDE_*_<FAC>.pdf.

USO:
    py tools\\auditar_devoluciones_eps.py NUEVA_EPS_DEV.xlsx
    py tools\\auditar_devoluciones_eps.py DEV.xlsx --facturas-base "\\\\172.16.32.83\\factura_electronica_net22" --soportes-base "Y:\\" --salida DEV_AUDITADO.xlsx

Requiere openpyxl y pymupdf (fitz). El repo ya los fija.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import Counter
from pathlib import Path

# Tipos de documento del sector salud (para reconocer "RC 12345"). Se dejan
# fuera 'DE' (choca con la preposicion) y 'SI' (choca con el "si"): darian
# falsos positivos en texto OCR en mayusculas.
TIPOS_DOC = ("RC", "TI", "CC", "CE", "PA", "MS", "AS", "PE", "CN", "NV", "SC", "PT", "CD")
# Números institucionales que NO son el documento del paciente.
_INSTITUCIONALES = {"900006037", "9000060374", "680010079201", "6800100792"}

_RE_DOC_TIPADO = re.compile(r"\b(" + "|".join(TIPOS_DOC) + r")\b[\s:\-.]{0,3}(\d{5,20})")
_RE_NUM = re.compile(r"\d{5,20}")
# etiquetas cerca de las que un número NO es el documento del paciente
_RE_NO_DOC = re.compile(
    r"(?:TELEFONO|CELULAR|\bTEL\b|\bCEL\b|FAX|VALOR|VIGENCIA|FECHA|RADICAD|NIT)", re.I
)
# contexto que sube/baja la probabilidad de ser el documento del PACIENTE
_CTX_MAS = re.compile(r"(?:PACIENTE|AFILIADO|USUARIO|NI[NÑ]O|BENEFICIARIO|IDENTIFICAC)", re.I)
_CTX_MENOS = re.compile(
    r"(?:MEDIC|PROFESIONAL|TRATANTE|ORDENA|SOLICITAD|PRESTADOR|AUTORIZAD|LIQUIDAD|REVISOR)", re.I
)


def normalizar_factura(valor: str) -> str:
    digitos = re.sub(r"\D", "", str(valor or ""))
    return digitos.lstrip("0") or digitos


def _patron_factura(norm: str) -> re.Pattern:
    """Reconoce la factura como token 'HUS<num>' (con ceros opcionales),
    sin confundir HUS532392 con HUS5323921 ni concatenar el NIT."""
    return re.compile(rf"HUS0*{re.escape(norm)}(?![0-9])", re.I)


def _texto_pdf(ruta: Path) -> str:
    try:
        import fitz
    except ImportError:
        return ""
    try:
        doc = fitz.open(str(ruta))
    except Exception:
        return ""
    partes = []
    try:
        for pg in doc:
            try:
                partes.append(pg.get_text())
            except Exception:
                continue  # página dañada/cifrada: se omite, no revienta
    finally:
        doc.close()
    return "\n".join(partes)


# --------------------------------------------------------------------------
#  RIPS - JSON
# --------------------------------------------------------------------------

_SERVICIOS_JSON = (
    "consultas",
    "procedimientos",
    "medicamentos",
    "urgencias",
    "hospitalizacion",
    "recienNacidos",
    "otrosServicios",
)


def _autoriz_json_bruto(s: dict) -> str:
    """Nº de autorización del servicio TAL CUAL viene en el JSON: si el campo
    existe y es null, se reporta 'null'; si no existe, ''."""
    if "numAutorizacion" not in s:
        return ""
    valor = s.get("numAutorizacion")
    if valor is None:
        return "null"
    return str(valor).strip()


def leer_rips(ruta: Path) -> dict:
    """Devuelve {factura, usuarios:[{tipo,doc,servicios:[...]}], error}."""
    datos = {"factura": "", "usuarios": [], "error": ""}
    try:
        crudo = ruta.read_bytes()
    except OSError as exc:
        datos["error"] = f"no se pudo leer el JSON ({type(exc).__name__})"
        return datos
    texto = None
    for cod in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            texto = crudo.decode(cod)
            break
        except UnicodeDecodeError:
            continue
    if texto is None:
        datos["error"] = "JSON con codificacion no reconocida"
        return datos
    try:
        raw = json.loads(texto)
    except Exception as exc:
        datos["error"] = f"JSON ilegible ({type(exc).__name__})"
        return datos
    try:
        if not isinstance(raw, dict):
            datos["error"] = "JSON con estructura inesperada (no es un objeto)"
            return datos
        datos["factura"] = str(raw.get("numFactura") or "")
        for u in raw.get("usuarios") or []:
            if not isinstance(u, dict):
                continue
            usuario = {
                "tipo": str(u.get("tipoDocumentoIdentificacion") or "").strip().upper(),
                "doc": str(u.get("numDocumentoIdentificacion") or "").strip(),
                "servicios": [],
            }
            servicios = u.get("servicios") or {}
            if isinstance(servicios, dict):
                for clave in _SERVICIOS_JSON:
                    for s in servicios.get(clave) or []:
                        if not isinstance(s, dict):
                            continue
                        usuario["servicios"].append(
                            {
                                "grupo": clave,
                                # si el JSON trae numAutorizacion en null, se
                                # reporta tal cual ("null"), no como vacío.
                                "autoriz": _autoriz_json_bruto(s),
                                "cod": str(
                                    s.get("codProcedimiento")
                                    or s.get("codConsulta")
                                    or s.get("codTecnologiaSalud")
                                    or ""
                                ).strip(),
                                "codServicio": str(s.get("codServicio") or "").strip(),
                                "valor": s.get("vrServicio") or s.get("vrServicios") or "",
                                "dx": str(s.get("codDiagnosticoPrincipal") or "").strip(),
                            }
                        )
            datos["usuarios"].append(usuario)
    except Exception as exc:
        datos["error"] = f"JSON con estructura inesperada ({type(exc).__name__})"
    return datos


def autorizaciones_del_json(rips: dict) -> list[str]:
    """Todas las autorizaciones del JSON tal cual (incluye 'null'), sin repetir
    y en orden. Se usa para la columna del Excel."""
    vistas, out = set(), []
    for u in rips["usuarios"]:
        for s in u["servicios"]:
            a = s["autoriz"]
            if a and a not in vistas:
                vistas.add(a)
                out.append(a)
    return out


# --------------------------------------------------------------------------
#  Soportes de autorización (PDF)
# --------------------------------------------------------------------------


# "N° Autorizacion: (POS) 5251-313608762" / "P071-314624922" / bare digits.
_RE_AUT = re.compile(
    r"(?:N[°ºo]?\s*)?Autorizaci[oó]n\b[^0-9A-Za-z]{0,8}(?:\(?POS\)?\s*)?"
    r"([0-9A-Za-z]{2,8}-\d{5,12}|\d{6,15})",
    re.I,
)


def _ultimos9(token: str) -> str:
    """Solo los últimos 9 dígitos del número de autorización (5251-313608762
    -> 313608762). Si trae menos de 9 dígitos, los que haya; '' si ninguno."""
    d = re.sub(r"\D", "", str(token or ""))
    return d[-9:] if len(d) >= 9 else d


def extraer_autorizaciones(texto: str) -> list[str]:
    """TODAS las autorizaciones del soporte, cada una reducida a sus últimos 9
    dígitos, sin repetir y en orden de aparición."""
    vistas, out = set(), []
    for m in _RE_AUT.finditer(texto):
        v = _ultimos9(m.group(1))
        if v and v not in _INSTITUCIONALES and v not in vistas:
            vistas.add(v)
            out.append(v)
    return out


def extraer_documento(texto: str) -> tuple[str, str]:
    """(tipo, numero) del PACIENTE en el soporte. Puntúa por contexto para no
    tomar el documento del médico, teléfonos, valores ni fechas."""
    candidatos = []
    for m in _RE_DOC_TIPADO.finditer(texto):
        tipo, num = m.group(1).upper(), m.group(2)
        if num in _INSTITUCIONALES:
            continue
        # el tipo (RC/CC/...) ya distingue un documento de un telefono/valor;
        # se puntua por lo que PRECEDE al documento (ahi va la etiqueta).
        ventana_prev = texto[max(0, m.start() - 25) : m.start()]
        score = 0
        if _CTX_MAS.search(ventana_prev):
            score += 2
        if _CTX_MENOS.search(ventana_prev):
            score -= 2
        candidatos.append((score, tipo, num))
    if candidatos:
        # el documento del paciente: mejor contexto; a igualdad, el mas repetido
        frec = Counter((t, n) for _, t, n in candidatos)
        mejor = max(candidatos, key=lambda c: (c[0], frec[(c[1], c[2])]))
        return mejor[1], mejor[2]
    # sin tipo explícito: tras "IDENTIFICACION"/"Historia clinica", evitando telefonos
    for m in re.finditer(
        r"(?:IDENTIFICACION|Historia\s+clinica)\b([\s\S]{0,40}?)(\d{5,20})", texto, re.I
    ):
        entre = m.group(1)
        if _RE_NO_DOC.search(entre):
            continue
        if m.group(2) not in _INSTITUCIONALES:
            return "", m.group(2)
    return "", ""


_STOP_NOMBRE = re.compile(
    r"\b(?:NUMERO|IDENTIFIC|CODIGO|EPS|EDAD|FECHA|DIRECCION|TIPO|HISTORIA|LIQUIDADOR|"
    r"REVISOR|AUXILIAR|INSTITUCION|INGRESO|NIT|NOMBRE|No|VALOR|AFILIADO|OrIGEN|DX)\b",
    re.I,
)


def _limpiar_nombre(bruto: str) -> str:
    nombre = re.sub(r"\s+", " ", bruto).strip()
    nombre = _STOP_NOMBRE.split(nombre)[0]
    tokens = [t for t in re.split(r"[\s/]+", nombre) if t]
    while tokens and len(re.sub(r"[^A-Za-zÁÉÍÓÚÑ]", "", tokens[-1])) <= 2:
        tokens.pop()
    limpio = " ".join(tokens).strip(" .,-/")
    return limpio if len(limpio.split()) >= 2 else ""


def extraer_nombre(texto: str) -> str:
    """Nombre del paciente en el soporte (varios formatos, sin cruzar líneas)."""
    patrones = (
        r"Nombre\s+del\s+usuario[:\s]+([^\n]{6,60})",
        r"Identificaci[oó]n\s+del\s+Ni[nñ]o[:\s]+([^\n]{6,60})",
        r"NOMBRES\s*\n\s*([^\n]{4,40})\s*\n\s*APELLIDOS\s*\n\s*([^\n]{4,40})",
        r"([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ ]{6,55})[^\n]{0,4}\n\s*NUMERO DE IDENTIFICACION",
    )
    for pat in patrones:
        m = re.search(pat, texto)
        if not m:
            continue
        bruto = " ".join(g for g in m.groups() if g) if m.re.groups > 1 else m.group(1)
        nombre = _limpiar_nombre(bruto)
        if nombre:
            return nombre
    return ""


def extraer_servicio(texto: str) -> str:
    """Descripción del servicio/programa autorizado."""
    for pat in (
        r"PAQUETE\s+MENSUAL[^\n]{0,60}",
        r"PROGRAMA\s+MADRE\s+CANGURO[^\n]{0,30}",
        r"Descripci[oó]n\s+Servicio\s*\n(?:[^\n]*\n){0,2}?\s*([A-ZÁÉÍÓÚÑ][^\n]{6,60})",
    ):
        m = re.search(pat, texto, re.I)
        if m:
            frag = m.group(1) if m.groups() else m.group(0)
            frag = re.sub(r"\s+", " ", frag).strip(" .:-")
            frag = re.sub(r"[^A-Za-zÁÉÍÓÚÑñáéíóú0-9 ]+.*$", "", frag).strip()
            if len(frag) >= 4:
                return frag.upper()
    return ""


def leer_soportes(pdfs: list[Path]) -> dict:
    """Combina el texto de los soportes (OPF/PDE) y extrae los campos clave.
    'autorizaciones' trae TODAS (últimos 9 dígitos); 'autoriz' es su unión."""
    res = {
        "autorizaciones": [],
        "autoriz": "",
        "tipo": "",
        "doc": "",
        "nombre": "",
        "servicio": "",
        "archivos": [],
        "sin_texto": False,
    }
    textos = []
    for p in pdfs:
        t = _texto_pdf(p)
        res["archivos"].append(p.name)
        if t.strip():
            textos.append(t)
    if not textos:
        res["sin_texto"] = bool(pdfs)  # había PDF pero sin texto (¿escaneado?)
        return res
    texto = "\n".join(textos)
    res["autorizaciones"] = extraer_autorizaciones(texto)
    res["autoriz"] = ", ".join(res["autorizaciones"])
    res["tipo"], res["doc"] = extraer_documento(texto)
    res["nombre"] = extraer_nombre(texto)
    res["servicio"] = extraer_servicio(texto)
    return res


# --------------------------------------------------------------------------
#  Índice de archivos (una sola pasada por cada base, tolerante a fallos)
# --------------------------------------------------------------------------


_RE_HUS_DIR = re.compile(r"^HUS\d+$", re.I)


def _es_dir(p: Path) -> bool:
    try:
        return p.is_dir()
    except OSError:
        return False


def _slot_vacio() -> dict:
    return {"json": None, "fv": None, "opf": [], "pde": [], "otros": []}


def _clasificar_archivo(slot: dict, p: Path) -> None:
    ext = p.suffix.lower()
    up = p.name.upper()
    if ext == ".json" and "RIPS" in up:
        slot["json"] = slot["json"] or p
    elif ext == ".pdf" and up.startswith("OPF"):
        slot["opf"].append(p)
    elif ext == ".pdf" and up.startswith("PDE"):
        slot["pde"].append(p)
    elif ext == ".pdf" and up.startswith("FV"):
        slot["fv"] = slot["fv"] or p
    elif ext == ".pdf":
        slot["otros"].append(p)


def _recolectar_carpeta(carpeta: Path, patron: re.Pattern, slot: dict) -> None:
    """Mete en `slot` los archivos de `carpeta` (y su subcarpeta RIPS) que
    correspondan a la factura (por el token del nombre; los que no traen HUS
    en el nombre, como fv, se aceptan por estar en la carpeta de la factura)."""
    for sub in (carpeta, carpeta / "RIPS", carpeta / "SOPORTES"):
        if not _es_dir(sub):
            continue
        try:
            entradas = list(os.scandir(sub))
        except OSError:
            continue
        for e in entradas:
            try:
                if not e.is_file():
                    continue
            except OSError:
                continue
            nombre = e.name
            ext = os.path.splitext(nombre)[1].lower()
            if ext not in (".json", ".pdf"):
                continue
            if patron.search(nombre) or "HUS" not in nombre.upper():
                _clasificar_archivo(slot, Path(e.path))


def indexar_directo(base: Path, facturas: list[tuple[str, str, str]]) -> dict[str, dict]:
    """Va DIRECTO a la carpeta de cada factura (sin recorrer todo el árbol).
    facturas: [(fac_original, norm, mes_yyyymm)]. Estructura esperada:
        <base>\\<AAAAMM>\\FACTURAS_SALUD\\<HUS...>\\(RIPS\\)
    Devuelve el índice y el conjunto de facturas que NO encontró así."""
    indice = {norm: _slot_vacio() for _, norm, _ in facturas}
    faltan = set()
    for fac, norm, mes in facturas:
        patron = _patron_factura(norm)
        # meses probables: el de la factura y el anterior/siguiente (por si se
        # radicó en un mes distinto al de la fecha de la factura).
        meses = [m for m in (mes, _mes_adyacente(mes, 1), _mes_adyacente(mes, -1)) if m]
        candidatas = []
        for mm in meses:
            for nombre_fac in (fac, f"HUS{norm}", f"HUS{norm.zfill(6)}"):
                candidatas.append(base / mm / "FACTURAS_SALUD" / nombre_fac)
        hallo = False
        for c in candidatas:
            if _es_dir(c):
                _recolectar_carpeta(c, patron, indice[norm])
                hallo = True
        if not (hallo and indice[norm]["json"]):
            faltan.add(norm)
    return indice, faltan


def indexar_arbol(base: Path, facturas_norm: set[str], etiqueta: str = "") -> dict[str, dict]:
    """Recorre `base` UNA vez. Para no tardar horas sobre una unidad de red se
    salta descender a las carpetas de facturas que NO son objetivo (las más
    numerosas), pero NADA más: así siempre encuentra los soportes, sin importar
    bajo qué envío/mes/gestor estén archivados. Muestra progreso."""
    indice = {n: _slot_vacio() for n in facturas_norm}
    patrones = {n: _patron_factura(n) for n in facturas_norm}
    if not _es_dir(base):
        return indice
    vistos = 0
    for root, dirs, files in os.walk(str(base), onerror=lambda _e: None):
        vistos += 1
        if vistos % 400 == 0:
            print(
                f"    ... {vistos} carpetas revisadas{(' en ' + etiqueta) if etiqueta else ''}",
                flush=True,
            )
        # ---- poda SEGURA: solo se salta descender a la carpeta de UNA factura
        #      que NO es objetivo (las mas numerosas). NO se poda por envio ni
        #      por mes: el envio de la devolucion no siempre coincide con el de
        #      la carpeta de soportes, y podar por eso hacia perder todo. ----
        conservar = []
        for d in dirs:
            if _RE_HUS_DIR.match(d) and not any(pat.search(d) for pat in patrones.values()):
                continue  # carpeta de otra factura: no entrar (aqui esta el ahorro)
            conservar.append(d)
        dirs[:] = conservar
        # ---- clasificar archivos de esta carpeta ----
        for fn in files:
            ext = os.path.splitext(fn)[1].lower()
            if ext not in (".json", ".pdf"):
                continue
            objetivo = next((n for n, pat in patrones.items() if pat.search(fn)), None)
            if objetivo is None and "HUS" not in fn.upper():
                objetivo = next((n for n, pat in patrones.items() if pat.search(root)), None)
            if objetivo is None:
                continue
            _clasificar_archivo(indice[objetivo], Path(root) / fn)
    return indice


def diagnostico_soportes(base: Path) -> None:
    """Cuando no se halla ningun soporte, dice qué se ve realmente bajo `base`
    (para saber si la ruta es la correcta y cómo se llaman los archivos)."""
    print("  [DIAGNOSTICO] Revisando qué hay en la carpeta de soportes...", flush=True)
    if not base or not _es_dir(base):
        print(f"  [DIAGNOSTICO] La ruta NO existe o no es accesible: {base}")
        print("               Verifica que la unidad (p. ej. Y:) este conectada.")
        return
    carpetas = pdfs = 0
    ejemplos = []
    for root, _dirs, files in os.walk(str(base), onerror=lambda _e: None):
        carpetas += 1
        for fn in files:
            if fn.lower().endswith(".pdf"):
                pdfs += 1
                if len(ejemplos) < 12:
                    ejemplos.append(str(Path(root) / fn))
        if carpetas >= 40000 or len(ejemplos) >= 12:
            break
    print(f"  [DIAGNOSTICO] Carpetas revisadas: {carpetas}   PDFs vistos: {pdfs}")
    if ejemplos:
        print("  [DIAGNOSTICO] Ejemplos de PDF encontrados (asi se llaman):")
        for e in ejemplos:
            print(f"                {e}")
        print("  [DIAGNOSTICO] Si estos NO son los soportes de estas facturas, la")
        print("               ruta de soportes esta equivocada; usa la carpeta correcta.")
    else:
        print("  [DIAGNOSTICO] No hay NINGUN PDF bajo esa ruta: es la carpeta equivocada.")


def _fundir(dst: dict, extra: dict) -> None:
    """Suma al índice `dst` lo que traiga `extra` (para completar lo que el
    modo directo no encontró usando el recorrido podado)."""
    for norm, slot in extra.items():
        if norm not in dst:
            dst[norm] = _slot_vacio()
        d = dst[norm]
        d["json"] = d["json"] or slot["json"]
        d["fv"] = d["fv"] or slot["fv"]
        for k in ("opf", "pde", "otros"):
            d[k] = d[k] or slot[k]


# --------------------------------------------------------------------------
#  Comparación y observación
# --------------------------------------------------------------------------


def observacion(rips_doc: str, autoriz_json: list[str], sop: dict) -> str:
    """Compara documento y autorizaciones (por sus últimos 9 dígitos) del JSON
    contra el soporte. Avisa cuando el JSON trae la autorización en null."""
    partes = []
    sop_doc = sop.get("doc", "")
    set_sop = {a for a in sop.get("autorizaciones", []) if a}
    set_json = {_ultimos9(a) for a in autoriz_json if _ultimos9(a)}
    hay_null = any(str(a).strip().lower() == "null" for a in autoriz_json)

    # --- autorización ---
    if not set_json and not set_sop:
        if hay_null:
            partes.append("JSON CON AUTORIZACION EN NULL")
        elif sop.get("sin_texto"):
            partes.append("SOPORTE SIN TEXTO (revisar/OCR)")
        else:
            partes.append("SIN AUTORIZACION EN JSON NI SOPORTE")
    elif not set_sop:
        partes.append(
            "SOPORTE SIN TEXTO (revisar/OCR)"
            if sop.get("sin_texto")
            else "SIN AUTORIZACION EN EL SOPORTE"
        )
    elif not set_json:
        partes.append(
            "JSON CON AUTORIZACION EN NULL" if hay_null else "SIN AUTORIZACION EN EL JSON"
        )
    elif set_json == set_sop or set_json <= set_sop or set_sop <= set_json:
        partes.append("OK")
        if hay_null:
            partes.append("OJO: alguna autorizacion del JSON en NULL")
    else:
        solo_json = sorted(set_json - set_sop)
        solo_sop = sorted(set_sop - set_json)
        det = []
        if solo_json:
            det.append("en JSON no en soporte: " + ", ".join(solo_json[:5]))
        if solo_sop:
            det.append("en soporte no en JSON: " + ", ".join(solo_sop[:5]))
        partes.append("AUTORIZACION DIFERENTE (" + "; ".join(det) + ")")

    # --- documento del paciente: JSON (RIPS) vs soporte (= FACTURA DGH) ---
    if not rips_doc:
        partes.append("SIN USUARIO EN EL JSON")
    elif not sop_doc:
        if not sop.get("sin_texto"):
            partes.append("SIN DOCUMENTO EN EL SOPORTE (no verificado)")
    elif normalizar_factura(rips_doc) != normalizar_factura(sop_doc):
        partes.append("DIFERENCIA DEL NUMERO DE DOCUMENTO DGH VS JSON")

    return " - ".join(partes) if partes else "REVISAR"


# --------------------------------------------------------------------------
#  Excel
# --------------------------------------------------------------------------


def _detectar_cabecera(ws) -> int | None:
    for fila in range(1, 8):
        vals = [
            str(ws.cell(fila, c).value or "").strip().upper() for c in range(1, ws.max_column + 1)
        ]
        if "FACTURA" in vals and "FAC" in vals:
            return fila
    return None


def _mes_yyyymm(valor) -> str:
    """Mes 'AAAAMM' de la FECHA FACTURA (datetime de Excel o texto dd/mm/aaaa)."""
    from datetime import date, datetime

    if isinstance(valor, (datetime, date)):
        return f"{valor.year:04d}{valor.month:02d}"
    s = str(valor or "").strip()
    m = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", s)
    if m:
        return f"{int(m.group(3)):04d}{int(m.group(2)):02d}"
    m = re.search(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}{int(m.group(2)):02d}"
    return ""


def _mes_adyacente(yyyymm: str, delta: int) -> str:
    if not re.fullmatch(r"\d{6}", yyyymm or ""):
        return ""
    y, mth = int(yyyymm[:4]), int(yyyymm[4:])
    mth += delta
    if mth == 0:
        y, mth = y - 1, 12
    elif mth == 13:
        y, mth = y + 1, 1
    return f"{y:04d}{mth:02d}"


def _es_factura(valor) -> bool:
    """La celda FAC parece una factura (no una fila de TOTAL/subtotal)."""
    s = str(valor or "").strip()
    if not s:
        return False
    if not re.search(r"\d", s):
        return False
    if re.search(r"TOTAL|SUBTOTAL|SUMA|GENERAL", s, re.I):
        return False
    return True


def procesar(excel: Path, facturas_base: Path, soportes_base: Path, salida: Path) -> int:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    try:
        wb = load_workbook(str(excel))
    except Exception as exc:
        sys.stderr.write(
            f"ERROR: no se pudo abrir el Excel ({type(exc).__name__}). ¿Está abierto?\n"
        )
        return 2
    ws = wb.active
    fila_cab = _detectar_cabecera(ws)
    if fila_cab is None:
        sys.stderr.write("ERROR: no encontre la cabecera (FACTURA / FAC) en el Excel.\n")
        return 2
    COL = {
        "fac": 2,
        "fecha": 4,
        "dgh_tipo": 7,
        "dgh_doc": 8,
        "dgh_nombre": 9,
        "dgh_serv": 10,
        "js_tipo": 11,
        "js_doc": 12,
        "js_aut": 13,
        "sop_aut": 14,
        "sop_tipo": 15,
        "sop_doc": 16,
        "obs": 17,
        "ruta_json": 18,
        "ruta_sop": 19,
    }
    # encabezados de las dos columnas nuevas (rutas de los archivos)
    ws.cell(fila_cab, COL["ruta_json"]).value = "RUTA DEL JSON"
    ws.cell(fila_cab, COL["ruta_sop"]).value = "RUTA DE LOS SOPORTES"

    facturas = []
    for fila in range(fila_cab + 1, ws.max_row + 1):
        fac = ws.cell(fila, COL["fac"]).value
        if _es_factura(fac):
            envio = str(ws.cell(fila, 6).value or "").strip()
            mes = _mes_yyyymm(ws.cell(fila, COL["fecha"]).value)
            facturas.append((fila, str(fac).strip(), envio, mes))

    print("=" * 66)
    print("  AUDITAR DEVOLUCIONES EPS — DGH vs RIPS(JSON) vs soportes")
    print("=" * 66)
    print(f"  Excel: {excel.name}   Facturas: {len(facturas)}")
    print(f"  Base facturas/JSON: {facturas_base}")
    print(f"  Base soportes:      {soportes_base}")

    norms = {normalizar_factura(fac) for _, fac, _, _ in facturas}

    # 1) FACTURAS/JSON: ir DIRECTO a la carpeta de cada factura (rapidísimo)
    print("  [1/2] Buscando los JSON de RIPS (directo por carpeta)...", flush=True)
    directo = [(fac, normalizar_factura(fac), mes) for _, fac, _, mes in facturas]
    idx_fact, faltan = indexar_directo(facturas_base, directo)
    if faltan:
        print(f"        {len(faltan)} sin ruta directa: buscando en el arbol...", flush=True)
        extra = indexar_arbol(facturas_base, faltan, etiqueta="facturas")
        _fundir(idx_fact, extra)

    # 2) SOPORTES: recorrido del arbol (salta solo las facturas ajenas). Los
    #    soportes tambien pueden estar en la base de facturas -> se fusiona.
    print("  [2/2] Buscando los soportes OPF/PDE (con progreso)...", flush=True)
    idx_sop = indexar_arbol(soportes_base, norms, etiqueta="soportes")
    _fundir(idx_sop, {n: idx_fact.get(n, _slot_vacio()) for n in norms})

    con_json = sum(1 for n in norms if idx_fact.get(n, {}).get("json"))
    con_sop = sum(
        1
        for n in norms
        if (
            idx_sop.get(n, {}).get("opf")
            or idx_sop.get(n, {}).get("pde")
            or idx_sop.get(n, {}).get("otros")
        )
    )
    print("-" * 66)
    print(f"  JSON de RIPS encontrados:   {con_json} de {len(norms)}")
    print(f"  Soportes OPF/PDE encontrados: {con_sop} de {len(norms)}")
    if con_sop == 0:
        print("  [OJO] No se hallo NINGUN soporte bajo la carpeta indicada.")
        diagnostico_soportes(soportes_base)
    print("-" * 66)

    verde = PatternFill("solid", fgColor="E2EFDA")
    amarillo = PatternFill("solid", fgColor="FFF2CC")
    rojo = PatternFill("solid", fgColor="F8CBAD")

    detalle = [
        [
            "Factura",
            "Usuario #",
            "Grupo servicio",
            "JSON tipo",
            "JSON documento",
            "Nº autorizacion JSON",
            "Cod servicio",
            "Cod procedimiento",
            "Valor",
            "Diagnostico",
        ]
    ]
    con_dif = con_ok = sin_datos = 0

    for fila, fac, _envio, _mes in facturas:
        norm = normalizar_factura(fac)
        arch_f = idx_fact.get(norm, {})  # el JSON viene de la base de facturas
        arch_s = idx_sop.get(norm, {})  # soportes (ya fusionados con facturas)
        pdfs, vistos_pdf = [], set()
        for k in ("opf", "pde", "otros"):  # tambien PDFs con otro nombre
            for p in arch_s.get(k) or []:
                if str(p) not in vistos_pdf:
                    vistos_pdf.add(str(p))
                    pdfs.append(p)

        rips = (
            leer_rips(arch_f["json"])
            if arch_f.get("json")
            else {"factura": "", "usuarios": [], "error": "sin JSON"}
        )
        autoriz_json = autorizaciones_del_json(rips)
        primer = (
            rips["usuarios"][0] if rips["usuarios"] else {"tipo": "", "doc": "", "servicios": []}
        )
        sop = leer_soportes(pdfs)

        ws.cell(fila, COL["dgh_tipo"]).value = sop["tipo"]
        ws.cell(fila, COL["dgh_doc"]).value = sop["doc"]
        ws.cell(fila, COL["dgh_nombre"]).value = sop["nombre"]
        ws.cell(fila, COL["dgh_serv"]).value = sop["servicio"]
        ws.cell(fila, COL["js_tipo"]).value = primer["tipo"]
        ws.cell(fila, COL["js_doc"]).value = primer["doc"]
        ws.cell(fila, COL["js_aut"]).value = ", ".join(autoriz_json)
        ws.cell(fila, COL["sop_aut"]).value = sop["autoriz"]
        ws.cell(fila, COL["sop_tipo"]).value = sop["tipo"]
        ws.cell(fila, COL["sop_doc"]).value = sop["doc"]

        # rutas de los archivos (para que sea facil ir a buscarlos)
        ws.cell(fila, COL["ruta_json"]).value = str(arch_f["json"]) if arch_f.get("json") else ""
        ws.cell(fila, COL["ruta_sop"]).value = "\n".join(str(p) for p in pdfs)

        if rips.get("error") and not pdfs:
            obs = "SIN JSON NI SOPORTES (revisar ruta)"
            sin_datos += 1
            pintura = amarillo
        else:
            obs = observacion(primer["doc"], autoriz_json, sop)
            if "DIFERENCIA" in obs or "DIFERENTE" in obs:
                con_dif += 1
                pintura = rojo
            elif obs == "OK":
                con_ok += 1
                pintura = verde
            else:
                sin_datos += 1
                pintura = amarillo
        n_serv = sum(len(u["servicios"]) for u in rips["usuarios"])
        if len(rips["usuarios"]) > 1 or n_serv > 1:
            obs += f"  (+{len(rips['usuarios'])} usuario(s)/{n_serv} servicio(s): ver hoja DETALLE)"
        ws.cell(fila, COL["obs"]).value = obs
        ws.cell(fila, COL["obs"]).fill = pintura

        for iu, u in enumerate(rips["usuarios"], 1):
            if not u["servicios"]:
                detalle.append([fac, iu, "", u["tipo"], u["doc"], "", "", "", "", ""])
            for s in u["servicios"]:
                detalle.append(
                    [
                        fac,
                        iu,
                        s["grupo"],
                        u["tipo"],
                        u["doc"],
                        s["autoriz"],
                        s["codServicio"],
                        s["cod"],
                        s["valor"],
                        s["dx"],
                    ]
                )

        marca = "!" if pintura is rojo else ("·" if pintura is amarillo else "OK")
        print(
            f"  {marca:2s} {fac:16s} JSON doc={primer['doc'] or '-':16s} sop doc={sop['doc'] or '-':12s} aut={', '.join(autoriz_json) or '-'}"
        )

    # estilo de las dos columnas de ruta (encabezado + ancho + ajuste de texto)
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    azul = PatternFill("solid", fgColor="1F4E79")
    for col in (COL["ruta_json"], COL["ruta_sop"]):
        cab = ws.cell(fila_cab, col)
        cab.font = Font(bold=True, color="FFFFFF")
        cab.fill = azul
        cab.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 60
        for fila in range(fila_cab + 1, ws.max_row + 1):
            ws.cell(fila, col).alignment = Alignment(vertical="top", wrap_text=True)

    if "DETALLE" in wb.sheetnames:
        del wb["DETALLE"]
    ws2 = wb.create_sheet("DETALLE")
    for r in detalle:
        ws2.append(r)
    for c in ws2[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = azul
    ws2.freeze_panes = "A2"

    try:
        wb.save(str(salida))
    except PermissionError:
        sys.stderr.write(
            f"\nERROR: no se pudo guardar {salida.name} (¿lo tienes abierto en Excel?).\n"
            "       Cierralo y vuelve a correr el bot.\n"
        )
        return 2
    except OSError as exc:
        sys.stderr.write(f"\nERROR: no se pudo guardar el Excel ({type(exc).__name__}).\n")
        return 2

    print("-" * 66)
    print(f"  Con diferencia: {con_dif}   OK: {con_ok}   Sin datos/revisar: {sin_datos}")
    print(f"  Informe: {salida}")
    print("=" * 66)
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Audita devoluciones EPS: DGH vs RIPS(JSON) vs soportes."
    )
    parser.add_argument("excel", help="Excel de devoluciones (NUEVA_EPS_DEV.xlsx).")
    parser.add_argument(
        "--facturas-base",
        default=r"\\172.16.32.83\factura_electronica_net22",
        help="Carpeta raíz de las facturas/JSON (red del HUS).",
    )
    parser.add_argument(
        "--soportes-base", default="Y:\\", help="Carpeta raíz de los soportes de autorización."
    )
    parser.add_argument("--salida", type=Path, help="Ruta del Excel auditado.")
    args = parser.parse_args(argv)

    excel = Path(args.excel).expanduser()
    if not excel.is_file():
        sys.stderr.write(f"ERROR: no existe el Excel: {excel}\n")
        return 2

    def _limpiar_base(v: str) -> Path:
        v = (v or "").strip().strip('"')  # quita comillas sueltas del batch
        return Path(os.path.normpath(v)) if v else Path(v)

    fbase = _limpiar_base(args.facturas_base)
    sbase = _limpiar_base(args.soportes_base)
    salida = args.salida or excel.with_name(excel.stem + "_AUDITADO.xlsx")
    return procesar(excel, fbase, sbase, salida)


if __name__ == "__main__":
    raise SystemExit(main())
