"""motor_evidencia_dispensario.py — Modulo 3: Motor de Evidencia.

No basta con decir "existe historia clinica". Este motor LEE los soportes
clinicos pagina por pagina y localiza, para cada glosa, EN QUE PAGINA esta la
prueba, con el fragmento textual exacto. Asi la respuesta pasa de
"encontre una historia clinica" a
"Historia clinica, pagina 18: '...se ordena procedimiento 890201 consulta...'".

Trabaja sobre el EXPEDIENTE unico (expedientes.json): sabe las rutas exactas de
los soportes de cada factura, asi que abre SOLO esos archivos (no recorre Y:\\).

REGLA: nunca inventa. Si el termino no aparece en ningun soporte, la evidencia
queda vacia y la glosa se marca como "sin evidencia localizada".

USO
---
    py tools\\motor_evidencia_dispensario.py ^
        --expedientes "D:\\...\\expedientes.json" ^
        --salida      "D:\\...\\expedientes_con_evidencia.json" ^
        --reporte     "D:\\...\\EVIDENCIA.xlsx"   # opcional, legible
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import sys
from pathlib import Path

logger = logging.getLogger("motor_evidencia")

# Maximo de paginas citadas por glosa. Evita el "disparo de escopeta" (citar
# toda la historia clinica: 300 paginas no son una prueba usable en un oficio).
# Se conservan las paginas MAS relevantes, no las primeras que aparezcan.
MAX_EVIDENCIAS_POR_GLOSA = 6

# Tipos de soporte donde tiene sentido buscar evidencia clinica.
CLINICOS = {
    "Historia clinica / Evolucion",
    "Historia clinica",
    "Epicrisis",
    "Descripcion quirurgica",
    "Orden medica / Formula",
    "Orden medica",
    "Hoja de administracion de medicamentos",
    "Resultados de laboratorio",
    "Descripcion / detalle de prestacion",
}

# Palabras vacias que no sirven como termino de busqueda.
STOP = {
    "DE",
    "DEL",
    "LA",
    "EL",
    "LOS",
    "LAS",
    "POR",
    "CON",
    "SIN",
    "PARA",
    "QUE",
    "SE",
    "GLOSA",
    "CODIGO",
    "VALOR",
    "PESOS",
    "UNIDAD",
    "SEGUN",
    "UNA",
    "UNO",
    "SERVICIO",
    "SOPORTE",
    "SOPORTES",
    "COBRO",
    "OBJETA",
    "OBJETAR",
    "FACTURA",
    # Palabras clinicas demasiado genericas: aparecen en casi toda HC y no
    # prueban por si solas el servicio glosado.
    "PROCEDIMIENTO",
    "PROCEDIMIENTOS",
    "ATENCION",
    "PACIENTE",
    "GENERAL",
    "DERECHO",
    "IZQUIERDO",
}


def setup_logging() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")


# ---------------------------------------------------------------------------
# Lectura pagina por pagina
# ---------------------------------------------------------------------------
def texto_por_pagina(ruta: Path, ocr: bool = False) -> list[str]:
    """Devuelve una cadena por pagina. OCR opcional para PDFs escaneados."""
    if ruta.suffix.lower() != ".pdf":
        try:
            return [ruta.read_text(encoding="utf-8", errors="ignore")]
        except OSError:
            return []
    try:
        import pdfplumber
    except ImportError:
        return []
    paginas: list[str] = []
    try:
        with pdfplumber.open(str(ruta)) as pdf:
            for p in pdf.pages:
                paginas.append(p.extract_text() or "")
    except Exception as e:  # noqa: BLE001
        logger.warning("No pude leer %s: %s", ruta.name, e)
        return []
    if sum(len(t.strip()) for t in paginas) >= 30 or not ocr:
        return paginas
    # Escaneado + OCR
    try:
        import pytesseract
        from pdf2image import convert_from_path

        return [pytesseract.image_to_string(im, lang="spa") for im in convert_from_path(str(ruta))]
    except Exception as e:  # noqa: BLE001
        logger.warning("OCR no disponible en %s: %s", ruta.name, e)
        return paginas


# ---------------------------------------------------------------------------
# Terminos de busqueda derivados de la glosa
# ---------------------------------------------------------------------------
def terminos_de_glosa(servicio: str, motivo: str) -> list[str]:
    """Extrae los terminos con que buscar la evidencia de una glosa.

    - Codigos (CUPS/CUM/servicio): 4-7 digitos con sufijo opcional (890201, 21519H).
    - Palabras significativas del servicio (mayusculas, >=5 letras, sin stopwords).
    """
    texto = f"{servicio} {motivo}".upper()
    terminos: list[str] = []
    for cod in re.findall(r"\b\d{4,7}[A-Z]?\b", texto):
        if cod not in terminos:
            terminos.append(cod)
    for pal in re.findall(r"\b[A-ZÁÉÍÓÚÑ]{5,}\b", (servicio or "").upper()):
        if pal not in STOP and pal not in terminos:
            terminos.append(pal)
    return terminos[:8]


def _snippet(texto: str, pos: int, ancho: int = 70) -> str:
    ini = max(0, pos - ancho)
    fin = min(len(texto), pos + ancho)
    frag = re.sub(r"\s+", " ", texto[ini:fin]).strip()
    return ("..." if ini > 0 else "") + frag + ("..." if fin < len(texto) else "")


# ---------------------------------------------------------------------------
# Localizacion de evidencia
# ---------------------------------------------------------------------------
def cachear_texto(clinicos: list[dict], ocr: bool = False) -> dict[str, list[str]]:
    """Lee cada soporte clinico UNA sola vez y devuelve {ruta: [paginas]}.

    Clave de rendimiento para el lote completo: sin esto, el motor releeria el
    mismo PDF una vez por cada glosa de la factura (una factura con 10 glosas
    leeria sus PDF 10 veces).
    """
    cache: dict[str, list[str]] = {}
    for doc in clinicos:
        ruta = Path(doc.get("ruta", ""))
        clave = str(ruta)
        if clave in cache or not ruta.exists():
            continue
        cache[clave] = texto_por_pagina(ruta, ocr)
    return cache


def evidencias_para_glosa(
    servicio: str,
    motivo: str,
    clinicos: list[dict],
    ocr: bool = False,
    cache: dict[str, list[str]] | None = None,
) -> list[dict]:
    """Localiza la evidencia de una glosa y devuelve SOLO las paginas mas
    relevantes (no todas las que contengan una palabra generica).

    Por cada pagina calcula un puntaje: cada codigo CUPS/CUM pesa 10 y cada
    palabra del servicio pesa 1. Ordena por evidencia fuerte (codigo) primero,
    luego por puntaje, y devuelve como maximo MAX_EVIDENCIAS_POR_GLOSA paginas.
    Asi la cita es usable en un oficio ("paginas 9-11"), no un listado de 300.

    clinicos: lista de {tipo, nombre, ruta}. `cache` (opcional): texto ya leido
    por ruta (ver cachear_texto) para no releer los PDF en cada glosa. Cada
    evidencia: {tipo, nombre, ruta, pagina, termino, fuerza, puntaje, fragmento}.
    Sin coincidencias -> [].
    """
    terminos = terminos_de_glosa(servicio, motivo)
    if not terminos:
        return []
    codigos = {t for t in terminos if re.fullmatch(r"\d{4,7}[A-Z]?", t)}
    candidatas: list[dict] = []
    for doc in clinicos:
        ruta = Path(doc.get("ruta", ""))
        clave = str(ruta)
        if cache is not None and clave in cache:
            paginas = cache[clave]
        elif ruta.exists():
            paginas = texto_por_pagina(ruta, ocr)
        else:
            continue
        for npag, texto in enumerate(paginas, 1):
            up = texto.upper()
            aciertos: list[tuple[str, int]] = []
            for term in terminos:
                pos = up.find(term)
                if pos >= 0:
                    aciertos.append((term, pos))
            if not aciertos:
                continue
            hay_codigo = any(t in codigos for t, _ in aciertos)
            puntaje = sum(10 if t in codigos else 1 for t, _ in aciertos)
            # Fragmento del termino mas fuerte (el codigo, si aparece).
            aciertos.sort(key=lambda a: (a[0] not in codigos, a[1]))
            term_best, pos_best = aciertos[0]
            candidatas.append(
                {
                    "tipo": doc.get("tipo", ""),
                    "nombre": doc.get("nombre", ""),
                    "ruta": str(ruta),
                    "pagina": npag,
                    "termino": term_best,
                    "fuerza": "fuerte" if hay_codigo else "debil",
                    "puntaje": puntaje,
                    "fragmento": _snippet(texto, pos_best),
                }
            )
    # Fuertes (con codigo) primero, luego por puntaje y numero de pagina.
    candidatas.sort(key=lambda e: (e["fuerza"] != "fuerte", -e["puntaje"], e["pagina"]))
    return candidatas[:MAX_EVIDENCIAS_POR_GLOSA]


def enriquecer_expedientes(payload: dict, ocr: bool = False) -> dict:
    """Agrega, a cada glosa de cada expediente, su lista 'evidencias' localizada."""
    total_ev = 0
    for exp in payload.get("expedientes", []):
        clinicos = [s for s in exp.get("soportes", []) if s.get("tipo") in CLINICOS]
        # Lee cada PDF de la factura UNA vez y reusa el texto en todas sus glosas.
        cache = cachear_texto(clinicos, ocr)
        for g in exp.get("glosas", []):
            ev = evidencias_para_glosa(
                g.get("servicio", ""), g.get("motivo", ""), clinicos, ocr, cache
            )
            g["evidencias"] = ev
            total_ev += len(ev)
    payload["evidencias_localizadas"] = total_ev
    return payload


# ---------------------------------------------------------------------------
# Reporte legible (opcional)
# ---------------------------------------------------------------------------
def escribir_reporte(payload: dict, salida: Path) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evidencia"
    ws.append(
        [
            "Factura",
            "Glosa",
            "Familia",
            "Servicio",
            "Soporte",
            "Pagina",
            "Termino",
            "Fragmento (evidencia)",
        ]
    )
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E5F")
    for exp in payload.get("expedientes", []):
        for g in exp.get("glosas", []):
            evs = g.get("evidencias", [])
            if not evs:
                ws.append(
                    [
                        exp["factura"],
                        g.get("codigo_corto", ""),
                        g.get("familia", ""),
                        str(g.get("servicio", ""))[:40],
                        "SIN EVIDENCIA LOCALIZADA",
                        "",
                        "",
                        "",
                    ]
                )
                continue
            for ev in evs:
                ws.append(
                    [
                        exp["factura"],
                        g.get("codigo_corto", ""),
                        g.get("familia", ""),
                        str(g.get("servicio", ""))[:40],
                        ev["tipo"],
                        ev["pagina"],
                        ev["termino"],
                        ev["fragmento"],
                    ]
                )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"
    for w, col in zip((16, 12, 12, 42, 34, 8, 12, 80), "ABCDEFGH"):
        ws.column_dimensions[col].width = w
    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(salida))


def main(argv: list[str] | None = None) -> int:
    setup_logging()
    ap = argparse.ArgumentParser(
        description="Motor de evidencia (pagina/parrafo) sobre el expediente."
    )
    ap.add_argument(
        "--expedientes", type=Path, required=True, help="expedientes.json (fuente de la verdad)"
    )
    ap.add_argument("--salida", type=Path, required=True, help="expedientes_con_evidencia.json")
    ap.add_argument(
        "--reporte", type=Path, default=None, help="Excel legible de evidencia (opcional)"
    )
    ap.add_argument("--ocr", action="store_true", help="OCR para soportes escaneados")
    args = ap.parse_args(argv)

    try:
        payload = json.loads(args.expedientes.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        logger.error("No pude leer %s: %s", args.expedientes, e)
        return 1

    payload = enriquecer_expedientes(payload, args.ocr)
    args.salida.parent.mkdir(parents=True, exist_ok=True)
    args.salida.write_text(json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")
    logger.info(
        "Evidencias localizadas: %d  ->  %s", payload.get("evidencias_localizadas", 0), args.salida
    )
    if args.reporte:
        escribir_reporte(payload, args.reporte)
        logger.info("Reporte legible: %s", args.reporte)
    return 0


if __name__ == "__main__":
    sys.exit(main())
