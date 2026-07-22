"""motor_verificacion_dispensario.py — Modulo 3.5: Verificacion de Evidencia.

Localizar una frase en un PDF NO prueba la glosa. Este motor, con REGLAS
DETERMINISTAS (no IA), convierte el expediente en un EXPEDIENTE PROBATORIO:
por cada glosa establece los HECHOS que hay que demostrar y verifica, con la
evidencia hallada (Modulo 3) y los soportes presentes, si cada hecho quedó
PROBADO, con su nivel de confianza y qué documento falta.

Ademas corre un MOTOR DE CONTRADICCIONES: cruza factura/paciente/CUPS/contrato
entre soportes y avisa inconsistencias ANTES de responder la glosa.

Filosofia: las reglas resuelven los casos claros; la IA (modulos posteriores)
solo interviene donde hay interpretacion. Todo es auditable y trazable.

Entrada: expedientes_con_evidencia.json (salida del Modulo 3).
Salida:  expedientes_probatorios.json (+ reporte opcional).

USO
---
    py tools\\motor_verificacion_dispensario.py ^
        --expedientes "D:\\...\\expedientes_con_evidencia.json" ^
        --salida      "D:\\...\\expedientes_probatorios.json" ^
        --reporte     "D:\\...\\HECHOS_PROBADOS.xlsx"
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import sys
from pathlib import Path

_TOOLS = Path(__file__).resolve().parent
if str(_TOOLS) not in sys.path:
    sys.path.insert(0, str(_TOOLS))
from asistente_conciliacion_dispensario import factura_corta  # noqa: E402

logger = logging.getLogger("motor_verificacion")

# ---------------------------------------------------------------------------
# Reglas: por cada familia de glosa, los HECHOS a demostrar y con qué
# documentos (basta uno del grupo). `por_codigo`=True exige que ademas el
# codigo CUPS/CUM aparezca en ese soporte (evidencia "fuerte" del Modulo 3).
# ---------------------------------------------------------------------------
HECHOS_POR_FAMILIA: dict[str, list[dict]] = {
    "SOPORTES": [
        {
            "clave": "ordenado",
            "hecho": "La prestacion fue ORDENADA",
            "requiere_alguno": [
                "Orden medica / Formula",
                "Orden medica",
                "Historia clinica / Evolucion",
                "Historia clinica",
            ],
            "por_codigo": True,
        },
        {
            "clave": "ejecutado",
            "hecho": "La prestacion fue EJECUTADA / registrada",
            "requiere_alguno": [
                "Descripcion quirurgica",
                "Epicrisis",
                "Hoja de administracion de medicamentos",
                "Resultados de laboratorio",
                "Historia clinica / Evolucion",
            ],
            "por_codigo": True,
        },
    ],
    "CALIDAD": [
        {
            "clave": "soporte_clinico",
            "hecho": "Existe soporte clinico de la atencion",
            "requiere_alguno": [
                "Historia clinica / Evolucion",
                "Historia clinica",
                "Epicrisis",
                "Descripcion quirurgica",
            ],
            "por_codigo": False,
        },
    ],
    "PERTINENCIA": [
        {
            "clave": "soporte_clinico",
            "hecho": "Existe soporte clinico que sustenta la pertinencia",
            "requiere_alguno": ["Historia clinica / Evolucion", "Epicrisis"],
            "por_codigo": False,
        },
    ],
    "TARIFAS": [
        {
            "clave": "facturado",
            "hecho": "El servicio fue facturado",
            "requiere_alguno": [
                "Factura electronica (XML/CUFE)",
                "Factura electronica (representacion PDF)",
                "RIPS",
            ],
            "por_codigo": False,
        },
        {
            "clave": "contrato",
            "hecho": "Hay contrato aplicable a la fecha de atencion",
            "requiere_contrato": True,
        },
    ],
    "FACTURACION": [
        {
            "clave": "facturado",
            "hecho": "El servicio fue facturado y registrado en RIPS",
            "requiere_alguno": [
                "RIPS",
                "Factura electronica (XML/CUFE)",
                "Descripcion / detalle de prestacion",
            ],
            "por_codigo": False,
        },
    ],
    "COBERTURA": [
        {
            "clave": "cobertura",
            "hecho": "Existe autorizacion / soporte de cobertura",
            "requiere_alguno": ["Autorizacion", "MIPRES"],
            "por_codigo": False,
        },
    ],
    "AUTORIZACION": [
        {
            "clave": "autorizacion",
            "hecho": "Existe autorizacion del servicio",
            "requiere_alguno": ["Autorizacion", "MIPRES"],
            "por_codigo": False,
        },
    ],
}


def setup_logging() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")


# ---------------------------------------------------------------------------
# Verificacion de hechos por glosa (regla determinista)
# ---------------------------------------------------------------------------
def verificar_hechos(glosa: dict, soportes: list[dict], contrato: dict) -> list[dict]:
    tipos_presentes = {s.get("tipo") for s in soportes}
    evid = glosa.get("evidencias", [])
    hechos: list[dict] = []
    for spec in HECHOS_POR_FAMILIA.get(glosa.get("familia", ""), []):
        if spec.get("requiere_contrato"):
            prob = contrato.get("codigo") in ("287", "440")
            hechos.append(
                {
                    "hecho": spec["hecho"],
                    "probado": prob,
                    "nivel_confianza": 0.90 if prob else 0.30,
                    "evidencia": (
                        [
                            f"Contrato {contrato.get('codigo')} ({contrato.get('base_tarifaria', '')})"
                        ]
                        if prob
                        else []
                    ),
                    "faltantes": [] if prob else ["Contrato/otrosí vigente a la fecha"],
                }
            )
            continue
        req = spec["requiere_alguno"]
        presentes = [t for t in req if t in tipos_presentes]
        if not presentes:
            hechos.append(
                {
                    "hecho": spec["hecho"],
                    "probado": False,
                    "nivel_confianza": 0.0,
                    "evidencia": [],
                    "faltantes": req,
                }
            )
            continue
        refs = [f"{e['tipo']} pag.{e['pagina']}" for e in evid if e.get("tipo") in presentes]
        fuerte = any(e.get("fuerza") == "fuerte" for e in evid if e.get("tipo") in presentes)
        if spec.get("por_codigo"):
            conf = 0.95 if fuerte else (0.80 if refs else 0.60)
        else:
            conf = 0.85 if refs else 0.70
        hechos.append(
            {
                "hecho": spec["hecho"],
                "probado": True,
                "nivel_confianza": conf,
                "evidencia": refs or [f"{presentes[0]} (presente)"],
                "faltantes": [],
            }
        )
    return hechos


# ---------------------------------------------------------------------------
# Motor de contradicciones (cruza factura / paciente / CUPS / contrato)
# ---------------------------------------------------------------------------
def _cargar_rips(ruta: Path) -> dict | None:
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return json.loads(ruta.read_text(encoding=enc))
        except (UnicodeDecodeError, json.JSONDecodeError, OSError):
            continue
    return None


def detectar_contradicciones(exp: dict) -> list[str]:
    alertas: list[str] = []
    cartera = exp.get("cartera", {})
    if not cartera.get("en_cartera", False):
        alertas.append("La factura no aparece en la cartera del HUS (verificar radicado real).")
    if exp.get("contrato", {}).get("pendiente_confirmar"):
        alertas.append(
            "Contrato 287 por confirmar: falta el acta de inicio para fijar la vigencia."
        )

    rips_sop = next((s for s in exp.get("soportes", []) if s.get("tipo") == "RIPS"), None)
    if rips_sop:
        ruta = Path(rips_sop.get("ruta", ""))
        data = _cargar_rips(ruta) if ruta.exists() else None
        if isinstance(data, dict):
            numf = str(data.get("numFactura") or "")
            if numf and factura_corta(numf) != exp.get("factura_corta"):
                alertas.append(f"El RIPS trae otra factura ({numf}) — posible cruce equivocado.")
            docs = {
                str(u.get("numDocumentoIdentificacion", "")) for u in (data.get("usuarios") or [])
            }
            ident = str(exp.get("paciente", {}).get("identificacion", ""))
            if ident and docs and ident not in docs:
                alertas.append(f"El documento del paciente ({ident}) no coincide con el RIPS.")
            # CUPS del servicio glosado presente en el RIPS
            blob = json.dumps(data)
            for g in exp.get("glosas", []):
                cods = re.findall(r"\b\d{4,7}\b", str(g.get("servicio", "")))
                if cods and not any(c in blob for c in cods):
                    alertas.append(
                        f"Glosa {g.get('codigo_corto', '')}: el codigo del servicio "
                        f"({cods[0]}) no aparece en el RIPS."
                    )
    return alertas


# ---------------------------------------------------------------------------
# Orquestacion
# ---------------------------------------------------------------------------
def verificar_expedientes(payload: dict) -> dict:
    total_hechos = probados = total_alertas = 0
    for exp in payload.get("expedientes", []):
        soportes = exp.get("soportes", [])
        contrato = exp.get("contrato", {})
        for g in exp.get("glosas", []):
            hechos = verificar_hechos(g, soportes, contrato)
            g["hechos_probados"] = hechos
            total_hechos += len(hechos)
            probados += sum(1 for h in hechos if h["probado"])
        exp["alertas"] = detectar_contradicciones(exp)
        total_alertas += len(exp["alertas"])
        # Estado sugerido (determinista): si todo hecho de todas las glosas esta
        # probado y no hay alertas -> defendible.
        todos = [h for g in exp.get("glosas", []) for h in g.get("hechos_probados", [])]
        exp["defendible"] = bool(todos) and all(h["probado"] for h in todos) and not exp["alertas"]
    payload["resumen_verificacion"] = {
        "hechos_totales": total_hechos,
        "hechos_probados": probados,
        "alertas": total_alertas,
        "expedientes_defendibles": sum(
            1 for e in payload.get("expedientes", []) if e.get("defendible")
        ),
    }
    return payload


# ---------------------------------------------------------------------------
# Reporte legible
# ---------------------------------------------------------------------------
def escribir_reporte(payload: dict, salida: Path) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hechos"
    ws.append(
        ["Factura", "Glosa", "Familia", "Hecho", "Probado", "Confianza", "Evidencia", "Falta"]
    )
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E5F")
    ok = PatternFill("solid", fgColor="D9EAD3")
    no = PatternFill("solid", fgColor="F4CCCC")
    for exp in payload.get("expedientes", []):
        for g in exp.get("glosas", []):
            for h in g.get("hechos_probados", []):
                ws.append(
                    [
                        exp.get("factura"),
                        g.get("codigo_corto", ""),
                        g.get("familia", ""),
                        h["hecho"],
                        "SI" if h["probado"] else "NO",
                        round(h["nivel_confianza"], 2),
                        "; ".join(h["evidencia"]),
                        "; ".join(h["faltantes"]),
                    ]
                )
                ws.cell(row=ws.max_row, column=5).fill = ok if h["probado"] else no
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"
    # Hoja de alertas
    wa = wb.create_sheet("Alertas")
    wa.append(["Factura", "Alerta / contradiccion"])
    for c in wa[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="B00000")
    for exp in payload.get("expedientes", []):
        for a in exp.get("alertas", []):
            wa.append([exp.get("factura"), a])
    for w, col in zip((16, 12, 12, 44, 8, 10, 40, 40), "ABCDEFGH"):
        ws.column_dimensions[col].width = w
    wa.column_dimensions["A"].width = 16
    wa.column_dimensions["B"].width = 90
    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(salida))


def main(argv: list[str] | None = None) -> int:
    setup_logging()
    ap = argparse.ArgumentParser(
        description="Motor de verificacion de evidencia (hechos probados)."
    )
    ap.add_argument(
        "--expedientes",
        type=Path,
        required=True,
        help="expedientes_con_evidencia.json (salida del Modulo 3)",
    )
    ap.add_argument("--salida", type=Path, required=True, help="expedientes_probatorios.json")
    ap.add_argument("--reporte", type=Path, default=None, help="Excel legible (opcional)")
    args = ap.parse_args(argv)

    try:
        payload = json.loads(args.expedientes.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        logger.error("No pude leer %s: %s", args.expedientes, e)
        return 1

    payload = verificar_expedientes(payload)
    args.salida.parent.mkdir(parents=True, exist_ok=True)
    args.salida.write_text(json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")
    r = payload["resumen_verificacion"]
    logger.info(
        "Hechos probados: %d/%d | Alertas: %d | Expedientes defendibles: %d  -> %s",
        r["hechos_probados"],
        r["hechos_totales"],
        r["alertas"],
        r["expedientes_defendibles"],
        args.salida,
    )
    if args.reporte:
        escribir_reporte(payload, args.reporte)
        logger.info("Reporte legible: %s", args.reporte)
    return 0


if __name__ == "__main__":
    sys.exit(main())
