#!/usr/bin/env python3
"""siifa_preparar_respuestas.py — Puente entre el informe de SIIFA y el bot
que sube las respuestas.

EL PROBLEMA QUE RESUELVE
El informe de seguimientos trae MILES de líneas, pero muchas son la misma
glosa repetida: la misma factura, el mismo código, la misma observación,
sólo que SIIFA las registró una por una. En la corrida del 03-08-2026 eran
2.579 líneas pendientes... que se responden con apenas 272 textos distintos.

Escribir 2.579 respuestas a mano es imposible. Escribir 272 es un día de
trabajo. Este script hace las dos mitades:

    1. --plantilla   : del informe saca UNA fila por respuesta a escribir
                       (agrupa factura + código + observación de la glosa).
                       El auditor llena dos columnas y nada más.
    2. --expandir    : toma esa plantilla ya llena y la reparte a todas las
                       líneas que le corresponden, dejando el Excel que
                       consume tools/responder_glosas_siifa.py.

USO:
    REM 1) Sacar la hoja de trabajo del informe
    py siifa_preparar_respuestas.py --informe "D:\\...\\informe_seguimientos.xlsx" ^
        --plantilla "D:\\...\\respuestas_PARA_LLENAR.xlsx"

    REM 2) (el auditor llena CODIGO_RESPUESTA y OBSERVACION_RESPUESTA)

    REM 3) Repartirla a todas las líneas, para el bot
    py siifa_preparar_respuestas.py --informe "D:\\...\\informe_seguimientos.xlsx" ^
        --expandir "D:\\...\\respuestas_PARA_LLENAR.xlsx" ^
        --salida "D:\\...\\respuestas_siifa.xlsx"

    REM 4) Piloto de 1 glosa y después el cargue (ver responder_glosas_siifa.py)

Se puede acotar el trabajo con --tipo GLOSA / --tipo DEVOLUCION y con
--factura, para ir por partes.

INSTALACIÓN (una vez):
    py -m pip install openpyxl
"""

from __future__ import annotations

import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path

logger = logging.getLogger("siifa_preparar")

# Columnas de la hoja de trabajo que llena el auditor. Las dos últimas son
# las únicas que se escriben a mano; el resto es contexto para decidir.
COLUMNAS_PLANTILLA = [
    "GRUPO",
    "TIPO",
    "NUMERO_FACTURA",
    "CODIGO_GLOSA",
    "DESCRIPCION_GLOSA",
    "OBSERVACION_GLOSA",
    "LINEAS_QUE_CUBRE",
    "VALOR_TOTAL",
    "FECHA_REPORTE_MAS_ANTIGUA",
    "DIAS_DESDE_EL_REPORTE",
    "CODIGO_RESPUESTA",
    "OBSERVACION_RESPUESTA",
]

# Columnas del Excel final, el que lee responder_glosas_siifa.py.
COLUMNAS_SALIDA = [
    "ID_SEGUIMIENTO_FACTURA_GLOSA",
    "NUMERO_FACTURA",
    "TIPO",
    "CODIGO_GLOSA",
    "VALOR_GLOSA",
    "CODIGO_RESPUESTA",
    "OBSERVACION_RESPUESTA",
]


def _texto(v) -> str:
    return "" if v is None else str(v).strip()


def clave_de_grupo(fila: dict) -> str:
    """Identifica una respuesta a escribir.

    Dos líneas van juntas si son la misma factura, el mismo código de glosa
    y la misma observación de la EPS: en ese caso la respuesta del hospital
    es necesariamente la misma. La clave se calcula igual al armar la
    plantilla y al expandirla, que es lo que permite volver a juntarlas.
    """
    return "|".join(
        (
            _texto(fila.get("numero_factura")),
            _texto(fila.get("codigo_glosa")),
            _texto(fila.get("observacion_glosa"))[:200],
        )
    )


def leer_informe(ruta: Path, hoja: str = "SEGUIMIENTOS") -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(ruta, read_only=True, data_only=True)
    if hoja not in wb.sheetnames:
        raise SystemExit(
            f"\nERROR: el archivo {ruta.name} no tiene la hoja «{hoja}».\n"
            f"Hojas que trae: {', '.join(wb.sheetnames)}\n"
            "¿Seguro que es el informe de tools/siifa_reporte_seguimientos.py?\n"
        )
    filas = wb[hoja].iter_rows(values_only=True)
    columnas = [_texto(c) for c in next(filas)]
    return [dict(zip(columnas, f)) for f in filas]


def _dias_desde(valor) -> int | None:
    if not valor:
        return None
    try:
        f = datetime.fromisoformat(str(valor)[:19])
    except ValueError:
        return None
    return (datetime.now() - f).days


def agrupar(
    datos: list[dict],
    tipo: str | None = None,
    factura: str | None = None,
    incluir_respondidas: bool = False,
) -> dict[str, list[dict]]:
    """Junta las líneas pendientes por respuesta a escribir."""
    grupos: dict[str, list[dict]] = {}
    for f in datos:
        if not incluir_respondidas and _texto(f.get("tiene_respuesta")).upper() == "SI":
            continue
        if tipo and _texto(f.get("tipo_seguimiento")).upper() != tipo.upper():
            continue
        if factura and _texto(f.get("numero_factura")).upper() != factura.upper():
            continue
        grupos.setdefault(clave_de_grupo(f), []).append(f)
    return grupos


def escribir_plantilla(grupos: dict[str, list[dict]], ruta: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "RESPUESTAS"

    encabezado = PatternFill("solid", fgColor="1F4E78")
    a_llenar = PatternFill("solid", fgColor="FFF2CC")
    for c, h in enumerate(COLUMNAS_PLANTILLA, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = encabezado
        cell.font = Font(bold=True, color="FFFFFF")

    # Primero lo más viejo y lo de más plata: es lo que hay que responder ya.
    def orden(par):
        lineas = par[1]
        dias = max((_dias_desde(x.get("fecha_reporte")) or 0) for x in lineas)
        return (-dias, -sum(x.get("valor_glosa") or 0 for x in lineas))

    for r, (clave, lineas) in enumerate(sorted(grupos.items(), key=orden), start=2):
        primera = lineas[0]
        fechas = [x.get("fecha_reporte") for x in lineas if x.get("fecha_reporte")]
        mas_antigua = min((str(x) for x in fechas), default="")
        valores = {
            "GRUPO": clave,
            "TIPO": primera.get("tipo_seguimiento"),
            "NUMERO_FACTURA": primera.get("numero_factura"),
            "CODIGO_GLOSA": primera.get("codigo_glosa"),
            "DESCRIPCION_GLOSA": primera.get("descripcion_glosa"),
            "OBSERVACION_GLOSA": primera.get("observacion_glosa"),
            "LINEAS_QUE_CUBRE": len(lineas),
            "VALOR_TOTAL": sum(x.get("valor_glosa") or 0 for x in lineas),
            "FECHA_REPORTE_MAS_ANTIGUA": mas_antigua[:10],
            "DIAS_DESDE_EL_REPORTE": _dias_desde(mas_antigua),
            "CODIGO_RESPUESTA": None,
            "OBSERVACION_RESPUESTA": None,
        }
        for c, h in enumerate(COLUMNAS_PLANTILLA, 1):
            cell = ws.cell(row=r, column=c, value=valores[h])
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=h in ("DESCRIPCION_GLOSA", "OBSERVACION_GLOSA", "OBSERVACION_RESPUESTA"),
            )
            if h in ("CODIGO_RESPUESTA", "OBSERVACION_RESPUESTA"):
                cell.fill = a_llenar
            if h == "VALOR_TOTAL":
                cell.number_format = '"$"#,##0'

    anchos = {
        "GRUPO": 34,
        "TIPO": 12,
        "NUMERO_FACTURA": 14,
        "CODIGO_GLOSA": 12,
        "DESCRIPCION_GLOSA": 50,
        "OBSERVACION_GLOSA": 50,
        "LINEAS_QUE_CUBRE": 10,
        "VALOR_TOTAL": 16,
        "FECHA_REPORTE_MAS_ANTIGUA": 13,
        "DIAS_DESDE_EL_REPORTE": 10,
        "CODIGO_RESPUESTA": 16,
        "OBSERVACION_RESPUESTA": 60,
    }
    for c, h in enumerate(COLUMNAS_PLANTILLA, 1):
        ws.column_dimensions[get_column_letter(c)].width = anchos[h]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNAS_PLANTILLA))}{ws.max_row}"

    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta)
    logger.info("Hoja de trabajo guardada: %s (%d respuestas por escribir)", ruta, len(grupos))


def leer_plantilla_llena(ruta: Path) -> dict[str, dict]:
    """Devuelve, por grupo, la respuesta que escribió el auditor."""
    from openpyxl import load_workbook

    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = ws.iter_rows(values_only=True)
    columnas = [_texto(c).upper() for c in next(filas)]
    faltan = {"GRUPO", "CODIGO_RESPUESTA"} - set(columnas)
    if faltan:
        raise SystemExit(
            f"\nERROR: a la hoja de trabajo le faltan columnas: {', '.join(sorted(faltan))}.\n"
            "Se llena sobre la plantilla que genera este mismo script con --plantilla.\n"
        )

    respuestas: dict[str, dict] = {}
    for f in filas:
        fila = dict(zip(columnas, f))
        clave = _texto(fila.get("GRUPO"))
        if not clave:
            continue
        respuestas[clave] = {
            "CODIGO_RESPUESTA": _texto(fila.get("CODIGO_RESPUESTA")),
            "OBSERVACION_RESPUESTA": _texto(fila.get("OBSERVACION_RESPUESTA")),
        }
    return respuestas


def expandir(
    grupos: dict[str, list[dict]], respuestas: dict[str, dict]
) -> tuple[list[dict], list[str]]:
    """Reparte cada respuesta a todas sus líneas.

    Devuelve las filas listas para el bot y la lista de grupos que quedaron
    sin responder — esos NO se suben: quedan para la siguiente pasada.
    """
    filas: list[dict] = []
    sin_responder: list[str] = []
    for clave, lineas in grupos.items():
        resp = respuestas.get(clave)
        if not resp or not resp["CODIGO_RESPUESTA"]:
            sin_responder.append(clave)
            continue
        for linea in lineas:
            filas.append(
                {
                    "ID_SEGUIMIENTO_FACTURA_GLOSA": linea.get("id_seguimiento_factura_glosa"),
                    "NUMERO_FACTURA": linea.get("numero_factura"),
                    "TIPO": linea.get("tipo_seguimiento"),
                    "CODIGO_GLOSA": linea.get("codigo_glosa"),
                    "VALOR_GLOSA": linea.get("valor_glosa"),
                    "CODIGO_RESPUESTA": resp["CODIGO_RESPUESTA"],
                    "OBSERVACION_RESPUESTA": resp["OBSERVACION_RESPUESTA"],
                }
            )
    return filas, sin_responder


def escribir_salida(filas: list[dict], ruta: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "RESPUESTAS"
    for c, h in enumerate(COLUMNAS_SALIDA, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(bold=True, color="FFFFFF")
    for r, fila in enumerate(filas, start=2):
        for c, h in enumerate(COLUMNAS_SALIDA, 1):
            ws.cell(row=r, column=c, value=fila.get(h))
    for c, h in enumerate(COLUMNAS_SALIDA, 1):
        ws.column_dimensions[get_column_letter(c)].width = 60 if "OBSERVACION" in h else 20
    ws.freeze_panes = "A2"

    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta)
    logger.info("Excel para el bot guardado: %s (%d respuestas)", ruta, len(filas))


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--informe", required=True, help="Informe de siifa_reporte_seguimientos.py.")
    ap.add_argument("--plantilla", help="Ruta donde dejar la hoja de trabajo por llenar.")
    ap.add_argument("--expandir", help="Hoja de trabajo YA LLENA, para repartir a cada línea.")
    ap.add_argument("--salida", help="Excel final para responder_glosas_siifa.py.")
    ap.add_argument("--tipo", choices=["GLOSA", "DEVOLUCION"], help="Trabajar solo un tipo.")
    ap.add_argument("--factura", help="Trabajar solo una factura (ej. HUS454747).")
    ap.add_argument("--verbose", action="store_true")
    args = ap.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s  %(message)s",
        datefmt="%H:%M:%S",
        stream=sys.stdout,
    )

    if not args.plantilla and not args.expandir:
        raise SystemExit(
            "\nERROR: falta --plantilla (para sacarla) o --expandir (para repartirla).\n"
        )
    if args.expandir and not args.salida:
        raise SystemExit("\nERROR: con --expandir hay que decir también --salida.\n")

    datos = leer_informe(Path(args.informe))
    grupos = agrupar(datos, tipo=args.tipo, factura=args.factura)
    lineas = sum(len(v) for v in grupos.values())
    if not grupos:
        raise SystemExit("\nNo quedó nada pendiente con esos filtros. Nada que hacer.\n")
    logger.info(
        "%d líneas pendientes se responden con %d textos distintos.",
        lineas,
        len(grupos),
    )

    if args.plantilla:
        escribir_plantilla(grupos, Path(args.plantilla))
        print(
            f"\nLlene las dos columnas amarillas (CODIGO_RESPUESTA y "
            f"OBSERVACION_RESPUESTA) de {Path(args.plantilla).name} y después corra "
            f"este mismo script con --expandir.\n"
        )
        return

    respuestas = leer_plantilla_llena(Path(args.expandir))
    filas, sin_responder = expandir(grupos, respuestas)
    if not filas:
        raise SystemExit(
            "\nNinguna fila de la hoja de trabajo tiene CODIGO_RESPUESTA. "
            "No hay nada que subir todavía.\n"
        )
    if sin_responder:
        logger.warning(
            "OJO: %d de %d respuestas quedaron en blanco — esas NO se van a subir. "
            "Quedan para la próxima pasada.",
            len(sin_responder),
            len(grupos),
        )
    escribir_salida(filas, Path(args.salida))
    print(
        f"\nListo. Ahora el piloto de UNA sola glosa antes del cargue "
        f"(regla del repo):\n"
        f'   py tools\\responder_glosas_siifa.py --excel "{args.salida}" '
        f"--solo-id {filas[0]['ID_SEGUIMIENTO_FACTURA_GLOSA']} "
        f'--reporte "piloto_siifa.csv"\n'
    )


if __name__ == "__main__":
    main()
