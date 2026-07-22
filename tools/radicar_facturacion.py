"""radicar_facturacion.py — Radicador maestro multi-entidad de la ESE HUS.

Es el paso AGUAS ARRIBA de los bots de glosas: toma el lote de facturas que
entrega el área de FACTURACIÓN y lo deja listo para radicar ante cada entidad
(EPS, regímenes especiales, SOAT/ADRES). En una sola corrida:

  1. INVENTARÍA y TIPIFICA cada soporte de cada factura según el catálogo
     ADRES (Resolución 2284/2023): FEV, RIP, CUV, HEV, EPI, HAU, PDX, OPF,
     DQX, RAN, CRC, HAM, TAP, FMO, FUR, SER, IPO, … Detecta los que están
     mal nombrados o sin reconocer.
  2. LEE el RIPS y la factura electrónica (FEV) — reutiliza tools/adres/ —
     para sacar número de factura, NIT del prestador, usuarios, servicios y
     el valor total facturado.
  3. RESUELVE la ENTIDAD/pagador de cada factura (desde el manifiesto de
     facturación, desde la carpeta padre o desde el nombre) y la cruza con su
     PERFIL de radicación (data/perfiles_radicacion.json): soportes
     obligatorios, canal/portal, si exige CUV, nomenclatura y formato de
     paquete.
  4. VALIDA la completitud: soportes base (FEV+RIP+CUV), soportes esperados
     según los servicios del RIPS, consistencia del número de factura entre
     RIPS / FEV / carpeta, y CUV cuando la entidad lo exige.
  5. ARMA el paquete (con --armar): crea <destino>/<ENTIDAD>/<factura>/,
     renombra los soportes a la nomenclatura de la entidad, escribe un
     manifiesto por factura y, con --zip, genera el ZIP
     <NIT>_<AAAAMMDD>_<consecutivo>.zip.
  6. REPORTA: CSV por factura, resumen por entidad y resumen global, con
     estados claros (LISTA / FALTAN_SOPORTES / SIN_CUV / SIN_RIPS / SIN_FEV /
     ENTIDAD_NO_RESUELTA / REVISAR_TIPIFICACION).

Por defecto NO toca archivos: solo diagnostica y reporta (modo auditoría).
Pasá --armar para construir los paquetes y --zip para comprimirlos.

USO RÁPIDO
----------

    # Solo auditar el lote (no arma nada): tipificación + completitud + reporte
    py radicar_facturacion.py --origen "D:\\LOTE_FACTURACION_2026-06" \\
        --reporte "D:\\salida\\radicacion_2026-06.csv" --xlsx "D:\\salida\\radicacion_2026-06.xlsx"

    # Armar los paquetes por entidad (carpetas + renombrado ADRES)
    py radicar_facturacion.py --origen "D:\\LOTE" --destino "D:\\RADICAR" --armar

    # Armar y comprimir en ZIP por factura, solo una entidad
    py radicar_facturacion.py --origen "D:\\LOTE" --destino "D:\\RADICAR" \\
        --armar --zip --entidad COOSALUD

    # El lote viene de un share anidado (ESCANEO o factura_electronica): el
    # layout 'auto' (por defecto) ancla en las carpetas HUS<factura>.
    py radicar_facturacion.py --origen "/mnt/radicacion_2026/.../ESCANEO"

    # Mapear factura → entidad con un Excel/CSV de facturación
    py radicar_facturacion.py --origen "D:\\LOTE" --manifiesto "D:\\facturacion.xlsx"

LAYOUTS DE ENTRADA
------------------

    auto (def.)             Ancla en carpetas cuyo nombre es una factura
                            (HUS520769) y les asigna TODO lo que cuelga debajo
                            (incluye subcarpetas como RIPS/ y los documentos DIAN
                            fv…/ad…/ar…/.zip). Lo no anclado se agrupa por el
                            token del nombre. Fusiona por número de factura. Es
                            el más robusto: cubre todos los shares reales.
    carpeta-factura         Cada subcarpeta es UNA factura; todos sus archivos
                            le pertenecen. Fusiona árboles paralelos con el mismo
                            nombre de carpeta.
    lote                    Una carpeta con archivos de MUCHAS facturas: se
                            agrupan por el token _<factura>_ del nombre. Los
                            compartidos (FURIPS) se reparten a su carpeta.

LOS SHARES REALES
-----------------

    1) ESCANEO (RIPS + CUV, organizado por EPS):

        …\\ESCANEO\\<EPS>\\RIPS\\ENV-<lote>\\HUS<factura>\\
            HUS<factura>.json        ← RIPS (¡nombrado SOLO con la factura!)
            CUV_HUS<factura>.json    ← validación MSPS

       El RIPS viene SIN token (es solo <factura>.json) y la EPS sale de la
       carpeta <EPS> de la ruta.

    2) factura_electronica (FEV DIAN, organizado por factura):

        …\\FACTURAS_SALUD\\HUS<factura>\\
            RIPS\\                    ← subcarpeta con el RIPS (y CUV)
            ad09…xml                 ← AttachedDocument DIAN (CUFE) = la FEV
            ar09…xml                 ← ApplicationResponse (acuse DIAN)
            fv09…xml / fv09…pdf      ← factura de venta (XML + representación)
            HUS<factura>.zip         ← paquete FE comprimido

       Los nombres DIAN (fv…/ad…/ar…) se tipifican solos. Como la ruta NO trae
       la EPS, el radicador la lee del adquiriente (AccountingCustomerParty) de
       la propia factura electrónica — no necesitás --manifiesto (aunque podés
       usarlo para forzar/corregir el mapeo). 'auto' recoge el RIPS de la
       subcarpeta y los documentos DIAN en una sola factura.

DEPENDENCIAS
------------

    Núcleo: solo librería estándar (corre en cualquier PC).
    --xlsx y manifiesto .xlsx: openpyxl (py -m pip install openpyxl).
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import logging
import os
import re
import shutil
import sys
import unicodedata
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from functools import lru_cache
from pathlib import Path, PureWindowsPath

logger = logging.getLogger("radicar")

# Reutilizamos los parsers ADRES (stdlib) que ya están vetados por el HUS.
_ADRES_DIR = Path(__file__).resolve().parent / "adres"
sys.path.insert(0, str(_ADRES_DIR))
try:
    from rips_lectura import (  # noqa: E402
        SERVICIOS_RIPS,
        cargar_rips,
        datos_generales,
        extraer_lineas_servicios,
    )

    _ADRES_OK = True
except ImportError:  # pragma: no cover - el módulo adres siempre acompaña al script
    _ADRES_OK = False
    SERVICIOS_RIPS = (
        "consultas",
        "procedimientos",
        "medicamentos",
        "otrosServicios",
        "urgencias",
        "hospitalizacion",
        "recienNacidos",
    )


# ─── Catálogo de soportes ADRES (Resolución 2284/2023) ───────────────────────
# TOKEN reconocido en el nombre del archivo → descripción del soporte.
# Mantiene los mismos códigos que tools/adres/inspeccionar_soportes.py y
# app/services/soportes_autodiscovery_service.py.
CODIGOS_SOPORTE: dict[str, str] = {
    "FEV": "Factura electrónica de venta (XML DIAN)",
    "FAC": "Factura de venta (PDF)",
    "FAT": "Factura / detalle de cobro",
    "FED": "Factura electrónica DIAN — documento adjunto (AttachedDocument/CUFE)",
    "ARD": "Acuse de la DIAN (ApplicationResponse)",
    "RIP": "RIPS — registro individual de prestación de servicios",
    "CUV": "Resultado de validación RIPS (MSPS / FEVRIPS)",
    "EPI": "Epicrisis",
    "HEV": "Resumen de atención u hoja de evolución (historia clínica)",
    "HAU": "Hoja de atención de urgencias",
    "HAO": "Hoja de atención odontológica",
    "HAM": "Hoja de administración de medicamentos",
    "OPF": "Orden o prescripción facultativa",
    "PDX": "Resultado de procedimientos de apoyo diagnóstico",
    "PDE": "Resultado de procedimientos / evolución",
    "DQX": "Descripción quirúrgica",
    "RAN": "Registro de anestesia",
    "CRC": "Comprobante de recibido del usuario / a cobro",
    "TAP": "Hoja de traslado asistencial de pacientes",
    "FMO": "Factura de material de osteosíntesis e insumos",
    "LDP": "Lista de precios",
    "FUR": "Formato Único de Reclamación (SOAT / ADRES)",
    "SER": "FUR Servicios (detalle de servicios reclamados)",
    "IPO": "Informe policial de accidente / aviso a autoridad",
    "CRA": "Constancia de atención / certificación",
    "VAL": "Validación RIPS — envío al validador (Doker), auxiliar",
    "ADM": "Documento administrativo / otros soportes",
}

# Tokens equivalentes que aparecen en los nombres reales del HUS → código ADRES.
_ALIAS_TOKEN: dict[str, str] = {
    "RIPS": "RIP",
    "FACTURA": "FEV",
    "EPICRIS": "EPI",
    "EPICRISIS": "EPI",
    "INFOPOL": "IPO",
    "FURIPS": "FUR",
    "RESULTADOSMSPS": "CUV",
    "RESULTADOSDOKER": "CUV",  # resultado de validación vía Doker (= CUV)
    "ENVIODOKER": "VAL",  # envío al validador Doker (auxiliar, no es soporte)
    "AD": "FEV",  # XML CUFE de la DIAN (ad0900...xml)
    "FACOSTE": "FAT",
}

# Tabla única token → código (códigos ADRES + alias) para resolver cada token
# con UNA búsqueda de dict. Los alias no colisionan con los códigos.
_TOKEN_A_CODIGO: dict[str, str] = {**{k: k for k in CODIGOS_SOPORTE}, **_ALIAS_TOKEN}

# Soportes esperados por tipo de servicio del RIPS (Res. 2284/2023). Defaults;
# el JSON de perfiles los puede sobrescribir.
SOPORTES_BASE_DEFAULT = ["FEV", "RIP", "CUV"]
SOPORTES_POR_SERVICIO_DEFAULT: dict[str, list[str]] = {
    "consultas": [],
    "procedimientos": ["HEV"],
    "urgencias": ["HAU"],
    "hospitalizacion": ["EPI"],
    "medicamentos": ["OPF"],
    "otrosServicios": ["OPF"],
    "recienNacidos": ["EPI"],
}

# Equivalencias de soporte: un documento "paraguas" satisface una exigencia más
# específica. En el HUS la epicrisis (EPI) y la atención de urgencias (HAU) se
# escanean DENTRO de la historia clínica / hoja de evolución (HEV) — el
# indexador de soportes de la app sólo reconoce HEV, no EPI ni HAU—, por eso un
# HEV presente cubre la exigencia de EPI y HAU. El JSON de perfiles lo puede
# sobrescribir (clave "equivalencias"); con {} se exige el código exacto.
EQUIVALENCIAS_SOPORTE_DEFAULT: dict[str, list[str]] = {
    "EPI": ["EPI", "HEV"],
    "HAU": ["HAU", "HEV"],
}

PRESTADOR_NIT_DEFAULT = "900006037"
PATRON_FACTURA_DEFAULT = r"HUS\d{4,12}"
# Versión del motor que queda registrada en cada corrida del Expediente Digital
# (HOSPIAI): permite saber qué código auditó cada factura y reproducir corridas.
VERSION_MOTOR = "1.0-fase1"

# Estados de salida (ordenados de peor a mejor para el resumen).
ESTADOS_PROBLEMA = (
    "ENTIDAD_NO_RESUELTA",
    "SIN_RIPS",
    "SIN_FEV",
    "SIN_CUV",
    "FALTAN_SOPORTES",
    "FACTURA_INCONSISTENTE",
    "REVISAR_TIPIFICACION",
)
ESTADOS_OK = ("LISTA",)


# ─── Normalización ───────────────────────────────────────────────────────────


def _norm(s: str | None) -> str:
    """Mayúsculas, sin tildes, espacios colapsados."""
    s = (s or "").strip().upper()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    return " ".join(s.split())


_RE_INVALIDOS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def sanitizar(nombre: str) -> str:
    return _RE_INVALIDOS.sub("_", nombre).strip()


_RE_HUS_CORTO = re.compile(r"^(HUS)0*(\d+)$", re.IGNORECASE)


def factura_corta(fac: str) -> str:
    """HUS0000409621 → HUS409621 (quita ceros de relleno). Algunos portales,
    como SIMED del Dispensario, rechazan el formato largo."""
    m = _RE_HUS_CORTO.match((fac or "").strip())
    return (m.group(1).upper() + m.group(2)) if m else fac


def normalizar_factura(fac: str) -> str:
    """Clave canónica para comparar facturas: HUS0000487523, HUS487523 y
    487523 colapsan a 487523 (mismo criterio que soportes_autodiscovery).
    Sin regex: se llama cientos de miles de veces al indexar los shares."""
    f = (fac or "").strip().upper()
    if f.startswith("HUS"):
        f = f[3:]
    f = f.lstrip("0")
    return f or "0"


# ─── Tipificación de soportes ────────────────────────────────────────────────


# Prefijos que en la realidad vienen pegados a dígitos (sin separador), típicos
# de los archivos compartidos por lote: FURIPS168001..., ResultadosMSPS_123,
# EPICRISIS123. Se evalúan por prefijo del nombre, del más largo al más corto
# para no confundir FURIPS con FUR ni RESULTADOSMSPS con RIPS.
_PREFIJOS_SOPORTE: tuple[tuple[str, str], ...] = (
    ("RESULTADOSMSPS", "CUV"),
    ("RESULTADOSDOKER", "CUV"),
    ("ENVIODOKER", "VAL"),
    ("FURIPS", "FUR"),
    ("EPICRISIS", "EPI"),
    ("EPICRIS", "EPI"),
    ("INFOPOL", "IPO"),
    ("FACOSTE", "FAT"),
    ("FACTURA", "FEV"),
)

# El RIPS "desnudo" del share real viene nombrado SOLO con el número de factura
# (HUS469401.json), sin token de soporte. Reconocemos ese patrón para no perder
# el RIPS ni reportarlo como SIN_RIPS. CUV_HUS469401.json no entra aquí porque
# se resuelve antes por el token CUV.
_RE_FACTURA_DESNUDA = re.compile(r"^(?:HUS)?\d{4,12}$", re.IGNORECASE)

# Factura electrónica DIAN: prefijo de 2 letras pegado al CUFE/consecutivo, como
# los del share factura_electronica_net22:
#   fv… = factura de venta (xml = FEV, pdf = representación gráfica)
#   ad… = AttachedDocument firmado con CUFE (la factura electrónica a radicar)
#   ar… = ApplicationResponse (acuse de la DIAN)
_RE_DIAN = re.compile(r"^(fv|ad|ar)\d{6,}", re.IGNORECASE)

# Separadores de tokens en los nombres de archivo (FEV_900006037_HUS487523).
_RE_SEP_TOKEN = re.compile(r"[ _\-.]+")


def clasificar_soporte(nombre: str) -> tuple[str, str, bool]:
    """Devuelve (codigo_adres, descripcion, reconocido) según el TOKEN del
    nombre del archivo. Recorre los tokens de atrás hacia adelante porque el
    tipo suele ir al final (FEV_900006037_HUS487523.pdf → FEV). Si no hay token
    separado, intenta por prefijo (FURIPS168...txt → FUR), por nombre DIAN
    (fv…/ad…/ar…) y, en último lugar, reconoce el RIPS desnudo (HUS469401.json)
    o el paquete FE comprimido (HUS520769.zip)."""
    raiz, punto, ext = nombre.rpartition(".")
    base = raiz if punto else nombre
    ext = ext.lower() if punto else ""
    for t in reversed(_RE_SEP_TOKEN.split(base)):
        cod = _TOKEN_A_CODIGO.get(t.upper())
        if cod is not None:
            return cod, CODIGOS_SOPORTE[cod], True
    base_up = base.upper()
    for pref, cod in _PREFIJOS_SOPORTE:
        if base_up.startswith(pref):
            return cod, CODIGOS_SOPORTE[cod], True
    # Factura electrónica DIAN nombrada por CUFE (fv…/ad…/ar… + dígitos).
    m = _RE_DIAN.match(base)  # el patrón ya es IGNORECASE
    if m:
        pref = m.group(1).lower()
        if pref == "fv":
            cod = "FEV" if ext == "xml" else "FAC"
        elif pref == "ad":
            cod = "FED"
        else:  # ar
            cod = "ARD"
        return cod, CODIGOS_SOPORTE[cod], True
    # Nombres "desnudos" (solo la factura): .json = RIPS, .zip = paquete FE.
    if _RE_FACTURA_DESNUDA.match(base):
        if ext == "json":
            return "RIP", CODIGOS_SOPORTE["RIP"], True
        if ext == "zip":
            return "FED", CODIGOS_SOPORTE["FED"], True
    return "ADM", "Sin clasificar (¿documento administrativo o nombre no ADRES?)", False


# ─── Perfiles de entidad ─────────────────────────────────────────────────────


@dataclass
class PerfilEntidad:
    id: str
    nombre: str
    alias: list[str] = field(default_factory=list)
    nit: str = ""
    eps_codigo: str = ""
    regimen: str = ""
    canal: str = ""
    portal: str = ""
    nomenclatura: str = "ADRES"
    cuv_obligatorio: bool = True
    soportes_extra: list[str] = field(default_factory=list)
    formato_paquete: str = "CARPETA"
    bot: str = ""
    observaciones: str = ""
    # Catálogo institucional (HOSPIAI Fase 1.5): la ficha operativa del pagador.
    # Los valores los suministra el área desde los contratos; 0/"" = sin definir.
    plazo_radicacion_dias: int = 0
    periodicidad: str = ""  # DIARIA / SEMANAL / QUINCENAL / MENSUAL / EVENTUAL
    tipo_radicacion: str = ""  # p.ej. INICIAL, COMPLEMENTARIA, RESPUESTA_GLOSA


@dataclass
class ConfigRadicacion:
    prestador_nombre: str = "ESE HOSPITAL UNIVERSITARIO DE SANTANDER"
    prestador_nit: str = PRESTADOR_NIT_DEFAULT
    patron_factura: str = PATRON_FACTURA_DEFAULT
    soportes_base: list[str] = field(default_factory=lambda: list(SOPORTES_BASE_DEFAULT))
    soportes_por_servicio: dict[str, list[str]] = field(
        default_factory=lambda: {k: list(v) for k, v in SOPORTES_POR_SERVICIO_DEFAULT.items()}
    )
    equivalencias_soporte: dict[str, list[str]] = field(
        default_factory=lambda: {k: list(v) for k, v in EQUIVALENCIAS_SOPORTE_DEFAULT.items()}
    )
    entidades: list[PerfilEntidad] = field(default_factory=list)
    # Trazabilidad HOSPIAI: versión de las reglas cargadas (data/
    # reglas_radicacion.json) y mapa código exigido → id de la regla que lo
    # sustenta, para que cada hallazgo del Expediente Digital cite su regla.
    version_reglas: str = "defaults-embebidos"
    regla_id_por_codigo: dict[str, str] = field(default_factory=dict, repr=False, compare=False)
    reglas_registro: list = field(default_factory=list, repr=False, compare=False)
    # Cache interno de resolver_entidad {pista: perfil|None}: en un lote las
    # mismas pistas (carpeta de EPS, razón social del FEV) se repiten miles de
    # veces y cada resolución escanea todo el catálogo normalizando alias.
    _cache_resolver: dict = field(default_factory=dict, repr=False, compare=False)


def cargar_perfiles(ruta: Path | None) -> ConfigRadicacion:
    """Carga el JSON de perfiles. Si no se pasa ruta, busca
    data/perfiles_radicacion.json relativo al repo. Si no existe, usa los
    defaults embebidos (sin entidades → todo cae a SIN_PERFIL)."""
    if ruta is None:
        candidato = Path(__file__).resolve().parent.parent / "data" / "perfiles_radicacion.json"
        ruta = candidato if candidato.is_file() else None
    if ruta is None:
        logger.warning(
            "No se encontró data/perfiles_radicacion.json — uso defaults sin catálogo de entidades."
        )
        return ConfigRadicacion()
    if not ruta.is_file():
        raise FileNotFoundError(f"No existe el archivo de perfiles: {ruta}")

    data = json.loads(ruta.read_text(encoding="utf-8"))
    prest = data.get("prestador") or {}
    sop = data.get("soportes") or {}
    cfg = ConfigRadicacion(
        prestador_nombre=prest.get("nombre", "ESE HOSPITAL UNIVERSITARIO DE SANTANDER"),
        prestador_nit=str(prest.get("nit", PRESTADOR_NIT_DEFAULT)),
        patron_factura=prest.get("patron_factura", PATRON_FACTURA_DEFAULT),
        soportes_base=list(sop.get("base_obligatorios", SOPORTES_BASE_DEFAULT)),
        soportes_por_servicio={
            k: list(v)
            for k, v in (sop.get("por_servicio") or SOPORTES_POR_SERVICIO_DEFAULT).items()
        },
        equivalencias_soporte={
            k: list(v)
            for k, v in (
                sop["equivalencias"]
                if isinstance(sop.get("equivalencias"), dict)
                else EQUIVALENCIAS_SOPORTE_DEFAULT
            ).items()
        },
    )
    for e in data.get("entidades", []):
        cfg.entidades.append(
            PerfilEntidad(
                id=e.get("id") or _norm(e.get("nombre", "")).replace(" ", "_"),
                nombre=e.get("nombre", ""),
                alias=list(e.get("alias", [])),
                nit=str(e.get("nit", "")),
                eps_codigo=str(e.get("eps_codigo", "")),
                regimen=e.get("regimen", ""),
                canal=e.get("canal", ""),
                portal=e.get("portal", ""),
                nomenclatura=e.get("nomenclatura", "ADRES"),
                cuv_obligatorio=bool(e.get("cuv_obligatorio", True)),
                soportes_extra=list(e.get("soportes_extra", [])),
                formato_paquete=e.get("formato_paquete", "CARPETA"),
                bot=e.get("bot", ""),
                observaciones=e.get("observaciones", ""),
                plazo_radicacion_dias=int(e.get("plazo_radicacion_dias", 0) or 0),
                periodicidad=e.get("periodicidad", ""),
                tipo_radicacion=e.get("tipo_radicacion", ""),
            )
        )
    logger.info(f"Perfiles cargados: {len(cfg.entidades)} entidades desde {ruta.name}")
    return cfg


def cargar_reglas(
    ruta: Path | None, cfg: ConfigRadicacion, fecha_referencia: str | None = None
) -> None:
    """Carga el motor de reglas DECLARATIVO (data/reglas_radicacion.json) sobre
    la configuración. Las reglas dejan de vivir en el código: el área puede
    editarlas, cada una lleva su fuente normativa, y la corrida registra la
    versión usada (HOSPIAI Fase 1; formato en docs/ARQUITECTURA_HOSPIAI.md §6).

    Versionamiento del conocimiento (Fase 1.5): cada regla puede llevar
    `vigencia_desde` / `vigencia_hasta` (AAAA-MM-DD). Solo las reglas VIGENTES a
    `fecha_referencia` (def.: hoy) entran a la validación — si mañana cambia la
    norma, conviven ambas versiones y aplica la correcta según la fecha. Todas
    quedan igual en el registro para el Expediente Digital.

    Si el archivo no existe, quedan los defaults embebidos (mismo contenido).
    Los perfiles por entidad (soportes_extra, cuv_obligatorio, equivalencias
    propias) siguen aplicando por encima, como siempre."""
    if ruta is None:
        cand = Path(__file__).resolve().parent.parent / "data" / "reglas_radicacion.json"
        ruta = cand if cand.is_file() else None
    if ruta is None:
        return
    if not ruta.is_file():
        logger.warning(f"--reglas no existe: {ruta} — sigo con los defaults embebidos.")
        return
    try:
        data = json.loads(ruta.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError) as e:
        logger.warning(f"No pude leer las reglas de {ruta}: {e} — sigo con los defaults.")
        return

    # Vigencia: comparación lexicográfica de fechas ISO (AAAA-MM-DD).
    fecha = fecha_referencia or datetime.now().strftime("%Y-%m-%d")
    base: list[str] = []
    por_servicio: dict[str, list[str]] = {}
    idx: dict[str, str] = {}
    registro: list = []
    n_fuera = 0
    for rg in data.get("reglas", []):
        registro.append(rg)  # TODAS quedan en el registro (memoria del expediente)
        vig_d = str(rg.get("vigencia_desde") or "")
        vig_h = str(rg.get("vigencia_hasta") or "")
        if (vig_d and fecha < vig_d) or (vig_h and fecha > vig_h):
            n_fuera += 1
            continue  # regla no vigente a la fecha: no valida, pero se conoce
        ambito = rg.get("ambito") or {}
        exige = [
            str(c).upper() for c in (rg.get("exige") or rg.get("documentos_obligatorios") or [])
        ]
        if ambito.get("aplica") == "todas":
            base = exige
        serv = ambito.get("servicio")
        if serv:
            dest = por_servicio.setdefault(serv, [])
            dest.extend(c for c in exige if c not in dest)
        for cod in exige:
            idx.setdefault(cod, rg.get("id", ""))

    if base:
        cfg.soportes_base = list(base)
    if por_servicio:
        cfg.soportes_por_servicio = {k: list(v) for k, v in por_servicio.items()}
    eq = data.get("equivalencias") or {}
    eq = {k: list(v) for k, v in eq.items() if not k.startswith("_")}
    if eq:
        cfg.equivalencias_soporte = eq
    cfg.version_reglas = str(data.get("version", "?"))
    cfg.regla_id_por_codigo = idx
    cfg.reglas_registro = registro
    vigencia_txt = f" ({n_fuera} fuera de vigencia al {fecha})" if n_fuera else ""
    logger.info(
        f"Reglas: v{cfg.version_reglas} — {len(registro) - n_fuera} regla(s) vigentes"
        f" desde {ruta.name}{vigencia_txt}"
    )


def resolver_entidad(hint: str | None, cfg: ConfigRadicacion) -> PerfilEntidad | None:
    """Resuelve un nombre/código de entidad contra el catálogo. Tolera código
    EPS embebido ('U220311 - DISPENSARIO ...'), tildes y truncamientos.
    Cachea por pista: en un lote las mismas pistas se repiten miles de veces."""
    if not hint:
        return None
    cache = cfg._cache_resolver
    if hint not in cache:
        cache[hint] = _resolver_entidad(hint, cfg)
    return cache[hint]


# Prefijo de código EPS estilo DGH: "U220311 - NOMBRE".
_RE_COD_EPS = re.compile(r"^\s*([A-Za-z]\d{5,})\s*[-–:·]\s*(.+)$")


def _resolver_entidad(hint: str, cfg: ConfigRadicacion) -> PerfilEntidad | None:
    h = hint.strip()
    # Quitar prefijo de código EPS estilo DGH: "U220311 - NOMBRE".
    m = _RE_COD_EPS.match(h)
    cod = m.group(1).upper() if m else ""
    if m:
        h = m.group(2)
    hn = _norm(h)
    if not hn:
        return None

    mejor: PerfilEntidad | None = None
    mejor_len = -1
    for ent in cfg.entidades:
        # Match por código EPS.
        if cod and ent.eps_codigo and cod == ent.eps_codigo.upper():
            return ent
        for a in [ent.nombre, *ent.alias]:
            an = _norm(a)
            if not an:
                continue
            if an == hn:
                return ent
            # Contención en cualquier dirección, gana el alias más largo (más
            # específico) para no confundir 'SALUD TOTAL' con 'SALUD'.
            if (an in hn or hn in an) and len(an) > mejor_len:
                mejor, mejor_len = ent, len(an)
    return mejor


# ─── Lectura de RIPS / FEV ───────────────────────────────────────────────────

_RE_CBC_ID = re.compile(r"<cbc:ID>([^<]+)</cbc:ID>")


@lru_cache(maxsize=8)
def _texto_fev(ruta: Path) -> str:
    """Texto del XML de la factura ('' si no se puede leer). Con caché chico:
    peek_num_factura_fev y peek_adquiriente_fev suelen leer el MISMO archivo en
    la misma factura, y así el segundo viaje a la red no se paga."""
    try:
        return ruta.read_text(encoding="utf-8", errors="ignore")
    except OSError:
        return ""


def peek_num_factura_fev(ruta: Path) -> str:
    """Número de factura del FEV XML (best-effort, tolera el wrapping DIAN)."""
    m = _RE_CBC_ID.search(_texto_fev(ruta))
    return m.group(1).strip() if m else ""


# Adquiriente (EPS) del AccountingCustomerParty del FEV. Permite resolver la
# entidad cuando la ruta no la trae (share de factura electrónica). Funciona con
# el Invoice directo y con el Invoice embebido (CDATA o escapado) de un
# AttachedDocument DIAN.
_RE_CLIENTE_REG = re.compile(
    r"AccountingCustomerParty.*?RegistrationName[^>]*>\s*([^<]+?)\s*<",
    re.IGNORECASE | re.DOTALL,
)
_RE_CLIENTE_NOM = re.compile(
    r"AccountingCustomerParty.*?PartyName.*?:Name[^>]*>\s*([^<]+?)\s*<",
    re.IGNORECASE | re.DOTALL,
)


def peek_eps_fev(ruta: Path) -> str:
    """Nombre del adquiriente (EPS) leído de la factura electrónica."""
    return peek_adquiriente_fev(ruta)[0]


# Tipo de adquiriente DIAN: AdditionalAccountID 1 = persona jurídica, 2 = persona
# natural. Sirve para marcar PARTICULAR (paciente self-pay) en lugar de
# ENTIDAD_NO_RESUELTA cuando la factura va a nombre de una persona.
_RE_ADQ_TIPO = re.compile(
    r"AccountingCustomerParty.*?AdditionalAccountID[^>]*>\s*([12])",
    re.IGNORECASE | re.DOTALL,
)
# Palabras (token completo) que delatan una persona JURÍDICA, no un particular.
_PISTAS_JURIDICA = {
    "EPS",
    "IPS",
    "ESE",
    "SA",
    "SAS",
    "LTDA",
    "SEGUROS",
    "ASEGURADORA",
    "FUNDACION",
    "INSTITUTO",
    "COMPANIA",
    "FONDO",
    "COOPERATIVA",
    "ASOCIACION",
    "TEMPORAL",
    "UT",
    "PATRIMONIO",
    "PATRIMONIOS",
    "FIDUCIARIA",
    "SECRETARIA",
    "GOBERNACION",
    "REGIONAL",
    "ADMINISTRADORA",
    "DEPARTAMENTO",
    "MUNICIPIO",
    "NACION",
    "SOCIEDAD",
    "MINISTERIO",
    "DIRECCION",
    "EMPRESA",
    "SALUD",
    "ENTIDAD",
    "POSITIVA",
    "PREVISORA",
    "SURAMERICANA",
    "COLPATRIA",
    "SANITAS",
    "COMPENSAR",
    "MUNDIAL",
    "EQUIDAD",
    "PROTEGER",
    "ICBF",
}


def _es_persona_natural(texto: str, nombre: str) -> bool:
    """Decide si el adquiriente es persona natural (particular). Primero por el
    indicador DIAN (AdditionalAccountID=2); si falta, por el nombre."""
    m = _RE_ADQ_TIPO.search(texto)
    if m:
        return m.group(1) == "2"
    palabras = [p.replace(".", "") for p in _norm(nombre).split()]
    if not palabras or any(p in _PISTAS_JURIDICA for p in palabras):
        return False
    return 2 <= len(palabras) <= 5 and all(p.isalpha() for p in palabras)


def peek_adquiriente_fev(ruta: Path) -> tuple[str, bool]:
    """(nombre del adquiriente, es_persona_natural) de la factura electrónica.
    Funciona con el Invoice directo y con el embebido (CDATA/escapado) de un AD."""
    texto = _texto_fev(ruta)
    if not texto:
        return "", False
    m = _RE_CLIENTE_REG.search(texto) or _RE_CLIENTE_NOM.search(texto)
    if (not m or "AdditionalAccountID" not in texto) and "&lt;" in texto:
        texto = html.unescape(texto)  # Invoice embebido y escapado en un AD.
        m = _RE_CLIENTE_REG.search(texto) or _RE_CLIENTE_NOM.search(texto)
    nombre = m.group(1).strip() if m else ""
    return nombre, _es_persona_natural(texto, nombre)


def _leer_rips(ruta: Path) -> dict | None:
    if not _ADRES_OK:
        try:
            return json.loads(ruta.read_text(encoding="utf-8-sig"))
        except (OSError, json.JSONDecodeError):
            return None
    try:
        return cargar_rips(ruta)
    except (ValueError, OSError):
        return None


# ─── Inventario de una factura ───────────────────────────────────────────────


@dataclass
class ResultadoFactura:
    factura: str = ""
    factura_norm: str = ""
    carpeta: str = ""
    entidad_hint: str = ""
    entidad_id: str = ""
    entidad_nombre: str = "SIN_PERFIL"
    canal: str = ""
    nit_prestador: str = ""
    num_usuarios: int = 0
    valor_total: float = 0.0
    servicios: dict[str, int] = field(default_factory=dict)
    soportes_presentes: list[str] = field(default_factory=list)
    soportes_faltantes: list[str] = field(default_factory=list)
    soportes_esperados_faltantes: list[str] = field(default_factory=list)
    archivos_sin_clasificar: list[str] = field(default_factory=list)
    # Rutas de los soportes clínicos anexados desde el share --soportes (gap-fill).
    soportes_clinicos: list[str] = field(default_factory=list)
    # Datos operativos leídos de la RUTA (HOSPIAI D5): quién radica y en qué lote.
    responsable: str = ""
    lote: str = ""
    anio: int = 0
    mes: int = 0
    n_archivos: int = 0
    estado: str = ""
    detalle: list[str] = field(default_factory=list)
    ruta_paquete: str = ""
    acciones: list[str] = field(default_factory=list)


def _archivos_de(carpeta: Path) -> list[Path]:
    """Archivos directos de la carpeta. os.scandir evita un stat por entrada
    (en un share de red el listado ya trae el tipo de cada entrada)."""
    if not carpeta.is_dir():
        return []
    archivos: list[Path] = []
    with os.scandir(carpeta) as it:
        for e in it:
            try:
                if e.is_file():
                    archivos.append(carpeta / e.name)
            except OSError:  # entrada ilegible: mismo criterio que Path.is_file()
                continue
    return sorted(archivos)


def procesar_factura(
    factura_hint: str,
    archivos: list[Path],
    carpeta: Path,
    entidad_hint: str | None,
    cfg: ConfigRadicacion,
    soportes_idx: dict[str, list[Path]] | None = None,
) -> ResultadoFactura:
    """Diagnóstico de UNA factura: tipifica soportes, lee RIPS/FEV, resuelve
    entidad y valida completitud. NO toca archivos.

    `soportes_idx` (opcional) es el índice del share de soportes CLÍNICOS
    {factura_norm: [rutas]}; si se pasa, los documentos que faltan en la carpeta
    de la factura (epicrisis, evolución, órdenes…) se rellenan desde ahí."""
    res = ResultadoFactura(carpeta=str(carpeta), entidad_hint=entidad_hint or "")
    res.n_archivos = len(archivos)

    # ── Tipificación de soportes ───────────────────────────────────────────
    presentes: set[str] = set()
    por_codigo: dict[str, list[Path]] = defaultdict(list)
    rips_path: Path | None = None
    fev_path: Path | None = None
    fed_path: Path | None = None
    for p in archivos:
        cod, _desc, reconocido = clasificar_soporte(p.name)
        por_codigo[cod].append(p)
        if reconocido:
            presentes.add(cod)
        else:
            res.archivos_sin_clasificar.append(p.name)
        sfx = p.suffix.lower()
        # El RIPS es el .json tipificado como RIP (incluye el desnudo
        # HUS469401.json). El CUV es otro .json y NO debe tomarse como RIPS.
        if sfx == ".json" and cod == "RIP" and rips_path is None:
            rips_path = p
        elif sfx == ".xml" and cod == "FEV" and fev_path is None:
            fev_path = p
        elif sfx == ".xml" and cod == "FED" and fed_path is None:
            fed_path = p
    # Preferir el Invoice (fv…, FEV) sobre el AttachedDocument (ad…, FED): el AD
    # lleva su PROPIO cbc:ID (no el número de la factura), lo que provocaba
    # falsos FACTURA_INCONSISTENTE. El AD sólo se usa si no hay Invoice.
    fev_es_invoice = fev_path is not None
    if fev_path is None:
        fev_path = fed_path
    # El documento adjunto DIAN (FED, ad…xml o el .zip) ES la factura
    # electrónica: cuenta como FEV para la completitud.
    if "FED" in presentes:
        presentes.add("FEV")
    # Fallback FEV: si ningún XML quedó tipificado como FEV/FED (nombre atípico),
    # tomar el primer XML que no sea la factura de costo (FACOSTE/FAT) ni un
    # acuse de la DIAN (ApplicationResponse).
    if fev_path is None:
        for p in archivos:
            if (
                p.suffix.lower() == ".xml"
                and p not in por_codigo["FAT"]
                and p not in por_codigo["ARD"]
            ):
                fev_path = p
                break
    res.soportes_presentes = sorted(presentes)

    # ── Datos del RIPS ──────────────────────────────────────────────────────
    num_rips = ""
    if rips_path is not None:
        data = _leer_rips(rips_path)
        if data is not None:
            gen = datos_generales(data) if _ADRES_OK else {}
            num_rips = (gen.get("num_factura") if _ADRES_OK else data.get("numFactura")) or ""
            res.nit_prestador = (
                gen.get("nit_prestador") if _ADRES_OK else data.get("numDocumentoIdObligado")
            ) or ""
            res.num_usuarios = (
                gen.get("num_usuarios") if _ADRES_OK else len(data.get("usuarios") or [])
            )
            res.servicios = _contar_servicios(data)
            res.valor_total = _valor_total(data)
    res.nit_prestador = res.nit_prestador or cfg.prestador_nit

    num_fev = peek_num_factura_fev(fev_path) if fev_path is not None else ""

    # ── Número de factura definitivo ────────────────────────────────────────
    res.factura = num_rips or factura_hint or num_fev or carpeta.name
    res.factura_norm = normalizar_factura(res.factura)

    # ── Resolución de entidad ───────────────────────────────────────────────
    perfil = resolver_entidad(entidad_hint, cfg)
    es_particular = False
    if perfil is None:
        # Sin pista por ruta/manifiesto (típico del share de factura
        # electrónica): leer el adquiriente de la propia FEV. Se prueba primero
        # el Invoice (fv…) y luego el AttachedDocument (ad…).
        eps_detectada = ""
        for cand in (*por_codigo.get("FEV", []), *por_codigo.get("FED", [])):
            if cand.suffix.lower() != ".xml":
                continue
            nombre_adq, persona = peek_adquiriente_fev(cand)
            if nombre_adq and not eps_detectada:
                eps_detectada = nombre_adq
            if persona:
                es_particular = True
            perfil = resolver_entidad(nombre_adq, cfg)
            if perfil is not None:
                eps_detectada = nombre_adq
                es_particular = False
                break
        # Guardar el nombre leído del FEV (aunque NO matchee el catálogo) para
        # diagnóstico: aparece en el detalle de las ENTIDAD_NO_RESUELTA.
        if eps_detectada and not res.entidad_hint:
            res.entidad_hint = eps_detectada
    if perfil is not None:
        res.entidad_id = perfil.id
        res.entidad_nombre = perfil.nombre
        res.canal = perfil.canal
    elif es_particular:
        res.entidad_id = "PARTICULAR"
        res.entidad_nombre = "PARTICULAR (paciente)"

    # ── Cruce con el share de soportes clínicos (si --soportes) ─────────────
    # Rellena SOLO los huecos: anexa los documentos clínicos (epicrisis,
    # evolución, órdenes, resultados…) cuyo código no venía ya en la carpeta de
    # la factura. No duplica FEV/RIP/CUV que el share de FE ya trae.
    if soportes_idx is not None:
        codigos_fe = set(presentes)
        vistos: set[str] = set()
        for sp in soportes_idx.get(res.factura_norm, []):
            ruta_sp = str(sp)
            # El mismo archivo puede venir dos veces (--soportes-indice y
            # --soportes sobre el mismo share, o listados concatenados): no se
            # anexa —ni luego se copia— duplicado.
            if ruta_sp in vistos:
                continue
            vistos.add(ruta_sp)
            cod, _desc, reconocido = clasificar_soporte(sp.name)
            if not reconocido or cod in codigos_fe:
                continue
            presentes.add(cod)
            por_codigo[cod].append(sp)
            res.soportes_clinicos.append(ruta_sp)
        res.soportes_presentes = sorted(presentes)

    # ── Ruta operativa: responsable, lote, año y mes (HOSPIAI D5) ───────────
    # Primero la carpeta del lote; si ahí no está (el share de FE no trae
    # funcionario), se completa desde las rutas de los soportes cruzados
    # (Y:\…\<EPS>\<RESPONSABLE>\ENV-…\…).
    info_ruta = analizar_ruta(str(carpeta))
    if res.soportes_clinicos and (not info_ruta.get("responsable") or not info_ruta.get("lote")):
        for sp in res.soportes_clinicos:
            extra = analizar_ruta(sp)
            for k, v in extra.items():
                info_ruta.setdefault(k, v)
            if info_ruta.get("responsable") and info_ruta.get("lote"):
                break
    res.responsable = info_ruta.get("responsable", "")
    res.lote = info_ruta.get("lote", "")
    res.anio = info_ruta.get("anio", 0)
    res.mes = info_ruta.get("mes", 0)

    # ── Completitud ─────────────────────────────────────────────────────────
    base = set(cfg.soportes_base)
    if perfil is not None and not perfil.cuv_obligatorio:
        base.discard("CUV")
    if perfil is not None:
        base |= set(perfil.soportes_extra)
    res.soportes_faltantes = sorted(base - presentes)

    esperados: set[str] = set()
    for serv, n in res.servicios.items():
        if n:
            esperados |= set(cfg.soportes_por_servicio.get(serv, []))

    # Un esperado se da por cumplido si está presente él mismo o cualquiera de
    # sus equivalentes (p.ej. la epicrisis EPI la cubre la historia clínica HEV).
    def _satisface(cod: str) -> bool:
        aceptados = set(cfg.equivalencias_soporte.get(cod, [cod])) or {cod}
        return bool(aceptados & presentes)

    res.soportes_esperados_faltantes = sorted(e for e in (esperados - base) if not _satisface(e))

    res.estado, res.detalle = _evaluar_estado(
        res, perfil, num_rips, num_fev, rips_path, fev_path, fev_es_invoice, es_particular
    )
    return res


def _contar_servicios(data: dict) -> dict[str, int]:
    conteo: dict[str, int] = dict.fromkeys(SERVICIOS_RIPS, 0)
    for u in data.get("usuarios") or []:
        servicios = u.get("servicios") or {}
        for tipo in SERVICIOS_RIPS:
            conteo[tipo] += len(servicios.get(tipo) or [])
    return {k: v for k, v in conteo.items() if v}


def _valor_total(data: dict) -> float:
    if _ADRES_OK:
        total = 0.0
        for ln in extraer_lineas_servicios(data):
            try:
                total += float(ln.get("vr_total") or 0)
            except (TypeError, ValueError):
                continue
        return round(total, 2)
    total = 0.0
    for u in data.get("usuarios") or []:
        for items in (u.get("servicios") or {}).values():
            for it in items or []:
                try:
                    total += float(it.get("vrServicio") or 0)
                except (TypeError, ValueError):
                    continue
    return round(total, 2)


def _evaluar_estado(
    res: ResultadoFactura,
    perfil: PerfilEntidad | None,
    num_rips: str,
    num_fev: str,
    rips_path: Path | None,
    fev_path: Path | None,
    fev_es_invoice: bool = True,
    es_particular: bool = False,
) -> tuple[str, list[str]]:
    detalle: list[str] = []

    if perfil is None and es_particular:
        detalle.append(
            f"Adquiriente persona natural (particular): '{res.entidad_hint or '—'}'. "
            "No se radica a una EPS."
        )
    elif perfil is None:
        detalle.append(
            f"No se pudo resolver la entidad (pista: '{res.entidad_hint or '—'}'). "
            "Usá --manifiesto o organizá el lote en carpetas por EPS."
        )

    if rips_path is None:
        detalle.append("Falta el RIPS (.json).")
    if fev_path is None:
        detalle.append("Falta la factura electrónica (FEV .xml).")

    # Nota (NO bloqueante) sobre el número de factura entre RIPS y FEV. La
    # factura electrónica DIAN suele numerarse con otra serie que el RIPS, así
    # que una diferencia es esperable y NO debe impedir radicar; la identidad de
    # la factura la da el RIPS / la carpeta. Solo se compara con un Invoice real
    # (el AttachedDocument lleva su propio cbc:ID) y con números plausibles.
    fuentes = {
        n
        for n in (normalizar_factura(num_rips), normalizar_factura(num_fev))
        if n != "0" and n.isdigit() and len(n) <= 12
    }
    if fev_es_invoice and len(fuentes) > 1:
        detalle.append(
            f"Nota: el número del FEV ({num_fev or '—'}) difiere del RIPS "
            f"({num_rips or '—'}); normal si la FE usa otra numeración."
        )

    if res.soportes_faltantes:
        detalle.append("Faltan soportes obligatorios: " + ", ".join(res.soportes_faltantes) + ".")
    if res.soportes_esperados_faltantes:
        detalle.append(
            "Soportes esperados según los servicios (revisar): "
            + ", ".join(res.soportes_esperados_faltantes)
            + "."
        )
    if res.archivos_sin_clasificar:
        detalle.append(
            f"{len(res.archivos_sin_clasificar)} archivo(s) sin tipificar (renombrar a la "
            "nomenclatura ADRES <TOKEN>_<NIT>_<factura>): "
            + ", ".join(res.archivos_sin_clasificar[:5])
            + ("…" if len(res.archivos_sin_clasificar) > 5 else "")
        )

    # Estado: el primer problema bloqueante manda.
    if perfil is None:
        return ("PARTICULAR" if es_particular else "ENTIDAD_NO_RESUELTA"), detalle
    if rips_path is None:
        return "SIN_RIPS", detalle
    if fev_path is None:
        return "SIN_FEV", detalle
    if "CUV" in res.soportes_faltantes:
        return "SIN_CUV", detalle
    if res.soportes_faltantes:
        return "FALTAN_SOPORTES", detalle
    if res.archivos_sin_clasificar or res.soportes_esperados_faltantes:
        return "REVISAR_TIPIFICACION", detalle
    detalle.append("Lista para radicar.")
    return "LISTA", detalle


def _debe_armar(estado: str, forzar: bool) -> bool:
    if estado == "LISTA":
        return True
    return estado == "REVISAR_TIPIFICACION" and forzar


# ─── Descubrimiento del lote ─────────────────────────────────────────────────


# Carpetas contenedoras del share que NO son entidades: ENV-…, ESCANEO, RIPS,
# SOPORTES, FE/FACTURA…, y prefijos numéricos tipo "1. DD FACTURACION".
_RE_CARPETA_CONTENEDORA = re.compile(
    r"(?i)^(env[-_].*|escaneo|soportes.*|rips|fe|factura.*|xml|pdf|\d+\..*)$"
)

# ── Analizador de Ruta (HOSPIAI D1/D5) ──────────────────────────────────────
# Cada nivel de la ruta del share significa algo:
#   Y:\7. JULIO 2026 - SOPORTES RADICACION\ASMET SALUD\LILIANA\ENV-230314-OKDGH\HUS528043
#        └── mes/año ──┘                    └─ entidad ─┘└ resp ┘└─── lote ────┘└ factura ┘
# De ahí salen el RESPONSABLE (funcionario que radica) y el LOTE (envío), que
# alimentan los indicadores por funcionario (Dominio 5 — Int. Operacional).

_RE_LOTE = re.compile(r"^ENV[-_][\w-]+$", re.IGNORECASE)
# Año suelto (2020–2039) no pegado a otros dígitos (evita leer "2076" de HUS520761).
_RE_ANIO = re.compile(r"(?<!\d)(20[2-3]\d)(?!\d)")
# Año+mes compactos estilo 202606.
_RE_ANIOMES = re.compile(r"(?<!\d)(20[2-3]\d)(0[1-9]|1[0-2])(?!\d)")
_MESES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
    "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}  # fmt: skip
# Tokens que delatan que una carpeta es una ENTIDAD u otra cosa, no un
# funcionario (el nivel responsable es un nombre de persona: KARIN, LILIANA…).
_NO_RESPONSABLE = frozenset(
    "EPS SALUD SEGUROS SEGURO SOAT ARL ADRES SA SAS LTDA IPS ESE UT UNION TEMPORAL "
    "COMPARTIDA FACTURACION RADICACION DIGITAL CARPETA SERVIDOR GLOSAS SINAC OK "
    "NUEVA COOSALUD ASMET FOMAG COMPENSAR SANITAS FAMISANAR DISPENSARIO POSITIVA "
    "PREVISORA MUNDIAL AXA HDI SURA SURAMERICANA SOLIDARIA EQUIDAD PPL REGIONAL".split()
)


def _es_responsable(nombre: str) -> bool:
    """¿La carpeta parece el NOMBRE de un funcionario (no una EPS/contenedora)?"""
    n = nombre.strip()
    if not (2 <= len(n) <= 40) or any(ch.isdigit() for ch in n):
        return False
    if _RE_CARPETA_CONTENEDORA.match(n):
        return False
    tokens = re.split(r"[\s._-]+", n.upper())
    return bool(tokens) and not any(t in _NO_RESPONSABLE for t in tokens)


def analizar_ruta(ruta: str) -> dict:
    """Extrae de una ruta del share: año, mes, lote (ENV-…), responsable y
    factura. Devuelve solo las claves halladas. Tolera rutas Windows y POSIX."""
    partes = [p for p in re.split(r"[\\/]+", str(ruta)) if p and not p.endswith(":")]
    info: dict = {}
    rx_fac = re.compile(PATRON_FACTURA_DEFAULT, re.IGNORECASE)
    for i, parte in enumerate(partes):
        pu = parte.upper()
        if "lote" not in info and _RE_LOTE.match(parte):
            info["lote"] = parte
            if i > 0 and "responsable" not in info and _es_responsable(partes[i - 1]):
                info["responsable"] = partes[i - 1].strip()
        if "anio" not in info:
            m = _RE_ANIOMES.search(pu)
            if m:
                info["anio"], info["mes"] = int(m.group(1)), int(m.group(2))
            else:
                m = _RE_ANIO.search(pu)
                if m:
                    info["anio"] = int(m.group(1))
        if "mes" not in info:
            for nombre_mes, num in _MESES.items():
                if nombre_mes in pu:
                    info["mes"] = num
                    break
        m = rx_fac.search(parte)
        if m:
            info["factura"] = m.group(0)  # la más profunda gana
    return info


def _entidad_desde_ruta(carpeta: Path, origen: Path, cfg: ConfigRadicacion) -> str | None:
    """Busca, de la carpeta hacia arriba hasta origen, un componente de ruta
    que resuelva a una entidad conocida (la carpeta EPS del share)."""
    try:
        rel = carpeta.relative_to(origen)
    except ValueError:
        rel = Path(carpeta.name)
    for parte in reversed(rel.parts):
        if _RE_CARPETA_CONTENEDORA.match(parte):
            continue
        if resolver_entidad(parte, cfg) is not None:
            return parte
    return None


def descubrir_carpeta_factura(
    origen: Path, manifiesto: dict[str, str], cfg: ConfigRadicacion
) -> list[tuple[str, list[Path], Path, str | None]]:
    """Layout carpeta-factura: cada carpeta con archivos directos es una
    factura. Devuelve [(factura_hint, archivos, carpeta, entidad_hint)].

    En el share real una misma factura aparece en árboles PARALELOS (p.ej.
    …/RIPS/ENV-…/HUS469401/ con el RIPS+CUV y …/SOPORTES/ENV-…/HUS469401/ con
    el FEV y los PDFs): dos carpetas con el MISMO nombre de factura. Se FUSIONAN
    por número de factura para no reportar la misma dos veces ni dejarla
    incompleta."""
    # Un solo listado por carpeta (los shares de red cobran cada round-trip).
    carpetas: list[tuple[Path, list[Path]]] = []
    raiz_archivos = _archivos_de(origen)
    if raiz_archivos:
        carpetas.append((origen, raiz_archivos))
    for d in sorted(p for p in origen.rglob("*") if p.is_dir()):
        archivos_d = _archivos_de(d)
        if archivos_d:
            carpetas.append((d, archivos_d))

    fusion: dict[str, list] = {}
    orden: list[str] = []
    for c, archivos in carpetas:
        fac_hint = c.name
        ent = manifiesto.get(normalizar_factura(fac_hint)) or _entidad_desde_ruta(c, origen, cfg)
        # Fusionar solo cuando el nombre de la carpeta ES una factura
        # (HUS469401); si no, cada carpeta es única (clave = su ruta).
        clave = (
            "F:" + normalizar_factura(fac_hint)
            if _RE_FACTURA_DESNUDA.match(fac_hint)
            else "D:" + str(c)
        )
        if clave in fusion:
            slot = fusion[clave]
            ya = {p.name for p in slot[1]}
            slot[1].extend(p for p in archivos if p.name not in ya)
            if not slot[3] and ent:
                slot[3] = ent
        else:
            fusion[clave] = [fac_hint, list(archivos), c, ent]
            orden.append(clave)
    return [tuple(fusion[k]) for k in orden]


def descubrir_lote(
    origen: Path, manifiesto: dict[str, str], cfg: ConfigRadicacion, patron: str
) -> list[tuple[str, list[Path], Path, str | None]]:
    """Layout lote: muchos archivos por carpeta, agrupados por el token
    _<factura>_ del nombre. Los compartidos (sin factura en el nombre, ej.
    FURIPS) se asocian a todas las facturas de su carpeta."""
    rx = re.compile(patron, re.IGNORECASE)
    grupos: dict[str, list[Path]] = defaultdict(list)
    carpeta_de: dict[str, Path] = {}
    compartidos_por_dir: dict[Path, list[Path]] = defaultdict(list)
    facturas_por_dir: dict[Path, set[str]] = defaultdict(set)

    # os.walk en vez de rglob + is_file: no hace un stat por entrada (caro en
    # shares de red enormes); el listado ya separa carpetas de archivos.
    todos: list[Path] = []
    for dirpath, _dirnames, filenames in os.walk(origen):
        d = Path(dirpath)
        todos.extend(d / fn for fn in filenames)
    todos.sort()

    for p in todos:
        m = rx.search(p.name)
        if m:
            clave = normalizar_factura(m.group(0))
            grupos[clave].append(p)
            carpeta_de.setdefault(clave, p.parent)
            facturas_por_dir[p.parent].add(clave)
        else:
            compartidos_por_dir[p.parent].append(p)

    # Repartir compartidos a cada factura de su carpeta.
    for d, comps in compartidos_por_dir.items():
        for clave in facturas_por_dir.get(d, set()):
            grupos[clave].extend(comps)

    salida = []
    for clave, archivos in grupos.items():
        carpeta = carpeta_de.get(clave, origen)
        # Nombre crudo de factura: el token tal como aparece en un archivo.
        fac_raw = clave
        for p in archivos:
            m = rx.search(p.name)
            if m:
                fac_raw = m.group(0)
                break
        ent = manifiesto.get(clave) or _entidad_desde_ruta(carpeta, origen, cfg)
        salida.append((fac_raw, sorted(set(archivos)), carpeta, ent))
    return salida


def descubrir_auto(
    origen: Path, manifiesto: dict[str, str], cfg: ConfigRadicacion, patron: str
) -> list[tuple[str, list[Path], Path, str | None]]:
    """Layout auto (def.): ancla en carpetas cuyo nombre ES una factura
    (HUS520769) y les asigna TODOS los archivos que cuelgan debajo,
    recursivamente — así recoge subcarpetas como RIPS/ y los documentos DIAN
    (fv…/ad…/ar…/.zip) que viven en la carpeta de la factura. Lo que no cae bajo
    ninguna carpeta-factura se agrupa por el token _<factura>_ del nombre
    (estilo lote). Fusiona por número de factura, así la misma factura repartida
    en árboles distintos queda UNA sola. Es el más robusto: cubre el share de FE
    (FACTURAS_SALUD\\HUS…\\), el de ESCANEO y los lotes sueltos."""
    rx = re.compile(patron, re.IGNORECASE)
    # Un solo recorrido del árbol con os.walk (no hace stat por entrada → mucho
    # más rápido en shares de red con miles de facturas que rglob + is_dir).
    # os.walk es top-down (el padre se visita antes que sus hijas), así que el
    # ancla de cada carpeta —la carpeta-factura más cercana hacia arriba— se
    # hereda en el mismo recorrido, sin re-caminar los parents de cada archivo.
    ancla_de_dir: dict[Path, Path | None] = {}
    archivos: list[tuple[Path, Path | None]] = []
    for dirpath, _dirnames, filenames in os.walk(origen):
        d = Path(dirpath)
        if d == origen:
            ancla = None
        elif _RE_FACTURA_DESNUDA.match(d.name):
            ancla = d  # la propia carpeta es una factura (la más cercana)
        else:
            ancla = ancla_de_dir.get(d.parent)  # hereda el ancla del padre
        ancla_de_dir[d] = ancla
        archivos.extend((d / fn, ancla) for fn in filenames)
    archivos.sort(key=lambda t: t[0])

    grupos: dict[str, list[Path]] = defaultdict(list)
    carpeta_de: dict[str, Path] = {}
    fac_raw_de: dict[str, str] = {}
    sueltos: list[Path] = []

    for p, ancla in archivos:
        if ancla is not None:
            clave = normalizar_factura(ancla.name)
            grupos[clave].append(p)
            carpeta_de.setdefault(clave, ancla)
            fac_raw_de.setdefault(clave, ancla.name)
        else:
            sueltos.append(p)

    # Lo no anclado: agrupar por el token del nombre y repartir los compartidos
    # (sin token) a las facturas de su misma carpeta — igual que el layout lote.
    compartidos_por_dir: dict[Path, list[Path]] = defaultdict(list)
    facturas_por_dir: dict[Path, set[str]] = defaultdict(set)
    for p in sueltos:
        m = rx.search(p.name)
        if m:
            clave = normalizar_factura(m.group(0))
            grupos[clave].append(p)
            carpeta_de.setdefault(clave, p.parent)
            fac_raw_de.setdefault(clave, m.group(0))
            facturas_por_dir[p.parent].add(clave)
        else:
            compartidos_por_dir[p.parent].append(p)
    for d, comps in compartidos_por_dir.items():
        for clave in facturas_por_dir.get(d, set()):
            grupos[clave].extend(comps)

    salida = []
    for clave, archivos in grupos.items():
        carpeta = carpeta_de.get(clave, origen)
        ent = manifiesto.get(clave) or _entidad_desde_ruta(carpeta, origen, cfg)
        salida.append((fac_raw_de.get(clave, clave), sorted(set(archivos)), carpeta, ent))
    return salida


# ─── Índice del share de soportes clínicos ───────────────────────────────────


def _clave_factura_en(texto: str, rx: re.Pattern) -> str | None:
    """Número de factura normalizado de la ÚLTIMA coincidencia VÁLIDA de `rx`
    en `texto`.

    En una ruta las coincidencias van de la raíz a la hoja, así que se prueba
    de la última (la más profunda = la más específica: …\\HUS521788\\… → 521788)
    hacia atrás, saltando las que normalizan a '0' (HUS0000 no identifica nada).
    Devuelve None si ninguna coincidencia da clave.

    findall solo devuelve la coincidencia completa si el patrón NO tiene grupos
    de captura; si el perfil trae un patrón con grupos, se cae a finditer +
    group(0) (más lento, pero correcto). Con el patrón por defecto va por la
    vía rápida: esto corre una vez por línea del índice (250k+)."""
    if rx.groups:
        ms = [m.group(0) for m in rx.finditer(texto)]
    else:
        ms = rx.findall(texto)
    for candidato in reversed(ms):
        clave = normalizar_factura(candidato)
        if clave != "0":
            return clave
    return None


def indexar_soportes_clinicos(raiz: Path, patron: str) -> dict[str, list[Path]]:
    """Recorre el share de soportes CLÍNICOS (epicrisis, evolución, urgencias,
    órdenes, resultados de apoyo diagnóstico…) y devuelve {factura_norm: [rutas]}.

    La llave es el número de factura. Se toma primero del NOMBRE del archivo
    (HEV_900006037_HUS487523.pdf → HUS487523); si el archivo no lo lleva —el HUS
    escanea muchos clínicos con nombres genéricos: EPICRISIS.pdf, 1.pdf— se toma
    de la CARPETA contenedora, que en el share va nombrada por factura
    (…\\SOPORTES\\HUS521788\\EPICRISIS.pdf → 521788). Se normaliza igual que el
    resto del motor (quita 'HUS' y ceros de relleno).

    Los documentos de lote sin factura en el nombre NI en una carpeta de factura
    (FURIPS/ACUSE que cuelgan de carpetas de entidad) no se reparten: sin clave,
    se saltan y no contaminan paquetes ajenos."""
    rx = re.compile(patron, re.IGNORECASE)
    indice: dict[str, list[Path]] = defaultdict(list)
    n = 0
    for dirpath, _dirnames, filenames in os.walk(raiz):
        if not filenames:  # nada que indexar: ahorra el escaneo de la ruta
            continue
        d = Path(dirpath)
        clave_carpeta = _clave_factura_en(dirpath, rx)  # factura embebida en la ruta
        for fn in filenames:
            clave = _clave_factura_en(fn, rx) or clave_carpeta
            if clave is None:
                continue
            indice[clave].append(d / fn)
            n += 1
    logger.info(
        f"Soportes clínicos: {n} archivo(s) indexados para {len(indice)} factura(s) desde {raiz}"
    )
    return indice


def indexar_soportes_desde_indice(ruta_indice: Path, patron: str) -> dict[str, list[Path]]:
    """Construye el índice de soportes desde un LISTADO pre-armado de rutas de
    ARCHIVOS (un 'dir /s /b' de los share(s) de soportes, una ruta por línea).

    No recorre la red: el listado ya hizo ese trabajo una vez, así que cubre de
    una TODAS las rutas que tenga adentro (X:, Y:, Z:…) — ideal cuando los
    soportes están repartidos en varios discos y recorrerlos sería lentísimo.
    Clasifica cada archivo por su nombre y lo cruza por número de factura. Las
    líneas que son CARPETAS (terminan en HUS<factura>, sin extensión) no aportan
    archivos: si el listado es de carpetas, regeneralo listando archivos."""
    rx = re.compile(patron, re.IGNORECASE)
    indice: dict[str, list[Path]] = defaultdict(list)
    n_arch = n_carpetas = 0
    # Codificación: la redirección de PowerShell ('dir > f.txt', Out-File) escribe
    # UTF-16 con BOM — leído como utf-8 el listado sale en garabatos y el índice
    # queda en 0 sin aviso. Se detecta el BOM UTF-16 y, si no, utf-8-sig (descarta
    # el BOM de 'Set-Content -Encoding UTF8' y no ensucia la primera ruta).
    with ruta_indice.open("rb") as fb:
        bom = fb.read(2)
    encoding = "utf-16" if bom in (b"\xff\xfe", b"\xfe\xff") else "utf-8-sig"
    # Las líneas de un 'dir /s /b' vienen agrupadas por carpeta: se cachea la
    # clave de la carpeta para no re-escanear la misma ruta por cada archivo.
    carpeta_prev: str | None = None
    clave_carpeta: str | None = None
    with ruta_indice.open(encoding=encoding, errors="replace") as fh:
        for raw in fh:
            linea = raw.strip().strip('"')
            if not linea:
                continue
            carpeta, _sep, base = linea.replace("/", "\\").rpartition("\\")
            if "." not in base:  # es una CARPETA, no un archivo
                if rx.search(base):
                    n_carpetas += 1
                continue
            # Clave por el NOMBRE del archivo; si no la lleva (EPICRISIS.pdf),
            # por la CARPETA contenedora (…\HUS521788\EPICRISIS.pdf → 521788).
            clave = _clave_factura_en(base, rx)
            if clave is None:
                if carpeta != carpeta_prev:
                    carpeta_prev = carpeta
                    clave_carpeta = _clave_factura_en(carpeta, rx)
                clave = clave_carpeta
            if clave is None:
                continue
            _cod, _desc, reconocido = clasificar_soporte(base)
            if reconocido:
                # El listado trae rutas de Windows (X:\…, \\srv\…); PureWindowsPath
                # interpreta '\\' como separador en cualquier SO, así .name y str()
                # quedan correctos (Path nativo en Linux no parte por '\\').
                indice[clave].append(PureWindowsPath(linea))
                n_arch += 1
    msg = (
        f"Índice de soportes: {n_arch} archivo(s) para {len(indice)} factura(s) "
        f"desde {ruta_indice.name}"
    )
    if n_carpetas and not n_arch:
        msg += (
            f" — OJO: las {n_carpetas} línea(s) son CARPETAS, no archivos. "
            "Regenerá el índice listando ARCHIVOS (dir /s /b o Get-ChildItem -File)."
        )
    logger.info(msg)
    return dict(indice)


# ─── Manifiesto factura → entidad ────────────────────────────────────────────


def cargar_manifiesto(ruta: Path) -> dict[str, str]:
    """CSV/XLSX con columnas de factura y entidad. Devuelve {factura_norm:
    entidad}. Reconoce encabezados flexibles (FACTURA, EPS, ENTIDAD, …)."""
    alias_fac = {"FACTURA", "FACTURA HUS", "NUMERO FACTURA", "NRO FACTURA", "NO FACTURA"}
    alias_ent = {"ENTIDAD", "EPS", "PAGADOR", "TERCERO", "ASEGURADORA", "RESPONSABLE"}
    filas = _leer_tabla(ruta)
    if not filas:
        return {}
    headers = [_norm(h) for h in filas[0]]
    i_fac = next((i for i, h in enumerate(headers) if h in alias_fac), None)
    i_ent = next((i for i, h in enumerate(headers) if h in alias_ent), None)
    if i_fac is None or i_ent is None:
        raise ValueError(
            f"--manifiesto: no encontré columnas de factura y entidad en {ruta.name} "
            f"(encabezados: {filas[0]})"
        )
    mapa: dict[str, str] = {}
    for fila in filas[1:]:
        if len(fila) <= max(i_fac, i_ent):
            continue
        fac = normalizar_factura(str(fila[i_fac] or ""))
        ent = str(fila[i_ent] or "").strip()
        if fac != "0" and ent:
            mapa[fac] = ent
    return mapa


def _leer_tabla(ruta: Path) -> list[list]:
    if ruta.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise SystemExit(
                "ERROR: para leer un manifiesto .xlsx falta openpyxl (py -m pip install openpyxl)."
            ) from e
        wb = load_workbook(filename=str(ruta), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        return [list(r) for r in ws.iter_rows(values_only=True)]
    with ruta.open("r", encoding="utf-8-sig", newline="") as fh:
        return list(csv.reader(fh))


# ─── Armado del paquete ──────────────────────────────────────────────────────


def _nombre_unico(nombre: str, usados: set[str]) -> str:
    """Evita pisar archivos cuando dos soportes mapean al mismo nombre destino
    (p.ej. fv…xml y ad…xml de un mismo paquete DIAN): agrega _2, _3, …"""
    if nombre not in usados:
        return nombre
    raiz, punto, ext = nombre.rpartition(".")
    base = raiz if punto else nombre
    sufijo = ("." + ext) if punto else ""
    i = 2
    while f"{base}_{i}{sufijo}" in usados:
        i += 1
    return f"{base}_{i}{sufijo}"


def nombre_soporte(archivo: Path, factura: str, nit_prestador: str, nomenclatura: str) -> str:
    """Nombre destino del soporte según la nomenclatura de la entidad.
    ADRES: <TOKEN>_<NIT>_<factura>.<ext>.  DISPENSARIO_HUS_CORTO: igual pero
    con la factura en formato corto (HUS487523). Archivos ya bien nombrados o
    sin clasificar se dejan con su nombre original."""
    cod, _desc, reconocido = clasificar_soporte(archivo.name)
    ext = archivo.suffix.lower().lstrip(".") or "dat"
    fac = factura_corta(factura) if nomenclatura == "DISPENSARIO_HUS_CORTO" else factura
    if not reconocido:
        return sanitizar(archivo.name)
    return sanitizar(f"{cod}_{nit_prestador}_{fac}.{ext}")


def armar_paquete(
    res: ResultadoFactura,
    archivos: list[Path],
    perfil: PerfilEntidad | None,
    destino: Path,
    fecha: str,
    consecutivo: int,
    mover: bool,
    hacer_zip: bool,
    forzar: bool,
    dry_run: bool,
    nit_prestador: str,
) -> bool:
    """Construye <destino>/<ENTIDAD>/<factura>/ con los soportes renombrados.
    Solo arma estados LISTA o REVISAR_TIPIFICACION (este último requiere
    --forzar). Registra acciones y la ruta del paquete en `res`. Devuelve True
    si se generó (o se generaría) un ZIP, para avanzar el consecutivo."""
    if not _debe_armar(res.estado, forzar):
        res.acciones.append(f"NO_ARMADO ({res.estado})")
        return False

    nomen = perfil.nomenclatura if perfil else "ADRES"
    formato = perfil.formato_paquete if perfil else "CARPETA"
    ent_id = res.entidad_id or "SIN_PERFIL"
    fac = factura_corta(res.factura) if nomen == "DISPENSARIO_HUS_CORTO" else res.factura
    dir_factura = destino / sanitizar(ent_id) / sanitizar(fac)
    res.ruta_paquete = str(dir_factura)

    if not dry_run:
        dir_factura.mkdir(parents=True, exist_ok=True)

    copiados = 0
    usados: set[str] = set()
    for p in archivos:
        nuevo = _nombre_unico(nombre_soporte(p, res.factura, nit_prestador, nomen), usados)
        usados.add(nuevo)
        destino_final = dir_factura / nuevo
        if dry_run:
            copiados += 1
            continue
        try:
            if mover:
                shutil.move(str(p), str(destino_final))
            else:
                shutil.copy2(str(p), str(destino_final))
            copiados += 1
        except OSError as e:
            res.acciones.append(f"ERROR_COPIA {p.name}: {e}")
    res.acciones.append(f"{'movería' if dry_run else 'copió'} {copiados} soporte(s)")

    # Soportes clínicos anexados desde --soportes: SIEMPRE se copian (nunca se
    # mueven: el share clínico es fuente y lo comparten otras facturas/la app).
    clinicos_ok = 0
    for ruta in res.soportes_clinicos:
        p = Path(ruta)
        nuevo = _nombre_unico(nombre_soporte(p, res.factura, nit_prestador, nomen), usados)
        usados.add(nuevo)
        if dry_run:
            clinicos_ok += 1
            continue
        try:
            shutil.copy2(str(p), str(dir_factura / nuevo))
            clinicos_ok += 1
        except OSError as e:
            res.acciones.append(f"ERROR_COPIA_CLINICO {p.name}: {e}")
    if res.soportes_clinicos:
        res.acciones.append(
            f"{'anexaría' if dry_run else 'anexó'} {clinicos_ok} soporte(s) clínico(s)"
        )

    if not dry_run:
        _escribir_manifiesto_factura(res, dir_factura, perfil, nit_prestador)

    if hacer_zip and formato.startswith("ZIP"):
        zip_nombre = f"{nit_prestador}_{fecha}_{consecutivo:04d}.zip"
        zip_ruta = destino / sanitizar(ent_id) / zip_nombre
        res.acciones.append(f"ZIP → {zip_nombre}")
        if not dry_run:
            _comprimir(dir_factura, zip_ruta)
            res.ruta_paquete = str(zip_ruta)
        return True
    return False


def _escribir_manifiesto_factura(
    res: ResultadoFactura, dir_factura: Path, perfil: PerfilEntidad | None, nit: str
) -> None:
    manifiesto = {
        "factura": res.factura,
        "entidad": res.entidad_nombre,
        "entidad_id": res.entidad_id,
        "canal": perfil.canal if perfil else "",
        "portal": perfil.portal if perfil else "",
        "nit_prestador": nit,
        "valor_total": res.valor_total,
        "usuarios": res.num_usuarios,
        "servicios": res.servicios,
        "soportes_presentes": res.soportes_presentes,
        "estado": res.estado,
        "generado": datetime.now().isoformat(timespec="seconds"),
    }
    (dir_factura / "_manifiesto_radicacion.json").write_text(
        json.dumps(manifiesto, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def _comprimir(carpeta: Path, zip_ruta: Path) -> None:
    """Comprime los soportes de la carpeta. Excluye los archivos de control
    internos (prefijo '_', como _manifiesto_radicacion.json) que no deben
    viajar a la entidad."""
    zip_ruta.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_ruta, "w", zipfile.ZIP_DEFLATED) as zf:
        for p in sorted(carpeta.iterdir()):
            if p.is_file() and not p.name.startswith("_"):
                zf.write(p, arcname=p.name)


# ─── Reportes ────────────────────────────────────────────────────────────────

CAMPOS_REPORTE = [
    "factura",
    "entidad",
    "entidad_id",
    "canal",
    "responsable",
    "lote",
    "valor_total",
    "usuarios",
    "servicios",
    "n_archivos",
    "soportes_presentes",
    "soportes_faltantes",
    "soportes_esperados_faltantes",
    "archivos_sin_clasificar",
    "soportes_clinicos",
    "estado",
    "detalle",
    "ruta_paquete",
    "acciones",
    "carpeta",
]


def _fila_reporte(res: ResultadoFactura) -> dict:
    serv = ", ".join(f"{k}:{v}" for k, v in res.servicios.items())
    return {
        "factura": res.factura,
        "entidad": res.entidad_nombre,
        "entidad_id": res.entidad_id,
        "canal": res.canal,
        "responsable": res.responsable,
        "lote": res.lote,
        "valor_total": f"{res.valor_total:.0f}",
        "usuarios": res.num_usuarios,
        "servicios": serv,
        "n_archivos": res.n_archivos,
        "soportes_presentes": " ".join(res.soportes_presentes),
        "soportes_faltantes": " ".join(res.soportes_faltantes),
        "soportes_esperados_faltantes": " ".join(res.soportes_esperados_faltantes),
        "archivos_sin_clasificar": "; ".join(res.archivos_sin_clasificar),
        # PureWindowsPath: las rutas del cruce pueden ser de Windows (X:\…) y el
        # basename debe salir bien aunque el reporte se genere en otro SO.
        "soportes_clinicos": "; ".join(PureWindowsPath(x).name for x in res.soportes_clinicos),
        "estado": res.estado,
        "detalle": " ".join(res.detalle),
        "ruta_paquete": res.ruta_paquete,
        "acciones": "; ".join(res.acciones),
        "carpeta": res.carpeta,
    }


def escribir_reporte_csv(resultados: list[ResultadoFactura], ruta: Path) -> None:
    ruta.parent.mkdir(parents=True, exist_ok=True)
    with ruta.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=CAMPOS_REPORTE)
        writer.writeheader()
        for r in resultados:
            writer.writerow(_fila_reporte(r))
    logger.info(f"Reporte CSV: {ruta}")


def escribir_reporte_xlsx(resultados: list[ResultadoFactura], ruta: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl no está instalado — omito el XLSX (el CSV sí se generó).")
        return

    colores = {
        "LISTA": "C6EFCE",
        "REVISAR_TIPIFICACION": "FFEB9C",
        "FACTURA_INCONSISTENTE": "FFEB9C",
        "FALTAN_SOPORTES": "FFC7CE",
        "SIN_CUV": "FFC7CE",
        "SIN_FEV": "FFC7CE",
        "SIN_RIPS": "FFC7CE",
        "ENTIDAD_NO_RESUELTA": "F4B084",
        "PARTICULAR": "D9D9D9",
    }
    wb = Workbook()
    ws = wb.active
    ws.title = "RADICACION"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(CAMPOS_REPORTE, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
    anchos = {
        "detalle": 60,
        "soportes_presentes": 22,
        "soportes_clinicos": 30,
        "carpeta": 40,
        "acciones": 30,
        "servicios": 24,
        "responsable": 14,
        "lote": 20,
    }
    for c, h in enumerate(CAMPOS_REPORTE, 1):
        ws.column_dimensions[get_column_letter(c)].width = anchos.get(h, 16)

    i_estado = CAMPOS_REPORTE.index("estado")
    for r, res in enumerate(resultados, start=2):
        fila = _fila_reporte(res)
        for c, h in enumerate(CAMPOS_REPORTE, 1):
            cell = ws.cell(row=r, column=c, value=fila[h])
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        color = colores.get(res.estado)
        if color:
            ws.cell(row=r, column=i_estado + 1).fill = PatternFill("solid", fgColor=color)
    ws.freeze_panes = "A2"

    # Hoja resumen por entidad.
    ws2 = wb.create_sheet("RESUMEN")
    ws2.append(["Entidad", "Facturas", "Listas", "Con problemas", "Valor total"])
    for c in range(1, 6):
        ws2.cell(row=1, column=c).fill = header_fill
        ws2.cell(row=1, column=c).font = header_font
    for ent, datos in _resumen_por_entidad(resultados).items():
        ws2.append(
            [ent, datos["total"], datos["listas"], datos["problemas"], f"{datos['valor']:.0f}"]
        )
    for col, ancho in (("A", 38), ("B", 10), ("C", 8), ("D", 14), ("E", 16)):
        ws2.column_dimensions[col].width = ancho

    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ruta))
    logger.info(f"Reporte XLSX: {ruta}")


def _resumen_por_entidad(resultados: list[ResultadoFactura]) -> dict[str, dict]:
    agg: dict[str, dict] = defaultdict(
        lambda: {"total": 0, "listas": 0, "problemas": 0, "valor": 0.0}
    )
    for r in resultados:
        d = agg[r.entidad_nombre]
        d["total"] += 1
        d["valor"] += r.valor_total
        if r.estado in ESTADOS_OK:
            d["listas"] += 1
        else:
            d["problemas"] += 1
    return dict(sorted(agg.items(), key=lambda kv: -kv[1]["total"]))


def imprimir_resumen(resultados: list[ResultadoFactura], n_indice: int | None = None) -> None:
    # Una sola pasada sobre resultados para todos los agregados del resumen.
    por_estado: Counter[str] = Counter()
    valor_total = 0.0
    revisar: list[ResultadoFactura] = []
    con_cruce = n_files = 0
    por_resp: dict[str, list] = {}  # responsable → [n, listas, valor]
    for r in resultados:
        por_estado[r.estado] += 1
        valor_total += r.valor_total
        if r.estado == "REVISAR_TIPIFICACION":
            revisar.append(r)
        if r.soportes_clinicos:
            con_cruce += 1
            n_files += len(r.soportes_clinicos)
        if r.responsable:
            acc = por_resp.setdefault(r.responsable, [0, 0, 0.0])
            acc[0] += 1
            acc[1] += 1 if r.estado == "LISTA" else 0
            acc[2] += r.valor_total
    total = len(resultados)
    logger.info("=" * 64)
    logger.info("RESUMEN DE RADICACIÓN")
    logger.info(f"  Facturas procesadas: {total:,}")
    logger.info(f"  Valor total facturado: $ {valor_total:,.0f}")
    logger.info("  Por estado:")
    orden = [*ESTADOS_OK, *ESTADOS_PROBLEMA]
    for estado in orden + [e for e in por_estado if e not in orden]:
        n = por_estado.get(estado, 0)
        if n:
            pct = 100 * n / max(1, total)
            logger.info(f"    {estado:<22} {n:>4} ({pct:.0f}%)")
    # Desglose de REVISAR_TIPIFICACION: distingue lo que el cruce --soportes puede
    # resolver (faltan soportes clínicos) de lo que necesita renombrar (archivos
    # sin tipificar). Así se ve dónde está el lever para bajar ese estado.
    if revisar:
        faltan_clin = [r for r in revisar if r.soportes_esperados_faltantes]
        solo_sin_clasif = [
            r for r in revisar if r.archivos_sin_clasificar and not r.soportes_esperados_faltantes
        ]
        cod_faltantes = Counter(c for r in faltan_clin for c in r.soportes_esperados_faltantes)
        logger.info("  REVISAR_TIPIFICACION — por qué:")
        logger.info(
            f"    faltan soportes clínicos   : {len(faltan_clin):>4}  (cruzables con --soportes)"
        )
        logger.info(
            f"    solo archivos sin tipificar: {len(solo_sin_clasif):>4}  (renombrar/aliasar)"
        )
        if cod_faltantes:
            top = ", ".join(f"{c}×{n}" for c, n in cod_faltantes.most_common(8))
            logger.info(f"    códigos clínicos más faltantes: {top}")
    # Qué hizo REALMENTE el cruce --soportes: cuántas facturas del lote recibieron
    # soportes del share. Se imprime aquí (final del resumen) para que el dato
    # quede a la vista aunque no se scrollee al inicio del log. Auto-diagnostica
    # los dos modos de falla: share vacío vs. numeración que no cruza.
    if n_indice is not None:
        logger.info("  Cruce de soportes clínicos (--soportes):")
        logger.info(f"    facturas indexadas en el/los share(s):     {n_indice:>5,}")
        logger.info(
            f"    facturas del lote que recibieron soportes: {con_cruce:>5,}  "
            f"({n_files:,} archivo(s) anexados)"
        )
        if n_indice == 0:
            logger.info(
                "    ⚠ el share no devolvió NINGÚN soporte: revisá la(s) ruta(s) de --soportes "
                "(¿carpeta correcta? ¿permiso de lectura? ¿el disco se ve desde Python?)."
            )
        elif con_cruce == 0:
            logger.info(
                "    ⚠ hay soportes indexados pero NINGUNO cruzó con el lote: la numeración "
                "del share no coincide con la del lote (revisar rango/normalización)."
            )
    # Por responsable (HOSPIAI D5): quién radica qué, leído de la ruta del share.
    if por_resp:
        logger.info("  Por responsable (leído de la ruta):")
        top = sorted(por_resp.items(), key=lambda kv: -kv[1][0])[:10]
        for nombre, (n, listas, valor) in top:
            logger.info(
                f"    {nombre[:24]:<24} {n:>5,} facturas | {listas:>5,} listas | $ {valor:,.0f}"
            )
    logger.info("  Por entidad:")
    for ent, d in _resumen_por_entidad(resultados).items():
        logger.info(
            f"    {ent[:34]:<34} {d['total']:>3} facturas | "
            f"{d['listas']:>3} listas | $ {d['valor']:,.0f}"
        )


# ─── CLI ─────────────────────────────────────────────────────────────────────


def _coincide_entidad(filtro_norm: str, res: ResultadoFactura) -> bool:
    return filtro_norm in _norm(res.entidad_nombre) or filtro_norm in _norm(res.entidad_hint)


def setup_logging(log_file: Path | None = None) -> None:
    fmt = "%(asctime)s [%(levelname)s] %(message)s"
    handlers: list[logging.Handler] = [logging.StreamHandler(sys.stdout)]
    if log_file is not None:
        log_file.parent.mkdir(parents=True, exist_ok=True)
        handlers.append(logging.FileHandler(log_file, encoding="utf-8"))
    logging.basicConfig(level=logging.INFO, format=fmt, handlers=handlers)


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    p.add_argument(
        "--origen", type=Path, required=True, help="Carpeta raíz del lote de facturación."
    )
    p.add_argument(
        "--destino", type=Path, default=None, help="Carpeta donde armar los paquetes (con --armar)."
    )
    p.add_argument(
        "--perfiles",
        type=Path,
        default=None,
        help="JSON de perfiles (def.: data/perfiles_radicacion.json).",
    )
    p.add_argument(
        "--reglas",
        type=Path,
        default=None,
        help="Motor de reglas declarativo (def.: data/reglas_radicacion.json). "
        "Cada regla lleva su fuente normativa; la corrida registra la versión usada.",
    )
    p.add_argument(
        "--db",
        type=Path,
        default=None,
        help="Expediente Digital HOSPIAI donde persistir la corrida "
        "(def.: data/hospiai.db del repo). El CSV/XLSX se generan igual.",
    )
    p.add_argument(
        "--sin-db",
        action="store_true",
        dest="sin_db",
        help="No escribir en el Expediente Digital (solo CSV/XLSX).",
    )
    p.add_argument(
        "--manifiesto",
        type=Path,
        default=None,
        help="CSV/XLSX que mapea factura → entidad (cuando el lote no está por carpetas de EPS).",
    )
    p.add_argument(
        "--soportes",
        type=Path,
        nargs="+",
        default=None,
        metavar="CARPETA",
        help="Una o VARIAS carpetas raíz del share de soportes CLÍNICOS (epicrisis, "
        "evolución, órdenes…). Se recorren recursivamente y se cruzan por número de "
        "factura para completar las que tienen procedimientos/urgencias/hospitalización "
        "y sacarlas de REVISAR_TIPIFICACION. Pasá todas las que apliquen (p.ej. cada "
        "carpeta mensual de Y:\\) o la raíz del disco. Si se omite, usa $SOPORTES_ROOT "
        "(que puede traer varias rutas separadas por ';').",
    )
    p.add_argument(
        "--soportes-indice",
        type=Path,
        default=None,
        dest="soportes_indice",
        metavar="ARCHIVO",
        help="Listado pre-armado de rutas de ARCHIVOS de soporte (un 'dir /s /b' o "
        "'Get-ChildItem -Recurse -File', una ruta por línea). El radicador lo lee SIN "
        "recorrer la red, así cubre de una todas las rutas del listado (X:, Y:, Z:…). "
        "Ideal cuando los soportes están repartidos en varios discos. Se combina con "
        "--soportes si pasás ambos.",
    )
    p.add_argument(
        "--layout",
        choices=["auto", "carpeta-factura", "lote"],
        default="auto",
        help="auto (def.): ancla en carpetas HUS<factura> y agrupa lo demás por token "
        "(cubre los shares anidados). carpeta-factura: una carpeta por factura. "
        "lote: muchos archivos por carpeta.",
    )
    p.add_argument(
        "--entidad", type=str, default=None, help="Filtrar solo esta entidad (substring)."
    )
    p.add_argument(
        "--reporte",
        type=Path,
        default=Path("radicacion_reporte.csv"),
        help="CSV de salida con el estado de cada factura.",
    )
    p.add_argument("--xlsx", type=Path, default=None, help="XLSX con formato y hoja de resumen.")
    p.add_argument(
        "--armar", action="store_true", help="Construir los paquetes (carpetas + renombrado)."
    )
    p.add_argument(
        "--zip", action="store_true", dest="hacer_zip", help="Comprimir cada paquete a ZIP."
    )
    p.add_argument(
        "--mover", action="store_true", help="Mover archivos en vez de copiar (con --armar)."
    )
    p.add_argument(
        "--forzar",
        action="store_true",
        help="Armar también las facturas en estado REVISAR_TIPIFICACION.",
    )
    p.add_argument("--fecha", type=str, default=None, help="Fecha AAAAMMDD para el nombre del ZIP.")
    p.add_argument(
        "--consecutivo-inicial",
        type=int,
        default=1,
        dest="consecutivo",
        help="Consecutivo inicial para los ZIP (def. 1).",
    )
    p.add_argument(
        "--dry-run", action="store_true", help="No toca archivos; solo reporta qué haría."
    )
    p.add_argument("--log", type=Path, default=None, help="Archivo de log opcional.")
    return p


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    setup_logging(args.log)

    if not args.origen.is_dir():
        logger.error(f"--origen no existe o no es carpeta: {args.origen}")
        return 1
    if args.armar and args.destino is None:
        logger.error("--armar requiere --destino.")
        return 1
    if not _ADRES_OK:
        logger.warning(
            "No pude importar tools/adres/ (rips_lectura, factura_lectura): los valores y "
            "servicios del RIPS saldrán limitados. Verificá que el script esté en tools/."
        )

    cfg = cargar_perfiles(args.perfiles)
    cfg.prestador_nit = cfg.prestador_nit or PRESTADOR_NIT_DEFAULT
    cargar_reglas(args.reglas, cfg)

    manifiesto: dict[str, str] = {}
    if args.manifiesto is not None:
        if not args.manifiesto.is_file():
            logger.error(f"--manifiesto no existe: {args.manifiesto}")
            return 1
        manifiesto = cargar_manifiesto(args.manifiesto)
        logger.info(f"Manifiesto: {len(manifiesto)} facturas mapeadas a entidad.")

    # Soportes clínicos (epicrisis, evolución, urgencias, órdenes…), que viven en
    # OTRO(S) share(s). Dos fuentes que se FUSIONAN en un solo índice por factura:
    #   --soportes-indice  un listado pre-armado de archivos (no recorre la red;
    #                      cubre de una todas las rutas del listado: X:, Y:, Z:…).
    #   --soportes         una o varias carpetas que se recorren recursivamente
    #                      (o $SOPORTES_ROOT, con rutas separadas por os.pathsep).
    soportes_idx: dict[str, list[Path]] | None = None
    fusion: dict[str, list[Path]] = defaultdict(list)
    if args.soportes_indice is not None:
        if not args.soportes_indice.is_file():
            logger.error(
                f"--soportes-indice no existe o no es archivo: {args.soportes_indice}\n"
                "  Armalo UNA vez (tarda unos minutos y sirve para todas las corridas):\n"
                '  Get-ChildItem "Y:\\","Z:\\SERVIDOR GLOSAS","X:\\SERVIDOR RADICACION\\2. '
                'SINAC SC SAS - 2026" -Recurse -File -ErrorAction SilentlyContinue | '
                "ForEach-Object FullName | "
                f'Set-Content "{args.soportes_indice}" -Encoding UTF8'
            )
            return 1
        for fac, rutas in indexar_soportes_desde_indice(
            args.soportes_indice, cfg.patron_factura
        ).items():
            fusion[fac].extend(rutas)
    rutas_soportes: list[Path] = list(args.soportes) if args.soportes else []
    if not rutas_soportes and args.soportes_indice is None and os.environ.get("SOPORTES_ROOT"):
        rutas_soportes = [
            Path(p) for p in os.environ["SOPORTES_ROOT"].split(os.pathsep) if p.strip()
        ]
    if rutas_soportes:
        invalidas = [r for r in rutas_soportes if not r.is_dir()]
        if invalidas:
            logger.error(
                "--soportes: no existe(n) o no es/son carpeta(s): "
                + ", ".join(str(x) for x in invalidas)
            )
            return 1
        for raiz in rutas_soportes:
            logger.info(f"Indexando soportes clínicos desde {raiz} …")
            for fac, rutas in indexar_soportes_clinicos(raiz, cfg.patron_factura).items():
                fusion[fac].extend(rutas)
    if fusion:
        soportes_idx = dict(fusion)
        logger.info(f"Soportes clínicos: {len(soportes_idx)} factura(s) con soporte para cruzar.")
    # Se pidió cruce (--soportes/--soportes-indice/$SOPORTES_ROOT) → reportá cuántas
    # facturas quedaron indexadas (0 incluido, para diagnosticar share vacío).
    cruce_pedido = bool(rutas_soportes) or args.soportes_indice is not None
    n_indice_soportes = len(fusion) if cruce_pedido else None

    if args.layout == "lote":
        items = descubrir_lote(args.origen, manifiesto, cfg, cfg.patron_factura)
    elif args.layout == "carpeta-factura":
        items = descubrir_carpeta_factura(args.origen, manifiesto, cfg)
    else:
        items = descubrir_auto(args.origen, manifiesto, cfg, cfg.patron_factura)

    if not items:
        logger.error(
            f"No encontré facturas en {args.origen} (layout={args.layout}). "
            "¿El lote tiene subcarpetas con soportes? Probá el otro --layout."
        )
        return 1

    filtro_ent = _norm(args.entidad or "")
    fecha = args.fecha or datetime.now().strftime("%Y%m%d")
    logger.info(
        f"Lote: {len(items)} factura(s) en {args.origen} (layout={args.layout})"
        + (f" | filtro entidad: '{args.entidad}'" if args.entidad else "")
        + ("  [MODO AUDITORÍA — no se arma nada]" if not args.armar else "")
        + ("  (DRY-RUN)" if args.dry_run else "")
    )

    resultados: list[ResultadoFactura] = []
    docs_por_factura: dict[str, list[Path]] = {}
    consecutivo = args.consecutivo
    for i, (fac_hint, archivos, carpeta, ent_hint) in enumerate(items, 1):
        res = procesar_factura(fac_hint, archivos, carpeta, ent_hint, cfg, soportes_idx)
        if filtro_ent and not _coincide_entidad(filtro_ent, res):
            continue
        docs_por_factura[res.factura_norm or res.factura] = archivos
        logger.info(f"[{i}/{len(items)}] {res.factura} → {res.entidad_nombre}  [{res.estado}]")
        if args.armar:
            perfil = resolver_entidad(ent_hint, cfg)
            uso_zip = armar_paquete(
                res,
                archivos,
                perfil,
                args.destino,
                fecha,
                consecutivo,
                mover=args.mover,
                hacer_zip=args.hacer_zip,
                forzar=args.forzar,
                dry_run=args.dry_run,
                nit_prestador=cfg.prestador_nit,
            )
            if uso_zip:
                consecutivo += 1
        resultados.append(res)

    if not resultados:
        logger.warning("Ninguna factura pasó el filtro de entidad.")
        return 0

    escribir_reporte_csv(resultados, args.reporte)
    if args.xlsx is not None:
        escribir_reporte_xlsx(resultados, args.xlsx)

    # ── Expediente Digital (HOSPIAI Fase 1) ─────────────────────────────────
    # Memoria ADICIONAL y persistente de la corrida: expedientes, documentos,
    # hallazgos y eventos con la versión de reglas usada. No reemplaza el CSV.
    if not args.sin_db:
        db_path = args.db or Path(__file__).resolve().parent.parent / "data" / "hospiai.db"
        try:
            import hospiai_db

            resumen_db = hospiai_db.persistir_corrida(
                db_path,
                resultados,
                docs_por_factura,
                version_reglas=cfg.version_reglas,
                version_motor=VERSION_MOTOR,
                regla_id_por_codigo=cfg.regla_id_por_codigo,
                reglas_registro=cfg.reglas_registro,
                catalogo=[
                    {
                        "id": e.id,
                        "nombre": e.nombre,
                        "nit": e.nit,
                        "regimen": e.regimen,
                        "canal": e.canal,
                        "plazo_radicacion_dias": e.plazo_radicacion_dias,
                        "periodicidad": e.periodicidad,
                        "tipo_radicacion": e.tipo_radicacion,
                    }
                    for e in cfg.entidades
                ],
                origen=str(args.origen),
                parametros=" ".join(argv if argv is not None else sys.argv[1:]),
                clasificar=lambda nombre: clasificar_soporte(nombre)[0],
            )
            logger.info(
                f"Expediente Digital: {resumen_db['expedientes']:,} expediente(s), "
                f"{resumen_db['documentos']:,} documento(s), {resumen_db['hallazgos']:,} "
                f"hallazgo(s) → {db_path} (corrida #{resumen_db['corrida_id']}, "
                f"reglas v{cfg.version_reglas})"
            )
        except Exception as e:  # la corrida no se pierde por un fallo de la BD
            logger.warning(
                f"No pude escribir el Expediente Digital en {db_path}: {e} — "
                "el CSV/XLSX sí quedaron generados."
            )

    imprimir_resumen(resultados, n_indice_soportes)

    listas = sum(1 for r in resultados if r.estado in ESTADOS_OK)
    logger.info(
        f"\n{listas}/{len(resultados)} factura(s) LISTAS para radicar. "
        "Revisá el reporte para las demás."
    )
    problemas = len(resultados) - listas
    return 0 if problemas == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
