"""Generador del Excel-respuesta que se envía a los gestores tras la
importación de recepción.

Toma el Excel original que subió el equipo de recepción, anota cada
fila con la respuesta IA + estado + ID glosa, y resalta visualmente
las filas asignadas al gestor destinatario.

El archivo resultante es lo que se adjunta al correo broadcast — el
gestor lo abre y ya tiene su tabla original más la columna nueva de
respuestas pre-generadas, sin necesidad de abrir la app para cada
glosa.

Diseño:
  • Re-usa los `COLUMN_ALIASES` y `_buscar_fila_encabezado` de
    `recepcion_service` para localizar dónde están factura, gestor,
    consecutivo y demás columnas en cualquier variante de Excel.
  • Lookup por clave (factura, consecutivo_dgh) para empatar la fila
    del Excel con la `GlosaRecord` persistida tras la importación.
  • Si una fila no tiene match (raro: solo si la importación falló
    parcialmente) se escribe "—" en RESPUESTA IA para que el gestor
    sepa que esa fila quedó sin procesar.
"""
from __future__ import annotations

from io import BytesIO
from typing import Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.logging_utils import logger
from app.services.recepcion_service import (
    COLUMN_ALIASES,
    _buscar_fila_encabezado,
)


_NUEVA_COL_RESPUESTA = "RESPUESTA IA"
_NUEVA_COL_ESTADO = "ESTADO IA"
_NUEVA_COL_GLOSA_ID = "ID GLOSA"

# Colores para resaltar al gestor destinatario y el estado IA.
_RESALTO_GESTOR_FILL = PatternFill("solid", fgColor="FEF3C7")  # amarillo suave
_RESALTO_GESTOR_FONT = Font(bold=True, color="92400E")
_RESPONDIDA_FILL = PatternFill("solid", fgColor="DCFCE7")  # verde claro
_REQUIERE_FILL = PatternFill("solid", fgColor="FEE2E2")  # rojo claro
_ERROR_FILL = PatternFill("solid", fgColor="E5E7EB")  # gris claro

_BORDE_FINO = Border(
    left=Side(border_style="thin", color="D1D5DB"),
    right=Side(border_style="thin", color="D1D5DB"),
    top=Side(border_style="thin", color="D1D5DB"),
    bottom=Side(border_style="thin", color="D1D5DB"),
)


def _normalizar_clave(valor) -> str:
    """Normaliza factura/consecutivo a string trimeado en mayúsculas
    para el match entre fila Excel y GlosaRecord."""
    if valor is None:
        return ""
    return str(valor).strip().upper()


def _truncar_para_celda(texto: str, max_chars: int = 32000) -> str:
    """Excel tiene un límite duro de 32767 caracteres por celda. Si el
    dictamen es más largo, lo cortamos avisando con sufijo."""
    if not texto:
        return ""
    if len(texto) <= max_chars:
        return texto
    return texto[: max_chars - 60] + "\n\n[... TRUNCADO — VER DICTAMEN COMPLETO EN LA APP]"


def _color_por_estado(estado: str) -> Optional[PatternFill]:
    estado = (estado or "").upper()
    if estado == "RESPONDIDA":
        return _RESPONDIDA_FILL
    if estado == "REQUIERE_SOPORTES":
        return _REQUIERE_FILL
    if estado in ("ERROR", "TEXTO_INSUFICIENTE", "NO_PROCESADA"):
        return _ERROR_FILL
    return None


def generar_excel_con_respuestas(
    contenido_original: bytes,
    respuestas_por_clave: dict[tuple[str, str], dict],
    gestor_destacar: Optional[str] = None,
) -> bytes:
    """Devuelve el Excel original anotado con las respuestas IA y el
    gestor destacado.

    Args:
      contenido_original: bytes del .xlsx original que subió recepción.
      respuestas_por_clave: dict {(factura_upper, consecutivo_upper):
        {"glosa_id", "estado", "dictamen", "modelo_ia"}}.
      gestor_destacar: si se da, las filas de ese gestor se pintan en
        amarillo suave para que el destinatario las identifique
        rápidamente.
    """
    wb = load_workbook(BytesIO(contenido_original), data_only=True)
    gestor_norm = (gestor_destacar or "").strip().upper()

    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        # Localizar encabezados RECEPCION en esta hoja.
        fila_h, idx = _buscar_fila_encabezado(
            ws, max_filas=5, mapa=COLUMN_ALIASES, min_aciertos=3,
        )
        if not idx or "factura" not in idx:
            # No es hoja de recepción — saltar (conceptos, vacías, etc.)
            continue

        col_factura_idx = idx["factura"]
        col_consec_idx = idx.get("consecutivo_dgh")
        col_gestor_idx = idx.get("gestor")

        # Determinar cuántas columnas tiene la hoja para añadir las
        # nuestras al final. Usamos ws.max_column que en modo writable
        # ya es preciso (a diferencia del read-only).
        col_existentes = ws.max_column
        col_resp = col_existentes + 1
        col_estado = col_existentes + 2
        col_id = col_existentes + 3

        # Escribir encabezados en la misma fila del header detectado
        ws.cell(row=fila_h, column=col_resp, value=_NUEVA_COL_RESPUESTA)
        ws.cell(row=fila_h, column=col_estado, value=_NUEVA_COL_ESTADO)
        ws.cell(row=fila_h, column=col_id, value=_NUEVA_COL_GLOSA_ID)
        for c_idx in (col_resp, col_estado, col_id):
            celda = ws.cell(row=fila_h, column=c_idx)
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", fgColor="1E40AF")
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.border = _BORDE_FINO

        # Iterar las filas de datos a partir de fila_h+1
        for fila in ws.iter_rows(
            min_row=fila_h + 1,
            max_row=ws.max_row,
            min_col=1,
            max_col=col_existentes,
        ):
            # Skip filas completamente vacías
            if all(c.value is None for c in fila):
                continue
            num_fila = fila[0].row

            factura_val = fila[col_factura_idx].value if col_factura_idx < len(fila) else None
            consec_val = (
                fila[col_consec_idx].value
                if col_consec_idx is not None and col_consec_idx < len(fila)
                else None
            )
            clave = (
                _normalizar_clave(factura_val),
                _normalizar_clave(consec_val),
            )
            if not clave[0]:
                continue  # fila sin factura, ignorar

            resp = respuestas_por_clave.get(clave)
            if resp is None:
                # Match degradado: solo por factura (sin consecutivo)
                for k, v in respuestas_por_clave.items():
                    if k[0] == clave[0]:
                        resp = v
                        break

            if resp:
                dictamen = _truncar_para_celda(resp.get("dictamen") or "")
                estado = (resp.get("estado") or "").upper()
                glosa_id = resp.get("glosa_id") or ""
            else:
                dictamen = "—"
                estado = "NO_PROCESADA"
                glosa_id = ""

            celda_resp = ws.cell(row=num_fila, column=col_resp, value=dictamen)
            celda_resp.alignment = Alignment(
                wrap_text=True, vertical="top", horizontal="left",
            )
            celda_resp.border = _BORDE_FINO
            celda_resp.font = Font(size=9)

            celda_estado = ws.cell(row=num_fila, column=col_estado, value=estado)
            celda_estado.alignment = Alignment(
                horizontal="center", vertical="center",
            )
            celda_estado.border = _BORDE_FINO
            celda_estado.font = Font(bold=True, size=9)
            fill_estado = _color_por_estado(estado)
            if fill_estado is not None:
                celda_estado.fill = fill_estado

            celda_id = ws.cell(row=num_fila, column=col_id, value=glosa_id)
            celda_id.alignment = Alignment(horizontal="center")
            celda_id.border = _BORDE_FINO
            celda_id.font = Font(size=9, color="6B7280")

            # Resaltar columna gestor si coincide con el destinatario
            if (
                gestor_norm
                and col_gestor_idx is not None
                and col_gestor_idx < len(fila)
            ):
                gestor_val = _normalizar_clave(fila[col_gestor_idx].value)
                if gestor_val and (gestor_val == gestor_norm or gestor_norm in gestor_val or gestor_val in gestor_norm):
                    fila[col_gestor_idx].fill = _RESALTO_GESTOR_FILL
                    fila[col_gestor_idx].font = _RESALTO_GESTOR_FONT

        # Ajustar ancho de la columna respuesta (la más larga)
        ws.column_dimensions[get_column_letter(col_resp)].width = 80
        ws.column_dimensions[get_column_letter(col_estado)].width = 22
        ws.column_dimensions[get_column_letter(col_id)].width = 10

    out = BytesIO()
    try:
        wb.save(out)
    finally:
        wb.close()
    return out.getvalue()


def construir_respuestas_por_clave(db, glosa_ids: list[int]) -> dict[tuple[str, str], dict]:
    """Consulta `GlosaRecord` por IDs y devuelve el diccionario que
    consume `generar_excel_con_respuestas`.

    Clave: (factura_upper, consecutivo_dgh_upper).
    """
    if not glosa_ids:
        return {}
    from app.models.db import GlosaRecord

    glosas = (
        db.query(GlosaRecord)
        .filter(GlosaRecord.id.in_(glosa_ids))
        .all()
    )
    fuera: dict[tuple[str, str], dict] = {}
    for g in glosas:
        clave = (
            _normalizar_clave(g.factura),
            _normalizar_clave(g.consecutivo_dgh),
        )
        fuera[clave] = {
            "glosa_id": g.id,
            "estado": g.estado or "",
            "dictamen": g.dictamen or "",
            "modelo_ia": g.modelo_ia or "",
        }
    if len(fuera) < len(glosa_ids):
        logger.warning(
            f"[excel-respuesta] {len(glosa_ids) - len(fuera)} glosas de "
            f"{len(glosa_ids)} no se encontraron al construir respuestas"
        )
    return fuera
