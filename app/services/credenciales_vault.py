"""Vault cifrado de credenciales de plataformas EPS/entidades.

El hospital mantiene un Excel maestro ("RADICAR FACTURAS ENTIDADES",
hoja CONSOLIDADO) con los accesos a las plataformas de radicación,
cartera y devoluciones de ~122 entidades. Este módulo:

1. Cifra/descifra los campos sensibles (usuario, contraseña, teléfono,
   correo) con Fernet (AES-128 AEAD) usando la clave de entorno
   ``CRED_VAULT_KEY``. Generar una con:

       python -c "from cryptography.fernet import Fernet; print(Fernet.generate_key().decode())"

2. Parsea el Excel maestro (``parsear_consolidado``) a registros por
   entidad y bloque (RADICACION | CARTERA | DEVOLUCIONES).

A DIFERENCIA de app/services/cifrado.py (que degrada a texto plano si
no hay clave), este vault es ESTRICTO: sin ``CRED_VAULT_KEY`` no se
importa ni se revela nada — los endpoints devuelven 503. Nunca se
persisten credenciales en claro.

Rotación de clave: fuera de alcance por ahora (follow-up). Si se cambia
la clave, los tokens viejos dejan de descifrarse (InvalidToken).

REGLA DE ORO: jamás loguear valores descifrados ni tokens — solo
conteos y metadatos.
"""

from __future__ import annotations

import io
import os
import re

from cryptography.fernet import Fernet

# Bloques soportados (orden = orden de las columnas en el Excel)
BLOQUE_RADICACION = "RADICACION"
BLOQUE_CARTERA = "CARTERA"
BLOQUE_DEVOLUCIONES = "DEVOLUCIONES"
BLOQUES = (BLOQUE_RADICACION, BLOQUE_CARTERA, BLOQUE_DEVOLUCIONES)

# Cache del Fernet keyed por la clave en uso: si la env var cambia
# (tests, rotación manual), se reconstruye en el siguiente uso.
_cache: dict = {"key": None, "fernet": None}


class VaultNoConfigurado(RuntimeError):
    """Se intentó cifrar/descifrar sin CRED_VAULT_KEY válida."""


def _get_fernet() -> Fernet | None:
    key = os.getenv("CRED_VAULT_KEY", "").strip()
    if not key:
        return None
    if _cache["key"] == key and _cache["fernet"] is not None:
        return _cache["fernet"]
    try:
        fernet = Fernet(key.encode("ascii"))
    except Exception:
        # Clave malformada (no es base64 urlsafe de 32 bytes) → vault apagado.
        return None
    _cache["key"] = key
    _cache["fernet"] = fernet
    return fernet


def vault_disponible() -> bool:
    """True si CRED_VAULT_KEY está seteada y es una clave Fernet válida."""
    return _get_fernet() is not None


def cifrar(texto: str) -> str:
    """Cifra un valor sensible. Lanza VaultNoConfigurado si no hay clave."""
    fernet = _get_fernet()
    if fernet is None:
        raise VaultNoConfigurado("CRED_VAULT_KEY no configurada o inválida")
    return fernet.encrypt(texto.encode("utf-8")).decode("ascii")


def descifrar(token: str) -> str:
    """Descifra un token Fernet. Lanza VaultNoConfigurado sin clave,
    o cryptography.fernet.InvalidToken si la clave no corresponde."""
    fernet = _get_fernet()
    if fernet is None:
        raise VaultNoConfigurado("CRED_VAULT_KEY no configurada o inválida")
    return fernet.decrypt(token.encode("ascii")).decode("utf-8")


# ─── Parser del Excel maestro (hoja CONSOLIDADO) ────────────────────────────
#
# Layout real (encabezados en FILA 2; fila 1 es banner de bloques):
#   A NIT | B EMPRESAS
#   C-I  RADICACION:   LINK PLATAFORMA, USUARIO, CONTRASEÑA, TELÉFONO,
#                      CORREO, NOMBRE, CARGO
#   J-L  CARTERA:      LINK PLATAFORMA, USUARIO, CLAVE
#   M-O  DEVOLUCIONES: PLATAFORMA/FISICO/CORREO, LINK PLATAFORMA DIRECCIÓN,
#                      USUARIO CONTRASEÑA (celda combinada, usualmente
#                      usuario y clave en líneas separadas)
#   P MANUAL DE RADICACION | Q MEDIO DE CONTACTO | R NOMBRE CONTACTO
# Columnas S+ (OBSERVACIONES, CIUDAD, CUV, bloque GLOSAS) se ignoran
# en esta tanda.

_NUM_COLS = 18  # A..R


def _celda(valor) -> str:
    """Normaliza una celda a str: números sin '.0', strip, None → ''."""
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


_LABEL_RE = re.compile(r"^(usuario|user|contraseñ?a|clave|password)\s*[:=]\s*", re.IGNORECASE)


def _separar_usuario_password(blob: str) -> tuple[str, str]:
    """Separa la celda combinada 'USUARIO CONTRASEÑA' de DEVOLUCIONES.

    Heurística (validada contra el archivo real, donde ~90% de las celdas
    son multilínea): primera línea = usuario, resto = contraseña. Si la
    celda es de una sola línea sin estructura, se guarda completa como
    usuario (sin pérdida: al revelar se ve el texto íntegro).
    """
    lineas = []
    for linea in blob.splitlines():
        linea = linea.strip()
        if not linea:
            continue
        sin_label = _LABEL_RE.sub("", linea).strip()
        lineas.append(sin_label or linea)
    if not lineas:
        return "", ""
    if len(lineas) >= 2:
        return lineas[0], " ".join(lineas[1:])
    return lineas[0], ""


def parsear_consolidado(contenido: bytes) -> dict:
    """Parsea el Excel maestro y devuelve entidades con sus bloques.

    Returns:
        {
          "entidades": [
            {
              "nit": str, "empresa": str,
              "manual_radicacion": str, "medio_contacto": str,
              "nombre_contacto": str,
              "bloques": {
                "RADICACION": {link_plataforma, usuario, password,
                               telefono, correo, nombre_contacto, cargo},
                "CARTERA": {link_plataforma, usuario, password},
                "DEVOLUCIONES": {link_plataforma, usuario, password,
                                 medio_contacto},
              },  # solo bloques con algún dato
            }, ...
          ],
          "filas_sin_nit": int,   # filas con datos pero sin NIT (se omiten)
        }

    Raises:
        ValueError: si falta la hoja CONSOLIDADO o el encabezado NIT.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    try:
        if "CONSOLIDADO" not in wb.sheetnames:
            raise ValueError(
                f"El Excel no tiene hoja 'CONSOLIDADO'. Hojas encontradas: {wb.sheetnames}"
            )
        ws = wb["CONSOLIDADO"]

        filas = ws.iter_rows(min_row=2, values_only=True)
        try:
            encabezados = next(filas)
        except StopIteration:
            raise ValueError("Hoja CONSOLIDADO vacía")
        if _celda(encabezados[0] if encabezados else "").upper() != "NIT":
            raise ValueError(
                "Encabezado inesperado: la celda A2 debe ser 'NIT' "
                "(fila 1 = banner de bloques, fila 2 = encabezados)"
            )

        entidades: list[dict] = []
        filas_sin_nit = 0

        for fila in filas:
            vals = [_celda(c) for c in fila[:_NUM_COLS]]
            vals += [""] * (_NUM_COLS - len(vals))
            if not any(vals):
                continue
            nit, empresa = vals[0], vals[1]
            if not nit:
                filas_sin_nit += 1
                continue

            manual = vals[15]
            medio = vals[16]
            nombre_entidad = vals[17]

            dev_usuario, dev_password = _separar_usuario_password(vals[14])
            bloques_raw = {
                BLOQUE_RADICACION: {
                    "link_plataforma": vals[2],
                    "usuario": vals[3],
                    "password": vals[4],
                    "telefono": vals[5],
                    "correo": vals[6],
                    "nombre_contacto": vals[7] or nombre_entidad,
                    "cargo": vals[8],
                },
                BLOQUE_CARTERA: {
                    "link_plataforma": vals[9],
                    "usuario": vals[10],
                    "password": vals[11],
                },
                BLOQUE_DEVOLUCIONES: {
                    "link_plataforma": vals[13],
                    "usuario": dev_usuario,
                    "password": dev_password,
                    "medio_contacto": vals[12] or medio,
                },
            }
            # Un bloque existe solo si trae algún dato PROPIO (los
            # fallbacks de entidad no cuentan para decidir presencia).
            presencia = {
                BLOQUE_RADICACION: any(vals[2:9]),
                BLOQUE_CARTERA: any(vals[9:12]),
                BLOQUE_DEVOLUCIONES: any(vals[12:15]),
            }
            bloques = {b: datos for b, datos in bloques_raw.items() if presencia[b]}

            entidades.append(
                {
                    "nit": nit,
                    "empresa": empresa or f"NIT {nit}",
                    "manual_radicacion": manual,
                    "medio_contacto": medio,
                    "nombre_contacto": nombre_entidad,
                    "bloques": bloques,
                }
            )

        return {"entidades": entidades, "filas_sin_nit": filas_sin_nit}
    finally:
        wb.close()
