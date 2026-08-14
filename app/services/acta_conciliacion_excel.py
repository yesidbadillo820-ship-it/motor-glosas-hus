"""El acta de conciliación que se trabaja a mano en Excel, leída y cuadrada.

EL TRABAJO QUE ESTO REEMPLAZA. En cada mesa de conciliación el área
diligencia un Excel (formato ACTA SINAC: una fila por glosa, con lo que
acepta la IPS, lo que levanta la entidad y lo que queda ratificado). Al
final alguien tiene que: cuadrar que lo repartido no supere lo glosado,
calcular el resultado de cada línea, refrescar los indicadores y armar el
PDF firmable. Todo eso era manual y por eso el acta de las 147 facturas
tardaba días en cerrarse.

Acá el mismo Excel entra tal cual (con sus macros, logos y hojas ocultas) y
sale: revisado (hallazgos concretos por fila), optimizado (resultado por
línea + hoja REVISION + indicadores recalculados) y el acta lista para
imprimir y firmar.

REGLA ARITMÉTICA DE CADA LÍNEA, que es lo que la mesa concilia:

    GLOSA INICIAL = ACEPTA IPS + LEVANTA ENTIDAD + RATIFICADO + PENDIENTE

Si una fila reparte más de lo glosado, o dice conciliada sin texto de
conciliación, eso es un hallazgo — se informa, no se corrige en silencio.

Lo que escribió el auditor NO se toca: la optimización agrega (columna de
resultado, hoja REVISION, indicadores), nunca sobreescribe una decisión.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any, Optional

from app.utils.moneda import parse_valor_cop

# Tolerancia de cuadre en pesos: las actas reales traen redondeos de $1.
_TOL = 1.0

# Cuántos caracteres de descripción se llevan al PDF (el Excel guarda
# párrafos enteros; el acta impresa necesita la esencia, no la novela).
_MAX_DESC_PDF = 420


@dataclass
class LineaActa:
    fila_excel: int
    item: str = ""
    radicado: str = ""
    factura: str = ""
    fecha_factura: str = ""
    tipo_glosa: str = ""
    tipificacion: str = ""
    cod_glosa: str = ""
    descripcion: str = ""
    valor_factura: float = 0.0
    glosa_inicial: float = 0.0
    pendiente: float = 0.0
    acepta_ips: float = 0.0
    levanta_entidad: float = 0.0
    ratificado: float = 0.0
    texto_conciliacion: str = ""

    @property
    def repartido(self) -> float:
        return self.acepta_ips + self.levanta_entidad + self.ratificado

    @property
    def resultado(self) -> str:
        """El veredicto de la línea, derivado de los valores de la mesa."""
        if self.glosa_inicial <= 0:
            return "SIN GLOSA"
        if self.repartido <= _TOL:
            return "PENDIENTE"
        if abs(self.levanta_entidad - self.glosa_inicial) <= _TOL:
            return "LEVANTADA TOTAL"
        if abs(self.ratificado - self.glosa_inicial) <= _TOL:
            return "RATIFICADA TOTAL"
        if abs(self.acepta_ips - self.glosa_inicial) <= _TOL:
            return "ACEPTADA POR LA IPS"
        if self.repartido < self.glosa_inicial - _TOL:
            return "PARCIAL (queda pendiente)"
        return "CONCILIADA"


@dataclass
class Acta:
    nit: str = ""
    razon_social: str = ""
    periodo: str = ""
    cantidad_facturas_declarada: Optional[int] = None
    valor_a_conciliar_declarado: Optional[float] = None
    lineas: list[LineaActa] = field(default_factory=list)
    total_excel: dict[str, float] = field(default_factory=dict)
    observaciones: list[str] = field(default_factory=list)
    firmas: list[dict[str, str]] = field(default_factory=list)
    hoja: str = "ACTA"
    fila_encabezado: int = 0
    columnas: dict[str, int] = field(default_factory=dict)
    fila_total: Optional[int] = None


# Etiquetas del encabezado de la tabla → campo. Se compara por prefijo
# normalizado, porque el formato real trae variantes («VALRO», dobles
# espacios, saltos de línea).
_MAPA_COLUMNAS: tuple[tuple[str, str], ...] = (
    ("ITEM", "item"),
    ("RADICADO", "radicado"),
    ("NUMERO FACTURA", "factura"),
    ("FECHA FACTURA", "fecha_factura"),
    ("TIPO DE GLOSA", "tipo_glosa"),
    ("TIPIFICACION", "tipificacion"),
    ("COD GLOSA", "cod_glosa"),
    ("DESCRIPCION GLOSA", "descripcion"),
    ("VALOR FACTURA", "valor_factura"),
    ("VALOR GLOSA INICIAL", "glosa_inicial"),
    ("VALOR PENDIENTE", "pendiente"),
    ("VALOR ACEPTA IPS", "acepta_ips"),
    ("VALOR LEVANTA", "levanta_entidad"),
    ("VALOR RATIFICADO", "ratificado"),
    ("DESCRIPCION DE CONCILIACION", "texto_conciliacion"),
)

_CAMPOS_DINERO = {
    "valor_factura",
    "glosa_inicial",
    "pendiente",
    "acepta_ips",
    "levanta_entidad",
    "ratificado",
}


def _norm(texto: Any) -> str:
    return re.sub(r"\s+", " ", str(texto or "")).strip().upper()


def _dinero(valor: Any) -> float:
    if valor is None or valor == "":
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    try:
        return parse_valor_cop(str(valor))
    except Exception:
        return 0.0


def _texto(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor == int(valor):
        return str(int(valor))
    s = str(valor).strip()
    return "" if s in ("0", "-", "None") else s


def _vecino_derecha(ws, fila: int, col: int, tope: int = 12) -> Any:
    """El valor no vacío más cercano a la derecha de una etiqueta."""
    for c in range(col + 1, min(col + tope, ws.max_column) + 1):
        v = ws.cell(row=fila, column=c).value
        if v is not None and str(v).strip() != "":
            return v
    return None


def leer_acta(contenido: bytes) -> Acta:
    """Lee el libro tal cual lo trabaja el área. Busca por etiquetas, no por
    posiciones fijas: si contratación mueve una columna, esto la encuentra."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True, read_only=False)
    nombre_hoja = "ACTA" if "ACTA" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[nombre_hoja]
    acta = Acta(hoja=nombre_hoja)

    # ── La fila del encabezado de la tabla: la que dice ITEM ──
    fila_hdr = 0
    for r in range(1, min(60, ws.max_row) + 1):
        for c in range(1, min(20, ws.max_column) + 1):
            if _norm(ws.cell(row=r, column=c).value) == "ITEM":
                fila_hdr = r
                break
        if fila_hdr:
            break
    if not fila_hdr:
        raise ValueError(
            f"La hoja {nombre_hoja} no trae la tabla del acta (no se encontró la columna ITEM)."
        )
    acta.fila_encabezado = fila_hdr

    columnas: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        etiqueta = _norm(ws.cell(row=fila_hdr, column=c).value)
        if not etiqueta:
            continue
        for prefijo, campo in _MAPA_COLUMNAS:
            if etiqueta.startswith(prefijo) and campo not in columnas:
                columnas[campo] = c
                break
    faltan = {"factura", "glosa_inicial"} - set(columnas)
    if faltan:
        raise ValueError(f"Al acta le faltan columnas esperadas: {', '.join(sorted(faltan))}.")
    acta.columnas = columnas

    # ── Datos del encabezado del acta (NIT, razón social, periodo…) ──
    for r in range(1, fila_hdr):
        for c in range(1, min(30, ws.max_column) + 1):
            et = _norm(ws.cell(row=r, column=c).value)
            if not et:
                continue
            if et.startswith("NIT") and not acta.nit:
                acta.nit = _texto(_vecino_derecha(ws, r, c))
            elif et.startswith("RAZON SOCIAL") and not acta.razon_social:
                acta.razon_social = _texto(_vecino_derecha(ws, r, c))
            elif et.startswith("PERIODO") and not acta.periodo:
                acta.periodo = _texto(_vecino_derecha(ws, r, c))
            elif et.startswith("CANTIDAD FACTURAS") and acta.cantidad_facturas_declarada is None:
                v = _vecino_derecha(ws, r, c)
                try:
                    acta.cantidad_facturas_declarada = int(_dinero(v))
                except Exception:
                    pass
            elif et.startswith("VALOR A CONCILIAR") and acta.valor_a_conciliar_declarado is None:
                acta.valor_a_conciliar_declarado = _dinero(_vecino_derecha(ws, r, c))

    # ── Las líneas, hasta la fila TOTAL o hasta que se acaben ──
    col_item = columnas.get("item", 0)
    col_factura = columnas["factura"]
    # La marca TOTAL puede vivir en cualquier columna de la izquierda de la
    # tabla (en el acta real está en la de código de glosa).
    cols_marca = sorted(set(range(1, 13)) | set(columnas.values()))
    for r in range(fila_hdr + 1, ws.max_row + 1):
        marcas = {_norm(ws.cell(row=r, column=c).value) for c in cols_marca}
        if any(m.startswith("TOTAL") for m in marcas if m):
            acta.fila_total = r
            for campo in _CAMPOS_DINERO:
                if campo in columnas:
                    acta.total_excel[campo] = _dinero(ws.cell(row=r, column=columnas[campo]).value)
            break
        factura = _texto(ws.cell(row=r, column=col_factura).value)
        item = _texto(ws.cell(row=r, column=col_item).value) if col_item else ""
        # Una línea del acta tiene factura, o al menos un ITEM numérico;
        # los textos sueltos de cierre no son líneas de glosa.
        if not factura and not re.fullmatch(r"\d+", item or ""):
            continue
        linea = LineaActa(fila_excel=r)
        for campo, c in columnas.items():
            crudo = ws.cell(row=r, column=c).value
            if campo in _CAMPOS_DINERO:
                setattr(linea, campo, _dinero(crudo))
            else:
                setattr(linea, campo, _texto(crudo))
        # Celdas combinadas: cuando la factura agrupa varias glosas, el
        # número solo está en la primera fila del grupo. Se hereda.
        if not linea.factura and acta.lineas:
            linea.factura = acta.lineas[-1].factura
        acta.lineas.append(linea)

    # ── Observaciones y firmas al pie (haya o no fila TOTAL) ──
    fin_tabla = acta.fila_total or (acta.lineas[-1].fila_excel if acta.lineas else fila_hdr)
    firma_actual: dict[str, str] = {}
    for r in range(fin_tabla + 1, ws.max_row + 1):
        for c in range(1, min(20, ws.max_column) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None or not str(v).strip():
                continue
            et = _norm(v)
            if et.startswith("NOMBRE"):
                if firma_actual.get("nombre"):
                    acta.firmas.append(firma_actual)
                    firma_actual = {}
                firma_actual["nombre"] = _texto(_vecino_derecha(ws, r, c))
            elif et.startswith("CARGO"):
                firma_actual["cargo"] = _texto(_vecino_derecha(ws, r, c))
            elif et.startswith("CORREO"):
                firma_actual["correo"] = _texto(_vecino_derecha(ws, r, c))
            elif et.startswith(("ESTA ACTA", "EN CONCILIACION")):
                texto = _texto(v)
                if len(texto) > 20:
                    acta.observaciones.append(texto)
    if firma_actual.get("nombre"):
        acta.firmas.append(firma_actual)

    return acta


# ─── Revisión: los hallazgos que hoy se buscan a ojo ─────────────────────


def revisar(acta: Acta) -> dict[str, Any]:
    """Cuadra el acta y devuelve el resumen con hallazgos concretos.

    Un hallazgo señala fila y factura: «la fila 87 reparte $200.000 más de
    lo glosado» se corrige en un minuto; «el acta no cuadra» se corrige en
    una tarde.
    """
    hallazgos: list[dict[str, Any]] = []

    def h(linea: Optional[LineaActa], problema: str, detalle: str = "") -> None:
        hallazgos.append(
            {
                "fila_excel": linea.fila_excel if linea else None,
                "factura": linea.factura if linea else "",
                "problema": problema,
                "detalle": detalle,
            }
        )

    tot = {c: 0.0 for c in _CAMPOS_DINERO}
    facturas: dict[str, float] = {}
    valor_por_factura: dict[str, float] = {}
    conciliadas = pendientes = 0

    for ln in acta.lineas:
        for campo in _CAMPOS_DINERO:
            tot[campo] += getattr(ln, campo)
        facturas.setdefault(ln.factura, 0.0)
        facturas[ln.factura] += ln.glosa_inicial
        # El valor de la factura viene repetido en cada una de sus glosas:
        # para el total se cuenta UNA vez por factura, como hace el acta.
        valor_por_factura[ln.factura] = max(
            valor_por_factura.get(ln.factura, 0.0), ln.valor_factura
        )

        if min(ln.acepta_ips, ln.levanta_entidad, ln.ratificado, ln.glosa_inicial) < 0:
            h(ln, "valor negativo", "ningún valor del acta puede ser negativo")
        if ln.repartido > ln.glosa_inicial + _TOL:
            h(
                ln,
                "reparte más de lo glosado",
                f"acepta+levanta+ratifica=${ln.repartido:,.0f} y la glosa inicial "
                f"es ${ln.glosa_inicial:,.0f}",
            )
        # El área usa dos convenciones válidas para la columna de pendiente:
        # lo que ENTRÓ a mesa (= glosa inicial) o lo que QUEDÓ después
        # (= glosa − repartido). Solo es hallazgo si no es ninguna de las dos.
        pendiente_calc = max(ln.glosa_inicial - ln.repartido, 0.0)
        if (
            abs(ln.pendiente - pendiente_calc) > _TOL
            and abs(ln.pendiente - ln.glosa_inicial) > _TOL
        ):
            h(
                ln,
                "el pendiente no cuadra",
                f"dice ${ln.pendiente:,.0f}; lo glosado fue ${ln.glosa_inicial:,.0f} "
                f"y tras la mesa quedaría ${pendiente_calc:,.0f}",
            )
        if ln.repartido > _TOL and not ln.texto_conciliacion.strip():
            h(ln, "conciliada sin texto", "tiene valores de mesa pero no dice qué se acordó")
        if ln.repartido > _TOL:
            conciliadas += 1
        else:
            pendientes += 1

    # El valor de la factura se cuenta una vez por factura, como hace el acta.
    tot["valor_factura"] = sum(valor_por_factura.values())

    # Cuadre contra lo que el acta declara de sí misma
    if acta.cantidad_facturas_declarada and acta.cantidad_facturas_declarada != len(facturas):
        h(
            None,
            "la cantidad de facturas no cuadra",
            f"el encabezado declara {acta.cantidad_facturas_declarada} y las líneas "
            f"suman {len(facturas)} facturas distintas",
        )
    if (
        acta.valor_a_conciliar_declarado
        and abs(acta.valor_a_conciliar_declarado - tot["glosa_inicial"]) > _TOL
    ):
        h(
            None,
            "el valor a conciliar no cuadra",
            f"el encabezado declara ${acta.valor_a_conciliar_declarado:,.0f} y las líneas "
            f"suman ${tot['glosa_inicial']:,.0f}",
        )
    for campo, declarado in acta.total_excel.items():
        if declarado and abs(declarado - tot[campo]) > _TOL:
            h(
                None,
                f"la fila TOTAL no cuadra en {campo}",
                f"dice ${declarado:,.0f} y las líneas suman ${tot[campo]:,.0f}",
            )

    return {
        "facturas": len(facturas),
        "lineas": len(acta.lineas),
        "lineas_conciliadas": conciliadas,
        "lineas_pendientes": pendientes,
        "glosado": tot["glosa_inicial"],
        "acepta_ips": tot["acepta_ips"],
        "levanta_entidad": tot["levanta_entidad"],
        "ratificado": tot["ratificado"],
        "pendiente_calculado": max(
            tot["glosa_inicial"] - tot["acepta_ips"] - tot["levanta_entidad"] - tot["ratificado"],
            0.0,
        ),
        "hallazgos": hallazgos,
        "lista_para_firmar": not hallazgos,
    }


# ─── Optimización: lo que se agrega al libro, sin tocar lo escrito ───────


def optimizar(contenido: bytes, es_xlsm: bool = True) -> tuple[bytes, dict[str, Any]]:
    """Devuelve el mismo libro con tres agregados y el resumen de revisión.

    1. Columna «RESULTADO DE LA MESA (auto)» al final de la tabla.
    2. Hoja REVISION con los hallazgos, para corregir antes de firmar.
    3. Indicadores del DASHBOARD recalculados (si la hoja existe).

    Los valores y textos del auditor quedan intactos.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    acta = leer_acta(contenido)
    resumen = revisar(acta)

    wb = openpyxl.load_workbook(io.BytesIO(contenido), keep_vba=es_xlsm)
    ws = wb[acta.hoja]

    # 1) Columna de resultado por línea (se reutiliza si ya existe)
    titulo = "RESULTADO DE LA MESA (auto)"
    col_res = None
    for c in range(1, ws.max_column + 1):
        if _norm(ws.cell(row=acta.fila_encabezado, column=c).value) == _norm(titulo):
            col_res = c
            break
    if col_res is None:
        col_res = ws.max_column + 1
    celda_hdr = ws.cell(row=acta.fila_encabezado, column=col_res, value=titulo)
    celda_hdr.font = Font(bold=True, size=9)
    celda_hdr.alignment = Alignment(wrap_text=True, vertical="center")
    colores = {
        "LEVANTADA TOTAL": "C6EFCE",
        "ACEPTADA POR LA IPS": "FFEB9C",
        "RATIFICADA TOTAL": "FFC7CE",
        "PARCIAL (queda pendiente)": "FFEB9C",
        "CONCILIADA": "C6EFCE",
        "PENDIENTE": "D9D9D9",
    }
    for ln in acta.lineas:
        celda = ws.cell(row=ln.fila_excel, column=col_res, value=ln.resultado)
        color = colores.get(ln.resultado)
        if color:
            celda.fill = PatternFill("solid", fgColor=color)

    # 2) Hoja REVISION (se rehace en cada corrida)
    if "REVISION" in wb.sheetnames:
        del wb["REVISION"]
    rev = wb.create_sheet("REVISION")
    rev["A1"] = "REVISIÓN AUTOMÁTICA DEL ACTA"
    rev["A1"].font = Font(bold=True, size=12)
    filas_resumen = [
        ("Facturas en el acta", resumen["facturas"]),
        ("Líneas de glosa", resumen["lineas"]),
        ("Líneas con resultado de mesa", resumen["lineas_conciliadas"]),
        ("Líneas pendientes", resumen["lineas_pendientes"]),
        ("Valor glosado", resumen["glosado"]),
        ("Acepta la IPS", resumen["acepta_ips"]),
        ("Levanta la entidad", resumen["levanta_entidad"]),
        ("Ratificado", resumen["ratificado"]),
        ("Pendiente por conciliar", resumen["pendiente_calculado"]),
        ("¿Lista para firmar?", "SÍ" if resumen["lista_para_firmar"] else "NO"),
    ]
    for i, (et, v) in enumerate(filas_resumen, start=3):
        rev.cell(row=i, column=1, value=et).font = Font(bold=True)
        rev.cell(row=i, column=2, value=v)
    base = len(filas_resumen) + 4
    rev.cell(row=base, column=1, value="HALLAZGOS (corregir antes de firmar)").font = Font(
        bold=True
    )
    encabezados = ("Fila en la hoja ACTA", "Factura", "Problema", "Detalle")
    for j, et in enumerate(encabezados, start=1):
        c = rev.cell(row=base + 1, column=j, value=et)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="FDE9D9")
    if resumen["hallazgos"]:
        for i, hz in enumerate(resumen["hallazgos"], start=base + 2):
            rev.cell(row=i, column=1, value=hz["fila_excel"])
            rev.cell(row=i, column=2, value=hz["factura"])
            rev.cell(row=i, column=3, value=hz["problema"])
            rev.cell(row=i, column=4, value=hz["detalle"])
    else:
        rev.cell(row=base + 2, column=1, value="Ninguno: el acta cuadra completa.")
    rev.column_dimensions["A"].width = 24
    rev.column_dimensions["B"].width = 18
    rev.column_dimensions["C"].width = 34
    rev.column_dimensions["D"].width = 70

    # 3) DASHBOARD: solo los indicadores que sabemos recalcular
    if "DASHBOARD" in wb.sheetnames:
        dash = wb["DASHBOARD"]
        actualizables = {
            "LEVANTADO": resumen["levanta_entidad"],
            "RATIFICADO": resumen["ratificado"],
            "PENDIENTE POR CONCILIAR": resumen["pendiente_calculado"],
        }
        for r in range(1, min(60, dash.max_row) + 1):
            for c in range(1, min(8, dash.max_column) + 1):
                et = _norm(dash.cell(row=r, column=c).value)
                if not et:
                    continue
                for clave, valor in actualizables.items():
                    if clave in et:
                        vecino_col = None
                        for c2 in range(c + 1, min(c + 8, dash.max_column) + 1):
                            if dash.cell(row=r, column=c2).value is not None:
                                vecino_col = c2
                                break
                        if vecino_col:
                            dash.cell(row=r, column=vecino_col, value=valor)
                        break

    salida = io.BytesIO()
    wb.save(salida)
    return salida.getvalue(), resumen


# ─── El acta imprimible (misma cara que el acta SINAC del sistema) ───────


def _cop(v: float) -> str:
    try:
        return "$" + f"{float(v or 0):,.0f}".replace(",", ".")
    except Exception:
        return "$0"


def _esc(s: str) -> str:
    return (
        str(s or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def html_acta(acta: Acta, resumen: Optional[dict[str, Any]] = None) -> str:
    """El acta firmable como HTML imprimible (Ctrl+P → PDF), con los datos
    y las firmas leídos del propio Excel de la mesa."""
    resumen = resumen or revisar(acta)

    filas_html = []
    for i, ln in enumerate(acta.lineas, start=1):
        desc = _esc(ln.descripcion[:_MAX_DESC_PDF])
        conc = _esc(ln.texto_conciliacion[:_MAX_DESC_PDF])
        filas_html.append(
            "<tr>"
            f"<td class='c'>{i}</td>"
            f"<td>{_esc(ln.radicado) or '—'}</td>"
            f"<td>{_esc(ln.factura)}</td>"
            f"<td>{_esc(ln.cod_glosa) or '—'}</td>"
            f"<td class='desc'>{desc}</td>"
            f"<td class='r'>{_cop(ln.glosa_inicial)}</td>"
            f"<td class='r'>{_cop(ln.acepta_ips)}</td>"
            f"<td class='r'>{_cop(ln.levanta_entidad)}</td>"
            f"<td class='r'>{_cop(ln.ratificado)}</td>"
            f"<td class='desc'>{conc}</td>"
            f"<td class='res'>{_esc(ln.resultado)}</td>"
            "</tr>"
        )

    # La cláusula de mérito ejecutivo va SIEMPRE (es el respaldo legal del
    # acta); si el Excel ya la trae en sus observaciones, no se repite.
    obs_propias = [o for o in acta.observaciones if "MERITO EJECUTIVO" not in _norm(o)]
    obs_html = "".join(f"<p>{_esc(o)}</p>" for o in obs_propias) or "<p>—</p>"
    clausula = (
        "Esta acta presta mérito ejecutivo para todos los efectos civiles, en el sentido "
        "del Art. 422 del Código General del Proceso (Ley 1564 de 2012) y el Art. 56 de la "
        "Ley 1438 de 2011. Las partes reconocen el contenido aquí consignado como una "
        "expresión consensuada y definitiva sobre los conceptos conciliados. De no lograrse "
        "acuerdo en alguno de los ítems, las partes podrán elevar el conflicto ante la "
        "Superintendencia Nacional de Salud (Art. 126 Ley 1438/2011)."
    )

    firmas_html = []
    for f in acta.firmas[:4]:
        firmas_html.append(
            "<div class='firma'>"
            "<div class='titulo'>FIRMA</div>"
            f"<div class='campo'><span class='lbl'>Nombre:</span> {_esc(f.get('nombre', ''))}</div>"
            f"<div class='campo'><span class='lbl'>Cargo:</span> {_esc(f.get('cargo', ''))}</div>"
            f"<div class='campo'><span class='lbl'>Correo:</span> {_esc(f.get('correo', ''))}</div>"
            "<div class='campo' style='height:46px'><span class='lbl'>Firma:</span></div>"
            "</div>"
        )
    if not firmas_html:
        firmas_html = [
            "<div class='firma'><div class='titulo'>POR LA IPS</div>"
            "<div class='campo' style='height:70px'></div></div>",
            "<div class='firma'><div class='titulo'>POR LA ENTIDAD</div>"
            "<div class='campo' style='height:70px'></div></div>",
        ]

    aviso = ""
    if resumen["hallazgos"]:
        aviso = (
            f"<div class='alerta'>⚠ Esta acta tiene {len(resumen['hallazgos'])} "
            "hallazgo(s) de cuadre sin resolver (ver hoja REVISION del Excel "
            "optimizado). Se puede imprimir, pero conviene corregir antes de firmar.</div>"
        )

    return f"""<!doctype html>
<html lang="es"><head><meta charset="utf-8">
<title>Acta de conciliación · {_esc(acta.razon_social) or "Entidad"}</title>
<style>
  @page {{ size: Letter landscape; margin: 1.2cm 1cm; }}
  body {{ font-family: Arial, sans-serif; color: #0f172a; font-size: 9pt; line-height: 1.35; }}
  .hdr {{ text-align: center; margin-bottom: 12px; padding-bottom: 10px; border-bottom: 2px solid #0b5d8a; }}
  .hdr .marca {{ color: #64748b; font-size: 8pt; letter-spacing: .4px; }}
  .hdr h1 {{ margin: 6px 0 2px; font-size: 13pt; color: #0b5d8a; }}
  .info {{ width: 100%; border-collapse: collapse; margin: 8px 0 12px; font-size: 9pt; }}
  .info td {{ padding: 4px 8px; border: 1px solid #cbd5e1; }}
  .info .lbl {{ background: #f1f5f9; font-weight: 700; color: #334155; width: 15%; }}
  .alerta {{ background: #fef3c7; border: 1px solid #f59e0b; padding: 8px 12px; margin: 8px 0; font-size: 9pt; }}
  .tabla {{ width: 100%; border-collapse: collapse; font-size: 7.5pt; margin: 4px 0 12px; }}
  .tabla th {{ background: #0b5d8a; color: white; padding: 5px 4px; border: 1px solid #0b5d8a; font-size: 7pt; text-transform: uppercase; }}
  .tabla td {{ padding: 3px 4px; border: 1px solid #cbd5e1; vertical-align: top; }}
  .tabla td.c {{ text-align: center; font-weight: 700; }}
  .tabla td.r {{ text-align: right; font-variant-numeric: tabular-nums; white-space: nowrap; }}
  .tabla td.desc {{ font-size: 7pt; }}
  .tabla td.res {{ font-size: 7pt; font-weight: 700; white-space: nowrap; }}
  .total td {{ background: #fef3c7; font-weight: 700; }}
  .obs {{ border: 1px solid #cbd5e1; padding: 8px 12px; margin: 6px 0 14px; background: #f8fafc; font-size: 8pt; }}
  .firmas {{ display: flex; gap: 24px; margin-top: 26px; }}
  .firma {{ flex: 1; }}
  .firma .titulo {{ background: #0b5d8a; color: white; padding: 5px 8px; font-weight: 700; font-size: 9pt; text-align: center; }}
  .firma .campo {{ border: 1px solid #cbd5e1; padding: 5px 8px; font-size: 9pt; }}
  .firma .lbl {{ font-weight: 700; color: #334155; display: inline-block; min-width: 62px; }}
  button.noprint {{ position: fixed; top: 10px; right: 10px; padding: 8px 14px; background: #0b5d8a; color: white; border: 0; border-radius: 6px; cursor: pointer; z-index: 10; }}
  @media print {{ button.noprint {{ display: none; }} }}
</style>
</head><body>
<button class="noprint" onclick="window.print()">Imprimir / Guardar PDF</button>

<div class="hdr">
  <div class="marca">SINAC S.C — ESE HOSPITAL UNIVERSITARIO DE SANTANDER · NIT 900.006.037-4</div>
  <h1>ACTA DE CONCILIACIÓN DE GLOSAS</h1>
</div>

<table class="info">
  <tr>
    <td class="lbl">Entidad</td><td>{_esc(acta.razon_social) or "—"}</td>
    <td class="lbl">NIT</td><td>{_esc(acta.nit) or "—"}</td>
    <td class="lbl">Periodo</td><td>{_esc(acta.periodo) or "—"}</td>
  </tr>
  <tr>
    <td class="lbl">Facturas</td><td>{resumen["facturas"]}</td>
    <td class="lbl">Líneas de glosa</td><td>{resumen["lineas"]}</td>
    <td class="lbl">Valor a conciliar</td><td>{_cop(resumen["glosado"])}</td>
  </tr>
  <tr>
    <td class="lbl">Acepta la IPS</td><td>{_cop(resumen["acepta_ips"])}</td>
    <td class="lbl">Levanta la entidad</td><td>{_cop(resumen["levanta_entidad"])}</td>
    <td class="lbl">Ratificado / pendiente</td>
    <td>{_cop(resumen["ratificado"])} / {_cop(resumen["pendiente_calculado"])}</td>
  </tr>
</table>
{aviso}
<table class="tabla">
  <thead><tr>
    <th>#</th><th>Radicado</th><th>Factura</th><th>Cód.</th><th>Descripción de la glosa</th>
    <th>Glosa inicial</th><th>Acepta IPS</th><th>Levanta entidad</th><th>Ratificado</th>
    <th>Conciliación</th><th>Resultado</th>
  </tr></thead>
  <tbody>
    {"".join(filas_html)}
    <tr class="total">
      <td colspan="5" style="text-align:right">TOTALES</td>
      <td class="r">{_cop(resumen["glosado"])}</td>
      <td class="r">{_cop(resumen["acepta_ips"])}</td>
      <td class="r">{_cop(resumen["levanta_entidad"])}</td>
      <td class="r">{_cop(resumen["ratificado"])}</td>
      <td colspan="2"></td>
    </tr>
  </tbody>
</table>

<div class="obs"><b>Observaciones:</b>{obs_html}</div>

<div class="obs" style="background:#fef2f2;border-left:3px solid #b91c1c">{_esc(clausula)}</div>

<div class="firmas">{"".join(firmas_html)}</div>
</body></html>"""
