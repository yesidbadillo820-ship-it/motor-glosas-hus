"""Parser de Excel de tarifas contratadas.

Soporta múltiples formatos:
  - Famisanar 3 hojas (Anexo 3 / 3.1 / 3.2)
  - Dispensario (hoja plana con CUPS + PRECIO DE REFERENCIA)
  - Otros formatos genéricos con CUPS + valor fijo

Usa openpyxl en modo ``read_only`` + ``iter_rows()`` para manejar archivos
grandes sin timeouts ni OOM (Famisanar real ~5k–15k filas).

Exporta:
  parsear_excel_tarifas(bytes, filename) -> dict con:
    - eps, contrato, vigencia_desde, vigencia_hasta
    - filas: list[dict] {codigo_cups, descripcion, valor_pactado, modalidad,
      tipo_tarifa, factor_ajuste, observacion}
    - hojas_detectadas: list[str]
    - errores: list[str]
"""
from __future__ import annotations

import re
import unicodedata
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


# ─── Normalización ──────────────────────────────────────────────────────────

def _normalizar_texto(s: Any) -> str:
    """Quita tildes, pasa a mayúsculas, colapsa espacios."""
    if s is None:
        return ""
    t = str(s).strip()
    t = unicodedata.normalize("NFKD", t)
    t = "".join(c for c in t if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", t).upper()


def _normalizar_valor(v: Any) -> float:
    """Parsea valores COP desde celda Excel (puede venir número o string)."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(" ", "")
    if not s or s.upper() in ("N/A", "NA", "-"):
        return 0.0
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        m = re.match(r"^(\d+)[\.,](\d{1,2})$", s)
        if m:
            s = f"{m.group(1)}.{m.group(2)}"
        else:
            s = s.replace(".", "").replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parsear_fecha(v: Any) -> datetime | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    # Fix typo común "14//04/2027"
    s = re.sub(r"/+", "/", s)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _parsear_porcentaje(v: Any) -> float:
    """Parsea '-5%', '-15%', '+10%', 0.05 → -5.0, -15.0, 10.0, 5.0."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        f = float(v)
        if -1 < f < 1 and f != 0:
            return f * 100
        return f
    s = str(v).strip().replace("%", "").replace(" ", "").replace(",", ".")
    if not s:
        return 0.0
    try:
        f = float(s)
        if -1 < f < 1 and f != 0:
            return f * 100
        return f
    except ValueError:
        return 0.0


# ─── Helpers que operan sobre rows (lista de tuplas) ────────────────────────

def _celda(fila: tuple, idx: int | None) -> Any:
    if idx is None or idx < 0 or idx >= len(fila):
        return None
    return fila[idx]


def _indice_columna(headers: list[str], *candidatos: str) -> int | None:
    """Busca el índice del primer header que coincida con algún candidato.

    Dos pasadas:
      1. Match exacto normalizado (ej. header "CUPS" == candidato "CUPS").
      2. Candidato aparece como palabra/frase completa dentro del header
         (ej. header "APLICA IVA (SI-NO)" contiene candidato "APLICA IVA").
         Solo aplica para candidatos de ≥6 caracteres, para evitar que
         "CUPS" (4 chars) coincida con "DESCRIPCION CUPS".
    """
    cands_norm = [_normalizar_texto(c) for c in candidatos if c]
    # Pasada 1: match exacto
    for i, h in enumerate(headers):
        if h and h in cands_norm:
            return i
    # Pasada 2: candidato como subcadena con frontera (solo candidatos largos)
    for cand in cands_norm:
        if len(cand) < 6:
            continue
        pat = re.compile(r"(?:^|\s|/|-)" + re.escape(cand) + r"(?:$|\s|/|-|\()")
        for i, h in enumerate(headers):
            if h and pat.search(h):
                return i
    return None


# ─── Detección de hojas ─────────────────────────────────────────────────────

def _tipo_hoja(headers_normalizados: list[str]) -> str | None:
    """Devuelve 'ANEXO3' | 'ANEXO31' | 'ANEXO32' | 'SIMPLE_FIJO'.

    Los formatos específicos se evalúan primero para evitar que un Excel
    tipo Famisanar caiga en el fallback SIMPLE_FIJO.
    """
    hset = {h for h in headers_normalizados if h}
    hunion = " ".join(hset)
    cups_like = any("CUPS" in h for h in hset)

    # 3.1 medicamentos — CODIGO DCI + CUM/IUM (muy distintivo)
    if {"CODIGO DCI", "DESCRIPCION DCI"} & hset and "CUM/IUM" in hset:
        return "ANEXO31"
    # 3.2 suministros — TARIFA FINAL + MAPIISS, o CODIGO DEL PRESTADOR + DESCRIPCION DEL PRESTADOR
    if "TARIFA FINAL" in hunion and "MAPIISS" in hset:
        return "ANEXO32"
    if "CODIGO DEL PRESTADOR" in hset and "DESCRIPCION DEL PRESTADOR" in hset:
        return "ANEXO32"
    # Anexo 3 servicios — CUPS + TIPO TARIFA + HOSPITALARIO/AMBULATORIO/URGENCIA
    if cups_like and "TIPO TARIFA" in hset and ("HOSPITALARIO" in hset or "AMBULATORIO" in hset):
        return "ANEXO3"
    # Fallback plano: hoja simple con CUPS + algún campo de valor fijo.
    # Cubre Dispensario (PRECIO DE REFERENCIA), Nueva EPS, Sanitas simple,
    # Compensar, y cualquier catálogo plano genérico.
    value_keywords = (
        "PRECIO DE REFERENCIA", "TARIFA UNITARIA", "VALOR PACTADO",
        "VALOR UNITARIO", "PRECIO UNITARIO",
    )
    has_value = any(k in hunion for k in value_keywords)
    if not has_value:
        has_value = any(h in {"VALOR", "PRECIO", "TARIFA"} for h in hset)
    if cups_like and has_value:
        return "SIMPLE_FIJO"
    return None


def _buscar_fila_encabezado(rows: list[tuple]) -> tuple[int, list[str]] | tuple[None, None]:
    """Escanea las primeras 200 filas buscando la fila de encabezados.

    Famisanar Anexo 3 tiene ~50 filas de metadata (directorio, sedes,
    agrupadores RIAS, agrupadores servicios) antes del encabezado real
    de la tabla CUPS. Además, la fila de encabezados tiene que contener
    palabras clave SUFICIENTES (≥2) para evitar falsos positivos de
    tablas auxiliares como AGRUPADORES (que tiene COD + DESCRIPCION).
    """
    keywords_fuertes = [
        "CUPS", "MAPIISS", "TIPO TARIFA",
        "CODIGO DEL PRESTADOR", "CODIGO DCI", "CUM/IUM",
        "DESCRIPCION DEL PRESTADOR", "PRECIO DE REFERENCIA",
        "TARIFA UNITARIA", "TARIFA FINAL",
    ]
    limite = min(200, len(rows))
    for idx, fila in enumerate(rows[:limite]):
        no_vacios = [c for c in fila if c is not None and str(c).strip()]
        if len(no_vacios) < 4:
            continue
        fila_norm = [_normalizar_texto(c) for c in fila]
        matches = sum(
            1 for celda in fila_norm
            if any(kw in celda for kw in keywords_fuertes)
        )
        if matches >= 2:
            return idx, fila_norm
    return None, None


def _extraer_metadata(rows: list[tuple]) -> dict:
    """Extrae eps, contrato, vigencia desde las primeras 50 filas."""
    meta = {"eps": None, "contrato": None, "vigencia_desde": None, "vigencia_hasta": None}
    labels_ruido = {
        "CORREO:", "CEDULA:", "CEDULA", "TELEFONO:", "TELEFONO",
        "CORREO", "HORARIO DE ATENCION", "CORREO ELECTRONICO",
    }
    for fila in rows[:50]:
        for col_idx, celda in enumerate(fila):
            if not celda:
                continue
            etiqueta = _normalizar_texto(celda)

            def valor_derecha(offset_min: int = 1) -> Any:
                fin = min(len(fila), col_idx + offset_min + 14)
                for c in range(col_idx + offset_min, fin):
                    v = fila[c]
                    if v is not None and str(v).strip() and _normalizar_texto(v) not in labels_ruido:
                        return v
                return None

            if not meta["eps"] and "NOMBRE DE LA EPS" in etiqueta:
                v = valor_derecha()
                if v:
                    meta["eps"] = str(v).strip()
            elif not meta["contrato"] and (
                "NUMERO DE CONTRATO" in etiqueta or "NÚMERO DE CONTRATO" in etiqueta
            ):
                v = valor_derecha()
                if v and _normalizar_texto(v) not in ("N/A", "NA", "-"):
                    meta["contrato"] = str(v).strip()
            elif "INICIO DE CONTRATO" in etiqueta or ("VIGENCIA" in etiqueta and "INICIO" in etiqueta):
                v = valor_derecha()
                f = _parsear_fecha(v)
                if f and not meta["vigencia_desde"]:
                    meta["vigencia_desde"] = f
            elif "FINAL" in etiqueta and ("VIGENCIA" in etiqueta or "CONTRATO" in etiqueta):
                v = valor_derecha()
                f = _parsear_fecha(v)
                if f and not meta["vigencia_hasta"]:
                    meta["vigencia_hasta"] = f
    return meta


# ─── Parsers por tipo de hoja ───────────────────────────────────────────────

def _parsear_anexo3(rows: list[tuple], hdr_idx: int, headers: list[str]) -> list[dict]:
    """Anexo 3 — Servicios CUPS con fórmula SOAT ± %."""
    idx_cups = _indice_columna(headers, "CUPS / CUMS / MIPRES", "CUPS/CUMS/MIPRES", "CUPS")
    idx_desc_cups = _indice_columna(headers, "DESCRIPCION CUPS / CUMS / MIPRES", "DESCRIPCION CUPS")
    idx_desc_propio = _indice_columna(headers, "DESCRIPCION CODIGO PROPIO")
    idx_desc_reps = _indice_columna(headers, "DESCRIPCION REPS")
    idx_desc = _indice_columna(headers, "DESCRIPCION", "DESCRIPCIÓN")
    idx_tipo = _indice_columna(headers, "TIPO TARIFA")
    idx_hosp = _indice_columna(headers, "HOSPITALARIO")
    idx_amb = _indice_columna(headers, "AMBULATORIO")
    idx_urg = _indice_columna(headers, "URGENCIA")
    idx_obs = _indice_columna(headers, "OBSERVACION", "OBSERVACIÓN")

    if idx_cups is None:
        return []

    filas: list[dict] = []
    for fila in rows[hdr_idx + 1:]:
        cups_raw = _celda(fila, idx_cups)
        if not cups_raw:
            continue
        cups = str(cups_raw).strip()
        # Filtro estricto: CUPS reales tienen letras/dígitos y ≤30 chars.
        # Filas de cierre como "SE SUSCRIBE EL PRESENTE ANEXO..." caen acá.
        if not _es_codigo_cups_valido(cups):
            continue
        desc = _primera_desc_no_vacia(
            fila, idx_desc_cups, idx_desc_propio, idx_desc_reps, idx_desc
        )
        tipo = _limpiar_descripcion(str(_celda(fila, idx_tipo) or ""))
        tipo_norm = _normalizar_texto(tipo)
        # Distinguir contrato SOAT (porcentajes) vs PROPIAS (valores absolutos).
        # Familia SOAT: "SOAT UVB VIGENTE", "SOAT UVB", "SOAT ISS", "SOAT"
        # Familia absoluta: "PROPIAS", "PROPIA", "MANUAL HUS", "NEGOCIADA", "TARIFA PLENA"
        es_soat = "SOAT" in tipo_norm or "UVB" in tipo_norm or "ISS" in tipo_norm
        if es_soat:
            # Los cols HOSP/AMB/URG son PORCENTAJES (-5%, +10%, 0%)
            factor = 0.0
            for idx in (idx_hosp, idx_amb, idx_urg):
                if idx is None:
                    continue
                f = _parsear_porcentaje(_celda(fila, idx))
                if f != 0:
                    factor = f
                    break
            tipo_tarifa = "SOAT_PORCENTAJE"
            valor_pactado = 0.0
            factor_ajuste = factor
        else:
            # Los cols HOSP/AMB/URG son VALORES ABSOLUTOS en COP
            valor = 0.0
            for idx in (idx_hosp, idx_amb, idx_urg):
                if idx is None:
                    continue
                v = _normalizar_valor(_celda(fila, idx))
                if v > 0:
                    valor = v
                    break
            if valor <= 0:
                continue  # sin valor, fila inválida
            tipo_tarifa = "VALOR_FIJO"
            valor_pactado = round(valor, 2)
            factor_ajuste = 0.0
        obs = _limpiar_descripcion(str(_celda(fila, idx_obs) or ""))
        filas.append({
            "codigo_cups": cups[:30],
            "codigo_ips": None,  # formato Famisanar no trae codigo_ips propio
            "descripcion": desc[:500] if desc else None,
            "valor_pactado": valor_pactado,
            "modalidad": (tipo or "SOAT UVB VIGENTE")[:80],
            "tipo_tarifa": tipo_tarifa,
            "factor_ajuste": factor_ajuste,
            "observacion": obs[:300] if obs else None,
        })
    return filas


def _es_codigo_cups_valido(cups: str) -> bool:
    """True si el string parece un código CUPS/CUM/MAPIISS/MIPRES real.

    Reglas:
      - No vacío
      - ≤30 chars (límite BD, ya no cabe en VARCHAR(30))
      - Sin espacios internos (los reales son tokens como '890202', 'FMQ6296',
        '19914262-04'). Excluye filas de cierre tipo 'SE SUSCRIBE...'.
      - Tiene al menos 1 letra o dígito
    """
    if not cups or len(cups) > 30:
        return False
    if " " in cups.strip():
        return False
    if not re.search(r"[A-Za-z0-9]", cups):
        return False
    return True


def _limpiar_descripcion(s: str) -> str:
    """Normaliza una descripción: vacía si es 'N/A', '-', 'NINGUNO' o similar."""
    if not s:
        return ""
    t = s.strip()
    if t.upper() in ("N/A", "NA", "-", "NINGUNO", "NONE", "#N/A"):
        return ""
    return t


def _primera_desc_no_vacia(fila: tuple, *indices: int | None) -> str:
    """Devuelve la primera descripción no-vacía entre los índices dados."""
    for idx in indices:
        if idx is None:
            continue
        v = _celda(fila, idx)
        d = _limpiar_descripcion(str(v) if v is not None else "")
        if d:
            return d
    return ""


def _parsear_anexo31(rows: list[tuple], hdr_idx: int, headers: list[str]) -> list[dict]:
    """Anexo 3.1 — Medicamentos, valor fijo."""
    idx_cod_prest = _indice_columna(headers, "CODIGO DEL PRESTADOR")
    idx_cum = _indice_columna(headers, "CUM/IUM", "CUM", "IUM")
    idx_mapiiss = _indice_columna(headers, "MAPIISS")
    idx_desc_dci = _indice_columna(headers, "DESCRIPCION DCI")
    idx_desc = _indice_columna(headers, "DESCRIPCION", "DESCRIPCIÓN")
    idx_desc_reps = _indice_columna(headers, "DESCRIPCION REPS")
    idx_agrup = _indice_columna(headers, "AGRUPADOR")
    idx_tarifa = _indice_columna(headers, "TARIFA UNITARIA")
    idx_iva = _indice_columna(headers, "APLICA IVA")

    idx_codigo = idx_cod_prest if idx_cod_prest is not None else (
        idx_cum if idx_cum is not None else idx_mapiiss
    )
    if idx_codigo is None or idx_tarifa is None:
        return []

    filas: list[dict] = []
    for fila in rows[hdr_idx + 1:]:
        cod_raw = _celda(fila, idx_codigo)
        if not cod_raw:
            continue
        codigo = str(cod_raw).strip()
        if not _es_codigo_cups_valido(codigo):
            continue
        # Fallback: DCI → DESCRIPCION → DESCRIPCION REPS
        desc = _primera_desc_no_vacia(fila, idx_desc_dci, idx_desc, idx_desc_reps)
        tarifa = _normalizar_valor(_celda(fila, idx_tarifa))
        if tarifa <= 0:
            continue
        iva_si = False
        if idx_iva is not None:
            iva_val = str(_celda(fila, idx_iva) or "").strip().upper()
            iva_si = iva_val in ("SI", "SÍ", "S", "YES", "Y", "1")
        valor_final = tarifa * 1.19 if iva_si else tarifa
        agrup = _limpiar_descripcion(str(_celda(fila, idx_agrup) or ""))
        filas.append({
            "codigo_cups": codigo[:30],
            "codigo_ips": None,
            "descripcion": desc[:500] if desc else None,
            "valor_pactado": round(valor_final, 2),
            "modalidad": (agrup or "MEDICAMENTOS")[:80],
            "tipo_tarifa": "VALOR_FIJO",
            "factor_ajuste": 0.0,
            "observacion": None,
        })
    return filas


def _parsear_anexo32(rows: list[tuple], hdr_idx: int, headers: list[str]) -> list[dict]:
    """Anexo 3.2 — Suministros, valor fijo (TARIFA FINAL si trae IVA)."""
    idx_cod_prest = _indice_columna(headers, "CODIGO DEL PRESTADOR")
    idx_mapiiss = _indice_columna(headers, "MAPIISS")
    idx_desc_prest = _indice_columna(headers, "DESCRIPCION DEL PRESTADOR")
    idx_desc = _indice_columna(headers, "DESCRIPCION", "DESCRIPCIÓN")
    idx_desc_reps = _indice_columna(headers, "DESCRIPCION REPS")
    idx_agrup = _indice_columna(headers, "AGRUPADOR")
    idx_unitaria = _indice_columna(headers, "TARIFA UNITARIA")
    idx_final = _indice_columna(headers, "TARIFA FINAL")
    idx_iva = _indice_columna(headers, "APLICA IVA")

    idx_codigo = idx_cod_prest if idx_cod_prest is not None else idx_mapiiss
    if idx_codigo is None:
        return []

    filas: list[dict] = []
    for fila in rows[hdr_idx + 1:]:
        cod_raw = _celda(fila, idx_codigo)
        if not cod_raw:
            continue
        codigo = str(cod_raw).strip()
        if not _es_codigo_cups_valido(codigo):
            continue
        desc = _primera_desc_no_vacia(fila, idx_desc_prest, idx_desc, idx_desc_reps)
        iva_si = False
        if idx_iva is not None:
            iva_val = str(_celda(fila, idx_iva) or "").strip().upper()
            iva_si = iva_val in ("SI", "SÍ", "S", "YES", "Y", "1")

        valor = 0.0
        if iva_si and idx_final is not None:
            valor = _normalizar_valor(_celda(fila, idx_final))
        if valor <= 0 and idx_unitaria is not None:
            valor_unit = _normalizar_valor(_celda(fila, idx_unitaria))
            valor = valor_unit * (1.19 if iva_si else 1.0)
        if valor <= 0:
            continue

        agrup = _limpiar_descripcion(str(_celda(fila, idx_agrup) or ""))
        filas.append({
            "codigo_cups": codigo[:30],
            "codigo_ips": None,
            "descripcion": desc[:500] if desc else None,
            "valor_pactado": round(valor, 2),
            "modalidad": (agrup or "SUMINISTROS")[:80],
            "tipo_tarifa": "VALOR_FIJO",
            "factor_ajuste": 0.0,
            "observacion": None,
        })
    return filas


def _parsear_simple_fijo(rows: list[tuple], hdr_idx: int, headers: list[str]) -> list[dict]:
    """Formato plano genérico: una sola hoja con CUPS + valor fijo.

    Cubre Dispensario (PRECIO DE REFERENCIA), Nueva EPS, Sanitas simple,
    Compensar, etc. La EPS y contrato normalmente se pasan por eps_override.
    """
    idx_cups = _indice_columna(headers, "CUPS")
    idx_desc = _indice_columna(
        headers, "DESCRIPCION CUPS", "DESCRIPCION IPS", "DESCRIPCION",
        "DESCRIPCIÓN", "NOMBRE"
    )
    idx_valor = _indice_columna(
        headers, "PRECIO DE REFERENCIA", "TARIFA UNITARIA", "VALOR PACTADO",
        "VALOR UNITARIO", "PRECIO UNITARIO", "VALOR", "PRECIO", "TARIFA"
    )
    idx_modalidad = _indice_columna(
        headers, "TARIFA A LA QUE CORRESPONDE EL PRECIO DE REFERENCIA",
        "TARIFA A LA QUE CORRESPONDE", "MODALIDAD", "TIPO TARIFA"
    )
    idx_cod_ips = _indice_columna(headers, "CODIGO IPS", "CODIGO PROPIO")

    if idx_cups is None or idx_valor is None:
        return []

    filas: list[dict] = []
    for fila in rows[hdr_idx + 1:]:
        cups_raw = _celda(fila, idx_cups)
        if not cups_raw:
            continue
        cups = str(cups_raw).strip()
        if not _es_codigo_cups_valido(cups):
            continue
        valor = _normalizar_valor(_celda(fila, idx_valor))
        if valor <= 0:
            continue
        desc = _limpiar_descripcion(str(_celda(fila, idx_desc) or "")) if idx_desc is not None else ""
        modalidad = _limpiar_descripcion(str(_celda(fila, idx_modalidad) or ""))
        # Si hay código IPS propio, guardarlo en observación (útil para trazabilidad)
        cod_ips = str(_celda(fila, idx_cod_ips) or "").strip() if idx_cod_ips is not None else ""
        obs = f"Código IPS: {cod_ips}" if cod_ips and cod_ips != cups else None

        filas.append({
            "codigo_cups": cups[:30],
            # Ronda 45: guardar el código IPS propio en campo indexado para
            # que el lookup pueda encontrar la tarifa cuando la EPS glose con
            # el código viejo (ej. '39147B-18' en vez de '890348').
            "codigo_ips": cod_ips[:30] if cod_ips and cod_ips != cups else None,
            "descripcion": desc[:500] if desc else None,
            "valor_pactado": round(valor, 2),
            "modalidad": (modalidad or "TARIFA PROPIA")[:80],
            "tipo_tarifa": "VALOR_FIJO",
            "factor_ajuste": 0.0,
            "observacion": obs[:300] if obs else None,
        })
    return filas


# ─── API pública ────────────────────────────────────────────────────────────

def parsear_excel_tarifas(contenido: bytes, filename: str = "") -> dict:
    """Parsea un Excel de tarifas contratadas.

    Soporta Famisanar 3 hojas, Dispensario plano, y formatos genéricos con
    columnas CUPS + valor. Modo ``read_only`` para evitar OOM con archivos
    de 10k+ filas.
    """
    errores: list[str] = []
    try:
        wb = load_workbook(BytesIO(contenido), data_only=True, read_only=True)
    except Exception as e:
        return {
            "eps": None, "contrato": None,
            "vigencia_desde": None, "vigencia_hasta": None,
            "filas": [], "hojas_detectadas": [],
            "errores": [f"No se pudo abrir el archivo: {type(e).__name__}: {e}"],
        }

    meta_global = {
        "eps": None, "contrato": None,
        "vigencia_desde": None, "vigencia_hasta": None,
    }
    filas_total: list[dict] = []
    hojas_detectadas: list[str] = []

    try:
        sheet_names = wb.sheetnames
        for sheet_name in sheet_names:
            try:
                ws = wb[sheet_name]
                # Stream rows → list una sola vez. iter_rows(values_only=True)
                # retorna tuplas de valores (no celdas), es la forma más rápida.
                rows: list[tuple] = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    continue

                meta = _extraer_metadata(rows)
                for k in meta_global:
                    if meta.get(k) and not meta_global[k]:
                        meta_global[k] = meta[k]

                hdr_idx, headers = _buscar_fila_encabezado(rows)
                if hdr_idx is None:
                    continue
                tipo = _tipo_hoja(headers)
                if tipo is None:
                    continue
                hojas_detectadas.append(f"{tipo}:{sheet_name}")
                if tipo == "ANEXO3":
                    nuevas = _parsear_anexo3(rows, hdr_idx, headers)
                elif tipo == "ANEXO31":
                    nuevas = _parsear_anexo31(rows, hdr_idx, headers)
                elif tipo == "ANEXO32":
                    nuevas = _parsear_anexo32(rows, hdr_idx, headers)
                elif tipo == "SIMPLE_FIJO":
                    nuevas = _parsear_simple_fijo(rows, hdr_idx, headers)
                else:
                    nuevas = []
                filas_total.extend(nuevas)
            except Exception as e:
                errores.append(f"Hoja '{sheet_name}': {type(e).__name__}: {e}")
                continue
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return {
        "eps": meta_global["eps"],
        "contrato": meta_global["contrato"],
        "vigencia_desde": meta_global["vigencia_desde"],
        "vigencia_hasta": meta_global["vigencia_hasta"],
        "filas": filas_total,
        "hojas_detectadas": hojas_detectadas,
        "errores": errores,
    }
