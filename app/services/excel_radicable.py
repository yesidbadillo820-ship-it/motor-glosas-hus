"""Generador del «Excel radicable» — formato institucional ESE HUS.

Jun-2026 (directiva del dueño): el archivo que se radica ante la entidad
pagadora debe replicar el formato del entregable RESPUESTA_GLOSAS_*_HUS
(dos hojas: «RESPUESTA GLOSAS» + «FUNDAMENTO JURÍDICO»), en vez del
Excel-respuesta interno (que solo anota el archivo original de
recepción). Este módulo lo construye 100 % desde la BD:

  Hoja «RESPUESTA GLOSAS»
    fila 1  título institucional merged (verde HUS, fuente blanca)
    fila 2  línea del contrato + fecha de entrega + área que elabora
    fila 3  header de 12 columnas (No., FACTURA, …, RESPUESTA ESE HUS)
    filas   UNA POR CONCEPTO/SERVICIO glosado (ConceptoGlosaRecord);
            si la glosa no tiene conceptos cargados, una fila con los
            datos de la propia GlosaRecord.
    última  TOTAL GLOSADO (suma de VR GLOSADO)

  Hoja «FUNDAMENTO JURÍDICO»
    6 secciones numeradas generadas desde ContratoRecord (NITs, SECOP,
    fechas, anexos, modalidades) + ClausulaContrato + el catálogo
    CONTRATOS_HUS como fallback. Nada de datos «quemados»: si un
    metadato no está poblado, la sección degrada a texto genérico sin
    inventar números.

Reutiliza (no duplica):
  - `dictamen_a_texto_plano` para volcar HTML → celda legible.
  - `_truncar_para_celda` (límite duro Excel 32.767 chars/celda).
  - `get_contrato` (catálogo CONTRATOS_HUS) como fallback de metadatos.
  - `_regimen_especial_texto` como única fuente de las normas por
    régimen (militar/FOMAG/PPL/ARL) — acá solo se filtran las líneas
    normativas, descartando las instrucciones dirigidas a la IA.
"""

import json
import re
from collections import Counter
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from app.core.logging_utils import logger
from app.utils.excel_seguro import sanear_celda_excel
from app.core.tz import ahora_utc
from app.models.db import (
    ClausulaContrato,
    ConceptoGlosaRecord,
    ContratoRecord,
    GlosaRecord,
    ImportacionRecepcionRecord,
)
from app.services.contexto_contractual_enriquecido import _regimen_especial_texto
from app.services.dictamen_secciones import extraer_seccion_por_codigo
from app.services.glosa_ia_prompts import get_contrato
from app.services.recepcion_excel_response import _truncar_para_celda
from app.utils.html_a_texto import dictamen_a_texto_plano

# ─── Paleta y estilos (calcados del archivo de referencia del dueño) ─────────
_VERDE_HUS = "FF1B5E3F"  # fill títulos/header
_VERDE_SUAVE = "FFE8F3EC"  # fill fila contrato + fila TOTAL
_ROJO_DECISION = "FFC00000"  # fuente columna DECISIÓN HUS (NO ACEPTA)
_GRIS_PENDIENTE = "FF6B7280"  # fuente DECISIÓN cuando aún no hay análisis
_BLANCO = "FFFFFFFF"

_FILL_VERDE = PatternFill("solid", fgColor=_VERDE_HUS)
_FILL_SUAVE = PatternFill("solid", fgColor=_VERDE_SUAVE)
_BORDE_FINO = Border(
    left=Side(border_style="thin"),
    right=Side(border_style="thin"),
    top=Side(border_style="thin"),
    bottom=Side(border_style="thin"),
)

# Anchos de columna de la hoja RESPUESTA GLOSAS — idénticos al archivo
# de referencia (la col E queda en ancho default, igual que el original).
_ANCHOS_RESPUESTA = {
    "A": 4,
    "B": 15,
    "C": 9,
    "D": 11,
    "F": 8,
    "G": 11,
    "H": 28,
    "I": 11,
    "J": 42,
    "K": 14,
    "L": 95,
}
_ALTO_FILA_DATOS = 129.95  # igual al de referencia (texto largo con wrap)

_NIT_ESE_HUS = "900.006.037-4"  # NIT institucional propio (constante pública)

_MESES_CORTOS = ["ENE", "FEB", "MAR", "ABR", "MAY", "JUN", "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"]
_MESES_LARGOS = [
    "ENERO",
    "FEBRERO",
    "MARZO",
    "ABRIL",
    "MAYO",
    "JUNIO",
    "JULIO",
    "AGOSTO",
    "SEPTIEMBRE",
    "OCTUBRE",
    "NOVIEMBRE",
    "DICIEMBRE",
]

# Alias → sigla corta para título de hoja jurídica y nombre de archivo.
# Solo entidades con sigla institucional conocida; el resto deriva de la
# primera palabra significativa de la razón social.
_SIGLAS_ENTIDAD: list[tuple[str, str]] = [
    ("DISPENSARIO", "DMBUG"),
    ("DMBUG", "DMBUG"),
    ("DIGSA", "DMBUG"),
    ("U220311", "DMBUG"),
    ("POLICIA", "PONAL"),
    ("POLICÍA", "PONAL"),
    ("NUEVA EPS", "NUEVAEPS"),
    ("FOMAG", "FOMAG"),
    ("MAGISTERIO", "FOMAG"),
    ("PPL", "PPL"),
]


# Palabras genéricas que no sirven como sigla (tipos societarios,
# artículos y prefijos de los export del DGH).
_TOKENS_GENERICOS = {"EPS", "IPS", "ESE", "ARL", "SAS", "LTDA", "DEL", "LAS", "LOS", "PARA", "CON"}


def _entidad_corta(eps: str) -> str:
    """Sigla corta de la entidad para el nombre de archivo y el header
    de la columna OBSERVACIÓN AUDITORÍA (p. ej. DMBUG, FOMAG)."""
    eu = (eps or "").upper().strip()
    for alias, sigla in _SIGLAS_ENTIDAD:
        if alias in eu:
            return sigla
    # Primera palabra alfanumérica significativa (saltando códigos tipo
    # "U220181 -" que anteceden la razón social en los export del DGH).
    for token in re.split(r"[^A-ZÁÉÍÓÚÑ0-9]+", eu):
        if (
            len(token) >= 3
            and token not in _TOKENS_GENERICOS
            and not token.isdigit()
            and not re.match(r"^[A-Z]\d+$", token)
        ):
            return token[:12]
    return "ENTIDAD"


def _fmt_fecha(dt: datetime | None) -> str:
    return dt.strftime("%d/%m/%Y") if isinstance(dt, datetime) else ""


def _fecha_archivo(dt: datetime) -> str:
    """10/06/2026 → «10JUN2026» (sufijo del nombre de archivo)."""
    return f"{dt.day:02d}{_MESES_CORTOS[dt.month - 1]}{dt.year}"


def _rango_meses_facturas(glosas: list[GlosaRecord]) -> str:
    """«FACTURAS DE ABRIL Y MAYO DE 2026» a partir de fecha_documento_dgh.

    Devuelve "" si ninguna glosa trae fecha de factura (no se inventa)."""
    fechas = sorted(g.fecha_documento_dgh for g in glosas if g.fecha_documento_dgh)
    if not fechas:
        return ""
    a, b = fechas[0], fechas[-1]
    if (a.year, a.month) == (b.year, b.month):
        return f"FACTURAS DE {_MESES_LARGOS[a.month - 1]} DE {a.year}"
    if a.year == b.year:
        return (
            f"FACTURAS DE {_MESES_LARGOS[a.month - 1]} A {_MESES_LARGOS[b.month - 1]} DE {a.year}"
        )
    return (
        f"FACTURAS DE {_MESES_LARGOS[a.month - 1]} DE {a.year} "
        f"A {_MESES_LARGOS[b.month - 1]} DE {b.year}"
    )


# ─── Decisión HUS por fila ────────────────────────────────────────────────────
# Coherente con el catálogo de códigos RE (catalogo_glosas.RESPUESTAS_IPS):
#   RE97xx → la IPS ACEPTA la glosa al 100 %
#   RE98xx → la IPS ACEPTA PARCIALMENTE y subsana
#   RE95xx / RE96xx / RE99xx → defensas (NO ACEPTA)
# El sufijo «(VALIDAR SOPORTE HC ANTES DE RADICAR)» replica la marca del
# archivo del dueño: la defensa existe pero depende de soporte de HC aún
# no verificado (estado REQUIERE_SOPORTES o confianza IA baja < 0.50).
_DECISION_VALIDAR = "NO ACEPTA (VALIDAR SOPORTE HC ANTES DE RADICAR)"


def _decision_hus(
    codigo_respuesta: str,
    estado: str,
    score: float | None,
    tiene_respuesta: bool,
) -> str:
    cod = (codigo_respuesta or "").upper().strip()
    if cod.startswith("RE97"):
        return "ACEPTA"
    if cod.startswith("RE98"):
        return "ACEPTA PARCIAL"
    estado_u = (estado or "").upper().strip()
    if tiene_respuesta or cod:
        if estado_u == "REQUIERE_SOPORTES":
            return _DECISION_VALIDAR
        # score 0.0 = «sin calificar» (default de columna), no confianza baja.
        if score is not None and 0 < score < 0.50:
            return _DECISION_VALIDAR
        return "NO ACEPTA"
    if estado_u == "REQUIERE_SOPORTES":
        return _DECISION_VALIDAR
    return "PENDIENTE"


def _color_decision(decision: str) -> str:
    if decision.startswith("NO ACEPTA"):
        return _ROJO_DECISION
    if decision.startswith("ACEPTA"):
        return _VERDE_HUS
    return _GRIS_PENDIENTE


def _respuesta_fila(
    glosa: GlosaRecord,
    concepto: ConceptoGlosaRecord | None,
    multi_concepto: bool = False,
) -> str:
    """RESPUESTA ESE HUS de la fila: dictamen del concepto si está
    poblado; si no, dictamen de la glosa; si tampoco, marca PENDIENTE.

    Regresión 12-jun-2026 (export del dueño): en glosas con VARIOS
    conceptos el campo por-concepto casi nunca viene, así que todas las
    filas de la misma factura caían al `glosa.dictamen` completo y salían
    palabra por palabra idénticas (filas AU0101 y CO0701 mostrando ambas
    el bloque encabezado por TA0101). Fix de PRESENTACIÓN («Opción A»):
    cuando la glosa es multi-concepto, el dictamen se divide por los
    separadores «═══ RESPUESTA AL CÓDIGO … ═══» (multi_codigo.py) y cada
    fila recibe SOLO su sección — el flujo de análisis no se toca.
    """
    if concepto is not None and (concepto.dictamen_html or "").strip():
        return _truncar_para_celda(dictamen_a_texto_plano(concepto.dictamen_html))
    if (glosa.dictamen or "").strip():
        if concepto is not None and multi_concepto:
            cod_glosa = (glosa.codigo_glosa or "").strip().upper()
            # Mismo fallback que la columna CÓD. GLOSA: un concepto sin
            # código hereda el de la glosa (y con él, su sección principal).
            cod_concepto = (concepto.codigo_glosa or "").strip().upper() or cod_glosa
            seccion = extraer_seccion_por_codigo(
                glosa.dictamen,
                codigo_concepto=cod_concepto,
                codigo_principal=cod_glosa or None,
            )
            if seccion:
                return _truncar_para_celda(dictamen_a_texto_plano(seccion))
            if cod_concepto and cod_glosa and cod_concepto != cod_glosa:
                # El dictamen no trae sección para este código (la glosa se
                # analizó como un todo): nota corta y profesional en vez de
                # duplicar el dictamen completo en N filas.
                return (
                    "Este concepto se responde como parte del dictamen integral de la "
                    f"glosa #{glosa.id} (código principal {cod_glosa}). Ver la fila del "
                    "código principal para la argumentación completa."
                )
        # Glosa de un solo concepto (o concepto = código principal sin
        # sección divisible): status quo — dictamen completo, que incluye
        # la cabecera del documento.
        return _truncar_para_celda(dictamen_a_texto_plano(glosa.dictamen))
    return f"PENDIENTE DE ANÁLISIS — {(glosa.estado or 'SIN ESTADO').upper()}"


# ─── Contexto del lote (contrato + entidad + fechas) ─────────────────────────


def _buscar_contrato_record(db, eps: str) -> ContratoRecord | None:
    """Match flexible GlosaRecord.eps ↔ ContratoRecord.eps (el export del
    DGH trae razones sociales largas; la PK de contratos es la clave
    canónica corta tipo «DISPENSARIO MEDICO»)."""
    eu = (eps or "").upper().strip()
    if not eu:
        return None
    try:
        registros = db.query(ContratoRecord).all()
    except Exception as e:  # BD sin tabla (tests parciales) — degradar
        logger.debug(f"[excel-radicable] consulta contratos falló: {e}")
        return None
    exacto = next((r for r in registros if (r.eps or "").upper().strip() == eu), None)
    if exacto:
        return exacto
    for r in registros:
        clave = (r.eps or "").upper().strip()
        if clave and (clave in eu or eu in clave):
            return r
    return None


def _metadatos_contrato(contrato_rec: ContratoRecord | None) -> dict:
    """ContratoRecord → dict plano (fechas formateadas, JSON decodificado)."""
    if contrato_rec is None:
        return {}

    def _json_lista(crudo):
        try:
            val = json.loads(crudo) if crudo else None
        except (ValueError, TypeError):
            val = None
        return val if isinstance(val, list) else None

    return {
        "numero_contrato": (contrato_rec.numero_contrato or "").strip(),
        "nit_eps": (contrato_rec.nit_eps or "").strip(),
        "nit_ips": (contrato_rec.nit_ips or "").strip(),
        "razon_social_eps": (contrato_rec.razon_social_eps or "").strip(),
        "razon_social_ips": (contrato_rec.razon_social_ips or "").strip(),
        "numero_proceso_secop": (contrato_rec.numero_proceso_secop or "").strip(),
        "fecha_suscripcion": _fmt_fecha(contrato_rec.fecha_suscripcion),
        "fecha_inicio": _fmt_fecha(contrato_rec.fecha_inicio),
        "fecha_fin": _fmt_fecha(contrato_rec.fecha_fin),
        "objeto_contractual": (contrato_rec.objeto_contractual or "").strip(),
        "anexos": _json_lista(contrato_rec.anexos_descripcion),
        "modalidades": _json_lista(contrato_rec.modalidades_tarifarias),
    }


class _ContextoLote:
    """Datos compartidos por ambas hojas (entidad, contrato, fechas)."""

    def __init__(self, db, glosas: list[GlosaRecord]):
        eps_counter = Counter((g.eps or "").strip() for g in glosas if (g.eps or "").strip())
        self.eps = eps_counter.most_common(1)[0][0] if eps_counter else ""
        self.entidad_corta = _entidad_corta(self.eps)

        contrato_rec = _buscar_contrato_record(db, self.eps) if self.eps else None
        self.md = _metadatos_contrato(contrato_rec)
        # Fallback al catálogo CONTRATOS_HUS (glosa_ia_prompts) — siempre
        # devuelve dict; «SIN CONTRATO PACTADO» cuando la EPS no está.
        self.catalogo = get_contrato(self.eps) if self.eps else {}

        # Fecha de entrega de la glosa inicial: la trae el Excel del DGH
        # (fecha_entrega); fallback recepción → ahora.
        fechas_entrega = [g.fecha_entrega for g in glosas if g.fecha_entrega]
        fechas_rec = [g.fecha_recepcion for g in glosas if g.fecha_recepcion]
        self.fecha_entrega = (
            max(fechas_entrega)
            if fechas_entrega
            else (max(fechas_rec) if fechas_rec else ahora_utc())
        )
        self.rango_facturas = _rango_meses_facturas(glosas)

        try:
            self.clausulas = (
                db.query(ClausulaContrato)
                .filter(ClausulaContrato.eps == (contrato_rec.eps if contrato_rec else self.eps))
                .order_by(ClausulaContrato.numero_clausula, ClausulaContrato.id)
                .limit(20)
                .all()
            )
        except Exception as e:
            logger.debug(f"[excel-radicable] consulta cláusulas falló: {e}")
            self.clausulas = []

    # — piezas de texto reutilizadas por las dos hojas —

    def nombre_entidad(self) -> str:
        nombre = self.md.get("razon_social_eps") or (self.eps or "ENTIDAD PAGADORA").upper()
        nit = self.md.get("nit_eps") or (self.catalogo.get("nit") or "")
        if nit and nit.upper() != "N/D":
            return f"{nombre} (NIT {nit})"
        return nombre

    def numero_contrato(self) -> str:
        numero = self.md.get("numero_contrato") or (self.catalogo.get("numero") or "")
        if numero.upper().strip() in ("", "SIN CONTRATO PACTADO", "N/A", "N/D"):
            return ""
        return numero

    def linea_contrato(self) -> str:
        """Fila 2 de la hoja RESPUESTA GLOSAS. El segmento de contrato se
        omite elegantemente si no hay número conocido."""
        partes: list[str] = []
        numero = self.numero_contrato()
        if numero:
            etiqueta = numero if "CONTRATO" in numero.upper() else f"CONTRATO {numero}"
            secop = self.md.get("numero_proceso_secop") or ""
            if secop and secop.upper() not in etiqueta.upper():
                etiqueta += f" (PROCESO {secop})"
            ini, fin = self.md.get("fecha_inicio"), self.md.get("fecha_fin")
            if ini and fin:
                etiqueta += f" — PLAZO DEL {ini} AL {fin}"
            partes.append(etiqueta.upper())
        partes.append(f"ENTREGA GLOSA INICIAL {_fmt_fecha(self.fecha_entrega)}")
        partes.append("RESPUESTA ELABORADA: AUDITORÍA DE CUENTAS MÉDICAS – ESE HUS")
        return "  |  ".join(partes)


# ─── Hoja 1: RESPUESTA GLOSAS ────────────────────────────────────────────────


def _encabezados(entidad_corta: str) -> list[str]:
    return [
        "No.",
        "FACTURA",
        "CONSEC. DGH",
        "FECHA FACTURA",
        "VR FACTURA",
        "CÓD. GLOSA",
        "CÓD. SERVICIO",
        "DESCRIPCIÓN SERVICIO",
        "VR GLOSADO",
        f"OBSERVACIÓN AUDITORÍA {entidad_corta}",
        "DECISIÓN HUS",
        "RESPUESTA ESE HUS",
    ]


def _consec_celda(consecutivo: str | None):
    """El consecutivo DGH va como número si es puramente numérico (igual
    que en el archivo de referencia), si no como texto."""
    s = (consecutivo or "").strip()
    if s.isdigit():
        return int(s)
    return s


def _orden_glosa(g: GlosaRecord):
    s = (g.consecutivo_dgh or "").strip()
    return (0, int(s)) if s.isdigit() else (1, g.factura or ""), g.id


def _filas_del_lote(
    glosas: list[GlosaRecord],
    conceptos_por_glosa: dict[int, list[ConceptoGlosaRecord]],
) -> list[dict]:
    """Aplana glosas → una fila por concepto (o una por glosa sin
    conceptos). Cada dict trae las 12 columnas ya resueltas."""
    filas: list[dict] = []
    for g in sorted(glosas, key=_orden_glosa):
        conceptos = conceptos_por_glosa.get(g.id) or [None]
        # Con 2+ conceptos la columna RESPUESTA se divide por código (fix
        # 12-jun-2026: filas duplicadas); con uno solo, status quo.
        multi = len(conceptos) > 1
        for c in conceptos:
            decision = _decision_hus(
                g.codigo_respuesta,
                g.estado,
                g.score,
                tiene_respuesta=bool(
                    (g.dictamen or "").strip()
                    or (c is not None and (c.dictamen_html or "").strip())
                ),
            )
            filas.append(
                {
                    "factura": g.factura or "",
                    "consec": _consec_celda(g.consecutivo_dgh),
                    "fecha_factura": (
                        g.fecha_documento_dgh.strftime("%Y-%m-%d") if g.fecha_documento_dgh else ""
                    ),
                    "vr_factura": int(round(g.valor_factura)) if g.valor_factura else 0,
                    "cod_glosa": (c.codigo_glosa if c else None) or g.codigo_glosa or "",
                    "cod_servicio": (c.cups_codigo if c else None) or g.cups_servicio or "",
                    "descripcion": (c.cups_descripcion if c else None)
                    or g.servicio_descripcion
                    or "",
                    "vr_glosado": int(
                        round((c.valor_objetado if c else None) or g.valor_objetado or 0)
                    ),
                    "observacion": _truncar_para_celda(
                        (c.observacion_eps if c else None)
                        or g.observacion_eps
                        or (c.nombre_glosa if c else None)
                        or g.concepto_glosa
                        or ""
                    ),
                    "decision": decision,
                    "respuesta": _respuesta_fila(g, c, multi_concepto=multi),
                }
            )
    return filas


def _armar_hoja_respuesta(ws: Worksheet, ctx: _ContextoLote, filas: list[dict]) -> None:
    for col, ancho in _ANCHOS_RESPUESTA.items():
        ws.column_dimensions[col].width = ancho

    # Fila 1 — título institucional (merged, verde HUS)
    ws.merge_cells("A1:L1")
    c = ws.cell(row=1, column=1)
    c.value = (
        "ESE HOSPITAL UNIVERSITARIO DE SANTANDER – RESPUESTA A GLOSA INICIAL – "
        + ctx.nombre_entidad().upper()
    )
    c.font = Font(name="Arial", size=11, bold=True, color=_BLANCO)
    c.fill = _FILL_VERDE
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Fila 2 — contrato + entrega + área (el segmento de contrato se
    # omite si no hay número conocido; la fila queda para no mover el
    # header de la fila 3).
    ws.merge_cells("A2:L2")
    c = ws.cell(row=2, column=1)
    c.value = ctx.linea_contrato()
    c.font = Font(name="Arial", size=9, bold=True, color=_VERDE_HUS)
    c.fill = _FILL_SUAVE
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20.1

    # Fila 3 — header de 12 columnas
    for i, titulo in enumerate(_encabezados(ctx.entidad_corta), start=1):
        h = ws.cell(row=3, column=i, value=titulo)
        h.font = Font(name="Arial", size=9, bold=True, color=_BLANCO)
        h.fill = _FILL_VERDE
        h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        h.border = _BORDE_FINO
    ws.row_dimensions[3].height = 26.1
    ws.freeze_panes = "A4"

    # Filas de datos — una por concepto
    fuente_dato = Font(name="Arial", size=8)
    align_texto = Alignment(vertical="top", wrap_text=True)
    align_num = Alignment(horizontal="right", vertical="top", wrap_text=True)
    fila_excel = 4
    for f in filas:
        valores = [
            fila_excel - 3,  # No. consecutivo por fila
            f["factura"],
            f["consec"],
            f["fecha_factura"],
            f["vr_factura"],
            f["cod_glosa"],
            f["cod_servicio"],
            f["descripcion"],
            f["vr_glosado"],
            f["observacion"],
            f["decision"],
            f["respuesta"],
        ]
        for col, valor in enumerate(valores, start=1):
            # Ronda 30: sanear texto de EPS/dictamen (inyección de fórmulas
            # Excel + caracteres de control). Los números pasan intactos.
            celda = ws.cell(row=fila_excel, column=col, value=sanear_celda_excel(valor))
            celda.border = _BORDE_FINO
            if col in (5, 9):  # VR FACTURA / VR GLOSADO
                celda.font = fuente_dato
                celda.number_format = "#,##0"
                celda.alignment = align_num
            elif col == 11:  # DECISIÓN HUS
                celda.font = Font(
                    name="Arial", size=8, bold=True, color=_color_decision(f["decision"])
                )
                celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                celda.font = fuente_dato
                celda.alignment = align_texto
        ws.row_dimensions[fila_excel].height = _ALTO_FILA_DATOS
        fila_excel += 1

    # Fila TOTAL GLOSADO — banda suave de borde a borde, suma literal en
    # la col I (igual que el archivo de referencia).
    total = sum(f["vr_glosado"] for f in filas)
    fuente_total = Font(name="Arial", size=9, bold=True)
    for col in range(1, 13):
        celda = ws.cell(row=fila_excel, column=col)
        celda.fill = _FILL_SUAVE
        celda.border = _BORDE_FINO
        celda.font = fuente_total
    ws.cell(row=fila_excel, column=8, value="TOTAL GLOSADO")
    celda_total = ws.cell(row=fila_excel, column=9, value=total)
    celda_total.number_format = "#,##0"


# ─── Hoja 2: FUNDAMENTO JURÍDICO ─────────────────────────────────────────────


def _normas_regimen(eps: str) -> list[str]:
    """Extrae SOLO las líneas normativas del bloque de régimen especial
    (decretos/leyes/acuerdos), descartando las instrucciones de estilo
    que ese helper dirige a la IA («NO cites…», «refiérete…»)."""
    bloque = _regimen_especial_texto((eps or "").upper(), {})
    if not bloque:
        return []
    normas: list[str] = []
    for linea in bloque.splitlines():
        limpio = linea.strip().lstrip("•").strip()
        if not limpio or limpio.endswith(":"):
            continue
        if re.match(r"^(Decreto|Acuerdo|Plan|Resoluci|Ley|Para |Manual)", limpio, re.IGNORECASE):
            normas.append(limpio.rstrip("."))
    return normas


def _clausula_que_mencione(clausulas: list, *palabras: str) -> ClausulaContrato | None:
    """Primera cláusula cuyo título o texto contenga alguna palabra clave."""
    for cl in clausulas:
        blob = f"{cl.titulo or ''} {cl.texto_literal or ''}".upper()
        if any(p in blob for p in palabras):
            return cl
    return None


def _cita_clausula(cl: ClausulaContrato | None) -> str:
    if cl is None:
        return ""
    num = (cl.numero_clausula or "").strip()
    titulo = (cl.titulo or "").strip()
    if not (num or titulo):
        return ""
    partes = [p for p in (f"CLÁUSULA {num}" if num else "", titulo.upper()) if p]
    return " — ".join(partes)


def _secciones_fundamento(ctx: _ContextoLote, filas: list[dict]) -> list[tuple[str, str]]:
    """Las 6 secciones (título, texto) generadas desde datos del sistema."""
    md = ctx.catalogo or {}
    numero = ctx.numero_contrato()
    anio = ctx.fecha_entrega.year
    bloque_regimen = _regimen_especial_texto((ctx.eps or "").upper(), {}).upper()
    es_militar = "FUERZAS MILITARES" in bloque_regimen

    # 1. EXISTENCIA Y VIGENCIA DEL CONTRATO
    if numero:
        nit_ips = ctx.md.get("nit_ips") or _NIT_ESE_HUS
        s1 = (
            f"ENTRE {ctx.nombre_entidad().upper()} Y LA ESE HOSPITAL UNIVERSITARIO DE "
            f"SANTANDER (NIT {nit_ips}) SE SUSCRIBIÓ EL {numero.upper()}"
        )
        if ctx.md.get("fecha_suscripcion"):
            s1 += f" EL {ctx.md['fecha_suscripcion']}"
        if ctx.md.get("numero_proceso_secop"):
            s1 += f" (PROCESO {ctx.md['numero_proceso_secop'].upper()}, PUBLICADO EN SECOP II)"
        if ctx.md.get("fecha_inicio") and ctx.md.get("fecha_fin"):
            s1 += f", CON PLAZO DE EJECUCIÓN DEL {ctx.md['fecha_inicio']} AL {ctx.md['fecha_fin']}"
        elif md.get("vigencia") and str(md["vigencia"]).upper() not in ("N/A", "N/D"):
            s1 += f", CON VIGENCIA {str(md['vigencia']).upper()}"
        s1 += ". "
    else:
        s1 = (
            f"ENTRE {ctx.nombre_entidad().upper()} Y LA ESE HOSPITAL UNIVERSITARIO DE "
            "SANTANDER EXISTE VÍNCULO PARA LA ATENCIÓN DE SUS USUARIOS, ACREDITADO POR LAS "
            "AUTORIZACIONES Y REMISIONES EMITIDAS POR LA PROPIA ENTIDAD. "
        )
    s1 += "TODAS LAS ATENCIONES GLOSADAS"
    if ctx.rango_facturas:
        s1 += f" ({ctx.rango_facturas})"
    s1 += (
        " FUERON PRESTADAS DENTRO DEL PLAZO CONTRACTUAL, PREVIA AUTORIZACIÓN Y/O REMISIÓN "
        "DE LA ENTIDAD PAGADORA. POR TANTO, LA PREMISA «IPS SIN CONTRATO NI ACUERDO DE "
        "TARIFAS» CARECE DE FUNDAMENTO."
    )

    # 2. TARIFAS PACTADAS — anexos y modalidades reales del contrato
    cl_tarifas = _clausula_que_mencione(ctx.clausulas, "TARIFA", "ANEXO")
    partes2: list[str] = []
    cita2 = _cita_clausula(cl_tarifas)
    if cita2:
        partes2.append(
            f"LA {cita2} INCORPORA COMO PARTE INTEGRAL DEL CONTRATO LOS ANEXOS TARIFARIOS, "
            "DECLARÁNDOLOS REFERENTES ECONÓMICOS OBLIGATORIOS PARA LA FACTURACIÓN."
        )
    if ctx.md.get("anexos"):
        nombres = []
        for a in ctx.md["anexos"][:6]:
            if isinstance(a, dict):
                n = (a.get("nombre") or "").strip()
                d = (a.get("descripcion") or "").strip()
                nombres.append(f"{n}{f' ({d})' if d else ''}" if n else d)
            else:
                nombres.append(str(a))
        nombres = [n for n in nombres if n]
        if nombres:
            partes2.append(f"ANEXOS INTEGRALES DEL CONTRATO: {'; '.join(nombres).upper()}.")
    if ctx.md.get("modalidades"):
        partes2.append(
            "MODALIDADES TARIFARIAS PACTADAS: "
            + ", ".join(str(m) for m in ctx.md["modalidades"]).upper()
            + "."
        )
    if md.get("tarifa"):
        partes2.append(f"TARIFA PACTADA: {str(md['tarifa']).upper()}.")
    cups_lote = []
    for f in filas:
        cod = (f.get("cod_servicio") or "").strip()
        if cod and cod not in cups_lote:
            cups_lote.append(cod)
    if cups_lote:
        partes2.append(
            "LOS SERVICIOS GLOSADOS EN EL PRESENTE LOTE ESTÁN IDENTIFICADOS POR CÓDIGO "
            f"({', '.join(cups_lote[:15])}) Y SE LIQUIDARON CONFORME AL ANEXO TARIFARIO "
            "PACTADO Y/O A LAS RESOLUCIONES TARIFARIAS INSTITUCIONALES DE LA ESE HUS."
        )
    if not partes2:
        partes2.append(
            "LAS TARIFAS APLICADAS CORRESPONDEN A LAS PACTADAS ENTRE LAS PARTES Y/O A LAS "
            "RESOLUCIONES TARIFARIAS INSTITUCIONALES DE LA ESE HUS VIGENTES AL MOMENTO DE "
            "LA ATENCIÓN."
        )
    s2 = " ".join(partes2)

    # 3. AJUSTE TARIFARIO DE LA VIGENCIA
    s3 = (
        "LAS TARIFAS REFERENCIADAS AL MANUAL SOAT SE AJUSTAN A LA ANUALIDAD VIGENTE "
        f"{anio} (LIQUIDADA EN UVB CONFORME AL ART. 89 DE LA LEY 2277 DE 2022), Y LAS "
        "TARIFAS PROPIAS DE LA ESE SE ACTUALIZAN ANUALMENTE MEDIANTE RESOLUCIÓN INTERNA. "
        "LOS SERVICIOS NO CONTEMPLADOS EN LOS ANEXOS SE LIQUIDAN SEGÚN LAS RESOLUCIONES "
        "TARIFARIAS DE LA ESE HUS. EN NINGÚN CASO EL CONTRATO AUTORIZA AL PAGADOR A "
        "RELIQUIDAR UNILATERALMENTE LOS SERVICIOS A UNA TARIFA DISTINTA DE LA PACTADA."
    )

    # 4. AGOTAMIENTO PRESUPUESTAL ≠ AUSENCIA DE CONTRATO — misma doctrina
    # tripartita que inyecta contexto_contractual_enriquecido (§7) al
    # prompt de la IA, redactada acá en prosa radicable.
    cl_presupuesto = _clausula_que_mencione(
        ctx.clausulas, "PRESUPUEST", "APROPIACI", "OBLIGACIONES DEL CONTRATANTE"
    )
    cita4 = _cita_clausula(cl_presupuesto)
    s4 = "(I) "
    if cita4:
        s4 += f"LA {cita4} RADICA EN EL CONTRATANTE "
    else:
        s4 += "ES OBLIGACIÓN DEL CONTRATANTE "
    s4 += (
        "LA GESTIÓN DE LAS APROPIACIONES PRESUPUESTALES REQUERIDAS PARA SOLVENTAR LAS "
        "PRESTACIONES SURGIDAS A SU CARGO; LA ENTIDAD NO PUEDE TRASLADAR AL PRESTADOR LAS "
        "CONSECUENCIAS DE SU PROPIA GESTIÓN PRESUPUESTAL. (II) LA ENTIDAD CONTINUÓ "
        "EMITIENDO AUTORIZACIONES Y REMITIENDO USUARIOS: ACTOS PROPIOS QUE RECONOCEN LA "
        "VIGENCIA DEL VÍNCULO (VENIRE CONTRA FACTUM PROPRIUM NON VALET; BUENA FE: ART. 83 "
        "C.P., ARTS. 1602 Y 1603 C.C., ART. 871 C.CO.). (III) LA CONTINUIDAD DEL SERVICIO "
        "DE SALUD NO PUEDE INTERRUMPIRSE POR RAZONES ADMINISTRATIVAS O ECONÓMICAS "
        "(ART. 14, LEY ESTATUTARIA 1751 DE 2015)."
    )

    # 5. SIN CONTRATO: SOAT PLENO — adaptado al régimen de la entidad
    s5 = (
        "EN GRACIA DE DISCUSIÓN, SI SE ACEPTARA LA INEXISTENCIA DEL VÍNCULO, EL ART. 20 "
        "DEL DECRETO 4747 DE 2007 (COMPILADO EN EL DECRETO 780 DE 2016) ORDENA PAGAR LA "
        "ATENCIÓN DE URGENCIAS A LAS TARIFAS DEL DECRETO 2423 DE 1996 (SOAT) PLENAS, Y "
        "LOS SERVICIOS ELECTIVOS FUERON EXPRESAMENTE AUTORIZADOS POR LA ENTIDAD, QUE "
        "RECIBIÓ EL SERVICIO PARA SUS AFILIADOS; NEGAR SU PAGO O IMPONER VALORES "
        "INFERIORES CONFIGURARÍA ENRIQUECIMIENTO SIN JUSTA CAUSA (ACTIO IN REM VERSO). "
        "EN CONSECUENCIA, LA TESIS DE LA AUDITORÍA CONDUCE A UN RESULTADO JURÍDICAMENTE "
        "INADMISIBLE: PAGAR MENOS QUE LO PACTADO Y MENOS QUE LO LEGALMENTE DEBIDO."
    )
    normas = _normas_regimen(ctx.eps)
    if normas:
        s5 += f" MARCO DEL RÉGIMEN ESPECIAL APLICABLE: {'; '.join(normas).upper()}."
        if es_militar:
            s5 += (
                " EN ESTE RÉGIMEN LA ENTIDAD ACTÚA COMO CONTRATANTE/ENTIDAD PAGADORA DEL "
                "SUBSISTEMA DE SALUD DE LAS FUERZAS MILITARES, CON OBLIGACIÓN DE PAGO POR "
                "LOS SERVICIOS AUTORIZADOS A SUS USUARIOS."
            )

    # 6. TRÁMITE DE LA GLOSA — texto procesal estándar
    cl_glosas = _clausula_que_mencione(ctx.clausulas, "GLOSA", "OBJECI")
    cita6 = _cita_clausula(cl_glosas)
    s6 = "LA PRESENTE RESPUESTA SE RADICA DENTRO DEL TÉRMINO LEGAL Y CONTRACTUAL ("
    if cita6:
        s6 += f"{cita6} DEL CONTRATO; "
    s6 += (
        "DECRETO 4747 DE 2007; RESOLUCIÓN 2284 DE 2023; ART. 57 DE LA LEY 1438 DE 2011 — "
        "QUINCE (15) DÍAS HÁBILES SIGUIENTES A LA RECEPCIÓN DE LA GLOSA), CON LOS "
        "SOPORTES DOCUMENTALES PERTINENTES. SE SOLICITA EL LEVANTAMIENTO TOTAL DE LAS "
        "GLOSAS Y EL PAGO DE LOS VALORES FACTURADOS; EN SU DEFECTO, SE SOLICITA MESA DE "
        "CONCILIACIÓN CONFORME A LA NORMATIVIDAD VIGENTE."
    )

    return [
        ("1. EXISTENCIA Y VIGENCIA DEL CONTRATO", s1),
        ("2. TARIFAS PACTADAS – ANEXO No. 1 Y ANEXOS DE INSUMOS", s2),
        (f"3. AJUSTE TARIFARIO VIGENCIA {anio}", s3),
        ("4. EL AGOTAMIENTO PRESUPUESTAL NO EQUIVALE A AUSENCIA DE CONTRATO", s4),
        ("5. RÉGIMEN SUBSIDIARIO SIN CONTRATO: SOAT PLENO, NO TARIFAS INFERIORES", s5),
        ("6. TRÁMITE DE LA GLOSA", s6),
    ]


def _armar_hoja_fundamento(ws: Worksheet, ctx: _ContextoLote, filas: list[dict]) -> None:
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 130

    ws.merge_cells("A1:B1")
    c = ws.cell(row=1, column=1)
    c.value = f"FUNDAMENTO JURÍDICO GENERAL – DEFENSA HUS FRENTE A GLOSAS – {ctx.entidad_corta}"
    c.font = Font(name="Arial", size=11, bold=True, color=_BLANCO)
    c.fill = _FILL_VERDE
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    fila = 3  # fila 2 en blanco, como el archivo de referencia
    for titulo, texto in _secciones_fundamento(ctx, filas):
        ca = ws.cell(row=fila, column=1, value=titulo)
        ca.font = Font(name="Arial", size=10, bold=True, color=_VERDE_HUS)
        ca.alignment = Alignment(vertical="top", wrap_text=True)
        ca.border = _BORDE_FINO

        cb = ws.cell(row=fila, column=2, value=_truncar_para_celda(texto))
        cb.font = Font(name="Arial", size=9)
        cb.alignment = Alignment(horizontal="justify", vertical="top", wrap_text=True)
        cb.border = _BORDE_FINO
        # Alto proporcional al texto (col B ≈ 120 chars/línea a 9 pt).
        ws.row_dimensions[fila].height = max(90, min(410, 14 * (len(texto) // 120 + 1)))
        fila += 1


# ─── API pública ─────────────────────────────────────────────────────────────


def entidades_del_lote(db, recepcion_import_id: int | None = None, glosa_ids=None) -> list[dict]:
    """Qué entidades hay en el lote y cuántas facturas de cada una.

    20-08-2026. Sirve para no radicarle a una EPS las facturas de otra: el
    panel ofrece un archivo por entidad en vez de uno solo con todo mezclado.
    """
    ids = _ids_del_lote(db, recepcion_import_id, glosa_ids)
    if not ids:
        return []
    glosas = db.query(GlosaRecord).filter(GlosaRecord.id.in_(ids)).all()
    cuenta = Counter((g.eps or "").strip() for g in glosas if (g.eps or "").strip())
    return [
        {"eps": eps, "glosas": n, "sigla": _entidad_corta(eps)} for eps, n in cuenta.most_common()
    ]


def _ids_del_lote(db, recepcion_import_id: int | None, glosa_ids) -> list[int]:
    """Los IDs de glosa del lote, vengan de una importación o de una lista."""
    if recepcion_import_id is not None:
        rec = (
            db.query(ImportacionRecepcionRecord)
            .filter(ImportacionRecepcionRecord.id == recepcion_import_id)
            .first()
        )
        if rec is None:
            raise LookupError(f"Importación {recepcion_import_id} no encontrada")
        try:
            return [int(x) for x in json.loads(rec.glosa_ids or "[]")]
        except (ValueError, TypeError):
            return []
    return list(glosa_ids or [])


def generar_excel_radicable_con_nombre(
    db,
    recepcion_import_id: int | None = None,
    glosa_ids: list[int] | None = None,
    eps_filtro: str | None = None,
) -> tuple[bytes, str]:
    """Construye el Excel radicable y su nombre de archivo institucional.

    Args:
      db: sesión SQLAlchemy.
      recepcion_import_id: si se da, las glosas salen del lote de esa
        importación de recepción (LookupError si no existe).
      glosa_ids: alternativa directa por lista de IDs de glosa.

    Returns:
      (bytes del .xlsx, nombre tipo RESPUESTA_GLOSAS_DMBUG_10JUN2026_HUS.xlsx)
    """
    glosa_ids = _ids_del_lote(db, recepcion_import_id, glosa_ids)
    glosas = db.query(GlosaRecord).filter(GlosaRecord.id.in_(glosa_ids)).all() if glosa_ids else []

    # 20-08-2026 (caso real de Yesid). El lote del 19 de agosto traía 29
    # facturas de COOSALUD y 6 del DISPENSARIO MÉDICO / EJÉRCITO. El archivo
    # se rotulaba con la entidad MÁS FRECUENTE —COOSALUD, y su contrato
    # 68001C00060340-24 en el encabezado— pero adentro iban LAS 35. Es decir,
    # se le radicaban a COOSALUD $10.290.042 en facturas de otro pagador,
    # bajo un contrato que nada tiene que ver con ellas.
    #
    # Un Excel radicable se presenta ante UNA entidad. Si el lote trae varias,
    # cada una necesita el suyo: no se mezclan ni se descartan en silencio.
    if eps_filtro:
        objetivo = eps_filtro.strip().upper()
        glosas = [g for g in glosas if (g.eps or "").strip().upper() == objetivo]
    conceptos_por_glosa: dict[int, list[ConceptoGlosaRecord]] = {}
    if glosas:
        conceptos = (
            db.query(ConceptoGlosaRecord)
            .filter(ConceptoGlosaRecord.glosa_id.in_([g.id for g in glosas]))
            .order_by(ConceptoGlosaRecord.glosa_id, ConceptoGlosaRecord.id)
            .all()
        )
        for con in conceptos:
            conceptos_por_glosa.setdefault(con.glosa_id, []).append(con)

    ctx = _ContextoLote(db, glosas)
    filas = _filas_del_lote(glosas, conceptos_por_glosa)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "RESPUESTA GLOSAS"
    _armar_hoja_respuesta(ws1, ctx, filas)
    ws2 = wb.create_sheet("FUNDAMENTO JURÍDICO")
    _armar_hoja_fundamento(ws2, ctx, filas)

    out = BytesIO()
    try:
        wb.save(out)
    finally:
        wb.close()

    fecha_dt = ctx.fecha_entrega if isinstance(ctx.fecha_entrega, datetime) else ahora_utc()
    nombre = f"RESPUESTA_GLOSAS_{ctx.entidad_corta}_{_fecha_archivo(fecha_dt)}_HUS.xlsx"
    logger.info(f"[excel-radicable] generado {nombre}: {len(glosas)} glosas → {len(filas)} filas")
    return out.getvalue(), nombre


def generar_excel_radicable(
    db,
    recepcion_import_id: int | None = None,
    glosa_ids: list[int] | None = None,
) -> bytes:
    """Variante que devuelve solo los bytes del .xlsx."""
    contenido, _nombre = generar_excel_radicable_con_nombre(
        db, recepcion_import_id=recepcion_import_id, glosa_ids=glosa_ids
    )
    return contenido
