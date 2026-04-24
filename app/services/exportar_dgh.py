"""Exportador formato DGH (Ronda 35) — estructura EXACTA para cargar al sistema.

El Excel que descarga el coordinador y sube al DGH debe tener exactamente
las 26 columnas en este orden:

  EstadoCxCObjecion
  TipoObjecionTramite
  FacturaCartera.Factura
  FechaDocumento
  Consecutivo
  Observaciones
  EstadoActual
  FacturaCartera.Valor
  FacturaCartera.Fecha
  FacturaCartera.Tercero.Documento
  FacturaCartera.Tercero.NombreCompletoAN
  FechaObjecion
  FacturaCartera.Tercero.NombreCompletoNA
  ListadoConceptos.ConceptoObjecion.Codigo
  ListadoConceptos.Oid
  ListadoConceptos.ConceptoObjecion.Nombre
  ListadoConceptos.ServicioProductoFactura.Codigo
  ListadoConceptos.ServicioProductoFactura.Descripcion
  ListadoConceptos.ValorObjecion
  ListadoConceptos.ServicioProductoFactura.CentroCosto.CodigoNombreCentro
  ListadoConceptos.Observaciones
  FECHA DE CARGUE           ← agregadas por HUS
  CODIGO RESPUESTA          ← agregadas por HUS
  VALOR ACEPTADO            ← agregadas por HUS
  OBSERVACION               ← agregadas por HUS

Regla clave: **una fila por CONCEPTO**, no por glosa. Si una glosa
tiene 3 conceptos, son 3 filas en el Excel. Si no hay conceptos
registrados (glosa legacy sin detalle por concepto), emite 1 fila
usando los datos de la glosa principal.

Adicional, la columna OBSERVACION debe contener el DICTAMEN LIMPIO:
sin emojis, sin headers tipo "📋 Tarifa pactada encontrada" ni
"RATIFICADA EPS: U240061 - ... | Observación recepción: ...", solo el
argumento jurídico final.
"""
from __future__ import annotations

import html as _html
import re
from datetime import datetime
from io import BytesIO
from typing import Optional

from sqlalchemy.orm import Session

from app.models.db import ConceptoGlosaRecord, GlosaRecord

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    EXCEL_OK = True
except Exception:
    EXCEL_OK = False


COLUMNAS_DGH = [
    "EstadoCxCObjecion",
    "TipoObjecionTramite",
    "FacturaCartera.Factura",
    "FechaDocumento",
    "Consecutivo",
    "Observaciones",
    "EstadoActual",
    "FacturaCartera.Valor",
    "FacturaCartera.Fecha",
    "FacturaCartera.Tercero.Documento",
    "FacturaCartera.Tercero.NombreCompletoAN",
    "FechaObjecion",
    "FacturaCartera.Tercero.NombreCompletoNA",
    "ListadoConceptos.ConceptoObjecion.Codigo",
    "ListadoConceptos.Oid",
    "ListadoConceptos.ConceptoObjecion.Nombre",
    "ListadoConceptos.ServicioProductoFactura.Codigo",
    "ListadoConceptos.ServicioProductoFactura.Descripcion",
    "ListadoConceptos.ValorObjecion",
    "ListadoConceptos.ServicioProductoFactura.CentroCosto.CodigoNombreCentro",
    "ListadoConceptos.Observaciones",
    "FECHA DE CARGUE",
    "CODIGO RESPUESTA",
    "VALOR ACEPTADO",
    "OBSERVACION",
]


# ─── Limpieza del dictamen para la columna OBSERVACION ────────────────────

# Removemos emojis (rango BMP común + suplementario para pictogramas)
_EMOJI_RE = re.compile(
    "["
    "\U0001F300-\U0001F6FF"     # pictogramas, transporte, símbolos
    "\U0001F900-\U0001F9FF"     # símbolos y pictogramas suplementarios
    "\U0001FA00-\U0001FAFF"
    "☀-➿"              # símbolos varios
    "✀-➿"              # dingbats
    "\U0001F000-\U0001F2FF"
    "←-⇿"              # flechas
    "⬀-⯿"              # flechas suplementarias
    "]+",
    flags=re.UNICODE,
)

_HTML_TAG_RE = re.compile(r"<[^>]+>")

# Marcadores donde empieza el ARGUMENTO jurídico real (ignorar todo lo anterior)
_INICIO_ARGUMENTO = re.compile(r"ARGUMENTACI[ÓO]N\s+JUR[ÍI]DICA\s*", re.IGNORECASE)

# Marcadores donde termina el argumento (todo lo siguiente es decoración/footer)
_FIN_ARGUMENTO = re.compile(
    r"(RELACI[ÓO]N\s+DE\s+SOPORTES\s+APORTADOS|"
    r"#\s*Documento\s+Marco\s+legal|"
    r"Nota:\s*Generado\s+con\s+asistencia\s+de\s+IA|"
    r"Generado\s+con\s+asistencia\s+de\s+IA)",
    re.IGNORECASE,
)

# Headers que debemos ELIMINAR del inicio del dictamen (no son argumento)
_HEADERS_BASURA = (
    re.compile(r"^\s*RATIFICADA\s+EPS:.*?Observaci[óo]n\s+recepci[óo]n:\s*", re.IGNORECASE | re.DOTALL),
    re.compile(r"^\s*GLOSA\s+EXTEMPOR[AÁ]NEA\s*[—-]\s*\d+\s+D[IÍ]AS\s+H[AÁ]BILES\s*", re.IGNORECASE),
    re.compile(r"^\s*RESPUESTA\s+A\s+GLOSA\s+RATIFICADA\s*", re.IGNORECASE),
    # Banner "Tarifa pactada encontrada..." completo
    re.compile(
        r"Tarifa\s+pactada\s+encontrada\s+en\s+el\s+contrato.*?(?=ESE\s+HUS|DE\s+CONFORMIDAD|$)",
        re.IGNORECASE | re.DOTALL,
    ),
    # "EPS: ... | Factura: ... | Observación recepción: ..." (sin prefijo RATIFICADA)
    re.compile(r"^\s*EPS:\s*.*?\|\s*Factura:\s*.*?\|\s*Observaci[óo]n\s+recepci[óo]n:\s*", re.IGNORECASE | re.DOTALL),
    # Línea tipo "CUPS: 890750 EPS: ... Contrato: ..."
    re.compile(r"CUPS:\s*\S+\s+EPS:\s*[^|]*?\s+Contrato:\s*\S+.*?(?=ESE\s+HUS|DE\s+CONFORMIDAD|$)", re.IGNORECASE | re.DOTALL),
    # Tabla encabezado "CÓDIGO GLOSA VALOR OBJETADO CÓDIGO RESPUESTA TA0201 $ N RE9901"
    re.compile(r"C[ÓO]DIGO\s+GLOSA\s+VALOR\s+OBJETADO\s+C[ÓO]DIGO\s+RESPUESTA\s+\S+\s+\$\s*[\d\.,]+\s+RE\d+", re.IGNORECASE),
    # Banner "GLOSA RATIFICADA - SE MANTIENE RESPUESTA INICIAL, SE SOLICITA CONCILIACIÓN"
    re.compile(r"GLOSA\s+RATIFICADA\s*[-–]\s*SE\s+MANTIENE\s+RESPUESTA\s+INICIAL.*?CONCILIACI[ÓO]N", re.IGNORECASE),
    # Meta "N° Factura: HUS000... [EPS] RATIFICADA"
    re.compile(r"N[°º]\s*Factura:\s*\S+\s+.*?\s+RATIFICADA", re.IGNORECASE),
    # Nota pie "Generado con asistencia de IA. Verificar antes de radicar..."
    re.compile(r"Nota:\s*Generado\s+con\s+asistencia\s+de\s+IA.*", re.IGNORECASE | re.DOTALL),
    # Sección "RELACIÓN DE SOPORTES APORTADOS ... tabla ... Res. XXX/XXXX"
    re.compile(r"RELACI[ÓO]N\s+DE\s+SOPORTES\s+APORTADOS.*", re.IGNORECASE | re.DOTALL),
    # Título "ARGUMENTACIÓN JURÍDICA" suelto
    re.compile(r"ARGUMENTACI[ÓO]N\s+JUR[ÍI]DICA\s*(?=[A-Z])", re.IGNORECASE),
)


# Para extraer descripción de servicio desde observacion_eps
# Patrón típico: "CUPS 881434H - PERFIL BIOFISICO - Valor objetado"
_DESCRIPCION_SERVICIO_RE = re.compile(
    r"CUPS\s+[\w\-]+\s*-\s*(.+?)\s*-\s*Valor\s+objetado",
    re.IGNORECASE | re.DOTALL,
)


def extraer_descripcion_servicio(obs: str) -> str:
    """Ronda 39 fix: si el ConceptoGlosaRecord no tiene `cups_descripcion`
    pero la observación sí contiene 'CUPS NNN - NOMBRE SERVICIO - Valor objetado',
    extraer el nombre del servicio."""
    if not obs:
        return ""
    m = _DESCRIPCION_SERVICIO_RE.search(str(obs))
    if not m:
        return ""
    desc = m.group(1).strip()
    # Sacar saltos de línea internos
    desc = re.sub(r"\s+", " ", desc)
    # Limit razonable
    return desc[:300]


def _strip_html(s: str) -> str:
    if not s:
        return ""
    txt = _HTML_TAG_RE.sub(" ", str(s))
    txt = _html.unescape(txt)
    txt = re.sub(r"[ \t]+", " ", txt)
    txt = re.sub(r"\n{3,}", "\n\n", txt)
    return txt.strip()


def limpiar_dictamen_para_dgh(dictamen_html: str, glosa: GlosaRecord) -> str:
    """Devuelve el texto del dictamen sin emojis ni headers de debug,
    listo para la columna OBSERVACION del Excel DGH."""
    if not dictamen_html:
        return ""

    modelo = (getattr(glosa, "modelo_ia", "") or "").lower()

    # Atajo: si es texto fijo, devolver el canónico sin adornos
    if "texto_fijo/ratificada" in modelo:
        try:
            from app.services.glosa_service import TEXTO_RATIFICADA
            return TEXTO_RATIFICADA
        except Exception:
            pass
    if "texto_fijo/extemporanea" in modelo:
        try:
            from app.services.glosa_service import generar_texto_extemporanea
            dias = int(getattr(glosa, "dias_radicacion_dgh", 0) or 0)
            if dias <= 0:
                dias = 21
            return generar_texto_extemporanea(dias)
        except Exception:
            pass

    # Caso general: HTML → texto + quitar emojis + quitar headers de debug
    txt = _strip_html(dictamen_html)
    txt = _EMOJI_RE.sub("", txt)

    # Estrategia 1: si existe marcador 'ARGUMENTACIÓN JURÍDICA', cortar desde ahí
    m_inicio = _INICIO_ARGUMENTO.search(txt)
    if m_inicio:
        txt = txt[m_inicio.end():]
    # Estrategia 2: cortar antes de 'RELACIÓN DE SOPORTES APORTADOS' o 'Nota: Generado...'
    m_fin = _FIN_ARGUMENTO.search(txt)
    if m_fin:
        txt = txt[:m_fin.start()]

    # Remover headers sueltos residuales
    for rx in _HEADERS_BASURA:
        txt = rx.sub("", txt)
    txt = re.sub(r"\s{2,}", " ", txt).strip()
    # Algunas respuestas tienen un pipe delimiter · residual
    txt = re.sub(r"^\s*[·\|]\s*", "", txt).strip()
    return txt


# ─── Resolución del tercero (EPS) ─────────────────────────────────────────

def resolver_tercero(glosa: GlosaRecord) -> tuple[str, str, str]:
    """Retorna (NombreCompletoAN, NombreCompletoNA, Documento).

    Prioridad para NombreCompletoAN (con prefijo de código interno):
      1. eps_codigo + ' - ' + tercero_nombre (si ambos existen)
      2. eps (nombre institucional)
      3. tercero_nombre solo
      4. 'SIN DEFINIR' como último recurso

    Prioridad para NombreCompletoNA (corto sin prefijo):
      1. tercero_nombre (ya limpio)
      2. eps sin el prefijo de código si lo tiene
      3. 'SIN DEFINIR'
    """
    eps = (getattr(glosa, "eps", "") or "").strip()
    eps_codigo = (getattr(glosa, "eps_codigo", "") or "").strip()
    tercero_nombre = (getattr(glosa, "tercero_nombre", "") or "").strip()
    tercero_nit = (getattr(glosa, "tercero_nit", "") or "").strip()

    # Quitar prefijo tipo "U220181 - " de eps si lo trae
    eps_sin_prefijo = eps
    m = re.match(r"^\s*([A-Z]\d{6,})\s*[-–]\s*(.+)$", eps)
    if m:
        eps_sin_prefijo = m.group(2).strip()
        if not eps_codigo:
            eps_codigo = m.group(1).strip()

    # NombreCompletoAN
    if eps_codigo and tercero_nombre:
        nombre_an = f"{eps_codigo} - {tercero_nombre}"
    elif eps_codigo and eps_sin_prefijo:
        nombre_an = f"{eps_codigo} - {eps_sin_prefijo}"
    elif eps and eps.upper() not in ("OTRA / SIN DEFINIR", "OTRA", "SIN DEFINIR", ""):
        nombre_an = eps
    elif tercero_nombre:
        nombre_an = tercero_nombre
    else:
        nombre_an = "SIN DEFINIR"

    # NombreCompletoNA
    if tercero_nombre:
        nombre_na = tercero_nombre
    elif eps_sin_prefijo and eps_sin_prefijo.upper() not in ("OTRA / SIN DEFINIR", "SIN DEFINIR"):
        nombre_na = eps_sin_prefijo
    else:
        nombre_na = "SIN DEFINIR"

    return nombre_an, nombre_na, tercero_nit


# ─── Clasificadores derivados ─────────────────────────────────────────────

def estado_cxc_objecion(glosa: GlosaRecord) -> str:
    """Glosa_Inicial / Glosa_Ratificada / Glosa_Total (usado por DGH)."""
    est = (getattr(glosa, "estado", "") or "").upper()
    wf = (getattr(glosa, "workflow_state", "") or "").upper()
    etapa = (getattr(glosa, "etapa", "") or "").upper()
    tipo_exc = (getattr(glosa, "tipo_glosa_excel", "") or "").upper()

    if "RATIF" in est or "RATIF" in wf or "RATIF" in etapa or tipo_exc == "R":
        return "Glosa_Ratificada"
    if tipo_exc == "I" or "INICIAL" in etapa:
        return "Glosa_Inicial"
    if "TOTAL" in tipo_exc or "DEVOL" in tipo_exc:
        return "Glosa_Total"
    # Default conservador: Inicial
    return "Glosa_Inicial"


def tipo_objecion_tramite(codigo_glosa: str) -> str:
    """Administrativo / Clínico según el prefijo del código DGH."""
    if not codigo_glosa:
        return "Administrativo"
    pref = codigo_glosa[:2].upper()
    # Prefijos clínicos (manual único de glosas 2284/2023)
    if pref in ("CL", "PE"):  # CL = Calidad clínica, PE = Pertinencia
        return "Clínico"
    return "Administrativo"


def codigo_respuesta_efectivo(glosa: GlosaRecord, concepto: Optional[ConceptoGlosaRecord] = None) -> str:
    """Código RE9XXX de la respuesta. Si no lo tenemos, inferimos del flujo.

    Para RATIFICADAS el código canónico es RE9901 (no acepta ratificación).
    Para iniciales con dictamen de defensa, RE9901.
    Para aceptación total → RE9702. Para parcial → RE9801. Para devolución
    justificada → RE9502. Para glosa injustificada → RE9602.
    """
    c = (getattr(glosa, "codigo_respuesta", "") or "").strip().upper()
    if c:
        return c
    modelo = (getattr(glosa, "modelo_ia", "") or "").lower()
    if "texto_fijo/ratificada" in modelo:
        return "RE9901"
    if "texto_fijo/extemporanea" in modelo:
        return "RE9501"  # glosa extemporánea → no acepta
    # Default defensa
    return "RE9901"


# ─── Formateadores ────────────────────────────────────────────────────────

def _fmt_fecha(dt) -> str:
    """Devuelve DD/MM/AAAA en formato español."""
    if not dt:
        return ""
    if isinstance(dt, str):
        # Intentar parsear ISO
        try:
            dt = datetime.fromisoformat(dt.replace("Z", "+00:00"))
        except Exception:
            return dt
    try:
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return str(dt)


def _fmt_fecha_hora(dt) -> str:
    if not dt:
        return ""
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt.replace("Z", "+00:00"))
        except Exception:
            return dt
    try:
        return dt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(dt)


# ─── Generación del Excel ─────────────────────────────────────────────────

def generar_filas_dgh(
    db: Session,
    glosas: list[GlosaRecord],
    fecha_cargue: Optional[datetime] = None,
) -> list[dict]:
    """Arma las filas (una por concepto) listas para escribir al Excel."""
    if fecha_cargue is None:
        fecha_cargue = datetime.now()
    fcargue = _fmt_fecha(fecha_cargue)

    # Precargar todos los conceptos de estas glosas en una query
    ids = [g.id for g in glosas if g.id]
    conceptos_por_glosa: dict[int, list[ConceptoGlosaRecord]] = {}
    if ids:
        todos = (
            db.query(ConceptoGlosaRecord)
            .filter(ConceptoGlosaRecord.glosa_id.in_(ids))
            .all()
        )
        for c in todos:
            conceptos_por_glosa.setdefault(c.glosa_id, []).append(c)

    filas = []
    for g in glosas:
        nombre_an, nombre_na, nit = resolver_tercero(g)
        estado_cxc = estado_cxc_objecion(g)
        observacion_limpia = limpiar_dictamen_para_dgh(g.dictamen or "", g)
        codigo_resp = codigo_respuesta_efectivo(g)
        valor_aceptado = float(g.valor_aceptado or 0.0)
        fecha_doc = _fmt_fecha(g.fecha_documento_dgh or g.creado_en)
        fecha_fact = _fmt_fecha_hora(g.fecha_radicacion_factura or g.creado_en)
        fecha_obj = _fmt_fecha(g.fecha_objecion_eps or g.fecha_recepcion or g.creado_en)
        factura = (g.factura or "").strip()
        consec = (g.consecutivo_dgh or "").strip()
        valor_factura = float(g.valor_factura or 0.0)

        conceptos = conceptos_por_glosa.get(g.id, [])
        if not conceptos:
            # Fallback: 1 fila con los datos de la glosa misma
            # Ronda 42: también aplicar extracción de descripción de servicio
            desc_fallback = (g.servicio_descripcion or "").strip()
            if not desc_fallback:
                desc_fallback = extraer_descripcion_servicio(g.texto_glosa_original or g.observacion_eps or "")
            fila = {
                "EstadoCxCObjecion": estado_cxc,
                "TipoObjecionTramite": tipo_objecion_tramite(g.codigo_glosa or ""),
                "FacturaCartera.Factura": factura,
                "FechaDocumento": fecha_doc,
                "Consecutivo": consec,
                "Observaciones": "",
                "EstadoActual": "Confirmado",
                "FacturaCartera.Valor": valor_factura,
                "FacturaCartera.Fecha": fecha_fact,
                "FacturaCartera.Tercero.Documento": nit,
                "FacturaCartera.Tercero.NombreCompletoAN": nombre_an,
                "FechaObjecion": fecha_obj,
                "FacturaCartera.Tercero.NombreCompletoNA": nombre_na,
                "ListadoConceptos.ConceptoObjecion.Codigo": g.codigo_glosa or "",
                "ListadoConceptos.Oid": "",
                "ListadoConceptos.ConceptoObjecion.Nombre": g.concepto_glosa or "",
                "ListadoConceptos.ServicioProductoFactura.Codigo": g.cups_servicio or "",
                "ListadoConceptos.ServicioProductoFactura.Descripcion": desc_fallback,
                "ListadoConceptos.ValorObjecion": float(g.valor_objetado or 0.0),
                "ListadoConceptos.ServicioProductoFactura.CentroCosto.CodigoNombreCentro": "",
                "ListadoConceptos.Observaciones": g.texto_glosa_original or g.observacion_eps or "",
                "FECHA DE CARGUE": fcargue,
                "CODIGO RESPUESTA": codigo_resp,
                "VALOR ACEPTADO": valor_aceptado,
                "OBSERVACION": observacion_limpia,
            }
            filas.append(fila)
            continue

        # Una fila por concepto (formato canónico DGH)
        for c in conceptos:
            # Ronda 39: fallback para descripción del servicio
            desc_servicio = (c.cups_descripcion or "").strip()
            if not desc_servicio:
                desc_servicio = extraer_descripcion_servicio(c.observacion_eps or "")
            if not desc_servicio:
                # Último fallback: intentar desde el nombre_glosa del concepto
                desc_servicio = extraer_descripcion_servicio(g.texto_glosa_original or "")

            fila = {
                "EstadoCxCObjecion": estado_cxc,
                "TipoObjecionTramite": tipo_objecion_tramite(c.codigo_glosa or g.codigo_glosa or ""),
                "FacturaCartera.Factura": factura,
                "FechaDocumento": fecha_doc,
                "Consecutivo": consec,
                "Observaciones": "",
                "EstadoActual": "Confirmado",
                "FacturaCartera.Valor": valor_factura,
                "FacturaCartera.Fecha": fecha_fact,
                "FacturaCartera.Tercero.Documento": nit,
                "FacturaCartera.Tercero.NombreCompletoAN": nombre_an,
                "FechaObjecion": fecha_obj,
                "FacturaCartera.Tercero.NombreCompletoNA": nombre_na,
                "ListadoConceptos.ConceptoObjecion.Codigo": c.codigo_glosa or "",
                "ListadoConceptos.Oid": c.oid_dgh or "",
                "ListadoConceptos.ConceptoObjecion.Nombre": c.nombre_glosa or "",
                "ListadoConceptos.ServicioProductoFactura.Codigo": c.cups_codigo or "",
                "ListadoConceptos.ServicioProductoFactura.Descripcion": desc_servicio,
                "ListadoConceptos.ValorObjecion": float(c.valor_objetado or 0.0),
                "ListadoConceptos.ServicioProductoFactura.CentroCosto.CodigoNombreCentro": c.centro_costo or "",
                "ListadoConceptos.Observaciones": c.observacion_eps or "",
                "FECHA DE CARGUE": fcargue,
                "CODIGO RESPUESTA": codigo_resp,
                "VALOR ACEPTADO": valor_aceptado,
                "OBSERVACION": limpiar_dictamen_para_dgh(c.dictamen_html or g.dictamen or "", g),
            }
            filas.append(fila)

    return filas


def generar_excel_dgh(
    db: Session,
    glosas: list[GlosaRecord],
    fecha_cargue: Optional[datetime] = None,
) -> BytesIO:
    """Construye el Excel con exactamente las 26 columnas DGH."""
    if not EXCEL_OK:
        raise ImportError("openpyxl no instalado")

    wb = Workbook()
    ws = wb.active
    ws.title = "Glosas_DGH"

    # Encabezado
    header_fill = PatternFill("solid", fgColor="1e40af")
    header_font = Font(size=10, bold=True, color="FFFFFF")
    for col_idx, col in enumerate(COLUMNAS_DGH, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    filas = generar_filas_dgh(db, glosas, fecha_cargue=fecha_cargue)
    for i, fila in enumerate(filas, start=2):
        for col_idx, col in enumerate(COLUMNAS_DGH, start=1):
            v = fila.get(col, "")
            ws.cell(row=i, column=col_idx, value=v)

    # Ancho básico por columna
    ANCHOS = {
        "EstadoCxCObjecion": 16,
        "TipoObjecionTramite": 18,
        "FacturaCartera.Factura": 18,
        "FechaDocumento": 12,
        "Consecutivo": 12,
        "EstadoActual": 12,
        "FacturaCartera.Valor": 14,
        "FacturaCartera.Fecha": 18,
        "FacturaCartera.Tercero.Documento": 15,
        "FacturaCartera.Tercero.NombreCompletoAN": 50,
        "FechaObjecion": 12,
        "FacturaCartera.Tercero.NombreCompletoNA": 50,
        "ListadoConceptos.ConceptoObjecion.Codigo": 10,
        "ListadoConceptos.Oid": 10,
        "ListadoConceptos.ConceptoObjecion.Nombre": 40,
        "ListadoConceptos.ServicioProductoFactura.Codigo": 14,
        "ListadoConceptos.ServicioProductoFactura.Descripcion": 40,
        "ListadoConceptos.ValorObjecion": 14,
        "ListadoConceptos.ServicioProductoFactura.CentroCosto.CodigoNombreCentro": 30,
        "ListadoConceptos.Observaciones": 60,
        "FECHA DE CARGUE": 14,
        "CODIGO RESPUESTA": 14,
        "VALOR ACEPTADO": 14,
        "OBSERVACION": 80,
    }
    for col_idx, col in enumerate(COLUMNAS_DGH, start=1):
        letra = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letra].width = ANCHOS.get(col, 20)

    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
