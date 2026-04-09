"""
Servicio de exportación de reportes Excel para el Motor de Glosas HUS.
Genera reportes con formato institucional.
"""
from datetime import datetime
from typing import Optional
from io import BytesIO

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False

from app.core.logging_utils import logger


class ExcelExporter:
    """Genera reportes Excel con formato institucional HUS."""
    
    HEADER_FILL = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    TITLE_FONT = Font(bold=True, size=14, color="1e40af")
    SUBTITLE_FONT = Font(italic=True, size=10, color="64748b")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    COLORES_ESTADO = {
        "RESPONDIDA": "dcfce7",
        "ACEPTADA": "dcfce7",
        "PARCIALMENTE_ACEPTADA": "fef9c3",
        "RATIFICADA": "fee2e2",
        "LEVANTADA": "dcfce7",
        "CONCILIADA": "dcfce7",
        "RADICADA": "f1f5f9",
    }
    
    COLORES_PRIORIDAD = {
        "ALTA": "ef4444",
        "MEDIA": "f59e0b",
        "BAJA": "22c55e",
    }

    def __init__(self):
        if not EXCEL_DISPONIBLE:
            logger.warning("openpyxl no instalado. Exportación Excel no disponible.")
    
    def generar_reporte_glosas(
        self,
        glosas: list,
        titulo: str = "Reporte de Glosas",
        fecha_inicio: Optional[str] = None,
        fecha_fin: Optional[str] = None,
    ) -> BytesIO:
        """Genera un archivo Excel con el reporte de glosas."""
        if not EXCEL_DISPONIBLE:
            raise ImportError("openpyxl no está instalado. Ejecute: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Glosas"
        
        self._agregar_encabezado(ws, titulo, fecha_inicio, fecha_fin)
        self._agregar_encabezados_columnas(ws)
        self._agregar_datos(ws, glosas)
        self._ajustar_columnas(ws)
        self._agregar_totales(ws, glosas)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def _agregar_encabezado(self, ws, titulo: str, fecha_inicio: Optional[str], fecha_fin: Optional[str]):
        ws.merge_cells("A1:L1")
        ws["A1"] = f"ESE HOSPITAL UNIVERSITARIO DE SANTANDER - {titulo.upper()}"
        ws["A1"].font = self.TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25
        
        ws.merge_cells("A2:L2")
        fecha_texto = f"Período: {fecha_inicio or 'Inicio'} - {fecha_fin or 'Actual'}"
        if fecha_inicio or fecha_fin:
            ws["A2"] = fecha_texto
        ws["A2"].font = self.SUBTITLE_FONT
        ws["A2"].alignment = Alignment(horizontal="center")
        
        ws.merge_cells("A3:L3")
        ws["A3"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A3"].font = self.SUBTITLE_FONT
        ws["A3"].alignment = Alignment(horizontal="center")
    
    def _agregar_encabezados_columnas(self, ws):
        headers = [
            "ID", "Fecha", "EPS", "Paciente", "N° Factura", "N° Radicado",
            "Código", "Valor Objetado", "Valor Aceptado", "Estado", "Prioridad", "Días Rest."
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.BORDER
        ws.row_dimensions[5].height = 30
    
    def _agregar_datos(self, ws, glosas: list):
        for row_idx, glosa in enumerate(glosas, 6):
            estado = getattr(glosa, "estado", "RADICADA") or "RADICADA"
            prioridad = getattr(glosa, "prioridad", "BAJA") or "BAJA"
            
            valores = [
                getattr(glosa, "id", ""),
                self._formato_fecha(getattr(glosa, "creado_en", None)),
                getattr(glosa, "eps", ""),
                getattr(glosa, "paciente", ""),
                getattr(glosa, "factura", ""),
                getattr(glosa, "numero_radicado", ""),
                getattr(glosa, "codigo_glosa", ""),
                getattr(glosa, "valor_objetado", 0) or 0,
                getattr(glosa, "valor_aceptado", 0) or 0,
                estado,
                prioridad,
                getattr(glosa, "dias_restantes", 0) or 0,
            ]
            
            for col, valor in enumerate(valores, 1):
                cell = ws.cell(row=row_idx, column=col, value=valor)
                cell.border = self.BORDER
                cell.alignment = Alignment(horizontal="center" if col in [1, 7, 10, 11, 12] else "left")
                
                if col in [8, 9]:
                    cell.number_format = '"$" #,##0.00'
                
                if col == 10:
                    fill_color = self.COLORES_ESTADO.get(estado, "f1f5f9")
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    cell.font = Font(bold=True)
                
                if col == 11:
                    color = self.COLORES_PRIORIDAD.get(prioridad, "64748b")
                    cell.font = Font(color=color, bold=True)
    
    def _ajustar_columnas(self, ws):
        column_widths = [8, 12, 20, 25, 15, 15, 10, 15, 15, 18, 10, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    def _agregar_totales(self, ws, glosas: list):
        total_obj = sum(g.valor_objetado or 0 for g in glosas)
        total_acept = sum(g.valor_aceptado or 0 for g in glosas)
        
        last_row = 6 + len(glosas)
        ws.cell(row=last_row + 1, column=7, value="TOTALES:").font = Font(bold=True)
        ws.cell(row=last_row + 1, column=8, value=total_obj).number_format = '"$" #,##0.00'
        ws.cell(row=last_row + 1, column=8).font = Font(bold=True)
        ws.cell(row=last_row + 1, column=9, value=total_acept).number_format = '"$" #,##0.00'
        ws.cell(row=last_row + 1, column=9).font = Font(bold=True)
        ws.cell(row=last_row + 1, column=11, value=f"Recuperado: {total_acept/total_obj*100:.1f}%" if total_obj > 0 else "N/A").font = Font(bold=True, color="16a34a")
    
    def _formato_fecha(self, fecha) -> str:
        if fecha is None:
            return ""
        if isinstance(fecha, str):
            try:
                fecha = datetime.fromisoformat(fecha.replace("Z", "+00:00"))
            except:
                return fecha
        return fecha.strftime("%Y-%m-%d")
    
    def generar_resumen_mensual(
        self,
        tendencias: list,
        eps: str = "TODAS",
    ) -> BytesIO:
        """Genera reporte de tendencias mensuales."""
        if not EXCEL_DISPONIBLE:
            raise ImportError("openpyxl no está instalado.")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen Mensual"
        
        ws.merge_cells("A1:F1")
        ws["A1"] = f"RESUMEN MENSUAL DE GLOSAS - {eps.upper()}"
        ws["A1"].font = self.TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center")
        
        headers = ["Mes", "Cantidad", "Valor Objetado", "Valor Aceptado", "Recuperado", "Tasa Éxito %"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.BORDER
        
        for row_idx, t in enumerate(tendencias, 4):
            ws.cell(row=row_idx, column=1, value=t.get("mes", "")).border = self.BORDER
            ws.cell(row=row_idx, column=2, value=t.get("count", 0)).border = self.BORDER
            ws.cell(row=row_idx, column=3, value=t.get("objetado", 0)).border = self.BORDER
            ws.cell(row=row_idx, column=3).number_format = '"$" #,##0.00'
            ws.cell(row=row_idx, column=4, value=t.get("aceptado", 0)).border = self.BORDER
            ws.cell(row=row_idx, column=4).number_format = '"$" #,##0.00'
            ws.cell(row=row_idx, column=5, value=t.get("recuperado", 0)).border = self.BORDER
            ws.cell(row=row_idx, column=5).number_format = '"$" #,##0.00'
            
            tasa = t.get("objetado", 0)
            acept = t.get("aceptado", 0)
            pct = (acept / tasa * 100) if tasa > 0 else 0
            ws.cell(row=row_idx, column=6, value=f"{pct:.1f}%").border = self.BORDER
            ws.cell(row=row_idx, column=6).font = Font(color="16a34a" if pct > 70 else "dc2626", bold=True)
        
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 15
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
