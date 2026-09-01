"""El informe en Excel del paquete de Glosas ADRES.

POR QUÉ EXISTE (31-08-2026). Yesid, textual: «me gustaría que en esta opción
también esté poder descargar un excel con un informe así como el que tenemos en
el apartado de preauditoría», y después, mandando dos archivos del paquete
31068: «los archivos descargados deben ser así como estos».

Esas dos frases mandan el diseño:

· **La hoja de datos es la de siempre.** `Hoja1` sale con las 26 columnas de la
  macro, en el mismo orden y con los mismos títulos, encabezado en la fila 1 y
  los datos desde la 2 — igual que los archivos `RTA_GLOSA_ADRES_PAQ_*` con los
  que trabaja el área. Así el archivo bajado sirve para lo de siempre: la tabla
  dinámica de valores aceptados y los bots que lo leen
  (`organizar_objeciones_adres.py`, `respuestas_adres_por_factura.py`), que
  **buscan la hoja por sus encabezados en la primera fila**.
· **Encima va el informe.** Las otras hojas leen de `Hoja1` con fórmulas vivas:
  el resumen del paquete, las glosas agrupadas por causal, el reparto por área
  y centro de costos, el avance de cada gestor y el estado de cada factura. Si
  alguien corrige una fila, todo se recalcula solo.

LAS DOS REGLAS DE PLATA QUE NO SE PUEDEN ROMPER (vienen de la pantalla, no se
inventan acá):

1. El reporte del ADRES abre **una fila por cada causal** del mismo ítem. Todas
   las filas se conservan —el gestor decide causal por causal— pero solo una
   cuenta para la plata; si no, la glosa sale al doble o al triple. Es la
   columna «CUENTA PARA LA PLATA» y todas las sumas de este libro la usan.
2. Lo ACEPTADO se consolida por ítem y se topa en lo glosado
   (`aceptado_consolidado`), porque aceptar en dos causales del mismo servicio
   le declararía al ADRES el doble de lo que ese servicio tiene glosado.

Vive en services/ y no en el router porque es lógica del negocio: qué se
agrupa, qué cuenta y qué no.
"""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.core.tz import TZ_BOGOTA
from app.models.db import (
    FacturaAdresRecord,
    GlosaAdresRecord,
    PaqueteAdresRecord,
)

# El formato de la macro y el texto de la respuesta salen del script del bot:
# es la fuente única del criterio, para no tener dos copias.
from app.services.preauditoria_adres import (
    COLUMNAS_APOYO,
    COLUMNAS_MACRO,
    _clave_item,  # el criterio de «mismo ítem» vive allá; acá no se duplica
    _fila_macro,
    aceptado_consolidado,
    cantidad_glosada,
)

FUENTE = "Arial"
AZUL = "1F3864"
AZUL_CLARO = "D9E2F3"
GRIS = "F2F2F2"
AMBAR = "FFF2CC"
ROJO = "F8CECC"
VERDE = "E2EFDA"

HOJA_DATOS = "Hoja1"  # el mismo nombre de los archivos del área
SIN_GESTOR = "(sin gestor asignado)"
SIN_AREA = "(sin clasificar)"
SIN_CENTRO = "(sin centro de costos)"
ITEM = "ÍTEM"
GLOSA_TOTAL = "GLOSA TOTAL"
SI = "SÍ"
NO = "NO"

# Columnas propias, DESPUÉS de las de la macro y las de apoyo, para no correr
# ninguna: el VBA y los bots trabajan por posición sobre las 26 primeras.
COLUMNAS_CONTROL = [
    "CANTIDAD GLOSADA",
    "TIPO DE RENGLÓN",
    "CUENTA PARA LA PLATA",
    "VALOR ACEPTADO QUE SE DECLARA",
    "QUIÉN DECIDIÓ",
    "CUÁNDO DECIDIÓ",
    "FALTA REPARTIR ÁREA",
]
ENCABEZADOS = COLUMNAS_MACRO + COLUMNAS_APOYO + COLUMNAS_CONTROL
COL_DECISION = "OBSERVACION (SE ACEPTA - SE OBJETA -  SE SUBSANA )"

# Qué hace falta para responder las causales que más se repiten. Sale de lo que
# el equipo ya responde; lo que no esté acá sale en blanco — no se inventa.
ACCIONES_CAUSAL: dict[str, str] = {
    "3106": (
        "Anexar el soporte del material o insumo: factura de compra, hoja de gasto o "
        "registro de entrega."
    ),
    "3109": "Anexar el soporte del procedimiento con su registro en la historia clínica.",
    "3202": (
        "Sustentar la pertinencia de la consulta con el motivo de ingreso y los hallazgos "
        "del examen físico."
    ),
    "3209": (
        "Sustentar la pertinencia de la ayuda diagnóstica con el mecanismo del trauma, los "
        "hallazgos clínicos y la sospecha diagnóstica."
    ),
    "3301": "Revisar la tarifa contra el contrato o el manual que aplique y anexar la liquidación.",
    "4506": (
        "Definir primero qué área responde: material de otro servicio (facturación) o "
        "pertinencia (auditoría médica)."
    ),
    "4507": (
        "Revisar si el medicamento está incluido en el servicio o si se entregó aparte, y "
        "anexar el soporte de la entrega."
    ),
}


def _col(nombre: str) -> str:
    """La letra de esa columna en la hoja de datos."""
    return get_column_letter(ENCABEZADOS.index(nombre) + 1)


# ------------------------------------------------------------------
# Ayudas
# ------------------------------------------------------------------
def _texto(v) -> str:
    return (str(v).strip() if v is not None else "") or ""


def _fecha(v) -> Optional[datetime]:
    if isinstance(v, datetime):
        return v.astimezone(TZ_BOGOTA).replace(tzinfo=None) if v.tzinfo else v
    return None


def _codigo_causal(g: GlosaAdresRecord) -> str:
    """El código de la causal, del campo propio o del texto («3209- ...»)."""
    codigo = _texto(g.causal_codigo)
    if codigo:
        return codigo
    texto = _texto(g.causal_texto)
    return texto.split("-", 1)[0].strip() if "-" in texto else ""


def _titulo(ws, texto: str, sub: str, ancho: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ancho)
    c = ws.cell(1, 1, texto)
    c.font = Font(name=FUENTE, size=15, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ancho)
    c = ws.cell(2, 1, sub)
    c.font = Font(name=FUENTE, size=9.5, italic=True, color="404040")
    c.alignment = Alignment(indent=1, vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30


def _encabezado(ws, fila: int, cols: list[str], anchos: list[int]) -> None:
    borde = Border(bottom=Side(style="medium", color=AZUL))
    for i, v in enumerate(cols, 1):
        c = ws.cell(fila, i, v)
        c.font = Font(name=FUENTE, size=10, bold=True, color=AZUL)
        c.fill = PatternFill("solid", fgColor=AZUL_CLARO)
        c.border = borde
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[fila].height = 30
    for i, a in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = a


def _celda(
    ws, fila, col, valor, *, negrita=False, fmt=None, ajuste=False, centro=False, relleno=None
):
    c = ws.cell(fila, col, valor)
    c.font = Font(name=FUENTE, size=10, bold=negrita)
    c.alignment = Alignment(
        vertical="center" if not ajuste else "top",
        wrap_text=ajuste,
        horizontal="center" if centro else "general",
    )
    if fmt:
        c.number_format = fmt
    if relleno:
        c.fill = PatternFill("solid", fgColor=relleno)
    return c


# ------------------------------------------------------------------
# Los datos, en pocas consultas (nunca una por factura)
# ------------------------------------------------------------------
def _reunir(db: Session, paquete_id: int) -> tuple[list[dict], list[dict]]:
    """Las glosas y las facturas del paquete, ya con la plata consolidada."""
    glosas = (
        db.query(GlosaAdresRecord)
        .filter(GlosaAdresRecord.paquete_id == paquete_id)
        .order_by(GlosaAdresRecord.factura, GlosaAdresRecord.id)
        .all()
    )
    fichas = (
        db.query(FacturaAdresRecord)
        .filter(FacturaAdresRecord.paquete_id == paquete_id)
        .order_by(FacturaAdresRecord.factura)
        .all()
    )

    por_factura: dict[str, list[GlosaAdresRecord]] = defaultdict(list)
    for g in glosas:
        por_factura[g.factura_clave].append(g)
    aceptado_en_fila: dict[int, float] = {}
    aceptado_factura: dict[str, float] = {}
    for clave, grupo in por_factura.items():
        aceptado_factura[clave] = aceptado_consolidado(grupo)
        por_item: dict[tuple, list[GlosaAdresRecord]] = defaultdict(list)
        for g in grupo:
            por_item[_clave_item(g.factura_clave, g.codigo, g.valor_glosado)].append(g)
        for filas in por_item.values():
            # Todo lo aceptado del ítem se pone en el renglón que cuenta: así la
            # columna suma exactamente lo que la pantalla le declara al ADRES.
            portador = next((g for g in filas if g.cuenta_valor), filas[0])
            aceptado_en_fila[portador.id] = aceptado_consolidado(filas)

    datos = []
    for g in glosas:
        datos.append(
            {
                "cod_habilitacion": _texto(g.cod_habilitacion),
                "radicacion": _texto(g.radicacion),
                "factura": _texto(g.factura),
                "factura_clave": _texto(g.factura_clave),
                "cant_reclamada": g.cant_reclamada or 0,
                "valor_reclamado": g.valor_reclamado or 0,
                "cant_aprobada": g.cant_aprobada or 0,
                "valor_aprobado": g.valor_aprobado or 0,
                "valor_glosado": g.valor_glosado or 0,
                "paciente": _texto(g.doc_victima),
                "consecutivo": _texto(g.consecutivo),
                "elemento": _texto(g.tipo_elemento),
                "codigo": _texto(g.codigo),
                "servicio": _texto(g.descripcion),
                "causal_texto": _texto(g.causal_texto),
                "anotacion": _texto(g.anotacion),
                "causal": _codigo_causal(g),
                "area": _texto(g.clasificacion) or SIN_AREA,
                "decision": _texto(g.decision),
                "observacion": _texto(g.observacion_tecnico),
                "rta_completa": _fila_macro(g).rta_glosa_completa,
                "cantidad_aceptada": _texto(g.cantidad_aceptada),
                "aceptado_escrito": g.valor_aceptado or 0,
                "centro": _texto(g.centro_costos) or SIN_CENTRO,
                "gestor": _texto(g.gestor) or SIN_GESTOR,
                "medico": _texto(g.medico),
                "sugerencia": _texto(g.sugerencia),
                "confianza": _texto(g.confianza),
                "motivo": _texto(g.motivo),
                "estado_detallado": _texto(g.estado_detallado),
                "cant_glosada": cantidad_glosada(g.cant_reclamada, g.cant_aprobada),
                "tipo": GLOSA_TOTAL if g.glosa_total else ITEM,
                "cuenta": SI if g.cuenta_valor else NO,
                "aceptado_declarado": aceptado_en_fila.get(g.id, 0.0),
                "decidido_por": _texto(g.decidido_por),
                "decidido_en": _fecha(g.decidido_en),
                "falta_area": SI if g.requiere_asignacion and not g.area_asignada_por else "",
            }
        )

    # Los conteos de cada factura, del mismo recorrido (nada de N+1).
    resumen: dict[str, dict] = defaultdict(
        lambda: {"items": 0, "decididas": 0, "totales": 0, "glosado": 0.0}
    )
    for g in glosas:
        r = resumen[g.factura_clave]
        if g.cuenta_valor:
            r["glosado"] += g.valor_glosado or 0
        if g.glosa_total:
            r["totales"] += 1
            continue
        r["items"] += 1
        if _texto(g.decision):
            r["decididas"] += 1

    facturas = []
    for f in fichas:
        r = resumen.get(f.factura_clave, {"items": 0, "decididas": 0, "totales": 0, "glosado": 0.0})
        glosado = r["glosado"]
        aceptado = aceptado_factura.get(f.factura_clave, 0.0)
        oficial = f.valor_glosado_oficial
        facturas.append(
            {
                "factura": _texto(f.factura),
                "radicacion": _texto(f.radicacion),
                "paciente": _texto(f.doc_victima),
                "gestor": _texto(f.gestor) or SIN_GESTOR,
                "medico": _texto(f.medico),
                "estado": _texto(f.estado) or "PENDIENTE",
                "glosas": r["items"],
                "decididas": r["decididas"],
                "pendientes": r["items"] - r["decididas"],
                "avance": (r["decididas"] / r["items"]) if r["items"] else 1.0,
                "totales": r["totales"],
                "glosado": glosado,
                "oficial": oficial,
                "cuadra": (
                    "" if oficial is None else (SI if abs(glosado - oficial) < 1 else "NO CUADRA")
                ),
                "aceptado": aceptado,
                "sigue_glosado": max(0.0, round(glosado - aceptado, 2)),
                "cerrada_por": _texto(f.cerrada_por),
                "cerrada_en": _fecha(f.cerrada_en),
            }
        )
    return datos, facturas


# ------------------------------------------------------------------
# El libro
# ------------------------------------------------------------------
def construir_informe_paquete(db: Session, paquete_id: int, generado_por: str = "") -> bytes:
    """El paquete del ADRES completo: la hoja de siempre, con el informe encima."""
    paquete = db.get(PaqueteAdresRecord, paquete_id)
    numero = _texto(paquete.numero_paquete) if paquete else ""
    datos, facturas = _reunir(db, paquete_id)
    hoy = datetime.now(TZ_BOGOTA).strftime("%d/%m/%Y %H:%M")
    wb = Workbook()

    # ── Hoja1: el formato de siempre, con el encabezado en la fila 1 ──
    # No se toca el orden ni el título de las 26 primeras columnas: la macro
    # trabaja por posición y los bots buscan la hoja por esos encabezados.
    hd = wb.active
    hd.title = HOJA_DATOS
    borde = Border(bottom=Side(style="medium", color=AZUL))
    for j, titulo in enumerate(ENCABEZADOS, 1):
        c = hd.cell(1, j, titulo)
        c.font = Font(name=FUENTE, size=10, bold=True, color=AZUL)
        c.fill = PatternFill("solid", fgColor=AZUL_CLARO if j <= len(COLUMNAS_MACRO) else VERDE)
        c.border = borde
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hd.row_dimensions[1].height = 34
    anchos = {
        "Número Factura": 15,
        "Tipo Elemento": 18,
        "Cod Elemento": 16,
        "Descripción Elemento": 34,
        "Descripción Glosa": 34,
        "Descripción Anotación": 40,
        "CLASIFICACION DE LA GLOSA": 18,
        COL_DECISION: 16,
        "OBSERVACION TECNICO / PROFESIONAL": 60,
        "RTA GLOSA COMPLETA": 60,
        "CENTRO DE COSTOS": 24,
        "GESTOR": 18,
        "POR QUÉ": 34,
        "CAUSAL (TEXTO)": 34,
        "VALOR ACEPTADO QUE SE DECLARA": 15,
    }
    for i, titulo in enumerate(ENCABEZADOS, 1):
        hd.column_dimensions[get_column_letter(i)].width = anchos.get(titulo, 13)

    color_dec = {"SE ACEPTA": AMBAR, "SE OBJETA": VERDE, "SE SUBSANA": AZUL_CLARO}
    dinero = {
        "Valor Reclamado",
        "Valor Aprobado",
        "Valor Glosado",
        "VALOR ACEPTADO",
        "VALOR ACEPTADO QUE SE DECLARA",
    }
    ajusta = {
        "Descripción Elemento",
        "Descripción Glosa",
        "Descripción Anotación",
        "OBSERVACION TECNICO / PROFESIONAL",
        "RTA GLOSA COMPLETA",
        "POR QUÉ",
        "CAUSAL (TEXTO)",
    }
    for n, d in enumerate(datos):
        f = 2 + n
        valores = [
            d["cod_habilitacion"],
            d["radicacion"],
            d["factura"],
            numero,
            d["cant_reclamada"],
            d["valor_reclamado"],
            d["cant_aprobada"],
            d["valor_aprobado"],
            d["valor_glosado"],
            d["paciente"],
            d["consecutivo"],
            d["elemento"],
            d["codigo"],
            d["servicio"],
            d["causal_texto"],
            d["anotacion"],
            d["causal"],
            d["area"],
            d["decision"] or None,  # en blanco mientras no se decida, como la macro
            d["observacion"],
            d["rta_completa"],
            d["cantidad_aceptada"],
            d["aceptado_escrito"],
            d["centro"],
            d["gestor"],
            d["medico"],
            d["sugerencia"],
            d["confianza"],
            d["motivo"],
            d["estado_detallado"],
            d["causal_texto"],
            d["cant_glosada"],
            d["tipo"],
            d["cuenta"],
            d["aceptado_declarado"],
            d["decidido_por"],
            d["decidido_en"],
            d["falta_area"],
        ]
        for j, v in enumerate(valores, 1):
            titulo = ENCABEZADOS[j - 1]
            c = hd.cell(f, j, v)
            c.font = Font(name=FUENTE, size=9, bold=(titulo == COL_DECISION))
            c.alignment = Alignment(vertical="top", wrap_text=(titulo in ajusta))
            if titulo in dinero:
                c.number_format = '"$"#,##0'
            if titulo == "CANTIDAD GLOSADA":
                c.alignment = Alignment(horizontal="center", vertical="top")
            if titulo == "CUÁNDO DECIDIÓ":
                c.number_format = "DD/MM/YYYY HH:MM"
            if titulo == COL_DECISION:
                c.fill = PatternFill("solid", fgColor=color_dec.get(v, ROJO))
            if titulo == "CUENTA PARA LA PLATA" and v == NO:
                c.fill = PatternFill("solid", fgColor=GRIS)
        hd.row_dimensions[f].height = 28
    ult = 1 + len(datos)
    hd.freeze_panes = "D2"
    if datos:
        hd.auto_filter.ref = f"A1:{get_column_letter(len(ENCABEZADOS))}{ult}"
    fin = max(ult, 2)

    def rango(nombre: str) -> str:
        letra = _col(nombre)
        return f"{HOJA_DATOS}!${letra}$2:${letra}${fin}"

    GES = rango("GESTOR")
    ARE = rango("CLASIFICACION DE LA GLOSA")
    CEN = rango("CENTRO DE COSTOS")
    CAU = rango("CODIGO NUMERICO")
    VGL = rango("Valor Glosado")
    DEC = rango(COL_DECISION)
    TIP = rango("TIPO DE RENGLÓN")
    CUE = rango("CUENTA PARA LA PLATA")
    ACE = rango("VALOR ACEPTADO QUE SE DECLARA")
    ARP = rango("FALTA REPARTIR ÁREA")

    # ── Los grupos que se listan (los cálculos los hace Excel) ──
    plata_causal: dict[str, float] = defaultdict(float)
    facturas_causal: dict[str, set] = defaultdict(set)
    texto_causal: dict[str, str] = {}
    area_causal: dict[str, str] = {}
    for d in datos:
        cod = d["causal"]
        if not cod:
            continue
        if d["cuenta"] == SI:
            plata_causal[cod] += d["valor_glosado"]
        facturas_causal[cod].add(d["factura"])
        texto_causal.setdefault(cod, d["causal_texto"])
        area_causal.setdefault(cod, d["area"])
    causales = sorted(facturas_causal, key=lambda c: -plata_causal[c])
    areas = sorted({d["area"] for d in datos})
    centros = sorted({d["centro"] for d in datos})
    gestores = sorted({d["gestor"] for d in datos})

    # ── CÓMO LEER ──
    cl = wb.create_sheet("CÓMO LEER", 0)
    _titulo(cl, "CÓMO LEER ESTE ARCHIVO", "Cada hoja responde una pregunta.", 4)
    cl.column_dimensions["A"].width = 3
    cl.column_dimensions["B"].width = 30
    cl.column_dimensions["C"].width = 88
    guia = [
        ("RESUMEN", "Los números gruesos del paquete y el avance. La página para presentar."),
        (
            "POR QUÉ NOS GLOSAN",
            "Las glosas agrupadas por causal del ADRES, con cuánta plata pesa cada una y qué "
            "hace falta para responderla.",
        ),
        (
            "POR ÁREA Y CENTRO",
            "Quién tiene que responder: el área (pertinencia, facturación…) y el centro de "
            "costos del hospital.",
        ),
        ("POR GESTOR", "Cuántas glosas tiene cada gestor, cuántas lleva decididas y cuánto falta."),
        (
            "FACTURAS",
            "Una fila por factura: su avance, lo glosado, lo aceptado y si la cifra del "
            "sistema cuadra con la del ADRES.",
        ),
        (
            HOJA_DATOS,
            "LA HOJA DE SIEMPRE: las 26 columnas de la macro, en el mismo orden y con los "
            "mismos títulos, con el encabezado en la fila 1. Es la que se usa para la tabla "
            "dinámica y la que leen los bots. Después de la 26 van columnas de apoyo, que no "
            "corren nada.",
        ),
    ]
    _celda(cl, 4, 2, "LAS HOJAS", negrita=True).font = Font(
        name=FUENTE, size=12, bold=True, color=AZUL
    )
    fila = 5
    for hoja, desc in guia:
        fila += 1
        c = _celda(cl, fila, 2, hoja, negrita=True, relleno=AZUL_CLARO)
        c.font = Font(name=FUENTE, size=10.5, bold=True, color=AZUL)
        _celda(cl, fila, 3, desc, ajuste=True)
        cl.row_dimensions[fila].height = 44
    fila += 2
    avisos = [
        (
            "LA PLATA",
            "El reporte del ADRES abre una fila por cada causal del mismo servicio. Todas se "
            "conservan —el gestor decide causal por causal— pero solo una cuenta para la plata: "
            "es la columna «CUENTA PARA LA PLATA», y todas las sumas de este libro la usan. Por "
            "eso una causal puede tener renglones y poco valor: la plata de ese servicio ya la "
            "contó la otra causal.",
            74,
        ),
        (
            "LO ACEPTADO",
            "«VALOR ACEPTADO» es lo que tecleó el gestor en cada renglón. «VALOR ACEPTADO QUE SE "
            "DECLARA» es lo que de verdad se le declara al ADRES: se junta por servicio y se topa "
            "en lo que ese servicio tiene glosado, para no reconocer dos veces la misma plata. "
            "Esa es la columna que suman los resúmenes.",
            60,
        ),
        (
            "GLOSA TOTAL",
            "Los renglones marcados GLOSA TOTAL son el desglose de una reclamación que el ADRES "
            "glosó entera por el FURIPS. No se responden uno por uno, así que no cuentan como "
            "glosas a trabajar, pero su plata sí está en juego.",
            48,
        ),
    ]
    for etiqueta, texto, alto in avisos:
        _celda(cl, fila, 2, etiqueta, negrita=True).font = Font(
            name=FUENTE, size=10, bold=True, color="9C0006"
        )
        _celda(cl, fila, 3, texto, ajuste=True)
        cl.row_dimensions[fila].height = alto
        fila += 1
    fila += 1
    _celda(cl, fila, 2, "Paquete", negrita=True)
    _celda(cl, fila, 3, numero or "—")
    fila += 1
    _celda(cl, fila, 2, "Generado", negrita=True)
    _celda(cl, fila, 3, f"{hoy}" + (f" · por {generado_por}" if generado_por else ""))

    # ── RESUMEN ──
    rs = wb.create_sheet("RESUMEN", 1)
    _titulo(
        rs,
        f"GLOSAS ADRES — RESUMEN DEL PAQUETE {numero or '—'}",
        f"{len(facturas)} factura(s) · {len(datos)} renglón(es) del reporte · generado el {hoy}"
        f" · todos los números son fórmulas sobre la hoja {HOJA_DATOS}",
        5,
    )
    rs.column_dimensions["A"].width = 3
    for col, an in zip("BCDE", (44, 16, 18, 58)):
        rs.column_dimensions[col].width = an
    _encabezado(rs, 4, ["", "Indicador", "Glosas", "Valor", "Qué significa"], [3, 44, 16, 18, 58])
    kpis = [
        (
            "Facturas del paquete",
            len(facturas),
            f'=SUMIFS({VGL},{CUE},"{SI}")',
            "Cuántas facturas trae el paquete y cuánto suma lo glosado, sin repetir la plata "
            "del mismo servicio.",
        ),
        (
            "Glosas a responder",
            f'=COUNTIFS({TIP},"{ITEM}")',
            f'=SUMIFS({VGL},{CUE},"{SI}",{TIP},"{ITEM}")',
            "Renglones que el gestor tiene que decidir, uno por uno.",
        ),
        (
            "Ya decididas",
            f'=COUNTIFS({TIP},"{ITEM}",{DEC},"<>")',
            "",
            "Tienen aceptar, objetar o subsanar.",
        ),
        (
            "Sin decidir",
            f'=COUNTIFS({TIP},"{ITEM}",{DEC},"")',
            "",
            "Lo que falta por trabajar. Si llega a cero, el paquete está listo.",
        ),
        (
            "Avance del paquete",
            "=IFERROR(C7/C6,0)",
            "",
            "De cada 100 glosas a responder, cuántas ya tienen decisión.",
        ),
        (
            "Se objeta (se defiende la plata)",
            f'=COUNTIF({DEC},"SE OBJETA")',
            "",
            "El hospital sostiene el cobro con su sustento técnico.",
        ),
        (
            "Se acepta",
            f'=COUNTIF({DEC},"SE ACEPTA")',
            f"=SUM({ACE})",
            "Lo que se reconoce. Es la plata que NO se va a recuperar.",
        ),
        (
            "Se subsana",
            f'=COUNTIF({DEC},"SE SUBSANA")',
            "",
            "Se responde anexando el soporte que faltaba.",
        ),
        (
            "Sigue glosado",
            "",
            "=D5-D11",
            "Lo glosado menos lo aceptado: la plata que se está defendiendo.",
        ),
        (
            "Renglones de glosa total",
            f'=COUNTIFS({TIP},"{GLOSA_TOTAL}")',
            f'=SUMIFS({VGL},{CUE},"{SI}",{TIP},"{GLOSA_TOTAL}")',
            "El ADRES glosó la reclamación entera por el FURIPS: no se responden uno por uno.",
        ),
        (
            "Glosas sin gestor asignado",
            f'=COUNTIFS({GES},"{SIN_GESTOR}",{TIP},"{ITEM}")',
            "",
            "Nadie las tiene a cargo. Si este número no es cero, hay trabajo sin dueño.",
        ),
        (
            "Falta repartir el área",
            f'=COUNTIF({ARP},"{SI}")',
            "",
            "Causales que trabajan dos áreas (la 4506) y todavía no tienen dueño.",
        ),
    ]
    for n, (et, c2, c3, nota) in enumerate(kpis):
        f = 5 + n
        _celda(rs, f, 2, et, negrita=True)
        if c2 != "":
            c = _celda(
                rs,
                f,
                3,
                c2,
                negrita=True,
                centro=True,
                relleno=GRIS,
                fmt="0.0%" if "Avance" in et else "#,##0",
            )
            c.font = Font(name=FUENTE, size=11, bold=True)
        if c3:
            _celda(rs, f, 4, c3, fmt='"$"#,##0', relleno=GRIS).alignment = Alignment(
                horizontal="right", vertical="center"
            )
        _celda(rs, f, 5, nota, ajuste=True)
        rs.row_dimensions[f].height = 30

    f = 5 + len(kpis) + 1
    _celda(rs, f, 2, "LAS TRES CAUSALES QUE MÁS PLATA NOS GLOSAN", negrita=True).font = Font(
        name=FUENTE, size=12, bold=True, color=AZUL
    )
    for n, cod in enumerate(causales[:3]):
        f += 1
        _celda(rs, f, 2, f"{cod} · {texto_causal.get(cod, '')}"[:120], negrita=True, ajuste=True)
        _celda(rs, f, 3, f"='POR QUÉ NOS GLOSAN'!D{5 + n}", fmt="#,##0", centro=True)
        _celda(rs, f, 4, f"='POR QUÉ NOS GLOSAN'!F{5 + n}", fmt='"$"#,##0')
        _celda(rs, f, 5, ACCIONES_CAUSAL.get(cod, ""), ajuste=True)
        rs.row_dimensions[f].height = 34

    base = f + 3
    for n, (etiqueta, formula) in enumerate(
        (
            ("Se objeta", f'=COUNTIF({DEC},"SE OBJETA")'),
            ("Se acepta", f'=COUNTIF({DEC},"SE ACEPTA")'),
            ("Se subsana", f'=COUNTIF({DEC},"SE SUBSANA")'),
            ("Sin decidir", f'=COUNTIFS({TIP},"{ITEM}",{DEC},"")'),
        )
    ):
        _celda(rs, base + n, 2, etiqueta)
        _celda(rs, base + n, 3, formula, fmt="#,##0")
    torta = PieChart()
    torta.title = "Qué se decidió en las glosas"
    torta.height, torta.width = 9, 13
    torta.add_data(Reference(rs, min_col=3, min_row=base, max_row=base + 3), titles_from_data=False)
    torta.set_categories(Reference(rs, min_col=2, min_row=base, max_row=base + 3))
    rs.add_chart(torta, f"B{base + 5}")

    # ── POR QUÉ NOS GLOSAN ──
    pq = wb.create_sheet("POR QUÉ NOS GLOSAN", 2)
    _titulo(
        pq,
        "POR QUÉ NOS GLOSA EL ADRES",
        "Agrupadas por la causal del reporte. El valor no repite la plata del mismo servicio "
        "glosado con varias causales; por eso una causal puede tener renglones y poco valor.",
        8,
    )
    _encabezado(
        pq,
        4,
        [
            "Causal",
            "Descripción de la causal",
            "Área que responde",
            "Renglones",
            "Facturas",
            "Valor glosado",
            "Sin decidir",
            "Qué hace falta para responderla",
        ],
        [9, 40, 16, 11, 10, 16, 11, 56],
    )
    for n, cod in enumerate(causales):
        f = 5 + n
        _celda(pq, f, 1, cod, negrita=True, centro=True)
        _celda(pq, f, 2, texto_causal.get(cod, ""), ajuste=True)
        _celda(pq, f, 3, area_causal.get(cod, ""), centro=True)
        _celda(pq, f, 4, f"=COUNTIF({CAU},$A{f})", fmt="#,##0", centro=True)
        _celda(pq, f, 5, len(facturas_causal.get(cod, ())), fmt="#,##0", centro=True)
        _celda(pq, f, 6, f'=SUMIFS({VGL},{CAU},$A{f},{CUE},"{SI}")', fmt='"$"#,##0')
        _celda(
            pq, f, 7, f'=COUNTIFS({CAU},$A{f},{TIP},"{ITEM}",{DEC},"")', fmt="#,##0", centro=True
        )
        _celda(pq, f, 8, ACCIONES_CAUSAL.get(cod, ""), ajuste=True)
        pq.row_dimensions[f].height = 32
    t = 5 + len(causales)
    _celda(pq, t, 1, "TOTAL", negrita=True, relleno=AZUL_CLARO)
    for col in (2, 3, 5, 8):
        _celda(pq, t, col, "", relleno=AZUL_CLARO)
    _celda(
        pq, t, 4, f"=SUM(D5:D{t - 1})", negrita=True, fmt="#,##0", centro=True, relleno=AZUL_CLARO
    )
    _celda(
        pq, t, 6, f'=SUMIFS({VGL},{CUE},"{SI}")', negrita=True, fmt='"$"#,##0', relleno=AZUL_CLARO
    )
    _celda(
        pq,
        t,
        7,
        f'=COUNTIFS({TIP},"{ITEM}",{DEC},"")',
        negrita=True,
        fmt="#,##0",
        centro=True,
        relleno=AZUL_CLARO,
    )
    if causales:
        tope = min(t - 1, 4 + 12)  # el gráfico se lee con las 12 primeras
        g = BarChart()
        g.type = "bar"
        g.title = "Valor glosado por causal"
        g.height, g.width = 11, 22
        g.add_data(Reference(pq, min_col=6, min_row=4, max_row=tope), titles_from_data=True)
        g.set_categories(Reference(pq, min_col=1, min_row=5, max_row=tope))
        pq.add_chart(g, f"A{t + 3}")

    # ── POR ÁREA Y CENTRO ──
    pa = wb.create_sheet("POR ÁREA Y CENTRO", 3)
    _titulo(
        pa,
        "QUIÉN TIENE QUE RESPONDER",
        "Primero por área y después por centro de costos del hospital. Sirve para repartir el "
        "trabajo y para saber a qué servicio pedirle el sustento.",
        5,
    )
    _encabezado(
        pa,
        4,
        ["Área", "Glosas a responder", "Sin decidir", "Valor glosado", "Aceptado"],
        [26, 15, 13, 17, 17],
    )
    for n, a_ in enumerate(areas):
        f = 5 + n
        _celda(pa, f, 1, a_)
        _celda(pa, f, 2, f'=COUNTIFS({ARE},$A{f},{TIP},"{ITEM}")', fmt="#,##0", centro=True)
        _celda(
            pa, f, 3, f'=COUNTIFS({ARE},$A{f},{TIP},"{ITEM}",{DEC},"")', fmt="#,##0", centro=True
        )
        _celda(pa, f, 4, f'=SUMIFS({VGL},{ARE},$A{f},{CUE},"{SI}")', fmt='"$"#,##0')
        _celda(pa, f, 5, f"=SUMIF({ARE},$A{f},{ACE})", fmt='"$"#,##0')
    fc = 5 + len(areas) + 2
    _celda(pa, fc, 1, "POR CENTRO DE COSTOS", negrita=True).font = Font(
        name=FUENTE, size=12, bold=True, color=AZUL
    )
    _encabezado(
        pa,
        fc + 1,
        ["Centro de costos", "Glosas a responder", "Sin decidir", "Valor glosado", "Aceptado"],
        [30, 15, 13, 17, 17],
    )
    for n, c_ in enumerate(centros):
        f = fc + 2 + n
        _celda(pa, f, 1, c_)
        _celda(pa, f, 2, f'=COUNTIFS({CEN},$A{f},{TIP},"{ITEM}")', fmt="#,##0", centro=True)
        _celda(
            pa, f, 3, f'=COUNTIFS({CEN},$A{f},{TIP},"{ITEM}",{DEC},"")', fmt="#,##0", centro=True
        )
        _celda(pa, f, 4, f'=SUMIFS({VGL},{CEN},$A{f},{CUE},"{SI}")', fmt='"$"#,##0')
        _celda(pa, f, 5, f"=SUMIF({CEN},$A{f},{ACE})", fmt='"$"#,##0')

    # ── POR GESTOR ──
    pg = wb.create_sheet("POR GESTOR", 4)
    _titulo(
        pg,
        "QUÉ TIENE CADA GESTOR",
        f"Fórmulas sobre la hoja {HOJA_DATOS}. «{SIN_GESTOR}» es trabajo que no tiene dueño.",
        6,
    )
    _encabezado(
        pg,
        4,
        ["Gestor", "Glosas a responder", "Ya decididas", "Sin decidir", "Avance", "Valor glosado"],
        [28, 15, 13, 13, 11, 18],
    )
    for n, g_ in enumerate(gestores):
        f = 5 + n
        c = _celda(pg, f, 1, g_)
        if g_ == SIN_GESTOR:
            c.fill = PatternFill("solid", fgColor=AMBAR)
        _celda(pg, f, 2, f'=COUNTIFS({GES},$A{f},{TIP},"{ITEM}")', fmt="#,##0", centro=True)
        _celda(
            pg, f, 3, f'=COUNTIFS({GES},$A{f},{TIP},"{ITEM}",{DEC},"<>")', fmt="#,##0", centro=True
        )
        _celda(
            pg, f, 4, f'=COUNTIFS({GES},$A{f},{TIP},"{ITEM}",{DEC},"")', fmt="#,##0", centro=True
        )
        _celda(pg, f, 5, f"=IFERROR(C{f}/B{f},0)", fmt="0.0%", centro=True)
        _celda(pg, f, 6, f'=SUMIFS({VGL},{GES},$A{f},{CUE},"{SI}")', fmt='"$"#,##0')
    tg = 5 + len(gestores)
    _celda(pg, tg, 1, "TOTAL", negrita=True, relleno=AZUL_CLARO)
    for col, letra in ((2, "B"), (3, "C"), (4, "D"), (6, "F")):
        _celda(
            pg,
            tg,
            col,
            f"=SUM({letra}5:{letra}{tg - 1})",
            negrita=True,
            centro=True,
            fmt='"$"#,##0' if col == 6 else "#,##0",
            relleno=AZUL_CLARO,
        )
    _celda(
        pg,
        tg,
        5,
        f"=IFERROR(C{tg}/B{tg},0)",
        negrita=True,
        fmt="0.0%",
        centro=True,
        relleno=AZUL_CLARO,
    )
    if gestores:
        g2 = BarChart()
        g2.type = "col"
        g2.grouping = "stacked"
        g2.overlap = 100
        g2.title = "Decididas vs. sin decidir, por gestor"
        g2.height, g2.width = 10, 22
        g2.add_data(
            Reference(pg, min_col=3, max_col=4, min_row=4, max_row=tg - 1), titles_from_data=True
        )
        g2.set_categories(Reference(pg, min_col=1, min_row=5, max_row=tg - 1))
        pg.add_chart(g2, f"A{tg + 3}")

    # ── FACTURAS ──
    fa = wb.create_sheet("FACTURAS", 5)
    _titulo(
        fa,
        "FACTURA POR FACTURA",
        "El avance de cada una y su plata. «¿Cuadra?» compara lo que suma el sistema con la "
        "cifra oficial del archivo de facturas del ADRES: si dice NO CUADRA, hay que revisarla "
        "antes de radicar.",
        17,
    )
    _encabezado(
        fa,
        4,
        [
            "Factura",
            "Radicación",
            "Paciente",
            "Gestor",
            "Médico",
            "Estado",
            "Glosas a responder",
            "Decididas",
            "Sin decidir",
            "Avance",
            "Glosa total (renglones)",
            "Valor glosado (sistema)",
            "Valor glosado (ADRES)",
            "¿Cuadra?",
            "Aceptado",
            "Sigue glosado",
            "Cerrada por",
        ],
        [15, 13, 16, 20, 20, 12, 12, 11, 11, 9, 12, 16, 16, 11, 15, 15, 24],
    )
    color_estado = {"CERRADA": VERDE, "EN PROCESO": AMBAR, "PENDIENTE": GRIS}
    for n, d in enumerate(facturas):
        f = 5 + n
        valores = [
            d["factura"],
            d["radicacion"],
            d["paciente"],
            d["gestor"],
            d["medico"],
            d["estado"],
            d["glosas"],
            d["decididas"],
            d["pendientes"],
            d["avance"],
            d["totales"],
            d["glosado"],
            d["oficial"],
            d["cuadra"],
            d["aceptado"],
            d["sigue_glosado"],
            d["cerrada_por"],
        ]
        for j, v in enumerate(valores, 1):
            c = fa.cell(f, j, v)
            c.font = Font(name=FUENTE, size=9, bold=(j == 6))
            c.alignment = Alignment(
                vertical="center", horizontal="center" if j in range(7, 15) else "general"
            )
            if j in (12, 13, 15, 16):
                c.number_format = '"$"#,##0'
            if j == 10:
                c.number_format = "0%"
            if j == 6:
                c.fill = PatternFill("solid", fgColor=color_estado.get(v, GRIS))
            if j == 14 and v == "NO CUADRA":
                c.fill = PatternFill("solid", fgColor=ROJO)
                c.font = Font(name=FUENTE, size=9, bold=True, color="9C0006")
    fa.freeze_panes = "B5"
    if facturas:
        fa.auto_filter.ref = f"A4:Q{4 + len(facturas)}"
        tf = 5 + len(facturas)
        _celda(fa, tf, 1, "TOTAL", negrita=True, relleno=AZUL_CLARO)
        for col in (2, 3, 4, 5, 6, 10, 14, 17):
            _celda(fa, tf, col, "", relleno=AZUL_CLARO)
        for col, letra in ((7, "G"), (8, "H"), (9, "I"), (11, "K")):
            _celda(
                fa,
                tf,
                col,
                f"=SUM({letra}5:{letra}{tf - 1})",
                negrita=True,
                fmt="#,##0",
                centro=True,
                relleno=AZUL_CLARO,
            )
        for col, letra in ((12, "L"), (13, "M"), (15, "O"), (16, "P")):
            _celda(
                fa,
                tf,
                col,
                f"=SUM({letra}5:{letra}{tf - 1})",
                negrita=True,
                fmt='"$"#,##0',
                relleno=AZUL_CLARO,
            )

    # Sin datos, el libro igual sale: las hojas vacías dicen que no hay nada.
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def nombre_informe(numero_paquete: str = "") -> str:
    """El nombre con el que se baja el archivo, como los del área."""
    dia = datetime.now(TZ_BOGOTA).strftime("%d-%m-%Y")
    limpio = "".join(ch for ch in _texto(numero_paquete) if ch.isalnum())
    return f"RTA_GLOSA_ADRES_{('PAQ_' + limpio + '_') if limpio else ''}{dia}.xlsx"
