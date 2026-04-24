"""Export Excel gerencial (Ronda 24).

Genera un Excel multi-hoja listo para reuniones de Comité de Cartera:

  Hoja 1 "Resumen":
    KPIs grandes del periodo (radicadas, recuperado, tasa, vencidas)
    + estado_general de salud + conteo autopilot.

  Hoja 2 "Top EPS":
    Top 10 EPS por cantidad y por valor objetado del periodo.

  Hoja 3 "Autopilot":
    Lista de glosas PENDIENTE agrupadas por estado autopilot
    (LISTA_ENVIAR → CASI_LISTA → REVISAR → INTERVENIR) con formato
    condicional por color.

  Hoja 4 "Anomalías":
    Duplicados + patrones EPS sospechosos de la ventana.

El Excel usa openpyxl puro (sin matplotlib ni charts complejos) para
mantener dependencias mínimas.
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Literal

from sqlalchemy.orm import Session

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    EXCEL_OK = True
except Exception:
    EXCEL_OK = False


# ─── Estilos reutilizables ─────────────────────────────────────────────────

def _estilos():
    return {
        "titulo": Font(size=16, bold=True, color="FFFFFF"),
        "header": Font(size=11, bold=True, color="FFFFFF"),
        "kpi_label": Font(size=10, bold=True, color="1e3a8a"),
        "kpi_value": Font(size=18, bold=True, color="1e40af"),
        "fill_titulo": PatternFill("solid", fgColor="1e40af"),
        "fill_header": PatternFill("solid", fgColor="3b82f6"),
        "fill_ok": PatternFill("solid", fgColor="dcfce7"),
        "fill_warn": PatternFill("solid", fgColor="fef3c7"),
        "fill_err": PatternFill("solid", fgColor="fee2e2"),
        "center": Alignment(horizontal="center", vertical="center"),
        "right": Alignment(horizontal="right"),
        "borde": Border(
            left=Side(style="thin", color="cbd5e1"),
            right=Side(style="thin", color="cbd5e1"),
            top=Side(style="thin", color="cbd5e1"),
            bottom=Side(style="thin", color="cbd5e1"),
        ),
    }


def _color_estado(estado: str) -> str:
    """Devuelve hex RGB para cada estado autopilot."""
    return {
        "LISTA_ENVIAR": "dcfce7",   # verde claro
        "CASI_LISTA":   "d1fae5",    # verde más claro
        "REVISAR":      "fef3c7",    # amarillo
        "INTERVENIR":   "fee2e2",    # rojo claro
    }.get(estado, "FFFFFF")


# ─── Helpers de hojas ──────────────────────────────────────────────────────

def _hoja_resumen(wb, digest: dict, salud: dict, bandeja: dict):
    s = _estilos()
    ws = wb.active
    ws.title = "Resumen"

    ws.merge_cells("A1:F1")
    ws["A1"] = "ESE HUS — RESUMEN EJECUTIVO SINAC"
    ws["A1"].font = s["titulo"]
    ws["A1"].fill = s["fill_titulo"]
    ws["A1"].alignment = s["center"]
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        f"Periodo: {digest.get('periodo', 'dia')}  ·  "
        f"Estado general: {salud.get('estado_general', 'OK')}  ·  "
        f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    ws["A2"].alignment = s["center"]
    ws.row_dimensions[2].height = 20

    ind = digest.get("indicadores", {})
    op = digest.get("operativo", {})
    kpis = [
        ("Radicadas", ind.get("radicadas", 0), ""),
        ("Respondidas", ind.get("respondidas", 0), ""),
        ("Valor objetado", ind.get("valor_objetado", 0), "$"),
        ("Valor recuperado", ind.get("valor_recuperado", 0), "$"),
        ("Tasa recuperación", f"{ind.get('tasa_recuperacion', 0) * 100:.1f}%", ""),
        ("Pendientes", op.get("pendientes_total", 0), ""),
        ("Vencidas", op.get("vencidas", 0), ""),
    ]
    fila = 4
    for idx, (label, value, prefix) in enumerate(kpis):
        col = (idx % 3) * 2 + 1   # 1, 3, 5
        row = fila + (idx // 3) * 3
        ws.cell(row=row, column=col).value = label
        ws.cell(row=row, column=col).font = s["kpi_label"]
        cell_val = ws.cell(row=row + 1, column=col)
        if isinstance(value, (int, float)) and prefix == "$":
            cell_val.value = f"${int(value):,}"
        else:
            cell_val.value = value
        cell_val.font = s["kpi_value"]

    # Conteo autopilot
    fila_auto = fila + 9
    ws.cell(row=fila_auto, column=1).value = "Autopilot"
    ws.cell(row=fila_auto, column=1).font = s["header"]
    ws.cell(row=fila_auto, column=1).fill = s["fill_header"]
    headers_auto = ["LISTA_ENVIAR", "CASI_LISTA", "REVISAR", "INTERVENIR"]
    conteo = bandeja.get("conteo_por_estado", {}) if isinstance(bandeja, dict) else {}
    for i, h in enumerate(headers_auto):
        c = ws.cell(row=fila_auto + 1, column=i + 1)
        c.value = h
        c.font = s["kpi_label"]
        c.fill = PatternFill("solid", fgColor=_color_estado(h))
        c2 = ws.cell(row=fila_auto + 2, column=i + 1)
        c2.value = conteo.get(h, 0)
        c2.font = s["kpi_value"]
        c2.alignment = s["center"]

    # Anchos
    for col_letter in ("A", "B", "C", "D", "E", "F"):
        ws.column_dimensions[col_letter].width = 22


def _hoja_top_eps(wb, digest: dict):
    s = _estilos()
    ws = wb.create_sheet("Top EPS")
    headers = ["#", "EPS", "Cantidad", "Valor objetado"]
    for i, h in enumerate(headers):
        c = ws.cell(row=1, column=i + 1)
        c.value = h
        c.font = s["header"]
        c.fill = s["fill_header"]
        c.alignment = s["center"]
        c.border = s["borde"]
    for i, e in enumerate(digest.get("top_eps", []), start=1):
        ws.cell(row=i + 1, column=1).value = i
        ws.cell(row=i + 1, column=2).value = e.get("eps", "")
        ws.cell(row=i + 1, column=3).value = e.get("cantidad", 0)
        vc = ws.cell(row=i + 1, column=4)
        vc.value = f"${int(e.get('valor_objetado', 0)):,}"
        vc.alignment = s["right"]
    for col, w in zip(("A", "B", "C", "D"), (6, 50, 12, 18)):
        ws.column_dimensions[col].width = w


def _hoja_autopilot(wb, bandeja: dict):
    s = _estilos()
    ws = wb.create_sheet("Autopilot")
    headers = ["Estado", "Glosa ID", "EPS", "Código", "Valor", "Días rest.", "Confianza"]
    for i, h in enumerate(headers):
        c = ws.cell(row=1, column=i + 1)
        c.value = h
        c.font = s["header"]
        c.fill = s["fill_header"]
        c.alignment = s["center"]
        c.border = s["borde"]
    # Ordenar por estado para que LISTA_ENVIAR vaya primero
    orden = {"LISTA_ENVIAR": 0, "CASI_LISTA": 1, "REVISAR": 2, "INTERVENIR": 3}
    glosas = sorted(
        bandeja.get("glosas", []) if isinstance(bandeja, dict) else [],
        key=lambda g: orden.get(g.get("estado_autopilot", ""), 99),
    )
    for i, g in enumerate(glosas, start=2):
        estado = g.get("estado_autopilot", "")
        ws.cell(row=i, column=1).value = estado
        ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor=_color_estado(estado))
        ws.cell(row=i, column=2).value = g.get("glosa_id")
        ws.cell(row=i, column=3).value = g.get("eps", "")
        ws.cell(row=i, column=4).value = g.get("codigo", "")
        vc = ws.cell(row=i, column=5)
        vc.value = f"${int(g.get('valor', 0) or 0):,}"
        vc.alignment = s["right"]
        ws.cell(row=i, column=6).value = g.get("dias_restantes", 0)
        ws.cell(row=i, column=7).value = f"{g.get('confianza', 0) * 100:.0f}%"
    for col, w in zip(("A", "B", "C", "D", "E", "F", "G"), (16, 10, 40, 12, 18, 10, 12)):
        ws.column_dimensions[col].width = w


def _hoja_anomalias(wb, resumen_anom: dict):
    s = _estilos()
    ws = wb.create_sheet("Anomalías")
    ws.merge_cells("A1:D1")
    ws["A1"] = (
        f"Anomalías detectadas · Alta: {resumen_anom.get('totales', {}).get('alta', 0)} · "
        f"Media: {resumen_anom.get('totales', {}).get('media', 0)}"
    )
    ws["A1"].font = s["titulo"]
    ws["A1"].fill = s["fill_titulo"]
    ws["A1"].alignment = s["center"]
    ws.row_dimensions[1].height = 24

    # Duplicados
    ws.cell(row=3, column=1).value = "DUPLICADOS"
    ws.cell(row=3, column=1).font = s["header"]
    ws.cell(row=3, column=1).fill = s["fill_header"]
    headers = ["Severidad", "Descripción", "Factura", "EPS"]
    for i, h in enumerate(headers):
        c = ws.cell(row=4, column=i + 1)
        c.value = h
        c.font = s["kpi_label"]
    fila = 5
    for a in resumen_anom.get("duplicados", []):
        ent = a.get("entidad", {})
        ws.cell(row=fila, column=1).value = a.get("severidad", "")
        sev = a.get("severidad", "")
        ws.cell(row=fila, column=1).fill = (
            s["fill_err"] if sev == "ALTA" else s["fill_warn"]
        )
        ws.cell(row=fila, column=2).value = a.get("descripcion", "")
        ws.cell(row=fila, column=3).value = ent.get("factura", "")
        ws.cell(row=fila, column=4).value = ent.get("eps", "")
        fila += 1

    # Patrones EPS
    fila += 2
    ws.cell(row=fila, column=1).value = "PATRONES EPS"
    ws.cell(row=fila, column=1).font = s["header"]
    ws.cell(row=fila, column=1).fill = s["fill_header"]
    fila += 1
    for a in resumen_anom.get("patrones_eps", []):
        sev = a.get("severidad", "")
        ws.cell(row=fila, column=1).value = sev
        ws.cell(row=fila, column=1).fill = (
            s["fill_err"] if sev == "ALTA" else s["fill_warn"]
        )
        ws.cell(row=fila, column=2).value = a.get("descripcion", "")
        fila += 1

    for col, w in zip(("A", "B", "C", "D"), (14, 60, 18, 28)):
        ws.column_dimensions[col].width = w


# ─── API pública ───────────────────────────────────────────────────────────

Periodo = Literal["dia", "semana", "mes"]


def generar_reporte_gerencial(
    db: Session,
    periodo: Periodo = "semana",
    ventana_anomalias_dias: int = 30,
) -> BytesIO:
    """Arma el Excel multi-hoja y lo devuelve como BytesIO listo para descargar."""
    if not EXCEL_OK:
        raise ImportError("openpyxl no está instalado. pip install openpyxl")

    from app.services.autopilot_service import evaluar_bandeja
    from app.services.detector_anomalias import resumen_anomalias
    from app.services.digest_ejecutivo import generar_digest
    from app.services.health_monitor import checar_salud

    salud = checar_salud(db)
    digest = generar_digest(db, periodo=periodo)
    bandeja = evaluar_bandeja(db, auditor_email=None, limite=200)
    anom = resumen_anomalias(db, ventana_dias=ventana_anomalias_dias)

    wb = Workbook()
    _hoja_resumen(wb, digest, salud, bandeja)
    _hoja_top_eps(wb, digest)
    _hoja_autopilot(wb, bandeja)
    _hoja_anomalias(wb, anom)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
