"""Centro de Automatización: correr las herramientas desde la aplicación.

EL PROBLEMA QUE RESUELVE. El repositorio tiene 34 herramientas que convierten
el archivo de un pagador al formato del ERP, consolidan ZIP, arman informes.
Todas funcionan. Ninguna se podía usar desde la aplicación: el auditor tenía
que copiar el guion a un PC con Python, abrir una consola y escribir la orden
con sus parámetros. En la práctica eso significa que las usa una persona.

Aquí cada herramienta se declara **como datos** —qué archivo pide, qué
devuelve, qué parámetros acepta— y el ejecutor la corre en un directorio
temporal, sin que nadie toque una consola.

POR QUÉ COMO DATOS Y NO COMO CÓDIGO. Agregar la herramienta número 35 no
puede exigir tocar el router ni la pantalla: se agrega una ficha a
`CATALOGO` y aparece sola en el catálogo, en la pantalla y en la API. Es la
misma idea del perfil por pagador: lo que cambia por cliente es dato, no
código.

LO QUE NO HACE. No corre nada que abra un navegador ni toque un portal —esos
bots necesitan credenciales y una sesión con la EPS, y van por el camino de
las corridas con evidencia—. Aquí solo entran las que reciben un archivo y
devuelven otro.
"""

from __future__ import annotations

import importlib
import io
import shutil
import sys
import tempfile
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable

_TOOLS = Path(__file__).resolve().parent.parent.parent / "tools"

# Tamaño máximo del archivo de entrada. Los Excel de glosas del HUS más
# grandes rondan los 6 MB (18.378 filas del Dispensario); 25 MB deja margen
# sobrado sin permitir que alguien suba un volcado de la base.
MAX_BYTES = 25 * 1024 * 1024

# Tiempo máximo de una corrida. La producción es una máquina pequeña y estas
# conversiones son de segundos; si una se cuelga, no puede dejar sin CPU a los
# gestores que están respondiendo glosas.
LIMITE_SEGUNDOS = 120


@dataclass(frozen=True)
class Parametro:
    """Un dato que la herramienta pide además del archivo."""

    nombre: str
    etiqueta: str
    ayuda: str = ""
    obligatorio: bool = False
    valor_defecto: str = ""
    tipo: str = "texto"  # texto | fecha | numero | si_no

    def como_dict(self) -> dict[str, Any]:
        return {
            "nombre": self.nombre,
            "etiqueta": self.etiqueta,
            "ayuda": self.ayuda,
            "obligatorio": self.obligatorio,
            "valor_defecto": self.valor_defecto,
            "tipo": self.tipo,
        }


@dataclass(frozen=True)
class Automatizacion:
    """Una herramienta que el auditor puede correr desde la aplicación."""

    id: str
    nombre: str
    # Qué hace, dicho para el auditor: sin nombres de archivo ni de función.
    que_hace: str
    # Cuándo la usa. Es lo que más se pregunta al ver un catálogo.
    cuando_usarla: str
    modulo: str  # módulo de `tools/` que la implementa
    extensiones: tuple[str, ...] = (".xlsx",)
    salida: str = "carpeta"  # carpeta (varios archivos → zip) | archivo
    parametros: tuple[Parametro, ...] = ()
    grupo: str = "Conversión"
    # Varias herramientas de `tools/` reciben una CARPETA y escriben el
    # resultado al lado de cada archivo que encuentran —así se usan por doble
    # clic en el PC del auditor—. Para esas, el archivo subido se deja dentro
    # de la carpeta de salida y se recoge todo lo que aparezca ahí menos el
    # original. Lo demás sería pedirles que cambien su forma de trabajar solo
    # para poder correrlas desde la aplicación.
    trabaja_en_carpeta: bool = False
    # Cómo se arma la orden. Recibe (entrada, salida, opciones) y devuelve la
    # lista de argumentos que la herramienta espera.
    argumentos: Callable[[Path, Path, dict[str, str]], list[str]] = field(repr=False, default=None)

    def como_dict(self) -> dict[str, Any]:
        return {
            "id": self.id,
            "nombre": self.nombre,
            "que_hace": self.que_hace,
            "cuando_usarla": self.cuando_usarla,
            "grupo": self.grupo,
            "extensiones": list(self.extensiones),
            "salida": self.salida,
            "parametros": [p.como_dict() for p in self.parametros],
        }


def _opt(opciones: dict[str, str], clave: str, bandera: str) -> list[str]:
    """`--bandera valor` solo si el auditor escribió algo."""
    valor = (opciones.get(clave) or "").strip()
    return [bandera, valor] if valor else []


# ─── El catálogo ──────────────────────────────────────────────────────────
# Agregar una herramienta es agregar una ficha acá. No hay que tocar el
# router ni la pantalla.

CATALOGO: tuple[Automatizacion, ...] = (
    Automatizacion(
        id="objeciones-savia",
        nombre="Glosas de SAVIA SALUD → formato del ERP",
        que_hace=(
            "Toma el Excel de glosas que manda SAVIA SALUD y lo convierte al "
            "formato de 16 columnas que acepta el ERP, con un archivo por factura."
        ),
        cuando_usarla="Cuando llega el archivo de glosas de SAVIA y hay que subirlo al ERP.",
        modulo="organizar_objeciones_savia",
        parametros=(
            Parametro(
                "fecha",
                "Fecha del documento",
                "Si se deja vacía, se usa la de hoy.",
                tipo="fecha",
            ),
            Parametro(
                "codigo_sufijo",
                "Sufijo del código de glosa",
                "SAVIA manda códigos de 4 caracteres (TA08) y el ERP pide 6 (TA0801).",
                valor_defecto="01",
            ),
        ),
        argumentos=lambda e, s, o: (
            ["--entrada", str(e), "--salida", str(s)]
            + _opt(o, "fecha", "--fecha")
            + _opt(o, "codigo_sufijo", "--codigo-sufijo")
        ),
    ),
    Automatizacion(
        id="objeciones-vco",
        nombre="Glosas de VCO → formato del ERP",
        que_hace=(
            "Convierte el consolidado de objeciones de VCO al formato de 16 columnas del ERP."
        ),
        cuando_usarla="Cuando llega el consolidado de VCO y hay que armar el cargue.",
        modulo="organizar_objeciones_vco",
        parametros=(
            Parametro("fecha_documento", "Fecha del documento", tipo="fecha"),
            Parametro("hoja", "Hoja del Excel", "Si el archivo trae varias hojas."),
        ),
        argumentos=lambda e, s, o: (
            ["--entrada", str(e), "--salida", str(s)]
            + _opt(o, "fecha_documento", "--fecha-documento")
            + _opt(o, "hoja", "--hoja")
        ),
    ),
    Automatizacion(
        id="objeciones-emssanar",
        nombre="Glosas de EMSSANAR (PDF) → formato del ERP",
        que_hace=(
            "Lee el PDF de objeciones de EMSSANAR, saca las glosas y arma el "
            "Excel de cargue del ERP."
        ),
        cuando_usarla="Cuando EMSSANAR manda las glosas en PDF en vez de Excel.",
        modulo="organizar_objeciones_emssanar",
        extensiones=(".pdf",),
        salida="archivo",
        parametros=(
            Parametro("fecha", "Fecha del documento", tipo="fecha"),
            Parametro("usuario", "Usuario del ERP", valor_defecto="999"),
        ),
        argumentos=lambda e, s, o: (
            ["--pdf", str(e), "--salida", str(s / "OBJECIONES_EMSSANAR.xlsx")]
            + _opt(o, "fecha", "--fecha")
            + _opt(o, "usuario", "--usuario")
        ),
    ),
    Automatizacion(
        id="indice-soportes",
        nombre="Índice de soportes (TXT) → Excel",
        que_hace=(
            "Convierte el índice de soportes que exporta el share —un TXT con "
            "miles de líneas— a un Excel donde se puede filtrar y buscar."
        ),
        cuando_usarla="Cuando hay que revisar qué soportes existen para un lote de facturas.",
        modulo="txt_a_excel",
        extensiones=(".txt",),
        salida="archivo",
        grupo="Soportes",
        trabaja_en_carpeta=True,
        parametros=(
            Parametro(
                "delimitador",
                "Separador de columnas",
                "Se detecta solo. Escribilo solo si el archivo sale mal partido: ; o TAB.",
            ),
        ),
        argumentos=lambda e, s, o: (
            [str(s), "--sin-recursion"] + _opt(o, "delimitador", "--delimitador")
        ),
    ),
    Automatizacion(
        id="excel-a-csv",
        nombre="Excel → CSV",
        que_hace="Convierte un Excel a CSV, que es lo que piden varios portales de EPS.",
        cuando_usarla="Cuando el portal rechaza el .xlsx y pide .csv.",
        modulo="excel_a_csv",
        salida="archivo",
        grupo="Conversión",
        trabaja_en_carpeta=True,
        parametros=(
            Parametro(
                "todas",
                "Convertir todas las hojas",
                "Escribí 1 para convertir todas; vacío convierte solo la hoja activa.",
            ),
        ),
        argumentos=lambda e, s, o: (
            [str(s), "--sin-recursion", "--forzar"]
            + (["--todas"] if (o.get("todas") or "").strip() in {"1", "si", "sí", "true"} else [])
        ),
    ),
    Automatizacion(
        id="revisar-xml",
        nombre="Revisar el XML de las facturas",
        que_hace=(
            "Abre los XML de factura electrónica y arma un Excel con lo que trae "
            "cada uno: CUFE, NIT, valor y fecha."
        ),
        cuando_usarla="Antes de radicar, para confirmar que los XML están completos y coinciden.",
        modulo="revisar_xml_facturas",
        extensiones=(".xml", ".zip"),
        salida="archivo",
        grupo="Radicación",
        trabaja_en_carpeta=True,
        argumentos=lambda e, s, o: [str(s), "--salida", str(s / "REVISION_XML.xlsx")],
    ),
    Automatizacion(
        id="tramite-masivo",
        nombre="Trámite masivo → Excel de cargue",
        que_hace="Convierte el archivo de trámites al Excel que acepta el ERP.",
        cuando_usarla="Cuando hay que cargar muchos trámites de una vez.",
        modulo="convertir_tramite_masivo",
        salida="archivo",
        grupo="Cargue",
        argumentos=lambda e, s, o: [str(e), str(s / "TRAMITE_MASIVO.xlsx")],
    ),
)

_POR_ID = {a.id: a for a in CATALOGO}


def catalogo() -> list[dict[str, Any]]:
    return [a.como_dict() for a in CATALOGO]


def obtener(id_automatizacion: str) -> Automatizacion | None:
    return _POR_ID.get(id_automatizacion)


class ErrorAutomatizacion(RuntimeError):
    """Falla previsible que se le puede explicar al auditor."""


# ─── Resumen de lo que salió ──────────────────────────────────────────────
# El archivo que produce una conversión se carga al ERP con plata adentro. La
# regla del área es no hacer un cargue masivo sin haber mirado antes qué
# salió, y hasta ahora "mirar" significaba abrir el Excel a mano.
#
# Estas funciones leen el resultado y lo resumen en tres números que el
# auditor reconoce: cuántas facturas, cuántas objeciones y cuánta plata. Si
# el total no se parece al del archivo que mandó la EPS, hay algo mal y se
# ve antes de cargarlo, no después.

# Nombres de columna de valor y de factura en los formatos que produce el
# repositorio. `CROVALOBJ` y `CRNCXC` son los del layout de 16 columnas.
_COL_VALOR = {"CROVALOBJ", "VALOR_GLOSA", "VALOR OBJETADO", "VALOR", "VALOR_OBJETADO"}
_COL_FACTURA = {"CRNCXC", "NUMERO_FACTURA", "FACTURA", "NUMERO FACTURA"}


def _resumir_excel(datos: bytes) -> dict[str, Any] | None:
    """Facturas, filas y total en pesos de un Excel generado. None si no aplica."""
    try:
        import openpyxl
    except ImportError:  # pragma: no cover - openpyxl es dependencia del repo
        return None
    try:
        wb = openpyxl.load_workbook(io.BytesIO(datos), read_only=True, data_only=True)
    except Exception:
        return None
    try:
        hoja = wb["OBJECIONES"] if "OBJECIONES" in wb.sheetnames else wb[wb.sheetnames[0]]
        filas = hoja.iter_rows(values_only=True)
        try:
            encabezados = [str(c or "").strip().upper() for c in next(filas)]
        except StopIteration:
            return None
        i_valor = next((i for i, h in enumerate(encabezados) if h in _COL_VALOR), None)
        i_fact = next((i for i, h in enumerate(encabezados) if h in _COL_FACTURA), None)
        total, cuenta, facturas = 0.0, 0, set()
        for fila in filas:
            if not fila or all(c is None or str(c).strip() == "" for c in fila):
                continue
            cuenta += 1
            if i_valor is not None and i_valor < len(fila):
                try:
                    total += float(fila[i_valor] or 0)
                except (TypeError, ValueError):
                    pass
            if i_fact is not None and i_fact < len(fila) and fila[i_fact]:
                facturas.add(str(fila[i_fact]).strip())
        return {"filas": cuenta, "facturas": len(facturas), "valor_total": round(total, 2)}
    finally:
        wb.close()


def resumir(archivos: list[tuple[str, bytes]]) -> dict[str, Any]:
    """Suma los resúmenes de todos los archivos generados."""
    detalle, filas, facturas, total = [], 0, set(), 0.0
    for nombre, datos in archivos:
        if not nombre.lower().endswith((".xlsx", ".xlsm")):
            detalle.append({"archivo": nombre, "bytes": len(datos)})
            continue
        r = _resumir_excel(datos)
        if r is None:
            detalle.append({"archivo": nombre, "bytes": len(datos)})
            continue
        filas += r["filas"]
        total += r["valor_total"]
        # Cada archivo suele ser una factura; se cuentan por archivo para no
        # perder las que el propio archivo no nombra en una columna.
        facturas.add(nombre)
        detalle.append({"archivo": nombre, "bytes": len(datos), **r})
    return {
        "archivos": len(archivos),
        "facturas": len(facturas) or None,
        "filas": filas or None,
        "valor_total": round(total, 2) if total else None,
        "detalle": detalle,
    }


@dataclass
class Resultado:
    """Lo que produjo una corrida, ya en memoria.

    Los archivos van como `(nombre, bytes)` y no como rutas: el directorio
    temporal se borra al terminar, así que devolver rutas sería devolver
    caminos a archivos que ya no existen.
    """

    ok: bool
    mensaje: str
    archivos: list[tuple[str, bytes]]
    salida_texto: str
    segundos: float

    @property
    def nombres(self) -> list[str]:
        return [n for n, _ in self.archivos]

    @property
    def bytes_totales(self) -> int:
        return sum(len(b) for _, b in self.archivos)


def _cargar(modulo: str):
    """Importa la herramienta de `tools/` sin dejar `sys.path` sucio."""
    ruta = str(_TOOLS)
    agregado = ruta not in sys.path
    if agregado:
        sys.path.insert(0, ruta)
    try:
        return importlib.import_module(modulo)
    finally:
        if agregado and ruta in sys.path:
            sys.path.remove(ruta)


def ejecutar(
    id_automatizacion: str,
    contenido: bytes,
    nombre_archivo: str,
    opciones: dict[str, str] | None = None,
) -> Resultado:
    """Corre una automatización sobre el archivo que subió el auditor.

    Todo pasa en un directorio temporal que se borra al terminar: la
    herramienta nunca ve el disco de la aplicación ni deja restos.
    """
    auto = obtener(id_automatizacion)
    if auto is None:
        raise ErrorAutomatizacion(f"No existe la automatización '{id_automatizacion}'")

    if not contenido:
        raise ErrorAutomatizacion("El archivo llegó vacío")
    if len(contenido) > MAX_BYTES:
        raise ErrorAutomatizacion(
            f"El archivo pesa {len(contenido) / 1048576:.1f} MB y el máximo son "
            f"{MAX_BYTES // 1048576} MB"
        )

    sufijo = Path(nombre_archivo or "").suffix.lower()
    if sufijo not in auto.extensiones:
        esperadas = " o ".join(auto.extensiones)
        raise ErrorAutomatizacion(
            f"Esta automatización espera un archivo {esperadas} y llegó '{sufijo or 'sin extensión'}'"
        )

    faltan = [
        p.etiqueta for p in auto.parametros if p.obligatorio and not (opciones or {}).get(p.nombre)
    ]
    if faltan:
        raise ErrorAutomatizacion("Faltan datos obligatorios: " + ", ".join(faltan))

    modulo = _cargar(auto.modulo)
    if not hasattr(modulo, "main"):
        raise ErrorAutomatizacion(f"La herramienta '{auto.modulo}' no se puede correr desde aquí")

    inicio = time.monotonic()
    trabajo = Path(tempfile.mkdtemp(prefix="sinac_auto_"))
    try:
        salida = trabajo / "salida"
        salida.mkdir()
        # Las que trabajan sobre una carpeta escriben al lado de su entrada,
        # así que la entrada va DENTRO de la carpeta de salida.
        destino = salida if auto.trabaja_en_carpeta else trabajo
        entrada = destino / (Path(nombre_archivo).name or "entrada")
        entrada.write_bytes(contenido)

        argv = auto.argumentos(entrada, salida, opciones or {})
        import contextlib
        import io

        buffer = io.StringIO()
        try:
            with contextlib.redirect_stdout(buffer), contextlib.redirect_stderr(buffer):
                codigo = modulo.main(argv)
        except SystemExit as e:  # las CLI salen con sys.exit
            codigo = e.code if isinstance(e.code, int) else 1
        except Exception as e:
            raise ErrorAutomatizacion(f"La herramienta falló: {type(e).__name__}: {e}") from e

        segundos = time.monotonic() - inicio
        if segundos > LIMITE_SEGUNDOS:
            raise ErrorAutomatizacion(
                f"La corrida tardó {segundos:.0f} s, más del límite de {LIMITE_SEGUNDOS} s"
            )

        generados = sorted(
            p for p in salida.rglob("*") if p.is_file() and p.resolve() != entrada.resolve()
        )
        if codigo not in (0, None) and not generados:
            raise ErrorAutomatizacion(
                buffer.getvalue().strip()[-500:] or "La herramienta terminó con error"
            )
        if not generados:
            raise ErrorAutomatizacion(
                "La herramienta no produjo ningún archivo. "
                + (buffer.getvalue().strip()[-300:] or "Revise que el archivo sea el correcto.")
            )

        # Se leen a memoria antes de borrar el temporal.
        leidos = [(p.name, p.read_bytes()) for p in generados]
        return Resultado(
            ok=True,
            mensaje=f"{len(leidos)} archivo(s) generado(s)",
            archivos=leidos,
            salida_texto=buffer.getvalue().strip()[-2000:],
            segundos=segundos,
        )
    finally:
        shutil.rmtree(trabajo, ignore_errors=True)
