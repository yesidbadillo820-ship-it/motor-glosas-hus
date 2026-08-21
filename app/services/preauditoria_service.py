"""Lógica de negocio de PRE-AUDITORÍA SINAC (v2 — el consolidado como base de datos).

El módulo convierte el consolidado de pre-auditoría en una base de datos web:

  1. Se alimentan dos FUENTES subiendo Excel (upsert, nunca duplican):
     - RADICACIÓN DE CUENTAS: por factura, en qué ENVÍO va + F_RECIBIDO,
       F_FACTURA, VALOR, NIT, ENTIDAD.
     - DGREPORT: por factura, si tuvo correo de factura electrónica (CORREO F.E.).
  2. Al ESCRIBIR un número de envío, el sistema crea automáticamente una fila
     por cada factura de ese envío (autocompletando desde las fuentes) y, si la
     factura ya venía DEVUELTA, la reconoce como subsanación (sin duplicarla).
  3. Cada factura es una entidad CANÓNICA (una fila) + un historial de eventos.
  4. Los datos descriptivos se resuelven por JOIN a la fuente más reciente, así
     que corregir un Excel se refleja solo (auto-sync).

Conserva del módulo v1: el semáforo del plazo (3 días hábiles) y el consecutivo
SINAC de los oficios de devolución (DEV-PRE-AUD-####-AAAA).
"""

from __future__ import annotations

import re
import threading
import unicodedata
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Optional

from app.core.logging_utils import logger
from sqlalchemy import func as sa_func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.tz import TZ_BOGOTA, a_utc, ahora_utc
from app.models.db import (
    DgReportRecord,
    EnvioCargadoRecord,
    FacturaEventoRecord,
    FacturaPreauditoriaRecord,
    OficioDevolucionRecord,
    OficioRecepcionRecord,
    RadicacionCuentaRecord,
)

# Regla del proceso: una misma factura se acepta devuelta máximo 3 veces.
MAX_DEVOLUCIONES = 3
# Un mismo envío puede cargarse en hasta 3 oficios distintos: el original y
# las recargas de subsanación (facturación reenvía las devueltas con el
# MISMO número de envío dentro de un oficio nuevo — caso real 30-07-2026).
MAX_OFICIOS_POR_ENVIO = 3

# Plazo para auditar un oficio: 3 días hábiles (lunes a viernes),
# contados a partir del día siguiente al recibo.
PLAZO_DIAS_HABILES = 3

# resultado_actual
RESULTADO_PENDIENTE = "PENDIENTE"
RESULTADO_RADICAR = "RADICAR"
RESULTADO_DEVUELTA = "DEVUELTA"

# estado (máquina de estado que ve el auditor)
ESTADO_NUEVA = "NUEVA"
ESTADO_RADICADA = "RADICADA"
ESTADO_DEVUELTA_PEND = "DEVUELTA_PEND_SUBSANACION"
ESTADO_EN_SUBSANACION = "EN_SUBSANACION"
ESTADO_SUBSANADA = "SUBSANADA"
ESTADO_NUEV_DEVUELTA = "NUEVAMENTE_DEVUELTA"
ESTADO_BLOQUEADA = "BLOQUEADA_LIMITE"

PREFIJO_CONSECUTIVO = "DEV-PRE-AUD"


# ==================================================================
# Semáforo del plazo (3 días hábiles desde el día siguiente al recibo)
# ==================================================================


def _es_habil(d: date) -> bool:
    return d.weekday() < 5  # lunes=0 ... viernes=4


def _sumar_dias_habiles(desde: date, n: int) -> date:
    """El n-ésimo día hábil DESPUÉS de `desde` (sin contar `desde`)."""
    d = desde
    restantes = n
    while restantes > 0:
        d += timedelta(days=1)
        if _es_habil(d):
            restantes -= 1
    return d


def _dias_habiles_entre(desde: date, hasta: date) -> int:
    """Días hábiles de `desde` a `hasta`, ambos inclusive (0 si desde>hasta)."""
    if desde > hasta:
        return 0
    total = 0
    d = desde
    while d <= hasta:
        if _es_habil(d):
            total += 1
        d += timedelta(days=1)
    return total


def calcular_semaforo(
    fecha_recibido: datetime,
    hoy: Optional[date] = None,
    completado: bool = False,
) -> dict:
    """Semáforo del plazo de un oficio.

    Estados: COMPLETADO (todo auditado), VERDE (va bien), AMARILLO
    (penúltimo día), ROJO (último día) y VENCIDO (se pasó el plazo).
    """
    fecha_recibido = a_utc(fecha_recibido)
    recibido_local = fecha_recibido.astimezone(TZ_BOGOTA).date()
    if hoy is None:
        hoy = datetime.now(TZ_BOGOTA).date()
    fecha_limite = _sumar_dias_habiles(recibido_local, PLAZO_DIAS_HABILES)
    dias_restantes = _dias_habiles_entre(hoy, fecha_limite)

    if completado:
        estado = "COMPLETADO"
    elif hoy > fecha_limite:
        estado = "VENCIDO"
    elif dias_restantes >= 3:
        estado = "VERDE"
    elif dias_restantes == 2:
        estado = "AMARILLO"
    else:
        estado = "ROJO"

    return {
        "estado": estado,
        "fecha_limite": fecha_limite.isoformat(),
        "dias_habiles_restantes": dias_restantes if hoy <= fecha_limite else 0,
        "dias_vencido": (hoy - fecha_limite).days if hoy > fecha_limite else 0,
    }


# ==================================================================
# Consecutivo SINAC de los oficios de devolución
# ==================================================================


def siguiente_consecutivo(db: Session, anio: Optional[int] = None) -> tuple[str, int, int]:
    """Siguiente consecutivo DEV-PRE-AUD-####-AAAA (reinicia cada año).

    Es solo una SUGERENCIA: el consecutivo real lo lleva SINAC por fuera del
    sistema, así que el auditor puede escribir el que corresponda.
    """
    if anio is None:
        anio = datetime.now(TZ_BOGOTA).year
    ultimo = (
        db.query(sa_func.max(OficioDevolucionRecord.numero))
        .filter(OficioDevolucionRecord.anio == anio)
        .scalar()
    )
    numero = int(ultimo or 0) + 1
    return f"{PREFIJO_CONSECUTIVO}-{numero:04d}-{anio}", numero, anio


def interpretar_consecutivo(texto: str, anio_actual: Optional[int] = None) -> tuple[str, int, int]:
    """Interpreta el consecutivo que escribe el auditor.

    La numeración de los oficios de devolución la lleva SINAC internamente, así
    que el auditor escribe el que va. Se acepta:

      - solo el número          → "89"                    → DEV-PRE-AUD-0089-2026
      - el consecutivo completo → "DEV-PRE-AUD-0089-2026" → se respeta
      - cualquier otro texto    → se respeta tal cual (numeración especial)

    Devuelve (consecutivo, numero, anio). Para el texto libre, `numero` viene
    en 0: quien llama debe asignarle uno interno que no choque (ese campo solo
    sirve para ordenar y sugerir el siguiente, no se le muestra a nadie).
    """
    if anio_actual is None:
        anio_actual = datetime.now(TZ_BOGOTA).year
    limpio = " ".join((texto or "").split()).upper()
    if not limpio:
        raise ValueError("Escriba el consecutivo del oficio de devolución.")
    if len(limpio) > 60:
        raise ValueError("El consecutivo es demasiado largo (máx. 60 caracteres).")

    solo_numero = re.fullmatch(r"(\d{1,6})", limpio)
    if solo_numero:
        numero = int(solo_numero.group(1))
        return f"{PREFIJO_CONSECUTIVO}-{numero:04d}-{anio_actual}", numero, anio_actual

    completo = re.fullmatch(rf"{PREFIJO_CONSECUTIVO}-(\d{{1,6}})-(\d{{4}})", limpio)
    if completo:
        numero, anio = int(completo.group(1)), int(completo.group(2))
        return f"{PREFIJO_CONSECUTIVO}-{numero:04d}-{anio}", numero, anio

    return limpio, 0, anio_actual


def reservar_consecutivo(db: Session, texto: Optional[str]) -> tuple[str, int, int]:
    """Deja listo el consecutivo a usar, validando que no esté repetido.

    Sin texto, toma el siguiente sugerido. Con texto, respeta lo que escribió
    el auditor. Lanza ValueError con un mensaje claro si ya se usó.
    """
    if not (texto or "").strip():
        return siguiente_consecutivo(db)

    consecutivo, numero, anio = interpretar_consecutivo(texto)
    repetido = (
        db.query(OficioDevolucionRecord)
        .filter(OficioDevolucionRecord.consecutivo == consecutivo)
        .first()
    )
    if repetido:
        raise ValueError(f"El consecutivo {consecutivo} ya fue usado en otro oficio de devolución.")

    # `numero` solo ordena y alimenta la sugerencia siguiente; si el que se
    # dedujo ya está ocupado (o es texto libre), se toma el primero libre.
    ocupado = (
        db.query(OficioDevolucionRecord.id)
        .filter(
            OficioDevolucionRecord.anio == anio,
            OficioDevolucionRecord.numero == numero,
        )
        .first()
    )
    if numero <= 0 or ocupado:
        ultimo = (
            db.query(sa_func.max(OficioDevolucionRecord.numero))
            .filter(OficioDevolucionRecord.anio == anio)
            .scalar()
        )
        numero = int(ultimo or 0) + 1
    return consecutivo, numero, anio


# ==================================================================
# Parseo de los Excel fuente
# ==================================================================


def _normalizar_encabezado(valor) -> str:
    if valor is None:
        return ""
    texto = str(valor)
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^A-Za-z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip().upper()


# Alias de columnas (tras normalizar). Cubre el reporte real de DGH
# "Radicación de Cuentas" y también los encabezados del CONSOLIDADO histórico.
_ALIAS_RADICACION = {
    "envio": {
        "RADICACION CONSECUTIVO",
        "CONSECUTIVO",
        "ENVIO",
        "NO ENVIO",
        "NUMERO ENVIO",
        "NRO ENVIO",
    },
    "factura": {"CXC FACTURA", "FACTURA", "NO FACTURA", "NUMERO FACTURA", "DOCUMENTO"},
    "f_recibido": {
        "RADICACION FECHADOCUMENTO",
        "FECHADOCUMENTO",
        "FECHA DOCUMENTO",
        "F RECIBIDO",
        "FECHA RECIBIDO",
    },
    "f_factura": {"CXC FECHA", "F FACTURA", "FECHA FACTURA"},
    "valor": {"CXC VALOR", "VALOR", "VR FACTURA", "VALOR FACTURA", "VLR FACTURA"},
    "nit": {"RADICACION TERCERO DOCUMENTO", "TERCERO DOCUMENTO", "NIT", "NIT ENTIDAD"},
    "entidad": {
        "RADICACION TERCERO NOMBRECOMPLETONA",
        "TERCERO NOMBRECOMPLETONA",
        "NOMBRECOMPLETONA",
        "ENTIDAD",
        "EMPRESA",
        "EPS",
        "ASEGURADORA",
        "CLIENTE",
    },
    "estado_radicacion": {"RADICACION ESTADOACTUAL", "ESTADOACTUAL", "ESTADO"},
}

_ALIAS_DGREPORT = {
    "factura": {"NUMERO DE FACTURA", "FACTURA", "CXC FACTURA", "NO FACTURA"},
    "fecha_correo": {"FECHA DE ENVIO CORREO", "FECHA CORREO", "FECHA DE ENVIO FACTURA DIAN"},
    "numero_fe": {"CUFE", "NUMERO FE", "NUMERO DE FE", "NUMERO FACTURA ELECTRONICA"},
}


def _mapear_columnas(fila_encabezado, alias: dict) -> dict:
    mapa = {}
    for idx, celda in enumerate(fila_encabezado):
        nombre = _normalizar_encabezado(celda)
        if not nombre:
            continue
        for campo, nombres in alias.items():
            if campo not in mapa and nombre in nombres:
                mapa[campo] = idx
    return mapa


def _como_fecha(valor) -> Optional[datetime]:
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, date):
        return datetime(valor.year, valor.month, valor.day)
    texto = str(valor).strip()
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%Y-%m-%dT%H:%M:%S",
    ):
        try:
            return datetime.strptime(texto[:19], fmt)
        except ValueError:
            continue
    return None


def _como_valor(valor) -> float:
    if valor is None or valor == "":
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = re.sub(r"[^0-9,.\-]", "", str(valor))
    # Formato colombiano: punto de miles, coma decimal.
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return 0.0


def _texto(valor) -> Optional[str]:
    if valor is None:
        return None
    t = str(valor).strip()
    return t or None


def _misma_fecha(a, b) -> bool:
    """Compara solo la parte DÍA de dos fechas (las de la fuente son fechas,
    no timestamps con zona). Robusto a que la BD las devuelva aware y el parser
    naive: evita marcar 'actualizada' toda la fuente en cada re-subida."""
    da = a.date() if isinstance(a, datetime) else a
    db_ = b.date() if isinstance(b, datetime) else b
    return da == db_


_LOCK_OPENPYXL = threading.Lock()


def _abrir_libro(contenido: bytes):
    """Abre el Excel en modo lectura, sin el barrido previo de dimensiones.

    Los reportes de Dinámica Gerencial no declaran `<dimension>` en su XML.
    Cuando ese dato falta, openpyxl recorre el archivo ENTERO solo para
    averiguar su tamaño y después lo vuelve a recorrer para leerlo: con el
    reporte real de Radicación (191.859 filas) son 12 segundos perdidos de 36.
    Aquí solo se recorre con `iter_rows` y nunca se consultan `ws.max_row` ni
    `ws.max_column`, así que ese primer barrido no aporta nada.

    El ajuste es sobre la clase de openpyxl; se toma un candado y se restaura
    de inmediato, de modo que solo cubre la apertura de este libro y no afecta
    a otros cargues del sistema que sí usen las dimensiones.
    """
    from openpyxl import load_workbook  # import perezoso
    from openpyxl.worksheet._read_only import ReadOnlyWorksheet

    with _LOCK_OPENPYXL:
        original = ReadOnlyWorksheet._get_size
        ReadOnlyWorksheet._get_size = lambda self: None
        try:
            return load_workbook(BytesIO(contenido), data_only=True, read_only=True)
        finally:
            ReadOnlyWorksheet._get_size = original


# Valores de la columna FACTURA que en realidad son encabezados o totales
# repetidos dentro del reporte: se saltan.
_FACTURA_NO_ES_FACTURA = {"FACTURA", "TOTAL"}


def _es_fila_de_encabezado(factura: str) -> bool:
    """¿El valor de la columna FACTURA es en realidad un encabezado/total?

    Se evita normalizar (NFKD + expresiones regulares) en cada una de las
    191.859 filas: un número de factura real es ASCII alfanumérico y no cambia
    al normalizarse, así que basta compararlo directo.
    """
    if factura.isascii() and factura.isalnum():
        return factura in _FACTURA_NO_ES_FACTURA
    return _normalizar_encabezado(factura) in _FACTURA_NO_ES_FACTURA


def _hallar_encabezado(ws, alias: dict, requeridos: set, max_filas: int = 15):
    """Busca la fila de encabezados (primeras `max_filas`) y devuelve
    (numero_fila, mapa) o (None, None) si no aparece."""
    for numero_fila, fila in enumerate(ws.iter_rows(values_only=True), start=1):
        if numero_fila > max_filas:
            break
        mapa = _mapear_columnas(fila, alias)
        if requeridos <= set(mapa):
            return numero_fila, mapa
    return None, None


def parsear_excel_radicacion(contenido: bytes) -> dict:
    """Lee el Excel de RADICACIÓN DE CUENTAS.

    Devuelve una fila por factura (deduplicada; se descartan las radicaciones
    'Anulado' y, si una factura tiene varias radicaciones válidas, se conserva
    la de mayor F_RECIBIDO). {"facturas": [...], "advertencias": [...]}.
    """
    wb = _abrir_libro(contenido)
    por_factura: dict[str, dict] = {}
    advertencias: list[str] = []
    anuladas = 0
    leidas = 0
    _repetidos: dict = {}
    # Los estados se repiten (unos pocos valores distintos en todo el archivo):
    # se normaliza cada uno una sola vez, no una vez por fila.
    _anulado: dict = {}

    def _compartido(valor):
        """ENTIDAD, NIT, ENVÍO y ESTADO repiten el mismo texto en miles de
        filas. Se guarda una sola copia y todas las filas la comparten, en vez
        de una copia por fila (importa en archivos de decenas de miles)."""
        if valor is None:
            return None
        return _repetidos.setdefault(valor, valor)

    for ws in wb.worksheets:
        fila_hdr, mapa = _hallar_encabezado(ws, _ALIAS_RADICACION, {"factura", "envio"})
        if mapa is None:
            continue
        for numero_fila, fila in enumerate(ws.iter_rows(values_only=True), start=1):
            if numero_fila <= fila_hdr:
                continue

            def _celda(campo):
                idx = mapa.get(campo)
                if idx is None or idx >= len(fila):
                    return None
                return fila[idx]

            factura = (str(_celda("factura") or "").strip() or "").upper()
            if not factura or _es_fila_de_encabezado(factura):
                continue
            leidas += 1
            estado = _texto(_celda("estado_radicacion"))
            if estado is not None:
                es_anulado = _anulado.get(estado)
                if es_anulado is None:
                    es_anulado = _normalizar_encabezado(estado).startswith("ANULAD")
                    _anulado[estado] = es_anulado
                if es_anulado:
                    anuladas += 1
                    continue
            f_recibido = _como_fecha(_celda("f_recibido"))
            registro = {
                "factura": factura,
                "envio": _compartido(str(_celda("envio") or "").strip() or "SIN ENVÍO"),
                "f_recibido": f_recibido,
                "f_factura": _como_fecha(_celda("f_factura")),
                "valor": _como_valor(_celda("valor")),
                "nit": _compartido(_texto(_celda("nit"))),
                # ENTIDAD trae espacios al final en la fuente real: strip.
                "entidad": _compartido(_texto(_celda("entidad"))),
                "estado_radicacion": _compartido(estado),
            }
            previo = por_factura.get(factura)
            if previo is None:
                por_factura[factura] = registro
            else:
                # Conservar la de mayor F_RECIBIDO (re-radicación más reciente).
                pf = previo.get("f_recibido")
                if f_recibido and (pf is None or f_recibido > pf):
                    por_factura[factura] = registro
                advertencias.append(
                    f"Factura {factura} con varias radicaciones válidas: se tomó la más reciente"
                )
        if por_factura:
            break  # primera hoja con datos

    wb.close()
    if anuladas:
        advertencias.append(f"{anuladas} radicación(es) 'Anulado' descartada(s)")
    if not por_factura:
        advertencias.append(
            "No se encontraron facturas: revise que el archivo sea el reporte de "
            "Radicación de Cuentas (columnas Radicacion.Consecutivo y CxC.Factura)."
        )
    return {"facturas": list(por_factura.values()), "advertencias": advertencias, "leidas": leidas}


def parsear_excel_dgreport(contenido: bytes) -> dict:
    """Lee el Excel de DGREPORT. Devuelve una fila por factura (con correo F.E.)."""
    wb = _abrir_libro(contenido)
    por_factura: dict[str, dict] = {}
    advertencias: list[str] = []
    leidas = 0

    for ws in wb.worksheets:
        fila_hdr, mapa = _hallar_encabezado(ws, _ALIAS_DGREPORT, {"factura"})
        if mapa is None:
            continue
        for numero_fila, fila in enumerate(ws.iter_rows(values_only=True), start=1):
            if numero_fila <= fila_hdr:
                continue

            def _celda(campo):
                idx = mapa.get(campo)
                if idx is None or idx >= len(fila):
                    return None
                return fila[idx]

            factura = (str(_celda("factura") or "").strip() or "").upper()
            if not factura or _normalizar_encabezado(factura) in {"FACTURA", "TOTAL"}:
                continue
            leidas += 1
            if factura in por_factura:
                continue
            por_factura[factura] = {
                "factura": factura,
                "correo_fe": "SI",
                "fecha_correo": _como_fecha(_celda("fecha_correo")),
                "numero_fe": _texto(_celda("numero_fe")),
            }
        if por_factura:
            break

    wb.close()
    if not por_factura:
        advertencias.append(
            "No se encontraron facturas: revise que el archivo sea el DGReport "
            "(columna 'NUMERO DE FACTURA')."
        )
    return {"facturas": list(por_factura.values()), "advertencias": advertencias, "leidas": leidas}


# ==================================================================
# Upsert de las fuentes (nunca duplican; la última carga es la verdad)
# ==================================================================


# Cuántas facturas se guardan por bloque en un cargue masivo.
# El servidor del hospital tiene 1 GB de memoria: guardar las 36.723 filas de
# un solo golpe consumía +154 MB y el sistema mataba la aplicación a mitad del
# cargue (el auditor veía "Error 524"). Guardando por bloques el consumo se
# mantiene plano sin importar el tamaño del archivo.
TAM_BLOQUE_UPSERT = 2000


def _guardar_bloque(db: Session, tocados) -> None:
    """Confirma el bloque y suelta sus filas de la memoria.

    Solo se sueltan las filas que tocó ESTE bloque: no se toca nada más que
    la petición pueda estar usando (por ejemplo el usuario conectado).
    """
    db.commit()
    for obj in tocados:
        try:
            if obj in db:
                db.expunge(obj)
        except Exception:  # pragma: no cover - defensivo
            pass


def _upsert_por_bloques(db: Session, filas: list[dict], aplicar) -> dict:
    """Recorre las filas en bloques; cada bloque se guarda y se libera.

    Si el proceso se interrumpe a mitad de un cargue (corte, falta de memoria),
    lo ya guardado NO se pierde: al volver a subir el mismo archivo el cargue
    retoma donde quedó, porque el upsert nunca duplica.
    """
    totales = {"nuevas": 0, "actualizadas": 0, "sin_cambio": 0}
    procesadas = 0
    for i in range(0, len(filas), TAM_BLOQUE_UPSERT):
        bloque = filas[i : i + TAM_BLOQUE_UPSERT]
        try:
            parciales = aplicar(bloque)
        except Exception as e:
            # 20-08-2026. Antes la excepción salía derecho y el auditor veía un
            # error a secas, como si no se hubiera guardado NADA. Pero los
            # bloques anteriores YA están confirmados en la base: creerlos
            # perdidos lleva a rehacer trabajo que no hace falta, o peor, a
            # dudar de lo que sí quedó.
            #
            # No se traga el error: se devuelve junto con lo que sí entró, y en
            # qué fila se quedó. Volver a subir el mismo archivo retoma donde
            # quedó, porque el upsert nunca duplica.
            db.rollback()
            logger.warning(
                f"[preauditoria] el cargue se cortó en la fila {procesadas + 1} "
                f"de {len(filas)}: {type(e).__name__}: {e}"
            )
            totales["error"] = (
                f"El cargue se cortó en la fila {procesadas + 1} de {len(filas)}. "
                f"Las {procesadas} anteriores SÍ quedaron guardadas. "
                "Vuelva a subir el mismo archivo: retoma donde quedó y no duplica nada."
            )
            totales["filas_procesadas"] = procesadas
            totales["completo"] = False
            return totales
        for clave in ("nuevas", "actualizadas", "sin_cambio"):
            totales[clave] += parciales[clave]
        procesadas += len(bloque)
    totales["filas_procesadas"] = procesadas
    totales["completo"] = True
    return totales


def _indice_existentes(db: Session, modelo, numeros: list[str]) -> dict:
    """Pre-carga las filas ya guardadas de esas facturas (evita N consultas)."""
    indice = {}
    for i in range(0, len(numeros), 900):
        for row in db.query(modelo).filter(modelo.factura.in_(numeros[i : i + 900])):
            indice[row.factura] = row
    return indice


def upsert_radicacion(db: Session, filas: list[dict], archivo: str, usuario: str) -> dict:
    """Inserta/actualiza RADICACIÓN por factura. Reporta nuevas/actualizadas/sin cambio."""
    ahora = ahora_utc()

    def aplicar(bloque: list[dict]) -> dict:
        nuevas = actualizadas = sin_cambio = 0
        # Facturas que VINIERON en este Excel y no cambiaron ningún dato: hay
        # que sellarles igual la fecha. `actualizado_en` no significa "se le
        # cambió algo", significa "el último cargue del Excel la traía" — es
        # lo que permite reconocer después las filas que el Excel ya NO trae
        # porque facturación las sacó del envío (ver facturas_de_fuente_rezagada).
        vistas_sin_cambio: list[int] = []
        indice = _indice_existentes(db, RadicacionCuentaRecord, [r["factura"] for r in bloque])
        for r in bloque:
            existente = indice.get(r["factura"])
            if existente is None:
                nuevo = RadicacionCuentaRecord(
                    factura=r["factura"],
                    envio=r["envio"],
                    f_recibido=r["f_recibido"],
                    f_factura=r["f_factura"],
                    valor=r["valor"],
                    nit=r["nit"],
                    entidad=r["entidad"],
                    estado_radicacion=r.get("estado_radicacion"),
                    fuente_archivo=archivo,
                    importado_por=usuario,
                )
                db.add(nuevo)
                indice[r["factura"]] = nuevo  # misma factura repetida en el lote
                nuevas += 1
            else:
                cambio = (
                    existente.envio != r["envio"]
                    or existente.valor != r["valor"]
                    or existente.nit != r["nit"]
                    or (existente.entidad or None) != r["entidad"]
                    or not _misma_fecha(existente.f_recibido, r["f_recibido"])
                    or not _misma_fecha(existente.f_factura, r["f_factura"])
                )
                if cambio:
                    existente.envio = r["envio"]
                    existente.f_recibido = r["f_recibido"]
                    existente.f_factura = r["f_factura"]
                    existente.valor = r["valor"]
                    existente.nit = r["nit"]
                    existente.entidad = r["entidad"]
                    existente.estado_radicacion = r.get("estado_radicacion")
                    existente.fuente_archivo = archivo
                    existente.importado_por = usuario
                    existente.actualizado_en = ahora
                    actualizadas += 1
                else:
                    sin_cambio += 1
                    if existente.id is not None:
                        vistas_sin_cambio.append(existente.id)
        # Una sola sentencia por tanda: 22.000 filas no se pueden sellar una
        # por una sin volver lento el cargue.
        for i in range(0, len(vistas_sin_cambio), 900):
            db.query(RadicacionCuentaRecord).filter(
                RadicacionCuentaRecord.id.in_(vistas_sin_cambio[i : i + 900])
            ).update({"actualizado_en": ahora}, synchronize_session=False)
        _guardar_bloque(db, indice.values())
        return {"nuevas": nuevas, "actualizadas": actualizadas, "sin_cambio": sin_cambio}

    return _upsert_por_bloques(db, filas, aplicar)


def upsert_dgreport(db: Session, filas: list[dict], archivo: str, usuario: str) -> dict:
    """Inserta/actualiza DGREPORT por factura."""
    ahora = ahora_utc()

    def aplicar(bloque: list[dict]) -> dict:
        nuevas = actualizadas = sin_cambio = 0
        indice = _indice_existentes(db, DgReportRecord, [r["factura"] for r in bloque])
        for r in bloque:
            existente = indice.get(r["factura"])
            if existente is None:
                nuevo = DgReportRecord(
                    factura=r["factura"],
                    correo_fe=r.get("correo_fe", "SI"),
                    fecha_correo=r.get("fecha_correo"),
                    numero_fe=r.get("numero_fe"),
                    fuente_archivo=archivo,
                    importado_por=usuario,
                )
                db.add(nuevo)
                indice[r["factura"]] = nuevo
                nuevas += 1
            else:
                cambio = existente.correo_fe != r.get(
                    "correo_fe", "SI"
                ) or existente.numero_fe != r.get("numero_fe")
                if cambio:
                    existente.correo_fe = r.get("correo_fe", "SI")
                    existente.fecha_correo = r.get("fecha_correo")
                    existente.numero_fe = r.get("numero_fe")
                    existente.fuente_archivo = archivo
                    existente.importado_por = usuario
                    existente.actualizado_en = ahora
                    actualizadas += 1
                else:
                    sin_cambio += 1
        _guardar_bloque(db, indice.values())
        return {"nuevas": nuevas, "actualizadas": actualizadas, "sin_cambio": sin_cambio}

    return _upsert_por_bloques(db, filas, aplicar)


# ==================================================================
# Cruce con las fuentes (auto-sync por lectura)
# ==================================================================


def datos_fuente(db: Session, factura: str) -> dict:
    """Resuelve los datos descriptivos de una factura desde las fuentes."""
    rad = (
        db.query(RadicacionCuentaRecord)
        .filter(RadicacionCuentaRecord.factura == factura)
        .one_or_none()
    )
    dg = db.query(DgReportRecord).filter(DgReportRecord.factura == factura).one_or_none()
    return {
        "envio": rad.envio if rad else None,
        "f_recibido": rad.f_recibido if rad else None,
        "f_factura": rad.f_factura if rad else None,
        "valor": rad.valor if rad else 0.0,
        "nit": rad.nit if rad else None,
        "entidad": rad.entidad if rad else None,
        "correo_fe": (dg.correo_fe if dg else "NO"),
    }


def facturas_de_envio(db: Session, envio: str) -> list[RadicacionCuentaRecord]:
    return (
        db.query(RadicacionCuentaRecord)
        .filter(RadicacionCuentaRecord.envio == str(envio).strip())
        .order_by(RadicacionCuentaRecord.factura)
        .all()
    )


# Dos filas de la fuente que se tocaron con menos de esta diferencia se toman
# como parte del MISMO cargue del Excel de Radicación (un cargue grande tarda
# minutos, no horas).
MARGEN_MISMO_CARGUE = timedelta(hours=6)


def facturas_de_fuente_rezagada(src: list) -> list[dict]:
    """Facturas del envío que quedaron de un cargue ANTERIOR de la Radicación.

    POR QUÉ EXISTE (caso real 19-08-2026, envío 232047): la fuente de
    Radicación se actualiza factura por factura (upsert). Cuando facturación
    saca una factura de un envío, esa factura simplemente deja de venir en el
    Excel nuevo — pero su fila vieja se queda en el sistema con el envío
    anterior, y al escribir el envío la arrastra como si todavía perteneciera.
    En ese caso el sistema traía 13 facturas donde el DGH mostraba 12.

    No se puede borrar sola (el Excel puede ser parcial y borraríamos filas
    buenas), pero SÍ se puede señalar: si el resto del envío se actualizó en
    el último cargue y una fila quedó con fecha vieja, esa es la sospechosa.
    """
    fechas = [a_utc(r.actualizado_en or r.importado_en) for r in src]
    conocidas = [f for f in fechas if f is not None]
    if len(conocidas) < 2:
        return []
    ultima = max(conocidas)
    corte = ultima - MARGEN_MISMO_CARGUE
    rezagadas = []
    for r, f in zip(src, fechas):
        if f is None or f >= corte:
            continue
        rezagadas.append(
            {
                "factura": r.factura,
                "valor": r.valor,
                "actualizada_en": f.astimezone(TZ_BOGOTA).strftime("%d/%m/%Y"),
                "resto_actualizado_en": ultima.astimezone(TZ_BOGOTA).strftime("%d/%m/%Y"),
                "fuente_archivo": r.fuente_archivo,
            }
        )
    return rezagadas


def _avisos_de_fuente_rezagada(src: list) -> list[str]:
    """Los mismos hallazgos, redactados para que el auditor los entienda."""
    return [
        f"{r['factura']} quedó de un cargue anterior de la Radicación "
        f"(actualizada el {r['actualizada_en']}); el resto del envío se actualizó el "
        f"{r['resto_actualizado_en']}. Puede que facturación ya la haya sacado de este "
        "envío: verifíquela en el DGH antes de auditarla."
        for r in facturas_de_fuente_rezagada(src)
    ]


# ==================================================================
# ¿Por qué el oficio muestra más envíos que facturas?
# ==================================================================


def _donde_quedo(factura: str, ubicada, oficio_id: int) -> dict:
    """Explica, en una frase, por qué una factura del envío no está en el oficio."""
    if ubicada is None:
        return {
            "factura": factura,
            "oficio": None,
            "estado": None,
            "motivo": (
                "la fuente la trae en este envío pero no quedó registrada en el "
                "sistema: vuelva a escribir el envío para que entre."
            ),
        }
    canon, radicado = ubicada
    donde = radicado or canon.oficio_fhus or "otro oficio"
    if canon.oficio_actual_id == oficio_id:
        motivo = (
            f"sí está en este oficio, pero entró con el envío {canon.envio_actual}: "
            "la fuente la cambió de envío."
        )
    elif canon.resultado_actual == RESULTADO_PENDIENTE:
        motivo = (
            f"sigue pendiente de auditar en el oficio {donde}: resuélvala allá "
            "(radicar o devolver) y después vuelva a escribir el envío aquí."
        )
    elif canon.resultado_actual == RESULTADO_RADICAR:
        motivo = f"ya estaba radicada en el oficio {donde}: no vuelve a entrar."
    elif canon.resultado_actual == RESULTADO_DEVUELTA:
        motivo = (
            f"está devuelta en el oficio {donde}: vuelve a entrar cuando la entidad "
            "la reingrese y usted escriba otra vez el envío."
        )
    else:
        motivo = f"quedó en el oficio {donde}."
    return {
        "factura": factura,
        "oficio": radicado or canon.oficio_fhus,
        "estado": canon.estado,
        "motivo": motivo,
    }


def faltantes_por_envio(db: Session, oficio_id: int, facturas: list) -> dict:
    """Por cada envío escrito en el oficio, qué facturas suyas NO están aquí.

    POR QUÉ EXISTE (caso real 20-08-2026, oficio FHUS-AS-I01190-26): la
    pantalla mostraba 5 envíos cargados y solo 4 facturas. El chip del envío
    dice cuántas facturas traía la fuente, pero el sistema OMITE las que ya
    venían abiertas (o ya radicadas) en otro oficio — esas nunca entran aquí.
    El aviso salía una sola vez, al cargar, y después no quedaba ni rastro: el
    auditor contaba 5 y 4 sin ninguna explicación en pantalla.

    Recibe las facturas que YA están en el oficio (para no consultarlas dos
    veces) y devuelve {envio: [ {factura, oficio, estado, motivo}, ... ]} solo
    con los envíos que dejaron algún faltante.
    """
    cargados = db.query(EnvioCargadoRecord).filter(EnvioCargadoRecord.oficio_id == oficio_id).all()
    if not cargados:
        return {}
    aqui: dict = {}
    for f in facturas:
        aqui.setdefault(f.envio_actual, set()).add(f.factura)
    # Solo se revisan los envíos que PROMETIERON más facturas de las que hay:
    # si el auditor quitó una con la 🗑 el registro ya bajó su cuenta y el
    # sistema no debe insistirle en volver a subirla.
    prometidas: dict = {}
    for c in cargados:
        if c.envio:
            prometidas[c.envio] = prometidas.get(c.envio, 0) + (c.total_facturas or 0)
    envios = sorted(e for e, n in prometidas.items() if len(aqui.get(e, ())) < n)
    if not envios:
        return {}
    en_fuente: dict = {}
    for envio, factura in (
        db.query(RadicacionCuentaRecord.envio, RadicacionCuentaRecord.factura)
        .filter(RadicacionCuentaRecord.envio.in_(envios))
        .all()
    ):
        en_fuente.setdefault(envio, set()).add(factura)
    faltan = sorted({f for e in envios for f in en_fuente.get(e, set()) - aqui.get(e, set())})
    if not faltan:
        return {}
    # Dónde quedó cada una: una sola consulta (por bloques, por el tope de
    # parámetros de SQLite), nunca una consulta por factura.
    ubicacion = {}
    for i in range(0, len(faltan), 400):
        for canon, radicado in (
            db.query(FacturaPreauditoriaRecord, OficioRecepcionRecord.numero_radicado)
            .outerjoin(
                OficioRecepcionRecord,
                FacturaPreauditoriaRecord.oficio_actual_id == OficioRecepcionRecord.id,
            )
            .filter(FacturaPreauditoriaRecord.factura.in_(faltan[i : i + 400]))
            .all()
        ):
            ubicacion[canon.factura] = (canon, radicado)
    salida = {}
    for envio in envios:
        detalle = [
            _donde_quedo(f, ubicacion.get(f), oficio_id)
            for f in sorted(en_fuente.get(envio, set()) - aqui.get(envio, set()))
        ]
        if detalle:
            salida[envio] = detalle
    return salida


# ==================================================================
# Escribir un envío (el corazón del automatismo)
# ==================================================================


def _nuevo_evento(
    factura_row: FacturaPreauditoriaRecord,
    tipo: str,
    *,
    oficio: Optional[OficioRecepcionRecord],
    fuente: dict,
    usuario: str,
    motivo: str = None,
    observaciones: str = None,
    resultado: str = None,
    oficio_devolucion_id: int = None,
) -> FacturaEventoRecord:
    return FacturaEventoRecord(
        factura_id=factura_row.id,
        factura=factura_row.factura,
        tipo_evento=tipo,
        subsanacion_num=factura_row.num_subsanacion,
        ronda=factura_row.ronda_actual,
        estado_resultante=factura_row.estado,
        envio=factura_row.envio_actual,
        oficio_id=oficio.id if oficio else None,
        oficio_fhus=oficio.numero_radicado if oficio else None,
        f_recibido=oficio.fecha_recibido if oficio else None,
        resultado=resultado if resultado is not None else factura_row.resultado_actual,
        auditor=usuario,
        motivo=motivo,
        observaciones=observaciones,
        oficio_devolucion_id=oficio_devolucion_id,
        valor_snapshot=fuente.get("valor") or 0.0,
        nit_snapshot=fuente.get("nit"),
        entidad_snapshot=fuente.get("entidad"),
        correo_fe_snapshot=fuente.get("correo_fe"),
        fecha_factura_snapshot=fuente.get("f_factura"),
        creado_por=usuario,
    )


def preview_envio(db: Session, oficio: OficioRecepcionRecord, envio: str) -> dict:
    """Simula escribir un envío sin tocar la BD (para confirmar antes)."""
    envio = str(envio).strip()
    cargas = (
        db.query(EnvioCargadoRecord)
        .filter(EnvioCargadoRecord.envio == envio)
        .order_by(EnvioCargadoRecord.cargado_en.desc(), EnvioCargadoRecord.id.desc())
        .all()
    )
    ya = next((c for c in cargas if c.oficio_id == oficio.id), None)
    # Si el envío salió antes en OTRO oficio, la pantalla avisa en cuál: la
    # recarga aquí solo reingresará las devueltas (subsanación).
    otro = next((c for c in cargas if c.oficio_id != oficio.id), None)
    oficio_anterior = (
        db.get(OficioRecepcionRecord, otro.oficio_id) if otro and otro.oficio_id else None
    )
    src = facturas_de_envio(db, envio)
    facturas = []
    nuevas = reingresos = radicadas = pendientes_otro = 0
    alertas = []
    for r in src:
        canon = (
            db.query(FacturaPreauditoriaRecord)
            .filter(FacturaPreauditoriaRecord.factura == r.factura)
            .one_or_none()
        )
        if canon is None:
            clasif = "NUEVA"
            nuevas += 1
        elif canon.resultado_actual == RESULTADO_DEVUELTA:
            clasif = "REINGRESO"
            reingresos += 1
            if canon.num_devoluciones >= MAX_DEVOLUCIONES:
                alertas.append(
                    f"{r.factura} lleva {canon.num_devoluciones} devoluciones: solo radicar/escalar"
                )
        elif canon.resultado_actual == RESULTADO_RADICAR:
            clasif = "YA_RADICADA"
            radicadas += 1
        else:  # PENDIENTE
            clasif = "YA_ABIERTA"
            pendientes_otro += 1
        dg = db.query(DgReportRecord).filter(DgReportRecord.factura == r.factura).one_or_none()
        facturas.append(
            {
                "factura": r.factura,
                "f_factura": r.f_factura.isoformat() if r.f_factura else None,
                "valor": r.valor,
                "nit": r.nit,
                "entidad": r.entidad,
                "correo_fe": dg.correo_fe if dg else "NO",
                "clasificacion": clasif,
            }
        )
    return {
        "envio": envio,
        "existe_en_fuente": bool(src),
        "ya_cargado": bool(ya),
        "cargado_en_otro_oficio": (
            oficio_anterior.numero_radicado
            if oficio_anterior
            else ("otro oficio" if otro else None)
        ),
        "veces_cargado": len(cargas),
        "max_cargas_envio": MAX_OFICIOS_POR_ENVIO,
        "oficios_donde_cargado": _radicados_de_cargas(db, cargas),
        "total_en_fuente": len(src),
        "nuevas": nuevas,
        "reingresos": reingresos,
        "ya_radicadas": radicadas,
        "ya_abiertas": pendientes_otro,
        "alertas_limite_devoluciones": alertas,
        "avisos_fuente": _avisos_de_fuente_rezagada(src),
        "facturas_rezagadas": [r["factura"] for r in facturas_de_fuente_rezagada(src)],
        "facturas": facturas,
    }


def _radicados_de_cargas(db: Session, cargas: list) -> list[str]:
    """Números de radicado de los oficios donde ya se cargó un envío (para
    mensajes que el auditor pueda entender, en vez de ids internos)."""
    nombres = []
    for c in cargas:
        o = db.get(OficioRecepcionRecord, c.oficio_id) if c.oficio_id else None
        nombres.append(o.numero_radicado if o else "un oficio ya eliminado")
    return nombres


def escribir_envio(db: Session, oficio: OficioRecepcionRecord, envio: str, usuario: str) -> dict:
    """Vuelca al consolidado todas las facturas de un envío.

    - Dedup (punto 3): si el envío ya fue cargado EN ESTE OFICIO → no recrea.
    - Primera vez: crea la factura canónica + evento ESCRITA.
    - Reingreso (factura ya DEVUELTA): incrementa subsanación, no crea factura
      nueva + evento REINGRESO.
    - Recarga en un oficio posterior (facturación reenvía las subsanaciones
      con el MISMO número de envío): permitida hasta completar
      MAX_OFICIOS_POR_ENVIO oficios distintos. Al recargar, reingresan las
      facturas devueltas; las radicadas y las aún pendientes se omiten con
      su advertencia.
    """
    envio = str(envio).strip()
    cargas = (
        db.query(EnvioCargadoRecord)
        .filter(EnvioCargadoRecord.envio == envio)
        .order_by(EnvioCargadoRecord.cargado_en.desc(), EnvioCargadoRecord.id.desc())
        .all()
    )
    aqui = next((c for c in cargas if c.oficio_id == oficio.id), None)
    if aqui:
        return {
            "ya_cargado": True,
            "mensaje": "El envío ya fue cargado en este oficio.",
            "envio": envio,
            "cargado_en": a_utc(aqui.cargado_en).isoformat() if aqui.cargado_en else None,
        }
    if len(cargas) >= MAX_OFICIOS_POR_ENVIO:
        radicados = ", ".join(_radicados_de_cargas(db, cargas))
        return {
            "ya_cargado": True,
            "mensaje": (
                f"El envío {envio} ya fue cargado en {len(cargas)} oficios ({radicados}): "
                f"el proceso acepta máximo {MAX_OFICIOS_POR_ENVIO}."
            ),
            "envio": envio,
        }

    src = facturas_de_envio(db, envio)
    if not src:
        return {
            "existe_en_fuente": False,
            "mensaje": (
                f"El envío {envio} no existe en la Radicación de Cuentas cargada. "
                "Suba o actualice la fuente."
            ),
            "envio": envio,
        }

    nuevas = reingresos = omitidas = 0
    alertas = []
    advertencias = []
    # TODO el volcado va dentro del try: si otra petición simultánea escribe
    # el mismo envío (doble clic aquí, o el mismo envío desde otro oficio con
    # una factura nueva), el conflicto puede saltar en un flush del bucle, no
    # solo en el commit — y debe responderse con un mensaje claro, no un 500.
    try:
        for r in src:
            dg = db.query(DgReportRecord).filter(DgReportRecord.factura == r.factura).one_or_none()
            fuente = {
                "valor": r.valor,
                "nit": r.nit,
                "entidad": r.entidad,
                "correo_fe": dg.correo_fe if dg else "NO",
                "f_factura": r.f_factura,
            }
            canon = (
                db.query(FacturaPreauditoriaRecord)
                .filter(FacturaPreauditoriaRecord.factura == r.factura)
                .one_or_none()
            )
            if canon is None:
                # PRIMERA VEZ
                canon = FacturaPreauditoriaRecord(
                    factura=r.factura,
                    envio_actual=envio,
                    oficio_actual_id=oficio.id,
                    oficio_fhus=oficio.numero_radicado,
                    f_recibido=oficio.fecha_recibido,
                    estado=ESTADO_NUEVA,
                    resultado_actual=RESULTADO_PENDIENTE,
                    ronda_actual=1,
                    num_subsanacion=0,
                    num_devoluciones=0,
                    pendiente_subsanacion=0,
                    creado_por=usuario,
                )
                db.add(canon)
                db.flush()
                db.add(
                    _nuevo_evento(canon, "ESCRITA", oficio=oficio, fuente=fuente, usuario=usuario)
                )
                nuevas += 1
            elif canon.resultado_actual == RESULTADO_DEVUELTA:
                # REINGRESO / SUBSANACIÓN (no se crea factura nueva)
                canon.ronda_actual += 1
                canon.num_subsanacion = canon.ronda_actual - 1
                canon.envio_actual = envio
                canon.oficio_actual_id = oficio.id
                canon.oficio_fhus = oficio.numero_radicado
                canon.f_recibido = oficio.fecha_recibido
                canon.resultado_actual = RESULTADO_PENDIENTE
                canon.estado = ESTADO_EN_SUBSANACION
                canon.pendiente_subsanacion = 0
                canon.auditor = None
                canon.fecha_auditoria = None
                canon.oficio_devolucion_id = None
                db.flush()
                db.add(
                    _nuevo_evento(canon, "REINGRESO", oficio=oficio, fuente=fuente, usuario=usuario)
                )
                reingresos += 1
                if canon.num_devoluciones >= MAX_DEVOLUCIONES:
                    alertas.append(
                        f"{r.factura} ya lleva {canon.num_devoluciones} devoluciones: "
                        "solo debería radicarse o escalarse."
                    )
            elif canon.resultado_actual == RESULTADO_RADICAR:
                omitidas += 1
                advertencias.append(f"{r.factura} ya estaba radicada: no se reingresó.")
            else:  # PENDIENTE ya abierta
                omitidas += 1
                advertencias.append(
                    f"{r.factura} sigue pendiente en {canon.oficio_fhus or 'otro oficio'}: "
                    "resuélvala primero."
                )

        if cargas and not nuevas and not reingresos:
            # Recarga que no trajo nada: ninguna factura del envío estaba devuelta.
            # Se permite (la regla es hasta 3 oficios) pero se avisa con claridad.
            advertencias.append(
                f"La recarga no movió ninguna factura: ninguna del envío {envio} estaba "
                "devuelta (las radicadas y las pendientes se quedan donde están)."
            )
        # Facturas que la fuente arrastró de un cargue viejo: puede que ya no
        # pertenezcan al envío. Se cargan igual (la fuente manda) pero el
        # auditor queda avisado y puede quitarlas con el botón 🗑.
        advertencias += _avisos_de_fuente_rezagada(src)
        db.add(
            EnvioCargadoRecord(
                envio=envio,
                oficio_id=oficio.id,
                total_facturas=len(src),
                nuevas=nuevas,
                reingresos=reingresos,
                cargado_por=usuario,
            )
        )
        # Recuento del tope DENTRO de la transacción, con la fila propia ya
        # insertada: dos peticiones simultáneas sobre DOS oficios distintos
        # pasan ambas el chequeo de arriba (el candado compuesto no las choca
        # porque son filas distintas). Este recuento detiene a la que llegue
        # de segunda antes de confirmar.
        db.flush()
        total_cargas = (
            db.query(EnvioCargadoRecord).filter(EnvioCargadoRecord.envio == envio).count()
        )
        if total_cargas > MAX_OFICIOS_POR_ENVIO:
            db.rollback()
            return {
                "ya_cargado": True,
                "mensaje": (
                    f"El envío {envio} ya alcanzó el máximo de {MAX_OFICIOS_POR_ENVIO} "
                    "oficios (otra carga entró al mismo tiempo). Actualice la página."
                ),
                "envio": envio,
            }
        db.commit()
    except IntegrityError:
        # Carrera: otra petición escribió el mismo envío al mismo tiempo (doble
        # clic en este oficio, o el mismo envío con una factura nueva desde otro
        # oficio). Nada quedó a medias: se deshace todo y se avisa claro.
        db.rollback()
        return {
            "ya_cargado": True,
            "mensaje": (
                "Otra persona estaba cargando este envío en este mismo momento. "
                "Actualice la página y revise los envíos del oficio antes de reintentar."
            ),
            "envio": envio,
        }
    return {
        "ya_cargado": False,
        "existe_en_fuente": True,
        "envio": envio,
        "total_en_fuente": len(src),
        "nuevas": nuevas,
        "reingresos": reingresos,
        "omitidas": omitidas,
        "alertas_limite_devoluciones": alertas,
        "advertencias": advertencias,
    }


# ==================================================================
# Auditar (radicar / devolver) con contador de devoluciones y tope 3
# ==================================================================


def puede_revertir(
    factura_row: FacturaPreauditoriaRecord, usuario: str, es_coordinacion: bool
) -> tuple[bool, str]:
    """¿Esta persona puede volver a PENDIENTE una auditoría ya decidida?

    Regla acordada con Cartera (20-08-2026): **cada auditor deshace lo suyo**
    mientras no haya salido del hospital. Antes solo coordinación podía, y los
    auditores quedaban trancados por su propio error de dedo. Lo que NO cambia:
    la decisión de otra persona, y cualquiera que ya salió en un oficio de
    devolución con consecutivo emitido, siguen siendo de coordinación — ese PDF
    ya está en manos de la entidad.

    Devuelve (puede, motivo_si_no_puede).
    """
    if es_coordinacion:
        return True, ""
    if factura_row.oficio_devolucion_id:
        return False, (
            f"{factura_row.factura} ya salió en un oficio de devolución entregado: "
            "solo coordinación o administración puede deshacer esa auditoría."
        )
    mia = (factura_row.auditor or "").strip().casefold() == (usuario or "").strip().casefold()
    if not mia:
        quien = factura_row.auditor or "otra persona"
        return False, (
            f"La auditoría de {factura_row.factura} la hizo {quien}: solo esa persona, "
            "coordinación o administración puede deshacerla."
        )
    return True, ""


def _unir_observacion(observaciones: str = None, motivo: str = None) -> Optional[str]:
    """Junta lo escrito en "Observaciones" con lo escrito en "Motivo" al radicar.

    Al radicar no existe un "motivo de devolución", pero el auditor igual puede
    haber escrito ahí: ese recuadro está arriba y es el que más se ve. Antes su
    texto se descartaba en silencio y el historial quedaba vacío (caso real
    29-07-2026: se escribió "OKAY SOPORTES" al radicar y no quedó registrado en
    ninguna parte). Ahora nunca se pierde: se guarda como observación.
    """
    obs = (observaciones or "").strip()
    mot = (motivo or "").strip()
    if not mot:
        return observaciones
    if not obs:
        return mot
    if mot in obs:  # ya lo escribió en los dos recuadros: no duplicar
        return observaciones
    return f"{obs} — {mot}"


def auditar_factura(
    db: Session,
    canon: FacturaPreauditoriaRecord,
    resultado: str,
    usuario: str,
    motivo: str = None,
    observaciones: str = None,
) -> dict:
    """Aplica RADICAR / DEVUELTA / PENDIENTE (revertir).

    Devuelve {"ok": True, ...} o {"ok": False, "codigo": 400|409, "mensaje": ...}.
    """
    oficio = (
        db.get(OficioRecepcionRecord, canon.oficio_actual_id) if canon.oficio_actual_id else None
    )
    fuente = datos_fuente(db, canon.factura)

    # --- Validaciones del proceso (antes de tocar nada) ---
    if resultado == RESULTADO_RADICAR:
        # Regla del proceso: sin soporte de facturación electrónica (CORREO
        # F.E. = NO) la factura no se puede radicar — solo devolverse.
        if (fuente.get("correo_fe") or "NO") != "SI":
            return {
                "ok": False,
                "codigo": 409,
                "mensaje": (
                    f"La factura {canon.factura} no tiene soporte de facturación "
                    "electrónica (CORREO F.E. = NO): solo puede devolverse. Si el "
                    "soporte ya existe, suba el Formato Facturación Electrónica "
                    "actualizado en la pestaña Fuentes."
                ),
            }
    elif resultado == RESULTADO_DEVUELTA:
        if not (motivo or "").strip():
            return {
                "ok": False,
                "codigo": 400,
                "mensaje": (
                    "Para devolver debe escribir la observación: es el texto "
                    "que sale en el oficio de devolución."
                ),
            }
        if canon.num_devoluciones >= MAX_DEVOLUCIONES:
            return {
                "ok": False,
                "codigo": 409,
                "mensaje": (
                    f"La factura {canon.factura} ya fue devuelta {canon.num_devoluciones} veces; "
                    f"el proceso acepta máximo {MAX_DEVOLUCIONES} devoluciones. Debe radicarse "
                    "o escalarse al coordinador."
                ),
            }

    # Guard de estado: solo se puede radicar/devolver una factura PENDIENTE.
    # Para cambiar una decisión ya tomada hay que revertir primero (→ PENDIENTE).
    #
    # Se toma con una actualización condicional (un solo UPDATE ... WHERE
    # resultado_actual = 'PENDIENTE'), no leyendo y luego escribiendo: si el
    # auditor hace varios clics seguidos porque el servidor va lento, llegan
    # varias peticiones a la vez y todas leerían "PENDIENTE" antes de que la
    # primera guarde, duplicando el evento en el historial. Así solo una gana.
    if resultado in (RESULTADO_RADICAR, RESULTADO_DEVUELTA):
        tomada = (
            db.query(FacturaPreauditoriaRecord)
            .filter(
                FacturaPreauditoriaRecord.id == canon.id,
                FacturaPreauditoriaRecord.resultado_actual == RESULTADO_PENDIENTE,
            )
            .update({"resultado_actual": resultado}, synchronize_session=False)
        )
        if not tomada:
            db.rollback()
            db.refresh(canon)
            etiqueta = "radicada" if canon.resultado_actual == RESULTADO_RADICAR else "devuelta"
            return {
                "ok": False,
                "codigo": 409,
                "mensaje": (
                    f"La factura {canon.factura} ya fue {etiqueta} en esta ronda. "
                    "Para cambiar la decisión, primero use 'Dejar pendiente' (revertir)."
                ),
            }
        # El objeto en memoria quedó desfasado respecto de la actualización.
        canon.resultado_actual = resultado

    if resultado == RESULTADO_RADICAR:
        # Nada de lo que escribió el auditor se descarta: si escribió en el
        # recuadro de motivo (que al radicar no aplica), su texto pasa a la
        # observación en vez de perderse.
        observaciones = _unir_observacion(observaciones, motivo)
        canon.resultado_actual = RESULTADO_RADICAR
        canon.estado = ESTADO_RADICADA if canon.ronda_actual == 1 else ESTADO_SUBSANADA
        canon.pendiente_subsanacion = 0
        canon.motivo_ultima_devolucion = None
        canon.auditor = usuario
        canon.fecha_auditoria = ahora_utc()
        if observaciones is not None:
            canon.observaciones = observaciones
        db.flush()
        tipo = "SUBSANADA" if canon.ronda_actual >= 2 else "RADICADA"
        db.add(
            _nuevo_evento(
                canon,
                tipo,
                oficio=oficio,
                fuente=fuente,
                usuario=usuario,
                observaciones=canon.observaciones,
                resultado=RESULTADO_RADICAR,
            )
        )
        db.commit()
        return {"ok": True}

    if resultado == RESULTADO_DEVUELTA:
        canon.num_devoluciones += 1
        canon.resultado_actual = RESULTADO_DEVUELTA
        canon.motivo_ultima_devolucion = motivo.strip()
        canon.pendiente_subsanacion = 1
        if canon.num_devoluciones >= MAX_DEVOLUCIONES:
            canon.estado = ESTADO_BLOQUEADA
        elif canon.ronda_actual == 1:
            canon.estado = ESTADO_DEVUELTA_PEND
        else:
            canon.estado = ESTADO_NUEV_DEVUELTA
        canon.auditor = usuario
        canon.fecha_auditoria = ahora_utc()
        if observaciones is not None:
            canon.observaciones = observaciones
        db.flush()
        tipo = "NUEVAMENTE_DEVUELTA" if canon.ronda_actual >= 2 else "DEVUELTA"
        db.add(
            _nuevo_evento(
                canon,
                tipo,
                oficio=oficio,
                fuente=fuente,
                usuario=usuario,
                motivo=motivo.strip(),
                observaciones=canon.observaciones,
                resultado=RESULTADO_DEVUELTA,
            )
        )
        db.commit()
        return {"ok": True}

    if resultado == RESULTADO_PENDIENTE:
        # Revertir una auditoría (no si ya salió en un oficio de devolución PDF).
        if canon.oficio_devolucion_id:
            return {
                "ok": False,
                "codigo": 409,
                "mensaje": "Esta factura ya salió en un oficio de devolución; no se puede revertir.",
            }
        if canon.resultado_actual == RESULTADO_DEVUELTA and canon.num_devoluciones > 0:
            canon.num_devoluciones -= 1
        canon.resultado_actual = RESULTADO_PENDIENTE
        canon.estado = ESTADO_NUEVA if canon.ronda_actual == 1 else ESTADO_EN_SUBSANACION
        canon.pendiente_subsanacion = 0
        canon.auditor = None
        canon.fecha_auditoria = None
        canon.motivo_ultima_devolucion = None
        db.flush()
        db.add(
            _nuevo_evento(
                canon,
                "REVERTIDA",
                oficio=oficio,
                fuente=fuente,
                usuario=usuario,
                # La observación viaja también en la fila REVERTIDA: al revisar
                # el historial se ve qué había anotado quien decidió antes.
                observaciones=canon.observaciones,
                resultado=RESULTADO_PENDIENTE,
            )
        )
        db.commit()
        return {"ok": True}

    return {"ok": False, "codigo": 400, "mensaje": "Resultado inválido."}


# ==================================================================
# Anotar la observación de una factura YA auditada (sin revertirla)
# ==================================================================


def anotar_observacion(
    db: Session,
    canon: FacturaPreauditoriaRecord,
    observaciones: str,
    usuario: str,
) -> dict:
    """Agrega o corrige la observación sin cambiar la decisión de la factura.

    Sirve para las facturas que ya se decidieron: se corrige el texto ahora,
    sin tener que revertir y volver a auditar. Queda como un evento más del
    historial, con quién la escribió y cuándo.

    Si la factura está DEVUELTA, la observación ES el motivo que sale en el
    oficio de devolución, así que la corrección también actualiza el evento
    de devolución de esta ronda — esté o no ya sellado en un oficio. Como el
    PDF se arma al abrirlo, el oficio queda corregido de inmediato.
    """
    texto = (observaciones or "").strip()
    if not texto:
        return {"ok": False, "codigo": 400, "mensaje": "Escriba la observación."}

    oficio = (
        db.get(OficioRecepcionRecord, canon.oficio_actual_id) if canon.oficio_actual_id else None
    )
    fuente = datos_fuente(db, canon.factura)
    canon.observaciones = texto
    motivo_evento = None
    oficio_corregido = False
    if canon.resultado_actual == RESULTADO_DEVUELTA:
        canon.motivo_ultima_devolucion = texto
        motivo_evento = texto
        # El evento de devolución de ESTA ronda: si la factura ya salió en un
        # oficio, es el evento sellado con ese oficio (corregirlo corrige el
        # PDF); si aún no ha salido, es el que se sellará cuando se genere.
        # Los eventos de rondas anteriores (sellados en oficios viejos) no se
        # tocan: esos documentos ya se entregaron tal como estaban.
        consulta = db.query(FacturaEventoRecord).filter(
            FacturaEventoRecord.factura_id == canon.id,
            FacturaEventoRecord.tipo_evento.in_(["DEVUELTA", "NUEVAMENTE_DEVUELTA"]),
        )
        if canon.oficio_devolucion_id:
            consulta = consulta.filter(
                FacturaEventoRecord.oficio_devolucion_id == canon.oficio_devolucion_id
            )
        else:
            consulta = consulta.filter(FacturaEventoRecord.oficio_devolucion_id.is_(None))
        evt = consulta.order_by(
            FacturaEventoRecord.creado_en.desc(), FacturaEventoRecord.id.desc()
        ).first()
        if evt:
            evt.motivo = texto
            oficio_corregido = bool(evt.oficio_devolucion_id)
    db.flush()
    db.add(
        _nuevo_evento(
            canon,
            "OBSERVACION",
            oficio=oficio,
            fuente=fuente,
            usuario=usuario,
            motivo=motivo_evento,
            observaciones=texto,
            # No cambia la decisión: el evento conserva la que ya tenía.
            resultado=canon.resultado_actual,
        )
    )
    db.commit()
    return {"ok": True, "oficio_actualizado": oficio_corregido}


# ==================================================================
# Eliminar un oficio mal registrado (solo administradores)
# ==================================================================


def _facturas_en_pdf_emitido(db: Session, facturas: list, oficio_id: int) -> list[str]:
    """Facturas del lote que ya salieron en un oficio de devolución emitido.

    Mientras exista una de estas, no se puede deshacer nada: el PDF ya se
    entregó a la entidad y debe seguir cuadrando con el sistema.
    """
    con_pdf = [f.factura for f in facturas if f.oficio_devolucion_id]
    ids = {f.id for f in facturas}
    if ids:
        sellados = (
            db.query(FacturaEventoRecord.factura)
            .filter(
                FacturaEventoRecord.oficio_id == oficio_id,
                FacturaEventoRecord.oficio_devolucion_id.isnot(None),
                FacturaEventoRecord.factura_id.in_(ids),
            )
            .all()
        )
        con_pdf += [r[0] for r in sellados]
    return sorted(set(con_pdf))


def _deshacer_facturas(db: Session, facturas: list, oficio_id: int) -> tuple[int, int]:
    """Deshace la carga de esas facturas en ese oficio.

    - Ronda 1 (nacieron aquí): se borran con sus eventos.
    - Ronda 2+ (reingresaron aquí para subsanar): NO se borran; se revierte el
      reingreso y vuelven a su estado DEVUELTA anterior, con su historial
      intacto.

    Devuelve (borradas, revertidas).
    """
    borradas = 0
    revertidas = 0
    for f in facturas:
        if f.ronda_actual <= 1:
            db.query(FacturaEventoRecord).filter(FacturaEventoRecord.factura_id == f.id).delete(
                synchronize_session=False
            )
            db.delete(f)
            borradas += 1
        else:
            # Reingreso: revertir a la devolución anterior sin perder historial.
            evento_dev = (
                db.query(FacturaEventoRecord)
                .filter(
                    FacturaEventoRecord.factura_id == f.id,
                    FacturaEventoRecord.tipo_evento.in_(["DEVUELTA", "NUEVAMENTE_DEVUELTA"]),
                    FacturaEventoRecord.oficio_id != oficio_id,
                )
                .order_by(FacturaEventoRecord.creado_en.desc(), FacturaEventoRecord.id.desc())
                .first()
            )
            db.query(FacturaEventoRecord).filter(
                FacturaEventoRecord.factura_id == f.id,
                FacturaEventoRecord.oficio_id == oficio_id,
            ).delete(synchronize_session=False)
            f.ronda_actual -= 1
            f.num_subsanacion = max(0, f.ronda_actual - 1)
            f.resultado_actual = RESULTADO_DEVUELTA
            f.pendiente_subsanacion = 1
            if f.num_devoluciones >= MAX_DEVOLUCIONES:
                f.estado = ESTADO_BLOQUEADA
            elif f.ronda_actual == 1:
                f.estado = ESTADO_DEVUELTA_PEND
            else:
                f.estado = ESTADO_NUEV_DEVUELTA
            if evento_dev:
                f.envio_actual = evento_dev.envio
                f.oficio_actual_id = evento_dev.oficio_id
                f.oficio_fhus = evento_dev.oficio_fhus
                f.f_recibido = evento_dev.f_recibido
                f.auditor = evento_dev.auditor
                f.motivo_ultima_devolucion = evento_dev.motivo
                f.fecha_auditoria = evento_dev.creado_en
                # Si esa devolución ya había salido en un oficio con
                # consecutivo emitido, la factura vuelve amarrada a él. Sin
                # esto quedaba "devuelta sin oficio" y el sistema dejaba
                # emitir un SEGUNDO oficio de devolución por la misma
                # factura: el PDF salía diciendo "1 factura" con la tabla
                # vacía, porque su único evento ya estaba sellado en el
                # primero.
                f.oficio_devolucion_id = evento_dev.oficio_devolucion_id
            revertidas += 1
    return borradas, revertidas


def _facturas_migradas_con_historial(
    db: Session, oficio_id: int, ids_locales: set, envio: str = None
) -> list[str]:
    """Facturas cuyo historial referencia este oficio (o este envío en este
    oficio) pero que YA pertenecen a otro oficio (subsanaron con la recarga).

    Mientras existan, ni el oficio ni el envío se pueden deshacer: sus eventos
    son la historia de esas facturas — borrarlos rompería la trazabilidad, y
    borrar el oficio dejaría eventos apuntando a un oficio inexistente (la
    base lo rechaza con un error de integridad que antes salía como error 500).
    """
    q = db.query(FacturaEventoRecord.factura).filter(FacturaEventoRecord.oficio_id == oficio_id)
    if envio is not None:
        q = q.filter(FacturaEventoRecord.envio == envio)
    if ids_locales:
        q = q.filter(~FacturaEventoRecord.factura_id.in_(ids_locales))
    return sorted({f for (f,) in q.distinct().all() if f})


def renombrar_oficio(db: Session, oficio: OficioRecepcionRecord, nuevo_radicado: str) -> dict:
    """Corrige el número de radicado de un oficio ya registrado.

    POR QUÉ EXISTE (caso real 20-08-2026): un oficio quedó grabado como
    FHUS-AS-101190-26, con un UNO en vez de la I, y no había forma de
    corregirlo sin borrarlo y perder su historia. El número no es un dato
    suelto: se copia en cada factura del oficio y en cada evento del
    historial, así que corregirlo aquí lo corrige en todas partes (dos
    UPDATE, sin recorrer factura por factura).
    """
    nuevo = re.sub(r"\s+", "", (nuevo_radicado or "")).upper()
    if not nuevo:
        return {"ok": False, "codigo": 400, "mensaje": "Escriba el número correcto del oficio."}
    if len(nuevo) > 60:
        return {"ok": False, "codigo": 400, "mensaje": "El número del oficio es demasiado largo."}
    antes = oficio.numero_radicado
    if nuevo == antes:
        return {
            "ok": True,
            "sin_cambios": True,
            "antes": antes,
            "ahora": nuevo,
            "facturas": 0,
            "eventos": 0,
        }
    otro = (
        db.query(OficioRecepcionRecord)
        .filter(
            OficioRecepcionRecord.numero_radicado == nuevo,
            OficioRecepcionRecord.id != oficio.id,
        )
        .first()
    )
    if otro:
        return {
            "ok": False,
            "codigo": 409,
            "mensaje": (
                f"Ya hay otro oficio registrado como {nuevo} (recibido el "
                f"{otro.fecha_recibido:%d/%m/%Y}): revise cuál es el correcto."
            ),
        }
    oficio.numero_radicado = nuevo
    facturas = (
        db.query(FacturaPreauditoriaRecord)
        .filter(FacturaPreauditoriaRecord.oficio_actual_id == oficio.id)
        .update({FacturaPreauditoriaRecord.oficio_fhus: nuevo}, synchronize_session=False)
    )
    eventos = (
        db.query(FacturaEventoRecord)
        .filter(FacturaEventoRecord.oficio_id == oficio.id)
        .update({FacturaEventoRecord.oficio_fhus: nuevo}, synchronize_session=False)
    )
    try:
        db.commit()
    except IntegrityError:
        # Otra persona registró ese mismo número mientras tanto: nada queda
        # a medias y el aviso es claro, no un error de programa.
        db.rollback()
        return {
            "ok": False,
            "codigo": 409,
            "mensaje": (
                f"Otra persona acaba de registrar el oficio {nuevo}. "
                "Actualice la página y revise cuál es el correcto."
            ),
        }
    return {
        "ok": True,
        "sin_cambios": False,
        "antes": antes,
        "ahora": nuevo,
        "facturas": facturas or 0,
        "eventos": eventos or 0,
    }


def eliminar_oficio(db: Session, oficio: OficioRecepcionRecord) -> dict:
    """Elimina un oficio de recepción mal registrado, con salvaguardas.

    Si alguna de sus facturas ya salió en un oficio de devolución emitido, o
    ya subsanó hacia otro oficio (recarga del envío), no se elimina. Los
    envíos quedan libres para volver a escribirse.
    """
    facturas = (
        db.query(FacturaPreauditoriaRecord)
        .filter(FacturaPreauditoriaRecord.oficio_actual_id == oficio.id)
        .all()
    )
    con_pdf = _facturas_en_pdf_emitido(db, facturas, oficio.id)
    if con_pdf:
        return {
            "ok": False,
            "codigo": 409,
            "mensaje": (
                f"El oficio {oficio.numero_radicado} tiene factura(s) en un oficio de "
                f"devolución ya emitido ({', '.join(con_pdf[:5])}...): "
                "no se puede eliminar."
            ),
        }
    migradas = _facturas_migradas_con_historial(db, oficio.id, {f.id for f in facturas})
    if migradas:
        return {
            "ok": False,
            "codigo": 409,
            "mensaje": (
                f"El oficio {oficio.numero_radicado} no se puede eliminar: "
                f"{len(migradas)} factura(s) que pasaron por él ya subsanaron hacia "
                f"otro oficio ({', '.join(migradas[:5])}{'…' if len(migradas) > 5 else ''}) "
                "y su historial referencia este oficio."
            ),
        }

    borradas, revertidas = _deshacer_facturas(db, facturas, oficio.id)
    envios = db.query(EnvioCargadoRecord).filter(EnvioCargadoRecord.oficio_id == oficio.id).all()
    for e in envios:
        db.delete(e)
    radicado = oficio.numero_radicado
    db.delete(oficio)
    db.commit()
    return {
        "ok": True,
        "radicado": radicado,
        "facturas_borradas": borradas,
        "subsanaciones_revertidas": revertidas,
        "envios_liberados": len(envios),
    }


def eliminar_oficio_devolucion(db: Session, dev: OficioDevolucionRecord) -> dict:
    """Elimina un oficio de devolución ya generado (solo coordinación/admin).

    Sirve cuando el PDF salió con el consecutivo equivocado o se generó por
    error. Las facturas NO cambian de decisión: siguen devueltas, pero quedan
    libres para entrar en un oficio nuevo. El consecutivo vuelve a quedar
    disponible.

    El historial NO se borra: a los eventos solo se les quita el sello del
    oficio que deja de existir, así queda constancia de que la factura sí fue
    devuelta ese día.
    """
    facturas = (
        db.query(FacturaPreauditoriaRecord)
        .filter(FacturaPreauditoriaRecord.oficio_devolucion_id == dev.id)
        .all()
    )
    eventos = (
        db.query(FacturaEventoRecord)
        .filter(FacturaEventoRecord.oficio_devolucion_id == dev.id)
        .all()
    )
    for f in facturas:
        f.oficio_devolucion_id = None
    for e in eventos:
        e.oficio_devolucion_id = None
    consecutivo = dev.consecutivo
    db.flush()
    db.delete(dev)
    db.commit()
    return {
        "ok": True,
        "consecutivo": consecutivo,
        "facturas_liberadas": len(facturas),
    }


def eliminar_envio(db: Session, oficio: OficioRecepcionRecord, envio: str) -> dict:
    """Deshace UN envío cargado por error, sin tocar el resto del oficio.

    Sirve para cuando el gestor escribe un número de envío equivocado: se
    borran las facturas que entraron con ese envío (o se revierte su
    reingreso si venían de una devolución) y el envío queda libre para
    volver a escribirse, aquí o en otro oficio.

    No se puede deshacer si alguna de esas facturas ya salió en un oficio de
    devolución con consecutivo emitido.
    """
    envio = (envio or "").strip()
    cargado = (
        db.query(EnvioCargadoRecord)
        .filter(
            EnvioCargadoRecord.oficio_id == oficio.id,
            EnvioCargadoRecord.envio == envio,
        )
        .first()
    )
    facturas = (
        db.query(FacturaPreauditoriaRecord)
        .filter(
            FacturaPreauditoriaRecord.oficio_actual_id == oficio.id,
            FacturaPreauditoriaRecord.envio_actual == envio,
        )
        .all()
    )
    if cargado is None and not facturas:
        return {
            "ok": False,
            "codigo": 404,
            "mensaje": f"El envío {envio} no está cargado en el oficio {oficio.numero_radicado}.",
        }

    con_pdf = _facturas_en_pdf_emitido(db, facturas, oficio.id)
    if con_pdf:
        return {
            "ok": False,
            "codigo": 409,
            "mensaje": (
                f"El envío {envio} tiene factura(s) en un oficio de devolución ya "
                f"emitido ({', '.join(con_pdf[:5])}...): no se puede eliminar."
            ),
        }
    migradas = _facturas_migradas_con_historial(
        db, oficio.id, {f.id for f in facturas}, envio=envio
    )
    if migradas:
        # Si se dejara quitar, el envío recuperaría un cupo del tope de 3 y el
        # registro dejaría de contar la historia real de por dónde anduvo.
        return {
            "ok": False,
            "codigo": 409,
            "mensaje": (
                f"El envío {envio} no se puede quitar de {oficio.numero_radicado}: "
                f"sus facturas ({', '.join(migradas[:5])}"
                f"{'…' if len(migradas) > 5 else ''}) ya subsanaron hacia otro oficio "
                "y este registro es parte de su historial."
            ),
        }

    borradas, revertidas = _deshacer_facturas(db, facturas, oficio.id)
    if cargado is not None:
        db.delete(cargado)
    db.commit()
    return {
        "ok": True,
        "envio": envio,
        "radicado": oficio.numero_radicado,
        "facturas_borradas": borradas,
        "subsanaciones_revertidas": revertidas,
    }


def quitar_factura(
    db: Session, oficio: OficioRecepcionRecord, factura_row: FacturaPreauditoriaRecord
) -> dict:
    """Quita UNA factura que entró por error a un oficio, sin tocar el resto.

    POR QUÉ EXISTE (caso real 19-08-2026): al escribir un envío el sistema
    trae TODAS las facturas que la fuente de Radicación tenga con ese número.
    Si facturación ya sacó una factura del envío, su fila vieja se queda en la
    fuente y la factura entra igual — el oficio queda con una factura que no
    le corresponde. Antes tocaba quitar el envío entero y volverlo a escribir.

    Se comporta igual que quitar un envío, pero para una sola factura:
      - Si nació aquí (ronda 1): se borra con sus eventos.
      - Si venía reingresada de una devolución: NO se borra; se revierte el
        reingreso y vuelve a su estado devuelto anterior, con su historial.
    El registro del envío se queda (es la historia de por dónde anduvo) pero
    se le descuenta esta factura del conteo.

    No se puede quitar si ya salió en un oficio de devolución emitido: ese PDF
    ya se entregó a la entidad y el sistema debe seguir cuadrando con él.
    """
    if factura_row.oficio_actual_id != oficio.id:
        return {
            "ok": False,
            "codigo": 409,
            "mensaje": (
                f"La factura {factura_row.factura} ya no pertenece a "
                f"{oficio.numero_radicado} (está en "
                f"{factura_row.oficio_fhus or 'otro oficio'}): quítela desde allá."
            ),
        }
    con_pdf = _facturas_en_pdf_emitido(db, [factura_row], oficio.id)
    if con_pdf:
        return {
            "ok": False,
            "codigo": 409,
            "mensaje": (
                f"La factura {factura_row.factura} ya salió en un oficio de devolución "
                "emitido: no se puede quitar. Si el oficio se generó por error, "
                "elimínelo primero desde la pestaña «Oficios de devolución»."
            ),
        }

    factura = factura_row.factura
    envio = factura_row.envio_actual
    borradas, revertidas = _deshacer_facturas(db, [factura_row], oficio.id)
    # El envío sigue cargado (con una factura menos): así el registro conserva
    # por dónde anduvo y el tope de 3 oficios no se altera.
    cargado = (
        db.query(EnvioCargadoRecord)
        .filter(
            EnvioCargadoRecord.oficio_id == oficio.id,
            EnvioCargadoRecord.envio == envio,
        )
        .first()
        if envio
        else None
    )
    if cargado is not None:
        cargado.total_facturas = max(0, (cargado.total_facturas or 0) - 1)
        if borradas:
            cargado.nuevas = max(0, (cargado.nuevas or 0) - 1)
        else:
            cargado.reingresos = max(0, (cargado.reingresos or 0) - 1)
    db.commit()
    return {
        "ok": True,
        "factura": factura,
        "envio": envio,
        "radicado": oficio.numero_radicado,
        "borrada": bool(borradas),
        "reingreso_revertido": bool(revertidas),
    }


def limpiar_modulo(db: Session, incluir_fuentes: bool = False) -> dict:
    """Borra TODOS los datos del módulo de pre-auditoría (solo administradores).

    Deja la página lista para empezar a trabajar desde cero: elimina el
    historial de eventos, las facturas del consolidado, los envíos escritos,
    los oficios de devolución emitidos y los oficios de recepción.

    Con ``incluir_fuentes=True`` borra también las dos fuentes cargadas
    (Radicación de Cuentas y Formato Facturación Electrónica); por defecto se
    conservan para no tener que subirlas otra vez.

    Ojo: al borrar los oficios de devolución, el consecutivo SINAC
    (DEV-PRE-AUD-####-AAAA) vuelve a empezar en 0001 para el año.
    Devuelve el conteo de filas borradas por tabla.
    """
    conteos = {}
    # Orden seguro respecto a llaves foráneas: primero las tablas hijas.
    conteos["eventos"] = db.query(FacturaEventoRecord).delete(synchronize_session=False)
    conteos["facturas"] = db.query(FacturaPreauditoriaRecord).delete(synchronize_session=False)
    conteos["envios_cargados"] = db.query(EnvioCargadoRecord).delete(synchronize_session=False)
    conteos["oficios_devolucion"] = db.query(OficioDevolucionRecord).delete(
        synchronize_session=False
    )
    conteos["oficios_recepcion"] = db.query(OficioRecepcionRecord).delete(synchronize_session=False)
    if incluir_fuentes:
        conteos["fuente_radicacion"] = db.query(RadicacionCuentaRecord).delete(
            synchronize_session=False
        )
        conteos["fuente_facturacion_electronica"] = db.query(DgReportRecord).delete(
            synchronize_session=False
        )
    db.commit()
    conteos["total"] = sum(conteos.values())
    return conteos


# ==================================================================
# ADRES — los oficios de esta entidad exigen entregar un Excel con la
# información completa de las facturas.
# ==================================================================

NIT_ADRES = "901037916"


def es_adres(nit, entidad) -> bool:
    """¿La factura pertenece a ADRES (Administradora de los Recursos del
    Sistema General de Seguridad Social en Salud)? Se reconoce por NIT
    (901037916, con o sin dígito de verificación) o por el nombre."""
    if nit:
        digitos = re.sub(r"\D", "", str(nit))
        if digitos.startswith(NIT_ADRES):
            return True
    e = (entidad or "").upper()
    return "ADRES" in e or "ADMINISTRADORA DE LOS RECURSOS" in e
