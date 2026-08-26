"""El Excel del consolidado de PRE-AUDITORÍA, hecho para leerse.

POR QUÉ EXISTE (26-08-2026). El botón «Exportar Excel» sacaba una sola hoja de
15 columnas: los datos estaban, pero no decían nada. Yesid, textual: «quisiera
que el botón de exportar Excel me genere el Excel pero así como este, bien
detallado, bien definido, algo pulido».

Este módulo arma el libro completo: una fila por factura con **lo que escribió
el gestor**, las devoluciones agrupadas por causa —que es la pregunta que de
verdad importa: por qué se están devolviendo— y los resúmenes por gestor y por
entidad. Los totales son FÓRMULAS sobre la hoja de datos: si el auditor corrige
una fila, el resto se recalcula solo.

Vive en services/ y no en el router porque es lógica del negocio: qué se
muestra, cómo se agrupa y qué recomendación lleva cada causa.
"""

from __future__ import annotations

import re
import unicodedata
from collections import Counter
from datetime import date, datetime
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.core.tz import TZ_BOGOTA
from app.models.db import (
    DgReportRecord,
    FacturaEventoRecord,
    FacturaPreauditoriaRecord,
    OficioDevolucionRecord,
    RadicacionCuentaRecord,
)

FUENTE = "Arial"
AZUL = "1F3864"
AZUL_CLARO = "D9E2F3"
GRIS = "F2F2F2"
AMBAR = "FFF2CC"
ROJO = "F8CECC"
VERDE = "E2EFDA"

# ------------------------------------------------------------------
# ¿Por qué se devolvió? Se deduce de lo que escribió el gestor.
# El orden importa: primero lo más específico. Los patrones salen de los
# textos reales del equipo (se revisaron 1.085 devoluciones el 25-08-2026).
# ------------------------------------------------------------------
CAUSAS_DEVOLUCION: list[tuple[str, str, str]] = [
    (
        "Falta el comprobante de recibido del usuario",
        r"COMPROBANTE DE RECIBIDO|RECIBIDO DEL USUARIO",
        "Anexar el comprobante de recibido firmado por el usuario en el archivo de epicrisis.",
    ),
    (
        "Falta la certificación de paciente por servicios prestados",
        r"CERTIFICACION DE PACIENTE|CERTIFICACION DEL PACIENTE",
        "Adjuntar la certificación de servicios prestados, firmada.",
    ),
    (
        "Sin soporte de envío de la factura electrónica",
        r"CORREO.*(FACTURACION|F\.?E)|NO SE EVIDENCIA CORREO|SOPORTE DE ENVIO DE LA FACTURACION"
        r"|FACTURACION ELECTRONICA",
        "Facturación debe enviar la factura desde el correo oficial y anexar el soporte.",
    ),
    (
        "Falta FURIPS o SIRAS",
        r"FURIPS|SIRAS|\bFUR\b",
        "Adjuntar el FURIPS y el reporte SIRAS en PDF, con la nomenclatura que exige ADRES.",
    ),
    (
        "Errores de fechas (ingreso, egreso, terapias)",
        r"ERROR EN LA[S]? FECHA|FECHA DE EGRESO|FECHA DE INGRESO"
        r"|FECHAS? .*(NO COINCIDE|NO CORRESPONDE)|INCONSISTENCIA EN LA FECHA",
        "Cuadrar las fechas entre la factura de venta, la historia y la certificación.",
    ),
    (
        "Falta la representación gráfica de la factura",
        r"REPRESENTACION GRAFICA",
        "Anexar la representación gráfica de la factura electrónica.",
    ),
    (
        "Falta orden médica o solicitud de servicios",
        r"ORDEN MEDICA|ORDEN DE SOLICITUD|SOLICITUD DE SERVICIO",
        "Anexar la orden médica que soporta el servicio facturado.",
    ),
    (
        "Falta historia clínica o epicrisis",
        r"HISTORIA CLINICA|EPICRISIS|\bHC\b|ARCHIVO EPI",
        "Completar el expediente clínico antes de entregar el paquete.",
    ),
    (
        "Certificado médico / tipo de evento",
        r"CERTIFICADO MEDICO|EVENTO TERRORISTA|ACCIDENTE DE TRANSITO",
        "Corregir el certificado: el tipo de evento debe coincidir con la atención.",
    ),
    (
        "Falta factura de compra / MAOS",
        r"FACTURA DE COMPRA|MAOS|OSTEOSINTESIS|MATERIAL",
        "Anexar la factura de compra del material de osteosíntesis.",
    ),
    (
        "Datos del FURIPS mal diligenciados (placa, campos)",
        r"CAMPO \d|PLACA|DILIGENCIAD",
        "Corregir en el FURIPS los campos que no coinciden con el evento.",
    ),
    (
        "Soportes ilegibles, con tachones o mal separados",
        r"TACHON|ILEGIBLE|SEPARAR LOS SOPORTE|INFOPOL",
        "Rehacer los soportes: legibles, sin tachones y separados como exige el proceso.",
    ),
    (
        "Falta RIPS o archivo JSON",
        r"\bRIPS\b|JSON",
        "Generar y anexar el RIPS en JSON para poder radicar.",
    ),
    (
        "Firmas o registro médico",
        r"FIRMA|REGISTRO MEDICO",
        "Firmar y sellar con registro médico los soportes que lo exigen.",
    ),
    (
        "Autorización, póliza o SOAT",
        r"AUTORIZACION|POLIZA|SOAT|AMPARO",
        "Verificar amparo y autorización antes de radicar.",
    ),
    (
        "Datos del paciente o documento",
        r"DOCUMENTO DE IDENTIDAD|NOMBRE DEL PACIENTE|CEDULA|NO CORRESPONDE AL PACIENTE",
        "Corregir la identificación del paciente en los soportes.",
    ),
    (
        "Validación con sistemas / facturación",
        r"VALIDACION|SISTEMAS|NO CUMPLE",
        "Revisar con sistemas la validación que está fallando.",
    ),
]
SIN_TEXTO = "Sin motivo escrito"
OTROS = "Otros motivos"
ACCIONES = {nombre: accion for nombre, _, accion in CAUSAS_DEVOLUCION}
ACCIONES[OTROS] = "Revisar caso por caso en la hoja DEVUELTAS AL DETALLE."
ACCIONES[SIN_TEXTO] = (
    "OJO: se devolvió sin dejar por escrito el motivo. Así no se puede sustentar ante la entidad."
)


def _sin_tildes(texto: str) -> str:
    t = unicodedata.normalize("NFD", (texto or "").upper())
    return "".join(c for c in t if unicodedata.category(c) != "Mn")


def causa_de(texto: str) -> str:
    """La causa de la devolución, deducida de lo que escribió el gestor."""
    t = _sin_tildes(texto)
    if not t.strip():
        return SIN_TEXTO
    for nombre, patron, _ in CAUSAS_DEVOLUCION:
        if re.search(patron, t):
            return nombre
    return OTROS


def _fecha(v) -> Optional[datetime]:
    if isinstance(v, datetime):
        return v.astimezone(TZ_BOGOTA).replace(tzinfo=None) if v.tzinfo else v
    return v if isinstance(v, date) else None


def _texto(v) -> str:
    return (str(v).strip() if v is not None else "") or ""


# ------------------------------------------------------------------
# Ayudas de formato (que el libro se vea como un informe, no como un volcado)
# ------------------------------------------------------------------
def _titulo(ws, texto: str, sub: str, ancho: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ancho)
    c = ws.cell(1, 1, texto)
    c.font = Font(name=FUENTE, size=15, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ancho)
    c = ws.cell(2, 1, sub)
    c.font = Font(name=FUENTE, size=9.5, italic=True, color="404040")
    c.alignment = Alignment(indent=1, vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30


def _encabezado(ws, fila: int, cols: list[str], anchos: list[int]) -> None:
    borde = Border(bottom=Side(style="medium", color=AZUL))
    for i, v in enumerate(cols, 1):
        c = ws.cell(fila, i, v)
        c.font = Font(name=FUENTE, size=10, bold=True, color=AZUL)
        c.fill = PatternFill("solid", fgColor=AZUL_CLARO)
        c.border = borde
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[fila].height = 30
    for i, a in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = a


def _celda(
    ws, fila, col, valor, *, negrita=False, fmt=None, ajuste=False, centro=False, relleno=None
):
    c = ws.cell(fila, col, valor)
    c.font = Font(name=FUENTE, size=10, bold=negrita)
    c.alignment = Alignment(
        vertical="center" if not ajuste else "top",
        wrap_text=ajuste,
        horizontal="center" if centro else "general",
    )
    if fmt:
        c.number_format = fmt
    if relleno:
        c.fill = PatternFill("solid", fgColor=relleno)
    return c


# ------------------------------------------------------------------
# Los datos, en pocas consultas (nunca una por factura)
# ------------------------------------------------------------------
def _reunir(db: Session) -> list[dict]:
    filas = (
        db.query(FacturaPreauditoriaRecord, RadicacionCuentaRecord, DgReportRecord)
        .outerjoin(
            RadicacionCuentaRecord,
            RadicacionCuentaRecord.factura == FacturaPreauditoriaRecord.factura,
        )
        .outerjoin(DgReportRecord, DgReportRecord.factura == FacturaPreauditoriaRecord.factura)
        .order_by(FacturaPreauditoriaRecord.envio_actual, FacturaPreauditoriaRecord.factura)
        .all()
    )
    escribio: dict[str, tuple] = {}
    audito: dict[str, tuple] = {}
    motivo: dict[str, str] = {}
    decision = {"RADICADA", "DEVUELTA", "SUBSANADA", "NUEVAMENTE_DEVUELTA"}
    for ev in (
        db.query(FacturaEventoRecord)
        .order_by(FacturaEventoRecord.creado_en, FacturaEventoRecord.id)
        .all()
    ):
        if ev.tipo_evento == "ESCRITA" and ev.factura not in escribio:
            escribio[ev.factura] = (_texto(ev.auditor), _fecha(ev.creado_en))
        if ev.tipo_evento in decision:
            audito[ev.factura] = (_texto(ev.auditor), _fecha(ev.creado_en), ev.tipo_evento)
            if ev.motivo or ev.observaciones:
                motivo[ev.factura] = _texto(ev.motivo) or _texto(ev.observaciones)
    consecutivos = {d.id: d.consecutivo for d in db.query(OficioDevolucionRecord).all()}

    datos = []
    for f, rad, dg in filas:
        esc = escribio.get(f.factura, ("", None))
        aud = audito.get(f.factura, ("", None, ""))
        observado = _texto(f.motivo_ultima_devolucion) or motivo.get(f.factura, "")
        devuelta = f.resultado_actual == "DEVUELTA"
        datos.append(
            {
                "envio": f.envio_actual or (rad.envio if rad else ""),
                "oficio": _texto(f.oficio_fhus),
                "factura": f.factura,
                "f_recibido": _fecha(f.f_recibido) or _fecha(rad.f_recibido if rad else None),
                "f_factura": _fecha(rad.f_factura if rad else None),
                "entidad": _texto(rad.entidad if rad else ""),
                "nit": _texto(rad.nit if rad else ""),
                "valor": (rad.valor if rad else 0) or 0,
                "correo_fe": (dg.correo_fe if dg else "NO"),
                "estado": _texto(f.estado),
                "resultado": _texto(f.resultado_actual),
                "causa": causa_de(observado) if devuelta else "",
                "observado": observado,
                "observaciones": _texto(f.observaciones),
                "ronda": f.ronda_actual or 1,
                "subsanaciones": f.num_subsanacion or 0,
                "devoluciones": f.num_devoluciones or 0,
                "escribio": esc[0],
                "f_escribio": esc[1],
                "audito": _texto(f.auditor) or aud[0],
                "f_audito": _fecha(f.fecha_auditoria) or aud[1],
                "oficio_devolucion": consecutivos.get(f.oficio_devolucion_id, ""),
            }
        )
    return datos


# ------------------------------------------------------------------
# El libro
# ------------------------------------------------------------------
def construir_consolidado_excel(db: Session, generado_por: str = "") -> bytes:
    datos = _reunir(db)
    hoy = datetime.now(TZ_BOGOTA).strftime("%d/%m/%Y %H:%M")
    wb = Workbook()

    # ── FACTURAS: la hoja base, de la que salen todas las fórmulas ──
    fa = wb.active
    fa.title = "FACTURAS"
    _titulo(
        fa,
        "PRE-AUDITORÍA SINAC — UNA FILA POR FACTURA",
        "Todo lo que se sabe de cada factura: quién la recibió, qué se le observó, si se radicó o se "
        "devolvió, por qué y quién decidió. Use los filtros del encabezado para buscar por entidad, "
        "gestor, causa o resultado.",
        21,
    )
    cols = [
        ("Envío", 10),
        ("Oficio", 20),
        ("Factura", 17),
        ("Recibido", 12),
        ("Fecha factura", 13),
        ("Entidad", 30),
        ("NIT", 13),
        ("Valor", 14),
        ("Correo F.E.", 11),
        ("Estado", 26),
        ("Resultado", 13),
        ("Por qué se devolvió (causa)", 32),
        ("Lo que escribió el gestor", 62),
        ("Observaciones", 40),
        ("Ronda", 8),
        ("Subsanaciones", 12),
        ("Devoluciones", 12),
        ("Escribió el envío", 22),
        ("Fecha en que se escribió", 15),
        ("Auditó", 22),
        ("Oficio de devolución", 20),
    ]
    _encabezado(fa, 4, [c[0] for c in cols], [c[1] for c in cols])
    color = {"DEVUELTA": ROJO, "RADICAR": VERDE, "PENDIENTE": AMBAR}
    for n, d in enumerate(datos):
        f = 5 + n
        valores = [
            d["envio"],
            d["oficio"],
            d["factura"],
            d["f_recibido"],
            d["f_factura"],
            d["entidad"],
            d["nit"],
            d["valor"],
            d["correo_fe"],
            d["estado"],
            d["resultado"],
            d["causa"],
            d["observado"],
            d["observaciones"],
            d["ronda"],
            d["subsanaciones"],
            d["devoluciones"],
            d["escribio"],
            d["f_escribio"],
            d["audito"],
            d["oficio_devolucion"],
        ]
        for j, v in enumerate(valores, 1):
            c = fa.cell(f, j, v)
            c.font = Font(name=FUENTE, size=9, bold=(j == 11))
            c.alignment = Alignment(vertical="top", wrap_text=(j in (6, 12, 13, 14)))
            if j == 8:
                c.number_format = '"$"#,##0'
            if j in (4, 5, 19):
                c.number_format = "DD/MM/YYYY"
            if j in (15, 16, 17):
                c.alignment = Alignment(horizontal="center", vertical="top")
            if j == 11:
                c.fill = PatternFill("solid", fgColor=color.get(v, GRIS))
        fa.row_dimensions[f].height = 28
    ult = 4 + len(datos)
    fa.freeze_panes = "D5"
    if datos:
        fa.auto_filter.ref = f"A4:U{ult}"
    fin = max(ult, 5)
    RES = f"FACTURAS!$K$5:$K${fin}"
    CAU = f"FACTURAS!$L$5:$L${fin}"
    VAL = f"FACTURAS!$H$5:$H${fin}"
    ENT = f"FACTURAS!$F$5:$F${fin}"
    AUD = f"FACTURAS!$T$5:$T${fin}"

    devueltas = [d for d in datos if d["resultado"] == "DEVUELTA"]
    causas = [c for c, _ in Counter(d["causa"] for d in devueltas).most_common()]
    gestores = sorted({d["audito"] for d in datos if d["audito"]})
    entidades = [e for e, _ in Counter(d["entidad"] for d in datos if d["entidad"]).most_common(25)]

    # ── CÓMO LEER ──
    cl = wb.create_sheet("CÓMO LEER", 0)
    _titulo(cl, "CÓMO LEER ESTE INFORME", "Cada hoja responde una pregunta.", 4)
    cl.column_dimensions["A"].width = 3
    cl.column_dimensions["B"].width = 30
    cl.column_dimensions["C"].width = 86
    guia = [
        (
            "RESUMEN",
            "Los números gruesos y el gráfico de resultado. La primera página para presentar.",
        ),
        (
            "POR QUÉ SE DEVUELVEN",
            "Las devoluciones agrupadas por causa, con cuántas facturas, cuánta "
            "plata y qué hacer en cada una.",
        ),
        (
            "DEVUELTAS AL DETALLE",
            "Solo las devueltas, con el texto completo de lo observado. Es la "
            "hoja que se le pasa a Facturación.",
        ),
        ("POR GESTOR", "Cuántas auditó cada gestor, cuántas radicó, cuántas devolvió y su tasa."),
        ("POR ENTIDAD", "Lo mismo por aseguradora: quién nos frena más facturas y por cuánto."),
        (
            "FACTURAS",
            "Una fila por factura con TODO, incluida la observación del gestor. Los filtros "
            "del encabezado sirven para buscar.",
        ),
    ]
    _celda(cl, 4, 2, "LAS HOJAS", negrita=True).font = Font(
        name=FUENTE, size=12, bold=True, color=AZUL
    )
    fila = 5
    for hoja, desc in guia:
        fila += 1
        c = _celda(cl, fila, 2, hoja, negrita=True, relleno=AZUL_CLARO)
        c.font = Font(name=FUENTE, size=10.5, bold=True, color=AZUL)
        _celda(cl, fila, 3, desc, ajuste=True)
        cl.row_dimensions[fila].height = 40
    fila += 2
    _celda(cl, fila, 2, "OJO", negrita=True).font = Font(
        name=FUENTE, size=10, bold=True, color="9C0006"
    )
    _celda(
        cl,
        fila,
        3,
        "La causa de devolución se deduce del texto que escribió el gestor. Lo que no se puede "
        "clasificar queda como «Otros motivos», y lo que se devolvió sin escribir nada, como «Sin "
        "motivo escrito»: esas dos filas son en sí mismas un hallazgo.",
        ajuste=True,
    )
    cl.row_dimensions[fila].height = 46
    fila += 2
    _celda(cl, fila, 2, "Generado", negrita=True)
    _celda(cl, fila, 3, f"{hoy}" + (f" · por {generado_por}" if generado_por else ""))

    # ── RESUMEN ──
    rs = wb.create_sheet("RESUMEN", 1)
    _titulo(
        rs,
        "PRE-AUDITORÍA SINAC — RESUMEN",
        f"{len(datos)} factura(s) en el consolidado · generado el {hoy}"
        " · todos los números son fórmulas sobre la hoja FACTURAS",
        5,
    )
    rs.column_dimensions["A"].width = 3
    for col, an in zip("BCDE", (42, 16, 18, 56)):
        rs.column_dimensions[col].width = an
    _encabezado(rs, 4, ["", "Indicador", "Facturas", "Valor", "Qué significa"], [3, 42, 16, 18, 56])
    kpis = [
        (
            "Facturas en el consolidado",
            f"=COUNTA({RES})",
            f"=SUM({VAL})",
            "Todo lo que ha pasado por pre-auditoría.",
        ),
        (
            "Listas para radicar",
            f'=COUNTIF({RES},"RADICAR")',
            f'=SUMIF({RES},"RADICAR",{VAL})',
            "Pasaron la revisión y siguen su camino.",
        ),
        (
            "Devueltas a facturación",
            f'=COUNTIF({RES},"DEVUELTA")',
            f'=SUMIF({RES},"DEVUELTA",{VAL})',
            "Se frenaron por un soporte incompleto o un dato que no cuadra.",
        ),
        (
            "Pendientes de auditar",
            f'=COUNTIF({RES},"PENDIENTE")',
            f'=SUMIF({RES},"PENDIENTE",{VAL})',
            "Todavía no tienen respuesta.",
        ),
        (
            "Tasa de devolución",
            "=IFERROR(C7/C5,0)",
            "",
            "De cada 100 facturas revisadas, cuántas se devolvieron.",
        ),
        (
            "Devueltas sin motivo escrito",
            f'=COUNTIF({CAU},"{SIN_TEXTO}")',
            f'=SUMIF({CAU},"{SIN_TEXTO}",{VAL})',
            "Si este número sube, son devoluciones que no se pueden sustentar ante la entidad.",
        ),
    ]
    for n, (et, c2, c3, nota) in enumerate(kpis):
        f = 5 + n
        _celda(rs, f, 2, et, negrita=True)
        c = _celda(
            rs,
            f,
            3,
            c2,
            negrita=True,
            centro=True,
            relleno=GRIS,
            fmt="0.0%" if "Tasa" in et else "#,##0",
        )
        c.font = Font(name=FUENTE, size=11, bold=True)
        if c3:
            _celda(rs, f, 4, c3, fmt='"$"#,##0', relleno=GRIS).alignment = Alignment(
                horizontal="right", vertical="center"
            )
        _celda(rs, f, 5, nota, ajuste=True)
        rs.row_dimensions[f].height = 30
    f = 13
    _celda(rs, f, 2, "LAS TRES CAUSAS QUE MÁS DEVUELVEN", negrita=True).font = Font(
        name=FUENTE, size=12, bold=True, color=AZUL
    )
    for n, causa in enumerate(causas[:3]):
        f += 1
        _celda(rs, f, 2, causa, negrita=True)
        _celda(rs, f, 3, f"='POR QUÉ SE DEVUELVEN'!B{5 + n}", fmt="#,##0", centro=True)
        _celda(rs, f, 4, f"='POR QUÉ SE DEVUELVEN'!D{5 + n}", fmt='"$"#,##0')
        _celda(rs, f, 5, ACCIONES.get(causa, ""), ajuste=True)
        rs.row_dimensions[f].height = 30
    base = f + 3
    for n, (etiqueta, formula) in enumerate(
        (
            ("Listas para radicar", f'=COUNTIF({RES},"RADICAR")'),
            ("Devueltas", f'=COUNTIF({RES},"DEVUELTA")'),
            ("Pendientes", f'=COUNTIF({RES},"PENDIENTE")'),
        )
    ):
        _celda(rs, base + n, 2, etiqueta)
        _celda(rs, base + n, 3, formula, fmt="#,##0")
    torta = PieChart()
    torta.title = "Resultado de las facturas"
    torta.height, torta.width = 9, 13
    torta.add_data(Reference(rs, min_col=3, min_row=base, max_row=base + 2), titles_from_data=False)
    torta.set_categories(Reference(rs, min_col=2, min_row=base, max_row=base + 2))
    rs.add_chart(torta, f"B{base + 4}")

    # ── POR QUÉ SE DEVUELVEN ──
    pq = wb.create_sheet("POR QUÉ SE DEVUELVEN", 2)
    _titulo(
        pq,
        "POR QUÉ SE DEVUELVEN LAS FACTURAS",
        "Agrupadas por causa, deducida de lo que escribió el gestor. Fórmulas sobre la hoja FACTURAS.",
        6,
    )
    _encabezado(
        pq,
        4,
        [
            "Causa",
            "Facturas devueltas",
            "% de las devueltas",
            "Valor devuelto",
            "% del valor",
            "Qué hacer",
        ],
        [36, 16, 16, 18, 13, 56],
    )
    for n, causa in enumerate(causas):
        f = 5 + n
        c = _celda(pq, f, 1, causa, negrita=(causa == SIN_TEXTO), ajuste=True)
        if causa in (SIN_TEXTO, OTROS):
            c.fill = PatternFill("solid", fgColor=AMBAR)
        _celda(pq, f, 2, f"=COUNTIF({CAU},$A{f})", fmt="#,##0", centro=True)
        _celda(pq, f, 3, f"=IFERROR(B{f}/$B${5 + len(causas)},0)", fmt="0.0%", centro=True)
        _celda(pq, f, 4, f"=SUMIF({CAU},$A{f},{VAL})", fmt='"$"#,##0')
        _celda(pq, f, 5, f"=IFERROR(D{f}/$D${5 + len(causas)},0)", fmt="0.0%", centro=True)
        _celda(pq, f, 6, ACCIONES.get(causa, ""), ajuste=True)
        pq.row_dimensions[f].height = 32
    t = 5 + len(causas)
    _celda(pq, t, 1, "TOTAL DEVUELTAS", negrita=True, relleno=AZUL_CLARO)
    _celda(
        pq,
        t,
        2,
        f'=COUNTIF({RES},"DEVUELTA")',
        negrita=True,
        fmt="#,##0",
        centro=True,
        relleno=AZUL_CLARO,
    )
    _celda(pq, t, 3, "", relleno=AZUL_CLARO)
    _celda(
        pq,
        t,
        4,
        f'=SUMIF({RES},"DEVUELTA",{VAL})',
        negrita=True,
        fmt='"$"#,##0',
        relleno=AZUL_CLARO,
    )
    _celda(pq, t, 5, "", relleno=AZUL_CLARO)
    _celda(pq, t, 6, "", relleno=AZUL_CLARO)
    if causas:
        g = BarChart()
        g.type = "bar"
        g.title = "Facturas devueltas por causa"
        g.height, g.width = 11, 22
        g.add_data(Reference(pq, min_col=2, min_row=4, max_row=t - 1), titles_from_data=True)
        g.set_categories(Reference(pq, min_col=1, min_row=5, max_row=t - 1))
        pq.add_chart(g, f"A{t + 3}")

    # ── DEVUELTAS AL DETALLE ──
    dd = wb.create_sheet("DEVUELTAS AL DETALLE", 3)
    _titulo(
        dd,
        "DEVUELTAS: QUÉ HAY QUE CORREGIR, FACTURA POR FACTURA",
        "El texto completo de lo que observó el gestor. Esta es la hoja que se le pasa a Facturación.",
        9,
    )
    _encabezado(
        dd,
        4,
        [
            "Causa",
            "Oficio",
            "Envío",
            "Factura",
            "Entidad",
            "Valor",
            "Lo que escribió el gestor",
            "Quién la auditó",
            "Oficio de devolución",
        ],
        [32, 20, 10, 17, 28, 14, 66, 22, 20],
    )
    for n, d in enumerate(
        sorted(devueltas, key=lambda x: (x["causa"], x["entidad"], x["factura"]))
    ):
        f = 5 + n
        for j, v in enumerate(
            [
                d["causa"],
                d["oficio"],
                d["envio"],
                d["factura"],
                d["entidad"],
                d["valor"],
                d["observado"],
                d["audito"],
                d["oficio_devolucion"],
            ],
            1,
        ):
            c = dd.cell(f, j, v)
            c.font = Font(name=FUENTE, size=9)
            c.alignment = Alignment(vertical="top", wrap_text=(j in (1, 5, 7)))
            if j == 6:
                c.number_format = '"$"#,##0'
        dd.row_dimensions[f].height = 30
    dd.freeze_panes = "A5"
    if devueltas:
        dd.auto_filter.ref = f"A4:I{4 + len(devueltas)}"

    # ── POR GESTOR ──
    pg = wb.create_sheet("POR GESTOR", 4)
    _titulo(pg, "QUÉ AUDITÓ CADA GESTOR", "Fórmulas sobre la hoja FACTURAS.", 6)
    _encabezado(
        pg,
        4,
        [
            "Gestor",
            "Facturas auditadas",
            "Listas para radicar",
            "Devueltas",
            "% devolución",
            "Valor auditado",
        ],
        [26, 16, 16, 13, 13, 18],
    )
    for n, g_ in enumerate(gestores):
        f = 5 + n
        _celda(pg, f, 1, g_)
        _celda(pg, f, 2, f"=COUNTIF({AUD},$A{f})", fmt="#,##0", centro=True)
        _celda(pg, f, 3, f'=COUNTIFS({AUD},$A{f},{RES},"RADICAR")', fmt="#,##0", centro=True)
        _celda(pg, f, 4, f'=COUNTIFS({AUD},$A{f},{RES},"DEVUELTA")', fmt="#,##0", centro=True)
        _celda(pg, f, 5, f"=IFERROR(D{f}/B{f},0)", fmt="0.0%", centro=True)
        _celda(pg, f, 6, f"=SUMIF({AUD},$A{f},{VAL})", fmt='"$"#,##0')
    tg = 5 + len(gestores)
    _celda(pg, tg, 1, "TOTAL", negrita=True, relleno=AZUL_CLARO)
    for col, letra in ((2, "B"), (3, "C"), (4, "D"), (6, "F")):
        _celda(
            pg,
            tg,
            col,
            f"=SUM({letra}5:{letra}{tg - 1})",
            negrita=True,
            centro=True,
            fmt='"$"#,##0' if col == 6 else "#,##0",
            relleno=AZUL_CLARO,
        )
    _celda(
        pg,
        tg,
        5,
        f"=IFERROR(D{tg}/B{tg},0)",
        negrita=True,
        fmt="0.0%",
        centro=True,
        relleno=AZUL_CLARO,
    )
    if gestores:
        g2 = BarChart()
        g2.type = "col"
        g2.grouping = "stacked"
        g2.overlap = 100
        g2.title = "Listas para radicar vs devueltas, por gestor"
        g2.height, g2.width = 10, 22
        g2.add_data(
            Reference(pg, min_col=3, max_col=4, min_row=4, max_row=tg - 1), titles_from_data=True
        )
        g2.set_categories(Reference(pg, min_col=1, min_row=5, max_row=tg - 1))
        pg.add_chart(g2, f"A{tg + 3}")

    # ── POR ENTIDAD ──
    pe = wb.create_sheet("POR ENTIDAD", 5)
    _titulo(
        pe, "QUÉ PASA CON CADA ENTIDAD", "Las 25 con más facturas, con su tasa de devolución.", 6
    )
    _encabezado(
        pe,
        4,
        ["Entidad", "Facturas", "Listas para radicar", "Devueltas", "% devolución", "Valor total"],
        [40, 12, 16, 12, 13, 18],
    )
    for n, e_ in enumerate(entidades):
        f = 5 + n
        _celda(pe, f, 1, e_)
        _celda(pe, f, 2, f"=COUNTIF({ENT},$A{f})", fmt="#,##0", centro=True)
        _celda(pe, f, 3, f'=COUNTIFS({ENT},$A{f},{RES},"RADICAR")', fmt="#,##0", centro=True)
        _celda(pe, f, 4, f'=COUNTIFS({ENT},$A{f},{RES},"DEVUELTA")', fmt="#,##0", centro=True)
        _celda(pe, f, 5, f"=IFERROR(D{f}/B{f},0)", fmt="0.0%", centro=True)
        _celda(pe, f, 6, f"=SUMIF({ENT},$A{f},{VAL})", fmt='"$"#,##0')

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
