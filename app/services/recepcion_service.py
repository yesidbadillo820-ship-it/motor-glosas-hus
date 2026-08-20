"""
RecepcionService
================

Procesa el archivo Excel enviado por el equipo de recepción de glosas.

Columnas esperadas (en cualquier orden, por nombre de encabezado):
    GESTOR
    FECHA DE ENTREGA
    FECHA RADICACION          (cuando se radicó la factura a la EPS)
    FECHA DOCUMENTO DGH       (cuando la EPS emitió la glosa)
    FECHA RECEPCION           (cuando HUS recibió la glosa)
    ENTIDAD                   (nombre/código EPS)
    FACTURA
    CONSECUTIVO DGH           (identificador único de la glosa)
    VALOR GLOSA
    VENCE                     (fecha límite para responder)
    DEVOLUCION S/N
    DIAS RADICACION VS RECEPCION
    RADICADO                  (texto libre; si contiene "RATIFICADA", se aplica automáticamente el texto de respuesta para ratificadas)

Para cada fila:
- Si la glosa es RATIFICADA -> estado=RATIFICADA + dictamen con TEXTO_RATIFICADA.
- Si fue glosada extemporáneamente (>20 días hábiles entre radicación y DGH) ->
  estado=EXTEMPORANEA + dictamen con el texto estándar.
- Se calcula el semáforo por días hábiles restantes hasta VENCE.
- Upsert por CONSECUTIVO DGH (o por factura si no viene consecutivo).
"""

from __future__ import annotations

import re
from datetime import datetime, date, timedelta
from io import BytesIO
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.logging_utils import logger
from app.models.db import GlosaRecord, ConceptoGlosaRecord
from app.services.glosa_service import (
    FERIADOS_CO,
    DIAS_HABILES_LIMITE_EXTEMPORANEA,
    TEXTO_RATIFICADA,
    generar_texto_extemporanea,
)


# ─── Parámetros de semáforo (días hábiles restantes) ─────────────────────────
SEMAFORO_VERDE_MIN = 11  # >10 días
SEMAFORO_AMARILLO_MIN = 5  # 5-10 días
# <5 días → ROJO; <=0 → NEGRO


# ─── Mapeo de columnas del Excel -> campo interno ────────────────────────────
# Soporta dos hojas de cabecera:
#   • INICIAL:    GESTOR | FECHA DE ENTREGA | FECHA RADICACION | FECHA DOCUMENTO
#                  DGH | FECHA RECEPCION | ENTIDAD | FACTURA | ...
#   • RATIFICADA: RESPONSABLE | FECHA ENTREGA | FECHA DE DOCUMENTO (DGH) |
#                  FECHA NOTIFICACION OBJECION | EMPRESA | NUMERO DE FACTURA |
#                  FECHA VENCIMIENTO | OBSERVACION RECEPCION | ...
COLUMN_ALIASES: dict[str, list[str]] = {
    "gestor": ["gestor", "responsable"],
    "fecha_entrega": ["fecha de entrega", "fecha entrega"],
    "fecha_radicacion": ["fecha radicacion", "fecha de radicacion"],
    "fecha_documento_dgh": [
        "fecha documento dgh",
        "fecha dgh",
        "fecha de documento (dgh)",
        "fecha de documento dgh",
        "fecha documento (dgh)",
    ],
    "fecha_recepcion": [
        "fecha recepcion",
        "fecha de recepcion",
        "fecha notificacion objecion",
        "fecha de notificacion objecion",
    ],
    "entidad": ["entidad", "eps", "empresa"],
    "factura": ["factura", "numero de factura", "numero factura"],
    "consecutivo_dgh": ["consecutivo dgh", "consecutivo"],
    "valor_glosa": ["valor glosa", "valor"],
    "vence": ["vence", "fecha vence", "fecha vencimiento", "fecha de vencimiento"],
    "devolucion": ["devolucion s/n", "devolucion", "devolucion s", "s/n"],
    "dias_rad_rec": ["dias radicacion vs recepcion", "dias radicacion recepcion"],
    "radicado": ["radicado"],
    "referencia": ["referencia"],
    "observacion_tecnico": [
        "observacion tecnico",
        "observacion",
        "obs tecnico",
        "observacion recepcion",
        "observacion de recepcion",
    ],
    "tecnico_recepcion": [
        "tecnico que recepciono",
        "tecnico recepcion",
        "tecnico recepciono",
        "tecnico que recepciona",
    ],
    "tipo_glosa": ["tipo glosa", "tipo de glosa"],
    "profesional_medico": [
        "profesional(medico)",
        "profesional (medico)",
        "profesional medico",
        "profesional",
        "medico auditor",
    ],
}

# ─── Columnas de las hojas DETALLE (I / R) del DGH ───────────────────────────
# El DGH exporta los conceptos por factura en hojas con nombres literales
# "I" (Glosa_Inicial) y "R" (Glosa_Ratificada). Estas columnas son las que
# usa el parser de conceptos (procesar_hoja_conceptos).
CONCEPTO_COLS: dict[str, list[str]] = {
    "estado_dgh": ["estadocxcobjecion"],
    "tipo_tramite": ["tipoobjeciontramite"],
    "factura": ["facturacartera.factura"],
    "consecutivo": ["consecutivo"],
    "valor_factura": ["facturacartera.valor"],
    "saldo_factura": ["facturacartera.saldo"],
    "fecha_documento": ["fechadocumento"],
    "fecha_objecion": ["fechaobjecion"],
    "eps_plan": ["facturacartera.planbeneficio.codigonombreplanbeneficios"],
    "eps_codigo_entidad": ["facturacartera.planbeneficio.contrato.entidad.codigoentidad"],
    "eps_nombre": ["facturacartera.planbeneficio.contrato.entidad.nombreentidad"],
    "tercero_nit": ["facturacartera.tercero.documento"],
    # Nombre comercial corto de la entidad (Tercero.NombreCompletoNA).
    # Ej: "DISPENSARIO MEDICO BUCARAMANGA" en vez del plan EPS completo.
    "tercero_nombre": [
        "facturacartera.tercero.nombrecompletona",
        "facturacartera.tercero.nombrecompletoan",
        "facturacartera.tercero.nombre",
    ],
    "concepto_codigo": ["listadoconceptos.conceptoobjecion.codigo"],
    "concepto_oid": ["listadoconceptos.oid"],
    "concepto_nombre": ["listadoconceptos.conceptoobjecion.nombre"],
    "cups_codigo": ["listadoconceptos.servicioproductofactura.codigo"],
    "cups_descripcion": ["listadoconceptos.servicioproductofactura.descripcion"],
    "concepto_valor": ["listadoconceptos.valorobjecion"],
    "centro_costo": ["listadoconceptos.servicioproductofactura.centrocosto.codigonombrecentro"],
    "concepto_observacion": ["listadoconceptos.observaciones"],
}


def _normalizar(texto: str) -> str:
    import unicodedata

    t = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", t).strip().lower()


# ─── Texto de glosa compilado desde los conceptos (hojas I/R) ────────────────
# La hoja INICIAL/RATIFICADA no trae el "por qué" de la glosa; ese texto vive
# en las hojas I/R (ListadoConceptos.Observaciones + ConceptoObjecion.Nombre).
# Compilamos esos conceptos en texto_glosa_original para que la IA del
# auto-responder LEA EL CONCEPTO REAL de la EPS (antes recibía el dictamen
# placeholder "Pendiente de análisis..." — bug reportado 2026-06-10).
MARCA_TEXTO_CONCEPTOS = "[CONCEPTOS GLOSADOS"
_MAX_TEXTO_CONCEPTOS = 30_000  # tope < GlosaInput.tabla_excel (50K)


def _fmt_cop(valor) -> str:
    try:
        return f"${int(round(float(valor or 0))):,}".replace(",", ".")
    except (TypeError, ValueError):
        return "$0"


def _plan_de(
    *,
    codigo_glosa,
    tipo_glosa,
    dias_restantes,
    dias_radicacion,
    estado,
    valor,
    profesional_medico,
) -> dict:
    """El plan de trabajo de una glosa, en forma de diccionario simple.

    Se devuelve como dict —no como objeto— porque este resumen viaja al correo
    y a la pantalla tal cual, sin pasar por la base de datos.
    """
    from app.services.glosa_service import TEXTO_RATIFICADA
    from app.services.plan_de_trabajo import construir_plan

    p = construir_plan(
        codigo_glosa=codigo_glosa,
        tipo_glosa=tipo_glosa,
        dias_restantes=dias_restantes,
        dias_radicacion=dias_radicacion,
        ratificada=(estado == "RATIFICADA"),
        extemporanea=(estado == "EXTEMPORANEA"),
        valor_objetado=valor,
        profesional_medico=profesional_medico,
        texto_ratificada=TEXTO_RATIFICADA,
    )
    return {
        "prioridad": p.prioridad,
        "urgencia": p.urgencia,
        "titular": p.titular,
        "ruta": p.ruta,
        "respuesta_sugerida": p.respuesta_sugerida,
        "avisos": p.avisos,
        "con_medico": p.con_medico,
        "texto_listo": p.texto_listo,
    }


def _dias_desde_texto(vence: str | None):
    """«03/09/2026» → días que faltan. None si no se puede leer."""
    from datetime import date, datetime

    if not vence or vence == "N/A":
        return None
    try:
        d = datetime.strptime(str(vence), "%d/%m/%Y").date()
    except ValueError:
        return None
    return (d - date.today()).days


def componer_texto_desde_conceptos(db, glosa) -> str:
    """Compila el texto de la glosa desde sus ConceptoGlosaRecord.

    Devuelve "" si la glosa no tiene conceptos vinculados. El texto
    empieza con MARCA_TEXTO_CONCEPTOS para distinguirlo de texto
    cargado manualmente (nunca se pisa texto manual al reimportar).
    """
    if glosa is None or glosa.id is None:
        return ""
    conceptos = (
        db.query(ConceptoGlosaRecord)
        .filter(ConceptoGlosaRecord.glosa_id == glosa.id)
        .order_by(ConceptoGlosaRecord.id)
        .all()
    )
    if not conceptos:
        return ""
    bloques = [
        f"{MARCA_TEXTO_CONCEPTOS} — DETALLE DGH] Factura {glosa.factura or 's/n'} — "
        f"{len(conceptos)} concepto(s) objetado(s) por la EPS:"
    ]
    for i, c in enumerate(conceptos, start=1):
        encabezado = f"CONCEPTO {i}: {c.codigo_glosa or 'sin código'}"
        if c.nombre_glosa:
            encabezado += f" — {c.nombre_glosa}"
        lineas = [encabezado]
        servicio = " — ".join(x for x in (c.cups_codigo, c.cups_descripcion) if x)
        if servicio:
            lineas.append(f"Servicio/CUPS: {servicio}")
        if c.centro_costo:
            lineas.append(f"Centro de costo: {c.centro_costo}")
        if c.valor_objetado:
            lineas.append(f"Valor objetado: {_fmt_cop(c.valor_objetado)}")
        if c.observacion_eps:
            lineas.append(f"Observación EPS: {c.observacion_eps}")
        bloques.append("\n".join(lineas))
    texto = "\n\n".join(bloques)
    if len(texto) > _MAX_TEXTO_CONCEPTOS:
        texto = texto[:_MAX_TEXTO_CONCEPTOS] + "\n[... truncado]"
    return texto


def _fix_mojibake(texto: str) -> str:
    """Arregla texto UTF-8 leído como Latin-1 (mojibake) y limpia artefactos.

    - Mojibake: "OBJECIÃ³N" → "OBJECIÓN".
    - Artefactos Excel: "_x000D_" (Windows CRLF escapado por openpyxl) → " ".
    - Multiples espacios/saltos: colapsados a un solo espacio.
    - Truncamientos del DGH: nombres de entidad cortados a la mitad
      ("BUCARAMANG" → "BUCARAMANGA", "MANIZAL" → "MANIZALES", etc.).
    """
    if not texto or not isinstance(texto, str):
        return texto
    # 1. Fix mojibake latin1/utf8 si aplica
    if "Ã" in texto or "Â" in texto:
        try:
            texto = texto.encode("latin1", errors="strict").decode("utf8", errors="strict")
        except (UnicodeDecodeError, UnicodeEncodeError):
            pass
    # 2. Limpiar artefactos comunes de export desde Excel
    # openpyxl a veces deja literal "_x000D_" donde había \r (retorno de carro).
    texto = texto.replace("_x000D_", " ").replace("_x000A_", " ")
    # Quitar saltos de línea intermedios y espacios redundantes
    texto = re.sub(r"\s+", " ", texto).strip()
    # 3. Restaurar nombres de ciudad truncados que el DGH manda cortados.
    # El sistema DGH usa columnas de ancho fijo y trunca la cola de la
    # entidad. Mapeo inverso de fragmentos conocidos → nombre completo.
    _CIUDADES_TRUNCADAS = {
        "BUCARAMANG": "BUCARAMANGA",
        "FLORIDABLAN": "FLORIDABLANCA",
        "BARRANCABERMEJ": "BARRANCABERMEJA",
        "PIEDECUEST": "PIEDECUESTA",
        "MANIZAL": "MANIZALES",
        "VALLEDUPA": "VALLEDUPAR",
        "CARTAGEN": "CARTAGENA",
    }
    for cortado, completo in _CIUDADES_TRUNCADAS.items():
        # Solo reemplazar como palabra final (no parcial dentro de otra)
        texto = re.sub(rf"\b{cortado}\b(?!A)", completo, texto)
    return texto


# ─── Resolución GESTOR (Excel) → UsuarioRecord (asignación automática) ───────
# La columna GESTOR de la hoja INICIAL trae el nombre de pila del gestor
# ("YESID PEREZ"). Para que la glosa aparezca en "Mis glosas" del usuario
# correcto hay que setear auditor_email (gestor_nombre solo matchea por
# igualdad EXACTA con UsuarioRecord.nombre — tildes o mayúsculas distintas
# rompen la visibilidad). Matching tolerante: sin tildes/mayúsculas,
# containment, subconjunto de tokens y local-part del email
# ("yesid.perez@hus.com" ≈ "YESID PEREZ").


def _tokens_nombre(texto: str) -> frozenset[str]:
    return frozenset(t for t in _normalizar(texto or "").split(" ") if len(t) >= 2)


def construir_indice_usuarios(db) -> list[dict]:
    """Índice en memoria de usuarios activos para resolver gestores."""
    try:
        from app.models.db import UsuarioRecord

        usuarios = db.query(UsuarioRecord).filter(UsuarioRecord.activo == 1).all()
    except Exception as e:
        logger.warning(f"No se pudo cargar usuarios para asignar gestores: {e}")
        return []
    indice: list[dict] = []
    for u in usuarios:
        email = (u.email or "").strip().lower()
        if not email:
            continue
        local_como_nombre = re.sub(r"[._\-+]+", " ", email.split("@", 1)[0])
        indice.append(
            {
                "usuario": u,
                "email": email,
                "nombre_norm": _normalizar(u.nombre or ""),
                "tokens_nombre": _tokens_nombre(u.nombre or ""),
                "email_local_norm": _normalizar(local_como_nombre),
                "tokens_email": _tokens_nombre(local_como_nombre),
            }
        )
    return indice


def _email_con_delegacion(usuario, email: str) -> str:
    """Si el usuario está de vacaciones con delegado configurado, la
    asignación automática se redirige al delegado (mismo criterio que
    /glosas/bulk/asignar — ver UsuarioRecord.delega_a_email)."""
    try:
        vd = usuario.vacaciones_desde
        vh = usuario.vacaciones_hasta
        delegado = (usuario.delega_a_email or "").strip().lower()
        if not (vd and vh and delegado):
            return email
        from datetime import timezone as _tz

        ahora = datetime.now(_tz.utc)
        try:
            en_vacaciones = vd <= ahora <= vh
        except TypeError:
            # BD con datetimes naive (SQLite) — comparar naive vs naive
            en_vacaciones = vd <= datetime.now() <= vh
        if en_vacaciones:
            logger.info(
                f"Gestor {email} en vacaciones — asignación redirigida al delegado {delegado}"
            )
            return delegado
    except Exception as e:
        logger.debug(f"Chequeo de delegación falló para {email}: {e}")
    return email


def resolver_gestor_a_email(nombre_gestor: str, indice: list[dict]) -> tuple[Optional[str], str]:
    """Resuelve el nombre de gestor del Excel a un email de usuario.

    Devuelve (email | None, motivo) con motivo en:
      "exacto"    — nombre normalizado idéntico
      "parcial"   — containment o tokens del más corto ⊆ tokens del más largo
      "email"     — match contra el local-part del email
      "vacio"     — celda vacía / "SIN ASIGNAR"
      "ambiguo"   — varios usuarios distintos matchean (ej. equipos que
                    comparten nombre): NO se asigna a uno al azar; el
                    correo grupal del Excel-respuesta ya los cubre.
      "sin_match" — ningún usuario activo coincide (se reporta como
                    advertencia en el resumen de importación).
    """
    g_norm = _normalizar(nombre_gestor or "")
    if not g_norm or g_norm == "sin asignar":
        return None, "vacio"
    g_tokens = _tokens_nombre(nombre_gestor)

    def _matchea(e: dict) -> str | None:
        n = e["nombre_norm"]
        if n:
            if n == g_norm:
                return "exacto"
            if g_norm in n or n in g_norm:
                return "parcial"
            tn = e["tokens_nombre"]
            if g_tokens and tn and (g_tokens <= tn or tn <= g_tokens):
                return "parcial"
        el = e["email_local_norm"]
        if el:
            if el == g_norm or g_norm in el:
                return "email"
            te = e["tokens_email"]
            if g_tokens and te and g_tokens <= te:
                return "email"
        return None

    matches = [(e, m) for e in indice if (m := _matchea(e))]
    if not matches:
        return None, "sin_match"
    emails_distintos = {e["email"] for e, _ in matches}
    if len(emails_distintos) > 1:
        # Desempate: si hay UN solo match exacto, gana (los parciales
        # suelen ser homónimos/equipos).
        exactos = [(e, m) for e, m in matches if m == "exacto"]
        if len({e["email"] for e, _ in exactos}) == 1:
            e, _ = exactos[0]
            return _email_con_delegacion(e["usuario"], e["email"]), "exacto"
        return None, "ambiguo"
    e, motivo = matches[0]
    return _email_con_delegacion(e["usuario"], e["email"]), motivo


def _split_entidad(entidad: str) -> tuple[str, str]:
    """Separa 'U220181 - FAMISANAR EPS SUBSIDIADO' en ('U220181', 'FAMISANAR EPS SUBSIDIADO').

    Si no hay guion, el código queda vacío y todo va al nombre.
    """
    if not entidad:
        return "", ""
    m = re.match(r"^\s*([A-Z]\d{5,8})\s*[-–—]\s*(.+)$", entidad.strip())
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return "", entidad.strip()


# Ronda 10 (17-jun-2026) — el DGH manda la entidad en formato Syscafe muy
# verboso ("C260043 - ENTIDAD PROMOTORA DE SALUD FAMISANAR S A S
# CONTRIBUTIVO", "U240061 - FIDEICOMISOS PATRIMONIOS AUTONOMOS FIDUCIARIA
# LA PREVISORA S.A. FOMAG"). Antes guardábamos esa cadena cruda como
# `eps`, y la IA la copiaba al dictamen como cabecera ("ESE HUS NO ACEPTA
# GLOSA APLICADA POR C260043 - ENTIDAD PROMOTORA..."). Normalizamos a la
# clave canónica del catálogo CONTRATOS_HUS para que la IA use "FAMISANAR
# EPS", "FOMAG", "MUTUAL SER EPS" — los mismos nombres que conocen sus
# bloques de contexto contractual.
_TOKENS_EPS_CANONICA: tuple[tuple[str, str], ...] = (
    # (token de búsqueda en mayúsculas, nombre canónico final).
    # Orden importa: el más específico primero (NUEVA EPS antes que SOLO
    # "EPS", FOMAG/MAGISTERIO antes que FIDEICOMISOS).
    ("FAMISANAR", "FAMISANAR EPS"),
    ("FOMAG", "FOMAG"),
    ("MAGISTERIO", "FOMAG"),
    ("FIDUPREVISORA", "FOMAG"),
    ("NUEVA EPS", "NUEVA EPS"),
    ("COOSALUD", "COOSALUD"),
    ("COMPENSAR", "COMPENSAR"),
    ("POSITIVA", "POSITIVA"),
    ("SANITAS", "SANITAS"),
    ("MUTUAL SER", "MUTUAL SER EPS"),
    ("MUTUALSER", "MUTUAL SER EPS"),
    ("SALUD TOTAL", "SALUD TOTAL EPS"),
    ("SALUDTOTAL", "SALUD TOTAL EPS"),
    ("SURA", "SURA EPS"),
    ("ECOOPSOS", "ECOOPSOS"),
    ("EMSSANAR", "EMSSANAR"),
    ("ASMET SALUD", "ASMET SALUD"),
    ("ASMETSALUD", "ASMET SALUD"),
    ("CAPITAL SALUD", "CAPITAL SALUD EPS"),
    ("CAPRESOCA", "CAPRESOCA EPS"),
    ("DUSAKAWI", "DUSAKAWI EPS"),
    ("PIJAOS", "PIJAOS SALUD EPS"),
    ("MALLAMAS", "MALLAMAS EPS"),
    ("DMBUG", "DMBUG"),
    # Ronda 10 (17-jun-2026) — entidades de las fuerzas militares NO se
    # canonizan: "DIRECCION DE SANIDAD EJERCITO - DISPENSARIO MEDICO" pierde
    # "EJERCITO" si colapsamos a "DISPENSARIO MEDICO". Idem POLICIA / ARMADA
    # / SANIDAD MILITAR — cada una tiene su régimen normativo propio y el
    # dispensario específico es información importante. Para esos casos el
    # fallback (strip del prefijo Syscafe) preserva el nombre completo.
    ("CLINICA CHICAMOCHA", "CLINICA CHICAMOCHA"),
)


def _normalizar_eps_canonica(entidad_raw: str) -> str:
    """Devuelve el nombre canónico corto si reconoce la entidad, o el
    nombre limpio (sin código Syscafe) en otro caso.

    Caso real prod 17-jun: "C260043 - ENTIDAD PROMOTORA DE SALUD FAMISANAR
    S A S CONTRIBUTIVO" → "FAMISANAR EPS".
    "U240061 - FIDEICOMISOS PATRIMONIOS AUTONOMOS FIDUCIARIA LA PREVISORA
    S.A.  FOMAG" → "FOMAG".
    "U220251 - MUTUAL SER EPS" → "MUTUAL SER EPS".
    """
    if not entidad_raw:
        return ""
    _, nombre = _split_entidad(entidad_raw)
    nombre_up = (nombre or entidad_raw).upper()
    nombre_norm = re.sub(r"\s+", " ", nombre_up).strip()
    for token, canonico in _TOKENS_EPS_CANONICA:
        if token in nombre_norm:
            return canonico
    return nombre_norm or entidad_raw.strip()


def _mapear_cabeceras(
    fila_encabezado: tuple, mapa: dict[str, list[str]] | None = None
) -> dict[str, int]:
    """Devuelve {nombre_interno: índice_columna}.

    Por defecto usa COLUMN_ALIASES (hojas INICIAL/RATIFICADA). Pasa
    ``mapa=CONCEPTO_COLS`` para parsear hojas I/R de detalle.
    """
    mapa = mapa if mapa is not None else COLUMN_ALIASES
    indices: dict[str, int] = {}
    for idx, celda in enumerate(fila_encabezado):
        valor = _normalizar(str(celda or ""))
        if not valor:
            continue
        for nombre_interno, aliases in mapa.items():
            if valor in aliases and nombre_interno not in indices:
                indices[nombre_interno] = idx
                break
    return indices


def _buscar_fila_encabezado(
    ws, max_filas: int, mapa: dict[str, list[str]], min_aciertos: int = 3
) -> tuple[int, dict[str, int]]:
    """Busca la primera fila que parezca encabezado.

    Escanea hasta ``max_filas`` filas y devuelve (num_fila_1based, indices).
    Si ninguna fila tiene al menos ``min_aciertos`` columnas mapeadas,
    devuelve (0, {}).
    """
    for num_fila, fila in enumerate(ws.iter_rows(values_only=True), start=1):
        if num_fila > max_filas:
            break
        if all(c is None or str(c).strip() == "" for c in fila):
            continue
        indices = _mapear_cabeceras(fila, mapa)
        if len(indices) >= min_aciertos:
            return num_fila, indices
    return 0, {}


def _a_fecha(valor) -> Optional[datetime]:
    """Acepta datetime, date, o string con varios formatos comunes."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, date):
        return datetime(valor.year, valor.month, valor.day)
    # Serial de Excel (días desde 1899-12-30) que llegó como número o
    # como texto numérico — frecuente cuando la celda perdió formato.
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        try:
            if 1 < float(valor) < 100000:
                return datetime(1899, 12, 30) + timedelta(days=int(valor))
        except (ValueError, OverflowError):
            return None
    s = str(valor).strip()
    if not s:
        return None
    s_norm = re.sub(r"[.\s-]+", "/", s)
    for fmt in (
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%y",
        "%m/%d/%Y",
        "%Y/%m/%d",
        "%d-%m-%y",
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    # Segundo intento con separadores normalizados (ej. "2025.01.15",
    # "15 01 2025" → "15/01/2025").
    for fmt in ("%d/%m/%Y", "%Y/%m/%d", "%d/%m/%y"):
        try:
            return datetime.strptime(s_norm, fmt)
        except ValueError:
            continue
    # ISO con 'T' (ej. "2025-01-15T00:00:00").
    try:
        return datetime.fromisoformat(s.replace("Z", "").split("+")[0])
    except ValueError:
        pass
    # Serial de Excel que llegó como texto numérico ("45672").
    if s.isdigit() and 1 < int(s) < 100000:
        try:
            return datetime(1899, 12, 30) + timedelta(days=int(s))
        except (ValueError, OverflowError):
            return None
    return None


def _a_float(valor) -> float:
    """Convierte un valor monetario a float preservando decimales y signo.

    Maneja formato colombiano ("1.234.567,89" → 1234567.89), formato US
    ("1,234,567.89"), símbolos de moneda, espacios y negativos (incluido
    paréntesis contable "(1.234,50)"). El bug previo borraba TODO lo que
    no fuera dígito → "1.234,50" se convertía en 123450 (×100, sin signo).
    """
    if valor is None or valor == "":
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)

    s = str(valor).strip()
    if not s:
        return 0.0

    negativo = s.startswith("-") or (s.startswith("(") and s.endswith(")"))
    # Dejar solo dígitos y separadores . ,
    s = re.sub(r"[^\d.,]", "", s)
    if not s:
        return 0.0

    tiene_punto = "." in s
    tiene_coma = "," in s
    if tiene_punto and tiene_coma:
        # El separador decimal es el que aparece más a la derecha.
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")  # 1.234.567,89
        else:
            s = s.replace(",", "")  # 1,234,567.89
    elif tiene_coma:
        # Una sola coma con 1-2 decimales → decimal; si no, miles.
        if s.count(",") == 1 and len(s.split(",")[1]) in (1, 2):
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    elif tiene_punto:
        # Un solo punto con 1-2 decimales → decimal; si no (o varios
        # puntos), son separadores de miles.
        if s.count(".") == 1 and len(s.split(".")[1]) in (1, 2):
            pass
        else:
            s = s.replace(".", "")

    try:
        n = float(s)
    except ValueError:
        return 0.0
    return -n if negativo else n


def _dias_habiles(desde: datetime, hasta: datetime) -> int:
    """Cuenta días hábiles (excluye sábados, domingos y festivos Colombia)."""
    if desde >= hasta:
        return 0
    dias = 0
    curr = desde
    while curr < hasta:
        curr += timedelta(days=1)
        if curr.weekday() < 5 and curr.strftime("%Y-%m-%d") not in FERIADOS_CO:
            dias += 1
    return dias


def _semaforo(dias_restantes: int) -> str:
    if dias_restantes <= 0:
        return "NEGRO"
    if dias_restantes < SEMAFORO_AMARILLO_MIN:
        return "ROJO"
    if dias_restantes < SEMAFORO_VERDE_MIN:
        return "AMARILLO"
    return "VERDE"


# Mapa de palabras clave del nombre de glosa → prefijo canónico (Res. 2284/2023).
# Si el DGH trae código numérico (Syscafe) sin el canónico, usamos el nombre
# para inferirlo. Los genéricos se eligen como punto de partida — el auditor
# puede ajustar el código específico (TA0201, TA0801, etc.) si hace falta.
_MAPEO_NOMBRE_A_PREFIJO = [
    ("AUTORIZACION", "AU0101"),
    ("AUTORIZACIÓN", "AU0101"),
    ("TARIFAS", "TA0201"),
    ("TARIFA", "TA0201"),
    ("SOPORTES", "SO0101"),
    ("SOPORTE", "SO0101"),
    ("COBERTURA", "CO0101"),
    ("PERTINENCIA", "PE0101"),
    ("FACTURACION", "FA0101"),
    ("FACTURACIÓN", "FA0101"),
    ("CALIDAD", "CL0101"),
    ("INCONSISTENCIA", "IN0101"),
    ("MEDICAMENTOS", "ME0101"),
    ("INSUMOS", "IN0201"),
    ("SERVICIO", "SE0101"),
]


def _inferir_codigo_canonico(nombre_glosa: str) -> str | None:
    """Dada 'AUTORIZACION - PROCEDIMIENTO' devuelve 'AU0101'. Si no detecta
    ningún concepto conocido, devuelve None y el caller usa el código
    numérico como fallback."""
    if not nombre_glosa:
        return None
    texto = nombre_glosa.upper()
    for keyword, codigo in _MAPEO_NOMBRE_A_PREFIJO:
        if keyword in texto:
            return codigo
    return None


def _es_ratificada(*valores: str) -> bool:
    """True si CUALQUIERA de los valores contiene la palabra RATIFICADA."""
    for v in valores:
        if v and "RATIFICADA" in str(v).upper():
            return True
    return False


def _no_aplicar_extemporaneidad(observacion: str) -> bool:
    """True si la observación del técnico pide no aplicar extemporaneidad."""
    if not observacion:
        return False
    texto = str(observacion).upper()
    return (
        "NO APLICAR EXTEMPORANEIDAD" in texto
        or "NO APLICA EXTEMPORANEIDAD" in texto
        or "NO APLICAR EXTEMPORANEA" in texto
    )


def _dictamen_ratificada(eps: str, factura: str, radicado_info: str) -> str:
    return f"""
    <div style="background:#ede9fe;border-left:4px solid #7c3aed;padding:20px;margin:15px 0;border-radius:8px;">
        <h4 style="color:#5b21b6;margin:0 0 10px 0;">RESPUESTA A GLOSA RATIFICADA</h4>
        <p style="font-size:12px;color:#6d28d9;margin:0 0 10px 0;">
            <b>EPS:</b> {eps} | <b>Factura:</b> {factura} | <b>Observación recepción:</b> {radicado_info}
        </p>
        <p style="font-size:13px;line-height:1.8;color:#4c1d95;white-space:pre-wrap;">{TEXTO_RATIFICADA}</p>
    </div>
    """.strip()


def _dictamen_extemporanea(eps: str, factura: str, dias_transcurridos: int) -> str:
    texto = generar_texto_extemporanea(dias_transcurridos)
    return f"""
    <div style="background:#fee2e2;border-left:4px solid #dc2626;padding:20px;margin:15px 0;border-radius:8px;">
        <h4 style="color:#991b1b;margin:0 0 10px 0;">GLOSA EXTEMPORÁNEA ({dias_transcurridos} DÍAS HÁBILES)</h4>
        <p style="font-size:12px;color:#b91c1c;margin:0 0 10px 0;">
            <b>EPS:</b> {eps} | <b>Factura:</b> {factura}
        </p>
        <p style="font-size:13px;line-height:1.8;color:#7f1d1d;white-space:pre-wrap;">{texto}</p>
    </div>
    """.strip()


class ResumenImportacion:
    def __init__(self):
        self.total = 0
        self.creadas = 0
        self.actualizadas = 0
        self.duplicadas = 0  # mismo (factura+consecutivo+valor+fecha) — se saltan
        self.ratificadas = 0
        self.extemporaneas = 0
        self.errores: list[str] = []
        self.duplicadas_detalle: list[dict] = []
        self.por_gestor: dict[str, list[dict]] = {}
        self.semaforo: dict[str, int] = {"VERDE": 0, "AMARILLO": 0, "ROJO": 0, "NEGRO": 0}
        # Conceptos (hojas I/R)
        self.conceptos_creados = 0
        self.conceptos_actualizados = 0
        self.conceptos_huerfanos: list[dict] = []  # sin GlosaRecord que los ancle
        # IDs de glosas creadas/actualizadas para auto-procesamiento
        # post-importación (cerebro IA en background).
        self.glosas_ids_para_auto_responder: list[int] = []
        # IDs de TODAS las glosas tocadas por la importación (creadas +
        # actualizadas + duplicados exactos). Se usa para generar y enviar
        # el Excel-respuesta aunque la importación sea una reimportación
        # 100% duplicada (en ese caso glosas_ids_para_auto_responder queda
        # vacío pero el gestor igual debe recibir el Excel anotado).
        self.glosas_ids_todas: list[int] = []
        # Filas válidas pero saltadas (sin ENTIDAD/FACTURA, fechas inválidas)
        # — antes se descartaban en silencio y el usuario veía total=0 sin
        # entender por qué. Guardamos motivo + muestra acotada.
        self.filas_omitidas: int = 0
        self.filas_omitidas_detalle: list[dict] = []
        # Hojas descartadas por no tener columnas reconocibles.
        self.hojas_descartadas: list[dict] = []
        # Asignación automática GESTOR → usuario (auditor_email).
        # gestores_asignados: {nombre_gestor: email_asignado}
        # gestores_sin_usuario: nombres que no matchearon ningún usuario
        # activo (las glosas quedan sin auditor_email y el coordinador ve
        # la advertencia en el resumen).
        # advertencias: avisos no-fatales (separados de `errores`, que
        # significa "filas/hojas que no se pudieron importar").
        self.gestores_asignados: dict[str, str] = {}
        self.gestores_sin_usuario: list[str] = []
        self.advertencias: list[str] = []

    def registrar_omitida(self, fila: int, motivo: str) -> None:
        self.filas_omitidas += 1
        if len(self.filas_omitidas_detalle) < 50:
            self.filas_omitidas_detalle.append({"fila": fila, "motivo": motivo})

    def to_dict(self) -> dict:
        return {
            "total": self.total,
            "creadas": self.creadas,
            "actualizadas": self.actualizadas,
            "duplicadas": self.duplicadas,
            "ratificadas": self.ratificadas,
            "extemporaneas": self.extemporaneas,
            "errores": self.errores,
            "duplicadas_detalle": self.duplicadas_detalle[:50],
            "por_gestor": self.por_gestor,
            "semaforo": self.semaforo,
            "conceptos_creados": self.conceptos_creados,
            "conceptos_actualizados": self.conceptos_actualizados,
            "conceptos_huerfanos": self.conceptos_huerfanos[:50],
            "filas_omitidas": self.filas_omitidas,
            "filas_omitidas_detalle": self.filas_omitidas_detalle,
            "hojas_descartadas": self.hojas_descartadas,
            "gestores_asignados": self.gestores_asignados,
            "gestores_sin_usuario": self.gestores_sin_usuario,
            "advertencias": self.advertencias,
        }


class RecepcionService:
    def __init__(self, db: Session):
        self.db = db
        # Índice de usuarios y cache de resolución gestor→email; viven lo
        # que dura la importación (una instancia por archivo procesado).
        self._indice_usuarios: list[dict] | None = None
        self._cache_gestor: dict[str, Optional[str]] = {}

    def _resolver_gestor(self, gestor: str, resumen: "ResumenImportacion") -> Optional[str]:
        """Email del usuario asignable para este gestor (o None).

        Cachea por nombre normalizado y registra UNA advertencia por
        gestor sin usuario para que el coordinador la vea en el resumen.
        """
        clave = _normalizar(gestor or "")
        if clave in self._cache_gestor:
            return self._cache_gestor[clave]
        if self._indice_usuarios is None:
            self._indice_usuarios = construir_indice_usuarios(self.db)
        email, motivo = resolver_gestor_a_email(gestor, self._indice_usuarios)
        if email:
            resumen.gestores_asignados[gestor] = email
            logger.info(f"Gestor '{gestor}' → {email} (match {motivo})")
        elif motivo == "sin_match":
            resumen.gestores_sin_usuario.append(gestor)
            resumen.advertencias.append(
                f"El gestor '{gestor}' no coincide con ningún usuario activo "
                f"del sistema — sus glosas quedan sin asignar (auditor_email "
                f"vacío). Cree el usuario o corrija el nombre en el Excel."
            )
            logger.warning(f"Gestor '{gestor}' sin usuario activo que matchee — sin asignar")
        elif motivo == "ambiguo":
            logger.info(
                f"Gestor '{gestor}' matchea varios usuarios distintos "
                f"(¿equipo?) — no se asigna auditor_email a uno solo"
            )
        self._cache_gestor[clave] = email
        return email

    def procesar_excel(self, contenido: bytes) -> ResumenImportacion:
        """Procesa el archivo Excel completo (múltiples hojas).

        Detecta automáticamente el tipo de cada hoja:
          • "RECEPCION"  — hojas INICIAL/RATIFICADA con encabezados de gestor+factura.
          • "CONCEPTOS"  — hojas I/R del DGH con columnas FacturaCartera.* y
                           ListadoConceptos.* (detalle por concepto).
          • "SALTAR"     — hoja vacía o sin columnas reconocibles.

        Orden garantizado: primero RECEPCION (crea/actualiza GlosaRecord),
        después CONCEPTOS (upsert sobre glosas existentes). Así los conceptos
        siempre encuentran su glosa padre. Conceptos huérfanos se reportan.
        """
        resumen = ResumenImportacion()
        try:
            wb = load_workbook(BytesIO(contenido), data_only=True, read_only=True)
        except Exception as e:
            resumen.errores.append(f"Archivo Excel inválido: {e}")
            return resumen

        hojas_disponibles = wb.sheetnames if wb.sheetnames else []
        if not hojas_disponibles:
            resumen.errores.append("El archivo no tiene hojas")
            return resumen

        hoy = datetime.now()

        # Clasificar hojas por tipo antes de procesar
        plan: list[tuple[str, str, int, dict]] = []  # (tipo, nombre, fila_header, indices)
        for nombre_hoja in hojas_disponibles:
            try:
                ws = wb[nombre_hoja]
            except KeyError:
                continue

            # Escaneo rápido de los primeros 5 encabezados para detectar tipo.
            # CONCEPTOS gana si aparecen columnas ListadoConceptos.*;
            # RECEPCION si aparecen factura+vence o factura+fecha_recepcion.
            # Ventana amplia (20 filas): los export reales del DGH ponen
            # una fila título ("ENTREGA GLOSA INICIAL") + filas de
            # formato/banner antes del encabezado real. Con max_filas=5 el
            # encabezado de la hoja INICIAL quedaba fuera de la ventana y
            # la hoja se descartaba entera → las glosas INICIAL (las
            # nuevas) NO se importaban. Las filas de datos no matchean ≥3
            # alias de encabezado, así que ampliar es seguro.
            fila_h_rec, idx_rec = _buscar_fila_encabezado(
                ws, max_filas=20, mapa=COLUMN_ALIASES, min_aciertos=3
            )
            fila_h_con, idx_con = _buscar_fila_encabezado(
                ws, max_filas=20, mapa=CONCEPTO_COLS, min_aciertos=4
            )

            if idx_con and "concepto_codigo" in idx_con and "factura" in idx_con:
                plan.append(("CONCEPTOS", nombre_hoja, fila_h_con, idx_con))
            elif idx_rec and {"factura", "vence"}.issubset(set(idx_rec.keys())):
                plan.append(("RECEPCION", nombre_hoja, fila_h_rec, idx_rec))
            else:
                detectadas = sorted(set((idx_rec or {}).keys()))
                faltan = sorted({"factura", "vence"} - set(detectadas))
                resumen.hojas_descartadas.append(
                    {
                        "hoja": nombre_hoja,
                        "columnas_detectadas": detectadas,
                        "columnas_faltantes": faltan,
                    }
                )
                resumen.errores.append(
                    f"Hoja '{nombre_hoja}' descartada: no se reconocieron "
                    f"las columnas mínimas (faltan: "
                    f"{', '.join(faltan) or 'encabezados no mapeables'}). "
                    f"Detectadas: {', '.join(detectadas) or 'ninguna'}."
                )
                logger.warning(
                    f"Hoja '{nombre_hoja}' sin columnas reconocibles — saltando (faltan: {faltan})"
                )

        # Procesar RECEPCION primero, CONCEPTOS después
        plan.sort(key=lambda p: 0 if p[0] == "RECEPCION" else 1)

        total_procesadas = 0
        for tipo, nombre_hoja, fila_header, indices in plan:
            ws = wb[nombre_hoja]
            # Re-iterar desde la fila siguiente al encabezado detectado
            filas = ws.iter_rows(values_only=True)
            for _ in range(fila_header):
                try:
                    next(filas)
                except StopIteration:
                    break

            hoja_es_ratificada = (
                "RATIFIC" in (nombre_hoja or "").upper() or nombre_hoja.strip().upper() == "R"
            )

            if tipo == "RECEPCION":
                logger.info(
                    f"Procesando hoja '{nombre_hoja}' como RECEPCION "
                    f"{'(RATIFICADAS)' if hoja_es_ratificada else '(INICIALES)'}"
                )
                self._procesar_filas_hoja(
                    filas=filas,
                    indices=indices,
                    resumen=resumen,
                    hoy=hoy,
                    hoja_es_ratificada=hoja_es_ratificada,
                    nombre_hoja=nombre_hoja,
                )
            else:  # CONCEPTOS
                logger.info(
                    f"Procesando hoja '{nombre_hoja}' como CONCEPTOS "
                    f"{'(RATIFICADOS)' if hoja_es_ratificada else '(INICIALES)'}"
                )
                self._procesar_filas_conceptos(
                    filas=filas,
                    indices=indices,
                    resumen=resumen,
                    nombre_hoja=nombre_hoja,
                )
            total_procesadas += 1

        if total_procesadas == 0:
            resumen.errores.append(
                "Ninguna hoja tiene columnas reconocibles. El parser busca hojas "
                "tipo RECEPCION (con FACTURA+VENCE) o CONCEPTOS (con ListadoConceptos.*)."
            )

        # Las hojas de resumen (INICIAL/RATIFICADA) no traen la causal: esa vive
        # en las hojas de detalle, que se leen después. Acá, ya con todo dentro,
        # se completa el plan de cada glosa con su causal real — que es lo que
        # convierte «causal por clasificar» en «Soportes: relacione los que SÍ
        # están en el expediente». 20-08-2026.
        try:
            self._completar_planes_con_causal(resumen)
        except Exception as e:  # noqa: BLE001 - el resumen vale igual sin esto
            logger.warning(f"No se pudo completar el plan con las causales: {e}")
        return resumen

    def _completar_planes_con_causal(self, resumen) -> None:
        """Rellena la causal en el plan de cada glosa del resumen.

        Se toma la causal del concepto de MÁS VALOR: si una factura viene
        glosada por soportes y por tarifa, la que manda para saber cómo
        trabajarla es la que tiene más plata detrás.
        """
        from app.models.db import ConceptoGlosaRecord, GlosaRecord

        ids = list(resumen.glosas_ids_todas or [])
        if not ids:
            return
        conceptos = (
            self.db.query(ConceptoGlosaRecord).filter(ConceptoGlosaRecord.glosa_id.in_(ids)).all()
        )
        # glosa_id -> causal del concepto de más valor
        mayor: dict[int, tuple[float, str]] = {}
        for c in conceptos:
            cod = (getattr(c, "codigo_glosa", "") or "").strip()
            if not cod:
                continue
            val = float(getattr(c, "valor_objetado", 0) or 0)
            if val >= mayor.get(c.glosa_id, (-1.0, ""))[0]:
                mayor[c.glosa_id] = (val, cod)

        # factura -> causal, para encontrarlas desde `por_gestor`
        por_factura: dict[str, str] = {}
        for g in self.db.query(GlosaRecord).filter(GlosaRecord.id.in_(ids)).all():
            dato = mayor.get(g.id)
            if dato and g.factura:
                por_factura[str(g.factura).strip().upper()] = dato[1]

        for glosas in (resumen.por_gestor or {}).values():
            for item in glosas:
                cod = por_factura.get(str(item.get("factura") or "").strip().upper())
                if not cod:
                    continue
                item["causal"] = cod
                item["plan"] = _plan_de(
                    codigo_glosa=cod,
                    tipo_glosa=item.get("tipo_glosa"),
                    dias_restantes=_dias_desde_texto(item.get("vence")),
                    dias_radicacion=None,
                    estado=item.get("estado"),
                    valor=item.get("valor"),
                    profesional_medico=None,
                )
                # No se pierden los avisos que ya se habían calculado con los
                # datos de la hoja de resumen (días de radicación, médico).
                previos = [a for a in (item.get("_avisos_previos") or [])]
                for a in previos:
                    if a not in item["plan"]["avisos"]:
                        item["plan"]["avisos"].append(a)

    def _procesar_filas_hoja(
        self,
        filas,
        indices: dict,
        resumen: "ResumenImportacion",
        hoy: datetime,
        hoja_es_ratificada: bool,
        nombre_hoja: str,
    ):
        """Procesa las filas de una hoja individual."""
        for num_fila, fila in enumerate(filas, start=2):
            if all(c is None or str(c).strip() == "" for c in fila):
                continue

            def _get(key: str):
                i = indices.get(key)
                return fila[i] if i is not None and i < len(fila) else None

            try:
                entidad_raw = _fix_mojibake(str(_get("entidad") or "").strip())
                entidad = entidad_raw.upper()
                factura = str(_get("factura") or "").strip()
                if not entidad or not factura:
                    # Antes: skip silencioso → el usuario veía total=0 sin
                    # saber por qué. Ahora se reporta (salvo fila vacía,
                    # ya filtrada arriba).
                    falta = []
                    if not entidad:
                        falta.append("ENTIDAD/EPS")
                    if not factura:
                        falta.append("FACTURA")
                    resumen.registrar_omitida(
                        num_fila,
                        f"sin {' y '.join(falta)}",
                    )
                    continue

                # Separar código plan (U220181) del nombre para normalización
                eps_codigo, eps_nombre_limpio = _split_entidad(entidad)
                # Ronda 10 (17-jun-2026) — el Syscafe-DGH manda nombres
                # verbosos ("C260043 - ENTIDAD PROMOTORA DE SALUD FAMISANAR
                # S A S CONTRIBUTIVO"). Antes guardábamos esa cadena cruda
                # como `eps` y la IA la copiaba al dictamen como cabecera.
                # Normalizamos a la clave canónica del catálogo (FAMISANAR
                # EPS, FOMAG, MUTUAL SER EPS) — los mismos nombres que
                # conocen sus bloques de contexto contractual y el resto
                # del catálogo.
                eps_canonica = _normalizar_eps_canonica(entidad) or entidad

                consecutivo = str(_get("consecutivo_dgh") or "").strip()
                gestor = str(_get("gestor") or "").strip().upper() or "SIN ASIGNAR"
                # Asignación automática: GESTOR del Excel → usuario activo
                # (auditor_email). Si no hay match queda sin asignar y se
                # reporta advertencia (una vez por gestor).
                auditor_email_asignado = self._resolver_gestor(gestor, resumen)
                radicado_info = str(_get("radicado") or "").strip()
                referencia = str(_get("referencia") or "").strip()
                observacion_tecnico = _fix_mojibake(str(_get("observacion_tecnico") or "").strip())
                tipo_glosa_excel = str(_get("tipo_glosa") or "").strip()
                profesional_medico = str(_get("profesional_medico") or "").strip()
                tecnico_recepcion = str(_get("tecnico_recepcion") or "").strip()
                devolucion = str(_get("devolucion") or "").strip().upper()[:1]

                fecha_entrega = _a_fecha(_get("fecha_entrega"))
                fecha_rad = _a_fecha(_get("fecha_radicacion"))
                fecha_dgh = _a_fecha(_get("fecha_documento_dgh"))
                fecha_rec = _a_fecha(_get("fecha_recepcion"))
                fecha_vence = _a_fecha(_get("vence"))
                valor = _a_float(_get("valor_glosa"))

                if fecha_vence is None or fecha_rec is None:
                    resumen.errores.append(f"Fila {num_fila}: fechas VENCE/RECEPCION inválidas")
                    continue

                # Flag del técnico para saltar extemporaneidad (ej. PPL/FOMAG con régimen especial)
                skip_extemporaneidad = _no_aplicar_extemporaneidad(observacion_tecnico)

                # Extemporaneidad: días hábiles entre FECHA RADICACION y FECHA DOCUMENTO DGH
                dias_transcurridos = 0
                es_extemporanea = False
                if fecha_rad and fecha_dgh:
                    dias_transcurridos = _dias_habiles(fecha_rad, fecha_dgh)
                    es_extemporanea = (
                        dias_transcurridos > DIAS_HABILES_LIMITE_EXTEMPORANEA
                        and not skip_extemporaneidad
                    )

                # Semáforo: días hábiles restantes hasta VENCE
                dias_restantes = _dias_habiles(hoy, fecha_vence) if fecha_vence > hoy else 0
                semaforo = _semaforo(dias_restantes)

                # Ratificación: la hoja entera puede ser de ratificaciones (nombre
                # "RATIFICADA") o bien detectarse fila a fila en RADICADO/REFERENCIA.
                ratificada = hoja_es_ratificada or _es_ratificada(radicado_info, referencia)

                # numero_radicado: si RADICADO no es un texto de ratificación, es el radicado real
                if ratificada:
                    numero_radicado_real = None
                else:
                    numero_radicado_real = radicado_info or None

                # requiere_ia: SOLO las glosas INICIAL pendientes necesitan
                # que el cerebro IA redacte el dictamen (usando el concepto
                # de las hojas I/R como contexto). RATIFICADA y EXTEMPORÁNEA
                # ya traen su TEXTO FIJO definitivo → mandarlas a la IA solo
                # quema tokens y tiempo (directiva Yesid 2026-05-19). Igual
                # se incluyen en glosas_ids_todas para que salgan en el
                # Excel-respuesta con su texto fijo.
                if ratificada:
                    estado = "RATIFICADA"
                    texto_ref = radicado_info or referencia
                    dictamen = _dictamen_ratificada(eps_canonica, factura, texto_ref)
                    resumen.ratificadas += 1
                    requiere_ia = False
                elif es_extemporanea:
                    estado = "EXTEMPORANEA"
                    dictamen = _dictamen_extemporanea(eps_canonica, factura, dias_transcurridos)
                    resumen.extemporaneas += 1
                    requiere_ia = False
                else:
                    estado = "RADICADA"
                    requiere_ia = True
                    nota_obs = (
                        (
                            f'<div style="margin-top:10px;padding:10px;background:#fef3c7;border-left:3px solid #eab308;border-radius:6px;font-size:12px">'
                            f"<b>⚠ Observación técnico:</b> {observacion_tecnico}</div>"
                        )
                        if observacion_tecnico
                        else ""
                    )
                    dictamen = (
                        f'<div style="padding:15px;background:#f8fafc;border-radius:8px;">'
                        f"<b>Glosa importada desde recepción.</b><br>"
                        f"Pendiente de análisis y respuesta por el gestor asignado."
                        f"{nota_obs}"
                        f"</div>"
                    )

                # Upsert por (factura + consecutivo_dgh) o solo factura si no hay consecutivo
                q = self.db.query(GlosaRecord).filter(GlosaRecord.factura == factura)
                if consecutivo:
                    q = q.filter(GlosaRecord.consecutivo_dgh == consecutivo)
                existente = q.first()

                campos = dict(
                    eps=eps_canonica,
                    eps_codigo=eps_codigo or None,
                    paciente="N/A",
                    factura=factura,
                    numero_radicado=(numero_radicado_real or None)
                    and str(numero_radicado_real)[:50],
                    consecutivo_dgh=consecutivo or None,
                    gestor_nombre=gestor,
                    tecnico_recepcion=tecnico_recepcion or None,
                    fecha_radicacion_factura=fecha_rad,
                    fecha_documento_dgh=fecha_dgh,
                    fecha_recepcion=fecha_rec,
                    fecha_entrega=fecha_entrega,
                    fecha_vencimiento=fecha_vence,
                    es_devolucion=devolucion or None,
                    radicado_info=radicado_info or None,
                    referencia=referencia or None,
                    observacion_tecnico=observacion_tecnico or None,
                    tipo_glosa_excel=tipo_glosa_excel or None,
                    profesional_medico=profesional_medico or None,
                    valor_objetado=valor,
                    valor_aceptado=0.0,
                    etapa="RESPUESTA A GLOSA",
                    estado=estado,
                    dictamen=dictamen,
                    dias_restantes=dias_restantes,
                    # Dias habiles FECHA RADICACION -> FECHA DOCUMENTO DGH (excl. findes/festivos).
                    # Es lo que el auditor ve en la columna "Dias" de Mis glosas, usado como
                    # indicador de extemporaneidad (si > 20, EPS gloso fuera de termino).
                    dias_radicacion_dgh=dias_transcurridos,
                    prioridad=semaforo,
                    workflow_state=estado,
                    modelo_ia="importacion_recepcion",
                )
                # Solo si se resolvió un usuario: así una reimportación sin
                # match NUNCA borra una asignación manual previa (la clave
                # ausente no entra al setattr del upsert).
                if auditor_email_asignado:
                    campos["auditor_email"] = auditor_email_asignado

                if existente:
                    # Detectar duplicado exacto (misma factura+consecutivo+valor+fecha)
                    es_duplicado_exacto = (
                        abs(float(existente.valor_objetado or 0) - float(valor)) < 0.01
                        and (existente.fecha_recepcion == fecha_rec)
                        and ((existente.consecutivo_dgh or "") == (consecutivo or ""))
                    )
                    if es_duplicado_exacto:
                        resumen.duplicadas += 1
                        resumen.duplicadas_detalle.append(
                            {
                                "fila": num_fila,
                                "factura": factura,
                                "consecutivo_dgh": consecutivo,
                                "valor": valor,
                                "glosa_existente_id": existente.id,
                                "motivo": "Misma factura + consecutivo + valor + fecha recepción",
                            }
                        )
                        if existente.id is not None:
                            resumen.glosas_ids_todas.append(existente.id)
                        continue
                    # Distinto en algún campo → actualizar (posible reimportación con correcciones)
                    for k, v in campos.items():
                        setattr(existente, k, v)
                    resumen.actualizadas += 1
                    if existente.id is not None:
                        if requiere_ia:
                            resumen.glosas_ids_para_auto_responder.append(existente.id)
                        resumen.glosas_ids_todas.append(existente.id)
                else:
                    nueva = GlosaRecord(**campos)
                    self.db.add(nueva)
                    self.db.flush()  # asignar nueva.id antes del commit final
                    resumen.creadas += 1
                    if nueva.id is not None:
                        if requiere_ia:
                            resumen.glosas_ids_para_auto_responder.append(nueva.id)
                        resumen.glosas_ids_todas.append(nueva.id)

                resumen.total += 1
                resumen.semaforo[semaforo] = resumen.semaforo.get(semaforo, 0) + 1
                resumen.por_gestor.setdefault(gestor, []).append(
                    {
                        "factura": factura,
                        "consecutivo_dgh": consecutivo,
                        "eps": eps_canonica,
                        "valor": valor,
                        "vence": fecha_vence.strftime("%d/%m/%Y"),
                        "fecha_entrega": fecha_entrega.strftime("%d/%m/%Y")
                        if fecha_entrega
                        else "N/A",
                        "semaforo": semaforo,
                        "estado": estado,
                        "tipo_glosa": tipo_glosa_excel or "-",
                        "radicado": numero_radicado_real or "-",
                        # 20-08-2026. El correo al gestor decía QUÉ llegó pero
                        # no QUÉ HACER. El plan sale de datos que este mismo
                        # archivo ya trae: causal, tipo de glosa, médico y los
                        # días entre radicación y notificación.
                        "plan": _plan_de(
                            codigo_glosa=None,
                            tipo_glosa=tipo_glosa_excel,
                            dias_restantes=dias_restantes,
                            dias_radicacion=dias_transcurridos,
                            estado=estado,
                            valor=valor,
                            profesional_medico=profesional_medico,
                        ),
                    }
                )

            except Exception as e:
                # Ronda 30: sanear la sesión — un flush fallido la deja
                # envenenada y TODAS las filas siguientes (y el commit final)
                # fallan, perdiendo el lote entero.
                try:
                    self.db.rollback()
                except Exception:
                    pass
                resumen.errores.append(f"Fila {num_fila}: {e}")
                logger.warning(f"Error procesando fila {num_fila}: {e}")
                continue

        try:
            self.db.commit()
        except Exception as e:
            logger.error(f"Error guardando importación: {e}")
            self.db.rollback()
            resumen.errores.append(f"Error al guardar: {e}")

        return resumen

    def _procesar_filas_conceptos(
        self,
        filas,
        indices: dict,
        resumen: "ResumenImportacion",
        nombre_hoja: str,
    ):
        """Procesa hoja de conceptos (I/R del DGH).

        Cada fila = 1 concepto asociado a una factura+consecutivo DGH. Se hace
        upsert contra la tabla ``conceptos_glosa`` usando ``ListadoConceptos.Oid``
        como clave de idempotencia. La glosa padre debe existir (cargada antes
        desde INICIAL/RATIFICADA); si no, el concepto se reporta como huérfano.
        """
        # Cache de la glosa padre por (factura, consecutivo). Una hoja I/R
        # trae DECENAS de filas-concepto por glosa; sin cache se hacían
        # 3 queries/fila + se re-logueaba el mismo fallback cientos de
        # veces, dejando la importación corriendo MINUTOS. Resolviendo
        # una sola vez por clave: O(filas) → unas pocas queries.
        _MISS = object()
        _padre_cache: dict = {}
        _logged_facturas: set = set()

        for num_fila, fila in enumerate(filas, start=2):
            if all(c is None or str(c).strip() == "" for c in fila):
                continue

            def _get(key: str):
                i = indices.get(key)
                return fila[i] if i is not None and i < len(fila) else None

            try:
                factura = str(_get("factura") or "").strip()
                consecutivo = str(_get("consecutivo") or "").strip()
                raw_codigo = str(_get("concepto_codigo") or "").strip().upper()
                oid = str(_get("concepto_oid") or "").strip()
                if not factura or not raw_codigo:
                    # Sin factura+código no es un concepto válido. El
                    # 'Consecutivo' puede venir vacío: exports DGH/PowerQuery
                    # del DMBUG ponen el consecutivo solo en la cabecera y
                    # dejan vacías las 277 filas de detalle por concepto.
                    # En ese caso el match a la glosa padre cae a "factura
                    # sola" (fallback 2 más abajo).
                    continue

                nombre_glosa = _fix_mojibake(str(_get("concepto_nombre") or "").strip())

                # Ronda 50 (Bug #4): distinguir código Syscafe numérico del
                # canónico Res. 2284/2023. Si viene puramente numérico (ej.
                # "423"), lo guardamos como codigo_syscafe y derivamos el
                # canónico desde el nombre del concepto. Si viene canónico
                # (TA0201, SO0101...), lo dejamos donde está.
                codigo_syscafe = None
                codigo_glosa = raw_codigo
                if raw_codigo.isdigit():
                    codigo_syscafe = raw_codigo
                    codigo_glosa = _inferir_codigo_canonico(nombre_glosa) or raw_codigo
                cups_codigo = str(_get("cups_codigo") or "").strip()
                cups_desc = _fix_mojibake(str(_get("cups_descripcion") or "").strip())
                centro_costo = _fix_mojibake(str(_get("centro_costo") or "").strip())
                observacion = _fix_mojibake(str(_get("concepto_observacion") or "").strip())
                valor_obj = _a_float(_get("concepto_valor"))

                # Buscar la glosa padre por factura + consecutivo.
                # Si no se encuentra con esa combinación exacta, intentar
                # fallback por solo FACTURA cuando el consecutivo del padre
                # esté vacío (caso típico: la hoja Recepción no tenía el
                # campo "consecutivo dgh" pero la hoja I/R sí lo trae).
                # En ese caso, completamos el consecutivo del padre con el
                # de esta fila — así el resto de la metadata (NIT, valor
                # factura, tercero, fecha objecion) sí se llena.
                _ck = (factura, consecutivo)
                glosa_padre = _padre_cache.get(_ck, _MISS)
                if glosa_padre is _MISS:
                    # Match exacto solo si la fila trae consecutivo. Si viene
                    # vacío (export DGH del DMBUG: cabecera tiene consec, las
                    # filas de detalle no), salteamos al fallback factura-sola.
                    glosa_padre = None
                    if consecutivo:
                        glosa_padre = (
                            self.db.query(GlosaRecord)
                            .filter(
                                GlosaRecord.factura == factura,
                                GlosaRecord.consecutivo_dgh == consecutivo,
                            )
                            .first()
                        )
                    if not glosa_padre and consecutivo:
                        # Fallback 1: factura + consecutivo NULL (parent
                        # creado sin consecutivo). Se lo seteamos ahora.
                        glosa_padre = (
                            self.db.query(GlosaRecord)
                            .filter(
                                GlosaRecord.factura == factura,
                                (GlosaRecord.consecutivo_dgh.is_(None))
                                | (GlosaRecord.consecutivo_dgh == ""),
                            )
                            .order_by(GlosaRecord.id.desc())
                            .first()
                        )
                        if glosa_padre:
                            glosa_padre.consecutivo_dgh = consecutivo
                            if factura not in _logged_facturas:
                                _logged_facturas.add(factura)
                                logger.info(
                                    f"[I/R] Fallback match: glosa_id={glosa_padre.id} "
                                    f"factura={factura} sin consecutivo previo, "
                                    f"se le asigna {consecutivo}"
                                )
                    if not glosa_padre:
                        # Fallback 2: factura sola (la más reciente). Si
                        # tiene consecutivo distinto, asumimos misma glosa
                        # y logueamos la divergencia (una vez por factura).
                        glosa_padre = (
                            self.db.query(GlosaRecord)
                            .filter(GlosaRecord.factura == factura)
                            .order_by(GlosaRecord.id.desc())
                            .first()
                        )
                        if glosa_padre:
                            prev_consec = glosa_padre.consecutivo_dgh
                            if prev_consec and prev_consec != consecutivo:
                                if factura not in _logged_facturas:
                                    _logged_facturas.add(factura)
                                    logger.warning(
                                        f"[I/R] Divergencia consecutivo factura={factura}: "
                                        f"BD tiene '{prev_consec}', I/R trae '{consecutivo}'. "
                                        f"Vinculando igual y conservando el de BD."
                                    )
                            elif not prev_consec and consecutivo:
                                # Solo sobrescribir si la fila TRAE consecutivo.
                                # No pisar con cadena vacía cuando el export
                                # del DGH dejó el campo en blanco.
                                glosa_padre.consecutivo_dgh = consecutivo
                    # Flush para que la asignación de consecutivo sea
                    # visible a queries siguientes; cachear el resultado
                    # (incluido None = huérfano) para no re-resolver.
                    if glosa_padre is not None:
                        try:
                            self.db.flush()
                        except Exception:
                            pass
                    _padre_cache[_ck] = glosa_padre
                if not glosa_padre:
                    resumen.conceptos_huerfanos.append(
                        {
                            "fila": num_fila,
                            "hoja": nombre_hoja,
                            "factura": factura,
                            "consecutivo_dgh": consecutivo,
                            "codigo_glosa": codigo_glosa,
                            "cups": cups_codigo,
                            "valor": valor_obj,
                            "motivo": "No existe glosa con esa FACTURA (carga primero INICIAL/RATIFICADA)",
                        }
                    )
                    continue

                # Extra: completar metadatos de la glosa padre desde la hoja I/R
                # (saldo, valor factura, NIT) si venían vacíos de INICIAL/RATIFICADA.
                if _get("saldo_factura") is not None and not glosa_padre.saldo_factura:
                    glosa_padre.saldo_factura = _a_float(_get("saldo_factura"))
                if _get("valor_factura") is not None and not glosa_padre.valor_factura:
                    glosa_padre.valor_factura = _a_float(_get("valor_factura"))
                nit = str(_get("tercero_nit") or "").strip()
                if nit and not glosa_padre.tercero_nit:
                    glosa_padre.tercero_nit = nit
                # Nombre corto de la entidad (desde Tercero.NombreCompletoNA).
                # Limpiar trailing spaces y _x000D_ via _fix_mojibake.
                nombre_corto = _fix_mojibake(str(_get("tercero_nombre") or "").strip())
                if nombre_corto and not getattr(glosa_padre, "tercero_nombre", None):
                    glosa_padre.tercero_nombre = nombre_corto
                fecha_obj = _a_fecha(_get("fecha_objecion"))
                if fecha_obj and not glosa_padre.fecha_objecion_eps:
                    glosa_padre.fecha_objecion_eps = fecha_obj

                # Upsert del concepto por OID (idempotente)
                concepto_existente = None
                if oid:
                    concepto_existente = (
                        self.db.query(ConceptoGlosaRecord)
                        .filter(ConceptoGlosaRecord.oid_dgh == oid)
                        .first()
                    )

                campos = dict(
                    glosa_id=glosa_padre.id,
                    oid_dgh=oid or None,
                    consecutivo_dgh=consecutivo,
                    factura=factura,
                    codigo_glosa=codigo_glosa,
                    codigo_syscafe=codigo_syscafe,
                    nombre_glosa=nombre_glosa or None,
                    cups_codigo=cups_codigo or None,
                    cups_descripcion=cups_desc or None,
                    centro_costo=centro_costo or None,
                    valor_objetado=valor_obj,
                    observacion_eps=observacion or None,
                )

                if concepto_existente:
                    for k, v in campos.items():
                        setattr(concepto_existente, k, v)
                    resumen.conceptos_actualizados += 1
                else:
                    self.db.add(ConceptoGlosaRecord(**campos))
                    resumen.conceptos_creados += 1

            except Exception as e:
                resumen.errores.append(f"[Conceptos {nombre_hoja}] Fila {num_fila}: {e}")
                logger.warning(f"Error procesando concepto fila {num_fila}: {e}")
                continue

        # Compilar texto_glosa_original de cada glosa que recibió conceptos:
        # es lo que la IA del auto-responder lee como "texto de la glosa".
        # Sin esto, la IA recibía el dictamen placeholder ("Pendiente de
        # análisis...") en vez del CONCEPTO real de la EPS. Solo se pisa si
        # el campo está vacío o si lo compuso una importación anterior
        # (empieza con MARCA_TEXTO_CONCEPTOS) — el texto manual se respeta.
        try:
            self.db.flush()  # los conceptos recién agregados deben ser visibles
        except Exception as e:
            logger.warning(f"flush pre-compilación de conceptos falló: {e}")
        padres_unicos = {p.id: p for p in _padre_cache.values() if p is not None and p.id}
        for glosa_padre in padres_unicos.values():
            try:
                actual = (glosa_padre.texto_glosa_original or "").strip()
                if actual and not actual.startswith(MARCA_TEXTO_CONCEPTOS):
                    continue
                texto = componer_texto_desde_conceptos(self.db, glosa_padre)
                if texto:
                    glosa_padre.texto_glosa_original = texto
            except Exception as e:
                logger.warning(
                    f"No se pudo compilar texto de conceptos para glosa {glosa_padre.id}: {e}"
                )

        try:
            self.db.commit()
        except Exception as e:
            logger.error(f"Error guardando conceptos: {e}")
            self.db.rollback()
            resumen.errores.append(f"Error al guardar conceptos: {e}")
