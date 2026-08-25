"""API del módulo de PRE-AUDITORÍA SINAC (v2 — el consolidado como base de datos).

Flujo del auditor:
  1. Subir RADICACIÓN DE CUENTAS  → POST /preauditoria/fuentes/radicacion
  2. Subir DGREPORT               → POST /preauditoria/fuentes/dgreport
  3. Registrar el oficio recibido → POST /preauditoria/oficios
  4. Escribir un envío            → POST /preauditoria/oficios/{id}/envios
       (el sistema crea una fila por cada factura, autocompleta y reconoce
        subsanaciones sin duplicar)
  5. Radicar o devolver           → PATCH /preauditoria/facturas/{id}/auditar
  6. Oficio de devolución PDF + estadísticas + consolidado consultable/exportable

Acceso: las siete rutas que escriben exigen **AUDITOR o superior**. Antes
bastaba con estar autenticado, así que un VIEWER podía subir una fuente,
auditar una factura o emitir un oficio de devolución —un documento que sale
del hospital con un consecutivo—. Consultar el consolidado sigue abierto a
cualquier usuario autenticado: leer no cambia nada.
"""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import func as sa_func
from sqlalchemy.orm import Session

from app.api.deps import (
    get_auditor_o_superior,
    get_coordinador_o_admin,
    get_usuario_actual,
)
from app.core.tz import TZ_BOGOTA, a_utc
from app.database import get_db
from app.models.db import (
    ROL_COORDINADOR,
    ROL_SUPER_ADMIN,
    DgReportRecord,
    EnvioCargadoRecord,
    FacturaEventoRecord,
    FacturaPreauditoriaRecord,
    OficioDevolucionRecord,
    OficioRecepcionRecord,
    RadicacionCuentaRecord,
    UsuarioRecord,
)
from app.services import preauditoria_service as svc
from app.services.oficio_devolucion_pdf import generar_pdf_oficio_devolucion

router = APIRouter(prefix="/preauditoria", tags=["preauditoria"])


# ------------------------------------------------------------------
# Schemas
# ------------------------------------------------------------------


class OficioIn(BaseModel):
    numero_radicado: str = Field(..., min_length=3, max_length=60)
    fecha_recibido: str = Field(..., min_length=10)  # "2026-07-22T14:35" hora Bogotá
    observaciones: Optional[str] = Field(None, max_length=2000)


class EnvioIn(BaseModel):
    envio: str = Field(..., min_length=1, max_length=30)


class AuditarIn(BaseModel):
    resultado: str = Field(..., pattern="^(RADICAR|DEVUELTA|PENDIENTE)$")
    motivo_devolucion: Optional[str] = Field(None, max_length=4000)
    observaciones: Optional[str] = Field(None, max_length=2000)


class ObservacionIn(BaseModel):
    # Anotar la observación de una factura ya auditada, sin revertir la decisión.
    observaciones: str = Field(..., max_length=2000)


class OficioDevolucionIn(BaseModel):
    # La numeración de los oficios de devolución la lleva SINAC por fuera del
    # sistema: el auditor escribe el consecutivo que va. Si no lo escribe, se
    # usa el siguiente sugerido.
    consecutivo: Optional[str] = Field(None, max_length=60)


# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------


def _parsear_fecha_local(texto: str) -> datetime:
    """'2026-07-22T14:35' (hora Bogotá) → datetime TZ-aware en UTC."""
    limpio = texto.strip().replace("Z", "")
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            naive = datetime.strptime(limpio, fmt)
            return naive.replace(tzinfo=TZ_BOGOTA).astimezone(timezone.utc)
        except ValueError:
            continue
    raise HTTPException(400, "Fecha de recibido inválida (use AAAA-MM-DDTHH:MM)")


def _nombre_auditor(u: UsuarioRecord) -> str:
    return (u.nombre or u.email or "").strip() or "SIN NOMBRE"


def _fecha_iso(dt) -> Optional[str]:
    """Para fechas del proceso guardadas en UTC (ej. la fecha de recibido del
    oficio): convierte a hora Bogotá."""
    if dt is None:
        return None
    dt = a_utc(dt)
    return dt.astimezone(TZ_BOGOTA).isoformat()


def _fecha_fuente_iso(dt) -> Optional[str]:
    """Para fechas que vienen de las fuentes (F_FACTURA, F_RECIBIDO de la
    Radicación, fecha de correo): son fechas de calendario, no timestamps con
    zona. Se serializan SIN corrimiento de zona (el navegador las lee como
    locales) para que no se corran un día."""
    if dt is None:
        return None
    if isinstance(dt, datetime):
        return dt.strftime("%Y-%m-%dT%H:%M:%S")
    return dt.isoformat()


async def _leer_xlsx(archivo: UploadFile) -> bytes:
    nombre = archivo.filename or "archivo.xlsx"
    if not nombre.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "El archivo debe ser un Excel (.xlsx)")
    contenido = await archivo.read()
    if len(contenido) > 30 * 1024 * 1024:
        raise HTTPException(400, "Archivo demasiado grande (máx. 30 MB)")
    return contenido


def _factura_dict(db: Session, f: FacturaPreauditoriaRecord, fuente: dict = None) -> dict:
    if fuente is None:
        fuente = svc.datos_fuente(db, f.factura)
    return {
        "id": f.id,
        "factura": f.factura,
        "envio": f.envio_actual or fuente.get("envio"),
        "oficio_fhus": f.oficio_fhus,
        "f_recibido": _fecha_iso(f.f_recibido) or _fecha_fuente_iso(fuente.get("f_recibido")),
        "f_factura": _fecha_fuente_iso(fuente.get("f_factura")),
        "valor": fuente.get("valor") or 0.0,
        "nit": fuente.get("nit"),
        "entidad": fuente.get("entidad"),
        "correo_fe": fuente.get("correo_fe"),
        "estado": f.estado,
        "resultado": f.resultado_actual,
        "ronda": f.ronda_actual,
        "num_subsanacion": f.num_subsanacion,
        "num_devoluciones": f.num_devoluciones,
        "max_devoluciones": svc.MAX_DEVOLUCIONES,
        "pendiente_subsanacion": bool(f.pendiente_subsanacion),
        "en_limite": f.num_devoluciones >= svc.MAX_DEVOLUCIONES,
        "motivo_devolucion": f.motivo_ultima_devolucion,
        "observaciones": f.observaciones,
        "auditor": f.auditor,
        "fecha_auditoria": _fecha_iso(f.fecha_auditoria),
        "oficio_devolucion_id": f.oficio_devolucion_id,
    }


def _oficio_dict(db: Session, o: OficioRecepcionRecord, con_facturas: bool = False) -> dict:
    facturas = (
        db.query(FacturaPreauditoriaRecord)
        .filter(FacturaPreauditoriaRecord.oficio_actual_id == o.id)
        .order_by(FacturaPreauditoriaRecord.envio_actual, FacturaPreauditoriaRecord.factura)
        .all()
    )
    pendientes = sum(1 for f in facturas if f.resultado_actual == svc.RESULTADO_PENDIENTE)
    radicar = sum(1 for f in facturas if f.resultado_actual == svc.RESULTADO_RADICAR)
    devueltas = sum(1 for f in facturas if f.resultado_actual == svc.RESULTADO_DEVUELTA)
    # Cuántas devueltas entrarían REALMENTE en el próximo oficio de devolución:
    # las que ya salieron en uno emitido no se repiten (la entidad recibiría el
    # mismo cobro dos veces). Sin este dato la pantalla decía "Dev 3" y el PDF
    # salía con 1, sin explicar por qué.
    devueltas_sin_oficio = sum(
        1
        for f in facturas
        if f.resultado_actual == svc.RESULTADO_DEVUELTA and f.oficio_devolucion_id is None
    )
    # Un oficio se da por cumplido cuando no le queda nada pendiente. También
    # el que se quedó SIN facturas porque todas siguieron su camino (devueltas
    # aquí y reingresadas en otro oficio): antes esos quedaban en rojo para
    # siempre, como si nadie los hubiera auditado. La consulta extra solo
    # ocurre en los oficios vacíos.
    completado = pendientes == 0 and (bool(facturas) or svc.tuvo_facturas(db, o.id))
    envios = (
        db.query(EnvioCargadoRecord)
        .filter(EnvioCargadoRecord.oficio_id == o.id)
        .order_by(EnvioCargadoRecord.envio)
        .all()
    )
    # Cuántas facturas de cada envío quedaron DE VERDAD en este oficio. El chip
    # decía "(1)" aunque esa factura nunca entrara (venía abierta en otro
    # oficio y el sistema la omite): el auditor contaba 5 envíos y 4 facturas
    # sin ninguna explicación en pantalla. Se cuenta en memoria, sin consultas
    # extra, para que la lista de oficios no se vuelva lenta.
    por_envio: dict = {}
    for f in facturas:
        por_envio[f.envio_actual] = por_envio.get(f.envio_actual, 0) + 1
    escritos = [
        {
            "envio": e.envio,
            "total_facturas": e.total_facturas,
            "en_este_oficio": por_envio.get(e.envio, 0),
            "reingresos": e.reingresos,
        }
        for e in envios
    ]
    # En la ventana del oficio (una sola) sí se averigua dónde quedó cada
    # factura que falta, para poder decirlo con nombre propio.
    if con_facturas:
        faltantes = svc.faltantes_por_envio(db, o.id, facturas)
        for e in escritos:
            e["faltantes"] = faltantes.get(e["envio"], [])
    auditores = sorted({f.auditor for f in facturas if f.auditor})
    # ¿El oficio tiene facturas de ADRES? (habilita el Excel especial que esa
    # entidad exige). Se resuelve contra la fuente con una sola consulta.
    entidades = (
        db.query(RadicacionCuentaRecord.nit, RadicacionCuentaRecord.entidad)
        .join(
            FacturaPreauditoriaRecord,
            FacturaPreauditoriaRecord.factura == RadicacionCuentaRecord.factura,
        )
        .filter(FacturaPreauditoriaRecord.oficio_actual_id == o.id)
        .distinct()
        .all()
    )
    tiene_adres = any(svc.es_adres(nit, ent) for nit, ent in entidades)
    d = {
        "id": o.id,
        "numero_radicado": o.numero_radicado,
        "fecha_recibido": _fecha_iso(o.fecha_recibido),
        "observaciones": o.observaciones,
        "creado_por": o.creado_por,
        # Fase 1: quién RECEPCIONÓ el oficio (lo registró en el sistema).
        "recepcionado_por": o.creado_por,
        # Fase 2: quién(es) están AUDITANDO sus facturas.
        "auditores": auditores,
        "tiene_adres": tiene_adres,
        "total_facturas": len(facturas),
        "pendientes": pendientes,
        "radicar": radicar,
        "devueltas": devueltas,
        "devueltas_sin_oficio": devueltas_sin_oficio,
        "semaforo": svc.calcular_semaforo(o.fecha_recibido, completado=completado),
        "envios_escritos": escritos,
    }
    if con_facturas:
        # Consecutivo del oficio en que salió cada devuelta, para que la fila
        # lo muestre y se entienda por qué no entra en el próximo.
        ids_dev = {f.oficio_devolucion_id for f in facturas if f.oficio_devolucion_id}
        consecutivos = (
            {
                dv.id: dv.consecutivo
                for dv in db.query(OficioDevolucionRecord)
                .filter(OficioDevolucionRecord.id.in_(ids_dev))
                .all()
            }
            if ids_dev
            else {}
        )
        d["facturas"] = [
            {
                **_factura_dict(db, f),
                "consecutivo_devolucion": consecutivos.get(f.oficio_devolucion_id),
            }
            for f in facturas
        ]
    return d


# ------------------------------------------------------------------
# 1-2. Fuentes (Radicación de Cuentas + DGReport)
# ------------------------------------------------------------------


@router.post("/fuentes/radicacion")
async def subir_radicacion(
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_auditor_o_superior),
):
    contenido = await _leer_xlsx(archivo)
    try:
        parsed = svc.parsear_excel_radicacion(contenido)
    except Exception:
        raise HTTPException(400, "No se pudo leer el Excel de Radicación de Cuentas.")
    facturas = parsed["facturas"]
    if not facturas:
        raise HTTPException(422, "; ".join(parsed["advertencias"]) or "Sin facturas válidas.")
    res = svc.upsert_radicacion(db, facturas, archivo.filename or "", _nombre_auditor(current_user))
    return {
        "tipo": "RADICACION",
        "filas_leidas": parsed.get("leidas", 0),
        "facturas_validas": len(facturas),
        **res,
        "advertencias": parsed["advertencias"][:50],
    }


@router.post("/fuentes/dgreport")
async def subir_dgreport(
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_auditor_o_superior),
):
    contenido = await _leer_xlsx(archivo)
    try:
        parsed = svc.parsear_excel_dgreport(contenido)
    except Exception:
        raise HTTPException(400, "No se pudo leer el Excel de DGReport.")
    facturas = parsed["facturas"]
    if not facturas:
        raise HTTPException(422, "; ".join(parsed["advertencias"]) or "Sin facturas válidas.")
    res = svc.upsert_dgreport(db, facturas, archivo.filename or "", _nombre_auditor(current_user))
    return {
        "tipo": "DGREPORT",
        "filas_leidas": parsed.get("leidas", 0),
        "facturas_validas": len(facturas),
        **res,
        "advertencias": parsed["advertencias"][:50],
    }


def _ultimo_cargue(db: Session, modelo) -> tuple[Optional[str], Optional[str]]:
    """Archivo y fecha del último cargue de una fuente (una sola consulta)."""
    cuando = sa_func.coalesce(modelo.actualizado_en, modelo.importado_en)
    fila = db.query(modelo.fuente_archivo, cuando).order_by(cuando.desc()).first()
    if fila is None:
        return None, None
    return fila[0], _fecha_iso(fila[1])


@router.get("/fuentes/resumen")
def resumen_fuentes(
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    n_rad = db.query(sa_func.count(RadicacionCuentaRecord.id)).scalar() or 0
    n_envios = db.query(sa_func.count(sa_func.distinct(RadicacionCuentaRecord.envio))).scalar() or 0
    n_dg = db.query(sa_func.count(DgReportRecord.id)).scalar() or 0
    rad = _ultimo_cargue(db, RadicacionCuentaRecord)
    dgr = _ultimo_cargue(db, DgReportRecord)
    return {
        "radicacion_facturas": n_rad,
        "radicacion_envios": n_envios,
        "dgreport_facturas": n_dg,
        # Qué archivo entró de último y cuándo: sin esto no había forma de
        # saber en la página si el Excel que se acaba de bajar del DGH ya se
        # había subido o no (la pregunta del 20-08-2026).
        "radicacion_ultimo_archivo": rad[0],
        "radicacion_ultimo_cargue": rad[1],
        "dgreport_ultimo_archivo": dgr[0],
        "dgreport_ultimo_cargue": dgr[1],
    }


@router.get("/fuentes/radicacion")
def listar_radicacion(
    q: Optional[str] = Query(None),
    envio: Optional[str] = None,
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    consulta = db.query(RadicacionCuentaRecord)
    if q:
        patron = f"%{q.strip()}%"
        consulta = consulta.filter(
            RadicacionCuentaRecord.factura.ilike(patron)
            | RadicacionCuentaRecord.entidad.ilike(patron)
            | RadicacionCuentaRecord.nit.ilike(patron)
        )
    if envio:
        consulta = consulta.filter(RadicacionCuentaRecord.envio == envio.strip())
    total = consulta.count()
    filas = (
        consulta.order_by(RadicacionCuentaRecord.envio, RadicacionCuentaRecord.factura)
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )
    return {
        "items": [
            {
                "factura": r.factura,
                "envio": r.envio,
                "f_recibido": _fecha_fuente_iso(r.f_recibido),
                "f_factura": _fecha_fuente_iso(r.f_factura),
                "valor": r.valor,
                "nit": r.nit,
                "entidad": r.entidad,
            }
            for r in filas
        ],
        "total": total,
        "page": page,
        "per_page": per_page,
    }


# ------------------------------------------------------------------
# 3. Oficios de recepción (FHUS + fecha/hora de recibido)
# ------------------------------------------------------------------


@router.post("/oficios")
def crear_oficio(
    body: OficioIn,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_auditor_o_superior),
):
    radicado = body.numero_radicado.strip().upper()
    existe = (
        db.query(OficioRecepcionRecord)
        .filter(OficioRecepcionRecord.numero_radicado == radicado)
        .first()
    )
    if existe:
        raise HTTPException(409, f"El oficio {radicado} ya está registrado (id {existe.id})")
    o = OficioRecepcionRecord(
        numero_radicado=radicado,
        fecha_recibido=_parsear_fecha_local(body.fecha_recibido),
        observaciones=(body.observaciones or "").strip() or None,
        creado_por=_nombre_auditor(current_user),
    )
    db.add(o)
    db.commit()
    db.refresh(o)
    return _oficio_dict(db, o)


@router.get("/oficios")
def listar_oficios(
    estado: Optional[str] = Query(None, description="VERDE|AMARILLO|ROJO|VENCIDO|COMPLETADO"),
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    consulta = db.query(OficioRecepcionRecord)
    if q:
        consulta = consulta.filter(OficioRecepcionRecord.numero_radicado.ilike(f"%{q.strip()}%"))
    oficios = consulta.order_by(OficioRecepcionRecord.fecha_recibido.desc()).all()
    items = [_oficio_dict(db, o) for o in oficios]
    if estado:
        items = [i for i in items if i["semaforo"]["estado"] == estado.strip().upper()]
    total = len(items)
    inicio = (page - 1) * per_page
    return {
        "items": items[inicio : inicio + per_page],
        "total": total,
        "page": page,
        "per_page": per_page,
    }


@router.get("/oficios/{oficio_id}")
def ver_oficio(
    oficio_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    o = db.get(OficioRecepcionRecord, oficio_id)
    if not o:
        raise HTTPException(404, "Oficio no encontrado")
    return _oficio_dict(db, o, con_facturas=True)


class RenombrarOficioIn(BaseModel):
    numero_radicado: str = Field(..., min_length=3, max_length=60)


@router.patch("/oficios/{oficio_id}")
def renombrar_oficio(
    oficio_id: int,
    body: RenombrarOficioIn,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_coordinador_o_admin),
):
    """Corrige el número de radicado de un oficio mal digitado.

    El número queda copiado en cada factura y en cada evento del historial:
    el servicio los corrige todos en la misma operación.
    """
    o = db.get(OficioRecepcionRecord, oficio_id)
    if not o:
        raise HTTPException(404, "Oficio no encontrado")
    res = svc.renombrar_oficio(db, o, body.numero_radicado)
    if not res.get("ok"):
        raise HTTPException(res.get("codigo", 409), res.get("mensaje", "No se pudo corregir"))
    return {**res, "oficio": _oficio_dict(db, o)}


class EliminarOficiosIn(BaseModel):
    ids: list[int] = Field(..., min_length=1, max_length=100)


@router.delete("/oficios/{oficio_id}")
def eliminar_oficio(
    oficio_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_coordinador_o_admin),
):
    """Elimina un oficio mal registrado (solo SUPER_ADMIN/COORDINADOR)."""
    o = db.get(OficioRecepcionRecord, oficio_id)
    if not o:
        raise HTTPException(404, "Oficio no encontrado")
    res = svc.eliminar_oficio(db, o)
    if not res.get("ok"):
        raise HTTPException(res.get("codigo", 409), res.get("mensaje", "No se pudo eliminar"))
    return res


@router.post("/oficios/eliminar-masivo")
def eliminar_oficios_masivo(
    body: EliminarOficiosIn,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_coordinador_o_admin),
):
    """Elimina varios oficios a la vez (solo SUPER_ADMIN/COORDINADOR).

    Los que no se puedan eliminar (PDF de devolución emitido) se reportan
    sin frenar a los demás.
    """
    eliminados = []
    rechazados = []
    for oid in body.ids:
        o = db.get(OficioRecepcionRecord, oid)
        if not o:
            rechazados.append({"id": oid, "motivo": "No encontrado"})
            continue
        res = svc.eliminar_oficio(db, o)
        if res.get("ok"):
            eliminados.append(res["radicado"])
        else:
            rechazados.append({"id": oid, "motivo": res.get("mensaje", "")})
    return {"eliminados": eliminados, "rechazados": rechazados}


@router.get("/oficios/{oficio_id}/export-adres.xlsx")
def exportar_oficio_adres(
    oficio_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Excel de las facturas ADRES de un oficio, con el formato del consolidado
    que se maneja con esa entidad.

    SINAC diligencia desde "Item" hasta "Fecha_Entrega_Fact" (esta última
    queda en blanco: es la fecha de entrega física a Facturación). Las
    columnas siguientes (Observación_FACTURACIÓN … INFOPOL) pertenecen a las
    otras áreas y van incluidas pero vacías, para que continúen el mismo
    archivo. Solo salen las facturas cuya entidad es ADRES.
    """
    import re as _re

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    o = db.get(OficioRecepcionRecord, oficio_id)
    if not o:
        raise HTTPException(404, "Oficio no encontrado")

    filas = (
        db.query(FacturaPreauditoriaRecord, RadicacionCuentaRecord, DgReportRecord)
        .outerjoin(
            RadicacionCuentaRecord,
            RadicacionCuentaRecord.factura == FacturaPreauditoriaRecord.factura,
        )
        .outerjoin(DgReportRecord, DgReportRecord.factura == FacturaPreauditoriaRecord.factura)
        .filter(FacturaPreauditoriaRecord.oficio_actual_id == oficio_id)
        .order_by(FacturaPreauditoriaRecord.envio_actual, FacturaPreauditoriaRecord.factura)
        .all()
    )
    adres = [
        (f, rad, dg)
        for f, rad, dg in filas
        if svc.es_adres(rad.nit if rad else None, rad.entidad if rad else None)
    ]
    if not adres:
        raise HTTPException(
            404,
            "Este oficio no tiene facturas de ADRES: el Excel especial solo "
            "aplica a oficios de esa entidad.",
        )

    # Consecutivos de oficios de devolución ya emitidos (para la columna).
    dev_ids = {f.oficio_devolucion_id for f, _, _ in adres if f.oficio_devolucion_id}
    consecutivos = (
        {
            d.id: d.consecutivo
            for d in db.query(OficioDevolucionRecord)
            .filter(OficioDevolucionRecord.id.in_(dev_ids))
            .all()
        }
        if dev_ids
        else {}
    )

    # Tramo que diligencia SINAC + tramo de las otras áreas (va vacío).
    COLS_SINAC = [
        "Item",
        "Fecha_Recibido",
        "Envío",
        "AUD",
        "HUS",
        "Fecha_Factura",
        "Valor",
        "NIT",
        "Entidad",
        "Correo F.E.",
        "Observación Preauditoria Radicación SINAC",
        "Radicar_1",
        "Observaciones Adicionales",
        "Fecha_Entrega_Fact",
    ]
    COLS_OTRAS_AREAS = [
        "Observación_FACTURACIÓN",
        "Fecha_Dev_CARTERA",
        "Fecha_Segunda_Revisión",
        "Segunda_Observación_SINAC",
        "Fecha_Dev_FACTURACIÓN",
        "Segunda_Observación_FACTURACIÓN",
        "Fecha_Dev_CARTERA",
        "Radicar_2",
        "Fecha_Radicación",
        "Número_Radicado",
        "INFOPOL",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "ADRES"
    ws.append(COLS_SINAC + COLS_OTRAS_AREAS)
    for idx, (f, rad, dg) in enumerate(adres, start=1):
        if f.resultado_actual == svc.RESULTADO_DEVUELTA:
            obs = f.motivo_ultima_devolucion or "DEVUELTA"
            cons = consecutivos.get(f.oficio_devolucion_id)
            if cons:
                obs = f"{obs} — Oficio {cons}"
            radicar_1 = "NO"
        elif f.resultado_actual == svc.RESULTADO_RADICAR:
            obs = "SOPORTES COMPLETOS"
            radicar_1 = "SI"
        else:
            obs = ""
            radicar_1 = ""
        ws.append(
            [
                idx,
                (_fecha_iso(f.f_recibido) or "").replace("T", " ")[:16],
                f.envio_actual or (rad.envio if rad else None),
                f.auditor,
                f.factura,
                (_fecha_fuente_iso(rad.f_factura if rad else None) or "")[:10],
                (rad.valor if rad else 0) or 0,
                rad.nit if rad else None,
                rad.entidad if rad else None,
                dg.correo_fe if dg else "NO",
                obs,
                radicar_1,
                f.observaciones,
                "",  # Fecha_Entrega_Fact: la diligencia a mano quien entrega
            ]
            + [""] * len(COLS_OTRAS_AREAS)
        )

    # Presentación: encabezado azul (tramo SINAC) / gris (otras áreas),
    # primera fila congelada, valor con formato de miles.
    azul = PatternFill("solid", fgColor="1F4E79")
    gris = PatternFill("solid", fgColor="808080")
    blanco = Font(bold=True, color="FFFFFF", size=9)
    for c, _titulo in enumerate(COLS_SINAC + COLS_OTRAS_AREAS, start=1):
        celda = ws.cell(row=1, column=c)
        celda.fill = azul if c <= len(COLS_SINAC) else gris
        celda.font = blanco
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    anchos = {
        1: 5,
        2: 17,
        3: 9,
        4: 16,
        5: 15,
        6: 11,
        7: 13,
        8: 11,
        9: 34,
        10: 9,
        11: 40,
        12: 9,
        13: 24,
        14: 12,
    }
    for c in range(1, len(COLS_SINAC + COLS_OTRAS_AREAS) + 1):
        letra = ws.cell(row=1, column=c).column_letter
        ws.column_dimensions[letra].width = anchos.get(c, 16)
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=7).number_format = "#,##0"
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    nombre = _re.sub(r"[^A-Za-z0-9_-]+", "_", o.numero_radicado or "oficio")
    return StreamingResponse(
        BytesIO(buf.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="ADRES_{nombre}.xlsx"'},
    )


# ------------------------------------------------------------------
# 4. Escribir un envío (autocompleta facturas + subsanaciones)
# ------------------------------------------------------------------


@router.get("/oficios/{oficio_id}/envios/{envio}/preview")
def preview_envio(
    oficio_id: int,
    envio: str,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    o = db.get(OficioRecepcionRecord, oficio_id)
    if not o:
        raise HTTPException(404, "Oficio no encontrado")
    return svc.preview_envio(db, o, envio)


@router.delete("/oficios/{oficio_id}/envios/{envio}")
def eliminar_envio(
    oficio_id: int,
    envio: str,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_coordinador_o_admin),
):
    """Deshace un envío cargado por error (solo SUPER_ADMIN/COORDINADOR).

    Borra las facturas que entraron con ese envío y lo libera para volver a
    escribirlo. El resto del oficio no se toca.
    """
    o = db.get(OficioRecepcionRecord, oficio_id)
    if not o:
        raise HTTPException(404, "Oficio no encontrado")
    res = svc.eliminar_envio(db, o, envio)
    if not res.get("ok"):
        raise HTTPException(res.get("codigo", 409), res.get("mensaje", "No se pudo eliminar"))
    return {**res, "oficio": _oficio_dict(db, o)}


@router.delete("/oficios/{oficio_id}/facturas/{factura_id}")
def quitar_factura_del_oficio(
    oficio_id: int,
    factura_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_coordinador_o_admin),
):
    """Quita UNA factura que entró por error al oficio (SUPER_ADMIN/COORDINADOR).

    Para el caso en que la fuente de Radicación arrastra una factura que
    facturación ya sacó del envío: en vez de quitar el envío entero y volver
    a escribirlo, se quita solo esa factura. El resto del oficio no se toca.
    """
    o = db.get(OficioRecepcionRecord, oficio_id)
    if not o:
        raise HTTPException(404, "Oficio no encontrado")
    f = db.get(FacturaPreauditoriaRecord, factura_id)
    if not f:
        raise HTTPException(404, "Factura no encontrada")
    res = svc.quitar_factura(db, o, f)
    if not res.get("ok"):
        raise HTTPException(res.get("codigo", 409), res.get("mensaje", "No se pudo quitar"))
    return {**res, "oficio": _oficio_dict(db, o)}


@router.post("/oficios/{oficio_id}/envios")
def escribir_envio(
    oficio_id: int,
    body: EnvioIn,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_auditor_o_superior),
):
    o = db.get(OficioRecepcionRecord, oficio_id)
    if not o:
        raise HTTPException(404, "Oficio no encontrado")
    res = svc.escribir_envio(db, o, body.envio, _nombre_auditor(current_user))
    if res.get("ya_cargado"):
        return {**res, "oficio": _oficio_dict(db, o)}
    if res.get("existe_en_fuente") is False:
        raise HTTPException(404, res["mensaje"])
    return {**res, "oficio": _oficio_dict(db, o)}


# ------------------------------------------------------------------
# 5. Auditar factura (RADICAR / DEVOLVER)
# ------------------------------------------------------------------

_MAP_RESULTADO = {
    "RADICAR": svc.RESULTADO_RADICAR,
    "DEVUELTA": svc.RESULTADO_DEVUELTA,
    "PENDIENTE": svc.RESULTADO_PENDIENTE,
}


def _es_admin_o_coordinador(u: UsuarioRecord) -> bool:
    return u.rol in (ROL_SUPER_ADMIN, ROL_COORDINADOR)


@router.patch("/facturas/{factura_id}/auditar")
def auditar_factura(
    factura_id: int,
    body: AuditarIn,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_auditor_o_superior),
):
    f = db.get(FacturaPreauditoriaRecord, factura_id)
    if not f:
        raise HTTPException(404, "Factura no encontrada")
    # Deshacer una auditoría YA DECIDIDA: cada auditor puede con LO SUYO
    # mientras no haya salido en un oficio de devolución entregado; lo de otra
    # persona sigue siendo de coordinación (la regla vive en el servicio).
    if body.resultado == "PENDIENTE" and f.resultado_actual != svc.RESULTADO_PENDIENTE:
        puede, motivo = svc.puede_revertir(
            f, _nombre_auditor(current_user), _es_admin_o_coordinador(current_user)
        )
        if not puede:
            raise HTTPException(403, motivo)
    res = svc.auditar_factura(
        db,
        f,
        _MAP_RESULTADO[body.resultado],
        _nombre_auditor(current_user),
        motivo=body.motivo_devolucion,
        observaciones=body.observaciones,
    )
    if not res.get("ok"):
        raise HTTPException(res.get("codigo", 400), res.get("mensaje", "No se pudo auditar"))
    db.refresh(f)
    return _factura_dict(db, f)


@router.patch("/facturas/{factura_id}/observacion")
def anotar_observacion(
    factura_id: int,
    body: ObservacionIn,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_auditor_o_superior),
):
    """Agrega o corrige la observación sin tocar la decisión ya tomada.

    En una factura DEVUELTA la observación es el motivo que imprime el oficio
    de devolución: si ya salió en uno, la corrección también corrige el PDF.
    Por eso mismo, corregir una auditoría YA DECIDIDA es solo de
    coordinación/administración (el auditor sí escribe su observación al
    decidir, y puede anotarla mientras la factura siga pendiente).
    """
    f = db.get(FacturaPreauditoriaRecord, factura_id)
    if not f:
        raise HTTPException(404, "Factura no encontrada")
    if f.resultado_actual != svc.RESULTADO_PENDIENTE and not _es_admin_o_coordinador(current_user):
        raise HTTPException(
            403,
            "Solo coordinación o administración puede corregir la observación de "
            "una auditoría ya radicada o devuelta (la corrección también cambia "
            "el oficio de devolución).",
        )
    res = svc.anotar_observacion(db, f, body.observaciones, _nombre_auditor(current_user))
    if not res.get("ok"):
        raise HTTPException(
            res.get("codigo", 400), res.get("mensaje", "No se pudo guardar la observación")
        )
    db.refresh(f)
    return {**_factura_dict(db, f), "oficio_actualizado": bool(res.get("oficio_actualizado"))}


# ------------------------------------------------------------------
# Consolidado (una fila por factura) + historial + export
# ------------------------------------------------------------------


@router.get("/consolidado")
def consolidado(
    q: Optional[str] = Query(None, description="Factura, entidad o NIT"),
    oficio_id: Optional[int] = None,
    envio: Optional[str] = None,
    auditor: Optional[str] = None,
    resultado: Optional[str] = Query(None, description="PENDIENTE|RADICAR|DEVUELTA"),
    estado: Optional[str] = None,
    solo_reincidentes: bool = Query(False),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    consulta = db.query(FacturaPreauditoriaRecord)
    if oficio_id:
        consulta = consulta.filter(FacturaPreauditoriaRecord.oficio_actual_id == oficio_id)
    if envio:
        consulta = consulta.filter(FacturaPreauditoriaRecord.envio_actual == envio.strip())
    if auditor:
        consulta = consulta.filter(FacturaPreauditoriaRecord.auditor.ilike(f"%{auditor.strip()}%"))
    if resultado:
        consulta = consulta.filter(
            FacturaPreauditoriaRecord.resultado_actual == resultado.strip().upper()
        )
    if estado:
        consulta = consulta.filter(FacturaPreauditoriaRecord.estado == estado.strip().upper())
    if solo_reincidentes:
        consulta = consulta.filter(FacturaPreauditoriaRecord.num_devoluciones >= 1)
    if q:
        patron = f"%{q.strip()}%"
        # entidad/NIT viven en la fuente → subquery correlacionada (NO materializar
        # la lista: con 36k facturas rompería el límite de variables de SQLite).
        #
        # La fuente tiene ~190.000 filas y buscar texto ahí recorría la tabla
        # ENTERA: 1 segundo medido en la VM, y dos veces por búsqueda (una para
        # contar y otra para paginar) = ~2 segundos por cada letra buscada.
        #
        # Pero la fuente solo importa para las facturas que YA están en el
        # consolidado: el OR de abajo las intersecta de todos modos. Así que se
        # restringe primero por `factura` —que tiene índice único— y el filtro
        # de texto cae sobre ese puñado de filas en vez de sobre 190.000. El
        # resultado es idéntico; el tiempo baja de ~1 s a menos de 1 ms.
        sub = db.query(RadicacionCuentaRecord.factura).filter(
            RadicacionCuentaRecord.factura.in_(db.query(FacturaPreauditoriaRecord.factura)),
            RadicacionCuentaRecord.entidad.ilike(patron) | RadicacionCuentaRecord.nit.ilike(patron),
        )
        consulta = consulta.filter(
            FacturaPreauditoriaRecord.factura.ilike(patron)
            | FacturaPreauditoriaRecord.factura.in_(sub)
        )
    total = consulta.count()
    filas = (
        consulta.order_by(FacturaPreauditoriaRecord.id.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )
    # batch de datos-fuente
    numeros = [f.factura for f in filas]
    rad = {
        r.factura: r
        for r in db.query(RadicacionCuentaRecord)
        .filter(RadicacionCuentaRecord.factura.in_(numeros))
        .all()
    }
    dg = {
        r.factura: r
        for r in db.query(DgReportRecord).filter(DgReportRecord.factura.in_(numeros)).all()
    }
    items = []
    for f in filas:
        r = rad.get(f.factura)
        d = dg.get(f.factura)
        fuente = {
            "envio": r.envio if r else None,
            "f_recibido": r.f_recibido if r else None,
            "f_factura": r.f_factura if r else None,
            "valor": r.valor if r else 0.0,
            "nit": r.nit if r else None,
            "entidad": r.entidad if r else None,
            "correo_fe": d.correo_fe if d else "NO",
        }
        items.append(_factura_dict(db, f, fuente))
    return {"items": items, "total": total, "page": page, "per_page": per_page}


@router.get("/facturas/{factura_id}")
def ver_factura(
    factura_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    f = db.get(FacturaPreauditoriaRecord, factura_id)
    if not f:
        raise HTTPException(404, "Factura no encontrada")
    return _factura_dict(db, f)


@router.get("/facturas/{numero}/historial")
def historial_factura(
    numero: str,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    f = (
        db.query(FacturaPreauditoriaRecord)
        .filter(FacturaPreauditoriaRecord.factura == numero.strip().upper())
        .one_or_none()
    )
    if not f:
        raise HTTPException(404, "Factura no encontrada en el consolidado")
    eventos = (
        db.query(FacturaEventoRecord)
        .filter(FacturaEventoRecord.factura_id == f.id)
        .order_by(FacturaEventoRecord.creado_en, FacturaEventoRecord.id)
        .all()
    )
    return {
        "actual": _factura_dict(db, f),
        "eventos": [
            {
                "tipo": e.tipo_evento,
                "ronda": e.ronda,
                "num_subsanacion": e.subsanacion_num,
                "envio": e.envio,
                "oficio_fhus": e.oficio_fhus,
                "resultado": e.resultado,
                "estado_resultante": e.estado_resultante,
                "auditor": e.auditor,
                "motivo": e.motivo,
                "observaciones": e.observaciones,
                "valor": e.valor_snapshot,
                "fecha": _fecha_iso(e.creado_en),
            }
            for e in eventos
        ],
    }


@router.get("/consolidado/export.xlsx")
def exportar_consolidado(
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    from openpyxl import Workbook

    # Un solo JOIN a las fuentes (evita N+1 sobre miles de facturas).
    filas = (
        db.query(FacturaPreauditoriaRecord, RadicacionCuentaRecord, DgReportRecord)
        .outerjoin(
            RadicacionCuentaRecord,
            RadicacionCuentaRecord.factura == FacturaPreauditoriaRecord.factura,
        )
        .outerjoin(DgReportRecord, DgReportRecord.factura == FacturaPreauditoriaRecord.factura)
        .order_by(FacturaPreauditoriaRecord.envio_actual, FacturaPreauditoriaRecord.factura)
        .all()
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "CONSOLIDADO"
    encabezados = [
        "ENVIO",
        "F_RECIBIDO",
        "OFICIO FHUS",
        "FACTURA",
        "F_FACTURA",
        "VALOR",
        "NIT",
        "ENTIDAD",
        "CORREO F.E.",
        "ESTADO",
        "RESULTADO",
        "N SUBSANACION",
        "DEVOLUCIONES",
        "AUDITOR",
        "MOTIVO DEVOLUCION",
    ]
    ws.append(encabezados)
    for f, rad, dg in filas:
        ws.append(
            [
                f.envio_actual or (rad.envio if rad else None),
                (
                    _fecha_iso(f.f_recibido)
                    or _fecha_fuente_iso(rad.f_recibido if rad else None)
                    or ""
                )[:10],
                f.oficio_fhus,
                f.factura,
                (_fecha_fuente_iso(rad.f_factura if rad else None) or "")[:10],
                (rad.valor if rad else 0) or 0,
                rad.nit if rad else None,
                rad.entidad if rad else None,
                dg.correo_fe if dg else "NO",
                f.estado,
                f.resultado_actual,
                f.num_subsanacion,
                f.num_devoluciones,
                f.auditor,
                f.motivo_ultima_devolucion,
            ]
        )
    buf = BytesIO()
    wb.save(buf)
    return StreamingResponse(
        BytesIO(buf.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="CONSOLIDADO_PRE_AUDITORIA.xlsx"'},
    )


# ------------------------------------------------------------------
# 6. Oficio de devolución (consecutivo SINAC + PDF)
# ------------------------------------------------------------------


@router.get("/consecutivo-sugerido")
def consecutivo_sugerido(
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Consecutivo que seguiría según lo ya registrado aquí.

    Es solo una sugerencia: la numeración real la lleva SINAC por fuera del
    sistema, así que el auditor puede cambiarla al generar el oficio.
    """
    consecutivo, numero, anio = svc.siguiente_consecutivo(db)
    return {"consecutivo": consecutivo, "numero": numero, "anio": anio}


@router.post("/oficios/{oficio_id}/oficio-devolucion")
def generar_oficio_devolucion(
    oficio_id: int,
    body: Optional[OficioDevolucionIn] = None,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_auditor_o_superior),
):
    o = db.get(OficioRecepcionRecord, oficio_id)
    if not o:
        raise HTTPException(404, "Oficio no encontrado")
    devueltas = (
        db.query(FacturaPreauditoriaRecord)
        .filter(
            FacturaPreauditoriaRecord.oficio_actual_id == oficio_id,
            FacturaPreauditoriaRecord.resultado_actual == svc.RESULTADO_DEVUELTA,
            FacturaPreauditoriaRecord.oficio_devolucion_id.is_(None),
        )
        .order_by(FacturaPreauditoriaRecord.envio_actual, FacturaPreauditoriaRecord.factura)
        .all()
    )
    if not devueltas:
        # Dos situaciones muy distintas, y antes se explicaban con el mismo
        # mensaje: que no haya devoluciones, o que YA hayan salido todas.
        ya_salieron = (
            db.query(FacturaPreauditoriaRecord)
            .filter(
                FacturaPreauditoriaRecord.oficio_actual_id == oficio_id,
                FacturaPreauditoriaRecord.resultado_actual == svc.RESULTADO_DEVUELTA,
            )
            .count()
        )
        if ya_salieron:
            raise HTTPException(
                400,
                f"Las {ya_salieron} factura(s) devuelta(s) de este radicado ya salieron en un "
                "oficio de devolución. Una factura no se repite en dos oficios: la entidad "
                "recibiría el mismo cobro dos veces. Si necesita rehacerlo, elimine el oficio "
                "anterior desde la pestaña Oficios de devolución.",
            )
        raise HTTPException(
            400,
            "No hay facturas devueltas sin oficio en este radicado: "
            "primero marque las devoluciones con su motivo.",
        )
    try:
        consecutivo, numero, anio = svc.reservar_consecutivo(db, body.consecutivo if body else None)
    except ValueError as e:
        raise HTTPException(409 if "ya fue usado" in str(e) else 400, str(e))
    total_valor = sum(svc.datos_fuente(db, f.factura).get("valor") or 0 for f in devueltas)
    dev = OficioDevolucionRecord(
        consecutivo=consecutivo,
        anio=anio,
        numero=numero,
        oficio_recepcion_id=oficio_id,
        generado_por=_nombre_auditor(current_user),
        total_facturas=len(devueltas),
        total_valor=total_valor,
    )
    db.add(dev)
    db.flush()
    for f in devueltas:
        f.oficio_devolucion_id = dev.id
        # Sellar el evento de DEVOLUCIÓN de esta ronda con el oficio: el PDF se
        # arma desde estos eventos (snapshot inmutable), no desde la fila
        # canónica que un reingreso posterior modificaría.
        evt = (
            db.query(FacturaEventoRecord)
            .filter(
                FacturaEventoRecord.factura_id == f.id,
                FacturaEventoRecord.tipo_evento.in_(["DEVUELTA", "NUEVAMENTE_DEVUELTA"]),
                FacturaEventoRecord.oficio_devolucion_id.is_(None),
            )
            .order_by(FacturaEventoRecord.creado_en.desc(), FacturaEventoRecord.id.desc())
            .first()
        )
        if evt:
            evt.oficio_devolucion_id = dev.id
    db.commit()
    db.refresh(dev)
    return {
        "id": dev.id,
        "consecutivo": dev.consecutivo,
        "total_facturas": dev.total_facturas,
        "total_valor": dev.total_valor,
        "pdf_url": f"/preauditoria/oficios-devolucion/{dev.id}/pdf",
    }


@router.get("/oficios-devolucion")
def listar_oficios_devolucion(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    consulta = db.query(OficioDevolucionRecord).order_by(
        OficioDevolucionRecord.anio.desc(), OficioDevolucionRecord.numero.desc()
    )
    total = consulta.count()
    filas = consulta.offset((page - 1) * per_page).limit(per_page).all()
    recepciones = {
        o.id: o.numero_radicado
        for o in db.query(OficioRecepcionRecord)
        .filter(
            OficioRecepcionRecord.id.in_(
                {f.oficio_recepcion_id for f in filas if f.oficio_recepcion_id}
            )
        )
        .all()
    }
    return {
        "items": [
            {
                "id": d.id,
                "consecutivo": d.consecutivo,
                "fecha_generado": _fecha_iso(d.fecha_generado),
                "generado_por": d.generado_por,
                "oficio_radicado": recepciones.get(d.oficio_recepcion_id),
                "total_facturas": d.total_facturas,
                "total_valor": d.total_valor,
                "pdf_url": f"/preauditoria/oficios-devolucion/{d.id}/pdf",
            }
            for d in filas
        ],
        "total": total,
        "page": page,
        "per_page": per_page,
    }


@router.delete("/oficios-devolucion/{dev_id}")
def eliminar_oficio_devolucion(
    dev_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_coordinador_o_admin),
):
    """Elimina un oficio de devolución generado por error (coordinación/admin).

    Las facturas siguen devueltas: solo quedan libres para entrar en un oficio
    nuevo, y el consecutivo vuelve a estar disponible.
    """
    dev = db.get(OficioDevolucionRecord, dev_id)
    if not dev:
        raise HTTPException(404, "Oficio de devolución no encontrado")
    res = svc.eliminar_oficio_devolucion(db, dev)
    if not res.get("ok"):
        raise HTTPException(res.get("codigo", 409), res.get("mensaje", "No se pudo eliminar"))
    return res


@router.get("/oficios-devolucion/{dev_id}/pdf")
def descargar_pdf_oficio_devolucion(
    dev_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    dev = db.get(OficioDevolucionRecord, dev_id)
    if not dev:
        raise HTTPException(404, "Oficio de devolución no encontrado")
    recepcion = (
        db.get(OficioRecepcionRecord, dev.oficio_recepcion_id) if dev.oficio_recepcion_id else None
    )
    # El PDF se arma desde los EVENTOS sellados (snapshot inmutable del día de la
    # devolución): un reingreso posterior de la factura no altera este documento.
    eventos = (
        db.query(FacturaEventoRecord)
        .filter(FacturaEventoRecord.oficio_devolucion_id == dev_id)
        .order_by(FacturaEventoRecord.envio, FacturaEventoRecord.factura)
        .all()
    )
    fecha_local = a_utc(dev.fecha_generado)
    recibido_local = a_utc(recepcion.fecha_recibido) if recepcion else None
    filas_pdf = [
        {
            "envio": e.envio,
            "factura": e.factura,
            "fecha_factura": e.fecha_factura_snapshot,
            "valor": e.valor_snapshot,
            "nit": e.nit_snapshot,
            "entidad": e.entidad_snapshot,
            "oficio": e.oficio_fhus or (recepcion.numero_radicado if recepcion else ""),
            "motivo_devolucion": e.motivo,
        }
        for e in eventos
    ]
    pdf = generar_pdf_oficio_devolucion(
        consecutivo=dev.consecutivo,
        fecha_generado=fecha_local.astimezone(TZ_BOGOTA) if fecha_local else None,
        numero_radicado=recepcion.numero_radicado if recepcion else "",
        facturas=filas_pdf,
        generado_por=dev.generado_por or "",
        fecha_recibido=recibido_local.astimezone(TZ_BOGOTA) if recibido_local else None,
    )
    return StreamingResponse(
        BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{dev.consecutivo}.pdf"'},
    )


# ------------------------------------------------------------------
# Estadísticas
# ------------------------------------------------------------------


@router.get("/estadisticas/informe.xlsx")
def informe_gestion(
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_auditor_o_superior),
):
    """Informe de gestión descargable: qué se ha hecho, quién y cuándo.

    Cinco hojas: RESUMEN (totales y valores), POR AUDITOR, POR OFICIO,
    DEVOLUCIONES (oficios emitidos) e HISTORIAL (el registro completo de
    eventos — la trazabilidad del módulo, factura por factura).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    facturas = db.query(FacturaPreauditoriaRecord).all()
    valores = {
        fac: (val or 0)
        for fac, val in db.query(
            FacturaPreauditoriaRecord.factura, RadicacionCuentaRecord.valor
        ).outerjoin(
            RadicacionCuentaRecord,
            RadicacionCuentaRecord.factura == FacturaPreauditoriaRecord.factura,
        )
    }

    def _f(dt):
        loc = a_utc(dt)
        return loc.astimezone(TZ_BOGOTA).strftime("%d/%m/%Y %H:%M") if loc else ""

    wb = Workbook()
    azul = PatternFill("solid", fgColor="1F4E79")
    blanco = Font(bold=True, color="FFFFFF", size=10)

    def _hoja(nombre, encabezados, filas, anchos=None):
        ws = (
            wb.active
            if wb.active.max_row == 1 and wb.active.title == "Sheet"
            else wb.create_sheet()
        )
        ws.title = nombre
        ws.append(encabezados)
        for c in range(1, len(encabezados) + 1):
            celda = ws.cell(row=1, column=c)
            celda.fill = azul
            celda.font = blanco
        for fila in filas:
            ws.append(fila)
        for c, ancho in enumerate(anchos or [], start=1):
            ws.column_dimensions[get_column_letter(c)].width = ancho
        ws.freeze_panes = "A2"
        return ws

    # --- RESUMEN -----------------------------------------------------
    radicar = [f for f in facturas if f.resultado_actual == svc.RESULTADO_RADICAR]
    devueltas = [f for f in facturas if f.resultado_actual == svc.RESULTADO_DEVUELTA]
    pendientes = [f for f in facturas if f.resultado_actual == svc.RESULTADO_PENDIENTE]
    devs = db.query(OficioDevolucionRecord).all()
    ahora = datetime.now(TZ_BOGOTA)
    _hoja(
        "RESUMEN",
        ["Concepto", "Valor"],
        [
            ["Informe generado", ahora.strftime("%d/%m/%Y %H:%M")],
            ["Generado por", _nombre_auditor(current_user)],
            ["Oficios de recepción", db.query(OficioRecepcionRecord).count()],
            ["Facturas en el consolidado", len(facturas)],
            ["Radicadas / subsanadas", len(radicar)],
            ["Devueltas (estado actual)", len(devueltas)],
            ["Pendientes por auditar", len(pendientes)],
            ["Oficios de devolución emitidos", len(devs)],
            ["Valor radicado", sum(valores.get(f.factura, 0) for f in radicar)],
            ["Valor devuelto", sum(valores.get(f.factura, 0) for f in devueltas)],
            ["Valor pendiente", sum(valores.get(f.factura, 0) for f in pendientes)],
        ],
        anchos=[34, 24],
    )

    # --- POR AUDITOR -------------------------------------------------
    por_auditor: dict[str, dict] = {}
    for f in facturas:
        if not f.auditor:
            continue
        a = por_auditor.setdefault(
            f.auditor, {"auditadas": 0, "radicar": 0, "devueltas": 0, "valor": 0.0}
        )
        a["auditadas"] += 1
        a["valor"] += valores.get(f.factura, 0)
        if f.resultado_actual == svc.RESULTADO_RADICAR:
            a["radicar"] += 1
        elif f.resultado_actual == svc.RESULTADO_DEVUELTA:
            a["devueltas"] += 1
    _hoja(
        "POR AUDITOR",
        ["Auditor", "Auditadas", "Radicadas", "Devueltas", "Valor auditado"],
        [
            [nombre, a["auditadas"], a["radicar"], a["devueltas"], a["valor"]]
            for nombre, a in sorted(
                por_auditor.items(), key=lambda kv: kv[1]["auditadas"], reverse=True
            )
        ],
        anchos=[28, 12, 12, 12, 18],
    )

    # --- POR OFICIO --------------------------------------------------
    conteo_oficio: dict[int, dict] = {}
    for f in facturas:
        c = conteo_oficio.setdefault(
            f.oficio_actual_id, {"total": 0, "pend": 0, "rad": 0, "dev": 0}
        )
        c["total"] += 1
        if f.resultado_actual == svc.RESULTADO_RADICAR:
            c["rad"] += 1
        elif f.resultado_actual == svc.RESULTADO_DEVUELTA:
            c["dev"] += 1
        else:
            c["pend"] += 1
    _hoja(
        "POR OFICIO",
        ["Radicado", "Recibido", "Recepcionó", "Facturas", "Pendientes", "Radicadas", "Devueltas"],
        [
            [
                o.numero_radicado,
                _f(o.fecha_recibido),
                o.creado_por or "",
                conteo_oficio.get(o.id, {}).get("total", 0),
                conteo_oficio.get(o.id, {}).get("pend", 0),
                conteo_oficio.get(o.id, {}).get("rad", 0),
                conteo_oficio.get(o.id, {}).get("dev", 0),
            ]
            for o in db.query(OficioRecepcionRecord)
            .order_by(OficioRecepcionRecord.fecha_recibido)
            .all()
        ],
        anchos=[24, 18, 22, 10, 12, 12, 12],
    )

    # --- DEVOLUCIONES ------------------------------------------------
    recepciones = {
        o.id: o.numero_radicado
        for o in db.query(OficioRecepcionRecord)
        .filter(
            OficioRecepcionRecord.id.in_(
                {d.oficio_recepcion_id for d in devs if d.oficio_recepcion_id}
            )
        )
        .all()
    }
    _hoja(
        "DEVOLUCIONES",
        ["Consecutivo", "Generado", "Generado por", "Oficio origen", "Facturas", "Valor"],
        [
            [
                d.consecutivo,
                _f(d.fecha_generado),
                d.generado_por or "",
                recepciones.get(d.oficio_recepcion_id, ""),
                d.total_facturas,
                d.total_valor,
            ]
            for d in sorted(devs, key=lambda d: (d.anio, d.numero))
        ],
        anchos=[26, 18, 22, 24, 10, 18],
    )

    # --- HISTORIAL (trazabilidad completa) ---------------------------
    eventos = (
        db.query(FacturaEventoRecord)
        .order_by(FacturaEventoRecord.creado_en, FacturaEventoRecord.id)
        .all()
    )
    _hoja(
        "HISTORIAL",
        [
            "Fecha",
            "Factura",
            "Envío",
            "Oficio",
            "Evento",
            "Resultado",
            "Auditor",
            "Motivo",
            "Observaciones",
            "Valor",
        ],
        [
            [
                _f(e.creado_en),
                e.factura,
                e.envio or "",
                e.oficio_fhus or "",
                e.tipo_evento,
                e.estado_resultante or "",
                e.auditor or e.creado_por or "",
                e.motivo or "",
                e.observaciones or "",
                e.valor_snapshot or 0,
            ]
            for e in eventos
        ],
        anchos=[17, 16, 10, 22, 20, 16, 22, 40, 40, 14],
    )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = f"informe-preauditoria-{ahora.strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


@router.get("/estadisticas")
def estadisticas(
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    facturas = db.query(FacturaPreauditoriaRecord).all()
    total = len(facturas)
    radicar = sum(1 for f in facturas if f.resultado_actual == svc.RESULTADO_RADICAR)
    devueltas = sum(1 for f in facturas if f.resultado_actual == svc.RESULTADO_DEVUELTA)
    pendientes = total - radicar - devueltas
    auditadas = radicar + devueltas
    subsanadas = sum(1 for f in facturas if f.estado == svc.ESTADO_SUBSANADA)
    nuevamente = sum(1 for f in facturas if f.estado == svc.ESTADO_NUEV_DEVUELTA)

    # valores desde la fuente: un JOIN (evita IN masivo con miles de binds).
    valores = {
        fac: (val or 0)
        for fac, val in db.query(
            FacturaPreauditoriaRecord.factura, RadicacionCuentaRecord.valor
        ).outerjoin(
            RadicacionCuentaRecord,
            RadicacionCuentaRecord.factura == FacturaPreauditoriaRecord.factura,
        )
    }

    por_auditor: dict[str, dict] = {}
    for f in facturas:
        if not f.auditor:
            continue
        a = por_auditor.setdefault(
            f.auditor,
            {"auditor": f.auditor, "auditadas": 0, "radicar": 0, "devueltas": 0, "valor": 0.0},
        )
        a["auditadas"] += 1
        a["valor"] += valores.get(f.factura, 0)
        if f.resultado_actual == svc.RESULTADO_RADICAR:
            a["radicar"] += 1
        elif f.resultado_actual == svc.RESULTADO_DEVUELTA:
            a["devueltas"] += 1

    # Por entidad (top 10 por nº de facturas), con desglose de resultado.
    entidades: dict[str, dict] = {}
    ent_por_factura = {
        fac: (ent or "SIN ENTIDAD")
        for fac, ent in db.query(
            FacturaPreauditoriaRecord.factura, RadicacionCuentaRecord.entidad
        ).outerjoin(
            RadicacionCuentaRecord,
            RadicacionCuentaRecord.factura == FacturaPreauditoriaRecord.factura,
        )
    }
    for f in facturas:
        ent = ent_por_factura.get(f.factura, "SIN ENTIDAD")
        e = entidades.setdefault(
            ent,
            {
                "entidad": ent,
                "facturas": 0,
                "radicar": 0,
                "devueltas": 0,
                "pendientes": 0,
                "valor": 0.0,
            },
        )
        e["facturas"] += 1
        e["valor"] += valores.get(f.factura, 0)
        if f.resultado_actual == svc.RESULTADO_RADICAR:
            e["radicar"] += 1
        elif f.resultado_actual == svc.RESULTADO_DEVUELTA:
            e["devueltas"] += 1
        else:
            e["pendientes"] += 1
    por_entidad = sorted(entidades.values(), key=lambda e: e["facturas"], reverse=True)[:10]

    reincidentes = sorted(
        (
            {"factura": f.factura, "veces_devuelta": f.num_devoluciones}
            for f in facturas
            if f.num_devoluciones >= 2
        ),
        key=lambda r: r["veces_devuelta"],
        reverse=True,
    )
    en_limite = [
        {"factura": f.factura, "veces_devuelta": f.num_devoluciones}
        for f in facturas
        if f.num_devoluciones >= svc.MAX_DEVOLUCIONES
    ]

    # Conteos por oficio en UNA query agrupada (evita N+1 por oficio).
    oficios = db.query(OficioRecepcionRecord).all()
    totales = dict(
        db.query(
            FacturaPreauditoriaRecord.oficio_actual_id,
            sa_func.count(FacturaPreauditoriaRecord.id),
        ).group_by(FacturaPreauditoriaRecord.oficio_actual_id)
    )
    pend = dict(
        db.query(
            FacturaPreauditoriaRecord.oficio_actual_id,
            sa_func.count(FacturaPreauditoriaRecord.id),
        )
        .filter(FacturaPreauditoriaRecord.resultado_actual == svc.RESULTADO_PENDIENTE)
        .group_by(FacturaPreauditoriaRecord.oficio_actual_id)
    )
    conteo_semaforo = {"VERDE": 0, "AMARILLO": 0, "ROJO": 0, "VENCIDO": 0, "COMPLETADO": 0}
    for o in oficios:
        n_tot = totales.get(o.id, 0)
        completado = n_tot > 0 and pend.get(o.id, 0) == 0
        s = svc.calcular_semaforo(o.fecha_recibido, completado=completado)
        conteo_semaforo[s["estado"]] = conteo_semaforo.get(s["estado"], 0) + 1

    return {
        "total_facturas": total,
        "auditadas": auditadas,
        "pendientes": pendientes,
        "radicar": radicar,
        "devueltas": devueltas,
        "subsanadas": subsanadas,
        "nuevamente_devueltas": nuevamente,
        "tasa_devolucion": round(devueltas / auditadas, 4) if auditadas else 0,
        "valor_total": sum(valores.values()),
        "por_auditor": sorted(por_auditor.values(), key=lambda a: a["auditadas"], reverse=True),
        "por_entidad": por_entidad,
        "reincidentes": reincidentes[:50],
        "en_limite_devoluciones": en_limite,
        "semaforo_oficios": conteo_semaforo,
        "max_devoluciones": svc.MAX_DEVOLUCIONES,
    }


# ------------------------------------------------------------------
# Administración: borrar TODOS los datos del módulo
# ------------------------------------------------------------------


class LimpiarTodoIn(BaseModel):
    confirmacion: str = Field(..., description='Escriba exactamente "BORRAR TODO"')
    incluir_fuentes: bool = False


@router.post("/admin/limpiar-todo")
def limpiar_todo(
    body: LimpiarTodoIn,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_coordinador_o_admin),
):
    """Deja el módulo limpio para empezar a trabajar (solo SUPER_ADMIN/COORDINADOR).

    Borra oficios, facturas, historial, envíos y oficios de devolución.
    Con incluir_fuentes=true borra también las dos fuentes cargadas.
    Requiere escribir "BORRAR TODO" como confirmación.
    """
    if body.confirmacion.strip().upper() != "BORRAR TODO":
        raise HTTPException(400, 'Para confirmar escriba exactamente "BORRAR TODO".')
    conteos = svc.limpiar_modulo(db, incluir_fuentes=body.incluir_fuentes)
    return {"ok": True, "borrado": conteos}
