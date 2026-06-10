import re
import uuid
from typing import Optional
from datetime import datetime
from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, BackgroundTasks, Query
from sqlalchemy.orm import Session
from pydantic import BaseModel, Field

from app.core.tz import ahora_utc
from app.database import get_db, SessionLocal
from app.repositories.glosa_repository import GlosaRepository
from app.repositories.contrato_repository import ContratoRepository
from app.repositories.audit_repository import AuditRepository
from app.services.glosa_service import GlosaService
from app.core.config import get_settings
from app.core.logging_utils import set_request_id, logger
from app.api.deps import get_usuario_actual, get_auditor_o_superior, get_coordinador_o_admin
from app.services.rate_limit_ia import consumir_cupo_ia as _consumir_cupo_ia
from app.models.db import UsuarioRecord, GlosaRecord, ConceptoGlosaRecord, ContratoRecord
from app.utils.moneda import parse_valor_cop

router = APIRouter(prefix="/glosas", tags=["glosas"])


class GlosaFilaInput(BaseModel):
    fila: int
    texto: str
    eps: str
    fecha_radicacion: Optional[str] = None
    fecha_recepcion: Optional[str] = None


class ImportacionMasivaRequest(BaseModel):
    # Si se deja vacío o "AUTO", la EPS se detecta de la primera columna de cada fila.
    eps: Optional[str] = None
    texto_excel: str
    fecha_radicacion: Optional[str] = None
    fecha_recepcion: Optional[str] = None
    # IM F1.3: si se setea, las glosas creadas por el lote quedan
    # asignadas a ese gestor automáticamente. Si es None, quedan en
    # buzón colectivo (como antes).
    gestor_asignado_id: Optional[int] = None


# ─── Normalizador de nombres de EPS ──────────────────────────────────────────
# El Excel suele traer razones sociales completas (p. ej.
# "ENTIDAD PROMOTORA DE SALUD SANITAS S.A.S") que deben mapearse a la clave
# canónica que usa el resto del sistema (contratos, perfiles, aseguradoras).
_EPS_ALIASES: list[tuple[str, str]] = [
    # (substring a buscar en el texto upper, clave canónica)
    ("SANITAS", "SANITAS"),
    ("NUEVA EPS", "NUEVA EPS"),
    ("COOSALUD", "COOSALUD"),
    ("COMPENSAR", "COMPENSAR"),
    ("FAMISANAR", "FAMISANAR"),
    ("SALUD TOTAL", "SALUD TOTAL"),
    ("SURA", "SURA"),
    ("MUTUAL SER", "MUTUAL SER"),
    ("SAVIA SALUD", "SAVIA SALUD"),
    ("CAPITAL SALUD", "CAPITAL SALUD"),
    ("ASMET SALUD", "ASMET SALUD"),
    ("EMSSANAR", "EMSSANAR"),
    ("CAJACOPI", "CAJACOPI"),
    ("COMFAMILIAR", "COMFAMILIAR"),
    ("COMFENALCO", "COMFENALCO"),
    ("ECOOPSOS", "ECOOPSOS"),
    ("ALIANSALUD", "ALIANSALUD"),
    ("ANAS WAYUU", "ANAS WAYUU"),
    ("DUSAKAWI", "DUSAKAWI"),
    ("PIJAOS SALUD", "PIJAOS SALUD"),
    ("MALLAMAS", "MALLAMAS"),
    ("CAPRESOCA", "CAPRESOCA"),
    ("SERVICIO OCCIDENTAL DE SALUD", "SOS"),
    ("SOS ", "SOS"),
    # Aseguradoras (SOAT / ARL / pólizas)
    ("SEGUROS COMERCIALES BOLIVAR", "SEGUROS BOLIVAR"),
    ("SEGUROS BOLIVAR", "SEGUROS BOLIVAR"),
    ("SEGUROS DEL ESTADO", "SEGUROS DEL ESTADO"),
    ("SEGUROS GENERALES SURAMERICANA", "SURA"),
    ("MAPFRE", "MAPFRE"),
    ("AXA COLPATRIA", "AXA COLPATRIA"),
    ("LA PREVISORA", "FOMAG"),
    ("FIDEICOMISOS PATRIMONIOS", "FOMAG"),
    ("FOMAG", "FOMAG"),
    # Regímenes especiales
    ("DISPENSARIO MEDICO", "DISPENSARIO MEDICO"),
    ("FUERZAS MILITARES", "DISPENSARIO MEDICO"),
    ("POLICIA NACIONAL", "SANIDAD POLICIA"),
    ("SANIDAD POLICIA", "SANIDAD POLICIA"),
    ("UNIDAD DE SERVICIOS PENITENCIARIOS", "USPEC"),
    ("USPEC", "USPEC"),
    ("MAGISTERIO", "FOMAG"),
]


def _normalizar_eps(valor: str) -> str:
    """Convierte la razón social que viene en el Excel a la clave canónica
    usada por el sistema. Si no encuentra match, devuelve el texto tal cual
    en mayúsculas (sin perder información, el analizador luego trabaja con eso)."""
    if not valor:
        return ""
    texto = re.sub(r"\s+", " ", str(valor).upper().strip())
    for patron, clave in _EPS_ALIASES:
        if patron in texto:
            return clave
    return texto


class GenerarLoteRequest(BaseModel):
    glosa_ids: list[int]
    sobrescribir: bool = False  # si True regenera aunque ya tenga dictamen


class RefinarRequest(BaseModel):
    mensaje: str
    guardar: bool = False  # si True persiste el dictamen refinado en la BD


class ValidarRequest(BaseModel):
    forzar: bool = False


class ReanalizarRequest(BaseModel):
    """R60 P2: petición de re-análisis sobre glosa existente.
    Sin duplicar la fila — actualiza el dictamen de la glosa actual."""

    tono: Optional[str] = "conciliador"
    modo_respuesta: Optional[str] = "defender"


class BulkActualizarEstadoRequest(BaseModel):
    """R71 P1: cambio masivo de estado. Útil cuando llega un Excel
    de respuesta de la EPS con N decisiones (LEVANTADAS, RATIFICADAS)
    para procesar de un golpe."""

    glosa_ids: list[int] = Field(..., min_length=1, max_length=500)
    nuevo_estado: str = Field(..., min_length=3, max_length=50)
    nota: Optional[str] = Field(default=None, max_length=300)


class BulkMoverPapeleraRequest(BaseModel):
    """R71 P2: mueve N glosas a la papelera en una sola transacción.
    Soporta dry_run para preview antes de ejecutar."""

    glosa_ids: list[int] = Field(..., min_length=1, max_length=200)
    motivo: Optional[str] = Field(default=None, max_length=300)
    dry_run: bool = False


def _limpiar_observacion(dictamen_html: str) -> str:
    """Extrae solo el texto del argumento jurídico del dictamen, quitando la
    tabla superior (código/valor/respuesta), los badges, la tabla de resumen
    de valores y la nota al pie de 'asistencia de IA'."""
    if not dictamen_html:
        return ""
    from html import unescape
    import re as _re

    txt = _re.sub(r"<[^>]+>", " ", dictamen_html)
    txt = _re.sub(r"\s+", " ", unescape(txt)).strip()

    # Cortar desde "ARGUMENTACIÓN JURÍDICA" (siempre precede al argumento real)
    for marker in ("ARGUMENTACIÓN JURÍDICA", "RESPUESTA A GLOSA"):
        if marker in txt:
            parts = txt.split(marker, 1)
            # Solo tomar lo que va después si el marker está cerca del inicio
            # (es el header de la tabla) o si hay muy poco texto antes.
            if len(parts) == 2 and len(parts[0]) < 500:
                txt = parts[1].strip()
                break

    # Cortar ANTES de la nota al pie de IA o del resumen de valores
    for cierre in (
        "Nota: Generado con asistencia",
        "Nota: Este documento constituye",
        "Nota: Generado con IA",
        "RESUMEN DE VALORES",
        "Valor objetado Valor aceptado",
    ):
        if cierre in txt:
            txt = txt.split(cierre)[0].strip()

    return txt.strip()


@router.get("/historial", response_model=list)
def historial(
    # le=500: sin tope, un limit=10_000_000 serializa la tabla completa
    # (dictamenes HTML incluidos) en una sola respuesta — DoS trivial
    # autenticado (auditoría jun-2026, P2 #9).
    limit: int = Query(50, ge=1, le=500),
    eps: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Historial detallado con todos los campos relevantes para vista IPS."""
    from app.main import _extraer_motivo_glosa
    from app.services.resolver_entidad import resolver_entidad_mostrar

    repo = GlosaRepository(db)
    glosas = repo.listar(limit=limit, eps=eps)
    items = []
    for g in glosas:
        obs_texto = _limpiar_observacion(g.dictamen)
        entidad_real = resolver_entidad_mostrar(
            eps=g.eps,
            tercero_nombre=getattr(g, "tercero_nombre", None),
            eps_codigo=getattr(g, "eps_codigo", None),
        )
        items.append(
            {
                "id": g.id,
                "fecha": g.creado_en.isoformat() if g.creado_en else None,
                "fecha_recepcion": g.fecha_recepcion.isoformat() if g.fecha_recepcion else None,
                "fecha_entrega": g.fecha_entrega.isoformat() if g.fecha_entrega else None,
                "entidad": entidad_real,
                "eps": g.eps,  # alias para compatibilidad (valor raw, sin resolver)
                "paciente": g.paciente,
                "factura": g.factura,
                "codigo_glosa": g.codigo_glosa,
                "concepto_glosa": g.concepto_glosa,
                "cups": g.cups_servicio,
                "servicio": g.servicio_descripcion,
                "valor_objetado": g.valor_objetado,
                "valor_aceptado": g.valor_aceptado,
                "glosa_original": _extraer_motivo_glosa(g.texto_glosa_original or ""),
                "codigo_respuesta": g.codigo_respuesta,
                "observacion": obs_texto,
                "etapa": g.etapa,
                "estado": g.estado,
                "dictamen": g.dictamen,
                "dias_restantes": g.dias_restantes,
                "creado_en": g.creado_en.isoformat() if g.creado_en else None,
            }
        )
    return items


@router.get("/historial-paginado")
def historial_paginado(
    page: int = 1,
    per_page: int = Query(20, ge=1, le=100),
    eps: Optional[str] = None,
    estado: Optional[str] = None,
    search: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    valor_min: Optional[float] = None,
    valor_max: Optional[float] = None,
    tipo: Optional[str] = None,
    semaforo: Optional[str] = None,
    workflow: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Historial con paginación y filtros avanzados (vista detallada IPS)."""
    from app.main import _extraer_motivo_glosa

    repo = GlosaRepository(db)
    resultado = repo.listar_paginado(
        page=page,
        per_page=per_page,
        eps=eps,
        estado=estado,
        search=search,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        valor_min=valor_min,
        valor_max=valor_max,
        tipo=tipo,
        semaforo=semaforo,
        workflow=workflow,
    )

    from app.services.resolver_entidad import resolver_entidad_mostrar

    items = []
    for g in resultado["items"]:
        obs_texto = _limpiar_observacion(g.dictamen)
        entidad_real = resolver_entidad_mostrar(
            eps=g.eps,
            tercero_nombre=getattr(g, "tercero_nombre", None),
            eps_codigo=getattr(g, "eps_codigo", None),
        )
        items.append(
            {
                "id": g.id,
                "eps": g.eps,
                "entidad": entidad_real,
                "paciente": g.paciente,
                "factura": g.factura,
                "codigo_glosa": g.codigo_glosa,
                "concepto_glosa": g.concepto_glosa,
                "cups": g.cups_servicio,
                "servicio": g.servicio_descripcion,
                "valor_objetado": g.valor_objetado,
                "valor_aceptado": g.valor_aceptado,
                "glosa_original": _extraer_motivo_glosa(g.texto_glosa_original or ""),
                "codigo_respuesta": g.codigo_respuesta,
                "observacion": obs_texto,
                "etapa": g.etapa,
                "estado": g.estado,
                "dias_restantes": g.dias_restantes,
                "fecha_recepcion": g.fecha_recepcion.isoformat() if g.fecha_recepcion else None,
                "fecha_entrega": g.fecha_entrega.isoformat() if g.fecha_entrega else None,
                "creado_en": g.creado_en.isoformat() if g.creado_en else None,
            }
        )

    return {
        "items": items,
        "total": resultado["total"],
        "page": resultado["page"],
        "per_page": resultado["per_page"],
        "pages": resultado["pages"],
    }


@router.get("/exportar-json")
def exportar_json(
    eps: Optional[str] = None,
    estado: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    valor_min: Optional[float] = None,
    valor_max: Optional[float] = None,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R92 P1: export streaming en formato NDJSON (newline-delimited JSON).

    NDJSON > JSON-array para datos grandes:
      - Cada línea es un objeto independiente parseable
      - Permite streaming sin cargar todo en memoria
      - Compatible con jq, pandas.read_json(lines=True), etc.

    Útil para integrar con BI/data warehouse (Snowflake, BigQuery)
    que aceptan NDJSON nativo.

    Filtros opcionales: eps, estado, fecha_desde, fecha_hasta,
    valor_min, valor_max.
    """
    import json

    from fastapi.responses import StreamingResponse

    repo = GlosaRepository(db)
    glosas = repo.listar_para_export(
        eps=eps,
        estado=estado,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        valor_min=valor_min,
        valor_max=valor_max,
    )

    def _generar():
        for g in glosas:
            obj = {
                "id": g.id,
                "creado_en": g.creado_en.isoformat() if g.creado_en else None,
                "eps": g.eps,
                "paciente": g.paciente,
                "factura": g.factura,
                "codigo_glosa": g.codigo_glosa,
                "valor_objetado": float(g.valor_objetado or 0),
                "valor_aceptado": float(g.valor_aceptado or 0),
                "valor_recuperado": float(g.valor_recuperado or 0),
                "etapa": g.etapa,
                "estado": g.estado,
                "decision_eps": g.decision_eps,
                "dias_restantes": g.dias_restantes,
                "gestor_nombre": g.gestor_nombre,
                "fecha_vencimiento": (
                    g.fecha_vencimiento.isoformat() if g.fecha_vencimiento else None
                ),
                "fecha_decision_eps": (
                    g.fecha_decision_eps.isoformat() if g.fecha_decision_eps else None
                ),
            }
            yield json.dumps(obj, ensure_ascii=False) + "\n"

    fname = f"glosas-{ahora_utc().strftime('%Y%m%d-%H%M%S')}.ndjson"
    return StreamingResponse(
        _generar(),
        media_type="application/x-ndjson",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/exportar-xlsx")
def exportar_xlsx(
    eps: Optional[str] = None,
    estado: Optional[str] = None,
    search: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    valor_min: Optional[float] = None,
    valor_max: Optional[float] = None,
    tipo: Optional[str] = None,
    semaforo: Optional[str] = None,
    workflow: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Exporta el historial filtrado a XLSX con las 13 columnas IPS + observación."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    repo = GlosaRepository(db)
    glosas_raw = repo.listar_para_export(
        eps=eps,
        estado=estado,
        search=search,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        valor_min=valor_min,
        valor_max=valor_max,
        tipo=tipo,
        semaforo=semaforo,
        workflow=workflow,
    )

    # R-export 27-abr-2026: deduplicar por (factura, código, cups, etapa)
    # quedándonos con la versión más reciente. Antes el Excel exportaba
    # TODAS las versiones (13 filas para una sola glosa que se reanalizó
    # 13 veces). Ahora 1 fila por glosa única — la última.
    visto = {}
    for g in glosas_raw:
        clave = (
            (g.factura or "").strip().upper(),
            (g.codigo_glosa or "").strip().upper(),
            (g.cups_servicio or "").strip().upper(),
            (g.etapa or "").strip().upper(),
        )
        # Solo conservamos si NO había una entrada para esta clave o
        # si la nueva es más reciente (creado_en mayor).
        prev = visto.get(clave)
        if prev is None or (g.creado_en and prev.creado_en and g.creado_en > prev.creado_en):
            visto[clave] = g
    glosas = sorted(
        visto.values(),
        key=lambda x: x.creado_en or 0,
        reverse=True,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial Glosas HUS"

    headers = [
        "ID",
        "Fecha Creación",
        "EPS/Entidad",
        "Paciente",
        "Factura",
        "Código Glosa",
        "Concepto",
        "CUPS",
        "Servicio",
        "Valor Objetado",
        "Valor Aceptado",
        "Valor Recuperado",
        "Código Respuesta",
        "Observación EPS",
        "Dictamen HUS",
        "Etapa",
        "Estado",
        "Workflow",
        "Semáforo",
        "Días Restantes",
        "Fecha Recepción",
        "Fecha Entrega",
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0B5D8A", end_color="0B5D8A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for g in glosas:
        # Observación EPS: lo que la EPS registró al glosar (lo que el
        # auditor VE y REGISTRA en el sistema). Prioridad:
        # observacion_eps (campo explícito) → texto_glosa_original (texto del
        # Excel DGH o del editor) → concepto_glosa (descripción canónica).
        obs_eps_raw = (
            (getattr(g, "observacion_eps", None) or "").strip()
            or (g.texto_glosa_original or "").strip()
            or (g.concepto_glosa or "").strip()
        )
        # Dictamen: el texto limpio (sin HTML) generado por la defensa.
        dictamen_txt = _limpiar_observacion(g.dictamen) or ""
        recuperado = (g.valor_objetado or 0) - (g.valor_aceptado or 0)
        ws.append(
            [
                g.id,
                g.creado_en.strftime("%Y-%m-%d %H:%M") if g.creado_en else "",
                g.eps or "",
                g.paciente or "",
                g.factura or "",
                g.codigo_glosa or "",
                g.concepto_glosa or "",
                g.cups_servicio or "",
                g.servicio_descripcion or "",
                float(g.valor_objetado or 0),
                float(g.valor_aceptado or 0),
                float(recuperado),
                g.codigo_respuesta or "",
                obs_eps_raw[:600] if obs_eps_raw else "",
                dictamen_txt[:800] if dictamen_txt else "",
                g.etapa or "",
                g.estado or "",
                g.workflow_state or "",
                g.prioridad or "",
                g.dias_restantes if g.dias_restantes is not None else "",
                g.fecha_recepcion.strftime("%Y-%m-%d") if g.fecha_recepcion else "",
                g.fecha_entrega.strftime("%Y-%m-%d") if g.fecha_entrega else "",
            ]
        )

    # Ajuste de anchos (22 columnas: Observación EPS 60, Dictamen HUS 80)
    widths = [6, 18, 22, 28, 14, 12, 26, 10, 32, 14, 14, 14, 12, 60, 80, 14, 14, 14, 10, 10, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Formato condicional premium (requirement #8 del user):
    # - Días Restantes columna S (col 19): rojo si ≤ 3, amarillo si ≤ 7, verde si > 7
    # - Estado columna P (col 16): resaltado según valor
    # - Valor Recuperado columna L (col 12): verde si > 0, rojo si negativo
    try:
        from openpyxl.styles import PatternFill
        from openpyxl.formatting.rule import CellIsRule, FormulaRule

        # Aplicar desde fila 2 hasta el final
        last_row = ws.max_row
        if last_row > 1:
            rango_dias = f"S2:S{last_row}"
            rango_recup = f"L2:L{last_row}"
            rango_estado = f"P2:P{last_row}"
            fill_rojo = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
            fill_amarillo = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            fill_verde = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
            fill_verde_fuerte = PatternFill(
                start_color="A7F3D0", end_color="A7F3D0", fill_type="solid"
            )
            # Días restantes — semáforo
            ws.conditional_formatting.add(
                rango_dias, CellIsRule(operator="lessThanOrEqual", formula=["3"], fill=fill_rojo)
            )
            ws.conditional_formatting.add(
                rango_dias, CellIsRule(operator="between", formula=["4", "7"], fill=fill_amarillo)
            )
            ws.conditional_formatting.add(
                rango_dias, CellIsRule(operator="greaterThan", formula=["7"], fill=fill_verde)
            )
            # Valor recuperado
            ws.conditional_formatting.add(
                rango_recup,
                CellIsRule(operator="greaterThan", formula=["0"], fill=fill_verde_fuerte),
            )
            # Estado
            ws.conditional_formatting.add(
                rango_estado, FormulaRule(formula=['EXACT(P2,"CERRADA")'], fill=fill_verde)
            )
            ws.conditional_formatting.add(
                rango_estado, FormulaRule(formula=['EXACT(P2,"RATIFICADA")'], fill=fill_rojo)
            )
            ws.conditional_formatting.add(
                rango_estado, FormulaRule(formula=['EXACT(P2,"EXTEMPORANEA")'], fill=fill_amarillo)
            )
    except Exception:
        # Sin formato condicional el Excel sigue siendo válido.
        pass

    # Registrar auditoría de la exportación
    AuditRepository(db).registrar(
        usuario_email=current_user.email,
        usuario_rol=current_user.rol,
        accion="EXPORTAR_XLSX",
        tabla="historial",
        detalle=f"Registros exportados: {len(glosas)}",
    )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"historial_glosas_hus_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/codigos-respuesta-catalogo")
def codigos_respuesta_catalogo(
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R137 P1: catálogo de códigos de respuesta IPS (Res 2284/2023).

    Códigos RE oficiales que IPS usa para responder a una glosa
    o devolución de la EPS:
      - RE9901: Glosa no aceptada (defensa total) — más común
      - RE9801: Glosa aceptada parcialmente
      - RE9602: Glosa injustificada al 100%
      - RE9502: Glosa improcedente por extemporánea
      - etc.

    Útil para que el frontend renderice un dropdown contextual
    al responder una glosa.

    Devuelve los códigos con descripción + clasificación
    funcional (DEFENSA / ACEPTACION / EXTEMPORANEA / OTRO).
    """
    from app.services.catalogo_glosas import CODIGOS_RESPUESTA

    DEFENSA = {"RE9901", "RE9602", "RE9502", "RE9601", "RE9501"}
    ACEPTACION = {"RE9701", "RE9801"}
    EXTEMPORANEA = {"RE2201", "RE2202"}

    items = []
    for codigo, descripcion in sorted(CODIGOS_RESPUESTA.items()):
        if codigo in DEFENSA:
            tipo = "DEFENSA"
        elif codigo in ACEPTACION:
            tipo = "ACEPTACION"
        elif codigo in EXTEMPORANEA:
            tipo = "EXTEMPORANEA"
        else:
            tipo = "OTRO"
        items.append(
            {
                "codigo": codigo,
                "descripcion": descripcion,
                "tipo_funcional": tipo,
            }
        )

    return {
        "regulacion": "Resolución 2284/2023 — Códigos de respuesta IPS",
        "total_codigos": len(items),
        "items": items,
    }


@router.get("/codigos-glosa-catalogo")
def codigos_glosa_catalogo(
    grupo: Optional[str] = None,
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R136 P2: catálogo de códigos de glosa Resolución 2284/2023.

    Expone el catálogo oficial del Manual Único usado por la IA
    para que el frontend pueda:
      - Mostrar autocomplete de códigos
      - Renderizar tooltip con descripción al hover
      - Validar que un código sea oficial antes de enviar

    Param `grupo` opcional filtra por familia (FA, TA, SO, AU,
    CO, CL, SA).

    Devuelve:
      - total_codigos
      - por_grupo: counts
      - items: [{codigo, grupo, descripcion}]
    """
    from app.services.catalogo_glosas import (
        CODIGOS_AU,
        CODIGOS_CL,
        CODIGOS_CO,
        CODIGOS_FA,
        CODIGOS_SA,
        CODIGOS_SO,
        CODIGOS_TA,
    )

    grupos = {
        "FA": CODIGOS_FA,
        "TA": CODIGOS_TA,
        "SO": CODIGOS_SO,
        "AU": CODIGOS_AU,
        "CO": CODIGOS_CO,
        "CL": CODIGOS_CL,
        "SA": CODIGOS_SA,
    }

    if grupo:
        g_upper = grupo.upper()
        if g_upper not in grupos:
            raise HTTPException(
                400,
                f"grupo inválido. Válidos: {sorted(grupos.keys())}",
            )
        grupos = {g_upper: grupos[g_upper]}

    items = []
    por_grupo: dict[str, int] = {}
    for nombre_grupo, dic in grupos.items():
        for codigo, descripcion in dic.items():
            items.append(
                {
                    "codigo": codigo,
                    "grupo": nombre_grupo,
                    "descripcion": descripcion,
                }
            )
            por_grupo[nombre_grupo] = por_grupo.get(nombre_grupo, 0) + 1

    items.sort(key=lambda x: x["codigo"])

    return {
        "regulacion": "Resolución 2284/2023 — Manual Único de Glosas",
        "total_codigos": len(items),
        "por_grupo": por_grupo,
        "filtro_grupo": grupo,
        "items": items,
    }


@router.get("/estados-disponibles")
def estados_disponibles(
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R136 P1: catálogo machine-readable de estados de glosa.

    Devuelve la lista oficial de estados con:
      - clave (valor en BD)
      - nombre amigable
      - descripcion
      - es_cerrado (bool)
      - color sugerido para UI (semáforo)

    Útil para que el frontend renderice dropdowns y badges
    consistentes sin hardcodear listas que se desactualizan.
    """
    estados = [
        {
            "clave": "RADICADA",
            "nombre": "Radicada",
            "descripcion": "Glosa recién recibida, esperando respuesta HUS.",
            "es_cerrado": False,
            "color": "AMARILLO",
        },
        {
            "clave": "RESPONDIDA",
            "nombre": "Respondida",
            "descripcion": ("HUS ya respondió la glosa, esperando decisión EPS."),
            "es_cerrado": False,
            "color": "AZUL",
        },
        {
            "clave": "RATIFICADA",
            "nombre": "Ratificada por EPS",
            "descripcion": (
                "EPS sostuvo la glosa tras respuesta HUS. Pasa a siguiente etapa o se acepta."
            ),
            "es_cerrado": False,
            "color": "ROJO",
        },
        {
            "clave": "LEVANTADA",
            "nombre": "Levantada (HUS ganó)",
            "descripcion": "EPS retiró la glosa. HUS recupera el valor.",
            "es_cerrado": True,
            "color": "VERDE",
        },
        {
            "clave": "ACEPTADA",
            "nombre": "Aceptada por HUS",
            "descripcion": "HUS aceptó la glosa. EPS no paga ese ítem.",
            "es_cerrado": True,
            "color": "GRIS",
        },
        {
            "clave": "CONCILIADA",
            "nombre": "Conciliada bilateralmente",
            "descripcion": ("HUS y EPS llegaron a acuerdo en audiencia bilateral."),
            "es_cerrado": True,
            "color": "AZUL",
        },
        {
            "clave": "ARCHIVADA",
            "nombre": "Archivada",
            "descripcion": ("Glosa retirada del flujo activo (sin valor a defender)."),
            "es_cerrado": True,
            "color": "GRIS",
        },
        {
            "clave": "EXTEMPORANEA",
            "nombre": "Extemporánea",
            "descripcion": ("EPS objetó fuera del término legal. HUS puede rechazar."),
            "es_cerrado": False,
            "color": "AMARILLO",
        },
    ]

    return {
        "total": len(estados),
        "estados": estados,
    }


@router.get("/sin-codigo-glosa")
def glosas_sin_codigo_glosa(
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R209 P1: glosas no-cerradas sin codigo_glosa configurado.

    Detecta glosas que se importaron mal o que les falta el
    código canónico Resolución 2284. Sin código no se puede:
      - Buscar plantillas Gold
      - Aplicar reglas de la IA
      - Generar dictamen consistente

    Útil como cola de saneamiento.

    Devuelve hasta `limit` glosas no-cerradas sin codigo_glosa.

    Declarado ANTES de /{glosa_id} para evitar collision.
    """
    ESTADOS_CERRADOS = {"ACEPTADA", "LEVANTADA", "ARCHIVADA", "CONCILIADA"}

    glosas = (
        db.query(GlosaRecord)
        .filter(~GlosaRecord.estado.in_(ESTADOS_CERRADOS))
        .filter((GlosaRecord.codigo_glosa.is_(None)) | (GlosaRecord.codigo_glosa == ""))
        .order_by(GlosaRecord.creado_en.desc())
        .limit(int(limit))
        .all()
    )

    items = []
    for g in glosas:
        items.append(
            {
                "id": g.id,
                "creado_en": (g.creado_en.isoformat() if g.creado_en else None),
                "eps": g.eps,
                "factura": g.factura,
                "valor_objetado": float(g.valor_objetado or 0),
                "estado": g.estado,
            }
        )

    return {
        "total_sin_codigo": len(glosas),
        "items": items,
    }


@router.get("/factura-detalle")
def factura_detalle(
    factura: str = Query(..., min_length=1),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R204 P1: detalle de TODAS las glosas asociadas a una factura.

    Útil para: "esta factura tiene 8 glosas distintas, agruparlas
    para dar respuesta unificada".

    Diferente a /buscar-radicado (por radicado) y a
    /stats/facturas-hot (top facturas con muchas glosas):
    aquí drill-down a UNA factura específica.
    """
    glosas = (
        db.query(GlosaRecord)
        .filter(GlosaRecord.factura == factura)
        .order_by(GlosaRecord.creado_en.asc())
        .all()
    )

    valor_obj = sum(float(g.valor_objetado or 0) for g in glosas)
    valor_rec = sum(float(g.valor_recuperado or 0) for g in glosas)

    items = []
    for g in glosas:
        items.append(
            {
                "id": g.id,
                "creado_en": (g.creado_en.isoformat() if g.creado_en else None),
                "eps": g.eps,
                "codigo_glosa": g.codigo_glosa,
                "estado": g.estado,
                "valor_objetado": float(g.valor_objetado or 0),
                "valor_recuperado": float(g.valor_recuperado or 0),
            }
        )

    return {
        "factura_buscada": factura,
        "total_glosas": len(glosas),
        "valor_objetado_total": int(valor_obj),
        "valor_recuperado_total": int(valor_rec),
        "items": items,
    }


@router.get("/buscar-radicado")
def buscar_por_radicado(
    radicado: str = Query(..., min_length=1),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R152 P2: búsqueda exacta por número de radicado.

    Diferente a /buscar/{termino} (búsqueda fuzzy en factura/ID/
    consecutivo): aquí solo numero_radicado, exact match.

    Útil para auditor que tiene un número de radicado en mano
    y necesita la glosa rápido sin abrir el menú avanzado.

    Param `radicado`: número exacto.
    """
    glosas = db.query(GlosaRecord).filter(GlosaRecord.numero_radicado == radicado).all()

    items = []
    for g in glosas:
        items.append(
            {
                "id": g.id,
                "creado_en": (g.creado_en.isoformat() if g.creado_en else None),
                "eps": g.eps,
                "factura": g.factura,
                "numero_radicado": g.numero_radicado,
                "estado": g.estado,
                "valor_objetado": float(g.valor_objetado or 0),
            }
        )

    return {
        "radicado_buscado": radicado,
        "encontradas": len(items),
        "items": items,
    }


@router.get("/buscar-similares-texto")
def buscar_similares_texto(
    texto: str = Query(..., min_length=10, max_length=2000),
    top: int = Query(10, ge=1, le=50),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R103 P2: búsqueda de glosas con texto similar al dado.

    Útil para auditor que recibe glosa nueva: "¿hemos visto algo
    parecido antes?". Permite reusar respuestas/dictámenes
    previos como punto de partida.

    Algoritmo: Jaccard similarity sobre tokens del texto_glosa_original.
    Liviano, sin dependencias ML — bueno para datasets pequeños/medianos.

    Devuelve hasta `top` glosas con score 0-1 (1 = idéntico).
    Solo glosas con texto_glosa_original no-vacío.
    """
    import re

    def _tokenizar(s: str) -> set[str]:
        s = (s or "").lower()
        # Tokens alfanuméricos de >=3 caracteres (filtra "el", "de", etc.)
        return {t for t in re.findall(r"\w+", s) if len(t) >= 3}

    tokens_query = _tokenizar(texto)
    if not tokens_query:
        return {"total_evaluadas": 0, "items": []}

    candidatas = db.query(GlosaRecord).filter(GlosaRecord.texto_glosa_original.isnot(None)).all()

    items = []
    for g in candidatas:
        tokens_g = _tokenizar(g.texto_glosa_original or "")
        if not tokens_g:
            continue
        union = tokens_query | tokens_g
        inter = tokens_query & tokens_g
        score = len(inter) / len(union) if union else 0
        if score < 0.05:  # threshold mínimo
            continue
        items.append(
            {
                "id": g.id,
                "eps": g.eps,
                "codigo_glosa": g.codigo_glosa,
                "estado": g.estado,
                "score_similitud": round(score, 4),
                "preview": (g.texto_glosa_original or "")[:200],
            }
        )

    items.sort(key=lambda x: x["score_similitud"], reverse=True)

    return {
        "total_evaluadas": len(candidatas),
        "total_con_score_minimo": len(items),
        "items": items[:top],
    }


@router.get("/buscar-avanzado")
def buscar_avanzado(
    eps: Optional[str] = None,
    paciente: Optional[str] = None,
    factura: Optional[str] = None,
    codigo_glosa: Optional[str] = None,
    estado: Optional[str] = None,
    etapa: Optional[str] = None,
    gestor: Optional[str] = None,
    valor_min: Optional[float] = None,
    valor_max: Optional[float] = None,
    fecha_desde: Optional[str] = Query(None, description="ISO YYYY-MM-DD"),
    fecha_hasta: Optional[str] = Query(None, description="ISO YYYY-MM-DD"),
    limit: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R94 P1: búsqueda multi-campo combinable (AND entre filtros).

    Complementa /buscar/{termino} (un solo término en factura/ID)
    permitiendo consultas precisas tipo:
      "glosas SANITAS de Pedro entre 1M-5M de marzo"

    Filtros opcionales combinables vía AND. Strings usan ILIKE %x%
    (búsqueda parcial case-insensitive). Fechas en ISO YYYY-MM-DD.

    Devuelve hasta `limit` resultados (default 50, max 500), ordenados
    DESC por creado_en.
    """
    from datetime import datetime, timezone

    q = db.query(GlosaRecord)
    if eps:
        q = q.filter(GlosaRecord.eps.ilike(f"%{eps}%"))
    if paciente:
        q = q.filter(GlosaRecord.paciente.ilike(f"%{paciente}%"))
    if factura:
        q = q.filter(GlosaRecord.factura.ilike(f"%{factura}%"))
    if codigo_glosa:
        q = q.filter(GlosaRecord.codigo_glosa.ilike(f"%{codigo_glosa}%"))
    if estado:
        q = q.filter(GlosaRecord.estado == estado.upper())
    if etapa:
        q = q.filter(GlosaRecord.etapa.ilike(f"%{etapa}%"))
    if gestor:
        q = q.filter(GlosaRecord.gestor_nombre.ilike(f"%{gestor}%"))
    if valor_min is not None:
        q = q.filter(GlosaRecord.valor_objetado >= valor_min)
    if valor_max is not None:
        q = q.filter(GlosaRecord.valor_objetado <= valor_max)
    if fecha_desde:
        try:
            dt = datetime.strptime(fecha_desde, "%Y-%m-%d").replace(
                tzinfo=timezone.utc,
            )
            q = q.filter(GlosaRecord.creado_en >= dt)
        except ValueError:
            raise HTTPException(400, "fecha_desde debe ser YYYY-MM-DD")
    if fecha_hasta:
        try:
            dt = datetime.strptime(fecha_hasta, "%Y-%m-%d").replace(
                tzinfo=timezone.utc,
            )
            q = q.filter(GlosaRecord.creado_en <= dt)
        except ValueError:
            raise HTTPException(400, "fecha_hasta debe ser YYYY-MM-DD")

    total = q.count()
    glosas = q.order_by(GlosaRecord.creado_en.desc()).limit(limit).all()

    return {
        "total_coincidencias": total,
        "limit": int(limit),
        "items": [
            {
                "id": g.id,
                "creado_en": g.creado_en.isoformat() if g.creado_en else None,
                "eps": g.eps,
                "paciente": g.paciente,
                "factura": g.factura,
                "codigo_glosa": g.codigo_glosa,
                "valor_objetado": float(g.valor_objetado or 0),
                "estado": g.estado,
                "etapa": g.etapa,
                "gestor_nombre": g.gestor_nombre,
            }
            for g in glosas
        ],
    }


@router.get("/buscar/{termino}")
def buscar_por_id_o_factura(
    termino: str,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Busca una glosa por ID interno, factura, consecutivo DGH o radicado.
    Útil para formularios donde el auditor no sabe el ID interno.

    Devuelve lista de coincidencias (puede ser 0, 1 o varias)."""
    termino = (termino or "").strip()
    if not termino:
        raise HTTPException(400, "Término vacío")

    from sqlalchemy import or_

    q = db.query(GlosaRecord)

    # Si es número puro, intentar como ID interno primero
    matches = []
    if termino.isdigit():
        g = q.filter(GlosaRecord.id == int(termino)).first()
        if g:
            matches.append(g)

    # Además buscar por factura / consecutivo / radicado (incluye partial)
    extra = (
        db.query(GlosaRecord)
        .filter(
            or_(
                GlosaRecord.factura.ilike(f"%{termino}%"),
                GlosaRecord.consecutivo_dgh.ilike(f"%{termino}%"),
                GlosaRecord.numero_radicado.ilike(f"%{termino}%"),
            )
        )
        .order_by(GlosaRecord.creado_en.desc())
        .limit(10)
        .all()
    )
    ya = {m.id for m in matches}
    for g in extra:
        if g.id not in ya:
            matches.append(g)

    return [
        {
            "id": g.id,
            "eps": g.eps,
            "factura": g.factura,
            "consecutivo_dgh": g.consecutivo_dgh,
            "numero_radicado": g.numero_radicado,
            "codigo_glosa": g.codigo_glosa,
            "paciente": g.paciente,
            "valor_objetado": float(g.valor_objetado or 0),
            "estado": g.estado,
            "creado_en": g.creado_en.isoformat() if g.creado_en else None,
        }
        for g in matches[:10]
    ]


@router.post("/generar-lote")
async def generar_lote(
    data: GenerarLoteRequest,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_auditor_o_superior),
):
    """Genera respuestas IA en lote para varias glosas pendientes.

    Toma las glosas por ID, reconstruye el input a partir de los campos
    guardados (texto_glosa_original, eps, etapa, fechas, factura) y llama
    al servicio para producir el dictamen. Las glosas que ya tienen
    dictamen se saltan salvo que `sobrescribir=True`.

    Ejecuta hasta 5 en paralelo con `asyncio.Semaphore` para no saturar Groq.
    """
    import asyncio
    from app.models.schemas import GlosaInput

    if not data.glosa_ids:
        raise HTTPException(400, "Lista de IDs vacía")
    if len(data.glosa_ids) > 100:
        raise HTTPException(400, "Máximo 100 glosas por lote")

    repo = GlosaRepository(db)
    contratos = ContratoRepository(db).como_dict()

    cfg = get_settings()
    service = GlosaService(
        groq_api_key=cfg.groq_api_key,
        anthropic_api_key=cfg.anthropic_api_key,
        primary_ai=cfg.primary_ai,
        anthropic_model=cfg.anthropic_model,
        groq_model=cfg.groq_model,
        gemini_api_key=cfg.gemini_api_key,
        gemini_model=cfg.gemini_model,
    )

    sem = asyncio.Semaphore(5)
    resumen = {
        "total": len(data.glosa_ids),
        "procesadas": 0,
        "saltadas": 0,
        "fallidas": 0,
        "detalle_fallidas": [],
        "detalle_saltadas": [],
    }

    async def _procesar_una(gid: int):
        async with sem:
            g = repo.obtener_por_id(gid)
            if not g:
                resumen["fallidas"] += 1
                resumen["detalle_fallidas"].append({"id": gid, "error": "no encontrada"})
                return
            if g.dictamen and not data.sobrescribir:
                resumen["saltadas"] += 1
                resumen["detalle_saltadas"].append({"id": gid, "razon": "ya_tiene_dictamen"})
                return
            # Construir input desde los campos del registro
            texto = g.texto_glosa_original or ""
            if not texto:
                # Fallback robusto: ensamblar texto desde TODOS los campos
                # disponibles + conceptos vinculados (hojas I/R del DGH).
                # Necesario para glosas importadas sin texto_glosa_original
                # (caso típico cuando la importación masiva creó la glosa
                # con placeholder).
                partes = []
                if g.codigo_glosa:
                    partes.append(g.codigo_glosa)
                if g.concepto_glosa:
                    partes.append(g.concepto_glosa)
                if g.cups_servicio:
                    partes.append(f"CUPS {g.cups_servicio}")
                if g.servicio_descripcion:
                    partes.append(g.servicio_descripcion)
                if g.valor_objetado and float(g.valor_objetado) > 0:
                    partes.append(f"Valor objetado: ${int(g.valor_objetado):,}".replace(",", "."))
                if g.observacion_eps:
                    partes.append(g.observacion_eps)
                # Conceptos vinculados (hojas I/R del DGH)
                try:
                    from app.models.db import ConceptoGlosaRecord as _CG

                    conceptos = db.query(_CG).filter(_CG.glosa_id == g.id).limit(5).all()
                    for c in conceptos:
                        if c.codigo_glosa and c.codigo_glosa not in partes:
                            partes.append(c.codigo_glosa)
                        if c.cups_codigo:
                            partes.append(f"CUPS {c.cups_codigo}")
                        if c.cups_descripcion:
                            partes.append(c.cups_descripcion)
                        if c.observacion_eps:
                            partes.append(c.observacion_eps)
                        if c.valor_objetado and float(c.valor_objetado) > 0:
                            partes.append(
                                f"Valor objetado: ${int(c.valor_objetado):,}".replace(",", ".")
                            )
                except Exception:
                    pass
                texto = " — ".join([p for p in partes if p]).strip()
            if not texto:
                resumen["fallidas"] += 1
                resumen["detalle_fallidas"].append(
                    {"id": gid, "error": "sin texto ni conceptos vinculados"}
                )
                return
            try:
                gi = GlosaInput(
                    eps=g.eps or "SIN DEFINIR",
                    etapa=g.etapa or "RESPUESTA A GLOSA",
                    # Pydantic v2 GlosaInput valida date estricto (sin hora).
                    # El campo del modelo es DateTime (trae T00:01:00 del DGH),
                    # así que tomamos solo la fecha con .date().
                    fecha_radicacion=(
                        g.fecha_radicacion_factura.date().isoformat()
                        if g.fecha_radicacion_factura
                        else None
                    ),
                    fecha_recepcion=(
                        g.fecha_recepcion.date().isoformat() if g.fecha_recepcion else None
                    ),
                    valor_aceptado=str(int(g.valor_aceptado or 0)),
                    tabla_excel=texto,
                    numero_factura=g.factura,
                    numero_radicado=g.numero_radicado,
                )
                # Few-shots según (EPS, código)
                from app.api.routers.plantillas_gold import obtener_few_shot, marcar_usos

                pg = obtener_few_shot(db, eps=gi.eps, codigo_glosa=g.codigo_glosa or "", limite=2)
                # Pre-lookup de tarifa pactada — sin esto el LLM defaultea
                # al argumento "no existe contrato" aunque exista en BD.
                from app.services.tarifa_lookup_service import pre_lookup_tarifa

                info_tarifa_pre = pre_lookup_tarifa(
                    db=db,
                    cod_pref=g.codigo_glosa or "",
                    eps=gi.eps or "",
                    tabla_excel=texto,
                    contexto_pdf="",
                    req_id=f"lote-{gid}",
                )
                # Memoria del gestor para este auditor + caso similar
                hint_gestor = ""
                try:
                    from app.services.memoria_gestor import patron_gestor

                    pat = patron_gestor(
                        db,
                        autor_email=current_user.email,
                        codigo_glosa=g.codigo_glosa or "",
                        eps=gi.eps or "",
                    )
                    hint_gestor = pat.get("hint_para_prompt", "") or ""
                except Exception as e:
                    logger.debug(f"[lote {gid}] memoria_gestor falló: {e}")
                res = await service.analizar(
                    gi,
                    contexto_pdf="",
                    contratos_db=contratos,
                    few_shots=[p.argumento for p in pg],
                    info_tarifa=info_tarifa_pre,
                    hint_gestor=hint_gestor,
                )
                if pg:
                    marcar_usos(db, [p.id for p in pg])
                from datetime import datetime, timezone as _tz

                g.dictamen = res.dictamen
                g.dictamen_generado_en = datetime.now(_tz.utc)
                g.score = res.score
                g.modelo_ia = res.modelo_ia
                if not g.codigo_respuesta:
                    g.codigo_respuesta = res.tipo.replace("RESPUESTA ", "").strip() or None
                db.commit()
                resumen["procesadas"] += 1
            except Exception as e:
                resumen["fallidas"] += 1
                resumen["detalle_fallidas"].append({"id": gid, "error": str(e)[:200]})
                logger.error(f"Lote: falló glosa {gid}: {e}")

    await asyncio.gather(*[_procesar_una(gid) for gid in data.glosa_ids])

    AuditRepository(db).registrar(
        usuario_email=current_user.email,
        usuario_rol=current_user.rol,
        accion="GENERAR_LOTE",
        tabla="historial",
        detalle=(
            f"total={resumen['total']} procesadas={resumen['procesadas']} "
            f"saltadas={resumen['saltadas']} fallidas={resumen['fallidas']}"
        ),
    )
    return resumen


@router.post("/{glosa_id}/refinar")
async def refinar_dictamen_endpoint(
    glosa_id: int,
    data: RefinarRequest,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_auditor_o_superior),
    _cupo_ia: None = Depends(_consumir_cupo_ia),
):
    """Refina el dictamen de una glosa con instrucciones en lenguaje natural.

    Si `guardar=true`, reemplaza el argumento dentro del HTML actual y persiste.
    Si no, solo devuelve el texto refinado para preview en el modal.
    """
    if not data.mensaje or len(data.mensaje.strip()) < 3:
        raise HTTPException(400, "Mensaje demasiado corto")

    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")
    if not glosa.dictamen:
        raise HTTPException(400, "La glosa no tiene dictamen generado aún")

    cfg = get_settings()
    service = GlosaService(
        groq_api_key=cfg.groq_api_key,
        anthropic_api_key=cfg.anthropic_api_key,
        primary_ai=cfg.primary_ai,
        anthropic_model=cfg.anthropic_model,
        groq_model=cfg.groq_model,
        gemini_api_key=cfg.gemini_api_key,
        gemini_model=cfg.gemini_model,
    )
    nuevo_argumento = await service.refinar_dictamen(
        dictamen_actual_html=glosa.dictamen,
        mensaje_usuario=data.mensaje,
        eps=glosa.eps or "",
        codigo=glosa.codigo_glosa or "",
    )

    # Reemplazar el bloque de argumento dentro del HTML existente
    import re as _re

    nuevo_html = glosa.dictamen
    patron = _re.compile(
        r'(<div style="font-size:12px;line-height:1\.9;[^"]*">)(.*?)(</div>)',
        _re.DOTALL,
    )
    argumento_html = nuevo_argumento.replace("\n", "<br/>")
    nuevo_html, n = patron.subn(
        lambda m: m.group(1) + argumento_html + m.group(3),
        nuevo_html,
        count=1,
    )
    if n == 0:
        # Si no encontramos el bloque esperado, adjuntamos al final como fallback
        nuevo_html = glosa.dictamen + (
            "<div style='margin-top:12px;padding:12px;background:#ecfeff;"
            "border-left:4px solid #0891b2;border-radius:8px;font-size:12px;line-height:1.8;'>"
            "<b>REFINADO:</b><br/>" + argumento_html + "</div>"
        )

    if data.guardar:
        from datetime import datetime, timezone as _tz

        glosa.dictamen = nuevo_html
        glosa.dictamen_generado_en = datetime.now(_tz.utc)
        db.commit()
        AuditRepository(db).registrar(
            usuario_email=current_user.email,
            usuario_rol=current_user.rol,
            accion="REFINAR_IA",
            tabla="historial",
            registro_id=glosa_id,
            campo="dictamen",
            detalle=f"instrucción: {data.mensaje[:200]}",
        )
        # Guardar snapshot en historial de versiones
        try:
            from app.api.routers.versiones import guardar_version

            guardar_version(
                db=db,
                glosa_id=glosa_id,
                dictamen_html=nuevo_html,
                accion="REFINAR",
                autor_email=current_user.email,
                mensaje_refinar=data.mensaje[:500],
            )
        except Exception:
            pass

    return {
        "argumento_refinado": nuevo_argumento,
        "dictamen_html": nuevo_html,
        "guardado": data.guardar,
    }


class ValidarNormasInput(BaseModel):
    texto: str = Field(..., min_length=20, max_length=20000)


@router.post("/validar-normas")
def validar_normas_texto(
    data: ValidarNormasInput,
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Valida citas normativas en un texto libre (sin persistir).
    Útil para que el auditor chequee rápido un borrador."""
    from app.services.normativa import validar_citas

    return validar_citas(data.texto)


@router.post("/{glosa_id}/validar")
async def validar_pre_radicacion(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Valida el dictamen antes de radicarlo ante la EPS.

    Hace checks locales (placeholders, factura, normas esperadas,
    citas derogadas) + consulta a la IA para verificar solidez.
    Retorna score de calidad 0-100, hallazgos y si puede_radicar.
    """
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")
    if not glosa.dictamen:
        raise HTTPException(400, "La glosa aún no tiene dictamen generado")

    cfg = get_settings()
    service = GlosaService(
        groq_api_key=cfg.groq_api_key,
        anthropic_api_key=cfg.anthropic_api_key,
        primary_ai=cfg.primary_ai,
        anthropic_model=cfg.anthropic_model,
        groq_model=cfg.groq_model,
        gemini_api_key=cfg.gemini_api_key,
        gemini_model=cfg.gemini_model,
    )

    # Calcular días hábiles si hay fechas
    dias = glosa.dias_restantes if glosa.dias_restantes is not None else 0
    # dias_restantes es lo que queda; para el validador queremos días transcurridos
    # cuando no es extemporánea. Si es 0 o negativo asumimos vencida.
    dias_transcurridos = max(0, 20 - dias) if dias > 0 else 25

    resultado = await service.validar_pre_radicacion(
        dictamen_html=glosa.dictamen,
        eps=glosa.eps or "",
        codigo_glosa=glosa.codigo_glosa or "",
        valor_objetado=float(glosa.valor_objetado or 0),
        numero_factura=glosa.factura or "",
        dias_habiles=dias_transcurridos,
    )

    AuditRepository(db).registrar(
        usuario_email=current_user.email,
        usuario_rol=current_user.rol,
        accion="VALIDAR_PRE_RADICACION",
        tabla="historial",
        registro_id=glosa_id,
        detalle=f"score={resultado['score_calidad']} errores={resultado['errores']} warnings={resultado['warnings']}",
    )
    return resultado


@router.get("/alertas")
def alertas(
    dias: int = 5,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    repo = GlosaRepository(db)
    alertas = repo.alertas_proximas(dias_limite=dias)
    return [
        {
            "id": a.id,
            "eps": a.eps,
            "paciente": a.paciente,
            "codigo_glosa": a.codigo_glosa,
            "valor_objetado": a.valor_objetado,
            "dias_restantes": a.dias_restantes,
            "estado": a.estado,
        }
        for a in alertas
    ]


# ─── Sprint #6 — Vencen en próximas 24h (alerta máxima urgencia) ────────
@router.get("/vencen-24h")
def glosas_vencen_24h(
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Glosas que vencen HOY o MAÑANA y no están cerradas.

    Pensado para banner alerta + notificación push diaria. Un cron en
    admin invoca este endpoint cada 8h para alimentar las notifs.
    """
    terminales = [
        "LEVANTADA",
        "CONCILIADA",
        "ACEPTADA",
        "RATIFICADA",
        "ARCHIVADA",
        "DUPLICADA_OCULTA",
    ]
    base_q = db.query(GlosaRecord).filter(
        GlosaRecord.dias_restantes <= 1,
        GlosaRecord.estado.notin_(terminales),
    )
    # Auditor solo ve las suyas
    if current_user.rol == "AUDITOR":
        base_q = base_q.filter(
            (GlosaRecord.auditor_email == current_user.email)
            | (GlosaRecord.gestor_nombre == current_user.email)
        )

    rows = (
        base_q.order_by(
            GlosaRecord.dias_restantes.asc(),
            GlosaRecord.valor_objetado.desc(),
        )
        .limit(200)
        .all()
    )

    valor_total_riesgo = sum(float(g.valor_objetado or 0) for g in rows)

    return {
        "total": len(rows),
        "valor_total_riesgo": valor_total_riesgo,
        "glosas": [
            {
                "id": g.id,
                "factura": g.factura,
                "eps": g.eps,
                "codigo_glosa": g.codigo_glosa,
                "valor_objetado": float(g.valor_objetado or 0),
                "dias_restantes": int(g.dias_restantes or 0),
                "estado": g.estado,
                "fecha_vencimiento": (
                    g.fecha_vencimiento.isoformat() if g.fecha_vencimiento else None
                ),
                "gestor_nombre": g.gestor_nombre,
            }
            for g in rows
        ],
    }


@router.get("/metrics")
def metrics(
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    repo = GlosaRepository(db)
    return repo.metrics()


@router.get("/analitica-predictiva")
def analitica_predictiva(
    ventana_dias: int = Query(180, ge=7, le=730),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Analítica agregada: top EPS, tasa de éxito por código/tipo,
    distribución semanal y recomendaciones automáticas."""
    repo = GlosaRepository(db)
    return repo.analitica_predictiva(ventana_dias=ventana_dias)


# Rutas estáticas (sin parámetros) ANTES que rutas dinámicas /{glosa_id}
@router.get("/semaforo")
def semaforo(
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Retorna el conteo de glosas activas agrupadas por color de semáforo
    (VERDE / AMARILLO / ROJO / NEGRO). Útil para el dashboard."""
    repo = GlosaRepository(db)
    return repo.semaforo_counts()


# ─── Sprint #7 — Dashboard plata recuperada ─────────────────────────────
@router.get("/dashboard-plata-recuperada")
def dashboard_plata_recuperada(
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Resumen de plata facturada vs aceptada vs recuperada.

    Agrupa por EPS, código y mes. Útil para mostrar a gerencia
    cuánto plata el motor le ha hecho recuperar al hospital.
    Filtros opcionales: ?desde=YYYY-MM-DD&hasta=YYYY-MM-DD.
    """

    base_q = db.query(GlosaRecord).filter(GlosaRecord.estado.notin_(["DUPLICADA_OCULTA"]))

    if desde:
        try:
            d = datetime.fromisoformat(desde)
            base_q = base_q.filter(GlosaRecord.creado_en >= d)
        except ValueError:
            raise HTTPException(400, "desde debe ser YYYY-MM-DD")
    if hasta:
        try:
            h = datetime.fromisoformat(hasta)
            base_q = base_q.filter(GlosaRecord.creado_en <= h)
        except ValueError:
            raise HTTPException(400, "hasta debe ser YYYY-MM-DD")

    glosas = base_q.all()

    # Agregados globales
    total_objetado = sum(float(g.valor_objetado or 0) for g in glosas)
    total_aceptado = sum(float(g.valor_aceptado or 0) for g in glosas)
    total_recuperado = sum(float(g.valor_recuperado or 0) for g in glosas)
    n_total = len(glosas)
    n_levantadas = sum(1 for g in glosas if (g.estado or "").upper() == "LEVANTADA")
    n_ratificadas = sum(1 for g in glosas if (g.estado or "").upper() == "RATIFICADA")
    n_pendientes = sum(
        1
        for g in glosas
        if (g.estado or "").upper()
        not in {"LEVANTADA", "RATIFICADA", "ACEPTADA", "CONCILIADA", "ARCHIVADA"}
    )

    tasa_efectividad = (n_levantadas / max(1, n_levantadas + n_ratificadas)) * 100

    # Por EPS
    por_eps: dict[str, dict] = {}
    for g in glosas:
        eps_k = (g.eps or "—").upper()
        d = por_eps.setdefault(
            eps_k,
            {
                "eps": eps_k,
                "n_glosas": 0,
                "valor_objetado": 0.0,
                "valor_aceptado": 0.0,
                "valor_recuperado": 0.0,
                "n_levantadas": 0,
            },
        )
        d["n_glosas"] += 1
        d["valor_objetado"] += float(g.valor_objetado or 0)
        d["valor_aceptado"] += float(g.valor_aceptado or 0)
        d["valor_recuperado"] += float(g.valor_recuperado or 0)
        if (g.estado or "").upper() == "LEVANTADA":
            d["n_levantadas"] += 1
    por_eps_list = sorted(por_eps.values(), key=lambda x: -x["valor_recuperado"])[:20]

    # Por código de glosa
    por_codigo: dict[str, dict] = {}
    for g in glosas:
        cod = (g.codigo_glosa or "—").upper()
        d = por_codigo.setdefault(
            cod,
            {
                "codigo_glosa": cod,
                "n_glosas": 0,
                "valor_objetado": 0.0,
                "valor_recuperado": 0.0,
            },
        )
        d["n_glosas"] += 1
        d["valor_objetado"] += float(g.valor_objetado or 0)
        d["valor_recuperado"] += float(g.valor_recuperado or 0)
    por_codigo_list = sorted(por_codigo.values(), key=lambda x: -x["valor_recuperado"])[:15]

    # Por mes (últimos 12 meses)
    por_mes: dict[str, dict] = {}
    for g in glosas:
        if not g.creado_en:
            continue
        ym = g.creado_en.strftime("%Y-%m")
        d = por_mes.setdefault(
            ym,
            {
                "mes": ym,
                "n_glosas": 0,
                "valor_objetado": 0.0,
                "valor_recuperado": 0.0,
            },
        )
        d["n_glosas"] += 1
        d["valor_objetado"] += float(g.valor_objetado or 0)
        d["valor_recuperado"] += float(g.valor_recuperado or 0)
    por_mes_list = sorted(por_mes.values(), key=lambda x: x["mes"])[-12:]

    return {
        "totales": {
            "n_glosas": n_total,
            "n_levantadas": n_levantadas,
            "n_ratificadas": n_ratificadas,
            "n_pendientes": n_pendientes,
            "valor_objetado": total_objetado,
            "valor_aceptado": total_aceptado,
            "valor_recuperado": total_recuperado,
            "tasa_efectividad_pct": round(tasa_efectividad, 1),
        },
        "por_eps": por_eps_list,
        "por_codigo": por_codigo_list,
        "por_mes": por_mes_list,
    }


@router.get("/paciente-resumen")
def paciente_resumen(
    paciente: str = Query(..., min_length=2),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R130 P1: resumen de glosas asociadas a un paciente.

    Útil para investigar el histórico de glosas de un paciente
    específico (mismo paciente puede tener varias hospitalizaciones
    objetadas).

    Query param `paciente` se usa con ILIKE para tolerancia a
    variaciones en mayúsculas/acentos.

    Devuelve:
      - total_glosas
      - facturas_distintas
      - eps_distintas
      - valor_objetado_total / valor_recuperado_total
      - estados (mapa)
      - glosas: lista resumida (id, factura, codigo_glosa, valor)
    """
    glosas = (
        db.query(GlosaRecord)
        .filter(GlosaRecord.paciente.ilike(f"%{paciente}%"))
        .order_by(GlosaRecord.creado_en.desc())
        .all()
    )

    if not glosas:
        return {
            "paciente_buscado": paciente,
            "total_glosas": 0,
            "facturas_distintas": 0,
            "eps_distintas": 0,
            "valor_objetado_total": 0,
            "valor_recuperado_total": 0,
            "estados": {},
            "glosas": [],
        }

    facturas: set[str] = set()
    epss: set[str] = set()
    estados: dict[str, int] = {}
    valor_obj = 0.0
    valor_rec = 0.0

    for g in glosas:
        if g.factura and g.factura != "N/A":
            facturas.add(g.factura)
        if g.eps:
            epss.add(g.eps)
        e = g.estado or "?"
        estados[e] = estados.get(e, 0) + 1
        valor_obj += float(g.valor_objetado or 0)
        valor_rec += float(g.valor_recuperado or 0)

    return {
        "paciente_buscado": paciente,
        "total_glosas": len(glosas),
        "facturas_distintas": len(facturas),
        "eps_distintas": len(epss),
        "valor_objetado_total": int(valor_obj),
        "valor_recuperado_total": int(valor_rec),
        "estados": estados,
        "glosas": [
            {
                "id": g.id,
                "creado_en": (g.creado_en.isoformat() if g.creado_en else None),
                "factura": g.factura,
                "eps": g.eps,
                "codigo_glosa": g.codigo_glosa,
                "valor_objetado": float(g.valor_objetado or 0),
                "estado": g.estado,
            }
            for g in glosas[:50]  # cap a 50 para no inflar
        ],
    }


@router.get("/sin-actividad")
def glosas_sin_actividad(
    dias: int = Query(15, ge=1, le=180),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R99 P1: glosas abiertas sin actualizaciones recientes.

    Detecta glosas no-cerradas que llevan más de N días sin
    movimiento en audit_log. Útil para que el coordinador
    identifique:
      - Glosas "olvidadas" en el flujo
      - Casos que necesitan seguimiento o reasignación
      - Trabajo estancado por carga de un gestor

    Por glosa devuelve:
      - id, eps, factura, estado, dias_restantes
      - ultimo_movimiento_en (max(creado_en, max(audit.timestamp)))
      - dias_sin_movimiento
      - gestor_nombre

    Ordenado DESC por dias_sin_movimiento.
    """
    from datetime import timedelta, timezone

    from app.models.db import AuditLogRecord

    ESTADOS_CERRADOS = {"ACEPTADA", "LEVANTADA", "ARCHIVADA", "CONCILIADA"}

    abiertas = db.query(GlosaRecord).filter(~GlosaRecord.estado.in_(ESTADOS_CERRADOS)).all()

    # Última actividad por glosa según audit_log
    ultimas: dict[int, "object"] = {}
    rows = (
        db.query(AuditLogRecord.registro_id, AuditLogRecord.timestamp)
        .filter(AuditLogRecord.tabla == "glosas")
        .filter(AuditLogRecord.registro_id.isnot(None))
        .all()
    )
    for rid, ts in rows:
        if not ts:
            continue
        prev = ultimas.get(rid)
        if prev is None or ts > prev:
            ultimas[rid] = ts

    ahora = ahora_utc()
    corte = ahora - timedelta(days=int(dias))

    items = []
    for g in abiertas:
        creado = g.creado_en
        if creado and creado.tzinfo is None:
            creado = creado.replace(tzinfo=timezone.utc)

        ult = ultimas.get(g.id)
        if ult is not None and ult.tzinfo is None:
            ult = ult.replace(tzinfo=timezone.utc)

        # Última actividad = max(creación, último audit)
        ultimo_mov = creado
        if ult and (ultimo_mov is None or ult > ultimo_mov):
            ultimo_mov = ult

        if ultimo_mov is None or ultimo_mov >= corte:
            continue

        dias_sin = (ahora - ultimo_mov).days
        items.append(
            {
                "id": g.id,
                "eps": g.eps,
                "factura": g.factura,
                "estado": g.estado,
                "dias_restantes": g.dias_restantes,
                "gestor_nombre": g.gestor_nombre,
                "ultimo_movimiento_en": ultimo_mov.isoformat(),
                "dias_sin_movimiento": dias_sin,
            }
        )

    items.sort(key=lambda x: x["dias_sin_movimiento"], reverse=True)

    return {
        "umbral_dias": int(dias),
        "total_sin_actividad": len(items),
        "limit": int(limit),
        "items": items[:limit],
    }


@router.get("/incompletas")
def glosas_incompletas(
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R96 P2: lista glosas con datos críticos faltantes.

    Complementa /glosas/{id}/checklist (vista por glosa) con una
    vista agregada: ¿qué glosas tienen huecos que necesitan
    completarse?

    Filtra glosas no-cerradas a las que les falta AL MENOS UNO de:
      - texto_glosa_original
      - dictamen (vacío o muy corto)
      - factura (vacía o "N/A")
      - valor_objetado (0 o NULL)

    Útil para batch cleanup masivo del coordinador.

    Devuelve cada glosa con un campo "campos_faltantes" indicando
    cuáles específicamente, ordenadas DESC por número de huecos.
    """
    ESTADOS_CERRADOS = {"ACEPTADA", "LEVANTADA", "ARCHIVADA", "CONCILIADA"}

    # Pre-filtramos por estado en SQL; los criterios de huecos los
    # evaluamos en Python para tener semántica consistente (ej.
    # "dictamen corto" requiere len() y SQL LENGTH no es portable).
    candidatas = db.query(GlosaRecord).filter(~GlosaRecord.estado.in_(ESTADOS_CERRADOS)).all()

    items = []
    for g in candidatas:
        faltantes = []
        if not g.texto_glosa_original:
            faltantes.append("texto_glosa_original")
        if not g.dictamen or len(g.dictamen) <= 50:
            faltantes.append("dictamen")
        if not g.factura or g.factura == "N/A":
            faltantes.append("factura")
        if not g.valor_objetado or g.valor_objetado == 0:
            faltantes.append("valor_objetado")

        if not faltantes:
            continue

        items.append(
            {
                "id": g.id,
                "creado_en": (g.creado_en.isoformat() if g.creado_en else None),
                "eps": g.eps,
                "factura": g.factura,
                "estado": g.estado,
                "campos_faltantes": faltantes,
                "total_huecos": len(faltantes),
            }
        )

    items.sort(key=lambda x: x["total_huecos"], reverse=True)

    return {
        "total_incompletas": len(items),
        "limit": int(limit),
        "items": items[:limit],
    }


@router.get("/facetas")
def facetas_glosas(
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R88 P1: facetas únicas de las glosas para construir filtros UI.

    Devuelve los valores DISTINCT no-nulos de eps, etapa, estado,
    codigo_glosa y gestor_nombre. Útil para que el frontend renderice
    <select> con valores reales en lugar de inputs libres (mejor UX,
    menos typos al filtrar).

    Hace un solo round-trip por columna; cada columna tiene índice
    en BD así que es O(distinct) eficiente.
    """

    def _distinct(col):
        rows = db.query(col).filter(col.isnot(None)).distinct().order_by(col.asc()).all()
        return [r[0] for r in rows if r[0]]

    return {
        "eps": _distinct(GlosaRecord.eps),
        "etapas": _distinct(GlosaRecord.etapa),
        "estados": _distinct(GlosaRecord.estado),
        "codigos_glosa": _distinct(GlosaRecord.codigo_glosa),
        "gestores": _distinct(GlosaRecord.gestor_nombre),
    }


@router.get("/por-factura")
def glosas_por_factura(
    numero_factura: str = Query(..., min_length=1, max_length=60),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Retorna los conceptos asociados a un número de factura.

    Dos fuentes posibles:
      • Si la factura fue importada desde el Excel de recepción con sus
        hojas I/R, cada fila del listado es un ConceptoGlosaRecord
        (motivo, CUPS, servicio, valor parcial, observación de la EPS).
      • Si no hay conceptos (flujo legacy de masiva), cada fila es un
        GlosaRecord individual por concepto.

    La UI del Analizar usa esto para precargar automáticamente cada
    concepto al auditor sin pegar texto.
    """
    from app.models.db import GlosaRecord as _GR

    factura_limpio = numero_factura.strip()
    if not factura_limpio:
        return {"numero_factura": "", "glosas": []}

    def _nombre_corto_entidad(plan_eps: str, tercero: str = "") -> str:
        """Wrapper sobre `pagador_normalizer.nombre_corto` con preferencia
        por `tercero_nombre` (FacturaCartera.Tercero.NombreCompletoNA).
        El normalizador también repara truncamientos como BUCARAMANG → BUCARAMANGA.
        """
        from app.services import pagador_normalizer

        t = pagador_normalizer.nombre_corto(tercero)
        if t:
            return t
        return pagador_normalizer.nombre_corto(plan_eps)

    glosas_padre = (
        db.query(_GR).filter(_GR.factura == factura_limpio).order_by(_GR.id.asc()).limit(50).all()
    )
    items: list[dict] = []
    glosa_ids = [g.id for g in glosas_padre]

    # Conceptos del nuevo modelo (importación recepción hojas I/R)
    conceptos = []
    if glosa_ids:
        conceptos = (
            db.query(ConceptoGlosaRecord)
            .filter(ConceptoGlosaRecord.glosa_id.in_(glosa_ids))
            .order_by(ConceptoGlosaRecord.codigo_glosa.asc(), ConceptoGlosaRecord.id.asc())
            .all()
        )
    # Mapa glosa_id -> GlosaRecord para enriquecer cada concepto con eps/fechas
    mapa_padre = {g.id: g for g in glosas_padre}

    if conceptos:
        # Caso normal: Excel de recepción completo con hojas I/R
        for c in conceptos:
            padre = mapa_padre.get(c.glosa_id)
            items.append(
                {
                    "id": c.glosa_id,  # glosa padre (para analizar llamando al endpoint)
                    "concepto_id": c.id,  # identificador del concepto específico
                    "oid_dgh": c.oid_dgh or "",
                    "codigo_glosa": c.codigo_glosa or "",
                    "nombre_glosa": c.nombre_glosa or "",
                    "cups": c.cups_codigo or "",
                    "servicio": c.cups_descripcion or "",
                    "centro_costo": c.centro_costo or "",
                    "observacion_eps": c.observacion_eps or "",
                    "valor_objetado": c.valor_objetado or 0,
                    "valor_aceptado": 0,
                    "estado": (padre.estado if padre else "") or "",
                    "eps": (padre.eps if padre else "") or "",
                    # Nombre comercial corto (FacturaCartera.Tercero.NombreCompletoNA),
                    # ej: "DISPENSARIO MEDICO BUCARAMANGA". La UI lo prefiere sobre
                    # el plan EPS cuando existe.
                    "tercero_nombre": _nombre_corto_entidad(
                        padre.eps if padre else "",
                        getattr(padre, "tercero_nombre", None) if padre else "",
                    ),
                    "concepto_glosa": c.nombre_glosa or "",
                    "texto_glosa_original": (c.observacion_eps or "")[:400],
                    "fecha_radicacion_factura": padre.fecha_radicacion_factura.isoformat()
                    if padre and padre.fecha_radicacion_factura
                    else None,
                    "fecha_recepcion": padre.fecha_recepcion.isoformat()
                    if padre and padre.fecha_recepcion
                    else None,
                    "dictamen_generado": bool(c.dictamen_html),
                }
            )
    else:
        # Fallback legacy: 1 GlosaRecord por concepto (flujo importación masiva)
        for g in glosas_padre:
            items.append(
                {
                    "id": g.id,
                    "concepto_id": None,
                    "codigo_glosa": g.codigo_glosa or "",
                    "nombre_glosa": g.concepto_glosa or "",
                    "concepto_glosa": g.concepto_glosa or "",
                    "cups": g.cups_servicio or "",
                    "servicio": g.servicio_descripcion or "",
                    "centro_costo": "",
                    "observacion_eps": "",
                    "valor_objetado": g.valor_objetado or 0,
                    "valor_aceptado": g.valor_aceptado or 0,
                    "estado": g.estado or "",
                    "eps": g.eps or "",
                    "texto_glosa_original": (g.texto_glosa_original or "")[:400],
                    "fecha_radicacion_factura": g.fecha_radicacion_factura.isoformat()
                    if g.fecha_radicacion_factura
                    else None,
                    "fecha_recepcion": g.fecha_recepcion.isoformat() if g.fecha_recepcion else None,
                    "dictamen_generado": bool(g.dictamen),
                }
            )

    eps_unicas = list({g.eps for g in glosas_padre if g.eps})
    # Nombre comercial corto de la entidad (Tercero.NombreCompletoNA). Si todas
    # las glosas padre apuntan al mismo tercero, exponemos ese nombre en la
    # cabecera de la respuesta. Si no, None (la UI cae al plan EPS).
    # Usar el helper para fallback: si tercero_nombre esta vacio, extrae el
    # nombre corto del plan EPS. Así las glosas importadas antes del campo
    # tercero_nombre también muestran el nombre comercial limpio.
    terceros_unicos = list(
        {
            _nombre_corto_entidad(g.eps, getattr(g, "tercero_nombre", None))
            for g in glosas_padre
            if g.eps
        }
    )
    terceros_unicos = [t for t in terceros_unicos if t]
    total_objetado = sum(i["valor_objetado"] or 0 for i in items)
    return {
        "numero_factura": factura_limpio,
        "total_conceptos": len(items),
        "total_objetado": total_objetado,
        "eps": eps_unicas[0] if len(eps_unicas) == 1 else None,
        "eps_multiples": eps_unicas if len(eps_unicas) > 1 else None,
        "tercero_nombre": terceros_unicos[0] if len(terceros_unicos) == 1 else None,
        "glosa_id": glosa_ids[0] if glosa_ids else None,
        "glosas": items,
    }


@router.get("/facturas-pendientes")
def facturas_pendientes_agrupadas(
    limite: int = Query(30, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Lista facturas con glosas pendientes agrupadas.

    Útil para el flujo "responder por factura" — muestra facturas con
    N conceptos pendientes cada una, asignadas al usuario actual o su
    equipo.
    """
    from app.models.db import GlosaRecord as _GR
    from sqlalchemy import or_ as _or

    repo = GlosaRepository(db)
    # Filtrar por gestor/equipo (igual que mis-asignaciones)
    equipo = getattr(current_user, "equipo", None)
    emails = repo.emails_del_mismo_equipo(equipo) if equipo else [current_user.email]
    if not emails:
        emails = [current_user.email]
    prefijos_nombre = [e.split("@")[0] for e in emails]

    # SUPER_ADMIN/COORDINADOR ve todas
    if current_user.rol in ("SUPER_ADMIN", "COORDINADOR"):
        base_q = db.query(_GR).filter(_GR.factura.isnot(None))
    else:
        condiciones = [_GR.auditor_email.in_(emails)]
        if current_user.nombre:
            condiciones.append(_GR.gestor_nombre.ilike(f"%{current_user.nombre.strip()}%"))
        for p in prefijos_nombre:
            condiciones.append(_GR.gestor_nombre.ilike(f"%{p}%"))
        base_q = db.query(_GR).filter(_or(*condiciones), _GR.factura.isnot(None))
    # Solo estados activos (no LEVANTADA ni CONCILIADA)
    base_q = base_q.filter(_GR.estado.notin_(["LEVANTADA", "CONCILIADA"]))
    # Agrupar por factura
    agrupados: dict[str, list] = {}
    for g in base_q.limit(500).all():
        fact = g.factura
        if not fact:
            continue
        agrupados.setdefault(fact, []).append(g)

    resultado = []
    for fact, glosas in agrupados.items():
        eps_set = {g.eps for g in glosas if g.eps}
        total = sum(g.valor_objetado or 0 for g in glosas)
        codigos = [g.codigo_glosa for g in glosas if g.codigo_glosa]
        fecha_mas_reciente = max(
            (g.fecha_recepcion for g in glosas if g.fecha_recepcion), default=None
        )
        resultado.append(
            {
                "numero_factura": fact,
                "eps": list(eps_set)[0] if len(eps_set) == 1 else None,
                "cantidad_conceptos": len(glosas),
                "valor_total_objetado": total,
                "codigos": codigos[:10],
                "fecha_recepcion_mas_reciente": fecha_mas_reciente.isoformat()
                if fecha_mas_reciente
                else None,
            }
        )
    # Orden: mas conceptos primero, luego mayor valor
    resultado.sort(key=lambda x: (-x["cantidad_conceptos"], -x["valor_total_objetado"]))
    return {"total_facturas": len(resultado), "facturas": resultado[:limite]}


@router.get("/mis-asignaciones")
def mis_asignaciones(
    todas: bool = False,
    vista: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Lista las glosas asignadas al usuario actual.

    Los SUPER_ADMIN y COORDINADOR pueden pasar `?todas=true` para ver todas.

    Sprint #5 — Vistas guardadas (`?vista=...`):
      • `urgentes`        : score_urgencia >= 85 (vence pronto + alta cuantía)
      • `vencen_hoy`      : dias_restantes <= 1 y no terminales
      • `aprobadas`       : workflow_state == APROBADA (faltan radicar)
      • `requieren_soportes`: estado == REQUIERE_SOPORTES
      • `ta_sin_contrato` : codigo_glosa empieza por TA y EPS sin contrato
      • `alta_cuantia`    : valor_objetado >= 5_000_000
      • `respondidas`     : workflow_state RESPONDIDA / LEVANTADA / etc.
    """
    repo = GlosaRepository(db)
    if todas and current_user.rol in ("SUPER_ADMIN", "COORDINADOR"):
        from app.models.db import GlosaRecord as _GR

        glosas = (
            db.query(_GR)
            .filter(_GR.estado.notin_(["LEVANTADA", "CONCILIADA"]))
            .order_by(_GR.dias_restantes.asc())
            .limit(500)
            .all()
        )
    else:
        # Si el usuario pertenece a un equipo (ej. EQUIPO_ASEGURADORAS),
        # agrupar asignaciones de todos los miembros del equipo.
        equipo = getattr(current_user, "equipo", None)
        emails_equipo = repo.emails_del_mismo_equipo(equipo) if equipo else None
        glosas = repo.listar_por_gestor(
            current_user.email,
            current_user.nombre,
            emails_equipo=emails_equipo,
        )

    # R-UI 27-abr-2026: deduplicar por (factura, código, cups, etapa)
    # quedándonos con la versión MÁS RECIENTE. Si una glosa antigua
    # ya está RESPONDIDA, otra fila duplicada antigua RADICADA no
    # debe seguir apareciendo en pendientes — eso pasaba con
    # glosas creadas antes del fix anti-dup (commit d8b1d53).
    visto = {}
    for g in glosas:
        clave = (
            (g.factura or "").strip().upper(),
            (g.codigo_glosa or "").strip().upper(),
            (getattr(g, "cups_servicio", None) or "").strip().upper(),
            (g.etapa or "").strip().upper(),
        )
        prev = visto.get(clave)
        if prev is None:
            visto[clave] = g
            continue
        # Si una versión está cerrada (RESPONDIDA/CONCILIADA/LEVANTADA),
        # esa gana sobre la abierta. Caso contrario gana la más reciente.
        prev_estado = (prev.estado or "").upper()
        prev_wf = (getattr(prev, "workflow_state", None) or "").upper()
        cur_estado = (g.estado or "").upper()
        cur_wf = (getattr(g, "workflow_state", None) or "").upper()
        terminales = {"RESPONDIDA", "CONCILIADA", "LEVANTADA"}
        prev_es_terminal = (prev_estado in terminales) or (prev_wf in terminales)
        cur_es_terminal = (cur_estado in terminales) or (cur_wf in terminales)
        if cur_es_terminal and not prev_es_terminal:
            visto[clave] = g
        elif (cur_es_terminal == prev_es_terminal) and (
            g.creado_en and prev.creado_en and g.creado_en > prev.creado_en
        ):
            visto[clave] = g
    glosas = list(visto.values())

    # R-UI 27-abr-2026: scoring de urgencia para priorización automática.
    # El gestor abre Mis glosas y ve PRIMERO lo que necesita atención
    # inmediata, sin tener que filtrar/ordenar manualmente.
    def _score_urgencia(g) -> tuple[int, str]:
        """Devuelve (score 0-100, motivo legible). Más alto = más urgente."""
        estado = (g.estado or "").upper()
        wf = (getattr(g, "workflow_state", None) or "").upper()
        valor = float(g.valor_objetado or 0)
        dias = int(g.dias_restantes or 999)
        terminales = {"RESPONDIDA", "CONCILIADA", "LEVANTADA"}

        # Aprobada pero no radicada → prioridad alta
        if wf == "APROBADA":
            return (88, "Aprobada — falta radicar")
        # Vencidas
        if dias <= 0 and estado not in terminales:
            return (100, "VENCIDA")
        # Estados cerrados — sin urgencia
        if estado in terminales or wf in terminales:
            return (10, "")
        # Requiere soportes — bloquea hasta que el gestor actúe
        if estado == "REQUIERE_SOPORTES":
            return (75, "Requiere soportes")
        # Vence pronto + alta cuantía
        if dias <= 3 and valor >= 1_000_000:
            return (95, f"Vence en {dias}d · ${int(valor):,}".replace(",", "."))
        # Vence muy pronto
        if dias <= 3:
            return (85, f"Vence en {dias}d")
        # Alta cuantía
        if valor >= 5_000_000:
            return (70, f"Alta cuantía · ${int(valor):,}".replace(",", "."))
        # Por vencer (5d)
        if dias <= 5:
            return (55, f"Vence en {dias}d")
        # Resto: por antigüedad
        return (max(20, 50 - min(dias, 50)), "")

    # Anotar score y ordenar
    glosas_con_score = []
    for g in glosas:
        score, motivo = _score_urgencia(g)
        glosas_con_score.append((score, motivo, g))
    glosas_con_score.sort(key=lambda x: x[0], reverse=True)
    glosas = [t[2] for t in glosas_con_score]
    score_por_id = {t[2].id: (t[0], t[1]) for t in glosas_con_score}

    # Sprint #5 — Aplicar vista guardada después de scoring/ordenamiento
    if vista:
        vista_norm = vista.lower().strip()
        contratos_db = ContratoRepository(db).como_dict() or {}
        terminales = {
            "RESPONDIDA",
            "CONCILIADA",
            "LEVANTADA",
            "ACEPTADA",
            "RATIFICADA",
            "ARCHIVADA",
        }

        # Vista "dictamen_obsoleto" — recorre el módulo dedicado. Cargamos
        # los IDs antes del filtro para no llamar dictamen_stale.es_stale
        # por cada glosa (evita N+1 sobre tarifas_contratadas).
        ids_obsoletos: set[int] = set()
        if vista_norm == "dictamen_obsoleto":
            from app.services.dictamen_stale import es_stale as _es_stale

            for g in glosas:
                try:
                    if _es_stale(g, db):
                        ids_obsoletos.add(g.id)
                except Exception:
                    continue

        def _matches(g) -> bool:
            score, _ = score_por_id.get(g.id, (0, ""))
            estado = (g.estado or "").upper()
            wf = (getattr(g, "workflow_state", None) or "").upper()
            valor = float(g.valor_objetado or 0)
            dias = int(g.dias_restantes or 999)
            cod = (g.codigo_glosa or "").upper()
            eps_norm = (g.eps or "").upper().strip()

            if vista_norm == "urgentes":
                return score >= 85 and estado not in terminales
            if vista_norm == "vencen_hoy":
                return dias <= 1 and estado not in terminales and wf not in terminales
            if vista_norm == "aprobadas":
                return wf == "APROBADA"
            if vista_norm == "requieren_soportes":
                return estado == "REQUIERE_SOPORTES"
            if vista_norm == "ta_sin_contrato":
                return cod.startswith("TA") and eps_norm not in contratos_db
            if vista_norm == "alta_cuantia":
                return valor >= 5_000_000 and estado not in terminales
            if vista_norm == "respondidas":
                return estado in terminales or wf in terminales
            if vista_norm == "dictamen_obsoleto":
                return g.id in ids_obsoletos
            return True

        glosas = [g for g in glosas if _matches(g)]

    from app.services.resolver_entidad import resolver_entidad_mostrar

    # Soportes auto-detectados: cuántos PDFs hay en el servidor por
    # factura + tipos detectados. Un lookup por factura única (no por
    # glosa) para no penalizar listas largas. Silencioso si el
    # indexador no tiene raíz accesible (servidor desconectado).
    _soportes_por_factura: dict[str, int] = {}
    _soportes_tipos_por_factura: dict[str, str] = {}
    try:
        from app.services.soportes_autodiscovery_service import get_indexer

        _indexer = get_indexer()
        if _indexer.stats().get("raiz_existe"):
            _facturas_unicas = {(g.factura or "").strip() for g in glosas if g.factura}
            for _fact in _facturas_unicas:
                if not _fact:
                    continue
                _hits = _indexer.lookup(_fact)
                if _hits:
                    _soportes_por_factura[_fact] = len(_hits)
                    _tipos_set = sorted({h.get("tipo_codigo") or "OTRO" for h in _hits})
                    _soportes_tipos_por_factura[_fact] = ", ".join(_tipos_set[:6])
    except Exception as _e:
        logger.debug(f"[mis-asignaciones] indexer lookup falló: {_e}")
        _soportes_por_factura = {}
        _soportes_tipos_por_factura = {}

    items = []
    for g in glosas:
        score, motivo = score_por_id.get(g.id, (0, ""))
        items.append(
            {
                "id": g.id,
                "eps": resolver_entidad_mostrar(
                    eps=g.eps,
                    tercero_nombre=getattr(g, "tercero_nombre", None),
                    eps_codigo=getattr(g, "eps_codigo", None),
                ),
                "eps_raw": g.eps,
                "factura": g.factura,
                "numero_radicado": g.numero_radicado,
                "consecutivo_dgh": g.consecutivo_dgh,
                "gestor_nombre": g.gestor_nombre,
                "valor_objetado": g.valor_objetado,
                "estado": g.estado,
                "prioridad": g.prioridad,
                "dias_restantes": g.dias_restantes,
                "dias_radicacion_dgh": getattr(g, "dias_radicacion_dgh", None),
                "fecha_vencimiento": g.fecha_vencimiento.isoformat()
                if g.fecha_vencimiento
                else None,
                "fecha_entrega": g.fecha_entrega.isoformat() if g.fecha_entrega else None,
                "fecha_radicacion_factura": g.fecha_radicacion_factura.isoformat()
                if g.fecha_radicacion_factura
                else None,
                "fecha_documento_dgh": g.fecha_documento_dgh.isoformat()
                if g.fecha_documento_dgh
                else None,
                "fecha_recepcion": g.fecha_recepcion.isoformat() if g.fecha_recepcion else None,
                "radicado_info": g.radicado_info,
                "referencia": g.referencia,
                # Flag para que el front muestre el boton "Marcar Respondida" solo
                # si ya hay un dictamen generado (sino no hay nada que cerrar).
                "dictamen_generado": bool(g.dictamen),
                "observacion_tecnico": g.observacion_tecnico,
                "tipo_glosa_excel": g.tipo_glosa_excel,
                "profesional_medico": g.profesional_medico,
                "dictamen": g.dictamen,
                "workflow_state": g.workflow_state or "BORRADOR",
                "nota_workflow": g.nota_workflow,
                "valor_aceptado": float(g.valor_aceptado or 0.0),
                "numero_nota_credito": getattr(g, "numero_nota_credito", None),
                # Priorización automática (R-UI 27-abr-2026)
                "score_urgencia": int(score),
                "motivo_prioridad": motivo,
                # Campos para Sprint #3 (similares en bloque) y filtros UI
                "codigo_glosa": g.codigo_glosa,
                "cups_servicio": getattr(g, "cups_servicio", None),
                # Soportes auto-detectados en el servidor de archivos
                "soportes_disponibles": _soportes_por_factura.get((g.factura or "").strip(), 0),
                "soportes_tipos": _soportes_tipos_por_factura.get((g.factura or "").strip(), ""),
            }
        )
    return items


@router.patch("/{glosa_id}/estado")
def actualizar_estado(
    glosa_id: int,
    nuevo_estado: str,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    repo = GlosaRepository(db)
    glosa = repo.actualizar_estado(glosa_id, nuevo_estado, responsable="sistema")
    if not glosa:
        raise HTTPException(status_code=404, detail="Glosa no encontrada")
    logger.info(f"Estado actualizado | glosa_id={glosa_id} | nuevo_estado={nuevo_estado}")
    return {"message": "Estado actualizado", "glosa": glosa}


@router.get("/exportar-resumen-eps.csv")
def exportar_resumen_eps_csv(
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R236 P1: export CSV con resumen agregado por EPS.

    Una fila por EPS con columnas:
      eps, total_glosas, abiertas, cerradas, levantadas,
      valor_objetado, valor_recuperado, tasa_levantamiento_pct

    Útil para reporting en Excel/Tableau.

    StreamingResponse para no cargar todo en memoria.
    """
    import csv
    import io
    from datetime import datetime, timezone

    from fastapi.responses import StreamingResponse

    ESTADOS_CERRADOS = {"ACEPTADA", "LEVANTADA", "ARCHIVADA", "CONCILIADA"}

    glosas = db.query(GlosaRecord).all()

    por_eps: dict[str, dict] = {}
    for g in glosas:
        eps = (g.eps or "").strip()
        if not eps:
            continue
        if eps not in por_eps:
            por_eps[eps] = {
                "total": 0,
                "abiertas": 0,
                "cerradas": 0,
                "decididas": 0,
                "levantadas": 0,
                "obj": 0.0,
                "rec": 0.0,
            }
        b = por_eps[eps]
        b["total"] += 1
        b["obj"] += float(g.valor_objetado or 0)
        b["rec"] += float(g.valor_recuperado or 0)
        estado = (g.estado or "").upper()
        if estado in ESTADOS_CERRADOS:
            b["cerradas"] += 1
            if estado in {"LEVANTADA", "ACEPTADA", "RATIFICADA"}:
                b["decididas"] += 1
                if estado == "LEVANTADA":
                    b["levantadas"] += 1
        else:
            b["abiertas"] += 1

    def _generar():
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(
            [
                "eps",
                "total_glosas",
                "abiertas",
                "cerradas",
                "levantadas",
                "valor_objetado",
                "valor_recuperado",
                "tasa_levantamiento_pct",
            ]
        )
        yield buf.getvalue()
        buf.seek(0)
        buf.truncate(0)

        for eps, b in sorted(por_eps.items()):
            tasa = round(100 * b["levantadas"] / b["decididas"], 2) if b["decididas"] else 0.0
            w.writerow(
                [
                    eps,
                    b["total"],
                    b["abiertas"],
                    b["cerradas"],
                    b["levantadas"],
                    int(b["obj"]),
                    int(b["rec"]),
                    tasa,
                ]
            )
            yield buf.getvalue()
            buf.seek(0)
            buf.truncate(0)

    fname = f"resumen-eps-{datetime.now(timezone.utc).strftime('%Y%m%d')}.csv"
    return StreamingResponse(
        _generar(),
        media_type="text/csv",
        headers={
            "Content-Disposition": f'attachment; filename="{fname}"',
        },
    )


@router.get("/exportar-paquete-multi.zip")
def exportar_paquete_multi_zip(
    ids: str = Query(..., description="IDs CSV, ej '1,2,3'"),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R138 P2: ZIP con evidencias de múltiples glosas en una
    sola descarga.

    Complementa /glosas/{id}/exportar-evidencia.zip (1 glosa)
    con la versión multi: descarga un ZIP con subcarpetas
    glosa-{id}/ por cada ID solicitado.

    Útil para entregas masivas a legal/compliance:
      "manda las 50 glosas de SANITAS de marzo en un paquete"

    Param `ids`: lista CSV de IDs (max 100 por request).

    Cada subcarpeta: glosa.json + dictamen.txt (si existe),
    plus README.txt en raíz con índice general.

    Declarado ANTES de /{glosa_id} para evitar collisión con
    el path resolver de FastAPI.
    """
    import io
    import json
    import zipfile

    from fastapi.responses import StreamingResponse

    try:
        ids_list = [int(x.strip()) for x in ids.split(",") if x.strip()]
    except ValueError:
        raise HTTPException(400, "ids debe ser CSV de enteros")

    if not ids_list:
        raise HTTPException(400, "ids no puede estar vacío")
    if len(ids_list) > 100:
        raise HTTPException(400, "máximo 100 glosas por paquete")

    glosas = db.query(GlosaRecord).filter(GlosaRecord.id.in_(ids_list)).all()
    if not glosas:
        raise HTTPException(404, "Ninguna glosa encontrada")

    encontrados = {g.id for g in glosas}
    no_encontrados = [i for i in ids_list if i not in encontrados]

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        indice = (
            f"PAQUETE MULTI-GLOSA — {len(glosas)} glosas\n"
            f"Generado: {ahora_utc().isoformat()}\n"
            f"Por: {current_user.email}\n\n"
            f"IDs solicitados: {ids_list}\n"
            f"IDs encontrados: {sorted(encontrados)}\n"
            f"IDs no encontrados: {no_encontrados}\n"
        )
        zf.writestr("README.txt", indice)

        for g in glosas:
            subdir = f"glosa-{g.id}/"
            datos = {
                "id": g.id,
                "creado_en": (g.creado_en.isoformat() if g.creado_en else None),
                "eps": g.eps,
                "factura": g.factura,
                "codigo_glosa": g.codigo_glosa,
                "valor_objetado": float(g.valor_objetado or 0),
                "valor_recuperado": float(g.valor_recuperado or 0),
                "estado": g.estado,
                "decision_eps": g.decision_eps,
            }
            zf.writestr(
                f"{subdir}glosa.json",
                json.dumps(datos, ensure_ascii=False, indent=2),
            )
            if g.dictamen:
                zf.writestr(f"{subdir}dictamen.txt", g.dictamen)

    buf.seek(0)
    fname = f"glosas-paquete-{ahora_utc().strftime('%Y%m%d-%H%M%S')}.zip"
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/top-antiguas")
def top_glosas_antiguas(
    top: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R199 P1: top N glosas abiertas más antiguas.

    Diferente a /top-urgentes (por dias_restantes): aquí por
    fecha_creacion. Detecta glosas que llevan demasiado tiempo
    en el sistema sin cerrarse — riesgo regulatorio + costo
    operativo.

    Devuelve top N ordenado ASC por creado_en (la más vieja
    primero) con antiguedad_dias.

    Declarado ANTES de /{glosa_id} para evitar collision.
    """
    from datetime import timezone

    ESTADOS_CERRADOS = {"ACEPTADA", "LEVANTADA", "ARCHIVADA", "CONCILIADA"}

    glosas = (
        db.query(GlosaRecord)
        .filter(~GlosaRecord.estado.in_(ESTADOS_CERRADOS))
        .order_by(GlosaRecord.creado_en.asc())
        .limit(int(top))
        .all()
    )

    ahora = ahora_utc()
    items = []
    for g in glosas:
        cre = g.creado_en
        if cre and cre.tzinfo is None:
            cre = cre.replace(tzinfo=timezone.utc)
        antig = (ahora - cre).days if cre else None
        items.append(
            {
                "id": g.id,
                "eps": g.eps,
                "factura": g.factura,
                "estado": g.estado,
                "creado_en": cre.isoformat() if cre else None,
                "antiguedad_dias": antig,
                "valor_objetado": float(g.valor_objetado or 0),
            }
        )

    return {
        "top_solicitado": int(top),
        "items": items,
    }


@router.get("/top-urgentes")
def top_glosas_urgentes(
    top: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R198 P1: top N glosas más urgentes (menor dias_restantes).

    Solo cuenta abiertas. Las que tienen dias_restantes más
    bajos (incluyendo negativos = vencidas) aparecen primero.

    Útil para "fila de atención prioritaria":
      - vencidas hace mucho aparecen primero
      - luego críticas (0..3d)
      - luego próximas (4..7d)

    Devuelve top N ordenado ASC por dias_restantes.

    Declarado ANTES de /{glosa_id} para evitar collision.
    """
    ESTADOS_CERRADOS = {"ACEPTADA", "LEVANTADA", "ARCHIVADA", "CONCILIADA"}

    glosas = (
        db.query(GlosaRecord)
        .filter(~GlosaRecord.estado.in_(ESTADOS_CERRADOS))
        .order_by(GlosaRecord.dias_restantes.asc())
        .limit(int(top))
        .all()
    )

    items = []
    for g in glosas:
        items.append(
            {
                "id": g.id,
                "eps": g.eps,
                "factura": g.factura,
                "estado": g.estado,
                "dias_restantes": g.dias_restantes,
                "valor_objetado": float(g.valor_objetado or 0),
            }
        )

    return {
        "top_solicitado": int(top),
        "items": items,
    }


@router.get("/top-recuperadas")
def top_glosas_recuperadas(
    top: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R189 P1: top N glosas por valor recuperado.

    Diferente a /top-valor (objetado): aquí solo glosas
    cerradas con valor_recuperado > 0.

    Útil para reconocimiento: "estas son nuestras mejores
    defensas históricas".

    Ordenado DESC por valor_recuperado.
    Declarado ANTES de /{glosa_id} para evitar collision.
    """
    glosas = (
        db.query(GlosaRecord)
        .filter(GlosaRecord.valor_recuperado > 0)
        .order_by(GlosaRecord.valor_recuperado.desc())
        .limit(int(top))
        .all()
    )

    items = []
    for g in glosas:
        items.append(
            {
                "id": g.id,
                "eps": g.eps,
                "factura": g.factura,
                "estado": g.estado,
                "codigo_glosa": g.codigo_glosa,
                "valor_objetado": float(g.valor_objetado or 0),
                "valor_recuperado": float(g.valor_recuperado or 0),
            }
        )

    return {
        "top_solicitado": int(top),
        "items": items,
    }


@router.get("/top-valor")
def top_glosas_por_valor(
    top: int = Query(20, ge=1, le=100),
    abiertas_only: bool = True,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R183 P1: top N glosas por valor objetado.

    Útil para identificar las glosas de mayor cuantía que
    requieren atención prioritaria. Por defecto solo abiertas.

    Devuelve top N ordenado DESC por valor_objetado.

    Declarado ANTES de /{glosa_id} para evitar collision.
    """
    ESTADOS_CERRADOS = {"ACEPTADA", "LEVANTADA", "ARCHIVADA", "CONCILIADA"}

    q = db.query(GlosaRecord)
    if abiertas_only:
        q = q.filter(~GlosaRecord.estado.in_(ESTADOS_CERRADOS))

    glosas = q.order_by(GlosaRecord.valor_objetado.desc()).limit(int(top)).all()

    items = []
    for g in glosas:
        items.append(
            {
                "id": g.id,
                "eps": g.eps,
                "factura": g.factura,
                "estado": g.estado,
                "codigo_glosa": g.codigo_glosa,
                "valor_objetado": float(g.valor_objetado or 0),
                "dias_restantes": g.dias_restantes,
            }
        )

    return {
        "top_solicitado": int(top),
        "abiertas_only": abiertas_only,
        "items": items,
    }


@router.get("/buscar-por-cups")
def buscar_por_cups(
    cups: str = Query(..., min_length=1),
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R186 P1: glosas que mencionan un CUPS específico.

    Cruza ConceptoGlosaRecord.cups_codigo == X para devolver
    las glosas (DISTINCT) que tienen al menos un concepto con
    ese CUPS.

    Útil para investigar: "¿qué glosas hemos tenido por el
    procedimiento 906625?"

    Param `cups`: código exacto.

    Declarado ANTES de /{glosa_id} para evitar collision.
    """
    from app.models.db import ConceptoGlosaRecord

    glosa_ids = {
        row[0]
        for row in (
            db.query(ConceptoGlosaRecord.glosa_id)
            .filter(ConceptoGlosaRecord.cups_codigo == cups)
            .filter(ConceptoGlosaRecord.glosa_id.isnot(None))
            .distinct()
            .all()
        )
    }

    if not glosa_ids:
        return {
            "cups_buscado": cups,
            "encontradas": 0,
            "items": [],
        }

    glosas = (
        db.query(GlosaRecord)
        .filter(GlosaRecord.id.in_(glosa_ids))
        .order_by(GlosaRecord.creado_en.desc())
        .limit(int(limit))
        .all()
    )

    items = []
    for g in glosas:
        items.append(
            {
                "id": g.id,
                "creado_en": (g.creado_en.isoformat() if g.creado_en else None),
                "eps": g.eps,
                "factura": g.factura,
                "estado": g.estado,
                "valor_objetado": float(g.valor_objetado or 0),
            }
        )

    return {
        "cups_buscado": cups,
        "encontradas": len(glosa_ids),
        "items": items,
    }


@router.get("/cups-perfil")
def cups_perfil(
    cups: str = Query(..., min_length=1),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R140 P2: perfil 360º de un CUPS específico.

    Para un CUPS dado, agrega información de TODOS los conceptos
    de glosa que lo mencionan:
      - frecuencia, valor total y promedio
      - distribución por EPS
      - códigos de glosa asociados
      - centros de costo

    Útil para responder: "¿qué historia tiene este procedimiento
    en glosas?"

    Param `cups`: código CUPS (ej. "906625", "39143A").

    Declarado ANTES de /{glosa_id} para evitar collision con path
    resolver de FastAPI.
    """
    conceptos = db.query(ConceptoGlosaRecord).filter(ConceptoGlosaRecord.cups_codigo == cups).all()

    if not conceptos:
        return {
            "cups_codigo": cups,
            "sin_historial": True,
        }

    valores = [float(c.valor_objetado or 0) for c in conceptos]
    descripcion = next(
        (c.cups_descripcion for c in conceptos if c.cups_descripcion),
        "",
    )

    glosa_ids = {c.glosa_id for c in conceptos if c.glosa_id is not None}
    epss: dict[str, int] = {}
    if glosa_ids:
        for g in db.query(GlosaRecord).filter(GlosaRecord.id.in_(glosa_ids)).all():
            if g.eps:
                epss[g.eps] = epss.get(g.eps, 0) + 1

    por_codigo: dict[str, int] = {}
    centros: set[str] = set()
    for c in conceptos:
        if c.codigo_glosa:
            por_codigo[c.codigo_glosa] = por_codigo.get(c.codigo_glosa, 0) + 1
        if c.centro_costo:
            centros.add(c.centro_costo)

    return {
        "cups_codigo": cups,
        "cups_descripcion": descripcion[:200] if descripcion else "",
        "sin_historial": False,
        "frecuencia_total": len(conceptos),
        "valor_objetado_total": int(sum(valores)),
        "valor_promedio": round(sum(valores) / len(valores), 2),
        "por_eps": dict(sorted(epss.items(), key=lambda x: x[1], reverse=True)),
        "por_codigo_glosa": dict(sorted(por_codigo.items(), key=lambda x: x[1], reverse=True)),
        "centros_costo": sorted(centros),
    }


# ─── Sprint #3 — Detector de glosas similares (DEBE ir antes de /{glosa_id}) ───
@router.get("/similares-bloque")
def detectar_glosas_similares_en_bloque(
    factura: Optional[str] = None,
    eps: Optional[str] = None,
    codigo_glosa: Optional[str] = None,
    cups: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Agrupa glosas pendientes por (eps, código_glosa, cups) para responder
    en bloque con el mismo dictamen base.

    Sin filtros: devuelve todos los grupos con >= 2 glosas pendientes
    (RADICADA / REQUIERE_SOPORTES) del usuario actual. Con filtros:
    devuelve solo los grupos que coinciden.
    """
    base_q = db.query(GlosaRecord).filter(
        GlosaRecord.estado.notin_(
            ["LEVANTADA", "CONCILIADA", "ACEPTADA", "RATIFICADA", "ARCHIVADA", "DUPLICADA_OCULTA"]
        ),
    )
    if current_user.rol == "AUDITOR":
        base_q = base_q.filter(
            (GlosaRecord.auditor_email == current_user.email)
            | (GlosaRecord.gestor_nombre == current_user.email)
        )
    if factura:
        base_q = base_q.filter(GlosaRecord.factura == factura)
    if eps:
        base_q = base_q.filter(GlosaRecord.eps.ilike(f"%{eps.upper()}%"))
    if codigo_glosa:
        base_q = base_q.filter(GlosaRecord.codigo_glosa == codigo_glosa.upper())
    if cups:
        base_q = base_q.filter(GlosaRecord.cups_servicio == cups)

    todas = base_q.limit(2000).all()

    grupos: dict[tuple, list] = {}
    for g in todas:
        key = (
            (g.eps or "").upper().strip(),
            (g.codigo_glosa or "").upper().strip(),
            (g.cups_servicio or "").strip(),
        )
        grupos.setdefault(key, []).append(g)

    resultado = []
    for (eps_g, cod_g, cups_g), glosas in grupos.items():
        if len(glosas) < 2:
            continue
        valor_total = sum(float(g.valor_objetado or 0.0) for g in glosas)
        respondidas = [g for g in glosas if g.dictamen and len(g.dictamen) > 200]
        dictamen_modelo = (
            max(respondidas, key=lambda g: len(g.dictamen or "")).dictamen if respondidas else None
        )
        resultado.append(
            {
                "eps": eps_g,
                "codigo_glosa": cod_g,
                "cups": cups_g,
                "n_glosas": len(glosas),
                "valor_total": valor_total,
                "dictamen_modelo_glosa_id": (
                    max(respondidas, key=lambda g: len(g.dictamen or "")).id
                    if respondidas
                    else None
                ),
                "tiene_dictamen_modelo": dictamen_modelo is not None,
                "glosas": [
                    {
                        "id": g.id,
                        "factura": g.factura,
                        "valor_objetado": float(g.valor_objetado or 0.0),
                        "estado": g.estado,
                        "workflow_state": g.workflow_state,
                        "dias_restantes": g.dias_restantes,
                        "tiene_dictamen": bool(g.dictamen and len(g.dictamen) > 200),
                    }
                    for g in glosas
                ],
            }
        )

    resultado.sort(key=lambda x: -x["valor_total"])
    return {
        "grupos": resultado,
        "total_grupos": len(resultado),
        "total_glosas_agrupables": sum(g["n_glosas"] for g in resultado),
    }


@router.get("/{glosa_id}")
def obtener_glosa(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    repo = GlosaRepository(db)
    glosa = repo.obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(status_code=404, detail="Glosa no encontrada")
    from app.services.dictamen_stale import motivo_stale

    aviso_stale = motivo_stale(glosa, db)
    return {
        "id": glosa.id,
        "eps": glosa.eps,
        "paciente": glosa.paciente,
        "codigo_glosa": glosa.codigo_glosa,
        "valor_objetado": glosa.valor_objetado,
        "valor_aceptado": glosa.valor_aceptado,
        "etapa": glosa.etapa,
        "estado": glosa.estado,
        "dictamen": glosa.dictamen,
        "dictamen_generado_en": glosa.dictamen_generado_en.isoformat()
        if getattr(glosa, "dictamen_generado_en", None)
        else None,
        "dictamen_stale": bool(aviso_stale),
        "dictamen_stale_motivo": aviso_stale,
        "dias_restantes": glosa.dias_restantes,
        "factura": glosa.factura,
        "numero_radicado": glosa.numero_radicado,
        "consecutivo_dgh": glosa.consecutivo_dgh,
        "gestor_nombre": glosa.gestor_nombre,
        "fecha_radicacion_factura": glosa.fecha_radicacion_factura.isoformat()
        if glosa.fecha_radicacion_factura
        else None,
        "fecha_documento_dgh": glosa.fecha_documento_dgh.isoformat()
        if glosa.fecha_documento_dgh
        else None,
        "fecha_recepcion": glosa.fecha_recepcion.isoformat() if glosa.fecha_recepcion else None,
        "fecha_entrega": glosa.fecha_entrega.isoformat() if glosa.fecha_entrega else None,
        "fecha_vencimiento": glosa.fecha_vencimiento.isoformat()
        if glosa.fecha_vencimiento
        else None,
        "radicado_info": glosa.radicado_info,
        "referencia": glosa.referencia,
        "observacion_tecnico": glosa.observacion_tecnico,
        "tipo_glosa_excel": glosa.tipo_glosa_excel,
        "profesional_medico": glosa.profesional_medico,
        "creado_en": glosa.creado_en.isoformat() if glosa.creado_en else None,
    }


@router.get("/{glosa_id}/preparar-conciliacion")
def preparar_conciliacion(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Material táctico para llegar preparado a la audiencia de
    conciliación: contraargumentos probables de la EPS, respuesta
    sugerida a cada uno, valor mínimo aceptable y recomendación
    táctica basada en el histórico de esa EPS contra ese tipo de glosa.
    """
    from app.services.conciliador_ia import preparar_audiencia

    return preparar_audiencia(db, glosa_id)


@router.post("/autopiloto-niveles")
def autopiloto_niveles_batch(
    payload: dict,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Batch: clasifica un conjunto de glosas en niveles de auto-piloto.

    Recibe `{"glosa_ids": [1, 2, ...]}` (máx 300) y devuelve para cada una
    `{glosa_id, nivel, icono, color, etiqueta, razon, accion_sugerida}`.
    Permite a la UI pintar un badge por glosa indicando si es mecánica
    (1 click), revisable rápido o requiere editor manual.
    """
    from app.services.autopiloto_nivel import clasificar_nivel

    ids = payload.get("glosa_ids") or []
    if not isinstance(ids, list) or not ids:
        return {"items": []}
    ids = [int(x) for x in ids if isinstance(x, (int, float, str))][:300]
    glosas = db.query(GlosaRecord).filter(GlosaRecord.id.in_(ids)).all()
    items = []
    for g in glosas:
        try:
            n = clasificar_nivel(g, db)
            n["glosa_id"] = g.id
            items.append(n)
        except Exception:
            continue
    return {"total": len(items), "items": items}


@router.get("/mi-estilo")
def mi_estilo_gestor(
    codigo_glosa: Optional[str] = None,
    eps: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Patrones de refinamiento del gestor logueado. Si pasás codigo_glosa
    y/o eps, filtra al contexto similar.

    Útil para mostrar al gestor en el editor: '💭 Tu estilo: sueles
    agregar T-760 + tono conciliador en TA0201 contra FAMISANAR.'
    """
    from app.services.memoria_gestor import patron_gestor

    return patron_gestor(
        db,
        autor_email=current_user.email,
        codigo_glosa=codigo_glosa or "",
        eps=eps or "",
    )


@router.get("/{glosa_id}/auditoria")
def auditar_glosa(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Auto-auditoría suave del dictamen antes de marcar como RESPONDIDA.

    Devuelve hallazgos de calidad (cita normativa, código RE coherente,
    texto canónico de extemporáneas/ratificadas, soportes, etc.). NO
    bloquea — el gestor decide si corrige o radica de todos modos.
    """
    repo = GlosaRepository(db)
    glosa = repo.obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")
    from app.services.auditor_dictamen import auditar_dictamen

    return auditar_dictamen(glosa, db)


@router.delete("/{glosa_id}")
def eliminar_glosa(
    glosa_id: int,
    motivo: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_coordinador_o_admin),
):
    """Elimina una glosa del historial. Se mueve a la papelera (restaurable
    por 30 días) antes de borrarla de la tabla principal."""
    repo = GlosaRepository(db)
    glosa = repo.obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(status_code=404, detail="Glosa no encontrada")
    # Mover a papelera (soft-delete con snapshot)
    try:
        from app.api.routers.papelera import mover_a_papelera

        pap_id = mover_a_papelera(db, glosa, eliminado_por=current_user.email, motivo=motivo or "")
    except Exception as e:
        logger.warning(f"No se pudo mover a papelera: {e}")
        pap_id = None
    db.delete(glosa)
    db.commit()
    logger.info(f"Glosa eliminada ID={glosa_id} por {current_user.email} (papelera #{pap_id})")
    return {
        "message": f"Glosa {glosa_id} eliminada",
        "papelera_id": pap_id,
        "restaurable_hasta": "30 días",
    }


class DecisionEPSInput(BaseModel):
    decision_eps: str
    valor_recuperado: float = 0.0
    observacion_eps: Optional[str] = None


class AsignarAuditorInput(BaseModel):
    auditor_email: str


class WorkflowTransicionInput(BaseModel):
    nuevo_estado: str  # BORRADOR | EN_REVISION | APROBADA | RADICADA
    comentario: Optional[str] = None


# Transiciones válidas del workflow (from_estado -> set(to_estado))
_WORKFLOW_TRANSICIONES = {
    "BORRADOR": {"EN_REVISION"},
    "EN_REVISION": {"BORRADOR", "APROBADA"},
    "APROBADA": {"RADICADA", "EN_REVISION"},
    "RADICADA": set(),  # estado final
}


@router.patch("/{glosa_id}/workflow")
def cambiar_workflow(
    glosa_id: int,
    data: WorkflowTransicionInput,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Cambia el estado del workflow de aprobación.

    Transiciones permitidas:
      BORRADOR -> EN_REVISION       (auditor solicita revisión)
      EN_REVISION -> APROBADA       (coordinador/admin aprueba)
      EN_REVISION -> BORRADOR       (coordinador devuelve para corregir)
      APROBADA -> RADICADA          (una vez radicada ante la EPS)
      APROBADA -> EN_REVISION       (se detecta algo para revisar)

    Permisos:
    - AUDITOR puede mover BORRADOR -> EN_REVISION de sus propias glosas.
    - COORDINADOR y SUPER_ADMIN pueden hacer cualquier transición.
    """
    nuevo = data.nuevo_estado.upper().strip()
    if nuevo not in {"BORRADOR", "EN_REVISION", "APROBADA", "RADICADA"}:
        raise HTTPException(400, f"Estado inválido: {nuevo}")

    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    actual = (glosa.workflow_state or "BORRADOR").upper()

    # Si no existe transición desde el estado actual, inicializar como BORRADOR
    if actual not in _WORKFLOW_TRANSICIONES:
        actual = "BORRADOR"

    if nuevo not in _WORKFLOW_TRANSICIONES.get(actual, set()):
        raise HTTPException(
            400,
            f"Transición no permitida: {actual} -> {nuevo}. "
            f"Desde {actual} solo puedes ir a: {sorted(_WORKFLOW_TRANSICIONES.get(actual, set())) or 'ninguno (estado final)'}",
        )

    # Validar permisos por transición
    if current_user.rol == "AUDITOR":
        # Auditor solo puede enviar a revisión sus glosas
        if nuevo != "EN_REVISION" or actual != "BORRADOR":
            raise HTTPException(403, "Como AUDITOR solo puedes enviar glosas propias a revisión")
        if glosa.auditor_email and glosa.auditor_email != current_user.email:
            # Si está asignada a otro auditor, no puede
            raise HTTPException(403, "Esta glosa está asignada a otro auditor")
    elif current_user.rol == "VIEWER":
        raise HTTPException(403, "VIEWER no puede cambiar estados")

    glosa.workflow_state = nuevo
    if data.comentario:
        nota = glosa.nota_workflow or ""
        nueva_nota = f"[{ahora_utc().strftime('%Y-%m-%d %H:%M')} {current_user.email} {actual}->{nuevo}] {data.comentario}"
        glosa.nota_workflow = (nota + " | " + nueva_nota)[-500:] if nota else nueva_nota[:500]

    db.commit()
    db.refresh(glosa)

    AuditRepository(db).registrar(
        usuario_email=current_user.email,
        usuario_rol=current_user.rol,
        accion="WORKFLOW",
        tabla="historial",
        registro_id=glosa_id,
        campo="workflow_state",
        valor_anterior=actual,
        valor_nuevo=nuevo,
        detalle=data.comentario or f"Transición {actual} -> {nuevo}",
    )
    return {
        "message": "Workflow actualizado",
        "glosa_id": glosa_id,
        "estado_anterior": actual,
        "estado_nuevo": nuevo,
        "nota_workflow": glosa.nota_workflow,
    }


@router.patch("/{glosa_id}/decision-eps")
def registrar_decision_eps(
    glosa_id: int,
    data: DecisionEPSInput,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_auditor_o_superior),
):
    DECISIONES = {"LEVANTADA", "ACEPTADA", "RATIFICADA", "PENDIENTE"}
    decision = data.decision_eps.upper()
    if decision not in DECISIONES:
        raise HTTPException(400, f"Decisión inválida. Use: {', '.join(DECISIONES)}")
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")
    glosa.decision_eps = decision
    glosa.fecha_decision_eps = ahora_utc()
    glosa.valor_recuperado = data.valor_recuperado
    if data.observacion_eps:
        glosa.observacion_eps = data.observacion_eps
    if decision in ("LEVANTADA", "ACEPTADA", "RATIFICADA"):
        glosa.estado = decision
    db.commit()

    # Ronda 3 — Aprendizaje por retroalimentación:
    # Si la EPS LEVANTÓ la glosa, promover el argumento exitoso a Plantilla
    # Gold automáticamente (si no existe ya una para esa combinación EPS+código).
    # Si RATIFICÓ, desactivar cualquier Gold previa de esa combinación para
    # que la IA no la sugiera más.
    try:
        from app.services.aprendizaje_feedback import aprender_de_decision_eps

        aprender_de_decision_eps(
            db=db,
            glosa=glosa,
            decision=decision,
            creado_por=current_user.email,
        )
    except Exception as _e:
        # El aprendizaje nunca debe bloquear la decisión; solo logear.
        import logging as _l

        _l.getLogger("motor_glosas").warning(f"Aprendizaje feedback falló: {_e}")

    AuditRepository(db).registrar(
        usuario_email=current_user.email,
        usuario_rol=current_user.rol,
        accion="DECISION_EPS",
        tabla="glosas",
        registro_id=glosa_id,
        campo="decision_eps",
        valor_nuevo=decision,
        detalle=f"Decisión: {decision} | recuperado: ${data.valor_recuperado:,.0f}",
    )
    return {"message": "Decisión registrada", "glosa_id": glosa_id, "decision_eps": decision}


# ─── Sprint #4 — Decisión EPS en LOTE ────────────────────────────────────
class DecisionEPSLoteInput(BaseModel):
    glosa_ids: list[int] = Field(..., min_length=1, max_length=500)
    decision_eps: str
    observacion_eps: Optional[str] = None
    # Si la decisión es LEVANTADA, el usuario suele dejar valor_recuperado
    # = valor_objetado de cada glosa. Si manda explícito, se usa ese.
    valor_recuperado: Optional[float] = None


@router.patch("/decision-eps-lote")
def registrar_decision_eps_lote(
    data: DecisionEPSLoteInput,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_auditor_o_superior),
):
    """Aplica la misma decisión de la EPS a N glosas seleccionadas.

    Caso típico: la EPS responde el lote y todas quedan LEVANTADAS — el
    gestor selecciona checkboxes en /mis-asignaciones y marca las N de un
    solo click. Para LEVANTADA, si no se manda `valor_recuperado`, se
    asume valor_objetado de cada glosa (recupera el 100%).
    """
    DECISIONES = {"LEVANTADA", "ACEPTADA", "RATIFICADA", "PENDIENTE"}
    decision = data.decision_eps.upper()
    if decision not in DECISIONES:
        raise HTTPException(400, f"Decisión inválida. Use: {', '.join(DECISIONES)}")

    procesadas, fallidas, recuperado_total = 0, [], 0.0
    for gid in data.glosa_ids:
        glosa = GlosaRepository(db).obtener_por_id(gid)
        if not glosa:
            fallidas.append({"glosa_id": gid, "motivo": "no encontrada"})
            continue
        # No re-decidir glosas ya cerradas con la misma decisión
        if (glosa.decision_eps or "").upper() == decision:
            fallidas.append({"glosa_id": gid, "motivo": f"ya estaba {decision}"})
            continue
        # Para LEVANTADA, valor recuperado por defecto = valor_objetado
        if decision == "LEVANTADA":
            recuperado = (
                data.valor_recuperado
                if data.valor_recuperado is not None
                else float(glosa.valor_objetado or 0.0)
            )
        else:
            recuperado = data.valor_recuperado or 0.0

        glosa.decision_eps = decision
        glosa.fecha_decision_eps = ahora_utc()
        glosa.valor_recuperado = recuperado
        if data.observacion_eps:
            glosa.observacion_eps = data.observacion_eps
        if decision in ("LEVANTADA", "ACEPTADA", "RATIFICADA"):
            glosa.estado = decision

        # Aprendizaje feedback: promover argumento exitoso a Plantilla Gold
        try:
            from app.services.aprendizaje_feedback import aprender_de_decision_eps

            aprender_de_decision_eps(
                db=db,
                glosa=glosa,
                decision=decision,
                creado_por=current_user.email,
            )
        except Exception as _e:
            import logging as _l

            _l.getLogger("motor_glosas").warning(f"Aprendizaje feedback (lote) falló: {_e}")

        recuperado_total += recuperado
        procesadas += 1

    db.commit()

    AuditRepository(db).registrar(
        usuario_email=current_user.email,
        usuario_rol=current_user.rol,
        accion="DECISION_EPS_LOTE",
        tabla="glosas",
        registro_id=0,
        campo="decision_eps",
        valor_nuevo=decision,
        detalle=(
            f"Lote {decision}: {procesadas}/{len(data.glosa_ids)} procesadas | "
            f"recuperado total: ${recuperado_total:,.0f}"
        ),
    )

    return {
        "message": f"Lote procesado: {procesadas} glosas → {decision}",
        "decision_eps": decision,
        "procesadas": procesadas,
        "fallidas": fallidas,
        "valor_recuperado_total": recuperado_total,
    }


@router.patch("/{glosa_id}/asignar")
def asignar_auditor(
    glosa_id: int,
    data: AsignarAuditorInput,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_coordinador_o_admin),
):
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")
    anterior = glosa.auditor_email
    glosa.auditor_email = data.auditor_email
    db.commit()
    AuditRepository(db).registrar(
        usuario_email=current_user.email,
        usuario_rol=current_user.rol,
        accion="ASIGNAR",
        tabla="glosas",
        registro_id=glosa_id,
        valor_anterior=anterior,
        valor_nuevo=data.auditor_email,
    )
    return {"ok": True, "id": glosa_id, "auditor_email": data.auditor_email}


class BulkAsignarInput(BaseModel):
    glosa_ids: list[int]
    auditor_email: str


def _resolver_destino_con_vacaciones(db: Session, email: str) -> tuple[str, bool]:
    """Si el destino esta en vacaciones activas y tiene delega_a_email,
    retorna el delegado en lugar del original. Retorna (email_real,
    fue_redirigido).
    """
    if not email:
        return email, False
    from datetime import datetime, timezone as _tz

    ahora = datetime.now(_tz.utc)
    u = db.query(UsuarioRecord).filter(UsuarioRecord.email.ilike(email)).first()
    if not u:
        return email, False
    if (
        u.vacaciones_desde
        and u.vacaciones_hasta
        and u.delega_a_email
        and u.vacaciones_desde <= ahora <= u.vacaciones_hasta
    ):
        return u.delega_a_email, True
    return email, False


@router.post("/bulk/asignar")
def bulk_asignar(
    data: BulkAsignarInput,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_coordinador_o_admin),
):
    """Bulk: reasigna N glosas seleccionadas a otro gestor en una sola
    operacion. Solo COORDINADOR/SUPER_ADMIN.

    Aplica `auditor_email` Y `gestor_nombre` para que la glosa
    aparezca en "Mis glosas" del nuevo dueno.

    Si el destino esta en vacaciones activas y configuro delega_a, la
    asignacion se redirige automaticamente al delegado.
    """
    if not data.glosa_ids:
        raise HTTPException(400, "Lista vacia")
    if len(data.glosa_ids) > 200:
        raise HTTPException(400, "Maximo 200 glosas por bulk")

    destino_real, redirigido = _resolver_destino_con_vacaciones(db, data.auditor_email)

    actualizadas = 0
    no_encontradas = 0
    for gid in data.glosa_ids:
        g = GlosaRepository(db).obtener_por_id(gid)
        if not g:
            no_encontradas += 1
            continue
        anterior = g.auditor_email or g.gestor_nombre
        g.auditor_email = destino_real
        g.gestor_nombre = destino_real
        actualizadas += 1
        try:
            AuditRepository(db).registrar(
                usuario_email=current_user.email,
                usuario_rol=current_user.rol,
                accion="BULK_ASIGNAR",
                tabla="glosas",
                registro_id=gid,
                valor_anterior=anterior,
                valor_nuevo=destino_real,
                detalle=("redirigido_vacaciones" if redirigido else None),
            )
        except Exception:
            pass
    db.commit()
    return {
        "ok": True,
        "actualizadas": actualizadas,
        "no_encontradas": no_encontradas,
        "total_solicitadas": len(data.glosa_ids),
        "auditor_email": destino_real,
        "redirigido_vacaciones": redirigido,
        "destino_solicitado": data.auditor_email,
    }


class BulkIdsInput(BaseModel):
    glosa_ids: list[int]


class ReasignarDeGestorInput(BaseModel):
    gestor_origen: str
    gestor_destino: str
    solo_pendientes: bool = True
    limite: int = 100


@router.post("/bulk/reasignar-de-gestor")
def bulk_reasignar_de_gestor(
    data: ReasignarDeGestorInput,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_coordinador_o_admin),
):
    """Mueve TODAS las glosas (o solo las pendientes) de un gestor a
    otro en una sola operacion. Util cuando un gestor entra a
    vacaciones, esta inactivo o el coordinador rebalancea carga.

    Args:
        gestor_origen: email del gestor que actualmente tiene las
            glosas asignadas.
        gestor_destino: email del gestor que recibira las glosas.
        solo_pendientes: si True (default), solo mueve glosas en
            estado RADICADA/EN_REVISION/BORRADOR. Si False, mueve
            tambien las cerradas (rara vez util — para limpieza).
        limite: max glosas a mover en una sola llamada (default 100,
            max 500). Por seguridad operacional.

    Solo COORDINADOR/SUPER_ADMIN.
    """
    if not data.gestor_origen or not data.gestor_destino:
        raise HTTPException(400, "Origen y destino son requeridos")
    if data.gestor_origen.strip().lower() == data.gestor_destino.strip().lower():
        raise HTTPException(400, "Origen y destino no pueden ser iguales")
    limite = max(1, min(int(data.limite or 100), 500))

    q = db.query(GlosaRecord).filter(
        (GlosaRecord.auditor_email == data.gestor_origen)
        | (GlosaRecord.gestor_nombre == data.gestor_origen)
    )
    if data.solo_pendientes:
        q = q.filter(GlosaRecord.estado.in_(["RADICADA", "EN_REVISION", "BORRADOR"]))
    glosas = q.order_by(GlosaRecord.creado_en.asc()).limit(limite).all()

    actualizadas = 0
    valor_movido = 0.0
    for g in glosas:
        g.auditor_email = data.gestor_destino
        g.gestor_nombre = data.gestor_destino
        valor_movido += float(g.valor_objetado or 0)
        actualizadas += 1
        try:
            AuditRepository(db).registrar(
                usuario_email=current_user.email,
                usuario_rol=current_user.rol,
                accion="REASIGNAR_DE_GESTOR",
                tabla="glosas",
                registro_id=g.id,
                valor_anterior=data.gestor_origen,
                valor_nuevo=data.gestor_destino,
                detalle=("solo_pendientes" if data.solo_pendientes else "todas"),
            )
        except Exception:
            pass
    db.commit()
    return {
        "ok": True,
        "actualizadas": actualizadas,
        "limite": limite,
        "valor_movido": valor_movido,
        "gestor_origen": data.gestor_origen,
        "gestor_destino": data.gestor_destino,
        "solo_pendientes": data.solo_pendientes,
    }


class AutoAsignarInput(BaseModel):
    glosa_ids: list[int]
    candidatos: list[str] = []  # opcional: lista de emails candidatos
    incluir_carga_actual: bool = True


@router.post("/bulk/auto-asignar")
def bulk_auto_asignar(
    data: AutoAsignarInput,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_coordinador_o_admin),
):
    """Distribuye N glosas equitativamente entre auditores activos
    usando algoritmo de balanceo: el que tiene MENOS pendientes recibe
    primero, hasta nivelar.

    Args:
        glosa_ids: glosas a repartir (max 500).
        candidatos: si se pasa, solo se distribuye entre estos emails.
            Si esta vacio, se usan todos los AUDITOR/COORDINADOR
            activos del sistema.
        incluir_carga_actual: si True (default), considera la carga
            actual de pendientes de cada candidato como base, asi el
            balance es real (no solo entre las nuevas). Si False, solo
            balancea las nuevas (primer asignado = primero recibe).

    Solo COORDINADOR/SUPER_ADMIN.
    """
    from sqlalchemy import func as _func

    if not data.glosa_ids:
        raise HTTPException(400, "Lista vacia")
    if len(data.glosa_ids) > 500:
        raise HTTPException(400, "Maximo 500 glosas por auto-asignar")

    # Determinar candidatos: explicitos o todos los AUDITOR/COORDINADOR activos
    from datetime import datetime, timezone as _tz

    ahora_dt = datetime.now(_tz.utc)
    if data.candidatos:
        candidatos = [c.strip() for c in data.candidatos if c and "@" in c]
    else:
        users = (
            db.query(UsuarioRecord)
            .filter(UsuarioRecord.activo == 1)
            .filter(UsuarioRecord.rol.in_(["AUDITOR", "COORDINADOR", "SUPER_ADMIN"]))
            .all()
        )
        # Excluir usuarios actualmente en vacaciones (sin delegacion).
        # Si tienen delegacion, el delegado entra en su lugar via _resolver.
        candidatos = []
        for u in users:
            en_vacaciones = (
                u.vacaciones_desde
                and u.vacaciones_hasta
                and u.vacaciones_desde <= ahora_dt <= u.vacaciones_hasta
            )
            if en_vacaciones:
                continue  # se omite — su delegado (si existe) ya estara en la lista
            candidatos.append(u.email)

    # Resolver vacaciones para cada candidato (deduplica)
    candidatos_real = []
    seen = set()
    for c in candidatos:
        cr, _ = _resolver_destino_con_vacaciones(db, c)
        if cr and cr.lower() not in seen:
            seen.add(cr.lower())
            candidatos_real.append(cr)
    candidatos = candidatos_real

    if not candidatos:
        raise HTTPException(400, "Sin candidatos disponibles")

    # Carga actual por candidato (pendientes asignados)
    carga: dict[str, int] = {c: 0 for c in candidatos}
    if data.incluir_carga_actual:
        rows = (
            db.query(
                GlosaRecord.auditor_email,
                _func.count(GlosaRecord.id),
            )
            .filter(GlosaRecord.auditor_email.in_(candidatos))
            .filter(GlosaRecord.estado.in_(["RADICADA", "EN_REVISION", "BORRADOR"]))
            .group_by(GlosaRecord.auditor_email)
            .all()
        )
        for email, n in rows:
            if email in carga:
                carga[email] = int(n)

    # Asignar usando heap de menor carga
    import heapq

    heap = [(carga[c], c) for c in candidatos]
    heapq.heapify(heap)

    asignaciones: dict[str, list[int]] = {c: [] for c in candidatos}
    actualizadas = 0
    no_encontradas = 0

    for gid in data.glosa_ids:
        g = GlosaRepository(db).obtener_por_id(gid)
        if not g:
            no_encontradas += 1
            continue
        carga_actual, email_destino = heapq.heappop(heap)
        anterior = g.auditor_email or g.gestor_nombre
        g.auditor_email = email_destino
        g.gestor_nombre = email_destino
        asignaciones[email_destino].append(gid)
        actualizadas += 1
        heapq.heappush(heap, (carga_actual + 1, email_destino))
        try:
            AuditRepository(db).registrar(
                usuario_email=current_user.email,
                usuario_rol=current_user.rol,
                accion="AUTO_ASIGNAR",
                tabla="glosas",
                registro_id=gid,
                valor_anterior=anterior,
                valor_nuevo=email_destino,
                detalle="round-robin balanceado",
            )
        except Exception:
            pass
    db.commit()

    distribucion = [
        {"email": k, "asignadas": len(v), "ids_sample": v[:5]} for k, v in asignaciones.items() if v
    ]
    distribucion.sort(key=lambda x: -x["asignadas"])

    return {
        "ok": True,
        "actualizadas": actualizadas,
        "no_encontradas": no_encontradas,
        "total_solicitadas": len(data.glosa_ids),
        "candidatos_usados": len(candidatos),
        "distribucion": distribucion,
    }


@router.post("/bulk/exportar-csv")
def bulk_exportar_csv(
    data: BulkIdsInput,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Exporta a CSV las glosas seleccionadas. StreamingResponse para
    no cargar todo en memoria.
    """
    from fastapi.responses import StreamingResponse
    import csv as _csv
    import io as _io
    from datetime import datetime as _dt

    if not data.glosa_ids:
        raise HTTPException(400, "Lista vacia")
    if len(data.glosa_ids) > 1000:
        raise HTTPException(400, "Maximo 1000 glosas por export")

    glosas = (
        db.query(GlosaRecord)
        .filter(GlosaRecord.id.in_(data.glosa_ids))
        .order_by(GlosaRecord.id)
        .all()
    )

    def _gen():
        buf = _io.StringIO()
        w = _csv.writer(buf)
        w.writerow(
            [
                "id",
                "factura",
                "eps",
                "codigo_glosa",
                "valor_objetado",
                "valor_aceptado",
                "valor_recuperado",
                "estado",
                "etapa",
                "auditor_email",
                "gestor_nombre",
                "fecha_recepcion",
                "fecha_entrega",
                "decision_eps",
                "creado_en",
            ]
        )
        yield buf.getvalue()
        buf.seek(0)
        buf.truncate(0)
        for g in glosas:
            w.writerow(
                [
                    g.id,
                    g.factura or "",
                    g.eps or "",
                    g.codigo_glosa or "",
                    float(g.valor_objetado or 0),
                    float(g.valor_aceptado or 0),
                    float(g.valor_recuperado or 0),
                    g.estado or "",
                    g.etapa or "",
                    g.auditor_email or "",
                    g.gestor_nombre or "",
                    g.fecha_recepcion.isoformat() if g.fecha_recepcion else "",
                    g.fecha_entrega.isoformat() if g.fecha_entrega else "",
                    g.decision_eps or "",
                    g.creado_en.isoformat() if g.creado_en else "",
                ]
            )
            yield buf.getvalue()
            buf.seek(0)
            buf.truncate(0)

    fname = f"glosas-seleccion-{_dt.now().strftime('%Y%m%d-%H%M%S')}.csv"
    return StreamingResponse(
        _gen(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/casos-similares/{glosa_id}")
def casos_similares(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    from app.services.rag_service import RAGService

    glosa = db.query(GlosaRecord).filter(GlosaRecord.id == glosa_id).first()
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")
    casos = RAGService().buscar_casos_similares(
        texto_glosa=glosa.dictamen or "",
        eps=glosa.eps,
        codigo_glosa=glosa.codigo_glosa or "",
        db=db,
        top_k=5,
        solo_exitosos=False,
    )
    return {"glosa_id": glosa_id, "casos_similares": casos}


@router.get("/{glosa_id}/conceptos")
def listar_conceptos_glosa(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Lista el detalle por concepto de una glosa (cargados desde hojas I/R).

    Devuelve también el encabezado de la glosa para que el front pueda
    pintar de una sola llamada la factura completa (fechas, vencimiento,
    semáforo, y todos los conceptos precargados para analizar).
    """
    glosa = db.query(GlosaRecord).filter(GlosaRecord.id == glosa_id).first()
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    conceptos = (
        db.query(ConceptoGlosaRecord)
        .filter(ConceptoGlosaRecord.glosa_id == glosa_id)
        .order_by(ConceptoGlosaRecord.codigo_glosa, ConceptoGlosaRecord.cups_codigo)
        .all()
    )

    total_conceptos_valor = sum(float(c.valor_objetado or 0) for c in conceptos)

    return {
        "glosa": {
            "id": glosa.id,
            "factura": glosa.factura,
            "consecutivo_dgh": glosa.consecutivo_dgh,
            "eps": glosa.eps,
            "eps_codigo": glosa.eps_codigo,
            "gestor_nombre": glosa.gestor_nombre,
            "tecnico_recepcion": glosa.tecnico_recepcion,
            "tipo_glosa_excel": glosa.tipo_glosa_excel,
            "profesional_medico": glosa.profesional_medico,
            "estado": glosa.estado,
            "valor_objetado": glosa.valor_objetado,
            "valor_factura": glosa.valor_factura,
            "saldo_factura": glosa.saldo_factura,
            "tercero_nit": glosa.tercero_nit,
            "fecha_radicacion_factura": glosa.fecha_radicacion_factura.isoformat()
            if glosa.fecha_radicacion_factura
            else None,
            "fecha_documento_dgh": glosa.fecha_documento_dgh.isoformat()
            if glosa.fecha_documento_dgh
            else None,
            "fecha_recepcion": glosa.fecha_recepcion.isoformat() if glosa.fecha_recepcion else None,
            "fecha_entrega": glosa.fecha_entrega.isoformat() if glosa.fecha_entrega else None,
            "fecha_vencimiento": glosa.fecha_vencimiento.isoformat()
            if glosa.fecha_vencimiento
            else None,
            "fecha_objecion_eps": glosa.fecha_objecion_eps.isoformat()
            if glosa.fecha_objecion_eps
            else None,
            "dias_restantes": glosa.dias_restantes,
            "prioridad": glosa.prioridad,
        },
        "conceptos": [
            {
                "id": c.id,
                "oid_dgh": c.oid_dgh,
                "codigo_glosa": c.codigo_glosa,
                "nombre_glosa": c.nombre_glosa,
                "cups_codigo": c.cups_codigo,
                "cups_descripcion": c.cups_descripcion,
                "centro_costo": c.centro_costo,
                "valor_objetado": c.valor_objetado,
                "observacion_eps": c.observacion_eps,
                "dictamen_html": c.dictamen_html,
                "score": c.score,
                "respondido_en": c.respondido_en.isoformat() if c.respondido_en else None,
                "respondido_por": c.respondido_por,
            }
            for c in conceptos
        ],
        "totales": {
            "conceptos": len(conceptos),
            "valor_suma_conceptos": total_conceptos_valor,
            "valor_glosa_cabecera": glosa.valor_objetado or 0,
        },
    }


def _parsear_filas_excel(texto: str) -> list[dict]:
    """
    Parsea el texto pegado de Excel y extrae cada fila como diccionario.
    Formato esperado (8 columnas): ENTIDAD | FACTURA | VALOR | CODIGO |
    CONCEPTO GLOSA | CUPS | SERVICIO | MOTIVO

    Acepta como separador **Tab** (copy/paste directo del Excel) o **"|"**
    (pipe, cuando el usuario exporta desde Office y lo pega aquí). Si una
    fila trae más columnas que las esperadas (porque el MOTIVO contiene el
    mismo separador), las columnas extra se re-unen al final en `motivo`.
    """
    filas: list[dict] = []
    if not texto:
        return filas

    lineas = texto.strip().split("\n")
    CAMPOS = ["eps", "factura", "valor", "codigo", "descripcion", "cups", "servicio", "motivo"]

    for i, linea in enumerate(lineas):
        linea = linea.strip()
        if not linea:
            continue

        # Auto-detectar separador: Tab si existe, sino pipe.
        if "\t" in linea:
            partes = [p.strip() for p in linea.split("\t")]
        elif "|" in linea:
            partes = [p.strip() for p in linea.split("|")]
        else:
            # Sin separador válido → saltar
            continue

        if len(partes) < 4:
            continue

        # Si hay más de 8 columnas, re-unir el excedente al motivo (último campo)
        if len(partes) > len(CAMPOS):
            motivo_extendido = " ".join(partes[len(CAMPOS) - 1 :]).strip()
            partes = partes[: len(CAMPOS) - 1] + [motivo_extendido]

        fila_data: dict = {"fila": i + 1}
        # Campo legacy 'servicio' (col 7) no existe en downstream — se mapea
        # al 'descripcion' adicional cuando hay 8 columnas.
        for idx, campo in enumerate(CAMPOS):
            fila_data[campo] = partes[idx] if idx < len(partes) else ""

        # Si hay columna 'servicio' (col 7) y la 'descripcion' está vacía,
        # promover servicio a descripcion. Si ambas tienen valor, concatenar.
        if fila_data.get("servicio"):
            if fila_data.get("descripcion") and fila_data["descripcion"] != fila_data["servicio"]:
                fila_data["descripcion"] = f"{fila_data['descripcion']} — {fila_data['servicio']}"
            else:
                fila_data["descripcion"] = fila_data["servicio"]

        if fila_data["codigo"] and len(fila_data["codigo"]) >= 2:
            filas.append(fila_data)

    return filas


async def _procesar_fila_en_background(
    fila_data: dict, servicio_id: str, req_id: str, eps_formulario: str, lote_id=None
):
    """Procesa una fila individual en segundo plano.

    Si `eps_formulario` viene vacío o "AUTO", se usa la EPS detectada de la
    primera columna de la fila (razón social del Excel).

    Si `lote_id` viene seteado, actualiza LoteImportacionRecord al
    finalizar (incrementa procesadas, exitosas, fallidas; marca COMPLETO
    cuando procesadas == total_filas).
    """
    db = SessionLocal()
    fila_ok = False
    error_msg = None
    try:
        cfg = get_settings()
        service = GlosaService(
            groq_api_key=cfg.groq_api_key,
            anthropic_api_key=cfg.anthropic_api_key,
            primary_ai=cfg.primary_ai,
            anthropic_model=cfg.anthropic_model,
            groq_model=cfg.groq_model,
            gemini_api_key=cfg.gemini_api_key,
            gemini_model=cfg.gemini_model,
        )

        from app.models.schemas import GlosaInput

        contrato_repo = ContratoRepository(db)
        contratos = contrato_repo.como_dict()

        # Resolver EPS: formulario > detectada de la fila
        eps_formulario_limpio = (eps_formulario or "").strip().upper()
        usa_auto = (not eps_formulario_limpio) or eps_formulario_limpio == "AUTO"
        if usa_auto:
            eps_final = _normalizar_eps(fila_data.get("eps", "")) or "SIN EPS"
        else:
            eps_final = eps_formulario

        texto_glosa = f"{fila_data['codigo']} {fila_data['valor']} {fila_data['descripcion']} {fila_data['cups']} {fila_data['motivo']}"

        data = GlosaInput(
            eps=eps_final,
            etapa="RESPUESTA A GLOSA",
            tabla_excel=texto_glosa,
            numero_factura=fila_data.get("factura"),
            numero_radicado=servicio_id,
        )

        from app.services.tarifa_lookup_service import pre_lookup_tarifa

        info_tarifa_pre = pre_lookup_tarifa(
            db=db,
            cod_pref=fila_data.get("codigo", ""),
            eps=eps_final,
            tabla_excel=texto_glosa,
        )
        resultado = await service.analizar(data, "", contratos, info_tarifa=info_tarifa_pre)

        repo = GlosaRepository(db)
        # Campos adicionales para que el flujo "responder por factura"
        # los pueda listar con contexto (servicio, CUPS, concepto).
        concepto_excel = fila_data.get("motivo") or fila_data.get("descripcion") or ""
        kwargs_extra = {}
        if fila_data.get("descripcion"):
            kwargs_extra["servicio_descripcion"] = fila_data["descripcion"][:400]
        if concepto_excel:
            kwargs_extra["concepto_glosa"] = concepto_excel[:500]
        if fila_data.get("cups"):
            kwargs_extra["cups_servicio"] = fila_data["cups"][:20]
        # Texto glosa original para que el auditor pueda revisar
        kwargs_extra["texto_glosa_original"] = texto_glosa[:2000]

        # IM F2: si el lote tiene gestor_asignado_id, resolvemos el
        # email del usuario y lo pasamos como asignado_a_email para
        # que la glosa quede en su bandeja "Mis glosas" automaticamente.
        asignado_email = None
        if lote_id is not None:
            try:
                from app.models.db import LoteImportacionRecord, UsuarioRecord

                lote = (
                    db.query(LoteImportacionRecord)
                    .filter(LoteImportacionRecord.id == lote_id)
                    .first()
                )
                if lote and lote.gestor_asignado_id:
                    user = (
                        db.query(UsuarioRecord)
                        .filter(UsuarioRecord.id == lote.gestor_asignado_id)
                        .first()
                    )
                    if user and user.email:
                        asignado_email = user.email
                        kwargs_extra["asignado_a_email"] = asignado_email
            except Exception as _e_asg:
                logger.debug(f"[{req_id}] No se pudo resolver gestor del lote {lote_id}: {_e_asg}")

        try:
            repo.crear(
                eps=eps_final,
                paciente="N/A",
                codigo_glosa=resultado.codigo_glosa,
                # parse_valor_cop: "7.700,00" → 7700.0 (el patrón anterior
                # re.sub(r"[^\d]") inflaba 100× — auditoría jun-2026 P0 #1)
                valor_objetado=parse_valor_cop(fila_data.get("valor", "0")),
                valor_aceptado=0,
                etapa="RESPUESTA A GLOSA",
                estado="RESPONDIDA",
                dictamen=resultado.dictamen,
                dias_restantes=resultado.dias_restantes,
                modelo_ia=resultado.modelo_ia,
                score=resultado.score,
                numero_radicado=servicio_id,
                factura=fila_data.get("factura"),
                **kwargs_extra,
            )
        except TypeError:
            # Fallback si repo.crear no soporta los kwargs extra
            repo.crear(
                eps=eps_final,
                paciente="N/A",
                codigo_glosa=resultado.codigo_glosa,
                valor_objetado=parse_valor_cop(fila_data.get("valor", "0")),
                valor_aceptado=0,
                etapa="RESPUESTA A GLOSA",
                estado="RESPONDIDA",
                dictamen=resultado.dictamen,
                dias_restantes=resultado.dias_restantes,
                modelo_ia=resultado.modelo_ia,
                score=resultado.score,
                numero_radicado=servicio_id,
                factura=fila_data.get("factura"),
            )

        logger.info(f"[{req_id}] Fila {fila_data['fila']} procesada: {resultado.codigo_glosa}")
        fila_ok = True
    except Exception as e:
        logger.error(f"[{req_id}] Error procesando fila {fila_data['fila']}: {e}")
        error_msg = str(e)[:300]
    finally:
        # IM F1.3: actualizar el LoteImportacionRecord con el resultado
        # de esta fila — incrementos atómicos + marcar COMPLETO si es
        # la última. Sesión propia para no chocar con la principal.
        if lote_id is not None:
            try:
                from app.models.db import LoteImportacionRecord as _LIR
                from datetime import datetime as _dt, timezone as _tz
                import json as _json

                db_lote = SessionLocal()
                try:
                    lote = db_lote.query(_LIR).filter(_LIR.id == lote_id).first()
                    if lote:
                        lote.procesadas = (lote.procesadas or 0) + 1
                        if fila_ok:
                            lote.exitosas = (lote.exitosas or 0) + 1
                        else:
                            lote.fallidas = (lote.fallidas or 0) + 1
                            # Acumular error en el campo `errores` (cap a 100)
                            try:
                                actuales = _json.loads(lote.errores) if lote.errores else []
                            except Exception:
                                actuales = []
                            if len(actuales) < 100:
                                # F2: guardamos tambien la fila_data completa
                                # para soportar retry desde la UI sin que el
                                # usuario tenga que volver a pegar el Excel.
                                actuales.append(
                                    {
                                        "fila": fila_data.get("fila"),
                                        "error": error_msg or "Error desconocido",
                                        "fila_data": {
                                            k: (str(v)[:500] if v else "")
                                            for k, v in fila_data.items()
                                            if k != "fila"
                                        },
                                    }
                                )
                                lote.errores = _json.dumps(actuales, ensure_ascii=False)
                        # Si se procesaron todas, marcar COMPLETO
                        if (lote.procesadas or 0) >= (lote.total_filas or 0):
                            lote.estado = "COMPLETO"
                            lote.terminado_en = _dt.now(_tz.utc)
                        db_lote.commit()
                finally:
                    db_lote.close()
            except Exception as _e:
                logger.debug(f"[{req_id}] No se pudo actualizar lote {lote_id}: {_e}")
        db.close()


@router.post("/importar-masiva/preview")
async def preview_importar_masiva(
    request: ImportacionMasivaRequest,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """IM Fase 1.2: previsualiza el lote SIN procesar.

    Devuelve:
      - total_filas (parseadas vs descartadas)
      - filas_validas / filas_invalidas con detalle por fila
      - eps_detectadas {eps: count}
      - facturas_unicas (count)
      - posibles_duplicados con BD existente
      - costo_estimado_usd (segun flags multi-agent/tools)
      - umbral_alerta_usd (configurable via env var)
      - bloquear: bool — true si costo > umbral, requiere confirm

    El frontend usa esto para mostrar un modal con el summary antes de
    disparar /importar-masiva. Si el lote es muy chico (< 5 filas) o
    barato (< $1), se permite proceder sin confirmacion explicita.
    """
    import os
    from app.models.db import GlosaRecord

    eps_formulario = (request.eps or "").strip()
    modo_auto = not eps_formulario or eps_formulario.upper() == "AUTO"

    # Parseo crudo
    lineas_raw = (request.texto_excel or "").strip().split("\n")
    total_lineas = len([l for l in lineas_raw if l.strip()])

    filas_validas = _parsear_filas_excel(request.texto_excel)
    filas_invalidas: list[dict] = []

    # Re-parsear para detectar lineas que se descartaron
    for i, linea in enumerate(lineas_raw):
        linea_clean = linea.strip()
        if not linea_clean:
            continue
        if "\t" in linea_clean:
            partes = linea_clean.split("\t")
        elif "|" in linea_clean:
            partes = linea_clean.split("|")
        else:
            filas_invalidas.append(
                {
                    "fila": i + 1,
                    "razon": "Sin separador (TAB o pipe)",
                    "preview": linea_clean[:80],
                }
            )
            continue
        if len(partes) < 4:
            filas_invalidas.append(
                {
                    "fila": i + 1,
                    "razon": f"Solo {len(partes)} columnas (minimo 4)",
                    "preview": linea_clean[:80],
                }
            )
            continue
        codigo_field = (partes[3] if len(partes) > 3 else "").strip()
        if not codigo_field or len(codigo_field) < 2:
            filas_invalidas.append(
                {
                    "fila": i + 1,
                    "razon": "Codigo de glosa vacio o muy corto",
                    "preview": linea_clean[:80],
                }
            )

    # EPS detectadas
    eps_detectadas: dict[str, int] = {}
    facturas_unicas: set[str] = set()
    for f in filas_validas:
        clave = _normalizar_eps(f.get("eps", "")) if modo_auto else eps_formulario
        eps_detectadas[clave or "SIN EPS"] = eps_detectadas.get(clave or "SIN EPS", 0) + 1
        if f.get("factura"):
            facturas_unicas.add(f["factura"].strip())

    # Detectar posibles duplicados con BD existente
    posibles_duplicados: list[dict] = []
    if facturas_unicas:
        existentes = (
            db.query(GlosaRecord.factura, GlosaRecord.codigo_glosa, GlosaRecord.id)
            .filter(GlosaRecord.factura.in_(facturas_unicas))
            .limit(500)
            .all()
        )
        existentes_set = {(e[0] or "", e[1] or "") for e in existentes}
        for f in filas_validas:
            par = (f.get("factura") or "", f.get("codigo") or "")
            if par in existentes_set:
                posibles_duplicados.append(
                    {
                        "factura": par[0],
                        "codigo": par[1],
                        "razon": "Ya existe en BD",
                    }
                )

    # Estimacion de costo (USD) basada en flags activos
    n_validas = len(filas_validas)
    costo_base_por_glosa = 0.05  # Sonnet 4.5 promedio
    multiplicador = 1.0
    multi_agent_on = os.getenv("MULTI_AGENT_HABILITADO", "0").strip() in ("1", "true", "yes")
    tool_use_on = os.getenv("TOOL_USE_HABILITADO", "0").strip() in ("1", "true", "yes")
    if multi_agent_on:
        multiplicador += 0.5  # +50% por Auditor agent extra
    if tool_use_on:
        multiplicador += 0.3  # +30% por turnos multi-tool
    costo_estimado_usd = round(n_validas * costo_base_por_glosa * multiplicador, 2)

    umbral_usd = float(os.getenv("IMPORT_MASIVO_COST_LIMIT_USD", "20"))
    bloquear = costo_estimado_usd > umbral_usd

    return {
        "total_lineas_pegadas": total_lineas,
        "filas_validas": n_validas,
        "filas_invalidas": filas_invalidas[:50],  # cap para no inflar response
        "filas_invalidas_total": len(filas_invalidas),
        "eps_detectadas": eps_detectadas,
        "facturas_unicas": len(facturas_unicas),
        "posibles_duplicados": posibles_duplicados[:30],
        "posibles_duplicados_total": len(posibles_duplicados),
        "costo_estimado_usd": costo_estimado_usd,
        "costo_por_glosa_usd": round(costo_base_por_glosa * multiplicador, 4),
        "flags_activos": {
            "multi_agent": multi_agent_on,
            "tool_use": tool_use_on,
        },
        "umbral_alerta_usd": umbral_usd,
        "bloquear": bloquear,
        "ok_para_procesar": n_validas > 0 and not bloquear,
    }


@router.post("/importar-masiva")
async def importar_glosas_masiva(
    request: ImportacionMasivaRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """
    Importa glosas masivamente desde texto pegado de Excel.

    - Si se envía `eps` = nombre específico, todas las filas usan esa EPS.
    - Si se envía `eps` = null, "" o "AUTO", la EPS se detecta de la primera
      columna de cada fila (razón social) y se normaliza a la clave canónica.

    Recibe: texto_excel (con tabs), fechas opcionales, eps opcional.
    Procesa en segundo plano y retorna el ID del lote para seguimiento.
    """
    req_id = uuid.uuid4().hex[:8]
    eps_formulario = (request.eps or "").strip()
    modo_auto = not eps_formulario or eps_formulario.upper() == "AUTO"
    logger.info(
        f"[{req_id}] Importación masiva iniciada | modo={'AUTO' if modo_auto else eps_formulario}"
    )

    filas = _parsear_filas_excel(request.texto_excel)

    if not filas:
        raise HTTPException(status_code=400, detail="No se detectaron filas válidas en el texto")

    # Detectar EPS/facturas únicas para dar feedback inmediato en el response
    eps_detectadas: dict[str, int] = {}
    facturas_detectadas: set[str] = set()
    for f in filas:
        clave = _normalizar_eps(f.get("eps", "")) if modo_auto else eps_formulario
        eps_detectadas[clave or "SIN EPS"] = eps_detectadas.get(clave or "SIN EPS", 0) + 1
        if f.get("factura"):
            facturas_detectadas.add(f["factura"])

    servicio_id = f"BATCH-{req_id}"

    # IM F1.3: persistir el lote en BD para tracking + historial
    import json as _json
    import hashlib as _hashlib
    from app.models.db import LoteImportacionRecord

    texto_hash = _hashlib.sha256((request.texto_excel or "").encode("utf-8")).hexdigest()
    gestor_id = getattr(request, "gestor_asignado_id", None)
    try:
        lote = LoteImportacionRecord(
            batch_id=servicio_id,
            usuario_email=current_user.email,
            total_filas=len(filas),
            procesadas=0,
            exitosas=0,
            fallidas=0,
            estado="PROCESANDO",
            eps_detectadas=_json.dumps(eps_detectadas, ensure_ascii=False),
            texto_hash=texto_hash,
            gestor_asignado_id=int(gestor_id) if gestor_id else None,
        )
        db.add(lote)
        db.commit()
        db.refresh(lote)
        lote_id = lote.id
    except Exception as e:
        logger.warning(f"[{req_id}] No se pudo crear LoteImportacionRecord: {e}")
        lote_id = None

    for fila_data in filas:
        background_tasks.add_task(
            _procesar_fila_en_background,
            fila_data,
            servicio_id,
            req_id,
            eps_formulario if not modo_auto else "AUTO",
            lote_id,
        )

    logger.info(
        f"[{req_id}] {len(filas)} filas enviadas | batch_id={servicio_id} | "
        f"lote_id={lote_id} | EPS: {dict(eps_detectadas)} | facturas: {len(facturas_detectadas)}"
    )

    return {
        "message": f"{len(filas)} glosas procesándose en segundo plano",
        "batch_id": servicio_id,
        "lote_id": lote_id,
        "total_filas": len(filas),
        "eps": eps_formulario if not modo_auto else "AUTO",
        "eps_detectadas": eps_detectadas,
        "facturas_detectadas": sorted(facturas_detectadas),
        "estado": "PROCESANDO",
    }


@router.get("/importar-masiva/lotes")
def listar_lotes_importacion(
    page: int = 1,
    per_page: int = 20,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """IM F1.3: historial paginado de lotes anteriores.

    Filtra automaticamente por usuario actual a menos que sea admin
    (en cuyo caso ve todos los lotes).
    """
    from app.models.db import LoteImportacionRecord
    from sqlalchemy import desc as _desc

    if per_page < 1 or per_page > 100:
        per_page = 20
    if page < 1:
        page = 1

    q = db.query(LoteImportacionRecord).order_by(_desc(LoteImportacionRecord.iniciado_en))
    rol = (getattr(current_user, "rol", "") or "").upper()
    if rol not in ("SUPER_ADMIN", "COORDINADOR"):
        q = q.filter(LoteImportacionRecord.usuario_email == current_user.email)

    total = q.count()
    items = q.offset((page - 1) * per_page).limit(per_page).all()

    import json as _json

    return {
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page,
        "items": [
            {
                "id": l.id,
                "batch_id": l.batch_id,
                "usuario_email": l.usuario_email,
                "total_filas": l.total_filas,
                "procesadas": l.procesadas,
                "exitosas": l.exitosas,
                "fallidas": l.fallidas,
                "estado": l.estado,
                "iniciado_en": l.iniciado_en.isoformat() if l.iniciado_en else None,
                "terminado_en": l.terminado_en.isoformat() if l.terminado_en else None,
                "eps_detectadas": (_json.loads(l.eps_detectadas) if l.eps_detectadas else {}),
                "costo_estimado_usd": l.costo_estimado_usd or 0,
                "costo_real_usd": l.costo_real_usd or 0,
                "gestor_asignado_id": l.gestor_asignado_id,
            }
            for l in items
        ],
    }


@router.get("/importar-masiva/lote/{lote_id}/status")
def status_lote(
    lote_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """IM F1.3: estado actual de un lote — usado por el polling del
    frontend cada 3 segundos para actualizar la barra de progreso."""
    from app.models.db import LoteImportacionRecord
    import json as _json

    l = db.query(LoteImportacionRecord).filter(LoteImportacionRecord.id == lote_id).first()
    if not l:
        raise HTTPException(404, "Lote no encontrado")

    # Permisos: el dueño del lote o admin
    rol = (getattr(current_user, "rol", "") or "").upper()
    if l.usuario_email != current_user.email and rol not in ("SUPER_ADMIN", "COORDINADOR"):
        raise HTTPException(403, "Sin permisos para ver este lote")

    return {
        "id": l.id,
        "batch_id": l.batch_id,
        "estado": l.estado,
        "total_filas": l.total_filas,
        "procesadas": l.procesadas,
        "exitosas": l.exitosas,
        "fallidas": l.fallidas,
        "porcentaje": round(100.0 * (l.procesadas or 0) / max(l.total_filas, 1), 1),
        "iniciado_en": l.iniciado_en.isoformat() if l.iniciado_en else None,
        "terminado_en": l.terminado_en.isoformat() if l.terminado_en else None,
        "errores": _json.loads(l.errores) if l.errores else [],
        "glosas_creadas_ids": (_json.loads(l.glosas_creadas_ids) if l.glosas_creadas_ids else []),
        "costo_real_usd": l.costo_real_usd or 0,
    }


@router.post("/importar-masiva/lote/{lote_id}/cancelar")
def cancelar_lote(
    lote_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """IM F1.4: marca un lote como CANCELADO. Las filas que YA estaban
    en el queue de background_tasks van a seguir procesandose (FastAPI
    no permite cancelar tasks ya disparadas), pero el frontend deja de
    pollear y la UI marca el lote como cancelado.

    En la practica, si el usuario cancela un lote de 50 filas a la
    fila #20, las 30 restantes igual se ejecutan. Lo que cancela es la
    visibilidad y deja el registro marcado para auditoria.
    """
    from app.models.db import LoteImportacionRecord
    from datetime import datetime as _dt, timezone as _tz

    l = db.query(LoteImportacionRecord).filter(LoteImportacionRecord.id == lote_id).first()
    if not l:
        raise HTTPException(404, "Lote no encontrado")
    rol = (getattr(current_user, "rol", "") or "").upper()
    if l.usuario_email != current_user.email and rol not in ("SUPER_ADMIN", "COORDINADOR"):
        raise HTTPException(403, "Sin permisos")
    if l.estado != "PROCESANDO":
        raise HTTPException(400, f"Lote ya esta en estado {l.estado}, no se puede cancelar")
    l.estado = "CANCELADO"
    l.terminado_en = _dt.now(_tz.utc)
    db.commit()
    logger.info(f"[LOTE-CANCEL] lote_id={lote_id} cancelado por {current_user.email}")
    return {"ok": True, "lote_id": lote_id, "estado": "CANCELADO"}


@router.get("/importar-masiva/lote/{lote_id}/exportar")
def exportar_lote_csv(
    lote_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """IM F1.4: exporta a CSV los detalles del lote (params + errores +
    glosas creadas) para forensia / reporte SuperSalud."""
    from app.models.db import LoteImportacionRecord, GlosaRecord
    from fastapi.responses import StreamingResponse
    import csv
    import io
    import json as _json

    l = db.query(LoteImportacionRecord).filter(LoteImportacionRecord.id == lote_id).first()
    if not l:
        raise HTTPException(404, "Lote no encontrado")
    rol = (getattr(current_user, "rol", "") or "").upper()
    if l.usuario_email != current_user.email and rol not in ("SUPER_ADMIN", "COORDINADOR"):
        raise HTTPException(403, "Sin permisos")

    def _generar():
        buf = io.StringIO()
        w = csv.writer(buf)
        # Header section
        w.writerow(["LOTE IMPORTACION MASIVA"])
        w.writerow(["Batch ID", l.batch_id])
        w.writerow(["Usuario", l.usuario_email])
        w.writerow(["Iniciado", l.iniciado_en.isoformat() if l.iniciado_en else ""])
        w.writerow(["Terminado", l.terminado_en.isoformat() if l.terminado_en else ""])
        w.writerow(["Estado", l.estado])
        w.writerow(["Total filas", l.total_filas])
        w.writerow(["Exitosas", l.exitosas])
        w.writerow(["Fallidas", l.fallidas])
        w.writerow([])
        # EPS detectadas
        eps = _json.loads(l.eps_detectadas) if l.eps_detectadas else {}
        if eps:
            w.writerow(["EPS DETECTADAS"])
            w.writerow(["EPS", "Filas"])
            for k, v in eps.items():
                w.writerow([k, v])
            w.writerow([])
        # Errores por fila
        errores = _json.loads(l.errores) if l.errores else []
        if errores:
            w.writerow(["ERRORES POR FILA"])
            w.writerow(["Fila", "Error"])
            for e in errores:
                w.writerow([e.get("fila", ""), e.get("error", "")])
            w.writerow([])
        # Glosas creadas (busqueda por numero_radicado = batch_id)
        glosas = (
            db.query(GlosaRecord)
            .filter(GlosaRecord.numero_radicado == l.batch_id)
            .order_by(GlosaRecord.id)
            .all()
        )
        if glosas:
            w.writerow(["GLOSAS CREADAS"])
            w.writerow(["ID", "EPS", "Factura", "Codigo", "Valor objetado", "Estado"])
            for g in glosas:
                w.writerow(
                    [g.id, g.eps, g.factura, g.codigo_glosa, g.valor_objetado or 0, g.estado or ""]
                )
        yield buf.getvalue()

    fname = f"lote-{l.batch_id}-{lote_id}.csv"
    return StreamingResponse(
        _generar(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/importar-masiva/plantilla.csv")
def descargar_plantilla_masiva():
    """IM F2: descarga plantilla CSV con headers correctos y 3 filas
    ejemplo. El usuario abre con Excel, llena las filas, exporta como
    CSV o pega directamente desde Excel en el textarea.

    Endpoint PÚBLICO: el archivo no tiene PII, solo headers de campos
    documentados + 3 filas de ejemplo. El frontend usa <a href="..."> que
    no manda el JWT, así que requerir auth aquí lo rompía. La operación
    sensible (POST /importar-masiva) sí requiere auth.
    """
    from fastapi.responses import StreamingResponse
    import csv
    import io

    def _generar():
        buf = io.StringIO()
        w = csv.writer(buf, delimiter="\t")  # TAB para compat con paste-from-Excel
        # Header
        w.writerow(
            [
                "ENTIDAD",
                "FACTURA",
                "VALOR",
                "CODIGO",
                "CONCEPTO",
                "CUPS",
                "SERVICIO",
                "MOTIVO",
            ]
        )
        # 3 filas ejemplo cubriendo casos comunes
        w.writerow(
            [
                "FAMISANAR EPS",
                "HUS0000123456",
                "$ 150.000",
                "TA0801",
                "Tarifa diferente a la pactada en contrato",
                "890301",
                "Consulta de control",
                "El valor facturado no corresponde a la tarifa SOAT pactada",
            ]
        )
        w.writerow(
            [
                "NUEVA EPS",
                "HUS0000123457",
                "$ 75.000",
                "SO0101",
                "Soporte ilegible",
                "",
                "",
                "Historia clínica con tinta corrida en página 3",
            ]
        )
        w.writerow(
            [
                "COMPENSAR",
                "HUS0000123458",
                "$ 200.000",
                "FA0601",
                "Cargos no facturables",
                "FMQ0114",
                "Catéter intravenoso",
                "Cantidad facturada no coincide con registros de uso",
            ]
        )
        yield buf.getvalue()

    return StreamingResponse(
        _generar(),
        media_type="text/tab-separated-values",
        headers={
            "Content-Disposition": 'attachment; filename="plantilla-importacion-masiva.tsv"',
        },
    )


@router.post("/importar-masiva/lote/{lote_id}/retry-fila")
async def retry_fila_lote(
    lote_id: int,
    fila_payload: dict,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """IM F2: re-ejecuta una fila específica que fallo en un lote
    anterior. El payload debe traer las 8 campos: eps, factura, valor,
    codigo, descripcion, cups, servicio, motivo.

    Útil cuando el dictamen falló por un error transitorio (Anthropic
    overloaded, timeout) y el usuario quiere reintentar sin re-importar
    todo el lote.
    """
    from app.models.db import LoteImportacionRecord
    import json as _json

    l = db.query(LoteImportacionRecord).filter(LoteImportacionRecord.id == lote_id).first()
    if not l:
        raise HTTPException(404, "Lote no encontrado")
    rol = (getattr(current_user, "rol", "") or "").upper()
    if l.usuario_email != current_user.email and rol not in ("SUPER_ADMIN", "COORDINADOR"):
        raise HTTPException(403, "Sin permisos")

    # Validar minimos
    if not fila_payload.get("codigo") or not fila_payload.get("eps"):
        raise HTTPException(400, "Faltan campos minimos: codigo, eps")

    # Asegurar que tenga 'fila' para tracking
    if "fila" not in fila_payload:
        fila_payload["fila"] = 999  # sentinel para retry

    # Disparar background task con el lote_id existente
    req_id = f"retry-{l.batch_id}"
    background_tasks.add_task(
        _procesar_fila_en_background,
        fila_payload,
        l.batch_id,
        req_id,
        l.usuario_email,  # mantener "AUTO" o eps
        lote_id,
    )

    # Sumar 1 al total_filas y resetear estado a PROCESANDO si estaba COMPLETO
    l.total_filas = (l.total_filas or 0) + 1
    if l.estado in ("COMPLETO", "ERROR"):
        l.estado = "PROCESANDO"
        l.terminado_en = None
        # Limpiar el error correspondiente a esta fila si existe
        try:
            errores = _json.loads(l.errores) if l.errores else []
            errores = [e for e in errores if e.get("fila") != fila_payload.get("fila")]
            l.errores = _json.dumps(errores, ensure_ascii=False) if errores else None
        except Exception:
            pass
    db.commit()

    logger.info(f"[RETRY] lote_id={lote_id} fila={fila_payload.get('fila')} re-procesando")
    return {
        "ok": True,
        "lote_id": lote_id,
        "fila": fila_payload.get("fila"),
        "estado": "EN_PROCESO",
    }


@router.post("/importar-masiva/dedupe-check")
def check_lote_duplicado(
    request: ImportacionMasivaRequest,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """IM F1.4: detecta si el texto pegado coincide con un lote
    procesado anteriormente (texto_hash matching). Permite advertir
    al usuario antes de procesar dos veces el mismo Excel."""
    import hashlib as _hashlib
    from app.models.db import LoteImportacionRecord

    if not request.texto_excel:
        return {"duplicado": False}
    h = _hashlib.sha256(request.texto_excel.encode("utf-8")).hexdigest()
    lote_existente = (
        db.query(LoteImportacionRecord)
        .filter(LoteImportacionRecord.texto_hash == h)
        .order_by(LoteImportacionRecord.iniciado_en.desc())
        .first()
    )
    if not lote_existente:
        return {"duplicado": False}
    return {
        "duplicado": True,
        "lote_id": lote_existente.id,
        "batch_id": lote_existente.batch_id,
        "usuario": lote_existente.usuario_email,
        "iniciado_en": (
            lote_existente.iniciado_en.isoformat() if lote_existente.iniciado_en else None
        ),
        "estado": lote_existente.estado,
        "total_filas": lote_existente.total_filas,
    }


def _recepcion_base_dir() -> str:
    import os as _os

    d = _os.path.join(_os.getenv("SOPORTES_ROOT", "/data"), "recepcion")
    _os.makedirs(d, exist_ok=True)
    return d


def _guardar_resumen_json(rec_id: int, data: dict) -> None:
    """Persiste el resumen de la importación junto al Excel en disco para
    que el endpoint de estado / la UI lo muestren cuando termine el
    procesamiento en background."""
    import os as _os
    import json as _json

    try:
        ruta = _os.path.join(_recepcion_base_dir(), f"{rec_id}.resumen.json")
        with open(ruta, "w", encoding="utf-8") as fh:
            _json.dump(data, fh, ensure_ascii=False, default=str)
    except Exception as e:
        logger.warning(f"[recepcion bg] no se pudo guardar resumen json {rec_id}: {e}")


async def _procesar_recepcion_bg(
    rec_id: int,
    contenido: bytes,
    usuario_email: str,
    usuario_rol: str,
    usuario_id,
    req_id: str,
) -> None:
    """Procesa el Excel de recepción FUERA del request (BackgroundTasks).

    El parseo de hojas CONCEPTOS del DGH puede tardar minutos; hacerlo en
    el request bloqueaba el único worker y tumbaba la app (incidente
    2026-05-19). Acá:
      1. procesar_excel corre en un thread (no bloquea el event loop).
      2. Se actualiza ImportacionRecepcionRecord (estado, totales, ids).
      3. Se guarda el resumen en disco para que la UI lo muestre al hacer
         polling de /importar-recepcion/{id}/status.
      4. Se dispara el lote IA + envío del Excel-respuesta y el broadcast.
    """
    import asyncio as _asyncio
    import json as _json

    from app.database import SessionLocal
    from app.services.recepcion_service import RecepcionService
    from app.services.email_service import enviar_resumen_importacion_recepcion
    from app.repositories.audit_repository import AuditRepository
    from app.models.db import ImportacionRecepcionRecord

    db = SessionLocal()
    try:
        servicio = RecepcionService(db)
        resumen = await _asyncio.to_thread(servicio.procesar_excel, contenido)

        logger.info(
            f"[{req_id}] Importación recepción (bg) por {usuario_email} | "
            f"total={resumen.total} nuevas={resumen.creadas} "
            f"actualizadas={resumen.actualizadas} "
            f"ratificadas={resumen.ratificadas} "
            f"extemporaneas={resumen.extemporaneas}"
        )

        try:
            AuditRepository(db).registrar(
                usuario_email=usuario_email,
                usuario_rol=usuario_rol,
                accion="IMPORTAR_RECEPCION",
                tabla="historial",
                detalle=(
                    f"total={resumen.total} nuevas={resumen.creadas} "
                    f"actualizadas={resumen.actualizadas} "
                    f"ratificadas={resumen.ratificadas} "
                    f"extemporaneas={resumen.extemporaneas}"
                ),
            )
        except Exception as e:
            logger.warning(f"[{req_id}] audit recepción falló: {e}")

        glosas_para_auto = list(getattr(resumen, "glosas_ids_para_auto_responder", []) or [])
        glosas_todas = list(getattr(resumen, "glosas_ids_todas", []) or [])

        resumen_dict = resumen.to_dict()
        resumen_dict["recepcion_import_id"] = rec_id
        resumen_dict["auto_respuesta_lanzada"] = bool(glosas_para_auto)
        resumen_dict["excel_respuesta_programado"] = bool(glosas_todas)
        resumen_dict["glosas_en_auto_proceso"] = len(glosas_para_auto)

        rec = (
            db.query(ImportacionRecepcionRecord)
            .filter(ImportacionRecepcionRecord.id == rec_id)
            .first()
        )
        if rec:
            rec.total_glosas = resumen.total + resumen.duplicadas
            rec.glosa_ids = _json.dumps(glosas_todas)
            # PROCESANDO → LISTO; el envío del Excel luego lo pasa a
            # ENVIADO/PARCIAL/SIN_DESTINATARIOS (no pisar SIN_ARCHIVO).
            if rec.estado == "PROCESANDO":
                rec.estado = "LISTO"
            db.commit()

        # El resumen ya es mostrable al usuario (procesar_excel terminó).
        _guardar_resumen_json(rec_id, resumen_dict)

        # Lote IA + envío del Excel-respuesta a cada gestor. Secuencial
        # dentro de este background (mantiene el event loop libre porque
        # procesar_lote_y_enviar_excel ya es async/awaitable).
        if glosas_todas:
            try:
                from app.services.auto_responder_service import (
                    procesar_lote_y_enviar_excel,
                )

                await procesar_lote_y_enviar_excel(
                    glosas_todas,
                    contenido,
                    resumen_dict,
                    ids_ia=glosas_para_auto,
                    rec_id=rec_id,
                )
            except Exception as e:
                logger.error(
                    f"[{req_id}] lote/Excel recepción falló: {e}",
                    exc_info=True,
                )

        # Broadcast resumen (best-effort).
        try:
            enviados = await enviar_resumen_importacion_recepcion(
                resumen_dict,
                db=db,
            )
            resumen_dict["correos_enviados"] = enviados
        except Exception as e:
            logger.error(f"[{req_id}] Error enviando correo resumen: {e}")
            resumen_dict["correos_enviados"] = 0
            resumen_dict["email_error"] = str(e)
        _guardar_resumen_json(rec_id, resumen_dict)

        try:
            from app.services.posthog_service import capture

            capture(
                event="recepcion_importada",
                distinct_id=str(usuario_id if usuario_id is not None else "anonimo"),
                properties={
                    "total_glosas": resumen_dict.get("total", 0),
                    "nuevas": resumen_dict.get("creadas", 0),
                    "actualizadas": resumen_dict.get("actualizadas", 0),
                    "duplicadas": resumen_dict.get("duplicadas", 0),
                    "ratificadas": resumen_dict.get("ratificadas", 0),
                    "extemporaneas": resumen_dict.get("extemporaneas", 0),
                    "gestores_afectados": len(resumen_dict.get("por_gestor", {})),
                    "rojo": resumen_dict.get("semaforo", {}).get("ROJO", 0),
                    "negro": resumen_dict.get("semaforo", {}).get("NEGRO", 0),
                    "auto_respuesta_lanzada": resumen_dict.get("auto_respuesta_lanzada", False),
                },
            )
        except Exception:
            pass
    except Exception as e:
        logger.error(
            f"[{req_id}] Importación recepción (bg) falló: {type(e).__name__}: {e}",
            exc_info=True,
        )
        try:
            rec = (
                db.query(ImportacionRecepcionRecord)
                .filter(ImportacionRecepcionRecord.id == rec_id)
                .first()
            )
            if rec and rec.estado == "PROCESANDO":
                rec.estado = "ERROR"
                db.commit()
        except Exception:
            pass
        _guardar_resumen_json(
            rec_id,
            {"estado": "ERROR", "error": str(e)[:300], "total": 0},
        )
    finally:
        db.close()


@router.post("/importar-recepcion")
async def importar_recepcion(
    background_tasks: BackgroundTasks,
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Sube el Excel de recepción y lo procesa EN SEGUNDO PLANO.

    Devuelve de inmediato el id de la importación; el cliente hace polling
    a /importar-recepcion/{id}/status para ver el resultado. Esto evita
    que el parseo (minutos para hojas CONCEPTOS del DGH) bloquee el worker
    y tumbe la app (incidente 2026-05-19).
    """
    req_id = set_request_id()
    contenido = await archivo.read()
    if not contenido:
        raise HTTPException(status_code=400, detail="Archivo vacío")
    if len(contenido) > 15_000_000:
        raise HTTPException(status_code=413, detail="Archivo demasiado grande (>15 MB)")

    import os as _os
    from app.models.db import ImportacionRecepcionRecord

    rec_id = None
    try:
        rec = ImportacionRecepcionRecord(
            usuario_email=current_user.email,
            archivo_nombre=(archivo.filename or "recepcion.xlsx")[:300],
            total_glosas=0,
            glosa_ids="[]",
            estado="PROCESANDO",
        )
        db.add(rec)
        db.commit()
        db.refresh(rec)
        rec_id = rec.id

        ruta = _os.path.join(_recepcion_base_dir(), f"{rec.id}.xlsx")
        with open(ruta, "wb") as fh:
            fh.write(contenido)
        rec.ruta_original = ruta
        db.commit()

        # Prune: conservar solo las últimas 30 importaciones en disco.
        viejas = (
            db.query(ImportacionRecepcionRecord)
            .order_by(ImportacionRecepcionRecord.id.desc())
            .offset(30)
            .all()
        )
        for v in viejas:
            if v.ruta_original and _os.path.exists(v.ruta_original):
                try:
                    _os.remove(v.ruta_original)
                    rj = _os.path.join(
                        _recepcion_base_dir(),
                        f"{v.id}.resumen.json",
                    )
                    if _os.path.exists(rj):
                        _os.remove(rj)
                    v.ruta_original = None
                    v.estado = "SIN_ARCHIVO"
                except OSError:
                    pass
        if viejas:
            db.commit()
    except Exception as e:
        logger.error(
            f"[{req_id}] No se pudo iniciar la importación de recepción: {e}",
            exc_info=True,
        )
        raise HTTPException(
            status_code=500,
            detail="No se pudo iniciar la importación. Reintentá.",
        )

    background_tasks.add_task(
        _procesar_recepcion_bg,
        rec_id,
        contenido,
        current_user.email,
        current_user.rol,
        getattr(current_user, "id", None),
        req_id,
    )
    logger.info(
        f"[{req_id}] Importación recepción id={rec_id} encolada "
        f"(archivo {len(contenido)} bytes) — procesando en background"
    )

    return {
        "recepcion_import_id": rec_id,
        "estado": "PROCESANDO",
        "procesando": True,
        "archivo_nombre": (archivo.filename or "recepcion.xlsx"),
        "mensaje": (
            "Importación recibida. El Excel se está procesando en segundo "
            "plano (puede tardar varios minutos en archivos grandes). "
            "Podés esperar acá o cerrar — el resultado queda en "
            "'Importaciones recientes'."
        ),
    }


@router.get("/importar-recepcion/historial")
def historial_importaciones_recepcion(
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Últimas importaciones de recepción, para listar y ofrecer la
    descarga del Excel-respuesta anotado de cada una."""
    import json as _json
    from app.models.db import ImportacionRecepcionRecord, GlosaRecord

    regs = (
        db.query(ImportacionRecepcionRecord)
        .order_by(ImportacionRecepcionRecord.id.desc())
        .limit(30)
        .all()
    )
    out = []
    for r in regs:
        try:
            ids = _json.loads(r.glosa_ids or "[]")
        except (ValueError, TypeError):
            ids = []
        listos = 0
        if ids:
            listos = (
                db.query(GlosaRecord.id)
                .filter(
                    GlosaRecord.id.in_(ids),
                    GlosaRecord.dictamen.isnot(None),
                    GlosaRecord.dictamen != "",
                )
                .count()
            )
        pendientes = max(0, len(ids) - listos)
        out.append(
            {
                "id": r.id,
                "creado_en": r.creado_en.isoformat() if r.creado_en else None,
                "usuario_email": r.usuario_email,
                "archivo_nombre": r.archivo_nombre,
                "total_glosas": r.total_glosas,
                "estado": r.estado,
                "descargable": bool(r.ruta_original),
                "dictamenes_listos": listos,
                "dictamenes_pendientes": pendientes,
            }
        )
    return {"importaciones": out}


@router.get("/importar-recepcion/{rec_id}/status")
def estado_importacion_recepcion(
    rec_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Estado de una importación en background. La UI hace polling acá.

    `resumen` viene null mientras procesar_excel no termina; cuando el
    background lo guarda en disco, se devuelve para que la UI lo muestre.
    """
    import os as _os
    import json as _json
    from app.models.db import ImportacionRecepcionRecord

    rec = (
        db.query(ImportacionRecepcionRecord).filter(ImportacionRecepcionRecord.id == rec_id).first()
    )
    if not rec:
        raise HTTPException(404, "Importación no encontrada")

    resumen = None
    rp = _os.path.join(_recepcion_base_dir(), f"{rec_id}.resumen.json")
    if _os.path.exists(rp):
        try:
            with open(rp, "r", encoding="utf-8") as fh:
                resumen = _json.load(fh)
        except Exception:
            resumen = None

    # "Terminado para mostrar" = procesar_excel terminó (hay resumen) o
    # falló (estado ERROR). El envío IA/correo sigue después y se refleja
    # en el badge de estado del historial.
    procesando = rec.estado == "PROCESANDO" and resumen is None
    return {
        "id": rec.id,
        "estado": rec.estado,
        "procesando": procesando,
        "descargable": bool(rec.ruta_original),
        "total_glosas": rec.total_glosas,
        "resumen": resumen,
    }


@router.get("/importar-recepcion/{rec_id}/excel-respuesta")
def descargar_excel_respuesta_recepcion(
    rec_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Regenera y descarga el Excel-respuesta anotado de una importación.

    Se regenera al vuelo desde el Excel original guardado + el estado
    actual de las glosas en BD, así incluye los dictámenes que la IA
    haya producido después de la importación.
    """
    import os as _os
    import json as _json

    from fastapi.responses import StreamingResponse

    from app.models.db import ImportacionRecepcionRecord
    from app.services.recepcion_excel_response import (
        construir_respuestas_por_clave,
        generar_excel_con_respuestas,
    )

    rec = (
        db.query(ImportacionRecepcionRecord).filter(ImportacionRecepcionRecord.id == rec_id).first()
    )
    if not rec:
        raise HTTPException(404, "Importación no encontrada")
    if not rec.ruta_original or not _os.path.exists(rec.ruta_original):
        raise HTTPException(
            410,
            "El Excel original de esta importación ya no está disponible "
            "(se conservan solo las últimas 30). Volvé a subir el archivo.",
        )

    try:
        with open(rec.ruta_original, "rb") as fh:
            contenido_original = fh.read()
    except OSError as e:
        raise HTTPException(500, f"No se pudo leer el archivo original: {e}")

    try:
        ids = _json.loads(rec.glosa_ids or "[]")
    except (ValueError, TypeError):
        ids = []

    respuestas = construir_respuestas_por_clave(db, ids)
    xlsx = generar_excel_con_respuestas(
        contenido_original,
        respuestas,
        gestor_destacar=None,
    )

    fecha = (rec.creado_en or datetime.now()).strftime("%Y-%m-%d")
    nombre = f"glosas_recepcion_{fecha}_respuesta.xlsx"
    return StreamingResponse(
        iter([xlsx]),
        media_type=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


@router.get("/batch/{batch_id}")
def obtener_estado_batch(
    batch_id: str,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Obtiene el estado de un lote de importación."""
    glosas_batch = db.query(GlosaRecord).filter(GlosaRecord.numero_radicado == batch_id).all()

    return {
        "batch_id": batch_id,
        "total": len(glosas_batch),
        "glosas": [
            {
                "id": g.id,
                "codigo_glosa": g.codigo_glosa,
                "valor_objetado": g.valor_objetado,
                "estado": g.estado,
                "creado_en": g.creado_en.isoformat() if g.creado_en else None,
            }
            for g in glosas_batch
        ],
    }


@router.get("/duplicados")
def listar_duplicados_factura(
    factura: str,
    eps: Optional[str] = None,
    limite: int = 5,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R58 P2: lista glosas previamente registradas con la misma factura.

    Útil para detección de duplicados antes de cargar una glosa nueva.
    Match exacto sobre numero_factura, opcional filtro por EPS.

    Query params:
      factura  número de factura a buscar (case-insensitive, trim)
      eps      EPS opcional para restringir
      limite   máximo de resultados (default 5)

    Respuesta:
      {
        "factura_consultada": "FE-2026-001",
        "total": 1,
        "duplicados": [
          {"id": 42, "eps": "FAMISANAR", "creado_en": "...",
           "estado": "RADICADA", "valor_objetado": 168563.0,
           "auditor_email": "x@hus.com"}
        ]
      }
    """
    from app.repositories.glosa_repository import buscar_duplicados_factura

    duplicados = buscar_duplicados_factura(
        db,
        numero_factura=factura,
        eps=eps,
        limite=limite,
    )
    return {
        "factura_consultada": factura,
        "eps_filtro": eps,
        "total": len(duplicados),
        "duplicados": [
            {
                "id": g.id,
                "eps": g.eps,
                "factura": g.factura,
                "codigo_glosa": g.codigo_glosa,
                "valor_objetado": float(g.valor_objetado or 0),
                "valor_aceptado": float(g.valor_aceptado or 0),
                "estado": g.estado,
                "auditor_email": g.auditor_email,
                "creado_en": g.creado_en.isoformat() if g.creado_en else None,
            }
            for g in duplicados
        ],
    }


@router.post("/{glosa_id}/reanalizar")
async def reanalizar_glosa(
    glosa_id: int,
    data: ReanalizarRequest,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_auditor_o_superior),
    _cupo_ia: None = Depends(_consumir_cupo_ia),
):
    """R60 P2: re-corre el análisis IA sobre una glosa existente.

    Útil cuando el gestor hizo primero 'auditoria_previa' y ahora quiere
    el dictamen de defensa, o cuando quiere cambiar el tono. NO duplica
    la fila — sobreescribe el dictamen de la glosa actual y guarda
    snapshot en versiones con accion='REANALIZAR'.

    Reusa los datos persistidos: eps, texto_glosa_original, etapa,
    factura, radicado.
    """
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    # R-UI 27-abr-2026: si falta texto_glosa_original (glosa importada
    # masivamente), construirlo on-the-fly con los datos disponibles.
    # Excepción: glosas legacy con EPS/etapa inválidas no se pueden reanalizar.
    if not (glosa.texto_glosa_original or "").strip() and len(glosa.eps or "") < 2:
        raise HTTPException(
            400,
            "Glosa legacy sin texto_glosa_original ni datos mínimos para reanalizar.",
        )
    texto_para_ia = (glosa.texto_glosa_original or "").strip()
    if len(texto_para_ia) < 30:
        partes = []
        if glosa.codigo_glosa:
            partes.append(glosa.codigo_glosa)
        if glosa.concepto_glosa:
            partes.append(glosa.concepto_glosa)
        if glosa.cups_servicio:
            partes.append(f"CUPS {glosa.cups_servicio}")
        if glosa.servicio_descripcion:
            partes.append(glosa.servicio_descripcion)
        if glosa.valor_objetado and float(glosa.valor_objetado) > 0:
            partes.append(f"Valor objetado: ${int(glosa.valor_objetado):,}".replace(",", "."))
        if glosa.observacion_eps:
            partes.append(glosa.observacion_eps)
        # Conceptos vinculados (importación masiva nueva con hojas I/R)
        try:
            from app.models.db import ConceptoGlosaRecord

            conceptos = (
                db.query(ConceptoGlosaRecord)
                .filter(ConceptoGlosaRecord.glosa_id == glosa.id)
                .limit(5)
                .all()
            )
            for c in conceptos:
                if c.codigo_glosa and c.codigo_glosa not in partes:
                    partes.append(c.codigo_glosa)
                if c.cups_codigo:
                    partes.append(f"CUPS {c.cups_codigo}")
                if c.cups_descripcion:
                    partes.append(c.cups_descripcion)
                if c.observacion_eps:
                    partes.append(c.observacion_eps)
                if c.valor_objetado and float(c.valor_objetado) > 0:
                    partes.append(f"Valor objetado: ${int(c.valor_objetado):,}".replace(",", "."))
        except Exception:
            pass
        texto_para_ia = " - ".join(p for p in partes if p and str(p).strip())
        # Salvaguarda final
        if len(texto_para_ia) < 30:
            texto_para_ia = (
                f"{glosa.codigo_glosa or 'GLOSA'} - "
                f"Glosa interpuesta por {glosa.eps or 'la entidad pagadora'}, "
                f"valor objetado ${int(glosa.valor_objetado or 0):,}".replace(",", ".")
                + ". Defender con argumentos generales aplicables al código."
            )

    # Construir GlosaInput a partir de los campos persistidos
    from app.models.schemas import GlosaInput

    try:
        glosa_input = GlosaInput(
            eps=glosa.eps or "",
            etapa=glosa.etapa or "RESPUESTA",
            fecha_radicacion=None,  # opcional, ya pasaron los chequeos al crear
            fecha_recepcion=None,
            valor_aceptado=str(int(glosa.valor_aceptado or 0)),
            tabla_excel=texto_para_ia,
            numero_factura=glosa.factura,
            numero_radicado=glosa.numero_radicado,
            tono=data.tono or "conciliador",
            modo_respuesta=data.modo_respuesta or "defender",
        )
    except Exception as e:
        raise HTTPException(422, f"No se pudo reconstruir el GlosaInput: {e}")

    # Trazabilidad request-scoped (R56 P1)
    from app.core.logging_utils import glosa_id_var, user_email_var

    user_email_var.set(current_user.email or "")
    glosa_id_var.set(glosa.id)

    cfg = get_settings()
    service = GlosaService(
        groq_api_key=cfg.groq_api_key,
        anthropic_api_key=cfg.anthropic_api_key,
        primary_ai=cfg.primary_ai,
        anthropic_model=cfg.anthropic_model,
        groq_model=cfg.groq_model,
        gemini_api_key=cfg.gemini_api_key,
        gemini_model=cfg.gemini_model,
    )

    contrato_repo = ContratoRepository(db)
    contratos = contrato_repo.como_dict()

    # Pre-lookup de tarifa pactada antes de invocar al LLM. Sin esto el
    # motor produce el argumento genérico "no existe contrato" aunque sí
    # exista en BD — bug crítico que afectaba a #2513 y otras TA-DMBUG.
    from app.services.tarifa_lookup_service import pre_lookup_tarifa

    info_tarifa_pre = pre_lookup_tarifa(
        db=db,
        cod_pref=glosa.codigo_glosa or "",
        eps=glosa.eps or "",
        tabla_excel=texto_para_ia,
        contexto_pdf="",
        req_id=f"reanalizar-{glosa.id}",
    )
    # Memoria del gestor: hint de estilo personal aprendido del histórico
    # de refinamientos de este auditor para casos similares (mismo CUPS,
    # misma EPS).
    hint_gestor = ""
    try:
        from app.services.memoria_gestor import patron_gestor

        patron = patron_gestor(
            db,
            autor_email=current_user.email,
            codigo_glosa=glosa.codigo_glosa or "",
            eps=glosa.eps or "",
        )
        hint_gestor = patron.get("hint_para_prompt", "") or ""
    except Exception:
        pass
    resultado = await service.analizar(
        glosa_input,
        "",
        contratos,
        info_tarifa=info_tarifa_pre,
        hint_gestor=hint_gestor,
    )

    # Sobreescribir dictamen + metadata. NO crear nueva fila.
    from datetime import datetime, timezone as _tz

    glosa.dictamen = resultado.dictamen
    glosa.dictamen_generado_en = datetime.now(_tz.utc)
    glosa.tipo_analisis = resultado.tipo if hasattr(glosa, "tipo_analisis") else None
    glosa.modelo_ia = resultado.modelo_ia
    if hasattr(glosa, "score"):
        glosa.score = resultado.score
    glosa.actualizado_en = ahora_utc()
    db.commit()
    db.refresh(glosa)

    # Snapshot del dictamen como nueva versión
    try:
        from app.api.routers.versiones import guardar_version

        guardar_version(
            db=db,
            glosa_id=glosa.id,
            dictamen_html=resultado.dictamen,
            accion="REANALIZAR",
            autor_email=current_user.email,
        )
    except Exception as _e:
        logger.warning(f"No se pudo guardar version: {_e}")

    AuditRepository(db).registrar(
        usuario_email=current_user.email,
        usuario_rol=current_user.rol,
        accion="REANALIZAR_GLOSA",
        tabla="glosas",
        registro_id=glosa.id,
        detalle=f"tono={data.tono} modo={data.modo_respuesta}",
    )

    return {
        "message": "Glosa reanalizada",
        "glosa_id": glosa.id,
        "modo": data.modo_respuesta,
        "tono": data.tono,
        "modelo_ia": resultado.modelo_ia,
        "dictamen": resultado.dictamen,
        "tipo": resultado.tipo,
    }


@router.post("/{glosa_id}/clonar")
def clonar_glosa(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_auditor_o_superior),
):
    """R65 P1: clona una glosa existente como BORRADOR para acelerar
    captura de glosas similares (mismo paciente, mismo servicio, valor
    distinto, o multi-conceptos sobre la misma factura).

    Comportamiento:
      - Copia campos descriptivos: eps, paciente, codigo_glosa, etapa,
        factura, numero_radicado, texto_glosa_original, cups_servicio,
        servicio_descripcion, concepto_glosa, fecha_recepcion.
      - NO copia: dictamen, modelo_ia, score, valor_aceptado, decisiones
        EPS, fechas de respuesta. La nueva glosa empieza limpia para
        que el gestor decida tono/modo.
      - Estado inicial: BORRADOR (no RADICADA — no se ha generado dictamen).
      - valor_objetado y valor_aceptado quedan en 0 para forzar al gestor
        a digitarlos según el nuevo concepto.

    Audit log registrado con detalle de la glosa origen.
    """
    repo = GlosaRepository(db)
    original = repo.obtener_por_id(glosa_id)
    if not original:
        raise HTTPException(404, "Glosa origen no encontrada")

    nueva = repo.crear(
        eps=original.eps,
        paciente=original.paciente,
        codigo_glosa=original.codigo_glosa,
        valor_objetado=0,
        valor_aceptado=0,
        etapa=original.etapa,
        estado="BORRADOR",
        dictamen=None,
        dias_restantes=original.dias_restantes,
        modelo_ia=None,
        score=0,
        numero_radicado=original.numero_radicado,
        factura=original.factura,
        texto_glosa_original=original.texto_glosa_original,
        codigo_respuesta=None,
        cups_servicio=original.cups_servicio,
        servicio_descripcion=original.servicio_descripcion,
        concepto_glosa=original.concepto_glosa,
        fecha_recepcion=original.fecha_recepcion,
    )

    AuditRepository(db).registrar(
        usuario_email=current_user.email,
        usuario_rol=current_user.rol,
        accion="CLONAR_GLOSA",
        tabla="glosas",
        registro_id=nueva.id,
        detalle=f"Clonada desde glosa #{glosa_id}",
    )

    return {
        "message": "Glosa clonada como BORRADOR",
        "id_origen": glosa_id,
        "id_nueva": nueva.id,
        "estado": "BORRADOR",
    }


@router.get("/{glosa_id}/paquete-evidencia.json")
def descargar_paquete_evidencia(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R85 P2: paquete completo de evidencia para una disputa.

    Bundle JSON listo para entregar al equipo legal o EPS contraparte
    cuando se necesita demostrar:
      - Cuál fue el dictamen exacto (texto + hash)
      - Quién lo firmó y cuándo
      - Qué pasó con esa glosa (timeline completo)
      - Calls IA que la generaron (auditoría regulatoria)

    Estructura:
      {
        "metadata": {generado_en, generado_por, glosa_id},
        "glosa": {...campos descriptivos...},
        "dictamen_actual": {texto, hash, firma, alg, timestamp},
        "timeline": [...eventos cronológicos...],
        "ia_calls": [...calls atribuidos...]
      }
    """
    import json
    from fastapi.responses import Response

    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    # 1. Datos descriptivos
    glosa_data = {
        "id": glosa.id,
        "eps": glosa.eps,
        "paciente": glosa.paciente,
        "codigo_glosa": glosa.codigo_glosa,
        "factura": glosa.factura,
        "numero_radicado": glosa.numero_radicado,
        "valor_objetado": float(glosa.valor_objetado or 0),
        "valor_aceptado": float(glosa.valor_aceptado or 0),
        "estado": glosa.estado,
        "modelo_ia": glosa.modelo_ia,
        "creado_en": glosa.creado_en.isoformat() if glosa.creado_en else None,
    }

    # 2. Firma del dictamen actual (si existe)
    firma_info = None
    if glosa.dictamen:
        from app.services.firma_digital import firmar_dictamen

        firma_info = firmar_dictamen(
            texto_dictamen=glosa.dictamen,
            firmante_email=current_user.email,
            glosa_id=glosa.id,
        )
        firma_info["texto_dictamen_html"] = glosa.dictamen

    # 3. Timeline reusable: invocamos la función directamente
    from app.models.db import (
        AICallRecord,
        AuditLogRecord,
        ComentarioGlosaRecord,
        DictamenVersionRecord,
    )

    eventos = []
    for v in db.query(DictamenVersionRecord).filter_by(glosa_id=glosa_id).all():
        eventos.append(
            {
                "tipo": f"VERSION_{v.accion or 'CREAR'}",
                "actor": v.autor_email,
                "timestamp": v.creado_en.isoformat() if v.creado_en else None,
            }
        )
    for a in (
        db.query(AuditLogRecord)
        .filter(AuditLogRecord.tabla.in_(("glosas", "historial")))
        .filter(AuditLogRecord.registro_id == glosa_id)
        .all()
    ):
        eventos.append(
            {
                "tipo": f"AUDIT_{a.accion or 'ACCION'}",
                "actor": a.usuario_email,
                "timestamp": a.timestamp.isoformat() if a.timestamp else None,
                "detalle": (a.detalle or "")[:200],
            }
        )
    for c in db.query(ComentarioGlosaRecord).filter_by(glosa_id=glosa_id).all():
        eventos.append(
            {
                "tipo": "COMENTARIO",
                "actor": c.autor_email,
                "timestamp": c.creado_en.isoformat() if c.creado_en else None,
                "texto": (c.texto or "")[:200],
            }
        )
    eventos.sort(key=lambda e: e.get("timestamp") or "", reverse=True)

    # 4. Calls IA atribuidos
    ia_calls = []
    for c in db.query(AICallRecord).filter_by(glosa_id=glosa_id).all():
        ia_calls.append(
            {
                "modelo": c.modelo,
                "tokens_total": (c.input_tokens or 0)
                + (c.cache_creation_input_tokens or 0)
                + (c.cache_read_input_tokens or 0)
                + (c.output_tokens or 0),
                "cost_usd": c.cost_usd,
                "latency_ms": c.latency_ms,
                "creado_en": c.creado_en.isoformat() if c.creado_en else None,
            }
        )

    payload = {
        "metadata": {
            "generado_en": ahora_utc().isoformat(),
            "generado_por": current_user.email,
            "glosa_id": glosa_id,
            "version_paquete": "R85 P2",
        },
        "glosa": glosa_data,
        "dictamen_actual": firma_info,
        "timeline": eventos,
        "ia_calls": ia_calls,
    }
    fname = f"paquete-evidencia-glosa-{glosa.id}.json"
    return Response(
        content=json.dumps(payload, ensure_ascii=False, default=str).encode("utf-8"),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/{glosa_id}/firma-dictamen")
def obtener_firma_dictamen(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R84 P1: genera la firma digital del dictamen actual.

    Útil para evidenciar integridad antes de radicar a la EPS:
    si la EPS modifica el documento, el hash cambia y la firma
    deja de validar.

    Usa RSA asimétrica si FIRMA_DIGITAL_PRIVATE_KEY está configurada
    (R50 P8); fallback a HMAC con SECRET_KEY.

    Devuelve:
      {
        "glosa_id": 42,
        "hash": "sha256-hex",
        "firma": "base64",
        "timestamp": "ISO",
        "firmante": "auditor@hus.com",
        "alg": "RSA-PSS-SHA256-v1" | "HMAC-SHA256",
        "verificable": "Endpoint /firma/verificar (futuro)"
      }
    """
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")
    if not glosa.dictamen:
        raise HTTPException(400, "La glosa no tiene dictamen generado")

    from app.services.firma_digital import firmar_dictamen

    info = firmar_dictamen(
        texto_dictamen=glosa.dictamen,
        firmante_email=current_user.email,
        glosa_id=glosa.id,
    )
    return {
        "glosa_id": glosa.id,
        **info,
    }


@router.get("/{glosa_id}/dictamen.txt")
def descargar_dictamen_txt(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R81 P1: descarga el dictamen como texto plano sin formato.

    Más portable que .md para sistemas que no entienden Markdown
    (integraciones legacy, copia/pega a correo electrónico, etc.).

    Strip total de HTML + entidades + normalización de whitespace.
    """
    import re

    from fastapi.responses import Response

    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")
    if not glosa.dictamen:
        raise HTTPException(400, "La glosa no tiene dictamen generado")

    html = glosa.dictamen
    # Reemplazos para preservar estructura visual
    txt = re.sub(r"</p>", "\n\n", html, flags=re.IGNORECASE)
    txt = re.sub(r"</h[1-6]>", "\n\n", txt, flags=re.IGNORECASE)
    txt = re.sub(r"<br\s*/?>", "\n", txt, flags=re.IGNORECASE)
    txt = re.sub(r"</li>", "\n", txt, flags=re.IGNORECASE)
    txt = re.sub(r"</tr>", "\n", txt, flags=re.IGNORECASE)
    txt = re.sub(r"</td>", " | ", txt, flags=re.IGNORECASE)
    # Quitar todos los tags
    txt = re.sub(r"<[^>]+>", "", txt)
    # Decode entidades
    txt = (
        txt.replace("&nbsp;", " ")
        .replace("&amp;", "&")
        .replace("&lt;", "<")
        .replace("&gt;", ">")
        .replace("&quot;", '"')
        .replace("&#39;", "'")
    )
    # Normalizar líneas en blanco
    txt = re.sub(r"\n{3,}", "\n\n", txt)
    txt = re.sub(r"[ \t]+", " ", txt)
    txt = re.sub(r" +\n", "\n", txt).strip()

    cabecera = (
        f"DICTAMEN GLOSA #{glosa.id}\n"
        f"{'=' * 50}\n"
        f"EPS:              {glosa.eps or '—'}\n"
        f"Código glosa:     {glosa.codigo_glosa or '—'}\n"
        f"Valor objetado:   ${(glosa.valor_objetado or 0):,.0f}\n"
        f"Estado:           {glosa.estado or '—'}\n"
        f"Factura:          {glosa.factura or '—'}\n"
        f"{'=' * 50}\n\n"
    )
    payload = (cabecera + txt).encode("utf-8")
    fname = f"dictamen-glosa-{glosa.id}.txt"
    return Response(
        content=payload,
        media_type="text/plain; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/stats/por-codigo-respuesta")
def stats_por_codigo_respuesta(
    dias: int = Query(30, ge=1, le=365),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R68 P1: agregaciones por código de respuesta (RE97xx, RE98xx, RE99xx).

    Útil para reportes mensuales: "este mes el HUS aceptó 23% (RE9702),
    aceptó parcial 15% (RE9801) y defendió 62% (RE9901)."

    Ventana configurable. Devuelve:
      {
        "ventana_dias": 30,
        "total": 145,
        "por_codigo": [
          {"codigo": "RE9901", "descripcion": "Glosa no aceptada",
           "count": 90, "valor_total": 12_345_678, "porcentaje": 62.1}
        ]
      }
    """
    from datetime import timedelta

    from sqlalchemy import func as _f

    from app.core.tz import ahora_utc

    desde = ahora_utc() - timedelta(days=int(dias))

    rows = (
        db.query(
            GlosaRecord.codigo_respuesta,
            _f.count(GlosaRecord.id),
            _f.sum(GlosaRecord.valor_objetado),
        )
        .filter(GlosaRecord.creado_en >= desde)
        .group_by(GlosaRecord.codigo_respuesta)
        .all()
    )

    descripciones = {
        "RE9901": "Glosa no aceptada (defensa)",
        "RE9701": "Glosa aceptada total (texto fijo)",
        "RE9702": "Glosa aceptada al 100%",
        "RE9801": "Glosa aceptada y subsanada parcialmente",
        "RE9502": "Glosa extemporánea",
        "": "Sin código de respuesta",
        None: "Sin código de respuesta",
    }

    total = sum(r[1] for r in rows) or 0
    por_codigo = []
    for codigo, count, valor in rows:
        porcentaje = (count / total * 100) if total else 0
        por_codigo.append(
            {
                "codigo": codigo or "—",
                "descripcion": descripciones.get(codigo, "Otro"),
                "count": count,
                "valor_total": float(valor or 0),
                "porcentaje": round(porcentaje, 1),
            }
        )
    por_codigo.sort(key=lambda x: x["count"], reverse=True)

    return {
        "ventana_dias": dias,
        "total": total,
        "por_codigo": por_codigo,
    }


@router.get("/stats/por-eps")
def stats_por_eps(
    dias: int = Query(90, ge=1, le=365),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R68 P2: distribución de glosas por EPS con tasa de recuperación.

    Útil para identificar:
      - EPS más conflictivas (más glosas formuladas)
      - EPS que más recursos absorben (mayor valor objetado)
      - EPS donde tenemos peor tasa de éxito (rev pertinencia)
      - EPS donde defendemos bien (replicar argumentos)

    Devuelve por EPS:
      count, valor_objetado, valor_aceptado, valor_recuperado,
      tasa_exito_pct (= valor_recuperado / valor_objetado * 100)

    Ordenado DESC por valor_objetado.
    """
    from datetime import timedelta

    from sqlalchemy import func as _f

    from app.core.tz import ahora_utc

    desde = ahora_utc() - timedelta(days=int(dias))

    rows = (
        db.query(
            GlosaRecord.eps,
            _f.count(GlosaRecord.id),
            _f.sum(GlosaRecord.valor_objetado),
            _f.sum(GlosaRecord.valor_aceptado),
        )
        .filter(GlosaRecord.creado_en >= desde)
        .filter(GlosaRecord.eps.isnot(None))
        .group_by(GlosaRecord.eps)
        .all()
    )

    items = []
    for eps, count, v_obj, v_ac in rows:
        v_obj = float(v_obj or 0)
        v_ac = float(v_ac or 0)
        v_rec = v_obj - v_ac
        tasa = (v_rec / v_obj * 100) if v_obj > 0 else 0
        items.append(
            {
                "eps": eps,
                "count": count,
                "valor_objetado": v_obj,
                "valor_aceptado": v_ac,
                "valor_recuperado": v_rec,
                "tasa_exito_pct": round(tasa, 1),
            }
        )
    items.sort(key=lambda x: x["valor_objetado"], reverse=True)

    return {
        "ventana_dias": dias,
        "total_eps": len(items),
        "items": items,
    }


@router.get("/{glosa_id}/dictamen.md")
def descargar_dictamen_markdown(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R69 P2: descarga el dictamen de una glosa como archivo Markdown.

    HTML → Markdown legible para:
      - Compartir con equipo legal externo (no abre HTML stylado)
      - Integración con sistemas de gestión documental que solo
        aceptan texto plano
      - Diff manual entre versiones en herramientas estándar (VSCode,
        BBEdit, etc.)

    No requiere librería externa — conversión simple por regex que
    cubre los tags reales del dictamen (h3, p, b, ul, li, tabla
    códigos al inicio).
    """
    import re
    from fastapi.responses import Response

    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")
    if not glosa.dictamen:
        raise HTTPException(400, "Esta glosa aún no tiene dictamen generado")

    html = glosa.dictamen
    # 1) headers
    md = re.sub(r"<h1[^>]*>(.*?)</h1>", r"# \1\n", html, flags=re.IGNORECASE | re.DOTALL)
    md = re.sub(r"<h2[^>]*>(.*?)</h2>", r"## \1\n", md, flags=re.IGNORECASE | re.DOTALL)
    md = re.sub(r"<h3[^>]*>(.*?)</h3>", r"### \1\n", md, flags=re.IGNORECASE | re.DOTALL)
    md = re.sub(r"<h4[^>]*>(.*?)</h4>", r"#### \1\n", md, flags=re.IGNORECASE | re.DOTALL)
    # 2) negritas
    md = re.sub(r"<(b|strong)[^>]*>(.*?)</\1>", r"**\2**", md, flags=re.IGNORECASE | re.DOTALL)
    md = re.sub(r"<(i|em)[^>]*>(.*?)</\1>", r"*\2*", md, flags=re.IGNORECASE | re.DOTALL)
    # 3) listas
    md = re.sub(r"<li[^>]*>(.*?)</li>", r"- \1\n", md, flags=re.IGNORECASE | re.DOTALL)
    # 4) saltos
    md = re.sub(r"<br\s*/?>", "\n", md, flags=re.IGNORECASE)
    md = re.sub(r"</p>", "\n\n", md, flags=re.IGNORECASE)
    md = re.sub(r"</div>", "\n", md, flags=re.IGNORECASE)
    md = re.sub(r"</tr>", "\n", md, flags=re.IGNORECASE)
    md = re.sub(r"</td>", " | ", md, flags=re.IGNORECASE)
    # 5) tags restantes
    md = re.sub(r"<[^>]+>", "", md)
    # 6) entidades comunes
    md = (
        md.replace("&nbsp;", " ")
        .replace("&amp;", "&")
        .replace("&lt;", "<")
        .replace("&gt;", ">")
        .replace("&quot;", '"')
        .replace("&#39;", "'")
    )
    # 7) normalizar líneas en blanco múltiples
    md = re.sub(r"\n{3,}", "\n\n", md)
    md = re.sub(r"[ \t]+", " ", md)
    md = re.sub(r" +\n", "\n", md)
    md = md.strip()

    # Header informativo del archivo
    cabecera = (
        f"# Dictamen Glosa #{glosa.id}\n\n"
        f"- **EPS:** {glosa.eps or '—'}\n"
        f"- **Código glosa:** {glosa.codigo_glosa or '—'}\n"
        f"- **Valor objetado:** ${(glosa.valor_objetado or 0):,.0f}\n"
        f"- **Estado:** {glosa.estado or '—'}\n"
        f"- **Factura:** {glosa.factura or '—'}\n"
        f"- **Modelo IA:** {glosa.modelo_ia or '—'}\n\n"
        f"---\n\n"
    )
    payload = (cabecera + md).encode("utf-8")

    fname = f"dictamen-glosa-{glosa.id}.md"
    return Response(
        content=payload,
        media_type="text/markdown",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/bulk-actualizar-estado")
def bulk_actualizar_estado(
    data: BulkActualizarEstadoRequest,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_auditor_o_superior),
):
    """R71 P1: actualiza el estado de N glosas en una sola transacción.

    Útil cuando llega Excel de respuesta de la EPS con decisiones para
    múltiples glosas. En vez de llamar PATCH /glosas/{id}/estado N veces,
    una sola llamada con la lista.

    Estados válidos típicos:
      LEVANTADA, RATIFICADA, ACEPTADA, ACEPTADA_PARCIAL, RESUELTA,
      CONCILIADA, EN_REVISION

    Devuelve:
      {
        "actualizadas": N,
        "no_encontradas": [ids_no_encontrados],
        "estado": "LEVANTADA"
      }

    Audit log: 1 entry por glosa con accion=BULK_UPDATE_ESTADO.
    """
    estados_validos = {
        "RADICADA",
        "BORRADOR",
        "EN_REVISION",
        "RESPONDIDA",
        "LEVANTADA",
        "RATIFICADA",
        "ACEPTADA",
        "PARCIALMENTE_ACEPTADA",
        "RESUELTA",
        "CONCILIADA",
        "ARCHIVADA",
    }
    nuevo_estado_norm = data.nuevo_estado.strip().upper()
    if nuevo_estado_norm not in estados_validos:
        raise HTTPException(
            422,
            f"Estado '{nuevo_estado_norm}' inválido. Use uno de: "
            f"{', '.join(sorted(estados_validos))}",
        )

    repo = GlosaRepository(db)
    actualizadas = 0
    no_encontradas = []
    audit_repo = AuditRepository(db)

    for gid in data.glosa_ids:
        g = repo.obtener_por_id(gid)
        if not g:
            no_encontradas.append(gid)
            continue
        estado_anterior = g.estado
        g.estado = nuevo_estado_norm
        actualizadas += 1
        audit_repo.registrar(
            usuario_email=current_user.email,
            usuario_rol=current_user.rol,
            accion="BULK_UPDATE_ESTADO",
            tabla="glosas",
            registro_id=gid,
            campo="estado",
            valor_anterior=estado_anterior,
            valor_nuevo=nuevo_estado_norm,
            detalle=(data.nota or "Bulk update de estado")[:300],
        )
    db.commit()

    return {
        "actualizadas": actualizadas,
        "no_encontradas": no_encontradas,
        "estado": nuevo_estado_norm,
    }


@router.post("/bulk-mover-papelera")
def bulk_mover_papelera(
    data: BulkMoverPapeleraRequest,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_coordinador_o_admin),
):
    """R71 P2: mueve N glosas a la papelera (soft-delete) de un golpe.

    Útil cuando se importó un Excel duplicado por error y hay 50
    glosas para depurar. Auth restringido a coordinador/admin (no
    auditor) por riesgo.

    Soporta dry_run=true para PREVIEW: lista qué glosas se moverían
    sin tocar la BD. UI puede mostrar al usuario qué borrará y pedir
    confirm.

    Cada glosa movida queda en glosas_eliminadas (R52, papelera con
    TTL 30 días) y se borra de historial. Si una glosa no existe,
    se reporta en no_encontradas pero el batch continúa.
    """
    repo = GlosaRepository(db)
    movidas = 0
    no_encontradas = []
    fallidas = []

    for gid in data.glosa_ids:
        g = repo.obtener_por_id(gid)
        if not g:
            no_encontradas.append(gid)
            continue
        if data.dry_run:
            movidas += 1
            continue
        try:
            from app.api.routers.papelera import mover_a_papelera

            mover_a_papelera(
                db,
                g,
                eliminado_por=current_user.email,
                motivo=(data.motivo or "Bulk delete")[:300],
            )
            db.delete(g)
            movidas += 1
        except Exception as e:
            fallidas.append({"id": gid, "error": str(e)[:200]})

    if not data.dry_run:
        db.commit()
        logger.info(
            f"[BULK-PAPELERA] {movidas} glosas movidas a papelera por "
            f"{current_user.email} | {len(fallidas)} fallidas"
        )

    return {
        "dry_run": data.dry_run,
        "movidas_a_papelera": movidas,
        "no_encontradas": no_encontradas,
        "fallidas": fallidas,
    }


@router.get("/{glosa_id}/acciones-disponibles")
def acciones_disponibles_glosa(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R78 P1: lista las acciones que el gestor puede tomar sobre esta
    glosa según su estado actual y contexto.

    Combina:
      - Transiciones válidas del workflow (state machine)
      - Acciones operativas disponibles
      - Sugerencia principal heurística

    Útil para que la UI muestre solo las acciones aplicables.
    """
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    from app.services.workflow_service import WorkflowService

    transiciones = WorkflowService.obtener_transiciones_validas(
        glosa.estado or "RADICADA",
    )

    tiene_dictamen = bool(glosa.dictamen and len(glosa.dictamen) > 50)
    tiene_texto_original = bool(glosa.texto_glosa_original)
    tiene_factura = bool(glosa.factura)

    return {
        "glosa_id": glosa.id,
        "estado_actual": glosa.estado,
        "transiciones_workflow": [
            {"hacia": t.hacia, "accion": t.accion, "requiere_nota": t.requiere_nota}
            for t in transiciones
        ],
        "acciones_operativas": {
            "puede_descargar_pdf": tiene_dictamen,
            "puede_descargar_md": tiene_dictamen,
            "puede_refinar": tiene_dictamen,
            "puede_reanalizar": tiene_texto_original,
            "puede_clonar": True,
            "puede_validar_rapido": tiene_dictamen,
            "puede_ver_timeline": True,
            "puede_ver_metricas_ia": True,
            "puede_buscar_duplicados": tiene_factura,
            "puede_ver_versiones": tiene_dictamen,
        },
        "sugerencia_principal": _sugerir_accion_principal(glosa),
    }


def _sugerir_accion_principal(glosa) -> Optional[str]:
    """Heurística: sugiere la próxima acción más útil según contexto."""
    estado = (glosa.estado or "").upper()
    tiene_dict = bool(glosa.dictamen and len(glosa.dictamen) > 50)
    dias = glosa.dias_restantes or 0
    if not tiene_dict:
        if glosa.texto_glosa_original or "":
            return "Generar dictamen con IA"
        return "Pegar texto de la glosa para empezar"
    if estado == "BORRADOR":
        return "Marcar como respondida cuando esté lista"
    if estado == "RADICADA" and dias > 0 and dias <= 2:
        return "URGENTE: vence en 2 días o menos — radicar respuesta YA"
    if estado == "RESPONDIDA":
        return "Esperar decisión EPS · monitor de plazos activo"
    if estado == "RATIFICADA":
        return "Considerar conciliación o escalar a SuperSalud"
    if estado == "LEVANTADA":
        return "Glosa exitosa · considerar guardar argumento como Plantilla Gold"
    return None


@router.get("/{glosa_id}/validar-rapido")
def validar_rapido_glosa(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R70 P1: validación instantánea del dictamen sin IA (solo checks
    locales programáticos del validador_dictamen).

    Diferencia con POST /{id}/validar (validar_pre_radicacion):
      - validar_pre_radicacion       → llama IA, ~5 seg, ~$0.05 USD
      - validar-rapido (este)        → solo checks locales, <50 ms, $0

    Útil para feedback rápido al gestor mientras edita o como
    pre-check antes del validador completo.

    Aplica los 11 checks del validador_dictamen.evaluar_dictamen():
      apertura, cups_real, sin_cifras_inventadas, normas_citadas,
      enumeracion, invitacion_conciliacion, extension,
      codigo_respuesta_coherente, contrato_mencionado, placeholders,
      cita_literal_normativa, anti_rebatimiento.

    Devuelve score 0-100, total checks, aprobados, lista detallada.
    """
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")
    if not glosa.dictamen:
        raise HTTPException(400, "La glosa aún no tiene dictamen generado")

    from app.services.validador_dictamen import evaluar_dictamen

    resultado = evaluar_dictamen(
        glosa.dictamen or "",
        codigo_glosa=glosa.codigo_glosa or "",
        cups_esperado=glosa.cups_servicio,
        valor_original=str(int(glosa.valor_objetado or 0)),
        codigo_respuesta=glosa.codigo_respuesta,
        eps=glosa.eps or "",
    )
    return {
        "glosa_id": glosa.id,
        **resultado,
    }


# ─── ESTADÍSTICAS (extraídas a glosas_stats.py — Fase 2) ───────────────────
# Los ~170 endpoints /stats/* están en app/api/routers/glosas_stats.py.
# Se incluyen aquí vía include_router para mantener el mismo path /glosas/stats/*.
from app.api.routers.glosas_stats import stats_router as _stats_router  # noqa: E402

router.include_router(_stats_router)
# ─────────────────────────────────────────────────────────────────────────────


@router.get("/{glosa_id}/checklist")
def checklist_glosa(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R96 P1: checklist de progreso de una glosa en el ciclo.

    Útil para que el auditor vea de un vistazo qué falta en cada
    glosa. Cada item dice si está completo + si es opcional.

    Devuelve:
      {
        "glosa_id": int,
        "items": [
          {"id": "texto_original", "descripcion": "...",
           "completado": true, "opcional": false},
          ...
        ],
        "total_items": int,
        "completados": int,
        "obligatorios_pendientes": int,
        "porcentaje_avance": float (0-100, solo obligatorios)
      }
    """
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    ESTADOS_CERRADOS = {"ACEPTADA", "LEVANTADA", "ARCHIVADA", "CONCILIADA"}

    items = [
        {
            "id": "texto_original",
            "descripcion": "Texto de glosa original capturado",
            "completado": bool(glosa.texto_glosa_original),
            "opcional": False,
        },
        {
            "id": "factura",
            "descripcion": "Factura asociada (no N/A)",
            "completado": bool(glosa.factura and glosa.factura != "N/A"),
            "opcional": False,
        },
        {
            "id": "valor_objetado",
            "descripcion": "Valor objetado registrado",
            "completado": bool(glosa.valor_objetado and glosa.valor_objetado > 0),
            "opcional": False,
        },
        {
            "id": "dictamen",
            "descripcion": "Dictamen HUS generado",
            "completado": bool(glosa.dictamen and len(glosa.dictamen) > 50),
            "opcional": False,
        },
        {
            "id": "gestor",
            "descripcion": "Gestor asignado",
            "completado": bool(glosa.gestor_nombre),
            "opcional": True,
        },
        {
            "id": "auditor",
            "descripcion": "Auditor asignado",
            "completado": bool(glosa.auditor_email),
            "opcional": True,
        },
        {
            "id": "fecha_recepcion",
            "descripcion": "Fecha de recepción registrada",
            "completado": bool(glosa.fecha_recepcion),
            "opcional": True,
        },
        {
            "id": "respuesta_eps",
            "descripcion": "Decisión EPS registrada",
            "completado": bool(glosa.decision_eps),
            "opcional": False,
        },
        {
            "id": "cierre",
            "descripcion": "Glosa cerrada",
            "completado": (glosa.estado or "").upper() in ESTADOS_CERRADOS,
            "opcional": False,
        },
    ]

    total = len(items)
    completados = sum(1 for it in items if it["completado"])
    obligatorios = [it for it in items if not it["opcional"]]
    obl_completados = sum(1 for it in obligatorios if it["completado"])
    obl_pendientes = len(obligatorios) - obl_completados
    pct = round(100 * obl_completados / len(obligatorios), 2) if obligatorios else 0.0

    return {
        "glosa_id": glosa_id,
        "items": items,
        "total_items": total,
        "completados": completados,
        "obligatorios_pendientes": obl_pendientes,
        "porcentaje_avance": pct,
    }


@router.get("/{glosa_id}/contexto-completo")
def contexto_completo_glosa(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R94 P2: contexto agregado para vista detalle de una glosa.

    Combina en un solo round-trip:
      - glosa (campos clave)
      - sla (estado_sla, color_semaforo, dias_restantes)
      - audit_resumen (total_cambios, ultimo_cambio_en, usuarios)
      - relacionadas_count (sin items para no inflar — usar
        /relacionadas para detalle)

    Reduce N+1 calls del frontend al cargar la ficha de una glosa.
    Si el frontend necesita detalle de cada sección, puede invocar
    los endpoints individuales.
    """
    from datetime import timezone

    from app.models.db import AuditLogRecord

    ESTADOS_CERRADOS = {"ACEPTADA", "LEVANTADA", "ARCHIVADA", "CONCILIADA"}

    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    # ─── SLA ─────────────────────────────────────────────────
    estado = (glosa.estado or "").upper()
    cerrada = estado in ESTADOS_CERRADOS

    venc = glosa.fecha_vencimiento
    if venc and venc.tzinfo is None:
        venc = venc.replace(tzinfo=timezone.utc)
    dec = glosa.fecha_decision_eps
    if dec and dec.tzinfo is None:
        dec = dec.replace(tzinfo=timezone.utc)

    if not venc:
        estado_sla, color = "SIN_VENCIMIENTO", "GRIS"
    elif cerrada:
        if dec and dec <= venc:
            estado_sla, color = "CERRADA_A_TIEMPO", "VERDE"
        else:
            estado_sla, color = "CERRADA_TARDE", "NEGRO"
    else:
        dr = glosa.dias_restantes if glosa.dias_restantes is not None else 0
        if dr < 0:
            estado_sla, color = "VENCIDA", "ROJO"
        elif dr <= 3:
            estado_sla, color = "CRITICA", "AMARILLO"
        else:
            estado_sla, color = "EN_TIEMPO", "VERDE"

    # ─── Audit resumen ──────────────────────────────────────
    eventos = (
        db.query(AuditLogRecord)
        .filter(AuditLogRecord.tabla == "glosas")
        .filter(AuditLogRecord.registro_id == glosa_id)
        .all()
    )
    timestamps = [e.timestamp for e in eventos if e.timestamp]
    usuarios = sorted({e.usuario_email for e in eventos if e.usuario_email})

    # ─── Relacionadas (counts only) ─────────────────────────
    rel_factura = 0
    if glosa.factura and glosa.factura != "N/A":
        rel_factura = (
            db.query(GlosaRecord)
            .filter(GlosaRecord.factura == glosa.factura)
            .filter(GlosaRecord.id != glosa_id)
            .count()
        )
    rel_paciente = 0
    if glosa.paciente:
        rel_paciente = (
            db.query(GlosaRecord)
            .filter(GlosaRecord.paciente == glosa.paciente)
            .filter(GlosaRecord.id != glosa_id)
            .count()
        )
    rel_patron = 0
    if glosa.codigo_glosa and glosa.eps:
        rel_patron = (
            db.query(GlosaRecord)
            .filter(GlosaRecord.codigo_glosa == glosa.codigo_glosa)
            .filter(GlosaRecord.eps == glosa.eps)
            .filter(GlosaRecord.id != glosa_id)
            .count()
        )

    return {
        "glosa": {
            "id": glosa.id,
            "creado_en": (glosa.creado_en.isoformat() if glosa.creado_en else None),
            "eps": glosa.eps,
            "paciente": glosa.paciente,
            "factura": glosa.factura,
            "codigo_glosa": glosa.codigo_glosa,
            "valor_objetado": float(glosa.valor_objetado or 0),
            "valor_recuperado": float(glosa.valor_recuperado or 0),
            "estado": glosa.estado,
            "etapa": glosa.etapa,
            "decision_eps": glosa.decision_eps,
        },
        "sla": {
            "estado_sla": estado_sla,
            "color_semaforo": color,
            "cerrada": cerrada,
            "dias_restantes": glosa.dias_restantes,
            "fecha_vencimiento": venc.isoformat() if venc else None,
        },
        "audit_resumen": {
            "total_cambios": len(eventos),
            "ultimo_cambio_en": (max(timestamps).isoformat() if timestamps else None),
            "usuarios_que_intervinieron": usuarios,
        },
        "relacionadas_count": {
            "misma_factura": rel_factura,
            "mismo_paciente": rel_paciente,
            "mismo_codigo_y_eps": rel_patron,
        },
    }


@router.get("/{glosa_id}/score-prioridad")
def score_prioridad_glosa(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R112 P2: score individual de prioridad para UNA glosa.

    Misma fórmula que /admin/glosas-prioritarias (R112 P1) pero
    aplicada a una sola glosa, con desglose detallado de cada
    componente.

    Útil para mostrar en la ficha de la glosa: "esta glosa tiene
    score 130 porque está vencida + alto valor".

    Devuelve:
      - score_total
      - desglose: {[{componente, peso, razon}]}
      - banner_recomendado: "URGENTE" | "ALTA" | "MEDIA" | "BAJA"
    """
    ESTADOS_CERRADOS = {"ACEPTADA", "LEVANTADA", "ARCHIVADA", "CONCILIADA"}

    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    estado = (glosa.estado or "").upper()
    if estado in ESTADOS_CERRADOS:
        return {
            "glosa_id": glosa_id,
            "score_total": 0,
            "desglose": [],
            "banner_recomendado": "INFO",
            "razon": "Glosa cerrada — sin score de prioridad.",
        }

    desglose = []
    score = 0

    dr = glosa.dias_restantes if glosa.dias_restantes is not None else 0
    if dr < 0:
        desglose.append(
            {"componente": "vencimiento", "peso": 100, "razon": f"vencida hace {abs(dr)}d"}
        )
        score += 100
    elif dr <= 3:
        desglose.append(
            {"componente": "vencimiento", "peso": 50, "razon": f"crítica ({dr}d restantes)"}
        )
        score += 50
    elif dr <= 7:
        desglose.append(
            {"componente": "vencimiento", "peso": 20, "razon": f"próxima ({dr}d restantes)"}
        )
        score += 20

    v_obj = float(glosa.valor_objetado or 0)
    if v_obj > 10_000_000:
        desglose.append(
            {"componente": "valor", "peso": 30, "razon": f"alto valor ({int(v_obj):,} COP)"}
        )
        score += 30
    elif v_obj > 1_000_000:
        desglose.append(
            {"componente": "valor", "peso": 15, "razon": f"valor medio ({int(v_obj):,} COP)"}
        )
        score += 15

    if not glosa.dictamen or len(glosa.dictamen) < 50:
        desglose.append({"componente": "dictamen", "peso": 25, "razon": "sin dictamen generado"})
        score += 25

    if not glosa.gestor_nombre:
        desglose.append({"componente": "asignacion", "peso": 15, "razon": "sin gestor asignado"})
        score += 15

    if score >= 100:
        banner = "URGENTE"
    elif score >= 50:
        banner = "ALTA"
    elif score >= 25:
        banner = "MEDIA"
    elif score > 0:
        banner = "BAJA"
    else:
        banner = "INFO"

    return {
        "glosa_id": glosa_id,
        "estado": glosa.estado,
        "score_total": score,
        "desglose": desglose,
        "banner_recomendado": banner,
    }


@router.get("/{glosa_id}/versiones-resumen")
def versiones_resumen_glosa(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R129 P1: resumen del versionado del dictamen de una glosa.

    Cada refinación, regeneración o restauración del dictamen
    crea un DictamenVersionRecord. Este endpoint resume:
      - Cuántas versiones tiene el dictamen
      - Quién lo refinó cuándo
      - Cuántas veces se REFINO (con instrucción humana) vs
        REGENERO (con IA pura)

    Útil para entender la "historia editorial" de un dictamen
    sin tener que ir versión por versión.

    Devuelve:
      - total_versiones
      - por_accion: mapa {CREAR, REFINAR, REGENERAR, RESTAURAR}
      - autores_distintos
      - primera_version_en / ultima_version_en
      - ultima_accion
    """
    from datetime import timezone

    from app.models.db import DictamenVersionRecord

    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    versiones = (
        db.query(DictamenVersionRecord)
        .filter(DictamenVersionRecord.glosa_id == glosa_id)
        .order_by(DictamenVersionRecord.creado_en.asc())
        .all()
    )

    if not versiones:
        return {
            "glosa_id": glosa_id,
            "total_versiones": 0,
            "por_accion": {},
            "autores_distintos": [],
            "primera_version_en": None,
            "ultima_version_en": None,
            "ultima_accion": None,
        }

    por_accion: dict[str, int] = {}
    autores: set[str] = set()
    for v in versiones:
        if v.accion:
            por_accion[v.accion] = por_accion.get(v.accion, 0) + 1
        if v.autor_email:
            autores.add(v.autor_email)

    primera = versiones[0].creado_en
    ultima = versiones[-1].creado_en
    if primera and primera.tzinfo is None:
        primera = primera.replace(tzinfo=timezone.utc)
    if ultima and ultima.tzinfo is None:
        ultima = ultima.replace(tzinfo=timezone.utc)

    return {
        "glosa_id": glosa_id,
        "total_versiones": len(versiones),
        "por_accion": por_accion,
        "autores_distintos": sorted(autores),
        "primera_version_en": primera.isoformat() if primera else None,
        "ultima_version_en": ultima.isoformat() if ultima else None,
        "ultima_accion": versiones[-1].accion,
    }


@router.get("/{glosa_id}/dialogo-bilateral")
def dialogo_bilateral(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R138 P1: narrativa cronológica del intercambio HUS↔EPS.

    Construye un diálogo entre las partes basado en datos reales:
      1. EPS objeta: codigo_glosa + valor_objetado
      2. HUS responde: dictamen + codigo_respuesta
      3. EPS decide: decision_eps + valor_recuperado
      4. (Opcional) Conciliación bilateral

    Cada paso: {actor, fecha, mensaje, estado_resultante}.

    Útil para mostrar la "historia" completa de la glosa de
    forma legible para no-técnicos (legal, gerencia).
    """
    from datetime import timezone

    from app.models.db import ConciliacionRecord

    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    pasos = []

    # 1) EPS objeta
    if glosa.creado_en:
        ts = glosa.creado_en
        if ts.tzinfo is None:
            ts = ts.replace(tzinfo=timezone.utc)
        valor = float(glosa.valor_objetado or 0)
        pasos.append(
            {
                "actor": "EPS",
                "fecha": ts.isoformat(),
                "mensaje": (
                    f"Objeta con código {glosa.codigo_glosa or '?'} por ${int(valor):,} COP"
                ),
                "estado_resultante": "RADICADA",
            }
        )

    # 2) HUS responde
    if glosa.dictamen and len(glosa.dictamen) > 50:
        pasos.append(
            {
                "actor": "HUS",
                "fecha": None,
                "mensaje": (
                    f"Responde con código {glosa.codigo_respuesta or '?'} "
                    f"y dictamen técnico-jurídico "
                    f"({len(glosa.dictamen)} chars)"
                ),
                "estado_resultante": glosa.estado or "RESPONDIDA",
            }
        )

    # 3) EPS decide
    if glosa.decision_eps:
        ts = glosa.fecha_decision_eps
        if ts and ts.tzinfo is None:
            ts = ts.replace(tzinfo=timezone.utc)
        v_rec = float(glosa.valor_recuperado or 0)
        pasos.append(
            {
                "actor": "EPS",
                "fecha": ts.isoformat() if ts else None,
                "mensaje": (f"Decisión: {glosa.decision_eps}. Recuperado: ${int(v_rec):,} COP"),
                "estado_resultante": glosa.estado or "?",
            }
        )

    # 4) Conciliación si existe
    conciliaciones = (
        db.query(ConciliacionRecord)
        .filter(ConciliacionRecord.glosa_id == glosa_id)
        .order_by(ConciliacionRecord.creado_en.asc())
        .all()
    )
    for c in conciliaciones:
        ts = c.fecha_audiencia or c.creado_en
        if ts and ts.tzinfo is None:
            ts = ts.replace(tzinfo=timezone.utc)
        v_conc = float(c.valor_conciliado or 0)
        pasos.append(
            {
                "actor": "BILATERAL",
                "fecha": ts.isoformat() if ts else None,
                "mensaje": (
                    f"Conciliación: {c.resultado or 'pendiente'}. "
                    f"Valor conciliado: ${int(v_conc):,} COP"
                ),
                "estado_resultante": c.estado_bilateral or "?",
            }
        )

    return {
        "glosa_id": glosa_id,
        "estado_actual": glosa.estado,
        "total_pasos": len(pasos),
        "dialogo": pasos,
    }


@router.get("/{glosa_id}/json-completo")
def json_completo_glosa(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R142 P2: dump completo del GlosaRecord en JSON.

    Devuelve TODOS los campos de la fila, útil para:
      - Debugging: ver el estado completo sin filtros
      - Backup individual de la glosa
      - Migración a otro sistema

    No incluye relaciones (conceptos, comentarios, audit) — solo
    la fila plana del GlosaRecord. Para el paquete completo usar
    /glosas/{id}/exportar-evidencia.zip.
    """
    from datetime import datetime as _dt

    from sqlalchemy import inspect as _inspect

    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    out = {}
    for col in _inspect(glosa).mapper.column_attrs:
        val = getattr(glosa, col.key)
        if isinstance(val, _dt):
            out[col.key] = val.isoformat()
        else:
            out[col.key] = val
    return out


@router.get("/{glosa_id}/estado-resumen")
def estado_resumen(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R251 P1: vista ultra-compacta del estado de una glosa.

    Diferente a /{id}/dashboard (con contadores) y a
    /{id}/contexto-completo (con relaciones): aquí sólo los 5
    datos esenciales para tooltip o cards minimalistas:
      - id, eps, valor, estado, dias_restantes

    Útil para componentes UI muy pequeños.
    """
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    return {
        "id": glosa.id,
        "eps": glosa.eps,
        "valor_objetado": float(glosa.valor_objetado or 0),
        "estado": glosa.estado,
        "dias_restantes": glosa.dias_restantes,
    }


@router.get("/{glosa_id}/glosas-mismo-paciente")
def glosas_mismo_paciente(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R235 P1: glosas del mismo paciente (historial clínico).

    Para una glosa, devuelve otras glosas del mismo paciente
    (case-insensitive). Útil para investigar:
      "Este paciente tiene 5 glosas, todas de SANITAS — patrón"

    Excluye la glosa actual.

    Devuelve {paciente, total_otras, items}.
    """
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    if not glosa.paciente:
        return {
            "paciente": None,
            "total_otras": 0,
            "items": [],
        }

    otras = (
        db.query(GlosaRecord)
        .filter(GlosaRecord.paciente.ilike(glosa.paciente))
        .filter(GlosaRecord.id != glosa_id)
        .order_by(GlosaRecord.creado_en.desc())
        .all()
    )

    items = []
    for g in otras:
        items.append(
            {
                "id": g.id,
                "creado_en": (g.creado_en.isoformat() if g.creado_en else None),
                "eps": g.eps,
                "factura": g.factura,
                "estado": g.estado,
                "valor_objetado": float(g.valor_objetado or 0),
            }
        )

    return {
        "paciente": glosa.paciente,
        "total_otras": len(otras),
        "items": items[:50],
    }


@router.get("/{glosa_id}/glosas-misma-factura")
def glosas_misma_factura(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R234 P1: glosas hermanas de la misma factura.

    Para una glosa, devuelve TODAS las otras glosas con la
    misma `factura`. Útil para ver el contexto: "esta factura
    tiene 8 glosas, agruparlas para defender en bloque".

    Excluye la glosa actual del listado.
    """
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    if not glosa.factura or glosa.factura == "N/A":
        return {
            "factura": glosa.factura,
            "total_hermanas": 0,
            "items": [],
        }

    hermanas = (
        db.query(GlosaRecord)
        .filter(GlosaRecord.factura == glosa.factura)
        .filter(GlosaRecord.id != glosa_id)
        .order_by(GlosaRecord.creado_en.desc())
        .all()
    )

    items = []
    for g in hermanas:
        items.append(
            {
                "id": g.id,
                "creado_en": (g.creado_en.isoformat() if g.creado_en else None),
                "eps": g.eps,
                "codigo_glosa": g.codigo_glosa,
                "estado": g.estado,
                "valor_objetado": float(g.valor_objetado or 0),
            }
        )

    return {
        "factura": glosa.factura,
        "total_hermanas": len(hermanas),
        "items": items,
    }


@router.get("/{glosa_id}/dictamen-similar-anterior")
def dictamen_similar_anterior(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R231 P1: sugerir dictamen de una glosa similar cerrada.

    Para una glosa, busca otras del mismo (eps, codigo_glosa)
    que estén LEVANTADAS y devuelve el dictamen más reciente
    como base para copiar/pegar.

    Útil para acelerar redacción cuando un caso similar ya se
    ganó.

    Devuelve {glosa_id_origen, dictamen, valor_recuperado_origen}
    o sin_match=true.
    """
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    if not glosa.eps or not glosa.codigo_glosa:
        return {
            "glosa_id_origen": None,
            "dictamen": None,
            "sin_match": True,
            "razon": "Glosa sin eps o codigo_glosa",
        }

    similar = (
        db.query(GlosaRecord)
        .filter(GlosaRecord.id != glosa_id)
        .filter(GlosaRecord.eps == glosa.eps)
        .filter(GlosaRecord.codigo_glosa == glosa.codigo_glosa)
        .filter(GlosaRecord.estado == "LEVANTADA")
        .filter(GlosaRecord.dictamen.isnot(None))
        .order_by(GlosaRecord.fecha_decision_eps.desc())
        .first()
    )

    if not similar:
        return {
            "glosa_id_origen": None,
            "dictamen": None,
            "sin_match": True,
            "razon": (f"No hay glosas LEVANTADAS de eps={glosa.eps} codigo={glosa.codigo_glosa}"),
        }

    return {
        "glosa_id_origen": similar.id,
        "dictamen": similar.dictamen,
        "valor_recuperado_origen": float(similar.valor_recuperado or 0),
        "fecha_decision_origen": (
            similar.fecha_decision_eps.isoformat() if similar.fecha_decision_eps else None
        ),
        "sin_match": False,
    }


@router.get("/stats/historicos-ganadores")
def historicos_ganadores(
    eps: str,
    codigo: str,
    limit: int = 3,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Lista los N dictámenes LEVANTADOS más recientes del mismo (eps, codigo).

    Side-by-side con "el ganador histórico" — el gestor puede ver hasta 3
    casos previos donde un argumento similar ya ganó, leer el dictamen
    completo y copiar las frases que probaron funcionar.

    A diferencia de /{id}/dictamen-similar-anterior (que necesita una glosa
    ya guardada como pivote), este endpoint acepta eps+codigo directos
    para poder dispararse ANTES del primer "Analizar con IA" — el panel UI
    se carga apenas el gestor selecciona EPS y se detecta el código.

    Usa ilike+trim para tolerar variaciones de casing/espacios en la EPS
    (mismo criterio que few_shot_gold).
    """
    eps_norm = (eps or "").strip()
    cod_norm = (codigo or "").strip().upper()
    if not eps_norm or not cod_norm:
        return {"total": 0, "items": [], "razon": "Falta eps o codigo"}

    limit = max(1, min(int(limit or 3), 10))

    filas = (
        db.query(GlosaRecord)
        .filter(GlosaRecord.eps.ilike(eps_norm))
        .filter(GlosaRecord.codigo_glosa == cod_norm)
        .filter(GlosaRecord.estado == "LEVANTADA")
        .filter(GlosaRecord.dictamen.isnot(None))
        .order_by(GlosaRecord.fecha_decision_eps.desc().nullslast())
        .limit(limit)
        .all()
    )

    valor_recuperado_total = sum(float(g.valor_recuperado or 0) for g in filas)

    return {
        "total": len(filas),
        "eps": eps_norm,
        "codigo": cod_norm,
        "valor_recuperado_total": valor_recuperado_total,
        "items": [
            {
                "glosa_id": g.id,
                "dictamen": g.dictamen,
                "valor_recuperado": float(g.valor_recuperado or 0),
                "valor_objetado": float(g.valor_objetado or 0),
                "fecha_decision_eps": (
                    g.fecha_decision_eps.isoformat() if g.fecha_decision_eps else None
                ),
                "auditor_email": g.auditor_email or "",
                "paciente": g.paciente or "",
            }
            for g in filas
        ],
    }


@router.get("/{glosa_id}/dashboard")
def dashboard_glosa(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R217 P1: dashboard ejecutivo de UNA glosa específica.

    Single-call que reúne todas las métricas relevantes en
    counts (no listas largas):
      - Datos de la glosa
      - Conceptos asociados
      - Versiones de dictamen
      - Comentarios + menciones pendientes
      - Conciliaciones
      - Audit log count

    Útil para abrir vista detalle sin múltiples requests.
    """
    from sqlalchemy import func as _f

    from app.models.db import (
        AuditLogRecord,
        ComentarioGlosaRecord,
        ConceptoGlosaRecord,
        ConciliacionRecord,
        DictamenVersionRecord,
    )

    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    n_conceptos = (
        db.query(_f.count(ConceptoGlosaRecord.id))
        .filter(ConceptoGlosaRecord.glosa_id == glosa_id)
        .scalar()
        or 0
    )
    n_versiones = (
        db.query(_f.count(DictamenVersionRecord.id))
        .filter(DictamenVersionRecord.glosa_id == glosa_id)
        .scalar()
        or 0
    )
    n_comentarios = (
        db.query(_f.count(ComentarioGlosaRecord.id))
        .filter(ComentarioGlosaRecord.glosa_id == glosa_id)
        .scalar()
        or 0
    )
    menciones_pend = (
        db.query(_f.count(ComentarioGlosaRecord.id))
        .filter(ComentarioGlosaRecord.glosa_id == glosa_id)
        .filter(ComentarioGlosaRecord.mencion.isnot(None))
        .filter((ComentarioGlosaRecord.resuelto == 0) | (ComentarioGlosaRecord.resuelto.is_(None)))
        .scalar()
        or 0
    )
    n_conciliaciones = (
        db.query(_f.count(ConciliacionRecord.id))
        .filter(ConciliacionRecord.glosa_id == glosa_id)
        .scalar()
        or 0
    )
    n_audit = (
        db.query(_f.count(AuditLogRecord.id))
        .filter(AuditLogRecord.tabla == "glosas")
        .filter(AuditLogRecord.registro_id == glosa_id)
        .scalar()
        or 0
    )

    return {
        "glosa_id": glosa_id,
        "datos": {
            "eps": glosa.eps,
            "factura": glosa.factura,
            "estado": glosa.estado,
            "valor_objetado": float(glosa.valor_objetado or 0),
            "valor_recuperado": float(glosa.valor_recuperado or 0),
            "dias_restantes": glosa.dias_restantes,
        },
        "contadores": {
            "conceptos": int(n_conceptos),
            "versiones_dictamen": int(n_versiones),
            "comentarios": int(n_comentarios),
            "menciones_pendientes": int(menciones_pend),
            "conciliaciones": int(n_conciliaciones),
            "eventos_audit": int(n_audit),
        },
    }


@router.get("/{glosa_id}/checklist-pre-envio")
def checklist_pre_envio(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R215 P1: checklist específico previo al envío a la EPS.

    Diferente a /{id}/checklist (avance general): aquí solo
    los ítems críticos para enviar respuesta:
      - EPS configurada
      - Factura no es N/A
      - Código glosa válido
      - Dictamen >= 200 chars
      - Código respuesta configurado
      - Gestor asignado
      - No vencida (dr >= 0)

    Devuelve {checklist[], todos_ok, faltantes}.
    """
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    items = []

    items.append(
        {
            "item": "EPS configurada",
            "ok": bool(glosa.eps and glosa.eps.strip()),
        }
    )
    items.append(
        {
            "item": "Factura válida (no N/A)",
            "ok": bool(glosa.factura and glosa.factura != "N/A"),
        }
    )
    items.append(
        {
            "item": "Código glosa configurado",
            "ok": bool(glosa.codigo_glosa and glosa.codigo_glosa.strip()),
        }
    )
    items.append(
        {
            "item": "Dictamen sólido (>=200 chars)",
            "ok": bool(glosa.dictamen and len(glosa.dictamen) >= 200),
        }
    )
    items.append(
        {
            "item": "Código respuesta configurado",
            "ok": bool(glosa.codigo_respuesta and glosa.codigo_respuesta.strip()),
        }
    )
    items.append(
        {
            "item": "Gestor asignado",
            "ok": bool(glosa.gestor_nombre),
        }
    )
    dr = glosa.dias_restantes if glosa.dias_restantes is not None else 0
    items.append(
        {
            "item": "No vencida",
            "ok": dr >= 0,
            "detalle": f"dias_restantes={dr}",
        }
    )

    todos_ok = all(it["ok"] for it in items)
    faltantes = [it["item"] for it in items if not it["ok"]]

    return {
        "glosa_id": glosa_id,
        "checklist": items,
        "todos_ok": todos_ok,
        "faltantes": faltantes,
    }


@router.get("/{glosa_id}/conciliaciones-resumen")
def conciliaciones_resumen(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R165 P1: resumen de conciliaciones de una glosa.

    Para una glosa específica, muestra las conciliaciones
    bilaterales asociadas (puede tener múltiples a lo largo
    del proceso).

    Útil para identificar el ciclo de defensa de una glosa
    sin tener que ir a /glosas/stats/conciliaciones global.

    Devuelve:
      - total
      - en_curso (estado_bilateral != ACTA_FIRMADA/CERRADA)
      - valor_conciliado_total
      - items: lista resumida (máx 20)
    """
    from datetime import timezone

    from app.models.db import ConciliacionRecord

    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    conciliaciones = (
        db.query(ConciliacionRecord)
        .filter(ConciliacionRecord.glosa_id == glosa_id)
        .order_by(ConciliacionRecord.creado_en.desc())
        .all()
    )

    CERRADAS = {"ACTA_FIRMADA", "CERRADA"}
    en_curso = sum(1 for c in conciliaciones if (c.estado_bilateral or "") not in CERRADAS)
    valor_total = sum(float(c.valor_conciliado or 0) for c in conciliaciones)

    items = []
    for c in conciliaciones[:20]:
        cre = c.creado_en
        if cre and cre.tzinfo is None:
            cre = cre.replace(tzinfo=timezone.utc)
        items.append(
            {
                "id": c.id,
                "creado_en": cre.isoformat() if cre else None,
                "estado_bilateral": c.estado_bilateral,
                "resultado": c.resultado,
                "valor_conciliado": float(c.valor_conciliado or 0),
                "acta_numero": c.acta_numero,
            }
        )

    return {
        "glosa_id": glosa_id,
        "total": len(conciliaciones),
        "en_curso": en_curso,
        "valor_conciliado_total": int(valor_total),
        "items": items,
    }


@router.get("/{glosa_id}/whatsapp-mensaje")
def whatsapp_mensaje(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R179 P1: pre-formatea un mensaje listo para WhatsApp.

    Devuelve un texto plano corto con los datos clave de la
    glosa, listo para copiar/pegar en WhatsApp o usar con la
    integración Meta Business.

    No envía nada — solo formatea.

    Útil para coordinación rápida entre equipo HUS por WhatsApp.
    """
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    valor = float(glosa.valor_objetado or 0)
    dr = glosa.dias_restantes if glosa.dias_restantes is not None else 0

    if dr < 0:
        urgencia = f"VENCIDA hace {abs(dr)}d"
    elif dr <= 3:
        urgencia = f"CRITICA - {dr}d"
    elif dr <= 7:
        urgencia = f"PROXIMA - {dr}d"
    else:
        urgencia = f"EN TIEMPO - {dr}d"

    mensaje = (
        f"Glosa #{glosa.id} | {glosa.eps or '?'}\n"
        f"Factura: {glosa.factura or 'N/A'}\n"
        f"Codigo: {glosa.codigo_glosa or '?'}\n"
        f"Valor objetado: ${int(valor):,} COP\n"
        f"Estado: {glosa.estado or '?'}\n"
        f"{urgencia}"
    )

    return {
        "glosa_id": glosa_id,
        "mensaje_whatsapp": mensaje,
        "longitud_chars": len(mensaje),
    }


@router.get("/{glosa_id}/score-defensa")
def score_defensa(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R185 P1: score predictivo de probabilidad de defensa exitosa.

    Heurístico (sin IA): combina señales históricas para estimar
    qué tan probable es que esta glosa sea LEVANTADA.

    Señales:
      +1 EPS con tasa_levantamiento alta (>60%)
      +1 codigo_glosa con tasa_levantamiento alta histórica
      +1 dictamen presente (>200 chars)
      +1 codigo_respuesta efectivo (RE9501, RE9502, RE9602, RE9901)
      +1 con tiempo (dr>3) / -1 si vencida o crítica

    Score 0-5. Veredicto:
      PROBABLE_DEFENSA (>=4) / INCIERTA (2-3) / PROBABLE_RATIFICACION (<2)
    """
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    score = 0
    razones = []

    ESTADOS_CERRADOS = {"ACEPTADA", "LEVANTADA", "ARCHIVADA", "CONCILIADA"}

    if glosa.eps:
        misma_eps = (
            db.query(GlosaRecord)
            .filter(GlosaRecord.eps == glosa.eps)
            .filter(GlosaRecord.id != glosa_id)
            .filter(GlosaRecord.estado.in_(ESTADOS_CERRADOS))
            .all()
        )
        decididas = [
            g
            for g in misma_eps
            if (g.estado or "").upper() in {"LEVANTADA", "ACEPTADA", "RATIFICADA"}
        ]
        levantadas = [g for g in decididas if (g.estado or "").upper() == "LEVANTADA"]
        if decididas:
            tasa = 100 * len(levantadas) / len(decididas)
            if tasa > 60:
                score += 1
                razones.append(f"EPS con buena tasa histórica ({tasa:.0f}%)")

    if glosa.codigo_glosa:
        mismo_cod = (
            db.query(GlosaRecord)
            .filter(GlosaRecord.codigo_glosa == glosa.codigo_glosa)
            .filter(GlosaRecord.id != glosa_id)
            .filter(GlosaRecord.estado.in_(ESTADOS_CERRADOS))
            .all()
        )
        decididas = [
            g
            for g in mismo_cod
            if (g.estado or "").upper() in {"LEVANTADA", "ACEPTADA", "RATIFICADA"}
        ]
        levantadas = [g for g in decididas if (g.estado or "").upper() == "LEVANTADA"]
        if decididas:
            tasa = 100 * len(levantadas) / len(decididas)
            if tasa > 60:
                score += 1
                razones.append(
                    f"Código {glosa.codigo_glosa} con buena tasa histórica ({tasa:.0f}%)"
                )

    if glosa.dictamen and len(glosa.dictamen) > 200:
        score += 1
        razones.append(f"Dictamen sólido ({len(glosa.dictamen)} chars)")

    EFECTIVOS = {"RE9501", "RE9502", "RE9602", "RE9901", "RE9601"}
    if glosa.codigo_respuesta in EFECTIVOS:
        score += 1
        razones.append(f"Código respuesta {glosa.codigo_respuesta} es de defensa")

    dr = glosa.dias_restantes if glosa.dias_restantes is not None else 0
    if dr < 0:
        score -= 1
        razones.append(f"PENALIZACIÓN: vencida hace {abs(dr)}d")
    elif dr <= 3:
        score -= 1
        razones.append(f"PENALIZACIÓN: crítica ({dr}d)")
    else:
        score += 1
        razones.append(f"Aún en tiempo ({dr}d)")

    score = max(0, min(5, score))

    if score >= 4:
        veredicto = "PROBABLE_DEFENSA"
    elif score >= 2:
        veredicto = "INCIERTA"
    else:
        veredicto = "PROBABLE_RATIFICACION"

    return {
        "glosa_id": glosa_id,
        "score": score,
        "veredicto": veredicto,
        "razones": razones,
    }


@router.get("/{glosa_id}/comentarios-resumen")
def comentarios_resumen(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R161 P1: resumen de comentarios de una glosa específica.

    Diferente a /glosas/{id}/comentarios/ (lista completa con
    todos los textos): aquí solo agregados rápidos:
      - cuántos comentarios
      - cuántas menciones pendientes
      - autores distintos
      - último comentario

    Útil para mostrar un badge en la card de la glosa.
    """
    from datetime import timezone

    from app.models.db import ComentarioGlosaRecord

    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    coms = (
        db.query(ComentarioGlosaRecord)
        .filter(ComentarioGlosaRecord.glosa_id == glosa_id)
        .order_by(ComentarioGlosaRecord.creado_en.desc())
        .all()
    )

    autores: set[str] = set()
    menciones_pendientes = 0
    for c in coms:
        if c.autor_email:
            autores.add(c.autor_email)
        if c.mencion and not c.resuelto:
            menciones_pendientes += 1

    ultimo = None
    if coms:
        ts = coms[0].creado_en
        if ts and ts.tzinfo is None:
            ts = ts.replace(tzinfo=timezone.utc)
        if ts:
            ultimo = ts.isoformat()

    return {
        "glosa_id": glosa_id,
        "total_comentarios": len(coms),
        "autores_distintos": sorted(autores),
        "menciones_pendientes": menciones_pendientes,
        "ultimo_comentario_en": ultimo,
    }


@router.get("/{glosa_id}/conceptos-resumen")
def conceptos_resumen(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R139 P2: resumen agregado de los conceptos múltiples de una glosa.

    Una glosa puede agrupar N conceptos (ConceptoGlosaRecord), cada
    uno con su propio CUPS, valor objetado y observación EPS.

    Este endpoint da el "TL;DR" de los conceptos:
      - cuántos hay
      - valor total objetado (suma de los valores de conceptos)
      - cuántos ya tienen dictamen_html (respondidos)
      - distribución por código_glosa
      - centros de costo distintos

    Útil para resumir glosas grandes con muchos conceptos.
    """
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    conceptos = db.query(ConceptoGlosaRecord).filter(ConceptoGlosaRecord.glosa_id == glosa_id).all()

    if not conceptos:
        return {
            "glosa_id": glosa_id,
            "total_conceptos": 0,
            "valor_objetado_conceptos": 0,
            "respondidos": 0,
            "pendientes": 0,
            "por_codigo_glosa": {},
            "centros_costo_distintos": [],
        }

    valor_total = sum(float(c.valor_objetado or 0) for c in conceptos)
    respondidos = sum(1 for c in conceptos if c.dictamen_html and len(c.dictamen_html) > 50)

    por_codigo: dict[str, int] = {}
    centros: set[str] = set()
    for c in conceptos:
        if c.codigo_glosa:
            por_codigo[c.codigo_glosa] = por_codigo.get(c.codigo_glosa, 0) + 1
        if c.centro_costo:
            centros.add(c.centro_costo)

    return {
        "glosa_id": glosa_id,
        "total_conceptos": len(conceptos),
        "valor_objetado_conceptos": int(valor_total),
        "respondidos": respondidos,
        "pendientes": len(conceptos) - respondidos,
        "por_codigo_glosa": por_codigo,
        "centros_costo_distintos": sorted(centros),
    }


@router.get("/{glosa_id}/historial-workflow")
def historial_workflow(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R134 P2: historial específico de cambios de workflow_state.

    Filtra audit_log a los eventos de cambio de workflow_state /
    estado para una glosa, mostrando solo las transiciones de
    máquina de estados (no toda la auditoría).

    Útil para responder: "¿cómo evolucionó el estado de esta glosa?"

    Devuelve transiciones ordenadas ASC por timestamp:
      [{"timestamp", "usuario", "valor_anterior", "valor_nuevo",
        "accion"}]
    """
    from app.models.db import AuditLogRecord

    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    eventos = (
        db.query(AuditLogRecord)
        .filter(AuditLogRecord.tabla == "glosas")
        .filter(AuditLogRecord.registro_id == glosa_id)
        .filter(AuditLogRecord.campo.in_(["estado", "workflow_state"]))
        .order_by(AuditLogRecord.timestamp.asc())
        .all()
    )

    items = [
        {
            "timestamp": (e.timestamp.isoformat() if e.timestamp else None),
            "usuario": e.usuario_email,
            "campo": e.campo,
            "valor_anterior": e.valor_anterior,
            "valor_nuevo": e.valor_nuevo,
            "accion": e.accion,
        }
        for e in eventos
    ]

    return {
        "glosa_id": glosa_id,
        "estado_actual": glosa.estado,
        "workflow_state_actual": glosa.workflow_state,
        "total_transiciones": len(items),
        "items": items,
    }


@router.get("/{glosa_id}/comparar-con-promedio")
def comparar_con_promedio(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R133 P2: compara una glosa con el promedio histórico de su
    cohorte (mismo EPS + mismo codigo_glosa).

    Útil para responder: "¿es esta glosa típica o atípica?"

    Si el valor objetado es 5x el promedio del cohorte, podría
    indicar:
      - Caso extraordinario que requiere atención senior
      - Posible error de captura de datos
      - Glosa fraccionada (mala práctica EPS)

    Devuelve:
      - glosa: valor_objetado, dias_restantes
      - cohorte: count, valor_promedio, valor_mediano,
                 tasa_levantamiento_pct
      - posicion: percentil aproximado del valor en el cohorte
      - flags: {valor_atipico, vencimiento_atipico}
    """
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    if not glosa.eps or not glosa.codigo_glosa:
        return {
            "glosa_id": glosa_id,
            "razon_no_evaluable": "Glosa sin EPS o codigo_glosa",
        }

    cohorte = (
        db.query(GlosaRecord)
        .filter(GlosaRecord.eps == glosa.eps)
        .filter(GlosaRecord.codigo_glosa == glosa.codigo_glosa)
        .filter(GlosaRecord.id != glosa_id)
        .all()
    )

    if not cohorte:
        return {
            "glosa_id": glosa_id,
            "razon_no_evaluable": (
                f"No hay otras glosas con eps={glosa.eps} y codigo={glosa.codigo_glosa}"
            ),
        }

    valores = sorted(float(g.valor_objetado or 0) for g in cohorte)
    n = len(valores)
    valor_glosa = float(glosa.valor_objetado or 0)

    valor_promedio = sum(valores) / n
    if n % 2 == 0:
        valor_mediano = (valores[n // 2 - 1] + valores[n // 2]) / 2
    else:
        valor_mediano = valores[n // 2]

    decididas = [
        g for g in cohorte if (g.estado or "").upper() in {"LEVANTADA", "ACEPTADA", "RATIFICADA"}
    ]
    levantadas = [g for g in decididas if (g.estado or "").upper() == "LEVANTADA"]
    tasa = round(100 * len(levantadas) / len(decididas), 2) if decididas else 0.0

    # Percentil aproximado
    menores = sum(1 for v in valores if v < valor_glosa)
    percentil = round(100 * menores / n, 1)

    valor_atipico = (
        (valor_glosa > 3 * valor_promedio or valor_glosa < valor_promedio / 5)
        if valor_promedio > 0
        else False
    )

    return {
        "glosa_id": glosa_id,
        "glosa": {
            "eps": glosa.eps,
            "codigo_glosa": glosa.codigo_glosa,
            "valor_objetado": valor_glosa,
            "dias_restantes": glosa.dias_restantes,
        },
        "cohorte": {
            "count": n,
            "valor_promedio": round(valor_promedio, 2),
            "valor_mediano": round(valor_mediano, 2),
            "tasa_levantamiento_pct": tasa,
        },
        "posicion": {
            "percentil_valor": percentil,
            "ratio_vs_promedio": round(
                valor_glosa / valor_promedio,
                2,
            )
            if valor_promedio
            else None,
        },
        "flags": {
            "valor_atipico": valor_atipico,
        },
    }


@router.get("/{glosa_id}/recomendaciones")
def recomendaciones_glosa(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R111 P1: sugerencias heurísticas de próximas acciones.

    Sin IA — usa reglas determinísticas basadas en el estado actual
    de la glosa. Útil para guiar al auditor: "¿qué debería hacer
    a continuación?".

    Devuelve lista de recomendaciones con prioridad y descripción:
      - HIGH: vencidas, sin dictamen
      - MEDIUM: sin gestor, datos incompletos
      - LOW: enriquecer información

    Cada recomendación tiene: {prioridad, accion, descripcion, endpoint?}
    """
    ESTADOS_CERRADOS = {"ACEPTADA", "LEVANTADA", "ARCHIVADA", "CONCILIADA"}

    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    estado = (glosa.estado or "").upper()
    cerrada = estado in ESTADOS_CERRADOS
    recomendaciones = []

    if cerrada:
        recomendaciones.append(
            {
                "prioridad": "INFO",
                "accion": "ARCHIVAR",
                "descripcion": "Glosa cerrada — sin acciones pendientes.",
            }
        )
        return {
            "glosa_id": glosa_id,
            "total": len(recomendaciones),
            "items": recomendaciones,
        }

    # ── Reglas críticas ───────────────────────────────────────
    dr = glosa.dias_restantes if glosa.dias_restantes is not None else 0
    if dr < 0:
        recomendaciones.append(
            {
                "prioridad": "HIGH",
                "accion": "ATENDER_VENCIDA",
                "descripcion": (
                    f"Glosa vencida hace {abs(dr)} días. Responder "
                    "urgentemente para evitar ratificación automática."
                ),
            }
        )
    elif dr <= 3:
        recomendaciones.append(
            {
                "prioridad": "HIGH",
                "accion": "ATENDER_CRITICA",
                "descripcion": f"Faltan {dr} días para vencimiento.",
            }
        )

    if not glosa.dictamen or len(glosa.dictamen) < 50:
        recomendaciones.append(
            {
                "prioridad": "HIGH",
                "accion": "GENERAR_DICTAMEN",
                "descripcion": "No hay dictamen generado. Usar IA para crear uno.",
                "endpoint": f"POST /glosas/{glosa_id}/refinar",
            }
        )

    # ── Reglas medias ─────────────────────────────────────────
    if not glosa.gestor_nombre:
        recomendaciones.append(
            {
                "prioridad": "MEDIUM",
                "accion": "ASIGNAR_GESTOR",
                "descripcion": "Glosa sin gestor asignado.",
                "endpoint": f"PATCH /glosas/{glosa_id}/asignar",
            }
        )

    if not glosa.factura or glosa.factura == "N/A":
        recomendaciones.append(
            {
                "prioridad": "MEDIUM",
                "accion": "COMPLETAR_FACTURA",
                "descripcion": "Falta número de factura.",
            }
        )

    if not glosa.texto_glosa_original:
        recomendaciones.append(
            {
                "prioridad": "MEDIUM",
                "accion": "CAPTURAR_TEXTO_ORIGINAL",
                "descripcion": "Sin texto original — el contexto IA será débil.",
            }
        )

    # ── Reglas bajas ──────────────────────────────────────────
    if not glosa.cups_servicio:
        recomendaciones.append(
            {
                "prioridad": "LOW",
                "accion": "AGREGAR_CUPS",
                "descripcion": "Sin código CUPS — útil para validación normativa.",
            }
        )

    if not recomendaciones:
        recomendaciones.append(
            {
                "prioridad": "INFO",
                "accion": "MONITOREAR",
                "descripcion": "Glosa en buen estado — esperar respuesta EPS.",
            }
        )

    return {
        "glosa_id": glosa_id,
        "estado_actual": glosa.estado,
        "total": len(recomendaciones),
        "items": recomendaciones,
    }


@router.get("/{glosa_id}/resumen-pdf")
def resumen_pdf_glosa(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R119 P1: PDF de 1 página con resumen ejecutivo de una glosa.

    Útil para imprimir/adjuntar a expedientes físicos sin tener
    que armar el PDF manualmente.

    Contenido:
      - Header con logo HUS (texto)
      - Datos clave: id, EPS, factura, valor objetado
      - Estado y SLA
      - Resumen del dictamen (primeros 1500 chars)
      - Footer con fecha de generación + auditor

    Usa reportlab (ya instalado en el stack).
    """
    import io

    from fastapi.responses import StreamingResponse
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )
    from reportlab.lib import colors
    from reportlab.lib.units import inch

    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=letter,
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
    )
    styles = getSampleStyleSheet()
    story = []

    # Header
    story.append(
        Paragraph(
            f"<b>RESUMEN GLOSA #{glosa_id} — HUS</b>",
            styles["Title"],
        )
    )
    story.append(Spacer(1, 0.2 * inch))

    # Datos clave (tabla)
    valor_obj = float(glosa.valor_objetado or 0)
    valor_rec = float(glosa.valor_recuperado or 0)
    datos = [
        ["EPS", glosa.eps or "-"],
        ["Factura", glosa.factura or "-"],
        ["Código glosa", glosa.codigo_glosa or "-"],
        ["Valor objetado", f"${valor_obj:,.0f} COP"],
        ["Valor recuperado", f"${valor_rec:,.0f} COP"],
        ["Estado", glosa.estado or "-"],
        ["Etapa", glosa.etapa or "-"],
        ["Días restantes", str(glosa.dias_restantes or "-")],
        ["Gestor", glosa.gestor_nombre or "-"],
        ["Decisión EPS", glosa.decision_eps or "Pendiente"],
    ]
    tabla = Table(datos, colWidths=[2.2 * inch, 4.5 * inch])
    tabla.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#0B5D8A")),
                ("TEXTCOLOR", (0, 0), (0, -1), colors.white),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("PADDING", (0, 0), (-1, -1), 6),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]
        )
    )
    story.append(tabla)
    story.append(Spacer(1, 0.3 * inch))

    # Resumen del dictamen
    if glosa.dictamen:
        story.append(Paragraph("<b>Dictamen HUS:</b>", styles["Heading3"]))
        # Limpiar HTML básico
        import re

        texto_dict = re.sub(r"<[^>]+>", " ", glosa.dictamen)
        texto_dict = re.sub(r"\s+", " ", texto_dict).strip()
        story.append(Paragraph(texto_dict[:1500], styles["BodyText"]))
        story.append(Spacer(1, 0.2 * inch))

    # Footer
    story.append(Spacer(1, 0.3 * inch))
    story.append(
        Paragraph(
            f"<i>Generado por {current_user.email} el {ahora_utc().strftime('%Y-%m-%d %H:%M UTC')}</i>",
            styles["BodyText"],
        )
    )

    doc.build(story)
    buf.seek(0)

    fname = f"resumen-glosa-{glosa_id}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/{glosa_id}/exportar-evidencia.zip")
def exportar_evidencia_zip(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R108 P1: paquete ZIP completo de evidencia para una glosa.

    Complementa /paquete-evidencia.json (solo datos) con un ZIP
    multi-archivo listo para entregar a legal/compliance:

      glosa.json      — todos los datos de la glosa
      dictamen.txt    — texto plano del dictamen
      audit_log.json  — eventos de auditoría asociados
      README.txt      — explicación del contenido

    StreamingResponse con el ZIP en memoria (suficiente para
    glosas individuales — el límite es ~10MB típicamente).
    """
    import io
    import json
    import zipfile

    from fastapi.responses import StreamingResponse

    from app.models.db import AuditLogRecord

    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    # Datos de la glosa (campos públicos)
    glosa_dict = {
        "id": glosa.id,
        "creado_en": glosa.creado_en.isoformat() if glosa.creado_en else None,
        "eps": glosa.eps,
        "paciente": glosa.paciente,
        "factura": glosa.factura,
        "codigo_glosa": glosa.codigo_glosa,
        "valor_objetado": float(glosa.valor_objetado or 0),
        "valor_recuperado": float(glosa.valor_recuperado or 0),
        "etapa": glosa.etapa,
        "estado": glosa.estado,
        "decision_eps": glosa.decision_eps,
        "gestor_nombre": glosa.gestor_nombre,
        "auditor_email": glosa.auditor_email,
        "fecha_vencimiento": (
            glosa.fecha_vencimiento.isoformat() if glosa.fecha_vencimiento else None
        ),
    }

    # Audit log
    eventos = (
        db.query(AuditLogRecord)
        .filter(AuditLogRecord.tabla == "glosas")
        .filter(AuditLogRecord.registro_id == glosa_id)
        .order_by(AuditLogRecord.timestamp.asc())
        .all()
    )
    audit_list = [
        {
            "timestamp": e.timestamp.isoformat() if e.timestamp else None,
            "usuario_email": e.usuario_email,
            "accion": e.accion,
            "campo": e.campo,
            "valor_anterior": e.valor_anterior,
            "valor_nuevo": e.valor_nuevo,
        }
        for e in eventos
    ]

    readme = (
        f"PAQUETE DE EVIDENCIA — GLOSA #{glosa_id}\n"
        f"Generado: {ahora_utc().isoformat()}\n"
        f"Generado por: {current_user.email}\n\n"
        "Contenido:\n"
        "  - glosa.json: datos estructurados completos\n"
        "  - dictamen.txt: texto del dictamen HUS (si existe)\n"
        "  - audit_log.json: histórico de eventos sobre esta glosa\n"
    )

    # Construir ZIP en memoria
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("README.txt", readme)
        zf.writestr("glosa.json", json.dumps(glosa_dict, ensure_ascii=False, indent=2))
        zf.writestr("audit_log.json", json.dumps(audit_list, ensure_ascii=False, indent=2))
        if glosa.dictamen:
            zf.writestr("dictamen.txt", glosa.dictamen)

    buf.seek(0)
    fname = f"evidencia-glosa-{glosa_id}.zip"
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/{glosa_id}/duplicados-potenciales")
def duplicados_potenciales_glosa(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R99 P2: detecta posibles duplicados de UNA glosa específica.

    Diferente a /glosas/duplicados (que escanea TODAS): este se enfoca
    en una glosa puntual. Útil para validar al crear/clonar:
      "¿Estoy generando un duplicado?"

    Heurística (DEBE coincidir TODO):
      - Misma EPS
      - Misma factura (no N/A)
      - Mismo codigo_glosa
      - Diferencia de valor_objetado < 1% (tolera redondeos)

    Excluye la propia glosa. Devuelve hasta 20 candidatas con
    score de similitud (0-100).
    """
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    # Sin factura "real" no hay forma de identificar duplicado fiable
    if not glosa.factura or glosa.factura == "N/A":
        return {
            "glosa_id": glosa_id,
            "razon_no_evaluable": "factura ausente o N/A",
            "candidatos": [],
        }

    candidatos = (
        db.query(GlosaRecord)
        .filter(GlosaRecord.id != glosa_id)
        .filter(GlosaRecord.eps == glosa.eps)
        .filter(GlosaRecord.factura == glosa.factura)
        .filter(GlosaRecord.codigo_glosa == glosa.codigo_glosa)
        .limit(20)
        .all()
    )

    valor_origen = float(glosa.valor_objetado or 0)
    items = []
    for c in candidatos:
        v = float(c.valor_objetado or 0)
        # Score: 100 si valores idénticos, decrece según diferencia relativa
        if valor_origen == 0 and v == 0:
            score = 100.0
        elif valor_origen == 0 or v == 0:
            score = 50.0  # Uno tiene valor y otro no — sospechoso pero menos
        else:
            diff_pct = abs(v - valor_origen) / max(v, valor_origen) * 100
            score = round(max(0, 100 - diff_pct), 2)

        items.append(
            {
                "id": c.id,
                "creado_en": (c.creado_en.isoformat() if c.creado_en else None),
                "valor_objetado": v,
                "estado": c.estado,
                "score_similitud": score,
            }
        )

    items.sort(key=lambda x: x["score_similitud"], reverse=True)

    return {
        "glosa_id": glosa_id,
        "total_candidatos": len(items),
        "candidatos": items,
    }


@router.get("/{glosa_id}/relacionadas")
def glosas_relacionadas(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R93 P2: glosas relacionadas a una glosa dada.

    Identifica vínculos por:
      - Misma factura: glosas que objetan distintos servicios de
        la misma factura (suelen ir/venir juntas en el ciclo)
      - Mismo paciente: histórico clínico-administrativo del paciente
      - Mismo código_glosa + misma EPS: patrones repetidos

    Devuelve cada grupo limitado a 10 entradas para no inflar
    el response. Ordenado DESC por creado_en (más reciente primero).
    """
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    def _serializar(g):
        return {
            "id": g.id,
            "creado_en": g.creado_en.isoformat() if g.creado_en else None,
            "eps": g.eps,
            "factura": g.factura,
            "codigo_glosa": g.codigo_glosa,
            "valor_objetado": float(g.valor_objetado or 0),
            "estado": g.estado,
            "etapa": g.etapa,
        }

    LIMITE = 10

    # Misma factura (excluyendo la glosa actual)
    misma_factura = []
    if glosa.factura and glosa.factura != "N/A":
        misma_factura = (
            db.query(GlosaRecord)
            .filter(GlosaRecord.factura == glosa.factura)
            .filter(GlosaRecord.id != glosa_id)
            .order_by(GlosaRecord.creado_en.desc())
            .limit(LIMITE)
            .all()
        )

    # Mismo paciente
    mismo_paciente = []
    if glosa.paciente:
        mismo_paciente = (
            db.query(GlosaRecord)
            .filter(GlosaRecord.paciente == glosa.paciente)
            .filter(GlosaRecord.id != glosa_id)
            .order_by(GlosaRecord.creado_en.desc())
            .limit(LIMITE)
            .all()
        )

    # Mismo código + misma EPS (patrones repetidos)
    mismo_patron = []
    if glosa.codigo_glosa and glosa.eps:
        mismo_patron = (
            db.query(GlosaRecord)
            .filter(GlosaRecord.codigo_glosa == glosa.codigo_glosa)
            .filter(GlosaRecord.eps == glosa.eps)
            .filter(GlosaRecord.id != glosa_id)
            .order_by(GlosaRecord.creado_en.desc())
            .limit(LIMITE)
            .all()
        )

    return {
        "glosa_id": glosa_id,
        "misma_factura": [_serializar(g) for g in misma_factura],
        "mismo_paciente": [_serializar(g) for g in mismo_paciente],
        "mismo_codigo_y_eps": [_serializar(g) for g in mismo_patron],
        "limite_por_grupo": LIMITE,
    }


@router.get("/{glosa_id}/diff/{otra_id}")
def diff_entre_glosas(
    glosa_id: int,
    otra_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R93 P1: comparativa lado-a-lado entre dos glosas.

    Útil para:
      - Identificar casos similares con mismo código pero distinto
        outcome (¿por qué SANITAS aceptó la TA0201 de Pedro pero
        ratificó la de Juan?)
      - Entrenar a auditores nuevos mostrando ejemplos contrastantes
      - Investigar inconsistencias en decisiones EPS

    Devuelve los campos clave de ambas glosas y un set de campos
    "diferentes" para destacar.
    """
    repo = GlosaRepository(db)
    g1 = repo.obtener_por_id(glosa_id)
    if not g1:
        raise HTTPException(404, f"Glosa {glosa_id} no encontrada")
    g2 = repo.obtener_por_id(otra_id)
    if not g2:
        raise HTTPException(404, f"Glosa {otra_id} no encontrada")

    CAMPOS = [
        "eps",
        "paciente",
        "factura",
        "codigo_glosa",
        "valor_objetado",
        "valor_aceptado",
        "valor_recuperado",
        "etapa",
        "estado",
        "decision_eps",
        "gestor_nombre",
        "cups_servicio",
        "codigo_respuesta",
    ]

    snapshot1 = {c: getattr(g1, c, None) for c in CAMPOS}
    snapshot2 = {c: getattr(g2, c, None) for c in CAMPOS}

    diferentes = sorted(c for c in CAMPOS if snapshot1.get(c) != snapshot2.get(c))

    # Casteamos floats para serialización JSON consistente
    def _normalizar(d: dict) -> dict:
        out = {}
        for k, v in d.items():
            if isinstance(v, (int, float)) and v is not None:
                out[k] = float(v) if isinstance(v, float) else v
            else:
                out[k] = v if v is not None else None
        return out

    return {
        "glosa_a": {"id": g1.id, **_normalizar(snapshot1)},
        "glosa_b": {"id": g2.id, **_normalizar(snapshot2)},
        "campos_diferentes": diferentes,
        "total_diferencias": len(diferentes),
    }


@router.get("/{glosa_id}/sla")
def sla_glosa(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R92 P2: estado SLA detallado de una glosa individual.

    Útil para el panel de detalle: muestra de un vistazo si esta
    glosa específica está cumpliendo el SLA o está en riesgo.

    Devuelve:
      - estado_sla: VENCIDA | CRITICA | EN_TIEMPO |
                    CERRADA_A_TIEMPO | CERRADA_TARDE | SIN_VENCIMIENTO
      - color_semaforo: ROJO | AMARILLO | VERDE | NEGRO | GRIS
      - dias_restantes
      - dias_transcurridos (desde creación)
      - fecha_creado / fecha_vencimiento / fecha_decision_eps
      - tiempo_total_resolucion_dias (si cerrada)
      - cerrada (bool)
    """
    from datetime import timezone

    ESTADOS_CERRADOS = {"ACEPTADA", "LEVANTADA", "ARCHIVADA", "CONCILIADA"}

    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    ahora = ahora_utc()
    estado = (glosa.estado or "").upper()
    cerrada = estado in ESTADOS_CERRADOS

    creado = glosa.creado_en
    if creado and creado.tzinfo is None:
        creado = creado.replace(tzinfo=timezone.utc)
    venc = glosa.fecha_vencimiento
    if venc and venc.tzinfo is None:
        venc = venc.replace(tzinfo=timezone.utc)
    dec = glosa.fecha_decision_eps
    if dec and dec.tzinfo is None:
        dec = dec.replace(tzinfo=timezone.utc)

    dias_transcurridos = (ahora - creado).days if creado else None

    tiempo_total = None
    if cerrada and dec and creado:
        tiempo_total = (dec - creado).days

    # Determinar estado_sla
    if not venc:
        estado_sla = "SIN_VENCIMIENTO"
        color = "GRIS"
    elif cerrada:
        if dec and dec <= venc:
            estado_sla = "CERRADA_A_TIEMPO"
            color = "VERDE"
        else:
            estado_sla = "CERRADA_TARDE"
            color = "NEGRO"
    else:
        dr = glosa.dias_restantes if glosa.dias_restantes is not None else 0
        if dr < 0:
            estado_sla = "VENCIDA"
            color = "ROJO"
        elif dr <= 3:
            estado_sla = "CRITICA"
            color = "AMARILLO"
        else:
            estado_sla = "EN_TIEMPO"
            color = "VERDE"

    return {
        "glosa_id": glosa_id,
        "estado": glosa.estado,
        "cerrada": cerrada,
        "estado_sla": estado_sla,
        "color_semaforo": color,
        "dias_restantes": glosa.dias_restantes,
        "dias_transcurridos": dias_transcurridos,
        "fecha_creado": creado.isoformat() if creado else None,
        "fecha_vencimiento": venc.isoformat() if venc else None,
        "fecha_decision_eps": dec.isoformat() if dec else None,
        "tiempo_total_resolucion_dias": tiempo_total,
    }


@router.get("/{glosa_id}/borrador-respuesta")
def borrador_respuesta(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R395 P1: borrador automático de respuesta basado en mejor
    caso similar.

    Busca el caso LEVANTADO más reciente del par
    (eps, codigo_glosa) y devuelve su dictamen como
    borrador inicial. Útil para que el gestor parta de
    un texto que ya funcionó en lugar de página en
    blanco.

    No reemplaza el dictamen actual, solo SUGIERE.
    """
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    # Buscar el mejor caso histórico con dictamen
    candidato = (
        db.query(GlosaRecord)
        .filter(GlosaRecord.eps == glosa.eps)
        .filter(GlosaRecord.codigo_glosa == glosa.codigo_glosa)
        .filter(GlosaRecord.id != glosa.id)
        .filter(GlosaRecord.estado == "LEVANTADA")
        .filter(GlosaRecord.dictamen.isnot(None))
        .order_by(GlosaRecord.fecha_decision_eps.desc())
        .first()
    )

    if not candidato or not (candidato.dictamen or "").strip():
        return {
            "glosa_id": glosa.id,
            "eps": glosa.eps,
            "codigo_glosa": glosa.codigo_glosa,
            "borrador_disponible": False,
            "mensaje": (
                "No hay caso similar levantado con dictamen "
                "para usar como base. Redacta desde cero."
            ),
        }

    # Adaptar el texto: solo tomamos el dictamen como punto
    # de partida, no se reemplaza nada automáticamente
    texto = (candidato.dictamen or "").strip()
    # Limitar tamaño para que sea manejable
    if len(texto) > 3000:
        texto = texto[:3000] + "\n\n[…borrador truncado, ver caso original]"

    return {
        "glosa_id": glosa.id,
        "eps": glosa.eps,
        "codigo_glosa": glosa.codigo_glosa,
        "borrador_disponible": True,
        "fuente_caso_id": candidato.id,
        "fuente_fecha_decision": (
            candidato.fecha_decision_eps.isoformat() if candidato.fecha_decision_eps else None
        ),
        "fuente_resultado": "LEVANTADA",
        "borrador": texto,
        "mensaje": (
            "Texto base tomado del caso anterior LEVANTADO "
            "más reciente del mismo par (EPS, código). "
            "Adáptalo a este caso antes de enviar."
        ),
    }


@router.get("/{glosa_id}/playbook")
def playbook_glosa(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R375 P1: playbook táctico para esta glosa.

    Combina señales y devuelve un plan de acción
    accionable: tono, próximo paso, riesgo, recomendaciones.
    """
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    ESTADOS_DECIDIDOS = {"LEVANTADA", "ACEPTADA", "RATIFICADA"}
    eps = (glosa.eps or "").strip()
    cod = (glosa.codigo_glosa or "").strip()

    pares = (
        db.query(GlosaRecord)
        .filter(GlosaRecord.eps == eps)
        .filter(GlosaRecord.codigo_glosa == cod)
        .filter(GlosaRecord.id != glosa.id)
        .filter(GlosaRecord.estado.in_(ESTADOS_DECIDIDOS))
        .all()
    )
    n_par = len(pares)
    lev_par = sum(1 for g in pares if (g.estado or "").upper() == "LEVANTADA")
    tasa_par = (100.0 * lev_par / n_par) if n_par else None

    eps_q = (
        db.query(GlosaRecord)
        .filter(GlosaRecord.eps == eps)
        .filter(GlosaRecord.id != glosa.id)
        .filter(GlosaRecord.estado.in_(ESTADOS_DECIDIDOS))
        .all()
    )
    n_eps = len(eps_q)
    lev_eps = sum(1 for g in eps_q if (g.estado or "").upper() == "LEVANTADA")
    tasa_eps_global = (100.0 * lev_eps / n_eps) if n_eps else None

    # Tono
    if tasa_par is not None and tasa_par >= 70:
        tono = "conciliador"
        tono_motivo = (
            "Históricamente esta EPS levanta este código — tono profesional y directo basta."
        )
    elif tasa_par is not None and tasa_par <= 30:
        tono = "firme"
        tono_motivo = "Histórico desfavorable — argumenta con técnica fuerte y respaldo normativo."
    else:
        tono = "neutral"
        tono_motivo = "Sin patrón claro — defensa técnico-jurídica estándar."

    # Próximo paso
    dr = glosa.dias_restantes if glosa.dias_restantes is not None else None
    estado_actual = (glosa.estado or "").upper()
    if estado_actual in ("LEVANTADA", "ACEPTADA", "RATIFICADA", "ARCHIVADA", "CONCILIADA"):
        proximo = "Glosa cerrada — solo lectura"
    elif estado_actual == "RESPONDIDA":
        proximo = "Esperando decisión de EPS — monitorear fecha_decision_eps"
    elif dr is not None and dr < 0:
        proximo = "URGENTE: glosa vencida. Ejecutar respuesta definitiva HOY o se archiva."
    elif dr is not None and dr <= 3:
        proximo = "Crítico (≤3 días): redactar dictamen y enviar respuesta antes de cierre."
    elif not (glosa.dictamen or "").strip():
        proximo = "Redactar dictamen técnico-jurídico — comienza revisando casos similares."
    elif len(glosa.dictamen or "") < 200:
        proximo = "Reforzar dictamen actual con normativa específica antes de enviar."
    else:
        proximo = "Revisar dictamen y enviar respuesta a EPS."

    # Riesgo
    valor = float(glosa.valor_objetado or 0)
    riesgo_score = 0
    razones_riesgo = []
    if dr is not None and dr < 0:
        riesgo_score += 40
        razones_riesgo.append("vencida")
    elif dr is not None and dr <= 3:
        riesgo_score += 25
        razones_riesgo.append("vence en 3 días o menos")
    if tasa_par is not None and tasa_par < 30:
        riesgo_score += 25
        razones_riesgo.append(f"tasa par baja ({tasa_par:.0f}%)")
    if valor >= 10_000_000:
        riesgo_score += 20
        razones_riesgo.append("alto valor (>10M)")
    if not (glosa.dictamen or "").strip():
        riesgo_score += 10
        razones_riesgo.append("sin dictamen")

    if riesgo_score >= 60:
        riesgo_nivel = "ALTO"
    elif riesgo_score >= 30:
        riesgo_nivel = "MEDIO"
    else:
        riesgo_nivel = "BAJO"

    # Recomendaciones
    recomendaciones = []
    if tasa_par is not None and tasa_par >= 60 and not (glosa.dictamen or "").strip():
        recomendaciones.append(
            "🎯 Caso favorable según histórico: redactar dictamen y cerrar pronto."
        )
    if valor >= 10_000_000 and not (glosa.dictamen or "").strip():
        recomendaciones.append(
            "💰 Alto valor sin dictamen — enfoque cuidadoso en respaldo normativo."
        )
    if dr is not None and dr < 0:
        recomendaciones.append("🚨 Vencida — riesgo de archivo automático.")
    if (
        tasa_par is not None
        and tasa_par < 30
        and tasa_eps_global is not None
        and tasa_eps_global < 30
    ):
        recomendaciones.append("🤝 Considera conciliación bilateral antes de ratificación.")

    return {
        "glosa_id": glosa.id,
        "eps": eps,
        "codigo_glosa": cod,
        "estado_actual": estado_actual,
        "tono_recomendado": tono,
        "tono_motivo": tono_motivo,
        "proximo_paso": proximo,
        "riesgo": {
            "nivel": riesgo_nivel,
            "score": riesgo_score,
            "razones": razones_riesgo,
        },
        "tasa_par_pct": (round(tasa_par, 2) if tasa_par is not None else None),
        "tasa_eps_global_pct": (round(tasa_eps_global, 2) if tasa_eps_global is not None else None),
        "n_par": n_par,
        "n_eps_global": n_eps,
        "recomendaciones": recomendaciones,
    }


@router.get("/{glosa_id}/eps-comportamiento")
def eps_comportamiento(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R374 P1: perfil de comportamiento de la EPS de esta glosa.

    Resumen rápido para el gestor: cómo se comporta esta
    EPS en general (no solo este código). Útil para
    calibrar el tono y la estrategia.

    Devuelve:
      - tasa_levantamiento_global_pct (cuántas levantamos
        de las decididas con esta EPS)
      - tiempo_promedio_decision_dias (qué tan rápido
        decide)
      - codigos_top_3 (los 3 códigos que más usa)
      - codigos_respuesta_top_3
      - estilo_resumen: una etiqueta legible
        ("EPS difícil", "EPS conciliadora", etc.)
    """
    from datetime import timezone

    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    eps = (glosa.eps or "").strip()
    if not eps:
        return {
            "glosa_id": glosa.id,
            "eps": None,
            "estilo_resumen": "Sin EPS",
        }

    rows = (
        db.query(GlosaRecord)
        .filter(GlosaRecord.eps == eps)
        .filter(GlosaRecord.id != glosa.id)
        .all()
    )

    ESTADOS_DECIDIDOS = {"LEVANTADA", "ACEPTADA", "RATIFICADA"}

    decididas = [g for g in rows if (g.estado or "").upper() in ESTADOS_DECIDIDOS]
    n_dec = len(decididas)
    n_lev = sum(1 for g in decididas if (g.estado or "").upper() == "LEVANTADA")
    tasa_lev = round(100 * n_lev / n_dec, 2) if n_dec else 0.0

    # Tiempo promedio de decisión
    tiempos = []
    for g in decididas:
        cre = g.creado_en
        dec = g.fecha_decision_eps
        if cre and cre.tzinfo is None:
            cre = cre.replace(tzinfo=timezone.utc)
        if dec and dec.tzinfo is None:
            dec = dec.replace(tzinfo=timezone.utc)
        if cre and dec:
            d = (dec - cre).days
            if d >= 0:
                tiempos.append(d)
    tiempo_prom = round(sum(tiempos) / len(tiempos), 2) if tiempos else None

    # Top códigos de glosa que usa la EPS
    codigos: dict[str, int] = {}
    for g in rows:
        c = (g.codigo_glosa or "").strip()
        if c:
            codigos[c] = codigos.get(c, 0) + 1
    top_codigos = sorted(
        codigos.items(),
        key=lambda x: x[1],
        reverse=True,
    )[:3]

    # Top códigos respuesta cuando HUS responde
    cresp: dict[str, int] = {}
    for g in rows:
        cr = (g.codigo_respuesta or "").strip()
        if cr:
            cresp[cr] = cresp.get(cr, 0) + 1
    top_cresp = sorted(
        cresp.items(),
        key=lambda x: x[1],
        reverse=True,
    )[:3]

    # Estilo: combina tasa + tiempo
    if n_dec < 5:
        estilo = "Sin historial suficiente"
    elif tasa_lev >= 70:
        estilo = "EPS conciliadora — levanta la mayoría"
    elif tasa_lev >= 40:
        estilo = "EPS estándar — depende del argumento"
    elif tasa_lev >= 20:
        estilo = "EPS difícil — exige técnica fuerte"
    else:
        estilo = "EPS muy difícil — considere conciliación"

    if tiempo_prom is not None:
        if tiempo_prom <= 15:
            estilo += " · responde rápido"
        elif tiempo_prom <= 60:
            estilo += " · responde dentro del SLA"
        else:
            estilo += " · responde tarde"

    return {
        "glosa_id": glosa.id,
        "eps": eps,
        "n_glosas_historico": len(rows),
        "n_decididas": n_dec,
        "tasa_levantamiento_global_pct": tasa_lev,
        "tiempo_promedio_decision_dias": tiempo_prom,
        "codigos_top_3": [{"codigo_glosa": c, "count": n} for c, n in top_codigos],
        "codigos_respuesta_top_3": [{"codigo_respuesta": c, "count": n} for c, n in top_cresp],
        "estilo_resumen": estilo,
    }


@router.get("/{glosa_id}/asistente-ficha")
def asistente_ficha(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R370 P1: asistente inteligente al abrir una glosa.

    Single-call que devuelve TODO lo que el gestor
    necesita para decidir cómo actuar en esta glosa,
    sin que tenga que pedirlo:

      - probabilidad_levantamiento: heurística mixta
      - casos_similares: 3 casos resueltos misma EPS+código
      - codigo_respuesta_sugerido: el que mejor tasa tuvo
        en el par (eps, codigo_glosa)
      - alerta_dictamen: si está corto/vacío
      - factura_contexto: glosas hermanas en la factura
      - urgencia: nivel según dias_restantes
      - acciones_sugeridas: lista priorizada
    """
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    ESTADOS_DECIDIDOS = {"LEVANTADA", "ACEPTADA", "RATIFICADA"}
    ESTADOS_CERRADOS = {"ACEPTADA", "LEVANTADA", "ARCHIVADA", "CONCILIADA"}

    # ─── Probabilidad heurística par + gestor ─────
    pares = (
        db.query(GlosaRecord)
        .filter(GlosaRecord.eps == glosa.eps)
        .filter(GlosaRecord.codigo_glosa == glosa.codigo_glosa)
        .filter(GlosaRecord.id != glosa.id)
        .filter(GlosaRecord.estado.in_(ESTADOS_DECIDIDOS))
        .all()
    )
    n_par = len(pares)
    lev_par = sum(1 for g in pares if (g.estado or "").upper() == "LEVANTADA")
    tasa_par = round(100 * lev_par / n_par, 2) if n_par else None

    tasa_gestor = None
    n_gestor = 0
    if glosa.gestor_nombre:
        gpQ = (
            db.query(GlosaRecord)
            .filter(GlosaRecord.gestor_nombre == glosa.gestor_nombre)
            .filter(GlosaRecord.id != glosa.id)
            .filter(GlosaRecord.estado.in_(ESTADOS_DECIDIDOS))
            .all()
        )
        n_gestor = len(gpQ)
        lev_g = sum(1 for g in gpQ if (g.estado or "").upper() == "LEVANTADA")
        tasa_gestor = round(100 * lev_g / n_gestor, 2) if n_gestor else None

    if tasa_par is not None and tasa_gestor is not None:
        prob = round(tasa_par * 0.6 + tasa_gestor * 0.4, 2)
    elif tasa_par is not None:
        prob = tasa_par
    elif tasa_gestor is not None:
        prob = tasa_gestor
    else:
        prob = None

    # ─── Casos similares (3) ─────
    casos = []
    casos_q = (
        db.query(GlosaRecord)
        .filter(GlosaRecord.eps == glosa.eps)
        .filter(GlosaRecord.codigo_glosa == glosa.codigo_glosa)
        .filter(GlosaRecord.id != glosa.id)
        .filter(GlosaRecord.estado.in_(ESTADOS_DECIDIDOS))
        .order_by(GlosaRecord.fecha_decision_eps.desc())
        .limit(3)
        .all()
    )
    for c in casos_q:
        d = c.dictamen or ""
        casos.append(
            {
                "glosa_id": c.id,
                "estado": c.estado,
                "valor_objetado": int(float(c.valor_objetado or 0)),
                "valor_recuperado": int(float(c.valor_recuperado or 0)),
                "dictamen_extracto": (d[:160] + ("…" if len(d) > 160 else "")),
            }
        )

    # ─── codigo_respuesta sugerido (mejor tasa en el par) ─────
    cr_bucket: dict[str, dict] = {}
    for c in pares:
        cr = (c.codigo_respuesta or "").strip()
        if not cr:
            continue
        b = cr_bucket.setdefault(cr, {"dec": 0, "lev": 0})
        b["dec"] += 1
        if (c.estado or "").upper() == "LEVANTADA":
            b["lev"] += 1
    cr_sugerido = None
    if cr_bucket:
        ranked = []
        for cr, b in cr_bucket.items():
            if b["dec"] < 2:
                continue
            ranked.append((cr, 100 * b["lev"] / b["dec"], b["dec"]))
        if ranked:
            ranked.sort(key=lambda x: (x[1], x[2]), reverse=True)
            cr_sugerido = {
                "codigo_respuesta": ranked[0][0],
                "tasa_levantamiento_pct": round(ranked[0][1], 2),
                "muestras": ranked[0][2],
            }

    # ─── Alerta dictamen ─────
    dlen = len(glosa.dictamen or "")
    alerta_dictamen = None
    if dlen == 0:
        alerta_dictamen = (
            "No has escrito dictamen. Considera mirar los casos similares antes de redactar."
        )
    elif dlen < 50:
        alerta_dictamen = (
            f"Dictamen muy corto ({dlen} chars). "
            "Los dictámenes >200 chars históricamente ganan más."
        )

    # ─── Contexto factura ─────
    factura_ctx = None
    f = (glosa.factura or "").strip()
    if f and f != "N/A":
        hermanas = db.query(GlosaRecord).filter(GlosaRecord.factura == f).all()
        abiertas_h = sum(1 for g in hermanas if (g.estado or "").upper() not in ESTADOS_CERRADOS)
        factura_ctx = {
            "factura": f,
            "total_glosas_factura": len(hermanas),
            "glosas_abiertas": abiertas_h,
            "valor_objetado_total": int(sum(float(g.valor_objetado or 0) for g in hermanas)),
        }

    # ─── Urgencia ─────
    dr = glosa.dias_restantes if glosa.dias_restantes is not None else None
    if dr is None:
        urgencia = {"nivel": "DESCONOCIDA", "mensaje": "Sin SLA registrada"}
    elif dr < 0:
        urgencia = {
            "nivel": "VENCIDA",
            "mensaje": f"Vencida hace {abs(dr)} días",
        }
    elif dr == 0:
        urgencia = {"nivel": "HOY", "mensaje": "Vence hoy"}
    elif dr <= 3:
        urgencia = {
            "nivel": "CRITICA",
            "mensaje": f"Vence en {dr} días",
        }
    elif dr <= 7:
        urgencia = {
            "nivel": "PROXIMA",
            "mensaje": f"Vence en {dr} días",
        }
    else:
        urgencia = {
            "nivel": "NORMAL",
            "mensaje": f"Vence en {dr} días",
        }

    # ─── Acciones sugeridas (priorizadas) ─────
    acciones = []
    if (glosa.estado or "").upper() in ESTADOS_CERRADOS:
        acciones.append(
            {
                "prioridad": 1,
                "tipo": "INFO",
                "mensaje": "Esta glosa ya está cerrada — solo lectura",
            }
        )
    else:
        if dr is not None and dr < 0:
            acciones.append(
                {
                    "prioridad": 1,
                    "tipo": "URGENTE",
                    "mensaje": (f"Glosa vencida hace {abs(dr)}d. Cierra hoy."),
                }
            )
        elif dr is not None and dr <= 3:
            acciones.append(
                {
                    "prioridad": 1,
                    "tipo": "IMPORTANTE",
                    "mensaje": f"Vence en {dr}d — atender hoy",
                }
            )
        if alerta_dictamen:
            acciones.append(
                {
                    "prioridad": 2,
                    "tipo": "DICTAMEN",
                    "mensaje": alerta_dictamen,
                }
            )
        if cr_sugerido:
            acciones.append(
                {
                    "prioridad": 3,
                    "tipo": "SUGERENCIA",
                    "mensaje": (
                        f"Considera responder con "
                        f"{cr_sugerido['codigo_respuesta']} — "
                        f"{cr_sugerido['tasa_levantamiento_pct']}% de "
                        "éxito en casos iguales"
                    ),
                }
            )
        if prob is not None and prob >= 70:
            acciones.append(
                {
                    "prioridad": 4,
                    "tipo": "POSITIVA",
                    "mensaje": (f"Alta probabilidad de levantamiento ({prob}%) — vale el esfuerzo"),
                }
            )
        elif prob is not None and prob < 30:
            acciones.append(
                {
                    "prioridad": 4,
                    "tipo": "PRECAUCION",
                    "mensaje": (f"Baja probabilidad histórica ({prob}%) — considera conciliar"),
                }
            )

    return {
        "glosa_id": glosa.id,
        "eps": glosa.eps,
        "codigo_glosa": glosa.codigo_glosa,
        "estado": glosa.estado,
        "probabilidad": {
            "levantamiento_pct": prob,
            "tasa_par_eps_codigo_pct": tasa_par,
            "n_par": n_par,
            "tasa_gestor_pct": tasa_gestor,
            "n_gestor": n_gestor,
        },
        "codigo_respuesta_sugerido": cr_sugerido,
        "casos_similares": casos,
        "alerta_dictamen": alerta_dictamen,
        "factura_contexto": factura_ctx,
        "urgencia": urgencia,
        "acciones_sugeridas": acciones,
    }


@router.get("/{glosa_id}/casos-similares-resueltos")
def casos_similares_resueltos(
    glosa_id: int,
    limit: int = Query(10, ge=1, le=50),
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R349 P1: casos similares ya resueltos a esta glosa.

    Para una glosa abierta, busca otras del mismo
    (eps, codigo_glosa) que ya estén decididas. Devuelve
    sus dictámenes (extracto) y resultado para guiar
    el argumento.

    Útil cuando el gestor enfrenta un caso "típico" y
    quiere ver qué argumentos funcionaron antes.
    """
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    rows = (
        db.query(GlosaRecord)
        .filter(GlosaRecord.eps == glosa.eps)
        .filter(GlosaRecord.codigo_glosa == glosa.codigo_glosa)
        .filter(GlosaRecord.id != glosa.id)
        .filter(
            GlosaRecord.estado.in_(
                ["LEVANTADA", "ACEPTADA", "RATIFICADA"],
            )
        )
        .order_by(GlosaRecord.fecha_decision_eps.desc())
        .limit(int(limit))
        .all()
    )

    items = []
    for g in rows:
        d = g.dictamen or ""
        extracto = d[:200] + ("..." if len(d) > 200 else "")
        items.append(
            {
                "glosa_id": g.id,
                "estado": g.estado,
                "valor_objetado": int(float(g.valor_objetado or 0)),
                "valor_recuperado": int(float(g.valor_recuperado or 0)),
                "dictamen_extracto": extracto,
                "fecha_decision_eps": (
                    g.fecha_decision_eps.isoformat() if g.fecha_decision_eps else None
                ),
            }
        )

    return {
        "glosa_id": glosa.id,
        "eps": glosa.eps,
        "codigo_glosa": glosa.codigo_glosa,
        "total_casos_similares": len(items),
        "items": items,
    }


@router.get("/{glosa_id}/probabilidad-levantamiento")
def probabilidad_levantamiento(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R299 P1: estima probabilidad heurística de levantamiento.

    Combina dos señales históricas:
      - tasa_eps_codigo: tasa de levantamiento histórica
        para el par (eps, codigo_glosa) de esta glosa
      - tasa_gestor: tasa histórica del gestor asignado

    Probabilidad final = promedio ponderado (60% par,
    40% gestor) si ambos están disponibles. Si falta uno,
    usa el otro al 100%.

    Devuelve también las muestras (n) usadas para
    confianza.
    """
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    ESTADOS_DECIDIDOS = {"LEVANTADA", "ACEPTADA", "RATIFICADA"}

    # tasa par (eps, codigo)
    par_query = (
        db.query(GlosaRecord)
        .filter(GlosaRecord.eps == glosa.eps)
        .filter(GlosaRecord.codigo_glosa == glosa.codigo_glosa)
        .filter(GlosaRecord.id != glosa.id)
        .filter(GlosaRecord.estado.in_(ESTADOS_DECIDIDOS))
        .all()
    )
    n_par = len(par_query)
    lev_par = sum(1 for g in par_query if (g.estado or "").upper() == "LEVANTADA")
    tasa_par = round(100 * lev_par / n_par, 2) if n_par else None

    # tasa gestor
    tasa_gestor = None
    n_gestor = 0
    if glosa.gestor_nombre:
        gest_query = (
            db.query(GlosaRecord)
            .filter(GlosaRecord.gestor_nombre == glosa.gestor_nombre)
            .filter(GlosaRecord.id != glosa.id)
            .filter(GlosaRecord.estado.in_(ESTADOS_DECIDIDOS))
            .all()
        )
        n_gestor = len(gest_query)
        lev_gestor = sum(1 for g in gest_query if (g.estado or "").upper() == "LEVANTADA")
        tasa_gestor = round(100 * lev_gestor / n_gestor, 2) if n_gestor else None

    if tasa_par is not None and tasa_gestor is not None:
        prob = round(tasa_par * 0.6 + tasa_gestor * 0.4, 2)
    elif tasa_par is not None:
        prob = tasa_par
    elif tasa_gestor is not None:
        prob = tasa_gestor
    else:
        prob = None

    return {
        "glosa_id": glosa.id,
        "eps": glosa.eps,
        "codigo_glosa": glosa.codigo_glosa,
        "gestor_nombre": glosa.gestor_nombre,
        "tasa_par_eps_codigo_pct": tasa_par,
        "n_par": n_par,
        "tasa_gestor_pct": tasa_gestor,
        "n_gestor": n_gestor,
        "probabilidad_levantamiento_pct": prob,
    }


@router.get("/{glosa_id}/contexto-cartera")
def contexto_cartera_glosa(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R280 P1: contexto financiero (cartera) de una glosa.

    Drill-down económico para una glosa específica:
      - factura, valor_factura, saldo_factura
      - count de glosas en la misma factura
      - valor_objetado_total y abierto en la factura
      - eps, tercero_nit, tercero_nombre

    Útil al revisar una glosa para ver el peso económico
    real (no solo el valor objetado individual).
    """
    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    ESTADOS_CERRADOS = {"ACEPTADA", "LEVANTADA", "ARCHIVADA", "CONCILIADA"}

    factura = (glosa.factura or "").strip()
    if factura and factura != "N/A":
        otras = db.query(GlosaRecord).filter(GlosaRecord.factura == factura).all()
        count_factura = len(otras)
        obj_factura = sum(float(g.valor_objetado or 0) for g in otras)
        obj_abierto = sum(
            float(g.valor_objetado or 0)
            for g in otras
            if (g.estado or "").upper() not in ESTADOS_CERRADOS
        )
    else:
        count_factura = 1
        obj_factura = float(glosa.valor_objetado or 0)
        obj_abierto = obj_factura if (glosa.estado or "").upper() not in ESTADOS_CERRADOS else 0.0

    return {
        "glosa_id": glosa.id,
        "eps": glosa.eps,
        "factura": glosa.factura,
        "valor_factura": int(float(glosa.valor_factura or 0)),
        "saldo_factura": int(float(glosa.saldo_factura or 0)),
        "valor_objetado": int(float(glosa.valor_objetado or 0)),
        "valor_recuperado": int(float(glosa.valor_recuperado or 0)),
        "tercero_nit": glosa.tercero_nit,
        "tercero_nombre": glosa.tercero_nombre,
        "factura_resumen": {
            "count_glosas": count_factura,
            "valor_objetado_total": int(obj_factura),
            "valor_objetado_abierto": int(obj_abierto),
        },
    }


@router.get("/{glosa_id}/audit-resumen")
def audit_resumen_glosa(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R88 P2: resumen agregado del audit log de una glosa.

    Mientras /audit/glosa/{id} devuelve los eventos en bruto y
    /glosas/{id}/timeline los enriquece narrativamente, este
    endpoint da el "TL;DR":
      - total_cambios
      - primer_cambio_en / ultimo_cambio_en
      - usuarios_que_intervinieron (lista DISTINCT)
      - eventos_por_accion (conteo)
      - eventos_por_campo (qué columnas se modificaron y cuántas veces)

    Útil para mostrar un mini-widget "Actividad" en la ficha
    de la glosa sin tener que renderizar todo el audit raw.
    """
    from app.models.db import AuditLogRecord

    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    eventos = (
        db.query(AuditLogRecord)
        .filter(AuditLogRecord.tabla == "glosas")
        .filter(AuditLogRecord.registro_id == glosa_id)
        .all()
    )

    total = len(eventos)
    if total == 0:
        return {
            "glosa_id": glosa_id,
            "total_cambios": 0,
            "primer_cambio_en": None,
            "ultimo_cambio_en": None,
            "usuarios_que_intervinieron": [],
            "eventos_por_accion": {},
            "eventos_por_campo": {},
        }

    timestamps = [e.timestamp for e in eventos if e.timestamp]
    usuarios = sorted({e.usuario_email for e in eventos if e.usuario_email})

    por_accion: dict[str, int] = {}
    por_campo: dict[str, int] = {}
    for e in eventos:
        if e.accion:
            por_accion[e.accion] = por_accion.get(e.accion, 0) + 1
        if e.campo:
            por_campo[e.campo] = por_campo.get(e.campo, 0) + 1

    return {
        "glosa_id": glosa_id,
        "total_cambios": total,
        "primer_cambio_en": (min(timestamps).isoformat() if timestamps else None),
        "ultimo_cambio_en": (max(timestamps).isoformat() if timestamps else None),
        "usuarios_que_intervinieron": usuarios,
        "eventos_por_accion": por_accion,
        "eventos_por_campo": por_campo,
    }


@router.get("/{glosa_id}/timeline")
def timeline_glosa(
    glosa_id: int,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """R67 P2: cronología consolidada de eventos de una glosa.

    Combina eventos de múltiples tablas en un solo timeline ordenado:
      - Creación de la glosa (creado_en del GlosaRecord)
      - Snapshots de versiones (DictamenVersionRecord — CREAR, REFINAR,
        REANALIZAR, RESTAURAR)
      - Cambios de estado (audit_log accion=ACTUALIZAR_ESTADO)
      - Decisión EPS (audit_log accion=DECISION_EPS)
      - Comentarios resueltos
      - Calls IA con costo (AICallRecord)

    Útil para:
      - Investigación de glosas con dictamen extraño
      - Auditoría regulatoria (¿quién tocó esta glosa, cuándo, qué hizo?)
      - Trazabilidad post-decisión de la EPS

    Respuesta: lista de {timestamp, tipo, actor, detalle, metadata}
    ordenada DESC (más reciente primero).
    """
    from app.models.db import (
        AICallRecord,
        AuditLogRecord,
        ComentarioGlosaRecord,
        DictamenVersionRecord,
    )

    glosa = GlosaRepository(db).obtener_por_id(glosa_id)
    if not glosa:
        raise HTTPException(404, "Glosa no encontrada")

    eventos = []

    # 1. Creación de la glosa
    if glosa.creado_en:
        eventos.append(
            {
                "timestamp": glosa.creado_en.isoformat(),
                "tipo": "CREAR_GLOSA",
                "actor": glosa.auditor_email or "—",
                "detalle": f"Glosa creada · {glosa.eps} · {glosa.codigo_glosa}",
                "metadata": {
                    "valor_objetado": float(glosa.valor_objetado or 0),
                    "estado": glosa.estado,
                },
            }
        )

    # 2. Versiones del dictamen
    versiones = (
        db.query(DictamenVersionRecord).filter(DictamenVersionRecord.glosa_id == glosa_id).all()
    )
    for v in versiones:
        eventos.append(
            {
                "timestamp": v.creado_en.isoformat() if v.creado_en else None,
                "tipo": f"VERSION_{v.accion or 'CREAR'}",
                "actor": v.autor_email or "—",
                "detalle": v.mensaje_refinar or f"Snapshot del dictamen ({v.accion})",
                "metadata": {"version_id": v.id},
            }
        )

    # 3. Audit log para esta glosa
    auditorias = (
        db.query(AuditLogRecord)
        .filter(
            AuditLogRecord.tabla.in_(("glosas", "historial")),
            AuditLogRecord.registro_id == glosa_id,
        )
        .all()
    )
    for a in auditorias:
        eventos.append(
            {
                "timestamp": a.timestamp.isoformat() if a.timestamp else None,
                "tipo": f"AUDIT_{a.accion or 'ACCION'}",
                "actor": a.usuario_email or "—",
                "detalle": (a.detalle or "")[:300],
                "metadata": {
                    "campo": a.campo,
                    "valor_anterior": (a.valor_anterior or "")[:80],
                    "valor_nuevo": (a.valor_nuevo or "")[:80],
                    "ip": a.ip,
                },
            }
        )

    # 4. Comentarios resueltos
    comentarios = (
        db.query(ComentarioGlosaRecord).filter(ComentarioGlosaRecord.glosa_id == glosa_id).all()
    )
    for c in comentarios:
        if c.creado_en:
            eventos.append(
                {
                    "timestamp": c.creado_en.isoformat(),
                    "tipo": "COMENTARIO",
                    "actor": c.autor_email or "—",
                    "detalle": (c.texto or "")[:300],
                    "metadata": {"resuelto": bool(c.resuelto_en)},
                }
            )
        if c.resuelto_en:
            eventos.append(
                {
                    "timestamp": c.resuelto_en.isoformat(),
                    "tipo": "COMENTARIO_RESUELTO",
                    "actor": c.resuelto_por or "—",
                    "detalle": "Comentario marcado como resuelto",
                    "metadata": {"comentario_id": c.id},
                }
            )

    # 5. Calls IA con costo
    calls = db.query(AICallRecord).filter(AICallRecord.glosa_id == glosa_id).all()
    for c in calls:
        if c.creado_en:
            eventos.append(
                {
                    "timestamp": c.creado_en.isoformat(),
                    "tipo": "AI_CALL",
                    "actor": c.user_email or "—",
                    "detalle": f"{c.proveedor}/{c.modelo} · {c.latency_ms}ms · ${c.cost_usd:.5f}",
                    "metadata": {
                        "tokens_in": (c.input_tokens or 0)
                        + (c.cache_creation_input_tokens or 0)
                        + (c.cache_read_input_tokens or 0),
                        "tokens_out": c.output_tokens,
                        "cost_usd": c.cost_usd,
                    },
                }
            )

    # Ordenar DESC (más reciente primero)
    eventos.sort(key=lambda e: e.get("timestamp") or "", reverse=True)

    return {
        "glosa_id": glosa_id,
        "total_eventos": len(eventos),
        "eventos": eventos,
    }


class _PreviewAuditoriaIn(BaseModel):
    texto_glosa: str = Field(..., min_length=1, max_length=10_000)
    eps: str = Field(default="")


@router.post("/preview-auditoria")
def preview_auditoria(
    body: _PreviewAuditoriaIn,
    db: Session = Depends(get_db),
    current_user: UsuarioRecord = Depends(get_usuario_actual),
):
    """Analiza el texto de una glosa sin generar dictamen: detecta patrones
    problemáticos, calcula score de defensibilidad y sugiere acción."""
    texto = body.texto_glosa.upper()
    eps = (body.eps or "").upper().strip()

    # Detectar si HUS tiene contrato con la EPS
    tiene_contrato = False
    if eps:
        tiene_contrato = bool(
            db.query(ContratoRecord).filter(ContratoRecord.eps.ilike(f"%{eps[:15]}%")).first()
        )

    hallazgos: list[dict] = []
    score = 40  # base

    # Patrón: "sin contrato entre las partes" cuando SÍ existe contrato
    if tiene_contrato and re.search(r"SIN\s+CONTRATO", texto):
        hallazgos.append(
            {
                "id": "afirmacion_sin_contrato_falsa",
                "severidad": "ALTA",
                "descripcion": "La EPS afirma ausencia de contrato pero HUS tiene contrato activo.",
                "argumento": "Refuta directamente citando número y vigencia del contrato.",
            }
        )
        score += 25

    # Patrón: SOAT como sustituto unilateral
    if re.search(r"SE\s+RECONOCE\s+(A\s+)?SOAT", texto) or re.search(
        r"RECONOCE.*SOAT.*VIGENTE", texto
    ):
        hallazgos.append(
            {
                "id": "soat_sustituto_indebido",
                "severidad": "ALTA",
                "descripcion": "La EPS propone SOAT como tarifa sustituta sin soporte contractual.",
                "argumento": "PACTA SUNT SERVANDA: solo aplica la tarifa pactada, no la impuesta.",
            }
        )
        score += 20

    # Patrón: "diferencia sin referente" — la EPS dice "se glosa la diferencia"
    # sin especificar la base de cálculo contractual (siempre es vicio inmotivación)
    if re.search(r"SE\s+GLOSA\s+LA\s+DIFERENCIA", texto):
        hallazgos.append(
            {
                "id": "diferencia_sin_referente",
                "severidad": "MEDIA",
                "descripcion": "La EPS glosa una diferencia sin indicar el valor de referencia.",
                "argumento": "Inmotivación: Decreto 4747/2007 Art. 21 exige sustentación precisa.",
            }
        )
        score += 15

    # Patrón: MVC (manejo vía contrato) sin especificación
    if re.search(r"\bMVC\b", texto):
        hallazgos.append(
            {
                "id": "mvc_sin_especificacion",
                "severidad": "MEDIA",
                "descripcion": "Uso de código MVC sin especificar el contrato de referencia.",
                "argumento": "Solicitar especificación del contrato y cláusula aplicable.",
            }
        )
        score += 10

    # Determinar acción sugerida
    score = min(score, 100)
    if score >= 70:
        accion = "DEFENDER_FUERTE"
    elif score >= 50:
        accion = "DEFENDER_MODERADO"
    elif score >= 30:
        accion = "REVISAR_CON_AUDITOR"
    else:
        accion = "EVALUAR_ACEPTACION"

    return {
        "hallazgos": hallazgos,
        "score_evidencia": score,
        "accion_sugerida": accion,
        "tiene_contrato_detectado": tiene_contrato,
        "total_hallazgos": len(hallazgos),
    }
