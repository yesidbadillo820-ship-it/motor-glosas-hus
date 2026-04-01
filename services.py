import io
import re
import asyncio
import logging
from datetime import datetime, timedelta

import PyPDF2
from groq import AsyncGroq
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_JUSTIFY, TA_CENTER
from reportlab.lib import colors

from models import GlosaInput, GlosaResult

logger = logging.getLogger("motor_glosas_v2")

FERIADOS_CO = [
    "2025-01-01", "2025-01-06", "2025-03-24", "2025-04-17", "2025-04-18", "2025-05-01", "2025-06-02", "2025-06-23", "2025-06-30", "2025-07-20", "2025-08-07", "2025-08-18", "2025-10-13", "2025-11-03", "2025-11-17", "2025-12-08", "2025-12-25",
    "2026-01-01", "2026-01-12", "2026-03-23", "2026-04-02", "2026-04-03", "2026-05-01", "2026-05-18", "2026-06-08", "2026-06-15", "2026-06-29", "2026-07-20", "2026-08-07", "2026-08-17", "2026-10-12", "2026-11-02", "2026-11-16", "2026-12-08", "2026-12-25"
]

CONTRATOS_FIJOS = {
    "COOSALUD": "CONTRATOS: 68001S00060339-24 y 68001C00060340-24. TARIFA: SOAT -15% e Institucionales. OBS: MAOS por HUS, Oncológicos por EPS.",
    "COMPENSAR": "CONTRATO: CSS009-2024. TARIFA: SOAT -15% y Tarifas Propias. OBS: Excluye oncológicos. MAOS por EPS.",
    "FAMISANAR": "CARTA DE INTENCIÓN. TARIFA: SOAT UVB -5% e Institucionales.",
    "FOMAG": "CONTRATO: 12076-359-2025. TARIFA: SOAT -15%, Institucionales y Paquetes (Tórax, IVE, Columna, Terapias, Gastro).",
    "LA PREVISORA": "CONTRATO: 12076-359-2025. TARIFA: SOAT -15% y Paquetes.",
    "DISPENSARIO MEDICO": "CONTRATO: 440-DIGSA/DMBUG-2025. TARIFA: SOAT SMLV -20% e Institucionales.",
    "POLICIA NACIONAL": "CONTRATOS: 068-5-200004-26 y 068-5-200006-26. TARIFA: SOAT UVB -8% e Institucionales. OBS: Contrato 0006-26 INCLUYE medicamentos oncológicos.",
    "NUEVA EPS": "CONTRATO: 02-01-06-00077-2017. TARIFA: SOAT -20% e Institucionales. OBS: Meds Oncológicos por HUS.",
    "PPL": "CONTRATO: IPS-001B-2022 (Otrosí 26). TARIFA: SOAT -15%. OBS: MAOS y Meds por HUS.",
    "FIDUCIARIA CENTRAL": "CONTRATO: IPS-001B-2022 (Otrosí 26). TARIFA: SOAT -15%.",
    "POSITIVA": "CONTRATO: 525 - OTROSÍ 3. TARIFA: SOAT SMLV -15%. OBS: Solo accidentes/laboral.",
    "PRECIMED": "CONTRATO: 319 DE 2024. TARIFA: Tarifas anexos / Institucionales.",
    "SALUD MIA": "CONTRATOS: SSA2025EVE3A005 y CSA2025EVE3A005. TARIFA: SOAT -15%. OBS: Urgencias Circular 019/2023.",
    "AURORA": "CONTRATOS: GID ARL 0090 y GID AP 0090. TARIFA: SOAT -3%.",
    "SECRETARIA DE SANTANDER": "MARCO LEGAL: Resolución 15997 de 2017 (Tarifas obligatorias ente territorial).",
    "SUMIMEDICAL": "CONTRATO: FPS23-050. TARIFA: SOAT -15%. OBS: MAOS y Oncológicos por EPS.",
    "OTRA / SIN DEFINIR": "SIN CONTRATO PACTADO. TARIFA: SOAT PLENO (RESOLUCIÓN 054 DE 2026_0001 / DECRETO 441 DE 2022)."
}

def _div(texto): 
    return f'<div style="text-align:justify;line-height:1.8;font-size:11px;">{texto}</div>'

def _tabla_defensa(codigo, servicio, valor, cod_res, desc_res):
    return f'<table border="1" style="width:100%;border-collapse:collapse;text-transform:uppercase;font-size:11px;margin-bottom:15px;"><tr style="background-color:#1e3a8a;color:white;"><th style="padding:8px;border:1px solid #cbd5e1;">CÓDIGO GLOSA</th><th style="padding:8px;border:1px solid #cbd5e1;">SERVICIO RECLAMADO</th><th style="padding:8px;border:1px solid #cbd5e1;">VALOR OBJ.</th><th style="padding:8px;border:1px solid #cbd5e1;background-color:#10b981;">CONCEPTO</th></tr><tr><td style="padding:8px;border:1px solid #cbd5e1;text-align:center;">{codigo}</td><td style="padding:8px;border:1px solid #cbd5e1;">{servicio}</td><td style="padding:8px;border:1px solid #cbd5e1;text-align:center;">{valor}</td><td style="padding:8px;border:1px solid #cbd5e1;text-align:center;font-weight:bold;">{cod_res}<br><span style="font-size:9px;">{desc_res}</span></td></tr></table>'

def _tabla_aceptacion(codigo, valor_obj, valor_acep, cod_res, desc_res):
    return f'<table border="1" style="width:100%;border-collapse:collapse;text-transform:uppercase;font-size:11px;margin-bottom:15px;"><tr style="background-color:#1e3a8a;color:white;"><th style="padding:8px;border:1px solid #cbd5e1;">CÓDIGO GLOSA</th><th style="padding:8px;border:1px solid #cbd5e1;">VALOR OBJETADO</th><th style="padding:8px;border:1px solid #cbd5e1;background-color:#d97706;">VALOR ACEPTADO</th><th style="padding:8px;border:1px solid #cbd5e1;background-color:#10b981;">CONCEPTO</th></tr><tr><td style="padding:8px;border:1px solid #cbd5e1;text-align:center;">{codigo}</td><td style="padding:8px;border:1px solid #cbd5e1;text-align:center;">{valor_obj}</td><td style="padding:8px;border:1px solid #cbd5e1;text-align:center;font-weight:bold;color:#d97706;">{valor_acep}</td><td style="padding:8px;border:1px solid #cbd5e1;text-align:center;font-weight:bold;">{cod_res}<br><span style="font-size:9px;">{desc_res}</span></td></tr></table>'

def _tabla_simple(codigo, estado, valor, cod_res, desc_res, color_header="#1e3a8a", color_estado="#b91c1c"):
    return f'<table border="1" style="width:100%;border-collapse:collapse;text-transform:uppercase;font-size:11px;margin-bottom:15px;"><tr style="background-color:{color_header};color:white;"><th style="padding:8px;border:1px solid #cbd5e1;">CÓDIGO GLOSA</th><th style="padding:8px;border:1px solid #cbd5e1;">ESTADO</th><th style="padding:8px;border:1px solid #cbd5e1;">VALOR</th><th style="padding:8px;border:1px solid #cbd5e1;background-color:#10b981;">CONCEPTO</th></tr><tr><td style="padding:8px;border:1px solid #cbd5e1;text-align:center;">{codigo}</td><td style="padding:8px;border:1px solid #cbd5e1;text-align:center;background-color:{color_estado};color:white;"><b>{estado}</b></td><td style="padding:8px;border:1px solid #cbd5e1;text-align:center;">{valor}</td><td style="padding:8px;border:1px solid #cbd5e1;text-align:center;font-weight:bold;">{cod_res}<br><span style="font-size:9px;">{desc_res}</span></td></tr></table>'

def _procesar_pdf_sync(file_content: bytes) -> str:
    unido = ""
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(file_content)) as pdf:
            for i, page in enumerate(pdf.pages):
                txt = page.extract_text() or ""
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        txt += " | ".join([str(c).replace('\n', ' ') if c else "" for c in row]) + "\n"
                unido += f"\n--- PÁG {i+1} ---\n{txt}"
    except Exception:
        reader = PyPDF2.PdfReader(io.BytesIO(file_content))
        for i in range(len(reader.pages)):
            txt = reader.pages[i].extract_text()
            if txt: unido += f"\n--- PÁG {i+1} ---\n{txt}"
    return unido[:4000] + "\n...[RECORTADO]...\n" + unido[-4000:] if len(unido) > 8000 else unido

def _calcular_dias_habiles(f_rad, f_rec):
    try:
        d1, d2 = datetime.strptime(f_rad, "%Y-%m-%d"), datetime.strptime(f_rec, "%Y-%m-%d")
        dias, current = 0, d1
        while current < d2:
            current += timedelta(days=1)
            if current.weekday() < 5 and current.strftime("%Y-%m-%d") not in FERIADOS_CO: dias += 1
        return dias
    except: return 0

class GlosaService:
    def __init__(self, api_key: str):
        self.cliente = AsyncGroq(api_key=api_key)

    async def extraer_pdf(self, file_content: bytes) -> str:
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, _procesar_pdf_sync, file_content)

    def xml(self, tag: str, texto: str, default: str = "") -> str:
        m = re.search(fr'<{tag}>(.*?)</{tag}>', texto, re.IGNORECASE | re.DOTALL)
        return m.group(1).strip().replace("**", "") if m else default

    async def analizar(self, data: GlosaInput, contexto_pdf: str = "", contratos_db: dict = None) -> GlosaResult:
        if contratos_db is None: contratos_db = {}
        contratos_activos = {**CONTRATOS_FIJOS, **contratos_db}

        eps_segura = str(data.eps).upper() if data.eps else "OTRA"
        etapa_segura = str(data.etapa).strip().upper()
        
        eps_key = eps_segura.replace(" / SIN DEFINIR", "").strip()
        info_c = contratos_activos.get(eps_key, CONTRATOS_FIJOS["OTRA / SIN DEFINIR"])

        texto_base = str(data.tabla_excel).strip()
        val_ac_num = float(re.sub(r'[^\d]', '', str(data.valor_aceptado)) or 0)
        
        cod_m = re.search(r'\b([A-Z]{2,3}\d{3,4})\b', texto_base)
        codigo_detectado = cod_m.group(1) if cod_m else "N/A"
        
        prefijo = codigo_detectado[:2].upper() if codigo_detectado != "N/A" else "XX"
        if "MCV" in texto_base.upper() or prefijo == "MC" or prefijo == "MV":
            prefijo = "TA" 

        val_m = re.search(r'\$\s*([\d\.,]+)', texto_base)
        valor_obj_raw = f"$ {val_m.group(1)}" if val_m else "$ 0.00"

        nombres_glosa = {
            "TA": "TARIFAS", "MV": "TARIFAS", "MC": "TARIFAS", "SO": "SOPORTES", "FA": "FACTURACIÓN",
            "PE": "PERTINENCIA", "AU": "AUTORIZACIÓN", "CO": "COBERTURA"
        }
        
        if "MAYOR VALOR" in texto_base.upper() or prefijo == "TA":
            nombre_tipo = "TARIFAS Y MAYOR VALOR COBRADO"
            prefijo_evaluacion = "TA"
        else:
            nombre_tipo = nombres_glosa.get(prefijo, "OBJECIONES VARIAS")
            prefijo_evaluacion = prefijo

        dias = _calcular_dias_habiles(data.fecha_radicacion, data.fecha_recepcion) if data.fecha_radicacion and data.fecha_recepcion else 0
        es_extemporanea = dias > 20
        dias_restantes = max(0, 20 - dias)

        msg_tiempo = f"EXTEMPORÁNEA ({dias} DÍAS)" if es_extemporanea else f"EN TÉRMINOS ({dias} DÍAS)"
        color_tiempo = "bg-red-600" if es_extemporanea else "bg-emerald-500"

        if "RATIF" in str(data.etapa).upper() and val_ac_num <= 0:
            apertura_txt = f"ESE HUS NO ACEPTA LA GLOSA POR {nombre_tipo} ({codigo_detectado}) EN INSTANCIA DE RATIFICACIÓN. NO SE APORTAN NUEVOS ELEMENTOS DE JUICIO QUE DESVIRTÚEN LA DEFENSA INICIAL. SE SOLICITA CONCILIACIÓN SEGÚN LEY 1438 DE 2011."
            tabla_html = _tabla_simple(codigo_detectado, "RATIFICACIÓN", valor_obj_raw, "RE9901", "GLOSA INJUSTIFICADA", color_estado="#2563eb")
            return GlosaResult(tipo="LEGAL - RATIFICADA", resumen="RECHAZO RATIFICACIÓN", dictamen=tabla_html + _div(apertura_txt), codigo_glosa=codigo_detectado, valor_objetado=valor_obj_raw, paciente="N/A", mensaje_tiempo=msg_tiempo, color_tiempo="bg-blue-600", dias_restantes=dias_restantes)

        if es_extemporanea and val_ac_num <= 0:
            apertura_txt = f"ESE HUS NO ACEPTA LA GLOSA POR {nombre_tipo} ({codigo_detectado}) POR CONSIDERARLA INJUSTIFICADA DEBIDO A SU EXTEMPORANEIDAD. AL TRANSCURRIR {dias} DÍAS HÁBILES, OPERA DE PLENO DERECHO LA ACEPTACIÓN TÁCITA (ART. 57 LEY 1438/2011). SE EXIGE EL PAGO INMEDIATO."
            tabla_html = _tabla_simple(codigo_detectado, "EXTEMPORÁNEA", valor_obj_raw, "RE9502", "GLOSA INJUSTIFICADA")
            return GlosaResult(tipo="LEGAL - EXTEMPORÁNEA", resumen="RECHAZO EXTEMPORANEIDAD", dictamen=tabla_html + _div(apertura_txt), codigo_glosa=codigo_detectado, valor_objetado=valor_obj_raw, paciente="N/A", mensaje_tiempo=msg_tiempo, color_tiempo=color_tiempo, dias_restantes=0)

        estrategias = {
            "TA": "P1: Cita contrato o Res. 054. P2: Desvirtúa descuento abusivo de EPS y justifica que la tarifa facturada es la correcta según el contrato. P3: Exige pago por Buena Fe (Art 871 C.Co).",
            "SO": "P1: Cita contrato. P2: Demuestra que la Historia Clínica es plena prueba (Res 1995/99). P3: Exige pago.",
            "PE": "P1: Cita contrato. P2: Justifica pertinencia clínica del acto médico basado en autonomía profesional (Ley 1751/15). P3: Exige pago.",
            "AU": "P1: Cita contrato. P2: Demuestra urgencia vital o trámite de autorización (Decreto 4747/07). P3: Exige pago."
        }
        est_actual = estrategias.get(prefijo_evaluacion, estrategias["PE"])

        few_shot_examples = """
        <ejemplos_de_respuestas_perfectas>
        --- EJEMPLO 1 (GLOSA TARIFARIA / MAYOR VALOR) ---
        USUARIO: EPS: DISPENSARIO. GLOSA: MVC001 Mayor valor cobrado. Valor Glosado: $ 150.000.
        ASISTENTE:
        <paciente>CARLOS PEREZ</paciente>
        <factura>F-1020</factura>
        <autorizacion>N/A</autorizacion>
        <codigo_glosa>MVC001</codigo_glosa>
        <valor_objetado>$ 150.000</valor_objetado>
        <servicio_glosado>ATENCIÓN MÉDICA</servicio_glosado>
        <motivo_resumido>MAYOR VALOR COBRADO</motivo_resumido>
        <score>98</score>
        <argumento>SE DEBE DESTACAR QUE EL CONTRATO VIGENTE ENTRE LAS PARTES ESTABLECE CLARAMENTE LAS TARIFAS APLICABLES PARA LOS PROCEDIMIENTOS Y SERVICIOS FACTURADOS.
        
        SE DESVIRTÚA LA OBJECIÓN DE LA EPS, YA QUE EL HOSPITAL LIQUIDÓ EL VALOR DEL SERVICIO CORRECTAMENTE Y CON APEGO ESTRICTO AL MANUAL TARIFARIO PACTADO. EN ESTE SENTIDO, LA FACTURACIÓN SE AJUSTA A LAS TARIFAS ESTABLECIDAS, POR LO QUE NO EXISTE MAYOR VALOR COBRADO.
        
        EXIGIMOS EL PAGO INMEDIATO DEL VALOR OBJETADO, TODA VEZ QUE EL ARTÍCULO 871 DEL CÓDIGO DE COMERCIO COLOMBIANO OBLIGA A LAS PARTES A EJECUTAR SUS ACTUACIONES DE BUENA FE. LA NEGATIVA AL PAGO CONSTITUYE UN INCUMPLIMIENTO INJUSTIFICADO, REQUIRIENDO EL REINTEGRO TOTAL A FAVOR DE NUESTRA INSTITUCIÓN.</argumento>
        </ejemplos_de_respuestas_perfectas>
        """

        system_prompt = f"""Eres el DIRECTOR JURÍDICO de la ESE HUS. DEBES RESPONDER EN XML. TODO EN MAYÚSCULAS. EXACTAMENTE 3 PÁRRAFOS TÉCNICOS.
        MARCO A DEFENDER Y CITAR: {info_c}
        ESTRATEGIA APLICABLE: {est_actual}
        
        {few_shot_examples}
        
        DEVUELVE: <paciente>, <factura>, <autorizacion>, <codigo_glosa>, <valor_objetado>, <servicio_glosado>, <score> (0-100), <argumento>."""
        
        user_prompt = f"EPS: {eps_segura}. GLOSA: {texto_base}\nSOPORTES: {contexto_pdf[:5000]}"

        try:
            comp = await self.cliente.chat.completions.create(
                messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": user_prompt}],
                model="llama-3.3-70b-versatile", temperature=0.2, max_tokens=2000
            )
            res_ia = comp.choices[0].message.content
        except Exception as e:
            res_ia = f"<argumento>ERROR DE IA: {str(e)}</argumento>"

        paciente = self.xml("paciente", res_ia, "NO IDENTIFICADO")
        servicio = self.xml("servicio_glosado", res_ia, "SERVICIOS ASISTENCIALES")
        codigo_final = self.xml("codigo_glosa", res_ia, codigo_detectado)
        if len(codigo_final) > 10 or codigo_final == "N/A": codigo_final = codigo_detectado
        
        valor_xml = self.xml("valor_objetado", res_ia, valor_obj_raw)
        if "$" not in valor_xml and not any(char.isdigit() for char in valor_xml): valor_xml = valor_obj_raw
        
        argumento_ia = self.xml("argumento", res_ia, "").replace('\n\n', '<br/><br/>')
        
        if val_ac_num > 0:
            apertura = f"ESE HUS ACEPTA PARCIALMENTE LA GLOSA {codigo_final} POR $ {val_ac_num:,.0f}."
            tabla_html = _tabla_aceptacion(codigo_final, valor_xml, f"$ {val_ac_num:,.0f}", "RE9801", "GLOSA ACEPTADA")
            tipo = "AUDITORÍA - ACEPTADA"
        else:
            apertura = f"ESE HUS NO ACEPTA LA GLOSA POR {nombre_tipo} ({codigo_final}) POR CONSIDERARLA INJUSTIFICADA, SUSTENTANDO ASÍ:"
            tabla_html = _tabla_defensa(codigo_final, servicio, valor_xml, "RE9602", "GLOSA INJUSTIFICADA")
            tipo = "TÉCNICO-LEGAL"

        dictamen_final = tabla_html + _div(f"<b>{apertura}</b><br/><br/>{argumento_ia}")

        return GlosaResult(
            tipo=tipo, resumen=f"DEFENSA: {paciente}", dictamen=dictamen_final,
            codigo_glosa=codigo_final, valor_objetado=valor_xml, paciente=paciente,
            mensaje_tiempo=msg_tiempo, color_tiempo=color_tiempo, factura=self.xml("factura", res_ia, "N/A"), 
            autorizacion=self.xml("autorizacion", res_ia, "N/A"), score=int(self.xml("score", res_ia, "100") or 100), dias_restantes=dias_restantes
        )

def crear_oficio_pdf(eps: str, resumen: str, conclusion: str) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=50, leftMargin=50, topMargin=50, bottomMargin=50)
    estilos = getSampleStyleSheet()
    estilo_n = ParagraphStyle('n', alignment=TA_JUSTIFY, fontSize=11, leading=16)
    
    elements = [
        Paragraph("<b>ESE HOSPITAL UNIVERSITARIO DE SANTANDER</b>", ParagraphStyle('h1', alignment=TA_CENTER, fontSize=14)),
        Spacer(1, 20), Paragraph(f"<b>Señores:</b><br/>{eps.upper()}<br/><b>Ref:</b> {resumen}", estilo_n), Spacer(1, 15)
    ]
    
    clean_text = re.sub(r'<br\s*/?>', '\n', re.sub(r'<table.*?>.*?</table>', '', conclusion, flags=re.IGNORECASE | re.DOTALL)).strip()
    for p in clean_text.split('\n\n'):
        if p.strip(): elements.append(Paragraph(p.strip(), estilo_n))
        elements.append(Spacer(1, 8))

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()

def exportar_excel_pro(glosas: list) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Glosas HUS"

    headers = ["ID", "Fecha Procesamiento", "EPS / Pagador", "Paciente", "Codigo Glosa", "Valor Objetado", "Valor Aceptado", "Etapa Procesal", "Estado Final"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1C3460", end_color="1C3460", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 18

    ws.column_dimensions['C'].width = 30 
    ws.column_dimensions['D'].width = 30 

    for g in glosas:
        ws.append([
            g.id, g.creado_en.strftime("%d/%m/%Y %H:%M"), g.eps, g.paciente, 
            g.codigo_glosa, g.valor_objetado, g.valor_aceptado, g.etapa, g.estado
        ])

    ws.auto_filter.ref = ws.dimensions

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
